import tkinter as tk
from tkinter import ttk  # เปลี่ยนมาใช้ ttk จาก tkinter โดยตรง
from tkinter import filedialog, messagebox
import pandas as pd
import pyreadstat
import os
from openpyxl.styles import Font, PatternFill
import sys # เพิ่ม sys สำหรับ sys.exit

class SpssConverterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("โปรแกรมแปลงไฟล์ SPSS To Excel")

        # --- จัดหน้าต่างให้อยู่กลางจอ (โค้ดส่วนนี้ใช้ tkinter ปกติอยู่แล้ว) ---
        window_width = 700
        window_height = 420
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        center_x = int((screen_width / 2) - (window_width / 2))
        center_y = int((screen_height / 2) - (window_height / 2))
        self.root.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
        # --------------------------------

        # --- เปลี่ยนมาใช้ StringVar ของ tkinter ---
        self.file_path = tk.StringVar()
        self.conversion_type = tk.StringVar(value="code")
        self.status_text = tk.StringVar()
        self.status_text.set("กรุณาเลือกไฟล์ SPSS (.sav) ที่ต้องการแปลง")

        self._create_widgets()

    def _create_widgets(self):
        """
        สร้างและจัดวาง Widgets ใหม่ทั้งหมดโดยใช้ .grid()
        """
        # --- สร้าง Frame หลัก ---
        # ใช้ ttk.Frame จาก tkinter.ttk
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # กำหนดให้คอลัมน์ที่ 1 (ช่องใส่ Path) ขยายตามขนาดหน้าต่าง
        main_frame.columnconfigure(1, weight=1)

        # --- ส่วนหัวโปรแกรม ---
        # ใช้ ttk.Label และลบ bootstyle ออก
        header_label = ttk.Label(main_frame, text="โปรแกรมแปลงไฟล์ SPSS เป็น Excel", font=("Helvetica", 16, "bold"))
        header_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))

        # --- ขั้นตอนที่ 1: เลือกไฟล์ ---
        step1_label = ttk.Label(main_frame, text="ขั้นตอนที่ 1: เลือกไฟล์", font=("Helvetica", 11, "bold"))
        step1_label.grid(row=1, column=0, columnspan=3, sticky=tk.W, pady=(0, 5))

        # ใช้ ttk.Entry
        file_entry = ttk.Entry(main_frame, textvariable=self.file_path, state="readonly", font=("Helvetica", 10))
        file_entry.grid(row=2, column=0, columnspan=2, sticky="ew", padx=(0, 10))

        # ใช้ ttk.Button และลบ bootstyle ออก
        browse_button = ttk.Button(main_frame, text="เลือกไฟล์...", command=self.select_file)
        browse_button.grid(row=2, column=2, sticky="ew")

        # --- ขั้นตอนที่ 2: เลือกรูปแบบการแปลง ---
        step2_label = ttk.Label(main_frame, text="ขั้นตอนที่ 2: เลือกรูปแบบการแปลง", font=("Helvetica", 11, "bold"))
        step2_label.grid(row=3, column=0, columnspan=3, sticky=tk.W, pady=(20, 5))

        # จัด Radiobutton ใน Frame ย่อยเพื่อให้ดูเป็นกลุ่ม
        options_frame = ttk.Frame(main_frame)
        options_frame.grid(row=4, column=0, columnspan=3, sticky=tk.W, padx=15)

        # ใช้ ttk.Radiobutton และลบ bootstyle ออก
        ttk.Radiobutton(options_frame, text="Code (ข้อมูลตัวเลขดิบ)", variable=self.conversion_type, value="code").pack(anchor=tk.W, pady=2)
        ttk.Radiobutton(options_frame, text="Label (ข้อความกำกับ)", variable=self.conversion_type, value="label").pack(anchor=tk.W, pady=2)
        ttk.Radiobutton(options_frame, text="Code และ Label (สร้าง 2 ชีทแยกกัน)", variable=self.conversion_type, value="code_and_label_separate").pack(anchor=tk.W, pady=2)

        # --- ขั้นตอนที่ 3: เริ่มการแปลงไฟล์ ---
        # ใช้ ttk.Separator และลบ bootstyle ออก
        separator = ttk.Separator(main_frame, orient='horizontal')
        separator.grid(row=5, column=0, columnspan=3, pady=20, sticky="ew")

        # ใช้ ttk.Button และลบ bootstyle กับ padding ที่เป็นของ bootstrap ออก
        convert_button = ttk.Button(main_frame, text="แปลงไฟล์เป็น Excel", command=self.convert_file)
        convert_button.grid(row=6, column=0, columnspan=3, pady=5, ipady=5) # ใช้ ipady เพื่อเพิ่ม padding ภายในปุ่ม

        # --- ส่วนแสดงสถานะและ Progress Bar ---
        # ใช้ ttk.Progressbar และลบ bootstyle ออก
        self.progress = ttk.Progressbar(main_frame, mode='determinate')
        self.progress.grid(row=7, column=0, columnspan=3, sticky="ew", pady=10)

        status_label = ttk.Label(main_frame, textvariable=self.status_text, wraplength=650, justify=tk.CENTER, font=("Helvetica", 9))
        status_label.grid(row=8, column=0, columnspan=3, pady=5)

    def select_file(self):
        filename = filedialog.askopenfilename(
            title="เลือกไฟล์ SPSS",
            filetypes=[("SPSS Files", "*.sav"), ("All files", "*.*")]
        )
        if filename:
            self.file_path.set(filename)
            self.status_text.set(f"ไฟล์ที่เลือก: {os.path.basename(filename)}")

    def create_qnr_sheet(self, meta):
        # โค้ดส่วนนี้ไม่มีส่วนเกี่ยวข้องกับ UI ไม่ต้องแก้ไข
        qnr_data = [
            ['SbjNum', 'SbjNum'],
            ['Filter', 'Filter'],
            ['Status', 'Status'],
        ]
        processed_bases = set()
        for i, col_name in enumerate(meta.column_names):
            base_name = None
            if "_O" in col_name and col_name.split('_O')[-1].isdigit():
                base_name = col_name.split('_O')[0] + '_O'
            elif "$" in col_name and col_name.split('$')[-1].isdigit():
                base_name = col_name.split('$')[0] + '$'
            if base_name:
                if base_name in processed_bases:
                    continue
                else:
                    processed_bases.add(base_name)
            var_label = meta.column_labels[i] if meta.column_labels else col_name
            qnr_data.append([col_name, var_label])
            if col_name in meta.variable_value_labels:
                value_labels = meta.variable_value_labels[col_name]
                for code, label in value_labels.items():
                    try:
                        code_display = int(code)
                    except (ValueError, TypeError):
                        code_display = code
                    qnr_data.append([code_display, label])
        df = pd.DataFrame(qnr_data, columns=['Q.No', 'QNR'])
        return df, qnr_data

    def format_qnr_sheet(self, worksheet, qnr_data):
        # โค้ดส่วนนี้ไม่มีส่วนเกี่ยวข้องกับ UI ไม่ต้องแก้ไข
        bold_font = Font(bold=True)
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        for cell in worksheet[1]:
            cell.font = bold_font
            cell.fill = gray_fill

        for i, row_data in enumerate(qnr_data):
            q_no_value = row_data[0]
            if isinstance(q_no_value, str) and q_no_value:
                row_index = i + 2
                for cell in worksheet[row_index]:
                    cell.font = bold_font
                    cell.fill = blue_fill

        worksheet.freeze_panes = 'B1'
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 80

    def convert_file(self):
        # โค้ดส่วนนี้ไม่มีส่วนเกี่ยวข้องกับ UI ไม่ต้องแก้ไข
        spss_path = self.file_path.get()
        if not spss_path:
            messagebox.showerror("ข้อผิดพลาด", "กรุณาเลือกไฟล์ SPSS ก่อน")
            return

        output_path = os.path.splitext(spss_path)[0] + ".xlsx"
        conv_type = self.conversion_type.get()

        self.progress['value'] = 0
        self.root.update_idletasks()

        try:
            self.status_text.set("กำลังอ่านไฟล์ SPSS และ Metadata...")
            self.progress['value'] = 10
            self.root.update_idletasks()
            df, meta = pyreadstat.read_sav(spss_path, apply_value_formats=False, dates_as_pandas_datetime=True)

            self.progress['value'] = 30
            self.root.update_idletasks()

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                self.status_text.set("กำลังเขียนข้อมูลลง Excel...")
                self.root.update_idletasks()

                if conv_type == "code_and_label_separate":
                    df.copy().to_excel(writer, sheet_name='Code', index=False)
                    df_label = df.copy()
                    for col in meta.variable_value_labels:
                        if col in df_label.columns:
                            mapping = meta.variable_value_labels[col]
                            df_label[col] = df_label[col].map(mapping).fillna(df_label[col])
                    df_label.to_excel(writer, sheet_name='Label', index=False)
                else:
                    sheet_name = 'Code' if conv_type == 'code' else 'Label'
                    df_to_save = df.copy()
                    if conv_type == "label":
                        for col in meta.variable_value_labels:
                            if col in df_to_save.columns:
                                mapping = meta.variable_value_labels[col]
                                df_to_save[col] = df_to_save[col].map(mapping).fillna(df_to_save[col])
                    df_to_save.to_excel(writer, sheet_name=sheet_name, index=False)

                self.progress['value'] = 70
                self.root.update_idletasks()

                self.status_text.set("กำลังสร้างและจัดรูปแบบชีท QNR...")
                self.root.update_idletasks()
                df_qnr, qnr_data_list = self.create_qnr_sheet(meta)
                df_qnr.to_excel(writer, sheet_name='QNR', index=False)
                worksheet_qnr = writer.sheets['QNR']
                self.format_qnr_sheet(worksheet_qnr, qnr_data_list)

            self.progress['value'] = 100
            self.status_text.set(f"แปลงไฟล์สำเร็จ! บันทึกที่:\n{output_path}")
            messagebox.showinfo("สำเร็จ", f"ไฟล์ถูกแปลงเรียบร้อยแล้ว\nบันทึกที่: {output_path}")

        except Exception as e:
            self.progress['value'] = 0
            self.status_text.set(f"เกิดข้อผิดพลาด: {e}")
            messagebox.showerror("เกิดข้อผิดพลาด", f"ไม่สามารถแปลงไฟล์ได้:\n{e}")

def run_this_app(working_dir=None):
    """
    ฟังก์ชันหลักสำหรับสร้างและรัน SpssConverterApp.
    """
    print(f"--- SPSS_CONVERTER_INFO: Starting 'SpssConverterApp' via run_this_app() ---")
    try:
        # --- เปลี่ยนจาก ttk.Window เป็น tk.Tk ---
        root = tk.Tk()
        app = SpssConverterApp(root)
        root.mainloop()

        print(f"--- SPSS_CONVERTER_INFO: SpssConverterApp mainloop finished. ---")

    except Exception as e:
        print(f"SPSS_CONVERTER_ERROR: An error occurred during SpssConverterApp execution: {e}")
        # ใช้ try-except เพื่อสร้าง root ชั่วคราวในกรณีที่ root หลักยังไม่ถูกสร้าง
        try:
            # ตรวจสอบว่า root มีอยู่และยังไม่ถูกทำลายหรือไม่
            if 'root' in locals() and root.winfo_exists():
                parent_window = root
            else:
                # ถ้าไม่มี ให้สร้างหน้าต่างชั่วคราวเพื่อแสดงกล่องข้อความ
                root_temp = tk.Tk()
                root_temp.withdraw() # ซ่อนหน้าต่างชั่วคราว
                parent_window = root_temp
                messagebox.showerror("Application Error", f"An unexpected error occurred:\n{e}", parent=parent_window)
                root_temp.destroy() # ทำลายหน้าต่างชั่วคราวหลังใช้งาน
        except Exception as inner_e:
             # กรณีเกิดข้อผิดพลาดซ้ำซ้อน
            print(f"SPSS_CONVERTER_FATAL_ERROR: Could not display error message box. Original error: {e}, Inner error: {inner_e}")

        sys.exit(f"Error running SpssConverterApp: {e}")

if __name__ == "__main__":
    print("--- Running SpssConverterApp.py directly for testing ---")
    run_this_app()
    print("--- Finished direct execution of SpssConverterApp.py ---")