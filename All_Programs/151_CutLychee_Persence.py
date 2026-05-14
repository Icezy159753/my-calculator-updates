from __future__ import annotations

import sys
import traceback
import os
import time
import re
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor, as_completed
from copy import copy
from pathlib import Path
from typing import List, Tuple

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange
from openpyxl.worksheet.hyperlink import Hyperlink
from bisect import bisect_left

from PyQt6.QtCore import (
    QEasingCurve,
    QObject,
    QPropertyAnimation,
    QThread,
    Qt,
    pyqtSignal,
)
from PyQt6.QtGui import QColor, QPalette
from PyQt6.QtWidgets import (
    QApplication,
    QFrame,
    QGraphicsOpacityEffect,
    QHBoxLayout,
    QLabel,
    QListWidget,
    QMainWindow,
    QMessageBox,
    QRadioButton,
    QPushButton,
    QTextEdit,
    QVBoxLayout,
    QWidget,
    QFileDialog,
    QProgressBar,
)


# ----------------------------
# Excel processing core
# ----------------------------
KEEP_COUNT = "count"
KEEP_PERCENT = "percent"
KEEP_BOTH = "both"
TABLE_ONE_SHEET_SIG = "one_sheet_sig"
TABLE_ONE_SHEET_NOT_SIG = "one_sheet_not_sig"
TABLE_MULTI_SHEET_SIG = "multi_sheet_sig"
TABLE_MULTI_SHEET_NOT_SIG = "multi_sheet_not_sig"
QUESTION_CODE_RE = re.compile(r"\b[A-Z]{1,3}\d+[A-Z]?(?:_[0-9]+|Z\d+)?\b", re.IGNORECASE)
SIG_TEXT_RE = re.compile(r"^[A-Z]+$")


def unique_output_path(out_dir: Path, source_path: Path, keep_mode: str = KEEP_COUNT) -> Path:
    base_stem = build_output_stem(source_path.stem, keep_mode)
    candidate = out_dir / f"{base_stem}.xlsx"
    if not candidate.exists():
        return candidate

    idx = 1
    while True:
        candidate = out_dir / f"{base_stem}_{idx}.xlsx"
        if not candidate.exists():
            return candidate
        idx += 1


def sheet_has_table_legend(sheet) -> bool:
    for row in range(1, sheet.max_row + 1):
        for col in range(1, 3):
            value = sheet.cell(row=row, column=col).value
            if value is None:
                continue
            text = str(value)
            if re.search(r"(?i)count", text) and re.search(r"(?i)column\s*%", text):
                return True
    return False


def count_stub_blocks(sheet) -> int:
    count = 0
    for row in range(1, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=1).value or "").strip().lower().startswith("stub:"):
            count += 1
    return count


def detect_one_sheet_layout(wb) -> bool:
    table_like_sheets = []
    for sheet in wb.worksheets:
        if sheet.title.strip().lower() in {"contents", "itemlist"}:
            continue
        if sheet_has_table_legend(sheet) or count_stub_blocks(sheet):
            table_like_sheets.append(sheet)

    if not table_like_sheets:
        return False
    if any(count_stub_blocks(sheet) > 1 for sheet in table_like_sheets):
        return True
    return len(table_like_sheets) == 1 and table_like_sheets[0].max_row > 200


def detect_sig_layout(wb) -> bool:
    label_merge_heights: list[int] = []

    for sheet in wb.worksheets:
        if sheet.title.strip().lower() in {"contents", "itemlist"}:
            continue
        if not sheet_has_table_legend(sheet):
            continue

        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            if min_col <= 2 and max_col <= 2 and max_row > min_row:
                text = str(sheet.cell(row=min_row, column=min_col).value or "")
                if re.search(r"(?i)1st\s*row", text):
                    continue
                label_merge_heights.append(max_row - min_row + 1)

    if not label_merge_heights:
        return False
    return max(label_merge_heights) >= 3


def detect_workbook_table_type(wb) -> str:
    is_one_sheet = detect_one_sheet_layout(wb)
    has_sig = detect_sig_layout(wb)

    if is_one_sheet and has_sig:
        return TABLE_ONE_SHEET_SIG
    if is_one_sheet and not has_sig:
        return TABLE_ONE_SHEET_NOT_SIG
    if not is_one_sheet and has_sig:
        return TABLE_MULTI_SHEET_SIG
    return TABLE_MULTI_SHEET_NOT_SIG


def build_output_stem(source_stem: str, keep_mode: str = KEEP_COUNT) -> str:
    today = datetime.now().strftime("%Y%m%d")
    stem = source_stem

    # Remove trailing processed marker if present.
    stem = re.sub(r"_processed(?:_\d+)?$", "", stem, flags=re.IGNORECASE)

    # Normalize any N% token in filename to the selected output type and keep it
    # as a clear suffix before the date.
    output_token = "%" if keep_mode == KEEP_PERCENT else "N"
    stem = re.sub(r"(?i)\bN\s*%", output_token, stem)

    date_match = re.search(r"(?:\s+)?\d{8}$", stem)
    if date_match:
        stem = stem[: date_match.start()].rstrip()

    stem = re.sub(r"(?i)(?:\s+)(?:N|%)$", "", stem.strip()).rstrip()
    stem = f"{stem} {output_token}" if stem else output_token

    stem = f"{stem} {today}"

    return stem


def excel_sheet_location(sheet_name: str, cell_ref: str = "A1") -> str:
    if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", sheet_name):
        return f"{sheet_name}!{cell_ref}"
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'!{cell_ref}"


def set_internal_link(
    ws,
    coord: str,
    target_sheet: str | None,
    text: str | None = None,
    cell_ref: str = "A1",
) -> None:
    cell = ws[coord]
    if text is not None:
        cell.value = text
    if not target_sheet:
        cell.hyperlink = None
        return
    cell.hyperlink = Hyperlink(
        ref=coord,
        location=excel_sheet_location(target_sheet, cell_ref),
        display=str(cell.value) if cell.value is not None else None,
    )


def sheet_from_internal_location(location: str) -> str | None:
    quoted = re.match(r"'((?:[^']|'')+)'!", location)
    if quoted:
        return quoted.group(1).replace("''", "'")
    unquoted = re.match(r"([^!]+)!", location)
    if unquoted:
        return unquoted.group(1)
    return None


def normalize_question_code(value: object) -> str | None:
    text = str(value or "")
    matches = QUESTION_CODE_RE.findall(text)
    if not matches:
        return None
    code = matches[0].upper()
    code = re.sub(r"Z\d+$", "", code)
    return code


def collect_table_anchors(wb, contents_sheet: str = "Contents") -> tuple[list[tuple[str, str]], dict[str, list[tuple[str, str]]]]:
    ordered: list[tuple[str, str]] = []
    by_code: dict[str, list[tuple[str, str]]] = {}

    for ws in wb.worksheets:
        if ws.title == contents_sheet:
            continue
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=1).value
            text = str(value or "").strip()
            if not text.lower().startswith("stub:"):
                continue
            anchor_row = row_idx
            for candidate_row in range(row_idx - 1, max(0, row_idx - 5), -1):
                if str(ws.cell(row=candidate_row, column=1).value or "").strip().lower() == "contents":
                    anchor_row = candidate_row
                    break
            anchor = (ws.title, f"A{anchor_row}")
            ordered.append(anchor)
            code = normalize_question_code(text)
            if code:
                by_code.setdefault(code, []).append(anchor)

    return ordered, by_code


def repair_contents_table_links(wb, contents_sheet: str = "Contents") -> None:
    if contents_sheet not in wb.sheetnames:
        return

    ws = wb[contents_sheet]
    ordered_anchors, anchors_by_code = collect_table_anchors(wb, contents_sheet)
    used_by_code: dict[str, int] = {}
    sequential_idx = 0

    for row_idx in range(1, ws.max_row + 1):
        table_col = None
        for candidate_col in (1, 2):
            if str(ws.cell(row=row_idx, column=candidate_col).value or "").strip().lower() == "table":
                table_col = candidate_col
                break
        if table_col is None:
            continue

        cell = ws.cell(row=row_idx, column=table_col)
        question_text = ws.cell(row=row_idx, column=table_col + 1).value
        code = normalize_question_code(question_text)
        anchor = None

        if code and code in anchors_by_code:
            code_idx = used_by_code.get(code, 0)
            choices = anchors_by_code[code]
            anchor = choices[min(code_idx, len(choices) - 1)]
            used_by_code[code] = code_idx + 1

        if anchor is None and sequential_idx < len(ordered_anchors):
            anchor = ordered_anchors[sequential_idx]

        sequential_idx += 1
        if anchor is None:
            cell.hyperlink = None
            continue

        target_sheet, cell_ref = anchor
        set_internal_link(ws, cell.coordinate, target_sheet, cell_ref=cell_ref)


def repair_internal_links(wb, contents_sheet: str = "Contents") -> None:
    existing = set(wb.sheetnames)
    navigable_sheets: List[str] = []
    has_contents_sheet = contents_sheet in existing

    repair_contents_table_links(wb, contents_sheet)

    for name in wb.sheetnames:
        if name == contents_sheet:
            continue
        ws = wb[name]
        if str(ws["A1"].value or "").strip().lower() == "contents":
            navigable_sheets.append(name)

    for idx, name in enumerate(navigable_sheets):
        ws = wb[name]
        if has_contents_sheet:
            set_internal_link(ws, "A1", contents_sheet)
        else:
            set_internal_link(ws, "A1", name)
        if str(ws["B1"].value or "").strip().lower() == "info" and "Info" in existing:
            set_internal_link(ws, "B1", "Info")

        prev_name = navigable_sheets[idx - 1] if idx > 0 else None
        next_name = navigable_sheets[idx + 1] if idx < len(navigable_sheets) - 1 else None

        if str(ws["E1"].value or "").strip().lower() == "previous":
            set_internal_link(ws, "E1", prev_name)
        if str(ws["F1"].value or "").strip().lower() == "next":
            set_internal_link(ws, "F1", next_name)

    for ws in wb.worksheets:
        first_contents_ref = None
        for row in ws.iter_rows():
            for cell in row:
                if str(cell.value or "").strip().lower() == "contents":
                    first_contents_ref = cell.coordinate
                    break
            if first_contents_ref:
                break

        for row in ws.iter_rows():
            for cell in row:
                if str(cell.value or "").strip().lower() == "contents":
                    if has_contents_sheet:
                        set_internal_link(ws, cell.coordinate, contents_sheet)
                    elif first_contents_ref:
                        set_internal_link(ws, cell.coordinate, ws.title, cell_ref=first_contents_ref)
                    continue

                if not cell.hyperlink:
                    continue
                location = cell.hyperlink.location
                if not location and isinstance(cell.hyperlink.target, str):
                    target = cell.hyperlink.target
                    if target.startswith("#"):
                        location = target[1:]
                if not location:
                    continue
                location_text = str(location)
                target_sheet = sheet_from_internal_location(location_text)
                if not target_sheet:
                    continue
                if target_sheet not in existing:
                    cell.hyperlink = None
                elif isinstance(cell.hyperlink.target, str) and cell.hyperlink.target.startswith("#"):
                    cell.hyperlink = Hyperlink(
                        ref=cell.coordinate,
                        location=location_text,
                        display=str(cell.value) if cell.value is not None else None,
                    )


def unique_output_path_reserved(
    out_dir: Path, source_path: Path, reserved_names: set[str], keep_mode: str = KEEP_COUNT
) -> Path:
    base = build_output_stem(source_path.stem, keep_mode)
    candidate_name = f"{base}.xlsx"
    idx = 1
    while True:
        candidate = out_dir / candidate_name
        key = str(candidate.resolve()).lower()
        if key not in reserved_names and not candidate.exists():
            reserved_names.add(key)
            return candidate
        candidate_name = f"{base}_{idx}.xlsx"
        idx += 1


def process_one_file_task(src_path: str, dst_path: str, keep_mode: str = KEEP_COUNT) -> str:
    process_workbook(Path(src_path), Path(dst_path), keep_mode)
    return dst_path


def process_workbook(input_path: Path, save_path: Path, keep_mode: str = KEEP_COUNT) -> None:
    if keep_mode not in {KEEP_COUNT, KEEP_PERCENT}:
        raise ValueError(f"Unsupported keep mode: {keep_mode}")

    wb = openpyxl.load_workbook(input_path)
    table_type = detect_workbook_table_type(wb)

    def detect_effective_max_col(sheet) -> int:
        # Use worksheet dimension first, then trim based on table values.
        # Rows 1-3 can contain navigation links such as Previous/Next outside the table.
        dim = sheet.calculate_dimension()
        _, _, dim_max_col, _ = range_boundaries(dim)

        max_row = sheet.max_row
        sample_rows = [4, 5, 6, 7, 8, 9, 10, max_row]
        sample_rows.extend(range(4, min(max_row, 30) + 1))
        if max_row > 30:
            sample_rows.extend(range(max_row - 9, max_row + 1))
        sample_rows = sorted(set(r for r in sample_rows if 1 <= r <= max_row))

        for col in range(dim_max_col, 2, -1):
            for r in sample_rows:
                value = sheet.cell(row=r, column=col).value
                if value is not None and str(value).strip() != "":
                    return col
        return 3

    def delete_rows_desc(sheet, row_indexes: List[int]) -> None:
        if not row_indexes:
            return
        rows = sorted(set(row_indexes), reverse=True)
        start = rows[0]
        count = 1
        prev = rows[0]
        for r in rows[1:]:
            if r == prev - 1:
                count += 1
            else:
                sheet.delete_rows(start - count + 1, count)
                start = r
                count = 1
            prev = r
        sheet.delete_rows(start - count + 1, count)

    def compact_rows_once(sheet, row_indexes: List[int]) -> None:
        deleted_rows = sorted(set(row_indexes))
        if not deleted_rows:
            return
        deleted_set = set(deleted_rows)
        original_merged_ranges = []
        for merged_range in sheet.merged_cells.ranges:
            top_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            original_merged_ranges.append(
                (
                    merged_range.bounds,
                    top_cell.value,
                    copy(top_cell.border),
                    copy(top_cell.fill),
                    copy(top_cell.font),
                    copy(top_cell.alignment),
                    top_cell.number_format,
                )
            )
        new_cells = {}
        for (row_idx, col_idx), cell in list(sheet._cells.items()):
            if row_idx in deleted_set:
                continue
            new_row = row_idx - bisect_left(deleted_rows, row_idx)
            if new_row != row_idx:
                cell.row = new_row
            new_cells[(new_row, col_idx)] = cell
        sheet._cells = new_cells

        new_row_dimensions = {}
        for row_idx, dimension in list(sheet.row_dimensions.items()):
            if row_idx in deleted_set:
                continue
            new_row = row_idx - bisect_left(deleted_rows, row_idx)
            dimension.index = new_row
            new_row_dimensions[new_row] = dimension
        sheet.row_dimensions = new_row_dimensions
        sheet._current_row = max((row for row, _ in sheet._cells.keys()), default=1)

        shifted_merges = []
        for bounds, merge_label, merge_border, merge_fill, merge_font, merge_alignment, merge_number_format in original_merged_ranges:
            min_col, min_row, max_col, max_row = bounds
            is_label_merge = min_col <= 2 and max_col <= 2
            if not is_label_merge:
                continue
            remaining_rows = [row_idx for row_idx in range(min_row, max_row + 1) if row_idx not in deleted_set]
            if not remaining_rows:
                continue
            new_min_row = remaining_rows[0] - bisect_left(deleted_rows, remaining_rows[0])
            new_max_row = remaining_rows[-1] - bisect_left(deleted_rows, remaining_rows[-1])
            start = f"{get_column_letter(min_col)}{new_min_row}"
            end = f"{get_column_letter(max_col)}{new_max_row}"
            should_merge = not (new_min_row == new_max_row and min_col == max_col)
            if should_merge:
                shifted_merges.append(CellRange(f"{start}:{end}"))
            existing_cell = sheet._cells.get((new_min_row, min_col))
            if existing_cell is not None and type(existing_cell).__name__ == "MergedCell":
                del sheet._cells[(new_min_row, min_col)]
            target_cell = sheet.cell(row=new_min_row, column=min_col)
            target_cell.value = merge_label
            target_cell.border = copy(merge_border)
            target_cell.fill = copy(merge_fill)
            target_cell.font = copy(merge_font)
            target_cell.alignment = copy(merge_alignment)
            target_cell.number_format = merge_number_format
        new_merged_cells = MultiCellRange()
        new_merged_cells.ranges = set(shifted_merges)
        sheet.merged_cells = new_merged_cells

    def choose_value_and_format(upper_cell, lower_cell):
        upper_val = upper_cell.value
        lower_val = lower_cell.value

        if lower_val is None or str(lower_val).strip() == "":
            return upper_val, upper_cell.number_format

        if isinstance(upper_val, (int, float)) and isinstance(lower_val, (int, float)):
            upper_has_frac = abs(float(upper_val) - int(float(upper_val))) > 1e-12
            lower_has_frac = abs(float(lower_val) - int(float(lower_val))) > 1e-12

            if upper_has_frac and not lower_has_frac:
                return upper_val, upper_cell.number_format
            return lower_val, lower_cell.number_format

        return lower_val, lower_cell.number_format

    def append_sig_values_to_row(sheet, target_row: int, sig_row: int, max_col: int) -> None:
        appended = False
        label_cells = []
        for col in range(1, min(max_col, 2) + 1):
            target_cell = sheet.cell(row=target_row, column=col)
            sig_cell = sheet.cell(row=sig_row, column=col)
            label_cells.append(target_cell)
            target_cell.border = Border(
                left=target_cell.border.left,
                right=target_cell.border.right,
                top=target_cell.border.top,
                bottom=sig_cell.border.bottom,
            )
        for col in range(3, max_col + 1):
            target_cell = sheet.cell(row=target_row, column=col)
            sig_cell = sheet.cell(row=sig_row, column=col)
            sig_value = sig_cell.value
            if sig_value is not None and str(sig_value).strip() != "":
                appended = True
                base_value = target_cell.value
                if base_value is None or str(base_value).strip() == "":
                    target_cell.value = sig_value
                else:
                    target_cell.value = f"{base_value}\n{sig_value}"
                new_alignment = copy(target_cell.alignment)
                new_alignment.wrap_text = True
                target_cell.alignment = new_alignment
            target_cell.border = Border(
                left=target_cell.border.left,
                right=target_cell.border.right,
                top=target_cell.border.top,
                bottom=sig_cell.border.bottom,
            )
        if appended:
            target_height = sheet.row_dimensions[target_row].height or 15
            sig_height = sheet.row_dimensions[sig_row].height or 15
            sheet.row_dimensions[target_row].height = max(target_height, target_height + sig_height)
            thin_border = Side(border_style="thin", color="000000")
            for cell in label_cells:
                if cell.border.bottom.style is None:
                    cell.border = Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=thin_border,
                    )

    def normalize_multiline_rows(sheet, max_col: int) -> None:
        for row_idx in range(1, sheet.max_row + 1):
            has_multiline = False
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if "\n" not in str(cell.value or ""):
                    continue
                has_multiline = True
                new_alignment = copy(cell.alignment)
                new_alignment.wrap_text = True
                cell.alignment = new_alignment
            if has_multiline:
                sheet.row_dimensions[row_idx].height = max(sheet.row_dimensions[row_idx].height or 15, 30)

    def normalize_body_label_styles(sheet, max_col: int) -> None:
        thin_border = Side(border_style="thin", color="000000")
        label_fill = None
        index_fill = None
        for row_idx in range(1, sheet.max_row + 1):
            label_cell = sheet.cell(row=row_idx, column=2)
            if str(label_cell.value or "").strip() and label_cell.fill.fill_type:
                label_fill = copy(label_cell.fill)
                index_fill = copy(sheet.cell(row=row_idx, column=1).fill)
                break
        if label_fill is None:
            label_fill = PatternFill("solid", fgColor="FFF5E4")
        if index_fill is None:
            index_fill = PatternFill("solid", fgColor="FFFFFF")
        mean_index_fill = PatternFill("solid", fgColor="FFFF0000")

        def is_body_index(text: str) -> bool:
            return bool(re.fullmatch(r"\d+(?:\.0+)?", text))

        for row_idx in range(1, sheet.max_row + 1):
            code_text = str(sheet.cell(row=row_idx, column=1).value or "").strip()
            label_text = str(sheet.cell(row=row_idx, column=2).value or "").strip()
            if re.search(r"(?i)1st\s*row", code_text) or re.search(r"(?i)1st\s*row", label_text):
                continue
            if not is_body_index(code_text):
                continue
            if not label_text and not code_text:
                continue
            has_table_values = any(
                str(sheet.cell(row=row_idx, column=col_idx).value or "").strip()
                for col_idx in range(3, max_col + 1)
            )
            if not has_table_values:
                continue

            row_dimension = sheet.row_dimensions.get(row_idx)
            if row_dimension is not None and row_dimension.height is None:
                row_dimension.height = 13.5
            left_cell = sheet.cell(row=row_idx, column=1)
            label_cell = sheet.cell(row=row_idx, column=2)
            if code_text:
                left_cell.fill = copy(mean_index_fill if label_text.lower() == "mean" else index_fill)
            if label_text:
                label_cell.fill = copy(label_fill)
            for cell in (left_cell, label_cell):
                cell.border = Border(
                    left=cell.border.left if cell.border.left.style else thin_border,
                    right=cell.border.right if cell.border.right.style else thin_border,
                    top=cell.border.top if cell.border.top.style else thin_border,
                    bottom=cell.border.bottom if cell.border.bottom.style else thin_border,
                )

    def normalize_single_row_table_borders(sheet, max_col: int) -> None:
        thin_border = Side(border_style="thin", color="000000")
        merged_body_rows = set()
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, merge_max_col, max_row = merged_range.bounds
            if min_col <= 2 and merge_max_col <= 2 and max_row > min_row:
                for row_idx in range(min_row, max_row + 1):
                    merged_body_rows.add(row_idx)

        for row_idx in range(1, sheet.max_row + 1):
            if row_idx in merged_body_rows:
                continue
            label_text = str(sheet.cell(row=row_idx, column=2).value or "").strip()
            code_text = str(sheet.cell(row=row_idx, column=1).value or "").strip()
            if not re.fullmatch(r"\d+(?:\.0+)?", code_text):
                continue
            if not label_text and not code_text:
                continue
            has_table_values = any(
                str(sheet.cell(row=row_idx, column=col_idx).value or "").strip()
                for col_idx in range(3, max_col + 1)
            )
            if not has_table_values:
                continue
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.border = Border(
                    left=cell.border.left if cell.border.left.style else thin_border,
                    right=cell.border.right if cell.border.right.style else thin_border,
                    top=cell.border.top if cell.border.top.style else thin_border,
                    bottom=cell.border.bottom if cell.border.bottom.style else thin_border,
                )

    def remove_consecutive_duplicate_labels(sheet) -> None:
        rows_to_delete: List[int] = []
        row_idx = 7
        while row_idx < sheet.max_row:
            code_text = str(sheet.cell(row=row_idx, column=1).value or "").strip()
            label_text = str(sheet.cell(row=row_idx, column=2).value or "").strip()
            next_code_text = str(sheet.cell(row=row_idx + 1, column=1).value or "").strip()
            next_label_text = str(sheet.cell(row=row_idx + 1, column=2).value or "").strip()
            if (
                label_text
                and label_text == next_label_text
                and re.fullmatch(r"\d+(?:\.0+)?", code_text)
                and re.fullmatch(r"\d+(?:\.0+)?", next_code_text)
            ):
                for col_idx in range(3, sheet.max_column + 1):
                    target_cell = sheet.cell(row=row_idx, column=col_idx)
                    source_cell = sheet.cell(row=row_idx + 1, column=col_idx)
                    target_cell.value = source_cell.value
                    target_cell.number_format = source_cell.number_format
                rows_to_delete.append(row_idx + 1)
            row_idx += 1
        delete_rows_desc(sheet, rows_to_delete)

    def normalize_multi_sheet_sig_body_borders(sheet, max_col: int) -> None:
        thin_border = Side(border_style="thin", color="000000")
        merged_rows_by_start: set[int] = set()
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, merge_max_col, max_row = merged_range.bounds
            if min_col <= 2 and merge_max_col <= 2 and max_row == min_row + 1:
                merged_rows_by_start.add(min_row)

        for row_idx in sorted(merged_rows_by_start):
            if not re.fullmatch(r"\d+(?:\.0+)?", str(sheet.cell(row=row_idx, column=1).value or "").strip()):
                continue
            sig_row = row_idx + 1
            sheet.row_dimensions[row_idx].height = 13.5
            for col_idx in range(1, max_col + 1):
                value_cell = sheet.cell(row=row_idx, column=col_idx)
                sig_cell = sheet.cell(row=sig_row, column=col_idx)
                value_cell.border = Border(
                    left=value_cell.border.left if value_cell.border.left.style else thin_border,
                    right=value_cell.border.right if value_cell.border.right.style else thin_border,
                    top=value_cell.border.top if value_cell.border.top.style else thin_border,
                    bottom=value_cell.border.bottom,
                )
                sig_cell.border = Border(
                    left=sig_cell.border.left if sig_cell.border.left.style else thin_border,
                    right=sig_cell.border.right if sig_cell.border.right.style else thin_border,
                    top=sig_cell.border.top,
                    bottom=sig_cell.border.bottom if sig_cell.border.bottom.style else thin_border,
                )

    def normalize_one_sheet_header_fill(sheet, max_col: int) -> None:
        header_fill = None
        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(3, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.fill.fill_type and cell.fill.fgColor.rgb not in {None, "00000000", "FFFFFFFF"}:
                    header_fill = copy(cell.fill)
                    break
            if header_fill is not None:
                break
        if header_fill is None:
            header_fill = PatternFill("solid", fgColor="FFEBEBEB")

        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, 3):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if not re.search(r"(?i)1st\s*row", str(cell.value or "")):
                    continue

                applied = False
                for merged_range in sheet.merged_cells.ranges:
                    min_col, min_row, merge_max_col, max_row = merged_range.bounds
                    if min_col <= col_idx <= merge_max_col and min_row <= row_idx <= max_row:
                        if min_col <= 2 and merge_max_col <= 2:
                            top_left = sheet.cell(row=min_row, column=min_col)
                            top_left.fill = copy(header_fill)
                            applied = True
                        break
                if not applied:
                    cell.fill = copy(header_fill)

    def unmerge_body_label_ranges(sheet, max_col: int) -> None:
        for merged_range in list(sheet.merged_cells.ranges):
            min_col, min_row, merge_max_col, max_row = merged_range.bounds
            if min_col > 2 or merge_max_col > 2:
                continue
            top_left = sheet.cell(row=min_row, column=min_col)
            top_text = str(top_left.value or "")
            if re.search(r"(?i)1st\s*row", top_text):
                continue
            has_body_label = any(
                str(sheet.cell(row=row_idx, column=2).value or "").strip()
                for row_idx in range(min_row, max_row + 1)
            )
            if not has_body_label:
                continue
            value = top_left.value
            fill = copy(top_left.fill)
            font = copy(top_left.font)
            alignment = copy(top_left.alignment)
            number_format = top_left.number_format
            border = copy(top_left.border)
            sheet.unmerge_cells(str(merged_range))
            for row_idx in range(min_row, max_row + 1):
                cell = sheet.cell(row=row_idx, column=min_col)
                if row_idx == min_row and cell.value is None:
                    cell.value = value
                cell.fill = copy(fill)
                cell.font = copy(font)
                cell.alignment = copy(alignment)
                cell.number_format = number_format
                cell.border = copy(border)

    def is_sig_only_row(sheet, row_idx: int, max_col: int) -> bool:
        if str(sheet.cell(row=row_idx, column=1).value or "").strip():
            return False
        if str(sheet.cell(row=row_idx, column=2).value or "").strip():
            return False

        has_sig = False
        for col_idx in range(3, max_col + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            text = str(value or "").strip()
            if not text:
                continue
            if not re.fullmatch(r"[A-Z]+", text):
                return False
            has_sig = True
        return has_sig

    def merge_orphan_sig_rows(sheet, max_col: int) -> None:
        rows_to_delete: List[int] = []
        for row_idx in range(2, sheet.max_row + 1):
            if not is_sig_only_row(sheet, row_idx, max_col):
                continue
            previous_label = str(sheet.cell(row=row_idx - 1, column=2).value or "").strip()
            if not previous_label:
                continue
            append_sig_values_to_row(sheet, row_idx - 1, row_idx, max_col)
            rows_to_delete.append(row_idx)
        delete_rows_desc(sheet, rows_to_delete)
        normalize_multiline_rows(sheet, max_col)

    def normalize_row_legends(sheet) -> List[int]:
        legend_rows: List[int] = []
        for row in range(1, sheet.max_row + 1):
            for col in range(1, 3):
                cell = sheet.cell(row=row, column=col)
                if type(cell).__name__ == "MergedCell":
                    continue
                value = cell.value
                if value is None:
                    continue
                text = str(value)
                if re.search(r"(?i)count", text) and re.search(r"(?i)column\s*%", text):
                    if table_type != TABLE_MULTI_SHEET_NOT_SIG:
                        cell.value = "1st row:  Column %" if keep_mode == KEEP_PERCENT else "1st row:  Count"
                    legend_rows.append(row)
                    break
        return legend_rows

    def process_stacked_table_sheet(sheet, max_col: int) -> None:
        rows_to_delete: List[int] = []
        vertically_merged_following_rows: set[int] = set()
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, merge_max_col, max_row = merged_range.bounds
            if min_col <= 2 and merge_max_col <= 2 and max_row > min_row:
                vertically_merged_following_rows.update(range(min_row + 1, max_row + 1))

        def has_data_values(row_idx: int) -> bool:
            for col_idx in range(3, max_col + 1):
                value = sheet.cell(row=row_idx, column=col_idx).value
                if value is not None and str(value).strip() != "":
                    return True
            return False

        row = 1
        while row < sheet.max_row:
            label = sheet.cell(row=row, column=2).value
            label_text = str(label).strip() if label is not None else ""
            if label_text == "":
                row += 1
                continue

            next_label = sheet.cell(row=row + 1, column=2).value
            next_label_text = str(next_label).strip() if next_label is not None else ""
            possible_sig_row = row + 2
            has_following_sig_row = (
                possible_sig_row <= sheet.max_row
                and str(sheet.cell(row=possible_sig_row, column=1).value or "").strip() == ""
                and str(sheet.cell(row=possible_sig_row, column=2).value or "").strip() == ""
                and possible_sig_row in vertically_merged_following_rows
            )
            if (
                row + 1 in vertically_merged_following_rows
                and next_label_text == ""
                and has_data_values(row)
                and not has_data_values(row + 1)
                and not has_following_sig_row
            ):
                for col in range(1, max_col + 1):
                    target_cell = sheet.cell(row=row, column=col)
                    source_cell = sheet.cell(row=row + 1, column=col)
                    target_cell.border = Border(
                        left=target_cell.border.left,
                        right=target_cell.border.right,
                        top=target_cell.border.top,
                        bottom=source_cell.border.bottom,
                    )
                rows_to_delete.append(row + 1)
                row += 2
                continue

            if next_label_text not in {"", label_text} or (
                not has_data_values(row + 1) and not has_following_sig_row
            ):
                row += 1
                continue

            pct_row = row + 1
            is_total_row = label_text.upper() == "TOTAL"
            sig_row = pct_row + 1
            has_sig_row = (
                sig_row <= sheet.max_row
                and str(sheet.cell(row=sig_row, column=1).value or "").strip() == ""
                and str(sheet.cell(row=sig_row, column=2).value or "").strip() == ""
            )
            last_group_row = sig_row if has_sig_row else pct_row

            keep_row = row
            pct_has_values = has_data_values(pct_row)
            if keep_mode == KEEP_PERCENT and not is_total_row and pct_has_values:
                keep_row = pct_row

            for col in range(1, max_col + 1):
                target_cell = sheet.cell(row=keep_row, column=col)
                bottom_source = sheet.cell(row=last_group_row, column=col)
                bottom_border = bottom_source.border.bottom if col <= 2 else target_cell.border.bottom
                target_cell.border = Border(
                    left=target_cell.border.left,
                    right=target_cell.border.right,
                    top=target_cell.border.top,
                    bottom=bottom_border,
                )

            if keep_mode == KEEP_PERCENT and not is_total_row and pct_has_values:
                rows_to_delete.append(row)
            else:
                rows_to_delete.append(pct_row)

            row = last_group_row + 1

        compact_rows_once(sheet, rows_to_delete)
        normalize_body_label_styles(sheet, max_col)
        if table_type in {
            TABLE_ONE_SHEET_SIG,
            TABLE_ONE_SHEET_NOT_SIG,
            TABLE_MULTI_SHEET_SIG,
            TABLE_MULTI_SHEET_NOT_SIG,
        }:
            normalize_one_sheet_header_fill(sheet, max_col)
        if table_type == TABLE_MULTI_SHEET_SIG:
            normalize_multi_sheet_sig_body_borders(sheet, max_col)
        if table_type in {TABLE_ONE_SHEET_NOT_SIG, TABLE_MULTI_SHEET_NOT_SIG}:
            normalize_single_row_table_borders(sheet, max_col)
        normalize_multiline_rows(sheet, max_col)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet_name.strip().lower() == "contents":
            continue
        if sheet.max_row < 7:
            continue
        legend_rows = normalize_row_legends(sheet)
        if not legend_rows:
            continue

        max_col = detect_effective_max_col(sheet)

        if len(legend_rows) > 1:
            process_stacked_table_sheet(sheet, max_col)
            continue

        merged_ranges = list(sheet.merged_cells.ranges)
        for m_range in merged_ranges:
            min_col, min_row, m_max_col, m_max_row = m_range.bounds
            if min_row < 7:
                continue

            top_left_cell = sheet.cell(row=min_row, column=min_col)
            tl_val = top_left_cell.value
            tl_border = copy(top_left_cell.border)
            tl_fill = copy(top_left_cell.fill)
            tl_font = copy(top_left_cell.font)
            tl_alignment = copy(top_left_cell.alignment)
            tl_number_format = top_left_cell.number_format

            sheet.unmerge_cells(str(m_range))

            for r in range(min_row, m_max_row + 1):
                for c in range(min_col, m_max_col + 1):
                    cell = sheet.cell(row=r, column=c)
                    cell.value = tl_val
                    if tl_border:
                        cell.border = copy(tl_border)
                    if tl_fill:
                        cell.fill = copy(tl_fill)
                    if tl_font:
                        cell.font = copy(tl_font)
                    if tl_alignment:
                        cell.alignment = copy(tl_alignment)
                    cell.number_format = tl_number_format

        rows_to_delete_regular: List[int] = []
        pending_label_merges: List[Tuple[int, int]] = []

        def row_has_values(row_idx: int, start_col: int = 3) -> bool:
            for col_idx in range(start_col, max_col + 1):
                value = sheet.cell(row=row_idx, column=col_idx).value
                if value is not None and str(value).strip() != "":
                    return True
            return False

        row = 7
        while row <= sheet.max_row:
            label = sheet.cell(row=row, column=2).value
            label_text = str(label).strip() if label is not None else ""
            if label_text == "":
                row += 1
                continue

            group_start = row
            group_end = row
            code_value = sheet.cell(row=row, column=1).value
            while group_end + 1 <= sheet.max_row:
                next_label = sheet.cell(row=group_end + 1, column=2).value
                next_code = sheet.cell(row=group_end + 1, column=1).value
                if next_label != label or next_code != code_value:
                    break
                group_end += 1

            group_len = group_end - group_start + 1
            is_total_group = label_text.upper() == "TOTAL"

            if is_total_group:
                if table_type == TABLE_MULTI_SHEET_SIG and group_len >= 3:
                    for remove_row in range(group_start + 1, group_end):
                        rows_to_delete_regular.append(remove_row)
                    pending_label_merges.append((group_start, group_end))
                    row = group_end + 1
                    continue

                for remove_row in range(group_start + 1, group_end + 1):
                    for col in range(1, max_col + 1):
                        target_cell = sheet.cell(row=group_start, column=col)
                        source_cell = sheet.cell(row=remove_row, column=col)
                        target_cell.border = Border(
                            left=target_cell.border.left,
                            right=target_cell.border.right,
                            top=target_cell.border.top,
                            bottom=source_cell.border.bottom,
                        )
                    rows_to_delete_regular.append(remove_row)
            elif group_len >= 3:
                # N/%/Sig tables: remove the row that does not match the selected output type.
                remove_row = group_start if keep_mode == KEEP_PERCENT else group_start + 1
                rows_to_delete_regular.append(remove_row)
                keep_row = group_start + 1 if keep_mode == KEEP_PERCENT else group_start
                if table_type == TABLE_MULTI_SHEET_SIG:
                    pending_label_merges.append((keep_row, group_end))
                    row = group_end + 1
                    continue

                for sig_row in range(group_start + 2, group_end + 1):
                    if row_has_values(sig_row):
                        append_sig_values_to_row(sheet, keep_row, sig_row, max_col)
                        rows_to_delete_regular.append(sig_row)
                    else:
                        for col in range(1, max_col + 1):
                            target_cell = sheet.cell(row=keep_row, column=col)
                            source_cell = sheet.cell(row=sig_row, column=col)
                            target_cell.border = Border(
                                left=target_cell.border.left,
                                right=target_cell.border.right,
                                top=target_cell.border.top,
                                bottom=source_cell.border.bottom,
                            )
                        rows_to_delete_regular.append(sig_row)
            elif group_len == 2:
                pct_row = group_start + 1
                for col in range(1, max_col + 1):
                    target_cell = sheet.cell(row=group_start, column=col)
                    source_cell = sheet.cell(row=pct_row, column=col)
                    target_cell.border = Border(
                        left=target_cell.border.left,
                        right=target_cell.border.right,
                        top=target_cell.border.top,
                        bottom=source_cell.border.bottom,
                    )

                if keep_mode == KEEP_PERCENT:
                    for col in range(3, max_col + 1):
                        target_cell = sheet.cell(row=group_start, column=col)
                        source_cell = sheet.cell(row=pct_row, column=col)
                        chosen_val, chosen_fmt = choose_value_and_format(target_cell, source_cell)
                        target_cell.value = chosen_val
                        target_cell.number_format = chosen_fmt

                rows_to_delete_regular.append(group_start if keep_mode == KEEP_PERCENT else pct_row)

            row = group_end + 1

        delete_rows_desc(sheet, rows_to_delete_regular)
        if table_type == TABLE_MULTI_SHEET_NOT_SIG:
            remove_consecutive_duplicate_labels(sheet)
        merge_orphan_sig_rows(sheet, max_col)
        unmerge_body_label_ranges(sheet, max_col)
        normalize_body_label_styles(sheet, max_col)
        if table_type in {
            TABLE_ONE_SHEET_SIG,
            TABLE_ONE_SHEET_NOT_SIG,
            TABLE_MULTI_SHEET_SIG,
            TABLE_MULTI_SHEET_NOT_SIG,
        }:
            normalize_one_sheet_header_fill(sheet, max_col)
        if table_type == TABLE_MULTI_SHEET_SIG:
            normalize_multi_sheet_sig_body_borders(sheet, max_col)
        if table_type in {TABLE_ONE_SHEET_NOT_SIG, TABLE_MULTI_SHEET_NOT_SIG}:
            normalize_single_row_table_borders(sheet, max_col)
        normalize_multiline_rows(sheet, max_col)

        deleted_rows = sorted(set(rows_to_delete_regular))

        def shifted_row(original_row: int) -> int:
            deleted_before = sum(1 for deleted_row in deleted_rows if deleted_row < original_row)
            return original_row - deleted_before

        for start_row, end_row in pending_label_merges:
            shifted_start = shifted_row(start_row)
            shifted_end = shifted_row(end_row)
            if shifted_end <= shifted_start:
                continue
            for col in range(1, 3):
                top_cell = sheet.cell(row=shifted_start, column=col)
                bottom_cell = sheet.cell(row=shifted_end, column=col)
                top_cell.border = Border(
                    left=top_cell.border.left,
                    right=top_cell.border.right,
                    top=top_cell.border.top,
                    bottom=bottom_cell.border.bottom,
                )
                sheet.merge_cells(
                    start_row=shifted_start,
                    start_column=col,
                    end_row=shifted_end,
                    end_column=col,
                )

        if table_type == TABLE_MULTI_SHEET_SIG:
            normalize_multi_sheet_sig_body_borders(sheet, max_col)

        last_row = 7
        for r in range(sheet.max_row, 6, -1):
            row_has_data = False
            for c in range(1, max_col + 1):
                v = sheet.cell(row=r, column=c).value
                if v is not None and str(v).strip() != "":
                    row_has_data = True
                    break
            if row_has_data:
                last_row = r
                break
        if table_type == TABLE_MULTI_SHEET_SIG:
            last_row = sheet.max_row

        thin_border = Side(border_style="thin", color="000000")
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=last_row, column=col)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=thin_border,
            )

    repair_internal_links(wb)
    wb.save(save_path)


def resolve_input_items(items: List[str]) -> Tuple[List[Path], List[str]]:
    valid_paths: List[Path] = []
    invalid_items: List[str] = []

    for raw in items:
        p = raw.strip().strip('"').strip("'")
        if not p:
            continue
        path_obj = Path(p)

        if path_obj.is_file() and path_obj.suffix.lower() == ".xlsx":
            valid_paths.append(path_obj)
        elif path_obj.is_dir():
            files = sorted(path_obj.glob("*.xlsx"))
            if files:
                valid_paths.extend(files)
            else:
                invalid_items.append(f"{p} (no .xlsx inside folder)")
        else:
            invalid_items.append(p)

    seen = set()
    deduped: List[Path] = []
    for p in valid_paths:
        key = str(p.resolve()).lower()
        if key not in seen:
            seen.add(key)
            deduped.append(p)

    return deduped, invalid_items


# ----------------------------
# UI components
# ----------------------------
class AnimatedBackground(QWidget):
    pass


class FancyButton(QPushButton):
    def __init__(self, text: str) -> None:
        super().__init__(text)
        self.setCursor(Qt.CursorShape.PointingHandCursor)


class ProcessingWorker(QObject):
    progress = pyqtSignal(int, str)
    done = pyqtSignal(list)
    failed = pyqtSignal(str)

    def __init__(self, files: List[Path], output_dir: Path, keep_mode: str) -> None:
        super().__init__()
        self.files = files
        self.output_dir = output_dir
        self.keep_mode = keep_mode

    def run(self) -> None:
        try:
            saved_files: List[str] = []
            output_modes = [KEEP_COUNT, KEEP_PERCENT] if self.keep_mode == KEEP_BOTH else [self.keep_mode]
            total = len(self.files) * len(output_modes)
            reserved_names: set[str] = set()
            plans: List[Tuple[Path, Path, str]] = []
            for src in self.files:
                for mode in output_modes:
                    dst = unique_output_path_reserved(self.output_dir, src, reserved_names, mode)
                    plans.append((src, dst, mode))

            if total <= 1:
                src, dst, mode = plans[0]
                process_workbook(src, dst, mode)
                saved_files.append(str(dst))
                self.progress.emit(100, f"1/1 processed: {src.name}")
            else:
                max_workers = max(2, min(total, os.cpu_count() or 2))
                with ProcessPoolExecutor(max_workers=max_workers) as executor:
                    future_map = {
                        executor.submit(process_one_file_task, str(src), str(dst), mode): (src, dst, mode)
                        for src, dst, mode in plans
                    }
                    completed = 0
                    for fut in as_completed(future_map):
                        src, dst, mode = future_map[fut]
                        fut.result()
                        saved_files.append(str(dst))
                        completed += 1
                        pct = int((completed / total) * 100)
                        mode_label = "%" if mode == KEEP_PERCENT else "N"
                        self.progress.emit(pct, f"{completed}/{total} processed ({mode_label}): {src.name}")

            # Keep output list in input order for easier review.
            ordered = [str(dst) for _, dst, _ in plans]
            saved_files = ordered
            self.done.emit(saved_files)
        except Exception as exc:
            self.failed.emit(str(exc))


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.selected_files: List[Path] = []
        self.output_dir: Path | None = None
        self.worker_thread: QThread | None = None
        self.started_at: float | None = None
        self.keep_mode = KEEP_COUNT

        self.setWindowTitle("Excel Smart Formatter Pro")
        self.resize(980, 700)

        self.bg = AnimatedBackground()
        self.bg.setObjectName("bg")
        self.setCentralWidget(self.bg)

        self._build_ui()
        self._animate_intro()

    def _build_ui(self) -> None:
        root_layout = QVBoxLayout(self.bg)
        root_layout.setContentsMargins(40, 30, 40, 30)

        self.card = QFrame()
        self.card.setObjectName("card")
        card_layout = QVBoxLayout(self.card)
        card_layout.setSpacing(14)
        card_layout.setContentsMargins(24, 24, 24, 24)

        title = QLabel("Excel %/N Cleaner")
        title.setObjectName("title")
        subtitle = QLabel("Professional PyQt6 UI with multi-file workflow, folder/path paste, and smooth motion")
        subtitle.setObjectName("subtitle")
        subtitle.setWordWrap(True)

        self.btn_pick_files = FancyButton("Select .xlsx Files")
        self.btn_pick_files.clicked.connect(self.pick_files)

        self.btn_pick_output = FancyButton("Select Output Folder")
        self.btn_pick_output.clicked.connect(self.pick_output_folder)

        self.paths_edit = QTextEdit()
        self.paths_edit.setPlaceholderText("Paste file/folder paths here. One per line (or separated by semicolon).")
        self.paths_edit.setFixedHeight(120)

        self.btn_apply_paths = FancyButton("Apply Pasted Paths")
        self.btn_apply_paths.clicked.connect(self.apply_pasted_paths)

        mode_label = QLabel("Output rows")
        mode_label.setObjectName("meta")
        self.radio_keep_both = QRadioButton("N + %")
        self.radio_keep_count = QRadioButton("N only")
        self.radio_keep_percent = QRadioButton("% only")
        self.radio_keep_both.setChecked(True)
        self.radio_keep_both.toggled.connect(self.update_keep_mode)
        self.radio_keep_count.toggled.connect(self.update_keep_mode)
        self.radio_keep_percent.toggled.connect(self.update_keep_mode)

        mode_row = QHBoxLayout()
        mode_row.addWidget(mode_label)
        mode_row.addWidget(self.radio_keep_both)
        mode_row.addWidget(self.radio_keep_count)
        mode_row.addWidget(self.radio_keep_percent)
        mode_row.addStretch(1)

        self.list_files = QListWidget()
        self.list_files.setMinimumHeight(160)

        self.lbl_output = QLabel("Output folder: not selected")
        self.lbl_output.setObjectName("meta")
        self.lbl_status = QLabel("Ready")
        self.lbl_status.setObjectName("status")

        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setValue(0)

        self.btn_start = FancyButton("Process All Files")
        self.btn_start.clicked.connect(self.start_processing)
        self.btn_start.setEnabled(False)

        row = QHBoxLayout()
        row.addWidget(self.btn_pick_files)
        row.addWidget(self.btn_pick_output)

        card_layout.addWidget(title)
        card_layout.addWidget(subtitle)
        card_layout.addLayout(row)
        card_layout.addWidget(self.paths_edit)
        card_layout.addWidget(self.btn_apply_paths)
        card_layout.addLayout(mode_row)
        card_layout.addWidget(self.list_files)
        card_layout.addWidget(self.lbl_output)
        card_layout.addWidget(self.progress)
        card_layout.addWidget(self.lbl_status)
        card_layout.addWidget(self.btn_start)

        root_layout.addWidget(self.card)

        self.setStyleSheet(
            """
            QMainWindow, QWidget { color: #eaf0ff; font-family: 'Segoe UI', 'Tahoma'; font-size: 14px; }
            #bg {
                background: qradialgradient(cx:0.2, cy:0.1, radius:1.25,
                    fx:0.15, fy:0.05,
                    stop:0 #1f3866, stop:0.45 #142746, stop:1 #0b1321);
            }
            #card {
                background: rgba(11, 19, 31, 0.80);
                border: 1px solid rgba(255,255,255,0.16);
                border-radius: 20px;
            }
            #title { font-size: 34px; font-weight: 700; color: #ffffff; }
            #subtitle { font-size: 13px; color: #bfd3ff; }
            #meta { color: #98b7ff; font-size: 12px; }
            #status { color: #ffffff; font-size: 12px; }
            QTextEdit, QListWidget {
                background: rgba(255, 255, 255, 0.08);
                border: 1px solid rgba(255, 255, 255, 0.2);
                border-radius: 12px;
                padding: 8px;
                color: #f2f6ff;
            }
            QProgressBar {
                background: rgba(255, 255, 255, 0.10);
                border: 1px solid rgba(255,255,255,0.2);
                border-radius: 9px;
                text-align: center;
                height: 18px;
            }
            QProgressBar::chunk {
                border-radius: 8px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2dc3ff, stop:1 #4df0b7);
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #0ea5e9, stop:1 #10b981);
                border: none;
                border-radius: 12px;
                color: #062033;
                font-weight: 700;
                padding: 12px 16px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #38bdf8, stop:1 #34d399);
            }
            QPushButton:pressed {
                padding-top: 13px;
                padding-bottom: 11px;
            }
            QPushButton:disabled {
                background: rgba(255,255,255,0.25);
                color: rgba(255,255,255,0.6);
            }
            QRadioButton {
                color: #f2f6ff;
                spacing: 8px;
                font-weight: 600;
            }
            QRadioButton::indicator {
                width: 16px;
                height: 16px;
            }
            """
        )

    def _animate_intro(self) -> None:
        effect = QGraphicsOpacityEffect(self.card)
        self.card.setGraphicsEffect(effect)
        anim = QPropertyAnimation(effect, b"opacity", self)
        anim.setDuration(650)
        anim.setStartValue(0.0)
        anim.setEndValue(1.0)
        anim.setEasingCurve(QEasingCurve.Type.OutCubic)
        anim.start()
        self._intro_anim = anim

    def refresh_file_list(self) -> None:
        self.list_files.clear()
        for p in self.selected_files:
            self.list_files.addItem(str(p))
        self.btn_start.setEnabled(bool(self.selected_files and self.output_dir))

    def update_keep_mode(self) -> None:
        if self.radio_keep_both.isChecked():
            self.keep_mode = KEEP_BOTH
        elif self.radio_keep_percent.isChecked():
            self.keep_mode = KEEP_PERCENT
        else:
            self.keep_mode = KEEP_COUNT

    def pick_files(self) -> None:
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Excel Files",
            "",
            "Excel Files (*.xlsx);;All Files (*.*)",
        )
        if not files:
            return

        parsed, invalid = resolve_input_items(files)
        self.selected_files = parsed
        self.refresh_file_list()

        if invalid:
            QMessageBox.warning(self, "Invalid paths", "\n".join(invalid[:5]))
        self.lbl_status.setText(f"Loaded {len(self.selected_files)} files")

    def pick_output_folder(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if not folder:
            return

        self.output_dir = Path(folder)
        self.lbl_output.setText(f"Output folder: {self.output_dir}")
        self.refresh_file_list()

    def apply_pasted_paths(self) -> None:
        raw = self.paths_edit.toPlainText().replace(";", "\n")
        items = [line for line in raw.splitlines() if line.strip()]
        files, invalid = resolve_input_items(items)

        self.selected_files = files
        self.refresh_file_list()

        if invalid:
            show = "\n".join(invalid[:7])
            more = "" if len(invalid) <= 7 else f"\n... and {len(invalid) - 7} more"
            QMessageBox.warning(self, "Some inputs were not valid", f"{show}{more}")

        self.lbl_status.setText(f"Loaded {len(self.selected_files)} files from pasted paths")

    def start_processing(self) -> None:
        if not self.selected_files:
            QMessageBox.warning(self, "No files", "Please select files first")
            return
        if not self.output_dir:
            QMessageBox.warning(self, "No output folder", "Please choose an output folder")
            return

        self.btn_start.setEnabled(False)
        self.progress.setValue(0)
        self.update_keep_mode()
        if self.keep_mode == KEEP_BOTH:
            mode_label = "N + %"
        else:
            mode_label = "% only" if self.keep_mode == KEEP_PERCENT else "N only"
        self.lbl_status.setText(f"Starting processing ({mode_label})...")
        self.started_at = time.perf_counter()

        self.worker_thread = QThread()
        self.worker = ProcessingWorker(self.selected_files, self.output_dir, self.keep_mode)
        self.worker.moveToThread(self.worker_thread)

        self.worker_thread.started.connect(self.worker.run)
        self.worker.progress.connect(self.on_progress)
        self.worker.done.connect(self.on_done)
        self.worker.failed.connect(self.on_failed)
        self.worker.done.connect(self.worker_thread.quit)
        self.worker.failed.connect(self.worker_thread.quit)
        self.worker_thread.finished.connect(lambda: self.btn_start.setEnabled(True))

        self.worker_thread.start()

    def on_progress(self, pct: int, text: str) -> None:
        self.progress.setValue(pct)
        self.lbl_status.setText(text)

    def on_done(self, saved_files: List[str]) -> None:
        self.progress.setValue(100)
        self.lbl_status.setText("Completed successfully")
        elapsed = 0.0
        if self.started_at is not None:
            elapsed = time.perf_counter() - self.started_at
        avg = (elapsed / len(saved_files)) if saved_files else 0.0

        preview = "\n".join(Path(p).name for p in saved_files[:8])
        more = "" if len(saved_files) <= 8 else f"\n...และอีก {len(saved_files) - 8} ไฟล์"
        summary = (
            f"ประมวลผลเสร็จ {len(saved_files)} ไฟล์\n"
            f"ใช้เวลา {elapsed:.1f} วินาที\n"
            f"เฉลี่ย {avg:.1f} วินาที/ไฟล์"
        )
        detail = f"ไฟล์ที่สร้าง:\n{preview}{more}"

        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setWindowTitle("Done")
        msg.setText(summary)
        msg.setInformativeText(detail)
        msg.setStyleSheet(
            "QMessageBox { background: #f7fafc; }"
            "QLabel { color: #0f172a; font-size: 12px; }"
            "QPushButton { min-width: 84px; color: #0f172a; background: #bae6fd; border: 1px solid #7dd3fc; border-radius: 8px; padding: 6px 12px; }"
            "QPushButton:hover { background: #7dd3fc; }"
        )
        msg.exec()

    def on_failed(self, message: str) -> None:
        self.lbl_status.setText("Error occurred")
        QMessageBox.critical(self, "Processing failed", message)


def main() -> None:
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    pal = QPalette()
    pal.setColor(QPalette.ColorRole.WindowText, QColor("#eaf0ff"))
    app.setPalette(pal)

    def handle_uncaught_exception(exc_type, exc_value, exc_tb):
        err_text = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
        log_path = Path(__file__).with_name("6_TEST_error.log")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with log_path.open("a", encoding="utf-8") as f:
            f.write(f"\n[{timestamp}]\n{err_text}\n")
        QMessageBox.critical(
            None,
            "Application Error",
            f"โปรแกรมเกิดข้อผิดพลาดและถูกหยุดการทำงาน\nบันทึก log ที่:\n{log_path}\n\n{exc_value}",
        )

    sys.excepthook = handle_uncaught_exception

    global APP_WINDOW
    APP_WINDOW = MainWindow()
    APP_WINDOW.show()
    sys.exit(app.exec())


#if __name__ == "__main__":
    #APP_WINDOW = None
    #main()


# <<< START OF CHANGES >>>
# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
def run_this_app(working_dir=None): # ชื่อฟังก์ชันนี้จะถูกใช้ใน Launcher
    """
    ฟังก์ชันหลักสำหรับสร้างและรัน QuotaSamplerApp.
    """
    print(f"--- QUOTA_SAMPLER_INFO: Starting 'QuotaSamplerApp' via run_this_app() ---")
    try:
    # --- ส่วนที่ใช้รันโปรแกรม ---

    #if __name__ == "__main__":
        main()


        print(f"--- QUOTA_SAMPLER_INFO: QuotaSamplerApp mainloop finished. ---")

    except Exception as e:
        # ดักจับ Error ที่อาจเกิดขึ้นระหว่างการสร้างหรือรัน App
        print(f"QUOTA_SAMPLER_ERROR: An error occurred during QuotaSamplerApp execution: {e}")
        # แสดง Popup ถ้ามีปัญหา
        if 'root' not in locals() or not root.winfo_exists(): # สร้าง root ชั่วคราวถ้ายังไม่มี
            root_temp = tk.Tk()
            root_temp.withdraw()
            messagebox.showerror("Application Error (Quota Sampler)",
                                f"An unexpected error occurred:\n{e}", parent=root_temp)
            root_temp.destroy()
        else:
            messagebox.showerror("Application Error (Quota Sampler)",
                                f"An unexpected error occurred:\n{e}", parent=root) # ใช้ root ที่มีอยู่ถ้าเป็นไปได้
        sys.exit(f"Error running QuotaSamplerApp: {e}") # อาจจะ exit หรือไม่ก็ได้ ขึ้นกับการออกแบบ


# --- ส่วน Run Application เมื่อรันไฟล์นี้โดยตรง (สำหรับ Test) ---
if __name__ == "__main__":
    print("--- Running QuotaSamplerApp.py directly for testing ---")
    # (ถ้ามีการตั้งค่า DPI ด้านบน มันจะทำงานอัตโนมัติ)

    # เรียกฟังก์ชัน Entry Point ที่เราสร้างขึ้น
    run_this_app()

    print("--- Finished direct execution of QuotaSamplerApp.py ---")
# <<< END OF CHANGES >>>