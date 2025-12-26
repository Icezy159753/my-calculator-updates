import tkinter
from tkinter import filedialog, messagebox, Listbox, simpledialog, scrolledtext
import ttkbootstrap as bstrap
from ttkbootstrap.constants import *
from ttkbootstrap.tableview import Tableview
import pandas as pd
import re
import pyreadstat
import numpy as np
import os
import openpyxl

# --- (คงเดิม) Imports for Factor/Regression Analysis ---
import statsmodels.api as sm
from factor_analyzer import FactorAnalyzer
from collections import OrderedDict
import io
import sys
from scipy.linalg import inv, eigh
from sklearn.preprocessing import StandardScaler


class SpssProcessorApp(bstrap.Window):
    def __init__(self):
        # คุณสามารถเปลี่ยน Theme ได้ที่นี่ เช่น "litera", "cosmo", "flatly", "superhero"
        super().__init__(themename="superhero")
        self.title("BrandSence Model Processor")
        self.geometry("1200x850")

        # --- State Variables ---
        self.df = None
        self.spss_original_order = []
        self.computed_c_cols = []
        self.c_vars_to_compute = []
        self.vars_to_transform = {}
        self.transformed_df = None
        self.za_cols = []
        self.id_vars = []
        self.last_excel_filepath = None
        self.original_filepath = None
        self.save_all_sheets_var = bstrap.BooleanVar(value=True)
        self.t2b_choice_var = bstrap.StringVar(value="5+4")
        self.index1_labels = {}
        self.filter_labels = {}

        # --- GUI Setup ---
        self.setup_gui()
        self.center_window()

    def center_window(self):
        """จัดหน้าต่างหลักให้อยู่กึ่งกลางจอ"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')

    def _center_toplevel(self, toplevel_window):
        """Helper function สำหรับจัดหน้าต่างย่อยให้อยู่กึ่งกลาง"""
        toplevel_window.update_idletasks()
        main_x = self.winfo_x()
        main_y = self.winfo_y()
        main_width = self.winfo_width()
        main_height = self.winfo_height()
        pop_width = toplevel_window.winfo_width()
        pop_height = toplevel_window.winfo_height()
        x = main_x + (main_width - pop_width) // 2
        y = main_y + (main_height - pop_height) // 2
        toplevel_window.geometry(f'+{x}+{y}')

    def setup_gui(self):
        """ตั้งค่าหน้าตาโปรแกรมหลัก"""
        main_pane = bstrap.Panedwindow(self, orient=HORIZONTAL)
        main_pane.pack(fill=BOTH, expand=True, padx=10, pady=10)

        # --- Frame ซ้าย (Control Panel) ---
        left_frame = bstrap.Frame(main_pane, padding=20)
        main_pane.add(left_frame, weight=1)

        # --- Frame ขวา (Display Area) ---
        self.right_frame = bstrap.Frame(main_pane, padding=10)
        main_pane.add(self.right_frame, weight=4)

        # --- Widgets ใน Control Panel ---
        header = bstrap.Label(left_frame, text="BrandSence Model", font=("Helvetica", 20, "bold"), bootstyle="primary")
        header.pack(pady=(0, 20), anchor="center")

        # --- Step 1 Frame ---
        step1_frame = bstrap.Labelframe(left_frame, text=" ขั้นตอนที่ 1: เริ่มประมวลผลข้อมูล ", bootstyle="info", padding=10)
        step1_frame.pack(fill="x", pady=10)
        self.btn_start_process = bstrap.Button(step1_frame, text="เริ่ม (เลือกตัวแปรเอง)", command=self.start_full_process, bootstyle="success", width=25)
        self.btn_start_process.pack(pady=5, fill="x", ipady=5)
        self.btn_load_settings_process = bstrap.Button(step1_frame, text="เริ่ม (โหลดการตั้งค่าจากไฟล์)", command=self.start_process_with_settings, bootstyle="secondary", width=25)
        self.btn_load_settings_process.pack(pady=5, fill="x", ipady=5)
        
        self.btn_reanalyze = bstrap.Button(step1_frame, text="วิเคราะห์ซ้ำ (จากไฟล์ Compute C)", command=self.start_reanalyze_process, bootstyle="warning", width=25)
        self.btn_reanalyze.pack(pady=5, fill="x", ipady=5)

        # --- Step 2 Frame ---
        step2_frame = bstrap.Labelframe(left_frame, text=" ขั้นตอนที่ 2: วิเคราะห์และส่งออก ", bootstyle="info", padding=10)
        step2_frame.pack(fill="x", pady=10)
        bstrap.Label(step2_frame, text="Filter สำหรับสรุป Excel:").pack(side="top", anchor="w")
        self.filter_entry = bstrap.Entry(step2_frame, state="disabled")
        self.filter_entry.pack(side="top", fill="x", pady=(5, 10))

        # --- Label Definition Button ---
        self.btn_define_labels = bstrap.Button(step2_frame, text="กำหนด Label สำหรับ Index และ Filter", command=self.open_label_editor, bootstyle="primary", width=25, state="disabled")
        self.btn_define_labels.pack(pady=5, fill="x", ipady=5)

        self.cb_save_all_sheets = bstrap.Checkbutton(
            step2_frame,
            text="บันทึกเฉพาะชีท Summary เท่านั้น",
            variable=self.save_all_sheets_var,
            bootstyle="primary"
        )
        self.cb_save_all_sheets.pack(pady=5, anchor="w")
        self.btn_analyze_export = bstrap.Button(step2_frame, text="วิเคราะห์และส่งออก Excel", command=self.run_analysis_and_export, bootstyle="info", width=25, state="disabled")
        self.btn_analyze_export.pack(pady=10, fill="x", ipady=5)

        # --- Settings Frame ---
        settings_frame = bstrap.Labelframe(left_frame, text=" จัดการการตั้งค่า & เครื่องมือ ", bootstyle="info", padding=10)
        settings_frame.pack(fill="x", pady=10)
        self.btn_save_settings = bstrap.Button(settings_frame, text="บันทึกการตั้งค่าปัจจุบัน", command=self.save_settings, bootstyle="primary-outline", width=25, state="disabled")
        self.btn_save_settings.pack(pady=5, fill="x", ipady=5)
        
        # --- Progress Bar and Status Label ---
        self.status_label = bstrap.Label(left_frame, text="พร้อมทำงาน", bootstyle="secondary")
        self.status_label.pack(side="bottom", fill="x", pady=(10,0))
        self.progress = bstrap.Progressbar(left_frame, mode='indeterminate', bootstyle="success-striped")
        self.progress.pack(side="bottom", pady=5, fill="x")

        # --- Display Area initial message ---
        self.initial_message = bstrap.Label(self.right_frame, text="กรุณากด 'เริ่มกระบวนการ' เพื่อโหลดไฟล์ SPSS", font=("Helvetica", 16), bootstyle="secondary")
        self.initial_message.pack(expand=True)

    def update_status(self, text, bootstyle="info"):
        """อัปเดตข้อความสถานะ"""
        self.status_label.config(text=text, bootstyle=bootstyle)
        self.update_idletasks()

    def start_progress(self):
        """เริ่มการทำงาน Progress Bar"""
        self.progress.start()

    def stop_progress(self):
        """หยุดการทำงาน Progress Bar"""
        self.progress.stop()

    def reset_state(self):
        """รีเซ็ตสถานะของโปรแกรมทั้งหมดเพื่อเริ่มใหม่"""
        self.df = None
        self.spss_original_order = []
        self.computed_c_cols = []
        self.c_vars_to_compute = []
        self.vars_to_transform = {}
        self.transformed_df = None
        self.za_cols = []
        self.id_vars = []
        self.last_excel_filepath = None
        self.original_filepath = None
        self.t2b_choice_var.set("5+4")
        self.index1_labels = {}
        self.filter_labels = {}

        self.btn_analyze_export.config(state="disabled")
        self.btn_define_labels.config(state="disabled")
        self.btn_save_settings.config(state="disabled")
        self.filter_entry.config(state="disabled")
        self.filter_entry.delete(0, 'end')
        self.update_status("พร้อมทำงาน", "secondary")

        for widget in self.right_frame.winfo_children():
            widget.destroy()
        self.initial_message = bstrap.Label(self.right_frame, text="กรุณากด 'เริ่มกระบวนการ' เพื่อโหลดไฟล์ SPSS", font=("Helvetica", 16), bootstyle="secondary")
        self.initial_message.pack(expand=True)

    # ===================================================================
    # WORKFLOWS
    # ===================================================================
    def start_full_process(self):
        """Workflow 1: เริ่มต้นกระบวนการแบบเลือกตัวแปรเองทั้งหมด"""
        self.reset_state()
        if not self.load_spss_file():
            return
        self.open_c_variable_selector()

    def start_process_with_settings(self):
        """เริ่มต้นกระบวนการโดยโหลดการตั้งค่าและไฟล์ SPSS อัตโนมัติ"""
        self.reset_state()
        self.update_status("กำลังรอเลือกไฟล์การตั้งค่า...")
        settings_filepath = filedialog.askopenfilename(
            filetypes=[("Excel Settings File", "*.xlsx")],
            title="เลือกไฟล์การตั้งค่า"
        )
        if not settings_filepath:
            self.update_status("ยกเลิกการเลือกไฟล์ตั้งค่า", "warning")
            return

        spss_filepath_from_settings = None
        try:
            self.update_status("กำลังโหลดการตั้งค่า...")
            xls = pd.ExcelFile(settings_filepath)

            if 'Settings' in xls.sheet_names:
                settings_df = pd.read_excel(xls, sheet_name='Settings')

                if 'PathFile' in settings_df.columns and not pd.isna(settings_df['PathFile'].iloc[0]):
                    spss_filepath_from_settings = str(settings_df['PathFile'].iloc[0])
                else:
                    raise ValueError("ไม่พบ PathFile ในไฟล์การตั้งค่า")

                if 'Filter_Var' in settings_df.columns and not pd.isna(settings_df['Filter_Var'].iloc[0]):
                    filter_value = str(settings_df['Filter_Var'].iloc[0])
                    self.filter_entry.config(state="normal")
                    self.filter_entry.delete(0, 'end')
                    self.filter_entry.insert(0, filter_value)

                if 'T2B_Choice' in settings_df.columns and not pd.isna(settings_df['T2B_Choice'].iloc[0]):
                    self.t2b_choice_var.set(str(settings_df['T2B_Choice'].iloc[0]))

                self.c_vars_to_compute = settings_df['C'].dropna().tolist() if 'C' in settings_df.columns else []
                self.vars_to_transform = {}
                for key in ['A', 'S', 'P', 'E', 'AgreeS', 'AgreeP']:
                    self.vars_to_transform[key] = settings_df[key].dropna().tolist() if key in settings_df.columns else []
            else:
                raise ValueError("ไม่พบชีท 'Settings' ในไฟล์การตั้งค่า")

            if 'Label' in xls.sheet_names:
                labels_df = pd.read_excel(xls, sheet_name='Label')

                if 'Index1_Code' in labels_df.columns and 'Index1_Label' in labels_df.columns:
                    index1_labels_df = labels_df[['Index1_Code', 'Index1_Label']].dropna()
                    self.index1_labels = dict(zip(index1_labels_df['Index1_Code'].astype(int), index1_labels_df['Index1_Label']))

                filter_var = self.filter_entry.get().strip()
                if filter_var and 'Filter_Code' in labels_df.columns and 'Filter_Label' in labels_df.columns:
                     self.filter_labels['var_name'] = filter_var
                     filter_labels_df = labels_df[['Filter_Code', 'Filter_Label']].dropna()
                     self.filter_labels['labels'] = dict(zip(filter_labels_df['Filter_Code'].astype(int), filter_labels_df['Filter_Label']))

        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถโหลดไฟล์การตั้งค่าได้: {e}", parent=self)
            self.reset_state()
            return

        self.update_status(f"โหลดตั้งค่าสำเร็จ. กำลังโหลดไฟล์ SPSS...", "info")

        if not self.load_spss_file(filepath=spss_filepath_from_settings):
            self.reset_state()
            return

        self.run_processing_with_loaded_settings()

    def start_reanalyze_process(self):
        """
        Workflow 3: โหลดไฟล์ที่ผ่านการประมวลผลแล้ว (Compute C) เพื่อวิเคราะห์ซ้ำ
        """
        self.reset_state()
        if self.load_processed_spss_file():
            self._infer_variables_from_transformed_df()
            
            self.update_status("โหลดไฟล์ที่ประมวลผลแล้วสำเร็จ", "success")
            self.show_message_in_display("โหลดไฟล์สำเร็จ\n\nกรุณาใส่ Filter (ถ้ามี) และกด 'วิเคราะห์และส่งออก Excel'")

            self.btn_analyze_export.config(state="normal")
            self.btn_define_labels.config(state="normal")
            self.btn_save_settings.config(state="disabled")
            self.filter_entry.config(state="normal")

    def load_spss_file(self, filepath=None):
        """โหลดไฟล์ SPSS ดั้งเดิม โดยรับ Path หรือเปิด Dialog"""
        if filepath is None:
            self.update_status("กำลังรอเลือกไฟล์ SPSS...")
            filepath = filedialog.askopenfilename(filetypes=[("SPSS Data File", "*.sav")])
            if not filepath:
                self.update_status("ยกเลิกการเลือกไฟล์", "warning")
                return False

        if not os.path.exists(filepath):
            self.update_status("ไฟล์ SPSS ไม่พบ", "danger")
            messagebox.showerror("ผิดพลาด", f"ไม่พบไฟล์ที่ระบุ:\n{filepath}")
            return False

        self.start_progress()
        self.update_status(f"กำลังโหลด: {os.path.basename(filepath)}...")
        try:
            self.df, meta = pyreadstat.read_sav(filepath)
            self.original_filepath = filepath
            self.spss_original_order = meta.column_names
            self.df = self.df[self.spss_original_order]
            self.update_status(f"โหลดไฟล์สำเร็จ! {len(self.df)} แถว", "success")
            self.stop_progress()
            return True
        except Exception as e:
            self.update_status("โหลดไฟล์ผิดพลาด", "danger")
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถโหลดไฟล์ได้: {e}")
            self.stop_progress()
            self.reset_state()
            return False

    def load_processed_spss_file(self):
        """โหลดไฟล์ SPSS ที่ผ่านการประมวลผลแล้ว (* Compute C.sav)"""
        self.update_status("กำลังรอเลือกไฟล์ SPSS ที่ประมวลผลแล้ว...")
        filepath = filedialog.askopenfilename(
            title="เลือกไฟล์ SPSS ที่ผ่านการ Compute C แล้ว",
            filetypes=[("SPSS Data File", "*.sav")]
        )
        if not filepath:
            self.update_status("ยกเลิกการเลือกไฟล์", "warning")
            return False
            
        self.start_progress()
        self.update_status(f"กำลังโหลด: {os.path.basename(filepath)}...")
        try:
            self.transformed_df, _ = pyreadstat.read_sav(filepath)
            
            base, _ = os.path.splitext(filepath)
            self.original_filepath = base.replace(" Compute C", "")
            
            self.update_status(f"โหลดไฟล์สำเร็จ! {len(self.transformed_df)} แถว", "success")
            self.stop_progress()
            return True
        except Exception as e:
            self.update_status("โหลดไฟล์ผิดพลาด", "danger")
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถโหลดไฟล์ที่ประมวลผลแล้วได้: {e}")
            self.stop_progress()
            self.reset_state()
            return False
    
    def _infer_variables_from_transformed_df(self):
        """
        พยายามสร้าง state ของตัวแปร (เช่น id_vars, vars_to_transform)
        จากคอลัมน์ที่มีอยู่ใน DataFrame ที่โหลดเข้ามา เพื่อให้ส่วนแสดงผลทำงานได้
        """
        if self.transformed_df is None:
            return

        self.vars_to_transform = {'A':[], 'S':[], 'P':[], 'E':[], 'AgreeS':[], 'AgreeP':[]}
        self.c_vars_to_compute = []
        self.computed_c_cols = []
        self.id_vars = []

        known_patterns = [
            re.compile(r'^(S|P|E|C)_\d+$'),
            re.compile(r'^N_(S|P|E|C)$'),
            re.compile(r'^(A|ZA|Index1)$')
        ]

        for col in self.transformed_df.columns:
            if col.startswith('S_'): self.vars_to_transform['S'].append(col)
            elif col.startswith('P_'): self.vars_to_transform['P'].append(col)
            elif col.startswith('E_'): self.vars_to_transform['E'].append(col)
            elif col.startswith('C_'): self.computed_c_cols.append(col)
            elif col == 'A': self.vars_to_transform['A'].append(col)
        
        for col in self.transformed_df.columns:
            is_known = False
            for pattern in known_patterns:
                if pattern.match(col):
                    is_known = True
                    break
            if not is_known:
                self.id_vars.append(col)
        
        print("Infered ID Vars:", self.id_vars)
        print("Infered S Vars:", self.vars_to_transform['S'])
        print("Infered C Vars (Computed):", self.computed_c_cols)

    # ===================================================================
    # VARIABLE SELECTION GUI
    # ===================================================================
    def open_c_variable_selector(self):
        """เปิดหน้าต่างสำหรับเลือกตัวแปร C และแก้ไขการเรียงลำดับ"""
        selector_window = bstrap.Toplevel(self)
        selector_window.title("ขั้นตอนที่ 1.1: เลือกตัวแปรสำหรับ Compute C")
        selector_window.geometry("700x500")
        selector_window.transient(self)
        selector_window.grab_set()
        self._center_toplevel(selector_window)

        filter_frame = bstrap.Frame(selector_window, padding=(10,10))
        filter_frame.pack(fill="x")

        bstrap.Label(filter_frame, text="กรองด้วยคำนำหน้า:").pack(side="left", padx=(0, 5))
        prefix_entry = bstrap.Entry(filter_frame)
        prefix_entry.pack(side="left", fill="x", expand=True, padx=5)

        main_frame = bstrap.Frame(selector_window, padding=(10,0,10,10))
        main_frame.pack(fill="both", expand=True)
        main_frame.grid_columnconfigure((0, 2), weight=1)
        main_frame.grid_rowconfigure(1, weight=1)

        bstrap.Label(main_frame, text="Available Variables").grid(row=0, column=0, padx=5, pady=5)
        available_frame = bstrap.Frame(main_frame)
        available_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        available_lb = Listbox(available_frame, selectmode="extended", exportselection=False, bg="#333", fg="white", selectbackground="#007bff")
        available_lb.pack(side="left", fill="both", expand=True)
        available_sb = bstrap.Scrollbar(available_frame, orient="vertical", command=available_lb.yview, bootstyle="secondary-round")
        available_sb.pack(side="right", fill="y")
        available_lb.config(yscrollcommand=available_sb.set)

        bstrap.Label(main_frame, text="Selected for Compute C").grid(row=0, column=2, padx=5, pady=5)
        selected_frame = bstrap.Frame(main_frame)
        selected_frame.grid(row=1, column=2, sticky="nsew", padx=5, pady=5)
        selected_lb = Listbox(selected_frame, selectmode="extended", exportselection=False, bg="#333", fg="white", selectbackground="#007bff")
        selected_lb.pack(side="left", fill="both", expand=True)
        selected_sb = bstrap.Scrollbar(selected_frame, orient="vertical", command=selected_lb.yview, bootstyle="secondary-round")
        selected_sb.pack(side="right", fill="y")
        selected_lb.config(yscrollcommand=selected_sb.set)

        button_frame = bstrap.Frame(main_frame)
        button_frame.grid(row=1, column=1, padx=10, pady=5, sticky="ns")

        def update_available_list(filter_text=""):
            available_lb.delete(0, "end")
            all_vars = self.spss_original_order
            selected_vars = list(selected_lb.get(0, "end"))
            display_vars = [v for v in all_vars if v not in selected_vars]
            if filter_text:
                display_vars = [v for v in display_vars if v.startswith(filter_text)]
            for var in display_vars:
                available_lb.insert("end", var)

        def move_to_selected():
            selected_items = [available_lb.get(i) for i in available_lb.curselection()]
            for item in selected_items:
                if item not in selected_lb.get(0, "end"):
                    selected_lb.insert("end", item)
            for i in reversed(available_lb.curselection()):
                available_lb.delete(i)

        def move_to_available():
            for i in reversed(selected_lb.curselection()):
                selected_lb.delete(i)
            update_available_list(prefix_entry.get())

        def confirm_and_proceed():
            self.c_vars_to_compute = list(selected_lb.get(0, "end"))
            if not self.c_vars_to_compute:
                messagebox.showwarning("คำเตือน", "คุณยังไม่ได้เลือกตัวแปรใดๆ", parent=selector_window)
                return
            selector_window.destroy()
            self.after(100, self.run_c_compute_and_proceed)

        bstrap.Button(filter_frame, text="Filter", width=6, command=lambda: update_available_list(prefix_entry.get()), bootstyle="secondary").pack(side="left", padx=5)
        bstrap.Button(button_frame, text=">", width=4, command=move_to_selected, bootstyle="primary-outline").pack(pady=5)
        bstrap.Button(button_frame, text="<", width=4, command=move_to_available, bootstyle="primary-outline").pack(pady=5)
        bstrap.Button(selector_window, text="ยืนยันและดำเนินการต่อ", command=confirm_and_proceed, bootstyle="success").pack(pady=10, fill='x', padx=10, ipady=4)

        update_available_list()

    def run_c_compute_and_proceed(self):
        """รันการคำนวณ C และไปขั้นตอนเลือกตัวแปรอื่นๆ"""
        self.start_progress()
        self.update_status(f"เลือก {len(self.c_vars_to_compute)} ตัวแปร. กำลัง Compute C...")
        if self._compute_c_variables_logic():
            self.update_status(f"Compute C สำเร็จ! สร้าง {len(self.computed_c_cols)} ตัวแปร.", "success")
            self.open_aspe_selector()
        else:
            self.update_status("Compute C ผิดพลาด", "danger")
            self.reset_state()
        self.stop_progress()

    def open_aspe_selector(self):
        """เปิดหน้าต่างเลือก A,S,P,E และเพิ่ม AgreeS, AgreeP, T2B Option"""
        selector_window = bstrap.Toplevel(self)
        selector_window.title("ขั้นตอนที่ 1.2: เลือกตัวแปรสำหรับแปลงข้อมูล")
        selector_window.geometry("800x650")
        selector_window.transient(self)
        selector_window.grab_set()
        self._center_toplevel(selector_window)

        style = bstrap.Style()
        style.configure('TNotebook', tabposition='n')
        style.configure('TNotebook.Tab', padding=(25, 8), font=('Helvetica', 10))

        tab_view_container = bstrap.Frame(selector_window)
        tab_view_container.pack(fill="both", expand=True, padx=10, pady=(10,0))

        tab_view = bstrap.Notebook(tab_view_container, style='TNotebook')
        tab_view.pack(fill="both", expand=True)

        tab_colors = {"A": "info", "S": "success", "P": "warning", "E": "danger", "AgreeS": "primary", "AgreeP": "secondary"}
        tabs = {}
        for name, color in tab_colors.items():
            colored_frame = bstrap.Frame(tab_view, bootstyle=color)
            tab_view.add(colored_frame, text=name)
            tabs[name] = colored_frame

        listboxes = {}
        all_selected_vars = set()

        def create_selector_tab(parent_frame, name):
            main_frame = bstrap.Frame(parent_frame)
            main_frame.pack(fill="both", expand=True, padx=5, pady=5)
            main_frame.grid_columnconfigure((0, 2), weight=1); main_frame.grid_rowconfigure(0, weight=1)

            available_frame = bstrap.Frame(main_frame); available_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
            bstrap.Label(available_frame, text="Available Variables").pack()
            available_lb = Listbox(available_frame, selectmode="extended", exportselection=False, bg="#333", fg="white", selectbackground="#007bff"); available_lb.pack(side="left", fill="both", expand=True)
            available_sb = bstrap.Scrollbar(available_frame, orient="vertical", command=available_lb.yview, bootstyle="secondary-round"); available_sb.pack(side="right", fill="y"); available_lb.config(yscrollcommand=available_sb.set)

            selected_frame = bstrap.Frame(main_frame); selected_frame.grid(row=0, column=2, sticky="nsew", padx=5, pady=5)
            bstrap.Label(selected_frame, text=f"Selected for '{name}'").pack()
            selected_lb = Listbox(selected_frame, selectmode="extended", exportselection=False, bg="#333", fg="white", selectbackground="#007bff"); selected_lb.pack(side="left", fill="both", expand=True)
            selected_sb = bstrap.Scrollbar(selected_frame, orient="vertical", command=selected_lb.yview, bootstyle="secondary-round"); selected_sb.pack(side="right", fill="y"); selected_lb.config(yscrollcommand=selected_sb.set)

            button_frame = bstrap.Frame(main_frame); button_frame.grid(row=0, column=1, padx=10, pady=5, sticky="ns")

            def move_to_selected():
                selected_items = [available_lb.get(i) for i in available_lb.curselection()]

                for item in selected_items:
                    if item not in selected_lb.get(0, "end"):
                        selected_lb.insert("end", item)
                        all_selected_vars.add(item)

                for i in reversed(available_lb.curselection()):
                    available_lb.delete(i)

            def move_to_available():
                for i in reversed(selected_lb.curselection()):
                    item = selected_lb.get(i)
                    all_selected_vars.remove(item)
                    selected_lb.delete(i)
                update_available_list_for_tab(available_lb)

            bstrap.Button(button_frame, text=">", width=4, command=move_to_selected, bootstyle="primary-outline").pack(pady=5)
            bstrap.Button(button_frame, text="<", width=4, command=move_to_available, bootstyle="primary-outline").pack(pady=5)
            return {"available": available_lb, "selected": selected_lb}

        def update_available_list_for_tab(listbox):
            listbox.delete(0, 'end')
            original_vars = [v for v in self.spss_original_order if v not in self.computed_c_cols and v not in all_selected_vars]
            for var in original_vars: listbox.insert("end", var)

        for name, colored_parent_frame in tabs.items():
            listboxes[name] = create_selector_tab(colored_parent_frame, name)

        for name in tabs.keys(): update_available_list_for_tab(listboxes[name]["available"])

        def confirm_and_transform():
            for name, lbs in listboxes.items(): self.vars_to_transform[name] = list(lbs["selected"].get(0, "end"))
            selector_window.destroy()
            self.after(100, self.run_full_transformation_and_save)

        option_frame = bstrap.Labelframe(selector_window, text=" T2B Options ", bootstyle="info", padding=10)
        option_frame.pack(pady=5, padx=10, fill='x')

        bstrap.Label(option_frame, text="เลือก Code ด้านดี (T2B) สำหรับ AgreeS/AgreeP:").pack(side='left', padx=(0, 10))
        rb1 = bstrap.Radiobutton(option_frame, text="5+4 (Default)", variable=self.t2b_choice_var, value="5+4", bootstyle="primary-toolbutton")
        rb1.pack(side='left', padx=5)
        rb2 = bstrap.Radiobutton(option_frame, text="1+2", variable=self.t2b_choice_var, value="1+2", bootstyle="primary-toolbutton")
        rb2.pack(side='left', padx=5)

        bstrap.Button(selector_window, text="ยืนยัน, แปลงข้อมูล และบันทึก", command=confirm_and_transform, bootstyle="success").pack(pady=10, padx=10, fill='x', ipady=4)

    def run_full_transformation_and_save(self):
        self.start_progress()
        self.update_status("กำลัง Recode ตัวแปร A...")
        if not self._recode_a_variables_logic():
            self.update_status("Recode A ผิดพลาด", "danger"); self.stop_progress(); self.reset_state(); return

        self.update_status("กำลังแปลงข้อมูล (Variables to Cases)...")
        if not self._run_full_transformation_logic():
            self.update_status("แปลงข้อมูลผิดพลาด", "danger"); self.stop_progress(); self.reset_state(); return

        self.update_status("แปลงข้อมูลสำเร็จ. กำลังบันทึกไฟล์อัตโนมัติ...", "success")
        if not self._auto_save_spss(self.transformed_df):
            self.update_status("บันทึก .sav อัตโนมัติไม่สำเร็จ", "danger"); self.stop_progress(); self.reset_state(); return

        self.show_message_in_display("ประมวลผลข้อมูลสำเร็จ\n\nกรุณาใส่ Filter (ถ้ามี) และกด 'วิเคราะห์และส่งออก Excel'")

        self.btn_analyze_export.config(state="normal")
        self.btn_define_labels.config(state="normal")
        self.btn_save_settings.config(state="normal")
        self.filter_entry.config(state="normal")
        self.stop_progress()

    def run_processing_with_loaded_settings(self):
        self.start_progress()

        self.update_status(f"กำลัง Compute C จากการตั้งค่า...")
        if not self._compute_c_variables_logic():
            self.update_status("Compute C ผิดพลาด", "danger"); self.stop_progress(); self.reset_state(); return
        self.update_status(f"Compute C สำเร็จ! สร้าง {len(self.computed_c_cols)} ตัวแปร.", "success")

        self.update_status("กำลัง Recode ตัวแปร A...")
        if not self._recode_a_variables_logic():
            self.update_status("Recode A ผิดพลาด", "danger"); self.stop_progress(); self.reset_state(); return

        self.update_status("กำลังแปลงข้อมูล (Variables to Cases)...")
        if not self._run_full_transformation_logic():
            self.update_status("แปลงข้อมูลผิดพลาด", "danger"); self.stop_progress(); self.reset_state(); return

        self.update_status("แปลงข้อมูลสำเร็จ. กำลังบันทึกไฟล์อัตโนมัติ...", "success")
        if not self._auto_save_spss(self.transformed_df):
            self.update_status("บันทึก .sav อัตโนมัติไม่สำเร็จ", "danger"); self.stop_progress(); self.reset_state(); return

        self.update_status("ประมวลผลข้อมูลสำเร็จ. เริ่มการวิเคราะห์และส่งออกอัตโนมัติ...", "info")
        self.after(100, lambda: self.run_analysis_and_export(automated=True))

    # ===================================================================
    # PROCESSING LOGIC (Back-end)
    # ===================================================================
    def _compute_c_variables_logic(self):
        if not self.c_vars_to_compute:
            messagebox.showerror("ผิดพลาด", "ไม่มีตัวแปรที่ถูกเลือกสำหรับคำนวณ C")
            return False
        try:
            first_var = self.c_vars_to_compute[0]
            if '#' not in first_var:
                messagebox.showerror("รูปแบบผิดพลาด", f"ตัวแปรที่เลือก ({first_var}) ไม่มีรูปแบบที่ถูกต้อง (เช่น 'PREFIX#GROUP$ITEM')")
                return False
            deduced_prefix = first_var.split('#')[0]
            pattern = re.compile(rf"^{re.escape(deduced_prefix)}#(\d+)\$(\d+)")
            groups = {}
            for col in self.c_vars_to_compute:
                match = pattern.match(col)
                if match:
                    group_num = int(match.group(1))
                    if group_num not in groups:
                        groups[group_num] = []
                    groups[group_num].append(col)
            if not groups:
                messagebox.showerror("ผิดพลาด", f"ไม่พบตัวแปรที่ตรงกับรูปแบบ '{deduced_prefix}#Group$Item' จากตัวแปรที่คุณเลือก")
                return False

            self.computed_c_cols = []
            new_cols_data = {}

            max_item_num = max((int(m.group(2)) for c in self.c_vars_to_compute if (m := pattern.match(c))), default=0)
            if max_item_num == 0:
                messagebox.showerror("ผิดพลาด", "ไม่สามารถหา Item number สูงสุดจากตัวแปรที่เลือกสำหรับ C ได้")
                return False

            for j in range(1, max_item_num + 1):
                for i in sorted(groups.keys()):
                    group_vars = groups[i]
                    main_var = f"{deduced_prefix}#{i}${j}"
                    if main_var in group_vars:
                        other_vars = [v for v in group_vars if v != main_var]
                        if not other_vars:
                            continue
                        new_c_name = f"C{j}.{i}"
                        mean_of_others = self.df[other_vars].mean(axis=1)
                        new_cols_data[new_c_name] = ((self.df[main_var] - mean_of_others) + 1) / 2
                        self.computed_c_cols.append(new_c_name)

            if not self.computed_c_cols:
                messagebox.showwarning("ไม่สำเร็จ", "ไม่สามารถสร้างตัวแปร C ได้ อาจเพราะโครงสร้างตัวแปรไม่ถูกต้อง")
                return False

            if new_cols_data:
                self.df = pd.concat([self.df, pd.DataFrame(new_cols_data)], axis=1)

            return True
        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"เกิดข้อผิดพลาดระหว่างคำนวณตัวแปร C: {e}")
            return False

    def _recode_a_variables_logic(self):
        a_vars_to_process = self.vars_to_transform.get('A', [])
        self.za_cols = []
        if not a_vars_to_process:
            return True
        try:
            za_map = {0: 0, 1: 0.05, 2: 0.12, 3: 0.27, 4: 0.50, 5: 0.73, 6: 0.88, 7: 0.95, 8: 1.00}
            new_za_cols_data = {}

            for var in a_vars_to_process:
                if var in self.df.columns and pd.api.types.is_numeric_dtype(self.df[var]):
                    temp_series = self.df[var].copy()
                    temp_series.replace(9, 0, inplace=True)
                    self.df[var] = temp_series

                    za_var_name = 'Z' + var
                    new_za_cols_data[za_var_name] = self.df[var].map(za_map).fillna(self.df[var])
                    self.za_cols.append(za_var_name)

            if new_za_cols_data:
                self.df = pd.concat([self.df, pd.DataFrame(new_za_cols_data)], axis=1)

            return True
        except Exception as e:
            messagebox.showerror("Recode Error", f"เกิดข้อผิดพลาดขณะแปลงค่าตัวแปร A: {e}")
            return False

    def _run_full_transformation_logic(self):
        try:
            temp_df = self.df.copy()
            all_transform_vars = set(self.computed_c_cols)
            for key, var_list in self.vars_to_transform.items():
                if key not in ['AgreeS', 'AgreeP']:
                    all_transform_vars.update(var_list)
            all_transform_vars.update(self.za_cols)

            self.id_vars = [col for col in self.df.columns if col not in all_transform_vars]

            A_PAT, ZA_PAT = re.compile(r".*?#(\d+)$"), re.compile(r"Z.*?#(\d+)$")
            SPE_PAT, C_PAT = re.compile(r".*?#(\d+)\$(\d+)$"), re.compile(r"C(\d+)\.(\d+)$")

            maps = {'A': {}, 'S': {}, 'P': {}, 'E': {}, 'C': {}, 'ZA': {}}
            groups = {'S': set(), 'P': set(), 'E': set(), 'C': set()}
            max_index = 0

            for var in self.vars_to_transform.get('A', []):
                if match := A_PAT.match(var): idx = int(match.group(1)); maps['A'][idx] = var; max_index = max(max_index, idx)
            for var in self.za_cols:
                if match := ZA_PAT.match(var): idx = int(match.group(1)); maps['ZA'][idx] = var
            for key in ['S', 'P', 'E']:
                for var in self.vars_to_transform.get(key, []):
                    if match := SPE_PAT.match(var):
                        grp, idx = int(match.group(1)), int(match.group(2))
                        if grp not in maps[key]: maps[key][grp] = {}
                        maps[key][grp][idx] = var; groups[key].add(grp); max_index = max(max_index, idx)
            for var in self.computed_c_cols:
                if match := C_PAT.match(var):
                    idx, grp = int(match.group(1)), int(match.group(2))
                    if grp not in maps['C']: maps['C'][grp] = {}
                    maps['C'][grp][idx] = var; groups['C'].add(grp); max_index = max(max_index, idx)

            if max_index == 0: messagebox.showerror("ผิดพลาด", "ไม่สามารถหา Index สำหรับการแปลงข้อมูลได้\nกรุณาตรวจสอบรูปแบบของตัวแปรที่เลือก"); return False

            new_data = []
            for _, row in temp_df.iterrows():
                base_record = {k: row[k] for k in self.id_vars}
                for j in range(1, max_index + 1):
                    new_record = base_record.copy(); new_record['Index1'] = j
                    if (a_source := maps['A'].get(j)) and a_source in row: new_record['A'] = row[a_source]
                    if (za_source := maps['ZA'].get(j)) and za_source in row: new_record['ZA'] = row[za_source]
                    for key in ['S', 'P', 'E', 'C']:
                        for i in sorted(list(groups[key])):
                            if (source_var := maps[key].get(i, {}).get(j)) and source_var in row: new_record[f'{key}_{i}'] = row[source_var]
                    new_data.append(new_record)

            self.transformed_df = pd.DataFrame(new_data)

            value_cols = [col for col in ['A', 'ZA'] if col in self.transformed_df.columns]
            for key in ['S', 'P', 'E', 'C']: value_cols.extend([c for c in self.transformed_df.columns if c.startswith(f"{key}_")])
            if value_cols_in_df := [c for c in value_cols if c in self.transformed_df.columns]:
                self.transformed_df.dropna(subset=value_cols_in_df, how='all', inplace=True)

            for key, col_name in {'S':'N_S', 'P':'N_P', 'C':'N_C', 'E':'N_E'}.items():
                if cols := [c for c in self.transformed_df.columns if c.startswith(f'{key}_')]: self.transformed_df[col_name] = self.transformed_df[cols].mean(axis=1)

            final_ordered_cols = self.id_vars + ['Index1']
            for col in ['N_S', 'N_P', 'N_C', 'N_E', 'A', 'ZA']:
                if col in self.transformed_df.columns: final_ordered_cols.append(col)

            all_new_keys = {c for key in ['S', 'P', 'E', 'C'] for c in self.transformed_df.columns if c.startswith(f"{key}_")}
            sorted_new_keys = sorted(list(all_new_keys), key=lambda x: (x.split('_')[0], int(x.split('_')[1])))
            final_ordered_cols.extend(sorted_new_keys)

            self.transformed_df = self.transformed_df[[c for c in final_ordered_cols if c in self.transformed_df.columns]]
            return True
        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"เกิดข้อผิดพลาดระหว่างการแปลงข้อมูล: {e}"); return False

    def _auto_save_spss(self, dataframe_to_save):
        if dataframe_to_save is None:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลให้บันทึก")
            return False
        if not self.original_filepath:
            messagebox.showerror("ผิดพลาด", "ไม่พบ Path ของไฟล์ต้นฉบับ")
            return False

        try:
            base, _ = os.path.splitext(self.original_filepath)
            new_filepath = f"{base} Compute C.sav"

            pyreadstat.write_sav(dataframe_to_save, new_filepath)
            self.update_status(f"บันทึกไฟล์ใหม่ที่: {new_filepath}", "success")
            return True
        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถบันทึกไฟล์อัตโนมัติได้: {e}")
            return False

    def display_table(self, dataframe):
        """แสดง DataFrame ใน Tableview"""
        for widget in self.right_frame.winfo_children(): widget.destroy()
        coldata = [{"text": col, "stretch": True} for col in dataframe.columns]
        rowdata = dataframe.head(1000).fillna('').values.tolist()
        self.table = Tableview(master=self.right_frame, coldata=coldata, rowdata=rowdata, paginated=True, searchable=True, bootstyle="primary", pagesize=50)
        self.table.pack(fill="both", expand=True)

    def show_message_in_display(self, message_text):
        """แสดงข้อความในพื้นที่แสดงผลด้านขวา"""
        for widget in self.right_frame.winfo_children():
            widget.destroy()
        message_label = bstrap.Label(
            self.right_frame,
            text=message_text,
            font=("Helvetica", 16),
            bootstyle="secondary",
            justify="left"
        )
        message_label.pack(side="top", anchor="nw", padx=20, pady=20)

    def _load_factor_output_text_from_excel(self):
        filepath = self.last_excel_filepath
        if not filepath or not os.path.exists(filepath):
            return ""
        try:
            workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            if "Factor_Output" not in workbook.sheetnames:
                workbook.close()
                return ""
            worksheet = workbook["Factor_Output"]
            lines = []
            for row in worksheet.iter_rows(min_row=1, max_col=1, values_only=True):
                value = row[0]
                lines.append("" if value is None else str(value))
            workbook.close()
            return "\n".join(lines).strip()
        except Exception:
            return ""

    def display_analysis_tabs(self, analysis_text):
        for widget in self.right_frame.winfo_children():
            widget.destroy()

        notebook = bstrap.Notebook(self.right_frame)
        notebook.pack(fill="both", expand=True, padx=5, pady=5)

        analysis_frame = bstrap.Frame(notebook)
        notebook.add(analysis_frame, text=" ผลการวิเคราะห์ (Factor Output) ")

        text_area = scrolledtext.ScrolledText(analysis_frame, wrap="none")
        text_area.pack(fill="both", expand=True, padx=5, pady=5)
        display_text = analysis_text if analysis_text and analysis_text.strip() else self._load_factor_output_text_from_excel()
        if not display_text:
            display_text = "ไม่พบข้อความผลการวิเคราะห์ (ลองรันใหม่อีกครั้ง)"
        text_area.insert("1.0", display_text)
        text_area.configure(
            font=("Consolas", 10),
            foreground="#E6E6E6",
            background="#1F2D3A",
            insertbackground="#E6E6E6",
            state="disabled"
        )

        desc_tab_main_frame = bstrap.Frame(notebook)
        notebook.add(desc_tab_main_frame, text=" คำอธิบายตัวแปร ")

        desc_canvas = tkinter.Canvas(desc_tab_main_frame, borderwidth=0, highlightthickness=0)
        desc_scrollbar = bstrap.Scrollbar(desc_tab_main_frame, orient="vertical", command=desc_canvas.yview, bootstyle="round")
        desc_scrollable_frame = bstrap.Frame(desc_canvas, padding=20)

        desc_scrollable_frame.bind(
            "<Configure>",
            lambda e: desc_canvas.configure(scrollregion=desc_canvas.bbox("all"))
        )

        desc_canvas.create_window((0, 0), window=desc_scrollable_frame, anchor="nw")
        desc_canvas.configure(yscrollcommand=desc_scrollbar.set)

        desc_canvas.pack(side="left", fill="both", expand=True)
        desc_scrollbar.pack(side="right", fill="y")

        header = bstrap.Label(desc_scrollable_frame, text="คำอธิบายและตัวแปรที่เลือกใน Model", font=("Helvetica", 16, "bold"), bootstyle="primary")
        header.pack(pady=(0, 20), anchor="w")

        descriptions = {
            "S (Sense)": "การรับรู้ผ่านประสาทสัมผัส (Sensory Perception) เช่น รูป รส กลิ่น เสียง สัมผัส ที่เกี่ยวข้องกับแบรนด์",
            "P (Personality/People)": "บุคลิกภาพของแบรนด์ (Brand Personality) หรือการรับรู้เกี่ยวกับผู้คนที่เกี่ยวข้องกับแบรนด์ (เช่น พนักงาน, ผู้ใช้งานคนอื่น)",
            "C (Cognition)": "การรับรู้เชิงเหตุผล (Cognitive Perception) เช่น ความคิด ความเชื่อ ความรู้ ความเข้าใจ เกี่ยวกับคุณสมบัติ ประโยชน์ หรือหน้าที่ของแบรนด์",
            "A (Action/Attitude)": "พฤติกรรมหรือทัศนคติ (Behavioral Action/Attitude) ที่มีต่อแบรนด์ เช่น การซื้อ การใช้งาน การบอกต่อ ความตั้งใจซื้อ",
            "E (Emotion)": "อารมณ์ความรู้สึก (Emotional Connection) ที่มีต่อแบรนด์ เช่น ความสุข ความผูกพัน ความไว้วางใจ",
            "AgreeS / AgreeP": "ตัวแปรที่ใช้วัดความเห็นด้วย เพื่อคำนวณ %T2B"
        }
        
        c_vars_for_display = self.c_vars_to_compute if self.c_vars_to_compute else self.computed_c_cols
        all_vars = {
            "S": self.vars_to_transform.get('S', []),
            "P": self.vars_to_transform.get('P', []),
            "C": c_vars_for_display,
            "A": self.vars_to_transform.get('A', []),
            "E": self.vars_to_transform.get('E', []),
            "AgreeS": self.vars_to_transform.get('AgreeS', []),
            "AgreeP": self.vars_to_transform.get('AgreeP', []),
        }

        key_map = {
            "S (Sense)": "S",
            "P (Personality/People)": "P",
            "C (Cognition)": "C",
            "A (Action/Attitude)": "A",
            "E (Emotion)": "E",
            "AgreeS / AgreeP": ["AgreeS", "AgreeP"],
        }

        for var_display, desc in descriptions.items():
            if var_display == "C (Cognition)" and not self.c_vars_to_compute and self.computed_c_cols:
                desc += "\n(หมายเหตุ: ตัวแปร C ที่แสดงด้านล่างคือชื่อคอลัมน์จากไฟล์ที่ประมวลผลแล้ว)"
            
            var_label = bstrap.Label(desc_scrollable_frame, text=var_display, font=("Helvetica", 12, "bold"), bootstyle="info")
            var_label.pack(anchor="w", pady=(15, 2))
            desc_label = bstrap.Label(desc_scrollable_frame, text=desc, wraplength=700, justify="left")
            desc_label.pack(anchor="w", padx=(10, 0))

            data_key = key_map[var_display]
            variable_list = []
            if isinstance(data_key, list):
                for k in data_key:
                    if all_vars.get(k):
                        variable_list.extend([f"--- {k} ---"] + all_vars.get(k, []))
            else:
                 variable_list = all_vars.get(data_key, [])

            if variable_list:
                vars_frame = bstrap.Labelframe(desc_scrollable_frame, text="  ตัวแปรที่เลือก/ตรวจพบ  ", bootstyle="secondary", padding=5)
                vars_frame.pack(fill="x", anchor="w", padx=(10, 0), pady=(5, 10), expand=True)

                vars_string = "\n".join(variable_list)
                num_lines = len(variable_list)
                height = min(num_lines, 10)

                vars_text_area = tkinter.Text(vars_frame, height=height, wrap="none", relief="flat", borderwidth=0, highlightthickness=0)
                vars_text_area.pack(fill="x", expand=True, pady=5, padx=5)
                vars_text_area.insert("1.0", vars_string)
                vars_text_area.configure(state="disabled")

    # ===================================================================
    # ANALYSIS AND EXPORT (STEP 2)
    # ===================================================================
    def open_label_editor(self):
        if self.transformed_df is None:
            messagebox.showerror("ผิดพลาด", "ยังไม่มีข้อมูลที่ประมวลผลแล้ว")
            return

        editor_window = bstrap.Toplevel(self)
        editor_window.title("กำหนด Label")
        editor_window.geometry("600x500")
        editor_window.transient(self)
        editor_window.grab_set()
        self._center_toplevel(editor_window)

        main_frame = bstrap.Frame(editor_window, padding=10)
        main_frame.pack(fill="both", expand=True)

        canvas = tkinter.Canvas(main_frame, borderwidth=0, highlightthickness=0)
        scrollbar = bstrap.Scrollbar(main_frame, orient="vertical", command=canvas.yview, bootstyle="round")
        scrollable_frame = bstrap.Frame(canvas, padding=10)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        index1_entries = {}
        filter_entries = {}

        index1_frame = bstrap.Labelframe(scrollable_frame, text=" Labels for Index1 ", padding=10, bootstyle="info")
        index1_frame.pack(fill="x", pady=(0, 10))

        unique_index1 = sorted(self.transformed_df['Index1'].dropna().unique())
        bstrap.Label(index1_frame, text="Code", font="-weight bold").grid(row=0, column=0, padx=5, pady=2)
        bstrap.Label(index1_frame, text="Label", font="-weight bold").grid(row=0, column=1, padx=5, pady=2)

        for i, code in enumerate(unique_index1):
            code = int(code)
            bstrap.Label(index1_frame, text=str(code)).grid(row=i+1, column=0, padx=5, pady=2, sticky='e')
            entry = bstrap.Entry(index1_frame, width=40)
            entry.grid(row=i+1, column=1, padx=5, pady=2, sticky='w')
            if code in self.index1_labels:
                entry.insert(0, self.index1_labels[code])
            index1_entries[code] = entry

        filter_var = self.filter_entry.get().strip()
        if filter_var and filter_var in self.transformed_df.columns:
            filter_frame = bstrap.Labelframe(scrollable_frame, text=f" Labels for {filter_var} ", padding=10, bootstyle="info")
            filter_frame.pack(fill="x", pady=10)

            unique_filter_vals = sorted(self.transformed_df[filter_var].dropna().unique())
            bstrap.Label(filter_frame, text="Code", font="-weight bold").grid(row=0, column=0, padx=5, pady=2)
            bstrap.Label(filter_frame, text="Label", font="-weight bold").grid(row=0, column=1, padx=5, pady=2)

            for i, code in enumerate(unique_filter_vals):
                code = int(code)
                bstrap.Label(filter_frame, text=str(code)).grid(row=i+1, column=0, padx=5, pady=2, sticky='e')
                entry = bstrap.Entry(filter_frame, width=40)
                entry.grid(row=i+1, column=1, padx=5, pady=2, sticky='w')
                if code in self.filter_labels.get('labels', {}):
                     entry.insert(0, self.filter_labels['labels'][code])
                filter_entries[code] = entry

        def save_labels():
            self.index1_labels.clear()
            for code, entry in index1_entries.items():
                if label_text := entry.get().strip():
                    self.index1_labels[code] = label_text

            self.filter_labels.clear()
            if filter_var:
                self.filter_labels['var_name'] = filter_var
                self.filter_labels['labels'] = {}
                for code, entry in filter_entries.items():
                    if label_text := entry.get().strip():
                        self.filter_labels['labels'][code] = label_text

            messagebox.showinfo("สำเร็จ", "บันทึก Labels เรียบร้อยแล้ว", parent=editor_window)
            editor_window.destroy()

        save_button = bstrap.Button(editor_window, text="บันทึก Labels", command=save_labels, bootstyle="success")
        save_button.pack(pady=10, padx=10, fill='x')

    def run_analysis_and_export(self, automated=False):
        if self.transformed_df is None:
            messagebox.showerror("ผิดพลาด", "ไม่มีข้อมูลที่แปลงแล้ว (Transformed Data) สำหรับการวิเคราะห์")
            return

        self.update_status("กำลังเตรียมการวิเคราะห์...")
        self.start_progress()

        primary_filter = "Index1"
        cross_filter = self.filter_entry.get().strip()

        if not cross_filter and not automated:
            proceed = messagebox.askyesno("ยืนยัน", "ยังไม่ได้ระบุ Filter ไขว้\nต้องการดำเนินการต่อหรือไม่?", parent=self)
            if not proceed:
                self.stop_progress()
                self.update_status("ยกเลิกโดยผู้ใช้", "warning")
                return

        self.update_status("กำลังสร้างข้อมูลสรุป...")
        summary_df = self._create_summary_df_logic(primary_filter=primary_filter, cross_filter=cross_filter)
        if summary_df is None:
            self.stop_progress(); return

        try:
            summary_df = self._calculate_and_add_t2b_values(
                summary_df,
                primary_filter=primary_filter,
                cross_filter=cross_filter
            )
        except Exception as e:
            messagebox.showwarning("T2B Warning", f"ไม่สามารถคำนวณค่า T2B ได้: {e}")

        self.update_status("กำลังรัน Factor & Regression...")
        results_for_saving, full_output_text = self._run_factor_regression_logic(primary_filter=primary_filter, cross_filter=cross_filter)
        if results_for_saving is None:
            self.stop_progress(); return

        self.update_status("กำลังบันทึกผลลัพธ์ลง Excel...")
        self.save_all_results_to_excel(summary_df, results_for_saving, full_output_text)

        self.display_analysis_tabs(full_output_text)

        self.stop_progress()
        self.update_status("วิเคราะห์และส่งออกเสร็จสมบูรณ์", "success")


    def _create_summary_df_logic(self, primary_filter, cross_filter):
        """ตรรกะการสร้าง Summary DataFrame"""
        try:
            cols_to_average = [col for col in self.transformed_df.columns if re.match(r'^(S|P|C|E)_\d+$', col)]
            if not cols_to_average:
                messagebox.showwarning("คำเตือน", "ไม่พบข้อมูลคอลัมน์ S, P, C, E สำหรับสร้างสรุป")
                return None
            corr_df = self.transformed_df.copy()
            if 'E_4' in corr_df.columns and 'E_5' in corr_df.columns:
                corr_df['E_45'] = (corr_df['E_4'] + corr_df['E_5']) / 2

            df_for_summary = self.transformed_df
            groups_to_summarize = OrderedDict()
            corr_groups = OrderedDict()
            groups_to_summarize['Overall'] = df_for_summary
            corr_groups['Overall'] = corr_df

            primary_values = []
            if primary_filter and primary_filter in df_for_summary.columns:
                primary_values = sorted(df_for_summary[primary_filter].dropna().unique())
                for p_val in primary_values:
                    filter_name = f"{primary_filter}={p_val}"
                    if filter_name not in groups_to_summarize:
                            groups_to_summarize[filter_name] = df_for_summary[df_for_summary[primary_filter] == p_val]
                            corr_groups[filter_name] = corr_df[corr_df[primary_filter] == p_val]

            if cross_filter and cross_filter in df_for_summary.columns:
                cross_values = sorted(df_for_summary[cross_filter].dropna().unique())
                for c_val in cross_values:
                    filter_name_cross = f"{cross_filter}={c_val}"
                    if filter_name_cross not in groups_to_summarize:
                            groups_to_summarize[filter_name_cross] = df_for_summary[df_for_summary[cross_filter] == c_val]
                            corr_groups[filter_name_cross] = corr_df[corr_df[cross_filter] == c_val]

                    if primary_filter and primary_filter in df_for_summary.columns:
                        for p_val in primary_values:
                            nested_name = f"{primary_filter}={p_val}+{cross_filter}={c_val}"
                            subset = df_for_summary[(df_for_summary[primary_filter] == p_val) & (df_for_summary[cross_filter] == c_val)]
                            groups_to_summarize[nested_name] = subset
                            corr_groups[nested_name] = corr_df[(corr_df[primary_filter] == p_val) & (corr_df[cross_filter] == c_val)]

            summary_list = []
            avg_cols_base = cols_to_average.copy()
            if 'A' in df_for_summary.columns: avg_cols_base.append('A')
            if 'ZA' in df_for_summary.columns: avg_cols_base.append('ZA')

            for name, df_group in groups_to_summarize.items():
                    if not df_group.empty:
                        avg_values = df_group[avg_cols_base].mean()
                        summary_row_df = pd.DataFrame([avg_values])
                        summary_row_df['Filter'] = name

                        index1_val = 0
                        match = re.search(r'Index1=(\d+)', name)
                        if match:
                            index1_val = int(match.group(1))

                        summary_row_df['Index1'] = index1_val
                        summary_list.append(summary_row_df)

            if not summary_list:
                messagebox.showwarning("ไม่มีข้อมูล", "ไม่พบข้อมูลสำหรับสร้างสรุปตาม Filter ที่กำหนด")
                return None

            final_summary_df = pd.concat(summary_list, ignore_index=True)

            def map_labels(row):
                filter_str = row['Filter']
                if filter_str == 'Overall': return 'Overall'

                index1_val_match = re.search(r'Index1=(\d+)', filter_str)
                current_filter_var = self.filter_entry.get().strip()
                filter_val_match = None
                if current_filter_var and current_filter_var in filter_str:
                    filter_val_match = re.search(rf'{re.escape(current_filter_var)}=(\d+)', filter_str)

                index1_label = None
                if index1_val_match:
                    code = int(index1_val_match.group(1))
                    index1_label = self.index1_labels.get(code)

                filter_label = None
                if filter_val_match:
                    code = int(filter_val_match.group(1))
                    if self.filter_labels.get('var_name') == current_filter_var:
                        filter_label = self.filter_labels.get('labels', {}).get(code)

                if index1_label and filter_label: return f"{index1_label} - {filter_label}"
                elif index1_label: return index1_label
                elif filter_label: return filter_label
                else: return filter_str

            final_summary_df['Labe Index1'] = final_summary_df.apply(map_labels, axis=1)

            for col in cols_to_average:
                if col in final_summary_df.columns:
                    final_summary_df[col] *= 100

            s_cols = sorted([c for c in cols_to_average if c.startswith('S_')], key=lambda x: int(x.split('_')[1]))
            p_cols = sorted([c for c in cols_to_average if c.startswith('P_')], key=lambda x: int(x.split('_')[1]))

            if 'A' in corr_df.columns:
                e_corr_map = {'CorE_11': 'E_1', 'CorE_12': 'E_2', 'CorE_13': 'E_3', 'CorE_14': 'E_45', 'CorE_15': 'E_6'}
                source_e_cols = [col for col in e_corr_map.values() if col in corr_df.columns]
                rename_dict = {v: k for k, v in e_corr_map.items()}

                corr_rows = []
                for name, df_group in corr_groups.items():
                    if df_group.empty:
                        continue
                    row = {'Filter': name}
                    if s_cols:
                        s_corr = df_group[s_cols].corrwith(df_group['A'])
                        for col, val in s_corr.items():
                            row['cor_' + col] = val
                    if p_cols:
                        p_corr = df_group[p_cols].corrwith(df_group['A'])
                        for col, val in p_corr.items():
                            row['cor_' + col] = val
                    if source_e_cols:
                        e_corr = df_group[source_e_cols].corrwith(df_group['A'])
                        for col, val in e_corr.items():
                            row[rename_dict.get(col, col)] = val
                    corr_rows.append(row)

                if corr_rows:
                    corr_by_filter_df = pd.DataFrame(corr_rows)
                    final_summary_df = pd.merge(final_summary_df, corr_by_filter_df, on='Filter', how='left')

            return final_summary_df
        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถสร้างข้อมูลสรุปได้: {e}")
            return None

    def _calculate_and_add_t2b_values(self, summary_df, primary_filter="Index1", cross_filter=None):
        """คำนวณ %T2B สำหรับ AgreeS/P ตามลำดับการเลือก"""
        agree_s_vars = self.vars_to_transform.get('AgreeS', [])
        agree_p_vars = self.vars_to_transform.get('AgreeP', [])

        s_cols_in_summary = sorted([c for c in summary_df.columns if c.startswith('S_') and 'cor' not in c and 'agree' not in c], key=lambda x: int(x.split('_')[1]))
        p_cols_in_summary = sorted([c for c in summary_df.columns if c.startswith('P_') and 'cor' not in c and 'agree' not in c], key=lambda x: int(x.split('_')[1]))

        for s_col in s_cols_in_summary:
            agree_col_name = 'agree_' + s_col
            if agree_col_name not in summary_df.columns:
                summary_df[agree_col_name] = np.nan
        for p_col in p_cols_in_summary:
            agree_col_name = 'agree_' + p_col
            if agree_col_name not in summary_df.columns:
                summary_df[agree_col_name] = np.nan

        if not agree_s_vars and not agree_p_vars:
            if not self.c_vars_to_compute:
                print("คำเตือน: ข้ามการคำนวณ T2B เนื่องจากไม่ได้เริ่มจากไฟล์ SPSS ดั้งเดิม")
            return summary_df

        if self.df is None:
            raise ValueError("ไม่พบข้อมูล SPSS ดั้งเดิม (self.df) สำหรับคำนวณ T2B")

        if self.transformed_df is None:
            raise ValueError("ไม่พบข้อมูล SPSS ที่ผ่านการประมวลผล (self.transformed_df) สำหรับคำนวณ T2B")

        if not self.id_vars:
            raise ValueError("Identifier variables (id_vars) not found.")

        if not self.id_vars or not all(col in self.df.columns for col in self.id_vars):
            raise ValueError("One or more ID columns not found in original SPSS data.")

        t2b_choice = self.t2b_choice_var.get()
        good_codes = [5, 4] if t2b_choice == "5+4" else [1, 2]

        if cross_filter is None:
            cross_filter = self.filter_entry.get().strip()

        groups_to_summarize = OrderedDict()
        groups_to_summarize['Overall'] = self.transformed_df

        primary_values = []
        if primary_filter and primary_filter in self.transformed_df.columns:
            primary_values = sorted(self.transformed_df[primary_filter].dropna().unique())
            for p_val in primary_values:
                filter_name = f"{primary_filter}={p_val}"
                if filter_name not in groups_to_summarize:
                    groups_to_summarize[filter_name] = self.transformed_df[self.transformed_df[primary_filter] == p_val]

        if cross_filter and cross_filter in self.transformed_df.columns:
            cross_values = sorted(self.transformed_df[cross_filter].dropna().unique())
            for c_val in cross_values:
                filter_name_cross = f"{cross_filter}={c_val}"
                if filter_name_cross not in groups_to_summarize:
                    groups_to_summarize[filter_name_cross] = self.transformed_df[self.transformed_df[cross_filter] == c_val]

                if primary_filter and primary_filter in self.transformed_df.columns:
                    for p_val in primary_values:
                        nested_name = f"{primary_filter}={p_val}+{cross_filter}={c_val}"
                        subset = self.transformed_df[
                            (self.transformed_df[primary_filter] == p_val) &
                            (self.transformed_df[cross_filter] == c_val)
                        ]
                        groups_to_summarize[nested_name] = subset

        for name, df_group in groups_to_summarize.items():
            if df_group.empty:
                continue
            row_mask = summary_df['Filter'] == name
            if not row_mask.any():
                continue

            base_ids = df_group[self.id_vars].drop_duplicates()
            base_original_df = pd.merge(self.df, base_ids, on=self.id_vars, how='inner')
            total_base = len(base_original_df)
            if total_base == 0:
                continue

            for i, s_col in enumerate(s_cols_in_summary):
                agree_col_name = 'agree_' + s_col
                if i < len(agree_s_vars):
                    source_var = agree_s_vars[i]
                    if source_var in base_original_df.columns:
                        t2b_sum = base_original_df[source_var].isin(good_codes).sum()
                        t2b_value = (t2b_sum / total_base) * 100 if total_base > 0 else 0
                        summary_df.loc[row_mask, agree_col_name] = t2b_value

            for i, p_col in enumerate(p_cols_in_summary):
                agree_col_name = 'agree_' + p_col
                if i < len(agree_p_vars):
                    source_var = agree_p_vars[i]
                    if source_var in base_original_df.columns:
                        t2b_sum = base_original_df[source_var].isin(good_codes).sum()
                        t2b_value = (t2b_sum / total_base) * 100 if total_base > 0 else 0
                        summary_df.loc[row_mask, agree_col_name] = t2b_value

        return summary_df

    def _run_factor_regression_logic(self, primary_filter, cross_filter):
        """ตรรกะการรัน Factor และ Regression"""
        df_for_analysis = self.transformed_df
        all_cols = list(df_for_analysis.columns)

        if primary_filter and primary_filter not in all_cols: primary_filter = ""
        if cross_filter and cross_filter not in all_cols: cross_filter = ""
        if primary_filter and primary_filter == cross_filter: messagebox.showwarning("คำเตือน", "Filter หลัก และ Filter ไขว้ ต้องเป็นคนละคอลัมน์"); return None, None

        results_for_saving = OrderedDict()
        old_stdout = sys.stdout; sys.stdout = captured_output = io.StringIO()
        try:
            groups_to_analyze = OrderedDict()
            groups_to_analyze['Overall'] = df_for_analysis

            primary_values = []
            if primary_filter:
                primary_values = sorted(df_for_analysis[primary_filter].dropna().unique())
                for p_val in primary_values:
                    filter_name = f"{primary_filter}={p_val}"
                    groups_to_analyze[filter_name] = df_for_analysis[df_for_analysis[primary_filter] == p_val]

            if cross_filter:
                cross_values = sorted(df_for_analysis[cross_filter].dropna().unique())
                for c_val in cross_values:
                    filter_name_cross = f"{cross_filter}={c_val}"
                    if filter_name_cross not in groups_to_analyze:
                        groups_to_analyze[filter_name_cross] = df_for_analysis[df_for_analysis[cross_filter] == c_val]

                    if primary_filter:
                        for p_val in primary_values:
                            nested_name = f"{primary_filter}={p_val}+{cross_filter}={c_val}"
                            subset = df_for_analysis[(df_for_analysis[primary_filter] == p_val) & (df_for_analysis[cross_filter] == c_val)]
                            groups_to_analyze[nested_name] = subset

            for name, df_group in groups_to_analyze.items():
                sys.stdout.write(f"\n{'='*80}\n--- ผลการวิเคราะห์สำหรับ: {name} ---\n{'='*80}\n")
                if df_group.empty:
                    print("ไม่มีข้อมูลสำหรับกลุ่มนี้")
                    continue
                if results := self._run_single_analysis(df_group.copy()):
                    results_for_saving[name] = results

            full_output_text = captured_output.getvalue()
            sys.stdout = old_stdout
            captured_output.close()
            return results_for_saving, full_output_text
        except Exception as e:
            sys.stdout = old_stdout; captured_output.close()
            messagebox.showerror("ผิดพลาด", f"เกิดข้อผิดพลาดระหว่างการวิเคราะห์ Factor/Regression:\n{e}")
            return None, None

    def _run_single_analysis(self, target_df):
        """รันการวิเคราะห์ 1 ชุด (Factor -> Regression)"""
        try:
            factor_scores_df, sorted_loadings_df, factor_to_variable_map = self.perform_factor_analysis(target_df)
            if factor_scores_df is not None:
                analysis_df = target_df.join(factor_scores_df)
                beta_df, beta_sorted_df, _ = self.perform_regression_analysis(analysis_df, factor_to_variable_map)
                return {'loadings': sorted_loadings_df, 'beta': beta_df, 'beta_sorted': beta_sorted_df}
        except Exception as e:
            print(f"\n!!! เกิดข้อผิดพลาดในการวิเคราะห์กลุ่มนี้: {e}\n!!! ข้ามการวิเคราะห์กลุ่มนี้...\n")
        return {}

    def save_settings(self):
        """บันทึกการตั้งค่าทั้งหมดลงใน Excel สองชีทโดยอัตโนมัติ"""
        if not self.original_filepath:
            messagebox.showerror("ผิดพลาด", "ไม่สามารถบันทึกการตั้งค่าได้ เนื่องจากยังไม่ได้โหลดไฟล์ SPSS ต้นฉบับ")
            return
        if not self.c_vars_to_compute and not any(self.vars_to_transform.values()):
            messagebox.showerror("ผิดพลาด", "ยังไม่มีการตั้งค่าตัวแปรให้บันทึก")
            return

        try:
            directory = os.path.dirname(self.original_filepath)
            filepath = os.path.join(directory, "Setting BS.xlsx")

            # --- Part 1: Settings Sheet ---
            settings_lists = {
                'C': self.c_vars_to_compute,
                'A': self.vars_to_transform.get('A', []),
                'S': self.vars_to_transform.get('S', []),
                'P': self.vars_to_transform.get('P', []),
                'E': self.vars_to_transform.get('E', []),
                'AgreeS': self.vars_to_transform.get('AgreeS', []),
                'AgreeP': self.vars_to_transform.get('AgreeP', [])
            }
            settings_df = pd.DataFrame({k: pd.Series(v) for k, v in settings_lists.items()})

            settings_df.insert(0, 'Filter_Var', self.filter_entry.get().strip())
            settings_df.insert(0, 'T2B_Choice', self.t2b_choice_var.get())
            settings_df.insert(0, 'PathFile', self.original_filepath)
            settings_df.loc[1:, ['PathFile', 'T2B_Choice', 'Filter_Var']] = ''

            # --- Part 2: Label Sheet ---
            index1_label_data = list(self.index1_labels.items())
            filter_label_data = list(self.filter_labels.get('labels', {}).items())

            label_dict = {
                'Index1_Code': [item[0] for item in index1_label_data],
                'Index1_Label': [item[1] for item in index1_label_data],
                'Filter_Code': [item[0] for item in filter_label_data],
                'Filter_Label': [item[1] for item in filter_label_data]
            }
            labels_df = pd.DataFrame({k: pd.Series(v) for k, v in label_dict.items()})

            # --- Write to Excel ---
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                settings_df.to_excel(writer, sheet_name='Settings', index=False)
                if not labels_df.empty or not all(labels_df[col].isnull().all() for col in labels_df.columns):
                    labels_df.to_excel(writer, sheet_name='Label', index=False)

            self.update_status(f"บันทึกการตั้งค่าสำเร็จที่: {filepath}", "success")
            messagebox.showinfo("สำเร็จ", f"บันทึกการตั้งค่าทั้งหมดเรียบร้อยแล้ว\nที่: {filepath}")
        except Exception as e:
            self.update_status("บันทึกการตั้งค่าผิดพลาด", "danger")
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถบันทึกไฟล์การตั้งค่าได้: {e}")

    def save_all_results_to_excel(self, summary_df, results_dict, full_output_text):
        """บันทึกข้อมูลสรุปและผลวิเคราะห์ทั้งหมดลงในไฟล์ Excel ไฟล์เดียวโดยอัตโนมัติ"""
        if not self.original_filepath:
            messagebox.showerror("ผิดพลาด", "ไม่สามารถบันทึกผลลัพธ์ได้ เนื่องจากไม่พบ Path ของไฟล์ต้นฉบับ")
            return

        try:
            base, _ = os.path.splitext(self.original_filepath)
            filepath = f"{base} BS Output.xlsx"
            self.last_excel_filepath = filepath
            rawdata_df = None
            if self.df is not None:
                rawdata_df = self.df
            elif self.transformed_df is not None:
                rawdata_df = self.transformed_df
            elif self.original_filepath.lower().endswith(".sav"):
                try:
                    rawdata_df, _ = pyreadstat.read_sav(self.original_filepath)
                    if self.spss_original_order:
                        rawdata_df = rawdata_df[self.spss_original_order]
                except Exception as e:
                    messagebox.showwarning("Rawdata Warning", f"ไม่สามารถโหลด Rawdata จากไฟล์ SPSS ได้: {e}")

            expected_factors = ['N_S', 'N_P', 'N_C', 'N_E']
            template_rows = []

            for filter_name in summary_df['Filter']:
                row_data = {'Filter': filter_name}
                analysis_result = results_dict.get(filter_name)

                if analysis_result and analysis_result.get('beta_sorted') is not None:
                    betas = analysis_result['beta_sorted']['Beta'].to_dict()
                    for factor in expected_factors:
                        row_data[factor] = betas.get(factor, 0)
                else:
                    for factor in expected_factors:
                        row_data[factor] = 0

                template_rows.append(row_data)

            template_df = pd.DataFrame(template_rows)

            if not template_df.empty:
                for factor in expected_factors:
                    if factor not in template_df.columns:
                        template_df[factor] = 0

                template_df['Total'] = template_df[expected_factors].sum(axis=1)

                beta_ratio_cols_names = {'N_S': 'B.S', 'N_P': 'B.P', 'N_C': 'B.C', 'N_E': 'B.E'}
                for factor, ratio_name in beta_ratio_cols_names.items():
                    template_df[ratio_name] = np.where(
                        template_df['Total'] != 0,
                        (template_df[factor] / template_df['Total']) * 100,
                        0
                    )

            beta_cols_to_add = ['B.S', 'B.P', 'B.C', 'B.E']
            if 'Filter' in template_df.columns:
                cols_to_drop = [col for col in beta_cols_to_add if col in summary_df.columns]
                if cols_to_drop:
                    summary_df = summary_df.drop(columns=cols_to_drop)

                summary_df = pd.merge(summary_df, template_df[['Filter'] + beta_cols_to_add], on='Filter', how='left')

            excel_df = self._prepare_final_excel_df(summary_df)

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                excel_df.to_excel(writer, sheet_name='Summary', index=False)

                workbook = writer.book
                worksheet = writer.sheets['Summary']

                headers = [cell.value for cell in worksheet[1]]
                for col_idx, header in enumerate(headers, 1):
                    if header is None: continue

                    format_str = None
                    if header in ['S', 'P', 'A level', 'A score', 'C', 'E', 'B.S', 'B.P', 'B.C', 'B.E'] or \
                        (header.startswith(('S_', 'P_', 'E_')) and 'cor' not in header):
                        format_str = '0.00'
                    elif header.startswith('C_'):
                        format_str = '0'
                    elif header.startswith(('cor_', 'CorE_')):
                        format_str = '0.000'
                    elif header.startswith('agree_'):
                        format_str = '0.0'

                    if format_str:
                        for row in range(2, worksheet.max_row + 1):
                            worksheet.cell(row=row, column=col_idx).number_format = format_str

                self.transformed_df.to_excel(writer, sheet_name='Sheet Dummy', index=False)
                if rawdata_df is not None:
                    rawdata_df.to_excel(writer, sheet_name='Rawdata', index=False)
                if not template_df.empty:
                    final_template_cols = ['Filter', 'N_S', 'N_P', 'N_C', 'N_E', 'Total', 'B.S', 'B.P', 'B.C', 'B.E']
                    template_df = template_df.reindex(columns=[col for col in final_template_cols if col in template_df.columns])
                    template_df.to_excel(writer, sheet_name='Factor_Template', index=False)

                output_lines = full_output_text.splitlines()
                safe_lines = ["'" + line if line.strip().startswith(('=', '-', '+', '@')) else line for line in output_lines]
                output_df = pd.DataFrame(safe_lines, columns=["Analysis Log"])
                output_df.to_excel(writer, sheet_name="Factor_Output", index=False)

            if self.save_all_sheets_var.get():
                self.update_status("กำลังลบชีทที่ไม่จำเป็น...", "info")
                workbook = openpyxl.load_workbook(filepath)
                sheets_to_delete = [s for s in workbook.sheetnames if s not in ['Summary', 'Rawdata']]
                for sheet_name in sheets_to_delete:
                    workbook.remove(workbook[sheet_name])
                workbook.save(filepath)
                final_message = f"บันทึก Excel (Summary + Rawdata) เรียบร้อยแล้วที่:\n{filepath}"
            else:
                final_message = f"บันทึก Excel (Full Report) เรียบร้อยแล้วที่:\n{filepath}"

            self.update_status("บันทึก Excel สำเร็จ", "success")
            messagebox.showinfo("สำเร็จ", final_message)

        except Exception as e:
            self.update_status("บันทึก Excel ผิดพลาด", "danger")
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถบันทึกไฟล์ Excel ได้: {e}")

    def _prepare_final_excel_df(self, final_summary_df):
        """จัดเรียงคอลัมน์และเตรียม DataFrame สำหรับเขียนลง Excel"""
        if 'A' in final_summary_df.columns: final_summary_df.rename(columns={'A': 'A level'}, inplace=True)
        if 'ZA' in final_summary_df.columns: final_summary_df['A score'] = final_summary_df['ZA'] * 100
        else: final_summary_df['A score'] = np.nan

        s_cols = sorted([c for c in final_summary_df.columns if c.startswith('S_') and 'cor' not in c and 'agree' not in c], key=lambda x: int(x.split('_')[1]))
        p_cols = sorted([c for c in final_summary_df.columns if c.startswith('P_') and 'cor' not in c and 'agree' not in c], key=lambda x: int(x.split('_')[1]))
        c_cols = sorted([c for c in final_summary_df.columns if c.startswith('C_') and 'cor' not in c], key=lambda x: int(x.split('_')[1]))
        e_cols = sorted([c for c in final_summary_df.columns if c.startswith('E_') and 'cor' not in c], key=lambda x: int(x.split('_')[1]))

        if s_cols: final_summary_df['S'] = final_summary_df[s_cols].mean(axis=1)
        if p_cols: final_summary_df['P'] = final_summary_df[p_cols].mean(axis=1)
        if c_cols: final_summary_df['C'] = final_summary_df[c_cols].mean(axis=1)
        if e_cols: final_summary_df['E'] = final_summary_df[e_cols].mean(axis=1)

        main_order = ['Code Index1', 'Labe Index1', 'Filter', 'S', 'P', 'A level', 'A score', 'C', 'E', 'B.S', 'B.P', 'B.C', 'B.E']
        final_summary_df.rename(columns={'Index1':'Code Index1'}, inplace=True)

        core_cols = sorted([c for c in final_summary_df.columns if c.startswith('CorE_')])
        cor_s_cols = sorted([c for c in final_summary_df.columns if c.startswith('cor_S_')], key=lambda x: int(x.split('_')[-1]))
        cor_p_cols = sorted([c for c in final_summary_df.columns if c.startswith('cor_P_')], key=lambda x: int(x.split('_')[-1]))
        agree_s_names = sorted([c for c in final_summary_df.columns if c.startswith('agree_S_')], key=lambda x: int(x.split('_')[-1]))
        agree_p_names = sorted([c for c in final_summary_df.columns if c.startswith('agree_P_')], key=lambda x: int(x.split('_')[-1]))

        final_column_order = (
            main_order +
            s_cols + p_cols + c_cols + e_cols + core_cols +
            cor_s_cols + cor_p_cols +
            agree_s_names + agree_p_names
        )

        final_column_order_existing = [col for col in final_column_order if col in final_summary_df.columns]

        excel_df = final_summary_df[final_column_order_existing]

        return excel_df

    # ===================================================================
    # CORE ANALYSIS LOGIC (UNCHANGED)
    # ===================================================================
    def perform_factor_analysis(self, target_df):
        print("ส่วนที่ 1: การวิเคราะห์องค์ประกอบ (Factor Analysis)\n" + "-"*50 + "\n")
        factor_vars = ['N_S', 'N_P', 'N_C', 'N_E']
        if not all(col in target_df.columns for col in factor_vars): raise KeyError(f"ไม่พบคอลัมน์สำหรับ Factor Analysis: {', '.join(factor_vars)}")
        df_factor = target_df[factor_vars].dropna().copy()
        if len(df_factor) < len(factor_vars): raise ValueError("ข้อมูลไม่เพียงพอสำหรับ Factor Analysis หลังจากการลบค่าว่าง")
        print(f"ข้อมูลที่ใช้ในการวิเคราะห์องค์ประกอบ: {len(df_factor)} แถว\n")
        fa_rotated = FactorAnalyzer(n_factors=4, rotation='equamax', method='principal', rotation_kwargs={'kappa': 0.5, 'max_iter': 250}); fa_rotated.fit(df_factor)
        original_loadings = fa_rotated.loadings_
        ss_loadings = np.sum(original_loadings**2, axis=0)
        spss_col_order = np.argsort(ss_loadings)[::-1]
        L = original_loadings[:, spss_col_order]
        print("Rotation: Rotated Component Matrix (Equamax - SPSS Compatible):")
        loadings_rotated_df = pd.DataFrame(L, index=df_factor.columns, columns=[f'Factor{i+1}' for i in range(4)])
        abs_loadings = loadings_rotated_df.abs(); primary_factor_map = abs_loadings.idxmax(axis=1)
        factor_to_variable_map = {v: k for k, v in primary_factor_map.items()}
        sort_list = sorted([(int(primary_factor_map[var].replace('Factor', '')), -abs_loadings.loc[var].max(), var) for var in abs_loadings.index])
        sorted_loadings_df = loadings_rotated_df.loc[[var for _, _, var in sort_list]]
        print(sorted_loadings_df.applymap(lambda x: f"{x:.3f}" if abs(x) >= 0.4 else "")); print("\n" + "-"*50 + "\n")
        print("คำนวณ Factor Scores ด้วยวิธี Anderson-Rubin (PCA)...\n")
        Z = StandardScaler().fit_transform(df_factor); R = df_factor.corr().values; inv_R = inv(R)
        temp_matrix = L.T @ inv_R @ L; eigvals, eigvecs = eigh(temp_matrix)
        inv_sqrt_eigvals_arr = np.zeros_like(eigvals); positive_eigvals_mask = eigvals > 1e-12
        inv_sqrt_eigvals_arr[positive_eigvals_mask] = 1.0 / np.sqrt(eigvals[positive_eigvals_mask])
        inv_sqrt_temp = eigvecs @ np.diag(inv_sqrt_eigvals_arr) @ eigvecs.T
        C_AR = inv_R @ L @ inv_sqrt_temp; factor_scores = Z @ C_AR
        df_scores = pd.DataFrame(factor_scores, columns=[f'FAC{i+1}_1' for i in range(factor_scores.shape[1])], index=df_factor.index)
        return df_scores, sorted_loadings_df, factor_to_variable_map

    def perform_regression_analysis(self, target_df, factor_to_variable_map):
        print("\nส่วนที่ 2: การวิเคราะห์การถดถอย (Regression Analysis)\n" + "-"*50 + "\n")
        dependent_var = 'ZA'; independent_vars = ['FAC1_1', 'FAC2_1', 'FAC3_1', 'FAC4_1']
        required_cols = [dependent_var] + independent_vars
        if not all(col in target_df.columns for col in required_cols): raise KeyError(f"ไม่พบคอลัมน์สำหรับ Regression: {', '.join(required_cols)}")
        df_regression = target_df[required_cols].dropna().copy()
        if len(df_regression) < len(independent_vars) + 2: raise ValueError("ข้อมูลไม่เพียงพอสำหรับ Regression Analysis")
        print(f"ข้อมูลที่ใช้ในการวิเคราะห์ Regression: {len(df_regression)} แถว\n")
        Y = df_regression[dependent_var]; X_original = df_regression[independent_vars]; X = sm.add_constant(X_original)
        model = sm.OLS(Y, X).fit()
        print("Regression Model Summary:"); print(model.summary()); print("\n" + "-"*50 + "\n")
        print("Standardized Coefficients (Beta):")
        unstandardized_coeffs = model.params.drop('const')
        betas = unstandardized_coeffs * (X_original.std() / Y.std())
        beta_df = pd.DataFrame({'Beta': betas}); print(beta_df); print("\n" + "-"*50 + "\n")
        print("Standardized Coefficients (Beta) - Sort:")
        score_to_factor_map = {f'FAC{i+1}_1': f'Factor{i+1}' for i in range(4)}
        renamed_betas = betas.rename(index=lambda score_name: factor_to_variable_map.get(score_to_factor_map.get(score_name)))
        valid_order = [v for v in ['N_S', 'N_P', 'N_C', 'N_E'] if v in renamed_betas.index]
        beta_sorted_df = pd.DataFrame({'Beta': renamed_betas}).loc[valid_order]; print(beta_sorted_df); print("\n" + "-"*50 + "\n")
        zpred = model.predict(X)
        return beta_df, beta_sorted_df, zpred




# <<< START OF CHANGES >>>
# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
def run_this_app(working_dir=None): # ชื่อฟังก์ชันนี้จะถูกใช้ใน Launcher
    """
    ฟังก์ชันหลักสำหรับสร้างและรัน QuotaSamplerApp.
    """
    print(f"--- QUOTA_SAMPLER_INFO: Starting 'QuotaSamplerApp' via run_this_app() ---")
    try:
    # --- ส่วนที่ใช้รันโปรแกรม ---
    #if __name__ == "__main__":
        app = SpssProcessorApp()
        app.mainloop()

        
        print(f"--- QUOTA_SAMPLER_INFO: QuotaSamplerApp mainloop finished. ---")

    except Exception as e:
        # ดักจับ Error ที่อาจเกิดขึ้นระหว่างการสร้างหรือรัน App
        print(f"QUOTA_SAMPLER_ERROR: An error occurred during QuotaSamplerApp execution: {e}")
        # แสดง Popup ถ้ามีปัญหา
        if 'root' not in locals() or not root.winfo_exists(): # สร้าง root ชั่วคราวถ้ายังไม่มี
            root_temp = tk.Tk()
            root_temp.withdraw()
            messagebox.showerror("Application Error (Quota Sampler)",
                                f"An unexpected error occurred:\n{e}", parent=root_temp)
            root_temp.destroy()
        else:
            messagebox.showerror("Application Error (Quota Sampler)",
                                f"An unexpected error occurred:\n{e}", parent=root) # ใช้ root ที่มีอยู่ถ้าเป็นไปได้
        sys.exit(f"Error running QuotaSamplerApp: {e}") # อาจจะ exit หรือไม่ก็ได้ ขึ้นกับการออกแบบ


# --- ส่วน Run Application เมื่อรันไฟล์นี้โดยตรง (สำหรับ Test) ---
if __name__ == "__main__":
    print("--- Running QuotaSamplerApp.py directly for testing ---")
    # (ถ้ามีการตั้งค่า DPI ด้านบน มันจะทำงานอัตโนมัติ)

    # เรียกฟังก์ชัน Entry Point ที่เราสร้างขึ้น
    run_this_app()

    print("--- Finished direct execution of QuotaSamplerApp.py ---")
# <<< END OF CHANGES >>>
