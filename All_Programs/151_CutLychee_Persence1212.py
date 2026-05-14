from __future__ import annotations

import multiprocessing
import os
import re
import shutil
import sys
import time
import traceback
import zipfile
from concurrent.futures import ProcessPoolExecutor, as_completed
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Tuple
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.cell import range_boundaries

try:
    import win32com.client as win32

    HAS_WIN32COM = True
except Exception:
    win32 = None
    HAS_WIN32COM = False


# ----------------------------
# Excel processing core
# ----------------------------
def build_output_stem(source_stem: str) -> str:
    today = datetime.now().strftime("%Y%m%d")
    stem = source_stem

    stem = re.sub(r"_processed(?:_\d+)?$", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"(?i)N%", "%", stem)

    if re.search(r"\d{8}$", stem):
        stem = re.sub(r"\d{8}$", today, stem)
    else:
        stem = f"{stem} {today}"

    return stem


def unique_output_path_reserved(
    out_dir: Path, source_path: Path, reserved_names: set[str]
) -> Path:
    base = build_output_stem(source_path.stem)
    candidate_name = f"{base}.xlsx"
    idx = 1
    while True:
        candidate = out_dir / candidate_name
        key = str(candidate.resolve()).lower()
        if key not in reserved_names and not candidate.exists():
            reserved_names.add(key)
            return candidate
        candidate_name = f"{base}_{idx}.xlsx"
        idx += 1


COM_VBA_CODE = r'''
Option Explicit

Public Sub CodexProcessWorkbook()
    Dim ws As Worksheet

    Application.ScreenUpdating = False
    Application.EnableEvents = False
    Application.DisplayAlerts = False
    On Error Resume Next
    Application.Calculation = xlCalculationManual
    On Error GoTo CleanFail

    For Each ws In ThisWorkbook.Worksheets
        ProcessWorksheet ws
    Next ws

CleanExit:
    On Error Resume Next
    Application.Calculation = xlCalculationAutomatic
    Application.EnableEvents = True
    Application.ScreenUpdating = True
    Exit Sub

CleanFail:
    Resume CleanExit
End Sub

Private Sub ProcessWorksheet(ByVal ws As Worksheet)
    Dim lastRow As Long
    Dim maxCol As Long

    If LCase$(Trim$(ws.Name)) = "contents" Then Exit Sub

    lastRow = WorksheetMaxRow(ws)
    If lastRow < 7 Then Exit Sub

    maxCol = DetectEffectiveMaxCol(ws, lastRow)
    If Not ws.Range("A6").MergeCells Then
        ws.Range("A6").Value = "1st row: Column %"
    End If

    UnmergeDataRegion ws, lastRow, maxCol
    MergeDuplicateRows ws, maxCol

    lastRow = FindLastDataRow(ws, maxCol)
    ApplyLastRowBottomBorder ws, lastRow, maxCol
    NormaliseRowHeights ws, maxCol, lastRow
End Sub

Private Function WorksheetMaxRow(ByVal ws As Worksheet) As Long
    Dim area As Range
    Set area = ws.UsedRange
    WorksheetMaxRow = area.Row + area.Rows.Count - 1
End Function

Private Function WorksheetMaxCol(ByVal ws As Worksheet) As Long
    Dim area As Range
    Set area = ws.UsedRange
    WorksheetMaxCol = area.Column + area.Columns.Count - 1
End Function

Private Function CellText(ByVal value As Variant) As String
    If IsError(value) Then
        CellText = ""
    ElseIf IsNull(value) Then
        CellText = ""
    Else
        CellText = Trim$(CStr(value & ""))
    End If
End Function

Private Function DetectEffectiveMaxCol(ByVal ws As Worksheet, ByVal lastRow As Long) As Long
    Dim dimMaxCol As Long
    Dim sampleRows As Object
    Dim key As Variant
    Dim r As Long
    Dim col As Long

    dimMaxCol = WorksheetMaxCol(ws)
    If dimMaxCol < 3 Then
        DetectEffectiveMaxCol = 3
        Exit Function
    End If

    Set sampleRows = CreateObject("Scripting.Dictionary")
    For r = 1 To WorksheetFunction.Min(lastRow, 30)
        sampleRows(CStr(r)) = True
    Next r

    For Each key In Array(4, 5, 6, 7, 8, 9, 10, lastRow)
        r = CLng(key)
        If r >= 1 And r <= lastRow Then
            sampleRows(CStr(r)) = True
        End If
    Next key

    If lastRow > 30 Then
        For r = WorksheetFunction.Max(1, lastRow - 9) To lastRow
            sampleRows(CStr(r)) = True
        Next r
    End If

    For col = dimMaxCol To 3 Step -1
        For Each key In sampleRows.Keys
            r = CLng(key)
            If CellText(ws.Cells(r, col).Value2) <> "" Then
                DetectEffectiveMaxCol = col
                Exit Function
            End If
        Next key
    Next col

    DetectEffectiveMaxCol = dimMaxCol
End Function

Private Sub UnmergeDataRegion(ByVal ws As Worksheet, ByVal lastRow As Long, ByVal maxCol As Long)
    Dim mergedAreas As Object
    Dim scanRange As Range
    Dim cell As Range
    Dim area As Range
    Dim sourceCell As Range
    Dim targetCell As Range
    Dim key As Variant
    Dim formulaText As String
    Dim topValue As Variant
    Dim sourceRow As Long
    Dim sourceCol As Long

    Set mergedAreas = CreateObject("Scripting.Dictionary")
    Set scanRange = ws.Range(ws.Cells(7, 1), ws.Cells(lastRow, maxCol))

    For Each cell In scanRange.Cells
        If cell.MergeCells Then
            Set area = cell.MergeArea
            If area.Row >= 7 Then
                If Not mergedAreas.Exists(area.Address(False, False)) Then
                    mergedAreas.Add area.Address(False, False), True
                End If
            End If
        End If
    Next cell

    For Each key In mergedAreas.Keys
        Set area = ws.Range(CStr(key))
        formulaText = area.Cells(1, 1).Formula
        topValue = area.Cells(1, 1).Value2
        sourceRow = area.Row
        sourceCol = area.Column
        area.UnMerge
        Set area = ws.Range(CStr(key))
        Set sourceCell = ws.Cells(sourceRow, sourceCol)
        For Each targetCell In area.Cells
            CopyCellStyle sourceCell, targetCell
            If Len(formulaText) > 0 And Left$(formulaText, 1) = "=" Then
                targetCell.Formula = formulaText
            Else
                targetCell.Value2 = topValue
            End If
        Next targetCell
    Next key
End Sub

Private Sub MergeDuplicateRows(ByVal ws As Worksheet, ByVal maxCol As Long)
    Dim lastRow As Long
    Dim colB As Variant
    Dim rowsToDelete() As Long
    Dim deleteCount As Long
    Dim i As Long
    Dim curText As String
    Dim aboveText As String
    Dim col As Long

    lastRow = WorksheetMaxRow(ws)
    If lastRow < 8 Then Exit Sub

    colB = ws.Range(ws.Cells(7, 2), ws.Cells(lastRow, 2)).Value2
    ReDim rowsToDelete(1 To lastRow - 6)

    For i = UBound(colB, 1) To 2 Step -1
        curText = CellText(colB(i, 1))
        aboveText = CellText(colB(i - 1, 1))

        If curText <> "" And UCase$(curText) <> "TOTAL" And curText = aboveText Then
            For col = 1 To maxCol
                CopyBottomBorder ws.Cells(i + 6 - 1, col), ws.Cells(i + 6, col)
            Next col
            For col = 3 To maxCol
                ChooseValueAndFormat ws.Cells(i + 6 - 1, col), ws.Cells(i + 6, col)
            Next col
            deleteCount = deleteCount + 1
            rowsToDelete(deleteCount) = i + 6
        ElseIf UCase$(aboveText) = "TOTAL" And (UCase$(curText) = "TOTAL" Or curText = "") Then
            For col = 1 To maxCol
                CopyBottomBorder ws.Cells(i + 6 - 1, col), ws.Cells(i + 6, col)
            Next col
            deleteCount = deleteCount + 1
            rowsToDelete(deleteCount) = i + 6
        End If
    Next i

    DeleteRowsDesc ws, rowsToDelete, deleteCount
End Sub

Private Sub CopyBottomBorder(ByVal targetCell As Range, ByVal sourceCell As Range)
    CopyBorderSide targetCell, sourceCell, xlEdgeBottom
End Sub

Private Sub CopyCellStyle(ByVal sourceCell As Range, ByVal targetCell As Range)
    On Error Resume Next
    targetCell.NumberFormat = sourceCell.NumberFormat
    targetCell.HorizontalAlignment = sourceCell.HorizontalAlignment
    targetCell.VerticalAlignment = sourceCell.VerticalAlignment
    targetCell.WrapText = sourceCell.WrapText
    targetCell.Orientation = sourceCell.Orientation
    targetCell.AddIndent = sourceCell.AddIndent
    targetCell.IndentLevel = sourceCell.IndentLevel
    targetCell.ShrinkToFit = sourceCell.ShrinkToFit
    targetCell.ReadingOrder = sourceCell.ReadingOrder

    With targetCell.Font
        .Name = sourceCell.Font.Name
        .Size = sourceCell.Font.Size
        .Bold = sourceCell.Font.Bold
        .Italic = sourceCell.Font.Italic
        .Underline = sourceCell.Font.Underline
        .Strikethrough = sourceCell.Font.Strikethrough
        .Color = sourceCell.Font.Color
    End With

    With targetCell.Interior
        .Pattern = sourceCell.Interior.Pattern
        .Color = sourceCell.Interior.Color
        .TintAndShade = sourceCell.Interior.TintAndShade
        .PatternTintAndShade = sourceCell.Interior.PatternTintAndShade
    End With

    CopyBorderSide targetCell, sourceCell, xlEdgeLeft
    CopyBorderSide targetCell, sourceCell, xlEdgeRight
    CopyBorderSide targetCell, sourceCell, xlEdgeTop
    CopyBorderSide targetCell, sourceCell, xlEdgeBottom
    On Error GoTo 0
End Sub

Private Sub CopyBorderSide(ByVal targetCell As Range, ByVal sourceCell As Range, ByVal edgeId As Long)
    On Error Resume Next
    With targetCell.Borders(edgeId)
        .LineStyle = sourceCell.Borders(edgeId).LineStyle
        .Weight = sourceCell.Borders(edgeId).Weight
        .Color = sourceCell.Borders(edgeId).Color
        .TintAndShade = sourceCell.Borders(edgeId).TintAndShade
    End With
    On Error GoTo 0
End Sub

Private Function IsBlankValue(ByVal value As Variant) As Boolean
    If IsError(value) Then
        IsBlankValue = False
    ElseIf IsNull(value) Then
        IsBlankValue = True
    ElseIf IsEmpty(value) Then
        IsBlankValue = True
    ElseIf VarType(value) = vbString Then
        IsBlankValue = (Trim$(CStr(value)) = "")
    Else
        IsBlankValue = False
    End If
End Function

Private Function HasFraction(ByVal value As Double) As Boolean
    HasFraction = (Abs(value - Fix(value)) > 0.000000000001#)
End Function

Private Sub ChooseValueAndFormat(ByVal upperCell As Range, ByVal lowerCell As Range)
    Dim upperVal As Variant
    Dim lowerVal As Variant

    upperVal = upperCell.Value2
    lowerVal = lowerCell.Value2

    If IsBlankValue(lowerVal) Then Exit Sub

    If IsNumeric(upperVal) And IsNumeric(lowerVal) Then
        If HasFraction(CDbl(upperVal)) And Not HasFraction(CDbl(lowerVal)) Then
            Exit Sub
        End If
    End If

    If Len(lowerCell.Formula) > 0 And Left$(lowerCell.Formula, 1) = "=" Then
        upperCell.Formula = lowerCell.Formula
    Else
        upperCell.Value2 = lowerVal
    End If
    upperCell.NumberFormat = lowerCell.NumberFormat
End Sub

Private Sub DeleteRowsDesc(ByVal ws As Worksheet, ByRef rowIndexes() As Long, ByVal itemCount As Long)
    Dim startRow As Long
    Dim countRows As Long
    Dim prevRow As Long
    Dim idx As Long

    If itemCount <= 0 Then Exit Sub

    startRow = rowIndexes(1)
    countRows = 1
    prevRow = rowIndexes(1)

    For idx = 2 To itemCount
        If rowIndexes(idx) = prevRow - 1 Then
            countRows = countRows + 1
        Else
            ws.Rows(CStr(startRow - countRows + 1) & ":" & CStr(startRow)).Delete
            startRow = rowIndexes(idx)
            countRows = 1
        End If
        prevRow = rowIndexes(idx)
    Next idx

    ws.Rows(CStr(startRow - countRows + 1) & ":" & CStr(startRow)).Delete
End Sub

Private Function FindLastDataRow(ByVal ws As Worksheet, ByVal maxCol As Long) As Long
    Dim lastRow As Long
    Dim r As Long
    Dim c As Long

    lastRow = WorksheetMaxRow(ws)
    For r = lastRow To 7 Step -1
        For c = 1 To maxCol
            If CellText(ws.Cells(r, c).Value2) <> "" Then
                FindLastDataRow = r
                Exit Function
            End If
        Next c
    Next r

    FindLastDataRow = 7
End Function

Private Sub ApplyLastRowBottomBorder(ByVal ws As Worksheet, ByVal lastRow As Long, ByVal maxCol As Long)
    Dim c As Long
    For c = 1 To maxCol
        With ws.Cells(lastRow, c).Borders(xlEdgeBottom)
            .LineStyle = xlContinuous
            .Weight = xlThin
            .Color = RGB(0, 0, 0)
        End With
    Next c
End Sub

Private Function EstimateLines(ByVal text As String, ByVal colWidthChars As Double) As Long
    Dim parts() As String
    Dim part As Variant
    Dim total As Long
    Dim widthChars As Long
    Dim partText As String

    If Len(text) = 0 Then
        EstimateLines = 1
        Exit Function
    End If

    text = Replace$(text, vbCrLf, vbLf)
    text = Replace$(text, vbCr, vbLf)
    parts = Split(text, vbLf)
    widthChars = CLng(Int(colWidthChars))
    If widthChars < 1 Then widthChars = 1

    For Each part In parts
        partText = CStr(part)
        If Len(partText) = 0 Then
            total = total + 1
        Else
            total = total + ((Len(partText) + widthChars - 1) \ widthChars)
        End If
    Next part

    If total < 1 Then total = 1
    EstimateLines = total
End Function

Private Sub NormaliseRowHeights(ByVal ws As Worksheet, ByVal maxCol As Long, ByVal lastRow As Long)
    Dim colWidths() As Double
    Dim r As Long
    Dim c As Long
    Dim maxLines As Long
    Dim currentLines As Long
    Dim text As String

    ReDim colWidths(1 To maxCol)
    For c = 1 To maxCol
        If ws.Columns(c).ColumnWidth > 0 Then
            colWidths(c) = ws.Columns(c).ColumnWidth
        Else
            colWidths(c) = 10#
        End If
    Next c

    For r = 7 To lastRow
        With ws.Cells(r, 2)
            .WrapText = True
            If .VerticalAlignment = 0 Then
                .VerticalAlignment = xlCenter
            End If
        End With

        maxLines = 1
        For c = 1 To maxCol
            text = CellText(ws.Cells(r, c).Value)
            If text <> "" Then
                currentLines = EstimateLines(text, colWidths(c))
                If currentLines > maxLines Then
                    maxLines = currentLines
                End If
            End If
        Next c
        ws.Rows(r).RowHeight = maxLines * 15#
    Next r
End Sub
'''


def process_workbook_com(input_path: Path, save_path: Path) -> None:
    if not HAS_WIN32COM:
        raise RuntimeError("win32com is not available on this machine")

    if input_path.resolve() != save_path.resolve():
        shutil.copy2(input_path, save_path)

    excel = None
    workbook = None
    module_comp = None
    save_ok = False
    previous_calc = None

    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.EnableEvents = False
        excel.ScreenUpdating = False
        excel.AskToUpdateLinks = False
        try:
            previous_calc = excel.Calculation
            excel.Calculation = -4135
        except Exception:
            previous_calc = None

        workbook = excel.Workbooks.Open(str(save_path), UpdateLinks=0, ReadOnly=False)
        vb_project = workbook.VBProject
        module_comp = vb_project.VBComponents.Add(1)
        module_comp.Name = "CodexFastModule"
        module_comp.CodeModule.AddFromString(COM_VBA_CODE)

        excel.Run(f"'{workbook.Name}'!CodexProcessWorkbook")

        vb_project.VBComponents.Remove(module_comp)
        module_comp = None

        workbook.Save()
        save_ok = True
    finally:
        try:
            if module_comp is not None and workbook is not None:
                workbook.VBProject.VBComponents.Remove(module_comp)
        except Exception:
            pass
        try:
            if workbook is not None:
                workbook.Close(SaveChanges=save_ok)
        except Exception:
            pass
        try:
            if excel is not None and previous_calc is not None:
                excel.Calculation = previous_calc
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        if not save_ok and save_path.exists() and input_path.resolve() != save_path.resolve():
            try:
                save_path.unlink()
            except Exception:
                pass


def process_workbook_auto(
    input_path: Path,
    save_path: Path,
    prefer_fast: bool = False,
    progress_hook: Callable[[str, Dict[str, object]], None] | None = None,
) -> str:
    if prefer_fast and HAS_WIN32COM:
        try:
            process_workbook_com(input_path, save_path)
            return "Fast Excel"
        except Exception:
            if save_path.exists() and input_path.resolve() != save_path.resolve():
                try:
                    save_path.unlink()
                except Exception:
                    pass

    process_workbook(input_path, save_path, progress_hook=progress_hook)
    return "openpyxl"


def process_one_file_task(
    src_path: str,
    dst_path: str,
    prefer_fast: bool = False,
) -> Tuple[str, str]:
    backend = process_workbook_auto(
        Path(src_path),
        Path(dst_path),
        prefer_fast=prefer_fast,
    )
    return dst_path, backend


# ----------------------------------------------------------------
# Row-height helpers
# ----------------------------------------------------------------
def _estimate_lines(text: str, col_width_chars: float) -> int:
    if not text:
        return 1

    total = 0
    for ln in text.split("\n"):
        length = len(ln)
        if length == 0:
            total += 1
        else:
            total += max(1, -(-length // max(int(col_width_chars), 1)))
    return total


def _build_col_width_cache(sheet, max_col: int) -> List[float]:
    col_widths: List[float] = []
    for c in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(c)
        dim = sheet.column_dimensions.get(col_letter)
        col_widths.append(dim.width if (dim and dim.width) else 10.0)
    return col_widths


def _auto_row_height(
    sheet,
    row: int,
    col_widths: List[float],
    default_line_h: float = 15.0,
) -> float:
    max_lines = 1
    cell = sheet.cell
    for c, col_w in enumerate(col_widths, start=1):
        val = cell(row=row, column=c).value
        if val is None:
            continue
        lines = _estimate_lines(str(val), col_w)
        if lines > max_lines:
            max_lines = lines
    return max_lines * default_line_h


# ----------------------------------------------------------------
# Core workbook processor
# ----------------------------------------------------------------
def process_workbook(
    input_path: Path,
    save_path: Path,
    progress_hook: Callable[[str, Dict[str, object]], None] | None = None,
) -> None:
    if progress_hook is not None:
        progress_hook("loading", {"file_name": input_path.name})

    wb = openpyxl.load_workbook(
        input_path,
        keep_links=False,
        rich_text=False,
    )
    try:
        target_sheets = [
            sheet
            for sheet in wb.worksheets
            if sheet.title.strip().lower() != "contents" and sheet.max_row >= 7
        ]
        sheet_total = len(target_sheets)
        if progress_hook is not None:
            progress_hook(
                "loaded",
                {
                    "file_name": input_path.name,
                    "sheet_total": sheet_total,
                },
            )

        for sheet_index, sheet in enumerate(target_sheets, start=1):
            if progress_hook is not None:
                progress_hook(
                    "sheet_start",
                    {
                        "file_name": input_path.name,
                        "sheet_name": sheet.title,
                        "sheet_index": sheet_index,
                        "sheet_total": sheet_total,
                    },
                )

            max_col = _detect_effective_max_col(sheet)

            if not isinstance(sheet["A6"], MergedCell):
                sheet["A6"] = "1st row: Column %"

            _unmerge_data_region(sheet)

            rows_to_delete: List[int] = []
            current_max_row = sheet.max_row
            rows = list(
                sheet.iter_rows(
                    min_row=7,
                    max_row=current_max_row,
                    min_col=1,
                    max_col=max_col,
                )
            )

            col_b: List[str] = []
            for row in rows:
                value = row[1].value
                col_b.append(str(value).strip() if value is not None else "")

            for idx in range(len(rows) - 1, 0, -1):
                cur_text = col_b[idx]
                above_text = col_b[idx - 1]
                current_row = rows[idx]
                upper_row = rows[idx - 1]

                is_regular = (
                    cur_text != ""
                    and cur_text.upper() != "TOTAL"
                    and cur_text == above_text
                )
                if is_regular:
                    for target, source in zip(upper_row, current_row):
                        target.border = Border(
                            left=target.border.left,
                            right=target.border.right,
                            top=target.border.top,
                            bottom=source.border.bottom,
                        )

                    for target, source in zip(upper_row[2:], current_row[2:]):
                        chosen_val, chosen_fmt = _choose_value_and_format(target, source)
                        target.value = chosen_val
                        target.number_format = chosen_fmt

                    rows_to_delete.append(idx + 7)
                    continue

                if above_text.upper() == "TOTAL" and (
                    cur_text.upper() == "TOTAL" or cur_text == ""
                ):
                    for target, source in zip(upper_row, current_row):
                        target.border = Border(
                            left=target.border.left,
                            right=target.border.right,
                            top=target.border.top,
                            bottom=source.border.bottom,
                        )
                    rows_to_delete.append(idx + 7)

            _delete_rows_desc(sheet, rows_to_delete)

            last_row = _find_last_data_row(sheet, max_col)
            thin = Side(border_style="thin", color="000000")
            for row in sheet.iter_rows(
                min_row=last_row,
                max_row=last_row,
                min_col=1,
                max_col=max_col,
            ):
                for current in row:
                    current.border = Border(
                        left=current.border.left,
                        right=current.border.right,
                        top=current.border.top,
                        bottom=thin,
                    )

            _normalise_row_heights(sheet, max_col, last_row)

            if progress_hook is not None:
                progress_hook(
                    "sheet_done",
                    {
                        "file_name": input_path.name,
                        "sheet_name": sheet.title,
                        "sheet_index": sheet_index,
                        "sheet_total": sheet_total,
                    },
                )

        if progress_hook is not None:
            progress_hook(
                "saving",
                {
                    "file_name": input_path.name,
                    "sheet_total": sheet_total,
                },
            )
        wb.save(save_path)
        if progress_hook is not None:
            progress_hook(
                "saved",
                {
                    "file_name": input_path.name,
                    "sheet_total": sheet_total,
                },
            )
    finally:
        wb.close()


# ----------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------
def _detect_effective_max_col(sheet) -> int:
    dim = sheet.calculate_dimension()
    _, _, dim_max_col, _ = range_boundaries(dim)
    max_row = sheet.max_row

    sample_rows = set(range(1, min(max_row, 30) + 1))
    sample_rows.update({4, 5, 6, 7, 8, 9, 10, max_row})
    if max_row > 30:
        sample_rows.update(range(max(1, max_row - 9), max_row + 1))
    sample_rows = sorted(r for r in sample_rows if 1 <= r <= max_row)

    cell = sheet.cell
    for col in range(dim_max_col, 2, -1):
        for r in sample_rows:
            value = cell(row=r, column=col).value
            if value is not None and str(value).strip() != "":
                return col
    return max(3, dim_max_col)


def _unmerge_data_region(sheet) -> None:
    for m_range in list(sheet.merged_cells.ranges):
        min_col, min_row, m_max_col, m_max_row = m_range.bounds
        if min_row < 7:
            continue

        top = sheet.cell(row=min_row, column=min_col)
        tl_val = top.value
        tl_style = copy(top._style)

        sheet.unmerge_cells(str(m_range))

        for r in range(min_row, m_max_row + 1):
            for c in range(min_col, m_max_col + 1):
                current = sheet.cell(row=r, column=c)
                current.value = tl_val
                current._style = tl_style


def _choose_value_and_format(upper_cell, lower_cell):
    upper_val = upper_cell.value
    lower_val = lower_cell.value

    if lower_val is None or str(lower_val).strip() == "":
        return upper_val, upper_cell.number_format

    if isinstance(upper_val, (int, float)) and isinstance(lower_val, (int, float)):
        upper_has_frac = abs(float(upper_val) - int(float(upper_val))) > 1e-12
        lower_has_frac = abs(float(lower_val) - int(float(lower_val))) > 1e-12
        if upper_has_frac and not lower_has_frac:
            return upper_val, upper_cell.number_format
        return lower_val, lower_cell.number_format

    return lower_val, lower_cell.number_format


def _delete_rows_desc(sheet, row_indexes: List[int]) -> None:
    if not row_indexes:
        return

    rows = sorted(set(row_indexes), reverse=True)
    start = rows[0]
    count = 1
    prev = rows[0]

    for r in rows[1:]:
        if r == prev - 1:
            count += 1
        else:
            sheet.delete_rows(start - count + 1, count)
            start = r
            count = 1
        prev = r

    sheet.delete_rows(start - count + 1, count)


def _find_last_data_row(sheet, max_col: int) -> int:
    cell = sheet.cell
    for r in range(sheet.max_row, 6, -1):
        for c in range(1, max_col + 1):
            value = cell(row=r, column=c).value
            if value is not None and str(value).strip() != "":
                return r
    return 7


def _normalise_row_heights(sheet, max_col: int, last_row: int) -> None:
    uniform_height = (
        sheet.row_dimensions[7].height
        or sheet.sheet_format.defaultRowHeight
        or 15.0
    )
    cell = sheet.cell
    for r in range(7, last_row + 1):
        b_cell = cell(row=r, column=2)
        if b_cell.alignment:
            b_cell.alignment = Alignment(
                horizontal=b_cell.alignment.horizontal,
                vertical=b_cell.alignment.vertical or "center",
                wrap_text=False,
            )
        else:
            b_cell.alignment = Alignment(vertical="center", wrap_text=False)

        sheet.row_dimensions[r].height = uniform_height


# ----------------------------
# Input resolution
# ----------------------------
def resolve_input_items(items: List[str]) -> Tuple[List[Path], List[str]]:
    valid_paths: List[Path] = []
    invalid_items: List[str] = []

    for raw in items:
        p = raw.strip().strip('"').strip("'")
        if not p:
            continue

        path_obj = Path(p)
        if path_obj.is_file() and path_obj.suffix.lower() == ".xlsx":
            valid_paths.append(path_obj)
        elif path_obj.is_dir():
            files = sorted(path_obj.glob("*.xlsx"))
            if files:
                valid_paths.extend(files)
            else:
                invalid_items.append(f"{p} (no .xlsx inside folder)")
        else:
            invalid_items.append(p)

    seen = set()
    deduped: List[Path] = []
    for path in valid_paths:
        key = str(path.resolve()).lower()
        if key not in seen:
            seen.add(key)
            deduped.append(path)

    return deduped, invalid_items


def _count_processable_sheets_xlsx(path: Path) -> int:
    try:
        with zipfile.ZipFile(path) as zf:
            workbook_xml = zf.read("xl/workbook.xml")
        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        root = ET.fromstring(workbook_xml)
        count = 0
        for sheet in root.findall(".//main:sheets/main:sheet", ns):
            name = sheet.attrib.get("name", "").strip().lower()
            if name != "contents":
                count += 1
        return max(count, 1)
    except Exception:
        return 1


def main() -> None:
    multiprocessing.freeze_support()

    # Delay PyQt imports so spawned worker processes do not pay the UI import cost.
    from PyQt6.QtCore import (
        QEasingCurve,
        QObject,
        QPropertyAnimation,
        QThread,
        QTimer,
        Qt,
        pyqtSignal,
    )
    from PyQt6.QtGui import QColor, QPalette
    from PyQt6.QtWidgets import (
        QApplication,
        QFileDialog,
        QFrame,
        QGraphicsOpacityEffect,
        QHBoxLayout,
        QLabel,
        QListWidget,
        QMainWindow,
        QMessageBox,
        QProgressBar,
        QPushButton,
        QTextEdit,
        QVBoxLayout,
        QWidget,
    )

    class AnimatedBackground(QWidget):
        pass

    class FancyButton(QPushButton):
        def __init__(self, text: str) -> None:
            super().__init__(text)
            self.setCursor(Qt.CursorShape.PointingHandCursor)

    class ProcessingWorker(QObject):
        progress = pyqtSignal(int, str)
        log = pyqtSignal(str)
        done = pyqtSignal(list)
        failed = pyqtSignal(str)

        def __init__(self, files: List[Path], output_dir: Path) -> None:
            super().__init__()
            self.files = files
            self.output_dir = output_dir

        def run(self) -> None:
            try:
                saved_files: List[str] = []
                total = len(self.files)
                # Keep the proven openpyxl path as the default because the
                # experimental COM backend can preserve the input unchanged.
                prefer_fast = False
                if total == 0:
                    self.done.emit([])
                    return

                self.progress.emit(0, f"Preparing {total} file(s)...")
                self.log.emit(f"Preparing {total} file(s)")

                reserved_names: set[str] = set()
                plans: List[Tuple[Path, Path]] = []
                for src in self.files:
                    dst = unique_output_path_reserved(self.output_dir, src, reserved_names)
                    plans.append((src, dst))

                estimated_sheets: Dict[str, int] = {
                    str(src.resolve()): _count_processable_sheets_xlsx(src)
                    for src, _ in plans
                }
                total_units = sum(estimated_sheets.values()) + (len(plans) * 2)
                completed_units = 0

                def emit_progress(status: str) -> None:
                    pct = int((completed_units / total_units) * 100) if total_units > 0 else 0
                    pct = max(0, min(pct, 100))
                    self.progress.emit(pct, status)

                for file_index, (src, dst) in enumerate(plans, start=1):
                    file_key = str(src.resolve())
                    self.log.emit(f"[{file_index}/{total}] Starting {src.name}")

                    def report(event: str, payload: Dict[str, object]) -> None:
                        nonlocal completed_units, total_units
                        file_name = str(payload.get("file_name", src.name))

                        if event == "loaded":
                            actual_sheets = int(payload.get("sheet_total", estimated_sheets[file_key]))
                            total_units += actual_sheets - estimated_sheets[file_key]
                            estimated_sheets[file_key] = actual_sheets
                            completed_units += 1
                            emit_progress(
                                f"[{file_index}/{total}] Loaded {file_name} ({actual_sheets} sheets)"
                            )
                            self.log.emit(
                                f"[{file_index}/{total}] Loaded {file_name} ({actual_sheets} sheets)"
                            )
                            return

                        if event == "sheet_start":
                            sheet_index = int(payload.get("sheet_index", 0))
                            sheet_total = int(payload.get("sheet_total", estimated_sheets[file_key]))
                            sheet_name = str(payload.get("sheet_name", ""))
                            emit_progress(
                                f"[{file_index}/{total}] {file_name} - sheet {sheet_index}/{sheet_total}: {sheet_name}"
                            )
                            if (
                                sheet_index == 1
                                or sheet_index == sheet_total
                                or sheet_index % 25 == 0
                            ):
                                self.log.emit(
                                    f"[{file_index}/{total}] Working on sheet {sheet_index}/{sheet_total}: {sheet_name}"
                                )
                            return

                        if event == "sheet_done":
                            completed_units += 1
                            sheet_index = int(payload.get("sheet_index", 0))
                            sheet_total = int(payload.get("sheet_total", estimated_sheets[file_key]))
                            sheet_name = str(payload.get("sheet_name", ""))
                            emit_progress(
                                f"[{file_index}/{total}] Finished sheet {sheet_index}/{sheet_total}: {sheet_name}"
                            )
                            return

                        if event == "saving":
                            emit_progress(f"[{file_index}/{total}] Saving {file_name}...")
                            self.log.emit(f"[{file_index}/{total}] Saving {file_name}...")
                            return

                        if event == "saved":
                            completed_units += 1
                            emit_progress(f"[{file_index}/{total}] Saved {file_name}")
                            self.log.emit(f"[{file_index}/{total}] Saved {file_name}")

                    try:
                        backend = process_workbook_auto(
                            src,
                            dst,
                            prefer_fast=prefer_fast,
                            progress_hook=report,
                        )
                        saved_files.append(str(dst))
                        self.log.emit(f"[{file_index}/{total}] Completed with {backend}: {src.name}")
                    except Exception as exc:
                        self.log.emit(f"[{file_index}/{total}] ERROR in {src.name}: {exc}")
                        self.progress.emit(
                            max(0, min(int((completed_units / total_units) * 100), 100)),
                            f"ERROR: {src.name} - {exc}",
                        )
                        raise

                ordered_set = set(saved_files)
                ordered = [str(dst) for _, dst in plans if str(dst) in ordered_set]
                self.done.emit(ordered)
            except Exception as exc:
                self.failed.emit(str(exc))

    class MainWindow(QMainWindow):
        def __init__(self) -> None:
            super().__init__()
            self.selected_files: List[Path] = []
            self.output_dir: Path | None = None
            self.worker_thread: QThread | None = None
            self.worker: ProcessingWorker | None = None
            self.started_at: float | None = None
            self.root_layout: QVBoxLayout | None = None
            self.card_layout: QVBoxLayout | None = None

            self.setWindowTitle("Excel Smart Formatter Pro")
            self.resize(980, 700)

            self.bg = AnimatedBackground()
            self.bg.setObjectName("bg")
            self.setCentralWidget(self.bg)

            self._build_ui()
            QTimer.singleShot(0, self._fit_to_screen)
            self._animate_intro()

        def _build_ui(self) -> None:
            self.root_layout = QVBoxLayout(self.bg)
            self.root_layout.setContentsMargins(40, 30, 40, 30)

            self.card = QFrame()
            self.card.setObjectName("card")
            self.card_layout = QVBoxLayout(self.card)
            self.card_layout.setSpacing(14)
            self.card_layout.setContentsMargins(24, 24, 24, 24)

            title = QLabel("N% To %")
            title.setObjectName("title")
            self.title_label = title
            subtitle = QLabel(
                "Professional PyQt6 UI with multi-file workflow, folder/path paste, and smooth motion"
            )
            subtitle.setObjectName("subtitle")
            subtitle.setWordWrap(True)
            self.subtitle_label = subtitle

            self.btn_pick_files = FancyButton("Select .xlsx Files")
            self.btn_pick_files.clicked.connect(self.pick_files)

            self.btn_pick_output = FancyButton("Select Output Folder")
            self.btn_pick_output.clicked.connect(self.pick_output_folder)

            self.paths_edit = QTextEdit()
            self.paths_edit.setPlaceholderText(
                "Paste file/folder paths here. One per line (or separated by semicolon)."
            )
            self.paths_edit.setFixedHeight(120)

            self.btn_apply_paths = FancyButton("Apply Pasted Paths")
            self.btn_apply_paths.clicked.connect(self.apply_pasted_paths)

            self.list_files = QListWidget()
            self.list_files.setMinimumHeight(160)

            self.lbl_output = QLabel("Output folder: not selected")
            self.lbl_output.setObjectName("meta")
            self.lbl_status = QLabel("Ready")
            self.lbl_status.setObjectName("status")

            self.progress = QProgressBar()
            self.progress.setRange(0, 100)
            self.progress.setValue(0)
            self.progress.setFormat("0%")

            self.log_view = QTextEdit()
            self.log_view.setReadOnly(True)
            self.log_view.setPlaceholderText("Processing log will appear here.")
            self.log_view.setMinimumHeight(150)

            self.btn_start = FancyButton("Process All Files")
            self.btn_start.clicked.connect(self.start_processing)
            self.btn_start.setEnabled(False)

            row = QHBoxLayout()
            row.addWidget(self.btn_pick_files)
            row.addWidget(self.btn_pick_output)
            self.top_button_row = row

            self.card_layout.addWidget(title)
            self.card_layout.addWidget(subtitle)
            self.card_layout.addLayout(row)
            self.card_layout.addWidget(self.paths_edit)
            self.card_layout.addWidget(self.btn_apply_paths)
            self.card_layout.addWidget(self.list_files)
            self.card_layout.addWidget(self.lbl_output)
            self.card_layout.addWidget(self.progress)
            self.card_layout.addWidget(self.lbl_status)
            self.card_layout.addWidget(self.log_view)
            self.card_layout.addWidget(self.btn_start)

            self.root_layout.addWidget(self.card)

            self.setStyleSheet(
                """
                QMainWindow, QWidget { color: #eaf0ff; font-family: 'Segoe UI', 'Tahoma'; font-size: 14px; }
                #bg {
                    background: qradialgradient(cx:0.2, cy:0.1, radius:1.25,
                        fx:0.15, fy:0.05,
                        stop:0 #1f3866, stop:0.45 #142746, stop:1 #0b1321);
                }
                #card {
                    background: rgba(11, 19, 31, 0.80);
                    border: 1px solid rgba(255,255,255,0.16);
                    border-radius: 20px;
                }
                #title { font-size: 34px; font-weight: 700; color: #ffffff; }
                #subtitle { font-size: 13px; color: #bfd3ff; }
                #meta { color: #98b7ff; font-size: 12px; }
                #status { color: #ffffff; font-size: 12px; }
                QTextEdit, QListWidget {
                    background: rgba(255, 255, 255, 0.08);
                    border: 1px solid rgba(255, 255, 255, 0.2);
                    border-radius: 12px;
                    padding: 8px;
                    color: #f2f6ff;
                }
                QProgressBar {
                    background: rgba(255, 255, 255, 0.10);
                    border: 1px solid rgba(255,255,255,0.2);
                    border-radius: 9px;
                    text-align: center;
                    height: 18px;
                }
                QProgressBar::chunk {
                    border-radius: 8px;
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #2dc3ff, stop:1 #4df0b7);
                }
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                        stop:0 #0ea5e9, stop:1 #10b981);
                    border: none;
                    border-radius: 12px;
                    color: #062033;
                    font-weight: 700;
                    padding: 12px 16px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                        stop:0 #38bdf8, stop:1 #34d399);
                }
                QPushButton:pressed {
                    padding-top: 13px;
                    padding-bottom: 11px;
                }
                QPushButton:disabled {
                    background: rgba(255,255,255,0.25);
                    color: rgba(255,255,255,0.6);
                }
                """
            )

        def _animate_intro(self) -> None:
            effect = QGraphicsOpacityEffect(self.card)
            self.card.setGraphicsEffect(effect)
            anim = QPropertyAnimation(effect, b"opacity", self)
            anim.setDuration(650)
            anim.setStartValue(0.0)
            anim.setEndValue(1.0)
            anim.setEasingCurve(QEasingCurve.Type.OutCubic)
            anim.start()
            self._intro_anim = anim

        def _fit_to_screen(self) -> None:
            screen = self.screen() or QApplication.primaryScreen()
            if screen is None:
                return

            available = screen.availableGeometry()
            max_width = max(900, available.width() - 36)
            max_height = max(720, available.height() - 48)
            target_width = min(1120, max_width)
            target_height = min(960, max_height)

            compact = target_height < 900
            vertical_margin = 16 if compact else 22
            horizontal_margin = 24 if target_width < 960 else 36
            card_margin = 18 if compact else 22
            spacing = 10 if compact else 12

            if self.root_layout is not None:
                self.root_layout.setContentsMargins(
                    horizontal_margin,
                    vertical_margin,
                    horizontal_margin,
                    vertical_margin,
                )
            if self.card_layout is not None:
                self.card_layout.setContentsMargins(
                    card_margin,
                    card_margin,
                    card_margin,
                    card_margin,
                )
                self.card_layout.setSpacing(spacing)

            self.resize(target_width, target_height)
            self.setMaximumSize(max_width, max_height)
            self.card.adjustSize()
            self.card_layout.activate()

            fixed_height = 0
            fixed_height += self.title_label.sizeHint().height()
            fixed_height += self.subtitle_label.sizeHint().height()
            fixed_height += max(
                self.btn_pick_files.sizeHint().height(),
                self.btn_pick_output.sizeHint().height(),
            )
            fixed_height += self.btn_apply_paths.sizeHint().height()
            fixed_height += self.lbl_output.sizeHint().height()
            fixed_height += self.progress.sizeHint().height()
            fixed_height += self.lbl_status.sizeHint().height()
            fixed_height += self.btn_start.sizeHint().height()

            item_count = 11
            total_spacing = spacing * (item_count - 1)
            usable_height = (
                target_height
                - (vertical_margin * 2)
                - (card_margin * 2)
                - total_spacing
                - fixed_height
            )

            scale = min(1.0, target_height / 960.0)
            paths_height = int(110 * scale)
            files_height = int(160 * scale)
            log_height = int(155 * scale)

            min_paths = 82 if compact else 92
            min_files = 110 if compact else 130
            min_log = 105 if compact else 125

            paths_height = max(min_paths, paths_height)
            files_height = max(min_files, files_height)
            log_height = max(min_log, log_height)

            total_variable = paths_height + files_height + log_height
            if total_variable > usable_height:
                ratio = usable_height / total_variable if total_variable > 0 else 1.0
                paths_height = max(min_paths, int(paths_height * ratio))
                files_height = max(min_files, int(files_height * ratio))
                log_height = max(min_log, int(log_height * ratio))

                total_variable = paths_height + files_height + log_height
                overflow = total_variable - usable_height
                if overflow > 0:
                    trim_log = min(max(0, log_height - min_log), overflow)
                    log_height -= trim_log
                    overflow -= trim_log
                if overflow > 0:
                    trim_files = min(max(0, files_height - min_files), overflow)
                    files_height -= trim_files
                    overflow -= trim_files
                if overflow > 0:
                    trim_paths = min(max(0, paths_height - min_paths), overflow)
                    paths_height -= trim_paths

            self.paths_edit.setFixedHeight(paths_height)
            self.list_files.setFixedHeight(files_height)
            self.log_view.setFixedHeight(log_height)

            frame = self.frameGeometry()
            frame.moveCenter(available.center())
            self.move(frame.topLeft())

        def refresh_file_list(self) -> None:
            self.list_files.clear()
            for path in self.selected_files:
                self.list_files.addItem(str(path))
            self.btn_start.setEnabled(bool(self.selected_files and self.output_dir))

        def set_progress_value(self, pct: int) -> None:
            pct = max(0, min(int(pct), 100))
            self.progress.setValue(pct)
            self.progress.setFormat(f"{pct}%")

        def append_log(self, text: str) -> None:
            timestamp = datetime.now().strftime("%H:%M:%S")
            self.log_view.append(f"[{timestamp}] {text}")

        def pick_files(self) -> None:
            files, _ = QFileDialog.getOpenFileNames(
                self,
                "Select Excel Files",
                "",
                "Excel Files (*.xlsx);;All Files (*.*)",
            )
            if not files:
                return

            parsed, invalid = resolve_input_items(files)
            self.selected_files = parsed
            self.refresh_file_list()

            if invalid:
                QMessageBox.warning(self, "Invalid paths", "\n".join(invalid[:5]))
            self.lbl_status.setText(f"Loaded {len(self.selected_files)} files")

        def pick_output_folder(self) -> None:
            folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
            if not folder:
                return

            self.output_dir = Path(folder)
            self.lbl_output.setText(f"Output folder: {self.output_dir}")
            self.refresh_file_list()

        def apply_pasted_paths(self) -> None:
            raw = self.paths_edit.toPlainText().replace(";", "\n")
            items = [line for line in raw.splitlines() if line.strip()]
            files, invalid = resolve_input_items(items)

            self.selected_files = files
            self.refresh_file_list()

            if invalid:
                show = "\n".join(invalid[:7])
                more = "" if len(invalid) <= 7 else f"\n... and {len(invalid) - 7} more"
                QMessageBox.warning(self, "Some inputs were not valid", f"{show}{more}")

            self.lbl_status.setText(
                f"Loaded {len(self.selected_files)} files from pasted paths"
            )

        def start_processing(self) -> None:
            if not self.selected_files:
                QMessageBox.warning(self, "No files", "Please select files first")
                return
            if not self.output_dir:
                QMessageBox.warning(
                    self, "No output folder", "Please choose an output folder"
                )
                return

            self.btn_start.setEnabled(False)
            self.progress.setRange(0, 100)
            self.set_progress_value(0)
            self.lbl_status.setText("Starting processing...")
            self.started_at = time.perf_counter()
            self.log_view.clear()
            self.append_log(f"Queued {len(self.selected_files)} file(s) for processing")

            self.worker_thread = QThread()
            self.worker = ProcessingWorker(self.selected_files, self.output_dir)
            self.worker.moveToThread(self.worker_thread)

            self.worker_thread.started.connect(self.worker.run)
            self.worker.progress.connect(self.on_progress)
            self.worker.log.connect(self.append_log)
            self.worker.done.connect(self.on_done)
            self.worker.failed.connect(self.on_failed)
            self.worker.done.connect(self.worker_thread.quit)
            self.worker.failed.connect(self.worker_thread.quit)
            self.worker_thread.finished.connect(self.worker_thread.deleteLater)
            self.worker_thread.finished.connect(self.worker.deleteLater)
            self.worker_thread.finished.connect(lambda: self.btn_start.setEnabled(True))

            self.worker_thread.start()

        def on_progress(self, pct: int, text: str) -> None:
            self.set_progress_value(pct)
            self.lbl_status.setText(text)

        def on_done(self, saved_files: List[str]) -> None:
            self.progress.setRange(0, 100)
            self.set_progress_value(100)
            self.lbl_status.setText("Completed successfully")
            elapsed = 0.0
            if self.started_at is not None:
                elapsed = time.perf_counter() - self.started_at
            avg = (elapsed / len(saved_files)) if saved_files else 0.0
            self.append_log(
                f"Completed successfully: {len(saved_files)} file(s) in {elapsed:.1f} sec"
            )

            preview = "\n".join(Path(path).name for path in saved_files[:8])
            more = "" if len(saved_files) <= 8 else f"\n... and {len(saved_files) - 8} more"
            summary = (
                f"Processed {len(saved_files)} file(s)\n"
                f"Elapsed: {elapsed:.1f} sec\n"
                f"Average: {avg:.1f} sec/file"
            )
            detail = f"Created files:\n{preview}{more}"

            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setWindowTitle("Done")
            msg.setText(summary)
            msg.setInformativeText(detail)
            msg.setStyleSheet(
                "QMessageBox { background: #f7fafc; }"
                "QLabel { color: #0f172a; font-size: 12px; }"
                "QPushButton { min-width: 84px; color: #0f172a; background: #bae6fd; border: 1px solid #7dd3fc; border-radius: 8px; padding: 6px 12px; }"
                "QPushButton:hover { background: #7dd3fc; }"
            )
            msg.exec()

        def on_failed(self, message: str) -> None:
            self.progress.setRange(0, 100)
            self.set_progress_value(self.progress.value())
            self.lbl_status.setText("Error occurred")
            self.append_log(f"FAILED: {message}")
            QMessageBox.critical(self, "Processing failed", message)

    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    palette = QPalette()
    palette.setColor(QPalette.ColorRole.WindowText, QColor("#eaf0ff"))
    app.setPalette(palette)

    def handle_uncaught_exception(exc_type, exc_value, exc_tb):
        err_text = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
        log_path = Path(__file__).with_name("6_TEST_error.log")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with log_path.open("a", encoding="utf-8") as file_obj:
            file_obj.write(f"\n[{timestamp}]\n{err_text}\n")
        QMessageBox.critical(
            None,
            "Application Error",
            f"The application encountered an error.\nLog saved to:\n{log_path}\n\n{exc_value}",
        )

    sys.excepthook = handle_uncaught_exception

    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()

# <<< START OF CHANGES >>>
# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
def run_this_app(working_dir=None): # ชื่อฟังก์ชันนี้จะถูกใช้ใน Launcher
    """
    ฟังก์ชันหลักสำหรับสร้างและรัน QuotaSamplerApp.
    """
    print(f"--- QUOTA_SAMPLER_INFO: Starting 'QuotaSamplerApp' via run_this_app() ---")
    try:
    # --- ส่วนที่ใช้รันโปรแกรม ---

    #if __name__ == "__main__":
        main()


        print(f"--- QUOTA_SAMPLER_INFO: QuotaSamplerApp mainloop finished. ---")

    except Exception as e:
        # ดักจับ Error ที่อาจเกิดขึ้นระหว่างการสร้างหรือรัน App
        print(f"QUOTA_SAMPLER_ERROR: An error occurred during QuotaSamplerApp execution: {e}")
        # แสดง Popup ถ้ามีปัญหา
        if 'root' not in locals() or not root.winfo_exists(): # สร้าง root ชั่วคราวถ้ายังไม่มี
            root_temp = tk.Tk()
            root_temp.withdraw()
            messagebox.showerror("Application Error (Quota Sampler)",
                                f"An unexpected error occurred:\n{e}", parent=root_temp)
            root_temp.destroy()
        else:
            messagebox.showerror("Application Error (Quota Sampler)",
                                f"An unexpected error occurred:\n{e}", parent=root) # ใช้ root ที่มีอยู่ถ้าเป็นไปได้
        sys.exit(f"Error running QuotaSamplerApp: {e}") # อาจจะ exit หรือไม่ก็ได้ ขึ้นกับการออกแบบ


# --- ส่วน Run Application เมื่อรันไฟล์นี้โดยตรง (สำหรับ Test) ---
if __name__ == "__main__":
    print("--- Running QuotaSamplerApp.py directly for testing ---")
    # (ถ้ามีการตั้งค่า DPI ด้านบน มันจะทำงานอัตโนมัติ)

    # เรียกฟังก์ชัน Entry Point ที่เราสร้างขึ้น
    run_this_app()

    print("--- Finished direct execution of QuotaSamplerApp.py ---")
# <<< END OF CHANGES >>>
