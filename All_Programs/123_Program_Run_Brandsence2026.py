import pandas as pd
import re
import pyreadstat
import numpy as np
import os
import glob
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment)

# --- (คงเดิม) Imports for Factor/Regression Analysis ---
import statsmodels.api as sm
from factor_analyzer import FactorAnalyzer
from collections import OrderedDict
import io
import sys
import json
import inspect
import traceback
from scipy.linalg import inv, eigh
from sklearn.preprocessing import StandardScaler
import time


# ===================================================================
# COMPATIBILITY SHIMS
# ===================================================================
_UNSET = object()   # ใช้แยก 'ยังไม่ได้ตรวจ' ออกจาก 'ตรวจแล้วไม่พบ'


def _patch_factor_analyzer_sklearn_compat():
    """factor_analyzer (<=0.5.1) เรียก check_array(force_all_finite=...)
    แต่ scikit-learn >= 1.8 เปลี่ยนชื่อพารามิเตอร์เป็น ensure_all_finite
    ทำให้ FactorAnalyzer.fit() โยน TypeError ทันที

    แก้โดยห่อ check_array ที่ factor_analyzer ผูกไว้ในโมดูลตัวเอง
    ให้แปลงชื่อพารามิเตอร์เก่าเป็นชื่อใหม่ (no-op ถ้า sklearn เป็นรุ่นเก่า)
    """
    module_names = [
        'factor_analyzer.factor_analyzer',
        'factor_analyzer.confirmatory_factor_analyzer',
    ]
    for mod_name in module_names:
        try:
            module = __import__(mod_name, fromlist=['check_array'])
        except Exception:
            continue

        original = getattr(module, 'check_array', None)
        if original is None or getattr(
                original, '_bs_compat_wrapped', False):
            continue

        try:
            params = inspect.signature(original).parameters
        except (TypeError, ValueError):
            continue
        # sklearn เก่ายังรับ force_all_finite ได้ ไม่ต้องแพตช์
        if 'force_all_finite' in params \
                or 'ensure_all_finite' not in params:
            continue

        def _make_wrapper(orig):
            def check_array_compat(*args, **kwargs):
                if 'force_all_finite' in kwargs:
                    kwargs['ensure_all_finite'] = kwargs.pop(
                        'force_all_finite')
                return orig(*args, **kwargs)
            check_array_compat._bs_compat_wrapped = True
            return check_array_compat

        module.check_array = _make_wrapper(original)


_patch_factor_analyzer_sklearn_compat()


def _df_map(dataframe, func):
    """DataFrame.applymap ถูกถอดออกใน pandas 3.0 -> ใช้ DataFrame.map แทน"""
    if hasattr(dataframe, 'map'):
        return dataframe.map(func)
    return dataframe.applymap(func)


# --- PyQt6 GUI ---
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout,
    QHBoxLayout, QPushButton, QLabel, QLineEdit,
    QCheckBox, QRadioButton, QProgressBar, QTabWidget,
    QSplitter, QScrollArea, QGroupBox, QListWidget,
    QTextEdit, QFileDialog, QMessageBox, QDialog,
    QGridLayout, QFrame, QTableWidget,
    QTableWidgetItem, QButtonGroup,
    QAbstractItemView, QHeaderView, QSizePolicy,
    QSpacerItem)
from PyQt6.QtCore import Qt, QTimer, QThread, pyqtSignal
from PyQt6.QtGui import (
    QFont as QFontObj, QColor, QPainter, QPixmap)


# --- Wrappers to keep .get()/.set() API ---
class _Var:
    def __init__(self, value=""):
        self._v = str(value)
        self._w = None

    def link(self, widget):
        self._w = widget

    def get(self):
        if self._w and hasattr(self._w, 'text'):
            return self._w.text()
        return self._v

    def set(self, value):
        self._v = str(value)
        if self._w and hasattr(self._w, 'setText'):
            self._w.setText(str(value))


class _BoolVar:
    def __init__(self, value=False):
        self._v = bool(value)
        self._w = None

    def link(self, widget):
        self._w = widget

    def get(self):
        if self._w and hasattr(self._w, 'isChecked'):
            return self._w.isChecked()
        return self._v

    def set(self, value):
        self._v = bool(value)
        if self._w and hasattr(self._w, 'setChecked'):
            self._w.setChecked(bool(value))


# --- QSS Theme (Modern) ---
_QSS = """
/* ---- Global ---- */
* { font-family: 'Segoe UI', sans-serif; }
QMainWindow { background: #F0F2F5; }
QSplitter::handle { background:#E0E0E0; width:1px; }

/* ---- Left panel ---- */
#leftPanel {
    background: qlineargradient(
        x1:0, y1:0, x2:0, y2:1,
        stop:0 #C62828, stop:1 #7B1616);
    border-top-right-radius: 20px;
    border-bottom-right-radius: 20px;
}
#banner {
    background: transparent;
    border-bottom: 1px solid rgba(255, 255, 255, 0.1);
}

/* ---- Buttons ---- */
QPushButton[class="danger"] {
    background: qlineargradient(
        x1:0,y1:0,x2:0,y2:1,
        stop:0 #EF5350, stop:1 #C62828);
    color:#fff; border:none; border-radius:8px;
    padding:10px 18px; font-size:12px;
    font-weight:600; }
QPushButton[class="danger"]:hover {
    background: qlineargradient(
        x1:0,y1:0,x2:0,y2:1,
        stop:0 #F44336, stop:1 #B71C1C); }
QPushButton[class="danger"]:pressed {
    background:#9B1B1B; }
QPushButton[class="danger"]:disabled {
    background:#E0E0E0; color:#9E9E9E; }

QPushButton[class="outline"] {
    background:transparent; color:#C62828;
    border:1.5px solid #E57373; border-radius:8px;
    padding:9px 18px; font-size:12px;
    font-weight:500; }
QPushButton[class="outline"]:hover {
    background:#FFF5F5; border-color:#C62828; }
QPushButton[class="outline"]:pressed {
    background:#FFEBEE; }
QPushButton[class="outline"]:disabled {
    border-color:#D0D0D0; color:#BDBDBD;
    background:transparent; }

QPushButton[class="warning"] {
    background: qlineargradient(
        x1:0,y1:0,x2:0,y2:1,
        stop:0 #FFB74D, stop:1 #F57C00);
    color:#fff; border:none; border-radius:8px;
    padding:10px 18px; font-size:12px;
    font-weight:600; }
QPushButton[class="warning"]:hover {
    background: qlineargradient(
        x1:0,y1:0,x2:0,y2:1,
        stop:0 #FFA726, stop:1 #E65100); }

QPushButton[class="success"] {
    background: qlineargradient(
        x1:0,y1:0,x2:0,y2:1,
        stop:0 #66BB6A, stop:1 #2E7D32);
    color:#fff; border:none; border-radius:8px;
    padding:10px 18px; font-size:12px;
    font-weight:600; }
QPushButton[class="success"]:hover {
    background: qlineargradient(
        x1:0,y1:0,x2:0,y2:1,
        stop:0 #4CAF50, stop:1 #1B5E20); }

/* ---- Inputs ---- */
QLineEdit {
    border:1.5px solid #D0D0D0; border-radius:6px;
    padding:7px 10px; font-size:12px;
    background:#FAFAFA; color:#222;
    selection-background-color:#EF9A9A; }
QLineEdit:focus {
    border-color:#E57373;
    background:#fff; }
QLineEdit:disabled {
    background:#F0F0F0; color:#999;
    border-color:#E0E0E0; }

/* ---- GroupBox ---- */
QGroupBox {
    font-weight:600; font-size:12px;
    color:#333; border:1.5px solid #D8D8D8;
    border-radius:8px; margin-top:10px;
    padding:14px 8px 8px 8px;
    background:#FAFAFA; }
QGroupBox::title {
    subcontrol-origin:margin;
    subcontrol-position:top left;
    left:12px; padding:0 6px;
    background:#FAFAFA; color:#B71C1C;
    font-weight:700; }

/* ---- Radio / Checkbox ---- */
QRadioButton { spacing:6px; font-size:12px; color:#333; }
QRadioButton::indicator {
    width:16px; height:16px; border-radius:8px;
    border:2px solid #BDBDBD; }
QRadioButton::indicator:checked {
    width:16px; height:16px; border-radius:8px;
    border:2px solid #C62828;
    background: qradialgradient(
        cx:0.5,cy:0.5,radius:0.4,
        fx:0.5,fy:0.5,
        stop:0 #fff, stop:0.35 #fff,
        stop:0.36 #C62828, stop:1 #C62828); }
QRadioButton::indicator:hover {
    border-color:#E57373; }

QCheckBox { spacing:6px; font-size:12px; color:#333; }
QCheckBox::indicator {
    width:18px; height:18px; border-radius:4px;
    border:2px solid #BDBDBD; background:#fff; }
QCheckBox::indicator:checked {
    background:#43A047; border-color:#2E7D32; }
QCheckBox::indicator:hover {
    border-color:#66BB6A; }

/* ---- Progress ---- */
QProgressBar {
    border:none; background:#FFE0E0;
    border-radius:3px; max-height:5px;
    text-align:center; }
QProgressBar::chunk {
    background: qlineargradient(
        x1:0,y1:0,x2:1,y2:0,
        stop:0 #EF5350, stop:1 #C62828);
    border-radius:3px; }

/* ---- Lists ---- */
QListWidget {
    background:#263238; color:#ECEFF1;
    border:1px solid #37474F; border-radius:6px;
    padding:4px; font-size:11px;
    selection-background-color:#EF5350;
    selection-color:#fff; outline:none; }
QListWidget::item { padding:4px 6px;
    border-radius:3px; }
QListWidget::item:selected {
    background:#EF5350; color:#fff; }
QListWidget::item:hover:!selected {
    background:#37474F; }

/* ---- Tabs ---- */
QTabWidget::pane {
    border:1px solid #D8D8D8; border-radius:6px;
    background:#fff; top:-1px; }
QTabBar::tab {
    padding:9px 24px; font-size:12px;
    color:#666; border:none;
    border-bottom:2px solid transparent;
    margin-right:2px; }
QTabBar::tab:selected {
    color:#B71C1C; font-weight:bold;
    border-bottom:3px solid #C62828; }
QTabBar::tab:hover:!selected {
    color:#E57373;
    border-bottom:2px solid #FFCDD2; }

/* ---- Table ---- */
QTableWidget {
    gridline-color:#EEEEEE; font-size:11px;
    border:1px solid #D8D8D8; border-radius:4px;
    background:#fff; alternate-background-color:#FAFAFA;
    color:#222; }
QTableWidget::item { padding:4px; }
QHeaderView::section {
    background:#F5F5F5; color:#222;
    font-weight:600; font-size:11px;
    border:none; border-bottom:2px solid #D8D8D8;
    padding:6px 8px; }

/* ---- ScrollBar ---- */
QScrollBar:vertical {
    border:none; background:#F5F5F5;
    width:8px; border-radius:4px; }
QScrollBar::handle:vertical {
    background:#BDBDBD; border-radius:4px;
    min-height:30px; }
QScrollBar::handle:vertical:hover {
    background:#9E9E9E; }
QScrollBar::add-line:vertical,
QScrollBar::sub-line:vertical { height:0; }
QScrollBar:horizontal {
    border:none; background:#F5F5F5;
    height:8px; border-radius:4px; }
QScrollBar::handle:horizontal {
    background:#BDBDBD; border-radius:4px;
    min-width:30px; }
QScrollBar::handle:horizontal:hover {
    background:#9E9E9E; }
QScrollBar::add-line:horizontal,
QScrollBar::sub-line:horizontal { width:0; }

/* ---- TextEdit (log) ---- */
QTextEdit {
    border:1px solid #D8D8D8; border-radius:6px; }

/* ---- Dialog ---- */
QDialog {
    background:#F0F2F5;
    color:#333; }
QDialog QLabel {
    color:#333;
    font-size:12px; }
QDialog QListWidget {
    background:#263238; color:#ECEFF1;
    border:1px solid #37474F; border-radius:6px;
    padding:4px; font-size:11px;
    selection-background-color:#EF5350;
    selection-color:#fff; }
QDialog QListWidget::item { padding:4px 6px; border-radius:3px; }
QDialog QListWidget::item:selected { background:#EF5350; color:#fff; }
QDialog QListWidget::item:hover:!selected { background:#37474F; }
"""

_DLG_QSS = """
QDialog {
    background:#F0F2F5; color:#333;
}
QLabel[class="dlg-header"] {
    color:#B71C1C; font-size:14px;
    font-weight:700; padding:2px 0 6px 0;
}
QLabel[class="dlg-sub"] {
    color:#444; font-size:12px;
    font-weight:600; padding:2px 0;
}
QPushButton[class="arrow"] {
    background:qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #EF5350,stop:1 #C62828);
    color:#fff; border:none; border-radius:6px;
    font-size:14px; font-weight:bold;
    padding:8px 0; min-height:32px;
}
QPushButton[class="arrow"]:hover {
    background:qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #F44336,stop:1 #B71C1C);
}
QTabWidget::pane {
    border:1px solid #D8D8D8; border-radius:6px;
    background:#fff; top:-1px;
}
QTabBar::tab {
    padding:10px 20px; font-size:12px;
    color:#555; font-weight:600;
    background:#ECEFF1;
    border:1px solid #CFD8DC;
    border-bottom:none;
    border-top-left-radius:6px;
    border-top-right-radius:6px;
    margin-right:3px;
}
QTabBar::tab:selected {
    color:#fff; font-weight:bold;
    background:qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #EF5350,stop:1 #C62828);
    border-color:#C62828;
}
QTabBar::tab:hover:!selected {
    background:#FFCDD2; color:#B71C1C;
    border-color:#E57373;
}
"""

_BTN_STYLES = {
    "danger": (
        "QPushButton{"
        "background:qlineargradient(x1:0,y1:0,x2:0,y2:1,"
        "stop:0 #EF5350,stop:1 #C62828);"
        "color:#fff;border:none;border-radius:8px;"
        "padding:10px 18px;font-size:12px;font-weight:600;}"
        "QPushButton:hover{"
        "background:qlineargradient(x1:0,y1:0,x2:0,y2:1,"
        "stop:0 #F44336,stop:1 #B71C1C);}"
        "QPushButton:pressed{background:#9B1B1B;}"
        "QPushButton:disabled{"
        "background:#E0E0E0;color:#9E9E9E;}"
    ),
    "outline": (
        "QPushButton{"
        "background:transparent;color:#C62828;"
        "border:1.5px solid #E57373;border-radius:8px;"
        "padding:9px 18px;font-size:12px;font-weight:500;}"
        "QPushButton:hover{"
        "background:#FFF5F5;border-color:#C62828;}"
        "QPushButton:pressed{background:#FFEBEE;}"
        "QPushButton:disabled{"
        "border-color:#D0D0D0;color:#BDBDBD;"
        "background:transparent;}"
    ),
    "warning": (
        "QPushButton{"
        "background:qlineargradient(x1:0,y1:0,x2:0,y2:1,"
        "stop:0 #FFB74D,stop:1 #F57C00);"
        "color:#fff;border:none;border-radius:8px;"
        "padding:10px 18px;font-size:12px;font-weight:600;}"
        "QPushButton:hover{"
        "background:qlineargradient(x1:0,y1:0,x2:0,y2:1,"
        "stop:0 #FFA726,stop:1 #E65100);}"
    ),
    "success": (
        "QPushButton{"
        "background:qlineargradient(x1:0,y1:0,x2:0,y2:1,"
        "stop:0 #66BB6A,stop:1 #2E7D32);"
        "color:#fff;border:none;border-radius:8px;"
        "padding:10px 18px;font-size:12px;font-weight:600;}"
        "QPushButton:hover{"
        "background:qlineargradient(x1:0,y1:0,x2:0,y2:1,"
        "stop:0 #4CAF50,stop:1 #1B5E20);}"
    ),
}

class _ClickableCard(QFrame):
    """การ์ดตัวเลือกที่กดได้

    ใช้ QFrame แทน QPushButton เพราะ QPushButton.sizeHint() ไม่สนใจ
    layout ลูก ทำให้การ์ดยุบจนข้อความถูกตัด ส่วน QFrame คิดขนาดจาก
    layout ให้ถูกต้อง ข้อความยาวแค่ไหนก็ตัดบรรทัดได้เอง
    """
    clicked = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_Hover, True)
        self.setCursor(Qt.CursorShape.PointingHandCursor)

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton \
                and self.rect().contains(event.position().toPoint()):
            self.clicked.emit()
        super().mouseReleaseEvent(event)


class _Worker(QThread):
    """รันงานหนักนอก main thread เพื่อไม่ให้ UI ค้าง

    ฟังก์ชันที่ส่งเข้ามาต้องไม่แตะ widget ใดๆ โดยตรง
    ให้สื่อสารกลับผ่าน signal ของ SpssProcessorApp เท่านั้น
    """
    sig_done = pyqtSignal(object)
    sig_error = pyqtSignal(str, str)

    def __init__(self, fn, parent=None):
        super().__init__(parent)
        self._fn = fn

    def run(self):
        try:
            result = self._fn()
        except Exception as e:
            self.sig_error.emit(
                f"{type(e).__name__}: {e}",
                traceback.format_exc())
            return
        self.sig_done.emit(result)


class SpssProcessorApp(QMainWindow):
    # Signals ทำให้ log/status/progress ถูกเรียกจาก worker thread ได้อย่างปลอดภัย
    # (Qt จะ queue ไปทำงานบน main thread ให้เอง)
    sig_log = pyqtSignal(str)
    sig_status = pyqtSignal(str, str)
    sig_progress = pyqtSignal(int, int)

    def __init__(self):
        super().__init__()
        self.setWindowTitle(
            "BrandSence Model Processor — By DP")
        self.resize(1050, 720)
        self.setStyleSheet(_QSS)

        # --- State Variables ---
        self.df = None
        self.spss_original_order = []
        self.computed_c_cols = []
        self.c_vars_to_compute = []
        self.vars_to_transform = {}
        self.transformed_df = None
        self.za_cols = []
        self.id_vars = []
        self.last_excel_filepath = None
        self.original_filepath = None
        self.agree_json_filepath_override = None
        self.save_all_sheets_var = _BoolVar(value=True)
        self.t2b_choice_var = _Var(value="5+4")
        self.index1_labels = {}
        self.filter_labels = {}
        self.spss_value_labels = {}
        self.spss_variable_labels = {}
        self.compute_sav_column_labels = {}
        self.compute_sav_value_labels = {}
        self.sandp_label_overrides = []
        self.agree_summary_cache_df = None
        self.is_reanalyze_mode = False
        self.reanalyze_good_mode = False
        self.good_filter_stats = None
        self.good_filter_full_df = None
        self._good_reference_cache = None
        self._good_k_used = None
        self.agree_json_auto_found = False
        self.e_group_mode_var = _Var(value="default")
        self.e_group_entry_var = _Var(value="")
        self.log_text = None
        self._analysis_errors = []
        self._analysis_skipped = []
        self._beta_warnings = []
        self._beta_zero_groups = []
        self._beta_abs_used = False
        self._weak_beta_cells = []
        self._weak_model_groups = []
        self._sample_size_approx = False
        self._respondent_key = _UNSET

        # --- Worker thread state ---
        self._worker = None
        self._ui = {}

        # --- GUI Setup ---
        self.setup_gui()
        self.sig_log.connect(self._append_log_ui)
        self.sig_status.connect(self._apply_status_ui)
        self.sig_progress.connect(self._apply_progress_ui)
        self.center_window()

    def center_window(self):
        screen = QApplication.primaryScreen()
        if screen:
            sg = screen.geometry()
            self.move(
                (sg.width() - self.width()) // 2,
                (sg.height() - self.height()) // 2)

    def _center_toplevel(self, dlg):
        dlg.move(
            self.x() + (self.width()
                        - dlg.width()) // 2,
            self.y() + (self.height()
                        - dlg.height()) // 2)

    # --- helpers for right panel ---
    def _clear_layout(self, layout):
        if layout is None:
            return
        while layout.count():
            item = layout.takeAt(0)
            w = item.widget()
            if w:
                w.setParent(None)
                w.deleteLater()
            elif item.layout():
                self._clear_layout(item.layout())

    def _clear_right_panel(self):
        self._clear_layout(self.right_frame.layout())

    def update_idletasks(self):
        QApplication.processEvents()

    def setup_gui(self):
        central = QWidget()
        self.setCentralWidget(central)
        outer = QHBoxLayout(central)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        splitter = QSplitter(
            Qt.Orientation.Horizontal)
        outer.addWidget(splitter)

        # === Left Panel (RED gradient) ===
        left = QWidget()
        left.setObjectName("leftPanel")
        left.setFixedWidth(300)
        lv = QVBoxLayout(left)
        lv.setContentsMargins(0, 0, 0, 0)
        lv.setSpacing(0)

        # --- Banner ---
        banner = QWidget()
        banner.setObjectName("banner")
        banner.setFixedHeight(64)
        bh = QHBoxLayout(banner)
        bh.setContentsMargins(16, 10, 16, 10)

        logo = QLabel("DP")
        logo.setFixedSize(40, 40)
        logo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        logo.setStyleSheet(
            "background:#fff; color:#8E0000;"
            "border-radius:20px;"
            "font-size:15px; font-weight:bold;")
        bh.addWidget(logo)

        tv = QWidget()
        tvl = QVBoxLayout(tv)
        tvl.setContentsMargins(10, 0, 0, 0)
        tvl.setSpacing(1)
        t1 = QLabel("BrandSence Model")
        t1.setStyleSheet(
            "color:#fff; font-size:15px;"
            "font-weight:bold; background:transparent;")
        t2 = QLabel("Data Processing Tool")
        t2.setStyleSheet(
            "color:rgba(255,255,255,0.6);"
            "font-size:10px;"
            "background:transparent;")
        tvl.addWidget(t1)
        tvl.addWidget(t2)
        bh.addWidget(tv, 1)
        lv.addWidget(banner)

        # --- Scroll Area for controls ---
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(
            "QScrollArea{border:none; background:transparent;}"
            "QScrollBar:vertical { border:none; background:transparent; width:8px; margin:0; }"
            "QScrollBar::handle:vertical { background:rgba(255,255,255,0.25); border-radius:4px; min-height:40px; }"
            "QScrollBar::handle:vertical:hover { background:rgba(255,255,255,0.4); }"
            "QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height:0px; }"
            "QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical { background:none; }")
        scroll_w = QWidget()
        scroll_w.setStyleSheet(
            "background:transparent;")
        sl = QVBoxLayout(scroll_w)
        sl.setContentsMargins(14, 10, 14, 6)
        sl.setSpacing(0)
        scroll.setWidget(scroll_w)
        lv.addWidget(scroll, 1)

        def sec(txt):
            lb = QLabel(txt)
            lb.setStyleSheet(
                "color:#FFFFFF;"
                "font-size:11px; font-weight:700;"
                "letter-spacing:0.8px;"
                "background:transparent;"
                "padding:12px 0 6px 2px;"
                "text-transform:uppercase;")
            sl.addWidget(lb)

        def card():
            f = QFrame()
            f.setStyleSheet(
                "QFrame{background:#FFFFFF;"
                "border-radius:10px;"
                "margin:4px 0;"
                "border:1px solid rgba(0,0,0,0.06);}")
            fl = QVBoxLayout(f)
            fl.setContentsMargins(12, 10, 12, 10)
            fl.setSpacing(6)
            sl.addWidget(f)
            return fl


        def mkbtn(layout, text, cls, slot=None):
            b = QPushButton(text)
            b.setProperty("class", cls)
            b.setMinimumHeight(38)
            b.setCursor(
                Qt.CursorShape.PointingHandCursor)
            if cls in _BTN_STYLES:
                b.setStyleSheet(_BTN_STYLES[cls])
            if slot:
                b.clicked.connect(slot)
            layout.addWidget(b)
            return b

        # ===== STEP 1 =====
        sec("\u25B6  Step 1 : โหลดข้อมูล")
        c1 = card()
        self.btn_start_process = mkbtn(
            c1, "\U0001F680  เริ่ม (เลือกตัวแปรเอง)",
            "danger", self.start_full_process)
        self.btn_load_settings_process = mkbtn(
            c1, "\U0001F4C2  เริ่ม (โหลดการตั้งค่า)",
            "outline",
            self.start_process_with_settings)
        self.btn_reanalyze = mkbtn(
            c1,
            "\U0001F504  วิเคราะห์ซ้ำ (Compute C)",
            "warning",
            self.start_reanalyze_process)

        # ===== STEP 2 =====
        sec("\u25B6  Step 2 : วิเคราะห์ & ส่งออก")
        c2 = card()

        fl = QLabel("Filter (คั่นด้วย ,) :")
        fl.setStyleSheet(
            "background:transparent;"
            "color:#333; font-size:12px;"
            "font-weight:600;")
        c2.addWidget(fl)
        self.filter_entry = QLineEdit()
        self.filter_entry.setEnabled(False)
        c2.addWidget(self.filter_entry)

        eg = QGroupBox(" Part E : Correlation ")
        eg.setStyleSheet(
            "QGroupBox{background:#FAFAFA;"
            "border:1.5px solid #E8E8E8;"
            "border-radius:8px; margin-top:8px;"
            "padding:12px 8px 8px 8px;}"
            "QGroupBox::title{color:#C62828;"
            "background:#FAFAFA;}")
        egl = QVBoxLayout(eg)
        self._rb_e_default = QRadioButton(
            "Default (E แยกกัน)")
        self._rb_e_default.setChecked(True)
        self._rb_e_group = QRadioButton(
            "Group (เช่น 4+5)")
        bg_e = QButtonGroup(self)
        bg_e.addButton(self._rb_e_default)
        bg_e.addButton(self._rb_e_group)
        self._rb_e_default.toggled.connect(
            self._on_e_mode_changed)
        egl.addWidget(self._rb_e_default)
        egl.addWidget(self._rb_e_group)
        eh = QHBoxLayout()
        elb = QLabel("ระบุ E:")
        eh.addWidget(elb)
        self.e_group_entry = QLineEdit()
        self.e_group_entry.setEnabled(False)
        self.e_group_entry.setMaximumWidth(120)
        eh.addWidget(self.e_group_entry)
        hint = QLabel("(เช่น 4+5)")
        hint.setStyleSheet(
            "color:#888; font-size:10px;"
            "background:transparent;")
        eh.addWidget(hint)
        eh.addStretch()
        egl.addLayout(eh)
        c2.addWidget(eg)

        self.e_group_entry_var.link(
            self.e_group_entry)

        self.btn_define_labels = mkbtn(
            c2, "\U0001F3F7  กำหนด Label Index",
            "outline", self.open_label_editor)
        self.btn_define_labels.setEnabled(False)

        self.cb_save_all_sheets = QCheckBox(
            "บันทึกเฉพาะ Summary")
        self.cb_save_all_sheets.setChecked(True)
        self.cb_save_all_sheets.setStyleSheet(
            "background:transparent;")
        self.save_all_sheets_var.link(
            self.cb_save_all_sheets)
        c2.addWidget(self.cb_save_all_sheets)

        self.btn_analyze_export = mkbtn(
            c2,
            "\U0001F4CA  วิเคราะห์และส่งออก Excel",
            "danger",
            self.run_analysis_and_export)
        self.btn_analyze_export.setEnabled(False)

        # ===== SETTINGS =====
        sec("\u25B6  Settings & Tools")
        c3 = card()
        self.btn_save_settings = mkbtn(
            c3,
            "\U0001F4BE  บันทึกการตั้งค่าปัจจุบัน",
            "outline", self.save_settings)
        self.btn_save_settings.setEnabled(False)

        sl.addStretch()

        # --- Bottom (credit + progress) ---
        bot = QWidget()
        bot.setStyleSheet("background:transparent;")
        botl = QVBoxLayout(bot)
        botl.setContentsMargins(14, 6, 14, 10)
        botl.setSpacing(6)

        credit = QLabel("🛠 Credit By DP")
        credit.setAlignment(
            Qt.AlignmentFlag.AlignCenter)
        credit.setStyleSheet(
            "color:rgba(255,255,255,0.7);"
            "font-size:9px; font-style:italic;"
            "background:transparent;")
        botl.addWidget(credit)

        bc = QFrame()
        bc.setStyleSheet(
            "QFrame{background:rgba(255,255,255,0.92);"
            "border-radius:8px;}")
        bcl = QVBoxLayout(bc)
        bcl.setContentsMargins(10, 8, 10, 8)
        bcl.setSpacing(4)
        self.progress = QProgressBar()
        self.progress.setRange(0, 0)
        self.progress.setVisible(False)
        bcl.addWidget(self.progress)
        self.status_label = QLabel("พร้อมทำงาน")
        self.status_label.setStyleSheet(
            "color:#555; font-size:11px;"
            "font-weight:500;")
        bcl.addWidget(self.status_label)
        botl.addWidget(bc)
        lv.addWidget(bot)

        splitter.addWidget(left)

        # === Right Panel (Display) ===
        self.right_frame = QWidget()
        self.right_frame.setStyleSheet(
            "background:#fff; border-radius:0;")
        self.right_frame.setLayout(QVBoxLayout())
        self.right_frame.layout().setContentsMargins(
            16, 16, 16, 16)
        splitter.addWidget(self.right_frame)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)

        # Welcome
        ww = QWidget()
        ww.setStyleSheet("background:transparent;")
        wl = QVBoxLayout(ww)
        wl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        wl.setSpacing(8)

        icon_lb = QLabel("\U0001F4C2")
        icon_lb.setStyleSheet(
            "font-size:48px; background:transparent;")
        icon_lb.setAlignment(
            Qt.AlignmentFlag.AlignCenter)
        wl.addWidget(icon_lb)

        self.initial_message = QLabel(
            "กรุณากด 'เริ่ม' เพื่อโหลดไฟล์ SPSS")
        self.initial_message.setStyleSheet(
            "color:#B71C1C; font-size:22px;"
            "font-weight:700;"
            "background:transparent;")
        self.initial_message.setAlignment(
            Qt.AlignmentFlag.AlignCenter)
        wl.addWidget(self.initial_message)

        sub = QLabel(
            "BrandSence Model Processor  |  By DP")
        sub.setStyleSheet(
            "color:#888; font-size:12px;"
            "background:transparent;")
        sub.setAlignment(
            Qt.AlignmentFlag.AlignCenter)
        wl.addWidget(sub)
        self.right_frame.layout().addWidget(ww)

    def _on_e_mode_changed(self, checked):
        if self._rb_e_default.isChecked():
            self.e_group_mode_var.set("default")
            self.e_group_entry.setEnabled(False)
            self.e_group_entry_var.set("")
        else:
            self.e_group_mode_var.set("group")
            self.e_group_entry.setEnabled(True)

    # -----------------------------------------------------------------
    # Thread-safe UI updates
    # -----------------------------------------------------------------
    def update_status(self, text, bootstyle="info"):
        """อัปเดตข้อความสถานะ (เรียกจาก thread ไหนก็ได้)"""
        self.sig_status.emit(str(text), str(bootstyle))

    def _apply_status_ui(self, text, bootstyle):
        color_map = {
            "info": "#2196F3", "success": "#43A047",
            "warning": "#FF9800", "danger": "#D32F2F",
            "secondary": "#888"}
        c = color_map.get(bootstyle, "#888")
        self.status_label.setText(text)
        self.status_label.setStyleSheet(
            f"color:{c}; font-size:11px;")
        # ระหว่างมี worker ทำงาน event loop เดินอยู่แล้ว
        # การเรียก processEvents ซ้ำเสี่ยง re-entrancy
        if self._worker is None:
            QApplication.processEvents()

    def start_progress(self):
        """แสดง Progress Bar แบบวิ่งไปเรื่อยๆ (ไม่ทราบจำนวนขั้น)"""
        self.sig_progress.emit(-1, 0)

    def set_progress(self, current, total):
        """แสดง Progress Bar แบบทราบเปอร์เซ็นต์จริง"""
        self.sig_progress.emit(int(current), int(total))

    def stop_progress(self):
        """ซ่อน Progress Bar"""
        self.sig_progress.emit(-2, 0)

    def _apply_progress_ui(self, current, total):
        if current == -2:
            self.progress.setVisible(False)
            self.progress.setFormat("")
            return
        if current == -1 or total <= 0:
            self.progress.setRange(0, 0)
        else:
            self.progress.setRange(0, total)
            self.progress.setValue(min(current, total))
        self.progress.setVisible(True)

    # -----------------------------------------------------------------
    # Worker thread plumbing
    # -----------------------------------------------------------------
    def _snapshot_ui_inputs(self):
        """อ่านค่าจาก widget ทั้งหมดเก็บไว้ก่อนเริ่ม thread

        ห้ามอ่าน widget จาก worker thread จึงต้อง snapshot ไว้ล่วงหน้า
        """
        self._ui = {
            'filter_text': self.filter_entry.text().strip(),
            'e_group_mode': self.e_group_mode_var.get(),
            'e_group_expr':
                self.e_group_entry_var.get().strip(),
            't2b_choice': self.t2b_choice_var.get(),
            'summary_only': self.save_all_sheets_var.get(),
        }
        return self._ui

    def _cross_filters(self):
        """รายชื่อ filter ไขว้จากค่าที่ snapshot ไว้

        ห้าม fallback ไปอ่าน widget เพราะอาจถูกเรียกจาก worker thread
        """
        text = self._ui.get('filter_text', '')
        return [f.strip() for f in text.split(',') if f.strip()]

    def _lock_ui(self):
        """จำสถานะปุ่มแล้วปิดทั้งหมด"""
        self._btn_state = {}
        for name in ('btn_start_process',
                     'btn_load_settings_process',
                     'btn_reanalyze', 'btn_analyze_export',
                     'btn_define_labels', 'btn_save_settings'):
            b = getattr(self, name)
            self._btn_state[name] = b.isEnabled()
            b.setEnabled(False)
        self._btn_state['filter_entry'] = \
            self.filter_entry.isEnabled()
        self.filter_entry.setEnabled(False)

    def _unlock_ui(self):
        """คืนสถานะปุ่มตามที่จำไว้"""
        for name, was_enabled in getattr(
                self, '_btn_state', {}).items():
            getattr(self, name).setEnabled(was_enabled)

    def _run_in_thread(self, fn, on_done, on_error=None):
        """รัน fn บน worker thread แล้วเรียก on_done บน main thread"""
        if self._worker is not None:
            self._msg_warn(
                "กำลังประมวลผลอยู่",
                "มีงานกำลังทำงานค้างอยู่ กรุณารอให้เสร็จก่อน")
            return None

        self._lock_ui()
        worker = _Worker(fn, parent=self)
        self._worker = worker

        def _cleanup():
            self._worker = None
            self._unlock_ui()
            worker.deleteLater()

        def _handle_done(result):
            _cleanup()
            on_done(result)

        def _handle_error(message, tb_text):
            _cleanup()
            self.stop_progress()
            self.log_message("")
            self.log_message(f"✗ เกิดข้อผิดพลาด: {message}")
            for line in tb_text.strip().splitlines():
                self.log_message("    " + line)
            self.update_status(f"ผิดพลาด: {message}", "danger")
            if on_error:
                on_error(message, tb_text)
            else:
                self._msg_error(
                    "เกิดข้อผิดพลาด",
                    f"{message}\n\n"
                    "ดูรายละเอียดเพิ่มเติมได้ที่ Log ด้านขวา")

        worker.sig_done.connect(_handle_done)
        worker.sig_error.connect(_handle_error)
        worker.start()
        return worker

    def closeEvent(self, event):
        """กันปิดโปรแกรมกลางคันขณะ worker ยังทำงาน"""
        if self._worker is not None and self._worker.isRunning():
            wait_first = self._styled_confirm(
                "ยังประมวลผลไม่เสร็จ",
                "ยังมีงานประมวลผลค้างอยู่\n"
                "ต้องการรอให้เสร็จก่อนปิดหรือไม่?",
                yes_text="  รอให้เสร็จ  ",
                no_text="  ปิดทันที  ",
                kind='warning')
            if wait_first:
                event.ignore()
                return
            self._worker.wait(3000)
        event.accept()

    def _format_filter_val(self, var_name, value):
        """แปลงค่าตัวเลขเป็น SPSS value label (ถ้ามี)"""
        val_labels = self.spss_value_labels.get(var_name, {})
        label = val_labels.get(value)
        if label is None:
            try:
                label = val_labels.get(float(value))
            except (ValueError, TypeError):
                pass
        if label is None:
            try:
                label = val_labels.get(int(float(value)))
            except (ValueError, TypeError):
                pass
        if label:
            return f"{var_name}={label}"
        return f"{var_name}={value}"

    def _get_var_group_label(self, var_prefix, group_num):
        """ดึง SPSS variable label สำหรับ group ของตัวแปร S/P"""
        SPE_PAT = re.compile(r".*?#(\d+)\$(\d+)$")
        orig_vars = self.vars_to_transform.get(var_prefix, [])
        for var in orig_vars:
            match = SPE_PAT.match(var)
            if match and int(match.group(1)) == group_num:
                lbl = self.spss_variable_labels.get(var)
                if lbl:
                    return lbl
        return f"{var_prefix}_{group_num}"

    def _run_ca_for_subset(self, var_prefix, df_subset):
        """รัน CA บน subset ของข้อมูล คืน list of lists (rows)"""
        if df_subset is None or df_subset.empty:
            return None
        cols = sorted(
            [c for c in df_subset.columns
             if c.startswith(f'{var_prefix}_')
             and 'cor' not in c and 'agree' not in c],
            key=lambda x: int(x.split('_')[1]))
        if not cols or 'Index1' not in df_subset.columns:
            return None

        idx1_vals = sorted(
            df_subset['Index1'].dropna().unique())
        if len(idx1_vals) < 2 or len(cols) < 2:
            return None

        cont = np.zeros((len(cols), len(idx1_vals)))
        for j, iv in enumerate(idx1_vals):
            sub = df_subset[df_subset['Index1'] == iv]
            for i, col in enumerate(cols):
                cont[i, j] = sub[col].mean()

        cont = np.nan_to_num(cont, nan=0.0)
        if cont.sum() == 0:
            return None

        N = cont
        n = N.sum()
        P = N / n
        r = P.sum(axis=1)
        c = P.sum(axis=0)
        r[r == 0] = 1e-10
        c[c == 0] = 1e-10

        Dr = np.diag(1.0 / np.sqrt(r))
        Dc = np.diag(1.0 / np.sqrt(c))
        S = Dr @ (P - np.outer(r, c)) @ Dc
        U, sigma, Vt = np.linalg.svd(S, full_matrices=False)

        n_ax = min(2, len(sigma))
        sv = sigma[:n_ax]
        ev = sv ** 2
        ti = (sigma ** 2).sum()
        cr = ev / ti if ti > 0 else np.zeros(n_ax)

        row_sc = Dr @ U[:, :n_ax] @ np.diag(sv)
        col_sc = Dc @ Vt[:n_ax, :].T @ np.diag(sv)

        row_labels = []
        for cn in cols:
            g = int(cn.split('_')[1])
            row_labels.append(
                self._get_var_group_label(var_prefix, g))

        col_labels = []
        for iv in idx1_vals:
            code = int(iv)
            lbl = self.index1_labels.get(code, str(code))
            col_labels.append(f"({lbl})")

        axes = [f'Axis{i+1}' for i in range(n_ax)]

        rows = []
        rows.append(['Axis information', '', '', ''])
        rows.append(['', 'Singular value',
                      'Eigen value', 'Contribution ratio'])
        for i in range(n_ax):
            rows.append([axes[i], sv[i], ev[i], cr[i]])
        rows.append(['', '', '', ''])
        rows.append(['', '', '', ''])

        rows.append(['Row category score', '', '', ''])
        rh = [''] + axes
        while len(rh) < 4:
            rh.append('')
        rows.append(rh)
        for i, lbl in enumerate(row_labels):
            rw = [lbl]
            for ax in range(n_ax):
                rw.append(row_sc[i, ax])
            while len(rw) < 4:
                rw.append('')
            rows.append(rw)
        rows.append(['', '', '', ''])
        rows.append(['', '', '', ''])

        rows.append(['Column category score', '', '', ''])
        ch = [''] + axes
        while len(ch) < 4:
            ch.append('')
        rows.append(ch)
        for i, lbl in enumerate(col_labels):
            rw = [lbl]
            for ax in range(n_ax):
                rw.append(col_sc[i, ax])
            while len(rw) < 4:
                rw.append('')
            rows.append(rw)

        return rows

    def _get_filter_val_label(self, fvar, val):
        """ดึง SPSS value label ของค่า filter"""
        vl = self.spss_value_labels.get(fvar, {})
        lbl = vl.get(val)
        if lbl is None:
            try:
                lbl = vl.get(int(float(val)))
            except (ValueError, TypeError):
                pass
        if lbl is None:
            try:
                lbl = vl.get(float(val))
            except (ValueError, TypeError):
                pass
        if lbl:
            return str(lbl)
        try:
            return str(int(val))
        except (ValueError, TypeError):
            return str(val)

    def _write_ca_sheet(self, workbook, sheet_name, var_prefix):
        """เขียนผล CA แบบ side-by-side ตาม filter ลง worksheet
        พร้อมสีเหลือง header + เส้นตาราง"""
        df = self.transformed_df
        if df is None:
            return

        ws = workbook.create_sheet(title=sheet_name)

        cross_filters = self._cross_filters()

        # Total ต้องมีบล็อกเดียวเสมอ แล้วต่อด้วยค่าของแต่ละ filter
        # (เดิมใส่ Total ซ้ำทุก filter ทำให้คอลัมน์ซ้ำเมื่อระบุหลายตัว)
        blocks = [('Total', df)]
        for fvar in cross_filters:
            if fvar not in df.columns:
                continue
            uvals = sorted(df[fvar].dropna().unique())
            for val in uvals:
                lbl = self._get_filter_val_label(fvar, val)
                subset = df[df[fvar] == val]
                blocks.append((lbl, subset))

        if not blocks:
            return

        yellow = PatternFill(
            start_color='FFD700', end_color='FFD700',
            fill_type='solid')
        peach = PatternFill(
            start_color='FFDAB9', end_color='FFDAB9',
            fill_type='solid')
        bold_font = Font(bold=True)
        center_al = Alignment(horizontal='center')
        right_al = Alignment(horizontal='right')
        thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'))

        section_headers = {
            'Axis information',
            'Row category score',
            'Column category score'}
        sub_headers = {
            'Singular value', 'Eigen value',
            'Contribution ratio', 'Axis1', 'Axis2'}

        bw = 4
        gap = 1
        col_off = 0

        for title, subset in blocks:
            ca_rows = self._run_ca_for_subset(
                var_prefix, subset)
            if ca_rows is None:
                continue

            cell = ws.cell(
                row=1, column=col_off + 1,
                value=title)
            cell.fill = yellow
            cell.font = bold_font
            cell.border = thin

            for r_idx, row_data in enumerate(ca_rows):
                first_val = row_data[0] if row_data else ''
                is_section = first_val in section_headers
                is_sub = (first_val == '' and any(
                    v in sub_headers
                    for v in row_data if isinstance(v, str)))
                is_blank = all(
                    v == '' for v in row_data)

                is_axis_data = (first_val in
                    ('Axis1', 'Axis2') and not is_sub)

                for c_idx, val in enumerate(row_data):
                    cell = ws.cell(
                        row=r_idx + 2,
                        column=col_off + c_idx + 1)
                    if val != '':
                        cell.value = val
                    if not is_blank:
                        cell.border = thin
                    if is_section:
                        cell.font = bold_font
                        cell.fill = peach
                        if first_val == 'Axis information':
                            cell.alignment = center_al
                    elif is_sub:
                        cell.fill = peach
                        cell.font = bold_font
                        cell.alignment = center_al
                    elif is_axis_data and c_idx == 0:
                        cell.alignment = right_al
                    if isinstance(val, float):
                        cell.number_format = '0.0000000'

            cl = get_column_letter(col_off + 1)
            ws.column_dimensions[cl].width = 40
            for c in range(1, bw):
                cl = get_column_letter(col_off + c + 1)
                ws.column_dimensions[cl].width = 18

            col_off += bw + gap

    def reset_state(self):
        """รีเซ็ตสถานะของโปรแกรมทั้งหมดเพื่อเริ่มใหม่"""
        self.df = None
        self.spss_original_order = []
        self.computed_c_cols = []
        self.c_vars_to_compute = []
        self.vars_to_transform = {}
        self.transformed_df = None
        self.za_cols = []
        self.id_vars = []
        self.last_excel_filepath = None
        self.original_filepath = None
        self.agree_json_filepath_override = None
        self.t2b_choice_var.set("5+4")
        self.index1_labels = {}
        self.filter_labels = {}
        self.spss_value_labels = {}
        self.spss_variable_labels = {}
        self.compute_sav_column_labels = {}
        self.compute_sav_value_labels = {}
        self.sandp_label_overrides = []
        self.agree_summary_cache_df = None
        self.is_reanalyze_mode = False
        self.reanalyze_good_mode = False
        self.good_filter_stats = None
        self.good_filter_full_df = None
        self._good_reference_cache = None
        self._good_k_used = None
        self.agree_json_auto_found = False
        self.e_group_mode_var.set("default")
        self._rb_e_default.setChecked(True)
        self.e_group_entry_var.set("")
        self._analysis_errors = []
        self._analysis_skipped = []
        self._beta_warnings = []
        self._beta_zero_groups = []
        self._beta_abs_used = False
        self._weak_beta_cells = []
        self._weak_model_groups = []
        self._sample_size_approx = False
        self._respondent_key = _UNSET
        self._ui = {}
        self.log_text = None

        self.btn_analyze_export.setEnabled(False)
        self.btn_define_labels.setEnabled(False)
        self.btn_save_settings.setEnabled(False)
        self.filter_entry.setEnabled(False)
        self.filter_entry.clear()
        self.update_status("พร้อมทำงาน", "secondary")

        self._clear_right_panel()
        self.initial_message = QLabel(
            "กรุณากด 'เริ่มกระบวนการ' "
            "เพื่อโหลดไฟล์ SPSS")
        self.initial_message.setStyleSheet(
            "color:#555; font-size:16px;"
            "font-weight:500;")
        self.initial_message.setAlignment(
            Qt.AlignmentFlag.AlignCenter)
        self.right_frame.layout().addWidget(
            self.initial_message)

    # ===================================================================
    # WORKFLOWS
    # ===================================================================
    def start_full_process(self):
        """Workflow 1: เริ่มต้นกระบวนการแบบเลือกตัวแปรเองทั้งหมด"""
        self.reset_state()
        if not self.load_spss_file():
            return
        self.open_c_variable_selector()

    def start_process_with_settings(self):
        """เริ่มต้นกระบวนการโดยโหลดการตั้งค่าและไฟล์ SPSS อัตโนมัติ"""
        self.reset_state()
        if not self._prompt_before_settings():
            self.update_status("ยกเลิกการเลือกไฟล์ตั้งค่า", "warning")
            return

        self.update_status("กำลังรอเลือกไฟล์การตั้งค่า...")
        settings_filepath, _ = QFileDialog.getOpenFileName(
            self, "เลือกไฟล์การตั้งค่า", "",
            "Excel Settings File (*.xlsx)")
        if not settings_filepath:
            self.update_status("ยกเลิกการเลือกไฟล์ตั้งค่า", "warning")
            return

        try:
            spss_filepath_from_settings = self._load_settings_file(
                settings_filepath,
                require_pathfile=True)
        except Exception as e:
            self._msg_error("โหลดไฟล์การตั้งค่าไม่สำเร็จ", str(e))
            self.reset_state()
            return

        self.update_status(f"โหลดตั้งค่าสำเร็จ. กำลังโหลดไฟล์ SPSS...", "info")

        if not self.load_spss_file(filepath=spss_filepath_from_settings):
            self.reset_state()
            return

        self.run_processing_with_loaded_settings()

    def start_reanalyze_process(self):
        """
        Workflow 3: โหลดไฟล์ที่ผ่านการประมวลผลแล้ว (Compute C) เพื่อวิเคราะห์ซ้ำ
        """
        mode = self._ask_reanalyze_mode()
        if mode is None:
            self.update_status("ยกเลิกการวิเคราะห์ซ้ำ", "warning")
            return

        self.reset_state()
        self.is_reanalyze_mode = True
        self.reanalyze_good_mode = (mode == 'auto')

        if not self.load_processed_spss_file():
            return

        if self.reanalyze_good_mode:
            if not self._apply_good_filter_for_reanalyze():
                self.reset_state()
                return

        self._infer_variables_from_transformed_df()

        if not self._prompt_before_settings():
            self.update_status("ยกเลิกการวิเคราะห์ซ้ำ", "warning")
            self.reset_state()
            return

        self.update_status("กำลังรอเลือกไฟล์การตั้งค่า...")
        settings_filepath, _ = QFileDialog.getOpenFileName(
            self, "เลือกไฟล์การตั้งค่า (จำเป็น)",
            os.path.dirname(self.original_filepath or ""),
            "Excel Settings File (*.xlsx)")
        if not settings_filepath:
            self.update_status("ยกเลิกการเลือกไฟล์ตั้งค่า", "warning")
            self._msg_warn(
                "ต้องใช้ไฟล์การตั้งค่า",
                "โหมดวิเคราะห์ซ้ำจำเป็นต้องโหลดไฟล์ Setting")
            self.reset_state()
            return

        try:
            spss_filepath_from_settings = self._load_settings_file(
                settings_filepath,
                require_pathfile=False)
        except Exception as e:
            self._msg_error(
                "โหลดไฟล์การตั้งค่าไม่สำเร็จ", str(e))
            self.reset_state()
            return

        # หา Agree Original JSON ในโฟลเดอร์เดียวกันก่อน
        # ถ้าไม่เจอค่อยให้ผู้ใช้เลือกเอง
        json_filepath = self._auto_find_agree_json()
        self.agree_json_auto_found = bool(json_filepath)

        if json_filepath:
            self.update_status(
                f"พบ Agree JSON อัตโนมัติ: "
                f"{os.path.basename(json_filepath)}", "success")
        else:
            self._styled_notice(
                "ไม่พบไฟล์ Agree Original JSON",
                "หาไฟล์ JSON ในโฟลเดอร์เดียวกับไฟล์ Compute C ไม่เจอ\n\n"
                "กด \"ตกลง\" เพื่อเลือกไฟล์เอง",
                emoji="\U0001F50D", accent="#F57C00")
            json_filepath, _ = QFileDialog.getOpenFileName(
                self,
                "เลือกไฟล์ Agree Original JSON (จำเป็น)",
                os.path.dirname(self.original_filepath or ""),
                "JSON File (*.json)")
            if not json_filepath:
                self.update_status("ยกเลิกการเลือกไฟล์ JSON", "warning")
                self._msg_warn(
                    "ต้องใช้ไฟล์ JSON",
                    "โหมดวิเคราะห์ซ้ำต้องเลือกไฟล์ Agree Original JSON")
                self.reset_state()
                return
        self.agree_json_filepath_override = json_filepath

        self.update_status(
            "โหลดไฟล์ Compute C + Setting สำเร็จ", "success")
        if self._load_agree_summary_cache_from_json():
            cache_cols = len([
                c for c in self.agree_summary_cache_df.columns
                if c.startswith('agree_')
            ]) if self.agree_summary_cache_df is not None else 0
            self.update_status(
                f"Add Agree Original แล้ว ({cache_cols} คอลัมน์)",
                "success")
            self._msg_success(
                "Add Agree Original แล้ว",
                f"โหลดค่า Agree จาก JSON ({cache_cols} คอลัมน์)")
        self.show_reanalyze_ready_panel(self.original_filepath)

        self.btn_analyze_export.setEnabled(True)
        self.btn_define_labels.setEnabled(True)
        self.btn_save_settings.setEnabled(False)
        self.filter_entry.setEnabled(True)

    # ===================================================================
    # RE-ANALYZE : GOOD FILTER
    # ===================================================================
    # ไอคอน + สีหัวข้อ ของกล่องข้อความแต่ละชนิด
    _NOTICE_KINDS = {
        'info':     ("\U0001F4CB", "#1565C0"),
        'success':  ("✅", "#2E7D32"),
        'warning':  ("⚠", "#EF6C00"),
        'error':    ("⛔", "#C62828"),
        'question': ("❓", "#1565C0"),
        'search':   ("\U0001F50D", "#F57C00"),
    }

    def _build_notice_dialog(self, kind, heading, detail, emoji=None):
        """โครงกล่องข้อความสไตล์เดียวกับโปรแกรม (ยังไม่ใส่ปุ่ม)"""
        default_emoji, accent = self._NOTICE_KINDS.get(
            kind, self._NOTICE_KINDS['info'])

        dlg = QDialog(self)
        dlg.setWindowTitle(heading)
        dlg.setModal(True)
        dlg.setMinimumWidth(520)
        dlg.setMaximumWidth(680)
        dlg.setStyleSheet(_DLG_QSS)

        vl = QVBoxLayout(dlg)
        vl.setContentsMargins(24, 20, 24, 18)
        vl.setSpacing(14)

        row = QHBoxLayout()
        row.setSpacing(14)
        icon = QLabel(emoji or default_emoji)
        icon.setFixedWidth(50)
        icon.setStyleSheet(
            "font-size:34px; background:transparent;")
        icon.setAlignment(Qt.AlignmentFlag.AlignTop
                          | Qt.AlignmentFlag.AlignHCenter)
        row.addWidget(icon)

        tv = QVBoxLayout()
        tv.setSpacing(6)
        h = QLabel(str(heading))
        h.setWordWrap(True)
        h.setStyleSheet(
            f"color:{accent}; font-size:16px; font-weight:700;"
            "background:transparent;")
        tv.addWidget(h)
        if detail:
            d = QLabel(str(detail))
            d.setWordWrap(True)
            d.setTextInteractionFlags(
                Qt.TextInteractionFlag.TextSelectableByMouse)
            d.setStyleSheet(
                "color:#555; font-size:12px; background:transparent;")
            tv.addWidget(d)
        tv.addStretch()
        row.addLayout(tv, 1)
        vl.addLayout(row)
        return dlg, vl

    def _styled_notice(self, heading, detail, emoji=None,
                       kind='info', ok_text="  ตกลง  "):
        """กล่องแจ้งเตือนปุ่มเดียว — คืน True เมื่อกดตกลง"""
        dlg, vl = self._build_notice_dialog(
            kind, heading, detail, emoji)

        btn = QPushButton(ok_text)
        btn.setStyleSheet(_BTN_STYLES["danger"])
        btn.setMinimumHeight(38)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.setDefault(True)
        btn.clicked.connect(dlg.accept)
        vl.addWidget(btn)

        dlg.adjustSize()
        self._center_toplevel(dlg)
        return dlg.exec() == QDialog.DialogCode.Accepted

    def _styled_confirm(self, heading, detail,
                        yes_text="  ตกลง  ", no_text="  ยกเลิก  ",
                        kind='question'):
        """กล่องยืนยันสองปุ่ม — คืน True เมื่อกดปุ่มยืนยัน"""
        dlg, vl = self._build_notice_dialog(kind, heading, detail)

        bar = QHBoxLayout()
        bar.setSpacing(10)
        btn_no = QPushButton(no_text)
        btn_no.setStyleSheet(_BTN_STYLES["outline"])
        btn_no.setMinimumHeight(38)
        btn_no.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_no.clicked.connect(dlg.reject)
        bar.addWidget(btn_no, 1)

        btn_yes = QPushButton(yes_text)
        btn_yes.setStyleSheet(_BTN_STYLES["danger"])
        btn_yes.setMinimumHeight(38)
        btn_yes.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_yes.setDefault(True)
        btn_yes.clicked.connect(dlg.accept)
        bar.addWidget(btn_yes, 1)
        vl.addLayout(bar)

        dlg.adjustSize()
        self._center_toplevel(dlg)
        return dlg.exec() == QDialog.DialogCode.Accepted

    # --- ตัวแทน QMessageBox เดิม ให้หน้าตาเป็นชุดเดียวกันทั้งโปรแกรม ---
    def _msg_info(self, heading, detail=""):
        return self._styled_notice(heading, detail, kind='info')

    def _msg_success(self, heading, detail=""):
        return self._styled_notice(heading, detail, kind='success')

    def _msg_warn(self, heading, detail=""):
        return self._styled_notice(heading, detail, kind='warning')

    def _msg_error(self, heading, detail=""):
        return self._styled_notice(heading, detail, kind='error')

    def _msg_ask(self, heading, detail=""):
        return self._styled_confirm(heading, detail)

    def _prompt_before_settings(self):
        """แจ้งให้ทราบก่อนเปิดหน้าต่างเลือกไฟล์ Setting"""
        return self._styled_notice(
            "กรุณาเลือกไฟล์ Setting BS ก่อน",
            "ขั้นตอนต่อไปจะเปิดหน้าต่างให้เลือกไฟล์ตั้งค่า\n"
            "(ปกติชื่อ \"Setting BS.xlsx\" อยู่โฟลเดอร์เดียวกับไฟล์ SPSS)\n\n"
            "กด \"ตกลง\" เพื่อเปิดหน้าต่างเลือกไฟล์",
            emoji="\U0001F4CB")

    def _auto_find_agree_json(self):
        """หาไฟล์ Agree Original JSON ในโฟลเดอร์เดียวกันอัตโนมัติ

        ลำดับการหา:
        1) ชื่อตรงตามที่โปรแกรมบันทึกไว้ (<base> Agree Original.json)
        2) ไฟล์ *Agree Original.json ในโฟลเดอร์เดียวกัน ถ้าเจอไฟล์เดียว
        เจอหลายไฟล์หรือไม่เจอเลย -> คืน None เพื่อให้ผู้ใช้เลือกเอง
        """
        if not self.original_filepath:
            return None

        def _usable(path):
            """ต้องอ่านได้และมีคอลัมน์ agree_ จริง"""
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    payload = json.load(f)
            except Exception:
                return False
            return bool(payload.get('records')
                        and payload.get('agree_columns'))

        base = self._get_base_output_path()
        if base:
            exact = f"{base} Agree Original.json"
            if os.path.exists(exact) and _usable(exact):
                return exact

        folder = os.path.dirname(self.original_filepath)
        if not folder or not os.path.isdir(folder):
            return None
        matches = sorted(
            p for p in glob.glob(
                os.path.join(folder, "*Agree Original.json"))
            if _usable(p))
        if len(matches) == 1:
            return matches[0]
        return None

    def _ask_reanalyze_mode(self):
        """ถามว่าจะใช้ไฟล์ที่ตัดชุดมาแล้ว หรือให้ระบบตัดด้วยเงื่อนไข Good

        คืน 'manual' / 'auto' / None (ยกเลิก)
        """
        dlg = QDialog(self)
        dlg.setWindowTitle("โหมดวิเคราะห์ซ้ำ (Compute C)")
        dlg.setModal(True)
        dlg.setFixedWidth(580)
        dlg.setStyleSheet(_DLG_QSS)

        vl = QVBoxLayout(dlg)
        vl.setContentsMargins(22, 18, 22, 18)
        vl.setSpacing(6)

        title = QLabel("🔄  เลือกวิธีตัดชุดข้อมูล")
        title.setStyleSheet(
            "color:#B71C1C; font-size:17px; font-weight:700;"
            "background:transparent;")
        vl.addWidget(title)

        sub = QLabel(
            "ก่อนวิเคราะห์ซ้ำ ต้องการให้ใช้ข้อมูลชุดไหน?")
        sub.setStyleSheet(
            "color:#777; font-size:11px; background:transparent;"
            "padding-bottom:8px;")
        vl.addWidget(sub)

        choice = {"v": None}

        def make_card(emoji, heading, detail, accent, value):
            card = _ClickableCard()
            card.setStyleSheet(
                "_ClickableCard{"
                "background:#FFFFFF;"
                "border:1.5px solid #E0E0E0;"
                f"border-left:5px solid {accent};"
                "border-radius:10px;}"
                "_ClickableCard:hover{"
                f"background:#FFF7F7; border:1.5px solid {accent};"
                f"border-left:5px solid {accent};}}")

            cl = QVBoxLayout(card)
            cl.setContentsMargins(18, 14, 18, 14)
            cl.setSpacing(4)

            head_lb = QLabel(f"{emoji}   {heading}")
            head_lb.setStyleSheet(
                "font-size:13px; font-weight:700; color:#333;"
                "background:transparent; border:none;")
            detail_lb = QLabel(detail)
            detail_lb.setWordWrap(True)
            detail_lb.setStyleSheet(
                "font-size:11px; color:#777;"
                "background:transparent; border:none;")
            for lb in (head_lb, detail_lb):
                lb.setAttribute(
                    Qt.WidgetAttribute.WA_TransparentForMouseEvents)
                cl.addWidget(lb)

            def _pick():
                choice["v"] = value
                dlg.accept()

            card.clicked.connect(_pick)
            vl.addWidget(card)
            return card

        make_card(
            "📄", "1.  ใช้ไฟล์ที่ตัดชุดมาเองแล้ว",
            "วิเคราะห์ทุกแถวในไฟล์ตามปกติ (พฤติกรรมเดิม)",
            "#1976D2", "manual")

        make_card(
            "⚙️", "2.  ให้ระบบตัดชุดให้ (เงื่อนไข Good)",
            "สร้างตัวแปร Good แล้วใช้เฉพาะ Good = 1 มาวิเคราะห์",
            "#C62828", "auto")

        note = QLabel(
            "เงื่อนไข Good ตรวจความสอดคล้องระหว่าง A กับ "
            "N_S / N_P / N_C / N_E\n"
            "แถวที่ไม่สอดคล้อง (รวมถึง A = 0) จะถูกกำหนดเป็น Good = 2 "
            "และตัดออก")
        note.setWordWrap(True)
        note.setStyleSheet(
            "color:#888; font-size:10px; background:#F5F5F5;"
            "border-radius:6px; padding:9px 11px; margin-top:6px;")
        vl.addWidget(note)

        btn_cancel = QPushButton("ยกเลิก")
        btn_cancel.setStyleSheet(_BTN_STYLES["outline"])
        btn_cancel.setMinimumHeight(34)
        btn_cancel.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_cancel.clicked.connect(dlg.reject)
        vl.addWidget(btn_cancel)

        dlg.adjustSize()
        self._center_toplevel(dlg)
        dlg.exec()
        return choice["v"]

    _GOOD_METRICS = ('N_S', 'N_P', 'N_C', 'N_E')
    # ตัวคูณความเข้ม: เบี่ยงจาก "ค่าที่ควรเป็นตาม A" เกิน k เท่าของ
    # ความเบี่ยงเบนปกติ -> ตัด ปรับตัวนี้ตัวเดียวก็พอ (ไม่ต้องตั้งเกณฑ์
    # รายตัวชี้วัด/รายงานใหม่ทุกครั้ง) เพราะ median/spread คำนวณจาก
    # ข้อมูลจริงของแต่ละงานเอง สเกลเปลี่ยน (จำนวนข้อ, binary/likert)
    # เกณฑ์ก็ปรับตาม k=4.0 ค่อนข้างหลวม (เก็บคนตอบปกติไว้เกือบหมด)
    # จับเฉพาะคนที่ตอบขัดแย้งชัดเจน ใช้ยืนยันกับข้อมูลจริงแล้วว่า
    # กลุ่มย่อยยังคำนวณได้ครบ ไม่พังเหมือนเกณฑ์ผูกค่าคงที่แบบเดิม
    _GOOD_DEFAULT_K = 4.0

    @staticmethod
    def _good_spread(residual):
        """ความกว้างของการเบี่ยงเบนแบบทนค่าผิดปกติ (MAD -> IQR -> SD)

        คืน None ถ้าคำนวณไม่ได้ (ข้อมูลไม่มีความแปรปรวนเลย)
        """
        r = pd.Series(residual).dropna()
        if r.empty:
            return None
        mad = (r - r.median()).abs().median() * 1.4826
        iqr = (r.quantile(.75) - r.quantile(.25)) / 1.349
        for v in (mad, iqr, r.std()):
            if v and v > 1e-9:
                return float(v)
        return None

    def _good_reference(self, df):
        """คำนวณ (ค่าที่ควรเป็นต่อระดับ A, ความเบี่ยงเบนปกติ) ของแต่ละ
        ตัวชี้วัด จากข้อมูลจริงในไฟล์นี้ — ทำให้เกณฑ์ปรับตามสเกล/
        จำนวนข้อของแต่ละงานเองโดยอัตโนมัติ ไม่ต้องตั้งเลขต่องาน

        คืน dict: metric -> (median_series ต่อค่า A, spread หรือ None)
        """
        a = pd.to_numeric(df['A'], errors='coerce')
        ref = {}
        for m in self._GOOD_METRICS:
            if m not in df.columns:
                continue
            x = pd.to_numeric(df[m], errors='coerce')
            med = x.groupby(a).median()
            resid = x - a.map(med)
            ref[m] = (med, self._good_spread(resid))
        return ref

    def _compute_good_filter(self, df, k=None):
        """สร้าง Series ของตัวแปร Good (1 = ผ่าน, 2 = ตัดออก)

        เงื่อนไข:
          - A = 0 (ปฏิเสธแบรนด์) ตัดเสมอ — ตามกฎธุรกิจเดิม เพราะคนกลุ่มนี้
            เอาไปวางบนบันได consistency ของคนที่ "รู้จักแบรนด์" ไม่ได้
          - ตัวชี้วัดใดเบี่ยงจาก "ค่าที่ควรเป็นตามระดับ A" เกิน k เท่าของ
            ความเบี่ยงเบนปกติ -> ตัด (คนตอบไม่สอดคล้องกับพฤติกรรมจริง)
        """
        required = ('A',) + self._GOOD_METRICS
        missing = [c for c in required if c not in df.columns]
        if missing:
            raise RuntimeError(
                "ไม่สามารถสร้างตัวแปร Good ได้ "
                f"เนื่องจากไม่พบคอลัมน์: {', '.join(missing)}\n"
                "กรุณาตรวจสอบว่าเลือกไฟล์ที่ผ่านการ Compute C แล้ว")

        if k is None:
            k = self._GOOD_DEFAULT_K

        a = pd.to_numeric(df['A'], errors='coerce')
        bad = (a == 0)   # ปฏิเสธแบรนด์ -> ตัดเสมอ

        ref = self._good_reference(df)
        # เก็บไว้ให้ตาราง Cross ใช้เกณฑ์ชุดเดียวกับที่ตัดจริง
        self._good_reference_cache = ref
        self._good_k_used = k

        for m, (med, spread) in ref.items():
            if spread is None:
                continue
            x = pd.to_numeric(df[m], errors='coerce')
            resid = x - a.map(med)
            bad = bad | (resid.abs() > k * spread).fillna(False)

        return pd.Series(
            np.where(bad.to_numpy(), 2.0, 1.0),
            index=df.index, dtype=float)

    def _metric_rule_ok(self, metric, a_value, x_value, ref=None, k=None):
        """ค่า (A, X) คู่นี้ผ่านเงื่อนไข Good ของตัวชี้วัดเดียวหรือไม่

        ใช้เกณฑ์ชุดเดียวกับ _compute_good_filter (median/spread) แต่ดู
        ทีละตัวชี้วัด (แถวจริงอาจยังถูกตัดจากตัวชี้วัดอื่นได้)
        """
        if a_value is None or x_value is None:
            return True
        try:
            a_value = float(a_value)
            x_value = float(x_value)
        except (TypeError, ValueError):
            return True
        if np.isnan(a_value) or np.isnan(x_value):
            return True
        if a_value == 0:
            return False

        if ref is None:
            ref = getattr(self, '_good_reference_cache', None) or {}
        if metric not in ref:
            return True
        med, spread = ref[metric]
        if spread is None:
            return True
        expected = med.get(a_value)
        if expected is None or (
                isinstance(expected, float) and np.isnan(expected)):
            return True
        if k is None:
            # getattr(..., default) ไม่ช่วยตรงนี้ เพราะ _good_k_used
            # ถูก init เป็น None เสมอ (attribute มีอยู่แล้ว ไม่ใช่ไม่มี)
            # ต้องเช็ค None explicit ถึงจะ fallback ไป default ได้จริง
            k = getattr(self, '_good_k_used', None)
            if k is None:
                k = self._GOOD_DEFAULT_K
        return abs(x_value - expected) <= k * spread

    def _metric_rule_pass_series(self, metric, a_series, x_series,
                                 ref=None, k=None):
        """เวกเตอร์: แต่ละแถวผ่านเงื่อนไขของตัวชี้วัดนี้หรือไม่"""
        bad = (a_series == 0)
        if ref is None:
            ref = getattr(self, '_good_reference_cache', None) or {}
        if metric in ref:
            med, spread = ref[metric]
            if spread is not None:
                if k is None:
                    k = getattr(self, '_good_k_used', None)
                    if k is None:
                        k = self._GOOD_DEFAULT_K
                resid = x_series - a_series.map(med)
                bad = bad | (resid.abs() > k * spread).fillna(False)
        return ~bad

    def _build_good_crosstab(self, metric):
        """ตาราง Cross: แถว = ค่า metric (ปัดทศนิยม 2), คอลัมน์ = A

        คืน (total, passed, a_values, x_values, ref) หรือ None ถ้าทำไม่ได้
        นับจากข้อมูลเต็มก่อนตัด เพื่อให้เห็นว่าเงื่อนไขตัดตรงไหนไปบ้าง
        คำนวณ ref (median/spread) สดจาก good_filter_full_df เสมอ
        เพื่อให้ตรงกับข้อมูลที่กำลังแสดง ไม่พึ่งพา cache จากรอบก่อน
        """
        df = self.good_filter_full_df
        if df is None or df.empty:
            return None
        if 'A' not in df.columns or metric not in df.columns:
            return None

        ref = self._good_reference(df)
        a = pd.to_numeric(df['A'], errors='coerce')
        x = pd.to_numeric(df[metric], errors='coerce')
        ok = self._metric_rule_pass_series(metric, a, x, ref=ref)

        sub = pd.DataFrame({
            'A': a, 'X': x.round(2), 'ok': ok.astype(int),
        }).dropna(subset=['A', 'X'])
        if sub.empty:
            return None

        total = pd.crosstab(sub['X'], sub['A'])
        passed = pd.crosstab(
            sub['X'], sub['A'],
            values=sub['ok'], aggfunc='sum')
        passed = passed.reindex(
            index=total.index, columns=total.columns).fillna(0)
        return (total, passed,
                list(total.columns), list(total.index), ref)

    def _build_crosstab_table(self, metric):
        """สร้าง QTableWidget ของตาราง Cross A x metric"""
        built = self._build_good_crosstab(metric)
        if built is None:
            lb = QLabel(f"ไม่มีข้อมูลสำหรับ {metric}")
            lb.setStyleSheet("color:#888; padding:20px;")
            return lb

        total, passed, a_vals, x_vals, ref = built
        n_rows, n_cols = len(x_vals), len(a_vals)

        tw = QTableWidget(n_rows + 1, n_cols + 1)
        tw.setHorizontalHeaderLabels(
            [(str(int(a)) if float(a).is_integer() else f"{a:g}")
             for a in a_vals] + ["Grand Total"])
        tw.setVerticalHeaderLabels(
            [f"{x:.2f}" for x in x_vals] + ["Grand Total"])
        tw.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        tw.setAlternatingRowColors(False)

        green = QColor('#A9D08E')
        yellow = QColor('#FFE699')
        mixed = QColor('#F4B183')
        totals_bg = QColor('#EDEDED')

        for r, xv in enumerate(x_vals):
            for c, av in enumerate(a_vals):
                cnt = int(total.iat[r, c])
                n_ok = int(passed.iat[r, c])
                rule_ok = self._metric_rule_ok(metric, av, xv, ref=ref)

                item = QTableWidgetItem(str(cnt) if cnt else "")
                item.setTextAlignment(
                    Qt.AlignmentFlag.AlignRight
                    | Qt.AlignmentFlag.AlignVCenter)

                # ปกติสีตามกฎ แต่ถ้าแถวจริงในช่องไม่ตรงกับกฎ
                # (เกิดจากการปัดทศนิยม) ให้เป็นสีส้มเตือน
                expected_ok = cnt if rule_ok else 0
                if cnt and n_ok != expected_ok:
                    item.setBackground(mixed)
                    item.setToolTip(
                        f"ผ่าน {n_ok} / {cnt} แถว "
                        "(ค่าจริงคร่อมเกณฑ์หลังปัดทศนิยม)")
                else:
                    item.setBackground(green if rule_ok else yellow)
                tw.setItem(r, c, item)

            gt = QTableWidgetItem(str(int(total.iloc[r].sum())))
            gt.setBackground(totals_bg)
            gt.setFont(QFontObj('Segoe UI', 9, QFontObj.Weight.Bold))
            gt.setTextAlignment(
                Qt.AlignmentFlag.AlignRight
                | Qt.AlignmentFlag.AlignVCenter)
            tw.setItem(r, n_cols, gt)

        for c in range(n_cols):
            gt = QTableWidgetItem(str(int(total.iloc[:, c].sum())))
            gt.setBackground(totals_bg)
            gt.setFont(QFontObj('Segoe UI', 9, QFontObj.Weight.Bold))
            gt.setTextAlignment(
                Qt.AlignmentFlag.AlignRight
                | Qt.AlignmentFlag.AlignVCenter)
            tw.setItem(n_rows, c, gt)

        grand = QTableWidgetItem(str(int(total.to_numpy().sum())))
        grand.setBackground(totals_bg)
        grand.setFont(QFontObj('Segoe UI', 9, QFontObj.Weight.Bold))
        grand.setTextAlignment(
            Qt.AlignmentFlag.AlignRight
            | Qt.AlignmentFlag.AlignVCenter)
        tw.setItem(n_rows, n_cols, grand)

        self._fit_table_to_content(tw)
        return tw

    def _fit_table_to_content(self, tw, row_height=22):
        """ขยายตารางให้พอดีเนื้อหา จะได้ไม่มี scrollbar ในตัวตารางเอง

        - คอลัมน์ยืดเต็มความกว้างที่มี (คอลัมน์ Grand Total จึงไม่โดนตัด)
        - ล็อกความสูงเท่าจำนวนแถวจริง (เห็นครบทุกแถว)
        ให้หน้าเพจด้านนอกเป็นตัวเลื่อนแทน
        """
        tw.setVerticalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        tw.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        header = tw.horizontalHeader()
        header.setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        header.setMinimumSectionSize(38)

        tw.verticalHeader().setDefaultSectionSize(row_height)
        for r in range(tw.rowCount()):
            tw.setRowHeight(r, row_height)

        # ล็อกความสูงหัวตารางไว้ จะได้คำนวณความสูงรวมได้แน่นอน
        # (ถ้าอ่านจาก sizeHint ตอนยังไม่แสดงผล จะได้ค่าน้อยกว่าจริง)
        header_h = max(header.sizeHint().height(), 28)
        header.setFixedHeight(header_h)

        total_h = (header_h
                   + row_height * tw.rowCount()
                   + 2 * tw.frameWidth() + 2)
        tw.setFixedHeight(total_h)
        tw.setSizePolicy(QSizePolicy.Policy.Expanding,
                         QSizePolicy.Policy.Fixed)
        return total_h

    def _save_good_sav(self, df_with_good):
        """เขียนตัวแปร Good กลับลงไฟล์ SPSS ที่โหลดเข้ามา (ทับไฟล์เดิม)

        เขียนลงไฟล์ชั่วคราวก่อนแล้วค่อย os.replace ทับ เพื่อไม่ให้
        ไฟล์ต้นทางเสียหายถ้าเขียนไม่สำเร็จกลางคัน
        """
        target = self.original_filepath
        if not target:
            return None

        folder = os.path.dirname(target) or "."
        tmp_path = os.path.join(
            folder, f"~good_tmp_{os.path.basename(target)}")
        try:
            column_labels = [
                'Good filter (1=Good, 2=Not Good)' if col == 'Good'
                else str(self.spss_variable_labels.get(col, col))
                for col in df_with_good.columns
            ]
            value_labels = {'Good': {1.0: 'Good', 2.0: 'Not Good'}}
            for col, labels in self.spss_value_labels.items():
                if col in df_with_good.columns \
                        and isinstance(labels, dict) and labels:
                    value_labels[col] = labels

            pyreadstat.write_sav(
                df_with_good, tmp_path,
                column_labels=column_labels,
                variable_value_labels=value_labels)
            os.replace(tmp_path, target)
            return target
        except Exception as e:
            # เขียนไม่สำเร็จ -> ทิ้งไฟล์ชั่วคราว ไฟล์เดิมยังอยู่ครบ
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except OSError:
                pass
            self.update_status(
                f"บันทึกตัวแปร Good ลงไฟล์ไม่สำเร็จ: {e}", "warning")
            return None

    def _apply_good_filter_for_reanalyze(self):
        """สร้างตัวแปร Good แล้วคัดเหลือเฉพาะ Good = 1 ไว้วิเคราะห์"""
        if self.transformed_df is None or self.transformed_df.empty:
            self._msg_error(
                "ไม่มีข้อมูล", "ไม่มีข้อมูลสำหรับสร้างตัวแปร Good")
            return False

        self.update_status("กำลังสร้างตัวแปร Good...")
        try:
            good = self._compute_good_filter(self.transformed_df)
        except Exception as e:
            self._msg_error("ผิดพลาด", str(e))
            return False

        df = self.transformed_df.copy()
        df['Good'] = good

        total = len(df)
        n_good = int((df['Good'] == 1).sum())
        n_drop = total - n_good

        saved_path = self._save_good_sav(df)

        if n_good == 0:
            self._msg_error(
                "ไม่เหลือข้อมูล",
                "เงื่อนไข Good ตัดข้อมูลออกทั้งหมด (Good = 1 เหลือ 0 แถว)\n\n"
                "กรุณาตรวจสอบข้อมูล หรือเลือกโหมด "
                "'ใช้ไฟล์ที่ตัดชุดมาเองแล้ว' แทน")
            return False

        # เก็บข้อมูลเต็ม (ก่อนตัด) ไว้ทำตาราง Cross ให้ตรวจสอบ
        self.good_filter_full_df = df
        self.transformed_df = df[df['Good'] == 1].reset_index(drop=True)
        self._respondent_key = _UNSET   # ข้อมูลเปลี่ยน ต้องตรวจรหัสผู้ตอบใหม่
        self.good_filter_stats = {
            'total': total,
            'good': n_good,
            'dropped': n_drop,
            'good_pct': (n_good / total * 100) if total else 0.0,
            'saved_path': saved_path,
        }

        detail = (
            f"ข้อมูลทั้งหมด : {total:,} แถว\n"
            f"Good = 1 (ใช้วิเคราะห์) : {n_good:,} แถว "
            f"({self.good_filter_stats['good_pct']:.1f}%)\n"
            f"Good = 2 (ตัดออก) : {n_drop:,} แถว")
        if saved_path:
            detail += (
                "\n\nบันทึกตัวแปร Good ลงไฟล์ SPSS ที่โหลดเข้ามาแล้ว:\n"
                f"{saved_path}")
        else:
            detail += ("\n\n⚠ บันทึกตัวแปร Good ลงไฟล์ไม่สำเร็จ "
                       "(ไฟล์เดิมยังอยู่ครบ)")
        self._msg_success("สร้างตัวแปร Good เรียบร้อยแล้ว", detail)

        self.update_status(
            f"ตัดชุดด้วยเงื่อนไข Good: เหลือ {n_good:,}/{total:,} แถว",
            "success")
        return True

    def load_spss_file(self, filepath=None):
        """โหลดไฟล์ SPSS ดั้งเดิม โดยรับ Path หรือเปิด Dialog"""
        if filepath is None:
            self.update_status("กำลังรอเลือกไฟล์ SPSS...")
            filepath, _ = QFileDialog.getOpenFileName(
                self, "เลือกไฟล์ SPSS", "",
                "SPSS Data File (*.sav)")
            if not filepath:
                self.update_status("ยกเลิกการเลือกไฟล์", "warning")
                return False

        if not os.path.exists(filepath):
            self.update_status("ไฟล์ SPSS ไม่พบ", "danger")
            self._msg_error("ไม่พบไฟล์ที่ระบุ", filepath)
            return False

        self.start_progress()
        self.update_status(f"กำลังโหลด: {os.path.basename(filepath)}...")
        try:
            self.df, meta = pyreadstat.read_sav(filepath)
            self.original_filepath = filepath
            self.spss_original_order = meta.column_names
            var_labels, value_labels = \
                self._extract_spss_metadata(meta)
            self.spss_variable_labels = var_labels
            self.spss_value_labels = value_labels
            self.df = self.df[self.spss_original_order]
            self.update_status(f"โหลดไฟล์สำเร็จ! {len(self.df)} แถว", "success")
            self.stop_progress()
            return True
        except Exception as e:
            self.update_status("โหลดไฟล์ผิดพลาด", "danger")
            self._msg_error("โหลดไฟล์ไม่สำเร็จ", str(e))
            self.stop_progress()
            self.reset_state()
            return False

    def load_processed_spss_file(self):
        """โหลดไฟล์ SPSS ที่ผ่านการประมวลผลแล้ว (* Compute C.sav)"""
        self.update_status("กำลังรอเลือกไฟล์ SPSS ที่ประมวลผลแล้ว...")
        filepath, _ = QFileDialog.getOpenFileName(
            self,
            "เลือกไฟล์ SPSS ที่ผ่านการ Compute C แล้ว",
            "", "SPSS Data File (*.sav)")
        if not filepath:
            self.update_status("ยกเลิกการเลือกไฟล์", "warning")
            return False
            
        self.start_progress()
        self.update_status(f"กำลังโหลด: {os.path.basename(filepath)}...")
        try:
            self.transformed_df, meta_cc = pyreadstat.read_sav(filepath)
            var_labels, value_labels = \
                self._extract_spss_metadata(meta_cc)
            self.spss_variable_labels = var_labels
            self.spss_value_labels = value_labels

            # โหมด Re-analyze ใช้ไฟล์ Compute C เป็นฐานหลัก
            # ไม่พึ่งพาไฟล์ SPSS Original
            self.original_filepath = filepath
            
            self.update_status(f"โหลดไฟล์สำเร็จ! {len(self.transformed_df)} แถว", "success")
            self.stop_progress()
            return True
        except Exception as e:
            self.update_status("โหลดไฟล์ผิดพลาด", "danger")
            self._msg_error(
                "โหลดไฟล์ที่ประมวลผลแล้วไม่สำเร็จ", str(e))
            self.stop_progress()
            self.reset_state()
            return False
    
    def _infer_variables_from_transformed_df(self):
        """
        พยายามสร้าง state ของตัวแปร (เช่น id_vars, vars_to_transform)
        จากคอลัมน์ที่มีอยู่ใน DataFrame ที่โหลดเข้ามา เพื่อให้ส่วนแสดงผลทำงานได้
        """
        if self.transformed_df is None:
            return

        self.vars_to_transform = {'A':[], 'S':[], 'P':[], 'E':[], 'AgreeS':[], 'AgreeP':[]}
        self.c_vars_to_compute = []
        self.computed_c_cols = []
        self.id_vars = []

        known_patterns = [
            re.compile(r'^(S|P|E|C)_\d+$'),
            re.compile(r'^N_(S|P|E|C)$'),
            re.compile(r'^(A|ZA|Index1)$')
        ]

        for col in self.transformed_df.columns:
            if col.startswith('S_'): self.vars_to_transform['S'].append(col)
            elif col.startswith('P_'): self.vars_to_transform['P'].append(col)
            elif col.startswith('E_'): self.vars_to_transform['E'].append(col)
            elif col.startswith('C_'): self.computed_c_cols.append(col)
            elif col == 'A': self.vars_to_transform['A'].append(col)
        
        for col in self.transformed_df.columns:
            is_known = False
            for pattern in known_patterns:
                if pattern.match(col):
                    is_known = True
                    break
            if not is_known:
                self.id_vars.append(col)
        
        print("Infered ID Vars:", self.id_vars)
        print("Infered S Vars:", self.vars_to_transform['S'])
        print("Infered C Vars (Computed):", self.computed_c_cols)

    def _load_settings_file(
            self, settings_filepath,
            require_pathfile=True):
        """โหลดไฟล์การตั้งค่าและอัปเดต state ภายในโปรแกรม"""
        def _clean_list_from_col(df_obj, col_name):
            if col_name not in df_obj.columns:
                return []
            vals = []
            for raw in df_obj[col_name].dropna().tolist():
                txt = str(raw).strip()
                if txt:
                    vals.append(txt)
            return vals

        self.update_status("กำลังโหลดการตั้งค่า...")
        xls = pd.ExcelFile(settings_filepath)

        if 'Settings' not in xls.sheet_names:
            raise ValueError("ไม่พบชีท 'Settings' ในไฟล์การตั้งค่า")

        settings_df = pd.read_excel(xls, sheet_name='Settings')

        spss_filepath_from_settings = None
        if 'PathFile' in settings_df.columns and \
                not settings_df.empty and \
                not pd.isna(settings_df['PathFile'].iloc[0]):
            spss_filepath_from_settings = str(
                settings_df['PathFile'].iloc[0]).strip()
        elif require_pathfile:
            raise ValueError("ไม่พบ PathFile ในไฟล์การตั้งค่า")

        self.filter_entry.clear()
        if 'Filter_Var' in settings_df.columns:
            filter_values = settings_df[
                'Filter_Var'].dropna().tolist()
            filter_values = [
                str(v).strip() for v in filter_values
                if str(v).strip()]
            if filter_values:
                self.filter_entry.setEnabled(True)
                self.filter_entry.setText(
                    ', '.join(filter_values))

        if 'T2B_Choice' in settings_df.columns and \
                not settings_df.empty and \
                not pd.isna(settings_df[
                    'T2B_Choice'].iloc[0]):
            self.t2b_choice_var.set(str(
                settings_df['T2B_Choice'].iloc[0]))

        e_group_val = ""
        if 'E_Group' in settings_df.columns and \
                not settings_df.empty and \
                not pd.isna(settings_df[
                    'E_Group'].iloc[0]):
            e_group_val = str(
                settings_df['E_Group'].iloc[0]).strip()

        if e_group_val.lower() == 'default' \
                or e_group_val == '':
            self.e_group_mode_var.set("default")
            self.e_group_entry_var.set("")
            self._rb_e_default.setChecked(True)
            self.e_group_entry.setEnabled(False)
        else:
            self.e_group_mode_var.set("group")
            self.e_group_entry_var.set(e_group_val)
            self._rb_e_group.setChecked(True)
            self.e_group_entry.setEnabled(True)

        self.c_vars_to_compute = (
            _clean_list_from_col(settings_df, 'C'))
        self.vars_to_transform = {}
        for key in ['A', 'S', 'P', 'E', 'AgreeS',
                    'AgreeP']:
            self.vars_to_transform[key] = _clean_list_from_col(
                settings_df, key)

        self.filter_labels = {}
        self.sandp_label_overrides = []
        if 'Label' in xls.sheet_names:
            labels_df = pd.read_excel(
                xls, sheet_name='Label')

            if 'Index1_Code' in labels_df.columns \
                    and 'Index1_Label' in labels_df.columns:
                index1_labels_df = labels_df[
                    ['Index1_Code', 'Index1_Label']
                ].dropna()
                self.index1_labels = dict(zip(
                    index1_labels_df[
                        'Index1_Code'].astype(int),
                    index1_labels_df['Index1_Label']))

            filter_text_for_label = \
                self.filter_entry.text().strip()
            filter_vars_list = [
                f.strip() for f in
                filter_text_for_label.split(',')
                if f.strip()]
            filter_var = (
                filter_vars_list[0]
                if filter_vars_list else '')
            if filter_var and \
                    'Filter_Code' in labels_df.columns \
                    and 'Filter_Label' in labels_df.columns:
                self.filter_labels['var_name'] = \
                    filter_var
                filter_labels_df = labels_df[
                    ['Filter_Code', 'Filter_Label']
                ].dropna()
                self.filter_labels['labels'] = dict(zip(
                    filter_labels_df[
                        'Filter_Code'].astype(int),
                    filter_labels_df['Filter_Label']))

            if 'SandP_Label' in labels_df.columns:
                self.sandp_label_overrides = [
                    str(v).strip()
                    for v in labels_df['SandP_Label'].tolist()
                    if pd.notna(v) and str(v).strip()
                ]

        return spss_filepath_from_settings

    def _try_load_original_spss_for_reanalyze(
            self, spss_filepath_from_settings=None):
        """พยายามโหลดไฟล์ต้นฉบับเพื่อรองรับการคำนวณ T2B"""
        loaded = False
        candidates = []
        if spss_filepath_from_settings:
            candidates.append(spss_filepath_from_settings)
        if self.original_filepath:
            candidates.append(self.original_filepath + ".sav")

        for cand in candidates:
            if not cand or not os.path.exists(cand):
                continue
            try:
                self.df, meta_orig = pyreadstat.read_sav(cand)
                self.spss_original_order = meta_orig.column_names
                var_labels, value_labels = \
                    self._extract_spss_metadata(meta_orig)
                self.spss_value_labels.update(
                    value_labels)
                self.spss_variable_labels.update(
                    var_labels)
                loaded = True
                return
            except Exception:
                continue
        if not loaded:
            self.update_status(
                "ไม่พบไฟล์ SPSS ต้นฉบับสำหรับคำนวณ T2B",
                "warning")

    def _get_base_output_path(self):
        """หา base path กลางจากไฟล์ต้นทางปัจจุบัน"""
        if not self.original_filepath:
            return None

        base, ext = os.path.splitext(self.original_filepath)
        if ext.lower() == '.sav' and base.endswith(" Compute C"):
            base = base[:-10]
        return base

    def _get_bs_output_filepath(self):
        """หา path ของไฟล์ Excel output หลักจากไฟล์ต้นทางปัจจุบัน"""
        base = self._get_base_output_path()
        if not base:
            return None
        return f"{base} BS Output.xlsx"

    def _get_agree_json_filepath(self):
        """หา path ของไฟล์ JSON ที่เก็บ Agree Original"""
        if self.agree_json_filepath_override and \
                os.path.exists(self.agree_json_filepath_override):
            return self.agree_json_filepath_override
        base = self._get_base_output_path()
        if not base:
            return None
        return f"{base} Agree Original.json"

    def _save_agree_summary_to_json(self, summary_df):
        """บันทึกค่า agree_* จาก Summary ลง JSON"""
        if summary_df is None or summary_df.empty:
            return False

        agree_cols = [
            c for c in summary_df.columns
            if str(c).startswith('agree_')]
        if not agree_cols:
            return False

        key_candidates = ['Filter', 'Index1']
        key_cols = [
            c for c in key_candidates
            if c in summary_df.columns]
        if not key_cols:
            return False

        payload_df = summary_df[
            key_cols + agree_cols
        ].reset_index(drop=True)
        payload_df = payload_df.where(
            pd.notna(payload_df), None)

        json_filepath = self._get_agree_json_filepath()
        if not json_filepath:
            return False

        # แปลง NaN เป็น None ตอนสร้าง record เพื่อให้เขียนเป็น null
        # (json.dump จะเขียน NaN ซึ่งไม่ใช่ JSON มาตรฐาน เครื่องมืออื่นอ่านไม่ได้)
        records = []
        for rec in payload_df.to_dict(orient='records'):
            records.append({
                k: (None if (v is not None and not isinstance(v, str)
                             and pd.isna(v)) else v)
                for k, v in rec.items()
            })

        payload = {
            'key_columns': key_cols,
            'agree_columns': agree_cols,
            'records': records
        }
        with open(
                json_filepath,
                'w',
                encoding='utf-8') as f:
            json.dump(
                payload,
                f,
                ensure_ascii=False,
                indent=2,
                allow_nan=False)

        self.agree_summary_cache_df = payload_df
        return True

    def _load_agree_summary_cache_from_json(self):
        """โหลดค่า agree_* จาก JSON มาไว้ใน cache"""
        self.agree_summary_cache_df = None
        json_filepath = self._get_agree_json_filepath()
        if not json_filepath or not os.path.exists(json_filepath):
            return False

        try:
            with open(
                    json_filepath,
                    'r',
                    encoding='utf-8') as f:
                payload = json.load(f)
        except Exception:
            return False

        records = payload.get('records', [])
        key_cols = payload.get('key_columns', [])
        agree_cols = payload.get('agree_columns', [])
        if not records or not agree_cols:
            return False

        if not key_cols:
            sample = records[0] if records else {}
            key_cols = [
                c for c in ['Filter', 'Index1']
                if c in sample]
        if not key_cols:
            return False

        cache_df = pd.DataFrame(records)
        required_cols = key_cols + agree_cols
        if not all(col in cache_df.columns for col in required_cols):
            return False

        # โหมดวางทั้งชุด: ต้องคงลำดับและจำนวนแถวเดิมจาก JSON
        self.agree_summary_cache_df = cache_df[
            required_cols
        ].reset_index(drop=True)
        return not self.agree_summary_cache_df.empty

    def _apply_agree_summary_cache(self, summary_df):
        """เติม/แทนค่า agree_* ใน summary จาก cache JSON"""
        if summary_df is None or summary_df.empty:
            return summary_df, 0
        if self.agree_summary_cache_df is None \
                or self.agree_summary_cache_df.empty:
            return summary_df, 0

        def _norm_filter(v):
            txt = str(v).strip().replace(" ", "").lower()
            # normalize numeric tokens in filter text (e.g. 1.0 -> 1)
            txt = re.sub(r'(?<=\=)(-?\d+)\.0\b', r'\1', txt)
            txt = re.sub(r'^(-?\d+)\.0$', r'\1', txt)
            return txt

        if 'Filter' not in summary_df.columns \
                or 'Filter' not in self.agree_summary_cache_df.columns:
            return summary_df, 0

        cache_agree_cols = [
            c for c in self.agree_summary_cache_df.columns
            if c.startswith('agree_')]
        if not cache_agree_cols:
            return summary_df, 0

        left = summary_df.copy()
        right = self.agree_summary_cache_df.copy()
        left['__fkey'] = left['Filter'].map(_norm_filter)
        right['__fkey'] = right['Filter'].map(_norm_filter)
        cache_base = right[
            ['__fkey'] + cache_agree_cols
        ].drop_duplicates(subset=['__fkey'])
        merged = left.merge(
            cache_base,
            on='__fkey',
            how='left',
            suffixes=('', '__json'))

        applied_count = 0
        for col in cache_agree_cols:
            json_col = f"{col}__json"
            if json_col not in merged.columns:
                continue
            if col not in merged.columns:
                merged[col] = np.nan
            before_notna = merged[col].notna()
            merged[col] = np.where(
                merged[json_col].notna(),
                merged[json_col],
                merged[col])
            applied_count += int(
                (merged[col].notna() & ~before_notna).sum())
            merged.drop(columns=[json_col], inplace=True)

        cleanup_cols = [
            c for c in ['__fkey', '__ikey']
            if c in merged.columns]
        if cleanup_cols:
            merged.drop(columns=cleanup_cols, inplace=True)

        return merged, applied_count

    def _summary_has_missing_agree_values(self, summary_df):
        if summary_df is None or summary_df.empty:
            return False
        expected = []
        expected.extend(
            f"agree_{c}" for c in summary_df.columns
            if c.startswith('S_')
            and 'cor' not in c and 'agree' not in c)
        expected.extend(
            f"agree_{c}" for c in summary_df.columns
            if c.startswith('P_')
            and 'cor' not in c and 'agree' not in c)
        if not expected:
            return False
        for col in expected:
            if col not in summary_df.columns:
                return True
            if summary_df[col].isna().any():
                return True
        return False

    @staticmethod
    def _agree_row_key(filter_text, index1_value):
        """คีย์จับคู่แถว Summary กับ record ใน Agree JSON

        ใช้ "รหัส Index1 (ตัวเลข) + ส่วน cross filter" เพราะข้อความ
        Index1=... ในไฟล์ Compute C เป็น value label (Index1=True Online)
        ต่างจากรอบปกติที่เป็นตัวเลข (Index1=1) จึงตัดส่วนนั้นทิ้ง
        แล้วใช้คอลัมน์ Index1 ที่เป็นตัวเลขแทน
        """
        try:
            if index1_value is None or pd.isna(index1_value):
                idx = 0
            else:
                idx = int(float(index1_value))
        except (TypeError, ValueError):
            idx = 0

        def _norm(part):
            txt = str(part).strip().replace(' ', '').lower()
            txt = re.sub(r'(?<==)(-?\d+)\.0$', r'\1', txt)
            return txt

        parts = [p for p in str(filter_text).split('+') if str(p).strip()]
        cross = sorted(
            _norm(p) for p in parts
            if not str(p).strip().startswith('Index1='))
        return f"{idx}|{'+'.join(cross)}"

    def _apply_agree_summary_cache_by_key(self, summary_df):
        """เติม agree_* จาก JSON โดยจับคู่ด้วยคีย์ ไม่ใช่ลำดับแถว

        ทนกรณีจำนวนแถวไม่เท่ากัน เช่นเงื่อนไข Good ตัดบางกลุ่มหายไปหมด
        คืน (df, จำนวนแถวที่จับคู่ได้, จำนวนค่าที่วาง, รายชื่อแถวที่ไม่เจอ)
        """
        cache = self.agree_summary_cache_df
        if summary_df is None or summary_df.empty:
            return summary_df, 0, 0, []
        if cache is None or cache.empty \
                or 'Filter' not in cache.columns:
            return summary_df, 0, 0, []

        agree_cols = [c for c in cache.columns
                      if str(c).startswith('agree_')]
        if not agree_cols:
            return summary_df, 0, 0, []

        cache_idx = (cache['Index1'] if 'Index1' in cache.columns
                     else pd.Series(0, index=cache.index))
        cache_keys = [
            self._agree_row_key(f, i)
            for f, i in zip(cache['Filter'], cache_idx)]

        lookup = {}
        for pos, key in enumerate(cache_keys):
            if key not in lookup:
                lookup[key] = {
                    c: cache.iloc[pos][c] for c in agree_cols}

        out = summary_df.copy()
        for c in agree_cols:
            if c not in out.columns:
                out[c] = np.nan

        out_idx = (out['Index1'] if 'Index1' in out.columns
                   else pd.Series(0, index=out.index))
        out_keys = pd.Series(
            [self._agree_row_key(f, i)
             for f, i in zip(out['Filter'], out_idx)],
            index=out.index)

        cells = 0
        for c in agree_cols:
            mapped = out_keys.map(
                {k: v[c] for k, v in lookup.items()})
            fill = mapped.notna()
            cells += int(fill.sum())
            out[c] = np.where(fill, mapped, out[c])

        found = out_keys.isin(lookup.keys())
        unmatched = [
            str(v) for v in out.loc[~found, 'Filter'].tolist()]
        return out, int(found.sum()), cells, unmatched

    def _apply_agree_summary_cache_by_position(self, summary_df):
        """โหมด Re-analyze: แปะ agree_* ทั้งชุดตามลำดับแถว (ไม่ map key)"""
        if summary_df is None or summary_df.empty:
            return summary_df, 0, "empty_summary"
        if self.agree_summary_cache_df is None \
                or self.agree_summary_cache_df.empty:
            return summary_df, 0, "empty_cache"

        cache_agree_cols = [
            c for c in self.agree_summary_cache_df.columns
            if c.startswith('agree_')]
        if not cache_agree_cols:
            return summary_df, 0, "no_agree_cols"

        out = summary_df.copy()
        for col in cache_agree_cols:
            if col not in out.columns:
                out[col] = np.nan

        cache_rows = len(self.agree_summary_cache_df)
        out_rows = len(out)
        if cache_rows != out_rows:
            return out, 0, f"row_mismatch:{cache_rows}:{out_rows}"

        out.loc[:, cache_agree_cols] = self.agree_summary_cache_df[
            cache_agree_cols
        ].to_numpy()
        return out, len(cache_agree_cols) * out_rows, "ok"

    # ===================================================================
    # VARIABLE SELECTION GUI
    # ===================================================================
    def open_c_variable_selector(self):
        """เปิดหน้าต่างสำหรับเลือกตัวแปร C"""
        dlg = QDialog(self)
        dlg.setWindowTitle(
            "ขั้นตอนที่ 1.1: เลือกตัวแปร Compute C")
        dlg.resize(700, 500)
        dlg.setModal(True)
        dlg.setStyleSheet(_DLG_QSS)
        vl = QVBoxLayout(dlg)

        fh = QHBoxLayout()
        fl_lbl = QLabel("กรองด้วยคำนำหน้า:")
        fl_lbl.setProperty("class", "dlg-sub")
        fh.addWidget(fl_lbl)
        prefix_entry = QLineEdit()
        fh.addWidget(prefix_entry, 1)
        btn_filter = QPushButton("  Filter  ")
        btn_filter.setStyleSheet(_BTN_STYLES["outline"])
        btn_filter.setMinimumHeight(34)
        btn_filter.setCursor(Qt.CursorShape.PointingHandCursor)
        fh.addWidget(btn_filter)
        vl.addLayout(fh)

        mid = QHBoxLayout()
        av_vl = QVBoxLayout()
        av_lbl = QLabel("Available Variables")
        av_lbl.setProperty("class", "dlg-header")
        av_vl.addWidget(av_lbl)
        available_lw = QListWidget()
        available_lw.setSelectionMode(
            QAbstractItemView.SelectionMode
            .ExtendedSelection)
        av_vl.addWidget(available_lw)
        mid.addLayout(av_vl, 1)

        bv = QVBoxLayout()
        bv.addStretch()
        btn_r = QPushButton("▶")
        btn_r.setProperty("class", "arrow")
        btn_r.setFixedSize(40, 36)
        btn_r.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_l = QPushButton("◀")
        btn_l.setProperty("class", "arrow")
        btn_l.setFixedSize(40, 36)
        btn_l.setCursor(Qt.CursorShape.PointingHandCursor)
        bv.addWidget(btn_r)
        bv.addWidget(btn_l)
        bv.addStretch()
        mid.addLayout(bv)

        sv_vl = QVBoxLayout()
        sv_lbl = QLabel("Selected for Compute C")
        sv_lbl.setProperty("class", "dlg-header")
        sv_vl.addWidget(sv_lbl)
        selected_lw = QListWidget()
        selected_lw.setSelectionMode(
            QAbstractItemView.SelectionMode
            .ExtendedSelection)
        sv_vl.addWidget(selected_lw)
        mid.addLayout(sv_vl, 1)
        vl.addLayout(mid, 1)

        def update_avail(ft=""):
            available_lw.clear()
            sel = set()
            for i in range(selected_lw.count()):
                sel.add(selected_lw.item(i).text())
            disp = [v for v in self.spss_original_order
                    if v not in sel]
            if ft:
                disp = [v for v in disp
                        if v.startswith(ft)]
            available_lw.addItems(disp)

        def move_right():
            for it in available_lw.selectedItems():
                txt = it.text()
                found = selected_lw.findItems(
                    txt, Qt.MatchFlag.MatchExactly)
                if not found:
                    selected_lw.addItem(txt)
            for it in reversed(
                    available_lw.selectedItems()):
                available_lw.takeItem(
                    available_lw.row(it))

        def move_left():
            for it in reversed(
                    selected_lw.selectedItems()):
                selected_lw.takeItem(
                    selected_lw.row(it))
            update_avail(prefix_entry.text())

        def confirm():
            items = []
            for i in range(selected_lw.count()):
                items.append(
                    selected_lw.item(i).text())
            self.c_vars_to_compute = items
            if not items:
                self._msg_warn(
                    "ยังไม่ได้เลือกตัวแปร",
                    "กรุณาเลือกตัวแปรอย่างน้อย 1 ตัว")
                return
            dlg.accept()
            QTimer.singleShot(
                100, self.run_c_compute_and_proceed)

        btn_filter.clicked.connect(
            lambda: update_avail(prefix_entry.text()))
        prefix_entry.returnPressed.connect(
            lambda: update_avail(prefix_entry.text()))
        btn_r.clicked.connect(move_right)
        btn_l.clicked.connect(move_left)

        ok = QPushButton("  ✔  ยืนยันและดำเนินการต่อ  ")
        ok.setStyleSheet(_BTN_STYLES["success"])
        ok.setMinimumHeight(40)
        ok.setCursor(Qt.CursorShape.PointingHandCursor)
        ok.clicked.connect(confirm)
        vl.addWidget(ok)

        update_avail()
        self._center_toplevel(dlg)
        dlg.exec()

    def run_c_compute_and_proceed(self):
        """รันการคำนวณ C และไปขั้นตอนเลือกตัวแปรอื่นๆ

        Compute C เป็นการคำนวณ pandas ล้วนๆ ที่เร็วมาก และต้องเปิด
        dialog ต่อทันที จึงรันบน main thread ต่อไป
        """
        self.start_progress()
        self.update_status(f"เลือก {len(self.c_vars_to_compute)} ตัวแปร. กำลัง Compute C...")
        try:
            self._compute_c_variables_logic()
        except Exception as e:
            self.stop_progress()
            self.update_status("Compute C ผิดพลาด", "danger")
            self._msg_error("ผิดพลาด", str(e))
            self.reset_state()
            return
        self.update_status(f"Compute C สำเร็จ! สร้าง {len(self.computed_c_cols)} ตัวแปร.", "success")
        self.stop_progress()
        self.open_aspe_selector()

    def open_aspe_selector(self):
        """เปิดหน้าต่างเลือก A,S,P,E + AgreeS,AgreeP + T2B"""
        dlg = QDialog(self)
        dlg.setWindowTitle(
            "ขั้นตอนที่ 1.2: เลือกตัวแปรแปลงข้อมูล")
        dlg.resize(800, 650)
        dlg.setModal(True)
        dlg.setStyleSheet(_DLG_QSS)
        vl = QVBoxLayout(dlg)

        tab_w = QTabWidget()
        vl.addWidget(tab_w, 1)

        tab_names = ["A", "S", "P", "E",
                     "AgreeS", "AgreeP"]
        listboxes = {}
        all_selected = set()

        def make_tab(name):
            w = QWidget()
            hl = QHBoxLayout(w)
            av_vl2 = QVBoxLayout()
            av_lbl2 = QLabel("Available")
            av_lbl2.setProperty("class", "dlg-header")
            av_vl2.addWidget(av_lbl2)
            a_lw = QListWidget()
            a_lw.setSelectionMode(
                QAbstractItemView.SelectionMode
                .ExtendedSelection)
            av_vl2.addWidget(a_lw)
            hl.addLayout(av_vl2, 1)

            bv2 = QVBoxLayout()
            bv2.addStretch()
            br = QPushButton("▶")
            br.setProperty("class", "arrow")
            br.setFixedSize(40, 36)
            br.setCursor(Qt.CursorShape.PointingHandCursor)
            bl = QPushButton("◀")
            bl.setProperty("class", "arrow")
            bl.setFixedSize(40, 36)
            bl.setCursor(Qt.CursorShape.PointingHandCursor)
            bv2.addWidget(br)
            bv2.addWidget(bl)
            bv2.addStretch()
            hl.addLayout(bv2)

            sv_vl2 = QVBoxLayout()
            sv_lbl2 = QLabel(f"Selected '{name}'")
            sv_lbl2.setProperty("class", "dlg-header")
            sv_vl2.addWidget(sv_lbl2)
            s_lw = QListWidget()
            s_lw.setSelectionMode(
                QAbstractItemView.SelectionMode
                .ExtendedSelection)
            sv_vl2.addWidget(s_lw)
            hl.addLayout(sv_vl2, 1)

            def mv_r():
                for it in a_lw.selectedItems():
                    t = it.text()
                    if not s_lw.findItems(
                            t,
                            Qt.MatchFlag.MatchExactly):
                        s_lw.addItem(t)
                        all_selected.add(t)
                for it in reversed(
                        a_lw.selectedItems()):
                    a_lw.takeItem(a_lw.row(it))

            def mv_l():
                for it in reversed(
                        s_lw.selectedItems()):
                    all_selected.discard(it.text())
                    s_lw.takeItem(s_lw.row(it))
                refresh_avail(a_lw)

            br.clicked.connect(mv_r)
            bl.clicked.connect(mv_l)
            tab_w.addTab(w, name)
            return {"available": a_lw,
                    "selected": s_lw}

        def refresh_avail(lw):
            lw.clear()
            orig = [
                v for v in self.spss_original_order
                if v not in self.computed_c_cols
                and v not in all_selected]
            lw.addItems(orig)

        for n in tab_names:
            listboxes[n] = make_tab(n)
        for n in tab_names:
            refresh_avail(listboxes[n]["available"])

        # T2B options
        og = QGroupBox(" T2B Options ")
        ogl = QHBoxLayout(og)
        ogl.addWidget(QLabel(
            "เลือก Code ด้านดี (T2B):"))
        rb1 = QRadioButton("5+4 (Default)")
        rb1.setChecked(True)
        rb2 = QRadioButton("1+2")
        t2b_bg = QButtonGroup(dlg)
        t2b_bg.addButton(rb1)
        t2b_bg.addButton(rb2)
        ogl.addWidget(rb1)
        ogl.addWidget(rb2)
        ogl.addStretch()
        vl.addWidget(og)

        def confirm():
            for nm, lbs in listboxes.items():
                items = []
                sl = lbs["selected"]
                for i in range(sl.count()):
                    items.append(sl.item(i).text())
                self.vars_to_transform[nm] = items
            if rb2.isChecked():
                self.t2b_choice_var.set("1+2")
            else:
                self.t2b_choice_var.set("5+4")
            dlg.accept()
            QTimer.singleShot(
                100,
                self.run_full_transformation_and_save)

        ok = QPushButton(
            "  ✔  ยืนยัน, แปลงข้อมูล และบันทึก  ")
        ok.setStyleSheet(_BTN_STYLES["success"])
        ok.setMinimumHeight(40)
        ok.setCursor(Qt.CursorShape.PointingHandCursor)
        ok.clicked.connect(confirm)
        vl.addWidget(ok)

        self._center_toplevel(dlg)
        dlg.exec()

    def _transform_pipeline(self, with_compute_c):
        """ขั้นตอนประมวลผลหนัก — รันบน worker thread เท่านั้น

        ห้ามแตะ widget โดยตรง ใช้ log_message/update_status/set_progress
        ซึ่งส่งผ่าน signal ให้ main thread ทำงานแทน
        """
        start_time = time.time()
        total_steps = 4 if with_compute_c else 3
        step = 0

        self.log_message("=" * 50)
        self.log_message(
            "เริ่มกระบวนการประมวลผลจากไฟล์ตั้งค่า"
            if with_compute_c
            else "เริ่มกระบวนการประมวลผลข้อมูล")
        self.log_message("=" * 50)
        if not with_compute_c:
            self.log_message(
                f"Compute C: {len(self.computed_c_cols)} ตัวแปร")
        self.log_message("")
        self.set_progress(step, total_steps)

        if with_compute_c:
            step += 1
            self.log_message(f"[{step}/{total_steps}] กำลัง Compute C...")
            self.update_status("กำลัง Compute C จากการตั้งค่า...")
            self._compute_c_variables_logic()
            self.log_message(
                f"   ✓ Compute C สำเร็จ ({len(self.computed_c_cols)} ตัวแปร)")
            self.set_progress(step, total_steps)
            self.log_message("")

        step += 1
        self.log_message(f"[{step}/{total_steps}] กำลัง Recode ตัวแปร A...")
        self.update_status("กำลัง Recode ตัวแปร A...")
        self._recode_a_variables_logic()
        self.log_message(
            f"   ✓ Recode A สำเร็จ ({len(self.za_cols)} ตัวแปร ZA)")
        self.set_progress(step, total_steps)

        self.log_message("")
        step += 1
        self.log_message(
            f"[{step}/{total_steps}] กำลังแปลงข้อมูล (Variables to Cases)...")
        self.update_status("กำลังแปลงข้อมูล (Variables to Cases)...")
        self._run_full_transformation_logic()
        self.log_message(
            f"   ✓ แปลงข้อมูลสำเร็จ ({len(self.transformed_df)} แถว)")
        self.set_progress(step, total_steps)

        self.log_message("")
        step += 1
        self.log_message(f"[{step}/{total_steps}] กำลังบันทึกไฟล์ .sav...")
        self.update_status(
            "แปลงข้อมูลสำเร็จ. กำลังบันทึกไฟล์อัตโนมัติ...", "success")
        self._auto_save_spss(self.transformed_df)
        self.log_message("   ✓ บันทึก .sav สำเร็จ")
        self.set_progress(step, total_steps)

        elapsed = time.time() - start_time
        self.log_message("")
        self.log_message("=" * 50)
        self.log_message(
            f"ประมวลผลข้อมูลเสร็จสมบูรณ์ (ใช้เวลา {elapsed:.1f} วินาที)")
        self.log_message("=" * 50)
        self.log_message("")
        return True

    def _on_transform_failed(self, message, tb_text):
        """เคลียร์ state เมื่อขั้นตอนแปลงข้อมูลล้มเหลว"""
        self._msg_error(
            "ประมวลผลไม่สำเร็จ",
            f"{message}\n\nดูรายละเอียดเพิ่มเติมได้ที่ Log ด้านขวา")
        self.reset_state()

    def run_full_transformation_and_save(self):
        self._snapshot_ui_inputs()
        self.show_log_panel("กำลังประมวลผลข้อมูล (Step 1)...")
        self.start_progress()

        def _done(_result):
            self.stop_progress()
            self.log_message(
                "กรุณาใส่ Filter (ถ้ามี) และกด 'วิเคราะห์และส่งออก Excel'")
            self.update_status("ประมวลผลข้อมูลสำเร็จ", "success")
            self.btn_analyze_export.setEnabled(True)
            self.btn_define_labels.setEnabled(True)
            self.btn_save_settings.setEnabled(True)
            self.filter_entry.setEnabled(True)

        self._run_in_thread(
            lambda: self._transform_pipeline(with_compute_c=False),
            _done,
            self._on_transform_failed)

    def run_processing_with_loaded_settings(self):
        self._snapshot_ui_inputs()
        self.show_log_panel("กำลังประมวลผลข้อมูล (จากไฟล์ตั้งค่า)...")
        self.start_progress()

        def _done(_result):
            self.stop_progress()
            self.log_message("กำลังเริ่มวิเคราะห์และส่งออกอัตโนมัติ...")
            self.update_status(
                "ประมวลผลข้อมูลสำเร็จ. เริ่มการวิเคราะห์และส่งออกอัตโนมัติ...",
                "info")
            self.btn_analyze_export.setEnabled(True)
            self.btn_define_labels.setEnabled(True)
            self.btn_save_settings.setEnabled(True)
            self.filter_entry.setEnabled(True)
            QTimer.singleShot(
                100,
                lambda: self.run_analysis_and_export(automated=True))

        self._run_in_thread(
            lambda: self._transform_pipeline(with_compute_c=True),
            _done,
            self._on_transform_failed)

    # ===================================================================
    # PROCESSING LOGIC (Back-end)
    # ===================================================================
    def _compute_c_variables_logic(self):
        """คำนวณตัวแปร C — โยน RuntimeError เมื่อผิดพลาด
        (ห้ามเปิด QMessageBox จากที่นี่ เพราะรันบน worker thread)"""
        if not self.c_vars_to_compute:
            raise RuntimeError(
                "ไม่มีตัวแปรที่ถูกเลือกสำหรับคำนวณ C")
        try:
            first_var = self.c_vars_to_compute[0]
            if '#' not in first_var:
                raise RuntimeError(
                    f"ตัวแปรที่เลือก ({first_var}) ไม่มีรูปแบบที่ถูกต้อง "
                    "(เช่น 'PREFIX#GROUP$ITEM')")
            deduced_prefix = first_var.split('#')[0]
            pattern = re.compile(rf"^{re.escape(deduced_prefix)}#(\d+)\$(\d+)")
            groups = {}
            for col in self.c_vars_to_compute:
                match = pattern.match(col)
                if match:
                    group_num = int(match.group(1))
                    if group_num not in groups:
                        groups[group_num] = []
                    groups[group_num].append(col)
            if not groups:
                raise RuntimeError(
                    f"ไม่พบตัวแปรที่ตรงกับรูปแบบ '{deduced_prefix}#Group$Item' "
                    "จากตัวแปรที่คุณเลือก")

            self.computed_c_cols = []
            new_cols_data = {}

            max_item_num = max((int(m.group(2)) for c in self.c_vars_to_compute if (m := pattern.match(c))), default=0)
            if max_item_num == 0:
                raise RuntimeError(
                    "ไม่สามารถหา Item number สูงสุดจากตัวแปรที่เลือกสำหรับ C ได้")

            for j in range(1, max_item_num + 1):
                for i in sorted(groups.keys()):
                    group_vars = groups[i]
                    main_var = f"{deduced_prefix}#{i}${j}"
                    if main_var in group_vars:
                        other_vars = [v for v in group_vars if v != main_var]
                        if not other_vars:
                            continue
                        new_c_name = f"C{j}.{i}"
                        mean_of_others = self.df[other_vars].mean(axis=1)
                        new_cols_data[new_c_name] = ((self.df[main_var] - mean_of_others) + 1) / 2
                        self.computed_c_cols.append(new_c_name)

            if not self.computed_c_cols:
                raise RuntimeError(
                    "ไม่สามารถสร้างตัวแปร C ได้ "
                    "อาจเพราะโครงสร้างตัวแปรไม่ถูกต้อง")

            if new_cols_data:
                self.df = pd.concat([self.df, pd.DataFrame(new_cols_data)], axis=1)

            return True
        except RuntimeError:
            raise
        except Exception as e:
            raise RuntimeError(
                f"เกิดข้อผิดพลาดระหว่างคำนวณตัวแปร C: {e}") from e

    def _recode_a_variables_logic(self):
        a_vars_to_process = self.vars_to_transform.get('A', [])
        self.za_cols = []
        if not a_vars_to_process:
            return True
        try:
            za_map = {0: 0, 1: 0.05, 2: 0.12, 3: 0.27, 4: 0.50, 5: 0.73, 6: 0.88, 7: 0.95, 8: 1.00}
            new_za_cols_data = {}

            for var in a_vars_to_process:
                if var in self.df.columns and pd.api.types.is_numeric_dtype(self.df[var]):
                    self.df[var] = self.df[var].replace(9, 0)

                    za_var_name = 'Z' + var
                    new_za_cols_data[za_var_name] = self.df[var].map(za_map).fillna(self.df[var])
                    self.za_cols.append(za_var_name)

            if new_za_cols_data:
                self.df = pd.concat([self.df, pd.DataFrame(new_za_cols_data)], axis=1)

            return True
        except Exception as e:
            raise RuntimeError(
                f"เกิดข้อผิดพลาดขณะแปลงค่าตัวแปร A: {e}") from e

    def _run_full_transformation_logic(self):
        try:
            temp_df = self.df.copy()
            all_transform_vars = set(self.computed_c_cols)
            for key, var_list in self.vars_to_transform.items():
                if key not in ['AgreeS', 'AgreeP']:
                    all_transform_vars.update(var_list)
            all_transform_vars.update(self.za_cols)

            self.id_vars = [col for col in self.df.columns if col not in all_transform_vars]

            A_PAT, ZA_PAT = re.compile(r".*?#(\d+)$"), re.compile(r"Z.*?#(\d+)$")
            SPE_PAT, C_PAT = re.compile(r".*?#(\d+)\$(\d+)$"), re.compile(r"C(\d+)\.(\d+)$")

            maps = {'A': {}, 'S': {}, 'P': {}, 'E': {}, 'C': {}, 'ZA': {}}
            groups = {'S': set(), 'P': set(), 'E': set(), 'C': set()}
            max_index = 0

            for var in self.vars_to_transform.get('A', []):
                if match := A_PAT.match(var): idx = int(match.group(1)); maps['A'][idx] = var; max_index = max(max_index, idx)
            for var in self.za_cols:
                if match := ZA_PAT.match(var): idx = int(match.group(1)); maps['ZA'][idx] = var
            for key in ['S', 'P', 'E']:
                for var in self.vars_to_transform.get(key, []):
                    if match := SPE_PAT.match(var):
                        grp, idx = int(match.group(1)), int(match.group(2))
                        if grp not in maps[key]: maps[key][grp] = {}
                        maps[key][grp][idx] = var; groups[key].add(grp); max_index = max(max_index, idx)
            for var in self.computed_c_cols:
                if match := C_PAT.match(var):
                    idx, grp = int(match.group(1)), int(match.group(2))
                    if grp not in maps['C']: maps['C'][grp] = {}
                    maps['C'][grp][idx] = var; groups['C'].add(grp); max_index = max(max_index, idx)

            if max_index == 0:
                raise RuntimeError(
                    "ไม่สามารถหา Index สำหรับการแปลงข้อมูลได้\n"
                    "กรุณาตรวจสอบรูปแบบของตัวแปรที่เลือก")

            # --- Variables to Cases (vectorized) ---
            # เดิมวนด้วย iterrows สร้าง dict ทีละเรคอร์ด ทำให้ช้าตาม
            # (จำนวนผู้ตอบ x max_index) ตอนนี้สร้างทีละ Index1 แบบทั้งคอลัมน์
            # แล้วค่อย concat ผลลัพธ์เหมือนเดิมทุกประการ
            base_df = temp_df.reset_index(drop=True)
            id_cols = [c for c in self.id_vars
                       if c in base_df.columns]

            frames = []
            for j in range(1, max_index + 1):
                part = base_df[id_cols].copy()
                part['Index1'] = j

                if (a_source := maps['A'].get(j)) \
                        and a_source in base_df.columns:
                    part['A'] = base_df[a_source]
                if (za_source := maps['ZA'].get(j)) \
                        and za_source in base_df.columns:
                    part['ZA'] = base_df[za_source]

                for key in ['S', 'P', 'E', 'C']:
                    for i in sorted(groups[key]):
                        source_var = maps[key].get(i, {}).get(j)
                        if source_var and source_var in base_df.columns:
                            part[f'{key}_{i}'] = base_df[source_var]

                frames.append(part)

            # sort_index แบบ stable ทำให้ได้ลำดับเดิม คือ
            # ผู้ตอบคนที่ 1 ครบทุก Index1 ก่อน แล้วจึงเป็นคนถัดไป
            self.transformed_df = (
                pd.concat(frames)
                .sort_index(kind='stable')
                .reset_index(drop=True))

            value_cols = [col for col in ['A', 'ZA'] if col in self.transformed_df.columns]
            for key in ['S', 'P', 'E', 'C']: value_cols.extend([c for c in self.transformed_df.columns if c.startswith(f"{key}_")])
            if value_cols_in_df := [c for c in value_cols if c in self.transformed_df.columns]:
                self.transformed_df.dropna(subset=value_cols_in_df, how='all', inplace=True)

            for key, col_name in {'S':'N_S', 'P':'N_P', 'C':'N_C', 'E':'N_E'}.items():
                if cols := [c for c in self.transformed_df.columns if c.startswith(f'{key}_')]: self.transformed_df[col_name] = self.transformed_df[cols].mean(axis=1)

            # --- E Group Mode: merge specified E groups ---
            # อ่านจาก snapshot เท่านั้น ห้ามแตะ widget จาก worker thread
            if self._ui.get('e_group_mode', 'default') == "group":
                e_group_expr = self._ui.get(
                    'e_group_expr', '').strip()
                if e_group_expr:
                    try:
                        e_group_nums = [int(x.strip()) for x in e_group_expr.split('+')]
                        e_cols_to_merge = [f'E_{g}' for g in e_group_nums if f'E_{g}' in self.transformed_df.columns]
                        if len(e_cols_to_merge) >= 2:
                            merged_name = 'E_' + ''.join(str(n) for n in e_group_nums)
                            self.transformed_df[merged_name] = self.transformed_df[e_cols_to_merge].mean(axis=1)
                            self.transformed_df.drop(columns=e_cols_to_merge, inplace=True)
                            self.log_message(f"   ✓ E Group: รวม {e_group_expr} → {merged_name}")
                            # Recompute N_E with merged columns
                            remaining_e = [c for c in self.transformed_df.columns if c.startswith('E_')]
                            if remaining_e:
                                self.transformed_df['N_E'] = self.transformed_df[remaining_e].mean(axis=1)
                    except ValueError:
                        self.log_message(f"   ⚠ E Group: ไม่สามารถแปลงค่า '{e_group_expr}' ได้")

            final_ordered_cols = self.id_vars + ['Index1']
            for col in ['N_S', 'N_P', 'N_C', 'N_E', 'A', 'ZA']:
                if col in self.transformed_df.columns: final_ordered_cols.append(col)

            all_new_keys = {c for key in ['S', 'P', 'E', 'C'] for c in self.transformed_df.columns if c.startswith(f"{key}_")}
            def _col_sort_key(x):
                prefix, suffix = x.split('_', 1)
                return (prefix, int(suffix))
            sorted_new_keys = sorted(list(all_new_keys), key=_col_sort_key)
            final_ordered_cols.extend(sorted_new_keys)

            self.transformed_df = self.transformed_df[[c for c in final_ordered_cols if c in self.transformed_df.columns]]
            self._respondent_key = _UNSET   # ข้อมูลเปลี่ยน ต้องตรวจใหม่
            self._build_compute_sav_metadata(maps)
            return True
        except RuntimeError:
            raise
        except Exception as e:
            raise RuntimeError(
                f"เกิดข้อผิดพลาดระหว่างการแปลงข้อมูล: {e}") from e

    def _build_compute_sav_metadata(self, maps):
        """เตรียม SPSS metadata (column labels + value labels) สำหรับไฟล์ Compute C"""
        self.compute_sav_column_labels = {}
        self.compute_sav_value_labels = {}
        if self.transformed_df is None or self.transformed_df.empty:
            return

        out_cols = set(self.transformed_df.columns)

        def _copy_meta(target_col, source_col):
            if target_col not in out_cols or not source_col:
                return
            src_lbl = self.spss_variable_labels.get(source_col)
            if src_lbl:
                self.compute_sav_column_labels[target_col] = str(src_lbl)
            src_vl = self.spss_value_labels.get(source_col)
            if isinstance(src_vl, dict) and src_vl:
                self.compute_sav_value_labels[target_col] = src_vl

        # 1) คอลัมน์ ID ดั้งเดิม
        for col in self.id_vars:
            _copy_meta(col, col)

        # 2) Index1 และ A/ZA
        if 'Index1' in out_cols:
            self.compute_sav_column_labels['Index1'] = 'Index1'
            if self.index1_labels:
                self.compute_sav_value_labels['Index1'] = {
                    int(k): str(v)
                    for k, v in self.index1_labels.items()
                }

        src_a = None
        if maps.get('A'):
            first_idx = sorted(maps['A'].keys())[0]
            src_a = maps['A'].get(first_idx)
        _copy_meta('A', src_a)
        if 'ZA' in out_cols:
            base_lbl = self.spss_variable_labels.get(src_a, 'A')
            self.compute_sav_column_labels['ZA'] = f"ZA from {base_lbl}"

        # 3) S/P/E: เอา label+value labels จากตัวแปรต้นทางตัวแรกของแต่ละกลุ่ม
        for prefix in ['S', 'P', 'E']:
            group_map = maps.get(prefix, {})
            for grp in sorted(group_map.keys()):
                tgt = f"{prefix}_{grp}"
                src_candidates = [
                    group_map[grp][idx]
                    for idx in sorted(group_map[grp].keys())
                ]
                src = src_candidates[0] if src_candidates else None
                _copy_meta(tgt, src)

        # 4) ค่าเฉลี่ยรวม
        for agg_col, agg_label in [
                ('N_S', 'Mean score of S'),
                ('N_P', 'Mean score of P'),
                ('N_C', 'Mean score of C'),
                ('N_E', 'Mean score of E')]:
            if agg_col in out_cols:
                self.compute_sav_column_labels[agg_col] = agg_label

    def _auto_save_spss(self, dataframe_to_save):
        if dataframe_to_save is None:
            raise RuntimeError("ไม่มีข้อมูลให้บันทึก")
        if not self.original_filepath:
            raise RuntimeError("ไม่พบ Path ของไฟล์ต้นฉบับ")

        try:
            base, _ = os.path.splitext(self.original_filepath)
            new_filepath = f"{base} Compute C.sav"

            write_kwargs = {}
            if self.compute_sav_column_labels:
                ordered_labels = [
                    self.compute_sav_column_labels.get(col, col)
                    for col in dataframe_to_save.columns
                ]
                write_kwargs['column_labels'] = ordered_labels
            if self.compute_sav_value_labels:
                valid_value_labels = {
                    col: labels
                    for col, labels in self.compute_sav_value_labels.items()
                    if col in dataframe_to_save.columns
                    and isinstance(labels, dict)
                    and labels
                }
                if valid_value_labels:
                    write_kwargs['variable_value_labels'] = valid_value_labels

            pyreadstat.write_sav(
                dataframe_to_save,
                new_filepath,
                **write_kwargs)
            self.update_status(f"บันทึกไฟล์ใหม่ที่: {new_filepath}", "success")
            return True
        except Exception as e:
            raise RuntimeError(
                f"ไม่สามารถบันทึกไฟล์อัตโนมัติได้: {e}") from e

    def _extract_spss_metadata(self, meta):
        """สกัด variable labels + value labels ให้ครบที่สุดจาก meta"""
        var_labels = {}
        value_labels = {}

        # Variable labels
        cn2l = getattr(meta, 'column_names_to_labels', None)
        if isinstance(cn2l, dict) and cn2l:
            var_labels.update(cn2l)
        else:
            col_names = getattr(meta, 'column_names', None) or []
            col_labels = getattr(meta, 'column_labels', None) or []
            if col_names and col_labels:
                for i, col in enumerate(col_names):
                    if i < len(col_labels):
                        lbl = col_labels[i]
                        if lbl is not None and str(lbl).strip():
                            var_labels[col] = str(lbl)

        # Value labels (preferred direct map)
        vvl = getattr(meta, 'variable_value_labels', None)
        if isinstance(vvl, dict) and vvl:
            value_labels.update(vvl)
        else:
            # Fallback: variable_to_label + value_labels
            var_to_label = getattr(meta, 'variable_to_label', None) or {}
            label_sets = getattr(meta, 'value_labels', None) or {}
            if isinstance(var_to_label, dict) and isinstance(label_sets, dict):
                for var_name, label_set_name in var_to_label.items():
                    labels = label_sets.get(label_set_name)
                    if isinstance(labels, dict) and labels:
                        value_labels[var_name] = labels

        return var_labels, value_labels

    def display_table(self, dataframe):
        """แสดง DataFrame ใน QTableWidget"""
        self._clear_right_panel()
        df = dataframe.head(1000).fillna('')
        tw = QTableWidget(
            df.shape[0], df.shape[1])
        tw.setHorizontalHeaderLabels(
            list(df.columns))
        tw.horizontalHeader().setStretchLastSection(
            True)
        for r in range(df.shape[0]):
            for c in range(df.shape[1]):
                tw.setItem(
                    r, c,
                    QTableWidgetItem(
                        str(df.iat[r, c])))
        tw.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.right_frame.layout().addWidget(tw)

    def show_message_in_display(self, message_text):
        """แสดงข้อความในพื้นที่แสดงผลด้านขวา"""
        self._clear_right_panel()
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(
            "QScrollArea{border:none; background:transparent;}")
        holder = QWidget()
        hl = QVBoxLayout(holder)
        hl.setContentsMargins(26, 24, 26, 24)
        lb = QLabel(message_text)
        lb.setStyleSheet(
            "color:#37474F; font-size:13px;"
            "line-height:170%;")
        lb.setWordWrap(True)
        lb.setTextInteractionFlags(
            Qt.TextInteractionFlag.TextSelectableByMouse)
        lb.setAlignment(
            Qt.AlignmentFlag.AlignTop
            | Qt.AlignmentFlag.AlignLeft)
        hl.addWidget(lb)
        hl.addStretch()
        scroll.setWidget(holder)
        self.right_frame.layout().addWidget(scroll)

    # -----------------------------------------------------------------
    # Rich "ready to analyse" panel
    # -----------------------------------------------------------------
    def _stat_tile(self, value, caption, sub, accent, bg):
        """กล่องตัวเลขสรุป 1 ช่อง"""
        card = QFrame()
        card.setStyleSheet(
            f"QFrame{{background:{bg}; border-radius:12px;"
            f"border:1px solid rgba(0,0,0,0.05);}}")
        cl = QVBoxLayout(card)
        cl.setContentsMargins(16, 14, 16, 14)
        cl.setSpacing(2)

        v = QLabel(value)
        v.setStyleSheet(
            f"color:{accent}; font-size:26px; font-weight:700;"
            "background:transparent; border:none;")
        cl.addWidget(v)

        c = QLabel(caption)
        c.setStyleSheet(
            "color:#455A64; font-size:12px; font-weight:600;"
            "background:transparent; border:none;")
        cl.addWidget(c)

        if sub:
            s = QLabel(sub)
            s.setStyleSheet(
                "color:#90A4AE; font-size:10px;"
                "background:transparent; border:none;")
            cl.addWidget(s)
        return card

    def show_reanalyze_ready_panel(self, source_path):
        """หน้าสรุปหลังโหลดไฟล์ในโหมดวิเคราะห์ซ้ำ"""
        self._clear_right_panel()

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(
            "QScrollArea{border:none; background:transparent;}")
        page = QWidget()
        pl = QVBoxLayout(page)
        pl.setContentsMargins(28, 24, 28, 24)
        pl.setSpacing(14)
        scroll.setWidget(page)
        self.right_frame.layout().addWidget(scroll)

        # ---- header ----
        head = QLabel("✅  โหลดไฟล์สำเร็จ")
        head.setStyleSheet(
            "color:#2E7D32; font-size:21px; font-weight:700;"
            "background:transparent;")
        pl.addWidget(head)

        if source_path:
            fname = QLabel(os.path.basename(source_path))
            fname.setStyleSheet(
                "color:#78909C; font-size:11px;"
                "background:transparent;")
            fname.setWordWrap(True)
            pl.addWidget(fname)

        st = self.good_filter_stats
        if st:
            # ---- section title ----
            sec = QLabel("ตัดชุดด้วยเงื่อนไข Good (System)")
            sec.setStyleSheet(
                "color:#B71C1C; font-size:13px; font-weight:700;"
                "background:transparent; padding-top:6px;")
            pl.addWidget(sec)

            # ---- stat tiles ----
            tiles = QHBoxLayout()
            tiles.setSpacing(10)
            tiles.addWidget(self._stat_tile(
                f"{st['total']:,}", "ข้อมูลทั้งหมด", "แถว",
                "#37474F", "#ECEFF1"))
            tiles.addWidget(self._stat_tile(
                f"{st['good']:,}", "Good = 1",
                f"ใช้วิเคราะห์  ({st['good_pct']:.1f}%)",
                "#2E7D32", "#E8F5E9"))
            tiles.addWidget(self._stat_tile(
                f"{st['dropped']:,}", "Good = 2",
                f"ตัดออก  ({100 - st['good_pct']:.1f}%)",
                "#C62828", "#FFEBEE"))
            pl.addLayout(tiles)

            # ---- ratio bar ----
            bar = QFrame()
            bar.setFixedHeight(10)
            bar.setStyleSheet(
                "QFrame{background:#ECEFF1; border-radius:5px;}")
            bl = QHBoxLayout(bar)
            bl.setContentsMargins(0, 0, 0, 0)
            bl.setSpacing(0)
            keep = QFrame()
            keep.setStyleSheet(
                "background:#66BB6A;"
                "border-top-left-radius:5px;"
                "border-bottom-left-radius:5px;")
            drop = QFrame()
            drop.setStyleSheet(
                "background:#EF9A9A;"
                "border-top-right-radius:5px;"
                "border-bottom-right-radius:5px;")
            bl.addWidget(keep, max(int(round(st['good_pct'] * 10)), 1))
            bl.addWidget(drop, max(int(round(
                (100 - st['good_pct']) * 10)), 1))
            pl.addWidget(bar)

            # ---- verification file ----
            if st.get('saved_path'):
                cap = QLabel(
                    "บันทึกตัวแปร Good ลงไฟล์ SPSS ที่โหลดเข้ามาแล้ว "
                    "(ครบทุกแถว)")
                cap.setStyleSheet(
                    "color:#607D8B; font-size:11px; font-weight:600;"
                    "background:transparent; padding-top:6px;")
                pl.addWidget(cap)

                box = QLabel(st['saved_path'])
                box.setWordWrap(True)
                box.setTextInteractionFlags(
                    Qt.TextInteractionFlag.TextSelectableByMouse)
                box.setStyleSheet(
                    "color:#455A64; font-size:10px;"
                    "font-family:Consolas,monospace;"
                    "background:#F5F7F8; border:1px solid #E0E0E0;"
                    "border-radius:8px; padding:10px 12px;")
                pl.addWidget(box)

        # ---- agree json ----
        if self.agree_json_filepath_override:
            cap = QLabel(
                "ไฟล์ Agree Original JSON"
                + ("  (พบอัตโนมัติ)" if self.agree_json_auto_found
                   else "  (เลือกเอง)"))
            cap.setStyleSheet(
                "color:#607D8B; font-size:11px; font-weight:600;"
                "background:transparent; padding-top:6px;")
            pl.addWidget(cap)

            jb = QLabel(os.path.basename(
                self.agree_json_filepath_override))
            jb.setWordWrap(True)
            jb.setTextInteractionFlags(
                Qt.TextInteractionFlag.TextSelectableByMouse)
            jb.setStyleSheet(
                "color:#455A64; font-size:10px;"
                "font-family:Consolas,monospace;"
                "background:#F5F7F8; border:1px solid #E0E0E0;"
                "border-radius:8px; padding:10px 12px;")
            pl.addWidget(jb)

        # ---- ตาราง Cross A x N_S/N_P/N_C/N_E ----
        if self.good_filter_full_df is not None:
            sec2 = QLabel("ตาราง Cross : A × N_S / N_P / N_C / N_E")
            sec2.setStyleSheet(
                "color:#B71C1C; font-size:13px; font-weight:700;"
                "background:transparent; padding-top:10px;")
            pl.addWidget(sec2)

            hint = QLabel(
                "นับจากข้อมูลเต็มก่อนตัด — แถว = ค่าตัวชี้วัด, "
                "คอลัมน์ = A\n"
                "🟩 อยู่ในช่วงที่เงื่อนไขยอมรับ   "
                "🟨 ถูกเงื่อนไขตัด   "
                "🟧 ค่าจริงคร่อมเกณฑ์หลังปัดทศนิยม")
            hint.setWordWrap(True)
            hint.setStyleSheet(
                "color:#78909C; font-size:10px;"
                "background:transparent;")
            pl.addWidget(hint)

            tabs = QTabWidget()
            tallest = 0
            for metric in ('N_S', 'N_P', 'N_C', 'N_E'):
                page_w = self._build_crosstab_table(metric)
                tabs.addTab(page_w, f"  {metric}  ")
                tallest = max(tallest, page_w.minimumHeight())
            # ล็อกความสูงตามแท็บที่แถวเยอะสุด ทุกแท็บจะเห็นครบโดยไม่ต้องเลื่อน
            tabs.setFixedHeight(
                tallest + tabs.tabBar().sizeHint().height() + 16)
            pl.addWidget(tabs)

        # ---- next step callout ----
        nxt = QLabel(
            "ขั้นตอนถัดไป\n"
            "ตรวจสอบช่อง Filter ทางซ้าย แล้วกด "
            "'วิเคราะห์และส่งออก Excel'")
        nxt.setWordWrap(True)
        nxt.setStyleSheet(
            "color:#0D47A1; font-size:12px; font-weight:600;"
            "background:#E3F2FD; border-left:4px solid #1976D2;"
            "border-radius:8px; padding:12px 14px; margin-top:8px;")
        pl.addWidget(nxt)

        pl.addStretch()

    def show_log_panel(self, title="Processing Log"):
        """แสดง Live Log Panel ในพื้นที่ด้านขวา"""
        self._clear_right_panel()
        hdr = QLabel(title)
        hdr.setStyleSheet(
            "color:#1565C0; font-size:15px;"
            "font-weight:700;")
        self.right_frame.layout().addWidget(hdr)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet(
            "background:#1F2D3A; color:#E6E6E6;"
            "font-family:Consolas; font-size:10px;"
            "border-radius:4px;")
        self.right_frame.layout().addWidget(
            self.log_text)

    def log_message(self, message):
        """เพิ่มข้อความลงใน Log Panel (เรียกจาก thread ไหนก็ได้)"""
        self.sig_log.emit(str(message))

    def _append_log_ui(self, message):
        if self.log_text is None:
            return
        try:
            self.log_text.append(message)
            sb = self.log_text.verticalScrollBar()
            sb.setValue(sb.maximum())
        except RuntimeError:
            # ป้องกันกรณี widget ถูกลบระหว่างงานยาว
            self.log_text = None

    def _load_factor_output_text_from_excel(self):
        filepath = self.last_excel_filepath
        if not filepath or not os.path.exists(filepath):
            return ""
        try:
            workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            if "Factor_Output" not in workbook.sheetnames:
                workbook.close()
                return ""
            worksheet = workbook["Factor_Output"]
            lines = []
            for row in worksheet.iter_rows(min_row=1, max_col=1, values_only=True):
                value = row[0]
                lines.append("" if value is None else str(value))
            workbook.close()
            return "\n".join(lines).strip()
        except Exception:
            return ""

    def display_analysis_tabs(self, analysis_text):
        self._clear_right_panel()
        tabs = QTabWidget()
        self.right_frame.layout().addWidget(tabs)

        # Tab 1 - Factor Output
        ta = QTextEdit()
        ta.setReadOnly(True)
        ta.setStyleSheet(
            "background:#1F2D3A; color:#E6E6E6;"
            "font-family:Consolas; font-size:10px;")
        dt = (analysis_text
              if analysis_text
              and analysis_text.strip()
              else self
              ._load_factor_output_text_from_excel())
        if not dt:
            dt = ("ไม่พบข้อความผลการวิเคราะห์ "
                  "(ลองรันใหม่อีกครั้ง)")
        ta.setPlainText(dt)
        tabs.addTab(ta, " ผลการวิเคราะห์ ")

        # Tab 2 - Variable descriptions
        desc_scroll = QScrollArea()
        desc_scroll.setWidgetResizable(True)
        desc_w = QWidget()
        desc_vl = QVBoxLayout(desc_w)
        desc_vl.setContentsMargins(20, 20, 20, 20)
        desc_scroll.setWidget(desc_w)
        tabs.addTab(desc_scroll, " คำอธิบายตัวแปร ")

        hdr = QLabel(
            "คำอธิบายและตัวแปรที่เลือกใน Model")
        hdr.setStyleSheet(
            "color:#1565C0; font-size:16px;"
            "font-weight:700;")
        desc_vl.addWidget(hdr)

        descriptions = {
            "S (Sense)": "การรับรู้ผ่านประสาทสัมผัส",
            "P (Personality/People)": "บุคลิกภาพของแบรนด์",
            "C (Cognition)": "การรับรู้เชิงเหตุผล",
            "A (Action/Attitude)": "พฤติกรรม/ทัศนคติ",
            "E (Emotion)": "อารมณ์ความรู้สึก",
            "AgreeS / AgreeP": "วัดความเห็นด้วย (%T2B)",
        }
        c_vars_d = (self.c_vars_to_compute
                    if self.c_vars_to_compute
                    else self.computed_c_cols)
        all_vars = {
            "S": self.vars_to_transform.get('S', []),
            "P": self.vars_to_transform.get('P', []),
            "C": c_vars_d,
            "A": self.vars_to_transform.get('A', []),
            "E": self.vars_to_transform.get('E', []),
            "AgreeS": self.vars_to_transform.get(
                'AgreeS', []),
            "AgreeP": self.vars_to_transform.get(
                'AgreeP', []),
        }
        key_map = {
            "S (Sense)": "S",
            "P (Personality/People)": "P",
            "C (Cognition)": "C",
            "A (Action/Attitude)": "A",
            "E (Emotion)": "E",
            "AgreeS / AgreeP": ["AgreeS", "AgreeP"],
        }
        for vd, desc in descriptions.items():
            vlb = QLabel(vd)
            vlb.setStyleSheet(
                "color:#1565C0; font-size:13px;"
                "font-weight:700;")
            desc_vl.addWidget(vlb)
            dlb = QLabel(desc)
            dlb.setWordWrap(True)
            dlb.setContentsMargins(10, 0, 0, 0)
            desc_vl.addWidget(dlb)

            dk = key_map[vd]
            vlist = []
            if isinstance(dk, list):
                for k in dk:
                    if all_vars.get(k):
                        vlist += [f"--- {k} ---"]
                        vlist += all_vars.get(k, [])
            else:
                vlist = all_vars.get(dk, [])
            if vlist:
                vte = QTextEdit()
                vte.setReadOnly(True)
                vte.setPlainText("\n".join(vlist))
                h = min(len(vlist), 10) * 18 + 10
                vte.setFixedHeight(h)
                desc_vl.addWidget(vte)
        desc_vl.addStretch()

    # ===================================================================
    # ANALYSIS AND EXPORT (STEP 2)
    # ===================================================================
    def open_label_editor(self):
        if self.transformed_df is None:
            self._msg_error(
                "ยังไม่มีข้อมูล",
                "ยังไม่มีข้อมูลที่ประมวลผลแล้ว")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("กำหนด Label")
        dlg.resize(600, 500)
        dlg.setModal(True)
        dlg.setStyleSheet(_DLG_QSS)
        vl = QVBoxLayout(dlg)

        sa = QScrollArea()
        sa.setWidgetResizable(True)
        sw = QWidget()
        gl = QGridLayout(sw)
        sa.setWidget(sw)
        vl.addWidget(sa, 1)

        index1_entries = {}
        hdr_code = QLabel("<b>Code</b>")
        hdr_code.setProperty("class", "dlg-header")
        gl.addWidget(hdr_code, 0, 0)
        hdr_label = QLabel("<b>Label</b>")
        hdr_label.setProperty("class", "dlg-header")
        gl.addWidget(hdr_label, 0, 1)

        unique_idx = sorted(
            self.transformed_df['Index1']
            .dropna().unique())
        for i, code in enumerate(unique_idx):
            code = int(code)
            gl.addWidget(QLabel(str(code)), i + 1, 0)
            entry = QLineEdit()
            if code in self.index1_labels:
                entry.setText(
                    self.index1_labels[code])
            gl.addWidget(entry, i + 1, 1)
            index1_entries[code] = entry

        def save_labels():
            self.index1_labels.clear()
            for cd, ent in index1_entries.items():
                t = ent.text().strip()
                if t:
                    self.index1_labels[cd] = t
            self._msg_success(
                "บันทึก Labels เรียบร้อยแล้ว")
            dlg.accept()

        btn = QPushButton("  ✔  บันทึก Labels  ")
        btn.setStyleSheet(_BTN_STYLES["success"])
        btn.setMinimumHeight(40)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.clicked.connect(save_labels)
        vl.addWidget(btn)

        self._center_toplevel(dlg)
        dlg.exec()

    def run_analysis_and_export(self, automated=False):
        """เตรียมการบน main thread แล้วส่งงานหนักไปให้ worker thread"""
        if self.transformed_df is None:
            self._msg_error("ยังไม่มีข้อมูลสำหรับวิเคราะห์",
                "ไม่พบข้อมูลที่แปลงแล้ว (Transformed Data)")
            return

        self.update_status("กำลังเตรียมการวิเคราะห์...")

        # snapshot ค่าจาก widget ก่อน แล้ว logic จะอ่านจาก snapshot เท่านั้น
        self._snapshot_ui_inputs()
        cross_filters = self._cross_filters()

        if not cross_filters and not automated:
            ret = self._msg_ask(
                "ยังไม่ได้ระบุ Filter ไขว้",
                "จะวิเคราะห์เฉพาะ Index1 อย่างเดียว\n"
                "ต้องการดำเนินการต่อหรือไม่?")
            if not ret:
                self.update_status("ยกเลิกโดยผู้ใช้", "warning")
                return

        self.show_log_panel("กำลังวิเคราะห์ข้อมูล...")
        self.start_progress()

        def _done(payload):
            self.stop_progress()
            if payload is None:
                return
            final_output, final_message = payload
            self.update_status(
                "วิเคราะห์และส่งออกเสร็จสมบูรณ์", "success")

            alerts = []
            if self._analysis_errors:
                alerts.append(
                    f"⚠ Factor/Regression ล้มเหลว "
                    f"{len(self._analysis_errors)} กลุ่ม "
                    "(B.S/B.P/B.C/B.E ของกลุ่มนั้นเป็น 0)")
            if self._beta_zero_groups:
                alerts.append(
                    f"⚠ {len(self._beta_zero_groups)} กลุ่มมี "
                    "B.S–B.E = 0 ทั้งหมด → Index = 0 ด้วย")

            if self._weak_model_groups:
                alerts.append(
                    f"⚠ {len(self._weak_model_groups)} กลุ่มที่โมเดลอธิบาย"
                    f"ข้อมูลได้น้อย (R² < {self._RELIABILITY_R2:.2f})\n"
                    "     ทั้งแถวเชื่อถือไม่ได้ ไม่ควรนำไปตีความ")
            if self._weak_beta_cells:
                alerts.append(
                    f"⚠ {len(self._weak_beta_cells)} ช่องที่ค่า B "
                    f"แยกจากศูนย์ไม่ได้ทางสถิติ (p > {self._RELIABILITY_P})\n"
                    "     คำนวณได้แต่เป็น noise — อย่าตีความว่า "
                    "'ปัจจัยนี้ไม่สำคัญ'")

            info = []
            if self._beta_warnings:
                info.append(
                    f"ℹ {len(self._beta_warnings)} กลุ่มมี beta ติดลบ "
                    "— แปลงเป็นค่าสัมบูรณ์ (ABS) ให้แล้ว B.S–B.E "
                    "จึงอยู่ในช่วง 0–100 ปกติ")

            if alerts:
                self._msg_warn(
                    "เสร็จแล้ว แต่มีข้อควรระวัง",
                    f"{final_message}\n\n"
                    + "\n\n".join(alerts + info)
                    + "\n\nดูรายละเอียดรายกลุ่มได้ที่ Log ด้านขวา")
            elif info:
                self._msg_success(
                    "วิเคราะห์และส่งออกเสร็จสมบูรณ์",
                    f"{final_message}\n\n" + "\n\n".join(info))
            else:
                self._msg_success(
                    "วิเคราะห์และส่งออกเสร็จสมบูรณ์", final_message)
            QTimer.singleShot(
                300,
                lambda: self.display_analysis_tabs(final_output))

        self._run_in_thread(
            lambda: self._analysis_pipeline(cross_filters),
            _done)

    def _analysis_pipeline(self, cross_filters):
        """งานวิเคราะห์+ส่งออกทั้งหมด — รันบน worker thread เท่านั้น"""
        primary_filter = "Index1"
        if not cross_filters:
            cross_filters = ['']

        self._analysis_errors = []
        self._analysis_skipped = []
        self._beta_warnings = []
        self._beta_zero_groups = []
        self._beta_abs_used = False
        self._weak_beta_cells = []
        self._weak_model_groups = []
        self._sample_size_approx = False
        self._respondent_key = _UNSET
        start_time = time.time()

        self.log_message("=" * 50)
        self.log_message("เริ่มกระบวนการวิเคราะห์และส่งออก")
        self.log_message("=" * 50)
        self.log_message(f"Primary Filter: {primary_filter}")
        cf_display = ', '.join(cross_filters) if cross_filters[0] else '(ไม่ระบุ)'
        self.log_message(f"Cross Filter(s): {cf_display}")
        self.log_message(
            f"E Correlation Mode: {self._ui.get('e_group_mode', 'default')}")
        self.log_message(
            "Correlation Mode: ABS (cor_S_* / cor_P_* / CorE_* "
            "ใช้ค่าสัมบูรณ์)")
        self.log_message(
            "B.S–B.E Mode: ABS (ใช้ |beta| — ไม่มีค่าติดลบ/เกิน 100)")
        self.log_message("")

        total_filters = len(cross_filters)
        steps_per_filter = 3
        total_steps = total_filters * steps_per_filter + 1
        current_step = 0
        self.set_progress(0, total_steps)

        all_summary_parts = []
        all_results = OrderedDict()
        all_output_parts = []
        use_json_agree_cache = False
        if self.is_reanalyze_mode:
            use_json_agree_cache = \
                self._load_agree_summary_cache_from_json()
            if use_json_agree_cache:
                self.log_message(
                    "โหลด Agree Original จาก JSON สำเร็จ")
            else:
                self.log_message(
                    "ไม่พบ Agree Original JSON (ยกเลิก: โหมดนี้ห้ามคำนวณใหม่)")
                raise RuntimeError(
                    "โหมดวิเคราะห์ซ้ำต้องใช้ค่า Agree จาก JSON เท่านั้น")

        for idx, cross_filter in enumerate(cross_filters):
            f_label = cross_filter if cross_filter else '(ไม่ระบุ)'
            if total_filters > 1:
                self.log_message(f"━━━ Filter {idx+1}/{total_filters}: {f_label} ━━━")
                self.log_message("")

            # --- Summary ---
            current_step += 1
            self.update_status(f"สร้าง Summary ({f_label})...")
            self.log_message(f"[{current_step}/{total_steps}] กำลังสร้าง Summary ({f_label})...")
            try:
                part_summary = self._create_summary_df_logic(
                    primary_filter=primary_filter,
                    cross_filter=cross_filter
                )
            except Exception as e:
                self.log_message(f"   ✗ สร้าง Summary ไม่สำเร็จ: {e}")
                if total_filters == 1:
                    raise
                current_step += 2
                self.set_progress(current_step, total_steps)
                continue
            self.log_message(f"   ✓ Summary สำเร็จ ({len(part_summary)} แถว)")
            self.set_progress(current_step, total_steps)

            # --- T2B ---
            current_step += 1
            self.log_message("")
            self.log_message(f"[{current_step}/{total_steps}] กำลังคำนวณ T2B ({f_label})...")
            try:
                if use_json_agree_cache:
                    # Re-analyze ใช้ JSON อย่างเดียว (ไม่คำนวณใหม่)
                    self.log_message(
                        "   ✓ ข้ามการคำนวณ T2B (จะวาง Agree จาก JSON ทั้งชุดหลังรวมผล)")
                else:
                    part_summary = self._calculate_and_add_t2b_values(
                        part_summary,
                        primary_filter=primary_filter,
                        cross_filter=cross_filter
                    )
                    self.log_message("   ✓ T2B สำเร็จ")
            except Exception as e:
                self.log_message(f"   ⚠ ข้ามการคำนวณ T2B: {e}")
            self.set_progress(current_step, total_steps)

            # --- Factor & Regression ---
            current_step += 1
            self.log_message("")
            self.update_status(f"รัน Factor & Regression ({f_label})...")
            self.log_message(f"[{current_step}/{total_steps}] กำลังรัน Factor & Regression ({f_label})...")
            try:
                part_results, part_output = self._run_factor_regression_logic(
                    primary_filter=primary_filter,
                    cross_filter=cross_filter
                )
            except Exception as e:
                self.log_message(f"   ✗ Factor/Regression ไม่สำเร็จ: {e}")
                if total_filters == 1:
                    raise
                self.set_progress(current_step, total_steps)
                continue
            self.log_message(f"   ✓ วิเคราะห์สำเร็จ ({len(part_results)} กลุ่ม)")
            self.set_progress(current_step, total_steps)

            # --- ตัด Overall และ Index1-only ออกจาก filter ตัวที่ 2 เป็นต้นไป ---
            if idx > 0 and part_summary is not None:
                dup_mask = part_summary['Filter'].apply(
                    lambda x: x == 'Overall' or (x.startswith('Index1=') and '+' not in x)
                )
                part_summary = part_summary[~dup_mask].reset_index(drop=True)
                if part_results:
                    dup_keys = [k for k in part_results if k == 'Overall' or (k.startswith('Index1=') and '+' not in k)]
                    for k in dup_keys:
                        part_results.pop(k, None)

            all_summary_parts.append(part_summary)
            all_results.update(part_results or {})
            all_output_parts.append(part_output or '')
            self.log_message("")

        # --- รวมผลลัพธ์ทั้งหมด ---
        if not all_summary_parts:
            raise RuntimeError("ไม่มีผลลัพธ์ที่สร้างได้")

        final_summary = pd.concat(all_summary_parts, ignore_index=True)
        final_output = '\n'.join(all_output_parts)

        if use_json_agree_cache:
            n_json = len(self.agree_summary_cache_df) \
                if self.agree_summary_cache_df is not None else 0
            n_sum = len(final_summary)

            # จับคู่ด้วยคีย์ (Index1 + cross filter) ก่อน — วิธีนี้ทนกรณี
            # จำนวนแถวไม่เท่ากัน เช่นเงื่อนไข Good ตัดบางกลุ่มหายไปหมด
            keyed, matched, cells, unmatched = \
                self._apply_agree_summary_cache_by_key(final_summary)

            if matched:
                final_summary = keyed
                self.log_message(
                    f"✓ Add Agree Original แล้ว "
                    f"(จับคู่ด้วยคีย์ {matched}/{n_sum} แถว, {cells} ค่า)")
                if n_json != n_sum:
                    self.log_message(
                        f"   หมายเหตุ: JSON มี {n_json} แถว "
                        f"แต่ Summary มี {n_sum} แถว "
                        "(ปกติเมื่อเงื่อนไข Good ตัดบางกลุ่มออก)")
                if unmatched:
                    self.log_message(
                        f"   ⚠ ไม่พบค่า Agree ของ {len(unmatched)} แถว "
                        "(ปล่อยว่างไว้):")
                    for name in unmatched[:10]:
                        self.log_message(f"      - {name}")
                    if len(unmatched) > 10:
                        self.log_message(
                            f"      ... และอีก {len(unmatched) - 10} แถว")
            else:
                # จับคู่ด้วยคีย์ไม่ได้เลย -> ลองวางตามลำดับแบบเดิม
                # (รองรับไฟล์ JSON เก่าที่คีย์ไม่เข้ากัน)
                final_summary, copied_cells, copy_status = \
                    self._apply_agree_summary_cache_by_position(
                        final_summary)
                if copy_status != "ok":
                    self.update_status(
                        "วาง Agree จาก JSON ไม่สำเร็จ", "danger")
                    if copy_status.startswith("row_mismatch:"):
                        _, c_rows, s_rows = copy_status.split(":")
                        raise RuntimeError(
                            "ไม่สามารถจับคู่ค่า Agree จาก JSON กับ Summary ได้\n"
                            f"JSON={c_rows} แถว, Summary={s_rows} แถว\n\n"
                            "สาเหตุที่พบบ่อย: ไฟล์ JSON มาจากงานอื่น "
                            "หรือใช้ Filter ไม่ตรงกับรอบที่สร้าง JSON\n"
                            "กรุณาตรวจสอบว่าเลือกไฟล์ Agree Original JSON "
                            "ของงานเดียวกัน และตั้ง Filter ให้ตรงกัน")
                    raise RuntimeError(
                        "ไม่สามารถวาง Agree Original จาก JSON ได้")
                self.log_message(
                    f"✓ Add Agree Original แล้ว "
                    f"(วางตรงตามลำดับ {copied_cells} ค่า)")

        # --- บันทึก Excel ---
        current_step += 1
        self.log_message("")
        self.update_status("กำลังบันทึกผลลัพธ์ลง Excel...")
        self.log_message(f"[{current_step}/{total_steps}] กำลังบันทึกผลลัพธ์ลง Excel...")
        final_message = self.save_all_results_to_excel(
            final_summary, all_results, final_output)
        self.log_message("   ✓ บันทึก Excel สำเร็จ")
        self.set_progress(current_step, total_steps)

        # --- สรุปกลุ่มที่วิเคราะห์ไม่ได้ ---
        if self._analysis_skipped:
            self.log_message("")
            self.log_message(
                f"ℹ ข้าม {len(self._analysis_skipped)} กลุ่ม (ข้อมูลไม่พอ):")
            for name, reason in self._analysis_skipped[:10]:
                self.log_message(f"    - {name}: {reason}")
            if len(self._analysis_skipped) > 10:
                self.log_message(
                    f"    ... และอีก {len(self._analysis_skipped) - 10} กลุ่ม")

        if self._analysis_errors:
            self.log_message("")
            self.log_message(
                f"⚠ Factor/Regression ล้มเหลว "
                f"{len(self._analysis_errors)} กลุ่ม "
                "(B.S/B.P/B.C/B.E ของกลุ่มนั้นจะเป็น 0):")
            for name, reason in self._analysis_errors[:10]:
                self.log_message(f"    - {name}: {reason}")
            if len(self._analysis_errors) > 10:
                self.log_message(
                    f"    ... และอีก {len(self._analysis_errors) - 10} กลุ่ม")

        if self._sample_size_approx:
            self.log_message("")
            self.log_message(
                "⚠ SampleSize เป็นค่าประมาณ เพราะไฟล์ไม่มีคอลัมน์ระดับ"
                "ผู้ตอบ (เช่น RESPID) ให้ใช้ตัดข้อมูลซ้ำ")
            self.log_message(
                "   ใช้จำนวนแถวสูงสุดของ Index1 เดียวแทน "
                "ซึ่งจะถูกต้องเมื่อผู้ตอบทุกคนตอบครบทุก Index1")

        if self._beta_zero_groups:
            self.log_message("")
            self.log_message(
                f"⚠ {len(self._beta_zero_groups)} กลุ่มมี B.S–B.E = 0 "
                "ทั้งหมด (ไม่มีผล Regression) → Index = 0 ด้วย:")
            for name in self._beta_zero_groups[:10]:
                self.log_message(f"    - {name}")
            if len(self._beta_zero_groups) > 10:
                self.log_message(
                    f"    ... และอีก {len(self._beta_zero_groups) - 10} กลุ่ม")

        if self._beta_warnings:
            self.log_message("")
            self.log_message(
                f"ℹ {len(self._beta_warnings)} กลุ่มมี beta ติดลบ "
                "— แปลงเป็นค่าสัมบูรณ์ (ABS) แล้ว")
            self.log_message(
                "   B.x = |beta| / ผลรวม|beta| × 100 "
                "จึงอยู่ในช่วง 0–100 และรวมกันได้ 100")
            self.log_message(
                "   หมายเหตุ: ปัจจัยที่สัมพันธ์กลับทางกับ A "
                "จะถูกนับเป็น 'สำคัญ' เท่ากับปัจจัยที่สัมพันธ์ตามทาง")
            for w in self._beta_warnings[:10]:
                lo, hi = w['span']
                self.log_message(
                    f"    - {w['filter']}: beta ติดลบ {w['n_negative']} ตัว, "
                    f"ผลรวม|beta| = {w['total']:.4f}, "
                    f"ช่วง B = {lo:.1f} ถึง {hi:.1f}")
            if len(self._beta_warnings) > 10:
                self.log_message(
                    f"    ... และอีก {len(self._beta_warnings) - 10} กลุ่ม")

        if self._weak_model_groups:
            self.log_message("")
            self.log_message(
                f"⚠ {len(self._weak_model_groups)} กลุ่มที่โมเดลอธิบาย"
                f"ข้อมูลได้น้อย (R² < {self._RELIABILITY_R2:.2f})")
            self.log_message(
                "   ทั้งแถวเชื่อถือไม่ได้ ไม่ควรนำ B.S–B.E / Index "
                "ไปตีความเป็นข้อค้นพบ:")
            for g in self._weak_model_groups[:10]:
                self.log_message(
                    f"    - {g['filter']}: R² = {g['r2']:.3f} "
                    f"(n = {g['n']})")
            if len(self._weak_model_groups) > 10:
                self.log_message(
                    f"    ... และอีก {len(self._weak_model_groups) - 10} กลุ่ม")

        if self._weak_beta_cells:
            self.log_message("")
            self.log_message(
                f"⚠ {len(self._weak_beta_cells)} ช่องที่ค่า B "
                f"แยกจากศูนย์ไม่ได้ทางสถิติ (p > {self._RELIABILITY_P})")
            self.log_message(
                "   คำนวณได้ แต่เป็น noise ไม่ใช่สัญญาณจริง "
                "— มักเกิดเมื่อตัวชี้วัดนั้น")
            self.log_message(
                "   แทบไม่มีความหลากหลายในกลุ่มย่อยนั้น "
                "(อย่าตีความว่า 'ปัจจัยนี้ไม่สำคัญ'):")
            for c in sorted(self._weak_beta_cells,
                            key=lambda x: -x['p'])[:12]:
                self.log_message(
                    f"    - {c['filter']} | {c['ratio_col']} = "
                    f"{c['value']:.2f}  (p = {c['p']:.3f}, n = {c['n']})")
            if len(self._weak_beta_cells) > 12:
                self.log_message(
                    f"    ... และอีก {len(self._weak_beta_cells) - 12} ช่อง")

        elapsed = time.time() - start_time
        self.log_message("")
        self.log_message("=" * 50)
        self.log_message(f"เสร็จสมบูรณ์ (ใช้เวลา {elapsed:.1f} วินาที)")
        self.log_message("=" * 50)

        return final_output, final_message


    def _detect_respondent_key(self):
        """หาคอลัมน์ที่ทำหน้าที่เป็นรหัสผู้ตอบใน transformed_df

        เกณฑ์: ภายใน Index1 เดียวกัน ค่าต้องไม่ซ้ำเลย
        คอลัมน์อย่าง GENDER หรือ Good จะไม่ผ่านเกณฑ์นี้ ทำให้เลือกได้
        เฉพาะรหัสผู้ตอบจริง (เช่น RESPID) ผลลัพธ์ถูก cache ไว้ใช้ซ้ำ
        """
        if self._respondent_key is not _UNSET:
            return self._respondent_key

        self._respondent_key = None
        df = self.transformed_df
        if df is None or df.empty:
            return None

        candidates = [c for c in self.id_vars if c in df.columns]
        if not candidates:
            return None

        if 'Index1' in df.columns:
            slices = [sub for _, sub in df.groupby('Index1')]
        else:
            slices = [df]

        for col in candidates:
            if df[col].isna().any():
                continue
            if all(not sub[col].duplicated().any() for sub in slices):
                self._respondent_key = col
                break
        return self._respondent_key

    def _unique_sample_size(self, df_group):
        """จำนวนผู้ตอบจริงของกลุ่ม (ไม่นับซ้ำ)

        transformed_df เป็น long format คือ 1 แถวต่อผู้ตอบ 1 คน
        ต่อ 1 Index1 ดังนั้นจำนวนแถวจึงไม่ใช่จำนวนผู้ตอบ
        """
        if df_group is None or df_group.empty:
            return 0

        # 1) ถ้าหารหัสผู้ตอบได้ ให้นับค่าไม่ซ้ำของคอลัมน์นั้นตรงๆ
        key = self._detect_respondent_key()
        if key and key in df_group.columns:
            return int(df_group[key].nunique(dropna=True))

        # 2) ไม่มีรหัสผู้ตอบชัดเจน -> dedup ด้วย id_vars ทั้งชุด
        #    (วิธีเดียวกับที่ใช้หาฐาน %T2B) แต่ถ้ามี id_var ตัวใดไม่คงที่
        #    ภายในผู้ตอบคนเดียว ตัวเลขจะสูงเกินจริง จึงถือว่าเป็นค่าประมาณ
        keys = [c for c in self.id_vars if c in df_group.columns]
        if keys:
            self._sample_size_approx = True
            return int(len(df_group[keys].drop_duplicates()))

        # 3) ไม่มีคอลัมน์ระดับผู้ตอบเลย (ไฟล์มีแต่คอลัมน์โมเดล)
        #    ประมาณจากจำนวนแถวสูงสุดของ Index1 เดียว
        self._sample_size_approx = True
        if 'Index1' in df_group.columns:
            per_index = df_group.groupby('Index1').size()
            if len(per_index):
                return int(per_index.max())
        return int(len(df_group))

    def _create_summary_df_logic(self, primary_filter, cross_filter):
        """ตรรกะการสร้าง Summary DataFrame"""
        try:
            cols_to_average = [col for col in self.transformed_df.columns if re.match(r'^(S|P|C|E)_\d+$', col)]
            if not cols_to_average:
                raise RuntimeError(
                    "ไม่พบข้อมูลคอลัมน์ S, P, C, E สำหรับสร้างสรุป")
            corr_df = self.transformed_df.copy()

            # --- E Group Mode: E columns ถูก merge แล้วจาก transformation (เช่น E_45) ---
            # ไม่ต้อง merge ซ้ำที่นี่

            df_for_summary = self.transformed_df
            groups_to_summarize = OrderedDict()
            corr_groups = OrderedDict()
            groups_to_summarize['Overall'] = df_for_summary
            corr_groups['Overall'] = corr_df

            primary_values = []
            if primary_filter and primary_filter in df_for_summary.columns:
                primary_values = sorted(df_for_summary[primary_filter].dropna().unique())
                for p_val in primary_values:
                    filter_name = self._format_filter_val(primary_filter, p_val)
                    if filter_name not in groups_to_summarize:
                            groups_to_summarize[filter_name] = df_for_summary[df_for_summary[primary_filter] == p_val]
                            corr_groups[filter_name] = corr_df[corr_df[primary_filter] == p_val]

            if cross_filter and cross_filter in df_for_summary.columns:
                cross_values = sorted(df_for_summary[cross_filter].dropna().unique())
                for c_val in cross_values:
                    filter_name_cross = self._format_filter_val(cross_filter, c_val)
                    if filter_name_cross not in groups_to_summarize:
                            groups_to_summarize[filter_name_cross] = df_for_summary[df_for_summary[cross_filter] == c_val]
                            corr_groups[filter_name_cross] = corr_df[corr_df[cross_filter] == c_val]

                    if primary_filter and primary_filter in df_for_summary.columns:
                        for p_val in primary_values:
                            nested_name = f"{self._format_filter_val(primary_filter, p_val)}+{self._format_filter_val(cross_filter, c_val)}"
                            subset = df_for_summary[(df_for_summary[primary_filter] == p_val) & (df_for_summary[cross_filter] == c_val)]
                            groups_to_summarize[nested_name] = subset
                            corr_groups[nested_name] = corr_df[(corr_df[primary_filter] == p_val) & (corr_df[cross_filter] == c_val)]

            summary_list = []
            avg_cols_base = cols_to_average.copy()
            if 'A' in df_for_summary.columns: avg_cols_base.append('A')
            if 'ZA' in df_for_summary.columns: avg_cols_base.append('ZA')

            for name, df_group in groups_to_summarize.items():
                    if not df_group.empty:
                        avg_values = df_group[avg_cols_base].mean()
                        summary_row_df = pd.DataFrame([avg_values])
                        summary_row_df['Filter'] = name

                        index1_val = 0
                        if name != 'Overall' and 'Index1' in df_group.columns:
                            unique_idx = df_group['Index1'].dropna().unique()
                            if len(unique_idx) == 1:
                                try:
                                    index1_val = int(unique_idx[0])
                                except (ValueError, TypeError):
                                    pass

                        summary_row_df['Index1'] = index1_val
                        summary_row_df['SampleSize'] = \
                            self._unique_sample_size(df_group)
                        summary_list.append(summary_row_df)

            if not summary_list:
                raise RuntimeError(
                    "ไม่พบข้อมูลสำหรับสร้างสรุปตาม Filter ที่กำหนด")

            final_summary_df = pd.concat(summary_list, ignore_index=True)

            def map_labels(row):
                filter_str = row['Filter']
                if filter_str == 'Overall': return 'Overall'

                parts = filter_str.split('+')
                labels_found = []
                for part in parts:
                    if '=' in part:
                        var_name, val_part = part.split('=', 1)
                        if var_name == 'Index1':
                            try:
                                code = int(float(val_part))
                                lbl = self.index1_labels.get(code)
                                if lbl:
                                    labels_found.append(lbl)
                                    continue
                            except (ValueError, TypeError):
                                pass
                        labels_found.append(val_part)
                    else:
                        labels_found.append(part)

                if labels_found:
                    return ' - '.join(labels_found)
                return filter_str

            final_summary_df['Labe Index1'] = final_summary_df.apply(map_labels, axis=1)

            for col in cols_to_average:
                if col in final_summary_df.columns:
                    final_summary_df[col] *= 100

            s_cols = sorted([c for c in cols_to_average if c.startswith('S_')], key=lambda x: int(x.split('_')[1]))
            p_cols = sorted([c for c in cols_to_average if c.startswith('P_')], key=lambda x: int(x.split('_')[1]))

            if 'A' in corr_df.columns:
                e_cols_for_corr = sorted([c for c in corr_df.columns if re.match(r'^E_\d+$', c)], key=lambda x: int(x.split('_')[1]))
                e_corr_map = {f'CorE_{col.split("_")[1]}': col for col in e_cols_for_corr}
                source_e_cols = [col for col in e_corr_map.values() if col in corr_df.columns]
                rename_dict = {v: k for k, v in e_corr_map.items()}

                corr_rows = []
                for name, df_group in corr_groups.items():
                    if df_group.empty:
                        continue
                    # ใช้ค่าสัมบูรณ์ของ correlation ทุกรอบการรัน
                    # เพื่อไม่ให้ cor_S_* / cor_P_* / CorE_* ติดลบ
                    # (ดูความแรงของความสัมพันธ์ ไม่สนทิศทาง)
                    row = {'Filter': name}
                    if s_cols:
                        s_corr = df_group[s_cols].corrwith(
                            df_group['A']).abs()
                        for col, val in s_corr.items():
                            row['cor_' + col] = val
                    if p_cols:
                        p_corr = df_group[p_cols].corrwith(
                            df_group['A']).abs()
                        for col, val in p_corr.items():
                            row['cor_' + col] = val
                    if source_e_cols:
                        e_corr = df_group[source_e_cols].corrwith(
                            df_group['A']).abs()
                        for col, val in e_corr.items():
                            row[rename_dict.get(col, col)] = val
                    corr_rows.append(row)

                if corr_rows:
                    corr_by_filter_df = pd.DataFrame(corr_rows)
                    final_summary_df = pd.merge(final_summary_df, corr_by_filter_df, on='Filter', how='left')

            return final_summary_df
        except RuntimeError:
            raise
        except Exception as e:
            raise RuntimeError(
                f"ไม่สามารถสร้างข้อมูลสรุปได้: {e}") from e

    def _calculate_and_add_t2b_values(self, summary_df, primary_filter="Index1", cross_filter=None):
        """คำนวณ %T2B สำหรับ AgreeS/P ตามลำดับการเลือก"""
        agree_s_vars = self.vars_to_transform.get('AgreeS', [])
        agree_p_vars = self.vars_to_transform.get('AgreeP', [])

        s_cols_in_summary = sorted([c for c in summary_df.columns if c.startswith('S_') and 'cor' not in c and 'agree' not in c], key=lambda x: int(x.split('_')[1]))
        p_cols_in_summary = sorted([c for c in summary_df.columns if c.startswith('P_') and 'cor' not in c and 'agree' not in c], key=lambda x: int(x.split('_')[1]))

        for s_col in s_cols_in_summary:
            agree_col_name = 'agree_' + s_col
            if agree_col_name not in summary_df.columns:
                summary_df[agree_col_name] = np.nan
        for p_col in p_cols_in_summary:
            agree_col_name = 'agree_' + p_col
            if agree_col_name not in summary_df.columns:
                summary_df[agree_col_name] = np.nan

        if not agree_s_vars and not agree_p_vars:
            if not self.c_vars_to_compute:
                print("คำเตือน: ข้ามการคำนวณ T2B เนื่องจากไม่ได้เริ่มจากไฟล์ SPSS ดั้งเดิม")
            return summary_df

        if self.transformed_df is None:
            raise ValueError("ไม่พบข้อมูล SPSS ที่ผ่านการประมวลผล (self.transformed_df) สำหรับคำนวณ T2B")

        if not self.id_vars:
            raise ValueError("Identifier variables (id_vars) not found.")

        t2b_choice = self._ui.get(
            't2b_choice', self.t2b_choice_var.get())
        good_codes = [5, 4] if t2b_choice == "5+4" else [1, 2]
        # รองรับกรณีชื่อจากไฟล์ Setting มี space แฝง
        tr_col_lookup = {
            str(c).strip(): c for c in self.transformed_df.columns
        }

        if cross_filter is None:
            cross_filter = ''

        groups_to_summarize = OrderedDict()
        groups_to_summarize['Overall'] = self.transformed_df

        primary_values = []
        if primary_filter and primary_filter in self.transformed_df.columns:
            primary_values = sorted(self.transformed_df[primary_filter].dropna().unique())
            for p_val in primary_values:
                filter_name = self._format_filter_val(primary_filter, p_val)
                if filter_name not in groups_to_summarize:
                    groups_to_summarize[filter_name] = self.transformed_df[self.transformed_df[primary_filter] == p_val]

        if cross_filter and cross_filter in self.transformed_df.columns:
            cross_values = sorted(self.transformed_df[cross_filter].dropna().unique())
            for c_val in cross_values:
                filter_name_cross = self._format_filter_val(cross_filter, c_val)
                if filter_name_cross not in groups_to_summarize:
                    groups_to_summarize[filter_name_cross] = self.transformed_df[self.transformed_df[cross_filter] == c_val]

                if primary_filter and primary_filter in self.transformed_df.columns:
                    for p_val in primary_values:
                        nested_name = f"{self._format_filter_val(primary_filter, p_val)}+{self._format_filter_val(cross_filter, c_val)}"
                        subset = self.transformed_df[
                            (self.transformed_df[primary_filter] == p_val) &
                            (self.transformed_df[cross_filter] == c_val)
                        ]
                        groups_to_summarize[nested_name] = subset

        for name, df_group in groups_to_summarize.items():
            if df_group.empty:
                continue
            row_mask = summary_df['Filter'] == name
            if not row_mask.any():
                continue

            resolved_agree_s = [
                tr_col_lookup.get(str(v).strip(), str(v).strip())
                for v in agree_s_vars
            ]
            resolved_agree_p = [
                tr_col_lookup.get(str(v).strip(), str(v).strip())
                for v in agree_p_vars
            ]
            agree_cols_all = [
                c for c in (resolved_agree_s + resolved_agree_p)
                if c in df_group.columns
            ]
            dedup_keys = [
                c for c in self.id_vars
                if c in df_group.columns and c not in agree_cols_all
            ]

            if agree_cols_all:
                base_source_df = df_group[
                    dedup_keys + agree_cols_all
                ].drop_duplicates(subset=dedup_keys) \
                    if dedup_keys else \
                    df_group[agree_cols_all].drop_duplicates()
            else:
                base_source_df = df_group[
                    dedup_keys
                ].drop_duplicates() if dedup_keys else pd.DataFrame()

            total_base = len(base_source_df)
            if total_base == 0:
                continue

            for i, s_col in enumerate(s_cols_in_summary):
                agree_col_name = 'agree_' + s_col
                if i < len(agree_s_vars):
                    source_var = str(agree_s_vars[i]).strip()
                    source_var = tr_col_lookup.get(
                        source_var, source_var)
                    if source_var in base_source_df.columns:
                        t2b_sum = base_source_df[source_var].isin(good_codes).sum()
                        t2b_value = (t2b_sum / total_base) * 100 if total_base > 0 else 0
                        summary_df.loc[row_mask, agree_col_name] = t2b_value

            for i, p_col in enumerate(p_cols_in_summary):
                agree_col_name = 'agree_' + p_col
                if i < len(agree_p_vars):
                    source_var = str(agree_p_vars[i]).strip()
                    source_var = tr_col_lookup.get(
                        source_var, source_var)
                    if source_var in base_source_df.columns:
                        t2b_sum = base_source_df[source_var].isin(good_codes).sum()
                        t2b_value = (t2b_sum / total_base) * 100 if total_base > 0 else 0
                        summary_df.loc[row_mask, agree_col_name] = t2b_value

        return summary_df

    def _run_factor_regression_logic(self, primary_filter, cross_filter):
        """ตรรกะการรัน Factor และ Regression"""
        df_for_analysis = self.transformed_df
        all_cols = list(df_for_analysis.columns)

        if primary_filter and primary_filter not in all_cols: primary_filter = ""
        if cross_filter and cross_filter not in all_cols: cross_filter = ""
        if primary_filter and primary_filter == cross_filter:
            raise RuntimeError(
                "Filter หลัก และ Filter ไขว้ ต้องเป็นคนละคอลัมน์")

        results_for_saving = OrderedDict()
        old_stdout = sys.stdout; sys.stdout = captured_output = io.StringIO()
        try:
            groups_to_analyze = OrderedDict()
            groups_to_analyze['Overall'] = df_for_analysis

            primary_values = []
            if primary_filter:
                primary_values = sorted(df_for_analysis[primary_filter].dropna().unique())
                for p_val in primary_values:
                    filter_name = self._format_filter_val(primary_filter, p_val)
                    groups_to_analyze[filter_name] = df_for_analysis[df_for_analysis[primary_filter] == p_val]

            if cross_filter:
                cross_values = sorted(df_for_analysis[cross_filter].dropna().unique())
                for c_val in cross_values:
                    filter_name_cross = self._format_filter_val(cross_filter, c_val)
                    if filter_name_cross not in groups_to_analyze:
                        groups_to_analyze[filter_name_cross] = df_for_analysis[df_for_analysis[cross_filter] == c_val]

                    if primary_filter:
                        for p_val in primary_values:
                            nested_name = f"{self._format_filter_val(primary_filter, p_val)}+{self._format_filter_val(cross_filter, c_val)}"
                            subset = df_for_analysis[(df_for_analysis[primary_filter] == p_val) & (df_for_analysis[cross_filter] == c_val)]
                            groups_to_analyze[nested_name] = subset

            for name, df_group in groups_to_analyze.items():
                sys.stdout.write(f"\n{'='*80}\n--- ผลการวิเคราะห์สำหรับ: {name} ---\n{'='*80}\n")
                if df_group.empty:
                    print("ไม่มีข้อมูลสำหรับกลุ่มนี้")
                    continue
                if results := self._run_single_analysis(
                        df_group.copy(), group_name=name):
                    results_for_saving[name] = results

            full_output_text = captured_output.getvalue()
            sys.stdout = old_stdout
            captured_output.close()
            return results_for_saving, full_output_text
        except Exception as e:
            sys.stdout = old_stdout; captured_output.close()
            if isinstance(e, RuntimeError):
                raise
            raise RuntimeError(
                f"เกิดข้อผิดพลาดระหว่างการวิเคราะห์ Factor/Regression: {e}") from e

    def _run_single_analysis(self, target_df, group_name=''):
        """รันการวิเคราะห์ 1 ชุด (Factor -> Regression)

        แยกความผิดพลาด 2 ประเภท:
        - ValueError = ข้อมูลไม่พอสำหรับกลุ่มย่อยนี้ (ปกติ, แค่ข้าม)
        - อื่นๆ = ความผิดพลาดจริง เช่น library ไม่ compatible
          ต้องเก็บไว้แจ้งผู้ใช้ ไม่ใช่กลืนเงียบจนได้ Beta = 0 ทั้งไฟล์
        """
        try:
            factor_scores_df, sorted_loadings_df, factor_to_variable_map = self.perform_factor_analysis(target_df)
            if factor_scores_df is not None:
                analysis_df = target_df.join(factor_scores_df)
                beta_df, beta_sorted_df, _, diagnostics = \
                    self.perform_regression_analysis(
                        analysis_df, factor_to_variable_map)
                return {'loadings': sorted_loadings_df, 'beta': beta_df,
                        'beta_sorted': beta_sorted_df,
                        'diagnostics': diagnostics}
        except ValueError as e:
            print(f"\n!!! ข้อมูลไม่พอสำหรับกลุ่มนี้: {e}\n!!! ข้ามการวิเคราะห์กลุ่มนี้...\n")
            self._analysis_skipped.append((group_name, str(e)))
        except Exception as e:
            detail = f"{type(e).__name__}: {e}"
            print(f"\n!!! เกิดข้อผิดพลาดในการวิเคราะห์กลุ่มนี้: {detail}\n"
                  f"{traceback.format_exc()}\n")
            self._analysis_errors.append((group_name, detail))
        return {}

    def save_settings(self):
        """บันทึกการตั้งค่าทั้งหมดลงใน Excel สองชีทโดยอัตโนมัติ"""
        if not self.original_filepath:
            self._msg_error("บันทึกการตั้งค่าไม่ได้",
                            "ยังไม่ได้โหลดไฟล์ SPSS ต้นฉบับ")
            return
        if not self.c_vars_to_compute and not any(self.vars_to_transform.values()):
            self._msg_error("ไม่มีอะไรให้บันทึก",
                            "ยังไม่ได้ตั้งค่าตัวแปร")
            return

        try:
            directory = os.path.dirname(self.original_filepath)
            filepath = os.path.join(directory, "Setting BS.xlsx")

            # --- Part 1: Settings Sheet ---
            settings_lists = {
                'C': self.c_vars_to_compute,
                'A': self.vars_to_transform.get('A', []),
                'S': self.vars_to_transform.get('S', []),
                'P': self.vars_to_transform.get('P', []),
                'E': self.vars_to_transform.get('E', []),
                'AgreeS': self.vars_to_transform.get('AgreeS', []),
                'AgreeP': self.vars_to_transform.get('AgreeP', [])
            }
            settings_df = pd.DataFrame({k: pd.Series(v) for k, v in settings_lists.items()})

            # --- E Group setting ---
            e_group_setting = "Default"
            if self.e_group_mode_var.get() == "group" and self.e_group_entry_var.get().strip():
                e_group_setting = self.e_group_entry_var.get().strip()

            # --- Multiple Filter_Var (แต่ละตัวอยู่คนละแถว) ---
            filter_text = self.filter_entry.text().strip()
            cross_filters = [f.strip() for f in filter_text.split(',') if f.strip()]
            max_len = max(len(settings_df), len(cross_filters), 1)
            settings_df = settings_df.reindex(range(max_len))

            settings_df.insert(0, 'Filter_Var', pd.Series(cross_filters))
            settings_df.insert(0, 'E_Group', e_group_setting)
            settings_df.insert(0, 'T2B_Choice', self.t2b_choice_var.get())
            settings_df.insert(0, 'PathFile', self.original_filepath)
            settings_df.loc[1:, ['PathFile', 'T2B_Choice', 'E_Group']] = ''

            # --- Part 2: Label Sheet ---
            index1_label_data = list(self.index1_labels.items())
            filter_label_data = list(self.filter_labels.get('labels', {}).items())
            sandp_df_for_save = self._build_sandp_df()
            if self.sandp_label_overrides:
                sandp_label_data = list(self.sandp_label_overrides)
            else:
                sandp_label_data = sandp_df_for_save[
                    'DescriptionEN'
                ].dropna().astype(str).tolist() \
                    if 'DescriptionEN' in sandp_df_for_save.columns \
                    else []

            label_dict = {
                'Index1_Code': [item[0] for item in index1_label_data],
                'Index1_Label': [item[1] for item in index1_label_data],
                'Filter_Code': [item[0] for item in filter_label_data],
                'Filter_Label': [item[1] for item in filter_label_data],
                'SandP_Label': sandp_label_data
            }
            labels_df = pd.DataFrame({k: pd.Series(v) for k, v in label_dict.items()})

            # --- Write to Excel ---
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                settings_df.to_excel(writer, sheet_name='Settings', index=False)
                if not labels_df.empty or not all(labels_df[col].isnull().all() for col in labels_df.columns):
                    labels_df.to_excel(writer, sheet_name='Label', index=False)

            self.update_status(f"บันทึกการตั้งค่าสำเร็จที่: {filepath}", "success")
            self._msg_success("บันทึกการตั้งค่าเรียบร้อยแล้ว", filepath)
        except Exception as e:
            self.update_status("บันทึกการตั้งค่าผิดพลาด", "danger")
            self._msg_error("บันทึกไฟล์การตั้งค่าไม่สำเร็จ", str(e))

    def _build_sandp_df(self):
        """สร้างข้อมูลชีท SandP จากตัวแปรที่เลือก S/P"""
        spe_pat = re.compile(r".*?#(\d+)\$(\d+)$")

        def _group_to_label(var_list, prefix):
            group_src = {}
            for var in var_list:
                m = spe_pat.match(str(var))
                if not m:
                    continue
                grp = int(m.group(1))
                idx = int(m.group(2))
                group_src.setdefault(grp, []).append((idx, str(var)))

            rows = []
            sorted_groups = sorted(group_src.keys())
            for seq, grp in enumerate(sorted_groups, start=1):
                src_var = sorted(group_src[grp], key=lambda x: x[0])[0][1]
                src_label = self.spss_variable_labels.get(src_var, src_var)
                rows.append({
                    'Variable': f'{prefix.upper()}_{seq}',
                    'DescriptionTH': '',
                    'DescriptionEN': str(src_label) if src_label else '',
                    'Rank_list': '',
                    'Spcode': '',
                    'Important': ''
                })
            return rows

        rows = []
        rows.extend(_group_to_label(
            self.vars_to_transform.get('S', []), 'S'))
        rows.extend(_group_to_label(
            self.vars_to_transform.get('P', []), 'P'))

        if self.sandp_label_overrides:
            for idx, lbl in enumerate(self.sandp_label_overrides):
                if idx >= len(rows):
                    break
                if str(lbl).strip():
                    rows[idx]['DescriptionEN'] = str(lbl).strip()

        cols = [
            'Variable', 'DescriptionTH', 'DescriptionEN',
            'Rank_list', 'Spcode', 'Important'
        ]
        if not rows:
            return pd.DataFrame(columns=cols)
        return pd.DataFrame(rows, columns=cols)

    def _format_summary_sheet(self, worksheet):
        """หัวตารางชีท Summary: ตัวหนา + สีพื้น + เส้นตาราง + ตรึงที่ B2

        ใช้โทนเดียวกับหัวตารางชีท SandP เพื่อให้ทั้งไฟล์เป็นชุดเดียวกัน
        """
        header_fill = PatternFill(
            start_color='4F81BD', end_color='4F81BD',
            fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        thin = Border(
            left=Side(style='thin', color='7F7F7F'),
            right=Side(style='thin', color='7F7F7F'),
            top=Side(style='thin', color='7F7F7F'),
            bottom=Side(style='thin', color='7F7F7F'))
        center = Alignment(
            horizontal='center', vertical='center', wrap_text=True)

        for cell in worksheet[1]:
            if cell.value is None:
                continue
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin
            cell.alignment = center

        worksheet.row_dimensions[1].height = 30

        # คอลัมน์ข้อความด้านซ้ายกว้างพอให้อ่านหัวตารางออก
        for col_letter, width in (
                ('A', 12), ('B', 30), ('C', 12), ('D', 22)):
            worksheet.column_dimensions[col_letter].width = width

        worksheet.freeze_panes = 'B2'

    def _format_sandp_sheet(self, workbook):
        """จัดรูปแบบชีท SandP ให้เหมือนเทมเพลต"""
        if 'SandP' not in workbook.sheetnames:
            return

        ws = workbook['SandP']
        header_fill = PatternFill(
            start_color='4F81BD',
            end_color='4F81BD',
            fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin = Border(
            left=Side(style='thin', color='7F7F7F'),
            right=Side(style='thin', color='7F7F7F'),
            top=Side(style='thin', color='7F7F7F'),
            bottom=Side(style='thin', color='7F7F7F'))

        # Body colors by column (A..F)
        body_fills = {
            1: PatternFill(start_color='DCE6F1',
                           end_color='DCE6F1',
                           fill_type='solid'),
            2: PatternFill(start_color='F2E6D9',
                           end_color='F2E6D9',
                           fill_type='solid'),
            3: PatternFill(start_color='F2E6D9',
                           end_color='F2E6D9',
                           fill_type='solid'),
            4: PatternFill(start_color='E4DFEC',
                           end_color='E4DFEC',
                           fill_type='solid'),
            5: PatternFill(start_color='EBDDE2',
                           end_color='EBDDE2',
                           fill_type='solid'),
            6: PatternFill(start_color='F2E6D9',
                           end_color='F2E6D9',
                           fill_type='solid')
        }

        max_row = max(ws.max_row, 2)
        max_col = 6
        for c in range(1, max_col + 1):
            hcell = ws.cell(row=1, column=c)
            hcell.fill = header_fill
            hcell.font = header_font
            hcell.border = thin
            hcell.alignment = Alignment(
                horizontal='left',
                vertical='center')

            fill = body_fills[c]
            for r in range(2, max_row + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill
                cell.border = thin
                if c == 1:
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='center')
                elif c in (4, 5, 6):
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center')
                else:
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='center')

        widths = {
            1: 14,  # Variable
            2: 16,  # DescriptionTH
            3: 30,  # DescriptionEN
            4: 16,  # Rank_list
            5: 12,  # Spcode
            6: 12   # Important
        }
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w

        ws.auto_filter.ref = f"A1:F{max_row}"
        ws.freeze_panes = 'A2'

    def _apply_sheet_tab_colors(self, workbook):
        """ตั้งสีแท็บชีทให้ตรงโทนเทมเพลต"""
        tab_map = {
            'Summary': '00B0F0',
            'SandP': '00B0F0',
            'Correspondence(S)': 'FF0000',
            'Correspondence(P)': 'FF0000',
            'Rawdata': '7030A0'
        }
        for sname, color in tab_map.items():
            if sname in workbook.sheetnames:
                workbook[sname].sheet_properties.tabColor = color

    def save_all_results_to_excel(self, summary_df, results_dict, full_output_text):
        """บันทึกข้อมูลสรุปและผลวิเคราะห์ลง Excel — คืนข้อความสรุปผล

        รันบน worker thread จึงห้ามเปิด dialog จากที่นี่
        """
        if not self.original_filepath:
            raise RuntimeError(
                "ไม่สามารถบันทึกผลลัพธ์ได้ "
                "เนื่องจากไม่พบ Path ของไฟล์ต้นฉบับ")

        summary_only = self._ui.get('summary_only', True)

        try:
            base, _ = os.path.splitext(self.original_filepath)
            filepath = f"{base} BS Output.xlsx"
            self.last_excel_filepath = filepath
            rawdata_df = None
            if self.df is not None:
                rawdata_df = self.df
            elif self.transformed_df is not None:
                rawdata_df = self.transformed_df
            elif self.original_filepath.lower().endswith(".sav"):
                try:
                    rawdata_df, _ = pyreadstat.read_sav(self.original_filepath)
                    if self.spss_original_order:
                        rawdata_df = rawdata_df[self.spss_original_order]
                except Exception as e:
                    self.log_message(
                        f"   ⚠ ไม่สามารถโหลด Rawdata จากไฟล์ SPSS ได้: {e}")

            expected_factors = ['N_S', 'N_P', 'N_C', 'N_E']
            template_rows = []

            diag_by_filter = {}
            for filter_name in summary_df['Filter']:
                row_data = {'Filter': filter_name}
                analysis_result = results_dict.get(filter_name)

                if analysis_result and analysis_result.get('beta_sorted') is not None:
                    betas = analysis_result['beta_sorted']['Beta'].to_dict()
                    for factor in expected_factors:
                        row_data[factor] = betas.get(factor, 0)
                    if analysis_result.get('diagnostics'):
                        diag_by_filter[filter_name] = \
                            analysis_result['diagnostics']
                else:
                    for factor in expected_factors:
                        row_data[factor] = 0

                template_rows.append(row_data)

            template_df = pd.DataFrame(template_rows)

            if not template_df.empty:
                for factor in expected_factors:
                    if factor not in template_df.columns:
                        template_df[factor] = 0

                # โหมด ABS: ใช้ |beta| เป็นฐานคิดสัดส่วน ทำให้ B.S–B.E
                # ไม่มีค่าติดลบ และผลรวมยังเป็น 100 เหมือนเดิม
                use_abs = self._use_abs_beta()
                raw_betas = template_df[expected_factors].copy()
                basis = raw_betas.abs() if use_abs else raw_betas
                template_df[expected_factors] = basis

                template_df['Total'] = basis.sum(axis=1)

                beta_ratio_cols_names = {'N_S': 'B.S', 'N_P': 'B.P', 'N_C': 'B.C', 'N_E': 'B.E'}
                for factor, ratio_name in beta_ratio_cols_names.items():
                    template_df[ratio_name] = np.where(
                        template_df['Total'] != 0,
                        (template_df[factor] / template_df['Total']) * 100,
                        0
                    )

                self._collect_beta_warnings(
                    template_df, expected_factors,
                    raw_betas=raw_betas, use_abs=use_abs)
                self._collect_reliability_warnings(
                    template_df, expected_factors, diag_by_filter)

            beta_cols_to_add = ['B.S', 'B.P', 'B.C', 'B.E']
            if 'Filter' in template_df.columns:
                cols_to_drop = [col for col in beta_cols_to_add if col in summary_df.columns]
                if cols_to_drop:
                    summary_df = summary_df.drop(columns=cols_to_drop)

                summary_df = pd.merge(summary_df, template_df[['Filter'] + beta_cols_to_add], on='Filter', how='left')

            # เก็บค่า agree_* จาก Summary ลง JSON เฉพาะรอบปกติ
            # (โหมด Re-analyze ห้ามเขียนทับ JSON เดิม)
            saved_agree_json = False
            if not self.is_reanalyze_mode:
                saved_agree_json = self._save_agree_summary_to_json(
                    summary_df)

            excel_df = self._prepare_final_excel_df(summary_df)
            sandp_df = self._build_sandp_df()

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                excel_df.to_excel(writer, sheet_name='Summary', index=False)
                sandp_df.to_excel(writer, sheet_name='SandP', index=False)

                workbook = writer.book
                worksheet = writer.sheets['Summary']

                headers = [cell.value for cell in worksheet[1]]
                for col_idx, header in enumerate(headers, 1):
                    if header is None: continue

                    format_str = None
                    if header == 'SampleSize':
                        format_str = '#,##0'
                    elif header in ['S', 'P', 'A level', 'A score', 'Index', 'C', 'E', 'B.S', 'B.P', 'B.C', 'B.E'] or \
                        (header.startswith(('S_', 'P_', 'E_')) and 'cor' not in header):
                        format_str = '0.00'
                    elif header.startswith('C_'):
                        format_str = '0'
                    elif header.startswith(('cor_', 'CorE_')):
                        format_str = '0.000'
                    elif header.startswith('agree_'):
                        format_str = '0.0'

                    if format_str:
                        for row in range(2, worksheet.max_row + 1):
                            worksheet.cell(row=row, column=col_idx).number_format = format_str

                color_scale_cols = ['Index', 'B.S', 'B.P', 'B.C', 'B.E']
                max_row = worksheet.max_row
                for col_idx, header in enumerate(headers, 1):
                    if header in color_scale_cols and max_row > 1:
                        col_letter = get_column_letter(col_idx)
                        cell_range = f"{col_letter}2:{col_letter}{max_row}"
                        rule = ColorScaleRule(
                            start_type='min', start_color='F8696B',
                            mid_type='percentile', mid_value=50,
                            mid_color='FFEB84',
                            end_type='max', end_color='63BE7B'
                        )
                        worksheet.conditional_formatting.add(
                            cell_range, rule)

                # ชีท 3 ตัวนี้จะถูกลบทิ้งอยู่ดีในโหมด "บันทึกเฉพาะ Summary"
                # จึงข้ามไปเลย ประหยัดทั้งเวลาเขียนและรอบ load/save ซ้ำ
                if not summary_only:
                    if self.transformed_df is not None:
                        self.transformed_df.to_excel(writer, sheet_name='Sheet Dummy', index=False)

                if rawdata_df is not None:
                    rawdata_df.to_excel(writer, sheet_name='Rawdata', index=False)

                if not summary_only:
                    if not template_df.empty:
                        final_template_cols = ['Filter', 'N_S', 'N_P', 'N_C', 'N_E', 'Total', 'B.S', 'B.P', 'B.C', 'B.E']
                        template_df = template_df.reindex(columns=[col for col in final_template_cols if col in template_df.columns])
                        template_df.to_excel(writer, sheet_name='Factor_Template', index=False)

                    output_lines = full_output_text.splitlines()
                    safe_lines = ["'" + line if line.strip().startswith(('=', '-', '+', '@')) else line for line in output_lines]
                    output_df = pd.DataFrame(safe_lines, columns=["Analysis Log"])
                    output_df.to_excel(writer, sheet_name="Factor_Output", index=False)

                for ca_prefix, ca_sheet in [
                        ('S', 'Correspondence(S)'),
                        ('P', 'Correspondence(P)')]:
                    self._write_ca_sheet(
                        workbook, ca_sheet, ca_prefix)

                self._format_summary_sheet(worksheet)
                self._format_sandp_sheet(workbook)
                self._apply_sheet_tab_colors(workbook)

                desired = [
                    'Summary',
                    'SandP',
                    'Correspondence(S)',
                    'Correspondence(P)',
                    'Rawdata']
                for idx, name in enumerate(desired):
                    if name in workbook.sheetnames:
                        workbook.move_sheet(
                            name, offset=idx
                            - workbook.sheetnames.index(name))

            if summary_only:
                final_message = f"บันทึก Excel (Summary + SandP + Rawdata) เรียบร้อยแล้วที่:\n{filepath}"
            else:
                final_message = f"บันทึก Excel (Full Report) เรียบร้อยแล้วที่:\n{filepath}"

            if saved_agree_json:
                final_message += "\n\nบันทึก Agree Original JSON เรียบร้อยแล้ว"

            self.update_status("บันทึก Excel สำเร็จ", "success")
            return final_message

        except Exception as e:
            self.update_status("บันทึก Excel ผิดพลาด", "danger")
            raise RuntimeError(
                f"ไม่สามารถบันทึกไฟล์ Excel ได้: {e}") from e

    def _use_abs_beta(self):
        """ใช้ |beta| เป็นฐานคิดสัดส่วน B.S–B.E เสมอ

        สาเหตุที่ B.S–B.E ติดลบ/เกิน 100 คือ N_S/N_P/N_C/N_E สัมพันธ์กัน
        เองสูงมาก (multicollinearity) ทำให้ regression beta พลิกเป็น
        ลบได้ในบางกลุ่ม — เป็นปัญหาเชิงโครงสร้างที่เกิดได้ทุก workflow
        ไม่ใช่เฉพาะตอนตัดข้อมูลด้วย Good จึงเปิด ABS ไว้เสมอ

        B.x = |beta_x| / Σ|beta| × 100 การันตีทางคณิตศาสตร์ว่าอยู่ใน
        ช่วง 0–100 และรวมกันได้ 100 เสมอ ไม่กระทบ N_S/N_P/N_C/N_E,
        Factor Analysis, ค่า beta ดิบ, Correlation, T2B หรือ SampleSize
        เลย — กลุ่มที่ beta บวกอยู่แล้ว (ส่วนใหญ่) ค่า B.x จะเท่าเดิม
        ทุกประการ เปลี่ยนเฉพาะกลุ่มที่ beta เคยติดลบเท่านั้น
        """
        return True

    # เกณฑ์บอกว่า "เลขนี้เชื่อไม่ได้" -- ไม่ได้ใช้ตัดหรือแก้ค่าใดๆ
    # ใช้แจ้งเตือนเท่านั้น เพื่อไม่ให้เอาเลขที่เป็น noise ไปตีความ
    _RELIABILITY_P = 0.05     # beta ที่ p เกินนี้ = แยกจากศูนย์ไม่ได้
    _RELIABILITY_R2 = 0.30    # R2 ต่ำกว่านี้ = โมเดลอธิบายกลุ่มนี้ไม่ได้

    def _collect_reliability_warnings(self, template_df, factors,
                                      diag_by_filter):
        """หาแถว/เซลล์ที่ตัวเลขคำนวณได้ แต่ทางสถิติเชื่อถือไม่ได้

        ต่างจาก _collect_beta_warnings ที่ดูเรื่อง beta ติดลบ อันนี้ดูว่า
        beta "แยกจากศูนย์ได้จริงไหม" (p-value) และโมเดลอธิบายกลุ่มนั้น
        ได้แค่ไหน (R2) -- เพราะเซลล์ที่ n ผ่านเกณฑ์แล้วก็ยังอาจเป็น noise
        ได้ ถ้าตัวชี้วัดนั้นไม่มีความหลากหลายพอในกลุ่มย่อยนั้น

        ไม่แก้ไขตัวเลขใดๆ เก็บไว้รายงานเท่านั้น
        """
        self._weak_beta_cells = []
        self._weak_model_groups = []
        if template_df is None or template_df.empty or not diag_by_filter:
            return

        ratio_of = {'N_S': 'B.S', 'N_P': 'B.P',
                    'N_C': 'B.C', 'N_E': 'B.E'}
        for _, row in template_df.iterrows():
            name = row.get('Filter', '')
            diag = diag_by_filter.get(name)
            if not diag:
                continue

            r2 = diag.get('r_squared')
            n_rows = diag.get('n_rows')
            if r2 is not None and r2 < self._RELIABILITY_R2:
                self._weak_model_groups.append({
                    'filter': name, 'r2': float(r2),
                    'n': n_rows,
                })

            for f in factors:
                p = diag.get('p_values', {}).get(f)
                if p is None or (isinstance(p, float) and np.isnan(p)):
                    continue
                if p > self._RELIABILITY_P:
                    self._weak_beta_cells.append({
                        'filter': name, 'metric': f,
                        'ratio_col': ratio_of.get(f, f),
                        'value': float(row.get(ratio_of.get(f, f), 0) or 0),
                        'p': float(p), 'n': n_rows,
                    })

    def _collect_beta_warnings(self, template_df, factors,
                               raw_betas=None, use_abs=False):
        """หากลุ่มที่สัดส่วน B.S–B.E ต้องระวัง

        B.x = beta_x / Σbeta × 100 จึงรวมกันได้ 100 เสมอโดยนิยาม
        แต่ถ้ามี beta ตัวใดติดลบ ตัวหารจะเล็กลงจนสัดส่วนพองเกิน 100
        หรือติดลบ — เป็นสัญญาณว่ากลุ่มนั้น n น้อย/โมเดลไม่นิ่ง

        เมื่อเปิดโหมด ABS ค่าจะไม่หลุดช่วง 0–100 แล้ว แต่ยังรายงาน
        กลุ่มที่ beta เดิมติดลบไว้ให้ทราบ
        ฟังก์ชันนี้แค่ตรวจจับเพื่อรายงาน ไม่แก้ไขตัวเลขใดๆ
        """
        self._beta_warnings = []
        self._beta_zero_groups = []
        self._beta_abs_used = bool(use_abs)
        if template_df is None or template_df.empty:
            return

        ratio_cols = ['B.S', 'B.P', 'B.C', 'B.E']
        for pos, (_, row) in enumerate(template_df.iterrows()):
            name = row.get('Filter', '')
            total = row.get('Total', 0)

            if not total:
                self._beta_zero_groups.append(name)
                continue

            # นับ beta ติดลบจากค่าดิบก่อนทำ ABS เสมอ
            if raw_betas is not None and pos < len(raw_betas):
                raw = [raw_betas.iloc[pos].get(f, 0) for f in factors]
            else:
                raw = [row.get(f, 0) for f in factors]
            ratios = [row.get(c, 0) for c in ratio_cols
                      if c in template_df.columns]
            if not ratios:
                continue

            n_negative = sum(1 for v in raw if v < 0)
            out_of_range = any(v < 0 or v > 100 for v in ratios)
            if n_negative or out_of_range:
                self._beta_warnings.append({
                    'filter': name,
                    'total': float(total),
                    'n_negative': n_negative,
                    'span': (min(ratios), max(ratios)),
                })

    def _prepare_final_excel_df(self, final_summary_df):
        """จัดเรียงคอลัมน์และเตรียม DataFrame สำหรับเขียนลง Excel"""
        if 'A' in final_summary_df.columns: final_summary_df.rename(columns={'A': 'A level'}, inplace=True)
        if 'ZA' in final_summary_df.columns: final_summary_df['A score'] = final_summary_df['ZA'] * 100
        else: final_summary_df['A score'] = np.nan

        s_cols = sorted([c for c in final_summary_df.columns if c.startswith('S_') and 'cor' not in c and 'agree' not in c], key=lambda x: int(x.split('_')[1]))
        p_cols = sorted([c for c in final_summary_df.columns if c.startswith('P_') and 'cor' not in c and 'agree' not in c], key=lambda x: int(x.split('_')[1]))
        c_cols = sorted([c for c in final_summary_df.columns if c.startswith('C_') and 'cor' not in c], key=lambda x: int(x.split('_')[1]))
        e_cols = sorted([c for c in final_summary_df.columns if c.startswith('E_') and 'cor' not in c], key=lambda x: int(x.split('_')[1]))

        if s_cols: final_summary_df['S'] = final_summary_df[s_cols].mean(axis=1)
        if p_cols: final_summary_df['P'] = final_summary_df[p_cols].mean(axis=1)
        if c_cols: final_summary_df['C'] = final_summary_df[c_cols].mean(axis=1)
        if e_cols: final_summary_df['E'] = final_summary_df[e_cols].mean(axis=1)

        idx_val = 0
        for avg_col, beta_col in [('S', 'B.S'), ('P', 'B.P'), ('C', 'B.C'), ('E', 'B.E')]:
            if avg_col in final_summary_df.columns and beta_col in final_summary_df.columns:
                idx_val = idx_val + final_summary_df[avg_col] * final_summary_df[beta_col]
        final_summary_df['Index'] = idx_val / 100

        main_order = ['Code Index1', 'Labe Index1', 'SampleSize', 'Filter', 'S', 'P', 'A level', 'A score', 'Index', 'C', 'E', 'B.S', 'B.P', 'B.C', 'B.E']
        final_summary_df.rename(columns={'Index1':'Code Index1'}, inplace=True)

        core_cols = sorted([c for c in final_summary_df.columns if c.startswith('CorE_')], key=lambda x: int(x.split('_')[1]))
        cor_s_cols = sorted([c for c in final_summary_df.columns if c.startswith('cor_S_')], key=lambda x: int(x.split('_')[-1]))
        cor_p_cols = sorted([c for c in final_summary_df.columns if c.startswith('cor_P_')], key=lambda x: int(x.split('_')[-1]))
        agree_s_names = sorted([c for c in final_summary_df.columns if c.startswith('agree_S_')], key=lambda x: int(x.split('_')[-1]))
        agree_p_names = sorted([c for c in final_summary_df.columns if c.startswith('agree_P_')], key=lambda x: int(x.split('_')[-1]))

        final_column_order = (
            main_order +
            s_cols + p_cols + c_cols + e_cols + core_cols +
            cor_s_cols + cor_p_cols +
            agree_s_names + agree_p_names
        )

        final_column_order_existing = [col for col in final_column_order if col in final_summary_df.columns]

        excel_df = final_summary_df[final_column_order_existing]

        return excel_df

    # ===================================================================
    # CORE ANALYSIS LOGIC (UNCHANGED)
    # ===================================================================
    def perform_factor_analysis(self, target_df):
        print("ส่วนที่ 1: การวิเคราะห์องค์ประกอบ (Factor Analysis)\n" + "-"*50 + "\n")
        factor_vars = ['N_S', 'N_P', 'N_C', 'N_E']
        if not all(col in target_df.columns for col in factor_vars): raise KeyError(f"ไม่พบคอลัมน์สำหรับ Factor Analysis: {', '.join(factor_vars)}")
        df_factor = target_df[factor_vars].dropna().copy()
        if len(df_factor) < len(factor_vars): raise ValueError("ข้อมูลไม่เพียงพอสำหรับ Factor Analysis หลังจากการลบค่าว่าง")
        print(f"ข้อมูลที่ใช้ในการวิเคราะห์องค์ประกอบ: {len(df_factor)} แถว\n")
        fa_rotated = FactorAnalyzer(n_factors=4, rotation='equamax', method='principal', rotation_kwargs={'kappa': 0.5, 'max_iter': 250}); fa_rotated.fit(df_factor)
        original_loadings = fa_rotated.loadings_
        ss_loadings = np.sum(original_loadings**2, axis=0)
        spss_col_order = np.argsort(ss_loadings)[::-1]
        L = original_loadings[:, spss_col_order]
        print("Rotation: Rotated Component Matrix (Equamax - SPSS Compatible):")
        loadings_rotated_df = pd.DataFrame(L, index=df_factor.columns, columns=[f'Factor{i+1}' for i in range(4)])
        abs_loadings = loadings_rotated_df.abs(); primary_factor_map = abs_loadings.idxmax(axis=1)
        factor_to_variable_map = {v: k for k, v in primary_factor_map.items()}
        sort_list = sorted([(int(primary_factor_map[var].replace('Factor', '')), -abs_loadings.loc[var].max(), var) for var in abs_loadings.index])
        sorted_loadings_df = loadings_rotated_df.loc[[var for _, _, var in sort_list]]
        print(_df_map(sorted_loadings_df, lambda x: f"{x:.3f}" if abs(x) >= 0.4 else "")); print("\n" + "-"*50 + "\n")
        print("คำนวณ Factor Scores ด้วยวิธี Anderson-Rubin (PCA)...\n")
        Z = StandardScaler().fit_transform(df_factor); R = df_factor.corr().values; inv_R = inv(R)
        temp_matrix = L.T @ inv_R @ L; eigvals, eigvecs = eigh(temp_matrix)
        inv_sqrt_eigvals_arr = np.zeros_like(eigvals); positive_eigvals_mask = eigvals > 1e-12
        inv_sqrt_eigvals_arr[positive_eigvals_mask] = 1.0 / np.sqrt(eigvals[positive_eigvals_mask])
        inv_sqrt_temp = eigvecs @ np.diag(inv_sqrt_eigvals_arr) @ eigvecs.T
        C_AR = inv_R @ L @ inv_sqrt_temp; factor_scores = Z @ C_AR
        df_scores = pd.DataFrame(factor_scores, columns=[f'FAC{i+1}_1' for i in range(factor_scores.shape[1])], index=df_factor.index)
        return df_scores, sorted_loadings_df, factor_to_variable_map

    def perform_regression_analysis(self, target_df, factor_to_variable_map):
        print("\nส่วนที่ 2: การวิเคราะห์การถดถอย (Regression Analysis)\n" + "-"*50 + "\n")
        dependent_var = 'ZA'; independent_vars = ['FAC1_1', 'FAC2_1', 'FAC3_1', 'FAC4_1']
        required_cols = [dependent_var] + independent_vars
        if not all(col in target_df.columns for col in required_cols): raise KeyError(f"ไม่พบคอลัมน์สำหรับ Regression: {', '.join(required_cols)}")
        df_regression = target_df[required_cols].dropna().copy()
        if len(df_regression) < len(independent_vars) + 2: raise ValueError("ข้อมูลไม่เพียงพอสำหรับ Regression Analysis")
        print(f"ข้อมูลที่ใช้ในการวิเคราะห์ Regression: {len(df_regression)} แถว\n")
        Y = df_regression[dependent_var]; X_original = df_regression[independent_vars]; X = sm.add_constant(X_original)
        model = sm.OLS(Y, X).fit()
        print("Regression Model Summary:"); print(model.summary()); print("\n" + "-"*50 + "\n")
        print("Standardized Coefficients (Beta):")
        unstandardized_coeffs = model.params.drop('const')
        betas = unstandardized_coeffs * (X_original.std() / Y.std())
        beta_df = pd.DataFrame({'Beta': betas}); print(beta_df); print("\n" + "-"*50 + "\n")
        print("Standardized Coefficients (Beta) - Sort:")
        score_to_factor_map = {f'FAC{i+1}_1': f'Factor{i+1}' for i in range(4)}
        renamed_betas = betas.rename(index=lambda score_name: factor_to_variable_map.get(score_to_factor_map.get(score_name)))
        valid_order = [v for v in ['N_S', 'N_P', 'N_C', 'N_E'] if v in renamed_betas.index]
        beta_sorted_df = pd.DataFrame({'Beta': renamed_betas}).loc[valid_order]

        # p-value ของแต่ละ beta และ R2 -- statsmodels คำนวณไว้แล้วตอน fit
        # ดึงมาเก็บเพื่อบอกว่าเลขไหน "แยกจากศูนย์ไม่ได้" (ไม่มีนัยสำคัญ)
        # ไม่กระทบการคำนวณ Beta/B.x/Index ใดๆ เป็นข้อมูลประกอบเท่านั้น
        renamed_p = model.pvalues.drop('const').rename(
            index=lambda s: factor_to_variable_map.get(
                score_to_factor_map.get(s)))
        beta_sorted_df['p_value'] = [
            renamed_p.get(v, np.nan) for v in valid_order]
        print(beta_sorted_df); print("\n" + "-"*50 + "\n")

        diagnostics = {
            'r_squared': float(model.rsquared),
            'n_rows': int(len(df_regression)),
            'p_values': {v: float(renamed_p.get(v, np.nan))
                         for v in valid_order},
        }
        zpred = model.predict(X)
        return beta_df, beta_sorted_df, zpred, diagnostics




# <<< START OF CHANGES >>>
# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
def run_this_app(working_dir=None):
    """Entry point for launcher."""
    qt_app = QApplication.instance()
    if qt_app is None:
        qt_app = QApplication(sys.argv)
    try:
        win = SpssProcessorApp()
        win.show()
        qt_app.exec()
    except Exception as e:
        print(f"ERROR: {e}")
        QMessageBox.critical(
            None, "Application Error",
            f"An unexpected error occurred:\n{e}")
        sys.exit(1)


if __name__ == "__main__":
    run_this_app()
