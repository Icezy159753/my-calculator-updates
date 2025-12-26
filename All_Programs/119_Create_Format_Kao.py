import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side
import os
import datetime
# ด้านบนไฟล์ (รวมกับ import เดิม)
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
import re, unicodedata

# ==== APP ICON (วาง Base64 PNG ที่คุณให้มา) ====
APP_ICON_B64 = """iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAYAAAD0eNT6AABdeElEQVR42uzdXYjm4x/H8d/YRX/0zwEhRA48JMIKSSSbUkaJKVLjyBYOFGoRcTgHyHJgN5JdjigOhrCDUh7aSNmVbY2HXavZmd/3c93DtJunva6PmrmLg/XU7uxc932/X/U+maOZ++D7nX797utqAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPwV28sifILksyK8QvLKCF8X4ZsivCrC96Tkh6Q8llJem1Je92+S8tNSHovIj0heLfmOCN+WkkcifJ3klZIvnp31KVNTPqIBAAD7z/ZhET49JV8teTTC90bkx6T8glQ2SmWzVKalkqXiCtojlR1S2ZRSeU3Kz6eUH134vX1zSr5kZsbHNQAADLqU/P+29fkp+UbJq1PK66TytlS+lcpeqbgP2yOVz6UyLuUnU/Ldkq9vW583OenDGwAA+oXt5TMzPi3CwwuP1Mt4SuXrShf0Ura3+7mMS3lM8mhKPtv2IQ0AADVrWx8V4StT8t1Sfl4qn0rll0oXbq+0WyqbIvIzku+M8AW2lzcAACyVCJ8Q4WEpj0WU91n2B63dC593XpOSR3bt8rENAACLwfZQp+NzJN8p5Relsr3S5TiobZPy+givmpnxaQ0AAPv5SH+4+4Led5UuPtpHKZWvF77W6JHZWR/dAADwd1Ly2ZJXS2VCKr/WuuDoP7VXKp90Xy5cyTsEAIDud+59rZTXS6WtdIHRASyihJSf63R8je3DGgDAYLC9rNPxZVJew9If+GalvCHCw5xFAAB9yPYhf1r605UuI1r6g4rGJY9OT/vIBgDQuySfFZEfl8pMpUuH6uzHiPxMSr6kAQD0hslJH56SR7ov8pVKFwz1Tlslr+a8AQColOQzu4fyRKWLhHq7n6XyUoSHbS9rAABLx/byCN8ilQ8qXRrUhy3cW+AH5uZ8TAMA6DpIj/klj0rly1qXBA1EP0t5g+QzGwDA4l6jm5LvkspUpQuBBrMslfGUfGkDADhw2tbHd6/T/aHSBUA0X0R5PyWP8J4AAOznjXsp5bXctEc92BeSb+UfAQD4jxfxdM/kn6t0uBP927Z2nwgMNQCAfbN9aIRXcVof9WGbJF/VAAD+YHuoe3jPZKXDm+hANRHhCxsAGHQRvlAqH1c6rIkWoyKVlyN8RgMAg2ZqykdIeUwqeysd0kSL3a9SXtO2PqoBgEEQ4Wulsr3SoUx0sPte8mgDAP2qbX28lDdUOoSJlrrx2Vmf2gBAv7A9JPl2DvIh+sf2SL7P9qENAPQyySdKZWOlw5ao1r5oW1/eAEAvSsk3SEWVDlii2isp5XXT0z6yAYBesHOn/xeRn610qBL1WlslX9QAQM0ifHpE+azSQUrUq/2Wkh/mbgEAVUrJI5zfT7R4pVTeS8knNQBQA9tD3ct7Sq2Dk6iP+iEljzQAsIS6N/eVVyodlET9WpHyGr4uCGBJpOSTpbKl0gFJNAhNzM35mAYADpZOx+dIZWelQ3G+HTt+8rZtc978WccffbjL777znTe++a0n3to+39sTO+Z/3s2fb5md76uvdlf7NxHto+0RXtEAwGKL8JVLdapf22Zv2dzxG69/4xfWb/YTj33gB+/f6Ntve9Uj12/wVVes9WUXP+UV5/7O3pnFRHWFcfzGh6aPvjRpmjRp0qRJ7YPYgq1arUtdiMVUoiTFqq0tSrTFpSY0WtdGQ6p1AK0UbBDZBGFAYEhZZFeUAgVlGcouMAvcc6C1SSUs99/ekzQxJhKEO3Bm+H7J74XwMpnk/uc75/u+e2FaLl4UjrUro7F1cwKCgzIQejgPYWfKER1VjfQ0K6ru9cPhGJU1EMi55z+cY4tCEAThwvDfyJj2eCYeao2NfyIjvQWm85U4FGIRYfzeOxEioGXQx8uEjzbEYu/umwg7W46khAbcrrCjp2dY1pAgPVtNVcdPKgRBEEbDOQIY00Zc8fDq7n6MX3M7YPqxUlTzq5dH6SHrlvp4mcRpxMljRbiR0qz/kJE1MEiPdDwWwAsKQRCEETCG7Ua+u99hH0FhQTfOhd1GYECyHprSBroRrl8dI04xrkTX4reqfkmDg/QgC/QJHYUgCGKa4e/PmDY63YdSi/URYn+pQ9AuM5Z4R0ob1jOh79orOHW8CPl5XdRLQLrKaocDLykEQRBTgXNsYEwbnvI9fsOQqHp3BKbAe6GcYTzbLvWJFH0E8XH30d5GEwikoVo5x6sKQRDEc1b+a6YQ/iLEYqJq9HtwaUNXVvWrkD1fmMWkgcM+ImuokO7lw/5+vK4QBEFMsvJ/izFtaNJ3+o5RvWNfr2T1ETppA9adXLnssrgmqLzjkDVYSPfRwRjeVAiCICZCVfEKY1rPZB4sra1/67P4bt217w5u9ovD5YtVtKSInI4OzrFAIQiCeNZu/8m8zrekuBcHQyyyVftiT8DalVHY8vFV7NqRjIMhGfjmQCaOH7UIz5zOww9hhcKz3+eLvx39Nkf834Gvzfjy8+vYuS0Jm3yvYNli0ago3ec7EpqP32uZrCFDyq1NVfGGQhAE8fRb/VR1PHWiLXyZ5j8Q4J84iw1zEQjwj8P+fekIO5OPuNg7yM9rQH1dJ3p7neCcG6rN1o8H97tQVtqMTHMtYn4ux7EjFnz2aRJWLf9pVnsF9gVnoay0T9agIeW1d2gIrykEQRD/wxiOPCv4U1Oa4O93bUZDbs2Ky9i9K0VU6+a0GhHyqspEMMtiZ6cNJcXNiI+7i1PHc8WPkxneaSAmLCzZbfr3JGvgkPLZNjCAlxWCIAjOsY4xbeyp4Bfrbf18Y2ckyDaui0Ho4SwkJdwTVTfn3C3t6+tH0a0mRIaXIDgoFR8svTRjfQLmNCtUVdrQISWSc61+aAjzFYIg5i56JcCY5nzi4SAqSldX/O+/GymO8pMTq9Bi7ZE20KerqjLcrWxFhKkYgQHxLj8h+GRrEopuPZQ2eEh5VFWtoqsLLyoEQcw9AMzjTCv4T+hWlNmwc1uqCzfgRetH5eLe3uEYkDa0XWlHhw3Xk6pE4+GKJa47HQgOykRN9YC04UNK4w0A8xSCIOYWgwyhnGmwNv2F/V/luGJbn2iWO/FdrmikY0zOUJ4tnc4B5GTVI2RvOpZ4R7iiWVBMDbRYH8kaPqQEcj5+XiEIYu7AORY4nWPD+nz5ssUXDR9XOxSSAUtOPZzOuVnpP6/dXXZcu1qJ7YGJ8F5o/Lrh8AuV9N4BcgKxTyEIwvMBMM+S3d5o9D3/hg+jcSmiBB3tNmmD1h1seNCFc2GFho8a+m+6hvIym6QBRM6yY4OD8FUIgvBcvLxM87cFJNcYWGWKBTrZWXVQVSZtqLqjdvuAGDH094s19Frg9Ili9PYMyxpE5OzJ6b0BBOGhvL0wfL3PIpPdsGU0e9L0DndpA9STLCttFo2DRk0RrFsVg5uZrbIGEfkve2ceHVWV5/FL27bO9Izd08t0z3T30Z5hph0hr0i9ghDZooCCiKiIqN2goAgKDQqCCwgooiCpWyEbQkLCLhAia4CwGXbCEkKAAIGQQCBbVRLFHU1+U/WOnLZtkPtqy6+qvp9zPsfjH+o5kt+Sd+/9/ZrPE/X19DMBAAgPYmISb9E1me63cbTj1tCRwtB9qx/KFhSUGo2Av77gjB6VQ6Wln3EtRrB5zCaiFgIAENq0jZJddE2e88+N8jV0ojh83+yHkvvzT9OIYVl+aQLu7TrH2O9Q52r6R3kWKBhg6+roRQEACFWoha7JUW4v++GM31Nw2BbDSHbP7lPGUYwfGjxjs6OzttEo+j6KZiL0vVxXR+0FACC0iGmd8BurJjf7WhT+3H+BZ9Y92+IH/6bnz6nfQ5k+NwJPP5lFp0s+8RRr1jItmuHm+UuX6FcCABAa6Bb7XXqUrPRxEY9nRC8G94SYtbVOmpu2y+cJg107v0e5G8rYFn80E0F1De4DABACfPvJ/2sfPgMbU/vKyyrZFjl4XY05DK+MXe3LRUHjn42fvstzJMC2cKOZCI5OJw0VAACetGyZeJMtSmb48ltf/77z8aQvzNy5/YSxotiXn4tRw9dQ5cXLbAtuuOri5WdOJ/1JAAB4oevxv7JqMs/bBN8uOsGzcx8je8NUz7FAalIexVgTvG4CHn1oIZ06cYltsYRBaSYOEtGNAgDAA2vrBE3XZIX3o2EzaN9e/NYfCXpeC/Tple7DJsc0OpBfy7bIwSBYRxMFAECB4Lzvb/D2rH/625uoOkJX8kaqnhXMUyZv8PpuQMf2ybQhp5RncYLB8Kv6emotAADNh02z99U1+YWXN/w9+/jZFikYeLdsPkb3dn3P64uimWkFXAsUDLz5RHSDAAAEH5vFMdjbm/6DBiymU6cq2BYmGDzPnr1oDHjy9khgptzDtUDBQIopgQA0DzbNPt7b39rip2/2XAhjW5Bg81wQnDY119sjAc/kQJ4FCgbazxoa6FYBAAgOepT9Za/ObWMSae3qQrZFCDa/OWsLvR4eNP3tHVyLFAysHwgAQDCKv3zDm+R8X/c5dGD/GbaFB/LxcMFZevD+uV41AdPe2s61SMEAWl9L9wkAAL/i/5fHFtKZ0xfYFhzIzwsV1fTs4KVeNQGTxm8hl5NnoYKBsd7VdLqsjG4WAAD/o1vkRG+S8djRq6gKT/ygl08FR4/8wKsm4I1J29gWKxgYXS56VQAA/Iuu2Z/zKglPzCGn08W2wED+ulx1numQXjUBSY69bIsVDIif1NbSbwUAwD/YNMfjuiYbTSZfT9JmW1Rg6Jk+Z5fxgsTsEqElC49yLVYwALpcjSkCAOA7Vk320DX5ldlnfvMz97AtJDB0XbH8ILWLNrVHwNg7sDHnLNuCBf3uN3V1dIcAAHhPdGtp0zX5qdnin7XsANsCAkPfD7IPmW4COsYk0d7dVVwLFvT/hcBVAgDgHdHR0//TWOxjsvgve38/28IBw8eVXjQBd3eaRUWFdWyLFvT7hcB2AgBgDk2b8VNdkwVmi/+SRflsCwYMP5cvPWD6TkDvnhl0tvQztkUL+tG6phwBADDD5B/pFvtKsxet5mXsZlsoYPiaveKg6SbgmadWkLO2kWfRgn61vp46CgCAGrrFMc3sbf+093ayLRAw/F28cJ/p/QGOeCwPihA3CQCA0o3/3romm8zNXt/EtjDAyDHBvtX0QqoNOaVcixb0p3XUXgAAro21lWypa/IjM0n0pRdXYsgPZDMs6NVxa0w1AXEdUunY0QaeRQv60xUCAHB14m6bfLOuyUNmkufTTy7BeF/IypoaJw0xuTvg0YcWUlXlZa6FC/rHxtpa+h8BALjqmN80M0nz4d4ZVFFRzbYQwMj1/PkqeviBDFNNwOuvbuZauKC/rGucLQAAf4/VYu9v7rNpChUVlbMtABAWF5+j7nGzTDUBK5af4Fm4oL/84pMq+rUAAAgDi0X+TtdknbmLU0fYJn4Ir7hzx0mKsaoPCrqrUyqdKfmEa/GC2BQIgD+hFlZNrjfzW9LsWTvYJnwIv29qUp6prwAjhq5iW7ygXzxHRDcIACIdXbOPNJMcx4xayTbRQ3itlwEjhmWZagKWLT3OtXhB/3wFeFAAEMnoUfG365r8XH18arrnchXbRA/htSwvr6Re98xRv+PSMZVO4yggnMVgIBDJUAtdk9vMrFLdtfMk2wQP4fXcu6eEYm0zlZuA4c/iKCCMbfqohv5LABCJ6JrjGZz7w0hzbtouc0cBS45xLWDQV+toogAg0tC0Gf9u5tb/sCHLyOXimdAhNHsfYNCAxcoNQJc7U+jwIRfPAgZ909l0lohaCAAiCd1iX6SeAJOppKSCbUKH0KzHjpZTh3aJyk3Afd3TafuHFT9YTM6Vf0EFB520eWOZ8dVgVsp+mvHOTnrrjQ9p4mub6cWR6wzHvrje8/eGb0/ZTkmOvZSZfpiylhUb/42TxR9jQ2EQbXBSFwFApBCt2bua+ASK3f4wLM1M300mV10b64OTZ+6lBfOO0Ny0QzT59a00eGAWdYubrfjvUb9vc/+9c407CDOm7aLsrBNUcMiJxiAQOhvTBQCRQL9+y2/QNXlE+dP/M8vYJnAIfdBYXjVo4BLSNRkyepYWPT9kJaUk7qMdeRfQEPhHFxHdKAAId6xRjmdVk02n9kl0ovg82wQOoa8eLSrzHAWwLfgqDcHoketoyaKjVFr6GdXVNbllW2j5Wkf3CgDCmZiYxFt0TVarJpfMuXvYJm4I/WXm3N1sC7yhifHcQwZn0+KFRVRe/rnRCHhp5DUSOAYA4Y4eJd9RTSb9+87Hfn8YMY58fgXbwu6N7fWZ9MJf19Lm3DJyOht9bQQioZlwEdGPBQDhSNvW9j/omvxC9bLTju0n2CZrCP3tuXNVnqaXbUH3xft7zKXUpPwrXwVCX7wGAMAcuiZnqSaMl19azTZRQxgoy8sq6bFHwrMJ8NixfTJNnZJHJ098zLOwN3czUdf4rgAg3IiOjr9V1+RXKkmic2wynTl9gW2ShjCQVlXV0qQJOWyLuL+eGL72ci4VF3/Esyg3n8cFAOGGzSJnqyaH5MQ8tskZwmC5ZfMxerxf+H4NuHJP4M1J29wN/ydcC3LQbWig2wQA4UKbNo7bdE1eVkkI3eJm0cWLNWyTMoRBHhlsNAKvjVtDcR1Trhs/PbvPpmcHL6UJr64lR/xWmpexm1avOkxbNh2jndtP0L69Jd/VWKy1IecILV96wBhKFD99M40dvYoGPrHIiMVgNQId2iUZw42qqy6zLczBsr6ehgkAwgVdk3Pw7A9C/4wPXremkBbO30tps3fSgnl7KWdtoVHML1zwf+NcdraStm45biwvGjdmFfXoNjvglwXXrTnNtjgHx8ZlAoBw4NuFP1+oBX8aVVfXsk2+EMI6YzDXsvf305hRKymuQ0pAGoHhQ1fRqVOXmBbogFuD5UAgLLBp9smKQe/5DMk26UEIrz7K+MNtxfTGpPXUtUuqX5uAzrEptGhBEblcbAt1AKU7BAChTNxtk2/WNVmj+Ns/hv5AGMJ64ndT7lEa88JK43KfvxqBZwdn0+kIuyToctFwAUAoY42yD1EMcs9lJbaJDUJoztLSi5SalEf3dZ/jlyaga+f3aENOKduC7f8GoHGxACB0oRZ6lDyuFNxdUqm6Cmf/EIabNTVO42ivb58Mn5sAm0UaQ4Rqar5hW7j99xKg6bQAIFSJjkrorBrYKUl49w9hOOt5zrh2dSE9/IDvjcCgAcuotPRTtsXbTzZdukS/FACEIrrmmK+47tczA51t4oIQ+vWegPGCoNc9vh0N9OiWRvv2VnEt3n6SeggAQg1dn/YzXZOfqQTyGxPXs01WEMLAjTtOkNso1ub9ZcFYWyKtXlXCtHj75SLgBAFAqKFb5DDVID544AzbJAUhDKzHj5XT8KFZXjcBbds4KDO9gG0R983GpQKAUMOmyQMKwesZN8o2MUEIg+eypfuvjDr2ynen7QzHeQFHBQChRHR0/B2KQesJerYJCUIYVI0NoL58DXhj4tZwawIuE9FPBAChgurkv7s74ekfhPAfXgsYOw68HST0+mubyels5FrQMREQhDe6JovVuvUctkkIQtisepYbGdNBvWkCXhuXGzZNgMtFDwsAQgFr6wRNMUg9q0nZJh8IYfNbXlZJzwx639vjALZF3Zw0RgAQCugWOUUhOI03wC4Xz6QDIeRjba2T3py03qsmYKZjD9Oirm59fWOSACAU0DV5QiUw46dvZptwIITcNO4FGE/+zDYB8zIOsy3uiq4TAHCnnTbjj6pBeegg3v5DCM25emUBtbfNND0nYH1oLxE6LgDgjq7J51UCsk+vdLYJBkLI29yNR01PD+wcm0KHC1xcC/z1bBAAcMemyTUqwfjOW7lskwuEkL+bc803Ab17ZlB52edci/wP6uYmAQBXWrZMvEnX5Ccqgbht63G2iQVCGBpu2XTM9HHA8KGrQnRQEP1eAMCVtlGObmqf4pI9+8HZJhUIYei4bm0htYtOMNUEzErez7TIX1unk6wCAK7YNDlVJfhGj/yAbTKBEIaeixfuI5tFuQEwGobteRVsi/3VpXsEAFyxajJPJfiWLMpnm0gghKHpTMc2U18Bet2TThcqvmRa7K8m9RUAcETXZ9+ouvv/2NFytkkEQhiaeoaKvfTiSrM7A5gW+6uOAx4oAOCIVbO3Vwm4e7u+xzaBQAhD26qqWnri0QWmmoANITIfoL6enhcAcES32EerBNvY0avYJg8IYeh7ovg8xXVMUW4A7r17DlWEwFFAfT2NFQBwxKrJFSrBtmDeHraJA0IYHuasLTR1KXDqm3lsC//fpNcFABzRNXlWJdAKCkrZJg0IYfg4dcpGU68C9ufXMC38aAAAY2JiEm/RNdl0vSDrGJNITqeLbcKAEIaPnvsADz+QodwEDHhiKfMBQWgAAENsbRwdVAJs4BOL2CYLCGH4mb+vhGKs6kOCPsg+ybT4owEATNEtcphKcE1+fT3bRAEhDE/j392i3AA80DODamq+4foMcIIAgBu65khVCa6F8/eyTRIQwvDUcxTwwH3pyk1A5tzDXF8BjBMAcMOm2berBNae3afYJgkIYfi6IeeIcgPQLW42VVVdxhwAABRfAFy4TlAZT3IuXqxhmyAghOHtiOey1L8CpBdwPAJ4UgDAibjbJt+sa7LxegHVPW4W28QAIQx/i4rKlS8E9uiWRtXVX3O7BPiIAIATelT87SoB9eSfF7NNDBDCyHDi+BzlrwBLlxzj1gD0EABwwqrJHirB9MrY1WyTAoQwMiwpqaA7285UagAeeXABqwbA6SSrAIATuiafVwmmxIQP2SYFCGHkaGZC4Pa8Ck53AH4nAOCEbnFMUwmk5UsPsE0IEMLI8dTJCmqvzyS15WXruTQATUT0EwEAJ6xR9rkqgbRt63G2CQFCGFmOf2WtUgMQa5tJpaWfcmgA6gQA3LBqjtUqgXTo4Bm2yQBCGFkeLjirvC0wbfZBDg3AUQEAN6wWuVsliEpLL7JNBhDCyHPQgMVKDcDj/RYzGALUtEoAwA1dkyXXC6C2bRzYAgghZGX2ioPKlwGLjtQ1cxPQKAUA3NA1WXe94Lm7UyrbJAAhjEyrq2upa5dUpQYgQe5p7hcAIwQAvKAWKlMA+/RKZ5sEIISR65TJG5QagEcfXtjcewDuEwBwomXLxJtUgqd/3/lsEwCEMHLdtfOk8jHAieKPm60BaGigWwUAnIiJSbxFJXAGPL6QbQKAEEa2vXumKTUAC+Ydaa4GoIGIWggAOBEdnfhrlcAZPHAJ2+CHEEa206bmKjUAI55b3Uzn/015AgButG1t/4NK4Dz37HK2wQ8hjGw9Q8pU8lin9snkdDY2xwuABAEAN6ytZEuVwBk1fAXb4IcQRrY1NU7qcmeyUhOwP7+mObYADhIA8EJ9FfCLIz9gG/wQQjhqRLZSA5CRXtAcWwD/VwDAC/UvAH99Pott4EMI4dy0XUoNwOiR64LdANTgAiAwDe4AQAihmvn7SpQagJ7d04LdAGQLADjSttW7v1UJmiGDl7INfAgh9Iwq7xyrdg+gvPzzYJ7/jxYAcCS2lfyFSsAMwjNACCFzn3nqfaUG4MNtFcEcABQtAOBIq1Yp/6I0COiJRWyDHkIIPU6dslGpAcgM3kXAKpz/A7Zo2oyfqgTMn/svYBv0EELoccmifKUGYNKELcF6/58hAOBKdHT8rSoBMxBfACCEzN23V+0i4NCns4N1/t9PAMAVa5R9CAYBQQjDwYqKaqUGoE+vzGA0AJcbGujnAgCO6Pq0n+maLFcJmPh3t7ANegghvGJcx5Tr5rNYWyK5XAFvANYLABhyZQnQLrek4oacI2wDHkIIr/jYI/OVclpJyaUALwCiJwUAXIiLm/xja2vH/9k0+3hdky7V4t8xJpGqq2rZBjyEEJodCbx65alANgBfffQR/ZsAwN/o+uwb22kz/mjVHHfbLI7BusXxuq7JBFuUXGjV5Hpdk/lui92WGhrF3lGva7LRLZl10oQctsEOIYTf1RG/VSmvDR8WuNXA9fVNqwUAvhZ6mzajtU1zPK5HyXd0Tea4LXP7tVsKhu31mXS0qIxtsEMI4XddmX1IOb9lLSsO1O3//gIAM7S7PfmXVk32/rbY73D7uVtqTqdNzWUb6BBC+H1LSy9Su+gEpfzWto2DUpPyqarqsj8bABcR3SQAuO5v+G0ccbrFMU3X5BG3TW6Ji337ZODsH0IYcg59eqmpXHd3p1nGcKCctafp4sWvfB3+YxcAXGv0rueTvk2T2XqU/FjXJHG0a+dUOna0nG2AQwjhtcxZW+h17ou1zaQhg7Mpc+5hKj3zqTef/+8QAFwhNlb+k9UiH7VqcoXnsz7Xon/F7nGzqOBQKdvghhDCH9LlqqMnHl3gay40jggGD8yi+ZmFqhsEdwkAPES3sltsFnuyrskGrsX++/7lsYV08sR5toENIYQq7t1TQjHWBP9diLYl0pgXcmhTbhk5nY3Xevv/lACRS8uWiTcZT/M0mc+1yF/rrf+slO1UU+NkG9AQQmhCI6cFIl/2uied3ks98P2vAg2VlfTPAkQeuh7/Kz1KTtA1Wc21yF/NHt1mk5yxhc6evcg2iCGE0NujgDcm5gQsf3Zol0QTx2+hI4We/15jggCRRdvW9j/omkxx+xm34t7eNpPiOqQY9uw+m3r3TDPW+o4ZtZKSE/No186T5HS62AYvhBD6owmY+ubGQOZa467AXR1TN7j/2kaA8EfXHf9htcgkXZNfBrmwe5ZdGGf1Y15YSTOmbaYF8/ZQzrpC2rnjJBUdKaPyskq2wQghhM3hB9mHqHNscqDzc5PbtZ77XwKEH7Gt5C90TcYH6Ta/5zd3o9CnJuXRxvVFdOpUBdsAgxBCznqGBE2ZvMGYcBrg3N3odrHe+t3/FiD08SzSsUXJEbom69xSoHzw/rnGD6i7W0WxhxBC/2vcecqcu5sGDVisMDXQJy/bNIejTRvHzwUITXSLvEePkscDdU4/YlgWzcvYTcePYfAOhBAG0/LySuMXrlfHraG7O6UGqhGosWnyaSEm/0iA0CCmdcJvbJpjqb9/GGK/LfrLlu6niopqtoEBIYSRZG2tkzblHqUJr66luzqlBKIRyLe2TtAE4Ay10DX5pL8/9/d7KJMy03fTuXNVbAMAQghhHVVX19LqlQX0/LPLjVv+/jwW0C1yintmDBYGcSNGS/y9HiVz/fmJf/wra2nP7lNsf9AhhBBe25KSCnLEb6WuXfx6RFAc3VraBOCBTXM8ovBbv+oyHeMHprQUA3YghDAcrKqqpWXv7ze+5vrvkqB9fL9+y28QoHno8Kfp/2qLkhn++AO9t+t7xu3SKqzRhRDCsNQzXGjd2kJjsJqfGoFd0dHxtwoQXNpaZCtdk6fcki92i5tF6XN2YX8+hBBGkBs3FFH/vvP90QS4bFH2+wQIDlaLvb+uyU98XaKTPPNDqqxE4YcQwkjU80UgO+sg9e6Z7vMkQavmeAtHAgEe6qNHOey+zn5+ZexqKinBsB4IIYR1ni/Axt2vWJtvkwatmlwfE5N4iwABOO/X5AZf/nAefXgebvVDCCG8msZOlmeeet+nJsAzfA6jhP39xM8iC30Z4JOSlEe1tdiZDyGE8IfNWnbAuB/m072ANo4OAviGZ02jrskLbskbn/rLYioqwqheCCGEynoGvxmTBX1oAj63afYHBPAO3RLfUY+SH3vxP95YFJFg34q9+RBCCL12zerDFNfB6/HC39gs9qECmKNtlKObrslP3ZJZ+/RKp717Stj+QEEIIQwd3QvffJkd0KRb7KMFUENv4+ila/ILt2TWEc9l0fnzmNkPIYTQrwuHjK/KXu8XsMiJAvwwtijZx1i64MUn/9mzdpDLxfOHB0IIYeibu/Go9xsHLY7XBfiB/f2a/NItmdD4w/hwWzHbHxgIIYTho+e5YO+eaV4eCTheEuDvadvafqc3Z/7390ijwsNn2f6gQAghDD/LyyqNV2be3AnAxcDv4FmtqGvyklsy44AnFtHZs9jaByGEsHkmCL7w12xvmoBGz0h7Eel4NinpmqxyS2YcPiwLC3wghBA2p8ZT8+lvb/KmCfjSqslOIlJp08bxc12Tx9ySGceMWkk1NZjqByGEkIXGJXQvmoA6m5bwJxFp6PrsG22a3GL2f9hr49ZguA+EEEJ2pibleXMp8Ey725N/KSIJm8WebLr4v7wGz/z+n737AK7iytIA3DOutXdra7w7U7s1u7NhXFue2ZrFykIBJCRAJIPIkhBBBJNzzjkL6T0FQAiQAJGjTQ4SOYsgsEQSIBBRIDAG4zBgi7t9tcssZoV0rl6/906/93fVV65y2erue1s6p7tvnwMAAGzNTT5QnS6COW7TStjXI6mD4gDJhRa48wcAAPZka2EZtxQlaq6+6Xf+PrJJgk5QDey7EZ38AADANBLjc5STAH9PazvNVTcPj1m/1k/ypk5QdY1bLUqw2h8AAExm0vgdSgmAbH7nW8P6seaKm5+nZbXKYLRslimLLbCdXAAAgMr6B3TvskZ1UeCZjz9O/UBzpU0/sc46QVU/LE1cvIA+/gAAYF7yJra5YtlgXy+rVXOVTW/t+x8qlf6C/FLEkcNX2E4oAAAA1Zfnb4rw2koNhMr8vZPCNfNvk3+pn8xRnaBatuQY24kEAABQlbOnQHatVUkCbtT+z/hfaWbe/D2s/VWC/5iRW9lOIAAAQHVlLDqq9lWAl3WhZtbNr0bKv6s8+o9uvQz1/QEAwCXJQnZ9eq5X6hxo2n4Bvp5JW8jv/f1TxJnT19lOHAAAgK2Kiu7JRe4qScCFGjUmv6+ZafPzsETJg6damon3/gAA4Pp2bDuv+GmgZYxmli042Po3KgV/+vXagBr/AADgNiaOUyoS9F2gZ+q/ambY/LySJlBPrE7wPFFYeIftJAEAABjt/v1S0aJpBn1BoId1hcZ98/Ky/ot+sN9ST2pF1gm2EwQAAGAvhw9dFv5e9AWB+n8boHHefD0smdTg3y1uNR79AwCA2xo5bLPKq4DDGtfN2zvlD/oB/kit9pefj1K/AADgvq5duytCg+bSywR7WhtrHDf94FZRTyJ+VjbbCQEAAHCUhQsOKzUL0jTxC43TVtPLWkPWL6Y2+rlz5wHbyQAAAHCUhw8fiVaRmSoLAltonDZ/z6S11INfufwk24kAAABwtL3ZF1TWAuRqXDZv76SP5Lt/ao//R48es50EAAAAZ+jdYx09CfCy1NU4bPrBpFAPetPGM2wHHwAAwFmOHrlCfw3gad2lOXmTd/9/T234E9VqKe7+AQAA3qF7lzX0ugCeCZ9oztx8Pa0jqBnL1i3n2A46AACAsx3Yd1Hli4A0zXmb+IV+EFcpB9r80wzc/QMAAFSha6dV1CTgeWBg6oeaMzYfT0t9aqayfNlxtoMNAADARc6eAnphII+kfhphc9qnf/XrpImSklK2gw0AAMBJTJtl1CQgT3P05uOT+o/6jl9QDtCasJftIAMAAHCTtfQ4/SnAJ8memiM3fw9rf8qB1fROEpcu3WI7yAAAANzIarm1A1KJSYAlQXPkJrsSEQ5MFjZgO8AAAABcjRm5lfoUoCQqav17miM2H5/431Hr/m/bcp7t4AIAAHB16OAl8muAmh7WMM0Rm6+nZTB18V9p6SO2gwsAAMCZQpOgFAc9/k86QjmgSeN3sB1UAAAA7halk1sF37Z7m2APj1m/1nf0E+WA9u+7yHZQAQAAuLt86bZCTYBEP/s+/veyxFAOJCJ8ASr/AQAA2Ci69TJigyDLOPs+/vdKWorH/8DF48dfiRPHrwpLwl4xYuhmWUKzvOV0s8aLnSaySYboFLtCDOy7UUyfskvs2H5ePEAhLIeSNx9yAVWyZZ8Y1G+jiI3Kkn9E5T/L5yUpcZ84eOASblLAFBLn7KU+BThs79r/JZQD2b4Nq//BfmRAXTD/kGhUP11eb+yFBKaKCWO3i8LCO2zH1BWU/O910aTBQtK8yOtn/tyD4v59JGjA17GjhdS/NS/t1hvA95OkP1EOItg/Rdy795DtYIK5bd96Xv6BZxvsK1OrZoq8K8Wdpx1k7ykQjSOqd100qpcudu/MZ3tu4N7k34v6YWm0dQCe1kh7rf7vTjmAHt3Wsh1IMC/5uD81+YDw9+IZ3FXIAlm3b5ewHWuzmZdyQFYdtWVO5HUlkzO25wjubSy1KJBX0mxnvv+Xj+DYDiKY15SJO9gG9OqIarVU3L3zgO14m0VifI6h8zJr+h625wrua8sXedRr+Ki9yv8WUg7gyOErbAcRzGlp5jG2gdwW/XtvwOsAG6xemWuPeZHXG9tzBvd09eod6vX75/CPJv+1scHfL/Ef9B/8qqqdB/mnoPUvGOpU7jUR4JPMNohXBk/M7Ofkiavy741d5iTIL0UcPYIbGeCFvsYlOVAzcqvpkRRB2LH8DIvt4IE59ey2lm3wNkJYrXni5o37bMefo+Li+/KTS7vOi/xCoKjoHtsxAPczqP8m4joAa2+j2/8Ooex4zuwctoMH5nNg30W2gdtI8bOy2c4Bw8Wg8tWJI+ZFJp94RQNsLJh3iFgRMCndKQsAN64/w3bwwHz0b+fZBm2j7zYfP+Y5B8zI7/YdODf4MgD4OLCf3B0w1+gFgGcpO847W8R28MBc5J2XLClN+oTLxyKaTB8vYo8PEnHX+jldx/wBonnmKBEUPoccaPSKhmznggnZX0Tpc7+anhZhDR8l7rTuIl61ay/u6v9MrjtC/nvyz5D727O7gO2YgPuQXw0Rr/9vaI2B6BUAvycUAEL7XzBMXl4R+Y90yxUjZeBlRyYkAXUSSOewcMFhtnPBZBX062IoZJsaDRSiXez/s6XxAKWfUy80TTZlYTs24D5aRy6hXbd+Sf+sGbHJH0TYoay1zXbQwHz2Zl8gXeh1Os2QwZatpiljSecxY+putnPhbPLGokvHVUpBe3rYGBns3yk+fJTSz+sQs1w8eIAvnMCpZC8L2pMrD2uYZsTm45EUTNnh4AGb2A4amM+6NadIF/qniePYBn8p5sBg0nkMHfg527lwttkz9qgF64Bp4kV0+0oTgBcxHURc4DSlnztl0k62YwTuYc5sauGrpO7GPAHwsLYnrWSeiZXMYGTxn+OkCz1y/hi2wV+KzR1ELQ/Mdi6cace280rln8N9EuS7fhnkq/SgTWdR34e+TkNav/Y027EC17dqxUnatephmaoZsckew6ieBUgAkAA4WkF+sagTPI9ez9/TIvY37SuDO1luZC+lRYG1A1Kx2BmcZt/ei8RPAS2ZRn0BMJ+yw107vmQ7aGA+SADcm2z5HNMmS+nufF69ETKoK0uvN1xpP80/zRB30McBnODL8zeJ16llt0FPAJLWEnYoS7ayHTQwHyQA7u119zOqnkGTxU/tyt/7KyuLiRV9a01W2t/wIV+wHTtwXXfvPqQmAAWaEZuvpzWHssMrl/GZDCABQAJg2HtOska+8eJx2zgZzKvtSVScaOI3W7Fp0HG2YwiuKzRoLmENgPW+QVUArecJvwxoAgRIAJAA2Oz0qesiWKHJT4CnRZxt3lMGcZsVtOghgrzo6wECfZPF4UOX2Y4luKamDRdRrs/vjVoDcLuqndWqmcJ2sMCckAC4n1u3SkRkkwylu/CsBkMIwZ1udcPBSvtv0mAhmjmBI8maO6Rr8+OPUz8wIgH4qspHcPXS2Q4WmBMSAPdS3uSnj1qTn2EhE8Qr8nt/utEh45WOo3f3dWgaBI4iu+7SnlB9kvxbIxKA51XsSLbmZDtYYE5IANzL605nVC39Z4nn0R0JAV3dd9EdRGv/mUrHk5K0n+3YgmuhdsP08Un8vREJwMuqdtQqMpPtYIE5IQFwH7Ltc4BPMjnYBntbxOWW3WWwtpvi1l1FqHeiUtOg3bvy2Y4xuI5B/WjlgH1rWD+2uREQ+gAAEgAkAPZSdP2eaFg3Xelue0vj8iY/dre7SX+l4woPmS8uXrzFdqzBNQwb/AX1CcB/abZschEBZUexUVlsBwvMCQmA65NNfrrGrVarx1+nvMmPw8wMG610fJ1iV4iHD9EVFexn1PAt1KdS3jZ2Alz4V5QdtWuLBACQACABUCL7hygF13YB08UPMYa/96+yaVCngKlKxzlt8i62Yw7mN2LoZmJp7IRPNFu2r78W4XoWUeWO2rRYwnawwJyQALi27dVo8nOH0OTHHkradFFuGrRu7Sm2Yw/mNqj/JlqNDC/LH7Xqbo8eiV99/fWrYlmUg1Abm+1ggTkhAXBdly7dku/LlZr87CM0+bGno017i5oKCUtIYKrIy0PTIDDpVwBPnpQt1BMAERpUZTcuWZmI7WCBOSEBcE3/0+RnmdondoQmP44wv94IpeNu0TRD3EXTIDBYz25raWsAasz5J60629Onor4e/F/pRP06C6rcUf06aWwHC8wJCYBrGjd6m1IQ7RE0hdDkxzHK9OPoE6zWNGjksM1s5wLMqXMHWiGg4BrW32iqmxDifT3wF+qE1OLTJZTVhqiEBUgAkABURr4XVwqeDXzjRenrJj9MfNW2k2jiq9Y0aNmSY2znBMynZbNMynVXFhW1/r3q3P2PkIH/tU6xa0gXeXEx6mEDEgAkABU7c/q67BlCL6zjaRG5kb1k0GUnv0V35aZBRw5fYTs3YC71QtMo191Xmur27bfit3rQf6oTr/XrRfvkID+/mO2AgfkgAXAd8j3467sWqqyIoTLYsrWqwRDlpkE3btxjO0dgDvJJO+HLPKmw2gv/3jR21B7SBX4EbTEBCQASgAqa/Azsu1EpWA6pPVE2+WEb/CV5fCNCJiidV5+e6/GqFGwhO09Sr7cTqt/8f6QH/Jc68ab4mbQmHZs2nGE7aGA+SABcQ3raYaUg2dx/lvgmqiPbwP+mb6M7ilaKTYPmJh9gO1fAX+7Jq8RPZ61bFROAssUy4L9taWYeaYfz5x5kO2hgPkgAzE++9w70VWvyc6nFZ2wDfkWut+omanslKDUNyt5TwHbOgLctm8/RGgF5Weeq3P3/vqK7f2nXzuukHU4av4PtoIH5IAEwN9nkp1E9tSY/mxsPYBvoK7NLsWmQ/Gy68ModtnMHfC1cQHyi5mUdpvLuP0EG+4qcP/eIskNZnIDtoIH5IAEwL/meu5tik5/JdcaxDfAUM8LGKJ1vXPuVaBoEyqZM3EGtntlGo2wPHoi/1QP9E52oSEnJC+oqV7aDBuaDBMC8EmbnKAXDmIAZ4ocYc7z3r6xpUMfA6UrnPWPabrZzCDx91pmWWPt7WXyo3/33koG+MhHhC1ELAJAAIAEwvMlPmHeCuNW6K9vAruJem86irg99PYAcp82f57GdS+Cnfh1SDYBXfn6z/46YALw6rROV6fkZrfvQgf2X2A4cmAsSAPORTX7qhio2+fm0L9uAXh1HmvVRbhp0Lu8G2zkFPq5du0u9roo1yvbsmfiDDPBVmT2D9ing0gyUvAQkAO6YAMgmP+3aZslzIUuuO4JtILfFPMWmQbKd+r17D9nOLbAgvx6hXlPbqJ/+TaMkABvWXSDteMzIrWwHD8wFCYC5TBy3XSnofRY0VfwYw7vYjy1Ng3oFTVEaj1HDt7CdW2BBfmpPvJ6SZlIf/1+nJABnTj+gFfFospjt4IG5IAEwjw3rTis2+ZkjHrbtzDaAO6tp0PJlJ9jOMThf357rqUWA2mlVbc+fiz/J4E7x6NGPItg/lbSopagI9a4BCYC7JADVafJzimmTH6PlNe8pAjzpTYOC/FPEyRNX2c41OLekdngIbX1NgJflj5TV/8NlcKfq0nEdaedbt5xjO4hgHkgA+KtOk58lzJv8GG15g6HKTYNu3sDXVPBzZ89cp15DD6mP//frBFVi/BHSAUyfsovtIIJ5IAHgTd6RDOq/SSm4Da41iX2TH3s0DRpeW61pUL9eG/Tx5Tnv4Ky/h8do14+H9XOtqk0I8YEe1P+sEzT0ksAtmmawHUQwDyQAvC1KV2vy08xvtnjathPbQG1P30V3EG1q/qVpEHqrgLL+vTcQSwBbhmpVbd98I2rLoK7izu3vqX2IxYWCYrYDCeaABICvw4cuu3yTH6MVtvxM1PJKVGoatH/fRbbXADhOSUmprBdBvHaSAynNf4bJoK6qffQqWj2ATNQDACQArpgAFBWpN/nZ1Ggg28DsSDur0zSoEE2D3N3ebNpn+OG100RBwVf/Rnn/v0EnVCVbjrnUHzPgCwkAzyY/3busUQpiE0LHsg3IzjC1zlil8evcYRWaBrm5GVN3k66VYYO2iydPRB9KAnBZJ1QdO0IqRSgfD2IlKyABcLEEIDFerclPdM3ppm/yY7QX0e2VmwbNmr6H7TUBdk+6yU/c1qwqkL17sqtaAPi+Hsx/1AlVettgUTc0jXQwq1acZDuowB8SAF527viSvAZICvVOEDdadWMbiJ3pfnnToDloGgRVOnjgEvkauX79G5kAvHz2TPymkt7/4hMZzKtr9PBdpAPq2W0t20EF/pAA8HHx4i0RHqLW5CfHxZr8GO1g0z5ynMhjWid4nsjPx+JqdzN10k7S9SHX573u3qvH+I6VLQBsqR741T8HDPBJFtev3WU7sMAbEgA+jyA7tV8pj5HMUncU28DLSYpi06DYqCxRWor1AO5CznVE+ALStZGedup1AqDH6bJVlXUA7GdLAlBa+qMIqzUf37ICEgA3SABUv/fv6sJNfoz2U7v2okfwFNQHgApt23Ke/Pj/ypWnMvi/9kgI8ct3LAAsmy4DuS3GjtpDK/7ReLG8g2A7wMAXEgDnky1q64WmocmPXZsGxYnGvvH0tRVBc8WtWyVsf2/BOH2IzX86xa6RQf8NMk4L33clABm2JgDZe26QL9i9ORfYDjDwhQTA+TIWHVVq8pPrJk1+jHY2Uq1pUHraYba/t2CMwit3yItuMxadeTsBkOsARr/rE8CNOmEL+TVA44jFlIOT9cLZDjLwhQTA6ZTe/WfUH8Y2wJpBVsRQlbUAbH9vwRjWhL20DpJ+KeLmze8qeALwaue7EoCdOmGrFOtxcknLgvybbAcaeEIC4FTyMTP5DmRg7UmiLIZnYDUL2TRoaO2J1He+qLPiwu7fLyW/ehs6aLsM+BV5KoR4r6IE4JBO2Opq4TPyH4gpk3ayHWzgCQmA0+v9k447xDtRPG4bxzawmsmTqDgR6k3rF3Bg/yW2v7vgmM5/Uk72zXclAPI1gGdFCcBJnTBCn56fkw6ydkCqKL6JjBWQAJglAdi4/gzpuGeGjWYbUM0oIXwUadzXrj7F9ncXbPvsNrIJ7fV600aZ+v9TVlkC0KeCOgCvjhmVAOzeVUTNVOQ7DbaDDkgAkAD83OqVuaTjXtlwMNtgakbrGtKul+XLjrP93QWbE2+Vb/8rUba4oicAB3XCCHInbZpnUT9fwXsrQAJgkgRgw7rTpONODEfRHyMl1x2BJwBuSt79t2iaQYyn88Tduz/IGFyZsxUlANk6YZQ1q/LJGcuc2TlsBx+QACAB+Av5jpn47X+8+DYaDX+M8H10B9HEdzbt8+psfF7tamRSR28QdUgG+Kq8kL1/3qoE+GqzkQlA6cOXIiI8nbwWoOj6PbYTAEgAkACofwUwoNYk8RxJgE2+04P/4FqTyF8BFBXh76grkS2fZeE8YrdduQhfBngC4f12IaDFtOBOt2B+LjlzGT9mG9tJACQASAD+T7u2WUpVAKeHjRFLIoaKZRFDlMj33icie4sXMR3YBujKvIxpX14EaVOjgcrWNBwsZujj1lChGmDblkvZ/t5C9SxOP0Ke/zEjdxODf/lCwNi3E4BZRicAJSUvRERYOrkuwKnca2wnApAAIAEg/FGygyZ+s8WxZr3ZBvqKyMDf1G+WreeOfgBuTD7NCas1j3z3f+niE4V1emWT3+4GOIwe3OkyFp2mNwyJW812MgAJABKAcuWf7oYEpjo0uAV6WWRpXLYB/03nW/QQQV4Wh45PsH8KXqO6mEnjd5Dnf8LYbMXYXLb2rScAIsoeCUBp6UvRsO4i8ols2niG7YQAEgAkAOXEvNSD8vgcKiZgBtug/6aOgdMdPjZJifvY/s6CutyTV+VTcXLZ38LCp6qx+dzbCYCv2g+gW77sHPlCrh+WhuJAgASAeQLwoKRUvnN2eKC71bor28Av3W/T2eFj0ioyU5aJZfs7C8r9/pXW2Uybsr86cflr7c1N3++HdkoA9JP6SbRpQT+hsSO3sp0cQAKABKCcyM8vJryjNNYp5p0Fv2zew6HjIeuonD93g+01AspkV0fy/IeHpInbt76vZlwWH75dC6BEJ+xh/95i8kn5e+F7VkACwD0BkE4cvyrqhs53WMA72Yx3ApDXvKfDxkImX0ePXGF7bYC6CwXFolbNFPI1sCTjrA035qLG2wnALp2wlwF9t5BPrHHEQlFcjFcBgASA61y8Ju9Ao1otRQLgwASgdeQSkXe2iO01AdWr+Ben0Gq7dfNl4vHjn6odj589E43f+hKgbJo9E4CLF56IYH/66uGhAz9nO1mABAAJwM/XBCRb9omQoLlIAOx4/rJoWmJ8Dt75u6CUpP1K18K+nJs2luwXcW9/CtiS/ANsLw6E2taABMBFEoDXbty4V14nIDYqS65iRgJgADmOMW2y5LthVPpzUfJVToBPMvmaGD18lwE9e8QQ7c3t0SPxO/UfpL4gMLr1CvKJhmCRCyABMKW7dx+K/2bvvMOkqtI0ftTRGWdnZ2dndvfZmZ3dnd3ZXXd2V1SyBJEMNtMoiAhN1AFFAcmioBIlNlGRpEhGclQRyTRBkoDk0JI70FTdc6vI9357T9mNhO6mTnV11Vf3vr/n+f0zzzzSfe6tet++4Xwb1h+kpUt208zpW2nap1sKtM1LczxVAFo3n1noesyYtpWWLN5N69cdoDOnM9geYxidfTWSak8K/025pz6ikyeDUchja0B+Q4EOFHcJ2Lkjk8qVDL/tJD8zmU6dOs/2AEIUABSAojlowEpPFYC+765geyxgbO/7t2s7V+uK0Py5B6KUxdZH4k58PuuD4i4AyuFD9LYTfaP9ArpwgedBhCgAKAAoACgAUFO1gZNWDrZ/dTFFb2y/NS2fAkDPxaIAZGdfpyaNZmKva4gCgAKAAgA954pl32o9J6Om66anB6KZw/OF4s4NgQzDvuJY7CVAvRVQqdxYrf0B5s7ZzvaAQhQAFAAUABQAeO+tfo+qeRpa2ffF58eincGfi/wwDHuZI0XLwsrEjGnfal0FKF96tHqwiO2BhSgAKAAoACgAsCAPHz5NtauP18q9gf3WqryMtutEfkhJrVRox8punVZoLUa1yuNoz7fpbA8wRAFAAUABQAGAd6o2t2tY/xOtvGv6wkzKyrwW7fBX2buhgCsA9GvDsK/FqgBkZl51BotM1VoU1aD2f/c92wMNUQBQAFAAUABgnmrzppYpes+9qb3+Dx30qcCOWQHIex1wYSyvAhzYf5GqVNDbU/zPdSfTkSOn2R5wiAKAAoACgAIAMzKy6JWX52hvAPXVyhMaoa7telEQhkFJKphjqHrIQf3S2qMwjx/D7lheEQXAnaIAQLd6/nyW9rv+ynEfbNUI8yg+A6AgogcMwz4V6xIwacJ27YVSGwUdOniK7QkAUQBQAFAAUAC8p7rs3zZ3l0sdu3dZQT6fZqDr+5UoDMOw+qpQjrX9+67RXrBnak6kfXvxTIDbRQFwpygA0G2qrbBfbjlLO8tat5hLWVnXizv8VdYuFoVhmvQPfr99KdYFwOez6PVXFmkvXK2q42nH9mNsTwiIAoACgAKAAuB+1dAmNRBLN8MaJk91isNlFdAx0Jop7oWU1sR4XAXIyLhKLZrO1l7AyuXH0lcr97E9MSAKAAoACgAKgHtVw+vq1ZmknV11akyiw4f9KpxjpDVR3AvTpD8Zhm3HowScO3dZbRcc0fjMTz9JY3uCQBQAFAAUAK7HAkbuxvUH1T412pmlJvzt+faCCuYYag0T4SClPV8FchxUYw/VZRG1SNoO7P8lZWVlsz1ZIAoACgAKAHSH06duoSdLj9bOKfWu/66dWSqUYyx1D7MA0COGYV+PVwk4flxS/aQpEZWA1i1m4TVBF4kC4E5RAGCimnE+i3r1XBZRPqm9b7ZuOacCOeb6fNRChIthWFNVGMfL9BMmNXpuWkSLXLvaeNqI+QGuEAXAnaIAwER0//6TeQ/7RXTZf8f2DBXGcdEwqI4IF7+f/s0w7KvxLAFnzlyi5k1mR7TYZZ8YRUMHr6LMTNwSSGRRANwpCgBMNBcv3EVPV/wwojyqWXUC7d4VuuwfNy9epBJCB7/fGqKCOJ6eP3+FXmrxmVrEiGzRdAZmCCSwKADuFAUAJoqnT2dQj66LI86gpNofq63vVQjHuwD8jdAhK4t+4ffbZ+JdArKzrlG3zrdMENR/VdAJkjTKzr7A9iSDKAAoACgAkJcrv9xHSbUmRpw9KY1nUfqJgArgeJsjIsEwKEWFcLxVv8SYkWlqUSO2edMZtGvXcbYnG0QBQAFAAYA8xvi+22s5lX4s8rzp+PoSyswd68vAnSISiOg+Ke01KoQ5OGvGHipXclTEB0W9tjFi2Ndqz2a2Jx9EAXC7KACQoeoqsZMx26j6U6F3+yP2/f7r6OJFSwUvF+eKSPH56A+GYUsuJSBt0xmqXe3OyzL62whP+3QLbgswFwXAnaIAQG5u2XyEmjeZXqRcebLMGJoza58KXFYaBr0rioJh0OsqfLl44oRJrZqFHg4skuqAr11zgO1J6XVRANwpCgDk4t6931PXTovyudyv/7Dfzh2ZKnAZSg3FXejfCljLqQTk5NygQQPWFfnAKf/SejZt2niI7UnqVVEA3CkKAOTwTv/bPZaqV8aLnB9qmN3pU5eYhr9Naot/UVScNftnw7BzHImTq746od6zVAeiyLZrOxdXBFAAUABQAFAAXOqeb9PpnbeXU/lSoW18i3zJf8rHu8jn4xn8uQaJ6CciGhgGJWkMC4rpDIEOry1RByUqNm44lT6b/Q1mC6AAoACgAKAAuMCtW46o9/nV8LhoZISaWcP4kv9NVT5uENHEMKyRKnS56feH3hJQ+y1HrQjUqTGBRqWupkMHT7E9sd0sCoA7RQGAsdrEZ+qUzfTi81OjlgnqlsGwIRspK/Ma29C/XWu4iCZE9JBh2FsciaNqmFCn9kvDOJh6I4fV7YFFC3bS+fN4hRAFAAUABQAFgOurfGtX7w/d369YdkxUc0CNq9+5I6Ow/OH4BsCLItoEg/RbKe3TXEuActnSI1Sneuh1wahaqdwY6tZ5ES1f+i1loAygAKAAoACgAMTVCxd+mM3fv88X6hXvqH/nVyw7liaO304XL96IUX5FdQvgfxHFgc9HTxiGHeRaAJRZWddCOwhWKKPRBDU3FlJvEIwbu4527zrB9gOSqKIAuFMUAFhUv08/p67Iqh371ATYYvl+V2+YdXljOR054mebcffwiChO/H5qrB4KZPrL3/TwYb86kOqgFqvJdSfRWz2W0szpW1UhcJopzw9PoogC4E5RAKCuR4+eCU3kG9D3CzWSV+Nhvsj38d+y+SzbTAtHKa3xorgxDGrPdQHudHPaWWrTer46wDFRjY98qcWs0KWpGdO2Ov/+YUo/cY7thwwFAAUABQAFIJ6qbdp3bD9GC+bvoGGDV9Hrr8yjujUnxOw7u0Hyp7R08SHy+3lmmI5+PzUSsUBK632ui5CfG9adisZOgkWaUNjouSnUvt0858vh89B8gskTN9HcOdvpyy/20tdffUfr1x1Qr6542qGDV3mqADgDq9gei2j6ZtclnioAnTsuZHssYmXapsPqey3kwgU7QxNax4xaq/6iD+3El9J4GlWvMi5u38nJz3xC8+ftJ5/PYptbml6Xkn4jYoVhWKOZLkSBrl3zPbVrszB3N0GYqLqlAEB3FgDI10bPTaP5c/erB/zY5lQkqp17RSwhovsNw5rDdUEKc//+izSw31q1sxPbExWiAHhNFABYXLZMmUNffH7MFZf681NK6iI0iOYeAV9yXZR7mX7CpI8+3Ep/rvsJ2xMXogB4RRQAGE2frjQu9Ifet7uz2GZQ9KT/FPGAiH5mGPZynosS/juY6vZAt84rcFUgAUQBcKcoADAam7i93HIeLZh3gLKzrrHNnChf/t8j4gkRPSSlvYDrAmnuJaAuFanXCNV7/2xPdC+LAuBOUQBgUfbqV3vAHD7kY5stxaWU1EPEGyJ6wDCs6VwXKRLPnr0cemCke5cVVLXyOLYnv9dEAXCnKABQZzLfq39ZQJ9M3pnIG/dEQ9vnoz8IBuSVgGlMF6pI+nyW2igi1DJfavEZVcCtAhQAFAAUABizS/svNJgeuqe/auUJyvLI5f0wbl+vE4zIezvgQ64LFi1zcm6ENhqaNGF7aBhRvTof4/VCFAAUABQAGAVrVBlPbV6aT6NS02jVVyfo/PkrbLMgvlILwRG/n95MhG2Do2nG+SuhUjBz+h56v/9a6tBucei900rlxrL9oCWiyRN4F4CmOzqyXTvOogB4y/KlRoc25Gn78gJ6r/eq0B9Ta75Op+/TA2y/45npO3eOfi64IiW1VDsUMV28mHryZJB278qi9etO0pLFh2jap7vpgzFbQrOmBw9crz4ASvW8gXoAMSFUVz7icdXj+cXd2IZ/yMPtqWz54TFfl+ZNZrM8T+onTfFUAXBCje1nNha+2+uH77IBfdeo7zf1V7y6V6923gv9Nf/NtvN07Jh00y58cdJKFdyRkpL9fvsSzwWERVB9qGP/V0ONoU7AMg3+W6zRrU+s10ZdMuV4nqiC66kC0L/ParafWegabSnpEZEIGAaVk9LOlpLtYkL91yVj/mZE6ZIjqOFC5n/955qyqyM9WXNIzEvAurUnUQBQAKD7XSISCdOkP5mmfVKVgGiIMhFfP5u9L6bBVjF5ED2/tCvbwM/Ppt90pKpt+lHpx2O3Tl3fWI4CgAIAXa6UVFEkGsEg/U7tWhStAoAyET9fe2VR+O/s1hpKdYf0oudmdaeGC7pp2cCx8YZObEM+HFO2d6SGS7rp/+7zulO9sW/TU80Ghr3W6vXU7OzrKAAoANC9bhSJimHQr6W0NzmSG2V6wkRVNUmrcvkPwtuPu1V/anagPdtwThSTRrwdbglQ469RAFAAoGulJJHIENHDUtpLuIY4ykShqu02w9u8o9xwStndgW2oJppPtxwQ1rpPn7obBQAFALrTzcINENEDUloTuYavmzSM6Lpp4+mwvgirvdqPbZgmoskTe4a17iOGp6EAoABAF+r3UzXhFojoPsOw+nANTphvmVCXmMP6IqzZvQ/bME1En5vTPax1Hz5kIwoACgB0mX6/vV64EcOgzlLaNtfQg7e7/ZvzYX0RVm70PtswTUTrDuodzrqrDaZQAFAAoLu0/H4qLdyKYVCKlPY1rqEHf/TsmUthP5DWYH53toGaSKbs6UBlKw4La80Xzj+AAoACAF2k329NFm5HPd0opX2Za/DBH61bc1JYX4blqg6lpjs7sg3WhPBIe6ratn/YpWvvngsoACgA0D0agQD9o/ACfj/VktK+xDX44A/2e2912IH0VMpAanEYrwLG4hXApFqTsREQCgB0kVJSN+ElTJOektKWXMMP2pS26bTWDnV1+vVmG7CcfX5ZVyr9RGrY6zxi2CYUABQA6BLVxnlE9KDwGlJSRSltg2sAQptaNfss7GBS2+E2+AzPA+iOFS7/9NDwd1wsM4bS081CjhkKAAoATCAtKelJ4VUMg8pLafu5BqDX3bL5LJV5fGTYAaUeYmuy+Q22gcttpHAVjS2AlWNHpcXo2KMAoADA4tcaIbyOakBS2gGuIeh1hw3eoBVSlRoMwtbAYVi737ta69q44XTKybnO8BxBAYAwAo9mZdEvBAiVgHpS2te5hqCXVXMBWqbM0Qqrmj3eYxu8HHzusx5aEwQrlRtL+/ZeYHuOoABAqOV1NUJfgLv2CbC4fsl52ePHJFV/6iOtEpA8qSfbAI6j6hZJ3vv+Ybt08SG25wYKAIS6Um8B8t8xkOuXnNddt/ak1vMAZcqkUuM1ndkGcRxUt0bULRKt8B80cB3bc8KrBWBA39UJMdQLsnQDET0gQP5IaQ3m+kXndceM2qwVXk/WGULN9mJaYJ41uvbRWr/mTWazve+PAsB/qBdkZ0YwSL8ToGCI6H4p7WVcv+y8rN9vUbu2C/O+CDExUMP6U3pqrVuNKuPVrRe25wIKAM9jgTLBVrXXf00B7o3PR7+S0j7C9cPlZdWcgKTaH2uFWb3Rb7MN5ljYeHVndUsk7PVSt1rWrE6P9bG1HS8ahn1KSvv4reb+bz4UgMQpACgT7Mb89hJAa7fA/5PSNrl+KLzsN9vO0ZOlR4cbaGqnO7XjHduALk6b7etAFZ4ZrFWYPhyzJdrH7KrjIcflpmmNMgxq75ikpo85/tHxb0WY5OTQLy9dot+bJv2P+otGSmrZrs2iNBQAiCJRkNZsIrpPAO0S8ALGCPN02pRdWqFWvpo3hwZVb99Xa51ee2WRutVS1ONz1DStGYZBHRzLFPdWo6UeHZmKAgBRJvJ1GxE9LEBkSGkN5XoSet2e3b/QCrcqzbw1NChp5Nta66MmMJ4+HYzkWFx2XCYl/aWgqWIoACgAMOZF4lQwSL8VIHKI6EHTtHdwPYG8bFbWVWqYPFUv5AZ4Y2hQo8+7UumS4a9LuZKj1NbLOut/RV1aNE1qkJFBfyUKAAUABQDGpUzkmCb9rwBFR91vVH/lcD3wXvbA/hyqXP4DraFBz81199CglN0dqHz1oVrFaPrU3eGu+XEp6U3TpL8XYYACgAIAY+4lNexOgOghJXVlerA97+JFB7XCrmyF4fTi5k5sA7woNj/8OlVpMUBrPbp3WRHOOq81DKpDRPcLDVAAUABgTL0uJSULUCz7A6xjetA9r/OliKFBjnX69dZah2frTaHMjCuFre1e06RGgjEoABCGvGEY1FSA4sHvp3/Hq4E8jWhoUE93DQ1quKAblXo8VWvIz9692QWt6UHTpIaJ8PoQCgCEofBvIkDxYprUlukJ4HmPHTWoWmW9oUH1J7tjaFCTLfpDfhbM35/fOl4zDKs/ET0kEgQUAOhxVfinCBAbpLRXMD0RPO/aNd/rDQ0qO5war03s5wGaH2pPlV54Xyv8B/Zbm9/67VOb84gEAwUAetirpknPCxA71EAFtW0p0xPC844ZqTs0aHBCDw2q2eM9rd+32Yuz7hzyc8MwrL6hDXsSEBQA6FEDfj/VEiD2qE1PmJ4Unjc0NKiN3tCgqq8l5tCgZ6e8qfV7Pl1pHB054r91vQy1La9IYFAAoAfNwat+cYSI7pPSXsP05PC8oaFBtSbrDQ0am1hDg15c35nKlBuuNeRn9dfpt73T74bNQlAAoMc8LiX9twDxRUp6RO2IxvQk8bzbtp6j8qX0hgY1SpChQT8M+RmiVXDGjt586/psMk36B+ECUACgh9zils+tK5DSGsj0RIGOn36yU3toUMou/kODqnXQG/LT9qX5tw75WZJIT/mjAKAAQKU1jYh+JgAfiOhhKe1jPE8YqOzZTW9o0FPMhwbVG12kIT+L3BT+KADQA143DOokAE8Mg2ozPXGgY2ZmBEODBr3Dc8jPF130h/yknclbi4WJ+qQ/CgAKgEfNNE2qKgBv1GQ0picQdNz/XQ5V0hwa1GBed3ZDfp6srnfff+qUm0N+Nrr18iEKAHSpa9Qr5wLwR81Bl9L2MT2RoOPihfpDg5psfoPNkJ+nW2oO+el8c8jPYSnpN8KloABAl3ndNKl3aPAWSBxMk15nekLBXPu9pzk0qOH71Pxg/AtA3QF6Q37qJ02hjB+G/GT7/fRH4WJQAKCLPCIlVRAg8cidGLiN6YkFc4cGtWg6WytMa70V36FBDTSH/FQsO5b27gkN+bHU8ynC5XitAPTv8zXbzxeMWFtKa0JWFv1CgMTFMKi8+uJlepLBCIcGJX/8ZtyG/JSrpDnkZ94PQ34Mw+ojPIDXCkDdmhNnS2lNVfvAc/2MQS2PBQJUXQB3EAhYk03Tpshke5K6ylUrj0cwNKhzzIf8VNYc8tO/75rc8LdXeeUeotcKQOnHUj8QDpcu0e+ltIZJaRtcP2ewUK+pfWSI6GEB3INp0t+bpp3jSNEQZaJ4HDUiTStcK9QbRM2+ax+7IT899Yb8pDSeRRcuhIb85Hjp6WGvFoA8Ll6kv5GSekhpn+X6WYN3uc4N23CDgkvAqyqA3SLTD1GRhwa98vICrZCt9lrfmIR//akRDPk57Mv965+aCg/h9QKQBxE9JCW1NE17B9fPHLTTMb7XA+Q+ELida6CjTPzgyZMBqlNjolbY/vnDt4p5yE8n7SE/X391Iu93mi88BgrA3UhJFXL3JrnGNAi9piElve3WvThAPhgGlTVN2+IaxImulNFx65azWkODypRKVTvyFUv4q1sMFZIG6w35GZWW97v4vDgoBAWgYNStIMOw+klpZzANRldrmnZQSmuIm/fhAIVgmtZErgEKf/STSTv0hgZVL56hQdU76g35adN63s0hP4ZBHYUHQQG4N0T0U8OgZnhNOXbBb5rWaLVBnADeRTU/07QvcA0++KNdOy3XCt8qrQdQ88NRHPIzVm/Ij7p1cerUzSE/e4joJ8KDoABoX5ksJ6U1RYUU1wBNYP3qyX4vXokDBWCa1JZr6MEfzcq6Sg2SP9UK4WcG9YrSkJ+u6taC1pCfzT8O+bFNk54SHgUFIDJycuiX6mFl9awS0zBNJE9ISd3UmgoA7nwgMBCwt3ENPvije/dkqZ309IYGze9e5CE/5WsM1Rzys+uWn9uaITwMCkBUZpk8bprWSCnt80wDlqO22m9DSqrvlT03QIQEg1TKNO0bXIMP5qk/NKhc5WHUZNsbEQ/5qdJKb8hPt07Lb/15pZfe+UcB0C8A+m8vUSW1La2UtmQavPH2nJTWYMOg/xAAaDwQ+BHX0IO32+edVVqhXPn5yIYG1R30jvaQn8zMKzd/zkCAugmPgwJQPGRk0F+ZJr1gGNYcKW2TaRjHSqmutBkGJRHRAwIAXQyDfm2adjbX0IM/mpNznVIaz9QbGtRLb2hQw0X6Q3727c2+9ec8TkQPCY+DAlD8ENHPpKRkKa1PpbQzmYZ0tPWrvRRMkxrg/X0QFaSkl7mGHrzdo0d89HTFceGXgMdS6dnpPcIJf3XLQN060CoYC+cfuO3nCwSouQAoADGGiO73+6m0aVIvKe2NjteZBngkHpLSSjVNqkZEDwoAov3hMU17A9fQg7f71cpjzhdu2CGtdvBTO/ndc8hPJc0hP33f/frOn20vHjzKKwAjeoezhl/Wbc82/JWrk14Ls2iO6C8Y4fPRr6Skeuq+uJT2JscrTMM9P49KaU1WeyRcukT/JAAobqSk/zJN+zLX0IO3O2LYJq2wrpishgZ1KLAA1Hpbd8jPTHVL4o5dEClZgBClS4x4OZx1fKtSb7bhr3yncq8wz4nUdoIxubcLKjt2yR1bvIfBlsS2Y7rjctOk9wyDnpGS/k4AEA+kpDe5Bh68XcOw6C+t5mmFdtVX+lGLfDYJSh7fU90q0Bryc/So/86fKU2Amzh/EdcK6+rMYyPo62deZxn+65LaUZlwrzQ9PjJJJBhE9JDfTyUNg1JMk97JfZZgo2HYp6S0L0cp5K/k/vc2SGlNMwyrr5TUSm3Jnp1Nfy0A4AIR/cQ07Z1cQw/e7qmTAapdbaLelYD6gyl5wlvUaFlXenZGD6r2aj8V/lpDflatPJ7Pz0NVBLhJqVITfu6sVzCsNS2RSoOqvBW6336mQau4uzu5DQ15uieVLRH2eXGl4iNDXBdmuSON/9vxyUCAaqiJeYZBzdUmandqGNREPZin/n/q/28Y9J8IeJBwBAJUwjTta1xDD96qxtCgKDlqRFp+A5A+F+AunGK1SK2Z6310xHIBAHAHpmkN4Bp48G4nT9weky/6ti/NJ8Ow7vz37WCQSgqQz3MAqclsQzuKOr9nQwEAcAdqSpdp2t9xDTx49wjiLm8sK9Yv+ZpVJ9DJ7838/v15AhRIqRIjN3IN7ii5TQi6TwAA3IO6j2WatsU19ODtZmRcpvrPfKLzxa015GfL5jP5/bs3TJP+V4BCCsCocs4aXmca3kX1RunHR1YUAAD3EQhYQ7gGHrzbPd9mUeXyH0T9i3761N0FXHmwpgtwT0o+OvJ1pgFeJEuWSO0kAADuhIgelNL+hmvgwbtduzqdKpQZE60vebXfQEH/1g0p6REBwsJZy9FcgzwyR44TAAB3Y5r0J9O0g1wDD97t5rTTVKPK+CJ9wZd9YhRNmrC9kH/HmiiA7lsBXdRlc56BHrY3SpYY0V0AALyBaVI7rmEH81c9sNep/dKIvuQbJk/Nu+dfkFcuX6Z/ESCSElDVWeM9TMP9Hqbue6JEanUBAPAWpmnP5Rp2sPCrAZ07LFUP8oW1ve+8z/aR33/jHv9da6QARaDP/c56t3Tc4mjxDPub2o7bSj46onWjRnMxchYAL5KVRb8wTfsg16CDhavm9X+x4iiNHZVGvXqupK6dllPP7p/T0EHrVejnbe0bjj4p6TcCRGnHwJG/LfloaitnM52+TtBOLvXYyLnx1vl5PlY/jwr9J54Y8jsBAACBAD2K5wG8bSBA3QQAAADvoQZZcA0nWNzhb6erqWoCAACAN1H3gLmGFCzOAkBNBAAAAO9CRA8EAvZyrkEFi8VdRHS/AAAA4G1ycuiXmBfgJamaAAAAABRq5rVp2lk8AwtG0YUCAAAAuJVgkEqbpm0yDS5YdIM+H/1BAAAAAHeiLg+r3eGYBhgsglIStn0FAABQ6B4BKRgf7Dr3EtGDAgAAACgMKak1SoBrvB4MUhkBAAAAoAR4SaufAAAAAFACPOV+IvqpAAAAAHQJBKi5adrXmAYcLNigadL/CQAAAKAIJaC6adoG06CD+aiu3ggAAAAgSvsEZHINPHir1iwBAAAARAvDoP8IBOwjjqQn16B0pTuJ6OcCAAAAiPbsACfQlzpSUUWRiLpnL12i3wsAAACg+KYIWsOcQLYdKVFlGuKRagYC9LgAAAAAYvBcwLNOkF7kGvAeKhJXAgGqIwAAAIBYcfky/atp2pu5hnOiaZraXpWS6gkAAAAg1hDRTwIBetMJsEtcg9WlXlVXYQQAAAAQT/x++qMTSmuYhqWrDAZtE5f9AQAAsIGI7gsGqa0TUpJreLrAjGCQSgoAAACAG+rZACeoPmcaoInsdrW2AgAAAOBMIEA1nNDaxTRME0xrGhE9LAAAAIBEgIjuDwSohRNi3/MMVvb61foJAAAAIBEhoodynw/IZBq0HB/226QerhQAAABAouPz0a+CQertBNx5rsHLwOxgkNoQ0f0CAAAAcBNE9NNAgFo5YbeHaQjHwxumaY03DPq1AAAAANyOlFQpGLTnqgBkGszFre24LBCgxwQAAADgNaSkR0zTGuiEYTrToI62N1TxCQbpCQEAAOD/27djlSzDKIDjb0NLFNTcEq3eQVMtDtEVFHQJFnQD4dRiU0sRgrfg6qijeA+un8KnnvOqIM93fD9QUJxExU/8/eAPzx0ceDiHx66qnhwc1Lu+b38yJ6MZHd43aZzZfh8d1ZsOALiqqp72fX3MbCsPfHGwDa1l1hf3/ABwTcfH9XZ6Tni2M7A/o8P+4hf/RmZ9Ozys1x0AcDuXBBH1PqItZk5Wh7bveeCfDG1ltqW+r087O/WiAwDu3t5evYqoD5n1PbMtZ042h3bv5lZ/sp7Z/kfUwnRnoaqedQDA7BiN6nlEzWXW/NDniFqIaIuZ7VdE+zv0L7OtTN/nZbaliPYzs34MfR2aj6i58bhedgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAJecAkDdBxs29doZAAAAAElFTkSuQmCC"""

def set_app_icon(root, b64_str: str):
    """
    ตั้งไอคอนโปรแกรม (window icon) จาก Base64 PNG
    - รองรับทุกแพลตฟอร์มที่ Tk PhotoImage รองรับ PNG
    - กัน GC ด้วยการเก็บ ref ไว้ที่ root._app_icon
    """
    try:
        icon = tk.PhotoImage(data=b64_str)
        root.iconphoto(True, icon)   # True = ใช้กับทุกหน้าต่างของแอป
        root._app_icon = icon        # กันถูกเก็บขยะ
    except Exception as e:
        print("ตั้งไอคอนไม่สำเร็จ:", e)


def create_format_kao(df_original):
    """
    - ยึด Column (B) เป็นกรุ๊ป และกำหนด Item No. ของกรุ๊ปจาก 'แถวคำถาม' แรกที่ C ไม่ว่าง
      (พิจารณาเฉพาะแถวที่ D เป็นชนิดคำถามเท่านั้น; ไม่ใช้ค่า C ที่ถูก ffill จากแถวคำตอบ)
    - แถวคำถามต้องมี B เป็นตัวเลข
    - รองรับชนิดขึ้นต้นด้วย 'M' ทั้งหมดให้เป็น MA (M, ML, MR, ...)
    - G เว้นว่างเสมอ และรวมคำตอบตัวแรกไว้แถวเดียวกับคำถาม
    """
    import math
    import pandas as pd

    COL_IDX_NUM = 1        # B
    COL_IDX_QUESTION = 2   # C
    COL_IDX_TYPE = 3       # D
    COL_IDX_ANS_CODE = 4   # E
    COL_IDX_TITLE = 5      # F

    kao_columns = [
        'Column', '', 'Item No.', 'item type', 'Number of categories/digits',
        'item name', 'Question text/comments', 'Category Number', 'category label'
    ]

    def _clean(x):
        if x is None:
            return ''
        try:
            if isinstance(x, float) and math.isnan(x):
                return ''
        except Exception:
            pass
        return str(x).strip()

    def _to_int_or_none(x):
        try:
            v = float(str(x).strip())
            if math.isnan(v): return None
            return int(v)
        except Exception:
            return None

    def _normalize_type(raw):
        t = _clean(raw).upper()
        if not t: return None
        if t in ('SA', 'MA', 'RN'): return t
        if t in ('S', 'SL'): return 'SA'
        if t in ('X', 'N'):  return 'RN'
        if t.startswith('M'): return 'MA'  # M, ML, MR, ...
        return None

    def _col_to_int(x):
        s = _clean(x)
        if not s: return None
        try:
            v = float(s)
            if math.isnan(v): return None
            return int(v)
        except Exception:
            return None

    # ---------- 1) หาค่า Item No. ต่อคอลัมน์ "จากแถวคำถามเท่านั้น" ----------
    item_no_by_col = {}
    rows = list(df_original.itertuples(index=False, name=None))
    for r in rows:
        col_int = _col_to_int(r[COL_IDX_NUM] if len(r) > COL_IDX_NUM else None)
        if col_int is None:
            continue
        qtype_norm = _normalize_type(r[COL_IDX_TYPE] if len(r) > COL_IDX_TYPE else None)
        if qtype_norm is None:
            continue  # ไม่ใช่แถวคำถาม
        cand = _clean(r[COL_IDX_QUESTION] if len(r) > COL_IDX_QUESTION else None)
        if cand and col_int not in item_no_by_col:
            item_no_by_col[col_int] = cand  # เก็บตัวแรกที่ไม่ว่างของ "แถวคำถาม"

    # ---------- 2) สร้างตาราง Format_Kao ----------
    new_rows = []
    current = None

    for r in rows:
        col_raw = r[COL_IDX_NUM] if len(r) > COL_IDX_NUM else None
        col_int = _col_to_int(col_raw)
        qtype_norm = _normalize_type(r[COL_IDX_TYPE] if len(r) > COL_IDX_TYPE else None)
        title = _clean(r[COL_IDX_TITLE] if len(r) > COL_IDX_TITLE else None)
        ans_code = _to_int_or_none(r[COL_IDX_ANS_CODE] if len(r) > COL_IDX_ANS_CODE else None)

        # แถวคำถาม
        is_question = (qtype_norm is not None) and (title != '' or _clean(r[COL_IDX_QUESTION] if len(r) > COL_IDX_QUESTION else None) != '')
        if is_question:
            if col_int is None:       # ต้องมีเลขในคอลัมน์ B
                current = None
                continue
            item_no_use = item_no_by_col.get(col_int, _clean(r[COL_IDX_QUESTION] if len(r) > COL_IDX_QUESTION else None))
            current = {
                'Column': col_raw,
                '': '',
                'Item No.': item_no_use,       # ยึดค่าตามคอลัมน์นี้
                'item type': qtype_norm,
                'Number of categories/digits': '',
                'item name': title,
                'Question text/comments': '',
                'Category Number': '',
                'category label': ''
            }
            new_rows.append(current)
            continue

        # แถวคำตอบ
        if current and ans_code is not None:
            label = r[COL_IDX_TITLE] if len(r) > COL_IDX_TITLE else None
            if current['Category Number'] == '':
                current['Category Number'] = ans_code
                current['category label'] = label
            else:
                new_rows.append({
                    'Column': col_raw,
                    '': '',
                    'Item No.': '',
                    'item type': '',
                    'Number of categories/digits': '',
                    'item name': '',
                    'Question text/comments': '',
                    'Category Number': ans_code,
                    'category label': label
                })

    return pd.DataFrame(new_rows, columns=kao_columns)






def style_header_row(ws, header_fill_hex="D9E1F2"):
    """
    จัดรูปแบบแถวหัวตารางให้ 'หนา + เติมสี + กึ่งกลาง'
    โดย 'ไม่เปิด wrap text' เพื่อกันความสูงแถวเพิ่ม
    """
    header_fill = PatternFill(start_color=header_fill_hex, end_color=header_fill_hex, fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for cell in ws[1]:  # แถวที่ 1 = header
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # ย้ำให้ใช้ความสูงมาตรฐาน (เผื่อเคยถูกตั้งค่ามาก่อน)
    ws.row_dimensions[1].height = None


def style_format_kao_sheet(ws_kao):
    """
    - ใส่เส้นขอบทุกเซลล์
    - ลงสีเทาให้ 'แถวคำถาม' โดยดูที่คอลัมน์ D (item type) != ว่าง
    - ทำ header ให้เรียบร้อย + ปรับความกว้างคอลัมน์ F และ I
    """
    grey_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # เส้นขอบ + แถวคำถามเป็นสีเทา (ยกเว้น header)
    for row in ws_kao.iter_rows(min_row=1, max_col=ws_kao.max_column, max_row=ws_kao.max_row):
        is_question_row = row[3].value is not None and str(row[3].value).strip() != ""  # D = item type
        for cell in row:
            cell.border = thin_border
            if cell.row > 1 and is_question_row:
                cell.fill = grey_fill

    # จัด header แถวที่ 1 ให้หนา+สี (โทนเดียวกับภาพ)
    style_header_row(ws_kao, header_fill_hex="D9E1F2")

    # ปรับความกว้างคอลัมน์ตามภาพ: F (item name) กว้าง, I (category label) กว้างกว่า
    # **ปรับค่าได้ตามใจ** ถ้าอยากแคบ/กว้างกว่านี้
    ws_kao.column_dimensions["F"].width = 40
    ws_kao.column_dimensions["I"].width = 60






class ExcelRecodeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("โปรแกรม Create Format_Kao By_DP V3")
        
        # --- START: ส่วนที่ปรับแก้ให้โปรแกรมแสดงกลางจอ ---
        
        # 1. กำหนดขนาดหน้าต่างที่ต้องการ
        window_width = 650
        window_height = 600

        # 2. หาขนาดของหน้าจอคอมพิวเตอร์
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # 3. คำนวณตำแหน่งกึ่งกลาง (x, y) สำหรับมุมซ้ายบนของหน้าต่าง
        # หลักการคือ (ขนาดจอ / 2) - (ขนาดโปรแกรม / 2)
        center_x = int((screen_width / 2) - (window_width / 2))
        center_y = int((screen_height / 2) - (window_height / 2))

        # 4. ตั้งค่าขนาดและตำแหน่งของโปรแกรมในคำสั่งเดียว
        # รูปแบบคือ "widthxheight+x+y"
        self.root.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
        
        # --- END: สิ้นสุดส่วนที่ปรับแก้ ---

        # self.root.geometry("650x600") # <--- ลบบรรทัดเดิมนี้ออก หรือคอมเมนต์ไว้
        self.root.configure(bg='#F0F0F0')

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
        style = ttk.Style(self.root)
        style.theme_use('clam')

        self.BG_COLOR = '#F0F0F0'
        self.PRIMARY_BLUE = '#0078D7'
        self.LIGHT_BLUE = '#2091EB'
        self.SUCCESS_GREEN = '#107C10'
        self.LIGHT_GREEN = '#30c553'
        self.WARNING_YELLOW = '#ffc107'
        self.LIGHT_YELLOW = '#ffd24a'
        self.ERROR_RED = '#D32F2F'
        self.TEXT_COLOR_DARK = '#333333'
        self.BORDER_COLOR = '#CCCCCC'
        self.STATUS_PROCESSING_BLUE = '#005A9E'

        self.NA_FULLCELL = re.compile(r'^\s*N[\s\.\-_/\\]*A\s*$', re.IGNORECASE)

        style.configure('.', background=self.BG_COLOR, foreground=self.TEXT_COLOR_DARK, font=('Segoe UI', 10))
        style.configure('TLabel', background=self.BG_COLOR)
        style.configure('TLabelFrame', background=self.BG_COLOR, bordercolor=self.BORDER_COLOR)
        style.configure('TLabelFrame.Label', background=self.BG_COLOR, font=('Segoe UI', 10, 'bold'))
        style.configure('TEntry', fieldbackground='white', bordercolor=self.BORDER_COLOR, lightcolor='white', darkcolor='white')
        style.configure('Select.TButton', background=self.PRIMARY_BLUE, foreground='white', font=('Segoe UI', 10, 'bold'), borderwidth=0)
        style.map('Select.TButton', background=[('active', self.LIGHT_BLUE)])
        style.configure('Find.TButton', background=self.SUCCESS_GREEN, foreground='white', font=('Segoe UI', 10, 'bold'), borderwidth=0)
        style.map('Find.TButton', background=[('active', self.LIGHT_GREEN)])
        style.configure('Process.TButton', background=self.WARNING_YELLOW, foreground='black', font=('Segoe UI', 10, 'bold'), borderwidth=0)
        style.map('Process.TButton', background=[('active', self.LIGHT_YELLOW)])
        style.configure("green.Horizontal.TProgressbar", troughcolor='#E0E0E0', background=self.SUCCESS_GREEN, bordercolor=self.SUCCESS_GREEN)

        self.file_path = ""
        self.groups_to_fix = {}
        self.file_path_var = tk.StringVar(value="ยังไม่ได้เลือกไฟล์...")

        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.grid(row=0, column=0, sticky='nsew')
        main_frame.rowconfigure(2, weight=1)
        main_frame.columnconfigure(0, weight=1) # เพิ่มเพื่อให้ frame ขยายตามแนวนอน

        file_frame = ttk.LabelFrame(main_frame, text="ขั้นตอนที่ 1: เลือกไฟล์ Excel", padding="15")
        file_frame.grid(row=0, column=0, sticky='ew', pady=(0, 10))
        file_frame.columnconfigure(0, weight=1)
        self.path_entry = ttk.Entry(file_frame, textvariable=self.file_path_var, state='readonly', font=('Segoe UI', 9))
        self.path_entry.grid(row=0, column=0, sticky='ew', padx=(0, 15))
        browse_button = ttk.Button(file_frame, text="เลือกไฟล์...", style='Select.TButton', command=self.select_file)
        browse_button.grid(row=0, column=1)

        action_frame = ttk.Frame(main_frame)
        action_frame.grid(row=1, column=0, sticky='ew', pady=10)
        action_frame.columnconfigure((0, 1), weight=1)
        self.find_button = ttk.Button(action_frame, text="ขั้นตอนที่ 2: ค้นหาข้อที่ต้องแก้ไข", style='Find.TButton', command=self.find_issues)
        self.find_button.grid(row=0, column=0, sticky='ew', ipady=5, padx=(0, 5))
        self.process_button = ttk.Button(action_frame, text="ขั้นตอนที่ 3: ประมวลผลและบันทึก", style='Process.TButton', command=self.process_and_save, state="disabled")
        self.process_button.grid(row=0, column=1, sticky='ew', ipady=5, padx=(5, 0))

        result_frame = ttk.LabelFrame(main_frame, text="ผลลัพธ์การค้นหา", padding="1")
        result_frame.grid(row=2, column=0, sticky='nsew', pady=(10, 0))
        result_frame.rowconfigure(0, weight=1)
        result_frame.columnconfigure(0, weight=1)
        self.result_text = tk.Text(result_frame, wrap="word", height=10, font=('Segoe UI', 10), bg='white', fg=self.TEXT_COLOR_DARK, padx=10, pady=10, relief='solid', borderwidth=1, selectbackground=self.LIGHT_BLUE)
        self.result_text.grid(row=0, column=0, sticky='nsew')
        scrollbar = ttk.Scrollbar(result_frame, orient='vertical', command=self.result_text.yview)
        scrollbar.grid(row=0, column=1, sticky='ns')
        self.result_text.config(yscrollcommand=scrollbar.set)
        
        self.result_text.tag_configure("header", font=('Segoe UI', 11, 'bold'), foreground=self.PRIMARY_BLUE, spacing3=5)
        self.result_text.tag_configure("success", font=('Segoe UI', 10, 'italic'), foreground=self.SUCCESS_GREEN)
        self.result_text.tag_configure("bullet", lmargin1=20, lmargin2=20)
        
        # <<-- แก้ไข: ใช้ grid_propagate เพื่อล็อคขนาดของ status_frame -->>
        status_frame = ttk.Frame(self.root, padding=(10, 5, 10, 10))
        status_frame.grid(row=1, column=0, sticky='ew')
        status_frame.grid_propagate(False) # <--- คำสั่งสำคัญ: ปิดการขยายตัวตามเนื้อหา
        status_frame.config(height=45)     # <--- คำสั่งสำคัญ: กำหนดความสูงคงที่
        status_frame.columnconfigure(0, weight=1)

        self.status_label = ttk.Label(status_frame, text="พร้อมใช้งาน", font=('Segoe UI', 9))
        self.status_label.grid(row=0, column=0, sticky='w', pady=(0, 2))

        self.progress_bar = ttk.Progressbar(status_frame, orient='horizontal', mode='determinate', style="green.Horizontal.TProgressbar")



    def is_na_title(self, v):
        if pd.isna(v): return False
        s = unicodedata.normalize('NFKC', str(v)).strip()
        return bool(self.NA_FULLCELL.match(s))

    def is_literal_none(self, v):
        if pd.isna(v): return False
        s = unicodedata.normalize('NFKC', str(v)).strip().lower()
        return s == 'none'

    def select_file(self):
        path = filedialog.askopenfilename(title="กรุณาเลือกไฟล์ Excel", filetypes=(("Excel Files", "*.xlsx"), ("All files", "*.*")))
        if path:
            self.file_path = path; self.file_path_var.set(path); self.result_text.delete("1.0", tk.END); self.process_button.config(state="disabled"); self.groups_to_fix = {}
            self.status_label.config(text="ไฟล์ถูกเลือกแล้ว, กรุณากดขั้นตอนที่ 2 เพื่อค้นหา", foreground=self.TEXT_COLOR_DARK)

    def find_issues(self):
        if not self.file_path: messagebox.showwarning("คำเตือน", "กรุณาเลือกไฟล์ Excel ก่อน"); return
        self.result_text.delete("1.0", tk.END); self.process_button.config(state="disabled"); self.groups_to_fix = {}
        try:
            df = pd.read_excel(self.file_path, sheet_name="Format", header=None, keep_default_na=False, na_values=[""])
            col_c, col_d, col_e, col_f = 2, 3, 4, 5
            df[col_c] = df[col_c].ffill()
            is_d_blank = df[col_d].isna(); is_f_blank = df[col_f].isna() | (df[col_f].astype(str).str.strip() == ""); is_f_literal_none = df[col_f].astype(str).str.strip().str.lower().eq("none")
            condition = is_d_blank & is_f_blank & (~is_f_literal_none)
            blank_rows_df = df[condition].copy()
            if not blank_rows_df.empty:
                blank_rows_df[col_e] = pd.to_numeric(blank_rows_df[col_e], errors='coerce')
                blank_rows_df.dropna(subset=[col_e], inplace=True)
                summary = blank_rows_df.groupby(col_c, sort=False)[col_e].agg(['min', 'max'])
                self.result_text.insert("1.0", "พบข้อที่Codeกระโดด:\n", "header")
                for q, data in summary.iterrows():
                    min_c, max_c = int(data['min']), int(data['max']); self.groups_to_fix[q] = {'min': min_c, 'max': max_c}
                    range_text = f"{min_c}-{max_c}" if min_c != max_c else f"{min_c}"; self.result_text.insert(tk.END, f"\n• {q} มี Code กระโดดเริ่มที่ Code>> {range_text}", "bullet")
                self.status_label.config(text="พบข้อที่ต้องแก้ไข, กดขั้นตอนที่ 3 เพื่อดำเนินการ", foreground=self.TEXT_COLOR_DARK); self.process_button.config(state="normal")
            else:
                self.result_text.insert("1.0", "ตรวจสอบเรียบร้อย\n", "header"); self.result_text.insert(tk.END, "\nไม่พบข้อที่Codeกระโดด\nคุณสามารถกด 'ประมวลผลและบันทึก' ได้เลย", "success")
                self.status_label.config(text="ไม่พบข้อผิดพลาด, กดขั้นตอนที่ 3 เพื่อดำเนินการ", foreground=self.TEXT_COLOR_DARK); self.process_button.config(state="normal")
        except Exception as e:
            messagebox.showerror("เกิดข้อผิดพลาด", f"ไม่สามารถอ่านไฟล์ได้:\n{e}"); self.status_label.config(text="เกิดข้อผิดพลาดในการอ่านไฟล์", foreground=self.ERROR_RED)

    def start_processing_ui(self):
        self.progress_bar['value'] = 0
        self.progress_bar.grid(row=1, column=0, sticky='ew')
        self.root.config(cursor="watch"); self.find_button.config(state="disabled"); self.process_button.config(state="disabled")
        self.update_progress(0, "กำลังเริ่มต้น...")

    def end_processing_ui(self, success=True, message=""):
        self.progress_bar.grid_remove() 
        self.root.config(cursor="")
        self.find_button.config(state="normal"); self.process_button.config(state="disabled")
        if success:
            self.status_label.config(text="ประมวลผลสำเร็จ!", foreground=self.SUCCESS_GREEN)
            messagebox.showinfo("สำเร็จ", f"ไฟล์ถูกประมวลผลและบันทึกเป็น:\n{message}")
        else:
            self.status_label.config(text="เกิดข้อผิดพลาดในการประมวลผล", foreground=self.ERROR_RED)
            messagebox.showerror("เกิดข้อผิดพลาด", f"ไม่สามารถประมวลผลไฟล์ได้:\n{message}")

    def update_progress(self, value, text):
        self.progress_bar['value'] = value
        self.status_label.config(text=text, foreground=self.STATUS_PROCESSING_BLUE)
        self.root.update_idletasks()

    def process_and_save(self):
        self.start_processing_ui()
        self.processor = self.process_and_save_generator()
        self.run_next_step()

    def run_next_step(self):
        try:
            next(self.processor)
            self.root.after(20, self.run_next_step)
        except StopIteration:
            pass

    def process_and_save_generator(self):
        output_filename = ""
        try:
            self.update_progress(5, "กำลังอ่านไฟล์ Excel...")
            df_format = pd.read_excel(self.file_path, sheet_name="Format", header=None, keep_default_na=False, na_values=[''])
            df_data = pd.read_excel(self.file_path, sheet_name="data", keep_default_na=False, na_values=[''])
            yield
            self.update_progress(10, "อ่านไฟล์สำเร็จ, กำลังวิเคราะห์ข้อมูล...")
            
            # ... (ส่วนตรรกะที่เหลือทั้งหมดเหมือนเดิมทุกประการ ไม่มีการเปลี่ยนแปลง) ...
            col_b, col_c, col_d, col_e, col_f = 1, 2, 3, 4, 5
            work = df_format.copy(); work[col_c] = work[col_c].ffill(); work[col_e] = pd.to_numeric(work[col_e], errors='coerce'); yield
            na_rows, na_codes_by_q = [], {}
            for i, row in work.iterrows():
                if pd.isna(row[col_d]) and self.is_na_title(row[col_f]) and not self.is_literal_none(row[col_f]) and pd.notna(row[col_e]):
                    na_rows.append(i); q = str(row[col_c]); na_codes_by_q.setdefault(q, set()).add(int(row[col_e]))
            yield
            self.update_progress(20, "กำลังคำนวณการจัดเรียงรหัสใหม่...")
            recode_maps, logs = {}, []; final_format_df = df_format.copy()
            if self.groups_to_fix:
                for q in self.groups_to_fix.keys():
                    is_q, is_ans, is_e_number = (work[col_c] == q), (work[col_c] == q) & work[col_d].isna(), work[col_e].notna()
                    is_na_ans = is_ans & work.index.isin(na_rows); is_f_blank = work[col_f].isna() | (work[col_f].astype(str).str.strip() == "")
                    is_not_special_case = ~work[col_f].apply(lambda x: self.is_na_title(x) or self.is_literal_none(x))
                    is_blank_ans = is_ans & is_f_blank & is_e_number & is_not_special_case
                    rows_to_keep_mask = is_ans & is_e_number & (~is_na_ans) & (~is_blank_ans)
                    if rows_to_keep_mask.any():
                        rows_to_keep_indices = work[rows_to_keep_mask].index; old_codes = work.loc[rows_to_keep_indices, col_e].astype(int).tolist()
                        new_codes = list(range(1, len(old_codes) + 1)); recode_maps[q] = dict(zip(old_codes, new_codes))
                        final_format_df.loc[rows_to_keep_indices, col_e] = new_codes
                        logs.append({'Timestamp': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),'Question': q, 'Action': 'Re-coded choices due to blank rows','Details': f"Original choices {old_codes} were re-numbered to {new_codes}.",'Data Records Modified': 0,'KEY': ''})
            yield
            self.update_progress(30, "กำลังลบบรรทัดที่ไม่จำเป็น...")
            blank_rows_to_drop = []
            if self.groups_to_fix:
                is_d_blank, is_f_blank, is_e_number = work[col_d].isna(), work[col_f].isna() | (work[col_f].astype(str).str.strip() == ""), work[col_e].notna()
                is_not_special_case = ~work[col_f].apply(lambda x: self.is_na_title(x) or self.is_literal_none(x))
                blank_mask = is_d_blank & is_f_blank & is_e_number & is_not_special_case; blank_rows_to_drop = work[blank_mask].index.tolist()
            all_indices_to_drop = sorted(list(set(na_rows + blank_rows_to_drop)))
            if all_indices_to_drop: final_format_df = final_format_df.drop(index=all_indices_to_drop).reset_index(drop=True)
            yield
            self.update_progress(45, "กำลังปรับปรุงข้อมูลในชีท Data...")
            def recode_move(val, mapping):
                if pd.isna(val): return val, False
                changed, out = False, []
                for p in str(val).split(','):
                    t = p.strip();
                    if not t: continue
                    try: num = int(float(t))
                    except ValueError: out.append(t); continue
                    if num in mapping: out.append(str(mapping[num])); changed = True
                    else: out.append(str(num))
                return ','.join(out), changed
            def drop_codes(val, code_set):
                if pd.isna(val): return val, False
                parts = [x.strip() for x in str(val).split(',') if x.strip() != '']; kept, changed = [], False
                for t in parts:
                    try: num = int(float(t))
                    except ValueError: kept.append(t); continue
                    if num in code_set: changed = True
                    else: kept.append(str(num))
                return ('' if parts and not kept else ','.join(kept)), changed
            for lg in logs:
                q = lg['Question'];
                if q in df_data.columns and q in recode_maps:
                    mp = recode_maps[q]; res = df_data[q].apply(lambda x: recode_move(x, mp)); df_data[q] = res.apply(lambda x: x[0]); m = res.apply(lambda x: x[1]); cnt = int(m.sum()); lg['Data Records Modified'] = cnt
                    if cnt > 0 and 'KEY' in df_data.columns: lg['KEY'] = ','.join(df_data.loc[m, 'KEY'].astype(str).tolist())
            yield
            for q, codes in na_codes_by_q.items():
                if q in df_data.columns and codes:
                    res = df_data[q].apply(lambda x: drop_codes(x, codes)); df_data[q] = res.apply(lambda x: x[0]); m = res.apply(lambda x: x[1]); cnt = int(m.sum())
                    logs.append({'Timestamp': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),'Question': q,'Action': 'Removed NA category','Details': f"Deleted NA rows in Format and dropped code(s) {sorted(list(codes))} from data.",'Data Records Modified': cnt,'KEY': ','.join(df_data.loc[m, 'KEY'].astype(str).tolist()) if cnt > 0 and 'KEY' in df_data.columns else ''})
            yield
            self.update_progress(70, "กำลังสร้างชีท Format_Kao และ Log...")
            final_format_kao_df = create_format_kao(final_format_df)
            wb = openpyxl.load_workbook(self.file_path)
            ws_format = wb["Format"]; old_rows = ws_format.max_row
            for r_idx, row in enumerate(final_format_df.itertuples(index=False), 1):
                for c_idx, val in enumerate(row, 1): ws_format.cell(row=r_idx, column=c_idx, value=val)
            if old_rows > len(final_format_df): ws_format.delete_rows(len(final_format_df)+1, old_rows-len(final_format_df))
            yield
            ws_data = wb["data"]
            for r_idx, row in enumerate(df_data.itertuples(index=False), 2):
                for c_idx, val in enumerate(row, 1): ws_data.cell(row=r_idx, column=c_idx, value=val)
            yield
            if logs:
                log_df = pd.DataFrame(logs);
                if "Log" in wb.sheetnames: del wb["Log"]
                ws_log = wb.create_sheet("Log")
                for r in dataframe_to_rows(log_df, index=False, header=True): ws_log.append(r)
                for col in ws_log.columns: L = col[0].column_letter; mx = max((len(str(c.value)) if c.value is not None else 0) for c in col); ws_log.column_dimensions[L].width = mx + 2
                style_header_row(ws_log, header_fill_hex="D9E1F2"); ws_log.freeze_panes = "A2"
            yield
            if "Format_Kao" in wb.sheetnames: del wb["Format_Kao"]
            ws_kao = wb.create_sheet("Format_Kao")
            ws_kao.append(final_format_kao_df.columns.tolist())
            for r in dataframe_to_rows(final_format_kao_df, index=False, header=False): ws_kao.append(r)
            style_format_kao_sheet(ws_kao); ws_kao.freeze_panes = "A2"
            yield
            tgt = next((n for n in wb.sheetnames if n.strip().lower() == "survey overview"), None)
            if tgt: del wb[tgt]
            try: wb.move_sheet(ws_kao, -wb.worksheets.index(ws_kao))
            except Exception: sh = wb._sheets; sh.insert(0, sh.pop(sh.index(ws_kao)))
            wb.active = wb.worksheets.index(ws_kao)
            yield
            self.update_progress(95, "กำลังบันทึกไฟล์ผลลัพธ์...")
            base, ext = os.path.splitext(self.file_path)
            out = f"{base}_recoded.xlsx"
            wb.save(out)
            output_filename = out
            self.update_progress(100, "บันทึกไฟล์สำเร็จ!")
        except Exception as e:
            self.end_processing_ui(success=False, message=str(e))
        else:
            self.end_processing_ui(success=True, message=output_filename)





# <<< START OF CHANGES >>>
# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
def run_this_app(working_dir=None): # ชื่อฟังก์ชันนี้จะถูกใช้ใน Launcher
    """
    ฟังก์ชันหลักสำหรับสร้างและรัน QuotaSamplerApp.
    """
    print(f"--- QUOTA_SAMPLER_INFO: Starting 'QuotaSamplerApp' via run_this_app() ---")
    try:
    # --- ส่วนที่ใช้รันโปรแกรม ---
    #if __name__ == "__main__":
        root = tk.Tk()
        set_app_icon(root, APP_ICON_B64)   # <-- เพิ่มบรรทัดนี้
        app = ExcelRecodeApp(root)
        root.mainloop()
        print(f"--- QUOTA_SAMPLER_INFO: QuotaSamplerApp mainloop finished. ---")

    except Exception as e:
        # ดักจับ Error ที่อาจเกิดขึ้นระหว่างการสร้างหรือรัน App
        print(f"QUOTA_SAMPLER_ERROR: An error occurred during QuotaSamplerApp execution: {e}")
        # แสดง Popup ถ้ามีปัญหา
        if 'root' not in locals() or not root.winfo_exists(): # สร้าง root ชั่วคราวถ้ายังไม่มี
            root_temp = tk.Tk()
            root_temp.withdraw()
            messagebox.showerror("Application Error (Quota Sampler)",
                               f"An unexpected error occurred:\n{e}", parent=root_temp)
            root_temp.destroy()
        else:
            messagebox.showerror("Application Error (Quota Sampler)",
                               f"An unexpected error occurred:\n{e}", parent=root) # ใช้ root ที่มีอยู่ถ้าเป็นไปได้
        sys.exit(f"Error running QuotaSamplerApp: {e}") # อาจจะ exit หรือไม่ก็ได้ ขึ้นกับการออกแบบ


# --- ส่วน Run Application เมื่อรันไฟล์นี้โดยตรง (สำหรับ Test) ---
if __name__ == "__main__":
    print("--- Running QuotaSamplerApp.py directly for testing ---")
    # (ถ้ามีการตั้งค่า DPI ด้านบน มันจะทำงานอัตโนมัติ)

    # เรียกฟังก์ชัน Entry Point ที่เราสร้างขึ้น
    run_this_app()

    print("--- Finished direct execution of QuotaSamplerApp.py ---")
# <<< END OF CHANGES >>>