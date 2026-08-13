# -*- coding: utf-8 -*-
"""
ตัวแปลง SPSS เป็น Itemdef Excel  (PyQt6 Edition)
--------------------------------------------------
พอร์ตมาจาก Program_ItemdefSPSS_Log.py (Tkinter) โดย:
  - Logic การจับ Loop / Manual Group / เขียน Excel เหมือนเดิมทุกประการ
  - เปลี่ยนเฉพาะชั้น UI เป็น PyQt6 และจัด Layout ให้คล้ายของเดิม
"""

import os
import re
import sys
import traceback
from collections import defaultdict

import pandas as pd
import pyreadstat
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation

from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtCore import Qt


# =========================================================================
#  Helper
# =========================================================================
def resource_path(relative_name):
    """คืน path ที่ถูกต้อง ไม่ว่าจะ run จากไฟล์ .py หรือ bundle เป็น exe"""
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative_name)


LOOP_TYPES = ("SA", "MA", "Loop Text", "Loop Numeric")


def get_base_name_heuristic(var_name):
    """
    Tries to derive a base name by removing trailing/intermediate numbers/identifiers.
    Handles more complex patterns for grouping including _r<number> and prefix_num_suffix.
    """

    # --- REMOVED 1.57: กฎเดิม 1.55 (I_<loopIndex>_<stem>_<subIndex> -> "I_<loopIndex>_<stem>")
    # ถูกยกเลิก เพราะทำให้ base ยังมีเลข loop ติดอยู่ ตัวแปรอย่าง
    # I_1_s14_1, I_2_s14_1, ... I_13_s14_1 จึงได้ base ต่างกันทุกตัว
    # และถูกแยกเป็น Loop ID คนละอัน (s14_1, s14_2, ...) แทนที่จะรวมเป็น s14_1 อันเดียว
    # ตอนนี้ปล่อยให้ตกไปที่กฎ 1.6 (I_\d+_(.+)$) ซึ่งจะได้ base = "s14_1" / "a4_1" เหมือนกันทุก loop
    # =========================================================================
    # เดิม: 1.4 s13_1_1 -> s13_1
    # (ข้ามกรณี I_<n>_... เพื่อไม่ให้เลข loop ค้างอยู่ใน base เช่น I_1_s13_1_1)
    match_s_num_num = re.match(r'(.+_\d+)_(\d+)$', var_name)
    if match_s_num_num and not re.match(r'I_\d+_', var_name):
        base_part = match_s_num_num.group(1)
        if not base_part.endswith('_O') and not base_part.endswith('_r'):
            return base_part

    # 1.5 I_num_basename_Onum -> basename (เช่น I_1_q1_1_O1 -> q1_1)
    match_i_num_base_onum = re.match(r'I_\d+_(.+?)_O\d+$', var_name)
    if match_i_num_base_onum:
        base_part = match_i_num_base_onum.group(1)
        if base_part:
            return base_part

    # 1.6 I_num_basename -> basename (เช่น I_1_q4_Oth -> q4_Oth, I_1_s14_1 -> s14_1)
    match_i_num_base = re.match(r'I_\d+_(.+)$', var_name)
    if match_i_num_base and not re.search(r'_O\d+$', match_i_num_base.group(1)):
        base_part = match_i_num_base.group(1)
        if base_part:
            return base_part

    # 1.65 prefix_num_Onum -> prefix
    match_pnon = re.match(r'(.+?)_(\d+)_O(\d+)$', var_name)
    if match_pnon:
        prefix = match_pnon.group(1)
        if not re.match(r'I_\d+_', prefix):
            return prefix

    # 1.7 prefix_rNum_suffixNum -> prefix_rNum_suffix
    match_prnsn = re.match(r'(.+?)(_r\d+)(_[a-zA-Z]+)(\d+)$', var_name)
    if match_prnsn:
        return f"{match_prnsn.group(1)}{match_prnsn.group(2)}{match_prnsn.group(3)}"

    # 1.8 prefix_suffixBaseNum -> prefix_suffixBase
    match_psn = re.match(r'(.+?_)([a-zA-Z]+)(\d+)$', var_name)
    if match_psn:
        return f"{match_psn.group(1)}{match_psn.group(2)}"

    # 2. prefix_number_suffix -> prefix_suffix
    match_prefix_num_suffix = re.match(r'(.+?)_(\d+)_([a-zA-Z][a-zA-Z_]*)$', var_name)
    if match_prefix_num_suffix:
        prefix = match_prefix_num_suffix.group(1)
        suffix = match_prefix_num_suffix.group(3)
        if not prefix.startswith('I_') or not prefix[2:].isdigit():
            return f"{prefix}_{suffix}"

    # 3.5 name_r<number> -> name_r
    match_r_num = re.match(r'(.+_r)(\d+)$', var_name)
    if match_r_num:
        return match_r_num.group(1)

    # 1. name(number) -> name
    match_paren = re.match(r'(.+)\((\d+)\)$', var_name)
    if match_paren:
        return match_paren.group(1)

    # 4. name[_ ]number -> name
    match_name_num = re.match(r'(.+?)[_ ](\d+)$', var_name)
    if match_name_num:
        base_part = match_name_num.group(1).rstrip('_ ')
        if base_part and not base_part.endswith('_r'):
            if not re.match(r'.+?_\d+$', base_part):
                return base_part

    # 3/3.1 name + trailing letter
    match_name_letter = re.match(r'(.+?)[_ ]([A-Za-z])$', var_name)
    if match_name_letter:
        base_part = match_name_letter.group(1).rstrip('_ ')
        if base_part and not base_part[-1].isdigit():
            return base_part
    match_name_letter_direct = re.match(r'(.+)([A-Za-z])$', var_name)
    if match_name_letter_direct:
        base_part = match_name_letter_direct.group(1)
        if base_part:
            return base_part

    return var_name


# =========================================================================
#  Stylesheet
# =========================================================================
STYLESHEET = """
QWidget {
    background-color: #F5F7FA;
    color: #1A202C;
    font-family: "Segoe UI", "Tahoma", sans-serif;
    font-size: 10pt;
}
QGroupBox {
    background-color: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-radius: 8px;
    margin-top: 14px;
    padding: 10px 12px 12px 12px;
    font-weight: 600;
}
QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    left: 12px;
    padding: 0 6px;
    color: #2D3748;
}
QLabel#FieldLabel {
    color: #C53030;
    font-weight: 700;
}
QLabel#StatusLabel {
    color: #2B6CB0;
    font-weight: 600;
    padding: 4px 2px;
}
QLabel#HintLabel {
    color: #2C5282;
}
QLineEdit {
    background-color: #FFFFFF;
    border: 1px solid #CBD5E0;
    border-radius: 6px;
    padding: 6px 8px;
    selection-background-color: #90CDF4;
}
QLineEdit:read-only {
    background-color: #EDF2F7;
    color: #2D3748;
}
QPushButton {
    background-color: #E2E8F0;
    border: 1px solid #CBD5E0;
    border-radius: 6px;
    padding: 7px 14px;
    font-weight: 600;
}
QPushButton:hover   { background-color: #CBD5E0; }
QPushButton:pressed { background-color: #A0AEC0; }
QPushButton:disabled {
    background-color: #EDF2F7;
    color: #A0AEC0;
    border-color: #E2E8F0;
}
QPushButton[compact="true"] { padding: 6px 8px; font-size: 9pt; }
QPushButton[accent="green"]  { background-color: #48BB78; border-color:#38A169; color:#FFFFFF; }
QPushButton[accent="green"]:hover  { background-color: #38A169; }
QPushButton[accent="blue"]   { background-color: #3182CE; border-color:#2B6CB0; color:#FFFFFF; }
QPushButton[accent="blue"]:hover   { background-color: #2B6CB0; }
QPushButton[accent="orange"] { background-color: #FBD38D; border-color:#F6AD55; color:#7B341E; }
QPushButton[accent="orange"]:hover { background-color: #F6AD55; }
QPushButton[accent="gold"]   { background-color: #D69E2E; border-color:#B7791F; color:#FFFFFF; }
QPushButton[accent="gold"]:hover   { background-color: #B7791F; }
QPushButton[accent="grey"]   { background-color: #CBD5E0; border-color:#A0AEC0; }
QPushButton[accent="steel"]  { background-color: #4682B4; border-color:#3A6D96; color:#FFFFFF; }
QPushButton[accent="steel"]:hover  { background-color: #3A6D96; }
QPushButton[accent="cyan"]   { background-color: #BEE3F8; border-color:#90CDF4; color:#1A365D; }
QPushButton[accent="cyan"]:hover   { background-color: #90CDF4; }
QPushButton[accent="red"]    { background-color: #FED7D7; border-color:#FEB2B2; color:#742A2A; }
QPushButton[accent="red"]:hover    { background-color: #FEB2B2; }
QPlainTextEdit#LogView {
    background-color: #1A202C;
    color: #E2E8F0;
    border: 1px solid #2D3748;
    border-radius: 8px;
    font-family: "Consolas", "Courier New", monospace;
    font-size: 9pt;
}
QTableWidget, QTreeWidget {
    background-color: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-radius: 6px;
    gridline-color: #EDF2F7;
    selection-background-color: #BEE3F8;
    selection-color: #1A365D;
}
QHeaderView::section {
    background-color: #EDF2F7;
    color: #2D3748;
    border: none;
    border-right: 1px solid #E2E8F0;
    border-bottom: 1px solid #CBD5E0;
    padding: 6px 8px;
    font-weight: 600;
}
QScrollBar:vertical   { background: #EDF2F7; width: 12px; margin: 0; border-radius: 6px; }
QScrollBar:horizontal { background: #EDF2F7; height: 12px; margin: 0; border-radius: 6px; }
QScrollBar::handle:vertical, QScrollBar::handle:horizontal {
    background: #A0AEC0; border-radius: 6px; min-height: 24px; min-width: 24px;
}
QScrollBar::handle:vertical:hover, QScrollBar::handle:horizontal:hover { background: #718096; }
QScrollBar::add-line, QScrollBar::sub-line { height: 0; width: 0; }
QRadioButton { padding: 4px 2px; }
"""


# =========================================================================
#  stdout -> Log widget
# =========================================================================
class LogStream(QtCore.QObject):
    """เปลี่ยนทาง print() ไปยัง QPlainTextEdit (ผ่าน signal จึงปลอดภัยข้าม thread)"""
    text_written = QtCore.pyqtSignal(str)

    def __init__(self, log_widget):
        super().__init__()
        self._alive = True
        self.text_written.connect(self._append)
        self._log_widget = log_widget

    @QtCore.pyqtSlot(str)
    def _append(self, text):
        if not self._alive or self._log_widget is None:
            return
        try:
            cursor = self._log_widget.textCursor()
            cursor.movePosition(QtGui.QTextCursor.MoveOperation.End)
            cursor.insertText(text)
            self._log_widget.setTextCursor(cursor)
            self._log_widget.ensureCursorVisible()
        except RuntimeError:
            self._alive = False

    def write(self, text):
        if self._alive and text:
            self.text_written.emit(text)

    def flush(self):
        pass

    def isatty(self):
        return False

    def destroy(self):
        self._alive = False
        self._log_widget = None


# =========================================================================
#  Dialog: เลือก Loop Type ของ Manual Group
# =========================================================================
class LoopTypeDialog(QtWidgets.QDialog):
    def __init__(self, parent, default_type, row_count):
        super().__init__(parent)
        self.setWindowTitle("เลือก Loop Type ของ Manual Group")
        self.setModal(True)
        self._choice = default_type if default_type in LOOP_TYPES else "SA"

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 18)
        layout.setSpacing(10)

        title = QtWidgets.QLabel(f"รวม {row_count} รายการเป็น Loop เดียว")
        title.setStyleSheet("font-size: 12pt; font-weight: 700; color:#2C5282;")
        layout.addWidget(title)
        layout.addWidget(QtWidgets.QLabel("เลือกประเภท Loop ที่ต้องการ:"))

        self._buttons = {}
        box = QtWidgets.QGroupBox()
        box_layout = QtWidgets.QVBoxLayout(box)
        box_layout.setSpacing(4)
        for loop_type, desc in (("SA", "Single Answer"), ("MA", "Multiple Answer"),
                                ("Loop Text", "ข้อความ"), ("Loop Numeric", "ตัวเลข")):
            radio = QtWidgets.QRadioButton(f"{loop_type}   ({desc})")
            radio.setChecked(loop_type == self._choice)
            box_layout.addWidget(radio)
            self._buttons[loop_type] = radio
        layout.addWidget(box)

        btn_box = QtWidgets.QDialogButtonBox()
        ok_btn = btn_box.addButton("ตกลง", QtWidgets.QDialogButtonBox.ButtonRole.AcceptRole)
        ok_btn.setProperty("accent", "cyan")
        btn_box.addButton("ยกเลิก", QtWidgets.QDialogButtonBox.ButtonRole.RejectRole)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def selected_type(self):
        for loop_type, radio in self._buttons.items():
            if radio.isChecked():
                return loop_type
        return None

    @staticmethod
    def ask(parent, default_type, row_count):
        """คืน Loop Type ที่เลือก หรือ None ถ้ายกเลิก"""
        dlg = LoopTypeDialog(parent, default_type, row_count)
        if dlg.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            return dlg.selected_type()
        return None


# =========================================================================
#  Dialog: ตรวจสอบ Value Code
# =========================================================================
class CodeValidationDialog(QtWidgets.QDialog):
    """
    แสดง Value Code ที่เป็น 0 หรือ > 2000 ให้ผู้ใช้ติ๊กเลือกว่าจะลบตัวไหน
    display_items: {display_name: [actual_var, ...]}
    """

    def __init__(self, parent, ordered_display_names, display_items, codes_to_validate, all_value_labels):
        super().__init__(parent)
        self.setWindowTitle("ตรวจสอบ Value Code")
        self.setModal(True)
        self.resize(720, 560)
        self.result_pairs = set()

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(10)

        header = QtWidgets.QLabel("พบ Value Code ที่เป็น 0 หรือเกิน 2000\nติ๊กเลือก Code ที่ต้องการลบออกจากผลลัพธ์")
        header.setStyleSheet("font-size: 11pt; font-weight: 700; color:#C53030;")
        layout.addWidget(header)

        self.tree = QtWidgets.QTreeWidget()
        self.tree.setHeaderLabels(["ตัวแปร / กลุ่ม  →  Code", "Label"])
        self.tree.setColumnWidth(0, 320)
        self.tree.setAlternatingRowColors(True)
        layout.addWidget(self.tree, 1)

        for display_name in ordered_display_names:
            actual_vars_in_item = display_items.get(display_name, [])
            problematic = defaultdict(list)   # {formatted_code: [(actual_var, original_code, label)]}
            has_problem = False

            for actual_var_name in actual_vars_in_item:
                if actual_var_name not in codes_to_validate:
                    continue
                has_problem = True
                current_var_labels = all_value_labels.get(actual_var_name, {})
                for original_code_str in codes_to_validate[actual_var_name]:
                    try:
                        f_code = float(original_code_str)
                        formatted_code = str(int(f_code)) if f_code == int(f_code) else original_code_str
                    except (ValueError, TypeError):
                        formatted_code = original_code_str

                    label_text = "???"
                    try:
                        f_code = float(original_code_str)
                        if f_code == int(f_code):
                            label_text = current_var_labels.get(
                                int(f_code), current_var_labels.get(f_code, current_var_labels.get(original_code_str, "???")))
                        else:
                            label_text = current_var_labels.get(
                                f_code, current_var_labels.get(original_code_str, "???"))
                    except (ValueError, TypeError):
                        label_text = current_var_labels.get(original_code_str, "???")

                    problematic[formatted_code].append((actual_var_name, original_code_str, label_text))

            if not has_problem:
                continue

            parent_item = QtWidgets.QTreeWidgetItem(self.tree, [display_name, ""])
            font = parent_item.font(0)
            font.setBold(True)
            parent_item.setFont(0, font)
            parent_item.setExpanded(True)

            unique_codes = sorted(
                problematic.keys(),
                key=lambda x: float(x) if x.replace('.', '', 1).lstrip('-').isdigit() else float('inf'))
            for formatted_code in unique_codes:
                occurrences = problematic[formatted_code]
                representative_label = occurrences[0][2]
                child = QtWidgets.QTreeWidgetItem(parent_item, [f"Code: {formatted_code}", str(representative_label)])
                child.setFlags(child.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                child.setCheckState(0, Qt.CheckState.Checked)   # default: เลือกลบ
                # เก็บคู่ (actual_var, original_code) ทั้งหมดที่ Code นี้แทนอยู่
                child.setData(0, Qt.ItemDataRole.UserRole,
                              [(var, code) for var, code, _lbl in occurrences])

        btn_row = QtWidgets.QHBoxLayout()
        btn_select = QtWidgets.QPushButton("เลือกทั้งหมด")
        btn_deselect = QtWidgets.QPushButton("ไม่เลือกเลย")
        btn_confirm = QtWidgets.QPushButton("ตกลง (ลบรายการที่เลือก)")
        btn_confirm.setProperty("accent", "blue")
        btn_select.clicked.connect(lambda: self._set_all(Qt.CheckState.Checked))
        btn_deselect.clicked.connect(lambda: self._set_all(Qt.CheckState.Unchecked))
        btn_confirm.clicked.connect(self._on_confirm)
        btn_row.addWidget(btn_select)
        btn_row.addWidget(btn_deselect)
        btn_row.addStretch(1)
        btn_row.addWidget(btn_confirm)
        layout.addLayout(btn_row)

    def _iter_code_items(self):
        for i in range(self.tree.topLevelItemCount()):
            top = self.tree.topLevelItem(i)
            for j in range(top.childCount()):
                yield top.child(j)

    def _set_all(self, state):
        for item in self._iter_code_items():
            item.setCheckState(0, state)

    def _on_confirm(self):
        self.result_pairs.clear()
        for item in self._iter_code_items():
            if item.checkState(0) == Qt.CheckState.Checked:
                for pair in (item.data(0, Qt.ItemDataRole.UserRole) or []):
                    self.result_pairs.add(tuple(pair))
        print(f"User confirmed deletion for {len(self.result_pairs)} codes (actual var/code pairs).", flush=True)
        self.accept()


# =========================================================================
#  Dialog: กำหนดตัวแปร Loop (หน้าต่างที่ 1)
# =========================================================================
class LoopDefinitionDialog(QtWidgets.QDialog):
    COL_NAME, COL_TYPE, COL_GROUP = 0, 1, 2

    def __init__(self, parent, app, items_for_treeview, iid_to_actual_vars):
        super().__init__(parent)
        self.app = app
        self.items = items_for_treeview          # [(display_text, iid, initial_value, is_group, members)]
        self.iid_to_actual_vars = iid_to_actual_vars
        self.setWindowTitle("กำหนดตัวแปร Loop (SA / MA / Text / Numeric)")
        self.setModal(True)
        self.resize(900, 660)

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(8)

        hint1 = QtWidgets.QLabel(
            "เลือกแถว (Shift/Ctrl) แล้วกดปุ่มด้านล่าง  หรือคลิกที่ช่อง 'Loop Type' เพื่อสลับ "
            "(ว่าง → SA → MA → Loop Text → Loop Numeric)")
        hint2 = QtWidgets.QLabel(
            "Manual Group: เลือกช่วงตัวแปร → กด 'รวมเป็น Loop เดียว' → ติ๊กเลือก Loop Type "
            "(โปรแกรมจะไม่ใช้กฎอัตโนมัติกับกลุ่มนี้)")
        hint2.setObjectName("HintLabel")
        for lbl in (hint1, hint2):
            lbl.setWordWrap(True)
            layout.addWidget(lbl)

        self.table = QtWidgets.QTableWidget(len(self.items), 3)
        self.table.setHorizontalHeaderLabels(["Variable Name / Group", "Loop Type", "Manual Group"])
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setDefaultSectionSize(24)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(1, 150)
        self.table.setColumnWidth(2, 200)

        for row, (display_text, iid, initial_value, _is_group, _members) in enumerate(self.items):
            name_item = QtWidgets.QTableWidgetItem(display_text)
            name_item.setData(Qt.ItemDataRole.UserRole, iid)
            type_item = QtWidgets.QTableWidgetItem(initial_value)
            type_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            group_item = QtWidgets.QTableWidgetItem("")
            self.table.setItem(row, self.COL_NAME, name_item)
            self.table.setItem(row, self.COL_TYPE, type_item)
            self.table.setItem(row, self.COL_GROUP, group_item)

        self.table.cellClicked.connect(self._on_cell_clicked)
        layout.addWidget(self.table, 1)

        manual_row = QtWidgets.QHBoxLayout()
        btn_make = QtWidgets.QPushButton("รวมเป็น Loop เดียว (Manual Group)")
        btn_make.setProperty("accent", "cyan")
        btn_clear_mg = QtWidgets.QPushButton("ยกเลิก Manual Group")
        btn_clear_mg.setProperty("accent", "red")
        btn_make.clicked.connect(self._make_manual_group)
        btn_clear_mg.clicked.connect(self._clear_manual_group)
        manual_row.addStretch(1)
        manual_row.addWidget(btn_make)
        manual_row.addWidget(btn_clear_mg)
        manual_row.addStretch(1)
        layout.addLayout(manual_row)

        set_row = QtWidgets.QHBoxLayout()
        for text, value in (("Set Selected SA", "SA"), ("Set Selected MA", "MA"),
                            ("Set Selected Loop Text", "Loop Text"),
                            ("Set Selected Loop Numeric", "Loop Numeric"),
                            ("Clear Selected", "")):
            btn = QtWidgets.QPushButton(text)
            btn.clicked.connect(lambda _c=False, v=value: self._apply_to_selected(v))
            set_row.addWidget(btn)
        layout.addLayout(set_row)

        btn_box = QtWidgets.QDialogButtonBox()
        ok_btn = btn_box.addButton("ตกลง", QtWidgets.QDialogButtonBox.ButtonRole.AcceptRole)
        ok_btn.setProperty("accent", "blue")
        btn_box.addButton("ยกเลิก", QtWidgets.QDialogButtonBox.ButtonRole.RejectRole)
        btn_box.accepted.connect(self._on_confirm)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

        self.refresh_rows()

    # ---------- helpers ----------
    def _row_key_var(self, row):
        iid = self.table.item(row, self.COL_NAME).data(Qt.ItemDataRole.UserRole)
        members = self.iid_to_actual_vars.get(iid, [])
        return members[0] if members else iid

    def _selected_rows(self):
        return sorted({idx.row() for idx in self.table.selectedIndexes()})

    def refresh_rows(self):
        """ซิงค์คอลัมน์ Loop Type + Manual Group ของทุกแถวให้ตรงกับข้อมูลจริง"""
        manual_brush = QtGui.QBrush(QtGui.QColor("#E6F2FF"))
        plain_brush = QtGui.QBrush(Qt.GlobalColor.transparent)
        for row in range(self.table.rowCount()):
            key_var = self._row_key_var(row)
            cur_type = self.app.variable_loop_types.get(key_var, "")
            mg = self.app.get_manual_group_of_var(key_var)
            self.table.item(row, self.COL_TYPE).setText(cur_type)
            self.table.item(row, self.COL_GROUP).setText(mg['name'] if mg else "")
            brush = manual_brush if mg else plain_brush
            for col in range(3):
                self.table.item(row, col).setBackground(brush)

    # ---------- events ----------
    def _on_cell_clicked(self, row, column):
        if column != self.COL_TYPE:
            return
        current_type = self.table.item(row, self.COL_TYPE).text()
        cycle = {"": "SA", "SA": "MA", "MA": "Loop Text",
                 "Loop Text": "Loop Numeric", "Loop Numeric": ""}
        next_type = cycle.get(current_type, "SA")

        iid = self.table.item(row, self.COL_NAME).data(Qt.ItemDataRole.UserRole)
        actual_vars = self.iid_to_actual_vars.get(iid, [])
        print(f"  Click Update: IID='{iid}', NextType='{next_type}', ActualVars={len(actual_vars)}", flush=True)
        for actual_var in actual_vars:
            if next_type:
                self.app.variable_loop_types[actual_var] = next_type
            elif actual_var in self.app.variable_loop_types:
                del self.app.variable_loop_types[actual_var]
        mg = self.app.get_manual_group_of_var(actual_vars[0]) if actual_vars else None
        if mg and next_type:
            mg['type'] = next_type
        self.refresh_rows()

    def _apply_to_selected(self, loop_type_to_set):
        rows = self._selected_rows()
        if not rows:
            QtWidgets.QMessageBox.warning(self, "ไม่มีรายการที่เลือก", "กรุณาเลือกตัวแปร/กลุ่ม ในตารางก่อน")
            return
        print(f"Applying '{loop_type_to_set}' to {len(rows)} selected items/groups...", flush=True)
        updated = 0
        for row in rows:
            iid = self.table.item(row, self.COL_NAME).data(Qt.ItemDataRole.UserRole)
            actual_vars = self.iid_to_actual_vars.get(iid, [])
            for actual_var in actual_vars:
                if loop_type_to_set:
                    self.app.variable_loop_types[actual_var] = loop_type_to_set
                elif actual_var in self.app.variable_loop_types:
                    del self.app.variable_loop_types[actual_var]
                updated += 1
            mg = self.app.get_manual_group_of_var(actual_vars[0]) if actual_vars else None
            if mg and loop_type_to_set:
                mg['type'] = loop_type_to_set
        print(f"  Applied to {len(rows)} rows, affecting {updated} actual variables.", flush=True)
        self.refresh_rows()

    def _make_manual_group(self):
        rows = self._selected_rows()
        if not rows:
            QtWidgets.QMessageBox.warning(self, "ไม่มีรายการที่เลือก",
                                          "กรุณาเลือกตัวแปร/กลุ่ม ที่ต้องการรวมเป็น Loop เดียวก่อน")
            return

        subs, all_group_vars = [], []
        for row in rows:   # rows เรียงตามลำดับที่แสดง = ลำดับในไฟล์ SPSS อยู่แล้ว
            iid = self.table.item(row, self.COL_NAME).data(Qt.ItemDataRole.UserRole)
            actual_vars = self.iid_to_actual_vars.get(iid, [])
            if not actual_vars:
                continue
            subs.append(actual_vars[0])
            all_group_vars.extend(actual_vars)

        if len(subs) < 2:
            QtWidgets.QMessageBox.warning(self, "เลือกไม่พอ", "Manual Group ต้องเลือกอย่างน้อย 2 แถว")
            return

        # ให้ผู้ใช้ติ๊กเลือก Loop Type ของกลุ่มนี้
        types_found = [self.app.variable_loop_types.get(v, "") for v in subs]
        default_type = next((t for t in types_found if t), "SA")
        group_type = LoopTypeDialog.ask(self, default_type, len(subs))
        if not group_type:
            print("  Manual Group cancelled by user (loop type dialog).", flush=True)
            return

        # เช็คว่าซ้อนทับกับ Manual Group เดิมหรือไม่
        overlapped = set()
        for v in all_group_vars:
            mg = self.app.get_manual_group_of_var(v)
            if mg:
                overlapped.add(mg['name'])
        if overlapped:
            answer = QtWidgets.QMessageBox.question(
                self, "ซ้อนทับกลุ่มเดิม",
                "ตัวแปรที่เลือกอยู่ใน Manual Group เดิมอยู่แล้ว:\n"
                f"{', '.join(sorted(overlapped))}\n\nต้องการลบกลุ่มเดิมแล้วสร้างใหม่หรือไม่?")
            if answer != QtWidgets.QMessageBox.StandardButton.Yes:
                return
            self.app.manual_loop_groups = [g for g in self.app.manual_loop_groups
                                           if g['name'] not in overlapped]

        for v in all_group_vars:
            self.app.variable_loop_types[v] = group_type

        default_name = get_base_name_heuristic(subs[0]) or subs[0]
        existing_names = {g['name'] for g in self.app.manual_loop_groups}
        final_name, n = default_name, 2
        while final_name in existing_names:
            final_name = f"{default_name}_g{n}"
            n += 1

        self.app.manual_loop_groups.append({
            'name': final_name, 'type': group_type,
            'subs': subs, 'vars': all_group_vars})
        print(f"  Manual Group created: '{final_name}' type={group_type}, "
              f"{len(subs)} subs, {len(all_group_vars)} vars", flush=True)
        self.refresh_rows()
        QtWidgets.QMessageBox.information(
            self, "สร้าง Manual Group สำเร็จ",
            f"รวม {len(subs)} รายการเป็น Loop เดียวแล้ว\n"
            f"Loop ID เริ่มต้น: {final_name}\nType: {group_type}\n\n"
            "(แก้ชื่อได้ที่ปุ่ม '2. กำหนดชื่อ Loop ID')")

    def _clear_manual_group(self):
        rows = self._selected_rows()
        if not rows:
            QtWidgets.QMessageBox.warning(self, "ไม่มีรายการที่เลือก",
                                          "กรุณาเลือกแถวที่อยู่ใน Manual Group ที่ต้องการยกเลิก")
            return
        names_to_remove = set()
        for row in rows:
            iid = self.table.item(row, self.COL_NAME).data(Qt.ItemDataRole.UserRole)
            for v in self.iid_to_actual_vars.get(iid, []):
                mg = self.app.get_manual_group_of_var(v)
                if mg:
                    names_to_remove.add(mg['name'])
        if not names_to_remove:
            QtWidgets.QMessageBox.information(self, "ไม่พบ Manual Group", "แถวที่เลือกไม่ได้อยู่ใน Manual Group")
            return
        self.app.manual_loop_groups = [g for g in self.app.manual_loop_groups
                                       if g['name'] not in names_to_remove]
        print(f"  Manual Groups removed: {sorted(names_to_remove)}", flush=True)
        self.refresh_rows()
        QtWidgets.QMessageBox.information(self, "ยกเลิกแล้ว",
                                          f"ยกเลิก Manual Group: {', '.join(sorted(names_to_remove))}")

    def _on_confirm(self):
        print(f"Loop definitions confirmed: {len(self.app.variable_loop_types)} vars defined.", flush=True)
        QtWidgets.QMessageBox.information(
            self, "บันทึกสำเร็จ",
            f"บันทึกการกำหนด Loop แล้ว ({len(self.app.variable_loop_types)} รายการ)")
        self.accept()


# =========================================================================
#  Dialog: กำหนดชื่อ Loop ID (หน้าต่างที่ 2)
# =========================================================================
class LoopNamingDialog(QtWidgets.QDialog):
    COL_REP, COL_TYPE, COL_NAME = 0, 1, 2

    def __init__(self, parent, items_to_display):
        super().__init__(parent)
        self.items_to_display = items_to_display   # [(first_var, loop_info, final_loop_id)]
        self.edited_names = {}
        self.setWindowTitle("กำหนดชื่อ Loop ID (Consolidated)")
        self.setModal(True)
        self.resize(760, 560)

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(10)

        header = QtWidgets.QLabel("แก้ไขชื่อ Loop ID ที่ต้องการในคอลัมน์ 'Loop ID Name (Editable)'")
        header.setStyleSheet("font-size: 11pt; font-weight: 600; color:#2C5282;")
        layout.addWidget(header)

        self.table = QtWidgets.QTableWidget(len(items_to_display), 3)
        self.table.setHorizontalHeaderLabels(["Representative Var", "Loop Type", "Loop ID Name (Editable)"])
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setDefaultSectionSize(26)
        head = self.table.horizontalHeader()
        head.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        head.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.Fixed)
        head.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(1, 150)

        manual_brush = QtGui.QBrush(QtGui.QColor("#E6F2FF"))
        for row, (first_var, loop_info, final_loop_id) in enumerate(items_to_display):
            is_manual = bool(loop_info.get('manual'))
            loop_type = loop_info['type'] + (" (Manual)" if is_manual else "")

            rep_item = QtWidgets.QTableWidgetItem(first_var)
            rep_item.setFlags(rep_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            type_item = QtWidgets.QTableWidgetItem(loop_type)
            type_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            type_item.setFlags(type_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            name_item = QtWidgets.QTableWidgetItem(final_loop_id)

            if is_manual:
                for item in (rep_item, type_item, name_item):
                    item.setBackground(manual_brush)

            self.table.setItem(row, self.COL_REP, rep_item)
            self.table.setItem(row, self.COL_TYPE, type_item)
            self.table.setItem(row, self.COL_NAME, name_item)

        layout.addWidget(self.table, 1)

        btn_box = QtWidgets.QDialogButtonBox()
        save_btn = btn_box.addButton("บันทึก", QtWidgets.QDialogButtonBox.ButtonRole.AcceptRole)
        save_btn.setProperty("accent", "blue")
        btn_box.addButton("ยกเลิก", QtWidgets.QDialogButtonBox.ButtonRole.RejectRole)
        btn_box.accepted.connect(self._on_save)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def _on_save(self):
        self.table.setCurrentCell(-1, -1)   # ปิด editor ที่ค้างอยู่ก่อนอ่านค่า
        self.edited_names = {}
        for row, (first_var, _info, _fid) in enumerate(self.items_to_display):
            name = self.table.item(row, self.COL_NAME).text().strip()
            if name:
                self.edited_names[first_var] = name
        self.accept()


# =========================================================================
#  Main Window
# =========================================================================
class SpssToExcelConverter(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ตัวแปลง SPSS เป็น Itemdef Excel  ·  PyQt6 Edition")
        self.resize(940, 660)

        # --- Data Storage (เหมือนเวอร์ชัน Tkinter) ---
        self.spss_file_path = ""
        self.excel_template_path = ""
        self.excel_output_path = ""
        self.codes_to_delete_confirmed = set()
        self.last_read_meta = None
        self.variable_loop_types = {}
        self.user_defined_loop_names = {}
        # Manual Loop Group: [{'name', 'type', 'subs', 'vars'}]
        self.manual_loop_groups = []
        self.ma_pattern = re.compile(r'(.+)_O\d+$')

        self._build_ui()

        # --- Redirect stdout ---
        self._original_stdout = sys.stdout
        self.redirector = LogStream(self.log_view)
        sys.stdout = self.redirector
        self._print_banner()

        self.check_fields()

    # ------------------------------------------------------------------
    #  UI
    # ------------------------------------------------------------------
    def _build_ui(self):
        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        root_layout = QtWidgets.QVBoxLayout(central)
        root_layout.setContentsMargins(14, 12, 14, 12)
        root_layout.setSpacing(10)

        # ----- กล่องเลือกไฟล์ -----
        files_box = QtWidgets.QGroupBox("ไฟล์ที่ใช้งาน")
        grid = QtWidgets.QGridLayout(files_box)
        grid.setHorizontalSpacing(8)
        grid.setVerticalSpacing(8)

        self.entry_spss = self._add_file_row(
            grid, 0, "เลือกไฟล์ SPSS (.sav) :", "Open SPSS", "green", self.browse_spss_file)
        self.entry_template = self._add_file_row(
            grid, 1, "เลือกไฟล์ Excel Template :", "Open Excel", "green", self.browse_excel_template)
        self.entry_output = self._add_file_row(
            grid, 2, "บันทึกเป็น Excel Itemdef :", "เลือกตำแหน่ง Save", "blue", self.browse_excel_output)
        grid.setColumnStretch(1, 1)
        root_layout.addWidget(files_box)

        # ----- แถวปุ่มควบคุม -----
        control_box = QtWidgets.QGroupBox("ขั้นตอนการทำงาน")
        control_layout = QtWidgets.QGridLayout(control_box)
        control_layout.setHorizontalSpacing(6)
        control_layout.setVerticalSpacing(6)

        self.btn_define_loop = self._make_button("1. กำหนดตัวแปร Loop", "orange", self.show_loop_definition_window)
        self.btn_name_loop = self._make_button("2. กำหนดชื่อ Loop ID", "orange", self.show_loop_naming_window)
        self.btn_save_loop = self._make_button("Save Loop", "grey", self.save_loop_settings)
        self.btn_load_loop = self._make_button("Load Setting Loop", "grey", self.load_loop_settings)
        self.btn_export_rawdata = self._make_button("Export Rawdata Excel", "steel", self.export_rawdata_excel)
        self.btn_convert = self._make_button("3. Run Itemdef", "gold", self.convert_file)

        # จัด 3 คอลัมน์ x 2 แถว เพื่อให้หน้าต่างไม่กว้างเกินไป
        # แถวบน = ขั้นตอนหลัก 1-2-3 / แถวล่าง = เครื่องมือเสริม
        control_grid = ((self.btn_define_loop, 0, 0), (self.btn_name_loop, 0, 1), (self.btn_convert, 0, 2),
                        (self.btn_save_loop, 1, 0), (self.btn_load_loop, 1, 1), (self.btn_export_rawdata, 1, 2))
        for btn, row, col in control_grid:
            btn.setProperty("compact", "true")
            btn.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding,
                              QtWidgets.QSizePolicy.Policy.Fixed)
            control_layout.addWidget(btn, row, col)
        for col in range(3):
            control_layout.setColumnStretch(col, 1)
        self.btn_export_rawdata.setToolTip("Export Rawdata Excel")
        root_layout.addWidget(control_box)

        # ----- Status -----
        self.status_label = QtWidgets.QLabel("")
        self.status_label.setObjectName("StatusLabel")
        root_layout.addWidget(self.status_label)

        # ----- Log -----
        log_box = QtWidgets.QGroupBox("Log")
        log_layout = QtWidgets.QVBoxLayout(log_box)
        log_layout.setContentsMargins(8, 8, 8, 8)
        self.log_view = QtWidgets.QPlainTextEdit()
        self.log_view.setObjectName("LogView")
        self.log_view.setReadOnly(True)
        self.log_view.setLineWrapMode(QtWidgets.QPlainTextEdit.LineWrapMode.NoWrap)
        self.log_view.setMaximumBlockCount(20000)
        log_layout.addWidget(self.log_view)
        root_layout.addWidget(log_box, 1)

    def _add_file_row(self, grid, row, label_text, button_text, accent, slot):
        label = QtWidgets.QLabel(label_text)
        label.setObjectName("FieldLabel")
        label.setMinimumWidth(190)
        entry = QtWidgets.QLineEdit()
        entry.setReadOnly(True)
        entry.setPlaceholderText("ยังไม่ได้เลือกไฟล์")
        button = self._make_button(button_text, accent, slot)
        button.setMinimumWidth(150)
        grid.addWidget(label, row, 0)
        grid.addWidget(entry, row, 1)
        grid.addWidget(button, row, 2)
        return entry

    @staticmethod
    def _make_button(text, accent, slot):
        btn = QtWidgets.QPushButton(text)
        btn.setProperty("accent", accent)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.clicked.connect(slot)
        return btn

    def _print_banner(self):
        self.log_view.appendHtml(
            '<span style="color:#68D391; font-weight:bold;">Application นี้ถูกพัฒนาโดย Team DP</span><br>'
            '<span style="color:#68D391; font-weight:bold;">---- App พร้อมเริ่มต้นทำงาน ----</span><br>')

    def set_status(self, text, color="#2B6CB0"):
        self.status_label.setText(text)
        self.status_label.setStyleSheet(f"color:{color}; font-weight:600; padding:4px 2px;")
        QtWidgets.QApplication.processEvents()

    def check_fields(self):
        spss_selected = bool(self.spss_file_path)
        can_convert = spss_selected and bool(self.excel_template_path) and bool(self.excel_output_path)
        self.btn_define_loop.setEnabled(spss_selected)
        self.btn_name_loop.setEnabled(spss_selected)
        self.btn_save_loop.setEnabled(spss_selected)
        self.btn_load_loop.setEnabled(spss_selected)
        self.btn_export_rawdata.setEnabled(spss_selected)
        self.btn_convert.setEnabled(can_convert)

    def closeEvent(self, event):
        self.cleanup_redirector()
        super().closeEvent(event)

    def cleanup_redirector(self):
        if getattr(self, 'redirector', None) is not None:
            self.redirector.destroy()
        if hasattr(self, '_original_stdout'):
            sys.stdout = self._original_stdout
        else:
            sys.stdout = sys.__stdout__

    # ------------------------------------------------------------------
    #  File pickers
    # ------------------------------------------------------------------
    def browse_spss_file(self):
        filepath, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "เลือกไฟล์ SPSS (.sav)", "", "SPSS files (*.sav);;All files (*.*)")
        if not filepath:
            return
        filepath = os.path.normpath(filepath)
        self.spss_file_path = filepath
        self.entry_spss.setText(filepath)

        # Reset related states when a new file is selected
        self.variable_loop_types.clear()
        self.codes_to_delete_confirmed = set()
        self.user_defined_loop_names.clear()
        self.manual_loop_groups.clear()
        self.last_read_meta = None

        # --- Set Default Output Path ---
        dir_name = os.path.dirname(filepath)
        stem = os.path.splitext(os.path.basename(filepath))[0]
        final_excel_path = os.path.join(dir_name, f"{stem}_Itemdef.xlsx")
        counter = 1
        while os.path.exists(final_excel_path):
            final_excel_path = os.path.join(dir_name, f"{stem}_output_{counter}.xlsx")
            counter += 1
        self.excel_output_path = final_excel_path
        self.entry_output.setText(final_excel_path)

        self.set_status("กำลังอ่าน Metadata และตรวจสอบ Code...", "#DD6B20")

        try:
            print(f"กำลังอ่านไฟล์จาก: {filepath}", flush=True)
            meta = None
            try:
                print("กำลังเข้ารหัสด้วย UTF-8 encoding...", flush=True)
                _df, meta = pyreadstat.read_sav(filepath, metadataonly=True, encoding='utf-8')
                print("เข้ารหัสและอ่านไฟล์สำเร็จด้วย: UTF-8.", flush=True)
            except pyreadstat.ReadstatError as e_utf8:
                print(f"UTF-8 ไม่สำเร็จ: {e_utf8}. กำลังลองด้วย: cp874...", flush=True)
                _df, meta = pyreadstat.read_sav(filepath, metadataonly=True, encoding='cp874')
                print("เข้ารหัสและอ่านไฟล์สำเร็จด้วย: cp874.", flush=True)

            if meta is None:
                raise ValueError("Metadata could not be read with attempted encodings.")
            self.last_read_meta = meta

            if hasattr(meta, 'variable_value_labels') and isinstance(meta.variable_value_labels, dict):
                initial_problems = self._validate_value_labels(meta.variable_value_labels)
                if initial_problems:
                    print(f"Found {len(initial_problems)} variables with problematic codes.", flush=True)
                    self.set_status("พบ Code ที่ต้องตรวจสอบ...", "#DD6B20")
                    self._show_validation_window(initial_problems, meta.variable_value_labels, meta)
                    self.set_status(
                        f"ตรวจสอบ Code เสร็จสิ้น ({len(self.codes_to_delete_confirmed)} รายการถูกเลือกที่จะลบ)")
                else:
                    print("No problematic codes (0 or >2000) found.", flush=True)
                    self.set_status("ไม่พบ Code ที่ต้องตรวจสอบ")
            else:
                print("No 'variable_value_labels' found in metadata.", flush=True)
                self.set_status("ไม่พบ Value Labels ในไฟล์")

        except FileNotFoundError:
            print(f"ERROR: SPSS file not found: {filepath}", flush=True)
            QtWidgets.QMessageBox.critical(self, "ข้อผิดพลาด", f"ไม่พบไฟล์ SPSS:\n{filepath}")
            self.set_status("เกิดข้อผิดพลาด: ไม่พบไฟล์ SPSS", "#C53030")
            self.spss_file_path = ""; self.entry_spss.clear(); self.last_read_meta = None
        except pyreadstat.ReadstatError as e_read:
            print(f"ERROR reading meta (tried UTF-8 and cp874): {e_read}", flush=True)
            QtWidgets.QMessageBox.critical(
                self, "ข้อผิดพลาดอ่าน SPSS",
                f"ไม่สามารถอ่าน Metadata จากไฟล์ SPSS ได้ (ลองทั้ง UTF-8 และ cp874):\n{e_read}")
            self.set_status("เกิดข้อผิดพลาดอ่าน SPSS", "#C53030")
            self.spss_file_path = ""; self.entry_spss.clear(); self.last_read_meta = None
        except Exception as e:
            print(f"ERROR during initial validation or metadata read: {e}", flush=True)
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "ข้อผิดพลาด", f"เกิดข้อผิดพลาดที่ไม่คาดคิด:\n{e}")
            self.set_status("เกิดข้อผิดพลาดไม่คาดคิด", "#C53030")
            self.spss_file_path = ""; self.entry_spss.clear(); self.last_read_meta = None
        finally:
            self.check_fields()

    def browse_excel_template(self):
        filepath, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "เลือกไฟล์ Excel Template (.xlsx)", "", "Excel files (*.xlsx *.xls);;All files (*.*)")
        if filepath:
            self.excel_template_path = os.path.normpath(filepath)
            self.entry_template.setText(self.excel_template_path)
            self.check_fields()

    def browse_excel_output(self):
        initial = self.excel_output_path
        if not initial and self.spss_file_path:
            stem = os.path.splitext(os.path.basename(self.spss_file_path))[0]
            initial = os.path.join(os.path.dirname(self.spss_file_path), f"{stem}_Itemdef.xlsx")
        filepath, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "บันทึกไฟล์ Excel Output (.xlsx)", initial, "Excel files (*.xlsx);;All files (*.*)")
        if filepath:
            if not filepath.lower().endswith(".xlsx"):
                filepath += ".xlsx"
            self.excel_output_path = os.path.normpath(filepath)
            self.entry_output.setText(self.excel_output_path)
            self.check_fields()

    # ------------------------------------------------------------------
    #  Value label validation
    # ------------------------------------------------------------------
    def _validate_value_labels(self, value_labels_dict):
        """Identifies variables with value codes that are 0 or > 2000."""
        problematic_codes = {}
        if not isinstance(value_labels_dict, dict):
            return problematic_codes
        for var_name, labels in value_labels_dict.items():
            if not isinstance(labels, dict):
                continue
            found_problems = []
            for code in labels.keys():
                try:
                    numeric_code = float(code)
                    if numeric_code == 0 or numeric_code > 2000:
                        found_problems.append(str(code))
                except (ValueError, TypeError):
                    continue
            if found_problems:
                problematic_codes[var_name] = sorted(found_problems, key=lambda x: float(x))
        return problematic_codes

    def _show_validation_window(self, codes_to_validate, all_value_labels, meta):
        """จัดกลุ่ม _O<n> แล้วเปิด dialog ให้ผู้ใช้เลือก Code ที่จะลบ"""
        if not meta or not hasattr(meta, 'column_names') or not meta.column_names:
            all_var_names = list(codes_to_validate.keys())
            display_items = {var: [var] for var in all_var_names}
            ordered_display_names = sorted(all_var_names)
            print("Warning: Metadata missing for grouping in validation window.", flush=True)
        else:
            all_var_names = meta.column_names
            print(f"Grouping {len(all_var_names)} variables for validation display...", flush=True)
            ma_pattern = re.compile(r'(.+)_O(\d+)$')
            groups = defaultdict(list)
            variable_order_keys = []
            processed = set()
            display_items = {}

            for i, var_name in enumerate(all_var_names):
                if i in processed:
                    continue
                match = ma_pattern.match(var_name)
                if match:
                    base_name = match.group(1)
                    current_group_vars = []
                    for j in range(i, len(all_var_names)):
                        if j in processed:
                            continue
                        inner = ma_pattern.match(all_var_names[j])
                        if inner and inner.group(1) == base_name:
                            current_group_vars.append(all_var_names[j])
                            processed.add(j)
                    if current_group_vars:
                        rep_name = f"{base_name}_O1"
                        if rep_name not in current_group_vars:
                            rep_name = sorted(current_group_vars)[0]
                        if base_name not in variable_order_keys:
                            variable_order_keys.append(base_name)
                        groups[base_name].extend(current_group_vars)
                        display_items[rep_name] = sorted(set(current_group_vars))
                else:
                    if var_name not in variable_order_keys:
                        variable_order_keys.append(var_name)
                    display_items[var_name] = [var_name]
                    processed.add(i)

            ordered_display_names = []
            processed_display = set()
            for key in variable_order_keys:
                if key in groups:
                    actual_rep = None
                    for dn in display_items:
                        dn_match = ma_pattern.match(dn)
                        if dn_match and dn_match.group(1) == key:
                            actual_rep = dn
                            break
                    if actual_rep and actual_rep not in processed_display:
                        ordered_display_names.append(actual_rep)
                        processed_display.add(actual_rep)
                elif key in display_items and key not in processed_display:
                    ordered_display_names.append(key)
                    processed_display.add(key)
            print(f"Displaying {len(ordered_display_names)} items/groups in validation window.", flush=True)

        dlg = CodeValidationDialog(self, ordered_display_names, display_items,
                                   codes_to_validate, all_value_labels)
        if dlg.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            self.codes_to_delete_confirmed = set(dlg.result_pairs)

    # ------------------------------------------------------------------
    #  Manual Loop Group helpers
    # ------------------------------------------------------------------
    def get_manual_group_of_var(self, var_name):
        """คืน dict ของ Manual Group ที่ตัวแปรนี้สังกัดอยู่ (ถ้ามี) ไม่งั้นคืน None"""
        if not var_name:
            return None
        for group in self.manual_loop_groups:
            if var_name in group['vars']:
                return group
        return None

    def get_manual_group_var_set(self):
        """คืน set ของตัวแปรทั้งหมดที่ถูกจองไว้โดย Manual Group"""
        reserved = set()
        for group in self.manual_loop_groups:
            reserved.update(group['vars'])
        return reserved

    def build_manual_group_entries(self, var_name_to_index):
        """
        เรียง subs/vars ตามลำดับในไฟล์ SPSS และตัดตัวแปรที่ไม่มีในไฟล์ปัจจุบันออก
        คืนค่า list ของ (first_var, group_dict, ordered_subs, ordered_vars)
        """
        entries = []
        for group in self.manual_loop_groups:
            ordered_vars = sorted([v for v in group['vars'] if v in var_name_to_index],
                                  key=lambda v: var_name_to_index[v])
            ordered_subs = sorted([v for v in group['subs'] if v in var_name_to_index],
                                  key=lambda v: var_name_to_index[v])
            if not ordered_vars or not ordered_subs:
                print(f"  Warning: Manual Group '{group['name']}' ไม่พบตัวแปรในไฟล์ปัจจุบัน ข้ามไป", flush=True)
                continue
            entries.append((ordered_vars[0], group, ordered_subs, ordered_vars))
        entries.sort(key=lambda e: var_name_to_index[e[0]])
        return entries

    # ------------------------------------------------------------------
    #  หน้าต่างที่ 1: กำหนดตัวแปร Loop
    # ------------------------------------------------------------------
    def _ensure_meta(self):
        """อ่าน metadata ถ้ายังไม่มี คืน meta หรือ None พร้อมแจ้ง error แล้ว"""
        if self.last_read_meta is not None:
            return self.last_read_meta
        try:
            print("Reading SPSS metadata...", flush=True)
            try:
                _df, meta = pyreadstat.read_sav(self.spss_file_path, metadataonly=True, encoding='utf-8')
            except pyreadstat.ReadstatError:
                _df, meta = pyreadstat.read_sav(self.spss_file_path, metadataonly=True, encoding='cp874')
            self.last_read_meta = meta
            return meta
        except Exception as e:
            print(f"ERROR reading meta: {e}", flush=True)
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "ข้อผิดพลาดอ่าน SPSS", f"อ่าน Metadata ไม่ได้:\n{e}")
            self.set_status("เกิดข้อผิดพลาดอ่าน SPSS", "#C53030")
            return None

    def _build_display_items(self, all_var_names):
        """
        จัดกลุ่ม _O<n> เป็นแถวเดียว (แทนด้วย <base>_O1)
        คืน (items_for_treeview, iid_to_actual_vars)
        """
        ma_pattern = re.compile(r'(.+)_O(\d+)$')
        groups = defaultdict(list)
        variable_order = []
        processed = set()

        for i, var_name in enumerate(all_var_names):
            if i in processed:
                continue
            match = ma_pattern.match(var_name)
            if match:
                base_name = match.group(1)
                current_group_vars = []
                for j in range(i, len(all_var_names)):
                    if j in processed:
                        continue
                    inner = ma_pattern.match(all_var_names[j])
                    if inner and inner.group(1) == base_name:
                        current_group_vars.append(all_var_names[j])
                        processed.add(j)
                if current_group_vars:
                    groups[base_name].extend(current_group_vars)
                    variable_order.append(base_name)
            else:
                variable_order.append(var_name)
                processed.add(i)

        items_for_treeview = []
        iid_to_actual_vars = {}
        processed_items = set()

        for item_key in variable_order:
            if item_key in processed_items:
                continue
            if item_key in groups:
                base_name = item_key
                group_members = sorted(groups[base_name])
                representative_iid = f"{base_name}_O1"
                initial_value = self.variable_loop_types.get(representative_iid, "")
                if not initial_value and group_members:
                    initial_value = self.variable_loop_types.get(group_members[0], "")
                items_for_treeview.append(
                    (representative_iid, representative_iid, initial_value, True, group_members))
                iid_to_actual_vars[representative_iid] = group_members
                processed_items.add(base_name)
            else:
                var_name = item_key
                initial_value = self.variable_loop_types.get(var_name, "")
                items_for_treeview.append((var_name, var_name, initial_value, False, [var_name]))
                iid_to_actual_vars[var_name] = [var_name]
                processed_items.add(var_name)

        return items_for_treeview, iid_to_actual_vars

    def show_loop_definition_window(self):
        print("--- Opening Loop Definition Window ---", flush=True)
        if not self.spss_file_path:
            QtWidgets.QMessageBox.warning(self, "คำเตือน", "กรุณาเลือกไฟล์ SPSS ก่อน")
            return
        self.set_status("กำลังอ่านตัวแปร...", "#DD6B20")
        meta = self._ensure_meta()
        if meta is None:
            return
        if not hasattr(meta, 'column_names') or not meta.column_names:
            QtWidgets.QMessageBox.critical(self, "ข้อผิดพลาด", "ไม่สามารถอ่านรายชื่อตัวแปรจากไฟล์ SPSS ได้")
            self.set_status("เกิดข้อผิดพลาด: ไม่พบชื่อคอลัมน์", "#C53030")
            return

        try:
            all_var_names = meta.column_names
            print(f"Found {len(all_var_names)} variables. Grouping display for _O<n>...", flush=True)
            items, iid_map = self._build_display_items(all_var_names)

            self.set_status("กรุณากำหนดตัวแปร Loop")
            dlg = LoopDefinitionDialog(self, self, items, iid_map)
            if dlg.exec() == QtWidgets.QDialog.DialogCode.Accepted:
                self.set_status("กำหนด Loop เสร็จสิ้น")
            else:
                self.set_status("ปิดหน้าต่างกำหนด Loop")
        except Exception as e:
            print(f"ERROR in Loop Window: {e}", flush=True)
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "ข้อผิดพลาด", f"เกิดข้อผิดพลาดในหน้าต่างกำหนด Loop:\n{e}")
            self.set_status("เกิดข้อผิดพลาดเปิดหน้าต่าง", "#C53030")

    # ------------------------------------------------------------------
    #  หน้าต่างที่ 2: กำหนดชื่อ Loop ID
    # ------------------------------------------------------------------
    def show_loop_naming_window(self):
        print("--- Opening Loop Naming Window ---", flush=True)
        if not self.spss_file_path:
            QtWidgets.QMessageBox.warning(self, "คำเตือน", "กรุณาเลือกไฟล์ SPSS ก่อน")
            return
        if not self.variable_loop_types:
            QtWidgets.QMessageBox.information(self, "ข้อมูล", "ยังไม่มีการกำหนดตัวแปร Loop (SA/MA/Text/Numeric)")
            return

        meta = self._ensure_meta()
        if meta is None:
            return
        if not hasattr(meta, 'column_names'):
            QtWidgets.QMessageBox.critical(self, "ข้อผิดพลาด", "ไม่พบชื่อคอลัมน์ใน Metadata")
            return

        all_vars = meta.column_names
        var_name_to_index = {name: i for i, name in enumerate(all_vars)}

        identified_loops = {}
        temp_processed_indices = set()

        # --- Manual Group มาก่อน แล้วจองตัวแปรไม่ให้ Heuristic แตะ ---
        manual_first_vars = set()
        for first_var, group, ordered_subs, ordered_vars in self.build_manual_group_entries(var_name_to_index):
            identified_loops[first_var] = {
                'type': group['type'],
                'default_name': group['name'],
                'vars': ordered_vars,
                'manual': True,
                'subs': ordered_subs,
            }
            manual_first_vars.add(first_var)
            for v in ordered_vars:
                temp_processed_indices.add(var_name_to_index[v])
            print(f"  Manual Group used: '{group['name']}' ({group['type']}) "
                  f"subs={len(ordered_subs)} vars={len(ordered_vars)}", flush=True)

        i = 0
        print("Identifying initial loop groups with heuristic...", flush=True)
        while i < len(all_vars):
            if i in temp_processed_indices:
                i += 1
                continue
            current_var = all_vars[i]
            current_type = self.variable_loop_types.get(current_var, "")
            if current_type in LOOP_TYPES:
                current_group_vars = [current_var]
                current_base = get_base_name_heuristic(current_var)
                m = re.match(r'(I_\d+_)', current_var)
                current_prefix = m.group(1) if m else None
                j = i + 1
                while j < len(all_vars):
                    if j in temp_processed_indices:
                        break   # ห้ามกินตัวแปรของ Manual Group
                    next_var = all_vars[j]
                    next_type = self.variable_loop_types.get(next_var, "")
                    next_base = get_base_name_heuristic(next_var)
                    stop_grouping = False
                    if next_type != current_type or next_base != current_base:
                        stop_grouping = True
                    elif current_prefix is not None or re.match(r'I_\d+_(.+?)(_O\d+)?$', current_var):
                        nm = re.match(r'(I_\d+_)', next_var)
                        next_prefix = nm.group(1) if nm else None
                        if current_prefix is not None and next_prefix is not None and current_prefix != next_prefix:
                            stop_grouping = True
                    if not stop_grouping:
                        current_group_vars.append(next_var)
                        j += 1
                    else:
                        break

                first_var_name = current_group_vars[0]
                m_base_i = re.match(r'^I_(\d+)_([A-Za-z]\w+)$', current_base)
                default_loop_name = f"{m_base_i.group(2)}_{m_base_i.group(1)}" if m_base_i else current_base
                identified_loops[first_var_name] = {
                    'type': current_type, 'default_name': default_loop_name, 'vars': current_group_vars}
                for k in range(i, j):
                    temp_processed_indices.add(k)
                i = j
                continue
            i += 1
        print(f"Finished initial identification. Found {len(identified_loops)} potential groups.", flush=True)

        if not identified_loops:
            QtWidgets.QMessageBox.information(self, "ข้อมูล", "ไม่พบกลุ่ม Loop ที่กำหนดไว้")
            return

        # --- รวมกลุ่มเพื่อการแสดงผล ---
        items_to_display = []
        sorted_identified_loops = sorted(
            identified_loops.items(), key=lambda item: var_name_to_index.get(item[0], float('inf')))
        processed_display_ids = set()
        for first_var_sorted, loop_info_sorted in sorted_identified_loops:
            final_loop_id = self.user_defined_loop_names.get(first_var_sorted, loop_info_sorted['default_name'])
            # Manual Group ต้องได้แถวของตัวเองเสมอ ไม่ยุบรวมกับกลุ่มอัตโนมัติ
            if loop_info_sorted.get('manual') or final_loop_id not in processed_display_ids:
                items_to_display.append((first_var_sorted, loop_info_sorted, final_loop_id))
                processed_display_ids.add(final_loop_id)
        print(f"Finished consolidating display. {len(items_to_display)} unique rows.", flush=True)

        self.set_status("กรุณากำหนดชื่อ Loop ID")
        dlg = LoopNamingDialog(self, items_to_display)
        if dlg.exec() != QtWidgets.QDialog.DialogCode.Accepted:
            self.set_status("ยกเลิกการกำหนดชื่อ Loop ID")
            return

        temp_edited_names = dlg.edited_names
        final_user_names = {}
        processed_original_first_vars = set()

        # --- Manual Group จับคู่ชื่อแบบตรงตัว ---
        for mf_var in manual_first_vars:
            if mf_var not in identified_loops:
                continue
            new_name = temp_edited_names.get(mf_var, identified_loops[mf_var]['default_name'])
            if new_name:
                final_user_names[mf_var] = new_name
                for grp in self.manual_loop_groups:
                    if mf_var in grp['vars']:
                        if grp['name'] != new_name:
                            print(f"  Manual Group renamed: '{grp['name']}' -> '{new_name}'", flush=True)
                        grp['name'] = new_name
                        break
            processed_original_first_vars.add(mf_var)

        # --- กลุ่มอัตโนมัติ ---
        for original_first_var, original_loop_info in identified_loops.items():
            if original_first_var in processed_original_first_vars:
                continue
            current_final_id = self.user_defined_loop_names.get(
                original_first_var, original_loop_info['default_name'])
            representative = None
            for rep_var_disp, rep_info_disp, disp_final_id in items_to_display:
                if rep_info_disp.get('manual'):
                    continue
                if disp_final_id == current_final_id:
                    representative = rep_var_disp
                    break
            final_name_to_set = current_final_id
            if representative and representative in temp_edited_names:
                final_name_to_set = temp_edited_names[representative]
            for ov, oi in identified_loops.items():
                if ov in manual_first_vars:
                    continue
                check_final_id = self.user_defined_loop_names.get(ov, oi['default_name'])
                if check_final_id == current_final_id:
                    if final_name_to_set:
                        final_user_names[ov] = final_name_to_set
                    processed_original_first_vars.add(ov)

        self.user_defined_loop_names = final_user_names
        print(f"Final loop names saved: {len(self.user_defined_loop_names)} mappings.", flush=True)
        QtWidgets.QMessageBox.information(
            self, "บันทึกสำเร็จ", f"บันทึกชื่อ Loop ID แล้ว ({len(self.user_defined_loop_names)} รายการ)")
        self.set_status("กำหนดชื่อ Loop ID เสร็จสิ้น")

    # ------------------------------------------------------------------
    #  Save / Load Loop Settings
    # ------------------------------------------------------------------
    def save_loop_settings(self):
        print("--- Saving Loop Settings (Display Structure) ---", flush=True)
        if not self.spss_file_path:
            QtWidgets.QMessageBox.warning(self, "คำเตือน", "กรุณาเลือกไฟล์ SPSS ก่อน")
            return
        meta = self.last_read_meta
        if meta is None or not hasattr(meta, 'column_names') or not meta.column_names:
            QtWidgets.QMessageBox.critical(
                self, "ข้อผิดพลาด",
                "ไม่สามารถอ่าน Metadata หรือรายชื่อตัวแปรจากไฟล์ SPSS ได้\nกรุณาลองเลือกไฟล์ SPSS ใหม่อีกครั้ง")
            return

        items, iid_map = self._build_display_items(meta.column_names)
        data_to_save = []
        for display_text, iid, _iv, _is_group, members in items:
            key_var = members[0] if members else iid
            current_loop_type = self.variable_loop_types.get(key_var, "")
            mg = self.get_manual_group_of_var(key_var)
            data_to_save.append((display_text, current_loop_type, mg['name'] if mg else ""))

        if not data_to_save:
            QtWidgets.QMessageBox.information(self, "ข้อมูล", "ไม่พบรายการตัวแปรที่จะบันทึก")
            return
        print(f"Prepared {len(data_to_save)} items for saving.", flush=True)

        stem = os.path.splitext(os.path.basename(self.spss_file_path))[0]
        suggested = os.path.join(os.path.dirname(self.spss_file_path), f"{stem}_loop_definitions.xlsx")
        save_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "บันทึกโครงสร้าง Loop ที่แสดง (.xlsx)", suggested, "Excel files (*.xlsx);;All files (*.*)")
        if not save_path:
            print("Save loop definitions cancelled.", flush=True)
            return
        if not save_path.lower().endswith(".xlsx"):
            save_path += ".xlsx"

        self.set_status("กำลังบันทึกโครงสร้าง Loop...", "#DD6B20")
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Loop Definitions"
            ws['A1'] = "Variable Name / Group"
            ws['B1'] = "Loop Type"
            ws['C1'] = "Manual Group"
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 25

            for row_num, (display_name, loop_type, manual_group_name) in enumerate(data_to_save, start=2):
                ws.cell(row=row_num, column=1).value = display_name
                ws.cell(row=row_num, column=2).value = loop_type
                ws.cell(row=row_num, column=3).value = manual_group_name or None

            dv = DataValidation(type="list", formula1='"SA,MA,Loop Text,Loop Numeric,"', allow_blank=True)
            dv.error = 'ค่าที่เลือกไม่ถูกต้อง'
            dv.errorTitle = 'Loop Type ไม่ถูกต้อง'
            dv.prompt = 'กรุณาเลือกประเภท Loop'
            dv.promptTitle = 'เลือกประเภท'
            ws.add_data_validation(dv)
            if ws.max_row >= 2:
                dv.add(f"B2:B{ws.max_row}")

            wb.save(save_path)
            print(f"Loop definitions saved to: {save_path}", flush=True)
            self.set_status(f"บันทึกโครงสร้าง Loop สำเร็จ: {os.path.basename(save_path)}", "#2F855A")
            QtWidgets.QMessageBox.information(
                self, "สำเร็จ", f"บันทึกโครงสร้าง Loop ที่แสดง สำเร็จ!\nไฟล์: {save_path}")
        except PermissionError:
            print(f"ERROR: Permission denied saving to {save_path}", flush=True)
            QtWidgets.QMessageBox.critical(
                self, "ข้อผิดพลาด",
                f"ไม่สามารถบันทึกไฟล์ได้:\n{save_path}\nอาจจะเปิดไฟล์นี้อยู่ หรือไม่มีสิทธิ์เขียน")
            self.set_status("เกิดข้อผิดพลาด: ไม่มีสิทธิ์บันทึกไฟล์", "#C53030")
        except Exception as e:
            print(f"ERROR saving loop definitions: {e}", flush=True)
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "ข้อผิดพลาด", f"เกิดข้อผิดพลาดในการบันทึก:\n{e}")
            self.set_status("เกิดข้อผิดพลาดในการบันทึก Loop Definitions", "#C53030")

    def load_loop_settings(self):
        print("--- Loading Loop Settings ---", flush=True)
        if not self.spss_file_path:
            QtWidgets.QMessageBox.warning(self, "คำเตือน", "กรุณาเลือกไฟล์ SPSS ก่อน")
            return
        meta = self._ensure_meta()
        if meta is None:
            return
        if not hasattr(meta, 'column_names') or not meta.column_names:
            QtWidgets.QMessageBox.critical(self, "ข้อผิดพลาด", "ไม่พบชื่อตัวแปรในไฟล์ SPSS ปัจจุบัน")
            return

        all_spss_vars_list = meta.column_names
        valid_spss_vars_set = set(all_spss_vars_list)
        print(f"- SPSS variables: {len(valid_spss_vars_set)}", flush=True)

        ma_pat = re.compile(r'^(?P<base>.+)_O(?P<opt>\d+)$')
        ma_index = {}
        for v in all_spss_vars_list:
            m = ma_pat.match(v)
            if m:
                ma_index.setdefault(m.group('base'), []).append(v)
        for base, members in ma_index.items():
            members.sort(key=lambda name: int(ma_pat.match(name).group('opt')))
        print(f"- MA groups indexed: {len(ma_index)}", flush=True)

        load_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "เลือกไฟล์ตั้งค่า Loop (.xlsx)", "", "Excel files (*.xlsx);;All files (*.*)")
        if not load_path:
            print("Load loop settings cancelled.", flush=True)
            return

        self.set_status("กำลังโหลดการตั้งค่า Loop...", "#DD6B20")

        loaded_settings_mapped = {}
        skipped_count = 0
        processed_rows = 0
        valid_loop_types = {"SA", "MA", "Loop Text", "Loop Numeric", ""}
        manual_rows = defaultdict(list)   # {group_name: [(loop_type, sub_rep_var, [members])]}

        try:
            wb = openpyxl.load_workbook(load_path, read_only=True, data_only=True)
            ws = wb.active
            print(f"- Reading Excel rows: {max(ws.max_row - 1, 0)}", flush=True)

            for row_num in range(2, ws.max_row + 1):
                processed_rows += 1
                if processed_rows % 200 == 0:
                    print(f"  ...processed {processed_rows} rows", flush=True)
                    QtWidgets.QApplication.processEvents()

                excel_var_name = ws.cell(row=row_num, column=1).value
                excel_loop_type = ws.cell(row=row_num, column=2).value
                excel_manual_group = ws.cell(row=row_num, column=3).value

                if not isinstance(excel_var_name, str) or not excel_var_name.strip():
                    skipped_count += 1
                    continue
                excel_var_name = excel_var_name.strip()
                excel_loop_type = (excel_loop_type or "").strip()
                excel_manual_group = str(excel_manual_group).strip() if excel_manual_group is not None else ""

                if excel_loop_type not in valid_loop_types:
                    print(f"  Skip row {row_num}: invalid Loop Type '{excel_loop_type}' for '{excel_var_name}'",
                          flush=True)
                    skipped_count += 1
                    continue

                # Case A: MA group representative ("<base>_O1")
                m = ma_pat.match(excel_var_name)
                if m and excel_loop_type == "MA":
                    base = m.group('base')
                    members = ma_index.get(base, [])
                    if not members:
                        print(f"  Warning row {row_num}: MA base '{base}' not found in SPSS, skip.", flush=True)
                        skipped_count += 1
                        continue
                    for spss_var in members:
                        loaded_settings_mapped[spss_var] = excel_loop_type
                    if excel_manual_group:
                        manual_rows[excel_manual_group].append((excel_loop_type, members[0], list(members)))
                    continue

                # Case B: single variable name
                if excel_var_name in valid_spss_vars_set:
                    if excel_loop_type:
                        loaded_settings_mapped[excel_var_name] = excel_loop_type
                    elif excel_var_name in loaded_settings_mapped:
                        del loaded_settings_mapped[excel_var_name]
                    if excel_manual_group and excel_loop_type:
                        manual_rows[excel_manual_group].append((excel_loop_type, excel_var_name, [excel_var_name]))
                else:
                    if excel_loop_type == "MA" and excel_var_name in ma_index:
                        for spss_var in ma_index[excel_var_name]:
                            loaded_settings_mapped[spss_var] = "MA"
                    else:
                        skipped_count += 1

            self.variable_loop_types.clear()
            self.variable_loop_types.update(loaded_settings_mapped)

            # --- สร้าง Manual Group กลับจากคอลัมน์ C ---
            var_index_lookup = {name: n for n, name in enumerate(all_spss_vars_list)}
            rebuilt_groups = []
            for mg_name, rows in manual_rows.items():
                if len(rows) < 2:
                    print(f"  Warning: Manual Group '{mg_name}' มีแค่ {len(rows)} แถว (ต้องมี >= 2) ข้ามไป",
                          flush=True)
                    continue
                rows_sorted = sorted(rows, key=lambda r: var_index_lookup.get(r[1], 10 ** 9))
                group_type = rows_sorted[0][0]
                subs = [r[1] for r in rows_sorted]
                group_vars = []
                for _t, _sub, members in rows_sorted:
                    group_vars.extend(members)
                group_vars = sorted(set(group_vars), key=lambda v: var_index_lookup.get(v, 10 ** 9))
                for v in group_vars:
                    self.variable_loop_types[v] = group_type
                rebuilt_groups.append({'name': mg_name, 'type': group_type, 'subs': subs, 'vars': group_vars})
                print(f"  Manual Group restored: '{mg_name}' ({group_type}) "
                      f"{len(subs)} subs / {len(group_vars)} vars", flush=True)
            self.manual_loop_groups = rebuilt_groups
            self.user_defined_loop_names = {g['vars'][0]: g['name'] for g in rebuilt_groups if g['vars']}

            loaded_count = len(self.variable_loop_types)
            print(f"Done. Set loop type for {loaded_count} variables. "
                  f"Skipped {skipped_count} rows of {processed_rows}.", flush=True)
            self.set_status(f"โหลดการตั้งค่า Loop สำเร็จ ({loaded_count} รายการ)", "#2F855A")
            QtWidgets.QMessageBox.information(
                self, "สำเร็จ",
                f"โหลดการตั้งค่า Loop สำเร็จ!\n"
                f"- กำหนด Loop Type ให้ {loaded_count} ตัวแปร\n"
                f"- Manual Group {len(rebuilt_groups)} กลุ่ม\n"
                f"- ข้าม {skipped_count} แถว (ชื่อไม่พบ/Loop Type ไม่ถูกต้อง)\n\n"
                "คุณสามารถเปิด '1. กำหนดตัวแปร Loop' เพื่อตรวจสอบได้")
        except FileNotFoundError:
            print(f"ERROR: Loop settings file not found: {load_path}", flush=True)
            QtWidgets.QMessageBox.critical(self, "ข้อผิดพลาด", f"ไม่พบไฟล์:\n{load_path}")
            self.set_status("เกิดข้อผิดพลาด: ไม่พบไฟล์ตั้งค่า", "#C53030")
        except Exception as e:
            print(f"ERROR loading loop settings: {e}", flush=True)
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "ข้อผิดพลาด", f"เกิดข้อผิดพลาดในการโหลด:\n{e}")
            self.set_status("เกิดข้อผิดพลาดในการโหลด Loop Settings", "#C53030")

    # ------------------------------------------------------------------
    #  Export Rawdata
    # ------------------------------------------------------------------
    def export_rawdata_excel(self):
        print("--- Starting export_rawdata_excel ---", flush=True)
        if not self.spss_file_path:
            QtWidgets.QMessageBox.warning(self, "คำเตือน", "กรุณาเลือกไฟล์ SPSS ก่อน")
            return

        stem = os.path.splitext(os.path.basename(self.spss_file_path))[0]
        initial_dir = os.path.dirname(self.spss_file_path)
        suggested = os.path.join(initial_dir, f"{stem}_Rawdata.xlsx")
        counter = 1
        while os.path.exists(suggested):
            suggested = os.path.join(initial_dir, f"{stem}_rawdata_{counter}.xlsx")
            counter += 1

        excel_output_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "บันทึก Rawdata เป็น Excel (.xlsx)", suggested, "Excel files (*.xlsx);;All files (*.*)")
        if not excel_output_path:
            print("Export rawdata cancelled.", flush=True)
            self.set_status("ยกเลิกการ Export Rawdata")
            return
        if not excel_output_path.lower().endswith(".xlsx"):
            excel_output_path += ".xlsx"

        self.set_status("กำลัง Export ข้อมูลดิบเป็น Excel...", "#DD6B20")
        try:
            print(f"Reading data with codes from: {self.spss_file_path}", flush=True)
            try:
                print("Attempting UTF-8 encoding for codes...", flush=True)
                df_codes, meta_codes = pyreadstat.read_sav(
                    self.spss_file_path, encoding='utf-8', apply_value_formats=False)
                print("Successfully read using UTF-8 without value labels.", flush=True)
            except pyreadstat.ReadstatError as e_utf8:
                print(f"UTF-8 failed: {e_utf8}. Trying cp874...", flush=True)
                df_codes, meta_codes = pyreadstat.read_sav(
                    self.spss_file_path, encoding='cp874', apply_value_formats=False)
                print("Successfully read using cp874 without value labels.", flush=True)

            if df_codes is None or meta_codes is None:
                raise ValueError("Failed to read data or metadata from SPSS file with codes.")

            columns_to_keep = [col for col in df_codes.columns
                               if (not col.endswith('_O')) or col.endswith('_O1')]
            df_codes = df_codes[columns_to_keep]

            print(f"Exporting {len(df_codes)} rows and {len(df_codes.columns)} columns to Excel...", flush=True)
            with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
                df_codes.to_excel(writer, sheet_name='Rawdata_Code', index=False)

            print(f"Rawdata exported successfully to: {excel_output_path}", flush=True)
            self.set_status(f"Export Rawdata สำเร็จ: {os.path.basename(excel_output_path)}", "#2F855A")
            QtWidgets.QMessageBox.information(
                self, "สำเร็จ", f"Export ข้อมูลดิบเป็น Excel สำเร็จ!\nไฟล์: {excel_output_path}")
        except FileNotFoundError:
            QtWidgets.QMessageBox.critical(self, "ข้อผิดพลาด", f"ไม่พบไฟล์ SPSS:\n{self.spss_file_path}")
            self.set_status("เกิดข้อผิดพลาด: ไม่พบไฟล์ SPSS", "#C53030")
        except PermissionError:
            QtWidgets.QMessageBox.critical(
                self, "ข้อผิดพลาด",
                f"ไม่สามารถบันทึกไฟล์ได้:\n{excel_output_path}\nอาจจะเปิดไฟล์นี้อยู่ หรือไม่มีสิทธิ์เขียน")
            self.set_status("เกิดข้อผิดพลาด: ไม่มีสิทธิ์บันทึกไฟล์", "#C53030")
        except Exception as e:
            print(f"ERROR during export: {e}", flush=True)
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "ข้อผิดพลาด", f"เกิดข้อผิดพลาดในการ Export:\n{e}")
            self.set_status("เกิดข้อผิดพลาดในการ Export Rawdata", "#C53030")
        finally:
            print("--- export_rawdata_excel สำเร็จ ---", flush=True)

    # ------------------------------------------------------------------
    #  Run Itemdef
    # ------------------------------------------------------------------
    def convert_file(self):
        """
        Two-phase conversion:
          PHASE 1 - pre-compute loop structures (Manual Group ก่อน แล้วค่อย Heuristic)
          PHASE 2 - write to the Excel template
        """
        print("--- Starting convert_file ---", flush=True)
        spss_path = self.spss_file_path
        excel_template_path = self.excel_template_path
        excel_output_path = self.excel_output_path

        if not spss_path or not excel_template_path or not excel_output_path:
            QtWidgets.QMessageBox.warning(self, "คำเตือน", "กรุณาเลือกไฟล์ให้ครบถ้วน")
            return
        if not os.path.exists(excel_template_path):
            QtWidgets.QMessageBox.critical(
                self, "ข้อผิดพลาด", f"ไม่พบไฟล์ Excel Template:\n{excel_template_path}")
            self.set_status("เกิดข้อผิดพลาด: ไม่พบ Template", "#C53030")
            return
        if not excel_output_path.lower().endswith(".xlsx"):
            QtWidgets.QMessageBox.critical(self, "ข้อผิดพลาด", "ไฟล์ Output ต้องเป็นนามสกุล .xlsx เท่านั้น")
            self.set_status("เกิดข้อผิดพลาด: Output .xlsx", "#C53030")
            return
        print("Input validation passed.", flush=True)

        self.set_status("กำลังประมวลผล...", "#DD6B20")
        processed_ma_base_names_auto = set()

        try:
            meta = self._ensure_meta()
            if meta is None:
                raise ValueError("ไม่สามารถอ่าน Metadata จากไฟล์ SPSS ได้")

            req_attrs = ['column_names', 'column_labels']
            miss = [a for a in req_attrs if not hasattr(meta, a) or not getattr(meta, a)]
            if miss:
                raise AttributeError(f"Meta ขาด Attribute ที่จำเป็น: {', '.join(miss)}")
            if not hasattr(meta, 'variable_value_labels'):
                meta.variable_value_labels = {}

            # --- กรอง Value Labels ---
            filtered_meta_value_labels = meta.variable_value_labels.copy() if meta.variable_value_labels else {}
            if self.codes_to_delete_confirmed:
                print(f"Filtering {len(self.codes_to_delete_confirmed)} codes...", flush=True)
                temp_filtered = {}
                count = 0
                for var, labels in filtered_meta_value_labels.items():
                    if not isinstance(labels, dict):
                        temp_filtered[var] = labels
                        continue
                    new_labels = {}
                    remove_codes = {cd for v, cd in self.codes_to_delete_confirmed if v == var}
                    for code, label in labels.items():
                        if str(code) not in remove_codes:
                            new_labels[code] = label
                        else:
                            count += 1
                    temp_filtered[var] = new_labels
                filtered_meta_value_labels = temp_filtered
                print(f"{count} codes filtered.", flush=True)

            # ============================================================ #
            #   PHASE 1: PRE-COMPUTATION OF ALL LOOP GROUPS                #
            # ============================================================ #
            print("\n--- PHASE 1: Pre-computing all loop structures ---", flush=True)
            all_vars = meta.column_names
            var_name_to_index = {name: i for i, name in enumerate(all_vars)}

            print("Step 1.1: Identifying initial user-defined loop groups...", flush=True)
            initial_user_loop_groups = {}
            processed_indices = set()

            # --- Manual Group มาก่อน และจองตัวแปรไม่ให้ Heuristic แตะ ---
            manual_loops_prepared = {}
            manual_first_vars_conv = set()
            for first_var, group, ordered_subs, ordered_vars in self.build_manual_group_entries(var_name_to_index):
                g_type = group['type']
                if any(re.search(r'_O\d+$', v) for v in ordered_vars):
                    g_type = "MA"
                manual_loops_prepared[group['name']] = {
                    'type': g_type,
                    'representative_vars': ordered_subs,
                    'label_source_var': ordered_subs[0],
                    'all_original_vars': ordered_vars,
                    'first_representative_var': ordered_vars[0],
                }
                initial_user_loop_groups[first_var] = {
                    'base_name': group['name'], 'type': g_type, 'vars': ordered_vars, 'manual': True}
                manual_first_vars_conv.add(first_var)
                for v in ordered_vars:
                    processed_indices.add(var_name_to_index[v])
                print(f"  Manual Group: '{group['name']}' type={g_type}, "
                      f"{len(ordered_subs)} sub-items, {len(ordered_vars)} vars", flush=True)

            i = 0
            while i < len(all_vars):
                if i in processed_indices:
                    i += 1
                    continue
                current_var = all_vars[i]
                current_type = self.variable_loop_types.get(current_var, "")
                if current_type in LOOP_TYPES:
                    group = [current_var]
                    base = get_base_name_heuristic(current_var)
                    pm = re.match(r'(I_\d+_)', current_var)
                    prefix = pm.group(1) if pm else None
                    j = i + 1
                    while j < len(all_vars):
                        if j in processed_indices:
                            break   # ห้ามกินตัวแปรของ Manual Group
                        next_var = all_vars[j]
                        next_type = self.variable_loop_types.get(next_var, "")
                        next_base = get_base_name_heuristic(next_var)
                        stop = (next_type != current_type) or (next_base != base)
                        if not stop and (prefix or re.match(r'I_\d+_(.+?)(_O\d+)?$', current_var)):
                            nm = re.match(r'(I_\d+_)', next_var)
                            next_prefix = nm.group(1) if nm else None
                            if prefix and next_prefix and prefix != next_prefix:
                                stop = True
                        if not stop:
                            group.append(next_var)
                            j += 1
                        else:
                            break
                    initial_user_loop_groups[group[0]] = {
                        'base_name': base, 'type': current_type, 'vars': group}
                    for k in range(i, j):
                        processed_indices.add(k)
                    i = j
                else:
                    i += 1
            print(f"Found {len(initial_user_loop_groups)} initial user groups.", flush=True)

            # --- 1.2: Consolidate ---
            print("Step 1.2: Consolidating groups with the same final Loop ID...", flush=True)
            consolidated_loops = {}
            groups_by_id = defaultdict(list)
            initial_loops_in_consolidation = set()

            for first_var, info in initial_user_loop_groups.items():
                if info.get('manual'):
                    continue   # Manual Group ไม่ต้องผ่าน logic นี้
                final_id = self.user_defined_loop_names.get(first_var, info['base_name'])
                if not final_id:
                    continue
                groups_by_id[final_id].append({'first_var': first_var, 'info': info})

            i_pat = re.compile(r'^(I_\d+_)(.+?)(_O\d+)?$')
            for final_id, groups_with_same_id in groups_by_id.items():
                if len(groups_with_same_id) <= 1:
                    continue
                can_consolidate = True
                is_o_suffix_present = False
                expected_base = None
                representative_vars_map = {}
                all_vars_in_consol_group = set()
                all_types_in_group = set()
                first_ever_data = None

                for data in groups_with_same_id:
                    if first_ever_data is None:
                        first_ever_data = data
                    first_var = data['first_var']
                    info = data['info']
                    all_types_in_group.add(info['type'])
                    match = i_pat.match(first_var)
                    if not match:
                        can_consolidate = False
                        break
                    prefix = match.group(1)
                    base = match.group(2)
                    o_suf = match.group(3)
                    if expected_base is None:
                        expected_base = base
                    elif base != expected_base:
                        can_consolidate = False
                        break
                    if o_suf:
                        is_o_suffix_present = True
                    rep_var = first_var
                    if o_suf:
                        for v in sorted(info['vars']):
                            if v.startswith(prefix) and v.endswith('_O1'):
                                rep_var = v
                                break
                    if prefix not in representative_vars_map:
                        representative_vars_map[prefix] = rep_var
                        all_vars_in_consol_group.update(info['vars'])
                    else:
                        can_consolidate = False
                        break

                if can_consolidate and representative_vars_map and first_ever_data:
                    if is_o_suffix_present or "MA" in all_types_in_group:
                        final_type = "MA"
                    elif "Loop Text" in all_types_in_group:
                        final_type = "Loop Text"
                    elif "Loop Numeric" in all_types_in_group:
                        final_type = "Loop Numeric"
                    else:
                        final_type = "SA"
                    print(f"  > Consolidating ID '{final_id}' as Type '{final_type}'", flush=True)
                    sorted_reps = sorted(representative_vars_map.values(),
                                         key=lambda v: int(re.match(r'I_(\d+)_', v).group(1)))
                    consolidated_loops[final_id] = {
                        'type': final_type,
                        'representative_vars': sorted_reps,
                        'label_source_var': first_ever_data['first_var'],
                        'all_original_vars': sorted(all_vars_in_consol_group,
                                                    key=lambda x: var_name_to_index.get(x, float('inf'))),
                        'first_representative_var': sorted_reps[0],
                    }
                    for data in groups_with_same_id:
                        initial_loops_in_consolidation.add(data['first_var'])

            # --- ใส่ Manual Group เข้าไปเป็น Loop สำเร็จรูป ---
            for mg_id, mg_data in manual_loops_prepared.items():
                if mg_id in consolidated_loops:
                    print(f"  Warning: Manual Group '{mg_id}' ชื่อซ้ำกับ Loop อัตโนมัติ "
                          "-> ใช้ของ Manual Group แทน", flush=True)
                consolidated_loops[mg_id] = mg_data
            initial_loops_in_consolidation.update(manual_first_vars_conv)
            print(f"Found {len(consolidated_loops)} consolidated loop groups "
                  f"({len(manual_loops_prepared)} จาก Manual Group).", flush=True)

            # ============================================================ #
            #   PHASE 2: WRITING                                           #
            # ============================================================ #
            print("\n--- PHASE 2: Writing to Excel based on pre-computed structures ---", flush=True)
            try:
                wb = openpyxl.load_workbook(excel_template_path)
                ws = wb.active
            except Exception as e:
                raise IOError(f"เปิดไฟล์ Excel Template ไม่ได้: {e}")

            cols = dict(item='A', fmt='B', code='C', type='D', disp='E', loopsub='F', id='G',
                        label='H', val_id='G', val_lbl='H', valid='I', cat='J', digit='K',
                        min='L', max='M', dec='N', stat='O')
            try:
                idx = {k: column_index_from_string(v) for k, v in cols.items()}
            except ValueError as e:
                raise ValueError(f"ชื่อคอลัมน์ใน Template ไม่ถูกต้อง: {e}")

            i_idx, fmt_idx, type_idx = idx['item'], idx['fmt'], idx['type']
            disp_idx, loopsub_idx, id_idx, label_idx = idx['disp'], idx['loopsub'], idx['id'], idx['label']
            val_id_idx, val_lbl_idx, valid_idx = idx['val_id'], idx['val_lbl'], idx['valid']
            digit_idx, min_idx, max_idx, dec_idx, stat_idx = (
                idx['digit'], idx['min'], idx['max'], idx['dec'], idx['stat'])

            def write_loop_header(row, out_type, loop_id):
                ws.cell(row=row, column=i_idx).value = "Item"
                ws.cell(row=row, column=fmt_idx).value = "Survey"
                if out_type == "Loop Text":
                    ws.cell(row=row, column=type_idx).value = "Loop(Text)"
                    ws.cell(row=row, column=digit_idx).value = 4000
                elif out_type == "Loop Numeric":
                    ws.cell(row=row, column=type_idx).value = "Loop(Numeric)"
                    ws.cell(row=row, column=digit_idx).value = 11
                    ws.cell(row=row, column=min_idx).value = -9999999999
                    ws.cell(row=row, column=max_idx).value = 9999999999
                    ws.cell(row=row, column=dec_idx).value = 0
                    ws.cell(row=row, column=stat_idx).value = None
                elif out_type == "SA":
                    ws.cell(row=row, column=type_idx).value = "Loop(SA)"
                elif out_type == "MA":
                    ws.cell(row=row, column=type_idx).value = "Loop(MA)"
                ws.cell(row=row, column=disp_idx).value = "O"
                ws.cell(row=row, column=id_idx).value = loop_id

            def write_value_labels(row, value_labels):
                """เขียน code 1..max เติมช่องว่างด้วย label ว่าง คืนเลขแถวถัดไป"""
                if not value_labels or not isinstance(value_labels, dict):
                    return row
                sorted_codes = sorted(k for k in value_labels.keys() if isinstance(k, (int, float)))
                if not sorted_codes:
                    return row
                max_c = int(sorted_codes[-1])
                code_map = {int(k): v for k, v in value_labels.items() if isinstance(k, (int, float))}
                print(f"   Writing codes from 1 to {max_c}, filling gaps.", flush=True)
                for code in range(1, max_c + 1):
                    ws.cell(row=row, column=val_id_idx).value = code
                    ws.cell(row=row, column=val_lbl_idx).value = code_map.get(code, "")
                    ws.cell(row=row, column=valid_idx).value = "Valid"
                    row += 1
                return row

            write_row = 3
            written_vars = set()
            has_types = hasattr(meta, 'variable_types') and isinstance(meta.variable_types, dict)
            has_orig_types = (hasattr(meta, 'original_variable_types')
                              and isinstance(meta.original_variable_types, dict))

            for index, var_name in enumerate(all_vars):
                if var_name in written_vars:
                    continue

                # --- 7.A CONSOLIDATED loop ---
                final_loop_id_to_write = None
                loop_data_to_write = None
                for final_id, loop_data in consolidated_loops.items():
                    if var_name == loop_data['first_representative_var']:
                        final_loop_id_to_write = final_id
                        loop_data_to_write = loop_data
                        break

                if loop_data_to_write:
                    print(f"[{index + 1}/{len(all_vars)}] '{var_name}' "
                          f"| Start CONSOLIDATED Loop '{final_loop_id_to_write}'", flush=True)
                    out_type = loop_data_to_write['type']
                    reps = loop_data_to_write['representative_vars']
                    lbl_src = loop_data_to_write['label_source_var']

                    write_loop_header(write_row, out_type, final_loop_id_to_write)
                    ws.cell(row=write_row, column=label_idx).value = None
                    write_row += 1

                    print(f"   Writing {len(reps)} consolidated sub-items...", flush=True)
                    for i_sub, rep_var_name in enumerate(reps, 1):
                        rep_var_index = var_name_to_index.get(rep_var_name, -1)
                        sub_label_value = meta.column_labels[rep_var_index] if rep_var_index != -1 else ""
                        ws.cell(row=write_row, column=loopsub_idx).value = "Loop sub"
                        ws.cell(row=write_row, column=id_idx).value = f"{final_loop_id_to_write}({i_sub})"
                        ws.cell(row=write_row, column=label_idx).value = sub_label_value
                        ws.cell(row=write_row, column=disp_idx).value = "O"
                        write_row += 1

                    if out_type in ("SA", "MA"):
                        write_row = write_value_labels(write_row, filtered_meta_value_labels.get(lbl_src))

                    written_vars.update(loop_data_to_write['all_original_vars'])
                    continue

                # --- 7.B UNCONSOLIDATED user loop ---
                if var_name in initial_user_loop_groups and var_name not in initial_loops_in_consolidation:
                    print(f"[{index + 1}/{len(all_vars)}] '{var_name}' | Start UNCONSOLIDATED Loop", flush=True)
                    group_info = initial_user_loop_groups[var_name]
                    g_type = group_info['type']
                    g_vars = group_info['vars']
                    disp_id = self.user_defined_loop_names.get(var_name, group_info['base_name'])
                    is_o_suf = any(re.search(r'_O\d+$', v) for v in g_vars)
                    out_type = "MA" if is_o_suf else g_type

                    write_loop_header(write_row, out_type, disp_id)
                    write_row += 1

                    pnon_pattern = re.compile(r'(.+?)_(\d+)_O(\d+)$')
                    is_pnon_loop = any(pnon_pattern.match(v) for v in g_vars)
                    sub_items_to_write = {}
                    if is_pnon_loop and out_type == "MA":
                        print(f"   Applying special sub-item logic for '{disp_id}'...", flush=True)
                        sub_groups = defaultdict(list)
                        for sub_var_name in g_vars:
                            match = pnon_pattern.match(sub_var_name)
                            if match:
                                sub_groups[int(match.group(2))].append((int(match.group(3)), sub_var_name))
                        for inter_num, o_list in sorted(sub_groups.items()):
                            if o_list:
                                sub_items_to_write[inter_num] = sorted(o_list)[0][1]
                    else:
                        print(f"   Writing {len(g_vars)} default sub-items...", flush=True)
                        for i_sub, sub_var_name in enumerate(g_vars, 1):
                            sub_items_to_write[i_sub] = sub_var_name

                    for i_sub, key in enumerate(sorted(sub_items_to_write.keys()), 1):
                        rep_var_name = sub_items_to_write[key]
                        sub_var_index = var_name_to_index.get(rep_var_name, -1)
                        sub_label_value = meta.column_labels[sub_var_index] if sub_var_index != -1 else ""
                        ws.cell(row=write_row, column=loopsub_idx).value = "Loop sub"
                        ws.cell(row=write_row, column=id_idx).value = f"{disp_id}({i_sub})"
                        ws.cell(row=write_row, column=label_idx).value = sub_label_value
                        ws.cell(row=write_row, column=disp_idx).value = "O"
                        write_row += 1

                    if out_type in ("SA", "MA"):
                        write_row = write_value_labels(write_row, filtered_meta_value_labels.get(var_name))

                    written_vars.update(g_vars)
                    continue

                # --- 7.C Non-Loop / Auto-Detect ---
                var_label = meta.column_labels[index]
                current_value_labels = filtered_meta_value_labels.get(var_name)

                col_d_value = None
                disp_name = var_name
                write_var = True
                type_code = meta.variable_types.get(var_name, -1) if has_types else -1
                orig_fmt = meta.original_variable_types.get(var_name, '') if has_orig_types else ''
                is_str = type_code > 0 or (isinstance(orig_fmt, str) and orig_fmt.strip().upper().startswith('A'))

                if is_str:
                    col_d_value = "Text"
                else:
                    ma_match = self.ma_pattern.match(var_name)
                    if ma_match:
                        auto_base = ma_match.group(1)
                        if auto_base in processed_ma_base_names_auto:
                            write_var = False
                        else:
                            col_d_value = "MA"
                            disp_name = auto_base
                            processed_ma_base_names_auto.add(auto_base)
                    elif current_value_labels:
                        col_d_value = "SA"
                    else:
                        col_d_value = "Numeric"

                if write_var:
                    ws.cell(row=write_row, column=i_idx).value = "Item"
                    ws.cell(row=write_row, column=fmt_idx).value = "Survey"
                    ws.cell(row=write_row, column=type_idx).value = col_d_value
                    ws.cell(row=write_row, column=disp_idx).value = "O"
                    ws.cell(row=write_row, column=id_idx).value = disp_name
                    ws.cell(row=write_row, column=label_idx).value = var_label
                    if col_d_value == "Text":
                        ws.cell(row=write_row, column=digit_idx).value = 4000
                    elif col_d_value == "Numeric":
                        ws.cell(row=write_row, column=digit_idx).value = 11
                        ws.cell(row=write_row, column=min_idx).value = -9999999999
                        ws.cell(row=write_row, column=max_idx).value = 9999999999
                        ws.cell(row=write_row, column=dec_idx).value = 0
                    write_row += 1

                    if col_d_value in ("SA", "MA") and current_value_labels:
                        print(f"   Writing labels for Auto-Detect '{disp_name}'...", flush=True)
                        write_row = write_value_labels(write_row, current_value_labels)

                written_vars.add(var_name)

            print("\n--- เพิ่ม 'End' marker ---", flush=True)
            ws.cell(row=write_row, column=i_idx).value = "End"
            for col_idx in range(fmt_idx, stat_idx + 2):
                ws.cell(row=write_row, column=col_idx).value = None

            print(f"Saving Excel file to: {excel_output_path}", flush=True)
            wb.save(excel_output_path)
            self.set_status(f"สำเร็จ! บันทึกที่: {os.path.basename(excel_output_path)}", "#2F855A")
            QtWidgets.QMessageBox.information(self, "สำเร็จ", f"สำเร็จ!\nบันทึกที่:\n{excel_output_path}")

        except (IOError, ValueError, AttributeError) as e:
            print(f"ERROR: {e}", flush=True)
            traceback.print_exc()
            self.set_status(f"ผิดพลาด: {e}", "#C53030")
            QtWidgets.QMessageBox.critical(self, "ผิดพลาด", f"เกิดข้อผิดพลาดในการประมวลผล:\n{e}")
        except Exception as e:
            print(f"ERROR: Unexpected error: {e}", flush=True)
            traceback.print_exc()
            self.set_status(f"ผิดพลาดไม่คาดคิด: {e}", "#C53030")
            QtWidgets.QMessageBox.critical(self, "ผิดพลาด", f"ผิดพลาดไม่คาดคิด:\n{e}")
        finally:
            print("--- สิ้นสุดการทำงานของ convert_file ---", flush=True)


# =========================================================================
#  Entry point
# =========================================================================
_MAIN_WINDOW = None   # กัน QMainWindow ถูก GC เมื่อถูกเรียกจาก QApplication ที่มีอยู่แล้ว


def run_this_app(working_dir=None):
    global _MAIN_WINDOW
    original_stdout = sys.stdout
    print("--- SPSS_LOG_INFO: Starting 'Program_ItemdefSPSS_Qt' via run_this_app() ---", file=original_stdout)

    if working_dir and os.path.isdir(working_dir):
        try:
            os.chdir(working_dir)
        except Exception as e:
            print(f"SPSS_LOG_WARNING: Could not chdir to {working_dir}: {e}", file=original_stdout)

    app = QtWidgets.QApplication.instance()
    owns_app = app is None
    if owns_app:
        app = QtWidgets.QApplication(sys.argv)
    app.setStyleSheet(STYLESHEET)

    window = None
    try:
        icon_path = resource_path("Peth.ico")
        if os.path.exists(icon_path):
            app.setWindowIcon(QtGui.QIcon(icon_path))
            print(f"SPSS_LOG_INFO: App icon loaded.", file=original_stdout)

        window = SpssToExcelConverter()
        _MAIN_WINDOW = window

        # จัดหน้าต่างกลางจอ
        screen = app.primaryScreen()
        if screen is not None:
            geo = screen.availableGeometry()
            frame = window.frameGeometry()
            frame.moveCenter(geo.center())
            window.move(frame.topLeft())

        window.show()
        if owns_app:
            app.exec()
    except Exception as e:
        print(f"\nSPSS_LOG_ERROR: An error occurred during application execution:", file=original_stdout)
        print(f"{type(e).__name__}: {e}", file=original_stdout)
        print(traceback.format_exc(), file=original_stdout)
        try:
            QtWidgets.QMessageBox.critical(None, "Application Error (SPSS Qt)",
                                           f"An unexpected error occurred:\n{e}")
        except Exception as popup_err:
            print(f"SPSS_LOG_ERROR: Could not show error popup: {popup_err}", file=original_stdout)
    finally:
        # ถ้าเราเป็นเจ้าของ QApplication แปลว่า exec() จบแล้ว = ปิดโปรแกรม -> คืน stdout ได้
        # ถ้ามี QApplication อยู่ก่อน (ถูก embed) หน้าต่างยังทำงานอยู่ ปล่อยให้ closeEvent คืนเอง
        if owns_app:
            if window is not None and hasattr(window, 'cleanup_redirector'):
                window.cleanup_redirector()
            else:
                sys.stdout = original_stdout
            print("--- SPSS_LOG_INFO: run_this_app() finished. stdout restored. ---", file=original_stdout)
        else:
            print("--- SPSS_LOG_INFO: run_this_app() attached to an existing QApplication. ---",
                  file=original_stdout)


if __name__ == "__main__":
    print("--- Running Program_ItemdefSPSS_Qt.py directly for testing ---")
    run_this_app()
    print("--- Finished direct execution of Program_ItemdefSPSS_Qt.py ---")
