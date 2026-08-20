import os
import re
import traceback

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PyQt6.QtCore import QElapsedTimer, QThread, QTimer, Qt, pyqtSignal
from PyQt6.QtGui import QColor, QIcon, QPainter, QPainterPath, QPixmap
from PyQt6.QtWidgets import (
    QApplication,
    QComboBox,
    QCompleter,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QStyle,
    QTableWidget,
    QTableWidgetItem,
    QProgressBar,
    QVBoxLayout,
    QWidget,
)


DONE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ADDED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
REMOVED_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")


def make_leaf_icon(size: int = 128) -> QIcon:
    pix = QPixmap(size, size)
    pix.fill(Qt.GlobalColor.transparent)

    p = QPainter(pix)
    p.setRenderHint(QPainter.RenderHint.Antialiasing, True)
    p.setPen(Qt.PenStyle.NoPen)
    p.setBrush(QColor("#e8f5e9"))
    p.drawEllipse(4, 4, size - 8, size - 8)

    path = QPainterPath()
    path.moveTo(size * 0.52, size * 0.18)
    path.cubicTo(size * 0.82, size * 0.26, size * 0.84, size * 0.62, size * 0.54, size * 0.80)
    path.cubicTo(size * 0.30, size * 0.70, size * 0.24, size * 0.40, size * 0.52, size * 0.18)
    p.setBrush(QColor("#43a047"))
    p.drawPath(path)
    p.setPen(QColor("#1b5e20"))
    p.drawLine(int(size * 0.53), int(size * 0.24), int(size * 0.48), int(size * 0.76))
    p.end()
    return QIcon(pix)


def txt(v) -> str:
    if pd.isna(v):
        return ""
    return str(v).strip()


def norm(v) -> str:
    return "".join(ch.lower() for ch in txt(v) if ch.isalnum())


def norm_sheet_name(v) -> str:
    s = txt(v).replace("\u3000", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def is_non_data_sbj(v) -> bool:
    return norm(v) in {"", "sbjnum", "idrds", "na"}


def detect_header_row_df(df: pd.DataFrame, must_keys: set[str], max_scan: int = 60) -> int:
    best_idx, best_score = -1, -1
    for i in range(min(len(df), max_scan)):
        keys = {norm(v) for v in df.iloc[i].tolist() if txt(v) != ""}
        score = len(must_keys.intersection(keys))
        if score > best_score:
            best_idx, best_score = i, score
    if best_idx < 0 or best_score < len(must_keys):
        raise ValueError(f"Header row not found (required: {', '.join(sorted(must_keys))})")
    return best_idx


def build_idx_map_from_header_row(row_values: list) -> dict[str, int]:
    idx = {}
    for c, v in enumerate(row_values):
        k = norm(v)
        if k and k not in idx:
            idx[k] = c
    return idx


def parse_code_cell(v) -> tuple[str | None, str | None]:
    s = txt(v)
    if s == "":
        return None, None

    if isinstance(v, (int, float)) and not pd.isna(v):
        n = float(v)
        if n.is_integer():
            code = str(int(n))
            if code == "0":
                return None, s
            return code, None
        return None, s

    m = re.match(r"^\s*(\d+)(?:\.0+)?\s*$", s)
    if m:
        code = m.group(1)
        if code == "0":
            return None, s
        return code, None
    return None, s


def is_numeric_only_text(v) -> bool:
    s = txt(v)
    if s == "":
        return False
    return re.fullmatch(r"\d+(?:\.\d+)?", s) is not None


def detect_code_block_range(header_vals: list, var_col: int) -> tuple[int, int] | None:
    # An explicit "Code" block belongs to THIS variable only when nothing but
    # blank headers sits between the variable column and the Code columns.
    # Scanning further would cross into the next variable and steal its codes.
    code_start = None
    for c in range(var_col + 1, len(header_vals)):
        k = norm(header_vals[c])
        if k == "code":
            code_start = c
            break
        if k != "":
            break

    # If no adjacent "Code" header, only use columns between this variable and
    # the next non-empty header. If none in-between, treat as no code block.
    if code_start is None:
        next_header_col = len(header_vals)
        for c in range(var_col + 1, len(header_vals)):
            if norm(header_vals[c]) != "":
                next_header_col = c
                break
        start = var_col + 1
        end = next_header_col - 1
        if end < start:
            return None
        return start, end

    code_end = len(header_vals) - 1
    for c in range(code_start + 1, len(header_vals)):
        k = norm(header_vals[c])
        if k != "" and k != "code":
            code_end = c - 1
            break
    if code_end < code_start:
        return None
    # Include any blank columns between the variable and its Code block -
    # they belong to the same answer block.
    return var_col + 1, code_end


def extract_code_slots_from_row(row_vals: list, code_start: int, code_end: int) -> list[tuple[str, str]]:
    # Return ordered slots so we can highlight the exact deleted position.
    # Each item is ("keep", code) or ("drop", original_text).
    slots: list[tuple[str, str]] = []
    upper = min(code_end, len(row_vals) - 1)
    for c in range(code_start, upper + 1):
        v = row_vals[c]
        code, bad = parse_code_cell(v)
        if code is not None:
            slots.append(("keep", code))
        elif bad is not None:
            slots.append(("drop", bad))
    return slots


def to_ns_base(var_name: str) -> str:
    s = txt(var_name)
    if s == "":
        return "NVAR"
    return f"N{s[0].upper()}{s[1:]}"


def merge_base_name(var_name: str) -> str:
    s = txt(var_name)
    if s == "":
        return s
    m = re.match(r"^(.+?)[_](\d+)$", s)
    if m:
        return m.group(1)
    return s


def parse_var_list(v: str) -> list[str]:
    parts = re.split(r"[,\n|]+", txt(v))
    out = []
    seen = set()
    for p in parts:
        t = txt(p)
        if t == "":
            continue
        k = norm(t)
        if k in seen:
            continue
        seen.add(k)
        out.append(t)
    return out


def load_items_from_settings_file(path: str, selected_sheets: list[str] | None = None) -> tuple[list[dict], list[str]]:
    xls = pd.ExcelFile(path)
    items: list[dict] = []
    seen = set()
    warnings: list[str] = []
    sheets = list(xls.sheet_names)
    if selected_sheets:
        resolved: list[str] = []
        by_exact = {txt(s): s for s in xls.sheet_names}
        by_norm = {norm_sheet_name(s): s for s in xls.sheet_names}
        for wanted in selected_sheets:
            picked = by_exact.get(txt(wanted)) or by_norm.get(norm_sheet_name(wanted))
            if picked and picked not in resolved:
                resolved.append(picked)
            elif txt(wanted):
                warnings.append(f"Settings sheet '{wanted}' not found")
        sheets = resolved

    for sheet in sheets:
        df = pd.read_excel(path, sheet_name=sheet, header=None, dtype=object)
        item_col = None
        merge_col = None
        raw_col = None
        use_col = None
        scan_rows = min(len(df), 40)
        for rr in range(scan_rows):
            row = df.iloc[rr].tolist()
            for cc, v in enumerate(row):
                if norm(v) == "item":
                    item_col = cc
                if norm(v) == "merge":
                    merge_col = cc
                if norm(v) in {"rawcolumn", "rawname", "rawvar", "raw"}:
                    raw_col = cc
                if norm(v) == "use":
                    use_col = cc
            if item_col is not None:
                start_row = rr + 1
                break
        if item_col is None:
            continue

        found_in_sheet = 0
        for rr in range(start_row, len(df)):
            row = df.iloc[rr].tolist()
            if item_col >= len(row):
                continue
            raw = txt(row[item_col])
            if raw == "":
                continue
            if norm(raw) == "item":
                continue
            key = norm(raw)
            if key == "" or key in seen:
                continue
            seen.add(key)
            merge_flag = False
            # Preferred: explicit "Merge" header column.
            # Fallback: next column after Item (common settings layout).
            merge_val = ""
            if merge_col is not None and merge_col < len(row):
                merge_val = txt(row[merge_col]).lower()
            elif (item_col + 1) < len(row):
                merge_val = txt(row[item_col + 1]).lower()
            merge_flag = merge_val in {"merge", "m", "1", "y", "yes", "true"}
            raw_name = txt(row[raw_col]) if raw_col is not None and raw_col < len(row) else ""
            # "Use" column: when the header exists, a blank cell means the item is
            # deselected. Without a Use header everything defaults to enabled.
            enabled = True
            if use_col is not None and use_col < len(row):
                enabled = txt(row[use_col]).lower() in {"1", "y", "yes", "true", "use", "x"}
            items.append({"name": raw, "raw_name": raw_name, "merge": merge_flag, "enabled": enabled})
            found_in_sheet += 1
        if found_in_sheet == 0:
            warnings.append(f"Sheet '{sheet}' has Item header but no values")

    if not items:
        raise ValueError("No Item values found in settings file")
    return items, warnings


def prepare_text_data(
    text_path: str,
    selected_sheets: list[str] | None = None,
) -> tuple[list[tuple[str, pd.DataFrame]], list[str]]:
    xls = pd.ExcelFile(text_path)
    warnings: list[str] = []
    sheets = xls.sheet_names
    if selected_sheets:
        resolved: list[str] = []
        seen = set()
        by_exact = {txt(s): s for s in xls.sheet_names}
        by_norm: dict[str, str] = {}
        for s in xls.sheet_names:
            k = norm_sheet_name(s)
            if k not in by_norm:
                by_norm[k] = s

        for wanted_raw in selected_sheets:
            wanted_text = txt(wanted_raw)
            if wanted_text == "":
                continue

            picked = by_exact.get(wanted_text)
            if picked is None:
                picked = by_norm.get(norm_sheet_name(wanted_text))
                if picked is not None and picked != wanted_text:
                    warnings.append(
                        f"Sheet '{wanted_text}' resolved as '{picked}' (normalized spaces/case)"
                    )

            if picked is None:
                warnings.append(f"Sheet '{wanted_text}' not found in Text file")
                continue

            key = norm_sheet_name(picked)
            if key in seen:
                continue
            seen.add(key)
            resolved.append(picked)

        sheets = resolved
    frames: list[tuple[str, pd.DataFrame]] = []
    for sheet in sheets:
        df = pd.read_excel(text_path, sheet_name=sheet, header=None, dtype=object)
        frames.append((sheet, df))
    return frames, warnings


RESERVED_HEADER_KEYS = {"sbjnum", "idrds", "na", "code", "item", "merge", "use", ""}


def detect_raw_columns(raw_path: str, raw_sheet: str, max_scan: int = 80) -> dict[str, str]:
    # Returns {normalized_key: original_display_name} for every non-empty header cell.
    df = pd.read_excel(raw_path, sheet_name=raw_sheet, header=None, nrows=max_scan, dtype=object)
    header_idx = detect_header_row_df(df, {"sbjnum"}, max_scan=max_scan)
    header_vals = df.iloc[header_idx].tolist()
    cols: dict[str, str] = {}
    for v in header_vals:
        name = txt(v)
        if name == "":
            continue
        k = norm(v)
        if k not in cols:
            cols[k] = name
    return cols


def prepare_text_headers(text_path: str, sheets: list[str], max_scan: int = 80) -> list[tuple[str, pd.DataFrame]]:
    # Only read the first `max_scan` rows of each sheet - enough to find the header
    # row without loading full sheet data, which matters when many sheets are selected.
    frames: list[tuple[str, pd.DataFrame]] = []
    for sheet in sheets:
        df = pd.read_excel(text_path, sheet_name=sheet, header=None, nrows=max_scan, dtype=object)
        frames.append((sheet, df))
    return frames


def detect_question_candidates(text_data: list[tuple[str, pd.DataFrame]]) -> list[str]:
    # Only surface columns that look like a genuine OE/verbatim question - i.e. they have
    # an actual code/verbatim block after them (explicit "Code" header, or a gap before the
    # next header). Plain data columns packed tight against the next header are excluded,
    # since those are ordinary Raw fields (quota, cell, flags, ...), not open-end questions.
    candidates: dict[str, str] = {}
    for _sheet, df in text_data:
        try:
            header_idx = detect_header_row_df(df, {"sbjnum"})
        except ValueError:
            continue
        header_vals = df.iloc[header_idx].tolist()
        for c, v in enumerate(header_vals):
            k = norm(v)
            if k in RESERVED_HEADER_KEYS:
                continue
            name = txt(v)
            if name == "":
                continue
            if detect_code_block_range(header_vals, c) is None:
                continue
            if k not in candidates:
                candidates[k] = name
    return sorted(candidates.values(), key=str.lower)


def load_verbatim_parts(
    text_data: list[tuple[str, pd.DataFrame]],
    var_name: str,
) -> tuple[dict, list[str], bool]:
    var_key = norm(var_name)
    merged: dict[str, dict] = {}
    warnings: list[str] = []
    found_question = False

    for sheet, df in text_data:
        try:
            header_idx = detect_header_row_df(df, {"sbjnum", var_key})
        except ValueError:
            warnings.append(f"Skip sheet '{sheet}' (missing SbjNum + {var_name})")
            continue

        header_vals = df.iloc[header_idx].tolist()
        idx = build_idx_map_from_header_row(header_vals)
        sbj_col = idx.get("sbjnum")
        var_col = idx.get(var_key)
        if sbj_col is None or var_col is None:
            warnings.append(f"Skip sheet '{sheet}' (required columns not complete)")
            continue
        found_question = True

        code_range = detect_code_block_range(header_vals, var_col)

        for r in range(header_idx + 1, len(df)):
            row = df.iloc[r].tolist()
            if sbj_col >= len(row):
                continue
            sbj = txt(row[sbj_col])
            if is_non_data_sbj(sbj):
                continue

            verb = txt(row[var_col]) if var_col < len(row) else ""
            # Verbatim column should be free text. If it is numeric-only, treat as no verbatim.
            if is_numeric_only_text(verb):
                verb = ""
            slots = extract_code_slots_from_row(row, code_range[0], code_range[1]) if code_range is not None else []
            if not verb and not slots:
                continue

            if sbj not in merged:
                merged[sbj] = {"verbatim": "", "slots": []}
            rec = merged[sbj]
            if verb:
                if rec["verbatim"] == "":
                    rec["verbatim"] = verb
                elif rec["verbatim"] != verb:
                    warnings.append(f"SbjNum {sbj}: duplicate verbatim found (keep first value)")
            rec["slots"].extend(slots)

    return merged, warnings, found_question


def build_raw_header_candidates(ws, max_scan: int = 80) -> list[tuple[int, dict[str, int]]]:
    cands: list[tuple[int, dict[str, int]]] = []
    for rr in range(1, min(ws.max_row, max_scan) + 1):
        idx = {}
        for cc in range(1, ws.max_column + 1):
            k = norm(ws.cell(row=rr, column=cc).value)
            if k and k not in idx:
                idx[k] = cc
        cands.append((rr, idx))
    return cands


def detect_header_row_ws(ws, must_keys: set[str], max_scan: int = 80) -> int:
    best_row, best_score = -1, -1
    for rr in range(1, min(ws.max_row, max_scan) + 1):
        keys = {norm(ws.cell(row=rr, column=cc).value) for cc in range(1, ws.max_column + 1)}
        score = len(must_keys.intersection(keys))
        if score > best_score:
            best_row, best_score = rr, score
    if best_row < 0 or best_score < len(must_keys):
        raise ValueError(f"Raw header row not found (required: {', '.join(sorted(must_keys))})")
    return best_row


def ws_header_map(ws, header_row: int) -> dict[str, int]:
    idx = {}
    for cc in range(1, ws.max_column + 1):
        k = norm(ws.cell(row=header_row, column=cc).value)
        if k and k not in idx:
            idx[k] = cc
    return idx


def run_merge(
    raw_path: str,
    raw_sheet: str,
    text_path: str,
    var_configs: list[dict],
    out_path: str,
    text_sheets: list[str] | None = None,
) -> dict:
    wb = load_workbook(raw_path)
    ws = wb[raw_sheet]
    text_data, text_data_warnings = prepare_text_data(text_path, text_sheets)

    all_warnings: list[str] = list(text_data_warnings)
    details: list[dict] = []
    raw_header_candidates = build_raw_header_candidates(ws)
    sbj_row_cache: dict[tuple[int, int], dict[str, int]] = {}

    code_groups: dict[str, dict] = {}
    group_order: list[str] = []

    for cfg in var_configs:
        var_name = txt(cfg.get("name"))
        raw_name = txt(cfg.get("raw_name")) or var_name
        merge_flag = bool(cfg.get("merge", False))
        if var_name == "":
            continue
        merged, warnings, found_question = load_verbatim_parts(text_data, var_name)
        all_warnings.extend([f"[{var_name}] {w}" for w in warnings])
        if not merged and not found_question:
            all_warnings.append(f"[{var_name}] no mergeable data found")
            details.append(
                {
                    "var": var_name,
                    "raw_var": raw_name,
                    "status": "FAILED",
                    "note": "No data in Text",
                    "matched_rows": 0,
                    "touched_verbatim": 0,
                    "max_codes": 0,
                    "removed_hits": 0,
                }
            )
            continue

        var_key = norm(raw_name)
        header_row = None
        idx = None
        for rr, m in raw_header_candidates:
            if "sbjnum" in m and var_key in m:
                header_row = rr
                idx = m
                break
        if header_row is None or idx is None:
            all_warnings.append(f"[{var_name}] mapped Raw column '{raw_name}' not found")
            details.append(
                {
                    "var": var_name,
                    "raw_var": raw_name,
                    "status": "FAILED",
                    "note": f"Header/column '{raw_name}' not found in Raw",
                    "matched_rows": 0,
                    "touched_verbatim": 0,
                    "max_codes": 0,
                    "removed_hits": 0,
                }
            )
            continue
        sbj_col = idx.get("sbjnum")
        var_col = idx.get(var_key)
        if sbj_col is None or var_col is None:
            all_warnings.append(f"[{var_name}] column SbjNum or '{raw_name}' not found in Raw")
            details.append(
                {
                    "var": var_name,
                    "raw_var": raw_name,
                    "status": "FAILED",
                    "note": f"Column '{raw_name}' not found in Raw",
                    "matched_rows": 0,
                    "touched_verbatim": 0,
                    "max_codes": 0,
                    "removed_hits": 0,
                }
            )
            continue

        matched_rows = 0
        touched_verbatim = 0
        removed_hits = 0

        cache_key = (header_row, sbj_col)
        if cache_key not in sbj_row_cache:
            sbj_row_map: dict[str, int] = {}
            for rr in range(header_row + 1, ws.max_row + 1):
                sbj = txt(ws.cell(row=rr, column=sbj_col).value)
                if is_non_data_sbj(sbj):
                    continue
                sbj_row_map[sbj] = rr
            sbj_row_cache[cache_key] = sbj_row_map
        sbj_row_map = sbj_row_cache[cache_key]

        for sbj, rec in merged.items():
            rr = sbj_row_map.get(sbj)
            if rr is None:
                continue
            matched_rows += 1
            if rec["verbatim"] != "":
                ws.cell(row=rr, column=var_col, value=rec["verbatim"]).fill = DONE_FILL
                touched_verbatim += 1

        max_codes = max((len(v["slots"]) for v in merged.values()), default=0)
        target_base = merge_base_name(raw_name) if merge_flag else raw_name
        ns_base = to_ns_base(target_base)

        if ns_base not in code_groups:
            code_groups[ns_base] = {"header_row": header_row, "rows": {}, "min_cols": 0}
            group_order.append(ns_base)
        grp = code_groups[ns_base]
        if found_question and max_codes <= 0:
            grp["min_cols"] = max(int(grp.get("min_cols", 0)), 1)
        row_slots_map: dict[int, list[tuple[str, str]]] = grp["rows"]

        for sbj, rr in sbj_row_map.items():
            rec = merged.get(sbj)
            if rec is None:
                continue
            row_has_removed = False
            if rr not in row_slots_map:
                row_slots_map[rr] = []
            row_slots_map[rr].extend(rec["slots"])
            for slot in rec["slots"]:
                if slot[0] == "drop":
                    row_has_removed = True
            if row_has_removed:
                removed_hits += 1

        details.append(
            {
                "var": var_name,
                "raw_var": raw_name,
                "status": "SUCCESS",
                "note": "Completed" if matched_rows > 0 else ("Question found; no code/verbatim data" if found_question else "No matched SbjNum"),
                "matched_rows": matched_rows,
                "touched_verbatim": touched_verbatim,
                "max_codes": max(max_codes, 1 if found_question and max_codes <= 0 else 0),
                "removed_hits": removed_hits,
            }
        )

    for ns_base in group_order:
        grp = code_groups[ns_base]
        header_row = int(grp["header_row"])
        row_slots_map: dict[int, list[tuple[str, str]]] = grp["rows"]
        min_cols = int(grp.get("min_cols", 0))
        max_codes = max(max((len(v) for v in row_slots_map.values()), default=0), min_cols)
        if max_codes <= 0:
            continue
        add_start_col = ws.max_column + 1
        for i in range(max_codes):
            ws.cell(row=header_row, column=add_start_col + i, value=f"{ns_base}_O{i+1}")
        for rr, slots in row_slots_map.items():
            for i, slot in enumerate(slots):
                kind, value = slot
                cell = ws.cell(row=rr, column=add_start_col + i)
                if kind == "keep":
                    cell.value = value
                    cell.fill = ADDED_FILL
                else:
                    cell.value = ""
                    cell.fill = REMOVED_FILL

    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(
        [
            "Item (Text)",
            "Raw Column",
            "Merge",
            "Output Base",
            "Status",
            "Note",
            "Matched Rows",
            "Overwritten Verbatim",
            "New Code Cols",
            "Removed-Highlight Rows",
        ]
    )
    dmap = {d["var"]: d for d in details}
    for cfg in var_configs:
        var_name = txt(cfg.get("name"))
        raw_name = txt(cfg.get("raw_name")) or var_name
        merge_flag = bool(cfg.get("merge", False))
        target_base = merge_base_name(raw_name) if merge_flag else raw_name
        d = dmap.get(var_name, {})
        ws_sum.append(
            [
                var_name,
                raw_name,
                "MERGE" if merge_flag else "",
                target_base,
                d.get("status", "FAILED"),
                d.get("note", "Unknown"),
                d.get("matched_rows", 0),
                d.get("touched_verbatim", 0),
                d.get("max_codes", 0),
                d.get("removed_hits", 0),
            ]
        )

    wb.save(out_path)
    return {
        "details": details,
        "warnings": all_warnings,
        "sheet": raw_sheet,
    }


class QuestionScanWorker(QThread):
    # Reading a large .xlsx (shared strings, zip decompression, etc.) can take a
    # noticeable amount of wall-clock time regardless of how few rows are requested.
    # Doing it on a background thread keeps the UI responsive instead of freezing.
    scanned = pyqtSignal(dict, list)
    failed = pyqtSignal(str)

    def __init__(self, raw_path: str, raw_sheet: str, text_path: str, text_sheets: list[str], parent=None):
        super().__init__(parent)
        self.raw_path = raw_path
        self.raw_sheet = raw_sheet
        self.text_path = text_path
        self.text_sheets = text_sheets

    def run(self):
        try:
            raw_map = detect_raw_columns(self.raw_path, self.raw_sheet)
            text_data = prepare_text_headers(self.text_path, self.text_sheets)
            candidates = detect_question_candidates(text_data)
        except Exception as e:
            self.failed.emit(str(e))
            return
        self.scanned.emit(raw_map, candidates)


class VerbatimMapperApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("VERBATIM + CODE Mapper")
        self.resize(1060, 800)
        self.setMinimumSize(920, 700)
        self.setWindowIcon(make_leaf_icon())

        self.raw_path = ""
        self.text_path = ""
        self.text_sheet_names: list[str] = []
        self.selected_text_sheets: list[str] = []
        self.settings_sheet_names: list[str] = []
        self.selected_settings_sheets: list[str] = []
        self.var_merge_map: dict[str, bool] = {}
        self.settings_items: list[dict] = []
        self.settings_path = ""
        self._last_dir = ""
        self._scan_worker: QuestionScanWorker | None = None
        self._build_ui()

    def _build_ui(self):
        self.setStyleSheet(
            """
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #0b132b, stop:0.45 #10213f, stop:1 #0a1a34);
            }
            QWidget {
                font-family: 'Segoe UI';
                font-size: 10pt;
                color: #e7eefc;
            }
            QFrame#HeroCard {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #06b6d4, stop:1 #14b8a6);
                border-radius: 16px;
                border: 1px solid rgba(255,255,255,0.28);
            }
            QLabel#Title {
                font-size: 28px;
                font-weight: 700;
                color: #042f3b;
            }
            QLabel#SubTitle {
                color: #0f3e4c;
                font-size: 11px;
            }
            QFrame#Card {
                background: rgba(10, 17, 33, 0.82);
                border: 1px solid rgba(255,255,255,0.16);
                border-radius: 14px;
            }
            QLabel#FieldLabel {
                color: #cfe5ff;
                font-weight: 600;
                min-width: 120px;
            }
            QLineEdit, QComboBox {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.22);
                border-radius: 9px;
                padding: 7px 10px;
                color: #f2f7ff;
            }
            QComboBox QAbstractItemView {
                background: #0f1b33;
                color: #eaf2ff;
                border: 1px solid rgba(255,255,255,0.22);
                selection-background-color: #2563eb;
                selection-color: #ffffff;
                outline: 0;
            }
            QLineEdit:focus, QComboBox:focus {
                border: 1px solid #38bdf8;
                background: rgba(56,189,248,0.10);
            }
            QPushButton {
                border: none;
                border-radius: 10px;
                padding: 8px 14px;
                font-weight: 700;
                color: #052234;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #67e8f9, stop:1 #2dd4bf);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #a5f3fc, stop:1 #5eead4);
            }
            QPushButton#Secondary {
                color: #dbecff;
                border: 1px solid rgba(255,255,255,0.25);
                background: rgba(255,255,255,0.10);
            }
            QPushButton#Secondary:hover {
                background: rgba(255,255,255,0.18);
            }
            QPushButton#Primary {
                padding: 10px 16px;
                min-height: 40px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #38bdf8, stop:1 #2dd4bf);
            }
            QLabel#StatusPill {
                background: rgba(56,189,248,0.16);
                border: 1px solid rgba(56,189,248,0.45);
                border-radius: 11px;
                color: #cffafe;
                font-weight: 700;
                padding: 7px 12px;
            }
            QProgressBar {
                min-height: 8px;
                max-height: 8px;
                border: 1px solid rgba(255,255,255,0.20);
                border-radius: 4px;
                background: rgba(255,255,255,0.08);
                text-align: center;
            }
            QProgressBar::chunk {
                border-radius: 3px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #22d3ee, stop:1 #34d399);
            }
            QLabel#StepBadge {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #38bdf8, stop:1 #2dd4bf);
                color: #052234;
                font-weight: 800;
                border-radius: 12px;
            }
            QLabel#StepTitle {
                font-size: 13px;
                font-weight: 700;
                color: #f2f7ff;
            }
            QLabel#StepHint {
                font-size: 10px;
                color: #9fb3d1;
            }
            QLabel#StepStatusPending {
                color: #9fb3d1;
                font-size: 10px;
                font-weight: 700;
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.14);
                border-radius: 9px;
                padding: 3px 10px;
            }
            QLabel#StepStatusDone {
                color: #052234;
                font-size: 10px;
                font-weight: 700;
                background: #6ee7b7;
                border: 1px solid rgba(255,255,255,0.14);
                border-radius: 9px;
                padding: 3px 10px;
            }
            """
        )

        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(18, 16, 18, 16)
        root.setSpacing(12)

        hero = QFrame(objectName="HeroCard")
        hero_lay = QVBoxLayout(hero)
        hero_lay.setContentsMargins(18, 14, 18, 14)
        hero_lay.setSpacing(3)
        title = QLabel("VERBATIM + CODE Mapper")
        title.setObjectName("Title")
        subtitle = QLabel("Map multi-sheet verbatim/code results into raw workbook with merge control and summary report.")
        subtitle.setObjectName("SubTitle")
        subtitle.setWordWrap(True)
        hero_lay.addWidget(title)
        hero_lay.addWidget(subtitle)
        root.addWidget(hero)

        # Step 1: Raw workbook
        raw_card, raw_body, self.raw_status_lbl = self._make_step_card(
            "1", "Raw Workbook", "Choose the Raw Excel file and the sheet you want to update."
        )
        r1 = QHBoxLayout()
        self.raw_edit = QLineEdit()
        self.raw_edit.setPlaceholderText("Raw file (.xlsx)")
        self.raw_edit.setToolTip("Path to the Raw data workbook. Its data will be updated and saved to a new output file.")
        b_raw = QPushButton("Browse...")
        b_raw.setObjectName("Secondary")
        b_raw.setToolTip("Browse for the Raw Excel file")
        b_raw.clicked.connect(self.pick_raw)
        r1.addWidget(self.raw_edit, 1)
        r1.addWidget(b_raw)
        raw_body.addLayout(r1)

        r2 = QHBoxLayout()
        lb2 = QLabel("Sheet:")
        lb2.setObjectName("FieldLabel")
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(260)
        self.sheet_combo.setToolTip("Sheet in the Raw file that contains the data rows (e.g. Diary)")
        self.sheet_combo.currentTextChanged.connect(self._update_step_statuses)
        r2.addWidget(lb2)
        r2.addWidget(self.sheet_combo)
        r2.addStretch(1)
        raw_body.addLayout(r2)
        root.addWidget(raw_card)

        # Step 2: Text source
        text_card, text_body, self.text_status_lbl = self._make_step_card(
            "2", "Text Source", "Choose the Text workbook and the sheet(s) that hold verbatim/code answers."
        )
        r3 = QHBoxLayout()
        self.text_edit = QLineEdit()
        self.text_edit.setPlaceholderText("Text file (.xlsx, many sheets)")
        self.text_edit.setToolTip("Path to the Text workbook containing verbatim answers and codes (one sheet per question or wave).")
        b_text = QPushButton("Browse...")
        b_text.setObjectName("Secondary")
        b_text.setToolTip("Browse for the Text Excel file")
        b_text.clicked.connect(self.pick_text)
        r3.addWidget(self.text_edit, 1)
        r3.addWidget(b_text)
        text_body.addLayout(r3)

        r3b = QHBoxLayout()
        self.text_sheets_edit = QLineEdit()
        self.text_sheets_edit.setReadOnly(True)
        self.text_sheets_edit.setPlaceholderText("Select one or more sheets from Text file")
        self.text_sheets_edit.setToolTip("Sheets from the Text file that will be scanned. All sheets are selected by default.")
        b_pick_sheets = QPushButton("Select Sheets...")
        b_pick_sheets.setObjectName("Secondary")
        b_pick_sheets.setToolTip("Choose which sheets of the Text file to include")
        b_pick_sheets.clicked.connect(self.pick_text_sheets)
        r3b.addWidget(self.text_sheets_edit, 1)
        r3b.addWidget(b_pick_sheets)
        text_body.addLayout(r3b)
        root.addWidget(text_card)

        # Step 3: Questions / items
        q_card, q_body, self.questions_status_lbl = self._make_step_card(
            "3",
            "Questions to Merge",
            "Pick questions from Text and map each to its Raw column (names can differ), "
            "type variable names manually, or load a Settings file for full control.",
        )
        r_pick = QHBoxLayout()
        self.b_pick_questions = QPushButton("Pick && Map Questions...")
        self.b_pick_questions.setObjectName("Primary")
        self.b_pick_questions.setToolTip(
            "Scan the selected Text sheet(s) for question columns, then map each one to its\n"
            "matching column in the Raw sheet (the two files don't need to use the same names)."
        )
        self.b_pick_questions.clicked.connect(self.pick_questions_from_files)
        r_pick.addWidget(self.b_pick_questions, 1)
        q_body.addLayout(r_pick)

        manual_hint = QLabel("or type variable names manually:")
        manual_hint.setObjectName("StepHint")
        q_body.addWidget(manual_hint)

        r4 = QHBoxLayout()
        self.var_edit = QLineEdit()
        self.var_edit.setPlaceholderText("e.g. s14b, s15, s16")
        self.var_edit.setToolTip("Comma-separated question variable names to merge. Ignored once questions are picked/loaded.")
        self.var_edit.textChanged.connect(self._update_step_statuses)
        r4.addWidget(self.var_edit, 1)
        q_body.addLayout(r4)

        r4btn = QHBoxLayout()
        self.b_load_settings = QPushButton("Load Settings...")
        self.b_load_settings.setObjectName("Secondary")
        self.b_load_settings.setToolTip("Load Item/Merge configuration from an Excel settings file")
        self.b_load_settings.clicked.connect(self.load_settings_items)
        r4btn.addWidget(self.b_load_settings)

        self.b_load_settings_sheet = QPushButton("Settings Sheet...")
        self.b_load_settings_sheet.setObjectName("Secondary")
        self.b_load_settings_sheet.setToolTip("Choose which sheet(s) of the settings file to read")
        self.b_load_settings_sheet.clicked.connect(self.pick_settings_sheets)
        r4btn.addWidget(self.b_load_settings_sheet)

        self.b_edit_items = QPushButton("Select Items/Merge...")
        self.b_edit_items.setObjectName("Secondary")
        self.b_edit_items.setToolTip("Review and edit which items to use, and which to merge")
        self.b_edit_items.clicked.connect(self.edit_settings_items)
        r4btn.addWidget(self.b_edit_items)

        self.b_save_settings = QPushButton("Save Settings...")
        self.b_save_settings.setObjectName("Secondary")
        self.b_save_settings.setToolTip("Save the current Item/Merge selection to a settings file")
        self.b_save_settings.clicked.connect(self.save_settings_items_as)
        r4btn.addWidget(self.b_save_settings)

        self.b_clear_settings = QPushButton("Clear")
        self.b_clear_settings.setObjectName("Secondary")
        self.b_clear_settings.setToolTip("Discard the loaded settings and switch back to manual entry")
        self.b_clear_settings.clicked.connect(self.clear_settings_items)
        r4btn.addWidget(self.b_clear_settings)
        r4btn.addStretch(1)
        q_body.addLayout(r4btn)
        root.addWidget(q_card)

        # Step 4: Merge mode & run
        run_card, run_body, _ = self._make_step_card(
            "4", "Merge Mode & Run", "Choose how per-item Merge flags are applied, then run and save.", show_status=False
        )
        r4b = QHBoxLayout()
        lb5 = QLabel("Merge Mode:")
        lb5.setObjectName("FieldLabel")
        self.merge_mode_combo = QComboBox()
        self.merge_mode_combo.addItem("Use Settings (Recommended)", userData="settings")
        self.merge_mode_combo.addItem("Force Merge All", userData="all")
        self.merge_mode_combo.addItem("No Merge", userData="none")
        self.merge_mode_combo.setToolTip(
            "Use Settings: apply each item's own Merge flag.\n"
            "Force Merge All: merge every item.\n"
            "No Merge: keep every item separate."
        )
        r4b.addWidget(lb5)
        r4b.addWidget(self.merge_mode_combo)
        r4b.addStretch(1)
        run_body.addLayout(r4b)

        self.status = QLabel("Ready")
        self.status.setObjectName("StatusPill")
        run_body.addWidget(self.status)
        self.progress = QProgressBar()
        self.progress.setTextVisible(False)
        self.progress.setRange(0, 1)
        self.progress.setValue(0)
        run_body.addWidget(self.progress)

        self.run_btn = QPushButton("Run and Save")
        self.run_btn.setObjectName("Primary")
        self.run_btn.setToolTip("Run the merge and save the result to a new Excel file")
        self.run_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton))
        self.run_btn.clicked.connect(self.run)
        run_body.addWidget(self.run_btn, alignment=Qt.AlignmentFlag.AlignRight)
        root.addWidget(run_card)

        self._busy_timer = QTimer(self)
        self._busy_timer.setInterval(500)
        self._busy_timer.timeout.connect(self._update_busy_elapsed)
        self._elapsed = QElapsedTimer()
        self._busy_base_text = "Processing..."

        self._controls_to_disable = [
            self.raw_edit,
            self.sheet_combo,
            self.text_edit,
            self.text_sheets_edit,
            self.var_edit,
            self.merge_mode_combo,
            self.run_btn,
            self.b_pick_questions,
            self.b_load_settings,
            self.b_load_settings_sheet,
            self.b_edit_items,
            self.b_save_settings,
            self.b_clear_settings,
        ]

        root.addStretch(1)
        self._update_settings_buttons_enabled()
        self._update_step_statuses()

    def _make_step_card(
        self, number: str, title: str, hint: str = "", show_status: bool = True
    ) -> tuple[QFrame, QVBoxLayout, QLabel | None]:
        card = QFrame(objectName="Card")
        outer = QVBoxLayout(card)
        outer.setContentsMargins(14, 12, 14, 12)
        outer.setSpacing(8)

        header = QHBoxLayout()
        badge = QLabel(number)
        badge.setObjectName("StepBadge")
        badge.setFixedSize(24, 24)
        badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header.addWidget(badge)

        title_box = QVBoxLayout()
        title_box.setSpacing(0)
        title_lbl = QLabel(title)
        title_lbl.setObjectName("StepTitle")
        title_box.addWidget(title_lbl)
        if hint:
            hint_lbl = QLabel(hint)
            hint_lbl.setObjectName("StepHint")
            hint_lbl.setWordWrap(True)
            title_box.addWidget(hint_lbl)
        header.addLayout(title_box, 1)

        status_lbl = None
        if show_status:
            status_lbl = QLabel("Not selected")
            status_lbl.setObjectName("StepStatusPending")
            header.addWidget(status_lbl, alignment=Qt.AlignmentFlag.AlignTop)

        outer.addLayout(header)
        body = QVBoxLayout()
        body.setSpacing(8)
        outer.addLayout(body)
        return card, body, status_lbl

    def _set_step_status(self, label: QLabel | None, done: bool, text: str):
        if label is None:
            return
        label.setText(text)
        label.setObjectName("StepStatusDone" if done else "StepStatusPending")
        label.style().unpolish(label)
        label.style().polish(label)

    def _update_step_statuses(self):
        raw_ready = bool(self.raw_path and txt(self.sheet_combo.currentText()))
        text_ready = bool(self.text_path and self.selected_text_sheets)

        if raw_ready:
            self._set_step_status(self.raw_status_lbl, True, f"Loaded: {os.path.basename(self.raw_path)}")
        else:
            self._set_step_status(self.raw_status_lbl, False, "Not selected")

        if text_ready:
            self._set_step_status(self.text_status_lbl, True, f"{len(self.selected_text_sheets)} sheet(s) selected")
        else:
            self._set_step_status(self.text_status_lbl, False, "Not selected")

        self.b_pick_questions.setEnabled(raw_ready and text_ready)
        self.b_pick_questions.setToolTip(
            "Scan the selected Text sheet(s) for question columns, then map each one to its\n"
            "matching column in the Raw sheet (the two files don't need to use the same names)."
            if raw_ready and text_ready
            else "Complete Step 1 (Raw sheet) and Step 2 (Text sheets) first."
        )

        if self.settings_items:
            enabled_count = sum(1 for it in self.settings_items if bool(it.get("enabled", True)))
            merge_count = sum(
                1 for it in self.settings_items if bool(it.get("enabled", True)) and bool(it.get("merge", False))
            )
            self._set_step_status(self.questions_status_lbl, True, f"{enabled_count} item(s), {merge_count} to merge")
        elif txt(self.var_edit.text()):
            names = parse_var_list(self.var_edit.text())
            self._set_step_status(self.questions_status_lbl, True, f"{len(names)} item(s) (manual)")
        else:
            self._set_step_status(self.questions_status_lbl, False, "Not selected")

    def _update_settings_buttons_enabled(self):
        has_settings = bool(self.settings_items)
        self.b_save_settings.setEnabled(has_settings)
        self.b_edit_items.setEnabled(has_settings)
        self.b_clear_settings.setEnabled(has_settings)

    def clear_settings_items(self):
        if not self.settings_items:
            return
        self.settings_items = []
        self.var_merge_map = {}
        self.settings_path = ""
        self.settings_sheet_names = []
        self.selected_settings_sheets = []
        self.var_edit.setReadOnly(False)
        self.var_edit.clear()
        self._update_settings_buttons_enabled()
        self._update_step_statuses()
        self.status.setText("Settings cleared - manual entry mode")

    def pick_questions_from_files(self):
        raw = txt(self.raw_edit.text())
        raw_sheet = txt(self.sheet_combo.currentText())
        textf = txt(self.text_edit.text())
        text_sheets = list(self.selected_text_sheets)

        if not raw or not os.path.exists(raw) or not raw_sheet:
            self.err("Error", "Please select the Raw file and sheet first (Step 1)")
            return
        if not textf or not os.path.exists(textf) or not text_sheets:
            self.err("Error", "Please select the Text file and sheet(s) first (Step 2)")
            return

        self._set_busy(True, "Scanning files...")
        self._scan_worker = QuestionScanWorker(raw, raw_sheet, textf, text_sheets, self)
        self._scan_worker.scanned.connect(self._on_questions_scanned)
        self._scan_worker.failed.connect(self._on_questions_scan_failed)
        self._scan_worker.finished.connect(self._scan_worker.deleteLater)
        self._scan_worker.start()

    def _on_questions_scan_failed(self, message: str):
        self._set_busy(False)
        self._scan_worker = None
        self.err("Error", f"Cannot scan the Raw/Text files:\n\n{message}")

    def _on_questions_scanned(self, raw_map: dict, candidates: list):
        self._set_busy(False)
        self._scan_worker = None

        if not candidates:
            self.err(
                "No Questions Found",
                "No open-end/code question columns were detected in the selected Text sheet(s).",
            )
            return

        items = self._choose_questions_dialog(candidates, raw_map)
        if items is None:
            return

        self.settings_items = items
        self.var_merge_map = {norm(it["name"]): bool(it.get("merge", False)) for it in items}
        self.settings_path = ""
        self.settings_sheet_names = []
        self.selected_settings_sheets = []
        self.refresh_selected_questions_text()

        enabled_count = sum(1 for it in items if bool(it.get("enabled", True)))
        merge_count = sum(1 for it in items if bool(it.get("enabled", True)) and bool(it.get("merge", False)))
        mapped_count = sum(1 for it in items if bool(it.get("enabled", True)) and txt(it.get("raw_name")))
        self.status.setText(f"Picked {enabled_count} question(s) from files")
        self.info(
            "Questions Picked",
            f"Found {len(candidates)} question column(s) in the Text file(s).\n"
            f"Selected: {enabled_count} to use ({mapped_count} mapped to a Raw column), {merge_count} to merge.",
        )

    def _choose_questions_dialog(self, candidates: list[str], raw_map: dict[str, str]) -> list[dict] | None:
        # Row-by-row mapping table: starts empty and the user adds one row per
        # question. Only the added rows have widgets (2 searchable dropdowns each),
        # so the dialog opens instantly no matter how many variables the files have.
        raw_options = sorted(raw_map.values(), key=str.lower)
        candidate_set = set(candidates)
        raw_set = set(raw_options)

        dlg = QDialog(self)
        dlg.setWindowTitle("Map Questions (OE Edit -> Raw)")
        dlg.resize(860, 560)
        self._apply_dialog_style(dlg)
        lay = QVBoxLayout(dlg)
        lay.addWidget(
            QLabel(
                f"OE Edit file: {len(candidates)} question column(s)  |  Raw sheet: {len(raw_options)} column(s).\n"
                "Add a row per question, then pick the OE Edit variable and its Raw column in each dropdown\n"
                "(type in the dropdown to search)."
            )
        )

        tbl = QTableWidget(0, 3)
        tbl.setHorizontalHeaderLabels(["OE Edit", "Raw Column", "Merge"])
        tbl.verticalHeader().setVisible(False)
        tbl.setColumnWidth(0, 340)
        tbl.setColumnWidth(1, 340)
        tbl.setColumnWidth(2, 70)
        tbl.setMinimumHeight(340)
        lay.addWidget(tbl)

        def make_combo(options: list[str]) -> QComboBox:
            combo = QComboBox()
            combo.setEditable(True)
            combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
            combo.addItem("")
            combo.addItems(options)
            comp = QCompleter(options, combo)
            comp.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
            comp.setFilterMode(Qt.MatchFlag.MatchContains)
            comp.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
            combo.setCompleter(comp)
            return combo

        def add_row():
            r = tbl.rowCount()
            tbl.insertRow(r)

            oe_combo = make_combo(candidates)
            raw_combo = make_combo(raw_options)
            tbl.setCellWidget(r, 0, oe_combo)
            tbl.setCellWidget(r, 1, raw_combo)

            merge_item = QTableWidgetItem("")
            merge_item.setFlags(merge_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            merge_item.setCheckState(Qt.CheckState.Unchecked)
            tbl.setItem(r, 2, merge_item)

            def on_oe_changed(name: str, row_combo=oe_combo):
                # Suggest Merge automatically for split-style names like s15_1.
                row = None
                for rr in range(tbl.rowCount()):
                    if tbl.cellWidget(rr, 0) is row_combo:
                        row = rr
                        break
                if row is None:
                    return
                m_item = tbl.item(row, 2)
                if m_item is not None and name in candidate_set:
                    m_item.setCheckState(
                        Qt.CheckState.Checked if merge_base_name(name) != name else Qt.CheckState.Unchecked
                    )

            oe_combo.currentTextChanged.connect(on_oe_changed)
            tbl.setCurrentCell(r, 0)
            oe_combo.setFocus()

        def remove_row():
            r = tbl.currentRow()
            if r < 0:
                self.info("Select a row", "Click a row first, then press Remove Row.")
                return
            tbl.removeRow(r)

        btn_row = QHBoxLayout()
        b_add = QPushButton("+ Add Row")
        b_add.clicked.connect(add_row)
        b_del = QPushButton("Remove Row")
        b_del.setObjectName("Secondary")
        b_del.clicked.connect(remove_row)
        btn_row.addWidget(b_add)
        btn_row.addWidget(b_del)
        btn_row.addStretch(1)
        lay.addLayout(btn_row)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)

        add_row()

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return None

        items = []
        seen_names = set()
        invalid_rows = 0
        for r in range(tbl.rowCount()):
            oe_combo = tbl.cellWidget(r, 0)
            raw_combo = tbl.cellWidget(r, 1)
            merge_item = tbl.item(r, 2)
            if oe_combo is None or raw_combo is None:
                continue
            oe_name = txt(oe_combo.currentText())
            raw_name = txt(raw_combo.currentText())
            if oe_name == "" and raw_name == "":
                continue
            if oe_name not in candidate_set or raw_name not in raw_set:
                invalid_rows += 1
                continue
            key = norm(oe_name)
            if key in seen_names:
                continue
            seen_names.add(key)
            items.append(
                {
                    "name": oe_name,
                    "raw_name": raw_name,
                    "enabled": True,
                    "merge": merge_item is not None and merge_item.checkState() == Qt.CheckState.Checked,
                }
            )

        if not items:
            self.err("No Mappings", "No valid OE Edit -> Raw pairs were selected.")
            return None
        if invalid_rows:
            self.info(
                "Some rows skipped",
                f"{invalid_rows} row(s) had a value that is not in the file's variable list and were skipped.",
            )
        return items

    def _apply_dialog_style(self, dlg: QDialog):
        dlg.setStyleSheet(
            """
            QDialog {
                background: #0f1b33;
                color: #e7eefc;
            }
            QLabel {
                color: #dbeafe;
            }
            QListWidget, QTableWidget {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.18);
                border-radius: 10px;
                color: #f2f7ff;
                gridline-color: rgba(255,255,255,0.15);
            }
            QHeaderView::section {
                background: rgba(56,189,248,0.20);
                color: #dbeafe;
                border: none;
                padding: 5px;
                font-weight: 700;
            }
            QPushButton {
                border: none;
                border-radius: 9px;
                padding: 7px 12px;
                font-weight: 700;
                color: #052234;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #67e8f9, stop:1 #2dd4bf);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #a5f3fc, stop:1 #5eead4);
            }
            QLineEdit {
                background: rgba(255,255,255,0.08);
                border: 1px solid rgba(255,255,255,0.22);
                border-radius: 8px;
                padding: 6px 10px;
                color: #f2f7ff;
            }
            QLineEdit:focus {
                border: 1px solid #38bdf8;
            }
            """
        )

    def info(self, title: str, text: str):
        self._show_message("info", title, text)

    def err(self, title: str, text: str):
        self._show_message("error", title, text)

    def _show_message(self, kind: str, title: str, text: str):
        msg = QMessageBox(self)
        if kind == "error":
            msg.setIcon(QMessageBox.Icon.Critical)
        else:
            msg.setIcon(QMessageBox.Icon.Information)
        msg.setWindowTitle(title)
        msg.setText(text)
        msg.setStyleSheet(
            """
            QMessageBox {
                background: #f8fafc;
            }
            QMessageBox QLabel {
                color: #0f172a;
                min-width: 360px;
                font-size: 11pt;
            }
            QMessageBox QPushButton {
                min-width: 88px;
                border: 1px solid #93c5fd;
                border-radius: 8px;
                padding: 6px 12px;
                color: #0f172a;
                background: #dbeafe;
                font-weight: 700;
            }
            QMessageBox QPushButton:hover {
                background: #bfdbfe;
            }
            """
        )
        msg.exec()

    def _set_busy(self, busy: bool, text: str = "Processing..."):
        if busy:
            self._busy_base_text = text
            self._elapsed.start()
            self._busy_timer.start()
            self.progress.setRange(0, 0)
            self.run_btn.setText("Running...")
            for w in self._controls_to_disable:
                w.setEnabled(False)
            self._update_busy_elapsed()
            QApplication.processEvents()
            return

        self._busy_timer.stop()
        self.progress.setRange(0, 1)
        self.run_btn.setText("Run and Save")
        for w in self._controls_to_disable:
            w.setEnabled(True)
        self._update_settings_buttons_enabled()
        self._update_step_statuses()

    def _update_busy_elapsed(self):
        secs = self._elapsed.elapsed() // 1000 if self._elapsed.isValid() else 0
        self.status.setText(f"{self._busy_base_text}  ({secs}s)")

    def pick_raw(self):
        p, _ = QFileDialog.getOpenFileName(self, "Select Raw file", self._last_dir, "Excel files (*.xlsx *.xls)")
        if not p:
            return
        self.raw_path = p
        self._last_dir = os.path.dirname(p)
        self.raw_edit.setText(p)
        try:
            xls = pd.ExcelFile(p)
            self.sheet_combo.clear()
            self.sheet_combo.addItems(xls.sheet_names)
            if "Diary" in xls.sheet_names:
                self.sheet_combo.setCurrentText("Diary")
        except Exception as e:
            self.sheet_combo.clear()
            self.err("Error", f"Cannot read sheets from Raw: {e}")
        self._update_step_statuses()

    def pick_text(self):
        p, _ = QFileDialog.getOpenFileName(self, "Select Text file", self._last_dir, "Excel files (*.xlsx *.xls)")
        if not p:
            return
        self.text_path = p
        self._last_dir = os.path.dirname(p)
        self.text_edit.setText(p)
        try:
            xls = pd.ExcelFile(p)
            self.text_sheet_names = list(xls.sheet_names)
            self.selected_text_sheets = list(self.text_sheet_names)
            self.text_sheets_edit.setText(", ".join(self.selected_text_sheets))
        except Exception as e:
            self.text_sheet_names = []
            self.selected_text_sheets = []
            self.text_sheets_edit.clear()
            self.err("Error", f"Cannot read sheets from Text: {e}")
        self._update_step_statuses()

    def pick_text_sheets(self):
        if not self.text_sheet_names:
            self.err("Error", "Please select Text file first")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Select Text Sheets")
        dlg.resize(420, 520)
        self._apply_dialog_style(dlg)
        lay = QVBoxLayout(dlg)

        search_edit = QLineEdit()
        search_edit.setPlaceholderText("Search sheets...")
        search_edit.setClearButtonEnabled(True)
        lay.addWidget(search_edit)

        lw = QListWidget()
        selected_set = set(self.selected_text_sheets) if self.selected_text_sheets else set(self.text_sheet_names)
        for name in self.text_sheet_names:
            item = QListWidgetItem(name)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Checked if name in selected_set else Qt.CheckState.Unchecked)
            lw.addItem(item)
        lay.addWidget(lw)

        def filter_rows(text: str):
            needle = text.strip().lower()
            for i in range(lw.count()):
                item = lw.item(i)
                item.setHidden(needle not in item.text().lower())

        search_edit.textChanged.connect(filter_rows)

        btn_row = QHBoxLayout()
        b_all = QPushButton("Select All")
        b_none = QPushButton("Clear All")
        b_all.clicked.connect(lambda: [lw.item(i).setCheckState(Qt.CheckState.Checked) for i in range(lw.count())])
        b_none.clicked.connect(lambda: [lw.item(i).setCheckState(Qt.CheckState.Unchecked) for i in range(lw.count())])
        btn_row.addWidget(b_all)
        btn_row.addWidget(b_none)
        btn_row.addStretch(1)
        lay.addLayout(btn_row)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        picked = []
        for i in range(lw.count()):
            item = lw.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                picked.append(item.text())

        self.selected_text_sheets = picked
        self.text_sheets_edit.setText(", ".join(picked) if picked else "(none selected)")
        self._update_step_statuses()

    def run(self):
        raw = txt(self.raw_edit.text())
        textf = txt(self.text_edit.text())
        text_sheets = list(self.selected_text_sheets)
        selected_items = []
        if self.settings_items:
            selected_items = [it for it in self.settings_items if bool(it.get("enabled", True))]
            var_names = [txt(it.get("name")) for it in selected_items if txt(it.get("name")) != ""]
        else:
            var_names = parse_var_list(self.var_edit.text())
        mode = self.merge_mode_combo.currentData()
        if self.settings_items:
            if mode == "all":
                var_configs = [
                    {"name": txt(it["name"]), "raw_name": txt(it.get("raw_name", "")), "merge": True}
                    for it in selected_items
                ]
            elif mode == "none":
                var_configs = [
                    {"name": txt(it["name"]), "raw_name": txt(it.get("raw_name", "")), "merge": False}
                    for it in selected_items
                ]
            else:
                var_configs = [
                    {
                        "name": txt(it["name"]),
                        "raw_name": txt(it.get("raw_name", "")),
                        "merge": bool(it.get("merge", False)),
                    }
                    for it in selected_items
                ]
        else:
            if mode == "all":
                var_configs = [{"name": v, "merge": True} for v in var_names]
            elif mode == "none":
                var_configs = [{"name": v, "merge": False} for v in var_names]
            else:
                var_configs = [{"name": v, "merge": self.var_merge_map.get(norm(v), False)} for v in var_names]
        merge_count = sum(1 for cfg in var_configs if bool(cfg.get("merge", False)))
        if self.settings_items and mode == "none":
            selected_merge_count = sum(1 for it in selected_items if bool(it.get("merge", False)))
            if selected_merge_count > 0:
                ans = QMessageBox.question(
                    self,
                    "Confirm No Merge",
                    f"You selected No Merge, which will ignore {selected_merge_count} checked Merge items.\nContinue?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No,
                )
                if ans != QMessageBox.StandardButton.Yes:
                    return
        sheet = txt(self.sheet_combo.currentText())
        if not raw or not os.path.exists(raw):
            self.err("Error", "Please select a valid Raw file")
            return
        if not textf or not os.path.exists(textf):
            self.err("Error", "Please select a valid Text file")
            return
        if not text_sheets:
            self.err("Error", "Please select at least one Text sheet")
            return
        if not var_names:
            self.err("Error", "Please select at least one question item")
            return
        if not sheet:
            self.err("Error", "Please select Raw sheet")
            return

        default_name = f"{os.path.splitext(os.path.basename(raw))[0]}_VERBATIM_MAPPED.xlsx"
        initial_path = os.path.join(self._last_dir, default_name) if self._last_dir else default_name
        out, _ = QFileDialog.getSaveFileName(
            self,
            "Save output file",
            initial_path,
            "Excel files (*.xlsx)",
        )
        if not out:
            return
        self._last_dir = os.path.dirname(out)

        self._set_busy(True, "Processing...")
        try:
            res = run_merge(raw, sheet, textf, var_configs, out, text_sheets)
            self._save_settings_items_to_file()
            has_fail = any(d.get("status") != "SUCCESS" for d in res.get("details", []))
            if has_fail:
                self.info("Completed", f"Completed (some items failed - check Summary sheet)\nMerge items used: {merge_count}/{len(var_configs)}")
            else:
                self.info("Completed", f"Completed\nMerge items used: {merge_count}/{len(var_configs)}")
            self.status.setText("Completed")
            self.progress.setRange(0, 1)
            self.progress.setValue(1)
        except Exception as e:
            self.status.setText("Error")
            traceback.print_exc()
            self.err("Error", f"Failed to run merge:\n\n{e}")
            self.progress.setRange(0, 1)
            self.progress.setValue(0)
        finally:
            self._set_busy(False)

    def load_settings_items(self):
        p, _ = QFileDialog.getOpenFileName(
            self,
            "Select settings file",
            self._last_dir,
            "Excel files (*.xlsx *.xls)",
        )
        if not p:
            return
        try:
            xls = pd.ExcelFile(p)
            self.settings_path = p
            self._last_dir = os.path.dirname(p)
            self.settings_sheet_names = list(xls.sheet_names)
            self.selected_settings_sheets = self._choose_settings_sheets(self.settings_sheet_names)
            if not self.selected_settings_sheets:
                return
            self._load_settings_from_path(p, self.selected_settings_sheets)
        except Exception as e:
            self.err("Error", f"Cannot load settings: {e}")

    def pick_settings_sheets(self):
        if not self.settings_path:
            p, _ = QFileDialog.getOpenFileName(
                self,
                "Select settings file",
                self._last_dir,
                "Excel files (*.xlsx *.xls)",
            )
            if not p:
                return
            self.settings_path = p
            self._last_dir = os.path.dirname(p)

        try:
            xls = pd.ExcelFile(self.settings_path)
            self.settings_sheet_names = list(xls.sheet_names)
        except Exception as e:
            self.err("Error", f"Cannot read settings sheets: {e}")
            return

        picked = self._choose_settings_sheets(self.settings_sheet_names)
        if not picked:
            return
        self.selected_settings_sheets = picked
        try:
            self._load_settings_from_path(self.settings_path, self.selected_settings_sheets)
        except Exception as e:
            self.err("Error", f"Cannot load settings: {e}")

    def _choose_settings_sheets(self, sheet_names: list[str]) -> list[str]:
        if not sheet_names:
            self.err("Error", "No sheets found in settings file")
            return []

        dlg = QDialog(self)
        dlg.setWindowTitle("Select Settings Sheet")
        dlg.resize(420, 480)
        self._apply_dialog_style(dlg)
        lay = QVBoxLayout(dlg)
        lay.addWidget(QLabel("Choose settings sheet(s) to load:"))

        lw = QListWidget()
        selected_set = set(self.selected_settings_sheets) if self.selected_settings_sheets else {sheet_names[0]}
        for name in sheet_names:
            item = QListWidgetItem(name)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Checked if name in selected_set else Qt.CheckState.Unchecked)
            lw.addItem(item)
        lay.addWidget(lw)

        btn_row = QHBoxLayout()
        b_all = QPushButton("Select All")
        b_none = QPushButton("Clear All")
        b_all.clicked.connect(lambda: [lw.item(i).setCheckState(Qt.CheckState.Checked) for i in range(lw.count())])
        b_none.clicked.connect(lambda: [lw.item(i).setCheckState(Qt.CheckState.Unchecked) for i in range(lw.count())])
        btn_row.addWidget(b_all)
        btn_row.addWidget(b_none)
        btn_row.addStretch(1)
        lay.addLayout(btn_row)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return []

        picked = []
        for i in range(lw.count()):
            item = lw.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                picked.append(item.text())

        if not picked:
            self.err("Error", "Please select at least one Settings sheet")
        return picked

    def _load_settings_from_path(self, path: str, sheets: list[str]):
        items, warnings = load_items_from_settings_file(path, sheets)
        self.settings_items = [
            {
                "name": txt(it["name"]),
                "raw_name": txt(it.get("raw_name", "")),
                "merge": bool(it.get("merge", False)),
                "enabled": bool(it.get("enabled", True)),
            }
            for it in items
        ]
        self.var_merge_map = {norm(it["name"]): bool(it.get("merge", False)) for it in items}
        self.refresh_selected_questions_text()

        merge_count = sum(1 for it in items if it.get("merge"))
        enabled_count = sum(1 for it in self.settings_items if it.get("enabled", True))
        mapped_count = sum(1 for it in self.settings_items if txt(it.get("raw_name", "")))
        sheet_text = ", ".join(sheets)
        self.status.setText(f"Loaded settings sheet: {sheet_text}")
        summary = (
            f"Loaded {len(items)} items from settings sheet(s): {sheet_text}\n"
            f"Use: {enabled_count} | Raw Column mapped: {mapped_count} | Merge: {merge_count}"
        )
        if warnings:
            msg = f"{summary}\n\nWarnings:\n"
            msg += "\n".join(f"- {w}" for w in warnings[:10])
            if len(warnings) > 10:
                msg += f"\n- ... and {len(warnings) - 10} more"
            self.info("Loaded with warnings", msg)
        else:
            self.info("Loaded", summary)

    def refresh_selected_questions_text(self):
        if self.settings_items:
            names = [txt(it["name"]) for it in self.settings_items if bool(it.get("enabled", True))]
            self.var_edit.setReadOnly(True)
            self.var_edit.setText(",".join(names))
        else:
            self.var_edit.setReadOnly(False)
        self._update_settings_buttons_enabled()
        self._update_step_statuses()

    def edit_settings_items(self):
        if not self.settings_items:
            self.err("Error", "Please load settings first")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Select Items and Merge")
        dlg.resize(880, 600)
        self._apply_dialog_style(dlg)
        lay = QVBoxLayout(dlg)
        lay.addWidget(
            QLabel(
                "Choose which items to use and merge. Edit Raw Column if the Raw file uses a\n"
                "different name than the Text file for that question (leave blank to use the same name)."
            )
        )

        search_edit = QLineEdit()
        search_edit.setPlaceholderText("Search items...")
        search_edit.setClearButtonEnabled(True)
        lay.addWidget(search_edit)

        tbl = QTableWidget(len(self.settings_items), 4)
        tbl.setHorizontalHeaderLabels(["Item (Text)", "Raw Column", "Use", "Merge"])
        tbl.verticalHeader().setVisible(False)
        tbl.setColumnWidth(0, 320)
        tbl.setColumnWidth(1, 320)
        tbl.setColumnWidth(2, 70)
        tbl.setColumnWidth(3, 70)
        tbl.setMinimumHeight(420)

        for r, it in enumerate(self.settings_items):
            name_item = QTableWidgetItem(txt(it.get("name")))
            name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            tbl.setItem(r, 0, name_item)

            raw_item = QTableWidgetItem(txt(it.get("raw_name")))
            raw_item.setToolTip("Leave blank to use the same name as the Text item.")
            tbl.setItem(r, 1, raw_item)

            use_item = QTableWidgetItem("")
            use_item.setFlags(use_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            use_item.setCheckState(Qt.CheckState.Checked if bool(it.get("enabled", True)) else Qt.CheckState.Unchecked)
            tbl.setItem(r, 2, use_item)

            merge_item = QTableWidgetItem("")
            merge_item.setFlags(merge_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            merge_item.setCheckState(Qt.CheckState.Checked if bool(it.get("merge", False)) else Qt.CheckState.Unchecked)
            tbl.setItem(r, 3, merge_item)

        lay.addWidget(tbl)

        def filter_rows(text: str):
            needle = text.strip().lower()
            for rr in range(tbl.rowCount()):
                name_cell = tbl.item(rr, 0)
                match = needle in name_cell.text().lower() if name_cell else True
                tbl.setRowHidden(rr, not match)

        search_edit.textChanged.connect(filter_rows)

        btn_row = QHBoxLayout()
        b_use_all = QPushButton("Use: Select All")
        b_use_none = QPushButton("Use: Clear All")
        b_merge_all = QPushButton("Merge: Select All")
        b_merge_none = QPushButton("Merge: Clear")

        def select_all_items():
            for rr in range(tbl.rowCount()):
                if tbl.item(rr, 2) is not None:
                    tbl.item(rr, 2).setCheckState(Qt.CheckState.Checked)

        def clear_all_items():
            for rr in range(tbl.rowCount()):
                if tbl.item(rr, 2) is not None:
                    tbl.item(rr, 2).setCheckState(Qt.CheckState.Unchecked)

        def select_all_merge():
            for rr in range(tbl.rowCount()):
                if tbl.item(rr, 3) is not None:
                    tbl.item(rr, 3).setCheckState(Qt.CheckState.Checked)

        def clear_all_merge():
            for rr in range(tbl.rowCount()):
                if tbl.item(rr, 3) is not None:
                    tbl.item(rr, 3).setCheckState(Qt.CheckState.Unchecked)

        b_use_all.clicked.connect(select_all_items)
        b_use_none.clicked.connect(clear_all_items)
        b_merge_all.clicked.connect(select_all_merge)
        b_merge_none.clicked.connect(clear_all_merge)
        btn_row.addWidget(b_use_all)
        btn_row.addWidget(b_use_none)
        btn_row.addWidget(b_merge_all)
        btn_row.addWidget(b_merge_none)
        btn_row.addStretch(1)
        lay.addLayout(btn_row)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        for r, it in enumerate(self.settings_items):
            raw_item = tbl.item(r, 1)
            use_item = tbl.item(r, 2)
            merge_item = tbl.item(r, 3)
            it["raw_name"] = txt(raw_item.text()) if raw_item is not None else ""
            it["enabled"] = use_item is not None and use_item.checkState() == Qt.CheckState.Checked
            it["merge"] = merge_item is not None and merge_item.checkState() == Qt.CheckState.Checked

        self.var_merge_map = {norm(it["name"]): bool(it.get("merge", False)) for it in self.settings_items}
        self.refresh_selected_questions_text()
        self._save_settings_items_to_file()

    def _save_settings_items_to_file(self, path: str | None = None, show_message: bool = False):
        if not self.settings_items:
            return
        target = txt(path) if path is not None else txt(self.settings_path)
        if target == "":
            return
        rows = []
        for it in self.settings_items:
            name = txt(it.get("name"))
            if name == "":
                continue
            rows.append(
                {
                    "Item": name,
                    "Raw Column": txt(it.get("raw_name", "")),
                    "Use": "1" if bool(it.get("enabled", True)) else "",
                    "Merge": "1" if bool(it.get("merge", False)) else "",
                }
            )
        out_df = pd.DataFrame(rows, columns=["Item", "Raw Column", "Use", "Merge"])
        out_df.to_excel(target, index=False, sheet_name="Sheet1")
        self.settings_path = target
        if show_message:
            self.info("Saved", f"Saved settings to:\n{target}")

    def save_settings_items_as(self):
        if not self.settings_items:
            self.err("Error", "Please load settings first")
            return
        default_name = os.path.basename(self.settings_path) if self.settings_path else "Seting OE.xlsx"
        initial_path = os.path.join(self._last_dir, default_name) if self._last_dir else default_name
        out, _ = QFileDialog.getSaveFileName(
            self,
            "Save settings file",
            initial_path,
            "Excel files (*.xlsx)",
        )
        if not out:
            return
        try:
            self._save_settings_items_to_file(path=out, show_message=True)
            self._last_dir = os.path.dirname(out)
        except Exception as e:
            self.err("Error", f"Cannot save settings: {e}")


def main():
    app = QApplication([])
    app.setWindowIcon(make_leaf_icon())
    win = VerbatimMapperApp()
    win.show()
    app.exec()


if __name__ == "__main__":
    main()



