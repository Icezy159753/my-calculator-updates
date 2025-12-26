import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox, font
import openpyxl
import os
import threading
from datetime import datetime

# --- ส่วนตรรกะการประมวลผลไฟล์ Excel (ไม่มีการเปลี่ยนแปลง) ---
def process_excel_files_inplace(file_paths, sheet_names_to_process, log_callback):
    """
    ฟังก์ชันหลักในการประมวลผลไฟล์ Excel โดยแก้ไขไฟล์เดิมโดยตรงและคง Format ไว้
    1. ลบแถวที่คอลัมน์ C เป็น 0 (ตั้งแต่แถวที่ 6 ลงไป) **ยกเว้นแถวสุดท้ายของชีท**
    2. จัดลำดับคอลัมน์ A ใหม่ โดยหา 0 แรกเจอแล้วเริ่มนับ 0, 1, 2... ใหม่จากจุดนั้น
    """
    all_missing_sheets_summary = set()

    try:
        for i, file_path in enumerate(file_paths):
            base_name = os.path.basename(file_path)
            log_callback(f"===== เริ่มประมวลผลไฟล์ {i+1}/{len(file_paths)}: {base_name} =====", "header")

            try:
                workbook = openpyxl.load_workbook(file_path)
            except Exception as e:
                log_callback(f"ไม่สามารถเปิดไฟล์ {base_name} ได้! อาจมีการป้องกันด้วยรหัสผ่าน หรือไฟล์เสียหาย.", "error")
                log_callback(f"ข้อผิดพลาด: {e}", "error")
                continue

            actual_sheet_names = workbook.sheetnames
            has_changes = False

            # ประมวลผลแต่ละชีทที่ระบุ
            for sheet_name in sheet_names_to_process:
                if sheet_name in actual_sheet_names:
                    sheet = workbook[sheet_name]
                    
                    rows_to_delete = []
                    if sheet.max_row > 6:
                        for row_index in range(sheet.max_row - 1, 5, -1): 
                            cell_value = sheet.cell(row=row_index, column=3).value
                            if cell_value == 0 or str(cell_value).strip() == '0':
                                rows_to_delete.append(row_index)
                    
                    if rows_to_delete:
                        has_changes = True
                        for row_idx in sorted(rows_to_delete, reverse=True):
                            sheet.delete_rows(row_idx)
                        log_callback(f"  [สำเร็จ] ชีท '{sheet_name}': ลบไป {len(rows_to_delete)} แถว (จากคอลัมน์ C)", "success")
                    else:
                        log_callback(f"  [ข้อมูล] ชีท '{sheet_name}': ไม่พบแถวที่เป็น 0 ในคอลัมน์ C (ไม่รวมแถวสุดท้าย)", "info")

                    start_renumber_row = -1
                    for row_idx in range(6, sheet.max_row + 1):
                        cell_a_value = sheet.cell(row=row_idx, column=1).value
                        if cell_a_value == 0 or str(cell_a_value).strip() == '0':
                            start_renumber_row = row_idx
                            log_callback(f"  [ข้อมูล] ชีท '{sheet_name}': พบค่า 0 ในคอลัมน์ A ที่แถว {row_idx}, เริ่มจัดลำดับใหม่", "info")
                            break
                    
                    if start_renumber_row != -1:
                        has_changes = True
                        counter = 0 
                        for row_to_update in range(start_renumber_row, sheet.max_row + 1):
                            sheet.cell(row=row_to_update, column=1).value = counter
                            counter += 1
                        log_callback(f"  [สำเร็จ] ชีท '{sheet_name}': จัดลำดับคอลัมน์ A ใหม่ (เริ่มจาก 0) จำนวน {counter} แถว", "success")
                    else:
                        log_callback(f"  [ข้อมูล] ชีท '{sheet_name}': ไม่พบค่า 0 ในคอลัมน์ A เพื่อเริ่มจัดลำดับใหม่", "info")
                else:
                    all_missing_sheets_summary.add(sheet_name)
                    log_callback(f"  [ไม่พบ] ชีท '{sheet_name}' ในไฟล์นี้", "error")
            
            if has_changes:
                log_callback(f"กำลังบันทึกการเปลี่ยนแปลงลงไฟล์ {base_name}...", "info")
                workbook.save(file_path)
                log_callback(f"บันทึกไฟล์ {base_name} เรียบร้อยแล้ว", "success_bold")
            else:
                log_callback(f"ไม่มีการเปลี่ยนแปลงในไฟล์ {base_name}, ไม่ต้องบันทึก", "info")
            
            log_callback("=" * 40, "header")

        if not all_missing_sheets_summary:
            final_message = "การประมวลผลเสร็จสมบูรณ์ทุกไฟล์!"
            log_callback(final_message, "success_bold")
        else:
            missing_sheets_str = ", ".join(sorted(list(all_missing_sheets_summary)))
            final_message = f"เสร็จสิ้น! แต่ไม่พบชีทเหล่านี้ในไฟล์ใดๆ เลย: {missing_sheets_str}"
            log_callback(final_message, "warning")
        
        messagebox.showinfo("เสร็จสิ้น", final_message)
    except Exception as e:
        final_message = f"เกิดข้อผิดพลาดร้ายแรงระหว่างการทำงาน: {e}"
        log_callback(final_message, "error")
        messagebox.showerror("ข้อผิดพลาดร้ายแรง", final_message)


# --- ส่วนหน้าจอโปรแกรม (GUI) - ปรับปรุงหน้าตาใหม่ทั้งหมด ---
class App:
    # ตั้งค่าสีและฟอนต์สำหรับ UI
    COLORS = {
        "bg": "#F5F5F5",
        "frame": "#FFFFFF",
        "text": "#212121",
        "primary": "#0078D4",
        "primary_hover": "#005A9E",
        "success": "#107C10",
        "success_hover": "#0F5B0F",
        "border": "#ACACAC", 
        "status_ok": "#107C10",
        "status_none": "#757575"
    }
    FONTS = {
        "title": ("Segoe UI", 16, "bold"),
        "header": ("Segoe UI", 11, "bold"),
        "body": ("Segoe UI", 10),
        "button": ("Segoe UI", 11, "bold"),
        "log": ("Consolas", 10)
    }

    def __init__(self, root):
        self.root = root
        self.root.title("โปรแกรมลบN=0 OEในLychee V1")
        self.root.configure(bg=self.COLORS["bg"])
        self.root.minsize(750, 600)

        # **การเปลี่ยนแปลง:** โค้ดสำหรับจัดหน้าต่างโปรแกรมให้อยู่กึ่งกลางจอ
        window_width = 750
        window_height = 700
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        position_x = int((screen_width / 2) - (window_width / 2))
        position_y = int((screen_height / 2) - (window_height / 2))
        self.root.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")


        self.file_paths = []

        # --- สร้างเฟรมหลัก ---
        main_frame = tk.Frame(root, bg=self.COLORS["bg"], padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        main_frame.grid_rowconfigure(4, weight=1) 
        main_frame.grid_columnconfigure(0, weight=1)

        # --- ส่วนหัวเรื่อง ---
        title_label = tk.Label(main_frame, text="โปรแกรมปรับปรุงไฟล์ Excel", 
                               font=self.FONTS["title"], bg=self.COLORS["bg"], fg=self.COLORS["text"])
        title_label.grid(row=0, column=0, sticky="w", pady=(0, 20))

        # --- ขั้นตอนที่ 1: เลือกไฟล์ ---
        step1_frame = tk.LabelFrame(main_frame, text=" ขั้นตอนที่ 1: เลือกไฟล์ Excel ", font=self.FONTS["header"], 
                                     bg=self.COLORS["frame"], fg=self.COLORS["text"], padx=15, pady=15,
                                     borderwidth=1, relief=tk.SOLID)
        step1_frame.grid(row=1, column=0, sticky="ew", pady=(0, 15))
        step1_frame.grid_columnconfigure(0, weight=1)

        self.btn_select_files = self.create_styled_button(step1_frame, "คลิกเพื่อเลือกไฟล์...", self.select_files, self.COLORS["primary"], self.COLORS["primary_hover"])
        self.btn_select_files.grid(row=0, column=0, sticky="ew")
        
        self.lbl_file_status = tk.Label(step1_frame, text="ยังไม่ได้เลือกไฟล์", font=self.FONTS["body"], 
                                        fg=self.COLORS["status_none"], bg=self.COLORS["frame"], wraplength=650, justify=tk.LEFT)
        self.lbl_file_status.grid(row=1, column=0, sticky="w", pady=(10, 0))

        # --- ขั้นตอนที่ 2: ระบุชื่อชีท ---
        step2_frame = tk.LabelFrame(main_frame, text=" ขั้นตอนที่ 2: ระบุชื่อชีท (หนึ่งชื่อต่อบรรทัด) ", font=self.FONTS["header"],
                                     bg=self.COLORS["frame"], fg=self.COLORS["text"], padx=15, pady=15,
                                     borderwidth=1, relief=tk.SOLID)
        step2_frame.grid(row=2, column=0, sticky="ew", pady=(0, 20))
        step2_frame.grid_columnconfigure(0, weight=1)

        self.sheet_entry = scrolledtext.ScrolledText(step2_frame, height=6, font=self.FONTS["body"],
                                                     borderwidth=1, relief=tk.SOLID)
        self.sheet_entry.pack(fill=tk.X, expand=True)
        self.sheet_entry.insert(tk.END, "Q2_OEZ1(1)\nQ2_OEZ1(2)")

        # --- ปุ่มเริ่มประมวลผล ---
        self.btn_process = self.create_styled_button(main_frame, "เริ่มประมวลผลและบันทึกทับไฟล์", self.start_processing_thread, self.COLORS["success"], self.COLORS["success_hover"])
        self.btn_process.grid(row=3, column=0, sticky="ew", pady=(0, 15))

        # --- ส่วนแสดง Log ---
        log_frame = tk.LabelFrame(main_frame, text=" หน้าต่างแสดงผลการทำงาน (Log) ", font=self.FONTS["header"],
                                  bg=self.COLORS["frame"], fg=self.COLORS["text"], padx=15, pady=10,
                                  borderwidth=1, relief=tk.SOLID)
        log_frame.grid(row=4, column=0, sticky="nsew")
        log_frame.grid_rowconfigure(0, weight=1)
        log_frame.grid_columnconfigure(0, weight=1)

        self.log_widget = scrolledtext.ScrolledText(log_frame, state='disabled', font=self.FONTS["log"], wrap=tk.WORD, 
                                                    bg=self.COLORS["frame"], borderwidth=0, highlightthickness=0)
        self.log_widget.grid(row=0, column=0, sticky="nsew")

        self.configure_log_tags()

    def create_styled_button(self, parent, text, command, bg_color, hover_color):
        """Helper function to create styled buttons with hover effects."""
        button = tk.Button(parent, text=text, command=command, font=self.FONTS["button"],
                           bg=bg_color, fg="#FFFFFF", relief=tk.FLAT,
                           activebackground=hover_color, activeforeground="#FFFFFF",
                           cursor="hand2", pady=8)
        
        button.bind("<Enter>", lambda e: button.config(bg=hover_color))
        button.bind("<Leave>", lambda e: button.config(bg=bg_color))
        return button

    def configure_log_tags(self):
        """ตั้งค่าสีสำหรับ Log Tags"""
        self.log_widget.tag_config('info', foreground='black')
        self.log_widget.tag_config('success', foreground=self.COLORS["success"])
        self.log_widget.tag_config('success_bold', foreground=self.COLORS["success"], font=(self.FONTS["log"][0], self.FONTS["log"][1], "bold"))
        self.log_widget.tag_config('error', foreground='red', font=(self.FONTS["log"][0], self.FONTS["log"][1], "bold"))
        self.log_widget.tag_config('warning', foreground='#E69138', font=(self.FONTS["log"][0], self.FONTS["log"][1], "bold"))
        self.log_widget.tag_config('header', foreground=self.COLORS["primary"], font=(self.FONTS["log"][0], self.FONTS["log"][1], "bold"))

    def log_message(self, message, tag='info'):
        def _update_log():
            self.log_widget.config(state='normal')
            timestamp = datetime.now().strftime("%H:%M:%S")
            full_message = f"[{timestamp}] {message}\n"
            self.log_widget.insert(tk.END, full_message, tag)
            self.log_widget.config(state='disabled')
            self.log_widget.see(tk.END)
        self.root.after(0, _update_log)

    def select_files(self):
        self.file_paths = filedialog.askopenfilenames(
            title="เลือกไฟล์ Excel", filetypes=(("Excel Files", "*.xlsx *.xls"), ("All files", "*.*"))
        )
        if self.file_paths:
            file_names = ", ".join([os.path.basename(p) for p in self.file_paths])
            self.lbl_file_status.config(text=f"เลือกแล้ว {len(self.file_paths)} ไฟล์: {file_names}", fg=self.COLORS["status_ok"])
        else:
            self.lbl_file_status.config(text="ยังไม่ได้เลือกไฟล์", fg=self.COLORS["status_none"])
    
    def get_sheet_names_from_input(self):
        text_content = self.sheet_entry.get("1.0", tk.END)
        sheet_names = [line.strip() for line in text_content.splitlines() if line.strip() and line.strip().lower() not in ['contents', 'info']]
        return sheet_names

    def start_processing_thread(self):
        if not self.file_paths:
            messagebox.showerror("ข้อผิดพลาด", "กรุณาเลือกไฟล์ Excel ก่อน")
            return
            
        sheet_names = self.get_sheet_names_from_input()
        if not sheet_names:
            messagebox.showerror("ข้อผิดพลาด", "กรุณาป้อนชื่อชีทที่ต้องการประมวลผล")
            return

        confirm = messagebox.askyesno("ยืนยันการทำงาน", 
            "โปรแกรมจะทำการแก้ไขและบันทึกทับไฟล์ต้นฉบับที่คุณเลือก\n\n"
            "**คำเตือน: การกระทำนี้ไม่สามารถย้อนกลับได้**\n\n"
            "แน่ใจหรือไม่ว่าต้องการดำเนินการต่อ?", icon='warning')
        
        if not confirm:
            return

        self.log_widget.config(state='normal')
        self.log_widget.delete('1.0', tk.END)
        self.log_widget.config(state='disabled')
        self.btn_process.config(state=tk.DISABLED)
        self.btn_select_files.config(state=tk.DISABLED)

        processing_thread = threading.Thread(
            target=self.run_process_and_reenable_button, 
            args=(self.file_paths, sheet_names, self.log_message)
        )
        processing_thread.daemon = True
        processing_thread.start()

    def run_process_and_reenable_button(self, file_paths, sheet_names, log_callback):
        try:
            process_excel_files_inplace(file_paths, sheet_names, log_callback)
        finally:
            def _reenable():
                self.btn_process.config(state=tk.NORMAL)
                self.btn_select_files.config(state=tk.NORMAL)
            self.root.after(0, _reenable)





# <<< START OF CHANGES >>>
# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
def run_this_app(working_dir=None): # ชื่อฟังก์ชันนี้จะถูกใช้ใน Launcher
    """
    ฟังก์ชันหลักสำหรับสร้างและรัน QuotaSamplerApp.
    """
    print(f"--- QUOTA_SAMPLER_INFO: Starting 'QuotaSamplerApp' via run_this_app() ---")
    try:
        # --- โค้ดที่ย้ายมาจาก if __name__ == "__main__": เดิมจะมาอยู่ที่นี่ ---
    #if __name__ == "__main__":
        root = tk.Tk()
        app = App(root)
        root.mainloop()

        print(f"--- QUOTA_SAMPLER_INFO: QuotaSamplerApp mainloop finished. ---")

    except Exception as e:
        # ดักจับ Error ที่อาจเกิดขึ้นระหว่างการสร้างหรือรัน App
        print(f"QUOTA_SAMPLER_ERROR: An error occurred during QuotaSamplerApp execution: {e}")
        # แสดง Popup ถ้ามีปัญหา
        if 'root' not in locals() or not root.winfo_exists(): # สร้าง root ชั่วคราวถ้ายังไม่มี
            root_temp = tk.Tk()
            root_temp.withdraw()
            messagebox.showerror("Application Error (Quota Sampler)",
                               f"An unexpected error occurred:\n{e}", parent=root_temp)
            root_temp.destroy()
        else:
            messagebox.showerror("Application Error (Quota Sampler)",
                               f"An unexpected error occurred:\n{e}", parent=root) # ใช้ root ที่มีอยู่ถ้าเป็นไปได้
        sys.exit(f"Error running QuotaSamplerApp: {e}") # อาจจะ exit หรือไม่ก็ได้ ขึ้นกับการออกแบบ


# --- ส่วน Run Application เมื่อรันไฟล์นี้โดยตรง (สำหรับ Test) ---
if __name__ == "__main__":
    print("--- Running QuotaSamplerApp.py directly for testing ---")
    # (ถ้ามีการตั้งค่า DPI ด้านบน มันจะทำงานอัตโนมัติ)

    # เรียกฟังก์ชัน Entry Point ที่เราสร้างขึ้น
    run_this_app()

    print("--- Finished direct execution of QuotaSamplerApp.py ---")
# <<< END OF CHANGES >>>