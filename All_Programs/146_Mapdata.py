import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QLabel, QFileDialog, 
                             QComboBox, QTableWidget, QTableWidgetItem, 
                             QHeaderView, QMessageBox, QFrame, QSplitter, QGroupBox)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QColor, QGuiApplication

class ExcelLoadWorker(QThread):
    loaded = pyqtSignal(int, object, str)
    failed = pyqtSignal(int, str)

    def __init__(self, file_num, file_path):
        super().__init__()
        self.file_num = file_num
        self.file_path = file_path

    def run(self):
        try:
            df = pd.read_excel(self.file_path, dtype=str, nrows=0)
            df.columns = df.columns.str.strip()
            self.loaded.emit(self.file_num, df, self.file_path)
        except Exception as exc:
            self.failed.emit(self.file_num, str(exc))


class ExcelComparerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("โปรแกรมตรวจเช็คข้อมูล Excel (V4 - Save/Load Settings)")
        self.resize(1200, 800)
        self.center_window()
        
        # ตัวแปรเก็บข้อมูล
        self.df1 = None
        self.df2 = None
        self.file1_path = ""
        self.file2_path = ""
        self.mapping_widgets = {} 
        self.loading_count = 0
        self.load_workers = {}
        self.header_mismatch = None

        # สร้างหน้าจอ
        self.init_ui()
        self.apply_stylesheet()

    def center_window(self):
        screen = QGuiApplication.primaryScreen()
        if screen is None:
            return
        screen_geometry = screen.availableGeometry()
        window_geometry = self.frameGeometry()
        window_geometry.moveCenter(screen_geometry.center())
        self.move(window_geometry.topLeft())

    def init_ui(self):
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)

        # --- ส่วนหัว: ปุ่ม Save/Load Setting ---
        setting_layout = QHBoxLayout()
        self.btn_save_setting = QPushButton("💾 บันทึกการตั้งค่า (Save Setting)")
        self.btn_save_setting.setObjectName("SettingBtn")
        self.btn_save_setting.clicked.connect(self.save_settings)
        
        self.btn_load_setting = QPushButton("📂 โหลดการตั้งค่าเดิม (Load Setting)")
        self.btn_load_setting.setObjectName("SettingBtn")
        self.btn_load_setting.clicked.connect(self.load_settings_file)
        
        setting_layout.addWidget(self.btn_load_setting)
        setting_layout.addWidget(self.btn_save_setting)
        setting_layout.addStretch()
        main_layout.addLayout(setting_layout)

        # --- ส่วนที่ 1: โหลดไฟล์ และ เลือก ID ---
        top_frame = QFrame()
        top_frame.setObjectName("Card")
        top_layout = QVBoxLayout(top_frame)

        # แถวโหลดไฟล์
        h_layout_files = QHBoxLayout()
        
        # ไฟล์ 1
        v_f1 = QVBoxLayout()
        self.btn_file1 = QPushButton("📂 1. เลือกไฟล์หลัก (Master)")
        self.btn_file1.clicked.connect(lambda: self.load_file_dialog(1))
        self.lbl_file1 = QLabel("ยังไม่ได้เลือกไฟล์")
        v_f1.addWidget(self.btn_file1)
        v_f1.addWidget(self.lbl_file1)
        
        # ไฟล์ 2
        v_f2 = QVBoxLayout()
        self.btn_file2 = QPushButton("📂 2. เลือกไฟล์ที่จะตรวจ (Target)")
        self.btn_file2.clicked.connect(lambda: self.load_file_dialog(2))
        self.lbl_file2 = QLabel("ยังไม่ได้เลือกไฟล์")
        v_f2.addWidget(self.btn_file2)
        v_f2.addWidget(self.lbl_file2)

        h_layout_files.addLayout(v_f1)
        h_layout_files.addLayout(v_f2)
        top_layout.addLayout(h_layout_files)

        # แถวเลือก ID (Key)
        key_group = QGroupBox("STEP 1: กำหนดตัวเชื่อมข้อมูล (Key ID)")
        key_layout = QHBoxLayout()
        
        key_layout.addWidget(QLabel("ID ไฟล์หลัก:"))
        self.combo_key1 = QComboBox()
        self.combo_key1.currentTextChanged.connect(self.update_run_button_state)
        key_layout.addWidget(self.combo_key1)
        
        key_layout.addWidget(QLabel("  จับคู่กับ  ➡  "))
        
        key_layout.addWidget(QLabel("ID ไฟล์ที่ตรวจ:"))
        self.combo_key2 = QComboBox()
        self.combo_key2.currentTextChanged.connect(self.update_run_button_state)
        key_layout.addWidget(self.combo_key2)

        key_group.setLayout(key_layout)
        top_layout.addWidget(key_group)

        main_layout.addWidget(top_frame)

        # --- ส่วนที่ 2: จับคู่คอลัมน์ และ ผลลัพธ์ ---
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # ซ้าย: พื้นที่เลือกคอลัมน์ที่จะตรวจ
        map_frame = QFrame()
        map_frame.setObjectName("Card")
        map_layout = QVBoxLayout(map_frame)
        map_layout.addWidget(QLabel("<b>STEP 2: Auto-Map (ไม่สร้าง UI)</b>"))
        self.lbl_mapping_status = QLabel("Auto-Map พร้อมใช้งาน")
        map_layout.addWidget(self.lbl_mapping_status)

        self.header_table = QTableWidget()
        self.header_table.setColumnCount(2)
        self.header_table.setHorizontalHeaderLabels(["Master เท่านั้น", "Target เท่านั้น"])
        self.header_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        map_layout.addWidget(self.header_table)

        self.btn_run = QPushButton("⚡ เริ่มตรวจสอบ (RUN)")
        self.btn_run.setObjectName("ActionBtn")
        self.btn_run.setEnabled(False)
        self.btn_run.clicked.connect(self.run_comparison)
        map_layout.addWidget(self.btn_run)

        splitter.addWidget(map_frame)

        # ขวา: ตารางผลลัพธ์
        result_frame = QFrame()
        result_frame.setObjectName("Card")
        result_layout = QVBoxLayout(result_frame)
        
        header_res = QHBoxLayout()
        header_res.addWidget(QLabel("<b>ผลลัพธ์ (เฉพาะข้อที่ไม่ตรง)</b>"))
        self.btn_export = QPushButton("💾 Export ผลลัพธ์")
        self.btn_export.setEnabled(False)
        self.btn_export.clicked.connect(self.export_results)
        header_res.addWidget(self.btn_export, alignment=Qt.AlignmentFlag.AlignRight)
        
        result_layout.addLayout(header_res)

        self.result_table = QTableWidget()
        self.result_table.setColumnCount(5)
        self.result_table.setHorizontalHeaderLabels(["ID", "ชื่อคอลัมน์", "ค่าไฟล์หลัก", "ค่าไฟล์ที่ตรวจ", "หมายเหตุ"])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        result_layout.addWidget(self.result_table)

        splitter.addWidget(result_frame)
        splitter.setSizes([450, 700])

        main_layout.addWidget(splitter)

    # --- File Loading Logic ---
    def load_file_dialog(self, file_num):
        file_name, _ = QFileDialog.getOpenFileName(self, "เปิดไฟล์ Excel ข้อมูล", "", "Excel Files (*.xlsx *.xls)")
        if file_name:
            self.start_load_file(file_num, file_name)

    def start_load_file(self, file_num, file_path):
        if not os.path.exists(file_path):
            QMessageBox.warning(self, "Error", f"หาไฟล์ไม่เจอ: {file_path}")
            return

        self.set_loading_state(True)
        worker = ExcelLoadWorker(file_num, file_path)
        self.load_workers[file_num] = worker
        worker.loaded.connect(self.on_file_loaded)
        worker.failed.connect(self.on_file_load_failed)
        worker.start()

    def on_file_loaded(self, file_num, df, file_path):
        self.load_workers.pop(file_num, None)
        self.apply_loaded_df(file_num, df, file_path)
        self.set_loading_state(False)

    def on_file_load_failed(self, file_num, error_message):
        self.load_workers.pop(file_num, None)
        self.set_loading_state(False)
        QMessageBox.critical(self, "Error", f"โหลดไฟล์ผิดพลาด: {error_message}")

    def apply_loaded_df(self, file_num, df, file_path):
        try:
            if file_num == 1:
                self.df1 = df
                self.file1_path = file_path
                self.lbl_file1.setText(os.path.basename(file_path))
                self.combo_key1.clear()
                self.combo_key1.addItems(df.columns)
                # Auto select ID default
                for p in ["SbjNum", "SBJNUM", "sbjnum", "KEY", "Key", "key", "ID", "id", "Id", "NO", "No"]:
                    if p in df.columns:
                        self.combo_key1.setCurrentText(p)
                        break
            else:
                self.df2 = df
                self.file2_path = file_path
                self.lbl_file2.setText(os.path.basename(file_path))
                self.combo_key2.clear()
                self.combo_key2.addItems(df.columns)
                for p in ["SbjNum", "SBJNUM", "sbjnum", "KEY", "Key", "key", "ID", "id", "Id", "NO", "No"]:
                    if p in df.columns:
                        self.combo_key2.setCurrentText(p)
                        break

            if self.df1 is not None and self.df2 is not None:
                self.lbl_mapping_status.setText("Auto-Map พร้อมใช้งาน")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"โหลดไฟล์ผิดพลาด: {str(e)}")

    def process_load_file(self, file_num, file_path):
        try:
            if not os.path.exists(file_path):
                QMessageBox.warning(self, "Error", f"หาไฟล์ไม่เจอ: {file_path}")
                return

            df = pd.read_excel(file_path, dtype=str, nrows=0)
            df.columns = df.columns.str.strip()
            self.apply_loaded_df(file_num, df, file_path)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"โหลดไฟล์ผิดพลาด: {str(e)}")

    def set_loading_state(self, is_loading):
        if is_loading:
            self.loading_count += 1
            QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
            self.btn_file1.setEnabled(False)
            self.btn_file2.setEnabled(False)
            self.btn_save_setting.setEnabled(False)
            self.btn_load_setting.setEnabled(False)
        else:
            self.loading_count = max(0, self.loading_count - 1)
            if self.loading_count == 0:
                QApplication.restoreOverrideCursor()
                self.btn_file1.setEnabled(True)
                self.btn_file2.setEnabled(True)
                self.btn_save_setting.setEnabled(True)
                self.btn_load_setting.setEnabled(True)
    def check_headers_before_mapping(self):
        if self.df1 is None or self.df2 is None:
            return False

        set1 = set(self.df1.columns)
        set2 = set(self.df2.columns)
        missing_in_target = sorted(set1 - set2)
        missing_in_master = sorted(set2 - set1)

        if not missing_in_target and not missing_in_master:
            QMessageBox.information(self, "ตรวจสอบคอลัมน์", "ตัวแปรตรงกัน 100% พร้อมทำการ Mapping")
            self.header_table.setRowCount(0)
            self.lbl_mapping_status.setText("ตัวแปรตรงกัน 100%")
            self.lbl_mapping_status.setStyleSheet("color: #1f4e79; font-weight: bold;")
            return True

        QMessageBox.warning(self, "ตรวจสอบคอลัมน์", "ตัวแปรไม่ตรงกัน 100%")
        self.header_table.setRowCount(max(len(missing_in_target), len(missing_in_master)))
        for i in range(self.header_table.rowCount()):
            val1 = missing_in_target[i] if i < len(missing_in_target) else ""
            val2 = missing_in_master[i] if i < len(missing_in_master) else ""
            self.header_table.setItem(i, 0, QTableWidgetItem(val1))
            self.header_table.setItem(i, 1, QTableWidgetItem(val2))
        self.lbl_mapping_status.setText("พบตัวแปรไม่ตรงกัน")
        self.lbl_mapping_status.setStyleSheet("color: #c0392b; font-weight: bold;")
        export_path, _ = QFileDialog.getSaveFileName(
            self, "บันทึกคอลัมน์ที่ไม่ตรงกัน", "Header_Mismatch.xlsx", "Excel Files (*.xlsx)"
        )
        if export_path:
            df_mismatch = pd.DataFrame({
                "Master_Only": pd.Series(missing_in_target, dtype="string"),
                "Target_Only": pd.Series(missing_in_master, dtype="string"),
            })
            df_mismatch.to_excel(export_path, index=False)

        reply = QMessageBox.question(
            self,
            "ยืนยันการทำ Mapping",
            "ต้องการทำ Mapping ต่อหรือไม่?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        return reply == QMessageBox.StandardButton.Yes

    def build_auto_mapping(self, key1):
        active_mapping = {}
        file2_map = {self._normalize_col(c): c for c in self.df2.columns}
        for col1 in self.df1.columns:
            if col1 == key1:
                continue
            norm = self._normalize_col(col1)
            if norm in file2_map:
                active_mapping[col1] = file2_map[norm]
        return active_mapping

    def _normalize_col(self, name):
        return "".join(ch for ch in name.strip().lower() if ch.isalnum())

    # --- Save & Load Settings Logic ---
    def save_settings(self):
        if not self.file1_path or not self.file2_path:
            QMessageBox.warning(self, "แจ้งเตือน", "กรุณาโหลดไฟล์ข้อมูลก่อนทำการบันทึก Setting")
            return

        save_path, _ = QFileDialog.getSaveFileName(self, "บันทึกการตั้งค่า", "MySetting.xlsx", "Excel Files (*.xlsx)")
        if save_path:
            try:
                # 1. Config Sheet: Paths & Keys
                config_data = {
                    'Param': ['File1', 'File2', 'Key1', 'Key2'],
                    'Value': [self.file1_path, self.file2_path, self.combo_key1.currentText(), self.combo_key2.currentText()]
                }
                df_config = pd.DataFrame(config_data)

                # 2. Mapping Sheet: Which column maps to which
                mapping_data = []
                for col1, combo in self.mapping_widgets.items():
                    mapping_data.append({
                        'Master_Col': col1,
                        'Target_Col': combo.currentText()
                    })
                df_mapping = pd.DataFrame(mapping_data)

                # Write to Excel
                with pd.ExcelWriter(save_path) as writer:
                    df_config.to_excel(writer, sheet_name='Config', index=False)
                    df_mapping.to_excel(writer, sheet_name='Mapping', index=False)
                
                QMessageBox.information(self, "สำเร็จ", "บันทึกการตั้งค่าเรียบร้อยแล้ว!")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"ไม่สามารถบันทึกได้: {e}")

    def load_settings_file(self):
        load_path, _ = QFileDialog.getOpenFileName(self, "โหลดไฟล์การตั้งค่า", "", "Excel Files (*.xlsx)")
        if load_path:
            try:
                # Read Config
                df_config = pd.read_excel(load_path, sheet_name='Config')
                df_mapping = pd.read_excel(load_path, sheet_name='Mapping')

                # Extract Params
                f1_path = df_config.loc[df_config['Param'] == 'File1', 'Value'].values[0]
                f2_path = df_config.loc[df_config['Param'] == 'File2', 'Value'].values[0]
                k1 = df_config.loc[df_config['Param'] == 'Key1', 'Value'].values[0]
                k2 = df_config.loc[df_config['Param'] == 'Key2', 'Value'].values[0]

                # 1. Load Files
                self.process_load_file(1, f1_path)
                self.process_load_file(2, f2_path)

                # 2. Set Keys
                self.combo_key1.setCurrentText(k1)
                self.combo_key2.setCurrentText(k2)

                # 3. Apply Mapping (Loop through saved mapping and set combo boxes)
                # Re-generate UI first to be sure
                if self.df1 is not None and self.df2 is not None:
                    # Create a dict for fast lookup from the loaded mapping
                    saved_map = dict(zip(df_mapping['Master_Col'], df_mapping['Target_Col']))
                    
                    for col1, combo in self.mapping_widgets.items():
                        if col1 in saved_map:
                            combo.setCurrentText(saved_map[col1])

                QMessageBox.information(self, "สำเร็จ", "โหลดการตั้งค่าเรียบร้อยแล้ว!")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"โหลดการตั้งค่าไม่สำเร็จ: {e}\n(ไฟล์อาจเสียหายหรือรูปแบบไม่ถูกต้อง)")

    # --- Comparison Logic ---
    def update_run_button_state(self):
        enabled = (self.df1 is not None and self.df2 is not None and 
                   self.combo_key1.currentText() != "" and 
                   self.combo_key2.currentText() != "")
        self.btn_run.setEnabled(enabled)

    def run_comparison(self):
        if not self.file1_path or not self.file2_path:
            QMessageBox.warning(self, "แจ้งเตือน", "กรุณาเลือกไฟล์ข้อมูลก่อนทำการตรวจสอบ")
            return

        self.set_loading_state(True)
        try:
            self.df1 = pd.read_excel(self.file1_path, dtype=str)
            self.df2 = pd.read_excel(self.file2_path, dtype=str)
            self.df1.columns = self.df1.columns.str.strip()
            self.df2.columns = self.df2.columns.str.strip()
        except Exception as e:
            self.set_loading_state(False)
            QMessageBox.critical(self, "Error", f"โหลดไฟล์ผิดพลาด: {str(e)}")
            return
        self.set_loading_state(False)

        key1 = self.combo_key1.currentText()
        key2 = self.combo_key2.currentText()
        
        try:
            df1_indexed = self.df1.set_index(key1)
            df2_indexed = self.df2.set_index(key2)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Key Error: {e}")
            return

        common_ids = df1_indexed.index.intersection(df2_indexed.index)
        
        if len(common_ids) == 0:
            QMessageBox.warning(self, "ไม่พบข้อมูล", "ไม่พบ ID ที่ตรงกันเลย")
            return

        if not self.check_headers_before_mapping():
            return
        active_mapping = self.build_auto_mapping(key1)

        if not active_mapping:
            QMessageBox.information(self, "แจ้งเตือน", "กรุณาเลือกคอลัมน์ที่จะเทียบอย่างน้อย 1 อัน")
            return

        self.result_table.setRowCount(0)
        mismatches = []

        for rec_id in common_ids:
            row1 = df1_indexed.loc[rec_id]
            row2 = df2_indexed.loc[rec_id]

            if isinstance(row1, pd.DataFrame): row1 = row1.iloc[0]
            if isinstance(row2, pd.DataFrame): row2 = row2.iloc[0]

            for col1, col2 in active_mapping.items():
                val1 = str(row1[col1]).strip()
                val2 = str(row2[col2]).strip()
                
                if val1.lower() == 'nan': val1 = ""
                if val2.lower() == 'nan': val2 = ""

                if val1 != val2:
                    mismatches.append({
                        "ID": rec_id,
                        "Column": col1,
                        "Val1": val1,
                        "Val2": val2,
                        "Desc": f"{col1} != {col2}"
                    })

        self.result_table.setRowCount(len(mismatches))
        for i, m in enumerate(mismatches):
            self.result_table.setItem(i, 0, QTableWidgetItem(str(m["ID"])))
            self.result_table.setItem(i, 1, QTableWidgetItem(m["Column"]))
            item1 = QTableWidgetItem(m["Val1"])
            item1.setBackground(QColor("#e8f5e9")) 
            self.result_table.setItem(i, 2, item1)
            item2 = QTableWidgetItem(m["Val2"])
            item2.setBackground(QColor("#ffebee")) 
            self.result_table.setItem(i, 3, item2)
            self.result_table.setItem(i, 4, QTableWidgetItem(m["Desc"]))

        if len(mismatches) == 0:
            QMessageBox.information(self, "ผลการตรวจ", "ข้อมูลถูกต้องทั้งหมด!")
        else:
            QMessageBox.warning(self, "ผลการตรวจ", f"พบข้อมูลไม่ตรงกัน {len(mismatches)} จุด")
            self.btn_export.setEnabled(True)
            self.current_mismatches = mismatches

    def export_results(self):
        path, _ = QFileDialog.getSaveFileName(self, "บันทึกไฟล์", "ผลการตรวจสอบ.xlsx", "Excel Files (*.xlsx)")
        if path:
            wb = Workbook()
            ws = wb.active
            headers = ["ID", "ชื่อคอลัมน์", "ค่าไฟล์หลัก", "ค่าไฟล์ที่ตรวจ", "หมายเหตุ"]
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center = Alignment(horizontal="center", vertical="center")

            ws.append(headers)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center

            green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

            for m in self.current_mismatches:
                ws.append([m["ID"], m["Column"], m["Val1"], m["Val2"], m["Desc"]])
                row = ws.max_row
                ws.cell(row=row, column=3).fill = green_fill
                ws.cell(row=row, column=4).fill = red_fill

            col_widths = [0] * len(headers)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(headers)):
                for idx, cell in enumerate(row):
                    cell_value = "" if cell.value is None else str(cell.value)
                    col_widths[idx] = max(col_widths[idx], len(cell_value))
            for idx, width in enumerate(col_widths, start=1):
                ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = max(12, min(50, width + 2))

            ws.freeze_panes = "B3"
            wb.save(path)
            QMessageBox.information(self, "สำเร็จ", "บันทึกไฟล์เรียบร้อยแล้ว!")

    def apply_stylesheet(self):
        self.setStyleSheet("""
            QMainWindow { background-color: #f0f4f8; }
            QGroupBox { font-weight: bold; font-size: 14px; border: 1px solid #bdc3c7; border-radius: 6px; margin-top: 10px; padding-top: 15px; background-color: white; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; color: #2c3e50; }
            QFrame#Card { background: white; border-radius: 8px; border: 1px solid #dcdcdc; }
            QPushButton { background-color: #3498db; color: white; border-radius: 5px; padding: 8px 15px; font-size: 13px; font-weight: bold; }
            QPushButton:hover { background-color: #2980b9; }
            QPushButton#SettingBtn { background-color: #607d8b; }
            QPushButton#SettingBtn:hover { background-color: #455a64; }
            QPushButton#ActionBtn { background-color: #27ae60; font-size: 16px; padding: 12px; }
            QPushButton#ActionBtn:hover { background-color: #2ecc71; }
            QPushButton:disabled { background-color: #95a5a6; }
            QLabel { color: #34495e; font-size: 14px; }
        """)



# <<< START OF CHANGES >>>
# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
def run_this_app(working_dir=None): # ชื่อฟังก์ชันนี้จะถูกใช้ใน Launcher
    """
    ฟังก์ชันหลักสำหรับสร้างและรัน QuotaSamplerApp.
    """
    print(f"--- QUOTA_SAMPLER_INFO: Starting 'QuotaSamplerApp' via run_this_app() ---")
    try:
    # --- ส่วนที่ใช้รันโปรแกรม ---
    #if __name__ == '__main__':
        app = QApplication(sys.argv)
        app.setFont(QFont("Tahoma", 10))
        window = ExcelComparerApp()
        window.show()
        sys.exit(app.exec())
        
        print(f"--- QUOTA_SAMPLER_INFO: QuotaSamplerApp mainloop finished. ---")

    except Exception as e:
        # ดักจับ Error ที่อาจเกิดขึ้นระหว่างการสร้างหรือรัน App
        print(f"QUOTA_SAMPLER_ERROR: An error occurred during QuotaSamplerApp execution: {e}")
        # แสดง Popup ถ้ามีปัญหา
        if 'root' not in locals() or not root.winfo_exists(): # สร้าง root ชั่วคราวถ้ายังไม่มี
            root_temp = tk.Tk()
            root_temp.withdraw()
            messagebox.showerror("Application Error (Quota Sampler)",
                                f"An unexpected error occurred:\n{e}", parent=root_temp)
            root_temp.destroy()
        else:
            messagebox.showerror("Application Error (Quota Sampler)",
                                f"An unexpected error occurred:\n{e}", parent=root) # ใช้ root ที่มีอยู่ถ้าเป็นไปได้
        sys.exit(f"Error running QuotaSamplerApp: {e}") # อาจจะ exit หรือไม่ก็ได้ ขึ้นกับการออกแบบ


# --- ส่วน Run Application เมื่อรันไฟล์นี้โดยตรง (สำหรับ Test) ---
if __name__ == "__main__":
    print("--- Running QuotaSamplerApp.py directly for testing ---")
    # (ถ้ามีการตั้งค่า DPI ด้านบน มันจะทำงานอัตโนมัติ)

    # เรียกฟังก์ชัน Entry Point ที่เราสร้างขึ้น
    run_this_app()

    print("--- Finished direct execution of QuotaSamplerApp.py ---")
# <<< END OF CHANGES >>>