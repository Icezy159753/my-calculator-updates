import os
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# =========================
# สีไฮไลท์
# =========================
FILL_ORANGE = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # เซลล์ Net ที่ผิด
FILL_RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # (สำรอง ถ้าจะใช้เติมแถว)

EPS = 1e-6  # กันความคลาดทศนิยม

# =========================
# Utilities: ตัวเลข
# =========================
def numeric_at(ws, row_idx, col_idx):
    """อ่านค่าเป็น float จากเซลล์ (รองรับ string ที่เป็นตัวเลข)"""
    val = ws.cell(row=row_idx, column=col_idx).value
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip().replace(",", "")
        if re.fullmatch(r"-?\d+(\.\d+)?", s):
            try:
                return float(s)
            except:
                return None
    return None


def first_numeric_col_in_row(ws, row_idx, start_col=3, end_col=50):
    """หา 'คอลัมน์ตัวเลขแรก' ในแถวนั้น (เริ่มจากคอลัมน์ start_col)"""
    for c in range(start_col, end_col + 1):
        v = numeric_at(ws, row_idx, c)
        if v is not None:
            return c, v
    return None, None


def is_net_header_text(text: str) -> bool:
    """
    ระบุตรงๆ ว่าเป็นหัว Net/Subnet/... เท่านั้น
    (กันเคสคำที่มี 'net' แทรกอยู่ เช่น cabinet)
    """
    import re
    if not isinstance(text, str):
        return False
    s = text.strip()
    pattern = re.compile(
        r'^\s*(?:net|subnet|subsubnet|subsubsubnet|subsubsubsubnet)\b',
        re.IGNORECASE
    )
    return bool(pattern.match(s))


# =========================
# Utilities: โครงสร้างตาราง
# =========================
def is_label_row(text_b: str) -> bool:
    """
    แถวโค้ด (leaf) = B มีข้อความ, ไม่ใช่หัว Net/Subnet/..., ไม่ใช่ TOTAL,
    และไม่ใช่บรรทัดหัวข้ออย่าง '1st row:' / '2nd row:'
    """
    if not isinstance(text_b, str):
        return False
    s = text_b.strip()
    if not s:
        return False
    if is_net_header_text(s):   # แก้จากเดิมที่เคยใช้ "net" in s.lower()
        return False
    if s.upper().startswith("TOTAL"):
        return False
    if "row:" in s.lower():     # "1st row:", "2nd row:"
        return False
    return True



def get_indent_level(ws, row_idx):
    """
    ประมาณระดับชั้นของหัวข้อจากคอลัมน์ B
    - ใช้ cell.alignment.indent (ถ้ามี) + จำนวนช่องว่างนำหน้า (รวม NBSP)
    - คืน int: 0=ชั้นบนสุด (Net), 1=Subnet, 2=Subsubnet, ...
    """
    cell = ws.cell(row=row_idx, column=2)
    txt = str(cell.value or "")
    lead_spaces = len(txt) - len(txt.lstrip(' \u00A0'))
    try:
        indent = cell.alignment.indent or 0
    except Exception:
        indent = 0
    return int(indent + (lead_spaces // 2))


def find_total_column_index(ws, first_net_row, start_col=3, end_col=200):
    """
    ยึดกติกา: TOTAL อยู่คอลัมน์ B เสมอ
    - หา 'แถวล่าสุด' ที่ B == 'TOTAL' ตั้งแต่แถว 1..first_net_row
    - ใช้คอลัมน์ตัวเลขแรกในแถวนั้นเป็นคอลัมน์อ้างอิง (Count/TOTAL)
    - ถ้าไม่พบแถว TOTAL ให้ fallback เป็นคอลัมน์ C
    """
    last_total_row = None
    for r in range(1, first_net_row + 1):
        b = ws.cell(row=r, column=2).value
        if isinstance(b, str) and b.strip().upper() == "TOTAL":
            last_total_row = r

    if last_total_row:
        col_idx, _ = first_numeric_col_in_row(ws, last_total_row, start_col=start_col, end_col=end_col)
        if col_idx is not None:
            return col_idx

    # fallback: C
    return 3




def get_total_row_index(ws, first_net_row):
    """
    หาแถว TOTAL สุดท้ายก่อน (หรือเท่ากับ) แถว Net แรก
    ใช้เป็น anchor บอกว่า 'แถว Count' คือแถวที่มี parity เดียวกับ TOTAL
    """
    last_total_row = None
    for r in range(1, first_net_row + 1):
        b = ws.cell(row=r, column=2).value
        if isinstance(b, str) and b.strip().upper() == "TOTAL":
            last_total_row = r
    return last_total_row

def is_count_row(r, total_row_idx):
    """
    คืน True ถ้า r เป็นแถว Count (อิง parity เทียบกับแถว TOTAL)
    ถ้าหา TOTAL ไม่ได้ ให้ถือว่าทุกแถวเป็น Count (fallback ปลอดภัย)
    """
    if total_row_idx is None:
        return True
    return ((r - total_row_idx) % 2) == 0


def nearest_count_row(r: int, total_row_idx: int | None) -> int | None:
    """
    คืนดัชนีแถว 'Count' ที่อยู่ใกล้ที่สุดรอบ ๆ r
    ลำดับพิจารณา: r -> r-1 -> r+1
    ถ้าหา TOTAL ไม่ได้ (total_row_idx=None) คืน r (ถือว่าเป็น Count)
    """
    if total_row_idx is None:
        return r
    for cand in (r, r - 1, r + 1):
        if cand >= 1 and is_count_row(cand, total_row_idx):
            return cand
    return None


def header_depth(text: str) -> int | None:
    """
    คำนวณระดับชั้นจากข้อความหัวข้อ:
      Net               -> 0
      Subnet            -> 1
      Subsubnet         -> 2
      Subsubsubnet      -> 3
      Subsubsubsubnet   -> 4
    ไม่พึ่ง indent; อาศัยจำนวน 'sub' ที่อยู่หน้า 'net' เท่านั้น
    ถ้าไม่ใช่หัวข้อพวกนี้ คืน None
    """
    if not isinstance(text, str):
        return None
    s = text.strip().lower()
    # ต้องขึ้นต้นด้วยเนื้อหาที่ is_net_header_text() มองว่าเป็นหัวจริง
    if not is_net_header_text(s):
        return None
    # นับจำนวน "sub" ที่มาก่อนคำว่า "net"
    prefix = s.split("net", 1)[0]  # ส่วนหน้าคำว่า net
    level = prefix.count("sub")
    return level


# =========================
# Core: ตรวจหลายระดับ (5 ระดับ)
# =========================
def check_sheet(ws, eps: float = EPS):
    """
    ตรวจหลายระดับตามกติกา:
      RULE A  (CODE_SUM_LT_NODE):             ผลรวม code ใต้ตัวเอง (Count/TOTAL)        >= ค่าโหนด
      RULE B  (SUBNETS_CODE_SUM_LT_NODE):     ผลรวม code_sum ของ 'หัวลูกโดยตรง'        >= ค่าโหนด
      RULE C  (SUBNET_HEADERS_SUM_LT_NODE):   ผลรวม 'ค่าในบรรทัดหัวของลูกโดยตรง'       >= ค่าโหนด
    เงื่อนไขใหม่:
      - ถ้าโหนดมี 'โค้ดลูกโดยตรง' (has_direct_code=True) → ข้าม RULE B/C สำหรับโหนดนั้น
        (เช็คเฉพาะ RULE A พอ)
    - ใช้คอลัมน์ TOTAL จากแถวที่ B="TOTAL"
    - Count อยู่บรรทัดเดียวกับ label เท่านั้น (ไม่อ่านบรรทัด %)
    - ระดับชั้นกำหนดจากข้อความหัวข้อ (Net/Subnet/...) ผ่าน header_depth() + is_net_header_text()
    """
    LEVEL_NAMES = ["Net", "Subnet", "Subsubnet", "Subsubsubnet", "Subsubsubsubnet"]

    errors = []
    has_any_error = False
    max_row = ws.max_row

    # หาแถวหัวข้อ Net/Sub* ทั้งหมด
    net_rows = []
    for r in range(1, max_row + 1):
        b = ws.cell(row=r, column=2).value
        if is_net_header_text(b):
            net_rows.append(r)
    if not net_rows:
        return errors, has_any_error

    # ระบุคอลัมน์ TOTAL (จากแถวที่ B='TOTAL')
    total_col_idx = find_total_column_index(ws, first_net_row=net_rows[0], start_col=3, end_col=200)

    # โครงสร้าง stack ของโหนด Net ตามระดับชั้น
    # เพิ่ม 'has_direct_code' เพื่อใช้ปิด RULE B/C เมื่อมี leaf อยู่ใต้โหนดโดยตรง
    stack = []  # [{depth,row,text,value,code_sum,subnet_sum,subnet_header_sum,has_subnet,has_direct_code}]

    def finalize_node(node):
        """ปิดโหนด: ตรวจ RULE A/B/C และดันยอดขึ้นให้พ่อสำหรับ RULE B/C"""
        nonlocal has_any_error
        level_name = LEVEL_NAMES[node["depth"]] if node["depth"] < len(LEVEL_NAMES) else f"Level{node['depth']+1}"

        # RULE A: รวม code ใต้ตัวเอง
        if node["code_sum"] + eps < node["value"]:
            has_any_error = True
            ws.cell(row=node["row"], column=2).fill = FILL_ORANGE
            errors.append({
                "sheet": ws.title,
                "level": level_name,
                "rule": "CODE_SUM_LT_NODE",
                "net_row": node["row"],
                "net_addr": f"B{node['row']}",
                "net_text": node["text"],
                "col_idx": total_col_idx,
                "net_value": float(node["value"]),
                "sum_code": float(node["code_sum"]),
                "sum_subnets": float(node["subnet_sum"]),
                "sum_subnet_headers": float(node["subnet_header_sum"])
            })

        # เช็ค RULE B/C เฉพาะกรณี 'ไม่มีโค้ดลูกโดยตรง'
        if node["has_subnet"] and (not node["has_direct_code"]):
            # RULE B: รวม code_sum ของลูกโดยตรง
            if node["subnet_sum"] + eps < node["value"]:
                has_any_error = True
                ws.cell(row=node["row"], column=2).fill = FILL_ORANGE
                errors.append({
                    "sheet": ws.title,
                    "level": level_name,
                    "rule": "SUBNETS_CODE_SUM_LT_NODE",
                    "net_row": node["row"],
                    "net_addr": f"B{node['row']}",
                    "net_text": node["text"],
                    "col_idx": total_col_idx,
                    "net_value": float(node["value"]),
                    "sum_code": float(node["code_sum"]),
                    "sum_subnets": float(node["subnet_sum"]),
                    "sum_subnet_headers": float(node["subnet_header_sum"])
                })

            # RULE C: รวม 'ค่าในหัวของลูกโดยตรง'
            if node["subnet_header_sum"] + eps < node["value"]:
                has_any_error = True
                ws.cell(row=node["row"], column=2).fill = FILL_ORANGE
                errors.append({
                    "sheet": ws.title,
                    "level": level_name,
                    "rule": "SUBNET_HEADERS_SUM_LT_NODE",
                    "net_row": node["row"],
                    "net_addr": f"B{node['row']}",
                    "net_text": node["text"],
                    "col_idx": total_col_idx,
                    "net_value": float(node["value"]),
                    "sum_code": float(node["code_sum"]),
                    "sum_subnets": float(node["subnet_sum"]),
                    "sum_subnet_headers": float(node["subnet_header_sum"])
                })

        # ---- ดันยอดขึ้นให้ 'พ่อ' เพื่อใช้ RULE B/C ของพ่อ ----
        if stack:
            parent = stack[-1]
            if parent["depth"] < node["depth"]:
                parent["has_subnet"] = True
                parent["subnet_sum"] += float(node["code_sum"])        # RULE B ใช้ code_sum ของลูก
                parent["subnet_header_sum"] += float(node["value"])    # RULE C ใช้ค่าหัวของลูก

    # เดินสแกนทั้งชีท
    r = 1
    while r <= max_row:
        bval = ws.cell(row=r, column=2).value

        # ---- เจอหัวข้อ Net/Sub* (ระดับจากข้อความโดยตรง)
        if is_net_header_text(bval):
            depth = header_depth(bval)  # ไม่ใช้ indent แล้ว
            text = str(bval).strip()

            net_val = numeric_at(ws, r, total_col_idx)  # Count อยู่บรรทัดนี้
            if net_val is None:
                r += 1
                continue  # อ่านค่าไม่ได้ ข้าม

            # ปิดโหนดที่ระดับ >= โหนดใหม่
            while stack and stack[-1]["depth"] >= depth:
                finalize_node(stack.pop())

            # เปิดโหนดใหม่
            stack.append({
                "depth": depth,
                "row": r,
                "text": text,
                "value": float(net_val),
                "code_sum": 0.0,
                "subnet_sum": 0.0,
                "subnet_header_sum": 0.0,
                "has_subnet": False,
                "has_direct_code": False,     # << ใหม่
            })

            r += 1
            continue

        # ---- แถวโค้ด (leaf): รวมเฉพาะค่าบรรทัดนี้ แล้วดันขึ้นทุก ancestor
        if is_label_row(bval) and stack:
            v = numeric_at(ws, r, total_col_idx)
            if v is not None:
                # โค้ดนี้เป็น 'ลูกโดยตรง' ของโหนดลึกสุดในสแตก
                stack[-1]["has_direct_code"] = True
                # ดันค่าขึ้นทุกระดับ
                for node in stack:
                    node["code_sum"] += float(v)
            r += 1
            continue

        r += 1

    # ปิดโหนดค้าง
    while stack:
        finalize_node(stack.pop())

    return errors, has_any_error









# =========================
# Summary Sheet
# =========================
def insert_summary_sheet(wb, summary_rows):
    """
    สร้างชีท 'Net_Check_Summary' (สรุปเฉพาะเคสที่ผิด)
    คอลัมน์: No | Sheet | Level | Cell(B) | Net text (B) | Result
    จัดรูปแบบ:
      - แท็บชีท: น้ำเงิน
      - หัวตาราง (แถวที่ 1): ตัวหนา + พื้นเขียวอ่อน
      - (ยกเลิก) ไม่ใส่พื้นส้มคอลัมน์ E แล้ว
      - ปรับความกว้างคอลัมน์ให้พอดี (เฉพาะชีทนี้)
    และย้ายชีทไปไว้หลัง 'Info' (fallback: 'Contents Info' → 'Contents')
    """
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    # ลบชีทเดิมถ้ามี
    if "Net_Check_Summary" in wb.sheetnames:
        wb.remove(wb["Net_Check_Summary"])

    # สร้างชีทใหม่
    ws = wb.create_sheet("Net_Check_Summary")

    # ตั้งสีแท็บชีทเป็นน้ำเงิน
    ws.sheet_properties.tabColor = "0000FF"

    # ส่วนหัวตามที่ต้องการ
    headers = ["No", "Sheet", "Level", "Cell(B)", "Net text (B)", "Result"]
    ws.append(headers)

    # เติมข้อมูล (สรุปเฉพาะรายการที่ผิดซึ่งถูกส่งมาแล้วใน summary_rows)
    for i, r in enumerate(summary_rows, start=1):
        ws.append([
            i,
            r.get("sheet", ""),
            r.get("level", ""),
            r.get("net_addr", ""),
            r.get("net_text", ""),
            "LESS THAN (ผิด)",
        ])

    # ----------------------------
    # จัดรูปแบบเฉพาะชีทนี้
    # ----------------------------
    last_row = ws.max_row
    last_col = ws.max_column

    # สไตล์
    fill_header_green = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")  # เขียวอ่อน
    bold_font = Font(bold=True)

    # 1) หัวตาราง: ตัวหนา + พื้นเขียวอ่อน + จัดกึ่งกลาง (ยกเว้น E ให้ชิดซ้าย)
    for c in range(1, last_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = bold_font
        cell.fill = fill_header_green
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.cell(row=1, column=5).alignment = Alignment(horizontal="left", vertical="center")

    # 2) ปรับความกว้างคอลัมน์อัตโนมัติ (พอดีเนื้อหา) เฉพาะชีทนี้
    for c in range(1, last_col + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for r in range(1, last_row + 1):
            val = ws.cell(row=r, column=c).value
            val_str = str(val) if val is not None else ""
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    # ----------------------------
    # ย้ายชีทไปไว้หลัง 'Info' (fallback: 'Contents Info' → 'Contents')
    # ----------------------------
    def _move_after(sheet_name: str) -> bool:
        try:
            anchor_idx = wb.sheetnames.index(sheet_name)
            cur_idx = wb.sheetnames.index("Net_Check_Summary")
            wb._sheets.insert(anchor_idx + 1, wb._sheets.pop(cur_idx))
            return True
        except ValueError:
            return False

    if not _move_after("Info"):
        if not _move_after("Contents Info"):
            _move_after("Contents")

# =========================
# Process ทั้งไฟล์
# =========================
def process_file(input_path):
    wb = load_workbook(input_path, data_only=True)
    any_error = False
    summary_rows = []

    for ws in wb.worksheets:
        if ws.title == "Contents Info":
            continue
        errors, has_err = check_sheet(ws)
        if has_err:
            any_error = True
            ws.sheet_properties.tabColor = "FF0000"
        summary_rows.extend(errors)

    dir_name, base_name = os.path.split(input_path)
    name, ext = os.path.splitext(base_name)
    out_path = os.path.join(dir_name, f"{name}_checked{ext}")

    if any_error and summary_rows:
        insert_summary_sheet(wb, summary_rows)
        wb.save(out_path)
        return out_path, True, summary_rows
    else:
        # ไม่มีผิด → ไม่สร้างชีทสรุป
        wb.save(out_path)
        return out_path, False, []


# =========================
# GUI
# =========================
class NetCheckerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("ตรวจเช็ค Net/Subnet 5 ระดับ (DP Team)")

        self.file_path = tk.StringVar()

        frm = ttk.Frame(root, padding=12)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="ไฟล์ Excel:").grid(row=0, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.file_path, width=60).grid(row=0, column=1, padx=6)
        ttk.Button(frm, text="เลือกไฟล์ Excel", command=self.browse).grid(row=0, column=2)

        ttk.Separator(frm).grid(row=1, column=0, columnspan=3, sticky="ew", pady=8)

        self.btn_run = ttk.Button(frm, text="เริ่มตรวจเช็ค", command=self.run_check)
        self.btn_run.grid(row=2, column=0, columnspan=3, pady=6)

        self.log = tk.Text(frm, height=14, width=90)
        self.log.grid(row=3, column=0, columnspan=3, pady=8, sticky="nsew")

        frm.rowconfigure(3, weight=1)
        frm.columnconfigure(1, weight=1)

    def browse(self):
        path = filedialog.askopenfilename(
            title="เลือกไฟล์ Excel",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")]
        )
        if path:
            self.file_path.set(path)

    def run_check(self):
        path = self.file_path.get().strip()
        if not path:
            messagebox.showwarning("แจ้งเตือน", "กรุณาเลือกไฟล์ Excel ก่อน")
            return

        try:
            out_path, has_error, summary = process_file(path)
            if has_error:
                self.log.insert("end", "พบ Net/Sub* ที่ผลรวมน้อยกว่าค่าเกณฑ์ในบางชีท → สร้างชีทสรุปแล้ว\n")
                for r in summary:
                    self.log.insert(
                        "end",
                        f"- [{r.get('level','')}/{r.get('rule','')}] ชีท: {r['sheet']} | "
                        f"แถว Net: {r['net_row']} | {r['net_text']} | Net={r['net_value']} "
                        f"| SumCode={r['sum_code']} | SumSubnets={r['sum_subnets']}\n"
                    )
                self.log.insert("end", f"\nบันทึกไฟล์: {out_path}\n")
                messagebox.showinfo("เสร็จสิ้น", f"ตรวจพบข้อผิดพลาดและสร้างชีทสรุปแล้ว\nบันทึกเป็น:\n{out_path}")
            else:
                self.log.insert("end", "ตรวจเสร็จ: Net ทุก Sheet ไม่มีผิดพลาด\n")
                self.log.insert("end", f"บันทึกไฟล์เป็น: {out_path}\n")
                messagebox.showinfo("เสร็จสิ้น", "Net ทุก Sheet ไม่มีผิดพลาด\n(ไม่มีชีทสรุปถูกสร้าง)")
        except Exception as e:
            messagebox.showerror("เกิดข้อผิดพลาด", str(e))


# =========================
# main
# =========================
if __name__ == "__main__":
    root = tk.Tk()
    # ปรับฟอนต์ GUI ให้อ่านง่ายหน่อย
    try:
        import tkinter.font as tkfont
        default_font = tkfont.nametofont("TkDefaultFont")
        default_font.configure(size=10)
    except Exception:
        pass

    app = NetCheckerGUI(root)
    root.mainloop()
