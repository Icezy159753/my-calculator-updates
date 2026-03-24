"""
Other Recode Tool  v2
=====================
Phase 1: Auto-detect "Other" codes per question from SPSS Value Labels
         (matches labels containing keywords like อื่น/ระบุ/other/specify)
         → Export to "Other Coding Sheet" Excel for manual coding

Phase 2: Read completed sheet → Apply new codes back to Rawdata
         → Save updated Rawdata + Recode Log
"""

from __future__ import annotations

import logging
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import pyreadstat
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

# Keywords to match "Other, please specify" labels in SPSS (case-insensitive)
# Matches any label that contains at least one of these patterns
OTHER_LABEL_KEYWORDS: list[str] = [
    "อื่น",     # Thai: other
    "ระบุ",     # Thai: specify
    "other",    # English
    "specify",  # English
    "others",
    "else",
]

OTH_SUFFIX: str = "_oth"        # Suffix pattern for open-text columns
NEW_CODE_COL: str = "New_Code"  # Column name coder fills in
SBJNUM_COL: str = "Sbjnum"     # Respondent ID column in rawdata

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# Pre-compile keyword pattern for speed
_OTHER_PATTERN = re.compile(
    "|".join(re.escape(k) for k in OTHER_LABEL_KEYWORDS),
    flags=re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def read_rawdata(excel_path: Path) -> pd.DataFrame:
    """Read the first sheet of the rawdata Excel file."""
    logger.info(f"Reading rawdata: {excel_path.name}")
    df = pd.read_excel(excel_path, dtype=str)
    logger.info(f"  → {len(df)} rows, {len(df.columns)} columns")
    return df


def find_oth_pairs(df: pd.DataFrame) -> list[tuple[str, str]]:
    """
    Detect pairs of (question_col, oth_col) by scanning for columns
    that end with OTH_SUFFIX and whose base name exists in df.columns.
    Returns list of (q_col, oth_col) tuples in original column order.
    """
    pairs: list[tuple[str, str]] = []
    for col in df.columns:
        if col.endswith(OTH_SUFFIX):
            base = col[: -len(OTH_SUFFIX)]
            if base in df.columns:
                pairs.append((base, col))
    logger.info(f"  → Found {len(pairs)} Q/_oth pair(s): {[p[0] for p in pairs]}")
    return pairs


def read_spss_labels(
    spss_path: Path,
) -> tuple[dict[str, str], dict[str, dict]]:
    """
    Read SPSS .sav file.
    Returns:
        variable_labels : {var_name: variable_label_text}
        value_labels    : {var_name: {code: label_text}}
    """
    logger.info(f"Reading SPSS labels: {spss_path.name}")
    _, meta = pyreadstat.read_sav(
        str(spss_path),
        apply_value_formats=False,
        formats_as_category=False,
    )
    var_label_map: dict[str, str] = dict(
        zip(meta.column_names, meta.column_labels, strict=False)
    )
    return var_label_map, meta.variable_value_labels


def detect_other_codes(
    q_col: str,
    value_labels: dict[str, dict],
) -> list[str]:
    """
    For a given question column, scan its Value Labels in SPSS and return
    all codes whose label text matches the OTHER_LABEL_KEYWORDS pattern.

    Returns list of code strings (e.g. ["98"], ["97", "99"]).
    Returns [] if no match found — column will be skipped.
    """
    vl = value_labels.get(q_col, {})
    matched: list[str] = []
    for code, label in vl.items():
        if _OTHER_PATTERN.search(str(label)):
            matched.append(str(int(code)) if isinstance(code, float) else str(code))
    return matched


# ---------------------------------------------------------------------------
# Phase 1 — Export Other Coding Sheet
# ---------------------------------------------------------------------------

def build_coding_sheet(
    df: pd.DataFrame,
    pairs: list[tuple[str, str]],
    variable_labels: dict[str, str],
    value_labels: dict[str, dict],
) -> pd.DataFrame:
    """
    For each (q_col, oth_col) pair:
      1. Auto-detect other codes from SPSS labels
      2. Collect rows where q_col value is one of those codes
      3. Build a unified coding table
    """
    records: list[dict] = []

    for q_col, oth_col in pairs:
        # --- Detect other codes for this question from SPSS ---
        other_codes = detect_other_codes(q_col, value_labels)

        if not other_codes:
            logger.info(f"  {q_col}: no 'other' label found in SPSS → skipped")
            continue

        logger.info(f"  {q_col}: other code(s) detected from SPSS = {other_codes}")

        # Build value labels reference string: "1=ชา, 2=กาแฟ, ..."
        vl = value_labels.get(q_col, {})
        labels_ref = _build_labels_ref(vl)

        # Variable label (question text from SPSS)
        var_label = variable_labels.get(q_col, "") or ""

        # Collect rows matching any other code
        col_values = df[q_col].astype(str).str.strip()
        mask = col_values.isin(other_codes)
        subset = df[mask].copy()

        if subset.empty:
            logger.info(f"  {q_col}: no rows with other code(s) {other_codes} in rawdata")
            continue

        logger.info(f"  {q_col}: {len(subset)} row(s) with other code(s) {other_codes}")

        for _, row in subset.iterrows():
            current_code = str(row[q_col]).strip()
            # Look up the label of that code for reference
            current_label = _get_label(vl, current_code)

            records.append(
                {
                    "Question": q_col,
                    "Variable_Label": var_label,
                    SBJNUM_COL: row.get(SBJNUM_COL, ""),
                    "Other_Code": current_code,
                    "Other_Label": current_label,   # e.g. "อื่นๆ ระบุ"
                    "Open_Text": row.get(oth_col, ""),
                    "Value_Labels_Reference": labels_ref,
                    NEW_CODE_COL: "",   # ← coder fills this in
                    "Remark": "",
                }
            )

    return pd.DataFrame(records)


def _build_labels_ref(vl: dict) -> str:
    """Build 'code=label, ...' reference string from value labels dict."""
    if not vl:
        return "(no labels)"
    return ", ".join(
        f"{int(k) if isinstance(k, float) else k}={v}"
        for k, v in sorted(vl.items(), key=lambda x: float(x[0]) if str(x[0]).replace(".", "").isdigit() else 0)
    )


def _get_label(vl: dict, code_str: str) -> str:
    """Return label for a given code string, searching flexibly."""
    for k, v in vl.items():
        k_str = str(int(k)) if isinstance(k, float) else str(k)
        if k_str == code_str:
            return str(v)
    return code_str


def style_coding_sheet(output_path: Path) -> None:
    """Apply Excel styling to the coding sheet for easy reading."""
    wb = load_workbook(output_path)
    ws = wb.active

    # Find column positions by header name
    col_idx: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            col_idx[str(cell.value)] = idx

    new_code_idx = col_idx.get(NEW_CODE_COL)
    open_text_idx = col_idx.get("Open_Text")
    labels_ref_idx = col_idx.get("Value_Labels_Reference")

    # Header row styling
    header_fill = PatternFill("solid", fgColor="2E4057")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Highlight New_Code column with yellow
    if new_code_idx:
        yellow = PatternFill("solid", fgColor="FFF176")
        for row in ws.iter_rows(min_row=2, min_col=new_code_idx, max_col=new_code_idx):
            for cell in row:
                cell.fill = yellow

    # Alternating row colors by Question group
    fills = [PatternFill("solid", fgColor="E3F2FD"), PatternFill("solid", fgColor="FFFFFF")]
    prev_q = None
    toggle = 0
    for row in ws.iter_rows(min_row=2):
        q_val = row[0].value
        if q_val != prev_q:
            toggle = 1 - toggle
            prev_q = q_val
        fill = fills[toggle]
        for cell in row:
            if cell.column != new_code_idx:
                cell.fill = fill
        # Wrap text for open text & labels
        if open_text_idx:
            row[open_text_idx - 1].alignment = Alignment(wrap_text=True)
        if labels_ref_idx:
            row[labels_ref_idx - 1].alignment = Alignment(wrap_text=True)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Column widths
    width_map: dict[str, int] = {
        "Question": 14,
        "Variable_Label": 35,
        SBJNUM_COL: 12,
        "Other_Code": 12,
        "Other_Label": 18,
        "Open_Text": 40,
        "Value_Labels_Reference": 50,
        NEW_CODE_COL: 14,
        "Remark": 25,
    }
    for col_name, width in width_map.items():
        if col_name in col_idx:
            letter = get_column_letter(col_idx[col_name])
            ws.column_dimensions[letter].width = width

    ws.row_dimensions[1].height = 30
    wb.save(output_path)
    logger.info(f"  Styling applied → {output_path.name}")


def phase1_export(
    rawdata_path: Path,
    spss_path: Path,
    output_path: Path,
) -> None:
    """
    Phase 1 entry point.
    Reads rawdata + SPSS, auto-detects other codes per question,
    builds the coding sheet, saves to output_path.
    """
    logger.info("=== PHASE 1: Export Other Coding Sheet ===")

    df = read_rawdata(rawdata_path)
    pairs = find_oth_pairs(df)
    if not pairs:
        logger.warning("No _oth column pairs found. Exiting.")
        return

    var_labels, val_labels = read_spss_labels(spss_path)
    coding_df = build_coding_sheet(df, pairs, var_labels, val_labels)

    if coding_df.empty:
        logger.warning("No rows matched any 'other' code. Nothing to export.")
        return

    coding_df.to_excel(output_path, index=False)
    style_coding_sheet(output_path)

    n_q = coding_df["Question"].nunique()
    logger.info(
        f"✓ Coding sheet saved: {output_path.name}  "
        f"({len(coding_df)} rows across {n_q} question(s))"
    )


# ---------------------------------------------------------------------------
# Phase 2 — Apply Recodes Back to Rawdata
# ---------------------------------------------------------------------------

def phase2_apply(
    rawdata_path: Path,
    coding_sheet_path: Path,
    output_rawdata_path: Path,
    output_log_path: Path,
) -> None:
    """
    Phase 2 entry point.
    Reads completed coding sheet → applies New_Code back to rawdata.
    Saves updated rawdata + recode log.
    """
    logger.info("=== PHASE 2: Apply Recodes to Rawdata ===")

    df = read_rawdata(rawdata_path)
    coding_df = pd.read_excel(coding_sheet_path, dtype=str)

    required_cols = {"Question", SBJNUM_COL, NEW_CODE_COL}
    missing = required_cols - set(coding_df.columns)
    if missing:
        raise ValueError(f"Coding sheet missing required columns: {missing}")

    # Separate coded vs skipped rows
    has_new_code = (
        coding_df[NEW_CODE_COL].notna()
        & (coding_df[NEW_CODE_COL].astype(str).str.strip() != "")
    )
    skipped_count = (~has_new_code).sum()
    coded_df = coding_df[has_new_code].copy()

    if skipped_count:
        logger.warning(f"  {skipped_count} row(s) skipped (New_Code is empty)")

    if SBJNUM_COL not in df.columns:
        raise ValueError(f"Column '{SBJNUM_COL}' not found in rawdata")

    # Build sbjnum → row index lookup
    sbjnum_index: dict[str, int] = {
        str(v).strip(): i for i, v in df[SBJNUM_COL].items()
    }

    log_records: list[dict] = []
    not_found: list[dict] = []

    for _, coding_row in coded_df.iterrows():
        q_col = str(coding_row["Question"]).strip()
        sbjnum = str(coding_row[SBJNUM_COL]).strip()
        new_code = str(coding_row[NEW_CODE_COL]).strip()
        open_text = str(coding_row.get("Open_Text", "")).strip()
        other_code = str(coding_row.get("Other_Code", "")).strip()
        other_label = str(coding_row.get("Other_Label", "")).strip()

        if q_col not in df.columns:
            logger.warning(f"  Column '{q_col}' not in rawdata — skipping")
            not_found.append({"Sbjnum": sbjnum, "Question": q_col, "Reason": "column not found"})
            continue

        row_idx = sbjnum_index.get(sbjnum)
        if row_idx is None:
            logger.warning(f"  Sbjnum '{sbjnum}' not found in rawdata — skipping")
            not_found.append({"Sbjnum": sbjnum, "Question": q_col, "Reason": "sbjnum not found"})
            continue

        old_val = str(df.at[row_idx, q_col]).strip()
        df.at[row_idx, q_col] = new_code

        log_records.append(
            {
                SBJNUM_COL: sbjnum,
                "Question": q_col,
                "Old_Code": old_val,
                "Old_Label (Other)": other_label,
                "New_Code": new_code,
                "Open_Text": open_text,
                "Recoded_At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
        )

    # Save updated rawdata
    df.to_excel(output_rawdata_path, index=False)
    logger.info(f"✓ Updated rawdata saved: {output_rawdata_path.name}")

    # Save recode log
    with pd.ExcelWriter(output_log_path, engine="openpyxl") as writer:
        pd.DataFrame(log_records).to_excel(writer, sheet_name="Recode_Log", index=False)
        if not_found:
            pd.DataFrame(not_found).to_excel(writer, sheet_name="Not_Found", index=False)
        if skipped_count:
            skipped_df = coding_df[~has_new_code][[SBJNUM_COL, "Question", "Open_Text"]]
            skipped_df.to_excel(writer, sheet_name="Skipped_No_Code", index=False)

    _style_log(output_log_path)

    logger.info(
        f"✓ Recode log saved: {output_log_path.name}  "
        f"({len(log_records)} recode(s) applied"
        + (f", {len(not_found)} not found" if not_found else "")
        + (f", {skipped_count} skipped)" if skipped_count else ")")
    )
    logger.info("=== PHASE 2 COMPLETE ===")


def _style_log(log_path: Path) -> None:
    """Apply green header styling to all sheets in the log file."""
    wb = load_workbook(log_path)
    sheet_colors = {"Recode_Log": "2E7D32", "Not_Found": "B71C1C", "Skipped_No_Code": "E65100"}
    for ws in wb.worksheets:
        color = sheet_colors.get(ws.title, "388E3C")
        header_fill = PatternFill("solid", fgColor=color)
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)
    wb.save(log_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _usage() -> None:
    print(
        """
Other Recode Tool v2  (SPSS-driven other code detection)
---------------------------------------------------------
Phase 1 — Export coding sheet:
  python other_recode_v2.py phase1 <rawdata.xlsx> <labels.sav> [output_coding.xlsx]

Phase 2 — Apply recodes to rawdata:
  python other_recode_v2.py phase2 <rawdata.xlsx> <coding_sheet.xlsx> [output_rawdata.xlsx] [output_log.xlsx]

Notes:
  • "Other" codes are detected automatically from SPSS Value Labels.
  • Any label containing: """ + str(OTHER_LABEL_KEYWORDS) + """
    will be treated as an "other/specify" code for that question.
"""
    )


def main() -> None:
    if len(sys.argv) < 2:
        _usage()
        sys.exit(1)

    phase = sys.argv[1].lower()

    if phase == "phase1":
        if len(sys.argv) < 4:
            _usage()
            sys.exit(1)
        rawdata = Path(sys.argv[2])
        spss = Path(sys.argv[3])
        output = (
            Path(sys.argv[4]) if len(sys.argv) > 4
            else rawdata.parent / "other_coding_sheet.xlsx"
        )
        phase1_export(rawdata, spss, output)

    elif phase == "phase2":
        if len(sys.argv) < 4:
            _usage()
            sys.exit(1)
        rawdata = Path(sys.argv[2])
        coding = Path(sys.argv[3])
        out_raw = (
            Path(sys.argv[4]) if len(sys.argv) > 4
            else rawdata.parent / "rawdata_recoded.xlsx"
        )
        out_log = (
            Path(sys.argv[5]) if len(sys.argv) > 5
            else rawdata.parent / "recode_log.xlsx"
        )
        phase2_apply(rawdata, coding, out_raw, out_log)

    else:
        print(f"Unknown phase: '{phase}'")
        _usage()
        sys.exit(1)


if __name__ == "__main__":
    main()