import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
import os
import pyreadstat # For reading SPSS files
from datetime import datetime
import openpyxl # For Excel export
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
# from openpyxl.comments import Comment # Not used yet, but can be for cell comments
import webbrowser
import tkinter as tk
import sys
import threading # For background tasks
import time # For potential small delays if needed
import traceback # For detailed error logging

# --- Imports for Google Sheets ---
import gspread
from google.oauth2.service_account import Credentials # Recommended for google-auth
# ---------------------------------


# Set default theme for customtkinter
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class SpssExporterApp(ctk.CTk):
    MAX_DISPLAY_ITEMS_IN_DIALOG = 100 # Max items to display in dialog lists

    def _center_window(self, window_to_center, width, height):
        """Centers the given window on the screen."""
        window_to_center.update_idletasks()
        screen_width = window_to_center.winfo_screenwidth()
        screen_height = window_to_center.winfo_screenheight()
        x_coordinate = int((screen_width / 2) - (width / 2))
        y_coordinate = int((screen_height / 2) - (height / 2))
        window_to_center.geometry(f"{width}x{height}+{x_coordinate}+{y_coordinate}")

    def __init__(self):
        super().__init__()

        self.title("เก็บข้อมูล Norm For DP")

        app_width = 1000
        app_height = 750 
        self._center_window(self, app_width, app_height)

        self.label_font_bold = ctk.CTkFont(weight="bold")
        self.panel_label_font_bold = ctk.CTkFont(size=14, weight="bold")
        self.dialog_info_font = ctk.CTkFont(size=10)
        
        # Fonts and colors for Log Dialog
        self.log_font_family = "Tahoma" # Or "Segoe UI", "Calibri" etc.
        self.log_font_size = 12 
        self.log_font = ctk.CTkFont(family=self.log_font_family, size=self.log_font_size)
        self.log_success_color = "green" 
        self.log_error_color = "red" 
        self.log_warning_color = "orange"


        self.spss_df_codes = None
        self.spss_df_labels = None
        self.spss_meta = None
        self.spss_variables = []
        self.spss_load_results = {"success": False} 

        # --- Google Sheet URL and Settings ---
        self.TARGET_GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1CU0KjP3eAYRuJk7uz5rBR0GQ8uDUH4zuYnr6rnSKMwU/edit?gid=106958934#gid=106958934"
        self.SERVICE_ACCOUNT_FILE_PATH = 'Test3.json' 
        self.INDEX_SHEET_NAME_FOR_COUNT = "Index"
        self.UNIQUE_JOB_IDENTIFIER_COLUMN = "Project #"
        self.TARGET_Norm_URL = "https://script.google.com/macros/s/AKfycbxv0suQxHaDNttsaHpBIo2zqf-QY8IjaXKloqYoKmKSSbOUmwLeNUxZdH7-bxQr8dOz/exec"

        self.current_year_str = str(datetime.now().year)
        year_implemented_options = [str(y) for y in range(int(self.current_year_str) - 4, int(self.current_year_str) + 5)]
        survey_method_options = ["CLT", "HUT", "Online survey", "D2D", "Intercept", "CATI", "Others"]
        category_options = [
            "Food","Snacks","Beverages","Alchoholic beverage","Household care",
            "Female Cosmetics and Fragrance","Male Cosmetics and Fragrance",
            "Facial care","Body & Haircare","Sanitary products","Diaper for baby",
            "Other baby care products","Diaper for adult","Pet care products",
            "Healthcare (OTC, functional foods & others)","Tobacco","Home appliances",
            "Banking","Telecom","Real Estate","Advertisement","Services Others","Others"
        ]
        evaluation_target_options = ["Concept","Product","Taste","Fragrance","Package","Brand","Ad/Campaign","Service"]
        scale_options = ["3-point scale", "4-point scale", "5-point scale","7-point scale","9-point scale","10-point scale", "11-point scale"]
        self.username_placeholder = "" 
        username_options_with_placeholder = [self.username_placeholder] + ["Songklod","Kanokrat","Pornpan","Ittichai","Chanida"]

        self.fixed_input_fields = {
            "Username": {"type": "combobox", "options": username_options_with_placeholder, "widget": None, "var": None, "default": self.username_placeholder},
            "Timestamp": {"type": "entry", "widget": None, "var": None, "default": "Auto (on export)", "hidden": True},
            "Project #": {"type": "entry", "widget": None, "var": None, "default": ""},
            "Project Name": {"type": "entry", "widget": None, "var": None, "default": ""},
            "Year Implemented": {"type": "combobox", "options": year_implemented_options, "widget": None, "var": None, "default": self.current_year_str},
            "Survey Method": {"type": "multiselect_dialog_button", "options": survey_method_options, "widget_button": None, "widget_display": None, "var": None, "selected_values": [], "default_selected": []},
            "Category": {"type": "multiselect_dialog_button", "options": category_options, "widget_button": None, "widget_display": None, "var": None, "selected_values": [], "default_selected": []},
            "Evaluation Target": {"type": "multiselect_dialog_button", "options": evaluation_target_options, "widget_button": None, "widget_display": None, "var": None, "selected_values": [], "default_selected": []},
            "Scale": {"type": "combobox", "options": scale_options, "widget": None, "var": None, "default": "5-point scale"},
        }

        self.demographic_fields = ["Gender", "Age (NA)", "Age (Range)", "Area", "SES"]
        self.demographic_var_selections = {field: None for field in self.demographic_fields}
        self.demographic_widgets = {}

        self.channel_custom_names = [
            "01 PI No price", "02 PI Price", "03 OL", "04 Attractive", "05 Overall Appealing",
            "06 Usage Intention", "07 Newness", "08 Uniqueness", "09 Relevancy", "10 Credibility",
            "11 Concept Match", "12 Compare Current", "13 Overall Satisfaction", "14 Fit Brand", "15 Recommend"
        ]
        self.num_channels = len(self.channel_custom_names)
        self.channel_selections = {i: [] for i in range(self.num_channels)}
        self.channel_widgets = []

        self.total_jobs_var = ctk.StringVar(value="Total Jobs: Loading...")

        self.spss_load_completion_event = threading.Event()
        self.current_job_project_name_for_prompt = ""
        self.is_batch_processing_active = False 

        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=2)
        self.grid_rowconfigure(1, weight=1)

        top_frame = ctk.CTkFrame(self)
        top_frame.grid(row=0, column=0, columnspan=2, padx=10, pady=(10,5), sticky="ew")

        self.load_button = ctk.CTkButton(top_frame, text="Load SPSS File (.sav)", command=self.load_spss_file)
        self.load_button.pack(side="left", padx=5, pady=5)

        self.total_jobs_label = ctk.CTkLabel(top_frame, textvariable=self.total_jobs_var, font=self.label_font_bold, text_color="red")
        self.total_jobs_label.pack(side="right", padx=(10, 5), pady=5)

        self.file_label = ctk.CTkLabel(top_frame, text="No file loaded.", anchor="w")
        self.file_label.pack(side="left", padx=5, pady=5, fill="x", expand=True)

        left_panel = ctk.CTkScrollableFrame(self, label_text="Project Info & Demographics", label_font=self.panel_label_font_bold)
        left_panel.grid(row=1, column=0, padx=(10, 5), pady=5, sticky="nsew")
        
        for field_name, config in self.fixed_input_fields.items():
            if config.get("hidden", False):
                config["var"] = ctk.StringVar(value=config.get("default", ""))
                continue
            item_frame = ctk.CTkFrame(left_panel); item_frame.pack(fill="x", padx=5, pady=(3, 4))
            label = ctk.CTkLabel(item_frame, text=field_name + ":", width=140, anchor="w", font=self.label_font_bold); label.pack(side="left", padx=(0,5))
            config["var"] = ctk.StringVar(); default_value_to_set = config.get("default", "")
            if config["type"] == "entry":
                widget = ctk.CTkEntry(item_frame, textvariable=config["var"]); config["var"].set(default_value_to_set); widget.pack(side="left", fill="x", expand=True); config["widget"] = widget
            elif config["type"] == "combobox":
                options = config.get("options", []); widget = ctk.CTkComboBox(item_frame, variable=config["var"], values=options, state="readonly")
                if default_value_to_set in options: config["var"].set(default_value_to_set)
                elif options: config["var"].set(options[0])
                else: config["var"].set("")
                widget.pack(side="left", fill="x", expand=True); config["widget"] = widget
            elif config["type"] == "multiselect_dialog_button":
                config["selected_values"] = list(config.get("default_selected", [])); config["var"].set(", ".join(config["selected_values"]))
                display_label = ctk.CTkEntry(item_frame, textvariable=config["var"], state="disabled", fg_color="gray80"); display_label.pack(side="left", fill="x", expand=True, padx=(0,5)); config["widget_display"] = display_label
                button = ctk.CTkButton(item_frame, text="Select...", width=80, command=lambda fn=field_name: self.open_multiselect_dialog(fn)); button.pack(side="left"); config["widget_button"] = button
        
        for field_name in self.demographic_fields:
            item_frame = ctk.CTkFrame(left_panel); item_frame.pack(fill="x", padx=5, pady=(3, 4))
            label = ctk.CTkLabel(item_frame, text=field_name + ":", width=140, anchor="w", font=self.label_font_bold); label.pack(side="left", padx=(0,5))
            display_spss_var_label = ctk.CTkLabel(item_frame, text="None selected", anchor="w"); display_spss_var_label.pack(side="left", fill="x", expand=True, padx=5)
            select_spss_var_button = ctk.CTkButton(item_frame, text="Select Var", width=100, command=lambda fn=field_name, dv_label=display_spss_var_label: self.open_demographic_var_selection_dialog(fn, dv_label)); select_spss_var_button.pack(side="left")
            self.demographic_widgets[field_name] = {"display": display_spss_var_label, "button": select_spss_var_button}

        right_panel = ctk.CTkScrollableFrame(self, label_text="Concept Evaluation Channels (1-15)", label_font=self.panel_label_font_bold)
        right_panel.grid(row=1, column=1, padx=(5, 10), pady=5, sticky="nsew")
        
        for i in range(self.num_channels):
            channel_frame = ctk.CTkFrame(right_panel); channel_frame.pack(fill="x", padx=5, pady=(3,4))
            label_text = self.channel_custom_names[i] + ":"; channel_label_widget = ctk.CTkLabel(channel_frame, text=label_text, anchor="w", width=170, font=self.label_font_bold); channel_label_widget.pack(side="left", padx=(0,5))
            selected_vars_display_textbox = ctk.CTkTextbox(channel_frame, height=40, state="disabled", wrap="word"); selected_vars_display_textbox.pack(side="left", fill="x", expand=True, padx=5)
            select_channel_vars_button = ctk.CTkButton(channel_frame, text="Select Vars", width=100, command=lambda ch_idx=i: self.open_channel_var_selection_dialog(ch_idx)); select_channel_vars_button.pack(side="left")
            self.channel_widgets.append({"display": selected_vars_display_textbox})

        export_button_frame = ctk.CTkFrame(self)
        export_button_frame.grid(row=2, column=0, columnspan=2, pady=(5,10), sticky="ew")
        for i in range(6): export_button_frame.grid_columnconfigure(i, weight=1)

        self.load_excel_process_button = ctk.CTkButton(export_button_frame, text="Load Excel & Process Jobs", command=self.process_jobs_from_excel_threaded, height=35, fg_color="#4A00E0", hover_color="#8E2DE2")
        self.load_excel_process_button.grid(row=0, column=0, padx=(10,5), pady=10, sticky="ew")
        self.export_excel_button = ctk.CTkButton(export_button_frame, text="Export to Excel", command=self.export_to_excel, height=35)
        self.export_excel_button.grid(row=0, column=1, padx=5, pady=10, sticky="ew")
        self.report_button = ctk.CTkButton(export_button_frame, text="Dashboard Report", command=self.open_Dashboard_link, height=35, fg_color="#00796B", hover_color="#004D40")
        self.report_button.grid(row=0, column=2, padx=5, pady=10, sticky="ew")
        self.save_template_button = ctk.CTkButton(export_button_frame, text="Save Excel Template", command=self.save_excel_template, height=35, fg_color="#FF8C00", hover_color="#FFA500")
        self.save_template_button.grid(row=0, column=3, padx=5, pady=10, sticky="ew")
        self.open_gsheet_button = ctk.CTkButton(export_button_frame, text="คลิกเพื่อเปิด Google Sheet", command=self.open_google_sheet_link, height=35, fg_color="#F00A0A", hover_color="#EB4646")
        self.open_gsheet_button.grid(row=0, column=4, padx=5, pady=10, sticky="ew")
        self.gspread_button = ctk.CTkButton(export_button_frame, text="Save To Google Sheet", command=self.save_to_google_sheet, height=35, fg_color="darkgreen", hover_color="#179F05")
        self.gspread_button.grid(row=0, column=5, padx=(5,10), pady=10, sticky="ew")

        self._clear_all_selections_and_inputs()
        self.fetch_total_jobs_from_gsheet_threaded()

    def _clear_all_selections_and_inputs(self, clear_project_info=True, clear_demographics=True, clear_channels=True):
        if clear_project_info:
            for field_name, config in self.fixed_input_fields.items():
                if config.get("hidden", False):
                    if "var" in config and config["var"] is not None: config["var"].set(config.get("default", ""))
                    continue
                if "var" in config and config["var"] is not None:
                    default_value = config.get("default", "")
                    if config["type"] == "combobox" and not default_value and config.get("options"): default_value = config["options"][0]
                    elif config["type"] == "multiselect_dialog_button":
                        config["selected_values"] = list(config.get("default_selected", []))
                        default_value = ", ".join(config["selected_values"])
                    config["var"].set(default_value)
        if clear_demographics:
            self.demographic_var_selections = {field: None for field in self.demographic_fields}
            for field_name in self.demographic_fields:
                if field_name in self.demographic_widgets and "display" in self.demographic_widgets[field_name]:
                    self.demographic_widgets[field_name]["display"].configure(text="None selected")
        if clear_channels:
            self.channel_selections = {i: [] for i in range(self.num_channels)}
            for i in range(self.num_channels):
                self.update_channel_display_text(i)

    def open_google_sheet_link(self):
        try:
            if self.TARGET_GOOGLE_SHEET_URL: webbrowser.open_new_tab(self.TARGET_GOOGLE_SHEET_URL)
            else: messagebox.showwarning("No URL", "Target Google Sheet URL is not defined.", parent=self)
        except Exception as e: messagebox.showerror("Error Opening Link", f"Could not open the Google Sheet link:\n{str(e)}", parent=self)

    def open_Dashboard_link(self):
        try:
            if self.TARGET_Norm_URL: webbrowser.open_new_tab(self.TARGET_Norm_URL)
            else: messagebox.showwarning("No URL", "Target Dashboard URL is not defined.", parent=self)
        except Exception as e: messagebox.showerror("Error Opening Link", f"Could not open the Dashboard link:\n{str(e)}", parent=self)

    def load_spss_file(self):
        self.loading_dialog_instance_created_flag = True
        file_path = filedialog.askopenfilename(title="Select SPSS File", filetypes=(("SPSS files", "*.sav"), ("All files", "*.*")))
        if not file_path:
            if hasattr(self, 'loading_dialog_instance_created_flag'): del self.loading_dialog_instance_created_flag
            return

        self.loading_dialog = ctk.CTkToplevel(self)
        self.loading_dialog.title("Loading SPSS")
        self._center_window(self.loading_dialog, 350, 120)
        self.loading_dialog.grab_set()
        self.loading_dialog.protocol("WM_DELETE_WINDOW", lambda: None)
        ctk.CTkLabel(self.loading_dialog, text="กำลังโหลดไฟล์ SPSS กรุณารอสักครู่...\nขั้นตอนนี้อาจใช้เวลาหลายนาทีสำหรับไฟล์ขนาดใหญ่",
                     font=self.label_font_bold, justify="center").pack(pady=20, padx=20)
        self.loading_dialog.update()

        self.spss_load_results = {"df_codes": None, "meta": None, "success": False, "used_encoding": "N/A", "last_exception": None}

        def _load_file_in_thread_interactive():
            encodings_to_try = [None, 'utf-8', 'cp874', 'tis-620'] 
            loaded_this_try = False
            for enc in encodings_to_try:
                try:
                    current_encoding_name = enc if enc else 'auto-detect by pyreadstat'
                    print(f"Attempting to load SPSS with encoding: {current_encoding_name}")
                    temp_df_codes, temp_meta = pyreadstat.read_sav(file_path, apply_value_formats=False, user_missing=False, encoding=enc)
                    if temp_df_codes is not None and not temp_df_codes.empty and temp_meta is not None:
                        self.spss_load_results.update({"df_codes": temp_df_codes, "meta": temp_meta, "success": True, "used_encoding": current_encoding_name})
                        loaded_this_try = True; break
                    else: self.spss_load_results["last_exception"] = RuntimeError(f"Encoding {current_encoding_name} resulted in empty/invalid data.")
                except Exception as e: self.spss_load_results["last_exception"] = e; print(f"Error with encoding {current_encoding_name}: {e}")
            if not loaded_this_try: print(f"Failed to load SPSS file '{os.path.basename(file_path)}' after trying all encodings.")
            self.after(0, self._finish_loading_spss, file_path)
        
        load_thread = threading.Thread(target=_load_file_in_thread_interactive, daemon=True)
        load_thread.start()

    def _finish_loading_spss(self, file_path_loaded):
        if hasattr(self, 'loading_dialog') and self.loading_dialog.winfo_exists():
            self.loading_dialog.destroy()

        loaded_successfully = self.spss_load_results.get("success", False)
        used_encoding = self.spss_load_results.get("used_encoding", "N/A")
        last_exception = self.spss_load_results.get("last_exception")

        if not loaded_successfully:
            error_message = f"Failed to load SPSS file: {os.path.basename(file_path_loaded)}."
            if last_exception: error_message += f"\nLast error ({type(last_exception).__name__}): {str(last_exception)}"
            if not (hasattr(self, 'is_batch_processing_active') and self.is_batch_processing_active):
                 messagebox.showerror("Error Loading File", error_message, parent=self)
            self.spss_df_codes, self.spss_df_labels, self.spss_meta, self.spss_variables = None, None, None, []
            self.file_label.configure(text="Failed to load file.")
            self._clear_all_selections_and_inputs(True, True, True)
            return False

        self.spss_df_codes = self.spss_load_results["df_codes"]
        self.spss_meta = self.spss_load_results["meta"]

        if self.spss_df_codes is not None and self.spss_meta is not None:
            self.spss_df_labels = self.spss_df_codes.copy()
            if hasattr(self.spss_meta, 'variable_value_labels') and self.spss_meta.variable_value_labels:
                for col_name, label_dict in self.spss_meta.variable_value_labels.items():
                    if col_name in self.spss_df_labels.columns and label_dict:
                        try: mapped_series = self.spss_df_labels[col_name].map(label_dict)
                        except TypeError:
                            try: mapped_series = self.spss_df_labels[col_name].astype(str).map({str(k): v for k, v in label_dict.items()})
                            except Exception: mapped_series = pd.Series([None] * len(self.spss_df_labels[col_name]), index=self.spss_df_labels[col_name].index)
                        self.spss_df_labels[col_name] = mapped_series.fillna(self.spss_df_codes[col_name].astype(str)).astype(str)
            else: self.spss_df_labels = self.spss_df_codes.astype(str)
        else: self.spss_df_labels = None

        self.spss_variables = list(self.spss_df_codes.columns) if self.spss_df_codes is not None else []
        self.file_label.configure(text=f"Loaded: {os.path.basename(file_path_loaded)} ({len(self.spss_variables)} vars) Encoding: {used_encoding}")
        self._clear_all_selections_and_inputs(True, True, True)

        if hasattr(self, 'loading_dialog_instance_created_flag'):
            messagebox.showinfo("Success", f"SPSS file loaded successfully using encoding '{used_encoding}'.", parent=self)
            del self.loading_dialog_instance_created_flag
        return True

    def spss_df_is_loaded(self):
        if self.spss_df_codes is None or self.spss_df_labels is None:
            messagebox.showwarning("No Data", "Please load an SPSS file first.", parent=self)
            return False
        return True

    def open_demographic_var_selection_dialog(self, field_name_ref, display_widget_ref):
        if not self.spss_df_is_loaded(): return
        dialog = ctk.CTkToplevel(self)
        dialog.title(f"Select SPSS Variable for {field_name_ref}")
        dialog_width, dialog_height = 500, 450
        self._center_window(dialog, dialog_width, dialog_height)
        dialog.grab_set()

        search_var = ctk.StringVar()
        selected_var_temp = ctk.StringVar(value=self.demographic_var_selections.get(field_name_ref, ""))

        ctk.CTkLabel(dialog, text="Search Variable:").pack(pady=(10,0))
        search_entry = ctk.CTkEntry(dialog, textvariable=search_var)
        search_entry.pack(pady=5, padx=10, fill="x")

        listbox_frame = ctk.CTkFrame(dialog)
        listbox_frame.pack(pady=5, padx=10, fill="both", expand=True)
        
        listbox = tk.Listbox(listbox_frame, selectmode="browse", exportselection=False, font=("Arial", 10)) 
        scrollbar_y = ctk.CTkScrollbar(listbox_frame, command=listbox.yview)
        scrollbar_x = ctk.CTkScrollbar(listbox_frame, command=listbox.xview, orientation="horizontal")
        listbox.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
        listbox.pack(side="left", fill="both", expand=True)

        all_vars_with_labels = []
        if hasattr(self.spss_meta, 'column_names_to_labels'):
            for var in self.spss_variables:
                label = self.spss_meta.column_names_to_labels.get(var, "No label")
                all_vars_with_labels.append(f"{var} -- {label}")
        else:
            all_vars_with_labels = self.spss_variables[:]


        def populate_listbox(filter_text=""):
            listbox.delete(0, "end")
            count = 0
            for item_text in all_vars_with_labels:
                if filter_text.lower() in item_text.lower():
                    listbox.insert("end", item_text)
                    count += 1
                    if count >= self.MAX_DISPLAY_ITEMS_IN_DIALOG and filter_text == "": 
                        listbox.insert("end", f"... and {len(all_vars_with_labels) - count} more (type to filter)")
                        break
            if selected_var_temp.get(): 
                try:
                    full_item_to_select = next((item for item in all_vars_with_labels if item.startswith(selected_var_temp.get() + " --")), None)
                    if full_item_to_select and full_item_to_select in listbox.get(0, "end"):
                        idx = listbox.get(0, "end").index(full_item_to_select)
                        listbox.selection_set(idx)
                        listbox.activate(idx)
                        listbox.see(idx)
                except (ValueError, tk.TclError): pass 

        populate_listbox()
        search_var.trace_add("write", lambda name, index, mode, sv=search_var: populate_listbox(sv.get()))

        def on_select():
            selected_indices = listbox.curselection()
            if selected_indices:
                full_selected_item = listbox.get(selected_indices[0])
                if "..." not in full_selected_item: 
                    actual_var_name = full_selected_item.split(" -- ")[0]
                    self.demographic_var_selections[field_name_ref] = actual_var_name
                    display_widget_ref.configure(text=actual_var_name)
            dialog.destroy()

        def on_clear():
            self.demographic_var_selections[field_name_ref] = None
            display_widget_ref.configure(text="None selected")
            dialog.destroy()

        button_frame = ctk.CTkFrame(dialog); button_frame.pack(pady=10, fill="x", padx=10)
        select_button = ctk.CTkButton(button_frame, text="Select", command=on_select); select_button.pack(side="left", padx=5, expand=True)
        clear_button = ctk.CTkButton(button_frame, text="Clear Selection", command=on_clear, fg_color="gray"); clear_button.pack(side="left", padx=5, expand=True)
        cancel_button = ctk.CTkButton(button_frame, text="Cancel", command=dialog.destroy); cancel_button.pack(side="right", padx=5, expand=True)

    def open_multiselect_dialog(self, field_name_ref):
        config = self.fixed_input_fields.get(field_name_ref)
        if not config or config["type"] != "multiselect_dialog_button": return
        
        dialog = ctk.CTkToplevel(self)
        dialog.title(f"Select {field_name_ref}")
        dialog_width, dialog_height = 400, 350
        self._center_window(dialog, dialog_width, dialog_height)
        dialog.grab_set()

        options = config.get("options", [])
        current_selections = list(config.get("selected_values", [])) 

        checkbox_vars = {}
        scrollable_frame = ctk.CTkScrollableFrame(dialog, label_text="Available Options")
        scrollable_frame.pack(pady=10, padx=10, fill="both", expand=True)

        for option_text in options:
            var = ctk.StringVar(value="on" if option_text in current_selections else "off")
            cb = ctk.CTkCheckBox(scrollable_frame, text=option_text, variable=var, onvalue="on", offvalue="off")
            cb.pack(anchor="w", padx=10, pady=2)
            checkbox_vars[option_text] = var

        def on_confirm_multiselect():
            new_selections = [opt for opt, var_cb in checkbox_vars.items() if var_cb.get() == "on"]
            config["selected_values"] = new_selections
            config["var"].set(", ".join(new_selections) if new_selections else "")
            dialog.destroy()

        button_frame = ctk.CTkFrame(dialog); button_frame.pack(pady=10, fill="x", padx=10)
        confirm_button = ctk.CTkButton(button_frame, text="Confirm", command=on_confirm_multiselect); confirm_button.pack(side="left", padx=5, expand=True)
        cancel_button = ctk.CTkButton(button_frame, text="Cancel", command=dialog.destroy); cancel_button.pack(side="right", padx=5, expand=True)

    def open_channel_var_selection_dialog(self, channel_index):
        if not self.spss_df_is_loaded(): return
        
        dialog = ctk.CTkToplevel(self)
        dialog_title = f"Select Variables for: {self.channel_custom_names[channel_index]}"
        dialog.title(dialog_title)
        dialog_width, dialog_height = 750, 650 
        self._center_window(dialog, dialog_width, dialog_height)
        dialog.grab_set()

        search_var = ctk.StringVar()
        initially_selected_vars_info = self.channel_selections.get(channel_index, [])
        temp_selected_var_details = {info['display_text_short']: info for info in initially_selected_vars_info}
        reverse_var = ctk.StringVar(value="off") 
        
        all_spss_vars_display = [] 
        if hasattr(self.spss_meta, 'column_names_to_labels'):
            for var_spss in self.spss_variables:
                label_spss = self.spss_meta.column_names_to_labels.get(var_spss, "No label")
                all_spss_vars_display.append(f"{var_spss} -- {label_spss}")
        else:
            all_spss_vars_display = self.spss_variables[:]

        available_listbox = None 
        selected_listbox = None
        value_codes_display_textbox = None 

        def populate_available_listbox(filter_text=""):
            nonlocal available_listbox 
            if available_listbox is None: return 
            available_listbox.delete(0, "end")
            current_selected_actual_names = [info['name'] for info in temp_selected_var_details.values()]
            count = 0
            for item_text_spss in all_spss_vars_display:
                actual_var_name_spss = item_text_spss.split(" -- ")[0]
                if actual_var_name_spss not in current_selected_actual_names: 
                    if filter_text.lower() in item_text_spss.lower():
                        available_listbox.insert("end", item_text_spss)
                        count += 1
                        if count >= self.MAX_DISPLAY_ITEMS_IN_DIALOG and filter_text == "":
                            available_listbox.insert("end", f"... and {len(all_spss_vars_display) - count} more (type to filter).")
                            break
        
        def populate_selected_listbox():
            nonlocal selected_listbox 
            if selected_listbox is None: return  
            selected_listbox.delete(0, "end")
            for display_text_short in temp_selected_var_details.keys():
                 selected_listbox.insert("end", display_text_short)

        def transfer_vars(adding_to_selected):
            nonlocal available_listbox, selected_listbox, reverse_var, search_var 
            if available_listbox is None or selected_listbox is None: return 

            if adding_to_selected:
                source_lb = available_listbox
                indices_to_move = source_lb.curselection()
                if not indices_to_move: return
                items_to_add_info = []
                for idx_av in indices_to_move:
                    full_item_text_av = source_lb.get(idx_av)
                    if "..." in full_item_text_av: continue 
                    actual_spss_name_av = full_item_text_av.split(" -- ")[0]
                    _, min_s, max_s, has_valid_s = self.get_variable_display_info(actual_spss_name_av)
                    should_reverse_av = reverse_var.get() == "on" and has_valid_s
                    display_text_short_av = f"{actual_spss_name_av}{' (R)' if should_reverse_av else ''}"
                    var_info_av = {'name': actual_spss_name_av, 'reverse': should_reverse_av, 
                                   'min_scale': min_s, 'max_scale': max_s, 
                                   'display_text_short': display_text_short_av}
                    items_to_add_info.append(var_info_av)
                for info_av in items_to_add_info:
                    if info_av['display_text_short'] not in temp_selected_var_details:
                         temp_selected_var_details[info_av['display_text_short']] = info_av
            else: 
                source_lb = selected_listbox
                indices_to_remove = source_lb.curselection()
                if not indices_to_remove: return
                items_to_remove_display_text = [source_lb.get(idx_sel) for idx_sel in reversed(indices_to_remove)]
                for display_text_sel in items_to_remove_display_text:
                    if display_text_sel in temp_selected_var_details:
                        del temp_selected_var_details[display_text_sel]
            populate_available_listbox(search_var.get())
            populate_selected_listbox()
            if value_codes_display_textbox:
                value_codes_display_textbox.configure(state="normal")
                value_codes_display_textbox.delete("1.0", "end")
                value_codes_display_textbox.insert("1.0", "Value Codes: (Select a variable from 'Available' list)")
                value_codes_display_textbox.configure(state="disabled")


        def update_selected_list_on_reverse_toggle():
            nonlocal selected_listbox, reverse_var 
            if selected_listbox is None: return 
            selected_indices_in_sel_lb = selected_listbox.curselection()
            if not selected_indices_in_sel_lb: return

            new_temp_selected_details = {}
            current_selection_texts_in_lb = [selected_listbox.get(i) for i in selected_indices_in_sel_lb]

            for old_display_text, info_dict in list(temp_selected_var_details.items()):
                if old_display_text in current_selection_texts_in_lb:
                    _, min_s, max_s, has_valid_s = self.get_variable_display_info(info_dict['name'])
                    new_reverse_status = reverse_var.get() == "on" and has_valid_s
                    if info_dict['reverse'] != new_reverse_status :
                        new_display_text = f"{info_dict['name']}{' (R)' if new_reverse_status else ''}"
                        new_info = info_dict.copy()
                        new_info['reverse'] = new_reverse_status
                        new_info['display_text_short'] = new_display_text
                        new_temp_selected_details[new_display_text] = new_info 
                    else:
                        new_temp_selected_details[old_display_text] = info_dict
                else:
                    new_temp_selected_details[old_display_text] = info_dict
            temp_selected_var_details.clear()
            temp_selected_var_details.update(new_temp_selected_details)
            populate_selected_listbox()

        def on_selected_listbox_selection_change(event):
            nonlocal selected_listbox, reverse_var 
            if selected_listbox is None: return 
            sel_indices = selected_listbox.curselection()
            if sel_indices:
                first_selected_item_text = selected_listbox.get(sel_indices[0])
                item_info = temp_selected_var_details.get(first_selected_item_text)
                if item_info:
                    reverse_var.set("on" if item_info['reverse'] else "off")
        
        def on_available_var_select(event):
            nonlocal available_listbox, value_codes_display_textbox
            if available_listbox is None or value_codes_display_textbox is None: return

            selected_indices = available_listbox.curselection()
            value_codes_display_textbox.configure(state="normal")
            value_codes_display_textbox.delete("1.0", "end")

            if not selected_indices:
                value_codes_display_textbox.insert("1.0", "Value Codes: (Select a variable)")
                value_codes_display_textbox.configure(state="disabled")
                return

            full_item_text = available_listbox.get(selected_indices[0])
            if "..." in full_item_text: 
                value_codes_display_textbox.insert("1.0", "Value Codes: (Select a specific variable)")
                value_codes_display_textbox.configure(state="disabled")
                return
            
            actual_spss_name = full_item_text.split(" -- ")[0]
            
            value_labels_dict = {}
            if self.spss_meta and hasattr(self.spss_meta, 'variable_value_labels'):
                value_labels_dict = self.spss_meta.variable_value_labels.get(actual_spss_name, {})

            if value_labels_dict:
                sorted_codes = []
                try:
                    sorted_codes = sorted(value_labels_dict.items(), key=lambda item: float(item[0]) if str(item[0]).replace('.','',1).isdigit() else str(item[0]))
                except ValueError: 
                     sorted_codes = sorted(value_labels_dict.items(), key=lambda item: str(item[0]))

                formatted_codes = [f"{code}: {label}" for code, label in sorted_codes]
                
                display_text_vc = f"Value Codes for {actual_spss_name}:\n" + "\n".join(formatted_codes)
                value_codes_display_textbox.insert("1.0", display_text_vc)
            else:
                value_codes_display_textbox.insert("1.0", f"Value Codes for {actual_spss_name}:\n(No value labels defined)")
            value_codes_display_textbox.configure(state="disabled")


        def on_confirm_channel_selection(dialog_ref, ch_idx, final_selected_details):
            self.channel_selections[ch_idx] = list(final_selected_details.values())
            self.update_channel_display_text(ch_idx)
            dialog_ref.destroy()

        top_controls_frame = ctk.CTkFrame(dialog); top_controls_frame.pack(pady=(10,5), padx=10, fill="x")
        ctk.CTkLabel(top_controls_frame, text="Search Available:").pack(side="left", padx=(0,5))
        search_entry = ctk.CTkEntry(top_controls_frame, textvariable=search_var); search_entry.pack(side="left", padx=5, fill="x", expand=True)

        main_paned_window = tk.PanedWindow(dialog, orient="horizontal", sashrelief="raised", sashwidth=6, bd=1) 
        main_paned_window.pack(fill="both", expand=True, padx=10, pady=5)

        left_pane_container = ctk.CTkFrame(main_paned_window, width=320) 
        main_paned_window.add(left_pane_container, stretch="first")

        available_frame = ctk.CTkFrame(left_pane_container) 
        available_frame.pack(fill="both", expand=True, pady=(0,5), padx=2)
        ctk.CTkLabel(available_frame, text="Available SPSS Variables:", font=self.label_font_bold).pack(pady=(5,2))
        
        available_listbox = tk.Listbox(available_frame, selectmode="extended", exportselection=False, font=("Arial", 10))
        av_scrollbar_y = ctk.CTkScrollbar(available_frame, command=available_listbox.yview)
        av_scrollbar_x = ctk.CTkScrollbar(available_frame, command=available_listbox.xview, orientation="horizontal")
        available_listbox.configure(yscrollcommand=av_scrollbar_y.set, xscrollcommand=av_scrollbar_x.set)
        av_scrollbar_y.pack(side="right", fill="y"); av_scrollbar_x.pack(side="bottom", fill="x")
        available_listbox.pack(side="left", fill="both", expand=True)
        
        value_codes_frame = ctk.CTkFrame(left_pane_container)
        value_codes_frame.pack(fill="x", pady=(5,0), padx=2)
        ctk.CTkLabel(value_codes_frame, text="Value Codes:", font=self.dialog_info_font).pack(anchor="w", pady=(0,2))
        value_codes_display_textbox = ctk.CTkTextbox(value_codes_frame, height=100, state="disabled", wrap="word", font=("Arial", 9)) 
        value_codes_display_textbox.pack(fill="x", expand=True)
        value_codes_display_textbox.configure(state="normal")
        value_codes_display_textbox.insert("1.0", "Select a variable from 'Available' list to see its codes.")
        value_codes_display_textbox.configure(state="disabled")

        action_buttons_frame = ctk.CTkFrame(main_paned_window, width=110) 
        main_paned_window.add(action_buttons_frame) 
        add_button = ctk.CTkButton(action_buttons_frame, text="Add >>", width=90, command=lambda: transfer_vars(True))
        add_button.pack(pady=(max(dialog_height // 5, 50), 10), padx=5) 
        remove_button = ctk.CTkButton(action_buttons_frame, text="<< Remove", width=90, command=lambda: transfer_vars(False))
        remove_button.pack(pady=10, padx=5)
        
        reverse_scale_frame = ctk.CTkFrame(action_buttons_frame) 
        reverse_scale_frame.pack(pady=(30,0), padx=5)
        ctk.CTkLabel(reverse_scale_frame, text="Reverse Scale:", font=self.dialog_info_font).pack(pady=(0,2))
        reverse_checkbox = ctk.CTkCheckBox(reverse_scale_frame, text="(R)", variable=reverse_var, onvalue="on", offvalue="off", command=update_selected_list_on_reverse_toggle)
        reverse_checkbox.pack()

        selected_frame = ctk.CTkFrame(main_paned_window, width=320) 
        main_paned_window.add(selected_frame, stretch="last")
        ctk.CTkLabel(selected_frame, text="Selected for Channel:", font=self.label_font_bold).pack(pady=(5,2))
        
        selected_listbox = tk.Listbox(selected_frame, selectmode="extended", exportselection=False, font=("Arial", 10))
        sel_scrollbar_y = ctk.CTkScrollbar(selected_frame, command=selected_listbox.yview)
        sel_scrollbar_x = ctk.CTkScrollbar(selected_frame, command=selected_listbox.xview, orientation="horizontal")
        selected_listbox.configure(yscrollcommand=sel_scrollbar_y.set, xscrollcommand=sel_scrollbar_x.set)
        sel_scrollbar_y.pack(side="right", fill="y"); sel_scrollbar_x.pack(side="bottom", fill="x")
        selected_listbox.pack(side="left", fill="both", expand=True)

        confirm_cancel_frame = ctk.CTkFrame(dialog); confirm_cancel_frame.pack(pady=10, fill="x", padx=10)
        confirm_button = ctk.CTkButton(confirm_cancel_frame, text="Confirm Selections", command=lambda: on_confirm_channel_selection(dialog, channel_index, temp_selected_var_details))
        confirm_button.pack(side="left", padx=5, expand=True)
        cancel_button = ctk.CTkButton(confirm_cancel_frame, text="Cancel", command=dialog.destroy)
        cancel_button.pack(side="right", padx=5, expand=True)

        populate_available_listbox() 
        populate_selected_listbox()  
        search_var.trace_add("write", lambda name, index, mode, sv=search_var: populate_available_listbox(sv.get()))
        
        available_listbox.bind("<Double-Button-1>", lambda event: transfer_vars(True))
        selected_listbox.bind("<Double-Button-1>", lambda event: transfer_vars(False))
        selected_listbox.bind("<<ListboxSelect>>", on_selected_listbox_selection_change)
        available_listbox.bind("<<ListboxSelect>>", on_available_var_select) 


    def get_variable_display_info(self, var_name):
        min_scale, max_scale, has_valid_scale = None, None, False
        if self.spss_meta and hasattr(self.spss_meta, 'variable_value_labels') and self.spss_meta.variable_value_labels:
            value_labels_for_var = self.spss_meta.variable_value_labels.get(var_name)
            if value_labels_for_var and isinstance(value_labels_for_var, dict) and value_labels_for_var:
                try:
                    codes = sorted([k for k in value_labels_for_var.keys() if isinstance(k, (int, float))])
                    if codes:
                        min_scale, max_scale = codes[0], codes[-1]
                        has_valid_scale = True
                except TypeError: pass
        return var_name, min_scale, max_scale, has_valid_scale

    def update_channel_display_text(self, channel_index):
        if channel_index < 0 or channel_index >= len(self.channel_widgets): return
        display_widget = self.channel_widgets[channel_index]["display"]
        selected_vars_info_list = self.channel_selections.get(channel_index, [])
        texts_to_display = [var_info['display_text_short'] for var_info in selected_vars_info_list]

        display_widget.configure(state="normal")
        display_widget.delete("1.0", "end")
        if texts_to_display: display_widget.insert("1.0", ", ".join(texts_to_display))
        else: display_widget.insert("1.0", "None selected")
        display_widget.configure(state="disabled")

    def generate_index_sheet_data(self):
        index_data_row = {}
        for field_name, config in self.fixed_input_fields.items():
            value = ""
            if config["type"] == "multiselect_dialog_button": value = ", ".join(config.get("selected_values",[]))
            elif "var" in config and config["var"] is not None: value = config["var"].get()
            if field_name == "Timestamp" and value == "Auto (on export)": index_data_row[field_name] = "Auto"
            else: index_data_row[field_name] = value if value and value != "None" else ""
        for field_name in self.demographic_fields:
            spss_var_name = self.demographic_var_selections.get(field_name)
            index_data_row[field_name] = spss_var_name if spss_var_name else ""
        for ch_idx in range(self.num_channels):
            channel_header_name = self.channel_custom_names[ch_idx]
            vars_info = self.channel_selections.get(ch_idx, [])
            texts = [var_info['display_text_short'] for var_info in vars_info] 
            index_data_row[channel_header_name] = ", ".join(texts) if texts else ""
        return pd.DataFrame([index_data_row])

    def _prepare_export_dataframes(self):
        if not self.spss_df_is_loaded():
            messagebox.showwarning("No SPSS Data", "Cannot prepare export data. Please load an SPSS file.", parent=self)
            return None, None, None
        output_data_channels_dict, max_len_from_channels = {}, 0
        if self.spss_df_codes is not None:
            for ch_idx in range(self.num_channels):
                channel_name_export = self.channel_custom_names[ch_idx]
                vars_info_for_channel = self.channel_selections.get(ch_idx, [])
                current_channel_data_list = []
                for var_info in vars_info_for_channel:
                    var_name_export = var_info['name']
                    if var_name_export not in self.spss_df_codes.columns: continue
                    series_data = self.spss_df_codes[var_name_export].copy()
                    if var_info.get('reverse') and var_info['min_scale'] is not None and var_info['max_scale'] is not None:
                        try:
                            min_s, max_s = float(var_info['min_scale']), float(var_info['max_scale'])
                            series_data_numeric = pd.to_numeric(series_data, errors='coerce')
                            reversed_values = (min_s + max_s) - series_data_numeric
                            series_data = reversed_values.where(series_data_numeric.notna(), series_data)
                        except (ValueError, TypeError) as e: print(f"Warning: Could not reverse scale for {var_name_export}: {e}")
                    current_channel_data_list.extend(series_data.tolist())
                output_data_channels_dict[channel_name_export] = pd.Series(current_channel_data_list)
                max_len_from_channels = max(max_len_from_channels, len(current_channel_data_list))
        if max_len_from_channels == 0 and self.spss_df_codes is not None and not self.spss_df_codes.empty: max_len_from_channels = len(self.spss_df_codes)
        elif max_len_from_channels == 0: max_len_from_channels = 1
        output_data_fixed_dict, timestamp_value_for_rawdata = {}, ""
        for field_name, config in self.fixed_input_fields.items():
            value_to_use = ""
            if config["type"] == "multiselect_dialog_button": value_to_use = ", ".join(config.get("selected_values",[]))
            elif "var" in config and config["var"] is not None: value_to_use = config["var"].get()
            if field_name == "Timestamp":
                if value_to_use == "Auto (on export)" or not value_to_use: timestamp_value_for_rawdata = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                else: timestamp_value_for_rawdata = value_to_use
                value_to_use = timestamp_value_for_rawdata
            output_data_fixed_dict[field_name] = pd.Series([value_to_use if value_to_use is not None else ""] * max_len_from_channels, dtype=str)
        output_data_demographics_dict = {}
        if self.spss_df_labels is not None and not self.spss_df_labels.empty:
            num_cases_spss = len(self.spss_df_labels)
            for field_name in self.demographic_fields:
                spss_var_name = self.demographic_var_selections.get(field_name)
                if spss_var_name and spss_var_name in self.spss_df_labels.columns:
                    demographic_series = self.spss_df_labels[spss_var_name].astype(str).fillna('')
                    if num_cases_spss > 0 and max_len_from_channels > 0 :
                        if max_len_from_channels >= num_cases_spss:
                            num_repeats = (max_len_from_channels + num_cases_spss - 1) // num_cases_spss
                            replicated_data = pd.concat([demographic_series] * num_repeats, ignore_index=True)[:max_len_from_channels]
                        else: replicated_data = demographic_series[:max_len_from_channels]
                    elif max_len_from_channels > 0: replicated_data = pd.Series([''] * max_len_from_channels, dtype=str)
                    else: replicated_data = pd.Series([], dtype=str)
                    output_data_demographics_dict[field_name] = replicated_data
                else: output_data_demographics_dict[field_name] = pd.Series([''] * max_len_from_channels, dtype=str)
        else:
            for field_name in self.demographic_fields: output_data_demographics_dict[field_name] = pd.Series([''] * max_len_from_channels, dtype=str)
        expected_headers_ordered = list(self.fixed_input_fields.keys()) + self.demographic_fields + self.channel_custom_names
        df_fixed = pd.DataFrame(output_data_fixed_dict).reindex(columns=list(self.fixed_input_fields.keys()), fill_value='')
        df_demographics = pd.DataFrame(output_data_demographics_dict).reindex(columns=self.demographic_fields, fill_value='')
        df_channels = pd.DataFrame(output_data_channels_dict).reindex(columns=self.channel_custom_names, fill_value='')
        df_rawdata = pd.concat([df_fixed, df_demographics, df_channels], axis=1)
        df_rawdata = df_rawdata.reindex(columns=expected_headers_ordered, fill_value='').astype(str).fillna('')
        df_index = self.generate_index_sheet_data()
        if "Timestamp" in df_index.columns and df_index.loc[0, "Timestamp"] == "Auto": df_index.loc[0, "Timestamp"] = timestamp_value_for_rawdata
        df_index = df_index.fillna('').astype(str)
        return df_index, df_rawdata, timestamp_value_for_rawdata

    def export_to_excel(self):
        df_index, df_rawdata, _ = self._prepare_export_dataframes()
        if df_index is None or df_rawdata is None: return
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=(("Excel files", "*.xlsx"),("All files", "*.*")), title="Save Exported Data As (Excel)")
        if not save_path: return
        if not save_path.lower().endswith(".xlsx"): save_path = os.path.splitext(save_path)[0] + ".xlsx"
        try:
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                df_index.to_excel(writer, sheet_name='Sheet Index', index=False)
                df_rawdata.to_excel(writer, sheet_name='RawdataNorm', index=False)
            messagebox.showinfo("Success", f"Data exported to Excel successfully!\n{save_path}", parent=self)
        except Exception as e: messagebox.showerror("Excel Export Error", str(e), parent=self)

    def fetch_total_jobs_from_gsheet_threaded(self):
        self.total_jobs_var.set("Total Jobs: Loading...")
        thread = threading.Thread(target=self.fetch_total_jobs_from_gsheet, daemon=True)
        thread.start()

    def fetch_total_jobs_from_gsheet(self):
        try:
            required_attrs = ['SERVICE_ACCOUNT_FILE_PATH', 'INDEX_SHEET_NAME_FOR_COUNT', 'TARGET_GOOGLE_SHEET_URL', 'UNIQUE_JOB_IDENTIFIER_COLUMN']
            for attr in required_attrs:
                if not hasattr(self, attr) or not getattr(self, attr):
                    self.total_jobs_var.set(f"Total Jobs: {attr} N/A"); print(f"GS Config Error: '{attr}' missing."); return
            if not os.path.exists(self.SERVICE_ACCOUNT_FILE_PATH):
                self.total_jobs_var.set("Total Jobs: Key File NF"); print(f"SA Key NF: {self.SERVICE_ACCOUNT_FILE_PATH}"); return
            scopes = ['https://www.googleapis.com/auth/spreadsheets.readonly']
            creds = Credentials.from_service_account_file(self.SERVICE_ACCOUNT_FILE_PATH, scopes=scopes)
            client = gspread.authorize(creds)
            spreadsheet = client.open_by_url(self.TARGET_GOOGLE_SHEET_URL)
            worksheet = spreadsheet.worksheet(self.INDEX_SHEET_NAME_FOR_COUNT)
            all_records = worksheet.get_all_records()
            if not all_records: self.total_jobs_var.set("Total Jobs: 0"); return
            unique_job_ids = set(str(rec.get(self.UNIQUE_JOB_IDENTIFIER_COLUMN,"")).strip() for rec in all_records if str(rec.get(self.UNIQUE_JOB_IDENTIFIER_COLUMN,"")).strip())
            self.total_jobs_var.set(f"Total Jobs: {len(unique_job_ids)}")
        except gspread.exceptions.APIError as api_e: self.total_jobs_var.set("Total Jobs: API Error"); print(f"GS API Error: {api_e}")
        except gspread.exceptions.SpreadsheetNotFound: self.total_jobs_var.set("Total Jobs: Sheet NF"); print(f"GS Not Found: {self.TARGET_GOOGLE_SHEET_URL}")
        except gspread.exceptions.WorksheetNotFound: self.total_jobs_var.set("Total Jobs: Tab NF"); print(f"Worksheet '{self.INDEX_SHEET_NAME_FOR_COUNT}' NF")
        except FileNotFoundError: self.total_jobs_var.set("Total Jobs: Key File NF"); print(f"SA Key NF: {self.SERVICE_ACCOUNT_FILE_PATH}")
        except Exception as e: self.total_jobs_var.set("Total Jobs: Error"); print(f"Error fetching total jobs: {type(e).__name__} - {str(e)}")

    def save_to_google_sheet(self):
        df_index, df_rawdata, _ = self._prepare_export_dataframes()
        if df_index is None or df_rawdata is None: return False
        project_num = self.fixed_input_fields["Project #"]["var"].get().strip()
        project_name = self.fixed_input_fields["Project Name"]["var"].get().strip()
        username = self.fixed_input_fields["Username"]["var"].get()
        if not project_num or not project_name or username == self.username_placeholder:
            warning_msg = "กรุณาระบุข้อมูลต่อไปนี้เป็นอย่างน้อย:\n"
            if not project_num: warning_msg += "- Project #\n"
            if not project_name: warning_msg += "- Project Name\n"
            if username == self.username_placeholder: warning_msg += "- Username\n"
            if not (hasattr(self, 'is_batch_processing_active') and self.is_batch_processing_active):
                 messagebox.showwarning("ข้อมูลไม่ครบถ้วน", warning_msg.strip(), parent=self)
            else: print(f"Batch GSheet Save Error for {project_num}: Missing required fields. {warning_msg.strip()}")
            return False
        
        all_appends_successful = True
        try:
            scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
            if not os.path.exists(self.SERVICE_ACCOUNT_FILE_PATH):
                err_msg_sa = f"ไม่พบไฟล์ Service Account Key:\n{self.SERVICE_ACCOUNT_FILE_PATH}"
                if not (hasattr(self, 'is_batch_processing_active') and self.is_batch_processing_active): messagebox.showerror("ข้อผิดพลาด Config", err_msg_sa, parent=self)
                else: print(f"Batch GSheet Save Error: {err_msg_sa}")
                return False
            creds = Credentials.from_service_account_file(self.SERVICE_ACCOUNT_FILE_PATH, scopes=scopes)
            client = gspread.authorize(creds)
            spreadsheet = client.open_by_url(self.TARGET_GOOGLE_SHEET_URL)
            def prepare_df_for_gspread(input_df):
                if input_df is None: return pd.DataFrame()
                df_copy = input_df.copy().astype(object).fillna('')
                for col in df_copy.columns: df_copy[col] = df_copy[col].astype(str).replace(['nan', 'NaN', 'None', 'none', '<NA>'], '', regex=False)
                return df_copy.fillna('')
            df_index_clean = prepare_df_for_gspread(df_index)
            df_rawdata_clean = prepare_df_for_gspread(df_rawdata)
            def append_data_to_sheet(ws_title, df_to_append):
                nonlocal all_appends_successful
                try:
                    worksheet = spreadsheet.worksheet(ws_title)
                    data_values_to_append = df_to_append.values.tolist()
                    headers_to_append = df_to_append.columns.tolist()
                    current_headers = []
                    try: current_headers = worksheet.row_values(1)
                    except gspread.exceptions.APIError as e_header:
                        if "exceeds grid limits" not in str(e_header).lower() and "empty" not in str(e_header).lower(): raise
                    if not current_headers and headers_to_append:
                        worksheet.append_row(headers_to_append, value_input_option='USER_ENTERED')
                        print(f"Appended headers to '{ws_title}'.")
                    elif current_headers and headers_to_append and current_headers != headers_to_append and not df_to_append.empty:
                         print(f"Warning: Header mismatch for sheet '{ws_title}'. Appending data below existing headers.")
                    if data_values_to_append:
                        worksheet.append_rows(data_values_to_append, value_input_option='USER_ENTERED')
                        print(f"Appended {len(data_values_to_append)} data rows to '{ws_title}'.")
                    else: print(f"No data rows to append to '{ws_title}'.")
                except gspread.exceptions.WorksheetNotFound: 
                    err_msg_wsnf = f"ไม่พบ Sheet '{ws_title}'"
                    if not (hasattr(self, 'is_batch_processing_active') and self.is_batch_processing_active): messagebox.showerror("ข้อผิดพลาด", err_msg_wsnf, parent=self)
                    else: print(f"Batch GSheet Save Error: {err_msg_wsnf}")
                    all_appends_successful = False
                except Exception as e_sheet: 
                    err_msg_esheet = f"เขียนข้อมูลลง Sheet '{ws_title}' ล้มเหลว:\n{type(e_sheet).__name__}: {e_sheet}"
                    if not (hasattr(self, 'is_batch_processing_active') and self.is_batch_processing_active): messagebox.showerror("ข้อผิดพลาด Sheet", err_msg_esheet, parent=self)
                    else: print(f"Batch GSheet Save Error: {err_msg_esheet}")
                    all_appends_successful = False

            if not df_index_clean.empty: append_data_to_sheet(self.INDEX_SHEET_NAME_FOR_COUNT, df_index_clean)
            else: print(f"Index DataFrame ('{self.INDEX_SHEET_NAME_FOR_COUNT}') is empty. Skipping.")
            
            if not all_appends_successful: return False 
            
            if not df_rawdata_clean.empty: append_data_to_sheet("RawdataNorm", df_rawdata_clean)
            else: print("Rawdata DataFrame ('RawdataNorm') is empty. Skipping.")
            
            if all_appends_successful:
                if not (hasattr(self, 'is_batch_processing_active') and self.is_batch_processing_active):
                    messagebox.showinfo("สำเร็จ", "ข้อมูลถูกประมวลผลสำหรับ Google Sheets เรียบร้อยแล้ว!", parent=self)
                if not (hasattr(self, 'is_batch_processing_active') and self.is_batch_processing_active):
                    self._clear_all_selections_and_inputs(clear_project_info=False, clear_demographics=False, clear_channels=True)
                self.fetch_total_jobs_from_gsheet_threaded()
                return True
            else: 
                return False
        except FileNotFoundError: 
            err_msg_fnf = f"ไม่พบไฟล์ SA Key:\n{self.SERVICE_ACCOUNT_FILE_PATH}"
            if not (hasattr(self, 'is_batch_processing_active') and self.is_batch_processing_active): messagebox.showerror("ข้อผิดพลาด Config", err_msg_fnf, parent=self)
            else: print(f"Batch GSheet Save Error: {err_msg_fnf}")
            return False
        except gspread.exceptions.SpreadsheetNotFound: 
            err_msg_gsnf = f"ไม่พบ GS หรือไม่มีสิทธิ์เข้าถึง:\n{self.TARGET_GOOGLE_SHEET_URL}"
            if not (hasattr(self, 'is_batch_processing_active') and self.is_batch_processing_active): messagebox.showerror("ข้อผิดพลาด", err_msg_gsnf, parent=self)
            else: print(f"Batch GSheet Save Error: {err_msg_gsnf}")
            return False
        except Exception as e_auth: 
            err_msg_eauth = f"ตั้งค่า GS ล้มเหลว:\n{type(e_auth).__name__}: {e_auth}"
            if not (hasattr(self, 'is_batch_processing_active') and self.is_batch_processing_active): messagebox.showerror("ข้อผิดพลาด GS", err_msg_eauth, parent=self)
            else: print(f"Batch GSheet Save Error: {err_msg_eauth}")
            return False
      
    def _update_gui_field(self, field_config_or_var, value_to_set, is_multiselect=False, selected_values_list=None):
        if is_multiselect:
            field_config_or_var["selected_values"] = list(selected_values_list if selected_values_list else [])
            field_config_or_var["var"].set(", ".join(field_config_or_var["selected_values"]))
        else: field_config_or_var.set(value_to_set)

    def _update_demographic_display(self, field_name, spss_var_name):
        display_text = spss_var_name if spss_var_name and pd.notna(spss_var_name) else "None selected"
        if field_name in self.demographic_widgets and "display" in self.demographic_widgets[field_name]:
            self.demographic_widgets[field_name]["display"].configure(text=display_text)

    def _update_channel_display_from_batch(self, channel_index): self.update_channel_display_text(channel_index)

    def _execute_spss_load_non_interactive(self, file_path):
        self.spss_load_completion_event.clear()
        if not file_path or not os.path.exists(file_path):
            print(f"Batch Error: SPSS file path invalid/not found: {file_path}")
            self.spss_load_results = {"success": False, "last_exception": FileNotFoundError(f"SPSS file not found: {file_path}")}
            self.spss_load_completion_event.set(); return
        self.file_label.configure(text=f"Loading (batch): {os.path.basename(file_path)}...")
        self.update_idletasks()
        self.spss_load_results = {"df_codes": None, "meta": None, "success": False, "used_encoding": "N/A", "last_exception": None}
        def _load_file_in_thread_batch():
            encodings_to_try = [None, 'utf-8', 'cp874', 'tis-620']
            loaded_ok = False
            for enc in encodings_to_try:
                try:
                    current_encoding_name = enc if enc else 'auto-detect'
                    temp_df_codes, temp_meta = pyreadstat.read_sav(file_path, apply_value_formats=False, user_missing=False, encoding=enc)
                    if temp_df_codes is not None and not temp_df_codes.empty and temp_meta is not None:
                        self.spss_load_results.update({"df_codes": temp_df_codes, "meta": temp_meta, "success": True, "used_encoding": current_encoding_name})
                        loaded_ok = True; break
                    else: self.spss_load_results["last_exception"] = RuntimeError(f"Encoding {current_encoding_name} resulted in empty data.")
                except Exception as e: self.spss_load_results["last_exception"] = e
            if not loaded_ok: print(f"Batch: Failed to load SPSS '{os.path.basename(file_path)}' after all encodings.")
            self.after(0, self._finish_loading_spss_and_signal, file_path)
        threading.Thread(target=_load_file_in_thread_batch, daemon=True).start()

    def _finish_loading_spss_and_signal(self, file_path_loaded):
        self._finish_loading_spss(file_path_loaded)
        self.spss_load_completion_event.set()

    def _finalize_batch_processing_ui(self):
        self.load_excel_process_button.configure(state="normal", text="Load Excel & Process Jobs")
        self.is_batch_processing_active = False
        print("Batch UI finalized: Button re-enabled, batch flag reset.")
        self.update_idletasks()

    def _handle_final_batch_summary_and_log(self, batch_log_entries, summary_title="Batch Processing Summary"):
        """Shows the batch log content in a dialog with styled text. Optionally allows saving."""
        
        log_dialog = ctk.CTkToplevel(self)
        log_dialog.title(summary_title)
        dialog_width, dialog_height = 800, 600 
        self._center_window(log_dialog, dialog_width, dialog_height)
        log_dialog.grab_set()

        ctk.CTkLabel(log_dialog, text="Batch Processing Log & Summary:", font=self.label_font_bold).pack(pady=(10,5))

        # For CTkTextbox, styling with tags is done on the underlying tk.Text widget
        # We need to access it via _textbox attribute.
        log_textbox_widget = ctk.CTkTextbox(log_dialog, wrap="word", activate_scrollbars=True, font=self.log_font) 
        log_textbox_widget.pack(pady=5, padx=10, fill="both", expand=True)
        
        # Access the internal tkinter Text widget to configure tags
        tk_text_widget = log_textbox_widget._textbox # This is how to access the Tk Text widget inside CTkTextbox

        tk_text_widget.tag_configure("success", foreground=self.log_success_color, font=(self.log_font_family, self.log_font_size, "bold"))
        tk_text_widget.tag_configure("error", foreground=self.log_error_color, font=(self.log_font_family, self.log_font_size, "bold"))
        tk_text_widget.tag_configure("warning", foreground=self.log_warning_color, font=(self.log_font_family, self.log_font_size))
        tk_text_widget.tag_configure("header", font=(self.log_font_family, self.log_font_size + 1, "bold", "underline")) # Slightly larger and underlined
        tk_text_widget.tag_configure("summary_header", font=(self.log_font_family, self.log_font_size, "bold"))
        tk_text_widget.tag_configure("summary_item", font=(self.log_font_family, self.log_font_size))

        log_textbox_widget.configure(state="normal")

        full_log_content_for_save = "" 

        for entry in batch_log_entries:
            full_log_content_for_save += entry + "\n" 
            entry_lower = entry.lower()
            applied_tag = None # Default to no specific tag

            # More specific keyword checks
            if entry_lower.startswith("===") and "===" in entry_lower.strip():
                applied_tag = "header"
            elif entry_lower.strip().startswith("สรุปการประมวลผล batch อัตโนมัติ (สิ้นสุดเวลา:"):
                applied_tag = "summary_header"
            elif "  - จำนวน job ที่บันทึกลง google sheets สำเร็จ:" in entry_lower:
                parts = entry.split(":", 1)
                log_textbox_widget.insert("end", parts[0] + ":", "summary_item")
                if len(parts) > 1:
                    try:
                        count_val_str = parts[1].strip()
                        count_val = int(count_val_str)
                        if count_val > 0: log_textbox_widget.insert("end", parts[1], ("summary_item", "success"))
                        else: log_textbox_widget.insert("end", parts[1], "summary_item")
                    except ValueError: log_textbox_widget.insert("end", parts[1], "summary_item") # If not a number, print normally
                log_textbox_widget.insert("end", "\n")
                continue 
            elif "  - จำนวน job ที่ตั้งค่าไม่สำเร็จ (spss/gui):" in entry_lower or \
                 "  - จำนวน job ที่บันทึกลง google sheets ไม่สำเร็จ (หลังตั้งค่าสำเร็จ):" in entry_lower:
                parts = entry.split(":", 1)
                log_textbox_widget.insert("end", parts[0] + ":", "summary_item")
                if len(parts) > 1:
                    try:
                        count_val_str = parts[1].strip()
                        count_val = int(count_val_str)
                        if count_val > 0: log_textbox_widget.insert("end", parts[1], ("summary_item", "error"))
                        else: log_textbox_widget.insert("end", parts[1], "summary_item")
                    except ValueError: log_textbox_widget.insert("end", parts[1], "summary_item")
                log_textbox_widget.insert("end", "\n")
                continue 
            elif entry_lower.strip().startswith("[สำเร็จ") or "เรียบร้อยแล้ว." in entry_lower : 
                applied_tag = "success"
            elif entry_lower.strip().startswith("[ผิดพลาด") or "ล้มเหลว)." in entry_lower or "error" in entry_lower.strip(): # Check .strip() for error
                applied_tag = "error"
            elif entry_lower.strip().startswith("[แจ้งเตือน") or "warning" in entry_lower.strip() or "ไม่พบ" in entry_lower :
                applied_tag = "warning"
            elif entry_lower.strip().startswith("  -"): # For other summary items not caught above
                applied_tag = "summary_item"
            
            # Insert text with tag if one was applied
            if applied_tag:
                log_textbox_widget.insert("end", entry + "\n", applied_tag)
            else:
                log_textbox_widget.insert("end", entry + "\n") # Insert with default font/color

        log_textbox_widget.configure(state="disabled") 

        button_frame = ctk.CTkFrame(log_dialog)
        button_frame.pack(pady=10, fill="x", padx=10)

        def save_log_to_file():
            log_save_path = filedialog.asksaveasfilename(
                defaultextension=".txt", filetypes=(("Text files", "*.txt"), ("All files", "*.*")),
                title="Save Batch Log As", initialfile=f"Batch_Process_Log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            )
            if log_save_path:
                try:
                    with open(log_save_path, "w", encoding="utf-8") as f: f.write(full_log_content_for_save)
                    messagebox.showinfo("Log Saved", f"Log saved to:\n{log_save_path}", parent=log_dialog)
                except Exception as e_save_log: messagebox.showerror("Error Saving Log", f"Could not save log:\n{e_save_log}", parent=log_dialog)
        
        save_button = ctk.CTkButton(button_frame, text="Save Log to File", command=save_log_to_file)
        save_button.pack(side="left", padx=10, expand=True)
        
        # Define finalize_ui_and_destroy lambda
        def finalize_ui_and_destroy():
            self._finalize_batch_processing_ui()
            log_dialog.destroy()

        close_button = ctk.CTkButton(button_frame, text="Close", command=finalize_ui_and_destroy)
        close_button.pack(side="right", padx=10, expand=True)
        
        log_dialog.protocol("WM_DELETE_WINDOW", finalize_ui_and_destroy)


    def process_jobs_from_excel_threaded(self):
        """Handles loading Excel and starting the background processing thread."""
        excel_file_path = filedialog.askopenfilename(
            title="Select Excel Template with Job Data", 
            filetypes=(
                ("Excel files", "*.xlsx *.xlsm"), # <--- เพิ่ม *.xlsm ที่นี่
                ("All files", "*.*")
            )
        )
        if not excel_file_path: 
            return
        
        try:
            # pandas.read_excel with openpyxl engine should handle .xlsm
            excel_df = pd.read_excel(excel_file_path, sheet_name='JobList', dtype=str, engine='openpyxl').fillna('')
            if excel_df.empty: 
                messagebox.showinfo("Info", "The 'JobList' sheet is empty or not found.", parent=self)
                return
        except Exception as e: 
            messagebox.showerror("Error Reading Excel", f"Could not read Excel file:\n{excel_file_path}\n\nError: {e}", parent=self)
            return
        
        # Ask for confirmation once before starting the batch
        num_jobs_to_process = len(excel_df)
        confirm_message = (
            f"คุณมี {num_jobs_to_process} Job (แถว) ในไฟล์ Excel ที่จะถูกประมวลผล\n"
            "การประมวลผลจะทำงานต่อเนื่องจนครบทุก Job หรือจนกว่าจะเกิดข้อผิดพลาดร้ายแรง\n\n"
            "ต้องการเริ่มการประมวลผล Batch หรือไม่?"
        )
        if not messagebox.askyesno("ยืนยันการประมวลผล Batch", confirm_message, parent=self, icon=messagebox.QUESTION):
            return 

        self.load_excel_process_button.configure(state="disabled", text="Processing Batch...")
        self.is_batch_processing_active = True
        self.update_idletasks()
        threading.Thread(target=self._process_excel_jobs_in_background, args=(excel_df,), daemon=True).start()

    def _process_excel_jobs_in_background(self, excel_df):
        total_jobs_in_excel = len(excel_df)
        jobs_setup_successful = 0 
        jobs_saved_to_gsheet = 0
        jobs_failed_setup = 0
        jobs_failed_gsheet_save = 0
        
        batch_log_entries = []
        
        try:
            log_timestamp_start = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            batch_log_entries.append(f"=== เริ่มการประมวลผล Batch อัตโนมัติ เวลา: {log_timestamp_start} ===")
            batch_log_entries.append(f"จำนวน Job ทั้งหมดในไฟล์ Excel: {total_jobs_in_excel}")
            batch_log_entries.append("-" * 50)

            for index, row in excel_df.iterrows():
                current_job_index_in_excel = index + 1
                proj_num_excel = str(row.get("Project #", f"แถวที่_{current_job_index_in_excel}")).strip()
                proj_name_excel = str(row.get("Project Name", "")).strip()
                current_job_identifier_for_log = f"Job Excel แถวที่ {current_job_index_in_excel} (Project #: {proj_num_excel}, Name: {proj_name_excel or 'N/A'})"
                
                print(f"\n--- Batch Auto: Processing {current_job_identifier_for_log} ---")
                batch_log_entries.append(f"\nกำลังประมวลผล: {current_job_identifier_for_log}")
                
                self.after(0, self._clear_all_selections_and_inputs, True, True, True) 
                spss_path = str(row.get("SPSS_File_Path", "")).strip()

                if not spss_path:
                    msg = f"ไม่พบที่อยู่ไฟล์ SPSS (SPSS_File_Path) ใน Excel"
                    print(f"{current_job_identifier_for_log}: {msg} - SKIPPING JOB"); batch_log_entries.append(f"  [ผิดพลาด SETUP] {msg}. ข้าม Job นี้.")
                    jobs_failed_setup += 1; continue
                
                self._execute_spss_load_non_interactive(spss_path)
                if not self.spss_load_completion_event.wait(timeout=300):
                    msg = f"หมดเวลาในการรอโหลดไฟล์ SPSS '{os.path.basename(spss_path)}'"
                    print(f"{current_job_identifier_for_log}: {msg} - SKIPPING JOB"); batch_log_entries.append(f"  [ผิดพลาด SETUP] {msg}. ข้าม Job นี้.")
                    jobs_failed_setup += 1; continue
                
                if not self.spss_load_results.get("success"):
                    err = self.spss_load_results.get('last_exception', 'ไม่ทราบสาเหตุ')
                    msg = f"ไม่สามารถโหลดไฟล์ SPSS '{os.path.basename(spss_path)}' ได้ สาเหตุ: {err}"
                    print(f"{current_job_identifier_for_log}: {msg} - SKIPPING JOB"); batch_log_entries.append(f"  [ผิดพลาด SETUP] {msg}. ข้าม Job นี้.")
                    jobs_failed_setup += 1; continue
                batch_log_entries.append(f"  [สำเร็จ SETUP] โหลดไฟล์ SPSS '{os.path.basename(spss_path)}' เรียบร้อยแล้ว.")

                setup_this_job_ok = True
                try:
                    for field_name, config in self.fixed_input_fields.items():
                        if config.get("hidden"): continue
                        excel_value_str = str(row.get(field_name, "")).strip()
                        if excel_value_str or (field_name == "Username" and excel_value_str == ""):
                            if config["type"] == "multiselect_dialog_button":
                                selected_list = [s.strip() for s in excel_value_str.split(',') if s.strip()]
                                self.after(0, self._update_gui_field, config, None, True, selected_list)
                            elif "var" in config and config["var"] is not None: self.after(0, self._update_gui_field, config["var"], excel_value_str)
                    
                    for demo_field in self.demographic_fields:
                        spss_var_from_excel = str(row.get(f"{demo_field}_SPSS_Var", "")).strip()
                        if spss_var_from_excel and hasattr(self, 'spss_variables') and self.spss_variables and spss_var_from_excel in self.spss_variables:
                            self.demographic_var_selections[demo_field] = spss_var_from_excel
                            self.after(0, self._update_demographic_display, demo_field, spss_var_from_excel)
                        else:
                            self.demographic_var_selections[demo_field] = None; self.after(0, self._update_demographic_display, demo_field, None)
                            if spss_var_from_excel: batch_log_entries.append(f"  [แจ้งเตือน SETUP] ตัวแปร Demo '{spss_var_from_excel}' ({demo_field}) ไม่พบใน SPSS ({current_job_identifier_for_log})"); print(f"Warn: Demo var '{spss_var_from_excel}' for '{demo_field}' not in SPSS for job {proj_num_excel}.")
                    
                    for ch_idx in range(self.num_channels):
                        ch_safe_name = self.channel_custom_names[ch_idx].replace(" ", "_").replace("#","Num").replace("/","_")
                        vars_string_from_excel = str(row.get(f"Channel_{ch_safe_name}_Vars", "")).strip()
                        current_channel_selections_for_job = []
                        if vars_string_from_excel:
                            var_items_from_excel = [v.strip() for v in vars_string_from_excel.split(',') if v.strip()]
                            for var_item_str in var_items_from_excel:
                                var_name_cleaned, reverse_this_var = var_item_str, False
                                if var_item_str.lower().endswith(" (r)"): var_name_cleaned, reverse_this_var = var_item_str[:-4].strip(), True
                                if hasattr(self, 'spss_variables') and self.spss_variables and var_name_cleaned in self.spss_variables:
                                    _, min_s, max_s, has_valid_s = self.get_variable_display_info(var_name_cleaned)
                                    should_be_reversed_final = reverse_this_var and has_valid_s
                                    current_channel_selections_for_job.append({'name': var_name_cleaned, 
                                                                       'reverse': should_be_reversed_final, 
                                                                       'min_scale': min_s, 'max_scale': max_s, 
                                                                       'display_text_short': f"{var_name_cleaned}{' (R)' if should_be_reversed_final else ''}"})
                                else: batch_log_entries.append(f"  [แจ้งเตือน SETUP] ตัวแปร Channel '{var_name_cleaned}' ({self.channel_custom_names[ch_idx]}) ไม่พบใน SPSS ({current_job_identifier_for_log})"); print(f"Warn: Channel var '{var_name_cleaned}' for '{self.channel_custom_names[ch_idx]}' not in SPSS for job {proj_num_excel}.")
                        self.channel_selections[ch_idx] = current_channel_selections_for_job 
                        self.after(0, self._update_channel_display_from_batch, ch_idx) 
                    
                    time.sleep(0.25) 
                    batch_log_entries.append(f"  [สำเร็จ SETUP] ตั้งค่าข้อมูล Job ในหน้า GUI เรียบร้อยแล้ว.")
                    jobs_setup_successful +=1
                except Exception as e_pop:
                    msg = f"เกิดข้อผิดพลาดขณะตั้งค่าข้อมูลในหน้า GUI สาเหตุ: {type(e_pop).__name__}: {e_pop}"
                    print(f"{current_job_identifier_for_log}: {msg} - SKIPPING SAVE FOR THIS JOB"); batch_log_entries.append(f"  [ผิดพลาด SETUP] {msg}. ข้ามการบันทึก Job นี้.")
                    jobs_failed_setup += 1; setup_this_job_ok = False; continue 

                if setup_this_job_ok:
                    print(f"Attempting to auto-save Job '{current_job_identifier_for_log}' to Google Sheets...")
                    save_success = self.save_to_google_sheet() 
                    if save_success:
                        jobs_saved_to_gsheet += 1
                        batch_log_entries.append(f"  [สำเร็จ GSHEET SAVE] Job ({current_job_identifier_for_log}) ถูกบันทึกลง Google Sheets เรียบร้อยแล้ว.")
                        print(f"Successfully auto-saved Job '{current_job_identifier_for_log}' to Google Sheets.")
                    else:
                        jobs_failed_gsheet_save += 1
                        batch_log_entries.append(f"  [ผิดพลาด GSHEET SAVE] ไม่สามารถบันทึก Job ({current_job_identifier_for_log}) ลง Google Sheets ได้.")
                        print(f"Failed to auto-save Job '{current_job_identifier_for_log}' to Google Sheets. See GSheet save errors.")
            
            log_timestamp_end = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            batch_log_entries.append("-" * 50)
            batch_log_entries.append(f"สรุปการประมวลผล Batch อัตโนมัติ (สิ้นสุดเวลา: {log_timestamp_end}):")
            batch_log_entries.append(f"  - จำนวน Job ทั้งหมดใน Excel: {total_jobs_in_excel}")
            batch_log_entries.append(f"  - จำนวน Job ที่ตั้งค่าข้อมูลสำเร็จ: {jobs_setup_successful}")
            batch_log_entries.append(f"  - จำนวน Job ที่บันทึกลง Google Sheets สำเร็จ: {jobs_saved_to_gsheet}")
            batch_log_entries.append(f"  - จำนวน Job ที่ตั้งค่าไม่สำเร็จ (SPSS/GUI): {jobs_failed_setup}")
            batch_log_entries.append(f"  - จำนวน Job ที่บันทึกลง Google Sheets ไม่สำเร็จ (หลังตั้งค่าสำเร็จ): {jobs_failed_gsheet_save}")
            batch_log_entries.append("=== สิ้นสุดการประมวลผล Batch อัตโนมัติ ===")
            
            print("\n" + "\n".join(batch_log_entries))
            self.after(0, self._handle_final_batch_summary_and_log, batch_log_entries, "สรุปผลการประมวลผล Batch อัตโนมัติ")
                
        except Exception as e_batch_main:
            critical_error_msg = f"ข้อผิดพลาดร้ายแรงที่ไม่คาดคิดเกิดขึ้นระหว่างการประมวลผล Batch: {type(e_batch_main).__name__}: {e_batch_main}"
            print(f"CRITICAL BATCH ERROR: {critical_error_msg}"); traceback.print_exc()
            if 'batch_log_entries' not in locals(): batch_log_entries = []
            batch_log_entries.append(f"\n[ข้อผิดพลาดร้ายแรงใน Batch] {critical_error_msg}")
            self.after(0, self._handle_final_batch_summary_and_log, batch_log_entries, "ข้อผิดพลาดร้ายแรงในการประมวลผล Batch")

    def save_excel_template(self):
        template_headers = ["SPSS_File_Path", "Project #", "Project Name", "Year Implemented", "Survey Method", "Category", "Evaluation Target", "Scale", "Username"]
        for demo_field in self.demographic_fields: template_headers.append(f"{demo_field}_SPSS_Var")
        channel_vars_col_names = [] 
        for i in range(self.num_channels):
            ch_safe = self.channel_custom_names[i].replace(" ", "_").replace("#","Num").replace("/","_"); col_name = f"Channel_{ch_safe}_Vars"
            template_headers.append(col_name); channel_vars_col_names.append(col_name)
        example_row = {header: "" for header in template_headers}
        example_row["SPSS_File_Path"] = "C:\\Data\\YourProjectFile.sav"; example_row["Project #"] = "P2025001"; example_row["Project Name"] = "Sample Project for Batch Import"
        example_row["Year Implemented"] = self.current_year_str; example_row["Survey Method"] = "Online survey"; example_row["Category"] = "Food"; example_row["Evaluation Target"] = "Product"; example_row["Scale"] = "5-point scale"
        username_options_for_template = [opt for opt in self.fixed_input_fields["Username"]["options"] if opt != self.username_placeholder]
        example_row["Username"] = username_options_for_template[0] if username_options_for_template else "YourName"
        if self.demographic_fields: example_row[f"{self.demographic_fields[0]}_SPSS_Var"] = "v_gender_spss_name" 
        if channel_vars_col_names: example_row[channel_vars_col_names[0]] = "Q1_Overall_Liking, Q2_Attribute_To_Reverse (R), Q3_Another_Attribute"
        df_example_row = pd.DataFrame([example_row], columns=template_headers)
        num_target_data_rows = 100; df_empty_rows = pd.DataFrame()
        if num_target_data_rows > 1: df_empty_rows = pd.DataFrame([{col: "" for col in template_headers}] * (num_target_data_rows - 1), columns=template_headers)
        df_template = pd.concat([df_example_row, df_empty_rows], ignore_index=True)
        instructions_data = {
            "Column Name": ["SPSS_File_Path", "Project #", "Project Name", "Year Implemented", "Survey Method", "Category", "Evaluation Target", "Scale", "Username", "[DemographicField]_SPSS_Var", "Channel_..._Vars", "Scale Reversal using (R) Suffix"],
            "Description": ["Full path to .sav SPSS data file (e.g., C:\\data\\mydata.sav).", "Unique Project Number.", "Name of the project.", "Year implemented.", "Survey methodology. Comma-separate for multiple.", "Product category. Comma-separate for multiple.", "What was evaluated. Comma-separate for multiple.", "Scale used.", "Username.", "SPSS variable name for demographic field.", "SPSS Variable names for channel. Comma-separate.", "To reverse scale, add ' (R)' after variable name in 'Channel_..._Vars'."],
            "Example": ["C:\\SPSS_Data\\ProjectX.sav", "P2024001", "Noodle Study", self.current_year_str, "CLT,HUT", "Snacks,Beverages", "Product,Taste", "7-point scale", "YourUsername", "q_gender (for Gender_SPSS_Var)", "q_overall_liking, q_uniqueness (R), q_pi", "q_price_perception (R)"]
        }
        max_len_instr = 0; 
        if instructions_data: max_len_instr = max(len(v) for v in instructions_data.values())
        for k_instr in instructions_data: instructions_data[k_instr].extend([""] * (max_len_instr - len(instructions_data[k_instr])))
        df_instructions = pd.DataFrame(instructions_data)
        try:
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=(("Excel files", "*.xlsx"),), title="Save Excel Job Template As", initialfile="Job_Import_Template_With_Example.xlsx")
            if not save_path: return
            if not save_path.lower().endswith(".xlsx"): save_path += ".xlsx"
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                df_template.to_excel(writer, sheet_name='JobList', index=False)
                df_instructions.to_excel(writer, sheet_name='Instructions', index=False)
                wb = writer.book; ws_joblist = wb['JobList']; ws_instr = wb['Instructions']
                for idx, col_name_instr in enumerate(df_instructions.columns):
                    max_l = len(str(col_name_instr)) 
                    for cell_val in df_instructions[col_name_instr].astype(str): max_l = max(max_l, len(cell_val))
                    ws_instr.column_dimensions[get_column_letter(idx + 1)].width = max_l + 2
                for idx, col_name_joblist in enumerate(df_template.columns):
                    max_l_job = len(str(col_name_joblist))
                    for cell_val_job in df_template[col_name_joblist].head(min(10, len(df_template))).astype(str): max_l_job = max(max_l_job, len(cell_val_job))
                    ws_joblist.column_dimensions[get_column_letter(idx + 1)].width = min(max_l_job + 5, 50) 
                helper_sheet_name = "DropdownOptions"; helper_ws = None; next_helper_col = 1
                val_start_row, val_end_row = 2, 2 + len(df_template) -1 
                dropdown_cfgs = [
                    {"col": "Year Implemented", "opts": self.fixed_input_fields["Year Implemented"]["options"]},
                    {"col": "Survey Method", "opts": self.fixed_input_fields["Survey Method"]["options"]},
                    {"col": "Category", "opts": self.fixed_input_fields["Category"]["options"]},
                    {"col": "Evaluation Target", "opts": self.fixed_input_fields["Evaluation Target"]["options"]},
                    {"col": "Scale", "opts": self.fixed_input_fields["Scale"]["options"]},
                    {"col": "Username", "opts": [opt for opt in self.fixed_input_fields["Username"]["options"] if opt != self.username_placeholder]},
                ]
                for cfg_dv in dropdown_cfgs:
                    col_dv, opts_dv = cfg_dv["col"], cfg_dv["opts"]
                    if not opts_dv: continue
                    try:
                        col_idx_df = df_template.columns.get_loc(col_dv); col_letter_dv = get_column_letter(col_idx_df + 1)
                        opts_str_dv = [str(o) for o in opts_dv]; needs_helper = any(',' in opt for opt in opts_str_dv)
                        formula_direct_dv = '"' + ",".join(opts_str_dv) + '"'; dv_formula_final = formula_direct_dv
                        if len(formula_direct_dv) > 250 or needs_helper: 
                            if helper_ws is None: helper_ws = wb[helper_sheet_name] if helper_sheet_name in wb.sheetnames else wb.create_sheet(helper_sheet_name)
                            helper_col_l_dv = get_column_letter(next_helper_col)
                            helper_ws.cell(row=1, column=next_helper_col, value=f"Options for {col_dv}")
                            for i_dv, opt_val_dv in enumerate(opts_str_dv): helper_ws.cell(row=i_dv + 2, column=next_helper_col, value=opt_val_dv)
                            dv_formula_final = f"'{helper_sheet_name}'!${helper_col_l_dv}$2:${helper_col_l_dv}${len(opts_str_dv) + 1}"; next_helper_col +=1
                        if dv_formula_final:
                            dv = DataValidation(type="list", formula1=dv_formula_final, allow_blank=True)
                            dv.errorTitle=f"Invalid {col_dv}"; dv.error="Select from list or type. For multiple, comma-separate."; dv.promptTitle=f"Select {col_dv}"; dv.prompt="Choose or type."
                            ws_joblist.add_data_validation(dv); dv.add(f"{col_letter_dv}{val_start_row}:{col_letter_dv}{val_end_row}")
                    except Exception as e_dv: print(f"Error DV for '{col_dv}': {e_dv}")
            messagebox.showinfo("Template Saved", f"Excel job template with example saved:\n{save_path}", parent=self)
        except Exception as e_save: messagebox.showerror("Error Saving Template", f"Could not save template:\n{e_save}", parent=self); print(f"Template save error: {e_save}"); traceback.print_exc()

def run_this_app(working_dir=None):
    print(f"--- SPSS_EXPORTER_INFO: Starting 'SpssExporterApp' ---")
    if working_dir and os.path.exists(working_dir): os.chdir(working_dir); print(f"--- Changed WD to: {os.getcwd()} ---")
    elif working_dir: print(f"--- WARNING: Provided working_dir '{working_dir}' does not exist. ---")
    try:
        app = SpssExporterApp()
        app.mainloop()
    except Exception as e:
        print(f"SPSS_EXPORTER_ERROR: An error occurred during SpssExporterApp execution: {e}")
        traceback.print_exc()
        try: 
            root_temp = tk.Tk(); root_temp.withdraw()
            messagebox.showerror("Application Error", f"Unexpected error, app closing:\n{e}\nCheck console.", parent=root_temp)
            root_temp.destroy()
        except Exception as e_msg: print(f"Could not display error in messagebox: {e_msg}")
        sys.exit(f"Error running SpssExporterApp: {e}")

if __name__ == "__main__":
    print("--- Running SpssExporterApp.py directly ---")
    run_this_app()
    print("--- Finished direct execution of SpssExporterApp.py ---")