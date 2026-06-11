import sys
import os
import re
import bisect
from collections import defaultdict
from datetime import datetime

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QListWidget, QRadioButton,
    QButtonGroup, QFileDialog, QMessageBox, QStatusBar,
    QFrame, QAbstractItemView, QProgressBar, QTextEdit, QSplitter
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QDragEnterEvent, QDropEvent

import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.cell_range import CellRange


def writable_cell(ws, row, col):
    """Return a writable cell. If (row, col) falls inside a merged range,
    return the top-left anchor of that range (the only writable member)."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                return ws.cell(row=mr.min_row, column=mr.min_col)
    return cell


# ===================== Utilities =====================

def clean_header(text):
    if not isinstance(text, str): return ""
    return "".join(re.findall(r"[A-Za-z]", text)).upper()

def parse_sig_groups(sig_input):
    groups = [g.strip().upper() for g in sig_input.split(",") if g.strip()]
    expanded = []
    for g in groups:
        if "-" in g and len(g) == 3:
            a,b = g.split("-")
            expanded.append("".join(chr(c) for c in range(ord(a), ord(b)+1)))
        else:
            expanded.append("".join(sorted(set(re.findall(r"[A-Z]", g)))))
    out = {}
    for rng in expanded:
        for ch in rng:
            out[ch] = rng
    return out

def detect_letter_header_row(ws, start_row, end_row, min_seq=3):
    scan_until = min(end_row, start_row + 120)
    for r in range(start_row, scan_until + 1):
        letters = []
        for c in range(3, ws.max_column + 1):
            t = clean_header(str(ws.cell(row=r, column=c).value))
            if len(t) == 1 and 'A' <= t <= 'Z':
                letters.append((c, t))
        if len(letters) < min_seq:
            continue
        letters.sort()
        vals = {c:v for c,v in letters}
        for c,v in letters:
            if v == 'A' and vals.get(c+1) == 'B' and vals.get(c+2) == 'C':
                return r, c
    return -1, -1

def build_col_letter_map(ws, header_row, start_col):
    mapping = {}
    for c in range(start_col, ws.max_column + 1):
        t = clean_header(str(ws.cell(row=header_row, column=c).value))
        if len(t) == 1 and 'A' <= t <= 'Z':
            mapping[c] = t
        else:
            if mapping: break
    return mapping

def find_total_row_in_colB(ws, header_row, end_row):
    for r in range(header_row, end_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and v.strip().upper() == "TOTAL":
            return r
    return -1

def find_last_data_row(ws, start_row, col_indexes):
    last = start_row
    for r in range(start_row, ws.max_row + 1):
        any_val = False
        for c in col_indexes:
            v = ws.cell(row=r, column=c).value
            if v not in (None, "") and str(v).strip() != "":
                any_val = True
                break
        if any_val: last = r
    return last

def filter_sig_text(text, allowed_letters, preserve_tokens=("ADJ",)):
    if not isinstance(text, str) or not text:
        return text
    out = text
    dynamic_preserve = tuple(
        tok for tok in preserve_tokens
        if tok and set(tok.upper()).issubset(set(allowed_letters))
    )
    placeholders = {}
    for idx, tok in enumerate(dynamic_preserve):
        if tok in out:
            ph = f"\u00a7{idx}\u00a7"
            placeholders[ph] = tok
            out = out.replace(tok, ph)
    out = out.replace("_", "")
    keep = set(allowed_letters)
    buf = []
    for ch in out:
        if 'A' <= ch.upper() <= 'Z':
            if ch.upper() in keep:
                buf.append(ch)
        else:
            buf.append(ch)
    out = "".join(buf)
    for ph, tok in placeholders.items():
        out = out.replace(ph, tok)
    return out.strip()


# ===================== Sig Beside Utilities =====================

def collect_sig_rows_and_merge(ws, start_row, end_row, label_cols, data_cols, keep_first_unlabeled=False):
    """
    Rule for Sig-beside mode:
    - If any label column has value -> value row.
    - If all label columns are empty -> paired/sig row.
      Merge only pure-letter cells (A-Z only, no digits) into previous value row,
      then delete that row (to compact layout like target output).
    - keep_first_unlabeled=True is used for N% layout (Count + Column % + Sig):
      keep first unlabeled row (Column %), then process/delete subsequent unlabeled rows.
    Return list of row indexes to delete.
    """
    data_cols = list(data_cols)
    label_cols = list(label_cols)
    rows_to_delete = []
    prev_value_row = None
    merge_target = None
    unlabeled_after_value = 0

    for r in range(start_row, end_row + 1):
        is_value_row = any(
            ws.cell(row=r, column=lc).value not in (None, "") and
            str(ws.cell(row=r, column=lc).value).strip()
            for lc in label_cols
        )

        if is_value_row:
            prev_value_row = r
            merge_target = r
            unlabeled_after_value = 0
            continue

        if prev_value_row is None:
            continue

        unlabeled_after_value += 1
        if keep_first_unlabeled and unlabeled_after_value == 1:
            # N% layout: first unlabeled row is the Column % row, which we keep.
            # The Sig must attach beside the % on THIS row, not the Count row.
            merge_target = r
            continue

        target = merge_target if merge_target is not None else prev_value_row
        for c in data_cols:
            sig_val = ws.cell(row=r, column=c).value
            if (sig_val and isinstance(sig_val, str) and sig_val.strip()
                    and re.search(r'[A-Za-z]', sig_val)
                    and not re.search(r'\d', sig_val)):
                val_cell = writable_cell(ws, target, c)
                existing = val_cell.value
                if existing is None or (isinstance(existing, str) and not str(existing).strip()):
                    val_cell.value = sig_val.strip()
                else:
                    val_cell.value = str(existing) + sig_val.strip()

        rows_to_delete.append(r)

    return rows_to_delete

def is_n_percent_layout(ws, start_row, end_row):
    scan_end = min(end_row, start_row + 30)
    texts = []

    for r in range(start_row, scan_end + 1):
        for c in (1, 2, 3):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                texts.append(v.upper())

    joined = " ".join(texts)
    joined = joined.replace("\r", " ").replace("\n", " ").replace("  ", " ")
    joined = re.sub(r"\s+", " ", joined)

    has_1st = ("1ST ROW" in joined)
    has_2nd = ("2ND ROW" in joined)
    has_count = ("COUNT" in joined)
    has_col_pct = (
        ("COLUMN %" in joined) or ("COLUMN%" in joined) or
        ("COLUM %" in joined) or ("COLUM%" in joined)
    )
    return has_1st and has_2nd and has_count and has_col_pct

def unmerge_ranges_touching_rows(ws, target_rows):
    if not target_rows:
        return
    targets = set(target_rows)
    for mr in list(ws.merged_cells.ranges):
        if any(r in targets for r in range(mr.min_row, mr.max_row + 1)):
            ws.unmerge_cells(str(mr))

def renumber_col_a_by_col_b(ws, crosstab_mode="NORMAL"):
    mode = (crosstab_mode or "").upper()
    if mode != "NORMAL":
        return

    stub_starts = [r for r, cell in enumerate(ws['A'], 1) if str(cell.value).startswith('Stub')]
    if not stub_starts:
        stub_starts = [1]
    table_boundaries = stub_starts + [ws.max_row + 2]

    for i in range(len(table_boundaries) - 1):
        start_row = table_boundaries[i]
        end_row = table_boundaries[i + 1] - 2
        if end_row <= start_row:
            continue

        data_total_row = -1
        for r in range(start_row, end_row + 1):
            c = ws.cell(row=r, column=3).value
            if str(c).strip().upper() == 'TOTAL':
                data_total_row = r
                break
        if data_total_row == -1:
            continue

        idx = 0
        for r in range(data_total_row + 1, end_row + 1):
            b = ws.cell(row=r, column=2).value
            has_label = b not in (None, "") and str(b).strip() != ""
            a_cell = ws.cell(row=r, column=1)
            if isinstance(a_cell, MergedCell):
                # Non-anchor of a merged A pair (the Column % row). Leave it as
                # part of the merge but keep the index sequence consistent.
                if has_label:
                    idx += 1
                continue
            if has_label:
                a_cell.value = idx
                idx += 1
            else:
                a_cell.value = None

def transfer_bottom_borders_before_delete(ws, rows_to_delete):
    """
    Before rows are deleted, push each deleted row's bottom border up to the
    nearest surviving row above it. The group-closing grid line lives on the
    Sig row (the row we delete), so without this the bottom border of every
    group would vanish from the data columns -> broken table grid.
    """
    if not rows_to_delete:
        return
    del_set = set(rows_to_delete)
    max_col = ws.max_column

    # Pre-map every merged non-anchor cell -> its anchor, so resolving the
    # writable target is an O(1) lookup instead of rescanning all merged ranges
    # for every border we move (sheets can have thousands of merges).
    anchor_of = {}
    for mr in ws.merged_cells.ranges:
        anchor = (mr.min_row, mr.min_col)
        for rr in range(mr.min_row, mr.max_row + 1):
            for cc in range(mr.min_col, mr.max_col + 1):
                if (rr, cc) != anchor:
                    anchor_of[(rr, cc)] = anchor

    for r in sorted(del_set):
        t = r - 1
        while t in del_set and t >= 1:
            t -= 1
        if t < 1:
            continue
        for c in range(1, max_col + 1):
            src = ws.cell(row=r, column=c).border
            if src.bottom and src.bottom.style:
                ar, ac = anchor_of.get((t, c), (t, c))
                dst = ws.cell(row=ar, column=ac)
                b = dst.border
                dst.border = Border(
                    left=b.left,
                    right=b.right,
                    top=b.top,
                    bottom=src.bottom,
                    diagonal=b.diagonal,
                    diagonal_direction=b.diagonal_direction,
                    outline=b.outline,
                    vertical=b.vertical,
                    horizontal=b.horizontal,
                )

def delete_rows_preserving_merges(ws, rows_to_delete):
    """
    Delete rows AND correctly shift/clip merged ranges and row heights.

    openpyxl's ws.delete_rows() does not adjust merged ranges or row heights at
    all, leaving stale merges and mis-indexed heights after the rows above them
    are removed. Stale merges then corrupt neighbouring values when written to
    (labels disappear); stale heights make random rows too tall/short. To avoid
    this we snapshot merges + heights, unmerge all, delete the rows, then
    re-apply each remapped to its new position:
      - merges living entirely on deleted rows are dropped; merges that lost
        only some rows (e.g. a 3-row Count/%/Sig label block that loses its Sig
        row) are clipped to the surviving rows
      - row heights are re-keyed to the surviving rows' new positions
    """
    if not rows_to_delete:
        return

    # openpyxl's delete_rows() mishandles hyperlinks: it can overwrite a cell's
    # value with the hyperlink location (e.g. the "Contents" links turn into
    # "Contents!R1C1"). Strip hyperlinks first; the visible values stay intact.
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink is not None:
                cell.hyperlink = None

    del_set = set(rows_to_delete)
    sorted_dels = sorted(del_set)

    def new_index(row):
        # surviving row's new position = row minus deleted rows above it
        return row - bisect.bisect_left(sorted_dels, row)

    snapshot = [(mr.min_row, mr.max_row, mr.min_col, mr.max_col)
                for mr in ws.merged_cells.ranges]

    # Snapshot the style of EVERY cell inside a merged range before unmerging.
    # ws.unmerge_cells() drops the non-anchor MergedCell objects from ws._cells,
    # so the rebuild below loses them and only the top-left anchor keeps a border.
    # Excel still PRINTS such a merged box (it derives the outline from the anchor)
    # but on SCREEN it draws borders per-cell -> the right/bottom edges have no
    # cell to draw them and the table grid looks broken in the app. We restore
    # these cells (with their per-edge borders) at their shifted positions below.
    merged_cell_styles = {}
    for (r1, r2, c1, c2) in snapshot:
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                src = ws._cells.get((rr, cc))
                if src is not None:
                    merged_cell_styles[(rr, cc)] = src._style

    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    # snapshot explicit row heights before deleting (openpyxl won't shift them)
    old_heights = {r: dim.height for r, dim in ws.row_dimensions.items()
                   if dim.height is not None}

    # Compact the rows in a SINGLE pass by rebuilding the cell map, instead of
    # calling ws.delete_rows() once per row. delete_rows() shifts every cell
    # below the deletion point each time -> O(deleted * cells), which is tens of
    # seconds on large sheets. Rebuilding is O(cells). Safe here because the
    # sheet has no conditional formatting / data validation that reference rows
    # (merges and row heights are handled explicitly below).
    new_cells = {}
    for (row, col), cell in ws._cells.items():
        if row in del_set:
            continue
        nr = new_index(row)
        if nr != row:
            cell.row = nr
        new_cells[(nr, col)] = cell
    ws._cells = new_cells
    ws._current_row = max((r for r, _ in new_cells), default=0)

    # re-key row heights: clear whatever delete_rows left behind, then re-apply
    # each surviving row's height at its shifted position.
    for r in list(ws.row_dimensions.keys()):
        ws.row_dimensions[r].height = None
    for old_r, h in old_heights.items():
        if old_r in del_set:
            continue
        ws.row_dimensions[new_index(old_r)].height = h

    # Recreate the merged-range cells that unmerge_cells dropped, at their shifted
    # positions, restoring each cell's original style. This keeps every edge of a
    # merged box (right/bottom included) carrying its own border so the grid lines
    # render on screen, not only in print. Done before re-adding the merges so the
    # targets are still writable plain cells.
    for (rr, cc), st in merged_cell_styles.items():
        if rr in del_set:
            continue
        nr = new_index(rr)
        existing = ws._cells.get((nr, cc))
        if existing is not None:
            existing._style = st
        else:
            ws.cell(row=nr, column=cc)._style = st

    # Re-apply via merged_cells.add (not ws.merge_cells): these cells came from
    # ranges that were already style-cleaned when the file was first merged, and
    # unmerging does not undo that, so the expensive _clean_merge_range step is
    # unnecessary here and only adds cost.
    for (r1, r2, c1, c2) in snapshot:
        surviving = [r for r in range(r1, r2 + 1) if r not in del_set]
        if not surviving:
            continue
        nr1 = new_index(surviving[0])
        nr2 = new_index(surviving[-1])
        if nr1 == nr2 and c1 == c2:
            continue  # collapsed to a single cell -> nothing to merge
        ws.merged_cells.add(CellRange(min_col=c1, min_row=nr1, max_col=c2, max_row=nr2))

def add_bottom_grid_to_last_used_row(ws):
    last_used_row = 0
    last_used_col = 0
    max_col = ws.max_column

    for r in range(ws.max_row, 0, -1):
        row_has_data = False
        row_last_col = 0
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, "") and str(v).strip() != "":
                row_has_data = True
                row_last_col = c
        if row_has_data:
            last_used_row = r
            last_used_col = row_last_col
            break

    if last_used_row == 0 or last_used_col == 0:
        return

    thin = Side(style="thin")
    for c in range(1, last_used_col + 1):
        cell = ws.cell(row=last_used_row, column=c)
        b = cell.border
        cell.border = Border(
            left=b.left,
            right=b.right,
            top=b.top,
            bottom=thin,
            diagonal=b.diagonal,
            diagonal_direction=b.diagonal_direction,
            outline=b.outline,
            vertical=b.vertical,
            horizontal=b.horizontal,
        )

def merge_col_ab_pairs_for_n_percent(ws, crosstab_mode="NORMAL"):
    mode = (crosstab_mode or "").upper()
    max_col = ws.max_column

    stub_starts = [r for r, cell in enumerate(ws['A'], 1) if str(cell.value).startswith('Stub')]
    if not stub_starts:
        stub_starts = [1]
    table_boundaries = stub_starts + [ws.max_row + 2]

    # Rows whose column A already belongs to a merge. delete_rows_preserving_merges
    # has normally already merged the Count/Column% A-B pairs, so this lets us skip
    # the expensive unmerge+remerge for pairs that are already correct.
    premerged_a_rows = set()
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= 1 <= mr.max_col:
            premerged_a_rows.update(range(mr.min_row, mr.max_row + 1))

    for i in range(len(table_boundaries) - 1):
        start_row = table_boundaries[i]
        end_row = table_boundaries[i + 1] - 2
        if end_row <= start_row:
            continue

        data_total_row = -1
        if mode == "MATRIX":
            data_total_row = find_total_row_in_colB(ws, start_row, end_row)
        else:
            for r in range(start_row, end_row + 1):
                c = ws.cell(row=r, column=3).value
                if str(c).strip().upper() == 'TOTAL':
                    data_total_row = r
                    break
        if data_total_row == -1:
            continue

        r = data_total_row + 1
        while r <= end_row:
            label = ws.cell(row=r, column=2).value
            has_label = label not in (None, "") and str(label).strip() != ""

            if not has_label:
                r += 1
                continue

            r2 = r + 1
            if r2 <= end_row:
                next_label = ws.cell(row=r2, column=2).value
                next_has_label = next_label not in (None, "") and str(next_label).strip() != ""
                if not next_has_label:
                    # Already merged by the delete step -> nothing to do.
                    if r in premerged_a_rows and r2 in premerged_a_rows:
                        r = r2 + 1
                        continue
                    row2_has_data = any(
                        ws.cell(row=r2, column=cc).value not in (None, "") and
                        str(ws.cell(row=r2, column=cc).value).strip() != ""
                        for cc in range(3, max_col + 1)
                    )
                    if row2_has_data:
                        for mr in list(ws.merged_cells.ranges):
                            touches_ab = not (mr.max_col < 1 or mr.min_col > 2)
                            overlaps_rows = not (mr.max_row < r or mr.min_row > r2)
                            if touches_ab and overlaps_rows:
                                ws.unmerge_cells(str(mr))
                        ws.cell(row=r2, column=1).value = None
                        ws.merge_cells(start_row=r, start_column=1, end_row=r2, end_column=1)
                        ws.merge_cells(start_row=r, start_column=2, end_row=r2, end_column=2)
                        r = r2 + 1
                        continue
            r += 1


# ===================== Core Processing =====================

def place_sig_stamp(ws, header_row, sig_text, col=3):
    stamp_row = header_row - 2 if header_row - 2 >= 1 else header_row
    cell = ws.cell(row=stamp_row, column=col)
    cell.value = sig_text
    cell.font = Font(name='Arial', size=9, color='FFFF0000')


def process_single_excel_file(file_path, sig_input, crosstab_mode="NORMAL", sig_beside=False):
    import openpyxl
    try:
        rules_by_char = parse_sig_groups(sig_input)
        wb = openpyxl.load_workbook(file_path)
        sheets_to_process = [n for n in wb.sheetnames if n.lower() not in ['contents', 'info']]
        cells_changed_count = 0

        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            rows_to_delete_in_sheet = []

            stub_starts = [r for r, cell in enumerate(ws['A'], 1) if str(cell.value).startswith('Stub')]
            if not stub_starts:
                stub_starts = [1]
            table_boundaries = stub_starts + [ws.max_row + 2]

            for i in range(len(table_boundaries) - 1):
                start_row = table_boundaries[i]
                end_row   = table_boundaries[i + 1] - 2
                if end_row <= start_row:
                    continue

                sig_text = f"Sig Used: {sig_input}"
                n_percent_layout = is_n_percent_layout(ws, start_row, end_row)

                if crosstab_mode.upper() == "NORMAL":
                    header_row, data_total_row = -1, -1
                    for r in range(start_row, end_row + 1):
                        d = ws.cell(row=r, column=4).value
                        e = ws.cell(row=r, column=5).value
                        c = ws.cell(row=r, column=3).value
                        if clean_header(str(d)) == 'A' and clean_header(str(e)) == 'B':
                            header_row = r
                        if str(c).strip().upper() == 'TOTAL':
                            data_total_row = r
                    if header_row == -1 or data_total_row == -1:
                        continue

                    place_sig_stamp(ws, header_row, sig_text, col=3)

                    for col in range(3, ws.max_column + 1):
                        hdr = clean_header(str(ws.cell(row=header_row, column=col).value))
                        if hdr in rules_by_char:
                            keep = set(rules_by_char[hdr])
                            for r in range(data_total_row + 1, end_row + 1):
                                v = ws.cell(row=r, column=col).value
                                if isinstance(v, str) and v.strip():
                                    nv = filter_sig_text(v, keep, preserve_tokens=("ADJ",))
                                    if nv != v:
                                        ws.cell(row=r, column=col).value = nv
                                        cells_changed_count += 1

                    if sig_beside:
                        data_cols_n = list(range(3, ws.max_column + 1))
                        sig_rows = collect_sig_rows_and_merge(
                            ws, data_total_row + 1, end_row, label_cols=[1, 2], data_cols=data_cols_n,
                            keep_first_unlabeled=n_percent_layout
                        )
                        rows_to_delete_in_sheet.extend(sig_rows)
                    continue

                # ================= MATRIX =================
                header_row, first_col = detect_letter_header_row(ws, start_row, end_row, min_seq=3)
                if header_row == -1:
                    continue

                place_sig_stamp(ws, header_row, sig_text, col=3)

                col_map = build_col_letter_map(ws, header_row, first_col)
                if not col_map:
                    continue

                total_row = find_total_row_in_colB(ws, header_row, end_row)
                if total_row == -1:
                    continue
                data_start = total_row
                data_end   = find_last_data_row(ws, data_start, list(col_map.keys()))
                if data_end < data_start:
                    continue

                for c, letter in col_map.items():
                    allowed = set(rules_by_char.get(letter, ""))
                    for r in range(data_start, data_end + 1):
                        v = ws.cell(row=r, column=c).value
                        if isinstance(v, str) and v.strip():
                            nv = filter_sig_text(v, allowed, preserve_tokens=("ADJ",))
                            if nv != v:
                                ws.cell(row=r, column=c).value = nv
                                cells_changed_count += 1

                if sig_beside:
                    sig_rows = collect_sig_rows_and_merge(
                        ws, data_start, data_end, label_cols=[1, 2], data_cols=list(col_map.keys()),
                        keep_first_unlabeled=n_percent_layout
                    )
                    rows_to_delete_in_sheet.extend(sig_rows)

            if rows_to_delete_in_sheet:
                transfer_bottom_borders_before_delete(ws, rows_to_delete_in_sheet)
                delete_rows_preserving_merges(ws, rows_to_delete_in_sheet)

                renumber_col_a_by_col_b(ws, crosstab_mode)
                add_bottom_grid_to_last_used_row(ws)

            if sig_beside:
                merge_col_ab_pairs_for_n_percent(ws, crosstab_mode)

        wb.save(file_path)
        return cells_changed_count, "Success"

    except Exception as e:
        return 0, f"Error: {e}"


# ===================== Pastel PyQt6 GUI =====================

STYLESHEET = """
/* ── Base ── */
QMainWindow {
    background-color: #f5f0eb;
}
QWidget#centralWidget {
    background-color: #f5f0eb;
}

/* ── Typography ── */
QLabel#titleLabel {
    font-size: 20px;
    font-weight: 700;
    color: #5b4a6b;
}
QLabel#subtitleLabel {
    font-size: 11px;
    color: #a89bb5;
    letter-spacing: 0.3px;
}
QLabel#sectionLabel {
    font-size: 12px;
    font-weight: 600;
    color: #6b5e7b;
}

/* ── Browse Button (soft mint) ── */
QPushButton#browseBtn {
    background-color: #b8e6d0;
    color: #2d6e4f;
    border: 1px solid #a0d4ba;
    border-radius: 8px;
    padding: 9px 22px;
    font-size: 13px;
    font-weight: 600;
}
QPushButton#browseBtn:hover {
    background-color: #a0d4ba;
}
QPushButton#browseBtn:pressed {
    background-color: #8cc5a8;
}

/* ── Delete Button (soft rose) ── */
QPushButton#deleteBtn {
    background-color: #f2c4c4;
    color: #8b3a3a;
    border: 1px solid #e8abab;
    border-radius: 8px;
    padding: 9px 22px;
    font-size: 13px;
    font-weight: 600;
}
QPushButton#deleteBtn:hover {
    background-color: #e8abab;
}
QPushButton#deleteBtn:pressed {
    background-color: #dc9595;
}

/* ── Process Button (soft lavender gradient) ── */
QPushButton#processBtn {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #c4b5d4, stop:1 #b8c6e8);
    color: #3d2e5c;
    border: 1px solid #b0a3c4;
    border-radius: 10px;
    padding: 13px 20px;
    font-size: 15px;
    font-weight: 700;
}
QPushButton#processBtn:hover {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #b5a3c8, stop:1 #a8b8de);
}
QPushButton#processBtn:pressed {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #a694bc, stop:1 #98aad4);
}
QPushButton#processBtn:disabled {
    background: #ddd8e4;
    color: #a099ab;
    border-color: #d0cad8;
}

/* ── Sig Input ── */
QLineEdit#sigInput {
    border: 1.5px solid #d8d0e0;
    border-radius: 8px;
    padding: 10px 14px;
    font-size: 13px;
    background-color: #fffcfa;
    color: #4a3f5c;
}
QLineEdit#sigInput:focus {
    border-color: #b5a3c8;
    background-color: #ffffff;
}

/* ── File List ── */
QListWidget#fileList {
    border: 1.5px solid #d8d0e0;
    border-radius: 10px;
    padding: 5px;
    font-size: 12px;
    background-color: #fffcfa;
    color: #4a3f5c;
    outline: none;
}
QListWidget#fileList::item {
    padding: 6px 10px;
    border-radius: 6px;
    margin: 1px 0px;
}
QListWidget#fileList::item:selected {
    background-color: #e8dff2;
    color: #5b4a6b;
}
QListWidget#fileList::item:hover {
    background-color: #f0eaf5;
}

/* ── Radio Buttons ── */
QRadioButton {
    font-size: 12px;
    color: #5b4a6b;
    spacing: 6px;
}
QRadioButton::indicator {
    width: 16px;
    height: 16px;
    border-radius: 8px;
    border: 2px solid #c4b5d4;
    background-color: #fffcfa;
}
QRadioButton::indicator:checked {
    background-color: #c4b5d4;
    border-color: #a694bc;
}
QRadioButton::indicator:hover {
    border-color: #a694bc;
}

/* ── Settings Card ── */
QFrame#card {
    background-color: #fffcfa;
    border: 1.5px solid #e4dced;
    border-radius: 14px;
}

/* ── Separator inside card ── */
QFrame#separator {
    background-color: #e8e0f0;
    max-height: 1px;
    border: none;
}

/* ── Status Bar ── */
QStatusBar {
    background-color: #ede7e0;
    color: #8a7e96;
    font-size: 11px;
    border-top: 1px solid #ddd6cc;
    padding: 2px 8px;
}

/* ── Progress Bar ── */
QProgressBar {
    border: none;
    border-radius: 4px;
    background-color: #e4dced;
    height: 6px;
    text-align: center;
}
QProgressBar::chunk {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #c4b5d4, stop:1 #b8c6e8);
    border-radius: 4px;
}

/* ── Log Panel (light pastel) ── */
QTextEdit#logPanel {
    border: 1.5px solid #d8d0e0;
    border-radius: 10px;
    padding: 8px;
    font-family: 'Cascadia Code', 'Consolas', 'Courier New', monospace;
    font-size: 11px;
    background-color: #faf8fc;
    color: #5b4a6b;
    selection-background-color: #e4dced;
    selection-color: #3d2e5c;
}

/* ── Clear Log Button ── */
QPushButton#clearLogBtn {
    background: #ede7f2;
    border: 1px solid #d8d0e0;
    border-radius: 6px;
    padding: 4px 14px;
    font-size: 11px;
    color: #8a7e96;
}
QPushButton#clearLogBtn:hover {
    background: #e0d8ec;
    color: #6b5e7b;
}

/* ── Splitter Handle ── */
QSplitter::handle {
    background-color: #ddd6e8;
    border-radius: 3px;
}
QSplitter::handle:hover {
    background-color: #c4b5d4;
}

/* ── Scrollbars ── */
QScrollBar:vertical {
    background: #f0eaf5;
    width: 8px;
    border-radius: 4px;
    margin: 0;
}
QScrollBar::handle:vertical {
    background: #d0c6dc;
    border-radius: 4px;
    min-height: 30px;
}
QScrollBar::handle:vertical:hover {
    background: #c4b5d4;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    height: 0px;
}
QScrollBar:horizontal {
    background: #f0eaf5;
    height: 8px;
    border-radius: 4px;
}
QScrollBar::handle:horizontal {
    background: #d0c6dc;
    border-radius: 4px;
    min-width: 30px;
}
QScrollBar::handle:horizontal:hover {
    background: #c4b5d4;
}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
    width: 0px;
}
"""


class FileListWidget(QListWidget):
    """QListWidget that accepts drag-and-drop .xlsx files."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.parent_app = None

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        urls = event.mimeData().urls()
        paths = []
        for url in urls:
            path = url.toLocalFile()
            if path.lower().endswith('.xlsx'):
                paths.append(path)
        if paths and self.parent_app:
            self.parent_app.add_files(paths)


class ProcessWorker(QThread):
    """Runs the Excel processing on a background thread so the window never freezes.

    All GUI updates happen in the main thread via these signals; the worker only
    crunches files and reports progress.
    """
    file_started = pyqtSignal(int, int, str)             # index, total, filename
    file_finished = pyqtSignal(int, int, str, int, str)  # index, total, filename, cells, status
    all_finished = pyqtSignal(int, int, int, bool, str)  # processed, total, cells, had_errors, last_error

    def __init__(self, file_paths, sig_groups, mode, sig_beside):
        super().__init__()
        self.file_paths = list(file_paths)   # snapshot so GUI edits can't disturb the run
        self.sig_groups = sig_groups
        self.mode = mode
        self.sig_beside = sig_beside

    def run(self):
        total = len(self.file_paths)
        processed = 0
        total_cells = 0
        had_errors = False
        last_error = ""
        for i, file_path in enumerate(self.file_paths):
            fname = os.path.basename(file_path)
            self.file_started.emit(i, total, fname)
            cells_changed, status = process_single_excel_file(
                file_path, self.sig_groups, self.mode, self.sig_beside
            )
            self.file_finished.emit(i, total, fname, cells_changed, status)
            if status == "Success":
                processed += 1
                total_cells += cells_changed
            else:
                had_errors = True
                last_error = status
                break
        self.all_finished.emit(processed, total, total_cells, had_errors, last_error)


class App(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Del Sig \u2014 Lychee Table Tool")
        self.setMinimumSize(740, 660)
        self.resize(780, 720)

        self.file_paths = []
        self.worker = None
        self.HINT_TEXT = "\u0e40\u0e0a\u0e48\u0e19 ABCDEF,GHIJKL,MNOPQR,STUVWX,YZ"

        self.setStyleSheet(STYLESHEET)
        self._center_on_screen()
        self._build_ui()

    def closeEvent(self, event):
        # Wait for the background processing to finish so the QThread is not
        # destroyed mid-run (which would crash).
        if self.worker is not None and self.worker.isRunning():
            self.worker.wait()
        super().closeEvent(event)

    def _center_on_screen(self):
        screen = QApplication.primaryScreen()
        if screen:
            geo = screen.availableGeometry()
            x = (geo.width() - self.width()) // 2 + geo.x()
            y = (geo.height() - self.height()) // 2 + geo.y()
            self.move(x, y)

    def _make_separator(self):
        sep = QFrame()
        sep.setObjectName("separator")
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setFixedHeight(1)
        return sep

    def _build_ui(self):
        central = QWidget()
        central.setObjectName("centralWidget")
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(26, 22, 26, 10)
        main_layout.setSpacing(12)

        # ── Header ──
        title = QLabel("\u0e42\u0e1b\u0e23\u0e41\u0e01\u0e23\u0e21\u0e25\u0e1a Sig \u0e08\u0e32\u0e01 Table Lychee")
        title.setObjectName("titleLabel")
        main_layout.addWidget(title)

        subtitle = QLabel("Crosstab & Matrix  \u00b7  Sig Beside  \u00b7  Drag & Drop")
        subtitle.setObjectName("subtitleLabel")
        main_layout.addWidget(subtitle)

        main_layout.addSpacing(2)

        # ── Top Buttons ──
        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)

        self.browse_btn = QPushButton("\u0e40\u0e25\u0e37\u0e2d\u0e01\u0e44\u0e1f\u0e25\u0e4c...")
        self.browse_btn.setObjectName("browseBtn")
        self.browse_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.browse_btn.clicked.connect(self.browse_files)
        btn_row.addWidget(self.browse_btn)

        self.delete_btn = QPushButton("\u0e25\u0e1a\u0e44\u0e1f\u0e25\u0e4c\u0e17\u0e35\u0e48\u0e40\u0e25\u0e37\u0e2d\u0e01")
        self.delete_btn.setObjectName("deleteBtn")
        self.delete_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.delete_btn.clicked.connect(self.delete_selected_files)
        btn_row.addWidget(self.delete_btn)

        btn_row.addStretch()
        main_layout.addLayout(btn_row)

        # ── Settings Card ──
        card = QFrame()
        card.setObjectName("card")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(20, 16, 20, 16)
        card_layout.setSpacing(10)

        # Crosstab mode
        mode_label = QLabel("\u0e1b\u0e23\u0e30\u0e40\u0e20\u0e17 Crosstab")
        mode_label.setObjectName("sectionLabel")
        card_layout.addWidget(mode_label)

        mode_row = QHBoxLayout()
        mode_row.setSpacing(24)
        self.radio_normal = QRadioButton("Crosstab \u0e18\u0e23\u0e23\u0e21\u0e14\u0e32")
        self.radio_matrix = QRadioButton("Matrix")
        self.radio_normal.setChecked(True)
        self.mode_group = QButtonGroup(self)
        self.mode_group.addButton(self.radio_normal, 0)
        self.mode_group.addButton(self.radio_matrix, 1)
        mode_row.addWidget(self.radio_normal)
        mode_row.addWidget(self.radio_matrix)
        mode_row.addStretch()
        card_layout.addLayout(mode_row)

        card_layout.addWidget(self._make_separator())

        # Sig mode
        sig_mode_label = QLabel("\u0e23\u0e39\u0e1b\u0e41\u0e1a\u0e1a Sig")
        sig_mode_label.setObjectName("sectionLabel")
        card_layout.addWidget(sig_mode_label)

        sig_mode_row = QHBoxLayout()
        sig_mode_row.setSpacing(24)
        self.radio_sig_normal = QRadioButton("Sig \u0e1b\u0e01\u0e15\u0e34")
        self.radio_sig_beside = QRadioButton("Sig\u0e02\u0e49\u0e32\u0e07 (\u0e22\u0e01 Sig \u0e02\u0e36\u0e49\u0e19\u0e1a\u0e23\u0e23\u0e17\u0e31\u0e14\u0e40\u0e14\u0e35\u0e22\u0e27\u0e01\u0e31\u0e1a\u0e15\u0e31\u0e27\u0e40\u0e25\u0e02 \u0e41\u0e25\u0e30\u0e25\u0e1a\u0e41\u0e16\u0e27 Sig \u0e40\u0e14\u0e34\u0e21)")
        self.radio_sig_normal.setChecked(True)
        self.sig_mode_group = QButtonGroup(self)
        self.sig_mode_group.addButton(self.radio_sig_normal, 0)
        self.sig_mode_group.addButton(self.radio_sig_beside, 1)
        sig_mode_row.addWidget(self.radio_sig_normal)
        sig_mode_row.addWidget(self.radio_sig_beside)
        sig_mode_row.addStretch()
        card_layout.addLayout(sig_mode_row)

        card_layout.addWidget(self._make_separator())

        # Sig input
        sig_label = QLabel("\u0e43\u0e2a\u0e48 Sig \u0e15\u0e23\u0e07\u0e19\u0e35\u0e49")
        sig_label.setObjectName("sectionLabel")
        card_layout.addWidget(sig_label)

        self.sig_input = QLineEdit()
        self.sig_input.setObjectName("sigInput")
        self.sig_input.setPlaceholderText(self.HINT_TEXT)
        card_layout.addWidget(self.sig_input)

        main_layout.addWidget(card)

        # ── File List + Log Splitter ──
        splitter = QSplitter(Qt.Orientation.Vertical)
        splitter.setHandleWidth(6)

        # File list panel
        file_panel = QWidget()
        file_layout = QVBoxLayout(file_panel)
        file_layout.setContentsMargins(0, 0, 0, 0)
        file_layout.setSpacing(4)

        file_label = QLabel("\u0e44\u0e1f\u0e25\u0e4c\u0e17\u0e35\u0e48\u0e40\u0e25\u0e37\u0e2d\u0e01 (0 \u0e44\u0e1f\u0e25\u0e4c)")
        file_label.setObjectName("sectionLabel")
        self.file_count_label = file_label
        file_layout.addWidget(file_label)

        self.file_list = FileListWidget()
        self.file_list.setObjectName("fileList")
        self.file_list.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.file_list.parent_app = self
        file_layout.addWidget(self.file_list)

        splitter.addWidget(file_panel)

        # Log panel
        log_panel = QWidget()
        log_layout = QVBoxLayout(log_panel)
        log_layout.setContentsMargins(0, 0, 0, 0)
        log_layout.setSpacing(4)

        log_header = QHBoxLayout()
        log_label = QLabel("Log")
        log_label.setObjectName("sectionLabel")
        log_header.addWidget(log_label)
        log_header.addStretch()

        clear_log_btn = QPushButton("\u0e25\u0e49\u0e32\u0e07 Log")
        clear_log_btn.setObjectName("clearLogBtn")
        clear_log_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        clear_log_btn.clicked.connect(lambda: self.log_view.clear())
        log_header.addWidget(clear_log_btn)
        log_layout.addLayout(log_header)

        self.log_view = QTextEdit()
        self.log_view.setObjectName("logPanel")
        self.log_view.setReadOnly(True)
        log_layout.addWidget(self.log_view)

        splitter.addWidget(log_panel)
        splitter.setSizes([180, 140])

        main_layout.addWidget(splitter, 1)

        # ── Progress Bar (hidden by default) ──
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setFixedHeight(6)
        main_layout.addWidget(self.progress_bar)

        # ── Process Button ──
        self.process_btn = QPushButton("\u0e40\u0e23\u0e34\u0e48\u0e21\u0e15\u0e49\u0e19\u0e25\u0e1a Sig")
        self.process_btn.setObjectName("processBtn")
        self.process_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.process_btn.clicked.connect(self.process_files)
        main_layout.addWidget(self.process_btn)

        # ── Status Bar ──
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("\u0e1e\u0e23\u0e49\u0e2d\u0e21\u0e43\u0e0a\u0e49\u0e07\u0e32\u0e19")

    def log(self, message, level="INFO"):
        time_str = datetime.now().strftime("%H:%M:%S")
        color_map = {
            "INFO":  "#7b8fc4",
            "OK":    "#5a9e6f",
            "WARN":  "#c49a3c",
            "ERROR": "#c45a5a",
            "START": "#8b6aad",
        }
        color = color_map.get(level, "#6b5e7b")
        level_tag = f'<span style="color:{color};font-weight:bold;">[{level}]</span>'
        time_tag = f'<span style="color:#a89bb5;">{time_str}</span>'
        self.log_view.append(f'{time_tag} {level_tag} {message}')
        self.log_view.verticalScrollBar().setValue(self.log_view.verticalScrollBar().maximum())

    def _update_file_count(self):
        n = len(self.file_paths)
        self.file_count_label.setText(f"\u0e44\u0e1f\u0e25\u0e4c\u0e17\u0e35\u0e48\u0e40\u0e25\u0e37\u0e2d\u0e01 ({n} \u0e44\u0e1f\u0e25\u0e4c)")

    def add_files(self, paths):
        existing = set(self.file_paths)
        added = 0
        for p in paths:
            if p not in existing:
                self.file_paths.append(p)
                self.file_list.addItem(os.path.basename(p))
                added += 1
        self._update_file_count()
        self.log(f"Drag & Drop เพิ่ม {added} ไฟล์ (รวม {len(self.file_paths)} ไฟล์)")
        self.status_bar.showMessage(f"\u0e40\u0e25\u0e37\u0e2d\u0e01\u0e44\u0e1f\u0e25\u0e4c\u0e41\u0e25\u0e49\u0e27 {len(self.file_paths)} \u0e44\u0e1f\u0e25\u0e4c \u0e1e\u0e23\u0e49\u0e2d\u0e21\u0e25\u0e1a Sig")

    def browse_files(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "\u0e40\u0e25\u0e37\u0e2d\u0e01\u0e44\u0e1f\u0e25\u0e4c Excel", "",
            "Excel Files (*.xlsx)"
        )
        if paths:
            self.file_paths = list(paths)
            self.file_list.clear()
            for p in self.file_paths:
                self.file_list.addItem(os.path.basename(p))
            self._update_file_count()
            self.log(f"\u0e40\u0e25\u0e37\u0e2d\u0e01\u0e44\u0e1f\u0e25\u0e4c {len(self.file_paths)} \u0e44\u0e1f\u0e25\u0e4c")
            for p in self.file_paths:
                self.log(f"  + {os.path.basename(p)}")
            self.status_bar.showMessage(f"\u0e40\u0e25\u0e37\u0e2d\u0e01\u0e44\u0e1f\u0e25\u0e4c\u0e41\u0e25\u0e49\u0e27 {len(self.file_paths)} \u0e44\u0e1f\u0e25\u0e4c \u0e1e\u0e23\u0e49\u0e2d\u0e21\u0e25\u0e1a Sig")

    def delete_selected_files(self):
        idxs = sorted([i.row() for i in self.file_list.selectedIndexes()], reverse=True)
        if not idxs:
            return
        for i in idxs:
            removed_name = os.path.basename(self.file_paths[i])
            self.file_list.takeItem(i)
            del self.file_paths[i]
            self.log(f"\u0e25\u0e1a\u0e44\u0e1f\u0e25\u0e4c: {removed_name}", "WARN")
        self._update_file_count()
        self.status_bar.showMessage(f"\u0e04\u0e07\u0e40\u0e2b\u0e25\u0e37\u0e2d {len(self.file_paths)} \u0e44\u0e1f\u0e25\u0e4c")

    def process_files(self):
        if getattr(self, "worker", None) is not None and self.worker.isRunning():
            return  # already processing
        if not self.file_paths:
            QMessageBox.warning(self, "Warning", "Please select at least one Excel file.")
            return

        sig_groups = self.sig_input.text().strip()
        if not sig_groups:
            QMessageBox.warning(self, "Warning", "Please enter Sig groups.")
            return

        mode = "MATRIX" if self.radio_matrix.isChecked() else "NORMAL"
        sig_beside = self.radio_sig_beside.isChecked()
        total_files = len(self.file_paths)

        # remember run settings for the summary message produced when finished
        self._run_mode = mode
        self._run_beside = sig_beside

        side_label = "Sig\u0e02\u0e49\u0e32\u0e07" if sig_beside else "Sig \u0e1b\u0e01\u0e15\u0e34"
        self.log(f"\u0e40\u0e23\u0e34\u0e48\u0e21\u0e1b\u0e23\u0e30\u0e21\u0e27\u0e25\u0e1c\u0e25 {total_files} \u0e44\u0e1f\u0e25\u0e4c  |  Mode: {mode}  |  {side_label}", "START")
        self.log(f"Sig Groups: {sig_groups}", "START")

        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(total_files)
        self.progress_bar.setValue(0)
        self.process_btn.setEnabled(False)
        self.browse_btn.setEnabled(False)
        self.delete_btn.setEnabled(False)

        # Run the heavy work off the GUI thread so the window stays responsive.
        self.worker = ProcessWorker(self.file_paths, sig_groups, mode, sig_beside)
        self.worker.file_started.connect(self._on_file_started)
        self.worker.file_finished.connect(self._on_file_finished)
        self.worker.all_finished.connect(self._on_all_finished)
        self.worker.start()

    def _on_file_started(self, i, total, fname):
        self.log(f"[{i+1}/{total}] \u0e01\u0e33\u0e25\u0e31\u0e07\u0e1b\u0e23\u0e30\u0e21\u0e27\u0e25\u0e1c\u0e25: {fname}...")
        self.status_bar.showMessage(f"Processing ({self._run_mode}) file {i+1}/{total}: {fname}...")

    def _on_file_finished(self, i, total, fname, cells_changed, status):
        if status == "Success":
            self.log(f"[{i+1}/{total}] {fname} \u0e2a\u0e33\u0e40\u0e23\u0e47\u0e08 \u2014 \u0e41\u0e01\u0e49\u0e44\u0e02 {cells_changed} cells", "OK")
            self.progress_bar.setValue(i + 1)
        else:
            self.log(f"[{i+1}/{total}] {fname} \u0e1c\u0e34\u0e14\u0e1e\u0e25\u0e32\u0e14: {status}", "ERROR")

    def _on_all_finished(self, processed_files, total_files, total_cells_changed, had_errors, last_error):
        self.process_btn.setEnabled(True)
        self.browse_btn.setEnabled(True)
        self.delete_btn.setEnabled(True)
        self.progress_bar.setVisible(False)

        mode = self._run_mode
        sig_beside = self._run_beside
        side_txt = " + Sig\u0e02\u0e49\u0e32\u0e07" if sig_beside else ""
        msg = (
            f"\u0e25\u0e1a Sig \u0e40\u0e23\u0e35\u0e22\u0e1a\u0e23\u0e49\u0e2d\u0e22 [{mode}{side_txt}]\n\n"
            f"\u0e2a\u0e33\u0e40\u0e23\u0e47\u0e08\u0e17\u0e31\u0e49\u0e07\u0e2b\u0e21\u0e14: {processed_files}/{total_files} files\n"
            f"\u0e41\u0e01\u0e49\u0e44\u0e02\u0e44\u0e1b\u0e17\u0e31\u0e49\u0e07\u0e2b\u0e21\u0e14: {total_cells_changed} cells"
        )

        self.log(f"\u0e40\u0e2a\u0e23\u0e47\u0e08\u0e2a\u0e34\u0e49\u0e19! \u0e2a\u0e33\u0e40\u0e23\u0e47\u0e08 {processed_files}/{total_files} \u0e44\u0e1f\u0e25\u0e4c, \u0e41\u0e01\u0e49\u0e44\u0e02\u0e23\u0e27\u0e21 {total_cells_changed} cells",
                 "OK" if not had_errors else "WARN")

        if had_errors:
            msg += f"\n\nSome files encountered errors and were not processed.\nDetails: {last_error}"
            QMessageBox.warning(self, "Process Finished with Errors", msg)
        else:
            QMessageBox.information(self, "Success", msg)

        self.file_list.clear()
        self.file_paths.clear()
        self.sig_input.clear()
        self._update_file_count()
        self.status_bar.showMessage("\u0e40\u0e23\u0e34\u0e48\u0e21\u0e15\u0e49\u0e19\u0e25\u0e1a Sig.....\u0e23\u0e2d\u0e1a\u0e15\u0e48\u0e2d\u0e44\u0e1b")


# -------- Entry point --------
def run_this_app(working_dir=None):
    try:
        app = QApplication(sys.argv)
        app.setFont(QFont("Segoe UI", 10))
        window = App()
        window.show()
        sys.exit(app.exec())
    except Exception as e:
        print(f"Error running app: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    run_this_app()
