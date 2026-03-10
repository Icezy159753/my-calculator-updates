import sys
import os
import pandas as pd
import numpy as np
import pyreadstat
import re
from io import BytesIO
import matplotlib
matplotlib.use('QtAgg')
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.lines import Line2D
from matplotlib.ticker import FuncFormatter
from matplotlib.patches import Patch
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QPushButton, QFileDialog, QLabel,
                             QComboBox, QListWidget, QListWidgetItem,
                             QMessageBox, QAbstractItemView, QGroupBox, QLineEdit,
                             QSplitter, QDialog, QScrollArea)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QIcon
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XlImage
try:
    # Preferred when loaded by Main_Program as package: All_Programs.147_Penalty
    from .penalty_logic import (
        build_summary_output,
        FilterQueryError,
        FilterSpec,
        HIGH_INFLUENCE_LABEL,
        HIGH_INFLUENCE_THRESHOLD,
        MEDIUM_INFLUENCE_LABEL,
        MEDIUM_INFLUENCE_THRESHOLD,
        NO_FILTER_LABEL,
        analyze_penalty,
    )
except ImportError:
    # Fallback when this file is run directly (without package context)
    from penalty_logic import (
        build_summary_output,
        FilterQueryError,
        FilterSpec,
        HIGH_INFLUENCE_LABEL,
        HIGH_INFLUENCE_THRESHOLD,
        MEDIUM_INFLUENCE_LABEL,
        MEDIUM_INFLUENCE_THRESHOLD,
        NO_FILTER_LABEL,
        analyze_penalty,
    )

EXCEL_DEFAULT_ROW_PIXELS = 20
EXCEL_ROW_HEIGHT_POINTS = 15
EXCEL_POINTS_PER_PIXEL = EXCEL_ROW_HEIGHT_POINTS / EXCEL_DEFAULT_ROW_PIXELS
EXCEL_IMAGE_SPACER_ROWS = 2
SUMMARY_BOTH_ISSUES_FILL = 'E2F0D9'
SUMMARY_TOO_LITTLE_FILL = 'FFF2CC'
SUMMARY_TOO_MUCH_FILL = 'FCE4D6'
SUMMARY_NO_ISSUE_FILL = 'F2F2F2'


def _resource_path(relative_path: str) -> str:
    """Resolve bundled resource paths for both source and PyInstaller builds."""
    base_path = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

def _build_single_chart(sub_df: pd.DataFrame, sorted_attrs: list,
                        filter_name: str, xl: float) -> Figure:
    """Build one butterfly chart for a single filter group."""
    bar_h = 0.55
    faint_alpha = 0.20
    bar_color = '#c0392b'
    medium_line_color = '#7c3aed'
    high_line_color = '#ef4444'
    n_a = len(sorted_attrs)
    fig_w = 8.5
    fig_h = max(2.8, n_a * 0.30 + 1.4)

    fig = Figure(figsize=(fig_w, fig_h), dpi=100, facecolor='white')
    gs = fig.add_gridspec(1, 2, wspace=0.06, width_ratios=[1.1, 0.9],
                          left=0.22, right=0.97, top=0.86, bottom=0.16)
    ax_l = fig.add_subplot(gs[0, 0])
    ax_r = fig.add_subplot(gs[0, 1], sharey=ax_l)
    medium_left_rows = []
    high_left_rows = []
    medium_right_rows = []
    high_right_rows = []

    def _threshold_text_bbox(value: float):
        abs_value = abs(float(value))
        if abs_value >= HIGH_INFLUENCE_THRESHOLD:
            return dict(
                boxstyle='square,pad=0.18',
                facecolor='white',
                edgecolor=high_line_color,
                linewidth=1.3,
                linestyle='--',
            )
        if abs_value >= MEDIUM_INFLUENCE_THRESHOLD:
            return dict(
                boxstyle='square,pad=0.18',
                facecolor='white',
                edgecolor=medium_line_color,
                linewidth=1.1,
                linestyle=':',
            )
        return None

    for ai, attr in enumerate(sorted_attrs):
        m = sub_df['Attribute'] == attr
        if not m.any():
            continue
        r = sub_df[m].iloc[0]

        # ── Left: Too Little ──
        w12 = float(r['Weighted Penalty 1+2'])
        s12 = str(r.get('Sig 1+2', ''))
        a12 = 1.0 if s12 in ('*', '**') else faint_alpha
        ax_l.barh(ai, w12, height=bar_h, color=bar_color, alpha=a12,
                  edgecolor='white', linewidth=0.5)
        if w12 != 0:
            ax_l.text(w12, ai, f' {w12:.2f} ', va='center', ha='right',
                      fontsize=6, color='#111',
                      fontweight='bold' if s12 in ('*', '**') else 'normal',
                      bbox=_threshold_text_bbox(w12))
        abs_w12 = abs(w12)
        if abs_w12 >= MEDIUM_INFLUENCE_THRESHOLD:
            medium_left_rows.append(ai)
        if abs_w12 >= HIGH_INFLUENCE_THRESHOLD:
            high_left_rows.append(ai)

        # ── Right: Too Much ──
        w45 = float(r['Weighted Penalty 4+5'])
        s45 = str(r.get('Sig 4+5', ''))
        a45 = 1.0 if s45 in ('*', '**') else faint_alpha
        ax_r.barh(ai, abs(w45), height=bar_h, color=bar_color, alpha=a45,
                  edgecolor='white', linewidth=0.5)
        if w45 != 0:
            ax_r.text(abs(w45), ai, f' {w45:.2f} ', va='center', ha='left',
                      fontsize=6, color='#111',
                      fontweight='bold' if s45 in ('*', '**') else 'normal',
                      bbox=_threshold_text_bbox(w45))
        abs_w45 = abs(w45)
        if abs_w45 >= MEDIUM_INFLUENCE_THRESHOLD:
            medium_right_rows.append(ai)
        if abs_w45 >= HIGH_INFLUENCE_THRESHOLD:
            high_right_rows.append(ai)

        # %TL / %TM
        ax_l.text(-xl + 0.01, ai, str(int(r['Base (%) 1+2'])),
                  va='center', ha='left', fontsize=6, color='#888')
        ax_r.text(xl - 0.01, ai, str(int(r['Base (%) 4+5'])),
                  va='center', ha='right', fontsize=6, color='#888')

    # Separators
    for ai in range(n_a - 1):
        ax_l.axhline(ai + 0.5, color='#eee', linewidth=0.5)
        ax_r.axhline(ai + 0.5, color='#eee', linewidth=0.5)

    # Axis format
    ax_l.set_xlim(-xl, 0)
    ax_r.set_xlim(0, xl)
    ax_l.set_ylim(-0.5, n_a - 0.5)
    ax_l.invert_yaxis()

    labels = [(a[:30] + '…') if len(a) > 30 else a for a in sorted_attrs]
    ax_l.set_yticks(range(n_a))
    ax_l.set_yticklabels(labels, fontsize=7)

    ax_l.set_title('Too little', fontsize=8, fontweight='bold', color='#1565c0', pad=4)
    ax_l.axvline(0, color='#bbb', linewidth=0.8)
    ax_l.grid(axis='x', alpha=0.10, linestyle='--')
    ax_l.spines[['right', 'top']].set_visible(False)
    ax_l.tick_params(axis='x', labelsize=6)
    ax_l.text(-xl + 0.01, -0.35, '%TL', fontsize=6, color='#aaa',
              fontstyle='italic', va='bottom')

    ax_r.yaxis.set_visible(False)
    ax_r.set_title('Too much', fontsize=8, fontweight='bold', color='#c62828', pad=4)
    ax_r.axvline(0, color='#bbb', linewidth=0.8)
    ax_r.grid(axis='x', alpha=0.10, linestyle='--')
    ax_r.spines[['left', 'top']].set_visible(False)
    ax_r.tick_params(axis='x', labelsize=6)
    ax_r.xaxis.set_major_formatter(FuncFormatter(
        lambda x, _: f'{-x:.1f}' if x > 0.001 else '0.0'))
    ax_r.text(xl - 0.01, -0.35, '%TM', fontsize=6, color='#aaa',
              fontstyle='italic', va='bottom', ha='right')

    # Legend
    handles = [
        Patch(facecolor=bar_color, alpha=1.0, label='Sig. (p < .05)'),
        Patch(facecolor=bar_color, alpha=faint_alpha, label='Not sig.'),
        Line2D([0], [0], color=high_line_color, linestyle='--', linewidth=1.2,
               label=(
                   f'Red dashed = ±{HIGH_INFLUENCE_THRESHOLD:.2f}: '
                   f'{HIGH_INFLUENCE_LABEL}'
               )),
        Line2D([0], [0], color=medium_line_color, linestyle=':', linewidth=1.2,
               label=(
                   f'Purple dotted = ±{MEDIUM_INFLUENCE_THRESHOLD:.2f}: '
                   f'{MEDIUM_INFLUENCE_LABEL}'
               )),
    ]
    fig.legend(
        handles=handles,
        loc='lower left',
        bbox_to_anchor=(0.58, -0.01),
        fontsize=5.5,
        framealpha=0.9,
        edgecolor='#ddd',
    )

    for row_idx in medium_left_rows:
        ax_l.vlines(
            -MEDIUM_INFLUENCE_THRESHOLD,
            row_idx - (bar_h / 2),
            row_idx + (bar_h / 2),
            color=medium_line_color,
            linewidth=1.0,
            linestyle=':',
            zorder=4,
        )
    for row_idx in high_left_rows:
        ax_l.vlines(
            -HIGH_INFLUENCE_THRESHOLD,
            row_idx - (bar_h / 2),
            row_idx + (bar_h / 2),
            color=high_line_color,
            linewidth=1.0,
            linestyle='--',
            zorder=4,
        )
    for row_idx in medium_right_rows:
        ax_r.vlines(
            MEDIUM_INFLUENCE_THRESHOLD,
            row_idx - (bar_h / 2),
            row_idx + (bar_h / 2),
            color=medium_line_color,
            linewidth=1.0,
            linestyle=':',
            zorder=4,
        )
    for row_idx in high_right_rows:
        ax_r.vlines(
            HIGH_INFLUENCE_THRESHOLD,
            row_idx - (bar_h / 2),
            row_idx + (bar_h / 2),
            color=high_line_color,
            linewidth=1.0,
            linestyle='--',
            zorder=4,
        )

    fig.suptitle(filter_name, fontsize=10, fontweight='bold', color='#222', y=0.96)
    return fig


def _build_all_penalty_charts(result_df: pd.DataFrame) -> list[tuple[str, Figure]]:
    """Build separate butterfly charts — one Figure per filter group."""
    df = result_df.copy()
    filters = df['Filter Condition'].unique().tolist()

    # Shared x-axis limit across all filters
    mx = max(df['Weighted Penalty 1+2'].abs().max(),
             df['Weighted Penalty 4+5'].abs().max(), 0.25)
    xl = float(np.ceil(mx * 4) / 4) + 0.25

    charts = []
    for filt in filters:
        sub = df[df['Filter Condition'] == filt]
        attr_impact = {}
        for attr in sub['Attribute'].unique():
            attr_rows = sub[sub['Attribute'] == attr]
            attr_impact[attr] = (
                attr_rows['Weighted Penalty 1+2'].abs().sum()
                + attr_rows['Weighted Penalty 4+5'].abs().sum()
            )
        sorted_attrs = sorted(attr_impact, key=attr_impact.get, reverse=True)

        if not sorted_attrs:
            fig = Figure(figsize=(8, 3))
            fig.text(0.5, 0.5, 'No data to chart', ha='center', va='center', fontsize=12)
            charts.append((filt, fig))
            continue

        fig = _build_single_chart(sub, sorted_attrs, filt, xl)
        charts.append((filt, fig))
    return charts


class PenaltyChartDialog(QDialog):
    """Dialog showing butterfly Penalty charts — one per filter, scrollable."""

    def __init__(self, result_df: pd.DataFrame, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("กราฟ Penalty")
        self.charts = _build_all_penalty_charts(result_df)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)

        # Scrollable area for stacked charts
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        container = QWidget()
        v_layout = QVBoxLayout(container)
        v_layout.setSpacing(10)

        total_h = 0
        for _, fig in self.charts:
            canvas = FigureCanvas(fig)
            ch = int(fig.get_figheight() * 82)
            canvas.setFixedHeight(ch)
            v_layout.addWidget(canvas)
            total_h += ch + 10

        scroll.setWidget(container)
        layout.addWidget(scroll)

        fig_w = int(self.charts[0][1].get_figwidth() * 82) if self.charts else 700
        self.resize(min(fig_w + 40, 900), min(total_h + 80, 750))

        # Save button
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_save = QPushButton("บันทึกกราฟทั้งหมดเป็น PNG")
        btn_save.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_save.clicked.connect(self._save_png)
        btn_row.addWidget(btn_save)
        layout.addLayout(btn_row)

    def _save_png(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "เลือกโฟลเดอร์สำหรับบันทึก PNG")
        if not folder:
            return
        saved = []
        for name, fig in self.charts:
            safe = re.sub(r'[\\/*?:"<>|]', '_', name)
            path = f"{folder}/Penalty_{safe}.png"
            fig.savefig(path, dpi=150, bbox_inches='tight',
                        facecolor='white', edgecolor='none')
            saved.append(safe)
        QMessageBox.information(self, "สำเร็จ",
            f"บันทึกกราฟ {len(saved)} รูปแล้วที่:\n{folder}")


class PenaltyAnalyzerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Penalty Analysis Pro")
        self.setWindowIcon(QIcon(_resource_path("PE.ico")))
        self.resize(920, 680)
        self._center_on_screen()

        self.df = None
        self.meta = None
        self.result_df = None
        self.dep_var_label = ""
        self.all_variables = []
        self.jar_scale_map = {}
        self.jar_order_from_settings = []

        self._apply_global_style()
        self.setup_ui()

    def _center_on_screen(self):
        screen = QApplication.primaryScreen()
        if screen:
            screen_geo = screen.availableGeometry()
            x = (screen_geo.width() - self.width()) // 2 + screen_geo.x()
            y = (screen_geo.height() - self.height()) // 2 + screen_geo.y()
            self.move(x, y)

    # ──────────────────────────────────────────────
    #  Pastel Theme
    # ──────────────────────────────────────────────
    def _apply_global_style(self):
        self.setStyleSheet("""
            QMainWindow { background-color: #faf8ff; }
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #e8e0f0;
                border-radius: 10px;
                margin-top: 12px;
                padding: 14px 10px 10px 10px;
                font-size: 11px; font-weight: 600; color: #5b4a8a;
            }
            QGroupBox::title {
                subcontrol-origin: margin; subcontrol-position: top left;
                padding: 2px 10px;
                background-color: #f3eefa;
                border-radius: 6px;
                color: #6c5ba7;
            }
            QLabel { color: #6b6085; font-size: 11px; }
            QLineEdit {
                border: 1.5px solid #e0d6f0;
                border-radius: 7px;
                padding: 5px 8px; font-size: 11px;
                background-color: #fdfbff; color: #3d3556;
            }
            QLineEdit:focus { border: 2px solid #a78bfa; padding: 4px 7px; }
            QComboBox {
                border: 1.5px solid #e0d6f0;
                border-radius: 7px;
                padding: 5px 8px; font-size: 11px;
                background-color: #fdfbff; color: #3d3556; min-height: 18px;
            }
            QComboBox:focus, QComboBox:on { border: 2px solid #a78bfa; }
            QComboBox::drop-down { border: none; width: 24px; }
            QComboBox QAbstractItemView {
                border: 1px solid #e0d6f0; border-radius: 6px;
                background-color: #ffffff;
                selection-background-color: #ede9fe; selection-color: #5b21b6;
                padding: 3px;
            }
            QListWidget {
                border: 1.5px solid #e0d6f0;
                border-radius: 7px;
                background-color: #fdfbff; font-size: 11px;
                padding: 3px; outline: none;
            }
            QListWidget::item {
                padding: 4px 6px; border-radius: 5px; margin: 1px 0;
            }
            QListWidget::item:hover { background-color: #f5f0ff; }
            QListWidget::item:selected { background-color: #ede9fe; color: #5b21b6; }
            QPushButton {
                border: 1.5px solid #e0d6f0;
                border-radius: 7px;
                padding: 6px 12px; font-size: 11px; font-weight: 500;
                background-color: #ffffff; color: #5b4a8a;
            }
            QPushButton:hover { background-color: #f5f0ff; border-color: #c4b5fd; }
            QPushButton:pressed { background-color: #ede9fe; }
            QSplitter::handle {
                background-color: #e8e0f0; width: 2px; margin: 3px 5px; border-radius: 1px;
            }
            QStatusBar {
                background-color: #f3eefa;
                border-top: 1px solid #e8e0f0;
                color: #8b7fad; font-size: 10px; padding: 2px 8px;
            }
        """)

    def _make_btn(self, text, bg, hover):
        btn = QPushButton(text)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {bg}; color: #ffffff; border: none;
                border-radius: 7px; padding: 7px 14px;
                font-size: 11px; font-weight: 600;
            }}
            QPushButton:hover {{ background-color: {hover}; }}
            QPushButton:pressed {{ background-color: {hover}; }}
        """)
        return btn

    # ──────────────────────────────────────────────
    #  UI Setup
    # ──────────────────────────────────────────────
    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(14, 8, 14, 6)
        main_layout.setSpacing(5)

        # Header
        header = QLabel("Penalty Analysis Pro")
        header.setStyleSheet("font-size: 17px; font-weight: 700; color: #5b21b6; padding: 0 0 1px 0;")
        subtitle = QLabel("วิเคราะห์แบบแยกตาม Filter จากไฟล์ SPSS (.sav)")
        subtitle.setStyleSheet("font-size: 10px; color: #a89ec4; padding: 0 0 2px 0;")
        main_layout.addWidget(header)
        main_layout.addWidget(subtitle)

        # ── Step 1: Load File ──
        group_file = QGroupBox("ขั้นตอนที่ 1: โหลดข้อมูล")
        layout_file = QHBoxLayout()
        layout_file.setContentsMargins(8, 4, 8, 4)
        self.lbl_file = QLabel("ยังไม่ได้เลือกไฟล์")
        self.lbl_file.setStyleSheet("color: #f472b6; font-weight: 600; font-size: 11px;")
        btn_browse = self._make_btn("เลือกไฟล์...", "#a78bfa", "#8b5cf6")
        btn_browse.clicked.connect(self.load_file)
        layout_file.addWidget(self.lbl_file)
        layout_file.addStretch()
        layout_file.addWidget(btn_browse)
        group_file.setLayout(layout_file)
        main_layout.addWidget(group_file)

        # ── Splitter ──
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # ── Left Panel ──
        left_widget = QWidget()
        layout_left = QVBoxLayout(left_widget)
        layout_left.setContentsMargins(0, 0, 3, 0)
        layout_left.setSpacing(5)

        # Step 2: Filters
        group_filter = QGroupBox("ขั้นตอนที่ 2: Filters")
        layout_filter = QVBoxLayout()
        layout_filter.setSpacing(4)

        row_label = QHBoxLayout()
        lbl_label = QLabel("Label:")
        lbl_label.setStyleSheet("font-weight: 600; min-width: 48px;")
        self.txt_filter_label = QLineEdit()
        self.txt_filter_label.setPlaceholderText("เช่น รวมทั้งหมด, Product I, Product J")
        row_label.addWidget(lbl_label)
        row_label.addWidget(self.txt_filter_label)
        layout_filter.addLayout(row_label)

        row_filter = QHBoxLayout()
        lbl_cond = QLabel("Query:")
        lbl_cond.setStyleSheet("font-weight: 600; min-width: 48px;")
        self.txt_filter_query = QLineEdit()
        self.txt_filter_query.setPlaceholderText("เช่น IndexPB==1 หรือ IndexPB==1 & Cell==2")
        btn_add_filter = self._make_btn("เพิ่ม", "#a78bfa", "#8b5cf6")
        btn_add_filter.clicked.connect(self.add_filter)
        row_filter.addWidget(lbl_cond)
        row_filter.addWidget(self.txt_filter_query, stretch=4)
        row_filter.addWidget(btn_add_filter, stretch=1)
        layout_filter.addLayout(row_filter)

        help_text = QLabel("แต่ละบรรทัดจะรันแยกกัน เช่น 3 บรรทัด = 3 กลุ่ม และถ้าใส่เฉพาะ Label จะถือว่าเป็นรวมทั้งหมด")
        help_text.setStyleSheet("color: #a78bfa; font-style: italic; font-size: 10px; padding-left: 2px;")
        layout_filter.addWidget(help_text)

        self.list_filters = QListWidget()
        self.list_filters.setFixedHeight(58)
        layout_filter.addWidget(self.list_filters)

        btn_remove_filter = QPushButton("ลบรายการที่เลือก")
        btn_remove_filter.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_remove_filter.setStyleSheet("""
            QPushButton {
                color: #f472b6; border: 1px solid #fbcfe8; border-radius: 6px;
                padding: 4px 10px; font-size: 10px; font-weight: 500; background: #fff;
            }
            QPushButton:hover { background-color: #fdf2f8; border-color: #f472b6; }
        """)
        btn_remove_filter.clicked.connect(self.remove_filter)
        layout_filter.addWidget(btn_remove_filter)

        group_filter.setLayout(layout_filter)
        layout_left.addWidget(group_filter)

        # Step 3: Dependent Variable
        group_overall = QGroupBox("ขั้นตอนที่ 3: Dependent Variable")
        layout_overall = QVBoxLayout()
        layout_overall.setSpacing(4)
        self.search_overall = QLineEdit()
        self.search_overall.setPlaceholderText("ค้นหาตัวแปร...")
        self.search_overall.textChanged.connect(self.filter_overall)
        layout_overall.addWidget(self.search_overall)
        self.combo_overall = QComboBox()
        layout_overall.addWidget(self.combo_overall)
        group_overall.setLayout(layout_overall)
        layout_left.addWidget(group_overall)

        # Settings
        group_settings = QGroupBox("จัดการ Settings")
        layout_settings = QHBoxLayout()
        layout_settings.setSpacing(6)
        btn_save_settings = QPushButton("บันทึก")
        btn_save_settings.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_save_settings.clicked.connect(self.save_settings)
        btn_load_settings = QPushButton("เปิด")
        btn_load_settings.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_load_settings.clicked.connect(self.load_settings)
        layout_settings.addWidget(btn_save_settings)
        layout_settings.addWidget(btn_load_settings)
        group_settings.setLayout(layout_settings)
        layout_left.addWidget(group_settings)

        layout_left.addStretch()
        splitter.addWidget(left_widget)

        # ── Right Panel ──
        right_widget = QWidget()
        layout_right = QVBoxLayout(right_widget)
        layout_right.setContentsMargins(3, 0, 0, 0)

        group_jar = QGroupBox("ขั้นตอนที่ 4: JAR Attributes")
        layout_jar = QVBoxLayout()
        layout_jar.setSpacing(4)
        self.search_jar = QLineEdit()
        self.search_jar.setPlaceholderText("ค้นหา JAR variables...")
        self.search_jar.textChanged.connect(self.filter_jar)
        layout_jar.addWidget(self.search_jar)
        hint_label = QLabel("ใช้ Ctrl / Shift + Click เพื่อเลือกหลายตัวแปร")
        hint_label.setStyleSheet("color: #a89ec4; font-size: 10px;")
        layout_jar.addWidget(hint_label)
        self.list_jar = QListWidget()
        self.list_jar.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        layout_jar.addWidget(self.list_jar)
        group_jar.setLayout(layout_jar)
        layout_right.addWidget(group_jar)

        splitter.addWidget(right_widget)
        splitter.setSizes([380, 500])
        main_layout.addWidget(splitter, stretch=1)

        # ── Bottom Action Bar ──
        group_action = QGroupBox("ขั้นตอนที่ 5: Run และ Export")
        layout_action = QHBoxLayout()
        layout_action.setSpacing(10)
        layout_action.setContentsMargins(8, 4, 8, 4)
        btn_run = self._make_btn("รันการวิเคราะห์", "#86efac", "#4ade80")
        btn_run.setStyleSheet(btn_run.styleSheet().replace("color: #ffffff", "color: #166534"))
        btn_run.clicked.connect(self.run_analysis)
        btn_chart = self._make_btn("ดูกราฟ Penalty", "#fbbf24", "#f59e0b")
        btn_chart.setStyleSheet(btn_chart.styleSheet().replace("color: #ffffff", "color: #78350f"))
        btn_chart.clicked.connect(self.show_penalty_chart)
        btn_export = self._make_btn("Export เป็น Excel", "#93c5fd", "#60a5fa")
        btn_export.setStyleSheet(btn_export.styleSheet().replace("color: #ffffff", "color: #1e3a5f"))
        btn_export.clicked.connect(self.export_excel)
        layout_action.addWidget(btn_run)
        layout_action.addWidget(btn_chart)
        layout_action.addWidget(btn_export)
        group_action.setLayout(layout_action)
        main_layout.addWidget(group_action)

        # Status Bar
        self.statusBar().showMessage("พร้อมใช้งาน: กรุณาโหลดไฟล์ SPSS เพื่อเริ่มต้น")

    # ──────────────────────────────────────────────
    #  Search / Filter helpers
    # ──────────────────────────────────────────────
    def filter_overall(self, text):
        current_selection = self.combo_overall.currentData()
        self.combo_overall.clear()
        search_text = text.lower()
        for display_text, col in self.all_variables:
            if search_text in display_text.lower():
                self.combo_overall.addItem(display_text, userData=col)
        idx = self.combo_overall.findData(current_selection)
        if idx != -1:
            self.combo_overall.setCurrentIndex(idx)

    def filter_jar(self, text):
        search_text = text.lower()
        for i in range(self.list_jar.count()):
            item = self.list_jar.item(i)
            item.setHidden(search_text not in item.text().lower())

    # ──────────────────────────────────────────────
    #  Load SPSS
    # ──────────────────────────────────────────────
    def load_file(self):
        filepath, _ = QFileDialog.getOpenFileName(self, "เปิดไฟล์ SPSS", "", "SPSS Files (*.sav)")
        if not filepath:
            return
        try:
            self.df, self.meta = pyreadstat.read_sav(filepath)
            filename = filepath.split('/')[-1]
            self.lbl_file.setText(f"โหลดแล้ว: {filename}")
            self.lbl_file.setStyleSheet("color: #34d399; font-weight: 600; font-size: 11px;")
            self.statusBar().showMessage(f"โหลดไฟล์สำเร็จ: พบตัวแปร {len(self.df.columns)} ตัว")

            self.combo_overall.clear()
            self.list_jar.clear()
            self.search_overall.clear()
            self.search_jar.clear()
            self.list_filters.clear()
            self.all_variables.clear()
            self.jar_scale_map.clear()
            self.jar_order_from_settings = []

            for col in self.df.columns:
                label = self.meta.column_names_to_labels.get(col, col)
                display_text = f"[{col}] {label}"
                self.all_variables.append((display_text, col))
                self.combo_overall.addItem(display_text, userData=col)
                item = QListWidgetItem(display_text)
                item.setData(Qt.ItemDataRole.UserRole, col)
                self.list_jar.addItem(item)

            QMessageBox.information(self, "สำเร็จ", f"โหลดไฟล์ SPSS สำเร็จ\nจำนวนตัวแปร: {len(self.df.columns)}")
        except Exception as e:
            QMessageBox.critical(self, "ผิดพลาด", f"โหลดไฟล์ไม่สำเร็จ:\n{str(e)}")

    def _build_jar_scale_map(self, jar_vars: list[str]) -> dict[str, int]:
        jar_scale_map: dict[str, int] = {}
        variable_value_labels = getattr(self.meta, 'variable_value_labels', {}) or {}
        value_labels = getattr(self.meta, 'value_labels', {}) or {}

        for jar_var in jar_vars:
            resolved_scale = None
            label_set_name = variable_value_labels.get(jar_var)
            labeled_values = None
            if isinstance(label_set_name, dict):
                labeled_values = label_set_name
            elif label_set_name and label_set_name in value_labels:
                labeled_values = value_labels[label_set_name]

            if labeled_values:
                numeric_codes = {
                    int(float(code))
                    for code in labeled_values.keys()
                    if pd.notna(code)
                }
                if 7 in numeric_codes:
                    resolved_scale = 7
                elif 5 in numeric_codes:
                    resolved_scale = 5

            if resolved_scale is None and self.df is not None and jar_var in self.df.columns:
                numeric_values = pd.to_numeric(self.df[jar_var], errors='coerce').dropna()
                if not numeric_values.empty and numeric_values.max() >= 7:
                    resolved_scale = 7

            jar_scale_map[jar_var] = resolved_scale or 5

        return jar_scale_map

    # ──────────────────────────────────────────────
    #  Filter management
    # ──────────────────────────────────────────────
    def add_filter(self):
        query_str = self.txt_filter_query.text().strip()
        label_str = self.txt_filter_label.text().strip()

        if not query_str and not label_str:
            QMessageBox.warning(self, "คำเตือน", "กรุณาใส่ Label หรือ Query อย่างน้อย 1 อย่าง")
            return

        if not query_str:
            query_str = ""
        if not label_str:
            label_str = query_str if query_str else NO_FILTER_LABEL

        display = f"{label_str}  [{query_str}]" if query_str else f"{label_str}  [ไม่กรองข้อมูล]"
        item = QListWidgetItem(display)
        item.setData(Qt.ItemDataRole.UserRole, query_str)
        item.setData(Qt.ItemDataRole.UserRole + 1, label_str)
        self.list_filters.addItem(item)
        self.txt_filter_query.clear()
        self.txt_filter_label.clear()

    def remove_filter(self):
        for item in self.list_filters.selectedItems():
            self.list_filters.takeItem(self.list_filters.row(item))

    # ──────────────────────────────────────────────
    #  Save / Load Settings (incl. filters)
    # ──────────────────────────────────────────────
    def save_settings(self):
        if self.combo_overall.count() == 0:
            QMessageBox.warning(self, "คำเตือน", "กรุณาโหลดไฟล์ข้อมูลก่อน")
            return

        dep_var = self.combo_overall.currentData()
        jar_vars = self._get_selected_jar_vars()

        filter_labels = []
        filter_queries = []
        for i in range(self.list_filters.count()):
            filter_queries.append(self.list_filters.item(i).data(Qt.ItemDataRole.UserRole) or "")
            filter_labels.append(self.list_filters.item(i).data(Qt.ItemDataRole.UserRole + 1) or "")

        filepath, _ = QFileDialog.getSaveFileName(self, "บันทึก Settings", "Penalty_Settings.xlsx", "Excel Files (*.xlsx)")
        if filepath:
            try:
                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    roles = ['Overall Liking'] + ['JAR Attribute'] * len(jar_vars)
                    variables = [dep_var] + jar_vars
                    config_df = pd.DataFrame({'Role': roles, 'Variable_Name': variables})
                    config_df.to_excel(writer, sheet_name='Variables', index=False)

                    if filter_labels:
                        filter_df = pd.DataFrame({'Label': filter_labels, 'Query': filter_queries})
                        filter_df.to_excel(writer, sheet_name='Filters', index=False)

                self.statusBar().showMessage(f"บันทึก Settings แล้ว: {filepath}")
                QMessageBox.information(self, "สำเร็จ", "บันทึก Settings แล้ว\n(Variables + Filters)")
            except Exception as e:
                QMessageBox.critical(self, "ผิดพลาด", f"บันทึก Settings ไม่สำเร็จ:\n{str(e)}")

    def load_settings(self):
        if len(self.all_variables) == 0:
            QMessageBox.warning(self, "คำเตือน", "กรุณาโหลดไฟล์ข้อมูลก่อน")
            return

        filepath, _ = QFileDialog.getOpenFileName(self, "เปิด Settings", "", "Excel Files (*.xlsx)")
        if filepath:
            try:
                xls = pd.ExcelFile(filepath)

                if 'Variables' in xls.sheet_names:
                    config_df = pd.read_excel(xls, sheet_name='Variables')
                else:
                    config_df = pd.read_excel(xls, sheet_name=0)

                overall_row = config_df[config_df['Role'] == 'Overall Liking']
                if not overall_row.empty:
                    dep_var = overall_row.iloc[0]['Variable_Name']
                    for i in range(self.combo_overall.count()):
                        if self.combo_overall.itemData(i) == dep_var:
                            self.combo_overall.setCurrentIndex(i)
                            break

                jar_vars = config_df[config_df['Role'] == 'JAR Attribute']['Variable_Name'].tolist()
                self.jar_order_from_settings = []
                seen_vars = set()
                available_vars = {
                    self.list_jar.item(i).data(Qt.ItemDataRole.UserRole)
                    for i in range(self.list_jar.count())
                }
                for var in jar_vars:
                    if var in available_vars and var not in seen_vars:
                        self.jar_order_from_settings.append(var)
                        seen_vars.add(var)
                self.list_jar.clearSelection()
                for i in range(self.list_jar.count()):
                    item = self.list_jar.item(i)
                    if item.data(Qt.ItemDataRole.UserRole) in jar_vars:
                        item.setSelected(True)

                if 'Filters' in xls.sheet_names:
                    filter_df = pd.read_excel(xls, sheet_name='Filters')
                    self.list_filters.clear()
                    for _, row in filter_df.iterrows():
                        label_str = str(row.get('Label', '')).strip()
                        query_str = str(row.get('Query', '')).strip()
                        if query_str == 'nan':
                            query_str = ""
                        if label_str == 'nan':
                            label_str = ""
                        if not label_str and not query_str:
                            continue
                        if not label_str:
                            label_str = query_str if query_str else NO_FILTER_LABEL

                        display = f"{label_str}  [{query_str}]" if query_str else f"{label_str}  [ไม่กรองข้อมูล]"
                        item = QListWidgetItem(display)
                        item.setData(Qt.ItemDataRole.UserRole, query_str)
                        item.setData(Qt.ItemDataRole.UserRole + 1, label_str)
                        self.list_filters.addItem(item)

                self.statusBar().showMessage(f"โหลด Settings แล้ว: {filepath}")
                QMessageBox.information(self, "สำเร็จ", "โหลด Settings แล้ว\n(Variables + Filters)")
            except Exception as e:
                QMessageBox.critical(self, "ผิดพลาด", f"โหลด Settings ไม่สำเร็จ:\n{str(e)}")

    # ──────────────────────────────────────────────
    #  Helpers
    # ──────────────────────────────────────────────
    def show_penalty_chart(self) -> None:
        if self.result_df is None or self.result_df.empty:
            QMessageBox.warning(self, "คำเตือน", "ยังไม่มีผลลัพธ์ กรุณารันการวิเคราะห์ก่อน")
            return
        dialog = PenaltyChartDialog(self.result_df, parent=self)
        dialog.exec()

    def _get_selected_jar_vars(self) -> list[str]:
        selected_vars = []
        for i in range(self.list_jar.count()):
            item = self.list_jar.item(i)
            if item.isSelected():
                selected_vars.append(item.data(Qt.ItemDataRole.UserRole))

        if self.jar_order_from_settings:
            selected_set = set(selected_vars)
            ordered_from_settings = [var for var in self.jar_order_from_settings if var in selected_set]
            ordered_set = set(ordered_from_settings)
            remaining = [var for var in selected_vars if var not in ordered_set]
            return ordered_from_settings + remaining

        return selected_vars

    def _get_filter_specs(self) -> list[FilterSpec]:
        if self.list_filters.count() == 0:
            return [FilterSpec(label=NO_FILTER_LABEL, query="")]

        filters = []
        for i in range(self.list_filters.count()):
            raw_query = self.list_filters.item(i).data(Qt.ItemDataRole.UserRole) or ""
            label = self.list_filters.item(i).data(Qt.ItemDataRole.UserRole + 1) or raw_query or NO_FILTER_LABEL
            filters.append(FilterSpec(label=label, query=raw_query))
        return filters

    # ──────────────────────────────────────────────
    #  Run Analysis
    # ──────────────────────────────────────────────
    def run_analysis(self):
        if self.df is None:
            QMessageBox.warning(self, "คำเตือน", "กรุณาโหลดไฟล์ข้อมูลก่อน")
            return

        dep_var = self.combo_overall.currentData()
        self.dep_var_label = self.combo_overall.currentText()
        jar_vars = self._get_selected_jar_vars()
        self.jar_scale_map = self._build_jar_scale_map(jar_vars)

        if not dep_var or not jar_vars:
            QMessageBox.warning(self, "คำเตือน", "กรุณาเลือก Dependent Variable และ JAR Attribute อย่างน้อย 1 ตัว")
            return

        filter_specs = self._get_filter_specs()

        try:
            self.result_df = analyze_penalty(
                df=self.df,
                dep_var=dep_var,
                jar_vars=jar_vars,
                filters=filter_specs,
                label_map=self.meta.column_names_to_labels,
                jar_scale_map=self.jar_scale_map,
            )
        except FilterQueryError as e:
            QMessageBox.warning(self, "Filter ไม่ถูกต้อง", f"เงื่อนไข Filter ไม่ถูกต้อง\n\n{str(e)}")
            return

        self.statusBar().showMessage(
            f"วิเคราะห์เสร็จแล้ว: {len(filter_specs)} กลุ่ม, {len(self.result_df)} แถว"
        )
        QMessageBox.information(
            self,
            "สำเร็จ",
            f"วิเคราะห์ Penalty เสร็จแล้ว\nจำนวนกลุ่ม: {len(filter_specs)}\nพร้อม Export ต่อได้เลย",
        )

    # ──────────────────────────────────────────────
    #  Export Excel
    # ──────────────────────────────────────────────
    def export_excel(self):
        if self.result_df is None or self.result_df.empty:
            QMessageBox.warning(self, "คำเตือน", "ยังไม่มีผลลัพธ์ กรุณารันการวิเคราะห์ก่อน")
            return

        filepath, _ = QFileDialog.getSaveFileName(self, "บันทึก Excel", "Penalty_Analysis_Result.xlsx", "Excel Files (*.xlsx)")

        if filepath:
            try:
                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    export_columns = [
                        'Filter Condition', 'Filter Query', 'Attribute',
                        'Base N',
                        'Base (%) 1+2', 'Base (%) JAR', 'Base (%) 4+5',
                        'Mean 1+2', 'Mean JAR', 'Mean 4+5',
                        'Weighted Penalty 1+2', 'Weighted Penalty 4+5',
                        'p-value 1+2', 'Sig 1+2', 'p-value 4+5', 'Sig 4+5',
                    ]
                    export_df = self.result_df[export_columns].copy()
                    export_df = export_df.rename(columns={'Base N': 'N'})
                    export_df.to_excel(writer, sheet_name='Penalty Report', index=False, startrow=3)

                    workbook = writer.book
                    worksheet = writer.sheets['Penalty Report']
                    ws_index = workbook.create_sheet('Index', 0)

                    filter_first_rows: dict[str, int] = {}
                    for row_num, filter_name in enumerate(export_df['Filter Condition'].tolist(), start=5):
                        filter_key = str(filter_name)
                        if filter_key not in filter_first_rows:
                            filter_first_rows[filter_key] = row_num

                    unique_filters = (
                        export_df[['Filter Condition', 'Filter Query']]
                        .drop_duplicates(subset=['Filter Condition'], keep='first')
                        .reset_index(drop=True)
                    )
                    index_row_by_filter: dict[str, int] = {}

                    bold_font = Font(bold=True)
                    header_font = Font(bold=True, color="000000")
                    red_bold_font = Font(bold=True, color="C00000")
                    dark_red_font = Font(bold=True, color="9C0006")
                    no_data_font = Font(color="7F6000")
                    header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    no_data_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                    base_mean_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

                    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    center_nowrap = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    left_nowrap = Alignment(horizontal='left', vertical='center', wrap_text=False)

                    thin_border = Border(
                        left=Side(style='thin', color='B0B0B0'),
                        right=Side(style='thin', color='B0B0B0'),
                        top=Side(style='thin', color='B0B0B0'),
                        bottom=Side(style='thin', color='B0B0B0')
                    )

                    ws_index.cell(row=1, column=1, value="Filter Index").font = Font(bold=True, size=13)
                    ws_index.cell(row=2, column=1, value="No")
                    ws_index.cell(row=2, column=2, value="Filter Condition")
                    ws_index.cell(row=2, column=3, value="Go to Penalty")
                    ws_index.cell(row=2, column=4, value="Go to Chart")
                    ws_index.cell(row=2, column=5, value="Filter Query")
                    ws_index.column_dimensions['A'].width = 6
                    ws_index.column_dimensions['B'].width = 24
                    ws_index.column_dimensions['C'].width = 12
                    ws_index.column_dimensions['D'].width = 14
                    ws_index.column_dimensions['E'].width = 44
                    ws_index.freeze_panes = 'A3'

                    for col_idx in range(1, 6):
                        idx_hdr = ws_index.cell(row=2, column=col_idx)
                        idx_hdr.font = header_font
                        idx_hdr.fill = header_fill
                        idx_hdr.alignment = center_wrap
                        idx_hdr.border = thin_border

                    for offset, (filter_name, filter_query) in enumerate(unique_filters.itertuples(index=False, name=None), start=3):
                        filter_name = str(filter_name)
                        filter_query = "" if pd.isna(filter_query) else str(filter_query)
                        target_row = filter_first_rows.get(filter_name, 5)
                        link_target = f"#'Penalty Report'!A{target_row}"
                        index_row_by_filter[filter_name] = offset

                        ws_index.cell(row=offset, column=1, value=offset - 2)
                        ws_index.cell(row=offset, column=2, value=filter_name)
                        go_cell = ws_index.cell(row=offset, column=3, value="Open")
                        chart_cell = ws_index.cell(row=offset, column=4, value="Open_Chart")
                        ws_index.cell(row=offset, column=5, value=filter_query)

                        go_cell.hyperlink = link_target
                        go_cell.style = "Hyperlink"
                        chart_cell.alignment = center_nowrap

                        for col_idx in range(1, 6):
                            idx_cell = ws_index.cell(row=offset, column=col_idx)
                            idx_cell.border = thin_border
                            idx_cell.alignment = center_nowrap if col_idx in (1, 3, 4) else left_nowrap

                    worksheet.cell(row=1, column=1, value="Penalty Analysis Report").font = Font(bold=True, size=14)
                    worksheet.cell(row=2, column=1, value="Dependent Variable:").font = bold_font
                    worksheet.cell(row=2, column=2, value=self.dep_var_label).font = Font(bold=True, color="1F497D")

                    detected_scales = set(self.jar_scale_map.values()) if getattr(self, "jar_scale_map", None) else {5}
                    if detected_scales == {7}:
                        low_group_label = "1-3"
                        high_group_label = "5-7"
                    elif detected_scales == {5}:
                        low_group_label = "1-2"
                        high_group_label = "4-5"
                    else:
                        low_group_label = "Low"
                        high_group_label = "High"

                    # A: Filter Condition, B: Filter Query, C: Attribute
                    # D: N (base count of filter), E-G: Base, H-J: Mean, K-L: Weighted Penalty
                    # M: p-value ฝั่ง low, N: Sig ฝั่ง low, O: p-value ฝั่ง high, P: Sig ฝั่ง high
                    total_cols = len(export_df.columns)
                    column_widths = {
                        'A': 14, 'B': 16, 'C': 34, 'D': 6,
                        'E': 9, 'F': 9, 'G': 9,
                        'H': 9, 'I': 9, 'J': 9,
                        'K': 12, 'L': 12,
                        'M': 10, 'N': 7, 'O': 10, 'P': 7,
                    }
                    for col_letter, width in column_widths.items():
                        worksheet.column_dimensions[col_letter].width = width

                    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                    n_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

                    # Header row
                    for col_idx in range(1, total_cols + 1):
                        cell = worksheet.cell(row=4, column=col_idx)
                        cell.font = header_font
                        if col_idx in [11, 12]:
                            cell.fill = yellow_fill
                        elif col_idx in [13, 14, 15, 16]:
                            cell.fill = green_fill
                        elif col_idx == 4:
                            cell.fill = n_fill
                        else:
                            cell.fill = header_fill
                        cell.alignment = center_wrap
                        cell.border = thin_border
                        worksheet.row_dimensions[4].height = 40

                    header_overrides = {
                        4: "N",
                        5: f"Base (%)\n{low_group_label}",
                        7: f"Base (%)\n{high_group_label}",
                        8: f"Mean\n{low_group_label}",
                        10: f"Mean\n{high_group_label}",
                        11: f"Weighted\nPenalty\n{low_group_label}",
                        12: f"Weighted\nPenalty\n{high_group_label}",
                        13: f"p-value\n{low_group_label}",
                        14: f"Sig\n{low_group_label}",
                        15: f"p-value\n{high_group_label}",
                        16: f"Sig\n{high_group_label}",
                    }
                    for col_idx, text in header_overrides.items():
                        worksheet.cell(row=4, column=col_idx, value=text)

                    # Freeze pane ที่ C5 → ตรึง header (row 1-4) + คอลัมน์ A,B
                    worksheet.freeze_panes = 'C5'

                    last_data_row = worksheet.max_row

                    # Data rows — format + thin border + ความสูงคงที่
                    for row in range(5, last_data_row + 1):
                        worksheet.row_dimensions[row].height = 15
                        base_12_value = worksheet.cell(row=row, column=5).value or 0
                        base_jar_value = worksheet.cell(row=row, column=6).value or 0
                        base_45_value = worksheet.cell(row=row, column=7).value or 0
                        for col_idx in range(1, total_cols + 1):
                            cell = worksheet.cell(row=row, column=col_idx)
                            cell.border = thin_border

                            if col_idx == 1:
                                cell.alignment = center_nowrap
                            elif col_idx in [2, 3]:
                                cell.alignment = left_nowrap
                            else:
                                cell.alignment = center_nowrap
                                if cell.value is not None:
                                    if col_idx in [4, 5, 6, 7]:
                                        cell.number_format = '0'
                                    elif col_idx in [8, 9, 10]:
                                        cell.number_format = '0.00'
                                    elif col_idx in [11, 12]:
                                        cell.number_format = '0.00'
                                    elif col_idx in [13, 15]:
                                        cell.number_format = '0.0000'

                            # ไฮไลท์เหลืองคอลัมน์ Weighted Penalty
                            if col_idx in [11, 12]:
                                cell.fill = yellow_fill
                                value = cell.value
                                if isinstance(value, (int, float, np.integer, np.floating)):
                                    abs_value = abs(float(value))
                                    if abs_value >= HIGH_INFLUENCE_THRESHOLD:
                                        cell.font = dark_red_font
                                    elif abs_value >= MEDIUM_INFLUENCE_THRESHOLD:
                                        cell.font = red_bold_font
                            # ไฮไลท์เขียวอ่อนคอลัมน์ Significance
                            elif col_idx in [13, 14, 15, 16]:
                                cell.fill = green_fill
                            elif col_idx == 4:
                                cell.fill = n_fill
                            elif col_idx in [5, 6, 7, 8, 9, 10]:
                                cell.fill = base_mean_fill

                            no_data_cols = set()
                            if base_12_value == 0:
                                no_data_cols.update([5, 8])
                            if base_jar_value == 0:
                                no_data_cols.update([6, 9])
                            if base_45_value == 0:
                                no_data_cols.update([7, 10])

                            if col_idx in no_data_cols:
                                cell.fill = no_data_fill
                                cell.font = no_data_font

                    # เส้นหนาแบ่งกลุ่ม Filter
                    for row in range(5, last_data_row + 1):
                        cell_val = worksheet.cell(row=row, column=1).value
                        next_val = worksheet.cell(row=row + 1, column=1).value if row < last_data_row else None
                        if cell_val != next_val:
                            for col_idx in range(1, total_cols + 1):
                                c = worksheet.cell(row=row, column=col_idx)
                                # รวม border เดิมกับเส้นหนาด้านล่าง
                                c.border = Border(
                                    left=c.border.left,
                                    right=c.border.right,
                                    top=c.border.top,
                                    bottom=Side(style='medium', color='000000')
                                )

                    note_start_row = last_data_row + 3
                    note_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                    note_title_fill = PatternFill(start_color='F4B183', end_color='F4B183', fill_type='solid')

                    worksheet.merge_cells(start_row=note_start_row, start_column=1, end_row=note_start_row, end_column=10)
                    note_title = worksheet.cell(row=note_start_row, column=1, value="วิธีอ่านคอลัมน์ M-P")
                    note_title.font = Font(bold=True, color="7F6000", size=11)
                    note_title.fill = note_title_fill
                    note_title.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    for col_idx in range(1, 11):
                        worksheet.cell(row=note_start_row, column=col_idx).border = thin_border
                        worksheet.cell(row=note_start_row, column=col_idx).fill = note_title_fill
                    worksheet.row_dimensions[note_start_row].height = 22

                    note_blocks = [
                        (
                            f"M (p-value {low_group_label})",
                            f"ใช้ดูว่าคะแนนของกลุ่ม {low_group_label} ต่างจากกลุ่ม JAR หรือไม่",
                            "ตัวอย่าง: ถ้า M = 0.0120 แปลว่าโอกาสที่ความต่างนี้เกิดจากความบังเอิญมีประมาณ 1.2%",
                        ),
                        (
                            f"N (Sig {low_group_label})",
                            "เป็นสรุประดับนัยสำคัญจากค่า p-value ของคอลัมน์ M",
                            "ตัวอย่าง: ถ้า N = ** แปลว่า p < .01, ถ้า N = * แปลว่า p < .05, ถ้า N = ns แปลว่าไม่แตกต่างอย่างมีนัยสำคัญ",
                        ),
                        (
                            f"O (p-value {high_group_label})",
                            f"ใช้ดูว่าคะแนนของกลุ่ม {high_group_label} ต่างจากกลุ่ม JAR หรือไม่",
                            "ตัวอย่าง: ถ้า O = 0.2840 แปลว่ายังไม่พบความแตกต่างอย่างมีนัยสำคัญ เพราะค่าสูงกว่า .05",
                        ),
                        (
                            f"P (Sig {high_group_label})",
                            "เป็นสรุประดับนัยสำคัญจากค่า p-value ของคอลัมน์ O",
                            f"ตัวอย่าง: ถ้า P = ns แปลว่ากลุ่ม {high_group_label} ยังไม่ต่างจาก JAR อย่างมีนัยสำคัญ, ถ้า P = * หรือ ** แปลว่าต่าง",
                        ),
                        (
                            "สรุปวิธีอ่าน",
                            "ให้ดู p-value ควบคู่กับ Sig เสมอ เพื่อแปลผลว่าแต่ละกลุ่มต่างจาก JAR จริงหรือไม่",
                            "ตัวอย่าง: ถ้า p-value = 0.0030 และ Sig = ** แปลว่าความแตกต่างมีนัยสำคัญชัดเจน",
                        ),
                    ]
                    for offset, (heading, explanation, example) in enumerate(note_blocks, start=1):
                        row_idx = note_start_row + offset
                        worksheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=10)
                        note_cell = worksheet.cell(
                            row=row_idx,
                            column=1,
                            value=f"{heading}: {explanation}\n{example}",
                        )
                        note_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        note_cell.fill = note_fill
                        note_cell.border = thin_border
                        for col_idx in range(1, 11):
                            worksheet.cell(row=row_idx, column=col_idx).border = thin_border
                            worksheet.cell(row=row_idx, column=col_idx).fill = note_fill
                        worksheet.row_dimensions[row_idx].height = 40

                    # ── Embed Penalty Charts (one per filter, stacked) ──
                    charts = _build_all_penalty_charts(self.result_df)
                    ws_chart = workbook.create_sheet('Chart')
                    ws_chart.column_dimensions['A'].width = 2
                    ws_chart.cell(
                        row=1,
                        column=2,
                        value=(
                            f"Penalty value at ±{HIGH_INFLUENCE_THRESHOLD:.2f}: "
                            f"{HIGH_INFLUENCE_LABEL}"
                        ),
                    )
                    ws_chart.cell(
                        row=2,
                        column=2,
                        value=(
                            f"Penalty value at ±{MEDIUM_INFLUENCE_THRESHOLD:.2f}: "
                            f"{MEDIUM_INFLUENCE_LABEL}"
                        ),
                    )
                    chart_anchor_rows: dict[str, int] = {}
                    anchor_row = 4
                    for chart_name, chart_fig in charts:
                        chart_anchor_rows[str(chart_name)] = anchor_row
                        buf = BytesIO()
                        chart_fig.savefig(buf, format='png', dpi=150,
                                         bbox_inches='tight', facecolor='white')
                        buf.seek(0)
                        img = XlImage(buf)
                        img.anchor = f'A{anchor_row}'
                        ws_chart.add_image(img)
                        pixel_height = int(img.height or 0)
                        rows_needed = max(int(np.ceil(pixel_height / EXCEL_DEFAULT_ROW_PIXELS)), 1)
                        for row_idx in range(anchor_row, anchor_row + rows_needed):
                            ws_chart.row_dimensions[row_idx].height = (
                                EXCEL_DEFAULT_ROW_PIXELS * EXCEL_POINTS_PER_PIXEL
                            )
                        ws_chart.cell(row=anchor_row, column=2, value=chart_name)
                        anchor_row += rows_needed + EXCEL_IMAGE_SPACER_ROWS

                    for filter_name, index_row in index_row_by_filter.items():
                        chart_row = chart_anchor_rows.get(filter_name)
                        if chart_row is None:
                            continue
                        chart_link_cell = ws_index.cell(row=index_row, column=4, value="Open_Chart")
                        chart_link_cell.hyperlink = f"#'Chart'!A{chart_row}"
                        chart_link_cell.style = "Hyperlink"
                        chart_link_cell.alignment = center_nowrap

                    summary_df = build_summary_output(self.result_df)
                    summary_df.to_excel(writer, sheet_name='Summary Output', index=False)
                    ws_summary = writer.sheets['Summary Output']
                    ws_summary.cell(row=1, column=4, value=f"ฝั่ง {low_group_label}")
                    ws_summary.cell(row=1, column=5, value=f"ฝั่ง {high_group_label}")
                    ws_summary.freeze_panes = 'A2'
                    summary_widths = {
                        'A': 16, 'B': 34, 'C': 28,
                        'D': 38, 'E': 38, 'F': 38,
                    }
                    for col_letter, width in summary_widths.items():
                        ws_summary.column_dimensions[col_letter].width = width

                    summary_header_fill = PatternFill(
                        start_color='D9E2F3',
                        end_color='D9E2F3',
                        fill_type='solid',
                    )
                    summary_wrap_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    summary_both_issues_fill = PatternFill(
                        start_color=SUMMARY_BOTH_ISSUES_FILL,
                        end_color=SUMMARY_BOTH_ISSUES_FILL,
                        fill_type='solid',
                    )
                    summary_too_little_fill = PatternFill(
                        start_color=SUMMARY_TOO_LITTLE_FILL,
                        end_color=SUMMARY_TOO_LITTLE_FILL,
                        fill_type='solid',
                    )
                    summary_too_much_fill = PatternFill(
                        start_color=SUMMARY_TOO_MUCH_FILL,
                        end_color=SUMMARY_TOO_MUCH_FILL,
                        fill_type='solid',
                    )
                    summary_no_issue_fill = PatternFill(
                        start_color=SUMMARY_NO_ISSUE_FILL,
                        end_color=SUMMARY_NO_ISSUE_FILL,
                        fill_type='solid',
                    )
                    for col_idx in range(1, len(summary_df.columns) + 1):
                        cell = ws_summary.cell(row=1, column=col_idx)
                        cell.font = header_font
                        cell.fill = summary_header_fill
                        cell.alignment = center_wrap
                        cell.border = thin_border

                    for row_idx in range(2, ws_summary.max_row + 1):
                        ws_summary.row_dimensions[row_idx].height = 42
                        key_message = str(ws_summary.cell(row=row_idx, column=3).value or "")
                        if "มีประเด็นทั้งฝั่ง Too little และ Too much" in key_message:
                            detail_fill = summary_both_issues_fill
                        elif "มีประเด็นฝั่ง Too little" in key_message:
                            detail_fill = summary_too_little_fill
                        elif "มีประเด็นฝั่ง Too much" in key_message:
                            detail_fill = summary_too_much_fill
                        else:
                            detail_fill = summary_no_issue_fill
                        for col_idx in range(1, len(summary_df.columns) + 1):
                            cell = ws_summary.cell(row=row_idx, column=col_idx)
                            cell.border = thin_border
                            cell.alignment = summary_wrap_left
                            if col_idx >= 3:
                                cell.fill = detail_fill

                    for row_idx in range(2, ws_summary.max_row + 1):
                        current_filter = ws_summary.cell(row=row_idx, column=1).value
                        next_filter = ws_summary.cell(row=row_idx + 1, column=1).value if row_idx < ws_summary.max_row else None
                        if current_filter != next_filter:
                            for col_idx in range(1, len(summary_df.columns) + 1):
                                cell = ws_summary.cell(row=row_idx, column=col_idx)
                                cell.border = Border(
                                    left=cell.border.left,
                                    right=cell.border.right,
                                    top=cell.border.top,
                                    bottom=Side(style='medium', color='000000'),
                                )

                self.statusBar().showMessage(f"Export Excel แล้ว: {filepath}")
                QMessageBox.information(self, "สำเร็จ", f"บันทึกไฟล์แล้ว:\n{filepath}\nพร้อมชีต Index, Chart และ Summary Output")
            except Exception as e:
                QMessageBox.critical(self, "ผิดพลาด", f"Export ไม่สำเร็จ:\n{str(e)}\n\n(ปิดไฟล์ Excel ก่อน Save)")


    
# <<< START OF CHANGES >>>
# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
def run_this_app(working_dir=None): # ชื่อฟังก์ชันนี้จะถูกใช้ใน Launcher
    """
    ฟังก์ชันหลักสำหรับสร้างและรัน QuotaSamplerApp.
    """
    print(f"--- QUOTA_SAMPLER_INFO: Starting 'QuotaSamplerApp' via run_this_app() ---")
    try:
    # --- ส่วนที่ใช้รันโปรแกรม ---
    #if __name__ == "__main__":
        app = QApplication(sys.argv)
        app.setWindowIcon(QIcon(_resource_path("PE.ico")))
        app.setStyle("Fusion")
        window = PenaltyAnalyzerApp()
        window.show()
        sys.exit(app.exec())
        print(f"--- QUOTA_SAMPLER_INFO: QuotaSamplerApp mainloop finished. ---")

    except Exception as e:
        # ดักจับ Error ที่อาจเกิดขึ้นระหว่างการสร้างหรือรัน App
        print(f"QUOTA_SAMPLER_ERROR: An error occurred during QuotaSamplerApp execution: {e}")
        # แสดง Popup ถ้ามีปัญหา
        if 'root' not in locals() or not root.winfo_exists(): # สร้าง root ชั่วคราวถ้ายังไม่มี
            root_temp = tk.Tk()
            root_temp.withdraw()
            messagebox.showerror("Application Error (Quota Sampler)",
                                f"An unexpected error occurred:\n{e}", parent=root_temp)
            root_temp.destroy()
        else:
            messagebox.showerror("Application Error (Quota Sampler)",
                                f"An unexpected error occurred:\n{e}", parent=root) # ใช้ root ที่มีอยู่ถ้าเป็นไปได้
        sys.exit(f"Error running QuotaSamplerApp: {e}") # อาจจะ exit หรือไม่ก็ได้ ขึ้นกับการออกแบบ


# --- ส่วน Run Application เมื่อรันไฟล์นี้โดยตรง (สำหรับ Test) ---
if __name__ == "__main__":
    print("--- Running QuotaSamplerApp.py directly for testing ---")
    # (ถ้ามีการตั้งค่า DPI ด้านบน มันจะทำงานอัตโนมัติ)

    # เรียกฟังก์ชัน Entry Point ที่เราสร้างขึ้น
    run_this_app()

    print("--- Finished direct execution of QuotaSamplerApp.py ---")
# <<< END OF CHANGES >>>
