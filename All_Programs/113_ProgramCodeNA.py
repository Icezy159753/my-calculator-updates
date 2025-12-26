import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import pyreadstat
# เพิ่มการ import สำหรับการจัดรูปแบบ Excel
from openpyxl.styles import Font, PatternFill
import re
import base64 # <-- เพิ่มบรรทัดนี้

class SPSSAnalyzerApp:
    """
    คลาสหลักสำหรับสร้างโปรแกรม GUI วิเคราะห์ไฟล์ SPSS
    """
    def __init__(self, root):
        """
        ตั้งค่าเริ่มต้นสำหรับหน้าต่างโปรแกรมและวิดเจ็ตต่างๆ
        """
        self.root = root
        
        # --- ส่วนที่เพิ่มเข้ามาสำหรับไอคอนBase64 ---
        try:
            # 1. ใส่ Base64 string ของไอคอน .png หรือ .gif ของคุณที่นี่
            # คุณสามารถหาเว็บแปลงไฟล์ภาพเป็น Base64 ได้ทั่วไป

            icon_base64 = """iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAYAAAD0eNT6AAA76ElEQVR42uzbvWsUQRjH8RE0MZWmSKGYLt2VR8g+ezvPDnhEN/PMnYoHKYwoEa0OsRD/BYlCDKiVWNgKvoCFkCa+VIIYo51eAnqFhYaANkG4EdQmNiohd7Ob3xc+/fI8C9PMqJDTOhuK2Y1H2jYjtjdIyzyxWyKWFrGsEss6sXiAP3wjlo/E8inQ7wOAYlj/fRa1iN3SrzNKrkfaNmN241pnQwr9W0SNgSgVR+xmieU1sXQCXToAAMDfdKJUFondLCVOjDG7FdrYWGLLEcscsXwOdIkAAACb5NZIy53I2KpSaofarmVZ1k/szpF278NcFAAAwJZ5F6f2bJZl/Wq7RNQYiLS7QCztQJcCAADQLW1iOU/UGFBFjhInxLIc6BIAAAB6pVVhmVBFa5Trw5TK/UCHDgAAEIp7YyY7oIpQxLZGLF8CHTQAAEBg3Fqs5bjKa8aYnRG7y3jKBwAA8N86EctcqdToU3kqSexgrOVZoEMFAADICffEmCN7VR7S+tC+KJXFMAcJAACQM6m8Df5eAJnaCLFdCXaIAAAAuWRXyNRGVIhVKrX9OPwBAAC2RszyYZTrwyqkytXqHmJ5FerQAAAACuJNkthBFUKlUqMPF/4AAAC6ZqFcLu9SvY5YrgU6IAAAgIKyV1Qvi7Wrd+Odv5444Q83Z/zRmYd+8vZLP/Wg7U/Pf/XTC9/9meceYIPpp52f/8epx6t+6u6yn7z1wh+7+sjbizf9wZOX/A/27u+lqTAOA/i6qv+hP6TO2Xb2enacTs9Rh5aTkxJOSiyKEjJDKshKCm9WhIsIyXTOSCSIbowICboTiqimXQR1UWk0+nUR9sR23ZV73/luPi987s8u9nzfc973+75Bp33L/rRieBb1uR9EVAOcmW9w7nxCNLMC++oziDNzsFIXYTb4FTknwAh5bmATQ87xvpb7VdWPCzV2oXloAn72rbaFhqpT6ukf+Nk3aB3NIeoPwhReBdt5PIiReW0DjYgkmP0OO70M60gaZiypMlPWDNvdHaj0UHW2f9jtQWJsAb1PfmtbQKi2HHy0jsTleYh9A5WbBJx/qGdwEZFUzvQaIsM5mPFuRXni3gtUchiiqVF6e4NoRfPpDHof/9K2UFDt65p+hYaBMRiRFrWTANEGe/y5tqFFRHI5d9cRGZyEWdcmPU8qdYtg6T5/2Vf6isQhHJh7p21RoO2n+/57xPpHYUQ8dT299Z2ITqxoG1hEJJ99/SWCXkp2nuSFELvU3+4X9k7IfPBY/wX0Lv7UthDQ9ubPvIadPK5sEhBs6YMz9UXbsCIi6Ur/+XDPObl5EnaPBlSOeDy+07DcD7IeuGkwjb6lDW3Dn6hkaQOJKw8QjHao6XJJXdI2qIhIkWwBVv+41FMCld4caFjeYbnF/6+egU/0H8X2U9s/qaY98OyCniFFRCq7BaROAvaGvb6AorHDCHurkj77882fqlKxhdAdysjfDxBtRzSzqmdIEZE62YLM5YB8sVYHZA/TarIkbfjjmj9VvY70Ikxb7m7eUPJU8Y1Az5AiImWcqc8wXTkbA/cIL6Tg83/zrbLfcurauNufakZychlBp4NLAURUNvvaCxiiVcIygHtTRetfodwH80ZuaxvmRJvhz+YRinXKWwpw9sOZ/KhtSBGROpFjN/CPvXtpbSKKAjj+IcSvItKSpLZp+kgidUKpltKqUUFr1ZUibnUTsdWCuCuKL0i11JKA1UWrWUrVIrjUhaALLQWxj1HD1VsIFB+lM+ckmST/xQ+66OQMszj3zMy5ZxTyyLLqlsDdkURSYcIf0/1Ql+xWweb2Xr1dAUczgU1QAMonevuLaeocUB4MJO/+HxOekB3vG9gEDkgduPlab8JXJGlary4ENkkBKJ/IufvyJ4mhxBXNT/4uCj/sw90/6l5qbFZvQJAzbLuDA5ukAJTJnSU7JVSaQ16qLP6hUNdO4Sd/7Vf9Apu0AU3xszf0GgIvPAhmggJQVuET16T5o7irbd8Ohe1/yZjwRGyjlFqCHZwrGifvmu6HqyaW/WZt/J3KuWZorvi/44iPikg/+25HB9MQCMC3tvFX8t0ALfGowuz/+Iiooal7QGXiX/o3J+ea9uzWF87+z5HCX8cTHxWdGNjc5tAQCMD3hMCmjn5p/hiWFwDh+HXJSXSezKgsfonptW1fvOT02uZFkPiouJ5Lk3oNgaMvgpmkAJRN6PBFYSNgclzeABhKPBV1/2ceiZOpk1v3fPGcnFs6nviouPTzn6alb4SGQAC+RM5npbljVmMHwBvJSeyfWJC+8y499vaq9E6c+KiK/rtv7R08DYEAPGvNFKR5Y1FjBPB7yUkMznyU3f3mXb8X0B5rf4P4qJquM2M0BALwyn4cTJo33mk8AViSnMShJ18lCdR2uPu+gN1Tq/Y3iI+qOfh42TQrjQoOH7sc2GQFQFf01ifZTUM48VmjAHAlJ5Ge/yFKoB3ZFd8XMJZdsb9BfFSVk5mhIRCAN/eWpTljXaMAMALi5BmTLICTK8QP6KLYUApFs6fvlE5DYM9xGgKBBiHNFzVfADT6I3heAdQHGgIBUAB4lMr5b4JL5V3iB3RBbERdp0eVGgJ7TXTiQ2CTFgAKAI0CwG5l833xhuaLxA/oYtiIaAgEQAHgkZNz/W6BIz5oCARQkygArIIxSQ+jcPdOr5l0YeNY4oOGQAA1iQKgpLC9O2En725e/IgPGgIB1CQKgD8MlT6HO7Vq98hb9m/b8Pavd97EBw2BAGoSBQBQh2gIBEABADQoGgIBUAAAjYiGQAAUAL/YuWOcqKIwDMMLo5lGO62mk0Y7YQdgJjZoMLZi0LXoAky0o7AgQAUFK4DiEBIKEioSbrjwPsXTk+8UvGH+AZocBAICAKIcBAICAIIcBAICAKIcBAICAIocBAICAJocBAICAKIcBAICAIIcBAICAKIcBAICAIocBAICAJocBAICAKIcBEKbAIAoB4HQJgAgzEEgdAkAKHMQCFkCAOIcBEKTAAAcBEKQAAAcBEKQAAAcBEKQAAAcBEKQAAAcBEKQAAAcBEKQAAAcBEKQAAAcBEKQAAAcBEKQAAAcBEKQAAAcBEKQAAAcBEKQAAAcBELQkw8AAODeBAAAIAAAAAEAAAgAAEAAAECVAACAIAEAAEECAACCBAAABAkAAAgSAAAQJAAAIEgAAECQAACAIAEAAEGPHgD/zwcAcD8CAACKBAAABAkAAAgSAAAQJAAAIEgAAECQAACAIAEAAEECAACCBAAABAkAAAgSAAAQJAAAIEgAAECQAACAIAEAAEECAACCBAAABAkAAAgSAAAQJAAAIEgAAECQAACAIAEAAEECAACCBAAABAkAAAgSAAAQJAAAIEgAAECQAACAIAHwwH79PRqr3f2xXN8YixfL658RAJ4dAXDj4OxibO/sjbXFq9k+FgAIgAf+5f92Y2u2jwQAAmACWzvfZvtAACAApvnM35/9AcjJB8Bqd3+2jwMAAmAiyzfvZ/s4ACAAJrJ46at+APQIAAEAQFA+AHwEAEBRPgBWX37M9nEAQABM5Pe/47G2eD3bBwIAATCR7U97s30gABAAEzk4vRzvNv0rYAA6BMCtCPjw+buPAwBIEAB3bgJOxsevP8dyfdNXBAF4tgTAE2d/+5fZ3/5lAiDO/vYvs7/9ywRAnP3tX2Z/+5cJgDj727/M/vYvEwBx9rd/mf3tXyYA4uxv/zL7279MAMTZ3/5l9rd/mQCIs7/9y+xv/zIBEGd/+5fZ3/5lAiDO/vYvs7/9ywRAnP3tX2Z/+5cJgDj727/M/vYvEwBx9rd/mf3tXyYA4uxv/zL7279MAMTZ3/5l9rd/mQCIs7/9y+xv/zIBEGd/+5fZ3/5lAiDO/vYvs7/9ywRAnP3tX2Z/+5cJgDj727/M/vYvEwBx9rd/mf3tXyYA4uxv/zL7279MAMTZ3/5l9rd/mQCIs7/9y+xv/zIBEGf/6f05vGLvTp+ivLIwgP9fwaVtBmSXnYYGgwooyMg2VcqqjuWADOMWiYLIKEpAUBb3dUpgJkrHSERRB6NZVWIGnUoUp8bomTofmJQaFfq+Defe+3z4VVkpg/1e6pzz9Lvc9zG19PRRVW0jZRdUUZxnOQVHeGhucCyvIQD4iWuIa4lrimtr7Z93U0tPP9ec2H5gUv9HANAc1j8wrnz1E9W3HCXvsmKxzRPAVB/Mj6b0rBL6eP8xrkWxfUL3/o8AoDmsv7MGR8dpXd1ucoUkim2OADbhWlxf10S+2+Ni+4au/R8BQHNYf2fc/OEX2rG3lxYsTBLbCAFs5gpJoE31rTTy4LnYPqJb/0cA0BzWX13/tXvk+XCV2MYHAL9KzSyg/uv3xfYTnfo/AoDmsP5q2k9cIndYsthmBwBv4pptPzUotq/o0v8RADSH9fffns5zFLRgkdgmBwBvF+SKoYa2k2L7iw79HwFAc1h//zS2neY7jcU2NwB4P67h3e1nxPYZ6f0fAUBzWH+/TvvztwexTQ0ApozP4ll7OQABwEKXRv9FO1tP0PLCtaS6/vwz+GcNjtrxiE3/tXu45g9gmOCwFBoYse/GQAQAS9x8+ILajl8kHthBruiAXE9bXriOvx3zvyV2HVTcGHuOu/0BDJW6pJAf5xXbfxAAEAD8GPwv6cCRAYpLXT5jhRSTtIxvrjGumLY0HRLbvABA3fbmHrH9BwEAAWBauv/2BcWnrpi1Ykr05lLvhati12eaO/xhdz8Aw/FGXr4vH4ntQwgACADvNfT1z1RUViumqIor6mjomydi12sKeHtfsU0LAJyzYXOz2D6EAIAA8E69fcMUFZ8prqiiEpbQkf5rYtftPS/2wbd/AEu4QhP5S5TYfoQAgADwm/i6+xy33FfM8o2CO/YdEbt+b8Fv9RO7pgDgvJ2tx8X2IwQABIBX3PrxJVVtahBbTK/h0+n8mcWu5+vSlhY5duze0DQ66Cmju1lbaSKvgSh/DwD4h2uIa4lrimvLqTrlVwmL7UcIAAgAr9zlv+aPH4kd9m9TtrFeixDgu/PYkR3/5rli6HBqBb1YKbORAujuRX4TdXrKudYc2SHw8t1/i+1LCAAIAIwHqdgh/z6VNbvEruuklp4+R4b/5cwasY0TwCSfZdY4EgL29w6I7UsIAAgAtHVPl9jhbspzt1W1jarHyN/8xTZLABN1eMqcuFQpti8hAFgeADrPXKYgl/4vo+FjOHT2sth1zilYq3rNn09Nim2UACbimksLTVXc5nyd2L6EAGBxAOBNaUKjvGKH+nQtjPaK3XxjUVK2yrHxzUlimySAydpSSpVqNzYlR+wMQACwOABkr6oUO8wVXiokcq1DIpW+RfAdymIbJIDJ7mRtUard0Kg0sTMAAcDSALC3+4LYIa6I31kgbr3nBqvtqzCR1yi2QQKY7Gmu2qPRc91xYucAAoCFAeDqdxMUtmhxYF6HGRJDVYUZ1FufRze6S+hxXyk991UQ4z+PdJdQb30uVRZm8N8NyGeIiP2Qhu89M6pApDZHABtgviAAGKOu4aDzW/RGxVHX9lx6drGc6PPKKeG/y0EhKT7B6c/Db9wzqkCkNkYAG2C+IAAY4Ytvn05ej3bEfHcMNVdn08Tk4PcDnx3YX5tD89zOPY3ANzcOf//MmAKR2hgBbID5ggBgBN5D36khGx4eS0MdRTzEHXHtcDFFRcY5tw/3gePGFIjUxghgA8wXBAAjJKavdGS4ZniSaOz8Gh7cjho7t5q8yUmOfMbkxfnGFIjUxghgA8wXBADtnfj0liODla/Zj/eVTg7tgISAyAhn3kZ4enDUiAKR2hgBbID5ggCgvQ1b9jpyzX/0SAkP6oAaPlTsyD0B1dtbjCgQqY0RwAaYLwgA2ktIy1X9RfENfzygZ8S+TTnqZysy8o0oEKmNEcAGmC8IAFrz3R534lG/adztr+6/gxWUEKf8iCC/jlP7ApHaGAFsgPmCAKC1tuMXVX9J/Jw/D+YZ1bND/axF+6lB7QtEamMEsAHmCwKA1v60rUV1h7/JTX5mFJ9xcCvuGFi944D2BSK1MQLYAPMFAUBrBaU1KuvD2/vyQJ4VFavSlT57YVmt9gUitTEC2ADzBQFAa2lLi1TWh/fvn7UA0P1RntJnT88q0b5ApDZGABtgviAAaI1fkKOwPvxin1kLANe7itVuXozP1L5ApDZGABtgviAAaC04PEVpfR73l85aAHh0oVTt/oUIj/YFIrUxAtgA8wUBQGtz3Eo76/EjebMVAPjfVnsfd3Cs9gUitTEC2ADzBQFAawgAeheI1MYIYAPMFwQAreESgN4FIrUxAtgA8wUBQGuqNwGOzOpNgCW4CVBoYwSwAeYLAoDWvMuK8RigxgUitTEC2ADzBQFAa6obAVXO4kZA5b/PwEZAQhsjgA0wXxAAtFa9XdOtgD8tJ/fvsBWw1MYIYAPMFwQArbWfuKT8S+qtz9PyZUAHT/m0LxCpjRHABpgvCABa8335SP3d+vEJ9NxXMcOvA45X+swfzI/G64ABAAFAcH9DAJgBiV71b9P7a3NmLADs3ZSt/HmTF+cbUSBSGyOADTBfEAC0t3HbPuVf1NwF0TTUURTw4T98qJjmuaP539T++j8CAIDeMF8QALR38uI/+ViVRUbE0tj5NQEb/g/OraaIyFhHPuvpwdtGFIjUxghgA8wXBAAjJGXkOzJYvclJAQkBY+dWU1pKoiOfMWXxKmMKRGpjBLAB5gsCgBHqW47y8ToiPDyWrhwscvK0P59dcOzz7Ww9YUyBSG2MADbAfEEAMMLV7yYoNCrNsSHL1+lbanNUXhbE/y/f8Dd5zd8RC6O9NHzvmTEFIrUxAtgA8wUBwBibGzv5mB2VEJdAvF3wxMXy6Wzyw8/5Tz7q56ite7qMKhCpjRHABpgvCADGGP7+GYXHLubjdpw7JIYqVqUT799/vauY+E1+/A2f8Z/5v3Vtz+Xtfd+1w5/yy3+G7/3HqAKR2hgBbID5ggBglJaefj5uI7Ue+4e49Z4brHZvw9PcBrHNEcBkT3J3qT067Y4TOwcQACwNACynYK3YIe6v3OL1Itc6JDJV6bjuZm0V2yABTHYna4tK7fI9V2JnAAKAxQGAtwdeGO0VO8ynKyw6nXx3Hotc69jkbLX3GXjKxDZIAJO1pZQq1W5sSo7YGYAAYHEAYJ1nLlOQK1rsUJ+qIFcMHT77udh1Vj3b4g1Noxf5TWKbJICJuObSQtXO3i0vXCe2LyEAWB4A2LbmbrGDfap27O0Vu76sqrZR+Rg7PeViGyWAiTo8Zcp1u65ut9i+hACAAECssmaX2OH+Hjxcxa6rkzddznPF0GeZNWKbJYBJfJnVNHe++pNKB44MiO1LCAAIAMRu/fiSyjbWix3yb1NRvZM/u9h1ncT3JvCriR0IAfytBJcDAAKDa4trbHL4G/E6cgQABICphAA+XSV22L9uw+ZmLYb/pPSsEseOna9L8s1JfIfyRF6j2GYKoAOuIa6lTzylk9f8HZGR/Qex/QgBAAHgNzW2n6I57lixgz9owSL6eP8xsev3jvcwiF1TAPg/I99HggCAADBlRweuU1TCEnEFFZ24lI4N3BC7bu9y5aufyBWSKLZZAYBzFixMoqGvfxbbjxAAEADeaeibJ1RcUSeimPhaWknlX2jo2ydi12sq1tc1iW1YAODkJcq/iu1DCAAIAFPWe+EqJabnzVohJaavpN6+a2LXZzp8t8fJFYqzAAAmc4cli92UDAEAAWDabj58Sfw4S3zqihkropikLGpoO0k3f/hF7Lr4Y0vTYbGNCwDM35cEAQABwM8g8ILaT16iFUXrefe9gOzox/v5t58apJsP9bnDfzpujD2n1MwCsc0LAPyXtrTIuC8tCADwhsHRcdr1yUnirS4V159/Bv8sPkUu9nid1H/9Pp8mFNvEAGD6gsNT6O8jD8T2HQQABIBAwPuyp4/PcvAjjWKbGQBMHddyx2mf2H4juf8jAGgO6++f3e1n6IP5+r+MCcBmXMNNHWfF9hnp/R8BQHNYf/81HzqPMwEAmgpyxfDmaWL7iw79HwFAc1h/9csBwWEpYpscALyBa9ba0/4IAIAA4KCBkfuUuqRQbLMDgFfu9rfyhj8EAEAACAx+fIifIcYTAgBC8Xbem+pbaeTB/9i7f5UgowCMw/cVRH9EiiIoKlAKCircMmpXa6tsy4QsCE1oVMr2bCmTloqg0KaKBilaDU6cC2hLPcf3GX4X4MN3vu910LPZ7Hukt/e/AdB5/P//9cE37jz0XwOlRqpnsd5A+mZt91/vawDIAGigepnI9OOlepWwvxaQtrl65uqVvtPzSzEX+xgAMgAabHX9V5lbfFkmJmfKyJXxcmJ4pAweHS57B9q9ulnqoXqG6lmqZ6qerXppV/3X56vrfts3AGQANB5//snx79vfAOg8/vyT488/OQMgPP78k+PPPzkDIDz+/JPjzz85AyA8/vyT488/OQMgPP78k+PPPzkDIDz+/JPjzz85AyA8/vyT488/OQMgPP78k+PPPzkDIDz+/JPjzz85AyA8/vyT488/OQMgPP78k+PPPzkDIDz+/JPjzz85AyA8/vyT488/OQMgPP78k+PPPzkDIDz+/JPjzz85AyA8/vyT488/OQMgPP78k+PPPzkDIDz+/JPjzz85AyA8/vyT488/OQMgPP78k+PPPzkDIDz+/JPjzz85AyA8/vyT488/OQMgPP78k+PPPzkDIDz+/JPjzz85AyA8/vyT488/OQMgPP78k+PPPzkDIDz+/JPjzz85AyA8/vyT488/OQMgPP78k+PPPzkDIDz+/JPjzz85AyA8/vyT488/OQMgPP78d7qVtZ9lduFFGbt1v1waHSsnhkbKwJGhsnfgeP0ZpX9Wn5H6rNRnpj4747dnyuzCcn2mmn3ed9P7xwDoPP78d6K3X36Xqdmn5cyFq81+XNRve/YdK2cvXiv35p7VZ63Zc9D7+8cA6Dz+/Lez1582ysTkTNk/eKrZj4d2V/VZuz75oKx83mj2XPT6/jEAOo8//+3o448/5e6jxXLg0OlmPxTa3e0fPFluTs2XD983mz0nvb1/DIDO489/q1t+97UMnbvc7IdBWQ2fHy3L7781e156ev8YAJ3Hn/9W9uT5q3Lw8F/2zv0piiuL4/8XiMJQIhhGHs4MgQGiiDJurAgCVbsqPmK5KoUioKsissagM4pGUd66rspGZSCLkvgIBjcvk6xZk62N4la5eLcOVVQlli+83T3fO/394VNlpQjcPn2/53y77+1zA7DFgLgTmZOR3kFY3ZiSf2gADIfxZ/zt4kD7OZWYkgVbBIi7SUxeoPaFe2D1Y0L+oQEwHMaf8beD5nCf7MSGTf6ECDJH90f6YXWEnn9oAAyH8Wf8bXjtL09XsEmfkF8hb6lcuxxAA+BCro79S+053K3KytdrTwD5HfK7Bsf4iQ0NwNSGP675E+NITc9TAzfctzGQBsAl3Lo/qcJdV6YKdmKy15b1tLLyDfL0J38LNg7xJkCka7n5wxPu9ifGkr+4XD5Xhc0VNADgCRCz8D9VbacHVE5+mWNCWuBfIptrXCcmtxuA+pbjsMmdkNehsfUUbK6gAQBPgGicPH9N5eaHYiYmX3C56rhwHTY+8SBAoA5/7O5HjEcaVUW/+Ak2X9AAgCZAJEa+/EVVrK6FEdWqtXVq5KuHsPEyWYAg1yHtfWGTOiEzYdOOVth8QQMAmgBR6Lg4qjJzF8GJKnPhYnX60qewcTNVgCAH+/Dpn8QNyWk+eYiCzRk0ADQAz0XW3Wd5cI9QlY2Cuz44DRs/EwUIcA1yqh/snCPkTdhzuAs2Z9AA0AD8hts/PlXrtu2DFdMzyOtiGTNsPE0SIMI1FJRUWDY3ggG/OlofUnc7q9X
            ElTVKfVJDyIuQOSJzReaMzB2r5qEcJQybM2gAaAB+s8v/9+83wRb7F7F6826agDgwANHxny3p+Dfb41UnGperySHMQkPwmRyuUe07QzKXLOkQOHz337B5gwaABkCQQgpb5F9Fzda9sHE1RYCxHv+hUxctKf7DkQrYwkLMYihcbokJ+LBjADZv0ADQAKidBz6CLe787tYdBmBdbbPuNciTP2wxIWZyrD5kxVIlbN6gAXC5AWjvH1aJyeYftiLXcPzsMGyc0QUY6/EvW7lee81/chiziBBzkTlVkOfTbHO+ATZv0AC42ABI05W0zCBsUZ8p87xBVzXfiCcDkOVfqjN+2bwFW0SI2YS3670FyM5bBps3aABcbACWvlcDW8w1DhWCjTeyAGM9/rlv5WuN/25XNWwBIWYz3lWlNTfTMgtg8wYNgEsNwMGTF2CLuC5tp+N/0028GYCkVL2+ExOX+akfsYdHl9dozc0kTw5s3qABcKEBuP7NhErPKrZnLT4tV2VUVyl/a5MqPndUlV7rUWVj54WpfxefjahAa5PKqK6Wn7VlDPOz31Gj9x7Dxp8GwPLxwxYPEh+Yri90/dIAOEjdvqOWF92kBW8rf0uDWnbznAqNX3wt5GfFKCT7Cq0ej5woBxt/GgAaAGIWpusLXb80AA5x7etH0+utlpCQmq1y6rdMF/43Qt4O5DbVqgRPlmXjks2No9+65y2A6fOfBoAgY7q+0PVLA+AM0kPfuqf++X5V1N0mRdwSivvDarY3z7o+3G3u6MNNA0ADQGgATNYvDYBD+ArftaS4eoKLVclQpxRuS1kSPaNS8q05hTBQvAL2PtAA0AAQczBdX+j6pQFwgO7Lty0prLJmXzrSM120bTEBSZkBS8baNzgGez9oAGgAiBmYri90/dIAOMCm+oOWrPm/89d2KdS2UtR3xJI9AVsaD8HeDxoAGgBiBqbrC12/NAAOsLBgue6Nkg1/UqAdIbdhm/Z4/UXuWAYwff7TABBkTNcXun5pAGwmeueBFZ/6Te/2d4Syz8+rOQu1PxHkcZwGzH8aAIKM6fpC1y8NgM2Eu67oP023NEhhdhR/a6P2uCO9g7D3hQaABoDgY7q+0PVLA2Azf2w4pNvhb/rp31HkbybOzdHbB7CrDfa+0ADQABB8TNcXun5pAGxm5R+26sRH2vtKQY4JGVWVWmMvX10Le19oAGgACD6m6wtdvzQANlNQUqETH+nfHzMD4D+gtwxQWFoJe19oAGgACD6m6wtdvzQA9iIH5OjERw72iZUBkA6BWmPPzF0Ee19oAGgACD6m6wtdvzQANpOaoddit/R6b6wMgDQd0hp76vy3Ye8LDQANAMHHdH2h65cGwGZmebTOW5dP8mJlAORv64xdzpqHvS80ADQABB/T9YWuXxoAGgAaABoAGgBCA2AgNADgcAkA877QANAAEHxM1xe6fmkA0DcBno3EcBNghJsAaQBoAAgNACg0AOAEl6ziZ4BxjOnznwaAIGO6vtD1SwMA3wioOnaNgCrZCIgGgAaA0ACgQgMAzpZGtgKOZ0yf/zQABBnT9YWuXxoAm4l0X9U/DKi1ycjDgI72RmHvCw0ADQDBx3R9oeuXBsBmol/8pH2Tkn2FqmzsvMPHAQe1xpwwx8vjgA2Y/zQABBnT9YWuXxoAB/AFl2vfqNymWscMQE7DNu3xBopXwN4PGgAaAGIGpusLXb80AA6wueED7RuVkJKlirrbbC/+RX1HVIInS/4m1/9pAGgACA0AMDQABtBz5XO5Vm2SMgOqZKjTtuJfEj0jf8OSsfYN3oG9HzQANADEDEzXF7p+aQAcwl+0wpLCmpK/yBYTsCR6Rn63JWPMK34P9j7QANAAEHMwXV/o+qUBcIjdh87I9VpC0ny/Kuxqs/K1v2VP/sKew92w94EGgAaAmIPp+kLXLw2AQ1z/ZkKlZRZYVmRlnT63sVbnsCD5f2XD3/SavyXM8wbV6L3HsPeBBoAGgJiD6fpC1y8NgIPsaG6Xa7aUOQsLpV3wTJoFyc/Kd/7Tn/pZys4DH8HGnwaABoCYhen6QtcvDYCDjH77WGVkF8t1W4507cuoqpzq31/cH5aT/OQJX5B/T/03f0uDtPd9WYc/7cN/Ru/9Fzb+NAA0AMQsTNcXun5pABzm0KlLct1xyeHOj2HjjirAWI8/KTVba/yPLq+BLR7EbB5+vFpvr5QnBzZv0AC41AAIy1auhy3ib8ryVRth440swFiPf+5b+Vrjv9tVDVtAiNmMd1XpzE3ZcwWbN2gAXGwApD3wPG8QtpjPlHRvoYqO/wwbb2QBxnr82YGleuc91IdgCwgxm/D2kNbczM5bBps3aABcbACE9v5hlZjshS3qr0ti8gJ14uwnsHFGF6Dpb6OCAb+aHMYsIMRcZE4VBHxac7OsfANs3qABcLkBEBpaT8IW9tdl18EO2PiaIMBYj39dbbP2NbTv5FsAYi3Hdoa05+WGuv2weYMGgAZgipqte2GL+yuQ4gEbV1MEGA+bUmd7vGooXA5bTIhZRMPlKilF/+1o2+kB2LxBA0ADMMXtH5+q1Zt3wxb5F7F2yx4ZO2xcTRFgjMcvezfk6GZLTMCx+hCXA4jOa3958p8u/jyOnAbAHUgh3VC3H7bYP8umHa0s/nFiAITC0krL5kZBnk82b03t4J7gJ4LkFcgckblyZHvZ9Jq/JRQtrYLNGTQANADPpTnSq2Z5smELf2JKlvrTh52w8TNRgPF2TgUhGvA8EhoAd3Nm4DOVuXAxnKC8vhLVOXATNm6mChDhGv7+j/+o5Lk+2GROyExImedXI1/+ApszaABoAF7KyFcP1aq1dRBikrW0yprtauTrh7DxMlmAKNexsa4FNqETMrMlyj/D5gsaANAEiEjHhevKV/i7mAnJV/iu6rj4KWx84kGAMM2p7jxQyWl8C0DMxpMecFVTMhqAOOfW/adTn7Pk5occE9ECf6naF+5Rt/75P9i4xIsAka6lvuUEbGInhH1JaABcy637kyrSc1WFKjZK9z1bOvpJP/9I76C6dZ87/N1oAG7+8ETlL1oJm9wJeRkFJRWue2ihAXAhg2MP1N4jPdLqUncCyO+Q3yWvgGGvF5l4m/+XPvtOXqPCJnlCnkdqRp76243vYfMEav6hATAcxp/xtxh5CySffMIme0Ke/Tz5WF8UNkcg5x8aAMNh/Bl/O9gf6VcJc8w/rIrENzJHW46dhdURev6hATAcxp/xt4vW43/hmwACi+xbao70wurHhPxDA2A4jD/jb/dyQGp6HmwRIO4kNT3Pta/9aQCIawoQOm6I/8CN71T+4nLYYkDcRUFJhSs3/NEAEFcWIGTcEn/5vGrXwQ5+IUBihrSr3rb7sLrx/RNYnZiWf2gADIfxZ/ydPj74/e0H2DWQOIbMNTmBdGg8/o/3pQEgLEAG4db4y2Er/2fv/lmqDOMwjr+vQLJEiiJI+oNSUFDhllG72lg5ZkI2hCY4KmW0dlpKpKUiKLKpokGK1oJf3FtDUx3hOV6f4fsGPtzP/VwOnmfhwUb7lLD/FlDfa2eqfdJ3YWUj5sM+BoC8gAYs/lVb299ref1Zzc4t1uSVmRqbmKzRoxM1NNLdT1urG7Uz0s5KOzPt7LSPUrWfPt/a9te+ASAvoI7Hn39y/Afb3wAY8PjzT44//+QMgPD480+OP//kDIDw+PNPjj//5AyA8PjzT44//+QMgPD480+OP//kDIDw+PNPjj//5AyA8PjzT44//+QMgPD480+OP//kDIDw+PNPjj//5AyA8PjzT44//+QMgPD480+OP//kDIDw+PNPjj//5AyA8PjzT44//+QMgPD480+OP//kDIDw+PNPjj//5AyA8PjzT44//+QMgPD480+OP//kDIDw+PNPjj//5AyA8PjzT44//+QMgPD480+OP//kDIDw+PNPjj//5AyA8PjzT44//+QMgPD480+OP//kDIDw+PNPjj//5AyA8PjzT44//+QMgPD480+OP//kDIDw+PNPjj//5AyA8PjzT44//+QMgPD480+OP//kDIDw+PNPjj//5AyA8PjzT44//+QMgPD4736bH77V0trTmr55ty5NTdfY+GSNHBmvoZHjzVD659oZamepnal2tmZuLdbSWq+duc4+D+6fveNvAAx4/Henlx9/1PzSwzpz4WpnXx7au+3bf6zOXrxWd5YftbPY2efE/TPY/gbAgMe/v714t1Ozc4s1PHqysy8HZdXO4vW5e7X5fsf9IwNAHsB+9/brr7p9f70OHDrV2ReBshsePVE35lfqzZef7h8ZAPIA9qPeq081fu5yZy9+6c8mzk9V7/Vn948MgPT4/1+rj5/XwcOnO3vZS3+rndnVJ7/bu9fgKqtzD+C7M+f09Hw558z0fDiXOZ/O+eZHpjUX9iU7MRD23kkgCTEXaQiQhMuRi61aDIg4VMAQBLmFANUABYvcVC4B2wmkrZKigUAAiWBBUSoitqVQpbC6n4zMtExFkrV28l9r/d+Z39RpbbL2s9/1PE/ey1qHmH88xwbAc4x//63d8rrKzCqATfJEd5MZyVdNm/Yw/3iMDYDnGP/+Wb2pVZ60hk3uRPf6tkDz5n3MP55iA+A5xr9fl/3lryfYpE7UB3IVS24HMP94iA2Ah948/YlauXG3qp32pO4JID9Dfpa8/gb7eQ0/8Md7/uSc7OGlqr37Q+sKEGqesAUbAE+c/Pim2rTrjd6CnRlJpOR+Yu20OfLXsfwu2DjoOHHxBp/2J2eVjZ0mr7NaVYBQc4Ut2AA47uTHt9SGV9rVyLLaAUsk+SXj5OEiSSawcemPxWtehk3eRCYsXbfDqgKEmitswQbAYdt/fkSNKqsbtGRSXDFR7Ww7ChufPq7wx9X9yHmykFXHu59aU4BQ84Ut2AA4qPM3f1DTZy6ESSoz6htU57mrsPG6B7K8L2zSJjLp6cZ11hQg1HxhCzYAjtl5oEvFR1XBJZV40Vj1ysFjsHH7mo19+Nc/eSOSWyx/RFhRgFBzhi3YADhE7rsPjeJuMSsPCi578RXY+H0F2dUPNqZEqbBy4y4rChBqzrAFGwAHnLp0S81e0ASbTO4gl9NlzLDxvFN59XRz91jLqlTxqkZVdXCrqjnxuqo7c4Cov+QcknNJzik5t0ydp7KVsBUFCDVn2IINgOXkKf/H5z4PW+y/ysx5y61oAjrOXDGy4l9GtECVNC9RtT1tsMWELNfTpoqbFsu5ZmSFwMNnP4MvQKh5wxZsACyXLKSwRf7rzJq/CjauQrTsOGCk+Fe+th6zaJBzKl9tMdIErN/ZDl+AUPOGLdgAWOy5tdtgizvqe8d9NXvhat3PKH/5wxYLclNxU6OJW3XwBQg1b9iCDYCltuw7rDIj9m9GI5/h5f2HYeNcM/VJ7Xv+dbzsTwOtp033mQBZ2RO+AKHmDVuwAbCQLEqTG6+ALep9NSxRcXvxETgFJRN0Pps8nIVZIMh5RSsWaZ27haU18AUINUfbgg2AhSY8PAu2mGtsKgQZ6wdiZVqfq6p9G2yBILfJ2wE6525uvBy+AKHmaFuwAbDMC9vbYIu4JtmzAC7ewWytdRX4qh8Nmpru/VrnbjA6Er4AoeZpW7ABsMjR96+p4QVjUrMdaFahmp1fp3aWz1InxjWoK3Ur1I3JzaL3n7urG3r/t1n5dfLvpmQMIwq/p7ouXHdqgqAWB/IDen5FH5/r2ABYpKHpJfNL9GYXq21l9er6pNVKTVl7T+TflWagJKfU9Hhkxz2nJghqYSA/oOdX9PG5jg2AJY6c/+Pt+9FGhMMFal3JD9S1SU1S1PtFrg6sL31MhcL5xsYlDzd2fXDdmQmCWhjID+j5FX18rmMDYAlZQ99Ukc3LGqU6xy6QIm7EsepnVTxabG4d8g27nJkgqIWB/ICeX9HH5zo2AJYorpxspLg+lFuuLtYuk8Jt1MWaZaoit9zIGEePmeLMBEEtDOQH9PyKPj7XsQGwwO5fnTJSWOWe/eW6FXcUb7NNQCxaZGSsrYdOOzFBUAsD+QE9v6KPz3VsACzw9OIXjNzzPz1+kRTqlOoau9DIMwHPLG1xYoKgFgbyA3p+RR+f69gAWKCofKLuFyUP/EmBHhAvjn5U/2rFQ1OcmCCohYH8gJ5f0cfnOjYA4Dp6Lpt41a8PT/vr+2JSsyrSf0VQtiO1foKgFgbyA3p+RR+f69gAgNu06w3dL0ne85fCPKB2lNdrj3vz3kPWTxDUwkB+QM+v6ONzHRsAcD9a0qK7wl8fFvkxR644RCMFes8BLNtg/QRBLQzkB/T8ij4+17EBADf1h/N14iPL+0pBHhT1iVqtsU+budD6CYJaGMgP6PkVfXyuYwMArrx6uk58ZMneQWsAtpfp7VpYOf4R6ycIamEgP6DnV/TxuY4NADjZIEcjPrKxz6A1AMern9Uae3xUlfUTBLUwkB/Q8yv6+FzHBgBcdp7e0/RXJq4ctAbg07oVes8vjHjQ+gmCWhjID+j5FX18rmMDAG5oVGvrXXklb9AagC8mN+uMXfbit36CoBYG8gN6fkUfn+vYAIBjA2D3BEEtDOQH9PyKPj7XsQEAx1sAdk8Q1MJAfkDPr+jjcx0bAHC6DwF2Vw/mQ4ANfAgQtDCQH9DzK/r4XMcGAFzFuBl8DdDiCYJaGMgP6PkVfXyuYwMATnchoFmDuBDQE1wICLYwkB/Q8yv6+FzHBgDcM0ttXQp4tYpGCrkUMGhhID+g51f08bmODQC4zbvflM9p3W0AE5sBvbS3w/oJgloYyA/o+RV9fK5jAwCu491P9ffWzylVNyY3D+jrf0U5o7XGnBFOcDtgIjYAXud/9PizARgAxRUTtb+o9aWPDVgD8MLoR7XHO3rMFCcmCGphID+g51f08bmODYAF5i15UfuLCobzVefYBSkv/l1jF6pQOF9+p/X3/9kAkO3Q8yv6+FzHBsACe954Rz6rtli0SF2sXZay4v9R7fNqRLTIyFhbD/U4MUFQCwP5AT2/oo/PdWwALFHy0BQjhbUitzwlTcDFmmWqPLfMyBhLxzzszARBLQzkB/T8ij4+17EBsMTyllfl8xqRlzVKvV21wOhl/5ihv/zFyo27nZkgqIWB/ICeX9HH5zo2AJY4+v41lRsvN1Zk5T59y+jHtDYLkv+vPPAXiuQbG9ewRIXqunDdmQmCWhjID+j5FX18rmMDYJHG1VvkMxtVlFPau07AtUlNfVnkR97zv/2qn1HPrd3m1ARBLQzkB/T8ij4+17EBsEjXB9dVXuEY+dzGRSMFqj5R27t+//HqZ2UnP/kLX8g/9/5328rqZXnfu63wp735T9eFPzk1QVALA/kBPb+ij891bAAs07LjoHxuJ2187Zdw8Q5m6zU7Nd37YYsDuW3C8X16rw5HR8IXINQ8bQs2ABaqmfokbBHvr4kznoKM9QMxvTcbqtq3wRYIclvVwa065648cwRfgFBztC3YAFhIlgcelqiALeZ9NTxRqTrOXIGMdeHoCVqfrXhVI2yBILcVrVikde4WltbAFyDUHG0LNgCW2rLvsMqMJGCL+r3KjOSrrfvfgo2z7tWWrLIqVdfTBlskyFE9bSrrwe9pnbu10+bAFyDUvGELNgAWW7JuO2xhv1fLXtgJG18xe+Fq7c9Y3LQYs0iQs+Sc0z1v5zQ0wxcg1LxhCzYAlps1fxVscf8aUlxh42ryocuMaIGqfLUFtliQWyqS51pGVoH2ebvhlXb4AoSaN2zBBsBypy7dUjPnLYct8l+l/pmVMnbYuAohzyZkhBNGmoDipkbeDqDU6WmTv/xvF38rtuNm/mcDwBNAvwmQy3Wwxf5OTzeus6L4C1E5/hFjn12eCZCHs+QJ7ZoTr2MWErKGnENyLsk5dfuevwkPTfi+FQUINWfYgg2AQ1Zv3quGRgthC39mVoFasf412PjdZR8G2JgS/RXr9uNg/mcDwBPAoFfbj6t40Vi4hJIorlavtZ+AjdvdvP3e71XkgWLYZE1kUtawEtX5mz9YUYBQc4Yt2AA4qPPcVTWjvgEimci9xEdmLVKd56/CxutePNWwBjZhE5m9RfdjawoQar6wBRsAh+1sO6qKKycNWiIprpysdh44BhufvujouawiubwKQG6LDh8tD75aU4BQ84Ut2AA47uTHt3pf5xlVVjdgSSS/ZLxq2rRHnfztn2Hj0h+L12yFTdxENq7LwfzPBoAnwIA0AjfV5j1vqrrpT8nqeylZ0U/W89+891Dyd9nzhH9fnLh4Q5VVTYVN3kQ6yqunS9NuVQFCzRW2YAPgoUOnL6tVP9kjS33qngDyM+RnySVy2M9r0sHjH8plUtgkTtQf2Xml6hfdH1lXgFDzhC3YAHiO8e8zucohrzTCJnOivr6e+9PWDuYfD7EB8Bzj3z/Nm/epjLD9mzGR3+QcXvPT/cw/nmID4DnGv//WvfwzXgkga2VG8mXxMOYfj7EB8Bzjr387IHt4KWySJ/o75JyVy/7MP55jA+A5xl9fe/eHqmzsNNhkT3TH0/7ywB/zD7EB8B3jb4S8PiXvUPMNAYIly1kvWL5RdX90g/mH2AAQJ2Aqtg+eu2gtVw0kGHIuyg6cvz7zGfMPsQEgTsBUk81UVm7YJVsJ820BGpSn+2VL35Ubd8m5CDtPmH/sjj8bAMsx/ql3+Oxnav3OdjWnoVnVTntSjSyrVQ/EylQwG3frZrJD8hySc0nOKTm3ZNMqWfpbzjnY+cD84078TTQAn2sMQO5nwQbXBpyAjL/PGH/G31fdF7/Qjf+fTDQAlzUGIHu1wwbYBpyAjL/PGH/G31eHz/5O7zZTKH7JQAMQe09jELI2O2yAbcAJyPj7jPFn/H118NgF3fifNXEF4JjOILbufws2wDbgBGT8fcb4M/6+2rLv17rxP6rfAATj+23av9o1nICMv88Yf8bfV0vX7dCNf6uJKwDLdAYxbeZC2ADbgBOQ8fcZ48/4++rhx+drxT4tFF8S0D3SgrH/1xnEsESFOnXpFmyQ0XECMv4+Y/wZfx9JzcyNV+iuNTEpoHtkhBK5uifB7l+dgg00Ok5Axt9njD/j76NdvzypHfu0SCwnoHsMiST+PfnDbukMpP6ZlbCBRscJyPj7jPFn/H00c95y3djf/G72yG8HTBxp4fgRncHkxMpU14XrsMFGxgnI+PuM8Wf8fXP0/Wsqe8SDurF/O2DqSA/HGjUHI080wgYcGScg4+8zxp/x983iNS9L7DTFnjXXAAxNxHUHNLxgjOr6gFcBOAHtwvgz/j5j/AfWkfN/VMMTlfr3/8OJvICpIxKJfCv5Q69oDkq2vIQNPCpOQMbfZ4w/4++T2QtXS9x0fZqXl/dPAZNHRjjerDuwodFC1drxLmzwOQGJ8Wf8ifEfDHvfPK0yswpMXP5fFTB9ZIRGhHQHJgpKJqjO81dhvwROQGL8GX9i/AdS57mrKr9kvMRMW1owPzOQguMbyR/eY2KAk2bMVSd/+2fYL4MTkBh/xp8Y/4EgtbBu+lMmir/sAPiO1Ook80dGOFZjYpDisblLuEIgJyA8xp/x9xnjn1pSAx+d85zEypDEuECqjiFDhvxj8pecMzXY789u5JUATkBojD/j7zPGP3VOfnxTPT73eYmTqb/+z993X8k3A6k8kr9oapIShm4H8JkATkBYjD/j7zPGPyWk5t2+7G/S5ECqj/T0kn9ODybeNTlweTCw9dBp2C+LE9BfjD/j7zPGPyVP+9/xwJ+Ze/99ePVP70gPxoaZHLyQ1x9mL2iShRBgvzhOQP8w/oy/zxh/c2Q5/AXLN6pgdqF2XPUX/tG/FbAtSZkmqyA9t3YbVw3kBGT8ifFn/K0ntWzxmq2yNb7EJAViLwUG+vhOqOB/kr/8cpJKBdkM4YkfrZCthL1+W4ATkPH3GePP+NtIapZs6fvEvBV3buxj+tL/pczM/P8KDMaRForl394qOJVy4xVq6g/n924otGXfYXXg6Afqrfd+r7o/ugF7AnACuoHxZ/x9xvjfndQgqUVSk6Q2LVm3XT38+HypWfL5U+1WZig+IjCYR0YwvkgGQ0RERAMjLZSYH9A4DK4NkDiAGiQiIiKXJIv/z6X2BhCO++/P+5fkoDpRg0VEROSGRFckUvhvAaRDHkRIDu4sZsCIiIisdzYYHPafAcQjPZL/f2wCiIiIjDubEc373wDy8Z3IiP/g7QAiIiJjjqVH4/8dsOGQ+xPJAbeBBpKIiMgKvQ/85eT8a8CmIxKJ/ENGKD4n+QFuogaWiIgI1K20UHxJ79P+th7pQxPx5Af5BDTAREREaD7JCMZiARcOuXeRHo5vAQ00ERERiNhLvcv7unZkhmNZyQ94AjPoREREg6YnPTJieMDlQ/YsTg/Gp2SE4udBvwQiIqKBci5pstTGgC/HffeVfDMtmBif/OCnQb8UIiKiVO3k9056KDFOamHA5+P+obEh8rSjbG2I+mURERHpSXyWHoy3pEViOcnS940Aj795dfBbsr3hlzsMdvIVQiIistjNpLfTQomGtHAiz6vL/LrHd7NHfls6Jbk/kvR8UmtaOH4k+Z9nki4nfQ76pRMRkfs+/7IWnfmyNrVmBBNLM8KJSVK7pIYFgI+/AB9668+eB2kRAAAAAElFTkSuQmCC
            """              
                                  
            # 2. แปลง Base64 เป็น PhotoImage ที่ Tkinter เข้าใจ
            icon_data = base64.b64decode(icon_base64)
            self.icon_photo = tk.PhotoImage(data=icon_data)
            
            # 3. ตั้งค่าไอคอนให้กับหน้าต่างโปรแกรม
            self.root.iconphoto(False, self.icon_photo)
        except Exception as e:
            # ป้องกันโปรแกรมแครชหากไฟล์ไอคอนมีปัญหา
            print(f"Could not load application icon: {e}") 
        # --- จบส่วนที่เพิ่ม Icon Base64 ---        
        
        self.root.title("โปรแกรมแตก CodeNA จากไฟล์ SPSS&Excel V1")
        self.root.minsize(700, 500)

        # --- โค้ดสำหรับจัดหน้าต่างให้อยู่กลางจอ ---
        # 1. กำหนดขนาดเริ่มต้นที่ต้องการ
        initial_width = 700
        initial_height = 700 # สามารถปรับความสูงเริ่มต้นได้ตามความเหมาะสม
        
        # 2. ดึงขนาดของหน้าจอมอนิเตอร์
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # 3. คำนวณตำแหน่ง (x, y)
        position_x = int((screen_width / 2) - (initial_width / 2))
        position_y = int((screen_height / 2) - (initial_height / 2))
        
        # 4. ตั้งค่าขนาดและตำแหน่งในคำสั่งเดียว
        self.root.geometry(f"{initial_width}x{initial_height}+{position_x}+{position_y}")
        # --- สิ้นสุดโค้ดสำหรับจัดกลางจอ ---

        # --- ตั้งค่าสไตล์และสี ---
        self.setup_styles()

        # ตัวแปรสำหรับเก็บข้อมูลและวิดเจ็ต
        self.df = None
        self.metadata = None
        self.variable_widgets = {} # เก็บ widget ของแต่ละตัวแปร
        self.grouped_vars = {} # เก็บข้อมูลการจัดกลุ่ม {'representative': ['var1', 'var2']}
        self.ordered_vars = [] # เก็บรายชื่อตัวแปรตามลำดับที่แสดงผล
        self.selected_vars = [] # เก็บรายชื่อตัวแปรที่ถูกเลือก
        self.last_clicked_index = None # สำหรับ Shift+Click

        # --- สร้าง Frame หลัก ---
        main_frame = ttk.Frame(self.root, padding=(20, 20, 20, 10))
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)


        # --- ส่วนของการเลือกไฟล์ ---
        file_frame = ttk.LabelFrame(main_frame, text="1. จัดการไฟล์และการตั้งค่า", padding=15)
        file_frame.grid(row=0, column=0, sticky="ew", pady=(0, 15))
        file_frame.columnconfigure(2, weight=1) # ให้คอลัมน์ที่แสดงชื่อไฟล์ขยาย

        # Grid สำหรับปุ่ม
        button_grid = ttk.Frame(file_frame)
        button_grid.grid(row=0, column=0, rowspan=2, sticky="ns", padx=(0, 15))

        self.load_spss_button = ttk.Button(button_grid, text="โหลดไฟล์ SPSS...", command=self.load_spss_file, style="Secondary.TButton")
        self.load_spss_button.grid(row=0, column=0, sticky="ew", pady=(0, 2))
        
        self.load_excel_button = ttk.Button(button_grid, text="โหลดไฟล์ Excel...", command=self.load_excel_file, style="Secondary.TButton")
        self.load_excel_button.grid(row=1, column=0, sticky="ew", pady=(2, 0))

        self.save_settings_button = ttk.Button(button_grid, text="Save Settings", command=self.save_settings, style="Info.TButton")
        self.save_settings_button.grid(row=0, column=1, sticky="ew", padx=5)

        self.load_settings_button = ttk.Button(button_grid, text="Load Settings", command=self.load_settings, style="Info.TButton")
        self.load_settings_button.grid(row=1, column=1, sticky="ew", padx=5)

        self.file_label = ttk.Label(file_frame, text="ยังไม่ได้เลือกไฟล์", anchor="w", style="File.TLabel")
        self.file_label.grid(row=0, column=2, rowspan=2, sticky="ew", padx=(10, 0))


        # --- ส่วนของรายการตัวแปรและใส่หน่วย ---
        list_frame = ttk.LabelFrame(main_frame, text="2. เลือก (Click, Ctrl+Click, Shift+Click), จัดกลุ่ม, ใส่หน่วย และ Export", padding=15)
        list_frame.grid(row=1, column=0, sticky="nsew", pady=(0, 15))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)

        self.canvas = tk.Canvas(list_frame, borderwidth=0, background="#FFFFFF", highlightthickness=0)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=scrollbar.set)
        self.scrollable_frame = ttk.Frame(self.canvas, style="Inner.TFrame")
        self.canvas_frame = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        
        self.scrollable_frame.bind("<Configure>", self.on_frame_configure)
        self.canvas.bind('<Configure>', self.on_canvas_configure)
        
        self.canvas.bind("<MouseWheel>", self._on_mousewheel)
        self.scrollable_frame.bind("<MouseWheel>", self._on_mousewheel)
        
        self.canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        # --- ปุ่มจัดการ ---
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=2, column=0, sticky="ew", pady=(0, 10))
        button_frame.columnconfigure((0, 1), weight=1)

        self.group_button = ttk.Button(button_frame, text="Group Selected", command=self.group_selected_variables, style="Secondary.TButton")
        self.group_button.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        self.reset_button = ttk.Button(button_frame, text="Reset Grouping", command=self.reset_grouping, style="Reset.TButton")
        self.reset_button.grid(row=0, column=1, sticky="ew", padx=(5, 0))
        
        export_button = ttk.Button(main_frame, text="Export to Excel...", command=self.export_to_excel, style="Primary.TButton")
        export_button.grid(row=3, column=0, sticky="ew")

    def setup_styles(self):
        """กำหนดสไตล์สำหรับวิดเจ็ตต่างๆ"""
        # Colors
        BG_COLOR = '#ECEFF1'
        PRIMARY_COLOR = '#007BFF'
        PRIMARY_ACTIVE = '#0056b3'
        SECONDARY_COLOR = '#28A745'
        SECONDARY_ACTIVE = '#1E7E34'
        RESET_COLOR = '#6C757D'
        RESET_ACTIVE = '#5A6268'
        INFO_COLOR = '#17A2B8'
        INFO_ACTIVE = '#117A8B'
        TEXT_COLOR = '#212529'
        FRAME_BG = '#FFFFFF'
        BORDER_COLOR = '#CFD8DC'
        self.SELECTED_BG_COLOR = '#D1E7FD' # สีพื้นหลังเมื่อถูกเลือก
        
        self.root.configure(bg=BG_COLOR)
        style = ttk.Style(self.root)
        style.theme_use('clam')

        # General Styles
        style.configure('.', background=BG_COLOR, foreground=TEXT_COLOR, font=('Segoe UI', 10))
        style.configure('TFrame', background=BG_COLOR)
        style.configure('Inner.TFrame', background=FRAME_BG)
        
        # LabelFrame
        style.configure('TLabelFrame', background=BG_COLOR, bordercolor=BORDER_COLOR, padding=10)
        style.configure('TLabelFrame.Label', background=BG_COLOR, foreground=TEXT_COLOR, font=('Segoe UI', 11, 'bold'))

        # Label
        style.configure('TLabel', background=BG_COLOR, foreground=TEXT_COLOR)
        style.configure('Inner.TLabel', background=FRAME_BG)
        style.configure('File.TLabel', background=BG_COLOR, font=('Segoe UI', 10, 'italic'))
        
        # Button Styles
        style.configure('TButton', padding=8, font=('Segoe UI', 10, 'bold'), borderwidth=0)
        style.map('TButton', background=[('active', BG_COLOR)])

        style.configure('Primary.TButton', background=PRIMARY_COLOR, foreground='white')
        style.map('Primary.TButton', background=[('active', PRIMARY_ACTIVE)])
        
        style.configure('Secondary.TButton', background=SECONDARY_COLOR, foreground='white')
        style.map('Secondary.TButton', background=[('active', SECONDARY_ACTIVE)])

        style.configure('Reset.TButton', background=RESET_COLOR, foreground='white')
        style.map('Reset.TButton', background=[('active', RESET_ACTIVE)])
        
        style.configure('Info.TButton', background=INFO_COLOR, foreground='white')
        style.map('Info.TButton', background=[('active', INFO_ACTIVE)])

    def on_frame_configure(self, event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas_frame, width=event.width)

    def _on_mousewheel(self, event):
        """Cross-platform mouse wheel scrolling for the canvas."""
        if event.num == 5 or event.delta < 0:
            self.canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta > 0:
            self.canvas.yview_scroll(-1, "units")

    def clear_variable_list_ui(self, message=""):
        """ล้างรายการตัวแปรและแสดงข้อความ"""
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        self.variable_widgets.clear()
        self.grouped_vars.clear()
        self.ordered_vars.clear()
        self.selected_vars.clear()
        self.last_clicked_index = None
        if message:
            ttk.Label(self.scrollable_frame, text=message, style="Inner.TLabel", font=('Segoe UI', 10, 'italic')).pack(pady=20, padx=10)


    def load_spss_file(self):
        file_path = filedialog.askopenfilename(title="เลือกไฟล์ SPSS", filetypes=(("SPSS Data Files", "*.sav"), ("All files", "*.*")))
        if not file_path: return
        try:
            self.df, self.metadata = pyreadstat.read_sav(file_path, apply_value_formats=False, formats_as_category=False)
            self.file_label.config(text=file_path.split('/')[-1])
            
            # ค้นหาตัวแปรที่จะแสดงสำหรับ SPSS
            vars_to_display = []
            ignore_list = ['sbjnum', 'filter', 'id']
            var_types = self.metadata.original_variable_types
            vars_with_labels_applied = {var.lower() for var in self.metadata.variable_value_labels.keys()}
            for var_name, var_type_str in var_types.items():
                if var_name.lower() in ignore_list: continue
                is_numeric = 'F' in var_type_str.upper() and 'A' not in var_type_str.upper()
                if is_numeric and var_name.lower() not in vars_with_labels_applied:
                    vars_to_display.append(var_name)
            
            self.populate_variable_list(vars_to_display)

        except Exception as e:
            messagebox.showerror("เกิดข้อผิดพลาด", f"ไม่สามารถโหลดไฟล์ได้:\n{e}")
            self.df, self.metadata = None, None
            self.clear_variable_list_ui()

    def load_excel_file(self):
        """โหลดไฟล์ข้อมูลจาก Excel และรอการโหลด Settings"""
        file_path = filedialog.askopenfilename(title="เลือกไฟล์ Excel", filetypes=(("Excel Files", "*.xlsx"), ("All files", "*.*")))
        if not file_path: return
        try:
            self.df = pd.read_excel(file_path)
            self.metadata = None # ข้อมูลจาก Excel ไม่มี metadata แบบ SPSS
            self.file_label.config(text=file_path.split('/')[-1])
            # ไม่แสดงผล แต่ให้ขึ้นข้อความแนะนำ
            self.clear_variable_list_ui("โหลดไฟล์ข้อมูลสำเร็จ กรุณาโหลดไฟล์ Settings เพื่อแสดงผล")
        except Exception as e:
            messagebox.showerror("เกิดข้อผิดพลาด", f"ไม่สามารถโหลดไฟล์ได้:\n{e}")
            self.df, self.metadata = None, None
            self.clear_variable_list_ui()


    def populate_variable_list(self, vars_to_display):
        """สร้าง UI จากรายการตัวแปรที่กำหนด"""
        self.clear_variable_list_ui() # เริ่มต้นด้วยการล้าง UI ก่อนเสมอ
        self.ordered_vars = vars_to_display

        if not self.ordered_vars:
            self.clear_variable_list_ui("ไม่พบตัวแปรที่เข้าเงื่อนไข")
            return

        header_frame = ttk.Frame(self.scrollable_frame, style="Inner.TFrame", padding=(10, 5))
        header_frame.columnconfigure(0, weight=1)
        header_frame.columnconfigure(1, minsize=200) # กำหนดความกว้างขั้นต่ำให้คอลัมน์หน่วย
        header_frame.columnconfigure(2, minsize=100) # กำหนดความกว้างขั้นต่ำให้คอลัมน์ Freq
        ttk.Label(header_frame, text="ชื่อตัวแปร", font=('Segoe UI', 10, 'bold'), style="Inner.TLabel").grid(row=0, column=0, sticky="w", padx=10)
        ttk.Label(header_frame, text="หน่วย (ถ้ามี)", font=('Segoe UI', 10, 'bold'), style="Inner.TLabel").grid(row=0, column=1, sticky="w", padx=10)
        ttk.Label(header_frame, text="Frequency", font=('Segoe UI', 10, 'bold'), style="Inner.TLabel").grid(row=0, column=2, sticky="w", padx=10)
        header_frame.pack(fill=tk.X)
        header_frame.bind("<MouseWheel>", self._on_mousewheel)
        for child in header_frame.winfo_children():
            child.bind("<MouseWheel>", self._on_mousewheel)

        ttk.Separator(self.scrollable_frame, orient='horizontal').pack(fill='x', padx=5, pady=5)

        for i, var_name in enumerate(self.ordered_vars):
            row_frame = ttk.Frame(self.scrollable_frame, style="Inner.TFrame", padding=(10, 5))
            row_frame.columnconfigure(0, weight=1)
            row_frame.columnconfigure(1, minsize=200) # ใช้ minsize เดียวกันกับ header
            row_frame.columnconfigure(2, minsize=100) # ใช้ minsize เดียวกันกับ header
            
            var_label = ttk.Label(row_frame, text=var_name, anchor="w", cursor="hand2", style="Inner.TLabel")
            var_label.grid(row=0, column=0, sticky="ew", padx=10)
            
            unit_entry = ttk.Entry(row_frame)
            unit_entry.grid(row=0, column=1, sticky="ew", padx=10)
            
            freq_count = self.df[var_name].dropna().nunique()
            freq_label = ttk.Label(row_frame, text=str(freq_count), anchor="w", style="Inner.TLabel")
            freq_label.grid(row=0, column=2, sticky="w", padx=10)
            
            widgets_to_bind = [row_frame, var_label, unit_entry, freq_label]
            for widget in widgets_to_bind:
                widget.bind("<Button-1>", lambda e, v=var_name: self.handle_click(e, v))
                widget.bind("<Control-Button-1>", lambda e, v=var_name: self.handle_ctrl_click(e, v))
                widget.bind("<Shift-Button-1>", lambda e, v=var_name: self.handle_shift_click(e, v))
                widget.bind("<Double-Button-1>", lambda e, v=var_name: self.show_frequency_for_var(v))
                widget.bind("<MouseWheel>", self._on_mousewheel)

            self.variable_widgets[var_name] = {'frame': row_frame, 'label': var_label, 'entry': unit_entry, 'freq_label': freq_label}
            row_frame.pack(fill=tk.X, expand=True)
    
    def handle_click(self, event, var_name):
        self.selected_vars = [var_name]
        self.last_clicked_index = self.ordered_vars.index(var_name)
        self.update_selection_visuals()

    def handle_ctrl_click(self, event, var_name):
        if var_name in self.selected_vars:
            self.selected_vars.remove(var_name)
        else:
            self.selected_vars.append(var_name)
        self.last_clicked_index = self.ordered_vars.index(var_name)
        self.update_selection_visuals()

    def handle_shift_click(self, event, var_name):
        current_index = self.ordered_vars.index(var_name)
        if self.last_clicked_index is None:
            self.handle_click(event, var_name)
            return
        
        start = min(self.last_clicked_index, current_index)
        end = max(self.last_clicked_index, current_index)
        
        self.selected_vars = self.ordered_vars[start:end+1]
        self.update_selection_visuals()

    def update_selection_visuals(self):
        for var_name, widgets in self.variable_widgets.items():
            if widgets['frame'].winfo_exists():
                bg_color = self.SELECTED_BG_COLOR if var_name in self.selected_vars else "#FFFFFF"
                widgets['frame'].configure(style="Selected.TFrame" if var_name in self.selected_vars else "Inner.TFrame")
                widgets['label'].configure(background=bg_color)
                widgets['freq_label'].configure(background=bg_color)
    
    def group_selected_variables(self):
        if len(self.selected_vars) < 2:
            messagebox.showwarning("เลือกไม่ถูกต้อง", "กรุณาเลือกตัวแปรตั้งแต่ 2 ตัวขึ้นไปเพื่อจัดกลุ่ม")
            return

        selected_to_group = list(self.selected_vars)
        selected_to_group.sort()
        representative = selected_to_group[0]
        self.grouped_vars[representative] = selected_to_group
        
        for var_name in selected_to_group:
            if var_name != representative:
                self.variable_widgets[var_name]['frame'].pack_forget() 
        
        rep_label = self.variable_widgets[representative]['label']
        rep_label.config(text=f"{representative} (Group of {len(selected_to_group)})", font=('Segoe UI', 10, 'bold'))
        
        # Update frequency label for the group representative
        combined_series = pd.concat([self.df[col].dropna() for col in selected_to_group])
        combined_freq = combined_series.nunique()
        self.variable_widgets[representative]['freq_label'].config(text=str(combined_freq))
        
        self.selected_vars.clear()
        self.update_selection_visuals()

    def reset_grouping(self):
        if self.metadata is not None: # SPSS file
            # This is a bit inefficient, but ensures a clean state
            file_path = self.file_label.cget("text")
            self.load_spss_file() # This will re-populate
            self.file_label.config(text=file_path) # Restore text
        else: # Excel file
            self.clear_variable_list_ui("กรุณาโหลดไฟล์ Settings เพื่อแสดงผล")
        messagebox.showinfo("สำเร็จ", "รีเซ็ตรายการเรียบร้อยแล้ว")

    def show_frequency_for_var(self, var_name):
            actual_var_name = var_name
            for rep, group in self.grouped_vars.items():
                if var_name in group:
                    actual_var_name = rep
                    break
            
            if self.df is None: return
            
            vars_to_process = self.grouped_vars.get(actual_var_name, [actual_var_name])
            combined_series = pd.concat([self.df[col].dropna() for col in vars_to_process])

            freq_window = tk.Toplevel(self.root)
            title = f"Frequency: {actual_var_name}" + (f" (Group)" if actual_var_name in self.grouped_vars else "")
            freq_window.title(title)
            freq_window.geometry("400x500")

            # --- จัดหน้าต่างย่อยให้อยู่กลางจอหลัก ---
            self.root.update_idletasks()
            root_x = self.root.winfo_x()
            root_y = self.root.winfo_y()
            root_width = self.root.winfo_width()
            root_height = self.root.winfo_height()
            win_width = 400
            win_height = 500
            x = root_x + (root_width - win_width) // 2
            y = root_y + (root_height - win_height) // 2
            freq_window.geometry(f'{win_width}x{win_height}+{x}+{y}')
            
            freq_frame = ttk.Frame(freq_window, padding="10")
            freq_frame.pack(fill=tk.BOTH, expand=True)
            tree = ttk.Treeview(freq_frame, columns=('Value', 'Count'), show='headings')
            tree.heading('Value', text='ค่า (Value)'); tree.heading('Count', text='จำนวน (Frequency)')
            tree.column('Value', anchor=tk.W, width=150); tree.column('Count', anchor=tk.E, width=100)
            scrollbar = ttk.Scrollbar(freq_frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y); tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            try:
                freq_data = combined_series.value_counts().reset_index()
                freq_data.columns = ['Value', 'Count']
                freq_data = freq_data.sort_values(by='Value').reset_index(drop=True)
                
                for _, row in freq_data.iterrows():
                    # พยายามแปลงค่า Value เป็นเลขจำนวนเต็ม
                    try:
                        display_value = int(row['Value'])
                    except (ValueError, TypeError):
                        display_value = row['Value']
                    
                    # ---ปรับปรุงโค้ดตรงนี้---
                    # แปลงค่า Count ให้เป็นเลขจำนวนเต็มด้วย
                    display_count = int(row['Count'])
                    
                    # นำค่าที่แปลงแล้วไปแสดงผล
                    tree.insert('', tk.END, values=(display_value, display_count))

            except Exception as e:
                messagebox.showerror("เกิดข้อผิดพลาด", f"ไม่สามารถคำนวณ Frequency ได้:\n{e}", parent=freq_window)

    def save_settings(self):
            """บันทึกการตั้งค่าลงไฟล์ Excel โดยบันทึกเฉพาะตัวแทนของกลุ่ม"""
            if not self.variable_widgets:
                messagebox.showwarning("ไม่มีข้อมูล", "ไม่มีการตั้งค่าให้บันทึก")
                return

            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Settings Files", "*.xlsx"), ("All Files", "*.*")],
                title="บันทึกการตั้งค่า"
            )
            if not file_path: return

            try:
                settings_data = []
                
                # 1. สร้าง Set ของตัวแปรทั้งหมดที่เป็นสมาชิกในกลุ่ม (ไม่รวมตัวแทน)
                # เพื่อใช้ในการตรวจสอบและข้ามตัวแปรเหล่านี้ในภายหลัง
                vars_in_groups = set()
                for group_members in self.grouped_vars.values():
                    # เพิ่มเฉพาะสมาชิกตัวที่ 2 เป็นต้นไป (ข้ามตัวแทน)
                    vars_in_groups.update(group_members[1:])

                # 2. วนลูปทุกตัวแปรตามลำดับที่แสดงผล
                for var_name in self.ordered_vars:
                    # 3. ตรวจสอบเงื่อนไขก่อนบันทึก
                    # - ตัวแปรต้องยังแสดงอยู่บนหน้าจอ (winfo_exists())
                    # - และตัวแปรนั้นต้อง "ไม่ใช่" สมาชิกในกลุ่ม (ที่เก็บไว้ใน vars_in_groups)
                    if var_name in self.variable_widgets and self.variable_widgets[var_name]['frame'].winfo_exists():
                        
                        # ถ้าตัวแปรเป็นสมาชิกในกลุ่ม (ที่ไม่ใช่ตัวแทน) ให้ข้ามไป
                        if var_name in vars_in_groups:
                            continue
                            
                        # ดึงข้อมูลสำหรับตัวแปรเดี่ยว หรือตัวแทนกลุ่ม
                        group = self.grouped_vars.get(var_name, [var_name])
                        unit = self.variable_widgets[var_name]['entry'].get().strip()
                        group_str = ",".join(group)
                        settings_data.append([var_name, group_str, unit])

                # สร้าง DataFrame และบันทึกเป็นไฟล์ Excel
                settings_df = pd.DataFrame(settings_data, columns=['Representative', 'GroupedVars', 'Unit'])
                settings_df.to_excel(file_path, index=False)
                messagebox.showinfo("สำเร็จ", "บันทึกการตั้งค่าเรียบร้อยแล้ว")

            except Exception as e:
                messagebox.showerror("เกิดข้อผิดพลาด", f"ไม่สามารถบันทึกการตั้งค่าได้:\n{e}")

    def load_settings(self):
        """โหลดการตั้งค่าจากไฟล์ Excel และนำมาใช้กับข้อมูลปัจจุบัน"""
        if self.df is None:
            messagebox.showwarning("ไม่มีข้อมูล", "กรุณาโหลดไฟล์ข้อมูล (SPSS หรือ Excel) ก่อน")
            return
            
        file_path = filedialog.askopenfilename(
            title="เลือกไฟล์การตั้งค่า",
            filetypes=(("Settings Files", "*.xlsx"), ("All files", "*.*"))
        )
        if not file_path: return

        try:
            settings_df = pd.read_excel(file_path)
            
            # ดึงรายการตัวแปรทั้งหมดที่ต้องการจากไฟล์ setting
            all_required_vars = set()
            for group_str in settings_df['GroupedVars']:
                all_required_vars.update(group_str.split(','))

            # ตรวจสอบว่าทุกคอลัมน์ที่ต้องการมีอยู่ใน DataFrame หรือไม่
            missing_vars = all_required_vars - set(self.df.columns)
            if missing_vars:
                messagebox.showwarning("ตัวแปรไม่ตรงกัน", f"ไม่พบตัวแปรต่อไปนี้ในไฟล์ข้อมูล:\n{', '.join(missing_vars)}")
                return

            # แสดงผลเฉพาะตัวแปร Representative จากไฟล์ setting
            vars_to_display = settings_df['Representative'].tolist()
            self.populate_variable_list(vars_to_display)
            
            # ใช้การตั้งค่า
            for index, row in settings_df.iterrows():
                rep_var = row['Representative']
                group_str = row['GroupedVars']
                unit = row['Unit'] if pd.notna(row['Unit']) else ""

                if rep_var not in self.variable_widgets:
                    continue 

                self.variable_widgets[rep_var]['entry'].delete(0, tk.END)
                self.variable_widgets[rep_var]['entry'].insert(0, unit)
                
                group_members = group_str.split(',')
                if len(group_members) > 1:
                    self.grouped_vars[rep_var] = group_members
                    rep_label = self.variable_widgets[rep_var]['label']
                    rep_label.config(text=f"{rep_var} (Group of {len(group_members)})", font=('Segoe UI', 10, 'bold'))
                    
                    # Update frequency for the newly formed group
                    combined_series = pd.concat([self.df[col].dropna() for col in group_members])
                    combined_freq = combined_series.nunique()
                    self.variable_widgets[rep_var]['freq_label'].config(text=str(combined_freq))

            self.selected_vars.clear()
            self.update_selection_visuals()
            messagebox.showinfo("สำเร็จ", "โหลดการตั้งค่าเรียบร้อยแล้ว")

        except Exception as e:
            messagebox.showerror("เกิดข้อผิดพลาด", f"ไม่สามารถโหลดการตั้งค่าได้:\n{e}")


    def export_to_excel(self):
        if not self.variable_widgets:
            messagebox.showwarning("ไม่มีข้อมูล", "ไม่มีตัวแปรให้ Export")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="บันทึกไฟล์ Excel")
        if not file_path: return

        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                items_to_export = []
                processed_vars = set()
                visible_vars = [var for var in self.ordered_vars if self.variable_widgets[var]['frame'].winfo_exists()]
                for var_name in visible_vars:
                    if var_name in processed_vars: continue
                    items_to_export.append(var_name)
                    if var_name in self.grouped_vars:
                        processed_vars.update(self.grouped_vars[var_name])
                    else:
                        processed_vars.add(var_name)
                
                index_data = []
                for i, var_name in enumerate(items_to_export, 1):
                    vars_to_process = self.grouped_vars.get(var_name, [var_name])
                    combined_series = pd.concat([self.df[col].dropna() for col in vars_to_process if col in self.df.columns])
                    freq_count = combined_series.nunique()
                    index_data.append([i, var_name, freq_count])
                
                index_df = pd.DataFrame(index_data, columns=['No', 'Sheet Name', 'Frequency'])
                index_df.to_excel(writer, sheet_name='Index', index=False)
                
                column_labels_map = {}
                if self.metadata:
                    column_labels_map = dict(zip(self.metadata.column_names, self.metadata.column_labels))

                for var_name in items_to_export:
                    vars_to_process = self.grouped_vars.get(var_name, [var_name])
                    combined_series = pd.concat([self.df[col].dropna() for col in vars_to_process if col in self.df.columns])
                    
                    var_label = column_labels_map.get(var_name, var_name)
                    unit = self.variable_widgets[var_name]['entry'].get().strip()

                    var_data = []
                    var_data.append([var_name, var_label, ''])
                    var_data.append(['ID', 'Category Label', 'Categorization Condition'])

                    freq = combined_series.value_counts().sort_index()
                    if freq.empty:
                        var_data.append(['(No data)', '', ''])
                    else:
                        for value, _ in freq.items():
                            try: display_value = int(value)
                            except (ValueError, TypeError): display_value = value
                            
                            category_label = f"{display_value} {unit}" if unit else str(display_value)
                            condition = f"{var_name}={display_value}"
                            var_data.append([display_value, category_label, condition])
                    
                    var_df = pd.DataFrame(var_data)
                    sheet_name = var_name[:31] 
                    var_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                workbook = writer.book
                index_sheet = workbook['Index']
                
                var_header_font = Font(bold=True, color="FFFFFF")
                var_header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
                index_header_font = Font(bold=True)
                index_header_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                white_font = Font(color="FFFFFF")
                white_hyperlink_font = Font(color="FFFFFF", underline="single")
                
                index_sheet.column_dimensions['A'].width = 5
                index_sheet.column_dimensions['B'].width = 40
                index_sheet.column_dimensions['C'].width = 15
                for cell in index_sheet[1]:
                    cell.font = index_header_font
                    cell.fill = index_header_fill

                for r_idx, row in index_df.iterrows():
                    excel_row_num = r_idx + 2
                    sheet_name = row['Sheet Name']
                    freq_count = row['Frequency']
                    
                    safe_sheet_name = sheet_name[:31]
                    link_target = f"#'{safe_sheet_name}'!A1"
                    cell_to_link = index_sheet.cell(row=excel_row_num, column=2)
                    cell_to_link.hyperlink = link_target

                    if freq_count == 0:
                        for col_num in range(1, 4):
                            cell = index_sheet.cell(row=excel_row_num, column=col_num)
                            cell.fill = black_fill
                            cell.font = white_font
                        cell_to_link.font = white_hyperlink_font
                        workbook[safe_sheet_name].sheet_properties.tabColor = "000000"
                    else:
                        cell_to_link.style = 'Hyperlink'

                for sheet_name in items_to_export:
                    safe_sheet_name = sheet_name[:31]
                    var_sheet = workbook[safe_sheet_name]
                    for row in var_sheet['A1:C2']:
                        for cell in row:
                            cell.font = var_header_font
                            cell.fill = var_header_fill
                    var_sheet.column_dimensions['A'].width = 12
                    var_sheet.column_dimensions['B'].width = 45
                    var_sheet.column_dimensions['C'].width = 35
                    var_sheet.column_dimensions['D'].width = 15
                    var_sheet.cell(row=1, column=4).value = 'Back to Index'
                    var_sheet.cell(row=1, column=4).hyperlink = "#'Index'!A1"
                    var_sheet.cell(row=1, column=4).style = 'Hyperlink'
                    var_sheet.freeze_panes = var_sheet['B4']

            messagebox.showinfo("สำเร็จ", f"ไฟล์ถูกบันทึกเรียบร้อยแล้วที่:\n{file_path}")

        except Exception as e:
            messagebox.showerror("เกิดข้อผิดพลาด", f"ไม่สามารถ Export ไฟล์ Excel ได้:\n{e}")



# <<< START OF CHANGES >>>
# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
def run_this_app(working_dir=None): # ชื่อฟังก์ชันนี้จะถูกใช้ใน Launcher
    """
    ฟังก์ชันหลักสำหรับสร้างและรัน QuotaSamplerApp.
    """
    print(f"--- QUOTA_SAMPLER_INFO: Starting 'QuotaSamplerApp' via run_this_app() ---")
    try:
        # --- โค้ดที่ย้ายมาจาก if __name__ == "__main__": เดิมจะมาอยู่ที่นี่ ---
    #if __name__ == '__main__':
        root = tk.Tk()
        app = SPSSAnalyzerApp(root)
        root.mainloop()

        print(f"--- QUOTA_SAMPLER_INFO: QuotaSamplerApp mainloop finished. ---")

    except Exception as e:
        # ดักจับ Error ที่อาจเกิดขึ้นระหว่างการสร้างหรือรัน App
        print(f"QUOTA_SAMPLER_ERROR: An error occurred during QuotaSamplerApp execution: {e}")
        # แสดง Popup ถ้ามีปัญหา
        if 'root' not in locals() or not root.winfo_exists(): # สร้าง root ชั่วคราวถ้ายังไม่มี
            root_temp = tk.Tk()
            root_temp.withdraw()
            messagebox.showerror("Application Error (Quota Sampler)",
                               f"An unexpected error occurred:\n{e}", parent=root_temp)
            root_temp.destroy()
        else:
            messagebox.showerror("Application Error (Quota Sampler)",
                               f"An unexpected error occurred:\n{e}", parent=root) # ใช้ root ที่มีอยู่ถ้าเป็นไปได้
        sys.exit(f"Error running QuotaSamplerApp: {e}") # อาจจะ exit หรือไม่ก็ได้ ขึ้นกับการออกแบบ


# --- ส่วน Run Application เมื่อรันไฟล์นี้โดยตรง (สำหรับ Test) ---
if __name__ == "__main__":
    print("--- Running QuotaSamplerApp.py directly for testing ---")
    # (ถ้ามีการตั้งค่า DPI ด้านบน มันจะทำงานอัตโนมัติ)

    # เรียกฟังก์ชัน Entry Point ที่เราสร้างขึ้น
    run_this_app()

    print("--- Finished direct execution of QuotaSamplerApp.py ---")
# <<< END OF CHANGES >>>