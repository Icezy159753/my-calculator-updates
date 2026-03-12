"""
Well-being Index Calculator — PyQt6 Desktop Application
คำนวณ Well-being Index จากไฟล์ SPSS (.sav) อัตโนมัติ
"""
import sys, os, re, traceback, warnings, json, itertools
from datetime import datetime
from dataclasses import dataclass, field
from typing import Any, Optional
import numpy as np
import pandas as pd
import pyreadstat
from scipy.stats import pearsonr
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QFileDialog, QLabel, QComboBox, QListWidget,
    QListWidgetItem, QMessageBox, QAbstractItemView, QGroupBox,
    QLineEdit, QSplitter, QStackedWidget, QTableWidget,
    QTableWidgetItem, QHeaderView, QTextEdit, QCheckBox,
    QProgressBar, QTabWidget, QScrollArea, QSpinBox, QGridLayout,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QLocale
from PyQt6.QtGui import QIcon, QFont, QColor, QDragEnterEvent, QDropEvent
from openpyxl.styles import Font as XlFont, Alignment, PatternFill, Border, Side, Protection

def _resource_path(relative_path: str) -> str:
    base_path = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

# ══════════════════════════════════════════════
#  Data Models
# ══════════════════════════════════════════════
DEFAULT_DIMENSIONS = [
    {"name": "Physical well-being",    "short": "Physical",      "prefix": "PWQ2",  "n_items": 5,  "q3_prefix": "PWQ3"},
    {"name": "Mental",                 "short": "Mental",        "prefix": "MWQ2",  "n_items": 4,  "q3_prefix": "MWQ4"},
    {"name": "Social",                 "short": "Social",        "prefix": "SWQ2",  "n_items": 8,  "q3_prefix": "SWQ3"},
    {"name": "Financial",              "short": "Financial",     "prefix": "FWQ2",  "n_items": 10, "q3_prefix": "FWQ3"},
    {"name": "Environmental",          "short": "Environmental", "prefix": "EWQ2",  "n_items": 7,  "q3_prefix": "EWQ3"},
    {"name": "Purpose & spiritual",    "short": "Purpose & spiritual", "prefix": "PSWQ2", "n_items": 6, "q3_prefix": "PSWQ3"},
]

@dataclass
class DimensionDef:
    name: str
    short_name: str
    prefix: str
    q2_variables: list
    q3_prefix: str
    q3_variables: list

@dataclass
class SubgroupDef:
    name: str
    filters: dict = field(default_factory=dict)
    main_group: str = ""
    filter_group: str = ""

@dataclass
class IndexResult:
    subgroup_name: str
    subgroup_main: str
    subgroup_filter: str
    index_value: float
    dim_means: dict
    dim_weights: dict
    regression_stats: dict
    n: int
    std_betas: dict = field(default_factory=dict)
    factor_output_log: str = ""
    rotated_component_matrix: Any = None
    coefficients_table: Any = None

# ══════════════════════════════════════════════
#  SPSS-compatible Orthomax Rotation
# ══════════════════════════════════════════════
def _orthomax_objective(normalized_loadings: np.ndarray, gamma: float) -> float:
    """Orthomax objective on Kaiser-normalized loadings (higher is better)."""
    p = normalized_loadings.shape[0]
    col_ss = np.sum(normalized_loadings ** 2, axis=0)
    return float(np.sum(np.sum(normalized_loadings ** 4, axis=0) - (gamma / p) * (col_ss ** 2)))


def _align_rotated_components_to_unrotated(unrotated: np.ndarray, rotated: np.ndarray) -> np.ndarray:
    """
    Canonicalize rotated component order against the original PCA component order.
    This removes equivalent column permutations that SPSS does not expose as renumbered factors.
    """
    k = rotated.shape[1]
    transform = np.linalg.pinv(unrotated) @ rotated
    best_perm = tuple(range(k))
    best_score = -np.inf

    for perm in itertools.permutations(range(k)):
        score = float(sum(abs(transform[i, perm[i]]) for i in range(k)))
        if score > best_score:
            best_score = score
            best_perm = perm

    return rotated[:, list(best_perm)]


def _normalize_dim_name(name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", " ", str(name).lower()).strip()
    if "environ" in s:
        return "environmental"
    if "mental" in s:
        return "mental"
    if "financ" in s:
        return "financial"
    if "social" in s:
        return "social"
    if "physical" in s:
        return "physical"
    if "purpose" in s or "spiritual" in s:
        return "purpose_spiritual"
    return s


def _brandspace_reference_order_indices(variable_names: list[str]) -> list[int] | None:
    expected = ["environmental", "mental", "financial", "social", "physical", "purpose_spiritual"]
    key_to_idx = {}
    for idx, name in enumerate(variable_names):
        key = _normalize_dim_name(name)
        if key in expected and key not in key_to_idx:
            key_to_idx[key] = idx
    if not all(k in key_to_idx for k in expected):
        return None
    return [key_to_idx[k] for k in expected]


def _reorder_components_for_brandspace(rotated: np.ndarray, variable_names: list[str]) -> np.ndarray:
    """
    Apply canonical component numbering for the BrandSpace 6-dimension matrix.
    Uses fuzzy name matching so minor label differences still map correctly.
    """
    target_rows = _brandspace_reference_order_indices(variable_names)
    if target_rows is None or rotated.shape[1] != len(target_rows):
        return rotated

    best_perm = tuple(range(rotated.shape[1]))
    best_score = -np.inf
    for perm in itertools.permutations(range(rotated.shape[1])):
        score = 0.0
        for j, row_idx in enumerate(target_rows):
            score += abs(float(rotated[row_idx, perm[j]]))
        if score > best_score:
            best_score = score
            best_perm = perm

    return rotated[:, list(best_perm)]


def _get_rotated_matrix_row_order(loadings: np.ndarray, variable_names: list[str]) -> list[int]:
    reference_rows = _brandspace_reference_order_indices(variable_names)
    if reference_rows is not None:
        return reference_rows

    row_order = []
    for j in range(loadings.shape[1]):
        max_idx = int(np.argmax(np.abs(loadings[:, j])))
        if max_idx not in row_order:
            row_order.append(max_idx)
    for idx in range(loadings.shape[0]):
        if idx not in row_order:
            row_order.append(idx)
    return row_order


def _spss_orthomax_rotation(loadings, gamma=1.0, max_iter=250, tol=1e-4, multi_start=False):
    """
    SPSS-compatible orthomax rotation using Jacobi pairwise algorithm
    with Kaiser normalization.

    Convergence: max absolute change in loading matrix < tol (SPSS criterion).
    gamma=0:   quartimax
    gamma=1:   varimax
    gamma=p/2: equamax  (p = number of variables)
    """
    p, k = loadings.shape

    # Kaiser normalization: divide each row by sqrt(communality)
    communalities = np.sum(loadings ** 2, axis=1)
    h = np.sqrt(communalities)
    h[h < 1e-12] = 1.0
    A_base = loadings.copy() / h[:, np.newaxis]

    # For <= 6 factors (the current app default), exhaustive sign starts are small (<=64).
    sign_starts = [np.ones(k)]
    if multi_start:
        sign_starts = []
        for signs in itertools.product([-1.0, 1.0], repeat=k):
            sign_starts.append(np.array(signs, dtype=float))

    best_A = None
    best_iter = 0
    best_obj = -np.inf
    best_start_idx = 0

    for start_idx, signs in enumerate(sign_starts):
        A = A_base.copy() * signs[np.newaxis, :]
        n_iter = 0
        for iteration in range(max_iter):
            n_iter = iteration + 1
            A_prev = A.copy()

            for j in range(k - 1):
                for m in range(j + 1, k):
                    u = A[:, j] ** 2 - A[:, m] ** 2
                    v = 2.0 * A[:, j] * A[:, m]

                    aa = np.sum(u)
                    bb = np.sum(v)
                    cc = np.sum(u ** 2 - v ** 2)
                    dd = np.sum(2.0 * u * v)

                    num = dd - 2.0 * gamma * aa * bb / p
                    den = cc - gamma * (aa ** 2 - bb ** 2) / p

                    if abs(num) < 1e-15 and abs(den) < 1e-15:
                        continue

                    theta = 0.25 * np.arctan2(num, den)

                    cos_t = np.cos(theta)
                    sin_t = np.sin(theta)
                    col_j = A[:, j].copy()
                    col_m = A[:, m].copy()
                    A[:, j] = cos_t * col_j + sin_t * col_m
                    A[:, m] = -sin_t * col_j + cos_t * col_m

            # SPSS convergence: max absolute change in loading matrix
            max_change = np.max(np.abs(A - A_prev))
            if max_change < tol:
                break

        obj = _orthomax_objective(A, gamma)
        if (obj > best_obj + 1e-12) or (abs(obj - best_obj) <= 1e-12 and n_iter < best_iter):
            best_obj = obj
            best_A = A.copy()
            best_iter = n_iter
            best_start_idx = start_idx

    # Denormalize
    rotated = best_A * h[:, np.newaxis]
    return rotated, best_iter, best_obj, len(sign_starts), best_start_idx


# ══════════════════════════════════════════════
#  Processing Pipeline
# ══════════════════════════════════════════════
class IndexCalculationPipeline:
    def __init__(self, df: pd.DataFrame, dimensions: list, q1_var: str = "Q1"):
        self.df = df
        self.dimensions = dimensions
        self.q1_var = q1_var

    def recode_q1(self, sub_df: pd.DataFrame) -> pd.DataFrame:
        if self.q1_var in sub_df.columns:
            sub_df = sub_df.copy()
            sub_df["NQ1"] = 8 - pd.to_numeric(sub_df[self.q1_var], errors="coerce")
        return sub_df

    def compute_dim_means(self, sub_df: pd.DataFrame) -> pd.DataFrame:
        dim_df = pd.DataFrame(index=sub_df.index)
        for dim in self.dimensions:
            cols = [c for c in dim.q2_variables if c in sub_df.columns]
            if cols:
                dim_df[dim.short_name] = sub_df[cols].apply(
                    lambda r: np.nanmean(r.values) if r.notna().any() else np.nan, axis=1
                )
            else:
                dim_df[dim.short_name] = np.nan
        return dim_df

    def run_factor_analysis(self, dim_df: pd.DataFrame, rotation: str = "equamax"):
        clean = dim_df.dropna()
        if len(clean) < 30:
            return None, clean.index, "Not enough data for Factor Analysis\n", {}
        n_vars = len(dim_df.columns)
        n_factors = min(6, n_vars)
        log_lines = []
        log_lines.append("--- Factor Analysis ---")
        log_lines.append(f"Data rows: {len(clean)}")
        try:
            from sklearn.preprocessing import StandardScaler
            from scipy.linalg import inv, eigh

            # ── Step 1: PCA extraction from correlation matrix ──
            R = clean.corr().values
            eigenvalues_all, eigenvectors_all = np.linalg.eigh(R)
            # Sort descending
            sort_idx = np.argsort(eigenvalues_all)[::-1]
            eigenvalues_all = eigenvalues_all[sort_idx]
            eigenvectors_all = eigenvectors_all[:, sort_idx]

            # Unrotated loadings = eigenvectors * sqrt(eigenvalues)
            eigenvalues = eigenvalues_all[:n_factors]
            eigenvectors = eigenvectors_all[:, :n_factors]
            unrotated = eigenvectors * np.sqrt(np.maximum(eigenvalues, 0))

            # Keep raw eigenvector signs for rotation start.
            # For orthogonal rotation, different sign starts can converge to different local optima.
            unrotated_for_rotation = unrotated.copy()

            log_lines.append(f"Rotation: {rotation.capitalize()} with Kaiser Normalization")
            log_lines.append(f"N Factors: {n_factors}")
            log_lines.append(f"Variables: {list(clean.columns)}")

            # Eigenvalues & Variance Explained
            log_lines.append("")
            log_lines.append("Total Variance Explained:")
            log_lines.append(f"  {'Component':<12} {'Eigenvalue':>12} {'% Variance':>12} {'Cumulative%':>12}")
            cum = 0.0
            total_var = float(n_vars)  # for correlation matrix, total = n_vars
            for i_ev, eigenval in enumerate(eigenvalues_all):
                pct = eigenval / total_var * 100
                cum += pct
                marker = " *" if i_ev < n_factors else ""
                log_lines.append(f"  {i_ev+1:<12} {eigenval:>12.4f} {pct:>11.2f}% {cum:>11.2f}%{marker}")

            # ── Step 2: Rotation (SPSS-compatible Jacobi pairwise) ──
            gamma_map = {
                "equamax": n_vars / 2.0,
                "varimax": 1.0,
                "quartimax": 0.0,
            }

            if rotation in gamma_map:
                gamma = gamma_map[rotation]
                L, n_rot_iter, _, _, _ = _spss_orthomax_rotation(
                    unrotated_for_rotation, gamma=gamma, max_iter=250, tol=1e-5, multi_start=False
                )
                log_lines.append(f"a. Rotation converged in {n_rot_iter} iterations.")

                if rotation == "equamax":
                    L = _reorder_components_for_brandspace(L, list(clean.columns))
                else:
                    L = _align_rotated_components_to_unrotated(unrotated, L)
            else:
                # Fallback to factor_analyzer for promax, oblimin
                from factor_analyzer import FactorAnalyzer
                rot_kwargs = {'max_iter': 250}
                fa = FactorAnalyzer(
                    n_factors=n_factors, rotation=rotation,
                    method="principal", rotation_kwargs=rot_kwargs
                )
                fa.fit(clean)
                L = fa.loadings_
                log_lines.append(f"Rotation: {rotation.capitalize()}")

            # ── Step 3: SPSS sign convention ──
            for col_idx in range(L.shape[1]):
                max_abs_idx = np.argmax(np.abs(L[:, col_idx]))
                if L[max_abs_idx, col_idx] < 0:
                    L[:, col_idx] *= -1

            # ── Step 4: Map each factor to its dominant dimension ──
            factor_dim_names = []
            for j in range(L.shape[1]):
                max_idx = np.argmax(np.abs(L[:, j]))
                factor_dim_names.append(clean.columns[max_idx])

            # ── Build Rotated Component Matrix DataFrame ──
            comp_cols = [f'Component {j+1}' for j in range(L.shape[1])]
            rotated_df = pd.DataFrame(L.copy(), index=list(clean.columns), columns=comp_cols)

            # Log: Rotated Component Matrix (FORMAT SORT BLANK(.4))
            log_lines.append("")
            log_lines.append("Rotated Component Matrix:")
            header = f"  {'Variable':<22}" + "".join([f"{'Comp'+str(j+1):>10}" for j in range(L.shape[1])])
            log_lines.append(header)
            row_order = _get_rotated_matrix_row_order(L, list(clean.columns))
            for i_row in row_order:
                col_name = clean.columns[i_row]
                vals = []
                for j in range(L.shape[1]):
                    v = L[i_row, j]
                    if abs(v) < 0.4:
                        vals.append(f"{'':>10}")
                    else:
                        vals.append(f"{v:>10.3f}")
                log_lines.append(f"  {col_name:<22}" + "".join(vals))
            log_lines.append("  Extraction Method: Principal Component Analysis.")
            log_lines.append(f"  Rotation Method: {rotation.capitalize()} with Kaiser Normalization.")

            # SS Loadings (component order is preserved as in SPSS output)
            ss_loadings = np.sum(L ** 2, axis=0)
            log_lines.append("")
            log_lines.append("  SS Loadings:          " + "".join(
                [f"{ss_loadings[j]:>10.4f}" for j in range(L.shape[1])]))

            # Communalities
            communalities = np.sum(L ** 2, axis=1)
            log_lines.append("")
            log_lines.append("Communalities:")
            log_lines.append(f"  {'Variable':<22} {'Extraction':>12}")
            for i_c, col_name in enumerate(clean.columns):
                log_lines.append(f"  {col_name:<22} {communalities[i_c]:>12.4f}")
            log_lines.append("")

            # ── Step 5: Anderson-Rubin factor scoring ──
            Z = StandardScaler().fit_transform(clean)
            inv_R = inv(R)
            temp_matrix = L.T @ inv_R @ L
            eigvals, eigvecs = eigh(temp_matrix)
            inv_sqrt_eigvals_arr = np.zeros_like(eigvals)
            positive_mask = eigvals > 1e-12
            inv_sqrt_eigvals_arr[positive_mask] = 1.0 / np.sqrt(eigvals[positive_mask])
            inv_sqrt_temp = eigvecs @ np.diag(inv_sqrt_eigvals_arr) @ eigvecs.T
            C_AR = inv_R @ L @ inv_sqrt_temp
            factor_scores = Z @ C_AR

            fa_extra = {"rotated_df": rotated_df, "factor_dim_names": factor_dim_names}
            return factor_scores, clean.index, "\n".join(log_lines) + "\n", fa_extra
        except Exception as e:
            return None, clean.index, f"Factor Analysis Error: {e}\n{traceback.format_exc()}", {}

    def run_regression(self, nq1: pd.Series, factor_scores: np.ndarray,
                       clean_dims: pd.DataFrame, factor_dim_names: list = None):
        try:
            import statsmodels.api as sm
            valid = nq1.notna()
            y = nq1[valid]
            X = factor_scores[valid.values[valid.index.isin(nq1.index)]] if len(factor_scores) != len(y) else factor_scores
            if len(y) < 10 or len(y) != len(X):
                return None, "Not enough valid data for Regression (-NQ1 vs Factors)\n"

            log_lines = []
            log_lines.append("=" * 70)
            log_lines.append("--- Regression Analysis (NQ1 on Factor Scores) ---")
            log_lines.append("=" * 70)
            log_lines.append(f"Valid N: {len(y)}")

            # Use dimension names mapped from factor analysis (SPSS-compatible labeling)
            if factor_dim_names and len(factor_dim_names) == X.shape[1]:
                fac_names = list(factor_dim_names)
            else:
                fac_names = [f'FAC{i+1}' for i in range(X.shape[1])]
            X_df = pd.DataFrame(X, index=y.index, columns=fac_names)
            X_const = sm.add_constant(X_df)
            model = sm.OLS(y, X_const).fit()

            # Full OLS Summary
            log_lines.append(str(model.summary()))

            # --- Standardized Beta ---
            unstandardized_coeffs = model.params.drop('const')
            std_betas_series = unstandardized_coeffs * (X_df.std() / y.std())

            # --- Collinearity Statistics ---
            try:
                corr_matrix = np.corrcoef(X, rowvar=False)
                from numpy.linalg import inv as np_inv
                inv_corr = np_inv(corr_matrix)
                vif_values = np.diag(inv_corr)
                tolerance_values = 1.0 / vif_values
            except Exception:
                vif_values = np.ones(X.shape[1])
                tolerance_values = np.ones(X.shape[1])

            # --- Full SPSS-style Coefficients Table ---
            log_lines.append("")
            log_lines.append("=" * 70)
            log_lines.append("Coefficients Table (SPSS-style)")
            log_lines.append("=" * 70)

            unstd_b = model.params
            std_err = model.bse
            t_vals = model.tvalues
            p_vals = model.pvalues
            ci = model.conf_int()

            header = f"  {'Variable':<22} {'B':>10} {'Std.Err':>10} {'Beta':>10} {'t':>10} {'Sig.':>10} {'Tolerance':>10} {'VIF':>10}"
            log_lines.append(header)
            log_lines.append("  " + "-" * 92)

            # Constant row
            log_lines.append(f"  {'(Constant)':<22} {unstd_b['const']:>10.4f} {std_err['const']:>10.4f} {'':>10} {t_vals['const']:>10.4f} {p_vals['const']:>10.4f} {'':>10} {'':>10}")

            # Build Coefficients DataFrame for Excel export
            coeff_rows = [{
                "Variable": "(Constant)",
                "B": round(float(unstd_b['const']), 4),
                "Std. Error": round(float(std_err['const']), 4),
                "Beta": "",
                "t": round(float(t_vals['const']), 4),
                "Sig.": round(float(p_vals['const']), 4),
                "Tolerance": "",
                "VIF": "",
            }]

            # Factor rows
            for i_fac, fac in enumerate(fac_names):
                beta_val = std_betas_series[fac]
                tol_val = float(tolerance_values[i_fac])
                vif_val = float(vif_values[i_fac])
                log_lines.append(f"  {fac:<22} {unstd_b[fac]:>10.4f} {std_err[fac]:>10.4f} {beta_val:>10.4f} {t_vals[fac]:>10.4f} {p_vals[fac]:>10.4f} {tol_val:>10.3f} {vif_val:>10.3f}")
                coeff_rows.append({
                    "Variable": fac,
                    "B": round(float(unstd_b[fac]), 4),
                    "Std. Error": round(float(std_err[fac]), 4),
                    "Beta": round(float(beta_val), 4),
                    "t": round(float(t_vals[fac]), 4),
                    "Sig.": round(float(p_vals[fac]), 4),
                    "Tolerance": round(tol_val, 3),
                    "VIF": round(vif_val, 3),
                })

            coefficients_df = pd.DataFrame(coeff_rows)
            log_lines.append("")

            # Model Summary
            log_lines.append("Model Summary:")
            log_lines.append(f"  R²     = {model.rsquared:.6f}")
            log_lines.append(f"  Adj R² = {model.rsquared_adj:.6f}")
            log_lines.append(f"  F      = {model.fvalue:.4f}  (p = {model.f_pvalue:.6f})")
            log_lines.append(f"  Std.Err of Estimate = {np.sqrt(model.mse_resid):.6f}")
            log_lines.append("")

            # Weight summary
            std_betas_arr = std_betas_series.values
            abs_betas = np.abs(std_betas_arr)
            total_abs = abs_betas.sum() if abs_betas.sum() > 0 else 1.0
            weights = abs_betas / total_abs

            log_lines.append("Weight Calculation (from Standardized Beta):")
            log_lines.append(f"  {'Dimension':<22} {'Std Beta':>12} {'|Beta|':>12} {'Weight':>12}")
            log_lines.append("  " + "-" * 58)
            for i_b, fac in enumerate(fac_names):
                log_lines.append(f"  {fac:<22} {std_betas_arr[i_b]:>12.6f} {abs_betas[i_b]:>12.6f} {weights[i_b]:>12.6f}")
            log_lines.append(f"  {'Sum':<22} {'':>12} {total_abs:>12.6f} {weights.sum():>12.6f}")
            log_lines.append("")

            # Build std_betas_dict: dimension_name -> beta value
            std_betas_dict = {}
            for i_fac, fac in enumerate(fac_names):
                std_betas_dict[fac] = float(std_betas_arr[i_fac])

            return {
                "std_betas": std_betas_arr,
                "std_betas_dict": std_betas_dict,
                "r_squared": model.rsquared,
                "adj_r_squared": model.rsquared_adj,
                "f_statistic": model.fvalue,
                "f_pvalue": model.f_pvalue,
                "params": model.params,
                "coefficients_df": coefficients_df,
            }, "\n".join(log_lines) + "\n"
        except Exception as e:
            return None, f"Regression Error: {e}\n"

    def compute_weights(self, std_betas: np.ndarray) -> np.ndarray:
        abs_betas = np.abs(std_betas)
        total = abs_betas.sum()
        if total == 0:
            return np.ones(len(std_betas)) / len(std_betas)
        return abs_betas / total

    def compute_index(self, dim_df: pd.DataFrame, weights: np.ndarray, spss_missing: bool = True) -> pd.Series:
        if len(dim_df.columns) == 0:
            return pd.Series(np.nan, index=dim_df.index)
        weighted = dim_df.mul(np.asarray(weights), axis=1)
        if spss_missing:
            # SPSS-like COMPUTE: if any referenced Dim is missing, Index becomes missing.
            return weighted.sum(axis=1, min_count=len(dim_df.columns))
        return weighted.fillna(0).sum(axis=1)

    def run_for_subgroup(self, subgroup: SubgroupDef, mask: pd.Series, rotation: str = "equamax") -> Optional[IndexResult]:
        sub_df = self.df[mask].copy()
        sub_df = self.recode_q1(sub_df)
        dim_df = self.compute_dim_means(sub_df)
        clean_dim = dim_df.dropna()
        if len(clean_dim) < 30:
            return None, "Not enough clean data for Factor Analysis\n"
        factor_scores, valid_idx, fa_log, fa_extra = self.run_factor_analysis(dim_df, rotation)
        if factor_scores is None:
            return None, fa_log

        factor_dim_names = fa_extra.get("factor_dim_names", [])
        rotated_df = fa_extra.get("rotated_df", None)

        nq1 = sub_df.loc[valid_idx, "NQ1"] if "NQ1" in sub_df.columns else None
        if nq1 is None or nq1.notna().sum() < 10:
            return None, "Q1 data is insufficient"
        reg_result, reg_log = self.run_regression(nq1, factor_scores, clean_dim.loc[valid_idx], factor_dim_names)
        if reg_result is None:
            return None, reg_log

        full_log = f"--- Analysis for Subgroup: {subgroup.name} ---\n{fa_log}{reg_log}\n{'='*50}\n"

        # Map betas and weights by dimension name (via factor-to-dimension mapping)
        std_betas_dict = reg_result.get("std_betas_dict", {})
        beta_dict = {}
        weight_dict = {}
        for d in self.dimensions:
            beta_dict[d.short_name] = std_betas_dict.get(d.short_name, 0.0)
            weight_dict[d.short_name] = abs(std_betas_dict.get(d.short_name, 0.0))

        # Normalize weights
        total_abs = sum(weight_dict.values())
        if total_abs > 0:
            for k in weight_dict:
                weight_dict[k] /= total_abs
        else:
            n_dims = len(weight_dict)
            for k in weight_dict:
                weight_dict[k] = 1.0 / n_dims if n_dims > 0 else 0.0

        # Keep Mean Score as subgroup-level means (same as original display),
        # then compute Index from those means and weights.
        mean_scores = {}
        for col in dim_df.columns:
            mean_scores[col] = float(dim_df[col].mean())

        index_value = 0.0
        for col in dim_df.columns:
            score = mean_scores.get(col, 0.0)
            if pd.isna(score):
                score = 0.0
            index_value += score * weight_dict.get(col, 0.0)

        return IndexResult(
            subgroup_name=subgroup.name,
            subgroup_main=subgroup.main_group or subgroup.name,
            subgroup_filter=subgroup.filter_group or "",
            index_value=float(index_value),
            dim_means=mean_scores,
            dim_weights=weight_dict,
            regression_stats={
                "r_squared": reg_result["r_squared"],
                "adj_r_squared": reg_result["adj_r_squared"],
                "f_statistic": reg_result["f_statistic"],
                "f_pvalue": reg_result["f_pvalue"],
            },
            n=int(mask.sum()),
            std_betas=beta_dict,
            factor_output_log=full_log,
            rotated_component_matrix=rotated_df,
            coefficients_table=reg_result.get("coefficients_df"),
        ), None

    def build_mask(self, subgroup: SubgroupDef) -> pd.Series:
        mask = pd.Series(True, index=self.df.index)
        for col, val in subgroup.filters.items():
            if col in self.df.columns:
                if pd.isna(val):
                    mask &= self.df[col].isna()
                else:
                    mask &= self.df[col] == val
        return mask

def compute_correlations(index_series: pd.Series, items_df: pd.DataFrame):
    results = []
    for col in items_df.columns:
        valid = index_series.notna() & items_df[col].notna()
        n = int(valid.sum())
        if n > 2:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                try:
                    r, p = pearsonr(index_series[valid], items_df[col][valid])
                except Exception:
                    r, p = np.nan, np.nan
        else:
            r, p = np.nan, np.nan
        results.append({"item": col, "r": r, "p": p, "n": n})
    return pd.DataFrame(results)


# ══════════════════════════════════════════════
#  QSS Style
# ══════════════════════════════════════════════
MAIN_STYLE = """
QMainWindow { background-color: #ffffff; }
QGroupBox {
    font-size: 15px; font-weight: bold;
    border: 2px solid #c5cae9; border-radius: 10px;
    margin-top: 16px; padding: 20px 12px 12px 12px;
    background-color: #fafbff;
}
QGroupBox::title {
    subcontrol-origin: margin; subcontrol-position: top left;
    padding: 4px 14px; background: #e8eaf6; border-radius: 6px;
    color: #283593; font-size: 14px;
}
QPushButton {
    background-color: #4361ee; color: white; border: none; border-radius: 8px;
    padding: 10px 24px; font-size: 14px; font-weight: 600;
}
QPushButton:hover { background-color: #3a56d4; }
QPushButton:pressed { background-color: #2f48b8; }
QPushButton:disabled { background-color: #bdbdbd; color: #eeeeee; }
QPushButton#exportBtn { background-color: #2e7d32; font-size: 15px; padding: 12px 28px; }
QPushButton#exportBtn:hover { background-color: #1b5e20; }
QPushButton#backBtn { background-color: #78909c; font-size: 14px; }
QPushButton#backBtn:hover { background-color: #546e7a; }
QTableWidget {
    border: 1px solid #cfd8dc; border-radius: 6px; gridline-color: #e0e0e0;
    selection-background-color: #e3f2fd; font-size: 13px;
    background-color: #ffffff; color: #212121;
}
QTableWidget::item { padding: 5px 10px; }
QHeaderView::section {
    background-color: #37474f; color: white; padding: 8px;
    border: none; font-weight: bold; font-size: 13px;
}
QProgressBar {
    border: 1px solid #cfd8dc; border-radius: 6px; text-align: center;
    height: 28px; font-size: 13px; background: #eceff1;
}
QProgressBar::chunk { background-color: #4361ee; border-radius: 5px; }
QTextEdit#logPanel {
    background-color: #263238; color: #b2ff59;
    font-family: 'Consolas', 'Courier New', monospace; font-size: 12px;
    border-radius: 6px; padding: 10px;
}
QLabel { font-size: 14px; color: #212121; }
QComboBox {
    border: 2px solid #c5cae9; border-radius: 8px; padding: 7px 10px;
    font-size: 14px; background: #ffffff; color: #212121;
    min-height: 22px;
}
QComboBox:focus { border: 2px solid #3f51b5; }
QComboBox QAbstractItemView {
    font-size: 13px; background: #ffffff; color: #212121;
    selection-background-color: #e8eaf6; selection-color: #1a237e;
}
QListWidget {
    border: 2px solid #c5cae9; border-radius: 8px;
    background: #ffffff; font-size: 14px; color: #212121;
}
QListWidget::item { padding: 5px 8px; border-radius: 5px; margin: 1px 0; }
QListWidget::item:hover { background-color: #e8eaf6; }
QListWidget::item:selected { background-color: #c5cae9; color: #1a237e; font-weight: bold; }
QLineEdit {
    border: 2px solid #c5cae9; border-radius: 8px; padding: 7px 10px;
    font-size: 14px; background: #ffffff; color: #212121;
}
QLineEdit:focus { border: 2px solid #3f51b5; }
QTabWidget::pane { border: 2px solid #c5cae9; border-radius: 6px; background: #ffffff; }
QTabBar::tab {
    background: #e8eaf6; border: 1px solid #c5cae9; padding: 8px 20px;
    border-top-left-radius: 8px; border-top-right-radius: 8px;
    font-size: 14px; color: #37474f;
}
QTabBar::tab:selected { background: #ffffff; border-bottom: none; font-weight: bold; color: #283593; }
QCheckBox { font-size: 14px; spacing: 8px; color: #212121; }
QSpinBox {
    border: 2px solid #c5cae9; border-radius: 8px; padding: 6px;
    font-size: 14px; background: #ffffff; color: #212121;
}
QScrollArea { background: #ffffff; border: none; }
QScrollBar:vertical { width: 10px; background: #eceff1; border-radius: 5px; }
QScrollBar::handle:vertical { background: #90a4ae; border-radius: 5px; min-height: 30px; }
QScrollBar::handle:vertical:hover { background: #78909c; }
QMessageBox {
    background-color: #ffffff;
}
QMessageBox QLabel {
    color: #212121;
    font-size: 14px;
}
QMessageBox QPushButton {
    min-width: 90px;
    padding: 8px 16px;
    background-color: #4361ee;
    color: #ffffff;
    border-radius: 8px;
}
QMessageBox QPushButton:hover {
    background-color: #3a56d4;
}
QMessageBox QPushButton:pressed {
    background-color: #2f48b8;
}
"""

# ══════════════════════════════════════════════
#  Worker Thread
# ══════════════════════════════════════════════
class CalculationWorker(QThread):
    progress = pyqtSignal(int, str)
    result_ready = pyqtSignal(object, object) # (results, corr_data)
    error = pyqtSignal(str)
    log_message = pyqtSignal(str)

    def __init__(self, df, dimensions, subgroups, q1_var, rotation, min_n,
                 q2_all_vars, q3_all_vars, country_var, gen_var, gender_var):
        super().__init__()
        self.df = df
        self.dimensions = dimensions
        self.subgroups = subgroups
        self.q1_var = q1_var
        self.rotation = rotation
        self.min_n = min_n
        self.q2_all_vars = q2_all_vars
        self.q3_all_vars = q3_all_vars
        self.country_var = country_var
        self.gen_var = gen_var
        self.gender_var = gender_var

    def run(self):
        try:
            pipeline = IndexCalculationPipeline(self.df, self.dimensions, self.q1_var)
            results = []
            total = len(self.subgroups)
            for i, sg in enumerate(self.subgroups):
                self.log_message.emit(f"Processing: {sg.name}...")
                self.progress.emit(int((i / total) * 80), sg.name)
                mask = pipeline.build_mask(sg)
                n = int(mask.sum())
                if n < self.min_n:
                    self.log_message.emit(f"  Skipped (n={n} < {self.min_n})")
                    continue
                result, err_msg = pipeline.run_for_subgroup(sg, mask, self.rotation)
                if result is None:
                    self.log_message.emit(f"  Failed: {err_msg}")
                    continue
                result.subgroup_name = sg.name
                result.subgroup_main = sg.main_group or sg.name
                result.subgroup_filter = sg.filter_group or ""
                results.append(result)
                self.log_message.emit(f"  Done: Index={result.index_value:.3f}, n={result.n}")

            # Correlations
            self.log_message.emit("Computing correlations...")
            self.progress.emit(85, "Correlations...")
            corr_data = {"q2": {}, "q3": {}}

            # Use TOTAL index weights as the single base for all correlation runs.
            total_result = next((x for x in results if str(x.subgroup_name).strip().lower() == "total"), None)
            if total_result is None:
                raise ValueError("Correlation requires Total weights, but Total subgroup was not found.")
            # Lock correlation index to TOTAL weights only (SPSS syntax style).
            # Round to 4 decimals as commonly used in COMPUTE syntax coefficients.
            total_weights = np.array([total_result.dim_weights.get(d.short_name, 0.0) for d in self.dimensions], dtype=float)
            total_weights = np.round(total_weights, 4)
            w_sum = float(total_weights.sum())
            if w_sum <= 0:
                raise ValueError("Correlation requires valid Total weights, but sum(weights) <= 0.")
            self.log_message.emit("Correlation base index: fixed to Total regression weights only.")
            self.log_message.emit(
                "Correlation Total weights used: " +
                ", ".join([f"{d.short_name}={w:.4f}" for d, w in zip(self.dimensions, total_weights)])
            )

            for sg in self.subgroups:
                mask = pipeline.build_mask(sg)
                if mask.sum() < self.min_n:
                    continue
                sub_df = self.df[mask].copy()
                sub_df = pipeline.recode_q1(sub_df)
                dim_df = pipeline.compute_dim_means(sub_df)
                # find the matching result
                matching = [r for r in results if r.subgroup_name == sg.name]
                if not matching:
                    continue
                r = matching[0]
                weights = total_weights
                idx_series = pipeline.compute_index(dim_df, weights, spss_missing=True)
                q3_cols = [c for c in self.q3_all_vars if c in sub_df.columns]
                if q3_cols:
                    q3_items = sub_df[q3_cols]
                    corr_data["q3"][sg.name] = compute_correlations(idx_series, q3_items)

            self.progress.emit(100, "Complete!")
            self.log_message.emit(f"All done. {len(results)} subgroups computed.")
            self.result_ready.emit(results, corr_data)
        except Exception as e:
            self.error.emit(f"{str(e)}\n{traceback.format_exc()}")


# ══════════════════════════════════════════════
#  Drop Zone Widget
# ══════════════════════════════════════════════
class DropZone(QLabel):
    file_dropped = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setText("📂  ลากไฟล์ .sav มาวางที่นี่\nหรือกดปุ่ม Browse ด้านล่าง")
        self.setMinimumHeight(100)
        self.setStyleSheet("""
            QLabel {
                border: 2px dashed #adb5bd; border-radius: 12px;
                background-color: #f1f3f5; color: #495057;
                font-size: 14px; padding: 20px;
            }
        """)

    def dragEnterEvent(self, e: QDragEnterEvent):
        if e.mimeData().hasUrls():
            for url in e.mimeData().urls():
                if url.toLocalFile().lower().endswith(".sav"):
                    self.setStyleSheet(self.styleSheet().replace("#adb5bd", "#4361ee").replace("#f1f3f5", "#eef2ff"))
                    e.acceptProposedAction()
                    return
        e.ignore()

    def dragLeaveEvent(self, e):
        self.setStyleSheet(self.styleSheet().replace("#4361ee", "#adb5bd").replace("#eef2ff", "#f1f3f5"))

    def dropEvent(self, e: QDropEvent):
        self.setStyleSheet(self.styleSheet().replace("#4361ee", "#adb5bd").replace("#eef2ff", "#f1f3f5"))
        for url in e.mimeData().urls():
            fp = url.toLocalFile()
            if fp.lower().endswith(".sav"):
                self.file_dropped.emit(fp)
                return


# ══════════════════════════════════════════════
#  Variable Picker Dialog
# ══════════════════════════════════════════════
class VariablePickerDialog(QWidget):
    """Popup dialog for picking variables — with search and checkboxes."""

    def __init__(self, all_vars, title="เลือกตัวแปร", multi=True,
                 pre_selected=None, parent=None):
        super().__init__(parent, Qt.WindowType.Window | Qt.WindowType.WindowStaysOnTopHint)
        self.setWindowTitle(title)
        self.resize(600, 520)
        self.setStyleSheet(MAIN_STYLE)
        self.all_vars = all_vars  # list of (display, var_name)
        self.multi = multi
        self.pre_selected = set(pre_selected or [])
        self._result = []
        self._accepted = False

        # Center on parent
        if parent:
            pg = parent.geometry()
            self.move(pg.center().x() - 300, pg.center().y() - 260)

        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(12, 10, 12, 10)
        lay.setSpacing(8)

        # Search
        self.search = QLineEdit()
        self.search.setPlaceholderText("🔍 ค้นหาตัวแปร (พิมพ์ชื่อหรือ label)...")
        self.search.textChanged.connect(self._filter)
        lay.addWidget(self.search)

        # Select all / deselect
        if self.multi:
            btn_row = QHBoxLayout()
            btn_all = QPushButton("✅ เลือกทั้งหมด (ที่แสดง)")
            btn_all.setStyleSheet("font-size: 12px; padding: 5px 12px; background: #66bb6a;")
            btn_all.clicked.connect(self._select_all_visible)
            btn_none = QPushButton("❌ ยกเลิกทั้งหมด")
            btn_none.setStyleSheet("font-size: 12px; padding: 5px 12px; background: #ef5350;")
            btn_none.clicked.connect(self._deselect_all)
            self.count_lbl = QLabel(f"เลือกแล้ว: {len(self.pre_selected)}")
            self.count_lbl.setStyleSheet("font-size: 13px; font-weight: bold; color: #283593;")
            btn_row.addWidget(btn_all)
            btn_row.addWidget(btn_none)
            btn_row.addStretch()
            btn_row.addWidget(self.count_lbl)
            lay.addLayout(btn_row)

        # Variable list
        self.var_list = QListWidget()
        self.var_list.setStyleSheet("font-size: 14px;")
        for display, var_name in self.all_vars:
            item = QListWidgetItem(display)
            item.setData(Qt.ItemDataRole.UserRole, var_name)
            if self.multi:
                item.setCheckState(
                    Qt.CheckState.Checked if var_name in self.pre_selected
                    else Qt.CheckState.Unchecked
                )
            self.var_list.addItem(item)

        if not self.multi:
            self.var_list.itemDoubleClicked.connect(self._on_double_click)

        if self.multi:
            self.var_list.itemChanged.connect(self._update_count)

        lay.addWidget(self.var_list, stretch=1)

        # OK / Cancel
        btn_row2 = QHBoxLayout()
        btn_ok = QPushButton("✔ ตกลง")
        btn_ok.setStyleSheet("font-size: 14px; padding: 10px 30px; background: #2e7d32;")
        btn_ok.clicked.connect(self._accept)
        btn_cancel = QPushButton("✘ ยกเลิก")
        btn_cancel.setStyleSheet("font-size: 14px; padding: 10px 30px; background: #78909c;")
        btn_cancel.clicked.connect(self.close)
        btn_row2.addStretch()
        btn_row2.addWidget(btn_cancel)
        btn_row2.addWidget(btn_ok)
        lay.addLayout(btn_row2)

    def _filter(self, text):
        s = text.lower()
        for i in range(self.var_list.count()):
            item = self.var_list.item(i)
            item.setHidden(s not in item.text().lower())

    def _select_all_visible(self):
        for i in range(self.var_list.count()):
            item = self.var_list.item(i)
            if not item.isHidden():
                item.setCheckState(Qt.CheckState.Checked)
        self._update_count()

    def _deselect_all(self):
        for i in range(self.var_list.count()):
            self.var_list.item(i).setCheckState(Qt.CheckState.Unchecked)
        self._update_count()

    def _update_count(self, _=None):
        n = sum(1 for i in range(self.var_list.count())
                if self.var_list.item(i).checkState() == Qt.CheckState.Checked)
        if hasattr(self, "count_lbl"):
            self.count_lbl.setText(f"เลือกแล้ว: {n}")

    def _on_double_click(self, item):
        self._result = [item.data(Qt.ItemDataRole.UserRole)]
        self._accepted = True
        self.close()

    def _accept(self):
        if self.multi:
            self._result = []
            for i in range(self.var_list.count()):
                item = self.var_list.item(i)
                if item.checkState() == Qt.CheckState.Checked:
                    self._result.append(item.data(Qt.ItemDataRole.UserRole))
        else:
            sel = self.var_list.currentItem()
            self._result = [sel.data(Qt.ItemDataRole.UserRole)] if sel else []
        self._accepted = True
        self.close()

    def get_selected(self):
        return self._result

    def exec(self):
        """Block until dialog closes (like QDialog.exec)."""
        from PyQt6.QtCore import QEventLoop
        loop = QEventLoop()
        self._loop = loop
        self.destroyed.connect(loop.quit)
        self.show()
        loop.exec()
        return self._accepted

    def closeEvent(self, event):
        if hasattr(self, '_loop') and self._loop.isRunning():
            self._loop.quit()
        super().closeEvent(event)


# ══════════════════════════════════════════════
#  Main Application Window
# ══════════════════════════════════════════════
class WellbeingIndexApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Brand Space - Index Caculator")
        try:
            self.setWindowIcon(QIcon(_resource_path("BrandS.ico")))
        except Exception:
            pass
        self.resize(1050, 720)
        self._center()
        self.setStyleSheet(MAIN_STYLE)

        self.df = None
        self.meta = None
        self.all_vars = []
        self.results = []
        self.corr_data = {}
        self.worker = None

        self._build_ui()

    def _center(self):
        scr = QApplication.primaryScreen()
        if scr:
            g = scr.availableGeometry()
            self.move((g.width() - self.width()) // 2 + g.x(),
                      (g.height() - self.height()) // 2 + g.y())

    # ── Build UI ──────────────────────────────
    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_lay = QVBoxLayout(central)
        main_lay.setContentsMargins(14, 8, 14, 6)
        main_lay.setSpacing(6)

        # Header
        hdr = QLabel("Brand Space - Index Caculator")
        hdr.setStyleSheet("font-size: 18px; font-weight: 700; color: #364fc7; padding: 2px;")
        main_lay.addWidget(hdr)

        # Steps indicator
        self.steps_label = QLabel("① Load Data  ──►  ② Configure  ──►  ③ Results")
        self.steps_label.setStyleSheet("font-size: 12px; color: #868e96; padding: 0 0 4px 0;")
        main_lay.addWidget(self.steps_label)

        # Stacked pages
        self.stack = QStackedWidget()
        self._build_page1()
        self._build_page2()
        self._build_page3()
        main_lay.addWidget(self.stack, stretch=1)

        # Log panel
        self.log_panel = QTextEdit()
        self.log_panel.setObjectName("logPanel")
        self.log_panel.setReadOnly(True)
        self.log_panel.setMaximumHeight(120)
        self.log_panel.setVisible(False)
        main_lay.addWidget(self.log_panel)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setLocale(QLocale(QLocale.Language.English, QLocale.Country.UnitedStates))
        self.progress_bar.setFormat("%p%")
        self.progress_bar.setVisible(False)
        main_lay.addWidget(self.progress_bar)

        # Nav buttons
        nav = QHBoxLayout()
        self.btn_back = QPushButton("◄ Back")
        self.btn_back.setObjectName("backBtn")
        self.btn_back.clicked.connect(self._go_back)
        self.btn_back.setVisible(False)
        self.btn_next = QPushButton("Next ►")
        self.btn_next.clicked.connect(self._go_next)
        self.btn_next.setVisible(False)
        nav.addWidget(self.btn_back)
        nav.addStretch()
        nav.addWidget(self.btn_next)
        main_lay.addLayout(nav)

        self.statusBar().showMessage("พร้อมใช้งาน: กรุณาโหลดไฟล์ SPSS เพื่อเริ่มต้น")

    # ── Page 1: Load Data ─────────────────────
    def _build_page1(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setSpacing(12)

        self.drop_zone = DropZone()
        self.drop_zone.file_dropped.connect(self._load_file)
        lay.addWidget(self.drop_zone)

        btn_row = QHBoxLayout()
        btn_browse = QPushButton("Browse .sav file...")
        btn_browse.clicked.connect(self._browse_file)
        btn_row.addStretch()
        btn_row.addWidget(btn_browse)
        btn_row.addStretch()
        lay.addLayout(btn_row)

        self.file_info_label = QLabel("")
        self.file_info_label.setStyleSheet("font-size: 15px; color: #2e7d32; padding: 8px; font-weight: 600;")
        lay.addWidget(self.file_info_label)

        guide_box = QGroupBox("วิธีคำนวณโดยสรุป (How It Works)")
        guide_box.setStyleSheet(
            "QGroupBox { font-size: 15px; font-weight: 700; color: #1f3a93; }"
            "QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 8px; }"
        )
        guide_lay = QVBoxLayout(guide_box)
        guide_lay.setContentsMargins(14, 12, 14, 12)
        guide_lay.setSpacing(6)
        guide_text = QLabel(
            "1) โหลดไฟล์ .sav และเลือกตัวแปร Q1, Q2 (6 มิติ), Subgroup และตัวแปร Correlation\n"
            "2) คำนวณคะแนนมิติรายคน (Dim1..Dim6) จากค่าเฉลี่ยของข้อ Q2 ในแต่ละมิติ\n"
            "3) รัน Factor Analysis (PCA + Equamax) และ Regression เพื่อหา Weight ของแต่ละมิติ\n"
            "4) คำนวณ Index แต่ละกลุ่มจากสูตร: Index = Σ(Mean Score ของมิติ × Weight)\n"
            "5) Correlation ใช้ Index ที่คูณด้วย Total Weight เท่านั้น แล้วหา Pearson กับตัวแปรที่เลือก\n"
            "6) ส่งออกผลลัพธ์เป็น Excel (Index, Correlations, Rotated Component, Regression Beta, Logs)"
        )
        guide_text.setWordWrap(True)
        guide_text.setStyleSheet(
            "font-size: 13px; color: #37474f; "
            "background: #f7f9fc; border: 1px solid #dbe3ef; border-radius: 10px; padding: 10px;"
        )
        guide_lay.addWidget(guide_text)
        lay.addWidget(guide_box)

        lay.addStretch()

        self.stack.addWidget(page)

    def _fill_log_viewer(self):
        if not self.results:
            return
        
        all_logs = []
        for r in self.results:
            if r.factor_output_log:
                all_logs.append(r.factor_output_log)
                
        if not all_logs:
            self.txt_factor_log.setText("No log available.")
            return

        self.txt_factor_log.setText("\n".join(all_logs))

    # ── Page 2: Configure ─────────────────────
    def _build_page2(self):
        page = QWidget()
        lay = QVBoxLayout(page)

        # Save / Load settings row
        settings_row = QHBoxLayout()
        btn_save_cfg = QPushButton("💾 บันทึกการตั้งค่า")
        btn_save_cfg.setStyleSheet("font-size: 13px; padding: 7px 16px; background: #5c6bc0;")
        btn_save_cfg.clicked.connect(self._save_settings)
        btn_load_cfg = QPushButton("📂 โหลดการตั้งค่า")
        btn_load_cfg.setStyleSheet("font-size: 13px; padding: 7px 16px; background: #78909c;")
        btn_load_cfg.clicked.connect(self._load_settings)
        settings_row.addWidget(btn_save_cfg)
        settings_row.addWidget(btn_load_cfg)
        settings_row.addStretch()
        lay.addLayout(settings_row)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        inner = QWidget()
        inner_lay = QVBoxLayout(inner)
        inner_lay.setSpacing(10)

        # Q1 mapping — click button to pick
        grp_q1 = QGroupBox("Q1 Variable (Overall Feeling, 7-point scale)")
        q1_lay = QHBoxLayout(grp_q1)
        self.lbl_q1 = QLabel("ยังไม่ได้เลือก")
        self.lbl_q1.setStyleSheet("font-size: 14px; color: #c62828; padding: 4px;")
        self._selected_q1 = None
        btn_q1 = QPushButton("เลือกตัวแปร Q1...")
        btn_q1.clicked.connect(lambda: self._pick_single_var("เลือกตัวแปร Q1", self._set_q1))
        q1_lay.addWidget(self.lbl_q1, stretch=1)
        q1_lay.addWidget(btn_q1)
        inner_lay.addWidget(grp_q1)

        # Dimensions — each has a pick button
        grp_dims = QGroupBox("Dimension Variable Mapping (Q2 sub-items)")
        dims_lay = QVBoxLayout(grp_dims)
        hint = QLabel("💡 กดปุ่ม 'เลือก...' เพื่อเปิดหน้าต่างเลือกตัวแปรสำหรับแต่ละ Dimension")
        hint.setStyleSheet("color: #5c6bc0; font-size: 13px; font-style: italic; padding: 2px 0 6px 0;")
        dims_lay.addWidget(hint)
        self._dim_selected = {}  # prefix → list of var names
        self._dim_labels = {}    # prefix → QLabel showing selections
        for d in DEFAULT_DIMENSIONS:
            row = QHBoxLayout()
            lbl = QLabel(f"{d['short']}:")
            lbl.setMinimumWidth(170)
            lbl.setStyleSheet("font-size: 14px; font-weight: bold; color: #283593;")
            sel_lbl = QLabel("ยังไม่ได้เลือก")
            sel_lbl.setStyleSheet("font-size: 13px; color: #757575; padding: 2px 8px;")
            sel_lbl.setWordWrap(True)
            btn = QPushButton("เลือก...")
            btn.setFixedWidth(100)
            prefix = d["prefix"]
            self._dim_selected[prefix] = []
            self._dim_labels[prefix] = sel_lbl
            btn.clicked.connect(lambda checked, p=prefix, s=d["short"]: self._pick_multi_var(
                f"เลือกตัวแปรสำหรับ {s}", p
            ))
            row.addWidget(lbl)
            row.addWidget(sel_lbl, stretch=1)
            row.addWidget(btn)
            dims_lay.addLayout(row)
        inner_lay.addWidget(grp_dims)

        # Subgroup main variable (separate from filter list)
        grp_main = QGroupBox("ตัวแปร Subgroup Main")
        main_lay = QHBoxLayout(grp_main)
        self._selected_subgroup_main = None
        self.lbl_subgroup_main = QLabel("ยังไม่ได้เลือก")
        self.lbl_subgroup_main.setStyleSheet("font-size: 13px; color: #757575; padding: 2px 8px;")
        self.lbl_subgroup_main.setWordWrap(True)
        btn_main = QPushButton("เลือก Subgroup Main...")
        btn_main.clicked.connect(self._pick_subgroup_main_var)
        main_lay.addWidget(self.lbl_subgroup_main, stretch=1)
        main_lay.addWidget(btn_main)
        inner_lay.addWidget(grp_main)

        # Subgroup filter variables
        grp_filter = QGroupBox("ตัวแปรสำหรับแบ่ง Subgroup (Filter)")
        filter_lay = QVBoxLayout(grp_filter)
        filter_hint = QLabel("เพิ่มตัวแปร demographic ที่ต้องการแบ่งกลุ่ม เช่น Country, Generation, Gender\n"
                             "โปรแกรมจะสร้าง subgroup จาก unique values ของตัวแปรเหล่านี้อัตโนมัติ")
        filter_hint.setStyleSheet("font-size: 12px; color: #5c6bc0; font-style: italic;")
        filter_lay.addWidget(filter_hint)
        self.filter_var_list = QListWidget()
        self.filter_var_list.setMaximumHeight(90)
        filter_lay.addWidget(self.filter_var_list)
        btn_filter_row = QHBoxLayout()
        btn_add_filter = QPushButton("+ เพิ่มตัวแปร Filter")
        btn_add_filter.clicked.connect(self._add_filter_var)
        btn_remove_filter = QPushButton("ลบที่เลือก")
        btn_remove_filter.setStyleSheet("background-color: #ef5350; font-size: 13px;")
        btn_remove_filter.clicked.connect(self._remove_filter_var)
        btn_filter_row.addWidget(btn_add_filter)
        btn_filter_row.addWidget(btn_remove_filter)
        btn_filter_row.addStretch()
        filter_lay.addLayout(btn_filter_row)
        inner_lay.addWidget(grp_filter)

        # Correlation variables — click to pick
        grp_q3 = QGroupBox("Correlation Variables (optional)")
        q3_lay = QHBoxLayout(grp_q3)
        self.lbl_q3 = QLabel("ยังไม่ได้เลือก")
        self.lbl_q3.setStyleSheet("font-size: 13px; color: #757575; padding: 2px;")
        self.lbl_q3.setWordWrap(True)
        self._selected_q3 = []
        btn_q3 = QPushButton("เลือกตัวแปร Correlation...")
        btn_q3.clicked.connect(self._pick_q3_vars)
        q3_lay.addWidget(self.lbl_q3, stretch=1)
        q3_lay.addWidget(btn_q3)
        inner_lay.addWidget(grp_q3)

        # Advanced
        grp_adv = QGroupBox("Advanced Settings")
        adv_lay = QGridLayout(grp_adv)
        adv_lay.addWidget(QLabel("Rotation:"), 0, 0)
        self.combo_rotation = QComboBox()
        self.combo_rotation.addItems(["equamax"])
        self.combo_rotation.setCurrentText("equamax")
        self.combo_rotation.setEnabled(False)
        adv_lay.addWidget(self.combo_rotation, 0, 1)
        adv_lay.addWidget(QLabel("Min N per subgroup:"), 1, 0)
        self.spin_min_n = QSpinBox()
        self.spin_min_n.setRange(10, 500)
        self.spin_min_n.setValue(30)
        adv_lay.addWidget(self.spin_min_n, 1, 1)
        inner_lay.addWidget(grp_adv)

        inner_lay.addStretch()
        scroll.setWidget(inner)
        lay.addWidget(scroll)
        self.stack.addWidget(page)

    # ── Variable Picker Dialogs ───────────────
    def _pick_single_var(self, title, callback):
        if not self.all_vars:
            QMessageBox.warning(self, "Warning", "กรุณาโหลดไฟล์ SPSS ก่อน")
            return
        dlg = VariablePickerDialog(self.all_vars, title=title, multi=False, parent=self)
        if dlg.exec():
            sel = dlg.get_selected()
            if sel:
                callback(sel[0])

    def _pick_multi_var(self, title, prefix):
        if not self.all_vars:
            QMessageBox.warning(self, "Warning", "กรุณาโหลดไฟล์ SPSS ก่อน")
            return
        pre_selected = self._dim_selected.get(prefix, [])
        dlg = VariablePickerDialog(self.all_vars, title=title, multi=True,
                                   pre_selected=pre_selected, parent=self)
        if dlg.exec():
            sel = dlg.get_selected()
            self._dim_selected[prefix] = sel
            lbl = self._dim_labels[prefix]
            if sel:
                lbl.setText(f"✅ {len(sel)} ตัวแปร: {', '.join(sel[:5])}{'...' if len(sel) > 5 else ''}")
                lbl.setStyleSheet("font-size: 13px; color: #2e7d32; padding: 2px 8px; font-weight: 600;")
            else:
                lbl.setText("ยังไม่ได้เลือก")
                lbl.setStyleSheet("font-size: 13px; color: #757575; padding: 2px 8px;")

    def _set_q1(self, var_name):
        self._selected_q1 = var_name
        labels = self.meta.column_names_to_labels if self.meta else {}
        lbl = labels.get(var_name, "")
        self.lbl_q1.setText(f"✅ [{var_name}] {lbl}")
        self.lbl_q1.setStyleSheet("font-size: 14px; color: #2e7d32; padding: 4px; font-weight: 600;")

    def _pick_q3_vars(self):
        if not self.all_vars:
            QMessageBox.warning(self, "Warning", "กรุณาโหลดไฟล์ SPSS ก่อน")
            return
        dlg = VariablePickerDialog(self.all_vars, title="เลือกตัวแปร Correlation",
                                   multi=True, pre_selected=self._selected_q3, parent=self)
        if dlg.exec():
            self._selected_q3 = dlg.get_selected()
            if self._selected_q3:
                self.lbl_q3.setText(f"✅ {len(self._selected_q3)} ตัวแปร: {', '.join(self._selected_q3[:5])}{'...' if len(self._selected_q3) > 5 else ''}")
                self.lbl_q3.setStyleSheet("font-size: 13px; color: #2e7d32; padding: 2px; font-weight: 600;")
            else:
                self.lbl_q3.setText("ยังไม่ได้เลือก")
                self.lbl_q3.setStyleSheet("font-size: 13px; color: #757575; padding: 2px;")

    def _pick_subgroup_main_var(self):
        if not self.all_vars:
            QMessageBox.warning(self, "Warning", "กรุณาโหลดไฟล์ SPSS ก่อน")
            return
        self._pick_single_var("เลือกตัวแปร Subgroup Main", self._set_subgroup_main_var)

    def _set_subgroup_main_var(self, var_name):
        self._selected_subgroup_main = var_name
        labels = self.meta.column_names_to_labels if self.meta else {}
        lbl = labels.get(var_name, "")
        self.lbl_subgroup_main.setText(f"✅ [{var_name}] {lbl}")
        self.lbl_subgroup_main.setStyleSheet("font-size: 13px; color: #2e7d32; padding: 2px 8px; font-weight: 600;")

    def _add_filter_var(self):
        if not self.all_vars:
            QMessageBox.warning(self, "Warning", "กรุณาโหลดไฟล์ SPSS ก่อน")
            return
        existing = []
        for i in range(self.filter_var_list.count()):
            existing.append(self.filter_var_list.item(i).data(Qt.ItemDataRole.UserRole))
        dlg = VariablePickerDialog(self.all_vars, title="เลือกตัวแปร Filter (Demographic)",
                                   multi=False, parent=self)
        if dlg.exec():
            sel = dlg.get_selected()
            if sel and sel[0] not in existing:
                var = sel[0]
                labels = self.meta.column_names_to_labels if self.meta else {}
                lbl = labels.get(var, "")
                item = QListWidgetItem(f"[{var}] {lbl}")
                item.setData(Qt.ItemDataRole.UserRole, var)
                self.filter_var_list.addItem(item)

    def _remove_filter_var(self):
        for item in self.filter_var_list.selectedItems():
            self.filter_var_list.takeItem(self.filter_var_list.row(item))

    # ── Page 3: Results ───────────────────────
    def _build_page3(self):
        page = QWidget()
        lay = QVBoxLayout(page)

        self.tabs = QTabWidget()

        # Tab 1: Weights
        tab1 = QWidget()
        t1_lay = QVBoxLayout(tab1)
        t1_lay.addWidget(QLabel("น้ำหนัก (Weight) ของแต่ละ Dimension — คำนวณจาก |Standardized Beta| ของ Regression"))
        self.tbl_weights = QTableWidget()
        t1_lay.addWidget(self.tbl_weights)
        self.tabs.addTab(tab1, "⚖ Weights")

        # Tab 2: Factor + Regression Output Log
        tab2 = QWidget()
        t2_lay = QVBoxLayout(tab2)
        self.txt_factor_log = QTextEdit()
        self.txt_factor_log.setReadOnly(True)
        self.txt_factor_log.setObjectName("logPanel")
        self.txt_factor_log.setStyleSheet("background-color: #263238; color: #b2ff59; font-family: 'Consolas', 'Courier New', monospace; font-size: 13px; padding: 10px;")
        t2_lay.addWidget(self.txt_factor_log)
        self.tabs.addTab(tab2, "📝 Factor+Regression Logs")

        lay.addWidget(self.tabs)

        btn_export = QPushButton("📥  Export เป็น Excel")
        btn_export.setObjectName("exportBtn")
        btn_export.clicked.connect(self._export_excel)
        lay.addWidget(btn_export)

        self.stack.addWidget(page)

    # ── Navigation ────────────────────────────
    def _go_back(self):
        idx = self.stack.currentIndex()
        if idx > 0:
            self.stack.setCurrentIndex(idx - 1)
            self._update_nav()

    def _go_next(self):
        idx = self.stack.currentIndex()
        if idx == 0:
            if self.df is None:
                QMessageBox.warning(self, "Warning", "กรุณาโหลดไฟล์ SPSS ก่อน")
                return
            self.stack.setCurrentIndex(1)
        elif idx == 1:
            self._run_analysis()
            return
        self._update_nav()

    def _update_nav(self):
        idx = self.stack.currentIndex()
        self.btn_back.setVisible(idx > 0)
        self.btn_next.setVisible(idx < 2)
        if idx == 1:
            self.btn_next.setText("▶  Run Analysis")
        else:
            self.btn_next.setText("Next ►")
        labels = ["① Load Data", "② Configure", "③ Results"]
        parts = []
        for i, l in enumerate(labels):
            if i == idx:
                parts.append(f"<b style='color:#4361ee'>{l}</b>")
            else:
                parts.append(l)
        self.steps_label.setText("  ──►  ".join(parts))

    # ── File Loading ──────────────────────────
    def _browse_file(self):
        fp, _ = QFileDialog.getOpenFileName(self, "เปิดไฟล์ SPSS", "", "SPSS Files (*.sav)")
        if fp:
            self._load_file(fp)

    def _load_file(self, filepath):
        try:
            self.df, self.meta = pyreadstat.read_sav(filepath)
            fname = os.path.basename(filepath)
            n_rec = len(self.df)
            n_var = len(self.df.columns)
            self.file_info_label.setText(f"✅ {fname} — {n_rec} records, {n_var} variables")
            self.drop_zone.setText(f"✅ โหลดแล้ว: {fname}")
            self.drop_zone.setStyleSheet(self.drop_zone.styleSheet().replace("#adb5bd", "#2d6a4f").replace("#f1f3f5", "#e6fcf5"))

            # Prepare variable list for pickers
            labels = self.meta.column_names_to_labels if self.meta else {}
            self.all_vars = []
            for i, col in enumerate(self.df.columns):
                lbl = labels.get(col, "")
                display = f"[{col}] {lbl}"
                self.all_vars.append((display, col))

            self.btn_next.setVisible(True)
            self.stack.setCurrentIndex(1)
            self._update_nav()
            self.statusBar().showMessage(f"โหลดสำเร็จ: {n_rec} records, {n_var} variables")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"โหลดไฟล์ไม่สำเร็จ:\n{str(e)}")

    # ── Run Analysis ──────────────────────────
    def _run_analysis(self):
        if self.df is None:
            QMessageBox.warning(self, "Warning", "กรุณาโหลดไฟล์ SPSS ก่อน")
            return

        # Validate Q1
        if not self._selected_q1:
            QMessageBox.warning(self, "Warning", "กรุณาเลือกตัวแปร Q1 ก่อน")
            return

        # Build dimensions using SPSS-style fixed ranges first (e.g., PWQ2#1..PWQ2#5).
        # If the full expected range is not present, fall back to user selection.
        dimensions = []
        q2_all = []
        for d_def in DEFAULT_DIMENSIONS:
            prefix = d_def["prefix"]
            expected = [f"{prefix}#{i}" for i in range(1, int(d_def["n_items"]) + 1)]
            has_full_expected = all(v in self.df.columns for v in expected)
            if has_full_expected:
                selected = expected
            else:
                selected = self._dim_selected.get(prefix, [])
            if not selected:
                continue
            q2_all.extend(selected)
            dimensions.append(DimensionDef(
                name=d_def["name"], short_name=d_def["short"],
                prefix=prefix, q2_variables=selected,
                q3_prefix=d_def["q3_prefix"], q3_variables=[],
            ))

        if len(dimensions) < 2:
            QMessageBox.warning(self, "Warning", "กรุณาเลือกตัวแปรอย่างน้อย 2 Dimensions")
            return

        q1_var = self._selected_q1
        rotation = "equamax"
        min_n = self.spin_min_n.value()

        # Build subgroups from filter variables
        subgroups = [SubgroupDef(name="Total", filters={}, main_group="Total", filter_group="")]
        filter_vars = []
        for i in range(self.filter_var_list.count()):
            fv = self.filter_var_list.item(i).data(Qt.ItemDataRole.UserRole)
            if fv and fv in self.df.columns:
                filter_vars.append(fv)

        val_labels = {}
        if self.meta and hasattr(self.meta, "variable_value_labels"):
            val_labels = self.meta.variable_value_labels or {}

        main_var = self._selected_subgroup_main if (self._selected_subgroup_main in self.df.columns) else None
        if main_var is None and filter_vars:
            main_var = filter_vars[0]

        if main_var:
            main_values = sorted(self.df[main_var].dropna().unique())
            main_items = [(v, str(val_labels.get(main_var, {}).get(v, v))) for v in main_values]

            for main_val, main_lbl in main_items:
                subgroups.append(SubgroupDef(
                    name=main_lbl,
                    filters={main_var: main_val},
                    main_group=main_lbl,
                    filter_group=""
                ))

            for profile_var in [v for v in filter_vars if v != main_var]:
                profile_values = sorted(self.df[profile_var].dropna().unique())
                profile_items = [(v, str(val_labels.get(profile_var, {}).get(v, v))) for v in profile_values]

                for profile_val, profile_lbl in profile_items:
                    subgroups.append(SubgroupDef(
                        name=f"Total - {profile_lbl}",
                        filters={profile_var: profile_val},
                        main_group="Total",
                        filter_group=profile_lbl
                    ))

                    for main_val, main_lbl in main_items:
                        subgroups.append(SubgroupDef(
                            name=f"{main_lbl} x {profile_lbl}",
                            filters={main_var: main_val, profile_var: profile_val},
                            main_group=main_lbl,
                            filter_group=profile_lbl
                        ))

        q3_all = self._selected_q3 or []

        self.log_panel.clear()
        self.log_panel.setVisible(True)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.btn_next.setEnabled(False)

        self.worker = CalculationWorker(
            self.df, dimensions, subgroups, q1_var, rotation, min_n,
            q2_all, q3_all, "", "", "",
        )
        self.worker.log_message.connect(self._on_log)
        self.worker.progress.connect(self._on_progress)
        self.worker.result_ready.connect(self._on_results)
        self.worker.error.connect(self._on_error)
        self.worker.start()

    def _on_log(self, msg):
        self.log_panel.append(f"[INFO] {msg}")

    def _on_progress(self, pct, msg):
        self.progress_bar.setValue(pct)
        self.statusBar().showMessage(msg)

    def _on_error(self, msg):
        self.progress_bar.setVisible(False)
        self.btn_next.setEnabled(True)
        QMessageBox.critical(self, "Error", f"เกิดข้อผิดพลาด:\n{msg}")

    def _on_results(self, results, corr_data):
        self.results = results
        self.corr_data = corr_data
        self.progress_bar.setVisible(False)
        self.btn_next.setEnabled(True)

        if not results:
            QMessageBox.warning(self, "Warning", "ไม่มี subgroup ที่คำนวณได้ (N น้อยเกินไป หรือ analysis failed)")
            return

        self._fill_weights_table()
        self._fill_log_viewer()
        self.stack.setCurrentIndex(2)
        self._update_nav()
        self.statusBar().showMessage(f"วิเคราะห์เสร็จ: {len(results)} subgroups")
        QMessageBox.information(self, "สำเร็จ", f"วิเคราะห์ Brand Space Index เสร็จแล้ว\n{len(results)} subgroups")

    # ── Fill Tables ───────────────────────────
    def _fill_index_table(self):
        if not self.results:
            return
        dim_names = list(self.results[0].dim_means.keys())
        n_dims = len(dim_names)
        cols = ["Subgroup", "N", "Index"] + [f"Score: {d}" for d in dim_names] + [f"Weight: {d}" for d in dim_names]
        self.tbl_index.setColumnCount(len(cols))
        self.tbl_index.setHorizontalHeaderLabels(cols)
        self.tbl_index.setRowCount(len(self.results))
        for i, r in enumerate(self.results):
            self.tbl_index.setItem(i, 0, QTableWidgetItem(r.subgroup_name))
            self.tbl_index.setItem(i, 1, QTableWidgetItem(str(r.n)))
            self.tbl_index.setItem(i, 2, QTableWidgetItem(f"{r.index_value:.3f}"))
            for j, d in enumerate(dim_names):
                self.tbl_index.setItem(i, 3 + j, QTableWidgetItem(f"{r.dim_means.get(d, 0):.3f}"))
                self.tbl_index.setItem(i, 3 + n_dims + j, QTableWidgetItem(f"{r.dim_weights.get(d, 0) * 100:.1f}%"))
        self.tbl_index.resizeColumnsToContents()

    def _fill_regression_table(self):
        if not self.results:
            return
        dim_names = list(self.results[0].std_betas.keys())
        cols = ["Subgroup", "R²", "Adj R²", "F", "Sig."] + [f"β: {d}" for d in dim_names]
        self.tbl_reg.setColumnCount(len(cols))
        self.tbl_reg.setHorizontalHeaderLabels(cols)
        self.tbl_reg.setRowCount(len(self.results))
        for i, r in enumerate(self.results):
            rs = r.regression_stats
            self.tbl_reg.setItem(i, 0, QTableWidgetItem(r.subgroup_name))
            self.tbl_reg.setItem(i, 1, QTableWidgetItem(f"{rs.get('r_squared',0):.4f}"))
            self.tbl_reg.setItem(i, 2, QTableWidgetItem(f"{rs.get('adj_r_squared',0):.4f}"))
            self.tbl_reg.setItem(i, 3, QTableWidgetItem(f"{rs.get('f_statistic',0):.2f}"))
            fp = rs.get('f_pvalue', 1)
            self.tbl_reg.setItem(i, 4, QTableWidgetItem(f"{fp:.4f}" if fp else ""))
            for j, d in enumerate(dim_names):
                self.tbl_reg.setItem(i, 5 + j, QTableWidgetItem(f"{r.std_betas.get(d, 0):.4f}"))
        self.tbl_reg.resizeColumnsToContents()

    def _fill_weights_table(self):
        if not self.results:
            return
        dim_names = list(self.results[0].dim_weights.keys())
        cols = ["Subgroup", "N", "Index"] + [f"{d}" for d in dim_names]
        self.tbl_weights.setColumnCount(len(cols))
        self.tbl_weights.setHorizontalHeaderLabels(cols)
        self.tbl_weights.setRowCount(len(self.results))
        for i, r in enumerate(self.results):
            self.tbl_weights.setItem(i, 0, QTableWidgetItem(r.subgroup_name))
            self.tbl_weights.setItem(i, 1, QTableWidgetItem(str(r.n)))
            item_idx = QTableWidgetItem(f"{r.index_value:.3f}")
            item_idx.setBackground(QColor("#e3f2fd"))
            self.tbl_weights.setItem(i, 2, item_idx)
            # Find max weight to highlight
            max_w = max(r.dim_weights.values()) if r.dim_weights else 0
            for j, d in enumerate(dim_names):
                w = r.dim_weights.get(d, 0)
                item = QTableWidgetItem(f"{w * 100:.1f}%")
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                # Color code: highest weight = green, others proportional
                if max_w > 0:
                    intensity = int(200 * (w / max_w))
                    item.setBackground(QColor(232 - intensity // 3, 245 - intensity // 8, 233 - intensity // 4))
                    if w == max_w:
                        item.setBackground(QColor("#a5d6a7"))
                        item.setFont(QFont("", -1, QFont.Weight.Bold))
                self.tbl_weights.setItem(i, 3 + j, item)
        self.tbl_weights.resizeColumnsToContents()

    def _fill_corr_table(self):
        if not self.corr_data or "q2" not in self.corr_data:
            return
        q2 = self.corr_data["q2"]
        if not q2:
            return
        sg_names = list(q2.keys())
        first = q2[sg_names[0]]
        items = first["item"].tolist()
        cols = ["Item"] + sg_names
        self.tbl_corr.setColumnCount(len(cols))
        self.tbl_corr.setHorizontalHeaderLabels(cols)
        self.tbl_corr.setRowCount(len(items) * 3)
        for i, item_name in enumerate(items):
            for k, stat in enumerate(["r", "p", "n"]):
                row = i * 3 + k
                label = f"{item_name}" if k == 0 else ""
                stat_label = ["Pearson r", "Sig. (2-tailed)", "N"][k]
                self.tbl_corr.setItem(row, 0, QTableWidgetItem(f"{item_name} — {stat_label}"))
                for j, sg in enumerate(sg_names):
                    df_c = q2[sg]
                    match = df_c[df_c["item"] == item_name]
                    if not match.empty:
                        val = match.iloc[0][stat]
                        if stat == "n":
                            self.tbl_corr.setItem(row, 1 + j, QTableWidgetItem(str(int(val))))
                        else:
                            self.tbl_corr.setItem(row, 1 + j, QTableWidgetItem(f"{val:.4f}" if pd.notna(val) else ""))
        self.tbl_corr.resizeColumnsToContents()

    # ── Settings Save / Load ──────────────────
    def _save_settings(self):
        if self.df is None:
            QMessageBox.warning(self, "Warning", "กรุณาโหลดไฟล์ SPSS และเลือกตัวแปรก่อนบันทึก")
            return
            
        fp, _ = QFileDialog.getSaveFileName(self, "Save Settings", "Excel Setting BrandSpace.xlsx", "Excel Files (*.xlsx)")
        if not fp:
            return
            
        try:
            filter_vars = [
                self.filter_var_list.item(i).data(Qt.ItemDataRole.UserRole)
                for i in range(self.filter_var_list.count())
            ]

            # General settings: one row per field, easy to edit directly in Excel.
            df_general = pd.DataFrame([
                {"Field": "q1_var", "Value": self._selected_q1 or ""},
                {"Field": "subgroup_main_var", "Value": self._selected_subgroup_main or ""},
                {"Field": "rotation", "Value": "equamax"},
                {"Field": "min_n", "Value": int(self.spin_min_n.value())},
            ])

            # Dimension mapping: each row contains one dimension and comma-separated variable names.
            dim_rows = []
            for d in DEFAULT_DIMENSIONS:
                prefix = d["prefix"]
                vars_sel = self._dim_selected.get(prefix, [])
                dim_rows.append({
                    "prefix": prefix,
                    "dimension": d["short"],
                    "variables_csv": ", ".join(vars_sel),
                })
            df_dims = pd.DataFrame(dim_rows)

            # Filters and correlation variables: one variable per row.
            df_filters = pd.DataFrame({"filter_var": [v for v in filter_vars if v]})
            df_corr = pd.DataFrame({"correlation_var": [v for v in self._selected_q3 if v]})

            with pd.ExcelWriter(fp, engine="openpyxl") as writer:
                df_general.to_excel(writer, sheet_name="General", index=False)
                df_dims.to_excel(writer, sheet_name="Dimensions", index=False)
                df_filters.to_excel(writer, sheet_name="Filters", index=False)
                df_corr.to_excel(writer, sheet_name="Correlation", index=False)

                # Basic formatting for easier manual editing.
                for sname in ["General", "Dimensions", "Filters", "Correlation"]:
                    ws = writer.sheets[sname]
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=1, column=col)
                        cell.font = XlFont(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    ws.freeze_panes = "A2"
                writer.sheets["General"].column_dimensions["A"].width = 24
                writer.sheets["General"].column_dimensions["B"].width = 44
                writer.sheets["Dimensions"].column_dimensions["A"].width = 16
                writer.sheets["Dimensions"].column_dimensions["B"].width = 22
                writer.sheets["Dimensions"].column_dimensions["C"].width = 90
                writer.sheets["Filters"].column_dimensions["A"].width = 38
                writer.sheets["Correlation"].column_dimensions["A"].width = 38

                # General sheet edits: keep Value column left-aligned and lock only rotation/min_n.
                ws_general = writer.sheets["General"]
                lock_fill = PatternFill(start_color="B0BEC5", end_color="B0BEC5", fill_type="solid")
                for r in range(2, ws_general.max_row + 1):
                    field = str(ws_general.cell(row=r, column=1).value or "").strip()
                    val_cell = ws_general.cell(row=r, column=2)
                    val_cell.alignment = Alignment(horizontal="left", vertical="center")
                    if field in ("rotation", "min_n"):
                        val_cell.fill = lock_fill
                        val_cell.protection = Protection(locked=True)
                    else:
                        val_cell.protection = Protection(locked=False)
                ws_general.protection.sheet = True
            QMessageBox.information(self, "สำเร็จ", "บันทึกการตั้งค่าสำเร็จ")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"บันทึกการตั้งค่าไม่สำเร็จ:\n{str(e)}")

    def _load_settings(self):
        if self.df is None:
            QMessageBox.warning(self, "Warning", "กรุณาโหลดไฟล์ SPSS ก่อนโหลดการตั้งค่า")
            return
            
        fp, _ = QFileDialog.getOpenFileName(self, "Load Settings", "", "Excel Files (*.xlsx)")
        if not fp:
            return
            
        try:
            xls = pd.ExcelFile(fp)
            sheet_names = set(xls.sheet_names)

            # Backward compatibility: old format with Settings(Key/Value)
            if "Settings" in sheet_names and "General" not in sheet_names:
                df_settings = pd.read_excel(fp, sheet_name="Settings")
                settings = {}
                for _, row in df_settings.iterrows():
                    key = str(row.get("Key", "")).strip()
                    raw_val = row.get("Value", "")
                    if not key:
                        continue
                    if pd.isna(raw_val):
                        settings[key] = None
                        continue
                    try:
                        settings[key] = json.loads(str(raw_val))
                    except Exception:
                        settings[key] = raw_val
                general = {
                    "q1_var": settings.get("q1_var", ""),
                    "subgroup_main_var": settings.get("subgroup_main_var", ""),
                    "rotation": settings.get("rotation", "equamax"),
                    "min_n": settings.get("min_n", 30),
                }
                dim_selected = settings.get("dim_selected", {}) or {}
                filter_vars = settings.get("filter_vars", []) or []
                corr_vars = settings.get("q3_vars", []) or []
            else:
                df_general = pd.read_excel(fp, sheet_name="General") if "General" in sheet_names else pd.DataFrame()
                df_dims = pd.read_excel(fp, sheet_name="Dimensions") if "Dimensions" in sheet_names else pd.DataFrame()
                df_filters = pd.read_excel(fp, sheet_name="Filters") if "Filters" in sheet_names else pd.DataFrame()
                df_corr = pd.read_excel(fp, sheet_name="Correlation") if "Correlation" in sheet_names else pd.DataFrame()

                general = {}
                for _, row in df_general.iterrows():
                    key = str(row.get("Field", "")).strip()
                    if not key:
                        continue
                    general[key] = row.get("Value", "")

                dim_selected = {}
                for d in DEFAULT_DIMENSIONS:
                    dim_selected[d["prefix"]] = []
                for _, row in df_dims.iterrows():
                    prefix = str(row.get("prefix", "")).strip()
                    vars_csv = row.get("variables_csv", "")
                    if not prefix:
                        continue
                    if pd.isna(vars_csv):
                        vars_list = []
                    else:
                        vars_list = [v.strip() for v in str(vars_csv).split(",") if v.strip()]
                    dim_selected[prefix] = vars_list

                filter_vars = []
                if "filter_var" in df_filters.columns:
                    for v in df_filters["filter_var"].tolist():
                        if pd.notna(v) and str(v).strip():
                            filter_vars.append(str(v).strip())

                corr_vars = []
                if "correlation_var" in df_corr.columns:
                    for v in df_corr["correlation_var"].tolist():
                        if pd.notna(v) and str(v).strip():
                            corr_vars.append(str(v).strip())

            # Restore Q1
            q1 = str(general.get("q1_var", "")).strip()
            if q1 and q1 in self.df.columns:
                self._set_q1(q1)
            else:
                self._set_q1(None)
                self.lbl_q1.setText("ยังไม่ได้เลือก")
                self.lbl_q1.setStyleSheet("font-size: 14px; color: #c62828; padding: 4px;")

            # Restore Dimensions
            for d in DEFAULT_DIMENSIONS:
                prefix = d["prefix"]
                sel = [v for v in dim_selected.get(prefix, []) if v in self.df.columns]
                self._dim_selected[prefix] = sel
                lbl = self._dim_labels[prefix]
                if sel:
                    lbl.setText(f"✅ {len(sel)} ตัวแปร: {', '.join(sel[:5])}{'...' if len(sel) > 5 else ''}")
                    lbl.setStyleSheet("font-size: 13px; color: #2e7d32; padding: 2px 8px; font-weight: 600;")
                else:
                    lbl.setText("ยังไม่ได้เลือก")
                    lbl.setStyleSheet("font-size: 13px; color: #757575; padding: 2px 8px;")

            # Restore Filter config
            self.filter_var_list.clear()
            labels = self.meta.column_names_to_labels if self.meta else {}
            for var in filter_vars:
                if var in self.df.columns:
                    lbl = labels.get(var, "")
                    item = QListWidgetItem(f"[{var}] {lbl}")
                    item.setData(Qt.ItemDataRole.UserRole, var)
                    self.filter_var_list.addItem(item)

            main_var = str(general.get("subgroup_main_var", "")).strip()
            if main_var and main_var in self.df.columns:
                self._set_subgroup_main_var(main_var)
            else:
                self._selected_subgroup_main = None
                self.lbl_subgroup_main.setText("ยังไม่ได้เลือก")
                self.lbl_subgroup_main.setStyleSheet("font-size: 13px; color: #757575; padding: 2px 8px;")

            # Restore Correlation variables
            self._selected_q3 = [v for v in corr_vars if v in self.df.columns]
            if self._selected_q3:
                self.lbl_q3.setText(f"✅ {len(self._selected_q3)} ตัวแปร: {', '.join(self._selected_q3[:5])}{'...' if len(self._selected_q3) > 5 else ''}")
                self.lbl_q3.setStyleSheet("font-size: 13px; color: #2e7d32; padding: 2px; font-weight: 600;")
            else:
                self.lbl_q3.setText("ยังไม่ได้เลือก")
                self.lbl_q3.setStyleSheet("font-size: 13px; color: #757575; padding: 2px;")

            # Restore advanced
            self.combo_rotation.setCurrentText("equamax")
            min_n = general.get("min_n", 30)
            try:
                self.spin_min_n.setValue(int(float(min_n)))
            except Exception:
                pass

            QMessageBox.information(self, "สำเร็จ", "โหลดการตั้งค่าสำเร็จ (ตัวแปรที่ไม่มีในไฟล์ปัจจุบันจะถูกข้าม)")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"โหลดการตั้งค่าไม่สำเร็จ:\n{str(e)}")

    # ── Export Excel ──────────────────────────
    def _export_excel(self):
        if not self.results:
            QMessageBox.warning(self, "Warning", "ยังไม่มีผลลัพธ์ กรุณารันการวิเคราะห์ก่อน")
            return
        date_text = datetime.now().strftime("%d %b")
        default_name = f"Brand Space (Well Bing) {date_text}.xlsx"
        fp, _ = QFileDialog.getSaveFileName(self, "บันทึก Excel", default_name, "Excel Files (*.xlsx)")
        if not fp:
            return
        try:
            self._do_export(fp)
            self.statusBar().showMessage(f"Export แล้ว: {fp}")
            QMessageBox.information(self, "สำเร็จ", f"บันทึกไฟล์แล้ว:\n{fp}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Export ไม่สำเร็จ:\n{str(e)}")

    def _do_export(self, filepath):
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # 1) Index first
            self._write_index_sheet(writer)
            # 2) Correlations right after Index
            self._write_corr_sheet(writer, "Correlations", self.corr_data.get("q3", {}))
            # 3) Rotated component matrix
            self._write_rotated_matrix_sheet(writer)
            # 4) Coefficients (renamed)
            self._write_coefficients_sheet(writer, sheet_name="Regression Beta")
            # 5) Factor + Regression output log (renamed)
            self._write_factor_output_sheet(writer, sheet_name="Factor_Regress_Output")

    def _write_rotated_matrix_sheet(self, writer):
        """Write Rotated Component Matrix (SPSS-style) for each subgroup."""
        if not self.results:
            return
        sheet_name = "Rotated_Component"
        all_rows = []
        for r in self.results:
            rdf = r.rotated_component_matrix
            if rdf is None:
                continue
            L_tmp = rdf.values
            L_tmp = _reorder_components_for_brandspace(L_tmp, list(rdf.index))
            rdf = pd.DataFrame(L_tmp, index=rdf.index, columns=[f"Component {j+1}" for j in range(L_tmp.shape[1])])
            # Subgroup header
            all_rows.append({"Variable": f"--- {r.subgroup_name} (N={r.n}) ---"})
            # Title row
            all_rows.append({"Variable": "Rotated Component Matrix"})
            # Determine sort order (by dominant component, SPSS FORMAT SORT)
            L = rdf.values
            n_comp = L.shape[1]
            row_order = _get_rotated_matrix_row_order(L, list(rdf.index))
            # Data rows — suppress values < 0.4 (SPSS FORMAT BLANK(.4))
            for i_row in row_order:
                row = {"Variable": rdf.index[i_row]}
                for j, col_name in enumerate(rdf.columns):
                    val = L[i_row, j]
                    row[col_name] = round(float(val), 3) if abs(val) >= 0.4 else ""
                all_rows.append(row)
            # Footer
            all_rows.append({"Variable": "Extraction Method: Principal Component Analysis."})
            all_rows.append({"Variable": "Rotation Method: Equamax with Kaiser Normalization."})
            all_rows.append({})  # blank row

        if not all_rows:
            pd.DataFrame({"Note": ["No data"]}).to_excel(writer, sheet_name=sheet_name, index=False)
            return
        df_out = pd.DataFrame(all_rows)
        df_out.to_excel(writer, sheet_name=sheet_name, index=False)

        # Style
        ws = writer.sheets[sheet_name]
        hdr_fill = PatternFill(start_color="495057", end_color="495057", fill_type="solid")
        hdr_font = XlFont(bold=True, color="FFFFFF")
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 28
        for col in range(2, ws.max_column + 1):
            ws.column_dimensions[chr(64 + col)].width = 14

        # Bold the subgroup header rows and title rows
        bold_font = XlFont(bold=True, color="1A237E")
        title_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
        for row_idx in range(2, ws.max_row + 1):
            cell_val = str(ws.cell(row=row_idx, column=1).value or "")
            if cell_val.startswith("---") or cell_val == "Rotated Component Matrix":
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col).font = bold_font
                    ws.cell(row=row_idx, column=col).fill = title_fill

    def _write_coefficients_sheet(self, writer, sheet_name="Coefficients"):
        """Write SPSS-style Coefficients table for each subgroup."""
        if not self.results:
            return
        all_rows = []
        for r in self.results:
            cdf = r.coefficients_table
            if cdf is None:
                continue
            # Subgroup header
            all_rows.append({col: "" for col in cdf.columns})
            all_rows[-1]["Variable"] = f"--- {r.subgroup_name} (N={r.n}) ---"
            # Title
            title_row = {col: "" for col in cdf.columns}
            title_row["Variable"] = "Coefficients (Dependent Variable: NQ1/Feeling)"
            all_rows.append(title_row)
            # Data rows
            for _, row in cdf.iterrows():
                all_rows.append(row.to_dict())
            # Model summary row
            rs = r.regression_stats
            summary_row = {col: "" for col in cdf.columns}
            summary_row["Variable"] = f"R²={rs.get('r_squared',0):.4f}, Adj R²={rs.get('adj_r_squared',0):.4f}, F={rs.get('f_statistic',0):.2f}, Sig.={rs.get('f_pvalue',1):.4f}"
            all_rows.append(summary_row)
            all_rows.append({col: "" for col in cdf.columns})  # blank row

        if not all_rows:
            pd.DataFrame({"Note": ["No data"]}).to_excel(writer, sheet_name=sheet_name, index=False)
            return
        df_out = pd.DataFrame(all_rows)
        df_out.to_excel(writer, sheet_name=sheet_name, index=False)

        # Style
        ws = writer.sheets[sheet_name]
        hdr_fill = PatternFill(start_color="495057", end_color="495057", fill_type="solid")
        hdr_font = XlFont(bold=True, color="FFFFFF")
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 24
        for col in range(2, ws.max_column + 1):
            letter = chr(64 + col) if col <= 26 else chr(64 + (col - 1) // 26) + chr(65 + (col - 1) % 26)
            ws.column_dimensions[letter].width = 12

        # Bold subgroup headers
        bold_font = XlFont(bold=True, color="1A237E")
        title_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
        for row_idx in range(2, ws.max_row + 1):
            cell_val = str(ws.cell(row=row_idx, column=1).value or "")
            if cell_val.startswith("---") or cell_val.startswith("Coefficients") or cell_val.startswith("R²="):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col).font = bold_font
                    ws.cell(row=row_idx, column=col).fill = title_fill

    def _write_corr_sheet(self, writer, sheet_name, corr_dict):
        if not corr_dict:
            pd.DataFrame({"Note": ["No data"]}).to_excel(writer, sheet_name=sheet_name, index=False)
            return
        sg_names = list(corr_dict.keys())
        first = corr_dict[sg_names[0]]
        items = first["item"].tolist()
        rows_data = []
        for item_name in items:
            row = {"": item_name, "Statistic": "Pearson Correlation"}
            for sg in sg_names:
                df_c = corr_dict[sg]
                match = df_c[df_c["item"] == item_name]
                if not match.empty:
                    val = match.iloc[0]["r"]
                    row[sg] = round(val, 3) if pd.notna(val) else ""
                else:
                    row[sg] = ""
            rows_data.append(row)
        pd.DataFrame(rows_data).to_excel(writer, sheet_name=sheet_name, index=False)

        # Style
        ws = writer.sheets[sheet_name]
        hdr_fill = PatternFill(start_color="495057", end_color="495057", fill_type="solid")
        hdr_font = XlFont(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="BDBDBD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 34

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 22
        for col in range(3, ws.max_column + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 11

        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=row, column=col)
                c.border = border
                if row >= 2 and col >= 3 and c.value not in ("", None):
                    c.number_format = "0.000"

        ws.freeze_panes = "C3"

    def _write_regression_sheet(self, writer):
        if not self.results:
            return
        dim_names = list(self.results[0].std_betas.keys())
        rows = []
        for r in self.results:
            rs = r.regression_stats
            row = {
                "Subgroup": r.subgroup_name,
                "R²": round(rs.get("r_squared", 0), 4),
                "Adj R²": round(rs.get("adj_r_squared", 0), 4),
                "F": round(rs.get("f_statistic", 0), 2),
                "Sig.": round(rs.get("f_pvalue", 1), 4),
            }
            for d in dim_names:
                row[f"Std Β: {d}"] = round(r.std_betas.get(d, 0), 4)
            rows.append(row)
            
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name="Regression", index=False)
        
        # Style
        ws = writer.sheets["Regression"]
        hdr_fill = PatternFill(start_color="495057", end_color="495057", fill_type="solid")
        hdr_font = XlFont(bold=True, color="FFFFFF")
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 22
        for col_letter in ["B", "C", "D", "E"]:
            ws.column_dimensions[col_letter].width = 10
        for col in range(6, ws.max_column + 1):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _write_index_sheet(self, writer):
        if not self.results:
            return
        dim_names = list(self.results[0].dim_means.keys())
        rows = []
        for r in self.results:
            row = {
                "Subgroup": r.subgroup_name,
                "N": r.n,
                "Index": round(r.index_value, 3),
            }
            for d in dim_names:
                row[f"Score: {d}"] = round(r.dim_means.get(d, 0), 3)
            for d in dim_names:
                row[f"Weight: {d}"] = round(r.dim_weights.get(d, 0), 4)
            row["Check Sum"] = ""
            rows.append(row)
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name="Index", index=False, startrow=1)

        ws = writer.sheets["Index"]
        # Merge header row for Score and Weight groups
        n_dims = len(dim_names)
        ws.cell(row=1, column=1, value="")
        ws.cell(row=1, column=2, value="")
        ws.cell(row=1, column=3, value="")
        if n_dims > 0:
            score_start = 4
            score_end = 3 + n_dims
            weight_start = 4 + n_dims
            weight_end = 3 + 2 * n_dims
            ws.merge_cells(start_row=1, start_column=score_start, end_row=1, end_column=score_end)
            ws.cell(row=1, column=score_start, value="score (1-7)").font = XlFont(bold=True)
            ws.cell(row=1, column=score_start).alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=1, start_column=weight_start, end_row=1, end_column=weight_end)
            ws.cell(row=1, column=weight_start, value="Weight").font = XlFont(bold=True)
            ws.cell(row=1, column=weight_start).alignment = Alignment(horizontal="center")
            ws.cell(row=1, column=weight_end + 1, value="Check Sum").font = XlFont(bold=True)
            ws.cell(row=1, column=weight_end + 1).alignment = Alignment(horizontal="center")

        # Style header row 2
        hdr_fill = PatternFill(start_color="495057", end_color="495057", fill_type="solid")
        hdr_font = XlFont(bold=True, color="FFFFFF")
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        # Score fill
        score_fill = PatternFill(start_color="E7F5FF", end_color="E7F5FF", fill_type="solid")
        weight_fill = PatternFill(start_color="FFF9DB", end_color="FFF9DB", fill_type="solid")
        check_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        check_col = 4 + 2 * n_dims
        for row in range(3, ws.max_row + 1):
            for col in range(4, 4 + n_dims):
                ws.cell(row=row, column=col).fill = score_fill
                ws.cell(row=row, column=col).number_format = "0.000"
            for col in range(4 + n_dims, 4 + 2 * n_dims):
                ws.cell(row=row, column=col).fill = weight_fill
                ws.cell(row=row, column=col).number_format = "0.0%"
            weight_start_col_letter = ws.cell(row=2, column=4 + n_dims).column_letter
            weight_end_col_letter = ws.cell(row=2, column=3 + 2 * n_dims).column_letter
            ws.cell(row=row, column=check_col).value = f"=SUM({weight_start_col_letter}{row}:{weight_end_col_letter}{row})"
            ws.cell(row=row, column=check_col).number_format = "0.0%"
            ws.cell(row=row, column=check_col).fill = check_fill

        # Index color scale (red -> yellow -> green) by rank
        if ws.max_row >= 3:
            idx_values = []
            for row in range(3, ws.max_row + 1):
                v = ws.cell(row=row, column=3).value
                try:
                    idx_values.append(float(v))
                except Exception:
                    idx_values.append(np.nan)
            valid_vals = [v for v in idx_values if pd.notna(v)]
            min_v = min(valid_vals) if valid_vals else None
            max_v = max(valid_vals) if valid_vals else None
            for row in range(3, ws.max_row + 1):
                v = idx_values[row - 3]
                if pd.isna(v) or min_v is None or max_v is None:
                    continue
                t = 0.5 if max_v == min_v else (v - min_v) / (max_v - min_v)
                if t <= 0.5:
                    # red -> yellow
                    a = t / 0.5
                    r, g, b = 244, int(67 + (197 - 67) * a), int(54 + (80 - 54) * a)
                else:
                    # yellow -> green
                    a = (t - 0.5) / 0.5
                    r, g, b = int(244 + (102 - 244) * a), int(197 + (187 - 197) * a), int(80 + (106 - 80) * a)
                ws.cell(row=row, column=3).fill = PatternFill(
                    start_color=f"{r:02X}{g:02X}{b:02X}",
                    end_color=f"{r:02X}{g:02X}{b:02X}",
                    fill_type="solid"
                )

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions[ws.cell(row=2, column=check_col).column_letter].width = 12
        ws.freeze_panes = "D3"

    def _write_factor_output_sheet(self, writer, sheet_name="Factor_Output"):
        if not self.results:
            return
        
        all_logs = []
        for r in self.results:
            if r.factor_output_log:
                all_logs.append(r.factor_output_log)
                
        if not all_logs:
            pd.DataFrame({"Note": ["No log available"]}).to_excel(writer, sheet_name=sheet_name, index=False)
            return

        full_log = "\n".join(all_logs)
        output_lines = full_log.splitlines()
        safe_lines = ["'" + line if line.strip().startswith(('=', '-', '+', '@')) else line for line in output_lines]
        output_df = pd.DataFrame(safe_lines, columns=["Analysis Log"])
        output_df.to_excel(writer, sheet_name=sheet_name, index=False)


# ══════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════
def run_this_app(working_dir=None):
    print("--- Starting Brand Space - Index Caculator ---")
    try:
        app = QApplication(sys.argv)
        app.setStyle("Fusion")
        try:
            app.setWindowIcon(QIcon(_resource_path("BrandS.ico")))
        except Exception:
            pass
        window = WellbeingIndexApp()
        window.show()
        sys.exit(app.exec())
    except Exception as e:
        print(f"ERROR: {e}")
        traceback.print_exc()
        try:
            import tkinter as tk
            from tkinter import messagebox
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Application Error", f"Error:\n{e}", parent=root)
            root.destroy()
        except Exception:
            pass
        sys.exit(1)


if __name__ == "__main__":
    run_this_app()

