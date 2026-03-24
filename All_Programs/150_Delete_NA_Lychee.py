import os
import sys
import threading
import time
import multiprocessing
from concurrent.futures import ProcessPoolExecutor, as_completed
from html import escape

MISSING_DEPENDENCIES = []

try:
    from PyQt6.QtCore import Qt, QObject, QThread, pyqtSignal
    from PyQt6.QtGui import QIcon, QPixmap, QPainter, QColor, QLinearGradient
    from PyQt6.QtWidgets import (
        QApplication,
        QMainWindow,
        QWidget,
        QLabel,
        QVBoxLayout,
        QHBoxLayout,
        QGroupBox,
        QListWidget,
        QListWidgetItem,
        QPushButton,
        QProgressBar,
        QTextEdit,
        QFileDialog,
        QMessageBox,
        QCheckBox,
        QAbstractItemView,
        QStyle,
    )
except ImportError:
    MISSING_DEPENDENCIES.append("PyQt6")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from copy import copy
except ImportError:
    openpyxl = None
    Font = PatternFill = Border = Side = None
    copy = None
    MISSING_DEPENDENCIES.append("openpyxl")

try:
    import win32com.client as win32
    HAS_WIN32COM = True
except Exception:
    win32 = None
    HAS_WIN32COM = False

if MISSING_DEPENDENCIES:
    missing = ", ".join(MISSING_DEPENDENCIES)
    raise SystemExit(
        f"Missing required dependencies: {missing}\n"
        f"Please install them first with: pip install {missing}"
    )

EXCLUDED_SHEETS = ("Contents", "Info", "Conte Tnfo")


# ================================================================
#  Worker function (runs in separate PROCESS for true parallelism)
# ================================================================
def process_single_file(fpath, log_q):
    """Process one Excel file with real-time logging via multiprocessing Queue."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side

    fname = os.path.basename(fpath)
    log_q.put((f"  [{fname}] Loading ...", None))

    wb = openpyxl.load_workbook(fpath, rich_text=False, keep_links=False)
    total_sheets = sum(1 for ws in wb.worksheets
                       if ws.title not in EXCLUDED_SHEETS)
    log_q.put((f"  [{fname}] Loaded ({total_sheets} sheets)", None))

    import traceback as _tb

    log_q.put((f"  [{fname}] Step 1: Unmerge & Shift ...", None))
    _step1(wb, log_q, fname)

    log_q.put((f"  [{fname}] Step 2: Delete TOTAL/NA columns ...", None))
    _step2(wb, log_q, fname)

    log_q.put((f"  [{fname}] Step 3: Delete rows & renumber ...", None))
    try:
        _step3(wb, log_q, fname)
    except Exception as e:
        log_q.put((f"  [{fname}] S3 ERROR: {e}\n{''.join(_tb.format_exc())}", "err"))
        raise

    log_q.put((f"  [{fname}] Saving ...", None))
    wb.save(fpath)
    wb.close()
    log_q.put((f"  [{fname}] Saved -> {fname}", "ok"))


def process_single_file_com(fpath, log_q):
    """Fast path via Excel COM + in-workbook VBA execution."""
    if not HAS_WIN32COM:
        raise RuntimeError("win32com is not available on this machine")

    fname = os.path.basename(fpath)
    log_q.put((f"  [{fname}] Fast COM mode: Loading ...", None))

    vba_code = r'''
Option Explicit

Public Sub CodexProcessWorkbook()
    Application.ScreenUpdating = False
    Application.EnableEvents = False
    On Error Resume Next
    Application.Calculation = xlCalculationManual
    On Error GoTo 0

    UnmergeAndShiftRight_FirstMergedRow
    DeleteExtraTotalColumns
    DeleteEmptyRowsAndRenumber

    On Error Resume Next
    Application.Calculation = xlCalculationAutomatic
    On Error GoTo 0
End Sub

Private Function IsExcludedSheet(ByVal wsName As String) As Boolean
    IsExcludedSheet = (wsName = "Contents" Or wsName = "Info" Or wsName = "Conte Tnfo")
End Function

Private Function CellText(ByVal v As Variant) As String
    CellText = Trim$(CStr(v & ""))
End Function

Private Function IsPercentSheet(ByVal ws As Worksheet) As Boolean
    Dim maxRowScan As Long
    Dim maxColScan As Long
    Dim r As Long, c As Long
    Dim txt As String

    maxRowScan = Application.WorksheetFunction.Min(8, ws.UsedRange.Rows.Count + ws.UsedRange.Row - 1)
    maxColScan = Application.WorksheetFunction.Min(3, ws.UsedRange.Columns.Count + ws.UsedRange.Column - 1)

    For r = 1 To maxRowScan
        For c = 1 To maxColScan
            txt = UCase$(CellText(ws.Cells(r, c).Value2))
            If InStr(txt, "1ST ROW:") > 0 And InStr(txt, "%") > 0 Then
                IsPercentSheet = True
                Exit Function
            End If
        Next c
    Next r
End Function

Private Sub FixPercentLastHeader(ByVal ws As Worksheet)
    Dim area As Range
    Dim targetRange As Range
    Dim sourceCell As Range
    Dim c As Long
    Dim lastHeaderRow As Long
    Dim lastHeaderCol As Long
    Dim lastHeaderWidth As Long
    Dim codeStartCol As Long
    Dim titleText As String
    Dim subRow As Long
    Dim codesRow As Long
    Dim current1 As String, current2 As String, current3 As String
    Dim subHeaderOk As Boolean
    Dim bannerOk As Boolean

    If Not IsPercentSheet(ws) Then Exit Sub

    For c = 3 To ws.Cells(5, ws.Columns.Count).End(xlToLeft).Column
        If ws.Cells(5, c).MergeCells Then
            Set area = ws.Cells(5, c).MergeArea
            If area.Row <= 5 And area.Row + area.Rows.Count - 1 <= 5 Then
                If UCase$(CellText(area.Cells(1, 1).Value2)) = "NUMBER OF CATS OWNED" Then
                    lastHeaderRow = area.Row
                    lastHeaderCol = area.Column
                    lastHeaderWidth = area.Columns.Count
                    Exit For
                End If
            End If
        End If
    Next c

    If lastHeaderCol = 0 Then Exit Sub

    titleText = UCase$(CellText(ws.Cells(lastHeaderRow, lastHeaderCol).Value2))
    subRow = lastHeaderRow + 1
    codesRow = lastHeaderRow - 1
    If subRow < 1 Or codesRow < 1 Then Exit Sub

    codeStartCol = lastHeaderCol
    For c = Application.WorksheetFunction.Max(1, lastHeaderCol - 1) To lastHeaderCol + 1
        If CellText(ws.Cells(codesRow, c).Value2) = "1" And _
           CellText(ws.Cells(codesRow, c + 1).Value2) = "2" And _
           CellText(ws.Cells(codesRow, c + 2).Value2) = "3" Then
            codeStartCol = c
            Exit For
        End If
    Next c

    current1 = CellText(ws.Cells(subRow, codeStartCol).Value2)
    current2 = CellText(ws.Cells(subRow, codeStartCol + 1).Value2)
    current3 = CellText(ws.Cells(subRow, codeStartCol + 2).Value2)

    subHeaderOk = (current1 = "1" And current2 = "2-3" And current3 = "4+")
    bannerOk = (lastHeaderCol = codeStartCol And lastHeaderWidth = 3)

    If titleText = "NUMBER OF CATS OWNED" Then
        Set sourceCell = ws.Cells(lastHeaderRow, lastHeaderCol)
        If Not bannerOk Then
            If ws.Cells(lastHeaderRow, lastHeaderCol).MergeCells Then
                ws.Cells(lastHeaderRow, lastHeaderCol).MergeArea.UnMerge
            End If
            Set targetRange = ws.Range(ws.Cells(lastHeaderRow, codeStartCol), ws.Cells(lastHeaderRow, codeStartCol + 2))
            targetRange.Merge
            targetRange.HorizontalAlignment = sourceCell.HorizontalAlignment
            targetRange.VerticalAlignment = sourceCell.VerticalAlignment
            targetRange.WrapText = sourceCell.WrapText
            targetRange.Orientation = sourceCell.Orientation
            targetRange.AddIndent = sourceCell.AddIndent
            targetRange.IndentLevel = sourceCell.IndentLevel
            targetRange.ShrinkToFit = sourceCell.ShrinkToFit
            targetRange.ReadingOrder = sourceCell.ReadingOrder
            targetRange.Font.Name = sourceCell.Font.Name
            targetRange.Font.Size = sourceCell.Font.Size
            targetRange.Font.Bold = sourceCell.Font.Bold
            targetRange.Font.Italic = sourceCell.Font.Italic
            targetRange.Font.Color = sourceCell.Font.Color
            targetRange.Interior.Color = sourceCell.Interior.Color
            targetRange.Borders.LineStyle = sourceCell.Borders.LineStyle
            targetRange.Borders.Weight = sourceCell.Borders.Weight
            ws.Cells(lastHeaderRow, codeStartCol).Value2 = "Number of cats owned"
        End If
        If Not subHeaderOk Then
            ws.Cells(subRow, codeStartCol).Value2 = "1"
            ws.Cells(subRow, codeStartCol + 1).Value2 = "2-3"
            ws.Cells(subRow, codeStartCol + 2).Value2 = "4+"
        End If
    End If
End Sub

Private Sub UnmergeAndShiftRight_FirstMergedRow()
    Dim ws As Worksheet
    Dim firstMergedRow As Long
    Dim refLastCol As Long
    Dim mergeCount As Long
    Dim c As Long
    Dim area As Range
    Dim key As String
    Dim seen As Object
    Dim cellValue As Variant
    Dim isPercent As Boolean
    Dim shouldShift As Boolean

    For Each ws In ActiveWorkbook.Worksheets
        If Not IsExcludedSheet(ws.Name) Then
            isPercent = IsPercentSheet(ws)
            firstMergedRow = 0
            Dim lastRowC As Long
            lastRowC = ws.Cells(ws.Rows.Count, 3).End(xlUp).Row
            For c = 2 To lastRowC
                If ws.Cells(c, 3).MergeCells Then
                    firstMergedRow = c
                    Exit For
                End If
            Next c
            If firstMergedRow = 0 Then GoTo NextSheetS1

            refLastCol = ws.Cells(firstMergedRow - 1, ws.Columns.Count).End(xlToLeft).Column
            Set seen = CreateObject("Scripting.Dictionary")
            mergeCount = 0

            For c = 1 To refLastCol
                If ws.Cells(firstMergedRow, c).MergeCells Then
                    Set area = ws.Cells(firstMergedRow, c).MergeArea
                    key = area.Address(False, False)
                    If Not seen.Exists(key) Then
                        seen.Add key, 1
                        mergeCount = mergeCount + 1
                        shouldShift = False
                        If isPercent Then
                            shouldShift = (UCase$(CellText(ws.Cells(firstMergedRow + 1, area.Column).Value2)) = "TOTAL")
                        ElseIf mergeCount >= 3 Then
                            shouldShift = True
                        End If
                        If shouldShift Then
                            cellValue = area.Cells(1, 1).Value2
                            Dim minCol As Long
                            minCol = area.Column
                            area.UnMerge
                            ws.Cells(firstMergedRow, minCol + 1).Value2 = cellValue
                            ws.Cells(firstMergedRow, minCol).Value2 = Empty
                        End If
                    End If
                End If
            Next c

            With ws.Range(ws.Cells(firstMergedRow, 1), ws.Cells(firstMergedRow, refLastCol))
                .Interior.Color = RGB(235, 235, 235)
                .Font.Bold = True
            End With
        End If
NextSheetS1:
    Next ws
End Sub

Private Sub DeleteExtraTotalColumns()
    Dim ws As Worksheet
    Dim targetRow As Long, lastCol As Long
    Dim vals As Variant, rowVals As Variant
    Dim i As Long, c As Long
    Dim cols() As Long, n As Long
    Dim seenTotal As Boolean
    Dim startC As Long, cnt As Long

    For Each ws In ActiveWorkbook.Worksheets
        If Not IsExcludedSheet(ws.Name) Then
            targetRow = 0
            vals = ws.Range("C1:C1000").Value2
            For i = 1 To UBound(vals, 1)
                If UCase$(Trim$(CStr(vals(i, 1)))) = "TOTAL" Then
                    targetRow = i
                    Exit For
                End If
            Next i
            If targetRow = 0 Then GoTo NextSheetS2

            If targetRow > 1 Then
                lastCol = ws.Cells(targetRow - 1, ws.Columns.Count).End(xlToLeft).Column
                rowVals = ws.Range(ws.Cells(targetRow - 1, 1), ws.Cells(targetRow - 1, lastCol)).Value2
                For c = 1 To lastCol
                    If Len(Trim$(CStr(rowVals(1, c)))) > 0 Then
                        ws.Cells(targetRow - 1, c).Borders(xlEdgeLeft).LineStyle = xlContinuous
                        ws.Cells(targetRow - 1, c).Borders(xlEdgeLeft).Weight = xlThin
                    End If
                Next c
            End If

            lastCol = ws.Cells(targetRow, ws.Columns.Count).End(xlToLeft).Column
            rowVals = ws.Range(ws.Cells(targetRow, 1), ws.Cells(targetRow, lastCol)).Value2
            seenTotal = False
            n = 0
            Erase cols
            For c = 1 To lastCol
                Dim s As String
                s = UCase$(Trim$(CStr(rowVals(1, c))))
                If s = "TOTAL" Then
                    If seenTotal Then
                        n = n + 1
                        ReDim Preserve cols(1 To n)
                        cols(n) = c
                    Else
                        seenTotal = True
                    End If
                ElseIf s = "NA" Then
                    n = n + 1
                    ReDim Preserve cols(1 To n)
                    cols(n) = c
                End If
            Next c

            If n > 0 Then
                startC = cols(n)
                cnt = 1
                For i = n - 1 To 1 Step -1
                    If cols(i) = startC - cnt Then
                        cnt = cnt + 1
                    Else
                        ws.Range(ws.Cells(1, startC - cnt + 1), ws.Cells(1, startC)).EntireColumn.Delete
                        startC = cols(i)
                        cnt = 1
                    End If
                Next i
                ws.Range(ws.Cells(1, startC - cnt + 1), ws.Cells(1, startC)).EntireColumn.Delete
            End If

            lastCol = ws.Cells(targetRow, ws.Columns.Count).End(xlToLeft).Column
            If lastCol > 0 And targetRow >= 3 Then
                With ws.Range(ws.Cells(targetRow - 2, lastCol), ws.Cells(targetRow - 1, lastCol)).Borders
                    .LineStyle = xlContinuous
                    .Weight = xlThin
                End With
            End If

            FixPercentLastHeader ws
        End If
NextSheetS2:
    Next ws
End Sub

Private Sub DeleteEmptyRowsAndRenumber()
    Dim ws As Worksheet
    Dim lastRow As Long, r As Long
    Dim lastGroupMin As Long, lastGroupMax As Long
    Dim hasLastGroup As Boolean
    Dim delRows As Object
    Dim area As Range
    Dim minR As Long, maxR As Long
    Dim val As String
    Dim rr As Long
    Dim keys As Variant
    Dim i As Long
    Dim startRun As Long, endRun As Long
    Dim totalRow As Long
    Dim rowNum As Long
    Dim startAfter As Long
    Dim visited As Object
    Dim nextVal As String
    Dim isPercent As Boolean
    Dim lastDataStartRow As Long

    For Each ws In ActiveWorkbook.Worksheets
        If Not IsExcludedSheet(ws.Name) Then
            isPercent = IsPercentSheet(ws)
            lastRow = ws.Cells(ws.Rows.Count, 2).End(xlUp).Row
            If lastRow < 8 Then GoTo NextSheetS3

            hasLastGroup = False
            r = lastRow
            Do While r > 7
                If ws.Cells(r, 2).MergeCells Then
                    Set area = ws.Cells(r, 2).MergeArea
                    minR = area.Row
                    maxR = minR + area.Rows.Count - 1
                    val = UCase$(Trim$(CStr(area.Cells(1, 1).Value2)))
                    If val <> "" And val <> "NA" Then
                        hasLastGroup = True
                        lastGroupMin = minR
                        lastGroupMax = maxR
                        Exit Do
                    End If
                    r = minR - 1
                Else
                    val = UCase$(Trim$(CStr(ws.Cells(r, 2).Value2)))
                    If val <> "" And val <> "NA" Then
                        hasLastGroup = True
                        lastGroupMin = r
                        lastGroupMax = r
                        Exit Do
                    End If
                    r = r - 1
                End If
            Loop

            Set delRows = CreateObject("Scripting.Dictionary")
            r = lastRow
            Do While r >= 8
                If ws.Cells(r, 2).MergeCells Then
                    Set area = ws.Cells(r, 2).MergeArea
                    minR = area.Row
                    maxR = minR + area.Rows.Count - 1
                    val = UCase$(Trim$(CStr(area.Cells(1, 1).Value2)))
                    If val = "" Or val = "NA" Then
                        For rr = minR To maxR
                            delRows(CStr(rr)) = 1
                        Next rr
                    End If
                    r = minR - 1
                Else
                    val = UCase$(Trim$(CStr(ws.Cells(r, 2).Value2)))
                    If val = "" Or val = "NA" Then
                        delRows(CStr(r)) = 1
                    End If
                    r = r - 1
                End If
            Loop

            If hasLastGroup Then
                For rr = lastGroupMin To lastGroupMax
                    If delRows.Exists(CStr(rr)) Then delRows.Remove CStr(rr)
                Next rr
            End If

            If delRows.Count > 0 Then
                keys = delRows.Keys
                ReDim Preserve keys(0 To UBound(keys))
                Dim nums() As Long
                ReDim nums(0 To UBound(keys))
                For i = 0 To UBound(keys)
                    nums(i) = CLng(keys(i))
                Next i
                QuickSortLong nums, LBound(nums), UBound(nums)

                startRun = nums(UBound(nums))
                endRun = startRun
                For i = UBound(nums) - 1 To LBound(nums) Step -1
                    If nums(i) = endRun - 1 Then
                        endRun = nums(i)
                    Else
                        ws.Rows(CStr(endRun) & ":" & CStr(startRun)).Delete
                        startRun = nums(i)
                        endRun = startRun
                    End If
                Next i
                ws.Rows(CStr(endRun) & ":" & CStr(startRun)).Delete
            End If

            lastRow = ws.Cells(ws.Rows.Count, 2).End(xlUp).Row
            If lastRow < 7 Then GoTo NextSheetS3

            On Error Resume Next
            ws.Range("A7:A" & CStr(lastRow)).UnMerge
            On Error GoTo 0

            totalRow = 0
            For r = 7 To lastRow
                val = UCase$(Trim$(CStr(ws.Cells(r, 2).Value2)))
                If val = "TOTAL" Then
                    totalRow = r
                    Exit For
                End If
            Next r

            rowNum = 0
            If totalRow > 0 Then
                ws.Cells(totalRow, 1).Value2 = rowNum
                If ws.Cells(totalRow, 2).MergeCells Then
                    Set area = ws.Cells(totalRow, 2).MergeArea
                    minR = area.Row
                    maxR = minR + area.Rows.Count - 1
                    If maxR > minR Then
                        ws.Range(ws.Cells(minR, 1), ws.Cells(maxR, 1)).Merge
                    End If
                Else
                    nextVal = Trim$(CStr(ws.Cells(totalRow + 1, 2).Value2))
                    If nextVal = "" Then
                        ws.Range(ws.Cells(totalRow, 1), ws.Cells(totalRow + 1, 1)).Merge
                    End If
                End If
                rowNum = rowNum + 1
            End If

            startAfter = IIf(totalRow > 0, totalRow + 1, 8)
            Set visited = CreateObject("Scripting.Dictionary")
            lastDataStartRow = 0
            For r = startAfter To lastRow
                If ws.Cells(r, 2).MergeCells Then
                    Set area = ws.Cells(r, 2).MergeArea
                    minR = area.Row
                    maxR = minR + area.Rows.Count - 1
                    If minR < startAfter Then
                        GoTo NextRowS3
                    End If
                    If Not visited.Exists(CStr(minR)) Then
                        visited(CStr(minR)) = 1
                        val = Trim$(CStr(area.Cells(1, 1).Value2))
                        If val <> "" Then
                            lastDataStartRow = minR
                            ws.Cells(minR, 1).Value2 = rowNum
                            If maxR > minR Then
                                ws.Range(ws.Cells(minR, 1), ws.Cells(maxR, 1)).Merge
                            End If
                            rowNum = rowNum + 1
                        End If
                    End If
                Else
                    val = Trim$(CStr(ws.Cells(r, 2).Value2))
                    If val <> "" Then
                        lastDataStartRow = r
                        ws.Cells(r, 1).Value2 = rowNum
                        nextVal = Trim$(CStr(ws.Cells(r + 1, 2).Value2))
                        If nextVal = "" And Not (isPercent And r = lastDataStartRow) Then
                            ws.Range(ws.Cells(r, 1), ws.Cells(r + 1, 1)).Merge
                        End If
                        rowNum = rowNum + 1
                    End If
                End If
            NextRowS3:
            Next r

            If isPercent And lastDataStartRow > 0 Then
                On Error Resume Next
                ws.Range(ws.Cells(lastDataStartRow, 1), ws.Cells(lastDataStartRow + 1, 1)).UnMerge
                On Error GoTo 0
            End If
        End If
NextSheetS3:
    Next ws
End Sub

Private Sub QuickSortLong(ByRef arr() As Long, ByVal lo As Long, ByVal hi As Long)
    Dim i As Long, j As Long, p As Long, t As Long
    i = lo: j = hi: p = arr((lo + hi) \ 2)
    Do While i <= j
        Do While arr(i) < p: i = i + 1: Loop
        Do While arr(j) > p: j = j - 1: Loop
        If i <= j Then
            t = arr(i): arr(i) = arr(j): arr(j) = t
            i = i + 1: j = j - 1
        End If
    Loop
    If lo < j Then QuickSortLong arr, lo, j
    If i < hi Then QuickSortLong arr, i, hi
End Sub
'''

    excel = None
    wb = None
    module_comp = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.EnableEvents = False
        excel.ScreenUpdating = False
        excel.AskToUpdateLinks = False
        wb = excel.Workbooks.Open(fpath, UpdateLinks=0, ReadOnly=False)
        total_sheets = sum(
            1 for ws in wb.Worksheets
            if ws.Name not in EXCLUDED_SHEETS
        )
        log_q.put((f"  [{fname}] Loaded ({total_sheets} sheets)", None))
        log_q.put((f"  [{fname}] Step 1/2/3 in VBA ...", None))

        vb_project = wb.VBProject
        module_comp = vb_project.VBComponents.Add(1)
        module_comp.Name = "CodexFastModule"
        module_comp.CodeModule.AddFromString(vba_code)

        excel.Run(f"'{wb.Name}'!CodexProcessWorkbook")
        vb_project.VBComponents.Remove(module_comp)
        module_comp = None

        log_q.put((f"  [{fname}] Saving ...", None))
        wb.Save()
        wb.Close(SaveChanges=True)
        wb = None
        excel.Quit()
        excel = None
        log_q.put((f"  [{fname}] Saved -> {fname}", "ok"))
    finally:
        try:
            if module_comp is not None and wb is not None:
                wb.VBProject.VBComponents.Remove(module_comp)
        except Exception:
            pass
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass


# ---- helpers ----------------------------------------------------

def _cell_text(value):
    return str(value).strip() if value is not None else ""


def _is_percent_sheet(ws):
    max_row = min(8, ws.max_row or 1)
    max_col = min(3, ws.max_column or 1)
    for row in ws.iter_rows(min_row=1, max_row=max_row,
                            min_col=1, max_col=max_col, values_only=True):
        for value in row:
            text = _cell_text(value).upper()
            if "1ST ROW:" in text and "%" in text:
                return True
    return False


def _fix_percent_last_header(ws, log_q, fname):
    if not _is_percent_sheet(ws):
        return

    header_ranges = [
        mr for mr in ws.merged_cells.ranges
        if mr.min_col >= 3 and mr.min_row <= 5 and mr.max_row <= 5 and mr.max_col > mr.min_col
    ]
    if not header_ranges:
        return

    target_headers = [
        mr for mr in header_ranges
        if _cell_text(ws.cell(mr.min_row, mr.min_col).value).upper() == "NUMBER OF CATS OWNED"
    ]
    if not target_headers:
        return
    last_header = target_headers[0]

    title = _cell_text(ws.cell(last_header.min_row, last_header.min_col).value).upper()
    subheader_row = last_header.max_row + 1
    if subheader_row > (ws.max_row or 1):
        return

    codes_row = max(1, last_header.min_row - 1)
    code_start_col = last_header.min_col
    for col in range(max(1, last_header.min_col - 1), min(ws.max_column - 2, last_header.min_col + 1) + 1):
        codes = [
            _cell_text(ws.cell(codes_row, scan_col).value)
            for scan_col in range(col, col + 3)
        ]
        if codes == ["1", "2", "3"]:
            code_start_col = col
            break

    current = [
        _cell_text(ws.cell(subheader_row, col).value)
        for col in range(code_start_col, code_start_col + 3)
    ]
    desired = ["1", "2-3", "4+"]
    if title != "NUMBER OF CATS OWNED":
        return
    subheader_ok = current == desired
    banner_ok = (
        last_header.min_col == code_start_col
        and (last_header.max_col - last_header.min_col + 1) == 3
    )

    if not banner_ok:
        source_cell = ws.cell(row=last_header.min_row, column=last_header.min_col)
        ws.unmerge_cells(str(last_header))
        for col in range(code_start_col, code_start_col + 3):
            cell = ws.cell(row=last_header.min_row, column=col)
            cell.font = copy(source_cell.font)
            cell.fill = copy(source_cell.fill)
            cell.border = copy(source_cell.border)
            cell.alignment = copy(source_cell.alignment)
            cell.protection = copy(source_cell.protection)
            cell.number_format = source_cell.number_format
        ws.merge_cells(
            start_row=last_header.min_row,
            end_row=last_header.max_row,
            start_column=code_start_col,
            end_column=code_start_col + 2,
        )
        ws.cell(row=last_header.min_row, column=code_start_col).value = "Number of cats owned"

    if not subheader_ok:
        for offset, value in enumerate(desired):
            ws.cell(row=subheader_row, column=code_start_col + offset).value = value

    if not banner_ok or not subheader_ok:
        log_q.put((f"    [{ws.title}] % header fixed -> 1 / 2-3 / 4+", None))


def _last_data_col(ws, row):
    last = 0
    row_data = list(ws.iter_rows(min_row=row, max_row=row,
                                  max_col=ws.max_column or 1,
                                  values_only=False))
    if row_data:
        for cell in row_data[0]:
            if cell.value is not None:
                last = cell.column
    return last


def _last_data_col_fast(ws, row, max_col):
    last = 0
    if max_col < 1:
        return 0
    row_data = list(ws.iter_rows(min_row=row, max_row=row,
                                  max_col=max_col, values_only=True))
    if row_data:
        for idx, v in enumerate(row_data[0], 1):
            if v is not None:
                last = idx
    return last


def _build_merge_cache(ws):
    cache = {}
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                cache[(r, c)] = mr
    return cache


def _build_col_merge_map(ws, col):
    merge_map = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= col <= mr.max_col:
            val = ws.cell(row=mr.min_row, column=mr.min_col).value
            val_s = str(val).strip() if val is not None else ""
            info = (mr.min_row, mr.max_row, val_s, mr)
            for r in range(mr.min_row, mr.max_row + 1):
                merge_map[r] = info
    return merge_map


def _last_data_row_in_col(ws, col, merge_map=None):
    for r in range((ws.max_row or 1), 0, -1):
        if ws.cell(row=r, column=col).value is not None:
            return r
        if merge_map and r in merge_map:
            if merge_map[r][2]:
                return r
    return 1


def _safe_write(ws, row, col, value):
    """Write value to cell, auto-unmerge if the cell is part of a merge."""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                ws.unmerge_cells(str(mr))
                break
    ws.cell(row=row, column=col).value = value


# ---- Step 1 : UnmergeAndShiftRight_FirstMergedRow -----------------
def _step1(wb, log_q, fname):
    exclude = EXCLUDED_SHEETS
    gray = PatternFill("solid", fgColor="EBEBEB")

    for ws in wb.worksheets:
        if ws.title in exclude:
            continue

        is_percent_sheet = _is_percent_sheet(ws)
        cache = _build_merge_cache(ws)

        first_merged_row = 0
        max_r = ws.max_row or 1
        for r in range(2, max_r + 1):
            if (r, 3) in cache:
                first_merged_row = r
                break
        if first_merged_row == 0:
            continue

        ref_last_col = _last_data_col(ws, first_merged_row - 1)
        log_q.put((f"    [{ws.title}] S1 merged={first_merged_row} cols={ref_last_col}", None))

        merge_count = 0
        processed_merges = set()
        for col in range(1, ref_last_col + 1):
            mr = cache.get((first_merged_row, col))
            if mr is None:
                continue
            mr_key = (mr.min_row, mr.min_col, mr.max_row, mr.max_col)
            if mr_key in processed_merges:
                continue
            merge_count += 1
            should_shift = False
            if is_percent_sheet:
                next_row_text = _cell_text(ws.cell(row=first_merged_row + 1, column=mr.min_col).value).upper()
                should_shift = next_row_text == "TOTAL"
            elif merge_count >= 3:
                should_shift = True
            if should_shift:
                val = ws.cell(row=mr.min_row, column=mr.min_col).value
                min_c = mr.min_col
                ws.unmerge_cells(str(mr))
                ws.cell(row=first_merged_row, column=min_c + 1).value = val
                ws.cell(row=first_merged_row, column=min_c).value = None
                processed_merges.add(mr_key)

        for c in range(1, ref_last_col + 1):
            cell = ws.cell(row=first_merged_row, column=c)
            cell.fill = gray
            f = cell.font
            cell.font = Font(name=f.name, size=f.size, bold=True,
                             italic=f.italic, color=f.color)


# ---- Step 2 : DeleteExtraTotalColumns -----------------------------
def _step2(wb, log_q, fname):
    exclude = EXCLUDED_SHEETS
    thin = Side(style="thin")

    for ws in wb.worksheets:
        if ws.title in exclude:
            continue

        target_row = 0
        max_r = min(1000, ws.max_row or 1)
        for row_data in ws.iter_rows(min_row=1, max_row=max_r,
                                      min_col=3, max_col=3, values_only=False):
            cell = row_data[0]
            if cell.value is not None and str(cell.value).strip().upper() == "TOTAL":
                target_row = cell.row
                break
        if target_row == 0:
            continue

        log_q.put((f"    [{ws.title}] S2 TOTAL row={target_row}", None))

        if target_row > 1:
            lc_prev = _last_data_col_fast(ws, target_row - 1, ws.max_column or 1)
            for row_data in ws.iter_rows(min_row=target_row - 1,
                                          max_row=target_row - 1,
                                          max_col=lc_prev, values_only=False):
                for cell in row_data:
                    if cell.value is not None and str(cell.value).strip():
                        b = cell.border
                        cell.border = Border(left=thin, right=b.right,
                                             top=b.top, bottom=b.bottom)

        last_col = _last_data_col_fast(ws, target_row, ws.max_column or 1)
        cols_to_delete = []
        found_total = False
        for row_data in ws.iter_rows(min_row=target_row, max_row=target_row,
                                      max_col=last_col, values_only=False):
            for cell in row_data:
                v = cell.value
                upper_v = str(v).strip().upper() if v is not None else ""
                if upper_v == "TOTAL":
                    if not found_total:
                        found_total = True
                    else:
                        cols_to_delete.append(cell.column)
                elif upper_v == "NA":
                    cols_to_delete.append(cell.column)

        if cols_to_delete:
            log_q.put((f"    [{ws.title}] S2 del {len(cols_to_delete)} cols", None))
            # Batch consecutive columns โ’ fewer delete_cols calls
            # (each call updates ALL merge ranges; fewer calls = faster)
            desc = sorted(cols_to_delete, reverse=True)
            s, cnt = desc[0], 1
            for c in desc[1:]:
                if c == s - cnt:
                    cnt += 1
                else:
                    ws.delete_cols(s - cnt + 1, cnt)
                    s, cnt = c, 1
            ws.delete_cols(s - cnt + 1, cnt)

        last_col = _last_data_col_fast(ws, target_row, ws.max_column or 1)
        if last_col > 0 and target_row >= 3:
            for r in range(target_row - 2, target_row):
                cell = ws.cell(row=r, column=last_col)
                cell.border = Border(left=thin, right=thin,
                                     top=thin, bottom=thin)

        _fix_percent_last_header(ws, log_q, fname)


# ---- Step 3 : DeleteEmptyRowsAndRenumber --------------------------
def _step3(wb, log_q, fname):
    exclude = EXCLUDED_SHEETS

    for ws in wb.worksheets:
        if ws.title in exclude:
            continue

        is_percent_sheet = _is_percent_sheet(ws)

        merge_map = _build_col_merge_map(ws, 2)
        last_row = _last_data_row_in_col(ws, 2, merge_map)

        last_group_min = last_group_max = None
        r = last_row
        while r > 7:
            if r in merge_map:
                min_r, max_r, val, _ = merge_map[r]
                if val and val.upper() != "NA":
                    last_group_min = min_r
                    last_group_max = max_r
                    break
                r = min_r - 1
            else:
                v = ws.cell(row=r, column=2).value
                if v is not None and str(v).strip() and \
                   str(v).strip().upper() != "NA":
                    last_group_min = r
                    last_group_max = r
                    break
                r -= 1

        # Save ALL data-area merges (row 7+) before deletion โ€” restore after
        saved_all_merges = []
        for mr in ws.merged_cells.ranges:
            if mr.min_row >= 7:
                saved_all_merges.append((
                    mr.min_row, mr.max_row, mr.min_col, mr.max_col))


        rows_to_delete = set()
        i = last_row
        while i >= 8:
            if i in merge_map:
                min_r, max_r, val, _ = merge_map[i]
                if val == "" or val.upper() == "NA":
                    for rr in range(min_r, max_r + 1):
                        rows_to_delete.add(rr)
                i = min_r - 1
            else:
                cell_val = ws.cell(row=i, column=2).value
                cell_val = str(cell_val).strip() if cell_val is not None else ""
                if cell_val == "" or cell_val.upper() == "NA":
                    rows_to_delete.add(i)
                i -= 1

        if last_group_min is not None:
            for rr in range(last_group_min, last_group_max + 1):
                rows_to_delete.discard(rr)

        if rows_to_delete:
            max_del = max(rows_to_delete)
            min_del = min(rows_to_delete)
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row > max_del or mr.max_row < min_del:
                    continue
                if last_group_min is not None and \
                   mr.min_row <= last_group_max and \
                   mr.max_row >= last_group_min:
                    continue
                for rr in range(mr.min_row, mr.max_row + 1):
                    if rr in rows_to_delete:
                        ws.unmerge_cells(str(mr))
                        break

            sorted_rows = sorted(rows_to_delete, reverse=True)
            batch_start = sorted_rows[0]
            batch_count = 1
            for idx in range(1, len(sorted_rows)):
                if sorted_rows[idx] == batch_start - batch_count:
                    batch_count += 1
                else:
                    ws.delete_rows(batch_start - batch_count + 1, batch_count)
                    batch_start = sorted_rows[idx]
                    batch_count = 1
            ws.delete_rows(batch_start - batch_count + 1, batch_count)

            log_q.put((f"    [{ws.title}] S3 del {len(rows_to_delete)} rows", None))

            # Restore ALL data-area merges โ€” O(m log d) with bisect
            if saved_all_merges:
                import bisect
                sorted_deleted = sorted(rows_to_delete)
                # Build set of surviving merges for O(1) duplicate check
                existing_merges = {
                    (m.min_row, m.max_row, m.min_col, m.max_col)
                    for m in ws.merged_cells.ranges
                }
                for (orig_mr, orig_maxr, mc, maxc) in saved_all_merges:
                    deleted_above = bisect.bisect_left(sorted_deleted, orig_mr)
                    del_to_max   = bisect.bisect_right(sorted_deleted, orig_maxr)
                    deleted_inside = del_to_max - deleted_above
                    if deleted_inside >= (orig_maxr - orig_mr + 1):
                        continue
                    new_mr   = orig_mr   - deleted_above
                    new_maxr = orig_maxr - deleted_above - deleted_inside
                    if new_mr >= new_maxr:
                        continue
                    key = (new_mr, new_maxr, mc, maxc)
                    if key not in existing_merges:
                        try:
                            ws.merge_cells(
                                start_row=new_mr, end_row=new_maxr,
                                start_column=mc, end_column=maxc)
                            existing_merges.add(key)
                        except Exception:
                            pass

        # Unmerge column A cells in DATA area only (row 7+)
        # Do NOT touch header merges (rows 1-6)
        # Renumber column A
        # Build merge_map once โ€” reused for last_row scan, total_row, and renumber
        merge_map = _build_col_merge_map(ws, 2)
        max_row = ws.max_row or 1

        last_row = 1
        for r in range(max_row, 0, -1):
            v = ws.cell(row=r, column=2).value
            if v is not None and str(v).strip():
                last_row = r
                break
            if r in merge_map and merge_map[r][2]:
                last_row = merge_map[r][1]
                break

        total_row = 0
        for r in range(7, last_row + 1):
            v = ws.cell(row=r, column=2).value
            if v is not None and str(v).strip().upper() == "TOTAL":
                total_row = r
                break

        # Reuse merge_map as merge_map_b (same data, no second build needed)
        merge_map_b = merge_map

        def write_and_merge_col_a(data_row):
            """Write to col A and merge it to match col B span."""
            _safe_write(ws, data_row, 1, row_num)
            # Find col B span at this row
            if data_row in merge_map_b:
                span_max = merge_map_b[data_row][1]
            else:
                # Col B not merged โ€” check if next row is a "%" row (empty B)
                next_val = ws.cell(row=data_row + 1, column=2).value
                if next_val is None or str(next_val).strip() == "":
                    span_max = data_row + 1
                else:
                    span_max = data_row
            if span_max > data_row:
                try:
                    ws.merge_cells(start_row=data_row, end_row=span_max,
                                   start_column=1, end_column=1)
                except Exception:
                    pass

        def col_a_span_end(data_row):
            if data_row in merge_map:
                return merge_map[data_row][1]
            if last_percent_row is not None and data_row == last_percent_row:
                return data_row
            next_val = ws.cell(row=data_row + 1, column=2).value
            if next_val is None or str(next_val).strip() == "":
                return data_row + 1
            return data_row

        desired_rows = []
        if total_row > 0:
            desired_rows.append(total_row)

        start_after = total_row + 1 if total_row > 0 else 8
        for r in range(start_after, last_row + 1):
            v = ws.cell(row=r, column=2).value
            if v is not None and str(v).strip():
                desired_rows.append(r)
            elif r in merge_map:
                min_r = merge_map[r][0]
                if r == min_r and merge_map[r][2]:
                    desired_rows.append(r)

        last_percent_row = desired_rows[-1] if is_percent_sheet and desired_rows else None

        desired_a_merges = {
            (data_row, col_a_span_end(data_row), 1, 1)
            for data_row in desired_rows
            if col_a_span_end(data_row) > data_row
        }

        # Most A-column merges already match after row deletes, so only
        # touch the ranges that actually changed.
        for mr in list(ws.merged_cells.ranges):
            if mr.min_col <= 1 <= mr.max_col and mr.min_row >= 7:
                key = (mr.min_row, mr.max_row, mr.min_col, mr.max_col)
                if key in desired_a_merges:
                    continue
                try:
                    ws.unmerge_cells(str(mr))
                except KeyError:
                    ws.merged_cells.ranges.discard(mr)

        current_a_merges = {
            (mr.min_row, mr.max_row, mr.min_col, mr.max_col)
            for mr in ws.merged_cells.ranges
            if mr.min_col == 1 and mr.max_col == 1 and mr.min_row >= 7
        }

        row_num = 0
        for data_row in desired_rows:
            _safe_write(ws, data_row, 1, row_num)
            span_end = col_a_span_end(data_row)
            key = (data_row, span_end, 1, 1)
            if span_end > data_row and key not in current_a_merges:
                try:
                    ws.merge_cells(start_row=data_row, end_row=span_end,
                                   start_column=1, end_column=1)
                    current_a_merges.add(key)
                except Exception:
                    pass
            row_num += 1

        log_q.put((f"    [{ws.title}] S3 renumbered ({row_num} items)", None))


# ================================================================
#  GUI Application
# ================================================================
class _LogAdapter:
    def __init__(self, emit_fn):
        self.emit_fn = emit_fn

    def put(self, item):
        if isinstance(item, tuple) and len(item) >= 2:
            msg, tag = item[0], item[1]
        else:
            msg, tag = str(item), None
        self.emit_fn(str(msg), tag or "")


class ProcessorWorker(QObject):
    log_signal = pyqtSignal(str, str)
    progress_signal = pyqtSignal(int, int, str)
    finished_signal = pyqtSignal(float, int, int)

    def __init__(self, files, use_fast_com):
        super().__init__()
        self.files = list(files)
        self.use_fast_com = bool(use_fast_com)

    def run(self):
        total = len(self.files)
        t0 = time.perf_counter()
        use_fast_com = bool(self.use_fast_com and HAS_WIN32COM)
        relay_running = None
        relay_thread = None
        mp_manager = None
        failed_files = []

        if self.use_fast_com and not HAS_WIN32COM:
            self.log_signal.emit(
                "  Fast COM mode is unavailable, fallback to openpyxl mode", "err"
            )

        try:
            if use_fast_com:
                log_q = _LogAdapter(self.log_signal.emit)
            else:
                mp_manager = multiprocessing.Manager()
                log_q = mp_manager.Queue()
                relay_running = threading.Event()
                relay_running.set()

                def relay_logs():
                    while relay_running.is_set():
                        try:
                            msg, tag = log_q.get(timeout=0.1)
                            self.log_signal.emit(str(msg), tag or "")
                        except Exception:
                            pass

                relay_thread = threading.Thread(target=relay_logs, daemon=True)
                relay_thread.start()

            self.log_signal.emit(f"{'=' * 52}", "head")
            if use_fast_com:
                self.log_signal.emit(
                    f"  Processing {total} file(s) in Fast COM mode", "head"
                )
            else:
                max_workers = min(total, max(1, os.cpu_count() or 2))
                self.log_signal.emit(
                    f"  Processing {total} file(s) with {max_workers} worker(s)", "head"
                )

            completed = 0
            if use_fast_com:
                for fpath in self.files:
                    fname = os.path.basename(fpath)
                    try:
                        process_single_file_com(fpath, log_q)
                    except Exception as e:
                        self.log_signal.emit(f"  ERROR ({fname}): {e}", "err")
                        failed_files.append(fname)
                    completed += 1
                    self.progress_signal.emit(completed, total, fname)
            else:
                with ProcessPoolExecutor(max_workers=max_workers) as executor:
                    futures = {executor.submit(process_single_file, fp, log_q): fp
                               for fp in self.files}
                    for fut in as_completed(futures):
                        fpath = futures[fut]
                        fname = os.path.basename(fpath)
                        try:
                            fut.result()
                        except Exception as e:
                            self.log_signal.emit(f"  ERROR ({fname}): {e}", "err")
                            failed_files.append(fname)
                        completed += 1
                        self.progress_signal.emit(completed, total, fname)
        finally:
            if relay_running is not None and relay_thread is not None:
                relay_running.clear()
                relay_thread.join(timeout=1)
                try:
                    while True:
                        msg, tag = log_q.get_nowait()
                        self.log_signal.emit(str(msg), tag or "")
                except Exception:
                    pass
            if mp_manager is not None:
                try:
                    mp_manager.shutdown()
                except Exception:
                    pass
            self.finished_signal.emit(
                time.perf_counter() - t0,
                len(failed_files),
                total,
            )


class ExcelProcessorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.files = []
        self.worker_thread = None
        self.worker = None
        self._build_ui()

    def _build_ui(self):
        self.setWindowTitle("โปรแกรมลบ NA+Codeกระโดด ในLychee")
        self.setMinimumSize(900, 680)
        self.resize(980, 760)
        self.setWindowIcon(self._make_app_icon())
        self.setStyleSheet(self._style_sheet())

        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(20, 20, 20, 20)
        root.setSpacing(12)

        title = QLabel("โปรแกรมลบ NA+Codeกระโดด ในLychee")
        title.setObjectName("Title")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        root.addWidget(title)

        subtitle = QLabel(
            "โหลดไฟล์ Table จาก Lychee เข้าไปพร้อมกันได้หลายไฟล์ทั้ง N%+%พร้อมกันได้"
        )
        subtitle.setObjectName("Subtitle")
        subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        root.addWidget(subtitle)

        file_group = QGroupBox("Excel Files")
        fg_layout = QVBoxLayout(file_group)
        self.file_list = QListWidget()
        self.file_list.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        fg_layout.addWidget(self.file_list, 1)

        btn_row = QHBoxLayout()
        self.add_btn = QPushButton("Add Files")
        self.add_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_FileIcon))
        self.add_btn.clicked.connect(self.add_files)
        btn_row.addWidget(self.add_btn)

        self.remove_btn = QPushButton("Remove Selected")
        self.remove_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_TrashIcon))
        self.remove_btn.clicked.connect(self.remove_files)
        btn_row.addWidget(self.remove_btn)

        self.clear_btn = QPushButton("Clear All")
        self.clear_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogResetButton))
        self.clear_btn.clicked.connect(self.clear_files)
        btn_row.addWidget(self.clear_btn)
        btn_row.addStretch(1)
        fg_layout.addLayout(btn_row)

        self.fast_com_check = QCheckBox("Fast COM Mode (Recommended)")
        self.fast_com_check.setChecked(HAS_WIN32COM)
        fg_layout.addWidget(self.fast_com_check)
        root.addWidget(file_group, 1)

        progress_group = QGroupBox("Progress")
        pg_layout = QVBoxLayout(progress_group)
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        pg_layout.addWidget(self.progress_bar)
        self.status_label = QLabel("Ready")
        self.status_label.setObjectName("Status")
        pg_layout.addWidget(self.status_label)
        root.addWidget(progress_group)

        log_group = QGroupBox("Log")
        lg_layout = QVBoxLayout(log_group)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        lg_layout.addWidget(self.log_text)
        root.addWidget(log_group, 2)

        self.process_btn = QPushButton("Process & Save")
        self.process_btn.setObjectName("AccentButton")
        self.process_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MediaPlay))
        self.process_btn.clicked.connect(self.start_processing)
        root.addWidget(self.process_btn, 0, Qt.AlignmentFlag.AlignHCenter)

    def _make_app_icon(self):
        pix = QPixmap(128, 128)
        pix.fill(Qt.GlobalColor.transparent)
        painter = QPainter(pix)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        grad = QLinearGradient(0, 0, 128, 128)
        grad.setColorAt(0.0, QColor("#ffd8e8"))
        grad.setColorAt(1.0, QColor("#bde9ff"))
        painter.setBrush(grad)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawRoundedRect(8, 8, 112, 112, 28, 28)
        painter.setBrush(QColor("#ffffff"))
        painter.drawEllipse(32, 26, 64, 64)
        painter.setBrush(QColor("#7bb7f7"))
        painter.drawEllipse(48, 42, 32, 32)
        painter.setBrush(QColor("#ffb8d3"))
        painter.drawEllipse(86, 20, 16, 16)
        painter.end()
        return QIcon(pix)

    def _style_sheet(self):
        return """
        QMainWindow { background: #fff9fc; }
        QLabel#Title { color: #355070; font-size: 30px; font-weight: 800; }
        QLabel#Subtitle { color: #6d597a; font-size: 14px; margin-bottom: 4px; }
        QGroupBox {
            border: 1px solid #f2d8e9;
            border-radius: 14px;
            margin-top: 8px;
            padding-top: 12px;
            background: #fffdfd;
            color: #5a4e63;
            font-weight: 700;
        }
        QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 6px; }
        QListWidget, QTextEdit {
            border: 1px solid #efdde8;
            border-radius: 10px;
            background: #ffffff;
            color: #2f2a36;
            font-size: 13px;
        }
        QListWidget::item:selected { background: #cde9ff; color: #274690; }
        QPushButton {
            border: 1px solid #ead5e6;
            border-radius: 10px;
            padding: 8px 14px;
            background: #fff2f8;
            color: #5a4e63;
            font-weight: 600;
        }
        QPushButton:hover { background: #ffe6f1; }
        QPushButton#AccentButton {
            background: #7fb3ff;
            border-color: #74a9f6;
            color: #ffffff;
            font-size: 16px;
            font-weight: 800;
            padding: 12px 24px;
        }
        QPushButton#AccentButton:hover { background: #6fa8fa; }
        QProgressBar {
            border: 1px solid #edd9e6;
            border-radius: 8px;
            text-align: center;
            background: #fff;
            color: #5a4e63;
            min-height: 20px;
        }
        QProgressBar::chunk { background-color: #9de0ad; border-radius: 7px; }
        QLabel#Status { color: #6a5f73; font-size: 13px; }
        QCheckBox { color: #5d5168; font-weight: 600; }
        """

    def _append_log(self, msg, tag=""):
        colors = {"ok": "#2aa07a", "err": "#cf4f6a", "head": "#6b8ce3"}
        color = colors.get(tag, "#4f4559")
        self.log_text.append(f"<span style='color:{color};'>{escape(str(msg))}</span>")

    def add_files(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "Select Excel Files", "", "Excel Files (*.xlsx *.xlsm)"
        )
        for p in paths:
            if p not in self.files:
                self.files.append(p)
                self.file_list.addItem(QListWidgetItem(os.path.basename(p)))
        if paths:
            QMessageBox.information(
                self,
                "Load Complete",
                f"Loaded {len(paths)} file(s) successfully.",
            )

    def remove_files(self):
        selected = self.file_list.selectedIndexes()
        for idx in sorted((i.row() for i in selected), reverse=True):
            self.file_list.takeItem(idx)
            self.files.pop(idx)

    def clear_files(self):
        self.file_list.clear()
        self.files.clear()

    def start_processing(self):
        if not self.files:
            QMessageBox.warning(self, "Warning", "Please add Excel files first!")
            return
        self.process_btn.setEnabled(False)
        self.progress_bar.setValue(0)
        self.status_label.setText("Processing...")

        self.worker_thread = QThread(self)
        self.worker = ProcessorWorker(self.files, self.fast_com_check.isChecked())
        self.worker.moveToThread(self.worker_thread)
        self.worker_thread.started.connect(self.worker.run)
        self.worker.log_signal.connect(self._append_log)
        self.worker.progress_signal.connect(self._on_progress)
        self.worker.finished_signal.connect(self._on_finished)
        self.worker.finished_signal.connect(self.worker_thread.quit)
        self.worker.finished_signal.connect(self.worker.deleteLater)
        self.worker_thread.finished.connect(self.worker_thread.deleteLater)
        self.worker_thread.start()

    def _on_progress(self, completed, total, fname):
        pct = int((completed / total) * 100) if total else 0
        self.progress_bar.setValue(pct)
        self.status_label.setText(f"[{completed}/{total}] Done: {fname}")

    def _on_finished(self, elapsed, failed_count, total):
        succeeded = max(0, total - failed_count)
        self._append_log(f"{'=' * 52}", "head")
        if failed_count:
            self.status_label.setText(
                f"Finished with errors ({succeeded}/{total} succeeded, {elapsed:.1f}s)"
            )
            self._append_log(
                f"  Finished with errors: {succeeded}/{total} succeeded ({elapsed:.1f}s)",
                "err",
            )
            QMessageBox.warning(
                self,
                "Finished with Errors",
                f"Processed {total} file(s)\n"
                f"Succeeded: {succeeded}\n"
                f"Failed: {failed_count}\n"
                f"Elapsed: {elapsed:.1f}s",
            )
        else:
            self.status_label.setText(f"Done! ({elapsed:.1f}s)")
            self._append_log(f"  All files processed! ({elapsed:.1f}s)", "ok")
            QMessageBox.information(self, "Done", f"All files processed! ({elapsed:.1f}s)")
        self.process_btn.setEnabled(True)




# <<< START OF CHANGES >>>
# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
def run_this_app(working_dir=None): # ชื่อฟังก์ชันนี้จะถูกใช้ใน Launcher
    """
    ฟังก์ชันหลักสำหรับสร้างและรัน QuotaSamplerApp.
    """
    print(f"--- QUOTA_SAMPLER_INFO: Starting 'QuotaSamplerApp' via run_this_app() ---")
    try:
    # --- ส่วนที่ใช้รันโปรแกรม ---
    #if __name__ == "__main__":
        multiprocessing.freeze_support()
        qt_app = QApplication(sys.argv)
        win = ExcelProcessorApp()
        win.show()
        sys.exit(qt_app.exec())


        print(f"--- QUOTA_SAMPLER_INFO: QuotaSamplerApp mainloop finished. ---")

    except Exception as e:
        # ดักจับ Error ที่อาจเกิดขึ้นระหว่างการสร้างหรือรัน App
        print(f"QUOTA_SAMPLER_ERROR: An error occurred during QuotaSamplerApp execution: {e}")
        # แสดง Popup ถ้ามีปัญหา
        if 'root' not in locals() or not root.winfo_exists(): # สร้าง root ชั่วคราวถ้ายังไม่มี
            root_temp = tk.Tk()
            root_temp.withdraw()
            messagebox.showerror("Application Error (Quota Sampler)",
                                f"An unexpected error occurred:\n{e}", parent=root_temp)
            root_temp.destroy()
        else:
            messagebox.showerror("Application Error (Quota Sampler)",
                                f"An unexpected error occurred:\n{e}", parent=root) # ใช้ root ที่มีอยู่ถ้าเป็นไปได้
        sys.exit(f"Error running QuotaSamplerApp: {e}") # อาจจะ exit หรือไม่ก็ได้ ขึ้นกับการออกแบบ


# --- ส่วน Run Application เมื่อรันไฟล์นี้โดยตรง (สำหรับ Test) ---
if __name__ == "__main__":
    print("--- Running QuotaSamplerApp.py directly for testing ---")
    # (ถ้ามีการตั้งค่า DPI ด้านบน มันจะทำงานอัตโนมัติ)

    # เรียกฟังก์ชัน Entry Point ที่เราสร้างขึ้น
    run_this_app()

    print("--- Finished direct execution of QuotaSamplerApp.py ---")
# <<< END OF CHANGES >>>
