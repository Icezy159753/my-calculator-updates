# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
import pyreadstat
import pandas as pd  # Uncommented or added this line
# import pandas as pd # Not strictly needed for this version based on pyreadstat usage
import openpyxl
from openpyxl.utils import column_index_from_string
import os
import re
import traceback
from collections import defaultdict # Used for grouping
from openpyxl.worksheet.datavalidation import DataValidation
import sys # ต้อง import sys เพิ่ม

# --- ฝัง ICON ---

def resource_path(relative_name):
    # คืน path ที่ถูกต้อง ไม่ว่าจะ run จากไฟล์ .py หรือ bundle เป็น exe
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative_name)
         

# --- Class to redirect stdout/stderr to a Tkinter Text widget ---
class TextRedirector(object):
    """A class to redirect stdout/stderr messages to a Tkinter Text widget."""
    def __init__(self, widget, tag="stdout"):
        """
        Initialize the redirector.
        :param widget: The Tkinter Text widget to redirect output to.
        :param tag: A tag to associate with the written text (optional).
        """
        self.widget = widget
        self.tag = tag

    def write(self, str_):
        """
        Writes the string to the Text widget.
        This method MUST be called from the main GUI thread.
        """
        # Check if widget is valid before writing
        if not self.widget or not self.widget.winfo_exists():
             # print(f"TextRedirector Warning: Target widget {self.widget} does not exist.", file=sys.__stdout__) # Log to original stdout if widget is gone
             return

        try:
            # Ensure updates are thread-safe if called from other threads
            # Using schedule here for robustness, even if current code is blocking
            self.widget.after_idle(self._write_to_widget, str_)
        except Exception as e:
            # Fallback if after_idle fails (e.g., root window destroyed)
            print(f"TextRedirector Error scheduling write: {e}", file=sys.__stdout__)


    def _write_to_widget(self, str_):
        """Internal method to perform the actual widget update."""
        # Check widget existence again just before update
        if not self.widget or not self.widget.winfo_exists():
             return

        try:
            self.widget.configure(state='normal')
            self.widget.insert(tk.END, str_, (self.tag,))
            self.widget.see(tk.END)  # Auto-scroll to the end
            self.widget.configure(state='disabled')
            # self.widget.update_idletasks() # Usually not needed with after_idle
        except tk.TclError as e:
             # Handle cases where the widget might be destroyed between check and update
             print(f"TextRedirector TclError writing to widget: {e}", file=sys.__stdout__)
        except Exception as e:
             print(f"TextRedirector Error writing to widget: {e}", file=sys.__stdout__)


    def flush(self):
        """
        Flush method for compatibility with stdout interface.
        No action needed for direct Text widget insertion.
        """
        pass # Nothing to do for this implementation

    def isatty(self):
        """
        Report as not being a terminal.
        """
        return False

# --- Updated Helper Function ---
def get_base_name_heuristic(var_name):
    """
    Tries to derive a base name by removing trailing/intermediate numbers/identifiers.
    Handles more complex patterns for grouping including _r<number> and prefix_num_suffix.
    """

    # --- NEW 1.55: I_<loopIndex>_<stem>_<subIndex>  -> base = "I_<loopIndex>_<stem>"
    # ตัวอย่าง: I_1_a4_1, I_1_a4_10, I_2_a4_3  -> I_1_a4, I_2_a4
    # (ต้องมาก่อน Heuristic 1.6)
    m_i_num_stem_num = re.match(r'^I_(\d+)_([A-Za-z]\w*)_(\d+)$', var_name)
    if m_i_num_stem_num:
        loop_idx = m_i_num_stem_num.group(1)
        stem = m_i_num_stem_num.group(2)
        return f"I_{loop_idx}_{stem}"

    # =========================================================================
    # เดิม: 1.4 s13_1_1 -> s13_1
    match_s_num_num = re.match(r'(.+_\d+)_(\d+)$', var_name)
    if match_s_num_num:
        base_part = match_s_num_num.group(1)
        if not base_part.endswith('_O') and not base_part.endswith('_r'):
            return base_part

    # 1.5 I_num_basename_Onum -> basename (เช่น I_1_q1_1_O1 -> q1_1)
    match_i_num_base_onum = re.match(r'I_\d+_(.+?)_O\d+$', var_name)
    if match_i_num_base_onum:
        base_part = match_i_num_base_onum.group(1)
        if base_part:
            return base_part

    # 1.6 I_num_basename -> basename (เช่น I_1_q4_Oth -> q4_Oth)
    # ** จะไม่โดนกรณี I_<n>_<stem>_<sub> เพราะถูกจับโดย 1.55 ไปก่อนแล้ว **
    match_i_num_base = re.match(r'I_\d+_(.+)$', var_name)
    if match_i_num_base and not re.search(r'_O\d+$', match_i_num_base.group(1)):
        base_part = match_i_num_base.group(1)
        if base_part:
            return base_part

    # 1.65 prefix_num_Onum -> prefix
    match_pnon = re.match(r'(.+?)_(\d+)_O(\d+)$', var_name)
    if match_pnon:
        prefix = match_pnon.group(1)
        if not re.match(r'I_\d+_', prefix):
            return prefix

    # 1.7 prefix_rNum_suffixNum -> prefix_rNum_suffix
    match_prnsn = re.match(r'(.+?)(_r\d+)(_[a-zA-Z]+)(\d+)$', var_name)
    if match_prnsn:
        prefix_part = match_prnsn.group(1)
        r_num_part = match_prnsn.group(2)
        suffix_part = match_prnsn.group(3)
        return f"{prefix_part}{r_num_part}{suffix_part}"

    # 1.8 prefix_suffixBaseNum -> prefix_suffixBase
    match_psn = re.match(r'(.+?_)([a-zA-Z]+)(\d+)$', var_name)
    if match_psn:
        prefix_part = match_psn.group(1)
        suffix_base = match_psn.group(2)
        return f"{prefix_part}{suffix_base}"

    # 2. prefix_number_suffix -> prefix_suffix
    match_prefix_num_suffix = re.match(r'(.+?)_(\d+)_([a-zA-Z][a-zA-Z_]*)$', var_name)
    if match_prefix_num_suffix:
        prefix = match_prefix_num_suffix.group(1)
        suffix = match_prefix_num_suffix.group(3)
        if not prefix.startswith('I_') or not prefix[2:].isdigit():
            return f"{prefix}_{suffix}"

    # 3.5 name_r<number> -> name_r
    match_r_num = re.match(r'(.+_r)(\d+)$', var_name)
    if match_r_num:
        return match_r_num.group(1)

    # 1. name(number) -> name
    match_paren = re.match(r'(.+)\((\d+)\)$', var_name)
    if match_paren:
        return match_paren.group(1)

    # 4. name[_ ]number -> name
    match_name_num = re.match(r'(.+?)[_ ](\d+)$', var_name)
    if match_name_num:
        base_part = match_name_num.group(1).rstrip('_ ')
        if base_part and not base_part.endswith('_r'):
            if not re.match(r'.+?_\d+$', base_part):
                return base_part

    # 3/3.1 name + trailing letter
    match_name_letter = re.match(r'(.+?)[_ ]([A-Za-z])$', var_name)
    if match_name_letter:
        base_part = match_name_letter.group(1).rstrip('_ ')
        if base_part and not base_part[-1].isdigit():
            return base_part
    match_name_letter_direct = re.match(r'(.+)([A-Za-z])$', var_name)
    if match_name_letter_direct:
        base_part = match_name_letter_direct.group(1)
        if base_part:
            return base_part

    return var_name




# --- Main Application Class ---
class SpssToExcelConverter:
    def __init__(self, root):
        self.root = root
        root.title("ตัวแปลง SPSS เป็น Itemdef Excel V3")
        # อาจจะขยายความกว้างหน้าต่างเล็กน้อยเพื่อให้ปุ่มพอดี
        root.geometry("880x500") # ปรับขนาดตามความเหมาะสม (เพิ่มความสูง)

        # File Path Variables
        self.spss_file_path = tk.StringVar()
        self.excel_template_path = tk.StringVar()
        self.excel_output_path = tk.StringVar()

        # Data Storage
        self.codes_to_delete_confirmed = set() # Stores (var_name, code_str) to delete
        self.last_read_meta = None # Cache for SPSS metadata
        self.variable_loop_types = {} # Stores {var_name: "SA"/"MA"/"Loop Text"/"Loop Numeric"}
        self.user_defined_loop_names = {} # Stores {first_var_name: "user_loop_id"}

        # --- GUI Widgets ---

        button_font = ('TkDefaultFont', 10, 'bold') # หรือ ('tahoma', 10, 'bold'), ('Arial', 10, 'bold') etc.button_font = ('TkDefaultFont', 10, 'bold') # หรือ ('tahoma', 10, 'bold'), ('Arial', 10, 'bold') etc.

        # SPSS File Selection Frame
        self.frame_spss = tk.Frame(root, padx=10, pady=5)
        self.frame_spss.pack(fill='x', padx=10, pady=2)
        # เพิ่ม width ให้ Label เพื่อให้ช่อง Entry ตรงกันแนวตั้ง
        self.lbl_spss = tk.Label(self.frame_spss, text="เลือกไฟล์ SPSS (.sav):", width=25, anchor='w',fg="red",font=button_font)
        self.lbl_spss.pack(side='left')
        self.entry_spss = tk.Entry(self.frame_spss, textvariable=self.spss_file_path, state='readonly', width=60) # ปรับ width ตามต้องการ
        self.entry_spss.pack(side='left', fill='x', expand=True, padx=5)
        self.btn_browse_spss = tk.Button(self.frame_spss, text="Open SPSS", command=self.browse_spss_file,font=button_font,bg="#4CAF50")
        self.btn_browse_spss.pack(side='left')

        # Excel Template Selection Frame
        self.frame_template = tk.Frame(root, padx=10, pady=5)
        self.frame_template.pack(fill='x', padx=10, pady=2)
        # เพิ่ม width ให้ Label
        self.lbl_template = tk.Label(self.frame_template, text="เลือกไฟล์ Excel Template (.xlsx):", width=25, anchor='w',fg="red",font=button_font)
        self.lbl_template.pack(side='left')
        self.entry_template = tk.Entry(self.frame_template, textvariable=self.excel_template_path, state='readonly', width=60) # ปรับ width ตามต้องการ
        self.entry_template.pack(side='left', fill='x', expand=True, padx=5)
        self.btn_browse_template = tk.Button(self.frame_template, text="Open Excel", command=self.browse_excel_template,font=button_font,bg="#4CAF50")
        self.btn_browse_template.pack(side='left')

        # Excel Output Selection Frame
        self.frame_excel_output = tk.Frame(root, padx=10, pady=5)
        self.frame_excel_output.pack(fill='x', padx=10, pady=2)
        # เพิ่ม width ให้ Label
        self.lbl_excel_output = tk.Label(self.frame_excel_output, text="บันทึกเป็น Excel Itemdef (.xlsx):", width=25, anchor='w',fg="red",font=button_font)
        self.lbl_excel_output.pack(side='left')
        self.entry_excel_output = tk.Entry(self.frame_excel_output, textvariable=self.excel_output_path, state='readonly', width=60) # ปรับ width ตามต้องการ
        self.entry_excel_output.pack(side='left', fill='x', expand=True, padx=5)
        self.btn_browse_excel_output = tk.Button(self.frame_excel_output, text="เลือกตำแหน่งSave", command=self.browse_excel_output,font=button_font,bg="blue")
        self.btn_browse_excel_output.pack(side='left')

        # Control Buttons Frame
        self.frame_control = tk.Frame(root, padx=10, pady=10)
        self.frame_control.pack(fill='x', padx=10, pady=10)

        # --- จัดเรียงปุ่มใหม่และปรับขนาด ---
        # ใช้ padx เท่าๆ กัน และ expand=True เพื่อให้กระจายตัว
        button_padx = 2 # ลดช่องว่างระหว่างปุ่ม

        # Button 1: กำหนดตัวแปร Loop
        self.btn_define_loop = tk.Button(self.frame_control, text="1. กำหนดตัวแปร Loop", command=self.show_loop_definition_window, state='disabled', width=18,font=button_font,bg="#FFE4B5") # ปรับ width
        self.btn_define_loop.pack(side=tk.LEFT, padx=button_padx, expand=True)

        # Button 2: กำหนดชื่อ Loop ID
        self.btn_name_loop = tk.Button(self.frame_control, text="2. กำหนดชื่อ Loop ID", command=self.show_loop_naming_window, state='disabled', width=18,font=button_font,bg="#FFE4B5") # ปรับ width
        self.btn_name_loop.pack(side=tk.LEFT, padx=button_padx, expand=True)

        # --- NEW Buttons ---
        # Button 3: Save Loop
        self.btn_save_loop = tk.Button(self.frame_control, text="Save Loop", command=self.save_loop_settings, state='disabled', width=12,font=button_font,bg="#BEBEBE") # ปุ่มใหม่
        self.btn_save_loop.pack(side=tk.LEFT, padx=button_padx, expand=True)

        # Button 4: Load Setting Loop
        self.btn_load_loop = tk.Button(self.frame_control, text="Load Setting Loop", command=self.load_loop_settings, state='disabled', width=16,font=button_font,bg="#BEBEBE") # ปุ่มใหม่
        self.btn_load_loop.pack(side=tk.LEFT, padx=button_padx, expand=True)
        # ------------------

        # หลังจากปุ่ม btn_load_loop และก่อน btn_convert
        self.btn_export_rawdata = tk.Button(
            self.frame_control, 
            text="Export Rawdata Excel", 
            command=self.export_rawdata_excel, 
            state='disabled', 
            width=18,
            font=button_font,
            bg="#4682B4")
        self.btn_export_rawdata.pack(side=tk.LEFT, padx=button_padx, expand=True)


        # Button 5: แปลงไฟล์ (จัดไว้ทางขวาเหมือนเดิม)
        self.btn_convert = tk.Button(self.frame_control, text="3. Run Itemdef", command=self.convert_file, state='disabled', width=12,font=button_font,bg="#DAA520") # ปรับ width
        # Pack ปุ่มนี้สุดท้ายเพื่อให้ไปอยู่ด้านขวาสุดของปุ่มที่ pack(side=tk.LEFT) ก่อนหน้า
        self.btn_convert.pack(side=tk.RIGHT, padx=button_padx, expand=True)

        # Status Label
        self.status_label = tk.Label(root, text="", fg="blue")
        # เพิ่ม fill='x' เพื่อให้ label ขยายเต็มความกว้าง เผื่อข้อความยาวๆ
        self.status_label.pack(pady=5, fill='x')

        # --- Log Area ---
        self.log_frame = tk.Frame(root, bd=1, relief=tk.SUNKEN) # เพิ่มกรอบให้เห็นชัดเจน
        # ใช้ fill='both' และ expand=True เพื่อให้ขยายเต็มพื้นที่ที่เหลือ
        self.log_frame.pack(fill='both', expand=True, padx=10, pady=(5, 10))

        # สร้าง Text widget สำหรับแสดง log
        # wrap='none' เพื่อไม่ให้ตัดคำเหมือนใน terminal
        self.log_text = tk.Text(self.log_frame, wrap='none', height=15, state='disabled', font=("Consolas", 9)) # เริ่มต้นเป็น disable ป้องกันการพิมพ์, ปรับ font ได้
        self.log_vsb = ttk.Scrollbar(self.log_frame, orient="vertical", command=self.log_text.yview)
        self.log_hsb = ttk.Scrollbar(self.log_frame, orient="horizontal", command=self.log_text.xview)
        self.log_text.configure(yscrollcommand=self.log_vsb.set, xscrollcommand=self.log_hsb.set)

        # จัดวาง Text widget และ scrollbars โดยใช้ grid ภายใน log_frame
        self.log_text.grid(row=0, column=0, sticky='nsew')
        self.log_vsb.grid(row=0, column=1, sticky='ns')
        self.log_hsb.grid(row=1, column=0, sticky='ew')

        # ทำให้ Text widget ขยายตามขนาด frame
        self.log_frame.grid_rowconfigure(0, weight=1)
        self.log_frame.grid_columnconfigure(0, weight=1)
        # --- End Log Area ---
        # สร้าง Text widget สำหรับแสดง log
        # wrap='none' เพื่อไม่ให้ตัดคำเหมือนใน terminal
        self.log_text = tk.Text(self.log_frame, wrap='none', height=15, state='disabled', font=("Consolas", 9)) # เริ่มต้นเป็น disable ป้องกันการพิมพ์, ปรับ font ได้
        self.log_vsb = ttk.Scrollbar(self.log_frame, orient="vertical", command=self.log_text.yview)
        self.log_hsb = ttk.Scrollbar(self.log_frame, orient="horizontal", command=self.log_text.xview)
        self.log_text.configure(yscrollcommand=self.log_vsb.set, xscrollcommand=self.log_hsb.set)

        # จัดวาง Text widget และ scrollbars โดยใช้ grid ภายใน log_frame
        self.log_text.grid(row=0, column=0, sticky='nsew')
        self.log_vsb.grid(row=0, column=1, sticky='ns')
        self.log_hsb.grid(row=1, column=0, sticky='ew')

        # ทำให้ Text widget ขยายตามขนาด frame
        self.log_frame.grid_rowconfigure(0, weight=1)
        self.log_frame.grid_columnconfigure(0, weight=1)
        # --- End Log Area ---


        # Trace file paths to enable/disable buttons
        self.spss_file_path.trace_add('write', self.check_fields)
        self.excel_template_path.trace_add('write', self.check_fields)
        self.excel_output_path.trace_add('write', self.check_fields)

        # --- Redirect stdout ---
        # ต้องทำหลังจากสร้าง self.log_text แล้ว
        self._original_stdout = sys.stdout # <<< เก็บ stdout ดั้งเดิมของ instance นี้
        self.redirector = TextRedirector(self.log_text)
        sys.stdout = self.redirector
        print(f"SPSS_LOG_INFO: SpssToExcelConverter stdout redirected to GUI log.") # <<< จะไปที่ GUI log
        # Optional: Redirect stderr ด้วยก็ได้ ถ้าต้องการเห็น error ใน log window เช่นกัน
        # sys.stderr = TextRedirector(self.log_text, tag="stderr")
        # self.log_text.tag_config("stderr", foreground="red") # ตั้งค่าสีสำหรับ stderr
        # ----------------------

        # พิมพ์ข้อความเริ่มต้นเพื่อทดสอบการ redirect
        #print("--- App เริ่มต้นทำงาน ---")
        #print(f"Application นี้ถูกพัฒนาโดย Team DP") # แก้ไขชื่อผู้พัฒนา
        #print("-" * 30)
        # --- START: เพิ่มข้อความแสดงผู้พัฒนา ---
        try:
            # 1. กำหนด Tag ชื่อ 'dev_credit' สำหรับตัวหนาสีน้ำเงิน
            #    (ใช้ font และ size ให้ใกล้เคียงกับ log_text ปกติ)
            log_font_family = "Consolas" # หรือ Tahoma, Arial ตามที่ใช้ใน log_text
            log_font_size = 9            # ขนาดเดียวกับ log_text
            self.log_text.tag_configure("dev_credit",
                                        font=(log_font_family, log_font_size, 'bold'),
                                        foreground="Green")

            # 2. เปิดให้แก้ไข log_text ชั่วคราว
            self.log_text.configure(state='normal')

            # 3. แทรกข้อความส่วนแรก (ปกติ)
            self.log_text.insert(tk.END, "Application นี้ถูกพัฒนาโดย ", ("dev_credit",))
            
            # 4. แทรกข้อความส่วนที่สอง (ตัวหนาสีน้ำเงิน) โดยใช้ tag
            self.log_text.insert(tk.END, "Team DP\n", ("dev_credit",))
          
            # 3. แทรกข้อความส่วนแรก (ปกติ)
            self.log_text.insert(tk.END, "---- App พร้อมเริ่มต้นทำงาน ----", ("dev_credit",))




            # 5. เพิ่มบรรทัดใหม่
            self.log_text.insert(tk.END, "\n")

            # 6. ปิดการแก้ไข และเลื่อนไปท้ายสุด
            self.log_text.configure(state='disabled')
            self.log_text.see(tk.END)

            # อาจจะเพิ่มเส้นคั่นอีกทีผ่าน print() เพื่อให้เห็นชัดเจน
            #print("-" * 30)

        except Exception as e:
            # กรณีเกิดข้อผิดพลาดในการแทรกโดยตรง (เผื่อไว้)
            print(f"[Error adding credit line: {e}]", file=sys.__stderr__) # พิมพ์ไปที่ stderr ดั้งเดิม
            print("Application นี้ถูกพัฒนาโดย Team DP") # พิมพ์แบบปกติเป็น fallback

        # --- END: เพิ่มข้อความแสดงผู้พัฒนา ---

        # Regex for Auto-Detect MA (e.g., Q5_O1, Q5_O2)
        # ต้อง import re ก่อนใช้งาน
        self.ma_pattern = re.compile(r'(.+)_O\d+$')
        # --- ผูก event การปิดหน้าต่างกับ on_closing ---
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    def on_closing(self):
        """Called when the user tries to close the window."""
        print("SPSS_LOG_INFO: Window closing event received.", file=self._original_stdout) # พิมพ์ไปที่ original
        self.cleanup_redirector() # เรียก cleanup ก่อน destroy
        if self.root.winfo_exists():
            self.root.destroy()

    def cleanup_redirector(self):
        """Restores stdout and tells the redirector its widget is gone."""
        print("SPSS_LOG_INFO: SpssToExcelConverter.cleanup_redirector() called.", file=self._original_stdout) # พิมพ์ไปที่ original
        if hasattr(self, 'redirector') and self.redirector:
            # (Optional) ถ้า TextRedirector มีเมธอด destroy() ให้เรียกด้วย
            if hasattr(self.redirector, 'destroy') and callable(self.redirector.destroy):
                self.redirector.destroy() # บอก redirector ว่า widget ไม่มีแล้ว

        if hasattr(self, '_original_stdout'): # คืนค่า stdout ที่ instance นี้เก็บไว้
            sys.stdout = self._original_stdout
            print("SPSS_LOG_INFO: SpssToExcelConverter stdout restored by cleanup_redirector.", file=self._original_stdout)
        else: # Fallback (ไม่ควรเกิดถ้า __init__ ทำงานถูกต้อง)
            sys.stdout = sys.__stdout__
            print("SPSS_LOG_WARNING: SpssToExcelConverter _original_stdout not found, restored to sys.__stdout__.", file=sys.__stdout__)
    def browse_spss_file(self):
        """Opens SPSS file dialog, sets output path, reads metadata, and validates codes."""
        filepath = filedialog.askopenfilename(
            initialdir=".",
            title="เลือกไฟล์ SPSS (.sav)",
            filetypes=(("SPSS files", "*.sav"), ("All files", "*.*"))
        )
        if filepath:
            self.spss_file_path.set(filepath)
            # Reset related states when a new file is selected
            self.variable_loop_types.clear()
            self.codes_to_delete_confirmed = set()
            self.user_defined_loop_names.clear()
            self.last_read_meta = None
            #print("Cleared previous loop definitions, confirmed deletions, and loop names.", flush=True)

            # --- Set Default Output Path ---
            dir_name = os.path.dirname(filepath)
            file_name_without_ext = os.path.splitext(os.path.basename(filepath))[0]
            default_excel_path = os.path.join(dir_name, f"{file_name_without_ext}_Itemdef.xlsx")
            counter = 1
            final_excel_path = default_excel_path
            while os.path.exists(final_excel_path):
                final_excel_path = os.path.join(dir_name, f"{file_name_without_ext}_output_{counter}.xlsx")
                counter += 1
            self.excel_output_path.set(final_excel_path)
            # ---------------------------------

            self.status_label.config(text="กำลังอ่าน Metadata และตรวจสอบ Code...", fg="orange")
            self.root.update_idletasks() # Force GUI update

            # --- Read Metadata and Validate ---
            try:
                print(f"กำลังอ่านไฟล์จาก: {filepath}", flush=True)
                meta = None
                try:
                    print("กำลังเข้ารหัสด้วย UTF-8 encoding...", flush=True)
                    _df_temp, meta_temp = pyreadstat.read_sav(filepath, metadataonly=True, encoding='utf-8')
                    meta = meta_temp
                    print("เข้ารหัสและอ่านไฟล์สำเร็จด้วย: UTF-8.", flush=True)
                except pyreadstat.ReadstatError as e_utf8:
                    print(f"UTF-8 ไม่สำเร็จ: {e_utf8}. กำลังลองด้วย: cp874...", flush=True)
                    try:
                         _df_temp, meta_temp = pyreadstat.read_sav(filepath, metadataonly=True, encoding='cp874')
                         meta = meta_temp
                         print("เข้ารหัสและอ่านไฟล์สำเร็จด้วย: cp874.", flush=True)
                    except pyreadstat.ReadstatError as e_cp874:
                        print(f"cp874 also failed: {e_cp874}. Could not determine encoding.", flush=True)
                        raise e_cp874

                if meta is None:
                    raise ValueError("Metadata could not be read with attempted encodings.")

                self.last_read_meta = meta # Cache the metadata

                # --- Validate Value Labels (Codes 0 or > 2000) ---
                if hasattr(meta, 'variable_value_labels') and isinstance(meta.variable_value_labels, dict):
                    initial_problems = self._validate_value_labels(meta.variable_value_labels)
                    if initial_problems:
                        print(f"Found {len(initial_problems)} variables with problematic codes. Showing validation window.", flush=True)
                        self.status_label.config(text="พบ Code ที่ต้องตรวจสอบ...", fg="orange")
                        self.root.update_idletasks()
                        self._show_validation_window(initial_problems, meta.variable_value_labels, meta)
                        self.status_label.config(text=f"ตรวจสอบ Code เสร็จสิ้น ({len(self.codes_to_delete_confirmed)} รายการถูกเลือกที่จะลบ)", fg="blue")
                    else:
                        print("No problematic codes (0 or >2000) found.", flush=True)
                        self.status_label.config(text="ไม่พบ Code ที่ต้องตรวจสอบ", fg="blue")
                else:
                    print("No 'variable_value_labels' found in metadata.", flush=True)
                    self.status_label.config(text="ไม่พบ Value Labels ในไฟล์", fg="blue")
                # ------------------------------------

            except FileNotFoundError:
                print(f"ERROR: SPSS file not found after selection: {filepath}", flush=True)
                messagebox.showerror("ข้อผิดพลาด", f"ไม่พบไฟล์ SPSS:\n{filepath}")
                self.status_label.config(text="เกิดข้อผิดพลาด: ไม่พบไฟล์ SPSS", fg="red")
                self.spss_file_path.set("")
                self.last_read_meta = None
            except pyreadstat.ReadstatError as e_read:
                print(f"ERROR reading meta after selection (tried UTF-8 and cp874): {e_read}", flush=True)
                messagebox.showerror("ข้อผิดพลาดอ่าน SPSS", f"ไม่สามารถอ่าน Metadata จากไฟล์ SPSS ได้ (ลองทั้ง UTF-8 และ cp874):\n{e_read}")
                self.status_label.config(text="เกิดข้อผิดพลาดอ่าน SPSS", fg="red")
                self.spss_file_path.set("")
                self.last_read_meta = None
            except Exception as e:
                print(f"ERROR during initial validation or metadata read: {e}", flush=True)
                traceback.print_exc()
                messagebox.showerror("ข้อผิดพลาด", f"เกิดข้อผิดพลาดที่ไม่คาดคิด:\n{e}")
                self.status_label.config(text="เกิดข้อผิดพลาดไม่คาดคิด", fg="red")
                self.spss_file_path.set("")
                self.last_read_meta = None
            finally:
                 self.check_fields() # Update button states

    def browse_excel_template(self):
        """Opens Excel template file dialog."""
        filepath = filedialog.askopenfilename(
            initialdir=".",
            title="เลือกไฟล์ Excel Template (.xlsx)",
            filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*"))
        )
        if filepath:
            self.excel_template_path.set(filepath)

    def browse_excel_output(self):
        """Opens Excel output file dialog."""
        initial_dir = "."
        suggested_filename = ""
        # Suggest filename based on input SPSS or previous output path
        if self.excel_output_path.get():
            initial_dir = os.path.dirname(self.excel_output_path.get())
            suggested_filename = os.path.basename(self.excel_output_path.get())
        elif self.spss_file_path.get():
            spss_filepath = self.spss_file_path.get()
            initial_dir = os.path.dirname(spss_filepath)
            file_name_without_ext = os.path.splitext(os.path.basename(spss_filepath))[0]
            suggested_filename = f"{file_name_without_ext}_Itemdef.xlsx"
            counter = 1
            final_suggested_path = os.path.join(initial_dir, suggested_filename)
            while os.path.exists(final_suggested_path): # Avoid overwriting existing files
                suggested_filename = f"{file_name_without_ext}_output_{counter}.xlsx"
                final_suggested_path = os.path.join(initial_dir, suggested_filename)
                counter += 1
        filepath = filedialog.asksaveasfilename(
            initialdir=initial_dir, initialfile=suggested_filename,
            title="บันทึกไฟล์ Excel Output (.xlsx)",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
            defaultextension=".xlsx"
        )
        if filepath:
            # Ensure .xlsx extension
            if not filepath.lower().endswith(".xlsx"): filepath += ".xlsx"
            self.excel_output_path.set(filepath)

    # แก้ไขส่วนนี้ใน check_fields:
    def check_fields(self, *args):
        """Enables/disables buttons based on whether files are selected."""
        spss_selected = bool(self.spss_file_path.get())
        template_selected = bool(self.excel_template_path.get())
        output_selected = bool(self.excel_output_path.get())

        can_define = spss_selected
        can_convert = spss_selected and template_selected and output_selected
        can_save_load_loop = spss_selected
        can_export_rawdata = spss_selected

        self.btn_define_loop.config(state='normal' if can_define else 'disabled')
        self.btn_name_loop.config(state='normal' if can_define else 'disabled')
        self.btn_save_loop.config(state='normal' if can_save_load_loop else 'disabled')
        self.btn_load_loop.config(state='normal' if can_save_load_loop else 'disabled')
        self.btn_export_rawdata.config(state='normal' if can_export_rawdata else 'disabled')
        self.btn_convert.config(state='normal' if can_convert else 'disabled')

    def _validate_value_labels(self, value_labels_dict):
        """Identifies variables with value codes that are 0 or > 2000."""
        problematic_codes = {}
        if not isinstance(value_labels_dict, dict):
            return problematic_codes
        for var_name, labels in value_labels_dict.items():
            if not isinstance(labels, dict):
                continue
            found_problems = []
            for code in labels.keys():
                try:
                    numeric_code = float(code)
                    if numeric_code == 0 or numeric_code > 2000:
                        found_problems.append(str(code))
                except (ValueError, TypeError):
                    continue
            if found_problems:
                problematic_codes[var_name] = sorted(found_problems, key=lambda x: float(x))
        return problematic_codes
    
    def _show_validation_window(self, codes_to_validate, all_value_labels, meta):
        """
        Displays a window allowing the user to select problematic codes for deletion.
        Groups _O<n> variables for display.
        Shows integer codes without decimals.
        Shows only unique codes within each displayed group/item.
        Selection applies to all occurrences of the code within the group.
        """
        if not meta or not hasattr(meta, 'column_names') or not meta.column_names:
            messagebox.showerror("ข้อผิดพลาด", "ไม่สามารถอ่านรายชื่อตัวแปรจาก Metadata เพื่อจัดกลุ่มได้", parent=self.root)
            all_var_names = list(codes_to_validate.keys()) # Fallback
            display_items = {var: [var] for var in all_var_names} # {display_name: [actual_vars]}
            ordered_display_names = sorted(all_var_names)
            print("Warning: Metadata missing for grouping in validation window. Displaying all variables individually.", flush=True)
        else:
            # --- Grouping Logic (same as before) ---
            all_var_names = meta.column_names
            print(f"Grouping {len(all_var_names)} variables for validation display...", flush=True)
            ma_pattern = re.compile(r'(.+)_O(\d+)$')
            groups = defaultdict(list)
            variable_order_keys = []
            processed_indices_for_grouping = set()
            display_items = {}

            for i, var_name in enumerate(all_var_names):
                if i in processed_indices_for_grouping: continue
                match = ma_pattern.match(var_name)
                if match:
                    base_name = match.group(1)
                    current_group_vars = []
                    for j in range(i, len(all_var_names)):
                        if j in processed_indices_for_grouping: continue
                        inner_match = ma_pattern.match(all_var_names[j])
                        if inner_match and inner_match.group(1) == base_name:
                            current_group_vars.append(all_var_names[j])
                            processed_indices_for_grouping.add(j)
                    if current_group_vars:
                        rep_name = f"{base_name}_O1"
                        if rep_name not in current_group_vars:
                            rep_name = sorted(current_group_vars)[0]
                        if base_name not in variable_order_keys:
                            variable_order_keys.append(base_name)
                        groups[base_name].extend(current_group_vars)
                        display_items[rep_name] = sorted(list(set(current_group_vars))) # Ensure unique vars in group list
                else:
                    if var_name not in variable_order_keys:
                        variable_order_keys.append(var_name)
                    display_items[var_name] = [var_name]
                    processed_indices_for_grouping.add(i)

            ordered_display_names = []
            processed_display = set()
            for key in variable_order_keys:
                if key in groups:
                    rep_name = f"{key}_O1"
                    actual_rep = None
                    for dn, av in display_items.items():
                        dn_match = ma_pattern.match(dn)
                        dn_base = dn_match.group(1) if dn_match else None
                        if key == dn_base:
                            actual_rep = dn; break
                    if actual_rep and actual_rep not in processed_display:
                        ordered_display_names.append(actual_rep)
                        processed_display.add(actual_rep)
                elif key in display_items and key not in processed_display:
                    ordered_display_names.append(key)
                    processed_display.add(key)
            print(f"Displaying {len(ordered_display_names)} items/groups in validation window.", flush=True)
            # --- End Grouping Logic ---

        validation_window = tk.Toplevel(self.root)
        validation_window.title("ตรวจสอบ Value Code")
        validation_window.geometry("600x450")
        validation_window.grab_set()

        lbl_instruction = tk.Label(validation_window, text="พบ Value Code ที่เป็น 0 หรือเกิน 2000:\nเลือก Code ที่ต้องการลบออกจากผลลัพธ์:")
        lbl_instruction.pack(pady=10)

        frame_list = tk.Frame(validation_window); frame_list.pack(fill='both', expand=True, padx=10)
        text_area = tk.Text(frame_list, wrap='none', height=15, width=70)
        v_scrollbar = ttk.Scrollbar(frame_list, orient='vertical', command=text_area.yview)
        h_scrollbar = ttk.Scrollbar(frame_list, orient='horizontal', command=text_area.xview)
        text_area.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        text_area.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        frame_list.grid_rowconfigure(0, weight=1); frame_list.grid_columnconfigure(0, weight=1)

        # ** IMPORTANT: checkbutton_vars still maps the ACTUAL var/code pair to its BooleanVar **
        checkbutton_vars = {} # {(actual_var_name, original_code_str): tk.BooleanVar}
        text_area.config(state='normal')

        # Iterate through the ordered DISPLAY names (groups or single vars)
        for display_name in ordered_display_names:
            actual_vars_in_item = display_items.get(display_name, [])
            problematic_codes_details = defaultdict(list) # {formatted_code: [(actual_var, original_code_str, label), ...]}
            has_problems_in_item = False

            # Collect all problematic codes and their details for this item
            for actual_var_name in actual_vars_in_item:
                if actual_var_name in codes_to_validate:
                    has_problems_in_item = True
                    current_var_labels = all_value_labels.get(actual_var_name, {})
                    for original_code_str in codes_to_validate[actual_var_name]:
                        # Format Code Display (Remove .0 for integers)
                        try:
                            f_code = float(original_code_str)
                            formatted_code = str(int(f_code)) if f_code == int(f_code) else original_code_str
                        except (ValueError, TypeError):
                            formatted_code = original_code_str

                        # Get Label
                        label_text = "???"
                        try:
                            f_code = float(original_code_str)
                            if f_code == int(f_code):
                                label_text = current_var_labels.get(int(f_code), current_var_labels.get(f_code, current_var_labels.get(original_code_str, "???")))
                            else:
                                label_text = current_var_labels.get(f_code, current_var_labels.get(original_code_str, "???"))
                        except (ValueError, TypeError):
                            label_text = current_var_labels.get(original_code_str, "???")

                        # Store details grouped by the formatted code
                        problematic_codes_details[formatted_code].append((actual_var_name, original_code_str, label_text))

            if not has_problems_in_item:
                continue # Skip this display item

            # Insert header for the group/variable
            text_area.insert('end', f"ตัวแปร / กลุ่ม: {display_name}\n", ('header',))

            # --- Create Checkbuttons for UNIQUE formatted codes ---
            # Sort unique codes numerically for display
            unique_formatted_codes = sorted(problematic_codes_details.keys(), key=lambda x: float(x) if x.replace('.', '', 1).isdigit() else float('inf'))

            for formatted_code in unique_formatted_codes:
                # Get all actual occurrences for this unique code
                actual_occurrences = problematic_codes_details[formatted_code]
                if not actual_occurrences: continue # Should not happen, but safety check

                # Use the label from the first occurrence for display (assuming labels are consistent for the same code)
                _first_actual_var, _first_original_code, representative_label = actual_occurrences[0]

                # Create the display string for the unique code
                display_string = f"  Code: {formatted_code} ({representative_label})"

                # **Create ONE BooleanVar for this unique code display**
                bool_var = tk.BooleanVar(value=True) # Default to selected

                # **Link this ONE BooleanVar to ALL underlying actual occurrences**
                for actual_var, original_code, _label in actual_occurrences:
                    checkbutton_vars[(actual_var, original_code)] = bool_var

                # Create and embed the single Checkbutton for this unique code
                cb = tk.Checkbutton(text_area, text=display_string, variable=bool_var)
                text_area.window_create('end', window=cb)
                text_area.insert('end', '\n')
            # --- End Checkbutton Creation Loop ---

            text_area.insert('end', '\n') # Add space between display items

        text_area.tag_configure('header', font=('TkDefaultFont', 10, 'bold'))
        text_area.config(state='disabled')

        frame_buttons = tk.Frame(validation_window); frame_buttons.pack(pady=10)

        # --- on_confirm, select_all, deselect_all work WITHOUT changes ---
        # They operate on checkbutton_vars, which correctly links actual pairs to the shared BooleanVars
        def on_confirm():
            self.codes_to_delete_confirmed.clear()
            for (v_name, cd_str), var_obj in checkbutton_vars.items(): # Iterate through the detailed map
                if var_obj.get(): # Check the state of the linked BooleanVar
                    self.codes_to_delete_confirmed.add((v_name, cd_str))
            print(f"User confirmed deletion for {len(self.codes_to_delete_confirmed)} codes (actual var/code pairs): {self.codes_to_delete_confirmed}", flush=True)
            validation_window.destroy()

        def select_all():
            # Create a set of unique BooleanVar objects to avoid setting the same var multiple times
            unique_bool_vars = set(checkbutton_vars.values())
            for var_obj in unique_bool_vars:
                var_obj.set(True)

        def deselect_all():
            unique_bool_vars = set(checkbutton_vars.values())
            for var_obj in unique_bool_vars:
                var_obj.set(False)
        # -------------------------------------------------------------

        btn_confirm = tk.Button(frame_buttons, text="ตกลง (ลบรายการที่เลือก)", command=on_confirm, width=20)
        btn_confirm.grid(row=0, column=0, padx=5)
        btn_select_all = tk.Button(frame_buttons, text="เลือกทั้งหมด", command=select_all, width=15)
        btn_select_all.grid(row=0, column=1, padx=5)
        btn_deselect_all = tk.Button(frame_buttons, text="ไม่เลือกเลย", command=deselect_all, width=15)
        btn_deselect_all.grid(row=0, column=2, padx=5)

        self.root.wait_window(validation_window)
    # --- End of Updated _show_validation_window ---

    def show_loop_definition_window(self):
        """แสดงหน้าต่างให้ผู้ใช้กำหนดประเภท Loop (SA/MA/Loop Text/Loop Numeric) สำหรับแต่ละตัวแปร (แสดง _O<n> เป็นกลุ่ม)"""
        print("--- Opening Loop Definition Window (Grouping _O<n> Display) ---", flush=True) # Updated message
        spss_path = self.spss_file_path.get()
        if not spss_path:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกไฟล์ SPSS ก่อน")
            return
        meta_to_use = self.last_read_meta
        if meta_to_use is None:
            self.status_label.config(text="กำลังอ่านตัวแปร...", fg="orange")
            self.root.update_idletasks()
            try:
                _df = None; meta_to_use = None
                try: _df, meta_to_use = pyreadstat.read_sav(spss_path, metadataonly=True, encoding='utf-8')
                except pyreadstat.ReadstatError: _df, meta_to_use = pyreadstat.read_sav(spss_path, metadataonly=True, encoding='cp874')
                self.last_read_meta = meta_to_use
            except Exception as e: print(f"ERROR reading meta: {e}", flush=True); traceback.print_exc(); messagebox.showerror("ข้อผิดพลาดอ่าน SPSS", f"อ่าน Metadata ไม่ได้:\n{e}"); self.status_label.config(text="เกิดข้อผิดพลาดอ่าน SPSS", fg="red"); return
        try:
            if not hasattr(meta_to_use, 'column_names') or not meta_to_use.column_names:
                messagebox.showerror("ข้อผิดพลาด", "ไม่สามารถอ่านรายชื่อตัวแปรจากไฟล์ SPSS ได้"); self.status_label.config(text="เกิดข้อผิดพลาด: ไม่พบชื่อคอลัมน์", fg="red"); return

            all_var_names = meta_to_use.column_names
            print(f"Found {len(all_var_names)} variables. Grouping display for _O<n>...", flush=True)

            # --- Grouping Logic ---
            ma_pattern = re.compile(r'(.+)_O(\d+)$')
            groups = defaultdict(list) # {base_name: [full_var_name1, full_var_name2,...]}
            variable_order = [] # Maintain original order for display

            processed_indices_for_grouping = set()
            for i, var_name in enumerate(all_var_names):
                if i in processed_indices_for_grouping: continue

                match = ma_pattern.match(var_name)
                if match:
                    base_name = match.group(1)
                    # Find all related _O variables starting from this point to maintain relative order
                    current_group_vars = []
                    for j in range(i, len(all_var_names)):
                        if j in processed_indices_for_grouping: continue
                        inner_match = ma_pattern.match(all_var_names[j])
                        if inner_match and inner_match.group(1) == base_name:
                            current_group_vars.append(all_var_names[j])
                            processed_indices_for_grouping.add(j)
                    if current_group_vars:
                        groups[base_name].extend(current_group_vars) # Add found members
                        variable_order.append(base_name) # Add base_name to maintain order
                else:
                    # Non-MA variable, add directly
                    variable_order.append(var_name)
                    processed_indices_for_grouping.add(i)

            # --- Prepare items for Treeview ---
            items_for_treeview = [] # [(display_text, iid, initial_value, is_group, group_members)]
            iid_to_actual_vars = {} # Map Treeview IID to list of actual variable names

            processed_items = set() # Track added items (base_name for groups, var_name for singles)

            for item_key in variable_order: # Iterate using the preserved order key
                if item_key in processed_items: continue

                if item_key in groups: # It's a base_name representing an _O group
                    base_name = item_key
                    group_members = sorted(groups[base_name]) # Sort members for consistency
                    representative_iid = f"{base_name}_O1" # Use _O1 as the representative ID/display text

                    # Determine initial value: Check _O1 first, then first member, then empty
                    initial_value = self.variable_loop_types.get(representative_iid, "")
                    if not initial_value and group_members:
                        initial_value = self.variable_loop_types.get(group_members[0], "")

                    items_for_treeview.append((representative_iid, representative_iid, initial_value, True, group_members))
                    iid_to_actual_vars[representative_iid] = group_members
                    processed_items.add(base_name)
                    print(f"  Grouped: Display='{representative_iid}', Represents={group_members}", flush=True)

                else: # It's a single, non-_O variable
                    var_name = item_key
                    initial_value = self.variable_loop_types.get(var_name, "")
                    items_for_treeview.append((var_name, var_name, initial_value, False, [var_name]))
                    iid_to_actual_vars[var_name] = [var_name]
                    processed_items.add(var_name)

            # --- Create GUI Window ---
            loop_window = tk.Toplevel(self.root); loop_window.title("กำหนดตัวแปร Loop (SA/MA/Text/Numeric)"); loop_window.geometry("650x550"); loop_window.grab_set()
            lbl_instruction = tk.Label(loop_window, text="เลือกแถว (ใช้ Shift/Ctrl) แล้วกดปุ่มด้านล่าง หรือคลิก 'Loop Type' เพื่อสลับ (SA<->MA<->Text<->Numeric<->ว่าง)"); lbl_instruction.pack(pady=(10, 5))
            tree_frame = ttk.Frame(loop_window); tree_frame.pack(fill='both', expand=True, padx=10, pady=5)
            columns = ('loop_type',); tree = ttk.Treeview(tree_frame, columns=columns, show='tree headings', selectmode='extended')
            tree.heading('#0', text='Variable Name / Group'); tree.heading('loop_type', text='Loop Type') # Updated heading
            tree.column('#0', width=350, stretch=tk.YES, anchor='w'); tree.column('loop_type', width=150, stretch=tk.NO, anchor='center')
            vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview); tree.configure(yscrollcommand=vsb.set); vsb.pack(side='right', fill='y'); tree.pack(side='left', fill='both', expand=True)

            # Populate Treeview using prepared items
            for display_text, iid, initial_value, is_group, _ in items_for_treeview:
                tree.insert('', tk.END, iid=iid, text=display_text, values=(initial_value,))

            # --- Event Handlers (Modified to handle groups) ---
            def on_tree_click(event):
                region = tree.identify("region", event.x, event.y); column_id = tree.identify_column(event.x)
                item_iid = tree.identify_row(event.y) # This is the IID (rep name or single var name)

                if region != "cell" or not item_iid or column_id != '#1': return

                current_values = tree.item(item_iid, 'values'); current_type = current_values[0] if current_values else ""
                if current_type == "": next_type = "SA"
                elif current_type == "SA": next_type = "MA"
                elif current_type == "MA": next_type = "Loop Text"
                elif current_type == "Loop Text": next_type = "Loop Numeric"
                elif current_type == "Loop Numeric": next_type = ""
                else: next_type = "SA"

                tree.item(item_iid, values=(next_type,)) # Update display

                # Update underlying dictionary for ALL actual variables represented by this IID
                actual_vars = iid_to_actual_vars.get(item_iid, [])
                print(f"  Click Update: IID='{item_iid}', NextType='{next_type}', ActualVars={actual_vars}", flush=True)
                for actual_var in actual_vars:
                    if next_type:
                        self.variable_loop_types[actual_var] = next_type
                    elif actual_var in self.variable_loop_types:
                        del self.variable_loop_types[actual_var]

            tree.bind('<ButtonRelease-1>', on_tree_click)

            def apply_to_selected(loop_type_to_set):
                selected_iids = tree.selection()
                if not selected_iids:
                    messagebox.showwarning("ไม่มีรายการที่เลือก", "กรุณาเลือกตัวแปร/กลุ่ม ในตารางก่อน", parent=loop_window); return

                print(f"Applying '{loop_type_to_set}' to {len(selected_iids)} selected items/groups...", flush=True)
                updated_var_count = 0
                for item_iid in selected_iids:
                    try:
                        tree.item(item_iid, values=(loop_type_to_set,)) # Update display
                        # Update underlying dictionary for ALL actual variables
                        actual_vars = iid_to_actual_vars.get(item_iid, [])
                        for actual_var in actual_vars:
                            if loop_type_to_set:
                                self.variable_loop_types[actual_var] = loop_type_to_set
                            elif actual_var in self.variable_loop_types:
                                del self.variable_loop_types[actual_var]
                            updated_var_count += 1 # Count actual vars updated
                    except tk.TclError: print(f"Warning: Item {item_iid} not found. Skipping.", flush=True)
                print(f"  Applied to {len(selected_iids)} rows, affecting {updated_var_count} actual variables.", flush=True)

            # --- Buttons (Remain the same visually) ---
            multi_select_frame = tk.Frame(loop_window); multi_select_frame.pack(pady=5)
            btn_set_sa = tk.Button(multi_select_frame, text="Set Selected SA", command=lambda: apply_to_selected("SA"), width=15); btn_set_sa.pack(side=tk.LEFT, padx=5)
            btn_set_ma = tk.Button(multi_select_frame, text="Set Selected MA", command=lambda: apply_to_selected("MA"), width=15); btn_set_ma.pack(side=tk.LEFT, padx=5)
            btn_set_loop_text = tk.Button(multi_select_frame, text="Set Selected Loop Text", command=lambda: apply_to_selected("Loop Text"), width=20); btn_set_loop_text.pack(side=tk.LEFT, padx=5)
            btn_set_loop_numeric = tk.Button(multi_select_frame, text="Set Selected Loop Numeric", command=lambda: apply_to_selected("Loop Numeric"), width=22); btn_set_loop_numeric.pack(side=tk.LEFT, padx=5)
            btn_clear_selected = tk.Button(multi_select_frame, text="Clear Selected", command=lambda: apply_to_selected(""), width=15); btn_clear_selected.pack(side=tk.LEFT, padx=5)

            button_frame = tk.Frame(loop_window); button_frame.pack(pady=10)
            def on_loop_confirm():
                # self.variable_loop_types is already updated by clicks/button presses
                print(f"Loop definitions confirmed: {len(self.variable_loop_types)} vars defined.")
                # print(self.variable_loop_types) # Optional debug print
                messagebox.showinfo("บันทึกสำเร็จ", f"บันทึกการกำหนด Loop แล้ว ({len(self.variable_loop_types)} รายการ)", parent=loop_window)
                loop_window.destroy(); self.status_label.config(text="กำหนด Loop เสร็จสิ้น", fg="blue")

            def on_loop_cancel():
                # Changes made ARE currently saved in self.variable_loop_types
                # To truly cancel, would need to store initial state and restore here
                loop_window.destroy(); self.status_label.config(text="ยกเลิกการกำหนด Loop", fg="blue")

            confirm_button = tk.Button(button_frame, text="ตกลง", command=on_loop_confirm, width=10); confirm_button.pack(side=tk.LEFT, padx=5)
            cancel_button = tk.Button(button_frame, text="ยกเลิก", command=on_loop_cancel, width=10); cancel_button.pack(side=tk.LEFT, padx=5)

            self.status_label.config(text="กรุณากำหนดตัวแปร Loop", fg="blue"); self.root.wait_window(loop_window)

        except Exception as e:
            print(f"ERROR in Loop Window: {e}", flush=True); traceback.print_exc()
            messagebox.showerror("ข้อผิดพลาด", f"เกิดข้อผิดพลาดในหน้าต่างกำหนด Loop:\n{e}"); self.status_label.config(text="เกิดข้อผิดพลาดเปิดหน้าต่าง", fg="red")

    def show_loop_naming_window(self):
        """แสดงหน้าต่างให้ผู้ใช้กำหนดชื่อ Loop ID โดยรวมกลุ่มตาม Base Name (ที่ปรับปรุงแล้ว) และแสดง Type ที่ User กำหนด"""
        print("--- Opening Loop Naming Window (Improved Grouping) ---", flush=True) # Updated message
        spss_path = self.spss_file_path.get()
        if not spss_path:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกไฟล์ SPSS ก่อน")
            return
        if not self.variable_loop_types:
            messagebox.showinfo("ข้อมูล", "ยังไม่มีการกำหนดตัวแปร Loop (SA/MA/Text/Numeric)", parent=self.root)
            return

        # --- 1. ระบุกลุ่ม Loop และหา Base Name (ใช้ Heuristic ที่ปรับปรุงแล้ว) ---
        meta_to_use = self.last_read_meta
        if meta_to_use is None:
            try:
                print("Reading metadata for loop naming window...", flush=True)
                _df = None; meta_to_use = None
                try: _df, meta_to_use = pyreadstat.read_sav(spss_path, metadataonly=True, encoding='utf-8')
                except pyreadstat.ReadstatError: _df, meta_to_use = pyreadstat.read_sav(spss_path, metadataonly=True, encoding='cp874')
                self.last_read_meta = meta_to_use
            except Exception as e: print(f"ERROR reading metadata: {e}", flush=True); messagebox.showerror("ข้อผิดพลาดอ่าน SPSS", f"อ่าน Metadata ไม่ได้:\n{e}"); return

        if not hasattr(meta_to_use, 'column_names'): messagebox.showerror("ข้อผิดพลาด", "ไม่พบชื่อคอลัมน์ใน Metadata"); return

        all_vars = meta_to_use.column_names
        var_name_to_index = {name: i for i, name in enumerate(all_vars)}

        identified_loops = {} # key: first_var_name, value: {'type': user_type, 'default_name': name, 'vars': list}
        temp_processed_indices = set()
        i = 0
        print("Identifying initial loop groups with updated heuristic...", flush=True)
        while i < len(all_vars):
            if i in temp_processed_indices: i += 1; continue
            current_var = all_vars[i]
            current_type = self.variable_loop_types.get(current_var, "")
            if current_type in ["SA", "MA", "Loop Text", "Loop Numeric"]:
                current_group_vars = [current_var]
                # *** ใช้ Heuristic ที่ปรับปรุงแล้ว ***
                current_base = get_base_name_heuristic(current_var)
                current_prefix_match = re.match(r'(I_\d+_)', current_var)
                current_prefix = current_prefix_match.group(1) if current_prefix_match else None
                j = i + 1
                while j < len(all_vars):
                    next_var = all_vars[j]
                    next_type = self.variable_loop_types.get(next_var, "")
                    # *** ใช้ Heuristic ที่ปรับปรุงแล้ว ***
                    next_base = get_base_name_heuristic(next_var)
                    stop_grouping = False
                    if next_type != current_type or next_base != current_base: # Group by type AND base name
                        stop_grouping = True
                    elif not stop_grouping and (current_prefix is not None or re.match(r'I_\d+_(.+?)(_O\d+)?$', current_var)):
                        next_prefix_match = re.match(r'(I_\d+_)', next_var); next_prefix = next_prefix_match.group(1) if next_prefix_match else None
                        if current_prefix is not None and next_prefix is not None and current_prefix != next_prefix: stop_grouping = True
                    if not stop_grouping: current_group_vars.append(next_var); j += 1
                    else: break

                first_var_name = current_group_vars[0]
                #default_loop_name = current_base # Default name comes from the new heuristic
                # ใหม่: ถ้า base เป็น I_<n>_<stem> ให้สลับเป็น <stem>_<n> เพื่อได้ a4_1, a4_2, ...
                m_base_i = re.match(r'^I_(\d+)_([A-Za-z]\w+)$', current_base)
                if m_base_i:
                    default_loop_name = f"{m_base_i.group(2)}_{m_base_i.group(1)}"
                else:
                    default_loop_name = current_base
                identified_loops[first_var_name] = {'type': current_type, 'default_name': default_loop_name, 'vars': current_group_vars}
                for k in range(i, j): temp_processed_indices.add(k)
                i = j; continue
            i += 1
        print(f"Finished initial identification. Found {len(identified_loops)} potential groups.", flush=True)

        if not identified_loops: messagebox.showinfo("ข้อมูล", "ไม่พบกลุ่ม Loop ที่กำหนดไว้", parent=self.root); return

        # --- 2. กรองและรวมกลุ่มเพื่อการแสดงผล (Logic เดิมทำงานได้ดีกับ Heuristic ใหม่) ---
        items_to_display = [] # List of (representative_first_var, loop_info, final_loop_id)
        print("Consolidating loops for naming display...", flush=True)
        sorted_identified_loops = sorted(identified_loops.items(), key=lambda item: var_name_to_index.get(item[0], float('inf')))

        # Group by final ID (user-defined or default from new heuristic)
        groups_by_final_id_for_display = defaultdict(list)
        for first_var, loop_info in sorted_identified_loops:
             # Final ID is either user-set or the default name from the *new* heuristic
             final_loop_id = self.user_defined_loop_names.get(first_var, loop_info['default_name'])
             groups_by_final_id_for_display[final_loop_id].append((first_var, loop_info))

        # Select the representative (first in SPSS order) for each final ID group
        processed_display_ids = set()
        # Iterate through the loops *in their original SPSS order* to ensure representative is correct
        for first_var_sorted, loop_info_sorted in sorted_identified_loops:
            final_loop_id = self.user_defined_loop_names.get(first_var_sorted, loop_info_sorted['default_name'])
            if final_loop_id not in processed_display_ids:
                # Add the first occurrence of this final_loop_id as the representative
                items_to_display.append((first_var_sorted, loop_info_sorted, final_loop_id))
                processed_display_ids.add(final_loop_id)
                print(f"  Adding display row: Rep='{first_var_sorted}', FinalID='{final_loop_id}', Type='{loop_info_sorted['type']}'", flush=True)

        print(f"Finished consolidating display. {len(items_to_display)} unique rows.", flush=True)

        # --- 3. สร้างหน้าต่างและ Widget แสดงผล (เหมือนเดิม) ---
        naming_window = tk.Toplevel(self.root); naming_window.title("กำหนดชื่อ Loop ID (Consolidated)"); naming_window.geometry("600x450"); naming_window.grab_set()
        lbl_instruction = tk.Label(naming_window, text="แก้ไขชื่อ Loop ID ที่ต้องการในคอลัมน์ 'Loop ID Name (Editable)'"); lbl_instruction.pack(pady=10)
        main_frame = tk.Frame(naming_window); main_frame.pack(fill='both', expand=True, padx=10, pady=5)
        canvas = tk.Canvas(main_frame); scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas); scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw"); canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True); scrollbar.pack(side="right", fill="y")

        # ----- Start: Add Mouse Wheel Binding -----
        def _on_mousewheel(event):
            # Determine scroll direction/amount based on platform
            # Use negative delta for yview_scroll because yview_scroll positive scrolls down
            scroll_units = 0
            # Linux uses event.num, Windows/Mac uses event.delta
            if event.num == 4 or event.delta > 0:  # Linux scroll up or Windows/Mac scroll up
                scroll_units = -1
            elif event.num == 5 or event.delta < 0: # Linux scroll down or Windows/Mac scroll down
                scroll_units = 1

            if scroll_units != 0:
                canvas.yview_scroll(scroll_units, "units")
                return "break" # Prevent default scroll behavior if necessary

        # --- Bind to multiple widgets to ensure coverage ---

        # 1. Bind to the canvas itself
        canvas.bind("<MouseWheel>", _on_mousewheel) # Windows/Mac
        canvas.bind("<Button-4>", _on_mousewheel)   # Linux scroll up
        canvas.bind("<Button-5>", _on_mousewheel)   # Linux scroll down

        # 2. Bind to the frame inside the canvas
        scrollable_frame.bind("<MouseWheel>", _on_mousewheel)
        scrollable_frame.bind("<Button-4>", _on_mousewheel)
        scrollable_frame.bind("<Button-5>", _on_mousewheel)

        # 3. Bind recursively to all children of the scrollable_frame
        #    This ensures scrolling works when the mouse is over labels, entries, etc.
        def bind_recursive(widget, event_name, callback):
             # Only bind if the widget doesn't already have a binding for this event
             # to avoid potential conflicts, although for scrolling it's usually fine.
             if not widget.bind(event_name):
                 widget.bind(event_name, callback)
             for child in widget.winfo_children():
                 bind_recursive(child, event_name, callback)

        # Apply recursive binding for all relevant events
        for event_name in ["<MouseWheel>", "<Button-4>", "<Button-5>"]:
             bind_recursive(scrollable_frame, event_name, _on_mousewheel)
             # Ensure the canvas itself is also bound (might be redundant but safe)
             canvas.bind(event_name, _on_mousewheel)

        # ----- End: Add Mouse Wheel Binding -----


        entry_widgets = {} # Key: representative first_var, Value: tk.StringVar
        header_frame = ttk.Frame(scrollable_frame); header_frame.pack(fill='x', pady=(0, 5))
        ttk.Label(header_frame, text="Representative Var", width=25, anchor='w').pack(side='left', padx=5)
        ttk.Label(header_frame, text="Loop Type", width=15, anchor='center').pack(side='left', padx=5)
        ttk.Label(header_frame, text="Loop ID Name (Editable)", width=40, anchor='w').pack(side='left', padx=5)

        for first_var, loop_info, final_loop_id in items_to_display:
            loop_type = loop_info['type']
            initial_name = final_loop_id # The default/user name for the whole group
            row_frame = ttk.Frame(scrollable_frame); row_frame.pack(fill='x', pady=2)
            ttk.Label(row_frame, text=first_var, width=25, anchor='w', wraplength=150).pack(side='left', padx=5)
            ttk.Label(row_frame, text=loop_type, width=15, anchor='center').pack(side='left', padx=5)
            entry_var = tk.StringVar(value=initial_name)
            entry = ttk.Entry(row_frame, textvariable=entry_var, width=40)
            entry.pack(side='left', padx=5, fill='x', expand=True)
            entry_widgets[first_var] = entry_var # Key is the representative var

        # --- ส่วนปุ่มกด (Logic การ Save ไม่ต้องเปลี่ยน) ---
        button_frame = tk.Frame(naming_window); button_frame.pack(pady=10)
        def on_name_save():
            temp_edited_names = {} # {representative_first_var: edited_name}
            print("Reading edited loop names...", flush=True)
            for rep_first_var, entry_var_value in entry_widgets.items():
                loop_name = entry_var_value.get().strip()
                if loop_name: temp_edited_names[rep_first_var] = loop_name

            final_user_names = {} # {original_first_var: final_name}
            processed_original_first_vars = set()

            # Apply edited names back to *all* original variables belonging to the group
            for original_first_var, original_loop_info in identified_loops.items():
                 if original_first_var in processed_original_first_vars: continue
                 current_final_id = self.user_defined_loop_names.get(original_first_var, original_loop_info['default_name'])
                 representative_for_this_group = None
                 for rep_var_disp, _, disp_final_id in items_to_display:
                     if disp_final_id == current_final_id:
                         representative_for_this_group = rep_var_disp; break

                 final_name_to_set = current_final_id # Default
                 if representative_for_this_group and representative_for_this_group in temp_edited_names:
                     final_name_to_set = temp_edited_names[representative_for_this_group] # Use edited name

                 # Find all other original vars that map to the same final_id and apply the name
                 for ov, oi in identified_loops.items():
                      check_final_id = self.user_defined_loop_names.get(ov, oi['default_name'])
                      if check_final_id == current_final_id:
                          if final_name_to_set: final_user_names[ov] = final_name_to_set
                          processed_original_first_vars.add(ov)

            self.user_defined_loop_names = final_user_names
            print(f"Final loop names saved: {len(self.user_defined_loop_names)} mappings.", flush=True)
            messagebox.showinfo("บันทึกสำเร็จ", f"บันทึกชื่อ Loop ID แล้ว ({len(self.user_defined_loop_names)} รายการ)", parent=naming_window)
            naming_window.destroy(); self.status_label.config(text="กำหนดชื่อ Loop ID เสร็จสิ้น", fg="blue")

        def on_name_cancel():
            naming_window.destroy(); self.status_label.config(text="ยกเลิกการกำหนดชื่อ Loop ID", fg="blue")

        save_button = tk.Button(button_frame, text="บันทึก", command=on_name_save, width=10); save_button.pack(side=tk.LEFT, padx=5)
        cancel_button = tk.Button(button_frame, text="ยกเลิก", command=on_name_cancel, width=10); cancel_button.pack(side=tk.LEFT, padx=5)
        self.status_label.config(text="กรุณากำหนดชื่อ Loop ID", fg="blue"); self.root.wait_window(naming_window)


    # --- NEW Function: Save Loop Settings (Modified) ---
    def save_loop_settings(self):
        """
        Saves the loop definition structure AS DISPLAYED in the definition window
        to an Excel file, including dropdowns for Loop Type selection.
        """
        print("--- Saving Loop Settings (Display Structure) ---", flush=True)
        spss_filepath = self.spss_file_path.get()
        if not spss_filepath:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกไฟล์ SPSS ก่อน", parent=self.root)
            return

        meta_to_use = self.last_read_meta
        if meta_to_use is None or not hasattr(meta_to_use, 'column_names') or not meta_to_use.column_names:
             messagebox.showerror("ข้อผิดพลาด", "ไม่สามารถอ่าน Metadata หรือรายชื่อตัวแปรจากไฟล์ SPSS ได้\nกรุณาลองเลือกไฟล์ SPSS ใหม่อีกครั้ง", parent=self.root)
             return

        # --- 1. Regenerate the display list structure ---
        # This replicates the logic from show_loop_definition_window to get the exact items displayed
        all_var_names = meta_to_use.column_names
        print(f"Regenerating display list from {len(all_var_names)} variables...", flush=True)

        ma_pattern = re.compile(r'(.+)_O(\d+)$')
        groups = defaultdict(list)
        variable_order = []
        processed_indices_for_grouping = set()

        for i, var_name in enumerate(all_var_names):
            if i in processed_indices_for_grouping: continue
            match = ma_pattern.match(var_name)
            if match:
                base_name = match.group(1)
                current_group_vars = []
                for j in range(i, len(all_var_names)):
                    if j in processed_indices_for_grouping: continue
                    inner_match = ma_pattern.match(all_var_names[j])
                    if inner_match and inner_match.group(1) == base_name:
                        current_group_vars.append(all_var_names[j])
                        processed_indices_for_grouping.add(j)
                if current_group_vars:
                    groups[base_name].extend(current_group_vars)
                    variable_order.append(base_name) # Use base name as the key in order
            else:
                variable_order.append(var_name) # Use var name as the key
                processed_indices_for_grouping.add(i)

        # Build the list of (display_text, current_loop_type) to save
        data_to_save = []
        processed_items_save = set()
        for item_key in variable_order:
             if item_key in processed_items_save: continue

             if item_key in groups: # It's a base_name representing an _O group
                 base_name = item_key
                 group_members = sorted(groups[base_name])
                 if not group_members: continue # Skip empty groups if they somehow occur

                 display_text = f"{base_name}_O1" # Representative display name
                 # Use the first actual member to look up the CURRENT loop type
                 # This assumes the type is consistent across the group members in self.variable_loop_types
                 first_member_key = group_members[0]
                 current_loop_type = self.variable_loop_types.get(first_member_key, "") # Get current type
                 data_to_save.append((display_text, current_loop_type))
                 processed_items_save.add(base_name)
                 # print(f"  Prepared Group: '{display_text}' -> '{current_loop_type}' (from {first_member_key})", flush=True)

             else: # It's a single, non-_O variable
                 var_name = item_key
                 display_text = var_name
                 current_loop_type = self.variable_loop_types.get(var_name, "") # Get current type
                 data_to_save.append((display_text, current_loop_type))
                 processed_items_save.add(var_name)
                 # print(f"  Prepared Single: '{display_text}' -> '{current_loop_type}'", flush=True)

        if not data_to_save:
             messagebox.showinfo("ข้อมูล", "ไม่พบรายการตัวแปรที่จะบันทึก", parent=self.root)
             return
        print(f"Prepared {len(data_to_save)} items for saving.", flush=True)

        # --- 2. Ask for save file path ---
        initial_dir = os.path.dirname(spss_filepath)
        file_name_without_ext = os.path.splitext(os.path.basename(spss_filepath))[0]
        # Changed suggested filename slightly
        suggested_filename = f"{file_name_without_ext}_loop_definitions.xlsx"

        save_path = filedialog.asksaveasfilename(
            initialdir=initial_dir,
            initialfile=suggested_filename,
            title="บันทึกโครงสร้าง Loop ที่แสดง (.xlsx)",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
            defaultextension=".xlsx"
        )

        if not save_path:
            print("Save loop definitions cancelled.", flush=True)
            return

        if not save_path.lower().endswith(".xlsx"):
            save_path += ".xlsx"

        self.status_label.config(text="กำลังบันทึกโครงสร้าง Loop...", fg="orange")
        self.root.update_idletasks()

        # --- 3. Write to Excel with Dropdown ---
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Loop Definitions"

            # Write Header
            ws['A1'] = "Variable Name / Group"
            ws['B1'] = "Loop Type"
            ws.column_dimensions['A'].width = 40 # Adjust width
            ws.column_dimensions['B'].width = 25 # Adjust width

            # Write Data
            row_num = 2
            for display_name, loop_type in data_to_save:
                ws.cell(row=row_num, column=1).value = display_name
                ws.cell(row=row_num, column=2).value = loop_type # Write the current value
                row_num += 1

            # --- Add Data Validation (Dropdown) ---
            # Define the list for the dropdown
            # Note: Excel requires the list as a comma-separated string within quotes
            allowed_values_string = '"SA,MA,Loop Text,Loop Numeric,"' # Trailing comma allows blank selection
            dv = DataValidation(type="list", formula1=allowed_values_string, allow_blank=True)
            dv.error = 'ค่าที่เลือกไม่ถูกต้อง' # Optional: Error message
            dv.errorTitle = 'Loop Type ไม่ถูกต้อง' # Optional: Error title
            dv.prompt = 'กรุณาเลือกประเภท Loop' # Optional: Input message
            dv.promptTitle = 'เลือกประเภท' # Optional: Input title

            # Add the validation to the sheet
            ws.add_data_validation(dv)

            # Apply the validation to all data rows in column B (Loop Type)
            last_data_row = ws.max_row
            if last_data_row >= 2: # Ensure there is data to apply to
                 cell_range = f"B2:B{last_data_row}"
                 dv.add(cell_range)
                 print(f"Added dropdown validation to range: {cell_range}", flush=True)
            # ------------------------------------

            wb.save(save_path)
            print(f"Loop definitions (display structure) saved to: {save_path}", flush=True)
            self.status_label.config(text=f"บันทึกโครงสร้าง Loop สำเร็จ: {os.path.basename(save_path)}", fg="green")
            messagebox.showinfo("สำเร็จ", f"บันทึกโครงสร้าง Loop ที่แสดง สำเร็จ!\nไฟล์: {save_path}", parent=self.root)

        except PermissionError:
            print(f"ERROR: Permission denied saving loop definitions to {save_path}", flush=True)
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถบันทึกไฟล์ได้:\n{save_path}\nอาจจะเปิดไฟล์นี้อยู่ หรือไม่มีสิทธิ์เขียน", parent=self.root)
            self.status_label.config(text="เกิดข้อผิดพลาด: ไม่มีสิทธิ์บันทึกไฟล์", fg="red")
        except Exception as e:
            print(f"ERROR saving loop definitions: {e}", flush=True)
            traceback.print_exc()
            messagebox.showerror("ข้อผิดพลาด", f"เกิดข้อผิดพลาดในการบันทึก:\n{e}", parent=self.root)
            self.status_label.config(text="เกิดข้อผิดพลาดในการบันทึก Loop Definitions", fg="red")

    # --- NEW Function: Load Loop Settings ---
    # --- NEW Function: Load Loop Settings (Modified to handle MA group representatives) ---
    def load_loop_settings(self):
        """
        Loads loop definitions from an Excel file and applies them to the actual underlying
        variables in the current SPSS file. Optimized to avoid O(N*M) scans that can hang.
        """
        print("--- Loading Loop Settings (Optimized / MA index) ---", flush=True)
        spss_filepath = self.spss_file_path.get()
        if not spss_filepath:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกไฟล์ SPSS ก่อน")
            return

        # Ensure metadata
        if self.last_read_meta is None:
            print("Metadata missing, reading SPSS metadata...", flush=True)
            try:
                _df = None; meta_temp = None
                try:
                    _df, meta_temp = pyreadstat.read_sav(spss_filepath, metadataonly=True, encoding='utf-8')
                except pyreadstat.ReadstatError:
                    _df, meta_temp = pyreadstat.read_sav(spss_filepath, metadataonly=True, encoding='cp874')
                self.last_read_meta = meta_temp
            except Exception as e:
                print(f"ERROR reading metadata: {e}", flush=True)
                messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถโหลด Metadata ของไฟล์ SPSS ปัจจุบันได้:\n{e}", parent=self.root)
                return

        if not hasattr(self.last_read_meta, 'column_names') or not self.last_read_meta.column_names:
            messagebox.showerror("ข้อผิดพลาด", "ไม่พบชื่อตัวแปรในไฟล์ SPSS ปัจจุบัน", parent=self.root)
            return

        all_spss_vars_list = self.last_read_meta.column_names
        valid_spss_vars_set = set(all_spss_vars_list)
        print(f"- SPSS variables: {len(valid_spss_vars_set)}", flush=True)

        # ---------- NEW: Build MA index once ----------
        # base: "xxx" for variables like "xxx_O1", "xxx_O2", ...
        ma_pat = re.compile(r'^(?P<base>.+)_O(?P<opt>\d+)$')
        ma_index = {}  # { base -> [members in file order] }
        for v in all_spss_vars_list:
            m = ma_pat.match(v)
            if m:
                base = m.group('base')
                ma_index.setdefault(base, []).append(v)

        # Keep order stable
        for base, members in ma_index.items():
            members.sort(key=lambda name: int(ma_pat.match(name).group('opt')) if ma_pat.match(name) else 10**9)
        print(f"- MA groups indexed: {len(ma_index)}", flush=True)
        # ------------------------------------------------

        load_path = filedialog.askopenfilename(
            initialdir=".",
            title="เลือกไฟล์ตั้งค่า Loop (.xlsx)",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        if not load_path:
            print("Load loop settings cancelled.", flush=True)
            return

        self.status_label.config(text="กำลังโหลดการตั้งค่า Loop...", fg="orange")
        self.root.update_idletasks()

        # Will store settings mapped to actual variables
        loaded_settings_mapped = {}
        skipped_count = 0
        processed_rows = 0
        valid_loop_types = {"SA", "MA", "Loop Text", "Loop Numeric", ""}

        try:
            wb = openpyxl.load_workbook(load_path, read_only=True, data_only=True)
            ws = wb.active

            # Optional: header sanity check (not strict)
            # Expect Col A: "Variable Name / Group", Col B: "Loop Type"
            hdr_a = (ws.cell(row=1, column=1).value or "").strip() if ws.max_row >= 1 else ""
            hdr_b = (ws.cell(row=1, column=2).value or "").strip() if ws.max_row >= 1 else ""
            if not hdr_a or not hdr_b:
                print("Warning: Header not found or empty. Proceeding anyway.", flush=True)

            # Main pass
            print(f"- Reading Excel rows: {ws.max_row - 1 if ws.max_row >= 2 else 0}", flush=True)
            for row_num in range(2, ws.max_row + 1):
                processed_rows += 1
                if processed_rows % 200 == 0:
                    print(f"  ...processed {processed_rows} rows", flush=True)
                    # ให้ UI ยังวาดอยู่ จะไม่ดูเหมือนค้าง
                    self.root.update_idletasks()

                excel_var_name = ws.cell(row=row_num, column=1).value
                excel_loop_type = ws.cell(row=row_num, column=2).value

                if not isinstance(excel_var_name, str) or not excel_var_name.strip():
                    skipped_count += 1
                    continue
                excel_var_name = excel_var_name.strip()
                excel_loop_type = (excel_loop_type or "").strip()

                if excel_loop_type not in valid_loop_types:
                    print(f"  Skip row {row_num}: invalid Loop Type '{excel_loop_type}' for '{excel_var_name}'", flush=True)
                    skipped_count += 1
                    continue

                # Case A: looks like MA group representative (we save as "<base>_O1" when exporting)
                m = ma_pat.match(excel_var_name)
                is_rep_candidate = bool(m) and excel_loop_type == "MA"
                if is_rep_candidate:
                    base = m.group('base')
                    # Use prebuilt index instead of scanning all vars
                    members = ma_index.get(base, [])
                    if not members:
                        print(f"  Warning row {row_num}: MA base '{base}' not found in SPSS, skip.", flush=True)
                        skipped_count += 1
                        continue
                    # Apply loop type to every member under that base
                    for spss_var in members:
                        if excel_loop_type:
                            loaded_settings_mapped[spss_var] = excel_loop_type
                        elif spss_var in loaded_settings_mapped:
                            del loaded_settings_mapped[spss_var]
                    continue

                # Case B: treat as a single variable name
                if excel_var_name in valid_spss_vars_set:
                    if excel_loop_type:
                        loaded_settings_mapped[excel_var_name] = excel_loop_type
                    elif excel_var_name in loaded_settings_mapped:
                        del loaded_settings_mapped[excel_var_name]
                else:
                    # เพิ่มกรณี: ถ้า user ใส่ชื่อ base ตรง ๆ (เช่น "I_1_a4") และ type=MA
                    # จะลองขยายด้วย index เช่นกัน
                    if excel_loop_type == "MA" and excel_var_name in ma_index:
                        for spss_var in ma_index[excel_var_name]:
                            loaded_settings_mapped[spss_var] = "MA"
                    else:
                        skipped_count += 1

            # Replace existing loop types atomically
            self.variable_loop_types.clear()
            self.variable_loop_types.update(loaded_settings_mapped)

            loaded_count = len(self.variable_loop_types)
            print(f"Done. Set loop type for {loaded_count} variables. Skipped {skipped_count} rows of {processed_rows}.", flush=True)
            self.status_label.config(text=f"โหลดการตั้งค่า Loop สำเร็จ ({loaded_count} รายการ)", fg="green")
            messagebox.showinfo(
                "สำเร็จ",
                f"โหลดการตั้งค่า Loop สำเร็จ!\n- กำหนด Loop Type ให้ {loaded_count} ตัวแปร\n- ข้าม {skipped_count} แถว (ชื่อไม่พบ/Loop Type ไม่ถูกต้อง)\n\nคุณสามารถเปิด '1. กำหนดตัวแปร Loop' เพื่อตรวจสอบได้",
                parent=self.root
            )

        except FileNotFoundError:
            print(f"ERROR: Loop settings file not found: {load_path}", flush=True)
            messagebox.showerror("ข้อผิดพลาด", f"ไม่พบไฟล์:\n{load_path}", parent=self.root)
            self.status_label.config(text="เกิดข้อผิดพลาด: ไม่พบไฟล์ตั้งค่า", fg="red")
        except Exception as e:
            print(f"ERROR loading loop settings: {e}", flush=True)
            traceback.print_exc()
            messagebox.showerror("ข้อผิดพลาด", f"เกิดข้อผิดพลาดในการโหลด:\n{e}", parent=self.root)
            self.status_label.config(text="เกิดข้อผิดพลาดในการโหลด Loop Settings", fg="red")



    # เพิ่มเมธอดใหม่ในคลาส SpssToExcelConverter
    def export_rawdata_excel(self):
        """Exports raw data from SPSS (.sav) to Excel (.xlsx) with only Rawdata_Code sheet."""
        print("--- Starting export_rawdata_excel ---", flush=True)
        spss_path = self.spss_file_path.get()
        
        if not spss_path:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกไฟล์ SPSS ก่อน", parent=self.root)
            print("No SPSS file selected.", flush=True)
            return

        # Suggest output filename
        initial_dir = os.path.dirname(spss_path)
        file_name_without_ext = os.path.splitext(os.path.basename(spss_path))[0]
        suggested_filename = f"{file_name_without_ext}_Rawdata.xlsx"
        counter = 1
        final_suggested_path = os.path.join(initial_dir, suggested_filename)
        while os.path.exists(final_suggested_path):
            suggested_filename = f"{file_name_without_ext}_rawdata_{counter}.xlsx"
            final_suggested_path = os.path.join(initial_dir, suggested_filename)
            counter += 1

        # Ask for save location
        excel_output_path = filedialog.asksaveasfilename(
            initialdir=initial_dir,
            initialfile=suggested_filename,
            title="บันทึก Rawdata เป็น Excel (.xlsx)",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
            defaultextension=".xlsx"
        )

        if not excel_output_path:
            print("Export rawdata cancelled.", flush=True)
            self.status_label.config(text="ยกเลิกการ Export Rawdata", fg="blue")
            return

        if not excel_output_path.lower().endswith(".xlsx"):
            excel_output_path += ".xlsx"

        self.status_label.config(text="กำลัง Export ข้อมูลดิบเป็น Excel...", fg="orange")
        self.root.update_idletasks()

        try:
            # Read SPSS data without labels (for Rawdata_Code sheet)
            print(f"Reading data with codes from: {spss_path}", flush=True)
            df_codes = None
            meta_codes = None
            try:
                print("Attempting UTF-8 encoding for codes...", flush=True)
                df_codes, meta_codes = pyreadstat.read_sav(spss_path, encoding='utf-8', apply_value_formats=False)
                print("Successfully read using UTF-8 without value labels.", flush=True)
            except pyreadstat.ReadstatError as e_utf8:
                print(f"UTF-8 failed: {e_utf8}. Trying cp874...", flush=True)
                try:
                    df_codes, meta_codes = pyreadstat.read_sav(spss_path, encoding='cp874', apply_value_formats=False)
                    print("Successfully read using cp874 without value labels.", flush=True)
                except pyreadstat.ReadstatError as e_cp874:
                    print(f"cp874 also failed: {e_cp874}.", flush=True)
                    raise e_cp874

            if df_codes is None or meta_codes is None:
                raise ValueError("Failed to read data or metadata from SPSS file with codes.")

            # Filter columns: Keep only '_O1' for variables ending with '_O'
            columns_to_keep = []
            for col in df_codes.columns:
                if col.endswith('_O'):
                    if col.endswith('_O1'):
                        columns_to_keep.append(col)
                else:
                    columns_to_keep.append(col)

            # Apply the filter to the dataframe
            df_codes = df_codes[columns_to_keep]

            # Export to Excel
            print(f"Exporting {len(df_codes)} rows and {len(df_codes.columns)} columns to Excel...", flush=True)
            
            with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
                # Export Rawdata_Code sheet (without value labels)
                df_codes.to_excel(writer, sheet_name='Rawdata_Code', index=False)

            print(f"Rawdata exported successfully to: {excel_output_path}", flush=True)
            self.status_label.config(text=f"Export Rawdata สำเร็จ: {os.path.basename(excel_output_path)}", fg="green")
            messagebox.showinfo("สำเร็จ", f"Export ข้อมูลดิบเป็น Excel สำเร็จ!\nไฟล์: {excel_output_path}", parent=self.root)

        except FileNotFoundError:
            print(f"ERROR: SPSS file not found: {spss_path}", flush=True)
            messagebox.showerror("ข้อผิดพลาด", f"ไม่พบไฟล์ SPSS:\n{spss_path}", parent=self.root)
            self.status_label.config(text="เกิดข้อผิดพลาด: ไม่พบไฟล์ SPSS", fg="red")
        except PermissionError:
            print(f"ERROR: Permission denied saving to {excel_output_path}", flush=True)
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถบันทึกไฟล์ได้:\n{excel_output_path}\nอาจจะเปิดไฟล์นี้อยู่ หรือไม่มีสิทธิ์เขียน", parent=self.root)
            self.status_label.config(text="เกิดข้อผิดพลาด: ไม่มีสิทธิ์บันทึกไฟล์", fg="red")
        except Exception as e:
            print(f"ERROR during export: {e}", flush=True)
            traceback.print_exc()
            messagebox.showerror("ข้อผิดพลาด", f"เกิดข้อผิดพลาดในการ Export:\n{e}", parent=self.root)
            self.status_label.config(text="เกิดข้อผิดพลาดในการ Export Rawdata", fg="red")
        finally:
            print("--- export_rawdata_excel สำเร็จ  ---", flush=True)



    def convert_file(self):
        """
        Handles file conversion with a corrected, robust, two-phase logic.
        This FINAL version standardizes value label writing for SA/MA loops
        to ALWAYS start from code 1 up to the maximum code found,
        filling any missing preceding codes with empty labels. This applies
        to ALL loop types (Consolidated, Unconsolidated, and Auto-Detect).
        """
        print("--- Starting convert_file (FINAL - Labels Always Start from 1) ---", flush=True)
        spss_path = self.spss_file_path.get()
        excel_template_path = self.excel_template_path.get()
        excel_output_path = self.excel_output_path.get()

        # --- 1. Input Validation ---
        if not spss_path or not excel_template_path or not excel_output_path: messagebox.showwarning("Warning", "กรุณาเลือกไฟล์ให้ครบถ้วน"); return
        if not os.path.exists(excel_template_path): messagebox.showerror("ข้อผิดพลาด", f"ไม่พบไฟล์ Excel Template:\n{excel_template_path}"); self.status_label.config(text="เกิดข้อผิดพลาด: ไม่พบ Template", fg="red"); return
        if not excel_output_path.lower().endswith(".xlsx"): messagebox.showerror("ข้อผิดพลาด", "ไฟล์ Output ต้องเป็นนามสกุล .xlsx เท่านั้น"); self.status_label.config(text="เกิดข้อผิดพลาด: Output .xlsx", fg="red"); return
        print("Input validation passed.", flush=True)

        # --- 2. Update Status & Initialize ---
        self.status_label.config(text="กำลังประมวลผล...", fg="orange"); self.root.update_idletasks()
        processed_ma_base_names_auto = set()

        try:
            # --- 3. อ่าน SPSS Metadata ---
            meta = self.last_read_meta
            if meta is None:
                print("Metadata not pre-read, reading again...", flush=True)
                try: _df, meta = pyreadstat.read_sav(spss_path, metadataonly=True, encoding='utf-8')
                except pyreadstat.ReadstatError: _df, meta = pyreadstat.read_sav(spss_path, metadataonly=True, encoding='cp874')
                self.last_read_meta = meta
            if meta is None: raise ValueError("ไม่สามารถอ่าน Metadata จากไฟล์ SPSS ได้")

            req_attrs=['column_names', 'column_labels']; miss=[a for a in req_attrs if not hasattr(meta, a) or not getattr(meta, a)]
            if miss: raise AttributeError(f"Meta ขาด Attribute ที่จำเป็น: {', '.join(miss)}")
            if not hasattr(meta, 'variable_value_labels'): meta.variable_value_labels={}
            
            # --- กรอง Value Labels ---
            filtered_meta_value_labels = meta.variable_value_labels.copy() if meta.variable_value_labels else {}
            if self.codes_to_delete_confirmed:
                print(f"Filtering {len(self.codes_to_delete_confirmed)} codes...", flush=True)
                temp_filtered={}; count=0
                for var, labels in filtered_meta_value_labels.items():
                    if not isinstance(labels, dict): temp_filtered[var]=labels; continue
                    new_labels={}; remove_codes={cd for v, cd in self.codes_to_delete_confirmed if v==var}
                    for code, label in labels.items():
                        if str(code) not in remove_codes: new_labels[code]=label
                        else: count+=1
                    temp_filtered[var]=new_labels
                filtered_meta_value_labels=temp_filtered; print(f"{count} codes filtered.", flush=True)

            # ================================================================= #
            #     PHASE 1: PRE-COMPUTATION & PREPARATION OF ALL LOOP GROUPS     #
            # ================================================================= #
            print("\n--- PHASE 1: Pre-computing all loop structures ---", flush=True)
            
            all_vars = meta.column_names
            var_name_to_index = {name: i for i, name in enumerate(all_vars)}

            # --- 1.1: ระบุกลุ่ม Loop เบื้องต้นตามที่ผู้ใช้กำหนด ---
            print("Step 1.1: Identifying initial user-defined loop groups...", flush=True)
            initial_user_loop_groups={}; processed_indices=set(); i=0
            while i < len(all_vars):
                if i in processed_indices: i+=1; continue
                current_var=all_vars[i]; current_type=self.variable_loop_types.get(current_var, "")
                if current_type in ["SA", "MA", "Loop Text", "Loop Numeric"]:
                    group=[current_var]; base=get_base_name_heuristic(current_var); prefix_match=re.match(r'(I_\d+_)', current_var); prefix=prefix_match.group(1) if prefix_match else None; j=i+1
                    while j < len(all_vars):
                        next_var=all_vars[j]; next_type=self.variable_loop_types.get(next_var, ""); next_base=get_base_name_heuristic(next_var); stop=(next_type != current_type) or (next_base != base)
                        if not stop and (prefix or re.match(r'I_\d+_(.+?)(_O\d+)?$', current_var)):
                            next_prefix_match=re.match(r'(I_\d+_)', next_var); next_prefix=next_prefix_match.group(1) if next_prefix_match else None
                            if prefix and next_prefix and prefix != next_prefix: stop=True
                        if not stop: group.append(next_var); j+=1
                        else: break
                    first_var = group[0]
                    initial_user_loop_groups[first_var]={'base_name': base, 'type': current_type, 'vars': group}
                    for k in range(i, j): processed_indices.add(k)
                    i=j
                else:
                    i+=1
            print(f"Found {len(initial_user_loop_groups)} initial user groups.", flush=True)

            # --- 1.2: รวมกลุ่ม (Consolidate) Loop ที่มี ID เดียวกัน ---
            print("Step 1.2: Consolidating groups with the same final Loop ID...", flush=True)
            consolidated_loops={}; groups_by_id=defaultdict(list); initial_loops_in_consolidation=set()
            
            # Group initial loops by their final ID
            for first_var, info in initial_user_loop_groups.items():
                final_id=self.user_defined_loop_names.get(first_var, info['base_name'])
                if not final_id: continue
                groups_by_id[final_id].append({'first_var': first_var, 'info': info})

            i_pat = re.compile(r'^(I_\d+_)(.+?)(_O\d+)?$')
            for final_id, groups_with_same_id in groups_by_id.items():
                if len(groups_with_same_id) <= 1: continue # Only consolidate if there's more than one group for an ID

                can_consolidate=True; is_o_suffix_present=False; expected_base=None; representative_vars_map={}; all_vars_in_consol_group=set(); all_types_in_group=set()
                first_ever_data = None
                
                for data in groups_with_same_id:
                    if first_ever_data is None: first_ever_data = data
                    first_var=data['first_var']; info=data['info']; all_types_in_group.add(info['type'])
                    match=i_pat.match(first_var)
                    if not match: can_consolidate=False; break
                    
                    prefix=match.group(1); base=match.group(2); o_suf=match.group(3)
                    if expected_base is None: expected_base=base
                    elif base != expected_base: can_consolidate=False; break
                    if o_suf: is_o_suffix_present=True
                    
                    rep_var = first_var # Default representative
                    if o_suf:
                        # Find the _O1 as representative if possible
                        for v in sorted(info['vars']):
                            if v.startswith(prefix) and v.endswith('_O1'):
                                rep_var=v; break
                    
                    if prefix not in representative_vars_map:
                        representative_vars_map[prefix] = rep_var
                        all_vars_in_consol_group.update(info['vars'])
                    else: # This case is unlikely with the current grouping logic but is a safeguard
                        can_consolidate = False; break

                if can_consolidate and representative_vars_map and first_ever_data:
                    # Determine final loop type
                    if is_o_suffix_present or "MA" in all_types_in_group: final_type="MA"
                    elif "Loop Text" in all_types_in_group: final_type="Loop Text"
                    elif "Loop Numeric" in all_types_in_group: final_type="Loop Numeric"
                    else: final_type="SA"
                    
                    print(f"  > Consolidating ID '{final_id}' as Type '{final_type}'", flush=True)
                    
                    sorted_reps=sorted(representative_vars_map.values(), key=lambda v: int(re.match(r'I_(\d+)_', v).group(1)))
                    first_representative_var = sorted_reps[0] # The very first var to trigger writing

                    consolidated_loops[final_id] = {
                        'type': final_type, 
                        'representative_vars': sorted_reps, 
                        'label_source_var': first_ever_data['first_var'], 
                        'all_original_vars': sorted(list(all_vars_in_consol_group), key=lambda x: var_name_to_index.get(x, float('inf'))),
                        'first_representative_var': first_representative_var
                    }
                    
                    for data in groups_with_same_id:
                        initial_loops_in_consolidation.add(data['first_var'])
            print(f"Found {len(consolidated_loops)} consolidated loop groups.", flush=True)


            # ================================================================= #
            #            PHASE 2: MAIN WRITING LOOP USING PRE-COMPUTED DATA     #
            # ================================================================= #
            print("\n--- PHASE 2: Writing to Excel based on pre-computed structures ---", flush=True)

            # --- 4. Open Template & Define Columns ---
            try: wb=openpyxl.load_workbook(excel_template_path); ws=wb.active
            except Exception as e: raise IOError(f"เปิดไฟล์ Excel Template ไม่ได้: {e}")

            item_c='A'; fmt_c='B'; code_c='C'; type_c='D'; disp_c='E'; loopsub_c='F'; id_c='G'; label_c='H'; val_id_c='G'; val_lbl_c='H'; valid_c='I'; cat_c='J'; digit_c='K'; min_c='L'; max_c='M'; dec_c='N'; stat_c='O'
            try: i_idx=column_index_from_string(item_c); fmt_idx=column_index_from_string(fmt_c); code_idx=column_index_from_string(code_c); type_idx=column_index_from_string(type_c); disp_idx=column_index_from_string(disp_c); loopsub_idx=column_index_from_string(loopsub_c); id_idx=column_index_from_string(id_c); label_idx=column_index_from_string(label_c); val_id_idx=column_index_from_string(val_id_c); val_lbl_idx=column_index_from_string(val_lbl_c); valid_idx=column_index_from_string(valid_c); cat_idx=column_index_from_string(cat_c); digit_idx=column_index_from_string(digit_c); min_idx=column_index_from_string(min_c); max_idx=column_index_from_string(max_c); dec_idx=column_index_from_string(dec_c); stat_idx=column_index_from_string(stat_c)
            except ValueError as e: raise ValueError(f"ชื่อคอลัมน์ใน Template ไม่ถูกต้อง: {e}")

            # --- 6. Prepare Writing ---
            start_row=3; write_row=start_row
            written_vars = set() # Keeps track of all vars that have been handled

            has_types=hasattr(meta, 'variable_types') and isinstance(meta.variable_types, dict); has_orig_types=hasattr(meta, 'original_variable_types') and isinstance(meta.original_variable_types, dict)

            # --- 7. Main Loop ---
            for index, var_name in enumerate(all_vars):
                if var_name in written_vars:
                    continue

                print(f"[{index+1}/{len(all_vars)}] Processing '{var_name}'...", end="")
                
                # --- 7.A Check if it's the start of a CONSOLIDATED loop ---
                final_loop_id_to_write = None
                loop_data_to_write = None

                for final_id, loop_data in consolidated_loops.items():
                    if var_name == loop_data['first_representative_var']:
                        final_loop_id_to_write = final_id
                        loop_data_to_write = loop_data
                        break
                
                if loop_data_to_write:
                    print(f" | Start CONSOLIDATED Loop '{final_loop_id_to_write}'")
                    out_type=loop_data_to_write['type']; reps=loop_data_to_write['representative_vars']; lbl_src=loop_data_to_write['label_source_var']; all_loop_vars=loop_data_to_write['all_original_vars']
                    
                    # --- Write Header Row ---
                    ws.cell(row=write_row, column=i_idx).value="Item"; ws.cell(row=write_row, column=fmt_idx).value="Survey"
                    if out_type=="Loop Text": ws.cell(row=write_row, column=type_idx).value="Loop(Text)"; ws.cell(row=write_row, column=digit_idx).value=4000
                    elif out_type=="Loop Numeric": 
                        ws.cell(row=write_row, column=type_idx).value="Loop(Numeric)"; ws.cell(row=write_row, column=digit_idx).value=11
                        ws.cell(row=write_row, column=min_idx).value=-9999999999; ws.cell(row=write_row, column=max_idx).value=9999999999
                        ws.cell(row=write_row, column=dec_idx).value=0; ws.cell(row=write_row, column=stat_idx).value=None
                    elif out_type=="SA": ws.cell(row=write_row, column=type_idx).value="Loop(SA)"
                    elif out_type=="MA": ws.cell(row=write_row, column=type_idx).value="Loop(MA)"
                    ws.cell(row=write_row, column=disp_idx).value="O"; ws.cell(row=write_row, column=id_idx).value=final_loop_id_to_write; ws.cell(row=write_row, column=label_idx).value=None; write_row+=1

                    # --- Write Sub-Items ---
                    print(f"   Writing {len(reps)} consolidated sub-items...", flush=True)
                    for i_sub, rep_var_name in enumerate(reps, 1):
                        sub_var_id=f"{final_loop_id_to_write}({i_sub})"
                        rep_var_index = var_name_to_index.get(rep_var_name, -1)
                        sub_label_value = meta.column_labels[rep_var_index] if rep_var_index != -1 else ""
                        ws.cell(row=write_row, column=loopsub_idx).value="Loop sub"; ws.cell(row=write_row, column=id_idx).value=sub_var_id; ws.cell(row=write_row, column=label_idx).value=sub_label_value
                        ws.cell(row=write_row, column=disp_idx).value="O"; write_row+=1
                    
                    # --- Write Value Labels ---
                    if out_type in ["SA", "MA"]:
                        loop_value_labels=filtered_meta_value_labels.get(lbl_src)
                        if loop_value_labels and isinstance(loop_value_labels, dict):
                            print(f"   Writing labels for CONSOLIDATED loop '{final_loop_id_to_write}'...", flush=True)
                            sorted_codes = sorted([k for k in loop_value_labels.keys() if isinstance(k, (int, float))])
                            if sorted_codes:
                                max_c = int(sorted_codes[-1])
                                code_map = {int(k): v for k, v in loop_value_labels.items() if isinstance(k, (int, float))}
                                print(f"   Writing codes from 1 to {max_c}, filling gaps.", flush=True)
                                for code in range(1, max_c + 1):
                                    ws.cell(row=write_row, column=val_id_idx).value = code
                                    ws.cell(row=write_row, column=val_lbl_idx).value = code_map.get(code, "")
                                    ws.cell(row=write_row, column=valid_idx).value = "Valid"; write_row += 1
                    
                    written_vars.update(all_loop_vars)
                    continue

                # --- 7.B Check if it's the start of an UNCONSOLIDATED user loop ---
                if var_name in initial_user_loop_groups and var_name not in initial_loops_in_consolidation:
                    print(f" | Start UNCONSOLIDATED Loop '{var_name}'")
                    group_info=initial_user_loop_groups[var_name]; g_type=group_info['type']; g_vars=group_info['vars']; disp_id=self.user_defined_loop_names.get(var_name, group_info['base_name']);
                    is_o_suf=any(re.search(r'_O\d+$', v) for v in g_vars); out_type="MA" if is_o_suf else g_type
                    
                    # --- Write Header Row ---
                    ws.cell(row=write_row, column=i_idx).value="Item"; ws.cell(row=write_row, column=fmt_idx).value="Survey"
                    if out_type=="Loop Text": ws.cell(row=write_row, column=type_idx).value="Loop(Text)"; ws.cell(row=write_row, column=digit_idx).value=4000
                    elif out_type=="Loop Numeric":
                        ws.cell(row=write_row, column=type_idx).value="Loop(Numeric)"; ws.cell(row=write_row, column=digit_idx).value=11
                        ws.cell(row=write_row, column=min_idx).value=-9999999999; ws.cell(row=write_row, column=max_idx).value=9999999999
                        ws.cell(row=write_row, column=dec_idx).value=0; ws.cell(row=write_row, column=stat_idx).value=None
                    elif out_type=="SA": ws.cell(row=write_row, column=type_idx).value="Loop(SA)"
                    elif out_type=="MA": ws.cell(row=write_row, column=type_idx).value="Loop(MA)"
                    ws.cell(row=write_row, column=disp_idx).value="O"; ws.cell(row=write_row, column=id_idx).value=disp_id; write_row+=1

                    # --- Write Sub-Items ---
                    pnon_pattern = re.compile(r'(.+?)_(\d+)_O(\d+)$'); is_pnon_loop = any(pnon_pattern.match(v) for v in g_vars)
                    sub_items_to_write = {}
                    if is_pnon_loop and out_type == "MA":
                        print(f"   Applying special sub-item logic for '{disp_id}'...", flush=True)
                        sub_groups = defaultdict(list)
                        for sub_var_name in g_vars:
                            match = pnon_pattern.match(sub_var_name)
                            if match: sub_groups[int(match.group(2))].append((int(match.group(3)), sub_var_name))
                        for inter_num, o_list in sorted(sub_groups.items()):
                            if o_list: sub_items_to_write[inter_num] = sorted(o_list)[0][1]
                    else:
                        print(f"   Writing {len(g_vars)} default sub-items...", flush=True)
                        for i_sub, sub_var_name in enumerate(g_vars, 1): sub_items_to_write[i_sub] = sub_var_name

                    for i_sub, key in enumerate(sorted(sub_items_to_write.keys()), 1):
                        rep_var_name = sub_items_to_write[key]
                        sub_var_id = f"{disp_id}({i_sub})"
                        sub_var_index = var_name_to_index.get(rep_var_name, -1)
                        sub_label_value = meta.column_labels[sub_var_index] if sub_var_index != -1 else ""
                        ws.cell(row=write_row, column=loopsub_idx).value="Loop sub"; ws.cell(row=write_row, column=id_idx).value=sub_var_id; ws.cell(row=write_row, column=label_idx).value=sub_label_value
                        ws.cell(row=write_row, column=disp_idx).value="O"; write_row+=1

                    # --- Write Value Labels ---
                    if out_type in ["SA", "MA"]:
                        loop_value_labels=filtered_meta_value_labels.get(var_name)
                        if loop_value_labels and isinstance(loop_value_labels, dict):
                            print(f"   Writing labels for UNCONSOLIDATED loop '{disp_id}'...", flush=True)
                            sorted_codes = sorted([k for k in loop_value_labels.keys() if isinstance(k, (int, float))])
                            if sorted_codes:
                                max_c = int(sorted_codes[-1])
                                code_map = {int(k): v for k, v in loop_value_labels.items() if isinstance(k, (int, float))}
                                print(f"   Writing codes from 1 to {max_c}, filling gaps.", flush=True)
                                for code in range(1, max_c + 1):
                                    ws.cell(row=write_row, column=val_id_idx).value = code
                                    ws.cell(row=write_row, column=val_lbl_idx).value = code_map.get(code, "")
                                    ws.cell(row=write_row, column=valid_idx).value = "Valid"; write_row += 1
                    
                    written_vars.update(g_vars)
                    continue

                # --- 7.C Non-Loop / Auto-Detect ---
                print(f" | Treating as Non-Loop/Auto-Detect")
                var_label=meta.column_labels[index]
                current_value_labels=filtered_meta_value_labels.get(var_name)

                col_d_value=None; disp_name=var_name; auto_base=None; write_var=True
                type_code = meta.variable_types.get(var_name, -1) if has_types else -1
                orig_fmt = meta.original_variable_types.get(var_name, '') if has_orig_types else ''
                is_str = type_code > 0 or (isinstance(orig_fmt, str) and orig_fmt.strip().upper().startswith('A'))

                if is_str:
                    col_d_value="Text"
                else:
                    ma_match=self.ma_pattern.match(var_name)
                    if ma_match:
                        auto_base=ma_match.group(1)
                        if auto_base in processed_ma_base_names_auto: write_var=False
                        else: col_d_value="MA"; disp_name=auto_base; processed_ma_base_names_auto.add(auto_base)
                    elif current_value_labels: col_d_value="SA"
                    else: col_d_value="Numeric"
                
                if write_var:
                    ws.cell(row=write_row, column=i_idx).value="Item"; ws.cell(row=write_row, column=fmt_idx).value="Survey"; ws.cell(row=write_row, column=type_idx).value=col_d_value; ws.cell(row=write_row, column=disp_idx).value="O"; ws.cell(row=write_row, column=id_idx).value=disp_name; ws.cell(row=write_row, column=label_idx).value=var_label
                    if col_d_value=="Text": ws.cell(row=write_row, column=digit_idx).value=4000
                    elif col_d_value=="Numeric":
                        ws.cell(row=write_row, column=digit_idx).value=11; ws.cell(row=write_row, column=min_idx).value=-9999999999
                        ws.cell(row=write_row, column=max_idx).value=9999999999; ws.cell(row=write_row, column=dec_idx).value=0
                    write_row+=1
                    
                    if col_d_value in ["SA", "MA"] and current_value_labels:
                        print(f"   Writing labels for Auto-Detect '{disp_name}'...", flush=True)
                        sorted_codes = sorted([k for k in current_value_labels.keys() if isinstance(k, (int, float))])
                        if sorted_codes:
                            max_c = int(sorted_codes[-1])
                            code_map = {int(k): v for k, v in current_value_labels.items() if isinstance(k, (int,float))}
                            print(f"   Writing codes from 1 to {max_c}, filling gaps.", flush=True)
                            for code in range(1, max_c + 1):
                                ws.cell(row=write_row, column=val_id_idx).value = code
                                ws.cell(row=write_row, column=val_lbl_idx).value = code_map.get(code, "")
                                ws.cell(row=write_row, column=valid_idx).value = "Valid"; write_row += 1
                
                written_vars.add(var_name)

            # --- End of Main Writing Loop ---
            print("\n--- เพิ่ม 'End' marker ---", flush=True)
            ws.cell(row=write_row, column=i_idx).value = "End"
            for col_idx in range(fmt_idx, stat_idx + 2): # Clear other cells
                ws.cell(row=write_row, column=col_idx).value = None

            # --- 8. Save Excel Output ---
            print(f"Saving Excel file to: {excel_output_path}", flush=True)
            wb.save(excel_output_path)
            self.status_label.config(text=f"สำเร็จ! บันทึกที่: {os.path.basename(excel_output_path)}", fg="green")
            messagebox.showinfo("สำเร็จ", f"สำเร็จ!\nบันทึกที่:\n{excel_output_path}")

        except (IOError, ValueError, AttributeError) as e:
            print(f"ERROR: {e}", flush=True); traceback.print_exc()
            self.status_label.config(text=f"ผิดพลาด: {e}", fg="red"); messagebox.showerror("ผิดพลาด", f"เกิดข้อผิดพลาดในการประมวลผล:\n{e}")
        except Exception as e:
            print(f"ERROR: Unexpected error: {e}", flush=True); traceback.print_exc()
            self.status_label.config(text=f"ผิดพลาดไม่คาดคิด: {e}", fg="red"); messagebox.showerror("ผิดพลาด", f"ผิดพลาดไม่คาดคิด:\n{e}")
        finally:
            print("--- สิ้นสุดการทำงานของ convert_file ---", flush=True)

# <<< START OF CHANGES >>>
# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
# --- ฟังก์ชัน Entry Point ---
# ใน run_this_app() ของ Program_ItemdefSPSS_Log.py

# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
def run_this_app(working_dir=None):
    # --- เก็บ stdout ดั้งเดิมไว้ตั้งแต่เริ่ม ---
    original_stdout = sys.stdout
    # original_stderr = sys.stderr # ถ้าคุณ redirect stderr ด้วย

    # --- พิมพ์ Log เริ่มต้นไปที่ original_stdout ---
    print(f"--- SPSS_LOG_INFO: Starting 'Program_ItemdefSPSS_Log' via run_this_app() ---", file=original_stdout)

    root = None
    app_instance = None # หรือ spss_converter_app
    # redirector_from_app = None # ไม่จำเป็นต้องใช้ตัวนี้แล้วถ้า App จัดการเอง

    try:
        root = tk.Tk()
        # --- Set Icon ---
        try:
            icon_filename = "Peth.ico" # ไอคอนสำหรับโปรแกรมนี้
            icon_path = resource_path(icon_filename)
            if os.path.exists(icon_path):
                root.iconbitmap(icon_path)
                print(f"SPSS_LOG_INFO: App icon '{icon_filename}' loaded.", file=original_stdout)
            else:
                print(f"SPSS_LOG_WARNING: App icon '{icon_filename}' not found at {icon_path}", file=original_stdout)
        except Exception as e:
            print(f"SPSS_LOG_WARNING: Could not set application icon: {e}", file=original_stdout)

        # --- จัดหน้าต่างกลางจอ ---
        window_width = 880
        window_height = 500 # ตรวจสอบว่าตรงกับใน Class
        root.update_idletasks() # ให้ Tkinter คำนวณขนาดจริงก่อน
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        center_x = int(screen_width/2 - window_width / 2)
        center_y = int(screen_height/2 - window_height / 2)
        root.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
        print(f"SPSS_LOG_INFO: Window geometry set.", file=original_stdout)

        # สร้าง Instance ของ App (ซึ่งใน __init__ ของมันจะ redirect sys.stdout)
        app_instance = SpssToExcelConverter(root)
        # ณ จุดนี้ sys.stdout ถูก redirect ไปยัง app_instance.log_text แล้ว
        # print("SPSS_LOG_INFO: SpssToExcelConverter initialized. stdout is now redirected to GUI log.") # อันนี้จะไปที่ GUI Log

        root.mainloop() # เริ่มการทำงานของหน้าต่าง GUI

        # --- เมื่อ mainloop จบ (ผู้ใช้ปิดหน้าต่าง) ---
        # ณ จุดนี้ self.log_text ใน app_instance อาจจะถูกทำลายไปพร้อมกับ root แล้ว
        # การ print() หลังจากนี้โดยที่ sys.stdout ยังเป็น redirector เก่า จะทำให้เกิด Error

    except Exception as e:
        # ถ้าเกิด Error ระหว่างการสร้าง App หรือระหว่าง mainloop
        # พยายาม print ไปที่ original_stdout ถ้าเป็นไปได้
        print(f"\nSPSS_LOG_ERROR: An error occurred during application execution:", file=original_stdout)
        print(f"Error Type: {type(e).__name__}", file=original_stdout)
        print(f"Error Message: {e}", file=original_stdout)
        print(f"Traceback:\n{traceback.format_exc()}", file=original_stdout)

        # แสดง Popup (สร้าง root ชั่วคราวถ้า root หลักอาจจะยังไม่มีหรือถูกทำลาย)
        try:
            if root is None or not root.winfo_exists(): # ตรวจสอบ root ก่อน
                root_temp = tk.Tk()
                root_temp.withdraw()
                messagebox.showerror("Application Error (SPSS Log)",
                                   f"An unexpected error occurred:\n{e}", parent=root_temp)
                root_temp.destroy()
            else: # ถ้า root ยังอยู่ ก็ใช้ root เดิมเป็น parent
                messagebox.showerror("Application Error (SPSS Log)",
                                   f"An unexpected error occurred:\n{e}", parent=root)
        except Exception as popup_err:
             print(f"SPSS_LOG_ERROR: Could not show error popup: {popup_err}", file=original_stdout)
             # ไม่ควร sys.exit จากตรงนี้ ปล่อยให้ Launcher จัดการ
    finally:
        # --- ส่วนนี้สำคัญมาก ---
        # ไม่ว่า try block จะจบลงอย่างไร (สำเร็จ หรือเกิด Exception)
        # เราต้องคืนค่า sys.stdout กลับไปเป็นค่าดั้งเดิม *ก่อน* ที่โปรเซสนี้จะจบลง

        # 1. บอก TextRedirector ภายใน SpssToExcelConverter (ถ้ามี) ว่าให้หยุดทำงาน
        #    (ถ้าคุณได้เพิ่มเมธอด destroy() ให้กับ TextRedirector และเมธอด cleanup_redirector ให้กับ SpssToExcelConverter)
        if app_instance and hasattr(app_instance, 'cleanup_redirector') and callable(app_instance.cleanup_redirector):
            print("SPSS_LOG_INFO: Calling app_instance.cleanup_redirector() in finally.", file=original_stdout)
            app_instance.cleanup_redirector() # เมธอดนี้ควรจะคืนค่า sys.stdout ที่ SpssToExcelConverter redirect ไว้
        else:
            # ถ้า SpssToExcelConverter ไม่ได้จัดการการคืนค่า stdout เอง
            # หรือถ้า app_instance สร้างไม่สำเร็จ เราต้องคืนค่า stdout ที่เก็บไว้ตอนต้นเอง
            print(f"SPSS_LOG_INFO: Restoring original stdout directly in finally.", file=original_stdout)
            sys.stdout = original_stdout # คืนค่า stdout หลัก

        # 2. (Optional) ทำลาย root window ถ้ามันยังอยู่ (อาจจะไม่จำเป็นถ้า mainloop จบเพราะผู้ใช้ปิดหน้าต่าง)
        # if root and root.winfo_exists():
        #     print("SPSS_LOG_INFO: Destroying root window in finally (if exists).", file=original_stdout)
        #     root.destroy()

        print(f"--- SPSS_LOG_INFO: run_this_app() finished. stdout restored. ---", file=original_stdout)

# --- ส่วน Run Application เมื่อรันไฟล์นี้โดยตรง (สำหรับ Test) ---
if __name__ == "__main__":
    # --- (Optional) เพิ่ม freeze_support() ถ้าจะทดสอบ multiprocessing จากไฟล์นี้โดยตรง ---
    # from multiprocessing import freeze_support
    # freeze_support()
    print("--- Running Program_ItemdefSPSS_Log.py directly for testing ---")
    run_this_app()
    print("--- Finished direct execution of Program_ItemdefSPSS_Log.py ---")
