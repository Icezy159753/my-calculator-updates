import os
import re
import traceback

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PyQt6.QtCore import QElapsedTimer, QTimer, Qt
from PyQt6.QtGui import QColor, QIcon, QPainter, QPainterPath, QPixmap
from PyQt6.QtWidgets import (
    QApplication,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QStyle,
    QTableWidget,
    QTableWidgetItem,
    QProgressBar,
    QVBoxLayout,
    QWidget,
)


DONE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ADDED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
REMOVED_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")


def make_leaf_icon(size: int = 128) -> QIcon:
    pix = QPixmap(size, size)
    pix.fill(Qt.GlobalColor.transparent)

    p = QPainter(pix)
    p.setRenderHint(QPainter.RenderHint.Antialiasing, True)
    p.setPen(Qt.PenStyle.NoPen)
    p.setBrush(QColor("#e8f5e9"))
    p.drawEllipse(4, 4, size - 8, size - 8)

    path = QPainterPath()
    path.moveTo(size * 0.52, size * 0.18)
    path.cubicTo(size * 0.82, size * 0.26, size * 0.84, size * 0.62, size * 0.54, size * 0.80)
    path.cubicTo(size * 0.30, size * 0.70, size * 0.24, size * 0.40, size * 0.52, size * 0.18)
    p.setBrush(QColor("#43a047"))
    p.drawPath(path)
    p.setPen(QColor("#1b5e20"))
    p.drawLine(int(size * 0.53), int(size * 0.24), int(size * 0.48), int(size * 0.76))
    p.end()
    return QIcon(pix)


def txt(v) -> str:
    if pd.isna(v):
        return ""
    return str(v).strip()


def norm(v) -> str:
    return "".join(ch.lower() for ch in txt(v) if ch.isalnum())


def norm_sheet_name(v) -> str:
    s = txt(v).replace("\u3000", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def is_non_data_sbj(v) -> bool:
    return norm(v) in {"", "sbjnum", "idrds", "na"}


def detect_header_row_df(df: pd.DataFrame, must_keys: set[str], max_scan: int = 60) -> int:
    best_idx, best_score = -1, -1
    for i in range(min(len(df), max_scan)):
        keys = {norm(v) for v in df.iloc[i].tolist() if txt(v) != ""}
        score = len(must_keys.intersection(keys))
        if score > best_score:
            best_idx, best_score = i, score
    if best_idx < 0 or best_score < len(must_keys):
        raise ValueError(f"Header row not found (required: {', '.join(sorted(must_keys))})")
    return best_idx


def build_idx_map_from_header_row(row_values: list) -> dict[str, int]:
    idx = {}
    for c, v in enumerate(row_values):
        k = norm(v)
        if k and k not in idx:
            idx[k] = c
    return idx


def parse_code_cell(v) -> tuple[str | None, str | None]:
    s = txt(v)
    if s == "":
        return None, None

    if isinstance(v, (int, float)) and not pd.isna(v):
        n = float(v)
        if n.is_integer():
            code = str(int(n))
            if code == "0":
                return None, s
            return code, None
        return None, s

    m = re.match(r"^\s*(\d+)(?:\.0+)?\s*$", s)
    if m:
        code = m.group(1)
        if code == "0":
            return None, s
        return code, None
    return None, s


def is_numeric_only_text(v) -> bool:
    s = txt(v)
    if s == "":
        return False
    return re.fullmatch(r"\d+(?:\.\d+)?", s) is not None


def detect_code_block_range(header_vals: list, var_col: int) -> tuple[int, int] | None:
    code_start = None
    for c in range(var_col + 1, len(header_vals)):
        if norm(header_vals[c]) == "code":
            code_start = c
            break
    # If no explicit "Code" header, only use columns between this variable and
    # the next non-empty header. If none in-between, treat as no code block.
    if code_start is None:
        next_header_col = len(header_vals)
        for c in range(var_col + 1, len(header_vals)):
            if norm(header_vals[c]) != "":
                next_header_col = c
                break
        start = var_col + 1
        end = next_header_col - 1
        if end < start:
            return None
        return start, end

    code_end = len(header_vals) - 1
    for c in range(code_start + 1, len(header_vals)):
        k = norm(header_vals[c])
        if k != "" and k != "code":
            code_end = c - 1
            break
    if code_end < code_start:
        return None
    return code_start, code_end


def extract_code_slots_from_row(row_vals: list, code_start: int, code_end: int) -> list[tuple[str, str]]:
    # Return ordered slots so we can highlight the exact deleted position.
    # Each item is ("keep", code) or ("drop", original_text).
    slots: list[tuple[str, str]] = []
    upper = min(code_end, len(row_vals) - 1)
    for c in range(code_start, upper + 1):
        v = row_vals[c]
        code, bad = parse_code_cell(v)
        if code is not None:
            slots.append(("keep", code))
        elif bad is not None:
            slots.append(("drop", bad))
    return slots


def to_ns_base(var_name: str) -> str:
    s = txt(var_name)
    if s == "":
        return "NVAR"
    return f"N{s[0].upper()}{s[1:]}"


def merge_base_name(var_name: str) -> str:
    s = txt(var_name)
    if s == "":
        return s
    m = re.match(r"^(.+?)[_](\d+)$", s)
    if m:
        return m.group(1)
    return s


def parse_var_list(v: str) -> list[str]:
    parts = re.split(r"[,\n|]+", txt(v))
    out = []
    seen = set()
    for p in parts:
        t = txt(p)
        if t == "":
            continue
        k = norm(t)
        if k in seen:
            continue
        seen.add(k)
        out.append(t)
    return out


def load_items_from_settings_file(path: str, selected_sheets: list[str] | None = None) -> tuple[list[dict], list[str]]:
    xls = pd.ExcelFile(path)
    items: list[dict] = []
    seen = set()
    warnings: list[str] = []
    sheets = list(xls.sheet_names)
    if selected_sheets:
        resolved: list[str] = []
        by_exact = {txt(s): s for s in xls.sheet_names}
        by_norm = {norm_sheet_name(s): s for s in xls.sheet_names}
        for wanted in selected_sheets:
            picked = by_exact.get(txt(wanted)) or by_norm.get(norm_sheet_name(wanted))
            if picked and picked not in resolved:
                resolved.append(picked)
            elif txt(wanted):
                warnings.append(f"Settings sheet '{wanted}' not found")
        sheets = resolved

    for sheet in sheets:
        df = pd.read_excel(path, sheet_name=sheet, header=None, dtype=object)
        item_col = None
        merge_col = None
        scan_rows = min(len(df), 40)
        for rr in range(scan_rows):
            row = df.iloc[rr].tolist()
            for cc, v in enumerate(row):
                if norm(v) == "item":
                    item_col = cc
                if norm(v) == "merge":
                    merge_col = cc
            if item_col is not None:
                start_row = rr + 1
                break
        if item_col is None:
            continue

        found_in_sheet = 0
        for rr in range(start_row, len(df)):
            row = df.iloc[rr].tolist()
            if item_col >= len(row):
                continue
            raw = txt(row[item_col])
            if raw == "":
                continue
            if norm(raw) == "item":
                continue
            key = norm(raw)
            if key == "" or key in seen:
                continue
            seen.add(key)
            merge_flag = False
            # Preferred: explicit "Merge" header column.
            # Fallback: next column after Item (common settings layout).
            merge_val = ""
            if merge_col is not None and merge_col < len(row):
                merge_val = txt(row[merge_col]).lower()
            elif (item_col + 1) < len(row):
                merge_val = txt(row[item_col + 1]).lower()
            merge_flag = merge_val in {"merge", "m", "1", "y", "yes", "true"}
            items.append({"name": raw, "merge": merge_flag})
            found_in_sheet += 1
        if found_in_sheet == 0:
            warnings.append(f"Sheet '{sheet}' has Item header but no values")

    if not items:
        raise ValueError("No Item values found in settings file")
    return items, warnings


def prepare_text_data(
    text_path: str,
    selected_sheets: list[str] | None = None,
) -> tuple[list[tuple[str, pd.DataFrame]], list[str]]:
    xls = pd.ExcelFile(text_path)
    warnings: list[str] = []
    sheets = xls.sheet_names
    if selected_sheets:
        resolved: list[str] = []
        seen = set()
        by_exact = {txt(s): s for s in xls.sheet_names}
        by_norm: dict[str, str] = {}
        for s in xls.sheet_names:
            k = norm_sheet_name(s)
            if k not in by_norm:
                by_norm[k] = s

        for wanted_raw in selected_sheets:
            wanted_text = txt(wanted_raw)
            if wanted_text == "":
                continue

            picked = by_exact.get(wanted_text)
            if picked is None:
                picked = by_norm.get(norm_sheet_name(wanted_text))
                if picked is not None and picked != wanted_text:
                    warnings.append(
                        f"Sheet '{wanted_text}' resolved as '{picked}' (normalized spaces/case)"
                    )

            if picked is None:
                warnings.append(f"Sheet '{wanted_text}' not found in Text file")
                continue

            key = norm_sheet_name(picked)
            if key in seen:
                continue
            seen.add(key)
            resolved.append(picked)

        sheets = resolved
    frames: list[tuple[str, pd.DataFrame]] = []
    for sheet in sheets:
        df = pd.read_excel(text_path, sheet_name=sheet, header=None, dtype=object)
        frames.append((sheet, df))
    return frames, warnings


def load_verbatim_parts(
    text_data: list[tuple[str, pd.DataFrame]],
    var_name: str,
) -> tuple[dict, list[str], bool]:
    var_key = norm(var_name)
    merged: dict[str, dict] = {}
    warnings: list[str] = []
    found_question = False

    for sheet, df in text_data:
        try:
            header_idx = detect_header_row_df(df, {"sbjnum", var_key})
        except ValueError:
            warnings.append(f"Skip sheet '{sheet}' (missing SbjNum + {var_name})")
            continue

        header_vals = df.iloc[header_idx].tolist()
        idx = build_idx_map_from_header_row(header_vals)
        sbj_col = idx.get("sbjnum")
        var_col = idx.get(var_key)
        if sbj_col is None or var_col is None:
            warnings.append(f"Skip sheet '{sheet}' (required columns not complete)")
            continue
        found_question = True

        code_range = detect_code_block_range(header_vals, var_col)

        for r in range(header_idx + 1, len(df)):
            row = df.iloc[r].tolist()
            if sbj_col >= len(row):
                continue
            sbj = txt(row[sbj_col])
            if is_non_data_sbj(sbj):
                continue

            verb = txt(row[var_col]) if var_col < len(row) else ""
            # Verbatim column should be free text. If it is numeric-only, treat as no verbatim.
            if is_numeric_only_text(verb):
                verb = ""
            slots = extract_code_slots_from_row(row, code_range[0], code_range[1]) if code_range is not None else []
            if not verb and not slots:
                continue

            if sbj not in merged:
                merged[sbj] = {"verbatim": "", "slots": []}
            rec = merged[sbj]
            if verb:
                if rec["verbatim"] == "":
                    rec["verbatim"] = verb
                elif rec["verbatim"] != verb:
                    warnings.append(f"SbjNum {sbj}: duplicate verbatim found (keep first value)")
            rec["slots"].extend(slots)

    return merged, warnings, found_question


def build_raw_header_candidates(ws, max_scan: int = 80) -> list[tuple[int, dict[str, int]]]:
    cands: list[tuple[int, dict[str, int]]] = []
    for rr in range(1, min(ws.max_row, max_scan) + 1):
        idx = {}
        for cc in range(1, ws.max_column + 1):
            k = norm(ws.cell(row=rr, column=cc).value)
            if k and k not in idx:
                idx[k] = cc
        cands.append((rr, idx))
    return cands


def detect_header_row_ws(ws, must_keys: set[str], max_scan: int = 80) -> int:
    best_row, best_score = -1, -1
    for rr in range(1, min(ws.max_row, max_scan) + 1):
        keys = {norm(ws.cell(row=rr, column=cc).value) for cc in range(1, ws.max_column + 1)}
        score = len(must_keys.intersection(keys))
        if score > best_score:
            best_row, best_score = rr, score
    if best_row < 0 or best_score < len(must_keys):
        raise ValueError(f"Raw header row not found (required: {', '.join(sorted(must_keys))})")
    return best_row


def ws_header_map(ws, header_row: int) -> dict[str, int]:
    idx = {}
    for cc in range(1, ws.max_column + 1):
        k = norm(ws.cell(row=header_row, column=cc).value)
        if k and k not in idx:
            idx[k] = cc
    return idx


def run_merge(
    raw_path: str,
    raw_sheet: str,
    text_path: str,
    var_configs: list[dict],
    out_path: str,
    text_sheets: list[str] | None = None,
) -> dict:
    wb = load_workbook(raw_path)
    ws = wb[raw_sheet]
    text_data, text_data_warnings = prepare_text_data(text_path, text_sheets)

    all_warnings: list[str] = list(text_data_warnings)
    details: list[dict] = []
    raw_header_candidates = build_raw_header_candidates(ws)
    sbj_row_cache: dict[tuple[int, int], dict[str, int]] = {}

    code_groups: dict[str, dict] = {}
    group_order: list[str] = []

    for cfg in var_configs:
        var_name = txt(cfg.get("name"))
        merge_flag = bool(cfg.get("merge", False))
        if var_name == "":
            continue
        merged, warnings, found_question = load_verbatim_parts(text_data, var_name)
        all_warnings.extend([f"[{var_name}] {w}" for w in warnings])
        if not merged and not found_question:
            all_warnings.append(f"[{var_name}] no mergeable data found")
            details.append(
                {
                    "var": var_name,
                    "status": "FAILED",
                    "note": "No data in Text",
                    "matched_rows": 0,
                    "touched_verbatim": 0,
                    "max_codes": 0,
                    "removed_hits": 0,
                }
            )
            continue

        var_key = norm(var_name)
        header_row = None
        idx = None
        for rr, m in raw_header_candidates:
            if "sbjnum" in m and var_key in m:
                header_row = rr
                idx = m
                break
        if header_row is None or idx is None:
            all_warnings.append(f"[{var_name}] header/column not found in Raw")
            details.append(
                {
                    "var": var_name,
                    "status": "FAILED",
                    "note": "Header/column not found in Raw",
                    "matched_rows": 0,
                    "touched_verbatim": 0,
                    "max_codes": 0,
                    "removed_hits": 0,
                }
            )
            continue
        sbj_col = idx.get("sbjnum")
        var_col = idx.get(var_key)
        if sbj_col is None or var_col is None:
            all_warnings.append(f"[{var_name}] column SbjNum or {var_name} not found in Raw")
            details.append(
                {
                    "var": var_name,
                    "status": "FAILED",
                    "note": "Column not found in Raw",
                    "matched_rows": 0,
                    "touched_verbatim": 0,
                    "max_codes": 0,
                    "removed_hits": 0,
                }
            )
            continue

        matched_rows = 0
        touched_verbatim = 0
        removed_hits = 0

        cache_key = (header_row, sbj_col)
        if cache_key not in sbj_row_cache:
            sbj_row_map: dict[str, int] = {}
            for rr in range(header_row + 1, ws.max_row + 1):
                sbj = txt(ws.cell(row=rr, column=sbj_col).value)
                if is_non_data_sbj(sbj):
                    continue
                sbj_row_map[sbj] = rr
            sbj_row_cache[cache_key] = sbj_row_map
        sbj_row_map = sbj_row_cache[cache_key]

        for sbj, rec in merged.items():
            rr = sbj_row_map.get(sbj)
            if rr is None:
                continue
            matched_rows += 1
            if rec["verbatim"] != "":
                ws.cell(row=rr, column=var_col, value=rec["verbatim"]).fill = DONE_FILL
                touched_verbatim += 1

        max_codes = max((len(v["slots"]) for v in merged.values()), default=0)
        target_base = merge_base_name(var_name) if merge_flag else var_name
        ns_base = to_ns_base(target_base)

        if ns_base not in code_groups:
            code_groups[ns_base] = {"header_row": header_row, "rows": {}, "min_cols": 0}
            group_order.append(ns_base)
        grp = code_groups[ns_base]
        if found_question and max_codes <= 0:
            grp["min_cols"] = max(int(grp.get("min_cols", 0)), 1)
        row_slots_map: dict[int, list[tuple[str, str]]] = grp["rows"]

        for sbj, rr in sbj_row_map.items():
            rec = merged.get(sbj)
            if rec is None:
                continue
            row_has_removed = False
            if rr not in row_slots_map:
                row_slots_map[rr] = []
            row_slots_map[rr].extend(rec["slots"])
            for slot in rec["slots"]:
                if slot[0] == "drop":
                    row_has_removed = True
            if row_has_removed:
                removed_hits += 1

        details.append(
            {
                "var": var_name,
                "status": "SUCCESS",
                "note": "Completed" if matched_rows > 0 else ("Question found; no code/verbatim data" if found_question else "No matched SbjNum"),
                "matched_rows": matched_rows,
                "touched_verbatim": touched_verbatim,
                "max_codes": max(max_codes, 1 if found_question and max_codes <= 0 else 0),
                "removed_hits": removed_hits,
            }
        )

    for ns_base in group_order:
        grp = code_groups[ns_base]
        header_row = int(grp["header_row"])
        row_slots_map: dict[int, list[tuple[str, str]]] = grp["rows"]
        min_cols = int(grp.get("min_cols", 0))
        max_codes = max(max((len(v) for v in row_slots_map.values()), default=0), min_cols)
        if max_codes <= 0:
            continue
        add_start_col = ws.max_column + 1
        for i in range(max_codes):
            ws.cell(row=header_row, column=add_start_col + i, value=f"{ns_base}_O{i+1}")
        for rr, slots in row_slots_map.items():
            for i, slot in enumerate(slots):
                kind, value = slot
                cell = ws.cell(row=rr, column=add_start_col + i)
                if kind == "keep":
                    cell.value = value
                    cell.fill = ADDED_FILL
                else:
                    cell.value = ""
                    cell.fill = REMOVED_FILL

    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(
        [
            "Item",
            "Merge",
            "Output Base",
            "Status",
            "Note",
            "Matched Rows",
            "Overwritten Verbatim",
            "New Code Cols",
            "Removed-Highlight Rows",
        ]
    )
    dmap = {d["var"]: d for d in details}
    for cfg in var_configs:
        var_name = txt(cfg.get("name"))
        merge_flag = bool(cfg.get("merge", False))
        target_base = merge_base_name(var_name) if merge_flag else var_name
        d = dmap.get(var_name, {})
        ws_sum.append(
            [
                var_name,
                "MERGE" if merge_flag else "",
                target_base,
                d.get("status", "FAILED"),
                d.get("note", "Unknown"),
                d.get("matched_rows", 0),
                d.get("touched_verbatim", 0),
                d.get("max_codes", 0),
                d.get("removed_hits", 0),
            ]
        )

    wb.save(out_path)
    return {
        "details": details,
        "warnings": all_warnings,
        "sheet": raw_sheet,
    }


class VerbatimMapperApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("VERBATIM + CODE Mapper")
        self.resize(1080, 620)
        self.setWindowIcon(make_leaf_icon())

        self.raw_path = ""
        self.text_path = ""
        self.text_sheet_names: list[str] = []
        self.selected_text_sheets: list[str] = []
        self.settings_sheet_names: list[str] = []
        self.selected_settings_sheets: list[str] = []
        self.var_merge_map: dict[str, bool] = {}
        self.settings_items: list[dict] = []
        self.settings_path = ""
        self._build_ui()

    def _build_ui(self):
        self.setStyleSheet(
            """
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #0b132b, stop:0.45 #10213f, stop:1 #0a1a34);
            }
            QWidget {
                font-family: 'Segoe UI';
                font-size: 10pt;
                color: #e7eefc;
            }
            QFrame#HeroCard {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #06b6d4, stop:1 #14b8a6);
                border-radius: 16px;
                border: 1px solid rgba(255,255,255,0.28);
            }
            QLabel#Title {
                font-size: 28px;
                font-weight: 700;
                color: #042f3b;
            }
            QLabel#SubTitle {
                color: #0f3e4c;
                font-size: 11px;
            }
            QFrame#Card {
                background: rgba(10, 17, 33, 0.82);
                border: 1px solid rgba(255,255,255,0.16);
                border-radius: 14px;
            }
            QLabel#FieldLabel {
                color: #cfe5ff;
                font-weight: 600;
                min-width: 120px;
            }
            QLineEdit, QComboBox {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.22);
                border-radius: 9px;
                padding: 7px 10px;
                color: #f2f7ff;
            }
            QComboBox QAbstractItemView {
                background: #0f1b33;
                color: #eaf2ff;
                border: 1px solid rgba(255,255,255,0.22);
                selection-background-color: #2563eb;
                selection-color: #ffffff;
                outline: 0;
            }
            QLineEdit:focus, QComboBox:focus {
                border: 1px solid #38bdf8;
                background: rgba(56,189,248,0.10);
            }
            QPushButton {
                border: none;
                border-radius: 10px;
                padding: 8px 14px;
                font-weight: 700;
                color: #052234;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #67e8f9, stop:1 #2dd4bf);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #a5f3fc, stop:1 #5eead4);
            }
            QPushButton#Secondary {
                color: #dbecff;
                border: 1px solid rgba(255,255,255,0.25);
                background: rgba(255,255,255,0.10);
            }
            QPushButton#Secondary:hover {
                background: rgba(255,255,255,0.18);
            }
            QPushButton#Primary {
                padding: 10px 16px;
                min-height: 40px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #38bdf8, stop:1 #2dd4bf);
            }
            QLabel#StatusPill {
                background: rgba(56,189,248,0.16);
                border: 1px solid rgba(56,189,248,0.45);
                border-radius: 11px;
                color: #cffafe;
                font-weight: 700;
                padding: 7px 12px;
            }
            QProgressBar {
                min-height: 8px;
                max-height: 8px;
                border: 1px solid rgba(255,255,255,0.20);
                border-radius: 4px;
                background: rgba(255,255,255,0.08);
                text-align: center;
            }
            QProgressBar::chunk {
                border-radius: 3px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #22d3ee, stop:1 #34d399);
            }
            """
        )

        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(18, 16, 18, 16)
        root.setSpacing(12)

        hero = QFrame(objectName="HeroCard")
        hero_lay = QVBoxLayout(hero)
        hero_lay.setContentsMargins(18, 14, 18, 14)
        hero_lay.setSpacing(3)
        title = QLabel("VERBATIM + CODE Mapper")
        title.setObjectName("Title")
        subtitle = QLabel("Map multi-sheet verbatim/code results into raw workbook with merge control and summary report.")
        subtitle.setObjectName("SubTitle")
        subtitle.setWordWrap(True)
        hero_lay.addWidget(title)
        hero_lay.addWidget(subtitle)
        root.addWidget(hero)

        card = QFrame(objectName="Card")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(14, 14, 14, 14)
        lay.setSpacing(10)

        r1 = QHBoxLayout()
        self.raw_edit = QLineEdit()
        self.raw_edit.setPlaceholderText("Raw file (.xlsx)")
        b_raw = QPushButton("Browse Raw")
        b_raw.setObjectName("Secondary")
        b_raw.clicked.connect(self.pick_raw)
        lb1 = QLabel("Raw File")
        lb1.setObjectName("FieldLabel")
        r1.addWidget(lb1)
        r1.addWidget(self.raw_edit, 1)
        r1.addWidget(b_raw)
        lay.addLayout(r1)

        r2 = QHBoxLayout()
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(260)
        lb2 = QLabel("Raw Sheet")
        lb2.setObjectName("FieldLabel")
        r2.addWidget(lb2)
        r2.addWidget(self.sheet_combo)
        r2.addStretch(1)
        lay.addLayout(r2)

        r3 = QHBoxLayout()
        self.text_edit = QLineEdit()
        self.text_edit.setPlaceholderText("Text file (.xlsx, many sheets)")
        b_text = QPushButton("Browse Text")
        b_text.setObjectName("Secondary")
        b_text.clicked.connect(self.pick_text)
        lb3 = QLabel("Text File")
        lb3.setObjectName("FieldLabel")
        r3.addWidget(lb3)
        r3.addWidget(self.text_edit, 1)
        r3.addWidget(b_text)
        lay.addLayout(r3)

        r3b = QHBoxLayout()
        self.text_sheets_edit = QLineEdit()
        self.text_sheets_edit.setReadOnly(True)
        self.text_sheets_edit.setPlaceholderText("Select one or more sheets from Text file")
        b_pick_sheets = QPushButton("Select Sheets...")
        b_pick_sheets.setObjectName("Secondary")
        b_pick_sheets.clicked.connect(self.pick_text_sheets)
        lb3b = QLabel("Text Sheets")
        lb3b.setObjectName("FieldLabel")
        r3b.addWidget(lb3b)
        r3b.addWidget(self.text_sheets_edit, 1)
        r3b.addWidget(b_pick_sheets)
        lay.addLayout(r3b)

        r4 = QHBoxLayout()
        self.var_edit = QLineEdit("s14b")
        self.var_edit.setReadOnly(True)
        lb4 = QLabel("Questions")
        lb4.setObjectName("FieldLabel")
        r4.addWidget(lb4)
        r4.addWidget(self.var_edit, 1)
        b_load_settings = QPushButton("Load Settings...")
        b_load_settings.setObjectName("Secondary")
        b_load_settings.clicked.connect(self.load_settings_items)
        r4.addWidget(b_load_settings)
        b_load_settings_sheet = QPushButton("Select Settings Sheet...")
        b_load_settings_sheet.setObjectName("Secondary")
        b_load_settings_sheet.clicked.connect(self.pick_settings_sheets)
        r4.addWidget(b_load_settings_sheet)
        b_save_settings = QPushButton("Save Settings...")
        b_save_settings.setObjectName("Secondary")
        b_save_settings.clicked.connect(self.save_settings_items_as)
        r4.addWidget(b_save_settings)
        b_edit_items = QPushButton("Select Items/Merge...")
        b_edit_items.setObjectName("Secondary")
        b_edit_items.clicked.connect(self.edit_settings_items)
        r4.addWidget(b_edit_items)
        lay.addLayout(r4)

        r4b = QHBoxLayout()
        self.merge_mode_combo = QComboBox()
        self.merge_mode_combo.addItem("Use Settings (Recommended)", userData="settings")
        self.merge_mode_combo.addItem("Force Merge All", userData="all")
        self.merge_mode_combo.addItem("No Merge", userData="none")
        lb5 = QLabel("Merge Mode")
        lb5.setObjectName("FieldLabel")
        r4b.addWidget(lb5)
        r4b.addWidget(self.merge_mode_combo)
        r4b.addStretch(1)
        lay.addLayout(r4b)

        self.status = QLabel("Ready")
        self.status.setObjectName("StatusPill")
        lay.addWidget(self.status)
        self.progress = QProgressBar()
        self.progress.setTextVisible(False)
        self.progress.setRange(0, 1)
        self.progress.setValue(0)
        lay.addWidget(self.progress)

        self.run_btn = QPushButton("Run and Save")
        self.run_btn.setObjectName("Primary")
        self.run_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton))
        self.run_btn.clicked.connect(self.run)
        lay.addWidget(self.run_btn, alignment=Qt.AlignmentFlag.AlignRight)

        self._busy_timer = QTimer(self)
        self._busy_timer.setInterval(500)
        self._busy_timer.timeout.connect(self._update_busy_elapsed)
        self._elapsed = QElapsedTimer()
        self._busy_base_text = "Processing..."

        self._controls_to_disable = [
            self.raw_edit,
            self.sheet_combo,
            self.text_edit,
            self.text_sheets_edit,
            self.var_edit,
            self.merge_mode_combo,
            self.run_btn,
        ]

        root.addWidget(card)
        root.addStretch(1)

    def _apply_dialog_style(self, dlg: QDialog):
        dlg.setStyleSheet(
            """
            QDialog {
                background: #0f1b33;
                color: #e7eefc;
            }
            QLabel {
                color: #dbeafe;
            }
            QListWidget, QTableWidget {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.18);
                border-radius: 10px;
                color: #f2f7ff;
                gridline-color: rgba(255,255,255,0.15);
            }
            QHeaderView::section {
                background: rgba(56,189,248,0.20);
                color: #dbeafe;
                border: none;
                padding: 5px;
                font-weight: 700;
            }
            QPushButton {
                border: none;
                border-radius: 9px;
                padding: 7px 12px;
                font-weight: 700;
                color: #052234;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #67e8f9, stop:1 #2dd4bf);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #a5f3fc, stop:1 #5eead4);
            }
            """
        )

    def info(self, title: str, text: str):
        self._show_message("info", title, text)

    def err(self, title: str, text: str):
        self._show_message("error", title, text)

    def _show_message(self, kind: str, title: str, text: str):
        msg = QMessageBox(self)
        if kind == "error":
            msg.setIcon(QMessageBox.Icon.Critical)
        else:
            msg.setIcon(QMessageBox.Icon.Information)
        msg.setWindowTitle(title)
        msg.setText(text)
        msg.setStyleSheet(
            """
            QMessageBox {
                background: #f8fafc;
            }
            QMessageBox QLabel {
                color: #0f172a;
                min-width: 360px;
                font-size: 11pt;
            }
            QMessageBox QPushButton {
                min-width: 88px;
                border: 1px solid #93c5fd;
                border-radius: 8px;
                padding: 6px 12px;
                color: #0f172a;
                background: #dbeafe;
                font-weight: 700;
            }
            QMessageBox QPushButton:hover {
                background: #bfdbfe;
            }
            """
        )
        msg.exec()

    def _set_busy(self, busy: bool, text: str = "Processing..."):
        if busy:
            self._busy_base_text = text
            self._elapsed.start()
            self._busy_timer.start()
            self.progress.setRange(0, 0)
            self.run_btn.setText("Running...")
            for w in self._controls_to_disable:
                w.setEnabled(False)
            self._update_busy_elapsed()
            QApplication.processEvents()
            return

        self._busy_timer.stop()
        self.progress.setRange(0, 1)
        self.run_btn.setText("Run and Save")
        for w in self._controls_to_disable:
            w.setEnabled(True)

    def _update_busy_elapsed(self):
        secs = self._elapsed.elapsed() // 1000 if self._elapsed.isValid() else 0
        self.status.setText(f"{self._busy_base_text}  ({secs}s)")

    def pick_raw(self):
        p, _ = QFileDialog.getOpenFileName(self, "Select Raw file", "", "Excel files (*.xlsx *.xls)")
        if not p:
            return
        self.raw_path = p
        self.raw_edit.setText(p)
        try:
            xls = pd.ExcelFile(p)
            self.sheet_combo.clear()
            self.sheet_combo.addItems(xls.sheet_names)
            if "Diary" in xls.sheet_names:
                self.sheet_combo.setCurrentText("Diary")
        except Exception as e:
            self.sheet_combo.clear()
            self.err("Error", f"Cannot read sheets from Raw: {e}")

    def pick_text(self):
        p, _ = QFileDialog.getOpenFileName(self, "Select Text file", "", "Excel files (*.xlsx *.xls)")
        if not p:
            return
        self.text_path = p
        self.text_edit.setText(p)
        try:
            xls = pd.ExcelFile(p)
            self.text_sheet_names = list(xls.sheet_names)
            self.selected_text_sheets = list(self.text_sheet_names)
            self.text_sheets_edit.setText(", ".join(self.selected_text_sheets))
        except Exception:
            self.text_sheet_names = []
            self.selected_text_sheets = []
            self.text_sheets_edit.clear()

    def pick_text_sheets(self):
        if not self.text_sheet_names:
            self.err("Error", "Please select Text file first")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Select Text Sheets")
        self._apply_dialog_style(dlg)
        lay = QVBoxLayout(dlg)
        lw = QListWidget()
        selected_set = set(self.selected_text_sheets) if self.selected_text_sheets else set(self.text_sheet_names)
        for name in self.text_sheet_names:
            item = QListWidgetItem(name)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Checked if name in selected_set else Qt.CheckState.Unchecked)
            lw.addItem(item)
        lay.addWidget(lw)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        picked = []
        for i in range(lw.count()):
            item = lw.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                picked.append(item.text())

        self.selected_text_sheets = picked
        self.text_sheets_edit.setText(", ".join(picked) if picked else "(none selected)")

    def run(self):
        raw = txt(self.raw_edit.text())
        textf = txt(self.text_edit.text())
        text_sheets = list(self.selected_text_sheets)
        selected_items = []
        if self.settings_items:
            selected_items = [it for it in self.settings_items if bool(it.get("enabled", True))]
            var_names = [txt(it.get("name")) for it in selected_items if txt(it.get("name")) != ""]
        else:
            var_names = parse_var_list(self.var_edit.text())
        mode = self.merge_mode_combo.currentData()
        if self.settings_items:
            if mode == "all":
                var_configs = [{"name": txt(it["name"]), "merge": True} for it in selected_items]
            elif mode == "none":
                var_configs = [{"name": txt(it["name"]), "merge": False} for it in selected_items]
            else:
                var_configs = [{"name": txt(it["name"]), "merge": bool(it.get("merge", False))} for it in selected_items]
        else:
            if mode == "all":
                var_configs = [{"name": v, "merge": True} for v in var_names]
            elif mode == "none":
                var_configs = [{"name": v, "merge": False} for v in var_names]
            else:
                var_configs = [{"name": v, "merge": self.var_merge_map.get(norm(v), False)} for v in var_names]
        merge_count = sum(1 for cfg in var_configs if bool(cfg.get("merge", False)))
        if self.settings_items and mode == "none":
            selected_merge_count = sum(1 for it in selected_items if bool(it.get("merge", False)))
            if selected_merge_count > 0:
                ans = QMessageBox.question(
                    self,
                    "Confirm No Merge",
                    f"You selected No Merge, which will ignore {selected_merge_count} checked Merge items.\nContinue?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No,
                )
                if ans != QMessageBox.StandardButton.Yes:
                    return
        sheet = txt(self.sheet_combo.currentText())
        if not raw or not os.path.exists(raw):
            self.err("Error", "Please select a valid Raw file")
            return
        if not textf or not os.path.exists(textf):
            self.err("Error", "Please select a valid Text file")
            return
        if not text_sheets:
            self.err("Error", "Please select at least one Text sheet")
            return
        if not var_names:
            self.err("Error", "Please select at least one question item")
            return
        if not sheet:
            self.err("Error", "Please select Raw sheet")
            return

        out, _ = QFileDialog.getSaveFileName(
            self,
            "Save output file",
            f"{os.path.splitext(os.path.basename(raw))[0]}_VERBATIM_MAPPED.xlsx",
            "Excel files (*.xlsx)",
        )
        if not out:
            return

        self._set_busy(True, "Processing...")
        try:
            res = run_merge(raw, sheet, textf, var_configs, out, text_sheets)
            self._save_settings_items_to_file()
            has_fail = any(d.get("status") != "SUCCESS" for d in res.get("details", []))
            if has_fail:
                self.info("Completed", f"Completed (some items failed - check Summary sheet)\nMerge items used: {merge_count}/{len(var_configs)}")
            else:
                self.info("Completed", f"Completed\nMerge items used: {merge_count}/{len(var_configs)}")
            self.status.setText("Completed")
            self.progress.setRange(0, 1)
            self.progress.setValue(1)
        except Exception as e:
            self.status.setText("Error")
            self.err("Error", "Failed")
            self.progress.setRange(0, 1)
            self.progress.setValue(0)
        finally:
            self._set_busy(False)

    def load_settings_items(self):
        p, _ = QFileDialog.getOpenFileName(
            self,
            "Select settings file",
            "",
            "Excel files (*.xlsx *.xls)",
        )
        if not p:
            return
        try:
            xls = pd.ExcelFile(p)
            self.settings_path = p
            self.settings_sheet_names = list(xls.sheet_names)
            self.selected_settings_sheets = self._choose_settings_sheets(self.settings_sheet_names)
            if not self.selected_settings_sheets:
                return
            self._load_settings_from_path(p, self.selected_settings_sheets)
        except Exception as e:
            self.err("Error", f"Cannot load settings: {e}")

    def pick_settings_sheets(self):
        if not self.settings_path:
            p, _ = QFileDialog.getOpenFileName(
                self,
                "Select settings file",
                "",
                "Excel files (*.xlsx *.xls)",
            )
            if not p:
                return
            self.settings_path = p

        try:
            xls = pd.ExcelFile(self.settings_path)
            self.settings_sheet_names = list(xls.sheet_names)
        except Exception as e:
            self.err("Error", f"Cannot read settings sheets: {e}")
            return

        picked = self._choose_settings_sheets(self.settings_sheet_names)
        if not picked:
            return
        self.selected_settings_sheets = picked
        try:
            self._load_settings_from_path(self.settings_path, self.selected_settings_sheets)
        except Exception as e:
            self.err("Error", f"Cannot load settings: {e}")

    def _choose_settings_sheets(self, sheet_names: list[str]) -> list[str]:
        if not sheet_names:
            self.err("Error", "No sheets found in settings file")
            return []

        dlg = QDialog(self)
        dlg.setWindowTitle("Select Settings Sheet")
        self._apply_dialog_style(dlg)
        lay = QVBoxLayout(dlg)
        lay.addWidget(QLabel("Choose settings sheet(s) to load:"))

        lw = QListWidget()
        selected_set = set(self.selected_settings_sheets) if self.selected_settings_sheets else {sheet_names[0]}
        for name in sheet_names:
            item = QListWidgetItem(name)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Checked if name in selected_set else Qt.CheckState.Unchecked)
            lw.addItem(item)
        lay.addWidget(lw)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return []

        picked = []
        for i in range(lw.count()):
            item = lw.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                picked.append(item.text())

        if not picked:
            self.err("Error", "Please select at least one Settings sheet")
        return picked

    def _load_settings_from_path(self, path: str, sheets: list[str]):
        items, warnings = load_items_from_settings_file(path, sheets)
        self.settings_items = [{"name": txt(it["name"]), "merge": bool(it.get("merge", False)), "enabled": True} for it in items]
        self.var_merge_map = {norm(it["name"]): bool(it.get("merge", False)) for it in items}
        self.refresh_selected_questions_text()

        merge_count = sum(1 for it in items if it.get("merge"))
        sheet_text = ", ".join(sheets)
        self.status.setText(f"Loaded settings sheet: {sheet_text}")
        if warnings:
            msg = f"Loaded {len(items)} settings items from sheet(s): {sheet_text}\n\nWarnings:\n"
            msg += "\n".join(f"- {w}" for w in warnings[:10])
            if len(warnings) > 10:
                msg += f"\n- ... and {len(warnings) - 10} more"
            self.info("Loaded with warnings", msg)
        else:
            self.info("Loaded", f"Loaded {len(items)} items from settings sheet(s): {sheet_text}\nMerge: {merge_count}")

    def refresh_selected_questions_text(self):
        if self.settings_items:
            names = [txt(it["name"]) for it in self.settings_items if bool(it.get("enabled", True))]
            self.var_edit.setText(",".join(names))
        else:
            self.var_edit.setText("")

    def edit_settings_items(self):
        if not self.settings_items:
            self.err("Error", "Please load settings first")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Select Items and Merge")
        dlg.resize(760, 560)
        self._apply_dialog_style(dlg)
        lay = QVBoxLayout(dlg)

        tbl = QTableWidget(len(self.settings_items), 3)
        tbl.setHorizontalHeaderLabels(["Item", "Use", "Merge"])
        tbl.verticalHeader().setVisible(False)
        tbl.setColumnWidth(0, 500)
        tbl.setColumnWidth(1, 90)
        tbl.setColumnWidth(2, 90)
        tbl.setMinimumHeight(420)

        for r, it in enumerate(self.settings_items):
            name_item = QTableWidgetItem(txt(it.get("name")))
            tbl.setItem(r, 0, name_item)

            use_item = QTableWidgetItem("")
            use_item.setFlags(use_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            use_item.setCheckState(Qt.CheckState.Checked if bool(it.get("enabled", True)) else Qt.CheckState.Unchecked)
            tbl.setItem(r, 1, use_item)

            merge_item = QTableWidgetItem("")
            merge_item.setFlags(merge_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            merge_item.setCheckState(Qt.CheckState.Checked if bool(it.get("merge", False)) else Qt.CheckState.Unchecked)
            tbl.setItem(r, 2, merge_item)

        lay.addWidget(tbl)

        btn_row = QHBoxLayout()
        b_use_all = QPushButton("Use: Select All")
        b_use_none = QPushButton("Use: Clear All")
        b_merge_all = QPushButton("Merge: Select All")
        b_merge_none = QPushButton("Merge: Clear")

        def select_all_items():
            for rr in range(tbl.rowCount()):
                if tbl.item(rr, 1) is not None:
                    tbl.item(rr, 1).setCheckState(Qt.CheckState.Checked)

        def clear_all_items():
            for rr in range(tbl.rowCount()):
                if tbl.item(rr, 1) is not None:
                    tbl.item(rr, 1).setCheckState(Qt.CheckState.Unchecked)

        def select_all_merge():
            for rr in range(tbl.rowCount()):
                if tbl.item(rr, 2) is not None:
                    tbl.item(rr, 2).setCheckState(Qt.CheckState.Checked)

        def clear_all_merge():
            for rr in range(tbl.rowCount()):
                if tbl.item(rr, 2) is not None:
                    tbl.item(rr, 2).setCheckState(Qt.CheckState.Unchecked)

        b_use_all.clicked.connect(select_all_items)
        b_use_none.clicked.connect(clear_all_items)
        b_merge_all.clicked.connect(select_all_merge)
        b_merge_none.clicked.connect(clear_all_merge)
        btn_row.addWidget(b_use_all)
        btn_row.addWidget(b_use_none)
        btn_row.addWidget(b_merge_all)
        btn_row.addWidget(b_merge_none)
        btn_row.addStretch(1)
        lay.addLayout(btn_row)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        for r, it in enumerate(self.settings_items):
            use_item = tbl.item(r, 1)
            merge_item = tbl.item(r, 2)
            it["enabled"] = use_item is not None and use_item.checkState() == Qt.CheckState.Checked
            it["merge"] = merge_item is not None and merge_item.checkState() == Qt.CheckState.Checked

        self.var_merge_map = {norm(it["name"]): bool(it.get("merge", False)) for it in self.settings_items}
        self.refresh_selected_questions_text()
        self._save_settings_items_to_file()

    def _save_settings_items_to_file(self, path: str | None = None, show_message: bool = False):
        if not self.settings_items:
            return
        target = txt(path) if path is not None else txt(self.settings_path)
        if target == "":
            return
        rows = []
        for it in self.settings_items:
            name = txt(it.get("name"))
            if name == "":
                continue
            rows.append(
                {
                    "Item": name,
                    "Use": "1" if bool(it.get("enabled", True)) else "",
                    "Merge": "1" if bool(it.get("merge", False)) else "",
                }
            )
        out_df = pd.DataFrame(rows, columns=["Item", "Use", "Merge"])
        out_df.to_excel(target, index=False, sheet_name="Sheet1")
        self.settings_path = target
        if show_message:
            self.info("Saved", f"Saved settings to:\n{target}")

    def save_settings_items_as(self):
        if not self.settings_items:
            self.err("Error", "Please load settings first")
            return
        default_name = os.path.basename(self.settings_path) if self.settings_path else "Seting OE.xlsx"
        out, _ = QFileDialog.getSaveFileName(
            self,
            "Save settings file",
            default_name,
            "Excel files (*.xlsx)",
        )
        if not out:
            return
        try:
            self._save_settings_items_to_file(path=out, show_message=True)
        except Exception as e:
            self.err("Error", f"Cannot save settings: {e}")


def main():
    app = QApplication([])
    app.setWindowIcon(make_leaf_icon())
    win = VerbatimMapperApp()
    win.show()
    app.exec()


if __name__ == "__main__":
    main()



