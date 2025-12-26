# -*- coding: utf-8 -*-
import sys
import platform # เพิ่ม import นี้

# --- เพิ่มโค้ดส่วนนี้ ---
# ทำให้ App รองรับ DPI Scaling บน Windows (คงไว้อย่างเดิม)
if platform.system() == "Windows":
    try:
        import ctypes
        # สำหรับ Windows 8.1 และใหม่กว่า (แนะนำ)
        ctypes.windll.shcore.SetProcessDpiAwareness(2) # PROCESS_PER_MONITOR_DPI_AWARE
        print("INFO: Set DPI awareness using shcore (Per-Monitor v2)")
    except (ImportError, AttributeError):
        try:
            # สำหรับ Windows Vista และ 7
            ctypes.windll.user32.SetProcessDPIAware()
            print("INFO: Set DPI awareness using user32 (System Aware)")
        except (ImportError, AttributeError):
            print("WARNING: Could not set DPI awareness.")
    except Exception as e:
        print(f"WARNING: Error setting DPI awareness - {e}")
# --- สิ้นสุดโค้ดที่เพิ่ม ---

# -*- coding: utf-8 -*-
import time
import tkinter as tk
import tkinter.filedialog as fd
import tkinter.ttk as ttk
import tkinter.messagebox as tkmb
import customtkinter as ctk
from customtkinter import *
import pandas as pd
from collections import defaultdict
# import sys # ซ้ำกับด้านบน เอาออกได้
import os
import traceback
import datetime # สำหรับ timestamp ใน log

# *** Import openpyxl with enhanced error handling ***
# (ส่วนนี้เหมือนเดิม)
try:
    import openpyxl
    from openpyxl.styles import Border, Side, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    root_temp = tk.Tk()
    root_temp.withdraw()
    tkmb.showerror("Missing Critical Library",
                   "The 'openpyxl' library is required for Excel operations.\n\n"
                   "Please install it by running this command in your terminal or command prompt:\n"
                   "pip install openpyxl\n\n"
                   "The application will now exit.")
    root_temp.destroy()
    sys.exit("Error: Critical 'openpyxl' library not found.")

# --- *** เพิ่มฟังก์ชัน resource_path ตรงนี้ *** ---
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
        # print(f"DEBUG: Running from MEIPASS: {base_path}") # Debug
    except Exception:
        # _MEIPASS not set, so running in normal Python environment
        base_path = os.path.abspath(os.path.dirname(__file__))
        # print(f"DEBUG: Running from script dir: {base_path}") # Debug
    return os.path.join(base_path, relative_path)
# --- *** สิ้นสุดฟังก์ชัน resource_path *** ---


# --- การตั้งค่า Theme --- (เหมือนเดิม)
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

# --- ค่าคงที่ Index คอลัมน์ --- (เหมือนเดิม)
COL_IDX = {
    'Segment': 0, 'Format': 1, 'Code': 2, 'ItemType': 3, 'Display': 4,
    'LoopSub': 5, 'ID': 6, 'Label': 7, 'CategoryType': 8, 'CategoryWeight': 9,
    'Digit': 10, 'Min': 11, 'Max': 12, 'AfterDecimal': 13, 'Statistic': 14, # Col O
    'Conditions': 15, # Col P
    'SpecialConditions': 16, # Col Q (สมมติว่ามี)
    'BaseType': 17      # Col R
}
EXPECTED_COL_COUNT = max(COL_IDX.values()) + 1
print(f"Expected minimum column count: {EXPECTED_COL_COUNT}")

# --- ค่าคงที่ตัวเลือก Making --- (เหมือนเดิม)
NEW_MAKING_OPTIONS = ["TB", "T2B", "BB", "B2B"] # 5 scale
NEW_MAKING_OPTIONS_7_10 = ["TB", "T2B", "T3B", "BB", "B2B", "B3B"] # 7/10 scale
DEFAULT_MAKING_OPTION_CATEGORY = "Valid"

# --- Application Class ---
class ExcelScaleExtractorApp(ctk.CTk):

    def __init__(self):
        super().__init__()
        self.title("โปรแกรมสร้าง TB/T2B จาก Itemdef")
        # --- *** เพิ่มโค้ดตั้งค่า Icon หลังจาก super().__init__() *** ---
        try:
            # ใช้ชื่อไฟล์ไอคอนของคุณที่นี่
            icon_filename = "T2B.ico"
            icon_path = resource_path(icon_filename)
            # ตรวจสอบว่าไฟล์มีอยู่จริงหรือไม่ก่อนเรียก iconbitmap (Optional แต่แนะนำ)
            if os.path.exists(icon_path):
                self.iconbitmap(icon_path)
                print(f"INFO: Loaded icon from {icon_path}")
            else:
                 print(f"WARNING: Icon file not found at {icon_path}")
        except tk.TclError as e:
            # ดักจับข้อผิดพลาด TclError ที่อาจเกิดจาก iconbitmap
             print(f"WARNING: Could not set icon (TclError): {e}")
        except Exception as e:
            # ดักจับข้อผิดพลาดอื่นๆ
            print(f"WARNING: Could not set icon (Other Error): {e}")
        # --- *** สิ้นสุดโค้ดตั้งค่า Icon *** ---
        # --- กำหนดขนาดหน้าต่างเบื้องต้น ---
        window_width = 1150
        window_height = 750
        self.geometry(f"{window_width}x{window_height}")

        # --- *** เพิ่มโค้ดจัดหน้าต่างกลางจอ *** ---
        self.update_idletasks() # สำคัญ: ทำให้ Tkinter รู้ขนาดที่แท้จริงของหน้าต่างก่อนคำนวณ
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        # print(f"Screen: {screen_width}x{screen_height}") # Debug (เอาออกได้)
        # print(f"Window requested: {window_width}x{window_height}") # Debug (เอาออกได้)

        # คำนวณตำแหน่ง x, y สำหรับมุมบนซ้ายของหน้าต่าง
        pos_x = (screen_width // 2) - (window_width // 2)
        pos_y = (screen_height // 2) - (window_height // 2)

        # ตั้งค่า geometry ใหม่พร้อมตำแหน่ง
        self.geometry(f"{window_width}x{window_height}+{pos_x}+{pos_y}")
        print(f"Set geometry to: {window_width}x{window_height}+{pos_x}+{pos_y}")
        # --- *** สิ้นสุดโค้ดจัดหน้าต่างกลางจอ *** ---


        # --- Frames --- (ส่วนนี้เหมือนเดิม)
        self.top_frame = ctk.CTkFrame(self)
        self.top_frame.pack(pady=10, padx=10, fill="x", side=tk.TOP)

        self.middle_frame = ctk.CTkFrame(self)
        self.middle_frame.pack(pady=(0, 5), padx=10, fill="both", expand=True)

        self.bottom_frame = ctk.CTkFrame(self)
        self.bottom_frame.pack(pady=(0, 10), padx=10, fill="x", side=tk.BOTTOM)
        self.bottom_frame.grid_columnconfigure(0, weight=1) # Column 0 expands

        # --- Widgets --- (ส่วนที่เหลือของ __init__ เหมือนเดิม)
        # --- Top Frame Widgets (Buttons Only) ---
        self.load_button = ctk.CTkButton(self.top_frame, text="โหลด Itemdef", command=self.load_excel)
        self.load_button.pack(side=tk.LEFT, padx=5, pady=5)
        self.generate_button = ctk.CTkButton(self.top_frame, text="สร้าง T2B Making", command=self.run_making_generation, state="disabled")
        self.generate_button.pack(side=tk.LEFT, padx=(10, 5), pady=5)
        self.save_button = ctk.CTkButton(self.top_frame, text="Save Itemdef Making...", command=self.save_as_new_file, state="disabled")
        self.save_button.pack(side=tk.LEFT, padx=(5, 5), pady=5)

        # --- Middle Frame: TabView & Treeviews ---
        self.tab_view = ctk.CTkTabview(self.middle_frame); self.tab_view.pack(fill="both", expand=True, padx=5, pady=5)
        self.tab_view.add("ตัวแปร Original"); self.tab_view.add("ตัวแปรที่ทำ Making")
        self.survey_tree_frame = ctk.CTkFrame(self.tab_view.tab("ตัวแปร Original"), fg_color="transparent"); self.survey_tree_frame.pack(fill="both", expand=True)
        self.survey_tree_frame.grid_columnconfigure(0, weight=1); self.survey_tree_frame.grid_rowconfigure(0, weight=1)
        survey_columns = ("Type", "ข้อแบบสอบถาม", "Code เริ่มต้น / Scale", "Code สุดท้าย / Scale", "Direction", "Condition Status")
        style = ttk.Style(self); style.theme_use("clam"); style.configure("Treeview", background="#FFFFFF", foreground="black", fieldbackground="#FFFFFF", rowheight=25, font=('TkDefaultFont', 10)); style.configure("Treeview.Heading", font=('TkDefaultFont', 10, 'bold')); style.map('Treeview', background=[('selected', '#BFDFFF')], foreground=[('selected', 'black')])
        self.survey_tree = ttk.Treeview(self.survey_tree_frame, columns=survey_columns, show="headings", style="Treeview")
        self.survey_tree.heading("Type", text="Type"); self.survey_tree.heading("ข้อแบบสอบถาม", text="ข้อแบบสอบถาม"); self.survey_tree.heading("Code เริ่มต้น / Scale", text="Code เริ่มต้น / Scale"); self.survey_tree.heading("Code สุดท้าย / Scale", text="Code สุดท้าย / Scale"); self.survey_tree.heading("Direction", text="คลิกขวาเลือก Scale"); self.survey_tree.heading("Condition Status", text="Condition Status")
        self.survey_tree.column("Type", width=100, anchor=tk.W); self.survey_tree.column("ข้อแบบสอบถาม", width=80, anchor=tk.W); self.survey_tree.column("Code เริ่มต้น / Scale", width=200, anchor=tk.W); self.survey_tree.column("Code สุดท้าย / Scale", width=200, anchor=tk.W); self.survey_tree.column("Direction", width=130, anchor=tk.CENTER); self.survey_tree.column("Condition Status", width=120, anchor=tk.CENTER)
        survey_vsb = ttk.Scrollbar(self.survey_tree_frame, orient="vertical", command=self.survey_tree.yview); survey_hsb = ttk.Scrollbar(self.survey_tree_frame, orient="horizontal", command=self.survey_tree.xview)
        self.survey_tree.configure(yscrollcommand=survey_vsb.set, xscrollcommand=survey_hsb.set); self.survey_tree.grid(row=0, column=0, sticky='nsew'); survey_vsb.grid(row=0, column=1, sticky='ns'); survey_hsb.grid(row=1, column=0, sticky='ew')
        self.making_tree_frame = ctk.CTkFrame(self.tab_view.tab("ตัวแปรที่ทำ Making"), fg_color="transparent"); self.making_tree_frame.pack(fill="both", expand=True)
        self.making_tree_frame.grid_columnconfigure(0, weight=1); self.making_tree_frame.grid_rowconfigure(0, weight=1)
        making_columns = ("Type", "ข้อแบบสอบถาม", "ตัวแปร Original", "Code เริ่มต้น / Scale", "Code สุดท้าย / Scale", "Condition Status")
        self.making_tree = ttk.Treeview(self.making_tree_frame, columns=making_columns, show="headings", style="Treeview")
        self.making_tree.heading("Type", text="Type"); self.making_tree.heading("ข้อแบบสอบถาม", text="ตัวแปร Making"); self.making_tree.heading("ตัวแปร Original", text="ตัวแปร Original"); self.making_tree.heading("Code เริ่มต้น / Scale", text="Code เริ่มต้น / Scale"); self.making_tree.heading("Code สุดท้าย / Scale", text="Code สุดท้าย / Scale"); self.making_tree.heading("Condition Status", text="Condition Status")
        self.making_tree.column("Type", width=100, anchor=tk.W); self.making_tree.column("ข้อแบบสอบถาม", width=80, anchor=tk.W); self.making_tree.column("ตัวแปร Original", width=80, anchor=tk.W); self.making_tree.column("Code เริ่มต้น / Scale", width=200, anchor=tk.W); self.making_tree.column("Code สุดท้าย / Scale", width=200, anchor=tk.W); self.making_tree.column("Condition Status", width=120, anchor=tk.CENTER)
        making_vsb = ttk.Scrollbar(self.making_tree_frame, orient="vertical", command=self.making_tree.yview); making_hsb = ttk.Scrollbar(self.making_tree_frame, orient="horizontal", command=self.making_tree.xview)
        self.making_tree.configure(yscrollcommand=making_vsb.set, xscrollcommand=making_hsb.set); self.making_tree.grid(row=0, column=0, sticky='nsew'); making_vsb.grid(row=0, column=1, sticky='ns'); making_hsb.grid(row=1, column=0, sticky='ew')
        # --- สิ้นสุดส่วน Middle Frame ---

        # --- Bottom Frame Widgets ---
        self.status_label = ctk.CTkLabel(self.bottom_frame, text="Please load an Itemdef Excel file.", anchor="w", height=1)
        self.status_label.grid(row=0, column=0, padx=5, pady=(2,0), sticky="ew")
        self.progress_bar = ttk.Progressbar(self.bottom_frame, orient='horizontal', mode='indeterminate')
        self.log_textbox = ctk.CTkTextbox(self.bottom_frame, height=80, wrap=tk.WORD, state=tk.DISABLED, font=('TkDefaultFont', 10))
        self.log_textbox.grid(row=2, column=0, padx=5, pady=(2,2), sticky="nsew")
        self.bottom_frame.grid_rowconfigure(2, weight=1) # แถว Log ขยายได้

        # --- Event Bindings ---
        self.survey_tree.bind("<Button-3>", self.on_survey_tree_right_click)

        # --- Instance Variables ---
        self.excel_data_original_rows = None
        self.file_path = None
        self.scales_data_store = {}
        self.making_data_generated = False
        self.survey_main_row_indices = {}
        self.template_file_path = "template.xlsx" # จะถูกตั้งค่าใหม่ใน __main__
        self.ordered_survey_qids = []
        self.ordered_making_qids = []


    # --- Logging Function --- (เหมือนเดิม)
    def log_message(self, message, level="INFO"):
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        formatted_message = f"[{timestamp} {level}] {message}\n"
        self.log_textbox.configure(state=tk.NORMAL)
        self.log_textbox.insert(tk.END, formatted_message)
        self.log_textbox.see(tk.END)
        self.log_textbox.configure(state=tk.DISABLED)
        status_prefix = f"[{level}] " if level != "INFO" else ""
        self.status_label.configure(text=f"{status_prefix}{message}")
        color = ctk.ThemeManager.theme["CTkLabel"]["text_color"]
        if level == "ERROR": color = "#FF4D4D"
        elif level == "WARNING": color = "#FFA500"
        elif level == "SUCCESS": color = "#4CAF50"
        self.status_label.configure(text_color=color)
        print(formatted_message.strip())

    # --- UX Helper Functions --- (เหมือนเดิม)
    def _start_processing(self, status_message="Processing..."):
        self.config(cursor="watch")
        self.load_button.configure(state="disabled")
        self.generate_button.configure(state="disabled")
        self.save_button.configure(state="disabled")
        self.status_label.configure(text=status_message, text_color=ctk.ThemeManager.theme["CTkLabel"]["text_color"])
        self.progress_bar.grid(row=1, column=0, padx=5, pady=(2, 2), sticky="ew")
        self.progress_bar.start(10)
        self.update_idletasks()
        self.update()

    def _end_processing(self, final_status_message=None, level="INFO"):
        self.progress_bar.stop()
        self.progress_bar.grid_forget()
        self.config(cursor="")
        if final_status_message:
            self.log_message(final_status_message, level=level)
        self.load_button.configure(state="normal")
        if self.file_path and self.excel_data_original_rows is not None:
            self.generate_button.configure(state="normal")
            if self.making_data_generated:
                self.save_button.configure(state="normal")
            else:
                self.save_button.configure(state="disabled")
        else:
            self.generate_button.configure(state="disabled")
            self.save_button.configure(state="disabled")
        self.update_idletasks()
        self.update()

    # --- ฟังก์ชัน Load Excel --- (เหมือนเดิม)
    def load_excel(self):
        self.status_label.configure(text="Selecting file...", text_color=ctk.ThemeManager.theme["CTkLabel"]["text_color"])
        self.update_idletasks()
        selected_file_path = fd.askopenfilename(title="Select Itemdef Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
        if not selected_file_path:
            self.status_label.configure(text="Ready. Please load an Excel file.")
            self.log_message("File selection cancelled.", level="INFO")
            return

        try: base_filename = os.path.basename(selected_file_path)
        except Exception as e:
            self.log_message(f"Error getting basename from selected path: {e}", level="WARNING")
            base_filename = "Selected File"

        self.clear_results()
        self.file_path = selected_file_path
        self._start_processing(f"Loading: {base_filename}...")
        success = False
        final_message = ""
        level = "INFO"

        try:
            self.after(100, lambda: None); self.update()
            self.log_message(f"Reading Excel file: {self.file_path}", level="INFO")
            start_time = time.time()
            try:
                df_full = pd.read_excel(self.file_path, header=None, sheet_name=0, keep_default_na=False, na_values=['']).fillna('')
            except Exception as read_err:
                error_type = type(read_err).__name__
                raise ValueError(f"Cannot read Excel file: [{error_type}] {read_err}")
            read_time = time.time()

            if len(df_full.columns) < EXPECTED_COL_COUNT:
                raise ValueError(f"File structure error: Expected at least {EXPECTED_COL_COUNT} columns, found {len(df_full.columns)}.")

            self.excel_data_original_rows = df_full.values.tolist()
            self.log_message(f"Excel file read successfully ({read_time - start_time:.2f}s).", level="INFO")
            self.status_label.configure(text=f"Processing data from '{base_filename}'..."); self.update()

            self.log_message(f"Processing data...", level="INFO")
            self.process_and_find_indices_combined() # This will populate self.ordered_survey_qids
            process_time = time.time()
            self.log_message(f"Data processed ({process_time - read_time:.2f}s). Found {len(self.ordered_survey_qids)} potential scales.", level="INFO") # Use ordered list count

            self.status_label.configure(text=f"Populating survey view..."); self.update()
            self.log_message(f"Populating ตัวแปร Original view...", level="INFO")

            survey_items_ordered = []
            for qid in self.ordered_survey_qids: # วนลูปตามลำดับที่ถูกต้อง
                if qid in self.scales_data_store:
                    data = self.scales_data_store[qid]
                    if not data.get('is_making'): # ตรวจสอบว่าเป็น survey item จริงๆ
                        survey_items_ordered.append((qid, data))
                else:
                    self.log_message(f"Warning: Ordered QID '{qid}' not found in data store during display.", "WARNING")

            self._populate_treeview(survey_items_ordered, target_tree=self.survey_tree) # ส่ง list ที่เรียงแล้วไปแสดง

            self.tab_view.set("ตัวแปร Original")

            if self.survey_tree.get_children():
                num_found = len(self.survey_tree.get_children())
                plural = "s" if num_found > 1 else ""
                final_message = f"Load complete. Displayed {num_found} survey scale{plural}. Right-click 'Direction' cell to edit."
                level = "SUCCESS"
            else:
                final_message = "Load complete. No valid ตัวแปร Original (5, 7, or 10 options) found."
                level = "WARNING"
            success = True
            self.making_data_generated = False

        except FileNotFoundError as e:
            final_message = f"Error: File '{base_filename}' not found."; level = "ERROR"
            self.show_error_popup("File Not Found", f"{final_message}\nPlease check the path:\n{selected_file_path}")
            self.clear_results()
        except ValueError as ve:
            final_message = f"Error processing file: {ve}"; level = "ERROR"
            self.show_error_popup("Data Structure/Format Error", f"{final_message}\nPlease ensure '{base_filename}' has the correct format.")
            self.clear_results()
        except Exception as e:
            error_type = type(e).__name__
            final_message = f"An unexpected error occurred: [{error_type}] {e}"; level = "ERROR"
            self.show_error_popup("Loading Error", f"{final_message}\nCheck log/console for details regarding '{base_filename}'.")
            print(traceback.format_exc())
            self.clear_results()
        finally:
            self._end_processing(final_status_message=final_message, level=level if success or level=="ERROR" else "INFO")

    # --- ฟังก์ชัน Clear Results --- (เหมือนเดิม)
    def clear_results(self):
        for item in self.survey_tree.get_children(): self.survey_tree.delete(item)
        for item in self.making_tree.get_children(): self.making_tree.delete(item)
        self.scales_data_store = {}
        self.excel_data_original_rows = None
        self.survey_main_row_indices = {}
        self.making_data_generated = False
        self.file_path = None
        self.ordered_survey_qids = []
        self.ordered_making_qids = []
        self.log_textbox.configure(state=tk.NORMAL)
        self.log_textbox.delete("1.0", tk.END)
        self.log_textbox.configure(state=tk.DISABLED)
        self.generate_button.configure(state="disabled")
        self.save_button.configure(state="disabled")
        self.log_message("Results cleared. Please load an Itemdef Excel file.", level="INFO")

    # --- ฟังก์ชัน Process Survey Data & Find Indices --- (เหมือนเดิม)
    def process_and_find_indices_combined(self):
        if self.excel_data_original_rows is None:
            self.log_message("Cannot process data: Original Excel data not loaded.", level="ERROR")
            return

        self.scales_data_store = defaultdict(lambda: {
            'type': 'N/A', 'sub_labels': [], 'scale_options': [], 'direction': '',
            'is_making': False, 'conditions': None, 'condition_status': ''
        })
        self.survey_main_row_indices = {} # Store {qid: row_index} for valid scales
        potential_main_row_indices = {} # Temporary store {qid: row_index} during scan
        self.ordered_survey_qids = []
        self.ordered_making_qids = []
        current_main_id = None
        current_main_is_loop = False
        row_errors = []

        for idx, row in enumerate(self.excel_data_original_rows):
            if idx < 2: continue # Skip headers

            try:
                if len(row) <= max(COL_IDX.values()):
                    row_errors.append(f"Row {idx + 1}: Insufficient columns (found {len(row)}, expected {EXPECTED_COL_COUNT}).")
                    continue

                id_val = str(row[COL_IDX['ID']]).strip()
                item_type_val = str(row[COL_IDX['ItemType']]).strip()
                label_val = str(row[COL_IDX['Label']]).strip()
                loop_sub_val = str(row[COL_IDX['LoopSub']]).strip()

                is_main_id = (id_val and not id_val.isdigit() and '(' not in id_val and id_val not in ['SbjNum', 'F_Status', 'team_FW'])

                if is_main_id:
                    current_main_id = id_val
                    stored_type = item_type_val if item_type_val else self.scales_data_store[current_main_id].get('type', 'N/A')
                    self.scales_data_store[current_main_id]['type'] = stored_type
                    self.scales_data_store[current_main_id]['is_making'] = False
                    current_main_is_loop = 'Loop' in stored_type
                    if item_type_val: # Only track potential main index if type is defined on this row
                       potential_main_row_indices[current_main_id] = idx

                elif current_main_id and label_val:
                    is_sub_label = current_main_is_loop and '(' in id_val and ')' in id_val
                    is_option = (not id_val or id_val.isdigit()) and not (current_main_is_loop and loop_sub_val == "Loop sub")

                    if is_sub_label:
                        self.scales_data_store[current_main_id]['sub_labels'].append(label_val)
                    elif is_option:
                        scale_text = label_val
                        parts_space = label_val.split(' ', 1)
                        parts_tab = label_val.split('\t', 1)
                        if id_val.isdigit() and len(parts_space) == 2 and parts_space[0].isdigit() and parts_space[0] == id_val:
                            scale_text = parts_space[1].strip()
                        elif len(parts_tab) == 2 and parts_tab[0].isdigit():
                           scale_text = parts_tab[1].strip()

                        if scale_text:
                            if scale_text is not None and str(scale_text).strip():
                                self.scales_data_store[current_main_id]['scale_options'].append(str(scale_text).strip())

            except IndexError as ie:
                 row_errors.append(f"Row {idx + 1}: Data access error (IndexError).")
                 continue
            except Exception as ex:
                error_type_name = type(ex).__name__
                row_errors.append(f"Row {idx + 1}: Unexpected error ({error_type_name}).")
                continue

        validated_qids_in_order = []
        for qid, idx in potential_main_row_indices.items():
            if qid in self.scales_data_store:
                num_options = len(self.scales_data_store[qid].get('scale_options', []))
                if num_options in [5, 7, 10]:
                    self.survey_main_row_indices[qid] = idx # เก็บ index ของ QID ที่ผ่านเงื่อนไข
                    validated_qids_in_order.append(qid) # เก็บ QID ตามลำดับที่เจอและผ่าน

        self.ordered_survey_qids = validated_qids_in_order

        if row_errors:
            self.log_message(f"Processing completed with {len(row_errors)} potential issues.", level="WARNING")

    # --- ฟังก์ชัน Populate Treeview --- (เหมือนเดิม)
    def _populate_treeview(self, items_to_display, target_tree):
        tree_name = 'Survey' if target_tree == self.survey_tree else 'Making'
        self.log_message(f"Populating {tree_name} view...", level="INFO")
        displayed_count = 0

        for item in target_tree.get_children(): target_tree.delete(item)

        for q_id, data in items_to_display: # วนตามลำดับที่ส่งเข้ามา
            is_making = data.get('is_making', False)
            original_num_scale_options = 0
            first_label = "N/A"; last_label = "N/A"
            scale_options = data.get('scale_options', [])
            valid_scale_options = [opt for opt in scale_options if opt is not None and str(opt).strip()]
            original_num_scale_options = len(valid_scale_options)

            if valid_scale_options:
                first_label = str(valid_scale_options[0]).replace('\t', ' ')
                if len(valid_scale_options) >= 1:
                    last_label = str(valid_scale_options[-1]).replace('\t', ' ')

            if original_num_scale_options in [5, 7, 10]:
                displayed_count += 1
                item_type = data.get('type', "N/A")
                condition_status = data.get('condition_status', '')
                first_val_text = f"1 = {first_label}"
                last_val_text = f"{original_num_scale_options} = {last_label}"

                if target_tree == self.survey_tree:
                    direction = data.get('direction', '')
                    if not is_making and not direction and condition_status != "Making Generated":
                        condition_status = "ไม่ทำ Making"
                    display_values = (item_type, q_id, first_val_text, last_val_text, direction, condition_status)
                    tags_list = ('parent_row',)
                elif target_tree == self.making_tree:
                    original_q_id = data.get('original_q_id', 'N/A')
                    display_values = (item_type, q_id, original_q_id, first_val_text, last_val_text, condition_status)
                    tags_list = ('making_row',)
                else: continue

                if target_tree.exists(q_id): target_tree.item(q_id, values=display_values, tags=tags_list)
                else: target_tree.insert("", tk.END, iid=q_id, values=display_values, tags=tags_list)

        self.survey_tree.tag_configure('parent_row', background='#FFFFFF')
        self.making_tree.tag_configure('making_row', background='#E6F2FF')
        self.log_message(f"Finished populating {tree_name} view ({displayed_count} items displayed).", level="INFO")

    # --- ฟังก์ชัน Generate Conditions --- (เหมือนเดิม)
    def _generate_conditions(self, q_id, direction, num_labels):
        conditions = []
        status = "OK"
        base_q_id = q_id
        for i in range(1, num_labels + 1): conditions.append(f"{base_q_id}={i}")
        if num_labels == 5: num_new_options = 4; current_new_options_list = NEW_MAKING_OPTIONS
        elif num_labels in [7, 10]: num_new_options = 6; current_new_options_list = NEW_MAKING_OPTIONS_7_10
        else: status = f"Invalid Scale ({num_labels})"; return [], status
        no_dir_placeholder = ["NO_DIRECTION"] * num_new_options
        generated_new_conditions = []
        if direction == "Scale น้อยดี(-)":
            if num_labels == 5: l1,l2,h1,h2 = 1,2,5,4; generated_new_conditions=[f"{base_q_id}={l1}",f"{base_q_id}={l1}|{base_q_id}={l2}",f"{base_q_id}={h1}",f"{base_q_id}={h2}|{base_q_id}={h1}"]
            elif num_labels==7: l1,l2,l3,h1,h2,h3 = 1,2,3,7,6,5; generated_new_conditions=[f"{base_q_id}={l1}",f"{base_q_id}={l1}|{base_q_id}={l2}",f"{base_q_id}={l1}|{base_q_id}={l2}|{base_q_id}={l3}",f"{base_q_id}={h1}",f"{base_q_id}={h2}|{base_q_id}={h1}",f"{base_q_id}={h3}|{base_q_id}={h2}|{base_q_id}={h1}"]
            elif num_labels==10:l1,l2,l3,h1,h2,h3 = 1,2,3,10,9,8; generated_new_conditions=[f"{base_q_id}={l1}",f"{base_q_id}={l1}|{base_q_id}={l2}",f"{base_q_id}={l1}|{base_q_id}={l2}|{base_q_id}={l3}",f"{base_q_id}={h1}",f"{base_q_id}={h2}|{base_q_id}={h1}",f"{base_q_id}={h3}|{base_q_id}={h2}|{base_q_id}={h1}"]
        elif direction == "Scale มากดี(+)":
            if num_labels == 5: l1,l2,h1,h2 = 1,2,5,4; generated_new_conditions=[f"{base_q_id}={h1}",f"{base_q_id}={h1}|{base_q_id}={h2}",f"{base_q_id}={l1}",f"{base_q_id}={l1}|{base_q_id}={l2}"]
            elif num_labels==7: l1,l2,l3,h1,h2,h3 = 1,2,3,7,6,5; generated_new_conditions=[f"{base_q_id}={h1}",f"{base_q_id}={h1}|{base_q_id}={h2}",f"{base_q_id}={h1}|{base_q_id}={h2}|{base_q_id}={h3}",f"{base_q_id}={l1}",f"{base_q_id}={l1}|{base_q_id}={l2}",f"{base_q_id}={l1}|{base_q_id}={l2}|{base_q_id}={l3}"]
            elif num_labels==10:l1,l2,l3,h1,h2,h3 = 1,2,3,10,9,8; generated_new_conditions=[f"{base_q_id}={h1}",f"{base_q_id}={h1}|{base_q_id}={h2}",f"{base_q_id}={h1}|{base_q_id}={h2}|{base_q_id}={h3}",f"{base_q_id}={l1}",f"{base_q_id}={l1}|{base_q_id}={l2}",f"{base_q_id}={l1}|{base_q_id}={l2}|{base_q_id}={l3}"]
        else: status = "No Direction"; generated_new_conditions.extend(no_dir_placeholder)
        conditions.extend(generated_new_conditions)
        expected_len = num_labels + num_new_options
        if len(conditions) != expected_len:
            self.log_message(f"CRITICAL ERROR: Condition length mismatch for {q_id}! Expected {expected_len}, got {len(conditions)}.", level="ERROR")
            status = "Length Error"
            while len(conditions) < expected_len: conditions.append("ERR_LEN_COND")
            conditions = conditions[:expected_len]
        return conditions, status

    # --- ฟังก์ชัน Run Making Generation --- (เหมือนเดิม)
    def run_making_generation(self):
        if not self.scales_data_store or not self.survey_main_row_indices:
            self.log_message("Cannot Generate: No survey data loaded or no valid scales found.", level="WARNING")
            self.show_warning_popup("Cannot Generate", "No survey data loaded or no valid scales found. Please load a file first.")
            return

        items_requiring_direction = []
        q_ids_to_process = []

        for q_id in self.ordered_survey_qids:
            if q_id in self.scales_data_store:
                 data = self.scales_data_store[q_id]
                 if not data.get('is_making'):
                    if not data.get('direction'):
                        items_requiring_direction.append(q_id)
                    else:
                        q_ids_to_process.append(q_id)

        if items_requiring_direction:
            num_missing = len(items_requiring_direction)
            plural = "s" if num_missing > 1 else ""
            self.log_message(f"{num_missing} item{plural} need direction.", level="WARNING")
            message = f"{num_missing} ตัวแปร 5/7/10 Scale.\n\nที่ไม่ได้ทำ Marking"
            self.show_warning_popup("Direction Needed", message)
            for qid in items_requiring_direction:
                if self.survey_tree.exists(qid): self.survey_tree.set(qid, "Condition Status", "ไม่ทำ Making")

        if not q_ids_to_process:
            self.log_message("No items ready for Making generation (check directions).", level="WARNING")
            if not items_requiring_direction and not any(d.get('is_making') for d in self.scales_data_store.values()):
                 self.log_message("No scale items found to process.", level="INFO")
            elif not items_requiring_direction and any(d.get('direction') for qid, d in self.scales_data_store.items() if not d.get('is_making')):
                 self.log_message("All items have direction, but none were queued for processing.", level="WARNING")
            return


        self._start_processing(f"Generating Making data for {len(q_ids_to_process)} item(s)...")
        self.after(100, lambda: None); self.update()

        making_items_store = {}
        items_with_errors = []
        generated_ids_this_run = set()
        generated_making_qids_ordered_this_run = []

        try:
            total_items = len(q_ids_to_process)
            for i, q_id in enumerate(q_ids_to_process):
                if i % 5 == 0 or i == total_items - 1:
                    progress_msg = f"Processing item {i+1} of {total_items}..."
                    self.status_label.configure(text=progress_msg); self.update()

                data = self.scales_data_store.get(q_id)
                if not data or data.get('is_making', False) or not data.get('direction'):
                    self.log_message(f"Skipping {q_id} during generation (unexpected state).", level="WARNING")
                    continue

                original_scale_options = data.get('scale_options', [])
                valid_original_scale_options = [opt for opt in original_scale_options if opt is not None and str(opt).strip()]
                original_num_scale_options = len(valid_original_scale_options)

                if original_num_scale_options not in [5, 7, 10]:
                    self.log_message(f"Skipping {q_id}: Invalid number of valid scale options ({original_num_scale_options}) found during generation.", level="WARNING")
                    items_with_errors.append(f"{q_id} (Invalid Scale Opts: {original_num_scale_options})")
                    if self.survey_tree.exists(q_id): self.survey_tree.set(q_id, "Condition Status", f"Err: Invalid Opts ({original_num_scale_options})")
                    continue

                direction = data.get('direction')
                making_q_id = f"N{q_id}"

                try:
                    original_type = data.get('type', 'N/A')
                    new_type = original_type
                    if 'Loop(SA)' in original_type: new_type = original_type.replace('Loop(SA)', 'Loop(MA)', 1)
                    elif 'SA' in original_type: new_type = original_type.replace('SA', 'MA', 1)

                    current_new_options_list = NEW_MAKING_OPTIONS_7_10 if original_num_scale_options in [7, 10] else NEW_MAKING_OPTIONS
                    sub_labels = data.get('sub_labels', [])
                    new_labels = sub_labels + valid_original_scale_options + current_new_options_list
                    conditions_list, condition_status = self._generate_conditions(q_id, direction, original_num_scale_options)

                    if condition_status != "OK":
                        self.log_message(f"ERROR generating conditions for {q_id}: {condition_status}", level="ERROR")
                        items_with_errors.append(f"{q_id} (Condition: {condition_status})")
                        if self.survey_tree.exists(q_id): self.survey_tree.set(q_id, "Condition Status", f"Cond. Error: {condition_status}")
                        continue

                    new_data = {
                        'type': new_type,
                        'sub_labels': sub_labels,
                        'scale_options': valid_original_scale_options,
                        'labels': new_labels,
                        'direction': '',
                        'conditions': conditions_list,
                        'condition_status': condition_status,
                        'is_making': True,
                        'original_q_id': q_id
                    }
                    making_items_store[making_q_id] = new_data
                    if making_q_id not in generated_ids_this_run:
                        generated_ids_this_run.add(making_q_id)
                        generated_making_qids_ordered_this_run.append(making_q_id)
                    if self.survey_tree.exists(q_id): self.survey_tree.set(q_id, "Condition Status", "Making Generated")

                except Exception as gen_err:
                    error_type_name = type(gen_err).__name__
                    self.log_message(f"CRITICAL ERROR generating {making_q_id}: {error_type_name} - {gen_err}", level="ERROR")
                    items_with_errors.append(f"{q_id} (Runtime Error: {error_type_name})")
                    if self.survey_tree.exists(q_id): self.survey_tree.set(q_id, "Condition Status", f"Gen Error: {error_type_name}")

            self.status_label.configure(text="Processing complete. Preparing results..."); self.update()
            self.scales_data_store.update(making_items_store)
            total_generated = len(generated_ids_this_run)
            final_status_message = ""
            level = "INFO"

            self.ordered_making_qids = generated_making_qids_ordered_this_run

            if total_generated > 0:
                self.making_data_generated = True
                all_making_items_ordered = []
                for making_qid in self.ordered_making_qids:
                    if making_qid in self.scales_data_store:
                        data = self.scales_data_store[making_qid]
                        if data.get('is_making'):
                            all_making_items_ordered.append((making_qid, data))
                    else:
                         self.log_message(f"Warning: Ordered Making QID '{making_qid}' not found in data store during making display.", "WARNING")
                self._populate_treeview(all_making_items_ordered, target_tree=self.making_tree)
                final_status_message = f"Generated/Updated {total_generated} 'Making' items."
                level = "SUCCESS"
                self.after(200, lambda: self.tab_view.set("ตัวแปรที่ทำ Making"))
            else:
                 if not items_with_errors: final_status_message = "No new 'Making' items were generated."
                 if not any(d.get('is_making') for d in self.scales_data_store.values()):
                    self.making_data_generated = False

            if items_with_errors:
                final_status_message += f" Encountered errors on {len(items_with_errors)} item(s)."
                level = "WARNING"
                error_list_str = "\n".join([f"- {e}" for e in items_with_errors[:5]])
                if len(items_with_errors) > 5: error_list_str += "\n..."

        except Exception as e:
            error_type = type(e).__name__
            final_status_message = f"Generation process failed: [{error_type}] {e}"
            level = "ERROR"
            self.show_error_popup("Generation Error", f"{final_status_message}\nCheck log for details.")
            print(traceback.format_exc())
            self.making_data_generated = False
        finally:
            self._end_processing(final_status_message=final_status_message, level=level)

    # --- ฟังก์ชัน Generate Making Rows for Excel --- (เหมือนเดิม)
    def _generate_making_rows_for_excel(self, original_q_id, making_data, original_survey_row):
        output_rows = []
        making_q_id = f"N{original_q_id}"
        sub_labels = making_data.get('sub_labels', [])
        scale_options = making_data.get('scale_options', [])
        conditions = making_data.get('conditions', [])
        making_type = making_data.get('type', 'N/A')
        is_loop_making = 'Loop' in making_type
        num_sub_labels = len(sub_labels)
        original_num_scale_options = len(scale_options)

        if original_num_scale_options in [7, 10]: current_new_options_list = NEW_MAKING_OPTIONS_7_10
        elif original_num_scale_options == 5: current_new_options_list = NEW_MAKING_OPTIONS
        else:
             self.log_message(f"ERROR: Cannot generate Excel rows for {making_q_id}. Invalid scale option count ({original_num_scale_options}).", level="ERROR")
             return []

        num_new_options = len(current_new_options_list)
        expected_conditions_count = original_num_scale_options + num_new_options

        if len(conditions) != expected_conditions_count:
            self.log_message(f"CRITICAL ERROR: Cond. count mismatch for {making_q_id}! Expected {expected_conditions_count}, got {len(conditions)}. Rows invalid.", level="ERROR")
            while len(conditions) < expected_conditions_count: conditions.append("ERR_COND_COUNT")
            conditions = conditions[:expected_conditions_count]

        main_making_row = [''] * EXPECTED_COL_COUNT
        main_making_row[COL_IDX['Segment']] = original_survey_row[COL_IDX['Segment']] if len(original_survey_row) > COL_IDX['Segment'] else "Item"
        main_making_row[COL_IDX['Format']] = "Making"
        main_making_row[COL_IDX['ItemType']] = making_type
        main_making_row[COL_IDX['Display']] = original_survey_row[COL_IDX['Display']] if len(original_survey_row) > COL_IDX['Display'] else "O"
        main_making_row[COL_IDX['ID']] = making_q_id

        # --- ส่วนที่ปรับปรุง: คัดลอก Label จากคำถาม Survey ต้นฉบับ ---
        # เพื่อให้แถวหลักของ "Making" มีโจทย์คำถามเดียวกันกับตัวแปรต้นฉบับ
        if COL_IDX['Label'] < EXPECTED_COL_COUNT and len(original_survey_row) > COL_IDX['Label']:
            main_making_row[COL_IDX['Label']] = original_survey_row[COL_IDX['Label']]
        # --- สิ้นสุดส่วนที่ปรับปรุง ---

        if COL_IDX['Statistic'] < EXPECTED_COL_COUNT:
             main_making_row[COL_IDX['Statistic']] = str(original_q_id).upper()
        if COL_IDX['Conditions'] < EXPECTED_COL_COUNT:
             main_making_row[COL_IDX['Conditions']] = ''
        if COL_IDX['BaseType'] < EXPECTED_COL_COUNT:
             main_making_row[COL_IDX['BaseType']] = "Follow the condition items"
        if is_loop_making and len(original_survey_row) > COL_IDX['LoopSub'] and original_survey_row[COL_IDX['LoopSub']]:
             if COL_IDX['LoopSub'] < EXPECTED_COL_COUNT:
                 main_making_row[COL_IDX['LoopSub']] = original_survey_row[COL_IDX['LoopSub']]
        output_rows.append(main_making_row)

        option_category_type = DEFAULT_MAKING_OPTION_CATEGORY
        if original_q_id in self.survey_main_row_indices:
            survey_main_idx = self.survey_main_row_indices[original_q_id]
            first_option_idx_0based = survey_main_idx + 1 + num_sub_labels
            if first_option_idx_0based < len(self.excel_data_original_rows):
                first_option_row = self.excel_data_original_rows[first_option_idx_0based]
                if len(first_option_row) > max(COL_IDX['ID'], COL_IDX['Label'], COL_IDX['CategoryType']):
                    is_option_row = (str(first_option_row[COL_IDX['ID']]).strip().isdigit() or
                                   (not str(first_option_row[COL_IDX['ID']]).strip() and str(first_option_row[COL_IDX['Label']]).strip()))
                    if is_option_row and first_option_row[COL_IDX['CategoryType']]:
                        option_category_type = first_option_row[COL_IDX['CategoryType']]

        if is_loop_making:
            for i, sub_label in enumerate(sub_labels):
                sub_label_row = [''] * EXPECTED_COL_COUNT
                if COL_IDX['LoopSub'] < EXPECTED_COL_COUNT: sub_label_row[COL_IDX['LoopSub']] = "Loop sub"
                if COL_IDX['ID'] < EXPECTED_COL_COUNT: sub_label_row[COL_IDX['ID']] = f"{making_q_id}({i+1})"
                if COL_IDX['Label'] < EXPECTED_COL_COUNT: sub_label_row[COL_IDX['Label']] = sub_label
                output_rows.append(sub_label_row)

        overall_option_index = 0
        for i in range(original_num_scale_options):
            overall_option_index += 1
            option_row = [''] * EXPECTED_COL_COUNT
            if COL_IDX['ID'] < EXPECTED_COL_COUNT: option_row[COL_IDX['ID']] = overall_option_index
            if COL_IDX['Label'] < EXPECTED_COL_COUNT: option_row[COL_IDX['Label']] = scale_options[i]
            if COL_IDX['CategoryType'] < EXPECTED_COL_COUNT: option_row[COL_IDX['CategoryType']] = option_category_type
            if COL_IDX['Conditions'] < EXPECTED_COL_COUNT:
                 option_row[COL_IDX['Conditions']] = conditions[i] if i < len(conditions) else "ERR_MISSING_COND"
            if is_loop_making and COL_IDX['LoopSub'] < EXPECTED_COL_COUNT:
                option_row[COL_IDX['LoopSub']] = ""
            output_rows.append(option_row)

        start_index_new_conditions = original_num_scale_options
        for i, new_label in enumerate(current_new_options_list):
            overall_option_index += 1
            option_row = [''] * EXPECTED_COL_COUNT
            if COL_IDX['ID'] < EXPECTED_COL_COUNT: option_row[COL_IDX['ID']] = overall_option_index
            if COL_IDX['Label'] < EXPECTED_COL_COUNT: option_row[COL_IDX['Label']] = new_label
            if COL_IDX['CategoryType'] < EXPECTED_COL_COUNT: option_row[COL_IDX['CategoryType']] = option_category_type
            condition_index = start_index_new_conditions + i
            if COL_IDX['Conditions'] < EXPECTED_COL_COUNT:
                option_row[COL_IDX['Conditions']] = conditions[condition_index] if condition_index < len(conditions) else "ERR_MISSING_COND"
            if is_loop_making and COL_IDX['LoopSub'] < EXPECTED_COL_COUNT:
                option_row[COL_IDX['LoopSub']] = ""
            output_rows.append(option_row)

        return output_rows

    # --- ฟังก์ชัน Save As New File --- (เหมือนเดิม)
    def save_as_new_file(self):
        if not self.making_data_generated:
            self.log_message("Cannot Save: No 'Making' data generated.", level="WARNING")
            self.show_error_popup("Cannot Save", "No 'Making' data has been generated yet.")
            return
        if not self.file_path:
            self.log_message("Cannot Save: Original file path missing.", level="ERROR")
            self.show_error_popup("Cannot Save", "Original file path is missing.")
            return
        if not self.excel_data_original_rows:
            self.log_message("Cannot Save: Internal data missing (original rows).", level="ERROR")
            self.show_error_popup("Cannot Save", "Internal data missing (original rows). Please reload the file.")
            return
        if not self.survey_main_row_indices: # Need this to know where to insert
            self.log_message("Cannot Save: Internal data missing (survey indices).", level="ERROR")
            self.show_error_popup("Cannot Save", "Internal data missing (survey indices). Please reload and regenerate.")
            return

        try:
            if not hasattr(self, 'template_file_path') or not self.template_file_path or not os.path.exists(self.template_file_path):
                missing_path = self.template_file_path if hasattr(self, 'template_file_path') else "Not Set"
                self.log_message(f"Template file not found or path not set: {missing_path}", level="ERROR")
                self.show_error_popup("Template Not Found", f"Template file not found or path invalid:\n{missing_path}")
                return
            directory = os.path.dirname(self.file_path)
            original_filename = os.path.basename(self.file_path)
            original_filename_no_ext, original_extension = os.path.splitext(original_filename)
            if not original_extension.lower().startswith(".xls"): original_extension = ".xlsx"
            new_filename = f"{original_filename_no_ext}_making{original_extension}"
            new_file_path = os.path.join(directory, new_filename)
        except Exception as path_err:
            self.log_message(f"Error preparing file paths: {path_err}", level="ERROR")
            self.show_error_popup("Path Error", f"Error preparing file paths:\n\n{path_err}")
            return

        confirm = tkmb.askyesno("Confirm Save As New File",
                            f"Create NEW file:\n'{new_filename}'?\n\nOriginal file ('{original_filename}') will NOT be changed.",
                            icon='question', parent=self)
        if not confirm:
            self.log_message("Save As operation cancelled by user.", level="INFO")
            return

        self._start_processing(f"Saving data to '{new_filename}'...")
        self.after(100, lambda: None); self.update()
        total_start_time = time.time()
        workbook = None
        final_rows_to_write_local = []
        level = "INFO" # Default level

        try:
            self.status_label.configure(text="Building data in memory..."); self.update()
            start_prep_time = time.time()
            self.log_message("Building final data rows in memory...", level="INFO")

            if not self.excel_data_original_rows: raise ValueError("Original data missing before prep.")
            current_rows = [row[:] for row in self.excel_data_original_rows] # Deep copy
            making_blocks_info = {} # {original_qid: {'rows': [...], 'insert_at_idx_0based': idx}}
            original_qids_processed_for_making = set()
            data_prep_errors = []

            items_to_process_for_save = {qid: data for qid, data in self.scales_data_store.items() if data.get('is_making')}
            total_items = len(items_to_process_for_save)

            for i, (making_q_id, data) in enumerate(items_to_process_for_save.items()):
                if i % 10 == 0 or i == total_items - 1:
                    prep_progress = f"Preparing item {i+1} of {total_items}..."; self.status_label.configure(text=prep_progress); self.update()

                original_q_id = data.get('original_q_id')
                if not original_q_id: data_prep_errors.append(f"{making_q_id}: Missing original ID."); continue

                if original_q_id in self.survey_main_row_indices:
                    original_row_idx = self.survey_main_row_indices[original_q_id]
                    if original_row_idx < len(current_rows):
                        try:
                            making_rows = self._generate_making_rows_for_excel(original_q_id, data, current_rows[original_row_idx])
                            if not making_rows:
                                data_prep_errors.append(f"{making_q_id}: Row generation returned empty list.")
                                continue

                            num_subs = len(data.get('sub_labels', []))
                            num_opts = len(data.get('scale_options', []))
                            insert_idx = original_row_idx + 1 + num_subs + num_opts

                            making_blocks_info[original_q_id] = {'rows': making_rows, 'insert_at_idx_0based': insert_idx}
                            original_qids_processed_for_making.add(original_q_id)
                        except Exception as e: data_prep_errors.append(f"{making_q_id}: Row gen failed ({type(e).__name__} - {e}).")
                    else: data_prep_errors.append(f"{original_q_id}: Original index {original_row_idx} out of bounds ({len(current_rows)}).")
                else: data_prep_errors.append(f"{making_q_id}: Linked survey index missing for '{original_q_id}'.")

            self.status_label.configure(text="Processing weights..."); self.update()
            TARGET_WEIGHT_COL_IDX = COL_IDX['CategoryWeight']
            weights_applied_count = 0

            for original_q_id in original_qids_processed_for_making:
                survey_data = self.scales_data_store.get(original_q_id)
                if not survey_data: continue

                direction = survey_data.get('direction','')
                scale_options_for_weight = survey_data.get('scale_options', [])
                num_options = len(scale_options_for_weight)

                if num_options > 0 and direction in ['Scale น้อยดี(-)','Scale มากดี(+)'] and original_q_id in self.survey_main_row_indices:
                    num_sub_labels = len(survey_data.get('sub_labels',[]))
                    main_row_idx = self.survey_main_row_indices[original_q_id]
                    first_option_row_idx = main_row_idx + 1 + num_sub_labels
                    weights = list(range(num_options, 0, -1)) if direction == 'Scale น้อยดี(-)' else list(range(1, num_options + 1))
                    applied_flag = False

                    for i in range(num_options):
                        target_row_idx = first_option_row_idx + i
                        if target_row_idx < len(current_rows) and len(current_rows[target_row_idx]) > TARGET_WEIGHT_COL_IDX:
                            try:
                                current_rows[target_row_idx][TARGET_WEIGHT_COL_IDX] = weights[i]
                                applied_flag = True
                            except IndexError: data_prep_errors.append(f"{original_q_id}: Weight apply IndexError row {target_row_idx + 1}.")
                            except TypeError: data_prep_errors.append(f"{original_q_id}: Weight apply TypeError row {target_row_idx + 1} (is target non-numeric?).")
                        else: data_prep_errors.append(f"{original_q_id}: Weight target row {target_row_idx + 1} invalid or too short.")
                    if applied_flag: weights_applied_count += 1

            self.status_label.configure(text="Finalizing data structure..."); self.update()
            if making_blocks_info:
                final_rows_after_insertion = []
                sorted_keys = sorted(making_blocks_info.keys(), key=lambda q: making_blocks_info[q]['insert_at_idx_0based'])
                last_copied_idx = 0
                for qid in sorted_keys:
                    info = making_blocks_info[qid]
                    insert_idx = info['insert_at_idx_0based']
                    final_rows_after_insertion.extend(current_rows[last_copied_idx : insert_idx])
                    final_rows_after_insertion.extend(info['rows'])
                    last_copied_idx = insert_idx
                final_rows_after_insertion.extend(current_rows[last_copied_idx:])
                final_rows_to_write_local = final_rows_after_insertion
            else:
                final_rows_to_write_local = current_rows

            if not final_rows_to_write_local and self.excel_data_original_rows:
                 self.log_message("Warning: Final row list was empty after processing, reverting to original data.", level="WARNING")
                 final_rows_to_write_local = [row[:] for row in self.excel_data_original_rows]

            prep_time = time.time() - start_prep_time
            self.log_message(f"Data preparation complete ({prep_time:.2f}s). Final total rows: {len(final_rows_to_write_local)}.", level="INFO")
            if weights_applied_count > 0:
                 self.log_message(f"Applied weights to {weights_applied_count} original survey items.", level="INFO")
            if data_prep_errors:
                self.log_message(f"Encountered {len(data_prep_errors)} issues during data prep/weighting. Check logs.", level="WARNING")

            self.status_label.configure(text="Creating Excel workbook..."); self.update()
            self.log_message(f"Loading template: {self.template_file_path}", level="INFO")
            workbook = openpyxl.load_workbook(self.template_file_path)
            ws = workbook.active

            self.log_message(f"Writing {len(final_rows_to_write_local)} rows to worksheet...", level="INFO")
            self.status_label.configure(text=f"Writing data ({len(final_rows_to_write_local)} rows)..."); self.update()
            write_start_time = time.time()

            ID_COL, LABEL_COL, WEIGHT_COL, FORMAT_COL = COL_IDX['ID'] + 1, COL_IDX['Label'] + 1, COL_IDX['CategoryWeight'] + 1, COL_IDX['Format'] + 1
            making_fill = PatternFill(start_color='FFE6F2FF', end_color='FFE6F2FF', fill_type='solid') # Light blue fill
            wrap_align = Alignment(wrap_text=True, vertical='top', horizontal='left')
            right_align = Alignment(horizontal='right', vertical='center')
            total_rows = len(final_rows_to_write_local)
            update_interval = 50

            for r_idx, row_data in enumerate(final_rows_to_write_local):
                if r_idx % update_interval == 0 or r_idx == total_rows - 1:
                    percent_done = min(100, int((r_idx + 1) / total_rows * 100))
                    self.status_label.configure(text=f"Writing data: {percent_done}% ({r_idx + 1}/{total_rows} rows)"); self.update()

                row_num_excel = r_idx + 1
                is_main_making = len(row_data) > FORMAT_COL-1 and str(row_data[FORMAT_COL-1]).strip() == "Making"

                for c_idx, cell_value in enumerate(row_data):
                    col_idx = c_idx + 1
                    try:
                        cell = ws.cell(row=row_num_excel, column=col_idx)
                        if isinstance(cell, openpyxl.cell.cell.MergedCell):
                             for merged_range in ws.merged_cells.ranges:
                                 if cell.coordinate in merged_range:
                                     top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                                     if cell.coordinate == top_left_cell.coordinate:
                                         top_left_cell.value = cell_value
                                     break
                             continue

                        cell.value = cell_value
                        if col_idx == LABEL_COL: cell.alignment = wrap_align
                        if is_main_making and col_idx == LABEL_COL: cell.fill = making_fill
                        if isinstance(cell_value, (int, float)) and col_idx in [ID_COL, WEIGHT_COL]:
                             cell.alignment = right_align
                             if isinstance(cell_value, int): cell.number_format = '0'

                    except Exception as write_err:
                        self.log_message(f"Warning: Error writing/styling cell ({row_num_excel},{col_idx}): {type(write_err).__name__}", level="WARNING")

            LABEL_COL_LETTER = get_column_letter(LABEL_COL)
            if LABEL_COL_LETTER: ws.column_dimensions[LABEL_COL_LETTER].width = 45

            write_time = time.time() - write_start_time
            self.log_message(f"Finished writing data ({write_time:.2f}s).", level="INFO")

            self.status_label.configure(text=f"Saving workbook to file..."); self.update()
            self.log_message(f"Saving workbook to: {new_file_path}", level="INFO")
            save_start_time = time.time()
            try:
                workbook.save(new_file_path)
            except Exception as save_err:
                 try: workbook.close()
                 except: pass
                 raise save_err

            save_time = time.time() - save_start_time
            self.log_message(f"Workbook saved successfully ({save_time:.2f}s).", level="SUCCESS")
            total_time = time.time() - total_start_time
            final_message = f"File saved successfully ({total_time:.2f}s): {new_filename}"
            level = "SUCCESS" # Set success level
            tkmb.showinfo("Save Successful", f"File saved as:\n{new_file_path}", parent=self)

        except PermissionError:
            msg = f"Cannot save '{new_filename}'. Permission Denied. (Is the file open or write-protected?)"
            self.log_message(msg, level="ERROR")
            self.show_error_popup("Permission Error", msg)
            final_message = "Save failed: Permission Error."
            level = "ERROR"
        except Exception as e:
            error_type_name = type(e).__name__
            msg = f"Error during Save As: [{error_type_name}] {e}"
            self.log_message(msg, level="ERROR")
            self.show_error_popup("Save Error", f"{msg}\nCheck log/console.")
            print(traceback.format_exc())
            final_message = "Save failed: Unexpected Error."
            level = "ERROR"
        finally:
            if workbook:
                try: workbook.close()
                except: pass
            self._end_processing(final_status_message=final_message if 'final_message' in locals() else "Save process finished.",
                                 level=level) # Use the determined level

    # --- ฟังก์ชันจัดการคลิกขวาและการตั้งค่า Direction --- (เหมือนเดิม)
    def on_survey_tree_right_click(self, event):
        row_id = self.survey_tree.identify_row(event.y)
        if not row_id: return
        column_id = self.survey_tree.identify_column(event.x)
        try: column_index = int(column_id.replace('#', '')) - 1
        except ValueError: return

        DIRECTION_COL_INDEX = 4
        item_tags = self.survey_tree.item(row_id, "tags")
        is_parent = 'parent_row' in item_tags

        if column_index == DIRECTION_COL_INDEX and is_parent:
            self.survey_tree.selection_set(row_id)
            context_menu = tk.Menu(self, tearoff=0)
            context_menu.add_command(label="Scale น้อยดี(-)", command=lambda r=row_id: self._set_direction(r, "Scale น้อยดี(-)"))
            context_menu.add_command(label="Scale มากดี(+)", command=lambda r=row_id: self._set_direction(r, "Scale มากดี(+)"))
            context_menu.add_separator()
            context_menu.add_command(label="Clear Direction", command=lambda r=row_id: self._set_direction(r, ""))
            context_menu.tk_popup(event.x_root, event.y_root)

    def _set_direction(self, row_id, new_direction):
        if not self.survey_tree.exists(row_id): return
        self.survey_tree.set(row_id, "Direction", new_direction)
        if row_id in self.scales_data_store and not self.scales_data_store[row_id].get('is_making'):
            self.scales_data_store[row_id]['direction'] = new_direction
            self.log_message(f"Set Direction for {row_id} to '{new_direction if new_direction else 'None'}'", level="INFO")
            current_status = self.survey_tree.set(row_id, "Condition Status")
            new_status = ""
            if not new_direction and current_status != "Making Generated": new_status = "ไม่ทำ Making"
            elif new_direction and current_status == "ไม่ทำ Making": new_status = ""
            elif new_direction and current_status != "Making Generated": new_status = current_status

            if new_status != current_status and current_status != "Making Generated":
                self.survey_tree.set(row_id, "Condition Status", new_status)
        else:
             self.log_message(f"Could not set direction for {row_id} (not found in store or is making item).", level="WARNING")

    # --- Popup Helper Functions --- (เหมือนเดิม)
    def show_error_popup(self, title, message): tkmb.showerror(title, message, parent=self)
    def show_warning_popup(self, title, message): tkmb.showwarning(title, message, parent=self)

# <<< START OF CHANGES >>>
# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
def run_this_app(working_dir=None): # ชื่อฟังก์ชันนี้จะถูกใช้ใน Launcher
    """
    ฟังก์ชันหลักสำหรับสร้างและรัน QuotaSamplerApp.
    """
    print(f"--- QUOTA_SAMPLER_INFO: Starting 'QuotaSamplerApp' via run_this_app() ---")
    try:
        # --- โค้ดที่ย้ายมาจาก if __name__ == "__main__": เดิมจะมาอยู่ที่นี่ ---
        # --- ส่วน Run Application ---
    #if __name__ == "__main__":
        if getattr(sys, 'frozen', False):
            # When running as a bundled app (e.g., via PyInstaller)
            application_path = sys._MEIPASS
            print(f"Running as frozen executable. MEIPASS: {application_path}")
        else:
            # When running as a normal Python script
            application_path = os.path.dirname(os.path.abspath(__file__))
            print(f"Running as script. Application path: {application_path}")

        template_filename = "template.xlsx"
        template_file_path_main = os.path.join(application_path, template_filename)
        print(f"Attempting to use template file at: {template_file_path_main}")

        # Check if template exists before starting GUI (เหมือนเดิม)
        if not os.path.exists(template_file_path_main):
            root_temp = tk.Tk(); root_temp.withdraw()
            tkmb.showerror("Template Missing", f"Template file '{template_filename}' not found in the application directory:\n{application_path}\n\nThe application cannot continue.")
            root_temp.destroy()
            sys.exit(f"Error: Template file '{template_filename}' not found.")


        app = ExcelScaleExtractorApp()
        app.template_file_path = template_file_path_main # Pass template path to app instance
        app.mainloop()

        print(f"--- QUOTA_SAMPLER_INFO: QuotaSamplerApp mainloop finished. ---")

    except Exception as e:
        # ดักจับ Error ที่อาจเกิดขึ้นระหว่างการสร้างหรือรัน App
        print(f"QUOTA_SAMPLER_ERROR: An error occurred during QuotaSamplerApp execution: {e}")
        # แสดง Popup ถ้ามีปัญหา
        if 'root' not in locals() or not root.winfo_exists(): # สร้าง root ชั่วคราวถ้ายังไม่มี
            root_temp = tk.Tk()
            root_temp.withdraw()
            messagebox.showerror("Application Error (Quota Sampler)",
                               f"An unexpected error occurred:\n{e}", parent=root_temp)
            root_temp.destroy()
        else:
            messagebox.showerror("Application Error (Quota Sampler)",
                               f"An unexpected error occurred:\n{e}", parent=root) # ใช้ root ที่มีอยู่ถ้าเป็นไปได้
        sys.exit(f"Error running QuotaSamplerApp: {e}") # อาจจะ exit หรือไม่ก็ได้ ขึ้นกับการออกแบบ


# --- ส่วน Run Application เมื่อรันไฟล์นี้โดยตรง (สำหรับ Test) ---
if __name__ == "__main__":
    print("--- Running QuotaSamplerApp.py directly for testing ---")
    # (ถ้ามีการตั้งค่า DPI ด้านบน มันจะทำงานอัตโนมัติ)

    # เรียกฟังก์ชัน Entry Point ที่เราสร้างขึ้น
    run_this_app()

    print("--- Finished direct execution of QuotaSamplerApp.py ---")
# <<< END OF CHANGES >>>

