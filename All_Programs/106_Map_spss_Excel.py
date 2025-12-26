import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import sys
import os
import re
import difflib
from openpyxl.styles import PatternFill

print(f"Python ที่กำลังใช้งานอยู่ที่: {sys.executable}") 

# --- ฟังก์ชันสำหรับจัดหน้าต่างให้อยู่กลางจอ (วางไว้ส่วนบนของไฟล์) ---
def center_window(win):
    """ฟังก์ชันสำหรับจัดหน้าต่าง (Tk หรือ Toplevel) ให้อยู่กึ่งกลางจอ"""
    win.update_idletasks() # ต้องเรียกใช้เพื่อบังคับให้ tkinter คำนวณขนาดที่แท้จริงของหน้าต่างก่อน
    width = win.winfo_width()
    height = win.winfo_height()
    x = (win.winfo_screenwidth() // 2) - (width // 2)
    y = (win.winfo_screenheight() // 2) - (height // 2)
    win.geometry(f'{width}x{height}+{x}+{y}')

class RenameConfirmationWindow(tk.Toplevel):
    def __init__(self, parent, potential_matches):
        super().__init__(parent)
        self.title("ยืนยันการเปลี่ยนชื่อตัวแปร")
        self.geometry("700x450") # เพิ่มความสูงเล็กน้อย
        self.transient(parent)
        self.grab_set()
        self.potential_matches = potential_matches
        self.confirmed = False

        # --- การตั้งค่า Style และพื้นหลังสีขาว ---
        self.configure(bg='white')
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure('.', background='white', foreground='black', font=('Tahoma', 9))
        style.configure('TFrame', background='white')
        style.configure('TLabel', background='white')
        style.configure('TButton', padding=5, font=('Tahoma', 10, 'bold'))
        style.configure('Treeview', rowheight=25, fieldbackground='white')
        style.configure('Treeview.Heading', font=('Tahoma', 10, 'bold'))

        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="โปรแกรมพบว่าชื่อตัวแปรบางส่วนอาจไม่ตรงกันแต่มีส่วนท้ายคล้ายกัน", foreground="blue", font=('Tahoma', 10, 'bold')).pack(pady=(0, 10))
        ttk.Label(main_frame, text="กรุณายืนยันการเปลี่ยนชื่อในไฟล์ Excel ให้ตรงกับ SPSS:").pack(pady=(0, 10))

        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        self.tree = ttk.Treeview(tree_frame, columns=("excel", "spss"), show="headings")
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.heading("excel", text="ชื่อปัจจุบันใน Excel")
        self.tree.heading("spss", text="ชื่อใหม่ที่แนะนำ (จาก SPSS)")
        self.tree.column("excel", anchor=tk.W, width=350)
        self.tree.column("spss", anchor=tk.W, width=300)

        # --- เพิ่มสีสลับแถวในตาราง ---
        self.tree.tag_configure('oddrow', background='#EAF3FF') # สีฟ้าอ่อน
        self.tree.tag_configure('evenrow', background='white')
        
        for i, (excel_name, spss_name) in enumerate(self.potential_matches):
            tag = 'oddrow' if i % 2 else 'evenrow'
            self.tree.insert("", "end", values=(excel_name, spss_name), tags=(tag,))

        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        ttk.Button(button_frame, text="ยืนยันและเปลี่ยนชื่อทั้งหมด", command=self.confirm_and_close).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="ยกเลิก", command=self.destroy).pack(side=tk.RIGHT, padx=5)
        # --- [NEW] ตัวแปรสำหรับเก็บ Metadata (การตั้งค่า) ของ SPSS Final ---
        self.spss_final_meta = None
        # --- จัดหน้าต่างให้อยู่กลางจอ ---
        center_window(self)

    def confirm_and_close(self):
        self.confirmed = True
        self.destroy()


class MissingVariableManagerWindow(tk.Toplevel):
    def __init__(self, parent, missing_vars):
        super().__init__(parent)
        self.title("ตรวจพบตัวแปรที่ไม่มีในไฟล์ Excel")
        self.geometry("700x500") #ปรับขนาดให้เหมาะสม
        self.transient(parent)
        self.grab_set()
        self.missing_vars = missing_vars
        self.confirmed = False
        self.decisions = {}

        # --- 1. ตั้งค่า Style ให้พื้นหลังเป็นสีขาวทั้งหมด ---
        self.configure(bg='white')
        style = ttk.Style(self)
        style.theme_use("clam") # ใช้ theme ที่ปรับแต่งง่าย

        # Style ทั่วไป
        style.configure('.', background='white', foreground='black', font=('Tahoma', 10))
        style.configure('TFrame', background='white')
        style.configure('TLabel', background='white')
        style.configure('Header.TLabel', font=('Tahoma', 10, 'bold'))

        # Style ปุ่ม
        style.configure('Green.TButton', background='#28a745', foreground='white', borderwidth=0, font=('Tahoma', 10, 'bold'))
        style.map('Green.TButton', background=[('active', '#32c955')]) # สีตอนเอาเมาส์ชี้

        style.configure('Outline.TButton', background='white', foreground='#6c757d', borderwidth=1, relief="solid", font=('Tahoma', 10))
        style.map('Outline.TButton',
                  foreground=[('active', 'white')],
                  background=[('active', '#6c757d')],
                  relief=[('pressed', 'sunken')])

        # Style ของ Combobox (Dropdown) ให้เป็นพื้นหลังสีขาว
        combo_options = {
            'fieldbackground': 'white',
            'background': 'white',
            'bordercolor': '#ACACAC',
            'darkcolor': 'white',
            'lightcolor': 'white',
            'arrowcolor': '#555555',
            'padding': (8, 4),
            'font': ('Tahoma', 10)
        }
        # Style แยกตามสีของข้อความ
        style.configure('RedText.TCombobox', foreground='#D32F2F', **combo_options) # สีแดงสำหรับ "ลบทิ้ง"
        style.configure('BlueText.TCombobox', foreground='#1E88E5', **combo_options) # สีน้ำเงิน
        style.configure('OrangeText.TCombobox', foreground='#F57C00', **combo_options) # สีส้ม

        # --- 2. Layout หลักของหน้าต่าง ---
        main_frame = ttk.Frame(self, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # ส่วนหัวข้อ
        ttk.Label(main_frame, text="ตรวจพบตัวแปรใน SPSS Final ที่ไม่มีในไฟล์ Excel", foreground="#C0392B", font=('Tahoma', 14, 'bold')).pack(pady=(0, 5))
        ttk.Label(main_frame, text="กรุณาเลือกการกระทำสำหรับแต่ละตัวแปร:", font=('Tahoma', 10)).pack(pady=(0, 15))

        # --- 3. Scrollable Frame (ส่วนที่เลื่อนได้) ---
        canvas_frame = ttk.Frame(main_frame)
        canvas_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # ใช้ Canvas ของ tk ปกติ เพราะ ttk.Canvas ไม่มี bg
        canvas = tk.Canvas(canvas_frame, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas) # Frame ที่จะใส่เนื้อหาจริงๆ

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # --- 4. การจัดกึ่งกลาง: ให้คอลัมน์ซ้ายสุด(0)และขวาสุด(3)ขยายตัวเพื่อดันเนื้อหาเข้ากลาง ---
        scrollable_frame.columnconfigure((0, 3), weight=1) # คอลัมน์เปล่าสำหรับดัน
        scrollable_frame.columnconfigure((1, 2), weight=0) # คอลัมน์เนื้อหา

        # สร้างรายการตัวแปรและ Combobox
        self.comboboxes = {}
        options = ["ลบทิ้ง", "เก็บไว้ (คงตำแหน่งเดิม)", "เก็บไว้ (ต่อท้ายสุด)"]
        
        for i, var in enumerate(self.missing_vars):
            row = i # เริ่มจากแถวที่ 0
            
            # Label ชื่อตัวแปร (วางในคอลัมน์ที่ 1)
            ttk.Label(scrollable_frame, text=var, font=('Consolas', 11, 'bold')).grid(row=row, column=1, sticky='w', pady=6, padx=(0, 20))
            
            # Combobox (วางในคอลัมน์ที่ 2)
            combo = ttk.Combobox(scrollable_frame, values=options, state="readonly", width=25)
            combo.set("ลบทิ้ง")
            combo.configure(style='RedText.TCombobox') # ตั้งค่า Style เริ่มต้น
            combo.bind("<<ComboboxSelected>>", self.on_combo_select)
            combo.grid(row=row, column=2, sticky='w', pady=6, padx=(20, 0))
            
            self.comboboxes[var] = combo

        # --- 5. ส่วนปุ่มด้านล่าง ---
        button_frame = ttk.Frame(self) # สร้าง Frame ใหม่นอก main_frame เพื่อให้ติดขอบล่าง
        button_frame.pack(fill=tk.X, padx=20, pady=(10, 20))

        ttk.Button(button_frame, text="ยืนยันและดำเนินการต่อ", command=self.confirm_and_close, style='Green.TButton').pack(side=tk.LEFT, ipadx=10, ipady=4)
        ttk.Button(button_frame, text="ยกเลิก", command=self.destroy, style='Outline.TButton').pack(side=tk.RIGHT, ipadx=10, ipady=3)

        center_window(self)

    def on_combo_select(self, event):
        """เปลี่ยนสีข้อความใน Combobox ตามตัวเลือก"""
        widget = event.widget
        selected_value = widget.get()
        if selected_value == "ลบทิ้ง":
            widget.configure(style='RedText.TCombobox')
        elif selected_value == "เก็บไว้ (คงตำแหน่งเดิม)":
            widget.configure(style='BlueText.TCombobox')
        else: # "เก็บไว้ (ต่อท้ายสุด)"
            widget.configure(style='OrangeText.TCombobox')
        self.focus_set() # คืน focus กลับให้หน้าต่างหลัก

    def confirm_and_close(self):
        """บันทึกการตัดสินใจและปิดหน้าต่าง"""
        for var, combo in self.comboboxes.items():
            self.decisions[var] = combo.get()
        self.confirmed = True
        self.destroy()

class SPSSExcelMapper(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("โปรแกรม แปลงCE Otherจาก Edit V2")
        self.geometry("850x650") 
        self.spss_df = None
        self.spss_vars = []
        self.excel_df = None
        self.excel_file_path = None
        self.df_ready_to_save = None
        self.selected_sheet = tk.StringVar()
        self.key_column = None
        self.missing_in_excel = []
        self.extra_in_excel = []
        self.renamed_matches = []
        self.dropped_vars_log = []
        self.df_code_sheet = None
        self.vars_added_log = []
        
        self._create_widgets()
        self._configure_styles()
        
        # --- จัดหน้าต่างให้อยู่กลางจอ ---
        center_window(self)

    def _configure_styles(self):
        """ฟังก์ชันสำหรับตั้งค่า Style ของ Widget ทั้งหมด"""
        BG_COLOR = '#ffffff'      # สีขาว
        TEXT_COLOR = '#000000'    # สีดำ
        BLUE = '#007bff'
        LIGHT_BLUE = '#3395ff'    # สีฟ้าสำหรับ Hover
        GREEN = '#28a745'
        LIGHT_GREEN = '#32c955'   # สีเขียวสำหรับ Hover
        ORANGE = '#fd7e14'
        LIGHT_ORANGE = '#ff9d47'  # สีส้มสำหรับ Hover
        
        self.configure(background=BG_COLOR)
        
        style = ttk.Style(self)
        style.theme_use("clam") 

        style.configure('.', 
                        background=BG_COLOR, 
                        foreground=TEXT_COLOR, 
                        fieldbackground=BG_COLOR, 
                        font=('Tahoma', 9))
                        
        style.configure('TFrame', background=BG_COLOR)
        style.configure('TLabel', background=BG_COLOR, foreground=TEXT_COLOR, font=('Tahoma', 9))
        style.configure('TLabelframe', background=BG_COLOR)
        style.configure('TLabelframe.Label', background=BG_COLOR, foreground=TEXT_COLOR, font=('Tahoma', 9, 'bold'))
        style.configure('TButton', padding=5, font=('Tahoma', 10, 'bold'))
        style.configure('Treeview', rowheight=25, fieldbackground=BG_COLOR)
        style.configure('Treeview.Heading', font=('Tahoma', 10, 'bold'))
        
        # ---> [ปรับปรุง] เพิ่ม anchor='w' เพื่อจัดข้อความชิดซ้าย <---
        style.configure('Blue.TButton', background=BLUE, foreground='white', anchor='w', padding=(10, 5))
        style.map('Blue.TButton', background=[('active', LIGHT_BLUE), ('disabled', '#cccccc')])

        style.configure('Orange.TButton', background=ORANGE, foreground='white')
        style.map('Orange.TButton', background=[('active', LIGHT_ORANGE), ('disabled', '#cccccc')])
        
        style.configure('Green.TButton', background=GREEN, foreground='white')
        style.map('Green.TButton', background=[('active', LIGHT_GREEN), ('disabled', '#cccccc')])

        self.tree.tag_configure('oddrow', background='#EAF3FF') 
        self.tree.tag_configure('evenrow', background=BG_COLOR)
        self.tree.tag_configure('renamed', foreground='#006400') # DarkGreen

    def _create_widgets(self):
            main_frame = ttk.Frame(self, padding="10")
            main_frame.pack(fill=tk.BOTH, expand=True)

            # --- [ปรับปรุง] เปลี่ยน Layout การโหลดไฟล์เป็นแบบ Grid (บน-ล่าง) ---
            load_frame = ttk.LabelFrame(main_frame, text="ขั้นตอนที่ 1 และ 2: โหลดไฟล์")
            load_frame.pack(fill=tk.X, pady=5)

            # กำหนดค่า Grid ให้คอลัมน์ที่ 1 (Label) ขยายตามขนาดหน้าต่าง
            load_frame.columnconfigure(0, weight=0) # คอลัมน์สำหรับปุ่ม
            load_frame.columnconfigure(1, weight=1) # คอลัมน์สำหรับสถานะ

            # แถวที่ 1: โหลด SPSS
            self.btn_load_spss = ttk.Button(load_frame, text="1. โหลดไฟล์ SPSS Original!!", command=self.load_spss, style='Blue.TButton')
            self.btn_load_spss.grid(row=0, column=0, sticky="ew", padx=5, pady=(5, 2))
            
            self.lbl_spss_status = ttk.Label(load_frame, text="ยังไม่โหลด", foreground="red")
            self.lbl_spss_status.grid(row=0, column=1, sticky="w", padx=10, pady=(5, 2))

            # แถวที่ 2: โหลด Excel
            self.btn_load_excel = ttk.Button(load_frame, text="2. โหลดไฟล์ Excel CE Other For Edit", command=self.load_excel, style='Blue.TButton')
            self.btn_load_excel.grid(row=1, column=0, sticky="ew", padx=5, pady=(2, 5))
            
            self.lbl_excel_status = ttk.Label(load_frame, text="ยังไม่โหลด", foreground="red")
            self.lbl_excel_status.grid(row=1, column=1, sticky="w", padx=10, pady=(2, 5))
            # --- สิ้นสุดส่วนปรับปรุง ---

            sheet_frame = ttk.LabelFrame(main_frame, text="ขั้นตอนที่ 3 และ 4: เลือกและMap ตัวแปร")
            sheet_frame.pack(fill=tk.X, pady=(10, 5))
            ttk.Label(sheet_frame, text="เลือก Sheet:").pack(side=tk.LEFT, padx=5, pady=5)
            self.sheet_menu = ttk.OptionMenu(sheet_frame, self.selected_sheet, "...")
            self.sheet_menu.pack(side=tk.LEFT, padx=5, pady=5)

            self.btn_check_sheet = ttk.Button(sheet_frame, text="4. เริ่ม Map ตัวแปร", command=self.perform_analysis, state=tk.DISABLED, style='Orange.TButton')
            self.btn_check_sheet.pack(side=tk.LEFT, padx=10, pady=5)
            self.lbl_sheet_status = ttk.Label(sheet_frame, text="กรุณาเลือก Sheet", foreground="blue")
            self.lbl_sheet_status.pack(side=tk.LEFT, padx=5, pady=5)
            self.selected_sheet.trace("w", self.on_sheet_selected)

            compare_frame = ttk.LabelFrame(main_frame, text="ผลการวิเคราะห์ (ตัวแปรที่แตกต่างกัน)")
            compare_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 5))
            
            tree_container = ttk.Frame(compare_frame)
            tree_container.pack(fill=tk.BOTH, expand=True)
            
            tree_scrollbar = ttk.Scrollbar(tree_container, orient="vertical")
            self.tree = ttk.Treeview(tree_container, columns=("status", "variable"), show="headings", yscrollcommand=tree_scrollbar.set)
            tree_scrollbar.config(command=self.tree.yview)
            tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            self.tree.heading("status", text="สถานะ")
            self.tree.heading("variable", text="ชื่อตัวแปร")
            self.tree.column("status", anchor=tk.W, width=250)
            self.tree.column("variable", anchor=tk.W)
            self.tree.tag_configure('missing', foreground='red')
            self.tree.tag_configure('extra', foreground='darkorange')

            action_frame = ttk.LabelFrame(main_frame, text="ขั้นตอนที่ 5: จัดเรียงและบันทึก")
            action_frame.pack(fill=tk.X, pady=5)
            self.btn_update_all = ttk.Button(action_frame, text="Get SPSS Final", command=self.perform_mapping, state=tk.DISABLED, style='Green.TButton')
            self.btn_update_all.pack(side=tk.LEFT, padx=5, pady=5)

            save_buttons_frame = ttk.Frame(action_frame)
            save_buttons_frame.pack(side=tk.RIGHT)

            # --- [NEW] ปุ่มบันทึกเป็น SPSS ---
            self.btn_save_excel = ttk.Button(save_buttons_frame, text="บันทึกเป็น Excel For Lychee", command=self.save_excel, state=tk.DISABLED, style='Green.TButton')
            self.btn_save_excel.pack(side=tk.RIGHT, padx=5, pady=5)
            
            self.btn_save_spss = ttk.Button(save_buttons_frame, text="บันทึก SPSS For Reporter", command=self.save_spss, state=tk.DISABLED, style='Green.TButton')
            self.btn_save_spss.pack(side=tk.RIGHT, padx=5, pady=5)

    def _get_excel_col_name(self, n):
        name = ""
        while n >= 0:
            name = chr(n % 26 + 65) + name
            n = n // 26 - 1
        return name

    def clear_analysis_results(self):
            """ล้างผลการวิเคราะห์และสถานะทั้งหมดเพื่อเริ่มใหม่"""
            for i in self.tree.get_children(): self.tree.delete(i)
            self.btn_update_all.config(state=tk.DISABLED)
            # ปิดปุ่มบันทึกทั้งสอง
            self.btn_save_excel.config(state=tk.DISABLED)
            self.btn_save_spss.config(state=tk.DISABLED)
            
            self.df_ready_to_save = None
            self.excel_df = None
            self.key_column = None
            self.missing_in_excel = []
            self.extra_in_excel = []
            self.renamed_matches = []
            self.dropped_vars_log = []
            self.df_code_sheet = None
            self.vars_added_log = []

    def on_sheet_selected(self, *args):
        sheet_name = self.selected_sheet.get()
        if sheet_name:
            self.lbl_sheet_status.config(text=f"เลือก Sheet '{sheet_name}' แล้ว", foreground="green")
            self.btn_check_sheet.config(state=tk.NORMAL)
            self.lbl_excel_status.config(text="ยังไม่ได้วิเคราะห์ชีตนี้", foreground="red")
            self.clear_analysis_results()
        else:
            self.lbl_sheet_status.config(text="กรุณาเลือก Sheet", foreground="blue")
            self.btn_check_sheet.config(state=tk.DISABLED)

    def load_spss(self):
        file_path = filedialog.askopenfilename(filetypes=[("SPSS Data File", "*.sav")])
        if not file_path: return
        try:
            self.spss_df = pd.read_spss(file_path)
            self.spss_vars = list(self.spss_df.columns)
            if not self.spss_vars:
                messagebox.showerror("ผิดพลาด", "ไฟล์ SPSS ไม่มีตัวแปร", parent=self)
                return
            self.key_column = self.spss_vars[0] 
            self.lbl_spss_status.config(text=f"โหลดแล้ว ({len(self.spss_vars)} ตัวแปร)", foreground="green")
            messagebox.showinfo("โหลด SPSS สำเร็จ", f"ไฟล์ SPSS โหลดสำเร็จ\nโปรแกรมจะใช้ '{self.key_column}' เป็นคอลัมน์หลัก (Key)", parent=self)
        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถโหลดไฟล์ SPSS:\n{e}", parent=self)
            self.spss_df = None
            self.spss_vars = []

    def load_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if not file_path: return
        self.excel_file_path = file_path
        try:
            xls = pd.ExcelFile(self.excel_file_path)
            self.lbl_excel_status.config(text=f"โหลดสำเร็จ", foreground="green")
            menu = self.sheet_menu["menu"]
            menu.delete(0, "end")
            sheet_names = xls.sheet_names
            if sheet_names:
                for sheet in sheet_names:
                    menu.add_command(label=sheet, command=lambda value=sheet: self.selected_sheet.set(value))
                self.selected_sheet.set(sheet_names[0])
            else:
                self.selected_sheet.set("")
        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถโหลดไฟล์ Excel:\n{e}", parent=self)

    def load_spss_final(self):
        file_path = filedialog.askopenfilename(filetypes=[("SPSS Data File", "*.sav")])
        if not file_path: return
        try:
            self.spss_final_df = pd.read_spss(file_path)
            self.spss_final_vars = list(self.spss_final_df.columns)
            if not self.spss_final_vars:
                messagebox.showerror("ผิดพลาด", "ไฟล์ SPSS Final ไม่มีตัวแปร", parent=self)
                return
            
            if self.key_column not in self.spss_final_vars:
                messagebox.showerror("ผิดพลาด", f"ไฟล์ SPSS Final ไม่มีคอลัมน์หลัก '{self.key_column}'\nกรุณาใช้ไฟล์ที่มีโครงสร้าง Key เดียวกัน", parent=self)
                self.spss_final_df = None
                self.spss_final_vars = []
                return

            self.lbl_spss_final_status.config(text=f"โหลดแล้ว ({len(self.spss_final_vars)} ตัวแปร)", foreground="green")
            messagebox.showinfo("โหลด SPSS Final สำเร็จ", "ไฟล์ SPSS Final โหลดสำเร็จแล้ว", parent=self)
        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"ไม่สามารถโหลดไฟล์ SPSS Final:\n{e}", parent=self)
            self.spss_final_df = None
            self.spss_final_vars = []

    def perform_analysis(self, analyze_again=False):
        if not analyze_again:
            self.clear_analysis_results()
        if not all([self.spss_df is not None, self.excel_file_path, self.selected_sheet.get()]):
            messagebox.showwarning("ข้อมูลไม่ครบ", "กรุณาโหลดไฟล์ SPSS และเลือก Sheet ในไฟล์ Excel ให้เรียบร้อยก่อน", parent=self)
            return
        if not self.spss_vars:
            messagebox.showerror("SPSS Error", "ไม่พบชื่อตัวแปรในข้อมูล SPSS กรุณาโหลดไฟล์ใหม่", parent=self)
            return
        self.key_column = self.spss_vars[0]
        self.lbl_sheet_status.config(text="กำลังวิเคราะห์ กรุณารอสักครู่...", foreground="orange")
        self.update_idletasks()
        try:
            if not analyze_again:
                sheet_name = self.selected_sheet.get()
                temp_df = pd.read_excel(self.excel_file_path, sheet_name=sheet_name, header=None)
                header_row_index = -1
                for index, row in temp_df.head(10).iterrows():
                    row_values_str = [str(c).strip().lower() for c in row.values]
                    if self.key_column.lower() in row_values_str:
                        header_row_index = index
                        break
                if header_row_index == -1:
                    messagebox.showerror("ผิดพลาดร้ายแรง", f"ไม่พบคอลัมน์หลัก '{self.key_column}' ใน 10 แถวแรกของไฟล์ Excel\nกรุณาตรวจสอบ Header", parent=self)
                    self.lbl_sheet_status.config(text="ผิดพลาด: หา Header ไม่พบ", foreground="red")
                    return
                self.excel_df = pd.read_excel(self.excel_file_path, sheet_name=sheet_name, header=header_row_index)

                # ตัด 3 ตัวแปรเจ้าปัญหาออกไปโดยตรงแบบเงียบๆ
                explicit_drop_list = [
                    'TrackedAnswerCodeNum',
                    'TrackedAnswerCodePercent',
                    'FlagsByAnswerCode'
                ]
                cols_found_to_drop = [col for col in explicit_drop_list if col in self.excel_df.columns]

                if cols_found_to_drop:
                    self.excel_df.drop(columns=cols_found_to_drop, inplace=True)
                    print(f"Explicitly dropped columns (silently): {cols_found_to_drop}")
                    
                    # ---> [ปรับปรุง] เอา Messagebox แจ้งเตือนออกตามที่ผู้ใช้ต้องการ <---
                    # messagebox.showinfo(
                    #     "แจ้งเตือนการตัดตัวแปร",
                    #     f"ได้ทำการตัดตัวแปรต่อไปนี้ออกจากไฟล์ Excel ก่อนทำการเปรียบเทียบ:\n\n- {', '.join(cols_found_to_drop)}",
                    #     parent=self
                    # )
                
                # เรียกใช้ฟังก์ชันจัดการคอลัมน์ Code และลบทิ้งก่อนเปรียบเทียบ
                code_columns_to_drop = self.handle_code_columns()
                if code_columns_to_drop:
                    self.excel_df.drop(columns=code_columns_to_drop, inplace=True)
                    print(f"Dropped {len(code_columns_to_drop)} code columns for comparison.")

            if not analyze_again:
                original_excel_cols = list(self.excel_df.columns)
                lower_cols = [str(c).lower() for c in original_excel_cols]
                
                try:
                    start_index = lower_cols.index('sbjnum') 
                except ValueError:
                    messagebox.showerror("หาคอลัมน์ไม่พบ", "ไม่พบคอลัมน์ 'Sbjnum' ในไฟล์ Excel\nกรุณาตรวจสอบชื่อคอลัมน์ (ต้องเป็น Sbjnum)", parent=self)
                    self.lbl_sheet_status.config(text="ผิดพลาด: ไม่พบ Sbjnum", foreground="red")
                    return
                
                try:
                    end_index = lower_cols.index('filter') 
                except ValueError:
                    messagebox.showerror("หาคอลัมน์ไม่พบ", "ไม่พบคอลัมน์ 'filter' ในไฟล์ Excel\nกรุณาตรวจสอบว่ามีคอลัมน์ filter อยู่ในข้อมูล", parent=self)
                    self.lbl_sheet_status.config(text="ผิดพลาด: ไม่พบ filter", foreground="red")
                    return

                cols_to_keep_before = original_excel_cols[:start_index + 1]
                cols_to_keep_after = original_excel_cols[end_index:]
                cols_to_keep = cols_to_keep_before + cols_to_keep_after
                self.excel_df = self.excel_df[cols_to_keep].copy()

            excel_vars_trimmed = list(self.excel_df.columns)
            
            self.lbl_excel_status.config(
                text=f"วิเคราะห์แล้ว ({len(excel_vars_trimmed)} ตัวแปร)", 
                foreground="green"
            )
            
            if len(self.spss_vars) != len(excel_vars_trimmed) and not analyze_again:
                error_msg = (f"จำนวนตัวแปรไม่ตรงกันหลังตัดคอลัมน์!\n\n"
                            f"SPSS มี: {len(self.spss_vars)} ตัวแปร\n"
                            f"Excel (หลังตัดส่วนกลางและ CODE ออก) มี: {len(excel_vars_trimmed)} ตัวแปร\n\n"
                            f"โปรแกรมไม่สามารถเทียบตำแหน่งต่อตำแหน่งได้ กรุณาตรวจสอบไฟล์")
                messagebox.showerror("จำนวนตัวแปรไม่ตรงกัน", error_msg, parent=self)
                self.lbl_sheet_status.config(text="ผิดพลาด: จำนวนตัวแปรไม่ตรงกัน", foreground="red")
                return
                
            potential_matches = []
            if not analyze_again:
                for i in range(len(self.spss_vars)):
                    spss_var = self.spss_vars[i]
                    excel_var = excel_vars_trimmed[i]
                    if spss_var != excel_var:
                        potential_matches.append((excel_var, spss_var))
            
            if potential_matches:
                rename_window = RenameConfirmationWindow(self, potential_matches)
                self.wait_window(rename_window)
                if rename_window.confirmed:
                    self.renamed_matches = potential_matches
                    rename_dict = {excel: spss for excel, spss in potential_matches}
                    self.excel_df.rename(columns=rename_dict, inplace=True)
                    messagebox.showinfo("สำเร็จ", f"เปลี่ยนชื่อตัวแปรจำนวน {len(rename_dict)} ตัวเรียบร้อยแล้ว\nกำลังแสดงผลการเปลี่ยนแปลง...", parent=self)
                    self.perform_analysis(analyze_again=True)
                    return
                else:
                    messagebox.showwarning("ยกเลิก", "การเปลี่ยนชื่อถูกยกเลิก โปรแกรมจะแสดงผลตามข้อมูลปัจจุบัน", parent=self)

            for i in self.tree.get_children(): self.tree.delete(i)
            
            if self.renamed_matches and analyze_again:
                self.tree.heading("status", text="ชื่อเดิมใน Excel")
                self.tree.heading("variable", text="ชื่อใหม่ที่ถูกเปลี่ยน")
                
                for i, (old_name, new_name) in enumerate(self.renamed_matches):
                    tag = 'oddrow' if i % 2 else 'evenrow'
                    self.tree.insert("", "end", values=(old_name, new_name), tags=('renamed', tag))

                self.lbl_sheet_status.config(text=f"เปลี่ยนชื่อ {len(self.renamed_matches)} ตัวแปรสำเร็จ! พร้อมสำหรับขั้นตอนต่อไป", foreground="green")
                self.btn_update_all.config(state=tk.NORMAL)

            else:
                self.tree.heading("status", text="สถานะ")
                self.tree.heading("variable", text="ชื่อตัวแปร")

                current_excel_vars = set(self.excel_df.columns)
                spss_vars_set = set(self.spss_vars)
                missing_in_excel = sorted(list(spss_vars_set - current_excel_vars))
                extra_in_excel = sorted(list(current_excel_vars - spss_vars_set))

                if not missing_in_excel and not extra_in_excel:
                    self.lbl_sheet_status.config(text="ยอดเยี่ยม! ตัวแปรทั้งหมดตรงกัน", foreground="green")
                    messagebox.showinfo("ผลการวิเคราะห์", "ยอดเยี่ยม! ตัวแปรทั้งหมดตรงกันและเรียงลำดับถูกต้อง", parent=self)
                    self.btn_update_all.config(state=tk.NORMAL)
                else:
                    self.lbl_sheet_status.config(text=f"พบข้อผิดพลาดหลังการเปรียบเทียบ", foreground="red")
                    row_count = 0
                    for var in missing_in_excel:
                        tag = 'oddrow' if row_count % 2 else 'evenrow'
                        self.tree.insert("", "end", values=("ตัวแปรขาดใน Excel", var), tags=('missing', tag))
                        row_count += 1
                    for var in extra_in_excel:
                        tag = 'oddrow' if row_count % 2 else 'evenrow'
                        self.tree.insert("", "end", values=("ตัวแปรเกินใน Excel", var), tags=('extra', tag))
                        row_count += 1
                    self.btn_update_all.config(state=tk.DISABLED)
            
        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"เกิดข้อผิดพลาดขณะวิเคราะห์: {e}", parent=self)
            self.lbl_sheet_status.config(text="เกิดข้อผิดพลาดในการวิเคราะห์", foreground="red")

    def handle_code_columns(self):
        """
        ตรวจหาคอลัมน์ Code ทั้งหมด (ทั้งแบบทั่วไปและ AnswerCode),
        สร้าง DataFrame สำหรับชีทใหม่, และคืนรายชื่อคอลัมน์ Code
        เพื่อนำไปลบออกจาก DataFrame หลัก
        """
        if self.excel_df is None: return [] # คืนค่าเป็น list ว่างถ้าไม่มีข้อมูล

        cols_for_code_sheet = []
        code_cols_to_drop = [] 

        # เพิ่มคอลัมน์ Key เข้าไปก่อนเสมอ
        if self.key_column in self.excel_df.columns:
            cols_for_code_sheet.append(self.key_column)
        else:
            return [] # ถ้าไม่มีคอลัมน์ Key ก็ไม่สามารถทำงานต่อได้

        all_cols = list(self.excel_df.columns)
        for i in range(len(all_cols)):
            current_col_name = all_cols[i]
            
            # ---> [ปรับปรุง] เปลี่ยนกลับไปมองหาคำว่า 'code' ทั่วไป เพื่อให้ครอบคลุมทั้งหมด <---
            if 'code' in str(current_col_name).lower():
                # เพิ่มชื่อคอลัมน์ code นี้เข้าไปใน list ที่จะลบทิ้ง
                code_cols_to_drop.append(current_col_name)

                # วนลูปย้อนกลับจากตำแหน่งปัจจุบันเพื่อหาคอลัมน์ข้อมูลที่คู่กัน
                for j in range(i - 1, -1, -1):
                    potential_data_col_name = all_cols[j]
                    # ---> [ปรับปรุง] หาคอลัมน์ก่อนหน้าที่ไม่มีคำว่า 'code' <---
                    if 'code' not in str(potential_data_col_name).lower():
                        # เมื่อเจอแล้ว ให้เพิ่มทั้งคอลัมน์ข้อมูลและคอลัมน์ code เข้าไป
                        cols_for_code_sheet.append(potential_data_col_name)
                        cols_for_code_sheet.append(current_col_name)
                        break # หยุดค้นหาเมื่อเจอคู่ของมันแล้ว

        # ถ้ามีคอลัมน์ที่ต้องทำชีท Code (มากกว่าแค่คอลัมน์ Key)
        if len(cols_for_code_sheet) > 1:
            # ทำให้รายการไม่ซ้ำกัน แต่ยังคงลำดับเดิมไว้
            unique_cols = list(dict.fromkeys(cols_for_code_sheet))
            # สร้าง DataFrame สำหรับชีท Code เก็บไว้ในตัวแปรของคลาส
            self.df_code_sheet = self.excel_df[unique_cols].copy()
            # ---> [ปรับปรุง] เปลี่ยนข้อความใน Messagebox ให้เป็นกลางมากขึ้น <---
            messagebox.showinfo(
                "ตรวจพบข้อมูล Code",
                f"โปรแกรมตรวจพบและแยกคอลัมน์ที่มีคำว่า 'Code' จำนวน {len(code_cols_to_drop)} คอลัมน์ออกไปเก็บไว้สำหรับชีทใหม่ 'ข้อที่ทำ Code'",
                parent=self
            )
        
        # คืนค่ารายชื่อคอลัมน์ Code ทั้งหมดที่เจอ
        return code_cols_to_drop
    
    def perform_mapping(self):
            if self.excel_df is None:
                messagebox.showerror("ผิดพลาด", "ข้อมูล Excel ยังไม่พร้อม", parent=self)
                return

            messagebox.showinfo(
                "โหลดไฟล์ SPSS Final", 
                "กรุณาเลือกไฟล์ SPSS Final (ที่นำเข้า Lychee) เพื่อใช้ในการ Get จัดเรียงข้อมูล",
                parent=self
            )
            file_path = filedialog.askopenfilename(filetypes=[("SPSS Data File", "*.sav")])
            if not file_path:
                messagebox.showwarning("ยกเลิก", "การจัดเรียงข้อมูลถูกยกเลิก", parent=self)
                return

            self.lbl_sheet_status.config(text="กำลังโหลดและจัดเรียง... กรุณารอสักครู่", foreground="blue")
            self.update_idletasks()

            try:
                # --- [NEW] โหลด Metadata (การตั้งค่าตัวแปร) เก็บไว้ด้วย ---
                import pyreadstat
                # เราโหลด 2 รอบ: 
                # 1. ใช้ pandas เพื่อความง่ายในการจัดการ Dataframe และ Logic เดิม
                # 2. ใช้ pyreadstat เพื่อเอา Metadata (ค่า Setting ต่างๆ)
                
                spss_final_df = pd.read_spss(file_path)
                _, self.spss_final_meta = pyreadstat.read_sav(file_path) # เก็บ Metadata ไว้ที่ตัวนี้
                
                spss_final_vars = list(spss_final_df.columns)
                
                if not spss_final_vars:
                    messagebox.showerror("ผิดพลาด", "ไฟล์ SPSS Final ที่เลือกไม่มีตัวแปร", parent=self)
                    return
                if self.key_column not in spss_final_vars:
                    messagebox.showerror("ผิดพลาด", f"ไฟล์ SPSS Final ไม่มีคอลัมน์หลัก '{self.key_column}'\nกรุณาใช้ไฟล์ที่มีโครงสร้าง Key เดียวกัน", parent=self)
                    return
                    
                excel_status_text = self.lbl_excel_status.cget("text")
                new_combined_status_text = f"{excel_status_text} | SPSS Final: {len(spss_final_vars)} ตัวแปร"
                self.lbl_excel_status.config(text=new_combined_status_text, foreground="green")

                final_vars_set = set(spss_final_vars)
                excel_vars_set = set(self.excel_df.columns)
                missing_vars_in_excel = final_vars_set - excel_vars_set

                ignore_list, add_in_place_list, append_list = [], [], []

                if missing_vars_in_excel:
                    manager_window = MissingVariableManagerWindow(self, sorted(list(missing_vars_in_excel)))
                    self.wait_window(manager_window)
                    if not manager_window.confirmed:
                        self.lbl_sheet_status.config(text="การจัดเรียงถูกยกเลิกโดยผู้ใช้", foreground="orange")
                        return
                    for var, decision in manager_window.decisions.items():
                        if decision == "ลบทิ้ง": ignore_list.append(var)
                        elif decision == "เก็บไว้ (คงตำแหน่งเดิม)": add_in_place_list.append(var)
                        elif decision == "เก็บไว้ (ต่อท้ายสุด)": append_list.append(var)
                
                self.vars_added_log = add_in_place_list + append_list
                for new_var in self.vars_added_log:
                    if new_var not in self.excel_df.columns:
                        self.excel_df[new_var] = pd.NA

                if spss_final_df[self.key_column].duplicated().any():
                    messagebox.showerror("Key Error", f"ข้อมูล '{self.key_column}' ใน SPSS Final มีค่าซ้ำซ้อน", parent=self)
                    return
                if self.excel_df[self.key_column].duplicated().any():
                    messagebox.showerror("Key Error", f"ข้อมูล '{self.key_column}' ใน Excel มีค่าซ้ำซ้อน", parent=self)
                    return

                final_column_order = [v for v in spss_final_vars if v not in ignore_list]
                for v_append in append_list:
                    if v_append in final_column_order:
                        final_column_order.remove(v_append)
                    final_column_order.append(v_append)

                spss_indexed = spss_final_df.set_index(self.key_column)
                excel_indexed = self.excel_df.set_index(self.key_column)
                df_aligned = excel_indexed.reindex(index=spss_indexed.index)
                df_aligned.reset_index(inplace=True)
                df_aligned = df_aligned.reindex(columns=final_column_order)
                self.df_ready_to_save = df_aligned
                
                if self.df_code_sheet is not None and self.key_column in self.df_code_sheet.columns:
                    code_sheet_indexed = self.df_code_sheet.set_index(self.key_column)
                    aligned_code_sheet = code_sheet_indexed.reindex(index=spss_indexed.index)
                    aligned_code_sheet.reset_index(inplace=True)
                    self.df_code_sheet = aligned_code_sheet

                final_cols_set = set(final_column_order)
                original_excel_cols_set = set(self.excel_df.columns)
                dropped_vars_set = original_excel_cols_set - final_cols_set
                self.dropped_vars_log = sorted(list(dropped_vars_set))

                self.btn_save_excel.config(state=tk.NORMAL)
                self.btn_save_spss.config(state=tk.NORMAL)
                self.btn_update_all.config(state=tk.DISABLED)
                
                self.lbl_sheet_status.config(text="จัดเรียงตาม SPSS Final สำเร็จ! พร้อมบันทึก", foreground="green")
                
                summary_message = "จัดเรียงข้อมูลสำเร็จ!\n"
                summary_message += f"- ข้อมูลถูกจัดเรียงตามลำดับของไฟล์ SPSS Final\n"
                summary_message += f"- เก็บการตั้งค่า (Metadata) จากไฟล์ Final ไว้เรียบร้อยแล้ว\n" # แจ้งให้ทราบ
                if self.vars_added_log:
                    summary_message += f"- เพิ่มตัวแปรใหม่ {len(self.vars_added_log)} ตัว\n"
                if self.dropped_vars_log:
                    summary_message += f"- ตัดตัวแปรเกิน {len(self.dropped_vars_log)} ตัว"
                messagebox.showinfo("สำเร็จ", summary_message, parent=self)
                
            except Exception as e:
                messagebox.showerror("ผิดพลาด", f"เกิดข้อผิดพลาดขณะจัดเรียงข้อมูล:\n{e}", parent=self)
                self.lbl_sheet_status.config(text="เกิดข้อผิดพลาดในการจัดเรียง", foreground="red")
    
    def save_excel(self):
        if self.df_ready_to_save is None:
            messagebox.showwarning("ไม่มีข้อมูล", "กรุณากด 'จัดเรียงข้อมูล' ก่อน", parent=self)
            return
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"Mapped_{self.selected_sheet.get()}.xlsx"
        )
        if not save_path: return
        try:
            light_blue_fill = PatternFill(start_color="EAF3FF", end_color="EAF3FF", fill_type="solid")

            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                main_sheet_name = self.selected_sheet.get()
                self.df_ready_to_save.to_excel(writer, sheet_name=main_sheet_name, index=False)
                worksheet_main = writer.sheets[main_sheet_name]
                
                if self.vars_added_log:
                    for cell in worksheet_main[1]: 
                        if cell.value in self.vars_added_log:
                            cell.fill = light_blue_fill

                worksheet_main.freeze_panes = 'B2'

                if self.df_code_sheet is not None:
                    sheet_name_code = 'ข้อที่ทำ Code'
                    self.df_code_sheet.to_excel(writer, sheet_name=sheet_name_code, index=False)
                    worksheet_code = writer.sheets[sheet_name_code]
                    worksheet_code.freeze_panes = 'B2'
                    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                    for cell in worksheet_code[1]:
                        if 'code' in str(cell.value).lower():
                            cell.fill = yellow_fill

                if self.renamed_matches:
                    df_renamed_log = pd.DataFrame(self.renamed_matches, columns=['ชื่อเดิมใน Excel', 'ชื่อใหม่ (จาก SPSS)'])
                    df_renamed_log.to_excel(writer, sheet_name='Log_Renamed_Vars', index=False)

                if self.dropped_vars_log:
                    df_dropped_log = pd.DataFrame(self.dropped_vars_log, columns=['ชื่อตัวแปรที่ถูกตัดทิ้ง'])
                    df_dropped_log.to_excel(writer, sheet_name='Log_Dropped_Vars', index=False)
            
            messagebox.showinfo("บันทึกสำเร็จ", f"ไฟล์ Excel ที่จัดเรียงแล้วถูกบันทึกที่:\n{save_path}\n\n(ตัวแปรที่เพิ่มใหม่จะมีไฮไลท์สีฟ้า)", parent=self)
            self.btn_save_excel.config(state=tk.DISABLED)
            
        except Exception as e: 
            messagebox.showerror("บันทึกผิดพลาด", f"ไม่สามารถบันทึกไฟล์ได้:\n{e}", parent=self)


    def save_spss(self):
            if self.df_ready_to_save is None:
                messagebox.showwarning("ไม่มีข้อมูล", "กรุณากด 'Get SPSS Final' ก่อน", parent=self)
                return

            save_path = filedialog.asksaveasfilename(
                defaultextension=".sav",
                filetypes=[("SPSS Data File", "*.sav")],
                initialfile=f"ForReporter_{self.selected_sheet.get()}.sav"
            )

            if not save_path: return

            try:
                import pyreadstat
                
                # --- [NEW] การบันทึกแบบคงค่า Setting เดิม ---
                # ใช้ข้อมูลจาก self.df_ready_to_save (ที่เป็นข้อมูลใหม่จาก Excel)
                # แต่ใช้ Metadata จาก self.spss_final_meta (ที่เป็นค่า Setting จาก SPSS Final)
                
                column_labels = None
                variable_value_labels = None
                variable_measure = None
                
                if self.spss_final_meta:
                    # ดึงค่า Label และ Values จากไฟล์ต้นฉบับ
                    column_labels = self.spss_final_meta.column_names_to_labels
                    variable_value_labels = self.spss_final_meta.variable_value_labels
                    variable_measure = self.spss_final_meta.variable_measure
                
                pyreadstat.write_sav(
                    self.df_ready_to_save, 
                    save_path, 
                    column_labels=column_labels, 
                    variable_value_labels=variable_value_labels,
                    variable_measure=variable_measure
                )
                
                messagebox.showinfo("บันทึกสำเร็จ", f"บันทึกไฟล์ SPSS เรียบร้อยแล้วที่:\n{save_path}\n\n(คงค่า Label และ Value เดิมจากไฟล์ Final ไว้ทั้งหมด)", parent=self)
                self.btn_save_spss.config(state=tk.DISABLED)
                
            except ImportError:
                messagebox.showerror("Missing Library", "ไม่พบ Library 'pyreadstat'\nกรุณาติดตั้ง: pip install pyreadstat", parent=self)
            except Exception as e:
                messagebox.showerror("บันทึกผิดพลาด", f"ไม่สามารถบันทึกไฟล์ SPSS ได้:\n{e}", parent=self)


def run_this_app(working_dir=None):
    """
    ฟังก์ชันหลักสำหรับสร้างและรัน SPSSExcelMapper.
    """
    print(f"--- APP_INFO: Starting 'SPSSExcelMapper' via run_this_app() ---")
    app = None
    try:
        app = SPSSExcelMapper()
        app.mainloop()
        print(f"--- APP_INFO: SPSSExcelMapper mainloop finished. ---")

    except Exception as e:
        print(f"APP_ERROR: An error occurred during SPSSExcelMapper execution: {e}")
        # ตรวจสอบว่าหน้าต่างหลักถูกสร้างและยังอยู่หรือไม่
        if app and app.winfo_exists():
             messagebox.showerror("Application Error",
                               f"An unexpected error occurred:\n{e}", parent=app)
        else:
            # ถ้าหน้าต่างหลักไม่มีอยู่ ให้สร้างหน้าต่างชั่วคราวเพื่อแสดง error
            root_temp = tk.Tk()
            root_temp.withdraw()
            messagebox.showerror("Application Error",
                               f"An unexpected error occurred:\n{e}", parent=root_temp)
            root_temp.destroy()
        sys.exit(f"Error running SPSSExcelMapper: {e}")

if __name__ == "__main__":
    print("--- Running SPSSExcelMapper.py directly for testing ---")
    run_this_app()
    print("--- Finished direct execution of SPSSExcelMapper.py ---")