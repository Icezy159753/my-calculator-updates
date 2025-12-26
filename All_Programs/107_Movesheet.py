import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import Font
import threading
from copy import copy
import customtkinter as ctk

# --- Set Appearance ---
ctk.set_appearance_mode("dark")  # Modes: "System" (default), "Dark", "Light"
ctk.set_default_color_theme("blue")  # Themes: "blue" (default), "green", "dark-blue"

class SheetMoverApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("โปรแกรมย้ายและจัดการชีท Excel")

        # --- [ปรับปรุง] โค้ดสำหรับจัดหน้าต่างโปรแกรมให้อยู่กึ่งกลางหน้าจอ ---
        window_width = 900
        window_height = 750

        # ดึงขนาดของหน้าจอ
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()

        # คำนวณหาตำแหน่งกึ่งกลาง
        center_x = int((screen_width / 2) - (window_width / 2))
        center_y = int((screen_height / 2) - (window_height / 2))

        # ตั้งค่าขนาดและตำแหน่งของหน้าต่าง
        self.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")
        # --- สิ้นสุดส่วนที่ปรับปรุง ---

        self.source_file_path = tk.StringVar()
        self.mappings = []

        # --- UI Elements ---
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        main_frame = ctk.CTkFrame(self, corner_radius=10)
        main_frame.grid(row=0, column=0, padx=15, pady=15, sticky="ew")
        main_frame.grid_columnconfigure(1, weight=1)
        
        # --- ส่วนที่ 1: เลือกไฟล์ต้นทาง ---
        ctk.CTkLabel(main_frame, text="1. เลือกไฟล์ Table Revised", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=2, padx=10, pady=(10,5), sticky="w")
        ctk.CTkButton(main_frame, text="เลือกไฟล์...", command=self.browse_source_file).grid(row=1, column=0, padx=10, pady=10)
        ctk.CTkEntry(main_frame, textvariable=self.source_file_path, state="readonly").grid(row=1, column=1, padx=10, pady=10, sticky="ew")

        # --- ส่วนใหม่: การตั้งค่า ---
        settings_frame = ctk.CTkFrame(self, corner_radius=10)
        settings_frame.grid(row=1, column=0, padx=15, pady=0, sticky="ew")
        settings_frame.grid_columnconfigure((0,1), weight=1)
        ctk.CTkLabel(settings_frame, text="การตั้งค่า (บันทึก/เรียกใช้)", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=2, padx=10, pady=(10,5), sticky="w")
        ctk.CTkButton(settings_frame, text="บันทึกการตั้งค่า (Excel)...", command=self.save_settings).grid(row=1, column=0, padx=10, pady=10, sticky="ew")
        ctk.CTkButton(settings_frame, text="เรียกใช้การตั้งค่า (Excel)...", command=self.load_settings, fg_color="transparent", border_width=2).grid(row=1, column=1, padx=10, pady=10, sticky="ew")

        # --- ส่วนที่ 2: ระบุชีทและไฟล์ปลายทาง ---
        frame2 = ctk.CTkFrame(self, corner_radius=10)
        frame2.grid(row=2, column=0, padx=15, pady=15, sticky="nsew")
        frame2.grid_rowconfigure(1, weight=1)
        frame2.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(frame2, text="2. กำหนดชีทที่ต้องการย้าย(T1 > T25)", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=(10,5), sticky="w")
        
        self.scrollable_frame = ctk.CTkScrollableFrame(frame2, label_text="")
        self.scrollable_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
        self.scrollable_frame.grid_columnconfigure(4, weight=1)

        # --- Buttons Frame ---
        buttons_frame = ctk.CTkFrame(self, corner_radius=10)
        buttons_frame.grid(row=3, column=0, padx=15, pady=0, sticky="ew")
        buttons_frame.grid_columnconfigure((0,1), weight=1)
        ctk.CTkButton(buttons_frame, text="เพิ่มรายการย้ายชีท +", command=self.add_mapping_row, fg_color="green", hover_color="#00A000").grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        self.start_button = ctk.CTkButton(buttons_frame, text="🚀 เริ่มย้ายชีท", command=self.start_processing_thread, font=ctk.CTkFont(size=16, weight="bold"))
        self.start_button.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        # --- Log Area ---
        self.log_area = ctk.CTkTextbox(self, height=150, state='disabled', corner_radius=10, font=("Courier New", 11))
        self.log_area.grid(row=4, column=0, padx=15, pady=15, sticky="ew")

        self.add_mapping_row()

    def log(self, message):
        self.log_area.configure(state='normal')
        self.log_area.insert(tk.END, message + '\n')
        self.log_area.configure(state='disabled')
        self.log_area.see(tk.END)
        self.update_idletasks()

    def save_settings(self):
        settings_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="บันทึกไฟล์การตั้งค่า"
        )
        if not settings_path: return

        try:
            wb = openpyxl.Workbook()
            ws_config = wb.active
            ws_config.title = "Config"
            ws_config['A1'] = "Source File Path:"
            ws_config['B1'] = self.source_file_path.get()
            
            ws_mappings = wb.create_sheet("Mappings")
            headers = ["Source Sheet", "Target Sheet", "Destination Files", "Use Above"]
            ws_mappings.append(headers)

            for mapping in self.mappings:
                dest_files = ";".join(mapping["dest_files_list"])
                use_above = "TRUE" if mapping["use_above_var"].get() else "FALSE"
                ws_mappings.append([mapping["source_sheet"].get(), mapping["target_sheet"].get(), dest_files, use_above])

            wb.save(settings_path)
            self.log(f"✔ บันทึกการตั้งค่าสำเร็จที่: {settings_path}")
            messagebox.showinfo("สำเร็จ", "บันทึกการตั้งค่าเรียบร้อย")
        except Exception as e:
            self.log(f"!! ERROR: ไม่สามารถบันทึกไฟล์ตั้งค่าได้: {e}")
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถบันทึกไฟล์ตั้งค่าได้:\n{e}")

    def load_settings(self):
        settings_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="เลือกไฟล์การตั้งค่า"
        )
        if not settings_path: return
        
        try:
            wb = openpyxl.load_workbook(settings_path)
            for mapping in self.mappings: mapping["row_frame"].destroy()
            self.mappings.clear()

            ws_config = wb["Config"]
            self.source_file_path.set(ws_config['B1'].value or "")
            self.log(f"เรียกใช้การตั้งค่าจาก: {settings_path}")

            ws_mappings = wb["Mappings"]
            for row in ws_mappings.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                source_sheet, target_sheet, dest_files_str, use_above_str = row
                self.add_mapping_row(source_sheet, target_sheet, dest_files_str, use_above_str)
            
            if not self.mappings: self.add_mapping_row()
        except Exception as e:
            self.log(f"!! ERROR: ไม่สามารถอ่านไฟล์ตั้งค่าได้: {e}")
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถอ่านไฟล์ตั้งค่าได้:\n{e}")

    def browse_source_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename:
            self.source_file_path.set(filename)
            self.log(f"เลือกไฟล์ต้นทาง: {filename}")

    def add_mapping_row(self, source_s="", target_s="", dest_files_s="", use_above_s=""):
        row_frame = ctk.CTkFrame(self.scrollable_frame, fg_color="transparent")
        row_frame.pack(fill=tk.X, expand=True, pady=4)
        
        mapping_data = {"row_frame": row_frame, "source_sheet": tk.StringVar(value=source_s), "target_sheet": tk.StringVar(value=target_s),
                        "dest_file_display": tk.StringVar(), "dest_files_list": [], "use_above_var": tk.BooleanVar()}

        ctk.CTkLabel(row_frame, text="ชื่อชีทRevised:").pack(side=tk.LEFT, padx=(5,0))
        ctk.CTkEntry(row_frame, textvariable=mapping_data["source_sheet"], width=120).pack(side=tk.LEFT, padx=5)
        ctk.CTkLabel(row_frame, text="แทนที่ชีทOriginal:").pack(side=tk.LEFT, padx=(10,0))
        ctk.CTkEntry(row_frame, textvariable=mapping_data["target_sheet"], width=120).pack(side=tk.LEFT, padx=5)
        
        cb = ctk.CTkCheckBox(row_frame, text="ใช้ไฟล์เดียวกับด้านบน", variable=mapping_data["use_above_var"], 
                             command=lambda m=mapping_data: self.toggle_dest_selection(m))
        cb.pack(side=tk.LEFT, padx=(10,0))

        dest_button = ctk.CTkButton(row_frame, text="เลือก...", width=60, command=lambda m=mapping_data: self.browse_multi_dest_files(m))
        dest_button.pack(side=tk.LEFT, padx=5)
        
        dest_label = ctk.CTkEntry(row_frame, textvariable=mapping_data["dest_file_display"], state="readonly")
        dest_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        mapping_data["dest_widgets"] = [dest_button, dest_label]
        mapping_data["checkbox"] = cb

        delete_button = ctk.CTkButton(row_frame, text="ลบ", width=40, command=lambda m=mapping_data: self.remove_mapping_row(m), fg_color="#D32F2F", hover_color="#B71C1C")
        delete_button.pack(side=tk.RIGHT, padx=5)

        self.mappings.append(mapping_data)
        
        # Load data if provided
        dest_files = dest_files_s.split(';') if dest_files_s else []
        mapping_data["dest_files_list"] = dest_files
        if len(dest_files) == 1: mapping_data["dest_file_display"].set(dest_files[0].split('/')[-1])
        elif len(dest_files) > 1: mapping_data["dest_file_display"].set(f"{len(dest_files)} files selected")
        use_above = (str(use_above_s).upper() == "TRUE")
        mapping_data["use_above_var"].set(use_above)
        self.toggle_dest_selection(mapping_data)
        
        if len(self.mappings) == 1: cb.configure(state=tk.DISABLED)

    def toggle_dest_selection(self, mapping_data):
        state = tk.DISABLED if mapping_data["use_above_var"].get() else tk.NORMAL
        for widget in mapping_data["dest_widgets"]: widget.configure(state=state)
        if state == tk.DISABLED: mapping_data["dest_file_display"].set("")

    def remove_mapping_row(self, mapping_data):
        if len(self.mappings) <= 1: return
        is_first = (self.mappings.index(mapping_data) == 0)
        mapping_data["row_frame"].destroy()
        self.mappings.remove(mapping_data)
        if is_first:
            new_first = self.mappings[0]
            new_first["use_above_var"].set(False)
            new_first["checkbox"].configure(state=tk.DISABLED)
            self.toggle_dest_selection(new_first)

    def browse_multi_dest_files(self, mapping_data):
        filenames = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
        if not filenames: return
        mapping_data["dest_files_list"] = list(filenames)
        if len(filenames) == 1: mapping_data["dest_file_display"].set(filenames[0].split('/')[-1])
        else: mapping_data["dest_file_display"].set(f"{len(filenames)} files selected")
        
        self.log(f"--- รายการที่ {self.mappings.index(mapping_data)+1} เลือกไฟล์ปลายทาง {len(filenames)} ไฟล์ ---")
        for f in filenames: self.log(f"  - {f}")

    def start_processing_thread(self):
        self.start_button.configure(state=tk.DISABLED, text="กำลังทำงาน...")
        self.log_area.configure(state='normal')
        self.log_area.delete('1.0', tk.END)
        self.log_area.configure(state='disabled')
        threading.Thread(target=self.process_sheets).start()

    def process_sheets(self):
        if not self.source_file_path.get():
            messagebox.showerror("ข้อผิดพลาด", "กรุณาเลือกไฟล์ต้นทางก่อน")
            self.on_process_complete()
            return

        try:
            self.log("--- เริ่มกระบวนการย้ายชีท ---")
            source_wb = openpyxl.load_workbook(self.source_file_path.get())
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถเปิดไฟล์ต้นทางได้:\n{e}")
            self.log(f"!! ERROR: ไม่สามารถเปิดไฟล์ต้นทางได้ {self.source_file_path.get()}")
            self.on_process_complete()
            return

        # STAGE 1 & 2: Resolve mappings and Group operations
        resolved_ops, previous_dest, config_error = [], [], False
        for i, m in enumerate(self.mappings):
            src_s, tgt_s = m["source_sheet"].get().strip(), m["target_sheet"].get().strip()
            final_dest = m["dest_files_list"] if not m["use_above_var"].get() else previous_dest
            if m["use_above_var"].get() and not previous_dest: config_error = True
            if not all([src_s, tgt_s, final_dest]): continue
            resolved_ops.append({'source': src_s, 'target': tgt_s, 'dests': final_dest, 'row': i+1})
            if not m["use_above_var"].get(): previous_dest = final_dest

        # STAGE 3: Process each group
        has_error = False
        grouped_ops = {}
        for op in resolved_ops:
            for dest in op['dests']:
                if dest not in grouped_ops: grouped_ops[dest] = []
                grouped_ops[dest].append(op)
        
        for dest_path, ops in grouped_ops.items():
            self.log(f"\n--- กำลังประมวลผลไฟล์: {dest_path.split('/')[-1]} ---")
            try:
                dest_wb = openpyxl.load_workbook(dest_path)
                for op in ops:
                    src_s, tgt_s = op['source'], op['target']
                    self.log(f"  -> ย้าย '{src_s}' ไปแทนที่ '{tgt_s}' (จากรายการที่ {op['row']})")
                    if src_s not in source_wb.sheetnames:
                        self.log(f"    !! ERROR: ไม่พบชีท '{src_s}' ในไฟล์ต้นทาง"); has_error = True; continue
                    idx = dest_wb.sheetnames.index(tgt_s) if tgt_s in dest_wb.sheetnames else None
                    if idx is not None: dest_wb.remove(dest_wb[tgt_s])
                    
                    new_sheet = dest_wb.create_sheet(title=tgt_s, index=idx)
                    source_sheet = source_wb[src_s]
                    for row in source_sheet.iter_rows():
                        for cell in row:
                            new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                            if cell.has_style:
                                new_cell.font, new_cell.border, new_cell.fill, new_cell.number_format, new_cell.protection, new_cell.alignment = copy(cell.font), copy(cell.border), copy(cell.fill), cell.number_format, copy(cell.protection), copy(cell.alignment)
                    
                    for letter, dim in source_sheet.column_dimensions.items(): new_sheet.column_dimensions[letter] = copy(dim)
                    for r_idx, dim in source_sheet.row_dimensions.items(): new_sheet.row_dimensions[r_idx] = copy(dim)
                    for merged in source_sheet.merged_cells.ranges: new_sheet.merge_cells(str(merged))
                    
                    new_sheet['E1'].value, new_sheet['E1'].hyperlink, new_sheet['E1'].font = "Index", f"#'Index'!A1", Font(color="5c91fa", underline="single")
                dest_wb.save(dest_path)
                self.log(f"  ✔ บันทึกไฟล์เรียบร้อย")
            except Exception as e:
                self.log(f"!! ERROR: เกิดข้อผิดพลาดกับไฟล์ '{dest_path}': {e}"); has_error = True

        self.log("\n--- กระบวนการเสร็จสิ้น ---")
        if not has_error and not config_error: messagebox.showinfo("สำเร็จ", "การย้ายชีททั้งหมดเสร็จเรียบร้อย!")
        else: messagebox.showwarning("เสร็จสิ้นพร้อมข้อผิดพลาด", "มีบางรายการเกิดข้อผิดพลาด\nกรุณาตรวจสอบ Log")
        self.on_process_complete()

    def on_process_complete(self):
        self.start_button.configure(state=tk.NORMAL, text="🚀 เริ่มย้ายชีท")



# <<< START OF CHANGES >>>
# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
def run_this_app(working_dir=None): # ชื่อฟังก์ชันนี้จะถูกใช้ใน Launcher
    """
    ฟังก์ชันหลักสำหรับสร้างและรัน QuotaSamplerApp.
    """
    print(f"--- QUOTA_SAMPLER_INFO: Starting 'QuotaSamplerApp' via run_this_app() ---")
    try:
        # --- โค้ดที่ย้ายมาจาก if __name__ == "__main__": เดิมจะมาอยู่ที่นี่ ---
    #if __name__ == "__main__":
        app = SheetMoverApp()
        app.mainloop()


        print(f"--- QUOTA_SAMPLER_INFO: QuotaSamplerApp mainloop finished. ---")

    except Exception as e:
        # ดักจับ Error ที่อาจเกิดขึ้นระหว่างการสร้างหรือรัน App
        print(f"QUOTA_SAMPLER_ERROR: An error occurred during QuotaSamplerApp execution: {e}")
        # แสดง Popup ถ้ามีปัญหา
        if 'root' not in locals() or not root.winfo_exists(): # สร้าง root ชั่วคราวถ้ายังไม่มี
            root_temp = tk.Tk()
            root_temp.withdraw()
            messagebox.showerror("Application Error (Quota Sampler)",
                               f"An unexpected error occurred:\n{e}", parent=root_temp)
            root_temp.destroy()
        else:
            messagebox.showerror("Application Error (Quota Sampler)",
                               f"An unexpected error occurred:\n{e}", parent=root) # ใช้ root ที่มีอยู่ถ้าเป็นไปได้
        sys.exit(f"Error running QuotaSamplerApp: {e}") # อาจจะ exit หรือไม่ก็ได้ ขึ้นกับการออกแบบ


# --- ส่วน Run Application เมื่อรันไฟล์นี้โดยตรง (สำหรับ Test) ---
if __name__ == "__main__":
    print("--- Running QuotaSamplerApp.py directly for testing ---")
    # (ถ้ามีการตั้งค่า DPI ด้านบน มันจะทำงานอัตโนมัติ)

    # เรียกฟังก์ชัน Entry Point ที่เราสร้างขึ้น
    run_this_app()

    print("--- Finished direct execution of QuotaSamplerApp.py ---")
# <<< END OF CHANGES >>>