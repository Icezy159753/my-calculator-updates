import tkinter
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.scrolled import ScrolledText
from tkinter import filedialog, messagebox
import pandas as pd
import pyreadstat
import os
import numpy as np
from openpyxl.styles import PatternFill, Font, Alignment

# --- ฟังก์ชันสำหรับจัดหน้าจอให้อยู่ตรงกลาง ---
def center_window(win):
    win.update_idletasks()
    width = win.winfo_width()
    height = win.winfo_height()
    x = (win.winfo_screenwidth() // 2) - (width // 2)
    y = (win.winfo_screenheight() // 2) - (height // 2)
    win.geometry(f'{width}x{height}+{x}+{y}')


class VariableSelector(ttk.Toplevel):
    # ... (คลาสนี้ไม่มีการเปลี่ยนแปลง) ...
    def __init__(self, parent, title, all_variables, previously_selected):
        super().__init__(parent)
        self.title(title)
        self.geometry("500x600")
        self.transient(parent)
        self.grab_set()

        self.all_variables = all_variables
        self.selection_set = set(previously_selected)
        self.result = None

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        
        search_frame = ttk.Frame(self, padding=(10, 10))
        search_frame.grid(row=0, column=0, padx=10, pady=(10,0), sticky="ew")
        search_frame.grid_columnconfigure(1, weight=1)
        ttk.Label(search_frame, text="ค้นหา:").grid(row=0, column=0, padx=(0,5))
        self.search_entry = ttk.Entry(search_frame)
        self.search_entry.grid(row=0, column=1, padx=5, sticky="ew")
        self.search_entry.bind("<KeyRelease>", self._filter_list)
        self.search_entry.insert(0, "พิมพ์เพื่อค้นหาตัวแปร...")
        self.search_entry.bind("<FocusIn>", lambda e: self.search_entry.delete(0, 'end'))

        list_frame = ttk.Frame(self)
        list_frame.grid(row=1, column=0, padx=10, pady=5, sticky="nsew")
        list_frame.grid_rowconfigure(0, weight=1)
        list_frame.grid_columnconfigure(0, weight=1)
        
        style = ttk.Style()
        listbox_bg = style.colors.get('bg')
        listbox_fg = style.colors.get('fg')
        select_bg = style.colors.get('primary')
        select_fg = style.colors.get('selectfg')

        self.listbox = tkinter.Listbox(
            list_frame, selectmode=tkinter.EXTENDED, background=listbox_bg,
            foreground=listbox_fg, selectbackground=select_bg, selectforeground=select_fg,
            borderwidth=0, highlightthickness=0, font=("Segoe UI", 11)
        )
        self.listbox.grid(row=0, column=0, sticky="nsew")

        scrollbar = ttk.Scrollbar(list_frame, orient=VERTICAL, command=self.listbox.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.listbox.configure(yscrollcommand=scrollbar.set)
        self._populate_listbox(self.all_variables)

        button_frame = ttk.Frame(self, padding=(10,10))
        button_frame.grid(row=2, column=0, padx=10, pady=(0,10), sticky="ew")
        button_frame.grid_columnconfigure((0,1), weight=1)
        ttk.Button(button_frame, text="ตกลง", command=self._on_ok, bootstyle=SUCCESS).grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        ttk.Button(button_frame, text="ยกเลิก", command=self.destroy, bootstyle=SECONDARY).grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        center_window(self)

    def _populate_listbox(self, var_list):
        self.listbox.delete(0, tkinter.END)
        for i, var in enumerate(var_list):
            self.listbox.insert(tkinter.END, var)
            if var in self.selection_set:
                self.listbox.selection_set(i)
    
    def _filter_list(self, event=None):
        search_term = self.search_entry.get().lower()
        if search_term == "พิมพ์เพื่อค้นหาตัวแปร...":
             search_term = ""
        filtered_list = [var for var in self.all_variables if search_term in var.lower()] if search_term else self.all_variables
        self._populate_listbox(filtered_list)

    def _on_ok(self):
        currently_displayed = self.listbox.get(0, tkinter.END)
        selected_indices = self.listbox.curselection()
        currently_selected = {currently_displayed[i] for i in selected_indices}
        unselected_and_displayed = set(currently_displayed) - currently_selected
        self.selection_set.difference_update(unselected_and_displayed)
        self.selection_set.update(currently_selected)
        self.result = sorted(list(self.selection_set))
        self.destroy()


class SPSSAnalyzerApp(ttk.Window):
    def __init__(self):
        super().__init__(themename="darkly") 
        self.title("โปรแกรมคำนวน BPI Brand Power Index v1")
        self.geometry("850x750")
        
        style = ttk.Style()
        style.configure('TLabelframe.Label', font=('Segoe UI', 12, 'bold'))
        style.configure('TLabel', font=('Segoe UI', 11))
        style.configure('TButton', font=('Segoe UI', 11))
        style.configure('TRadiobutton', font=('Segoe UI', 11))
        style.configure('Treeview.Heading', font=('Segoe UI', 11, 'bold'))
        style.configure('Treeview', rowheight=28, font=('Segoe UI', 10))

        self.WEIGHTS = {
            "1. Aided awareness": {"BC": 6.2, "FC": 7.8, "Other": 7.0}, "2. Currently use": {"BC": 2.6, "FC": 3.1, "Other": 2.9},
            "3. Purchase consideration": {}, "4. Favorability 1st": {"BC": 39.1,"FC": 37.6,"Other": 38.3},
            "5. Favorability 2nd": {"BC": 21.1,"FC": 20.7,"Other": 20.9}, "6. Recommendation": {"BC": 38.6,"FC": 40.8,"Other": 39.7},
            "Purchase consideration (composition)": {"BC": 13.5,"FC": 10.7,"Other": 12.1}
        }

        self.spss_file_path = None; self.df_spss = None; self.df_result = None
        self.spss_variable_names = []; self.selected_vars_map = {}
        
        self.grid_columnconfigure(0, weight=1); self.grid_rowconfigure(2, weight=1)

        # --- Section 1: Input Frame ---
        self.input_frame = ttk.Labelframe(self, text="ตั้งค่าการวิเคราะห์", padding=15)
        self.input_frame.grid(row=0, column=0, padx=10, pady=10, sticky="new")
        self.input_frame.grid_columnconfigure(1, weight=1)

        ttk.Label(self.input_frame, text="1. เลือกไฟล์ SPSS (.sav):").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.file_entry = ttk.Entry(self.input_frame, font=('Segoe UI', 10))
        self.file_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.file_entry.insert(0, "ยังไม่ได้เลือกไฟล์")
        self.file_entry.configure(state='readonly')
        ttk.Button(self.input_frame, text="Browse...", command=self.browse_file, bootstyle=SECONDARY).grid(row=0, column=2, padx=5, pady=5)

        ttk.Label(self.input_frame, text="2. กำหนด Brand และ Code:").grid(row=1, column=0, padx=5, pady=5, sticky="nw")
        self.brand_textbox = ScrolledText(self.input_frame, height=4, autohide=True, font=('Segoe UI', 10))
        self.brand_textbox.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky="ew")
        self.brand_textbox.insert("1.0", "Lifree 1\nCertainty 2\n")

        ttk.Label(self.input_frame, text="3. เลือก Type Weight:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.weight_var = tkinter.StringVar(value="No Weight")
        weight_frame = ttk.Frame(self.input_frame)
        weight_frame.grid(row=2, column=1, columnspan=2, padx=5, pady=5, sticky="w")
        ttk.Radiobutton(weight_frame, text="ไม่ใช้ Weight", variable=self.weight_var, value="No Weight").pack(side="left", padx=5)
        ttk.Radiobutton(weight_frame, text="BC (Type1)", variable=self.weight_var, value="BC").pack(side="left", padx=5)
        ttk.Radiobutton(weight_frame, text="FC (Type2)", variable=self.weight_var, value="FC").pack(side="left", padx=5)
        ttk.Radiobutton(weight_frame, text="Other (Type3)", variable=self.weight_var, value="Other").pack(side="left", padx=5)

        ttk.Label(self.input_frame, text="4. กำหนดตัวแปรสำหรับแต่ละข้อ:").grid(row=3, column=0, padx=5, pady=5, sticky="nw")
        self.question_ui_map = {}
        self.questions_order = ["1. Aided awareness", "2. Currently use", "3. Purchase consideration", "4. Favorability 1st", "5. Favorability 2nd", "6. Recommendation"]
        question_container_frame = ttk.Frame(self.input_frame)
        question_container_frame.grid(row=3, column=1, columnspan=2, padx=5, pady=5, sticky="ew")
        question_container_frame.grid_columnconfigure(1, weight=1)
        for i, q_text in enumerate(self.questions_order):
            self.selected_vars_map[q_text] = []
            label = ttk.Label(question_container_frame, text=q_text, width=25, anchor="w")
            label.grid(row=i, column=0, padx=5, pady=3, sticky="w")
            status_label = ttk.Label(question_container_frame, text="ยังไม่ได้เลือกตัวแปร", bootstyle="secondary")
            status_label.grid(row=i, column=1, padx=5, pady=3, sticky="ew")
            select_button = ttk.Button(question_container_frame, text="เลือกตัวแปร...", command=lambda q=q_text: self.open_variable_selector(q), state="disabled", bootstyle=INFO)
            select_button.grid(row=i, column=2, padx=5, pady=3)
            self.question_ui_map[q_text] = {'button': select_button, 'label': status_label}

        # --- Section 2: Action Frame ---
        self.action_frame = ttk.Frame(self)
        self.action_frame.grid(row=1, column=0, padx=10, pady=5, sticky="ew")
        
        ttk.Button(self.action_frame, text="Save Settings", command=self.save_settings, bootstyle="outline-info").pack(side="left", padx=(10, 5), pady=10)
        ttk.Button(self.action_frame, text="Load Settings", command=self.load_settings, bootstyle="outline-info").pack(side="left", padx=5, pady=10)
        
        ttk.Button(self.action_frame, text="ประมวลผล (Run Analysis)", command=self.run_analysis, bootstyle=PRIMARY).pack(side="left", padx=10, pady=10)
        self.export_button = ttk.Button(self.action_frame, text="Export to Excel", command=self.export_to_excel, state="disabled", bootstyle=SUCCESS)
        self.export_button.pack(side="left", padx=0, pady=10)
        self.status_label = ttk.Label(self.action_frame, text="สถานะ: พร้อมทำงาน", bootstyle=SECONDARY)
        self.status_label.pack(side="right", padx=10, pady=10)

        # --- Section 3: Result Frame ---
        self.result_frame = ttk.Labelframe(self, text="ผลลัพธ์", padding=10)
        self.result_frame.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")
        self.result_frame.grid_rowconfigure(0, weight=1)
        self.result_frame.grid_columnconfigure(0, weight=1)
        self.tree = ttk.Treeview(self.result_frame, show='headings', bootstyle=PRIMARY)
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb = ttk.Scrollbar(self.result_frame, orient="vertical", command=self.tree.yview, bootstyle=ROUND)
        vsb.grid(row=0, column=1, sticky='ns')
        self.tree.configure(yscrollcommand=vsb.set)
        hsb = ttk.Scrollbar(self.result_frame, orient="horizontal", command=self.tree.xview, bootstyle=ROUND)
        hsb.grid(row=1, column=0, sticky='ew')
        self.tree.configure(xscrollcommand=hsb.set)

        center_window(self)

    def save_settings(self):
        if not any(self.selected_vars_map.values()):
            messagebox.showwarning("คำเตือน", "ยังไม่มีการตั้งค่าตัวแปรให้บันทึก")
            return
            
        filepath = filedialog.asksaveasfilename(
            title="บันทึกการตั้งค่า",
            defaultextension=".xlsx",
            filetypes=(("Excel Files", "*.xlsx"), ("All files", "*.*"))
        )
        if not filepath:
            return

        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # 1. บันทึกการตั้งค่าตัวแปร
                data_to_save = []
                for question, var_list in self.selected_vars_map.items():
                    for var in var_list:
                        data_to_save.append({'Question': question, 'Variable': var})
                df_vars = pd.DataFrame(data_to_save)
                df_vars.to_excel(writer, sheet_name='Variable_Map', index=False)

                # 2. บันทึกการตั้งค่าอื่นๆ (Weight และ Brand Text)
                ### <<< ดึงค่าจาก Brand Textbox >>> ###
                brand_text = self.brand_textbox.get("1.0", "end-1c").strip()

                ### <<< สร้าง DataFrame สำหรับการตั้งค่าอื่นๆ >>> ###
                other_settings_data = [
                    {'Setting': 'Selected_Weight', 'Value': self.weight_var.get()},
                    {'Setting': 'Brand_Code_Text', 'Value': brand_text}
                ]
                df_other_settings = pd.DataFrame(other_settings_data)
                df_other_settings.to_excel(writer, sheet_name='Other_Settings', index=False)

            self.update_status(f"สถานะ: บันทึกการตั้งค่าสำเร็จ", "success")
            messagebox.showinfo("สำเร็จ", f"บันทึกการตั้งค่าเรียบร้อยแล้วที่:\n{filepath}")
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถบันทึกไฟล์ตั้งค่าได้:\n{e}")
            self.update_status("สถานะ: บันทึกการตั้งค่าไม่สำเร็จ", "danger")

    def load_settings(self):
        if self.df_spss is None:
            messagebox.showerror("ข้อผิดพลาด", "กรุณาเลือกและโหลดไฟล์ SPSS ก่อน")
            return

        filepath = filedialog.askopenfilename(
            title="เลือกไฟล์การตั้งค่า",
            filetypes=(("Excel Files", "*.xlsx"), ("All files", "*.*"))
        )
        if not filepath:
            return
            
        try:
            self.update_status("สถานะ: กำลังโหลดการตั้งค่า...", "warning")
            
            # 1. อ่านค่าการตั้งค่าตัวแปร
            df_vars = pd.read_excel(filepath, sheet_name='Variable_Map')
            if 'Question' not in df_vars.columns or 'Variable' not in df_vars.columns:
                raise ValueError("ไฟล์ตั้งค่าไม่ถูกต้อง รูปแบบคอลัมน์ต้องเป็น 'Question' และ 'Variable'")

            new_var_map = {q: [] for q in self.questions_order}
            loaded_vars_count = 0
            skipped_vars_count = 0

            for _, row in df_vars.iterrows():
                question = row['Question']
                variable = row['Variable']
                if variable in self.spss_variable_names:
                    if question in new_var_map:
                        new_var_map[question].append(variable)
                        loaded_vars_count += 1
                else:
                    skipped_vars_count += 1
            
            self.selected_vars_map = new_var_map
            self._update_all_question_labels()

            # 2. อ่านค่าอื่นๆ จากชีท 'Other_Settings' (ถ้ามี)
            try:
                df_other = pd.read_excel(filepath, sheet_name='Other_Settings')
                
                # โหลด Weight
                weight_row = df_other[df_other['Setting'] == 'Selected_Weight']
                if not weight_row.empty:
                    self.weight_var.set(weight_row['Value'].iloc[0])

                ### <<< โหลด Brand Text >>> ###
                brand_row = df_other[df_other['Setting'] == 'Brand_Code_Text']
                if not brand_row.empty:
                    # ดึงค่ามา, ตรวจสอบว่าเป็นค่าว่าง (NaN) หรือไม่
                    brand_text_to_load = brand_row['Value'].iloc[0]
                    if pd.isna(brand_text_to_load):
                        brand_text_to_load = "" # ถ้าเป็นค่าว่าง ให้เป็น string ว่าง
                    
                    self.brand_textbox.delete("1.0", "end")
                    self.brand_textbox.insert("1.0", str(brand_text_to_load))

            except Exception:
                # ไม่ต้องทำอะไรถ้าไม่มีชีทนี้ (สำหรับไฟล์เวอร์ชันเก่า)
                pass

            self.update_status("สถานะ: โหลดการตั้งค่าสำเร็จ", "success")
            final_message = f"โหลดการตั้งค่าเรียบร้อย!\n\n- โหลดตัวแปรสำเร็จ: {loaded_vars_count} ตัว"
            if skipped_vars_count > 0:
                final_message += f"\n- ข้ามตัวแปรที่ไม่มีในไฟล์ SPSS นี้: {skipped_vars_count} ตัว"
            messagebox.showinfo("สำเร็จ", final_message)

        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถโหลดไฟล์ตั้งค่าได้:\n{e}")
            self.update_status("สถานะ: โหลดการตั้งค่าไม่สำเร็จ", "danger")
    
    def _update_all_question_labels(self):
        # ... (เหมือนเดิม) ...
        for q_name, ui_elements in self.question_ui_map.items():
            count = len(self.selected_vars_map.get(q_name, []))
            if count > 0:
                ui_elements['label'].configure(text=f"เลือกแล้ว {count} ตัว", bootstyle="info")
            else:
                ui_elements['label'].configure(text="ยังไม่ได้เลือกตัวแปร", bootstyle="secondary")

    def update_status(self, message, style="secondary"):
        # ... (เหมือนเดิม) ...
        self.status_label.configure(text=message, bootstyle=style)
        self.update_idletasks()

    def browse_file(self):
        # ... (เหมือนเดิม) ...
        filepath = filedialog.askopenfilename(filetypes=(("SPSS Data File", "*.sav"), ("All files", "*.*")))
        if not filepath: return
        self.spss_file_path = filepath
        self.file_entry.configure(state='normal')
        self.file_entry.delete(0, "end"); self.file_entry.insert(0, os.path.basename(filepath))
        self.file_entry.configure(state='readonly')
        self.update_status("สถานะ: กำลังอ่านไฟล์...", "warning")
        try:
            self.df_spss, meta = pyreadstat.read_sav(self.spss_file_path)
            self.spss_variable_names = meta.column_names
            for q_ui in self.question_ui_map.values(): q_ui['button'].configure(state="normal")
            self.update_status(f"สถานะ: อ่านไฟล์สำเร็จ มี {len(self.spss_variable_names)} ตัวแปร", "success")
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถอ่านไฟล์ SPSS ได้:\n{e}")
            self.update_status("สถานะ: เกิดข้อผิดพลาด", "danger")

    def open_variable_selector(self, q_name):
        # ... (เหมือนเดิม) ...
        selector_window = VariableSelector(self, f"เลือกตัวแปรสำหรับ: {q_name}", self.spss_variable_names, self.selected_vars_map.get(q_name, []))
        self.wait_window(selector_window)
        if selector_window.result is not None:
            self.selected_vars_map[q_name] = selector_window.result
            self._update_all_question_labels()

    def run_analysis(self):
        # ... (ส่วน Logic การคำนวณเหมือนเดิมทุกประการ ไม่มีการเปลี่ยนแปลง) ...
        if self.df_spss is None: messagebox.showerror("ข้อผิดพลาด", "กรุณาเลือกและโหลดไฟล์ SPSS ก่อน"); return
        brand_text = self.brand_textbox.get("1.0", "end-1c").strip()
        if not brand_text: messagebox.showerror("ข้อผิดพลาด", "กรุณากำหนด Brand และ Code"); return
        brands = {};
        try:
            for line in brand_text.split('\n'):
                if line.strip(): parts = line.strip().rsplit(None, 1); brands[parts[0]] = float(parts[1])
        except (IndexError, ValueError): messagebox.showerror("ข้อผิดพลาด", "รูปแบบ Brand ไม่ถูกต้อง"); return
            
        self.update_status("สถานะ: กำลังประมวลผล...", "warning")
        
        id_column = 'Sbjnum' if 'Sbjnum' in self.df_spss.columns else self.df_spss.columns[0]
        self.df_result = self.df_spss[[id_column]].copy()
        
        selected_weight_type = self.weight_var.get()

        for q_name in self.questions_order:
            target_vars = self.selected_vars_map.get(q_name, [])
            for brand_name, brand_code in brands.items():
                new_col_name = f"{brand_name} - {q_name}"
                existing_vars = [v for v in target_vars if v in self.df_spss.columns]
                if not target_vars or not existing_vars:
                    self.df_result[new_col_name] = pd.Series(0, index=self.df_result.index)
                else:
                    self.df_result[new_col_name] = (self.df_spss[existing_vars] == brand_code).any(axis=1).astype(int)
        
        q_pc = "3. Purchase consideration"
        consideration_cols = [f"{bn} - {q_pc}" for bn in brands.keys() if f"{bn} - {q_pc}" in self.df_result.columns]
        total_consideration_per_row = self.df_result[consideration_cols].sum(axis=1)
        for brand_name in brands.keys():
            source_col = f"{brand_name} - {q_pc}"
            new_col_name = f"{brand_name} - Calculation of purchase consideration composition"
            self.df_result[new_col_name] = np.where(total_consideration_per_row == 0, 0.0, self.df_result[source_col] / total_consideration_per_row)
        
        if selected_weight_type != "No Weight":
            self.update_status("สถานะ: กำลังคำนวณ BPI...", "warning")
            standard_metrics = ["1. Aided awareness", "2. Currently use", "4. Favorability 1st", "5. Favorability 2nd", "6. Recommendation"]
            
            for brand_name in brands.keys():
                bpi_score = pd.Series(0.0, index=self.df_result.index)
                for metric_name in standard_metrics:
                    col_0_1_value = self.df_result[f"{brand_name} - {metric_name}"]
                    weight = self.WEIGHTS.get(metric_name, {}).get(selected_weight_type, 0.0)
                    bpi_score += col_0_1_value * weight
                
                composition_value = self.df_result[f"{brand_name} - Calculation of purchase consideration composition"]
                composition_weight = self.WEIGHTS.get("Purchase consideration (composition)", {}).get(selected_weight_type, 0.0)
                bpi_score += composition_value * composition_weight
                self.df_result[f"{brand_name} - BPI"] = bpi_score

        self.update_treeview(self.df_result)
        self.update_bpi_summary(brands.keys(), selected_weight_type)
        self.export_button.configure(state="normal")
        self.update_status("สถานะ: ประมวลผลเสร็จสิ้น", "success")


    def update_bpi_summary(self, brand_names, selected_weight_type):
        if self.df_result is None or selected_weight_type == "No Weight":
            messagebox.showwarning("สรุป BPI", "ยังไม่ได้คำนวณ BPI เพราะยังไม่ได้เลือก Weight")
            return

        summary_data = []
        has_over_100 = False

        for brand_name in brand_names:
            bpi_col = f"{brand_name} - BPI"
            if bpi_col not in self.df_result.columns:
                continue

            max_bpi = float(self.df_result[bpi_col].max())
            is_over_100 = max_bpi > 100
            has_over_100 = has_over_100 or is_over_100
            summary_data.append((brand_name, max_bpi, is_over_100))

        if not summary_data:
            messagebox.showwarning("สรุป BPI", "ไม่พบคอลัมน์ BPI")
            return

        self.show_bpi_summary_dialog(summary_data, has_over_100)

    def show_bpi_summary_dialog(self, summary_data, has_over_100):
        dialog = ttk.Toplevel(self)
        dialog.title("สรุป BPI ต่อ Brand")
        dialog.geometry("520x360")
        dialog.transient(self)
        dialog.grab_set()
        dialog.resizable(False, False)

        container = ttk.Frame(dialog, padding=16)
        container.pack(fill="both", expand=True)

        title_style = "danger" if has_over_100 else "success"
        title_text = "พบ Brand ที่ค่า BPI เกิน 100" if has_over_100 else "ทุก Brand มีค่า BPI ไม่เกิน 100"
        ttk.Label(container, text=title_text, bootstyle=title_style, font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 12))

        list_frame = ttk.Frame(container)
        list_frame.pack(fill="both", expand=True)

        for brand_name, max_bpi, is_over_100 in summary_data:
            brand_style = "danger" if is_over_100 else "light"
            status_text = "เกิน 100" if is_over_100 else "ไม่เกิน 100"
            row_text = f"{brand_name}: MAX {max_bpi:.2f} ({status_text})"
            ttk.Label(
                list_frame,
                text=row_text,
                bootstyle=brand_style,
                font=("Segoe UI", 11, "bold" if is_over_100 else "normal"),
                anchor="w",
                justify="left"
            ).pack(fill="x", anchor="w", pady=4)

        ttk.Button(container, text="ปิด", command=dialog.destroy, bootstyle="secondary").pack(anchor="e", pady=(12, 0))
        center_window(dialog)


    def update_treeview(self, df):
        # ... (เหมือนเดิม) ...
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = list(df.columns)
        for col in df.columns: 
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor='center', width=120)
        for _, row in df.iterrows(): 
            formatted_values = [f"{v:.4f}" if isinstance(v, (float, np.floating)) else v for v in row]
            self.tree.insert("", "end", values=formatted_values)

    
    def export_to_excel(self):
        # ... (ส่วน Export Excel เหมือนเดิมทุกประการ ไม่มีการเปลี่ยนแปลง) ...
        if self.df_result is None:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลให้ส่งออก")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=(("Excel Files", "*.xlsx"),)
        )

        if not filepath:
            return

        try:
            base_export_df = self.df_result.copy()
            export_df = base_export_df.copy()
            id_column_name = str(base_export_df.columns[0])
            bpi_columns = [col for col in base_export_df.columns if str(col).endswith(" - BPI")]

            def format_id_value(value):
                if isinstance(value, (int, np.integer)):
                    return str(int(value))
                if isinstance(value, (float, np.floating)):
                    return str(int(float(value)))
                return str(value)

            def format_syntax_value(value):
                if isinstance(value, (int, np.integer)):
                    return f"{float(value):.2f}"
                if isinstance(value, (float, np.floating)):
                    return f"{float(value):.2f}"
                return str(value)

            syntax_summary_rows = []
            syntax_detail_rows = []

            for idx, bpi_col in enumerate(bpi_columns, start=1):
                syntax_col_name = f"BPI_BRAND{idx}"
                brand_name = str(bpi_col).replace(" - BPI", "")
                syntax_texts = [
                    f"IF {id_column_name}={format_id_value(row_id)} THEN  {syntax_col_name}={format_syntax_value(bpi_value)} FI"
                    for row_id, bpi_value in zip(base_export_df.iloc[:, 0], base_export_df[bpi_col])
                ]
                max_bpi_value = float(base_export_df[bpi_col].max())
                syntax_summary_rows.append((idx, syntax_col_name, brand_name, max_bpi_value))
                syntax_detail_rows.extend(
                    (syntax_col_name, brand_name, syntax_text)
                    for syntax_text in syntax_texts
                )

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                export_df.to_excel(writer, index=False, sheet_name='BPI')

                worksheet = writer.sheets['BPI']
                workbook = writer.book

                header_font_white = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                header_font_black = Font(name='Calibri', size=11, bold=True, color='000000')
                header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                data_align = Alignment(horizontal='center', vertical='center')
                syntax_align = Alignment(horizontal='left', vertical='center')
                
                colors = {
                    "Aided": PatternFill(start_color='004225', end_color='004225', fill_type='solid'),
                    "Currently use": PatternFill(start_color='52003A', end_color='52003A', fill_type='solid'),
                    "Purchase consideration": PatternFill(start_color='833403', end_color='833403', fill_type='solid'),
                    "Favorability": PatternFill(start_color='0B2447', end_color='0B2447', fill_type='solid'),
                    "Recommendation": PatternFill(start_color='1A5D1A', end_color='1A5D1A', fill_type='solid'),
                    "BPI_BRAND": PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
                    "BPI": PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid'),
                }

                for col_idx, column_cell in enumerate(worksheet[1], 1):
                    column_cell.font = header_font_white
                    column_cell.alignment = header_align
                    header_text = column_cell.value
                    
                    column_cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')

                    for key, fill in colors.items():
                        if key in header_text:
                            column_cell.fill = fill
                            if key in {"BPI", "BPI_BRAND"}:
                                column_cell.font = header_font_black
                            break
                    
                    column_letter = column_cell.column_letter
                    if column_letter == 'A':
                        worksheet.column_dimensions[column_letter].width = 14
                    elif str(header_text).startswith("BPI_BRAND"):
                        worksheet.column_dimensions[column_letter].width = 34
                    else:
                        worksheet.column_dimensions[column_letter].width = 12

                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = data_align

                bpi_header_names = {str(col) for col in bpi_columns}
                for col_idx, header_cell in enumerate(worksheet[1], 1):
                    if str(header_cell.value) in bpi_header_names:
                        for row_idx in range(2, worksheet.max_row + 1):
                            worksheet.cell(row=row_idx, column=col_idx).number_format = '0.00'

                worksheet.freeze_panes = 'B2'
                worksheet.sheet_properties.tabColor = "5B9BD5"

                syntax_ws = workbook.create_sheet(title="Syntax Lychee")
                syntax_ws.sheet_properties.tabColor = "F4B183"

                syntax_header_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                syntax_header_font = Font(name='Calibri', size=11, bold=True, color='000000')

                syntax_ws["A1"] = "No."
                syntax_ws["B1"] = "ตัวแปรใหม่ Lychee"
                syntax_ws["C1"] = "VarBrand"
                syntax_ws["D1"] = "MAX นำเกินค่า100"
                syntax_ws["E1"] = "Var"
                syntax_ws["F1"] = "Brand"
                syntax_ws["G1"] = "Syntax"

                for cell in syntax_ws[1]:
                    cell.fill = syntax_header_fill
                    cell.font = syntax_header_font
                    cell.alignment = header_align

                for row_idx, (idx, syntax_var, brand_name, max_bpi_value) in enumerate(syntax_summary_rows, start=2):
                    syntax_ws.cell(row=row_idx, column=1, value=idx)
                    syntax_ws.cell(row=row_idx, column=2, value=syntax_var)
                    syntax_ws.cell(row=row_idx, column=3, value=brand_name)
                    syntax_ws.cell(row=row_idx, column=4, value=round(max_bpi_value))

                for row_idx, (syntax_var, brand_name, syntax_text) in enumerate(syntax_detail_rows, start=2):
                    syntax_ws.cell(row=row_idx, column=5, value=syntax_var)
                    syntax_ws.cell(row=row_idx, column=6, value=brand_name)
                    syntax_ws.cell(row=row_idx, column=7, value=syntax_text)

                for row in syntax_ws.iter_rows(min_row=2, max_row=syntax_ws.max_row):
                    for cell in row:
                        cell.alignment = syntax_align

                syntax_ws.column_dimensions['A'].width = 8
                syntax_ws.column_dimensions['B'].width = 18
                syntax_ws.column_dimensions['C'].width = 16
                syntax_ws.column_dimensions['D'].width = 18
                syntax_ws.column_dimensions['E'].width = 16
                syntax_ws.column_dimensions['F'].width = 16
                syntax_ws.column_dimensions['G'].width = 70
                syntax_ws.freeze_panes = 'A2'

            messagebox.showinfo("สำเร็จ", f"บันทึกไฟล์ Excel เรียบร้อยแล้วที่:\n{filepath}")

        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถบันทึกไฟล์ได้:\n{e}")



# <<< START OF CHANGES >>>
# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
def run_this_app(working_dir=None): # ชื่อฟังก์ชันนี้จะถูกใช้ใน Launcher
    """
    ฟังก์ชันหลักสำหรับสร้างและรัน QuotaSamplerApp.
    """
    print(f"--- QUOTA_SAMPLER_INFO: Starting 'QuotaSamplerApp' via run_this_app() ---")
    try:
    # --- ส่วนที่ใช้รันโปรแกรม ---
    #if __name__ == "__main__":
        app = SPSSAnalyzerApp()
        app.mainloop()
        
        print(f"--- QUOTA_SAMPLER_INFO: QuotaSamplerApp mainloop finished. ---")

    except Exception as e:
        # ดักจับ Error ที่อาจเกิดขึ้นระหว่างการสร้างหรือรัน App
        print(f"QUOTA_SAMPLER_ERROR: An error occurred during QuotaSamplerApp execution: {e}")
        # แสดง Popup ถ้ามีปัญหา
        if 'root' not in locals() or not root.winfo_exists(): # สร้าง root ชั่วคราวถ้ายังไม่มี
            root_temp = tk.Tk()
            root_temp.withdraw()
            messagebox.showerror("Application Error (Quota Sampler)",
                               f"An unexpected error occurred:\n{e}", parent=root_temp)
            root_temp.destroy()
        else:
            messagebox.showerror("Application Error (Quota Sampler)",
                               f"An unexpected error occurred:\n{e}", parent=root) # ใช้ root ที่มีอยู่ถ้าเป็นไปได้
        sys.exit(f"Error running QuotaSamplerApp: {e}") # อาจจะ exit หรือไม่ก็ได้ ขึ้นกับการออกแบบ


# --- ส่วน Run Application เมื่อรันไฟล์นี้โดยตรง (สำหรับ Test) ---
if __name__ == "__main__":
    print("--- Running QuotaSamplerApp.py directly for testing ---")
    # (ถ้ามีการตั้งค่า DPI ด้านบน มันจะทำงานอัตโนมัติ)

    # เรียกฟังก์ชัน Entry Point ที่เราสร้างขึ้น
    run_this_app()

    print("--- Finished direct execution of QuotaSamplerApp.py ---")
# <<< END OF CHANGES >>>
