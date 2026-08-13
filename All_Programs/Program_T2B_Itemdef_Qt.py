# -*- coding: utf-8 -*-
"""
โปรแกรมสร้าง TB/T2B จาก Itemdef  (PyQt6 Edition)
--------------------------------------------------
พอร์ตมาจาก Program_T2B_Itermdef.py (CustomTkinter) โดย:
  - Logic การอ่าน Itemdef / สร้าง Making / เขียน Excel เหมือนเดิมทุกประการ
  - UX เหมือนเดิม: ปุ่ม 3 ปุ่มด้านบน, 2 แท็บตาราง, คลิกขวาเลือก Direction, log ด้านล่าง
"""

import os
import sys
import time
import datetime
import traceback
from collections import defaultdict

import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: Critical 'openpyxl' library not found. -> pip install openpyxl")
    raise

from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtCore import Qt


# =========================================================================
#  Helper / Constants  (เหมือนเวอร์ชันเดิมทุกอย่าง)
# =========================================================================
def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    base_path = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
    return os.path.join(base_path, relative_path)


COL_IDX = {
    'Segment': 0, 'Format': 1, 'Code': 2, 'ItemType': 3, 'Display': 4,
    'LoopSub': 5, 'ID': 6, 'Label': 7, 'CategoryType': 8, 'CategoryWeight': 9,
    'Digit': 10, 'Min': 11, 'Max': 12, 'AfterDecimal': 13, 'Statistic': 14,  # Col O
    'Conditions': 15,          # Col P
    'SpecialConditions': 16,   # Col Q
    'BaseType': 17             # Col R
}
EXPECTED_COL_COUNT = max(COL_IDX.values()) + 1

NEW_MAKING_OPTIONS = ["TB", "T2B", "BB", "B2B"]                       # 5 scale
NEW_MAKING_OPTIONS_7_10 = ["TB", "T2B", "T3B", "BB", "B2B", "B3B"]    # 7/10 scale
DEFAULT_MAKING_OPTION_CATEGORY = "Valid"

DIR_LOW_GOOD = "Scale น้อยดี(-)"
DIR_HIGH_GOOD = "Scale มากดี(+)"


def strip_leading_code(label_text):
    """
    ตัดเลข Code ที่นำหน้า Label ออก ใช้ "เฉพาะการแสดงผลบนหน้าจอ" เท่านั้น
    เช่น "1 - Don't like at all" -> "- Don't like at all",  "1\tมากที่สุด" -> "มากที่สุด"
    *** ห้ามใช้กับ Label ที่จะเขียนลงไฟล์ Excel เพราะจะทำให้เลข Code หายไป ***
    """
    text = str(label_text or "").strip()
    for sep in ('\t', ' '):
        parts = text.split(sep, 1)
        if len(parts) == 2 and parts[0].isdigit():
            return parts[1].strip()
    return text


# =========================================================================
#  Stylesheet
# =========================================================================
STYLESHEET = """
QWidget {
    background-color: #F5F7FA;
    color: #1A202C;
    font-family: "Segoe UI", "Tahoma", sans-serif;
    font-size: 10pt;
}
QGroupBox {
    background-color: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-radius: 8px;
    margin-top: 14px;
    padding: 10px 12px 12px 12px;
    font-weight: 600;
}
QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    left: 12px;
    padding: 0 6px;
    color: #2D3748;
}
QPushButton {
    background-color: #E2E8F0;
    border: 1px solid #CBD5E0;
    border-radius: 6px;
    padding: 8px 18px;
    font-weight: 600;
}
QPushButton:hover   { background-color: #CBD5E0; }
QPushButton:pressed { background-color: #A0AEC0; }
QPushButton:disabled { background-color: #EDF2F7; color: #A0AEC0; border-color: #E2E8F0; }
QPushButton[accent="blue"]  { background-color: #3182CE; border-color:#2B6CB0; color:#FFFFFF; }
QPushButton[accent="blue"]:hover  { background-color: #2B6CB0; }
QPushButton[accent="green"] { background-color: #38A169; border-color:#2F855A; color:#FFFFFF; }
QPushButton[accent="green"]:hover { background-color: #2F855A; }
QPushButton[accent="gold"]  { background-color: #D69E2E; border-color:#B7791F; color:#FFFFFF; }
QPushButton[accent="gold"]:hover  { background-color: #B7791F; }
QPushButton:disabled[accent="blue"],
QPushButton:disabled[accent="green"],
QPushButton:disabled[accent="gold"] {
    background-color: #EDF2F7; color: #A0AEC0; border-color: #E2E8F0;
}
QTabWidget::pane {
    border: 1px solid #E2E8F0;
    border-radius: 8px;
    background: #FFFFFF;
    top: -1px;
}
QTabBar::tab {
    background: #E2E8F0;
    color: #4A5568;
    padding: 8px 22px;
    margin-right: 3px;
    border-top-left-radius: 7px;
    border-top-right-radius: 7px;
    font-weight: 600;
}
QTabBar::tab:selected { background: #FFFFFF; color: #2B6CB0; border: 1px solid #E2E8F0; border-bottom-color: #FFFFFF; }
QTabBar::tab:hover:!selected { background: #CBD5E0; }
QTableWidget {
    background-color: #FFFFFF;
    border: none;
    gridline-color: #EDF2F7;
    selection-background-color: #BFDFFF;
    selection-color: #1A202C;
}
QHeaderView::section {
    background-color: #EDF2F7;
    color: #2D3748;
    border: none;
    border-right: 1px solid #E2E8F0;
    border-bottom: 1px solid #CBD5E0;
    padding: 7px 8px;
    font-weight: 700;
}
QPlainTextEdit#LogView {
    background-color: #1A202C;
    color: #E2E8F0;
    border: 1px solid #2D3748;
    border-radius: 8px;
    font-family: "Consolas", "Courier New", monospace;
    font-size: 9pt;
}
QLabel#StatusLabel { font-weight: 600; padding: 2px 4px; }
QProgressBar {
    border: 1px solid #CBD5E0;
    border-radius: 5px;
    background: #EDF2F7;
    height: 10px;
    text-align: center;
}
QProgressBar::chunk { background-color: #3182CE; border-radius: 4px; }
QMenu {
    background-color: #FFFFFF;
    border: 1px solid #CBD5E0;
    border-radius: 6px;
    padding: 4px;
}
QMenu::item { padding: 6px 24px 6px 16px; border-radius: 4px; }
QMenu::item:selected { background-color: #BEE3F8; color: #1A365D; }
QMenu::separator { height: 1px; background: #E2E8F0; margin: 4px 8px; }
QScrollBar:vertical   { background: #EDF2F7; width: 12px; border-radius: 6px; }
QScrollBar:horizontal { background: #EDF2F7; height: 12px; border-radius: 6px; }
QScrollBar::handle:vertical, QScrollBar::handle:horizontal {
    background: #A0AEC0; border-radius: 6px; min-height: 24px; min-width: 24px;
}
QScrollBar::handle:vertical:hover, QScrollBar::handle:horizontal:hover { background: #718096; }
QScrollBar::add-line, QScrollBar::sub-line { height: 0; width: 0; }
"""

LEVEL_COLORS = {
    "ERROR": "#E53E3E",
    "WARNING": "#DD6B20",
    "SUCCESS": "#2F855A",
    "INFO": "#2D3748",
}
LOG_COLORS = {
    "ERROR": "#FC8181",
    "WARNING": "#F6AD55",
    "SUCCESS": "#68D391",
    "INFO": "#E2E8F0",
}


# =========================================================================
#  Main Window
# =========================================================================
class ExcelScaleExtractorApp(QtWidgets.QMainWindow):
    SURVEY_COLUMNS = ("Type", "ข้อแบบสอบถาม", "Code เริ่มต้น / Scale",
                      "Code สุดท้าย / Scale", "Direction", "Condition Status")
    SURVEY_HEADERS = ("Type", "ข้อแบบสอบถาม", "Code เริ่มต้น / Scale",
                      "Code สุดท้าย / Scale", "คลิกขวาเลือก Scale", "Condition Status")
    MAKING_COLUMNS = ("Type", "ข้อแบบสอบถาม", "ตัวแปร Original",
                      "Code เริ่มต้น / Scale", "Code สุดท้าย / Scale", "Condition Status")
    MAKING_HEADERS = ("Type", "ตัวแปร Making", "ตัวแปร Original",
                      "Code เริ่มต้น / Scale", "Code สุดท้าย / Scale", "Condition Status")
    DIRECTION_COL_INDEX = 4

    def __init__(self):
        super().__init__()
        self.setWindowTitle("โปรแกรมสร้าง TB/T2B จาก Itemdef  ·  PyQt6 Edition")
        # กว้างพอให้หัวคอลัมน์ภาษาไทยทุกช่องแสดงครบโดยไม่ต้องเลื่อนแนวนอน
        self.resize(1300, 760)

        # --- Instance Variables (เหมือนเวอร์ชันเดิม) ---
        self.excel_data_original_rows = None
        self.file_path = None
        self.scales_data_store = {}
        self.making_data_generated = False
        self.survey_main_row_indices = {}
        self.template_file_path = "template.xlsx"
        self.ordered_survey_qids = []
        self.ordered_making_qids = []

        # map qid -> row index ในตาราง (ใช้แทน iid ของ Treeview เดิม)
        self._survey_rows = {}
        self._making_rows = {}

        self._build_ui()
        self.log_message("Please load an Itemdef Excel file.", level="INFO")

    # ------------------------------------------------------------------
    #  UI
    # ------------------------------------------------------------------
    def _build_ui(self):
        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        root = QtWidgets.QVBoxLayout(central)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(8)

        # ----- Top: ปุ่ม 3 ปุ่ม -----
        top_box = QtWidgets.QGroupBox("ขั้นตอนการทำงาน")
        top_layout = QtWidgets.QHBoxLayout(top_box)
        top_layout.setSpacing(10)
        self.load_button = self._make_button("โหลด Itemdef", "blue", self.load_excel)
        self.generate_button = self._make_button("สร้าง T2B Making", "green", self.run_making_generation)
        self.save_button = self._make_button("Save Itemdef Making...", "gold", self.save_as_new_file)
        self.generate_button.setEnabled(False)
        self.save_button.setEnabled(False)
        for btn in (self.load_button, self.generate_button, self.save_button):
            top_layout.addWidget(btn)
        top_layout.addStretch(1)
        root.addWidget(top_box)

        # ----- Middle: TabWidget + ตาราง -----
        self.tab_view = QtWidgets.QTabWidget()
        self.survey_tree = self._make_table(self.SURVEY_HEADERS,
                                            widths=(100, 90, 220, 220, 150, 140),
                                            center_cols=(4, 5))
        self.making_tree = self._make_table(self.MAKING_HEADERS,
                                            widths=(100, 90, 110, 220, 220, 140),
                                            center_cols=(5,))
        self.tab_view.addTab(self.survey_tree, "ตัวแปร Original")
        self.tab_view.addTab(self.making_tree, "ตัวแปรที่ทำ Making")
        root.addWidget(self.tab_view, 1)

        # คลิกขวาที่คอลัมน์ Direction เพื่อเลือก Scale
        self.survey_tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.survey_tree.customContextMenuRequested.connect(self.on_survey_tree_right_click)

        # ----- Bottom: status + progress + log -----
        self.status_label = QtWidgets.QLabel("")
        self.status_label.setObjectName("StatusLabel")
        root.addWidget(self.status_label)

        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setRange(0, 0)     # indeterminate
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setVisible(False)
        root.addWidget(self.progress_bar)

        log_box = QtWidgets.QGroupBox("Log")
        log_layout = QtWidgets.QVBoxLayout(log_box)
        log_layout.setContentsMargins(8, 8, 8, 8)
        self.log_textbox = QtWidgets.QPlainTextEdit()
        self.log_textbox.setObjectName("LogView")
        self.log_textbox.setReadOnly(True)
        self.log_textbox.setMaximumBlockCount(20000)
        self.log_textbox.setMinimumHeight(120)
        log_layout.addWidget(self.log_textbox)
        root.addWidget(log_box)

    @staticmethod
    def _make_button(text, accent, slot):
        btn = QtWidgets.QPushButton(text)
        btn.setProperty("accent", accent)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.clicked.connect(slot)
        return btn

    @staticmethod
    def _make_table(headers, widths, center_cols=()):
        table = QtWidgets.QTableWidget(0, len(headers))
        table.setHorizontalHeaderLabels(list(headers))
        table.verticalHeader().setVisible(False)
        table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        table.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)
        table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setAlternatingRowColors(False)
        table.verticalHeader().setDefaultSectionSize(26)

        header = table.horizontalHeader()
        # หัวคอลัมน์ใน QSS เป็นตัวหนา ต้องวัดด้วยฟอนต์ตัวหนาจริง ไม่งั้นความกว้างจะขาด
        header_font = QtGui.QFont(header.font())
        header_font.setBold(True)
        metrics = QtGui.QFontMetrics(header_font)
        # เผื่อ padding ซ้าย/ขวาของ QHeaderView::section (7px x 2) + เส้นขอบ + กันเศษ
        header_padding = 26

        for col, width in enumerate(widths):
            needed = metrics.horizontalAdvance(str(headers[col])) + header_padding
            table.setColumnWidth(col, max(width, needed))
            header.setSectionResizeMode(col, QtWidgets.QHeaderView.ResizeMode.Interactive)
        header.setStretchLastSection(True)
        # ให้ผู้ใช้ดับเบิลคลิกเส้นแบ่งเพื่อขยายคอลัมน์ให้พอดีเนื้อหาได้
        header.setCascadingSectionResizes(True)
        table.setProperty("center_cols", list(center_cols))
        table.setProperty("header_min_widths",
                          [metrics.horizontalAdvance(str(h)) + header_padding for h in headers])
        return table

    MAX_AUTO_COL_WIDTH = 340

    def _fit_columns_to_contents(self, table):
        """ขยายคอลัมน์ให้พอดีเนื้อหา แต่ไม่แคบกว่าหัวคอลัมน์ และพยายามให้พอดีจอ

        ลำดับความสำคัญ: หัวคอลัมน์ต้องอ่านครบเสมอ (เป็นพื้นล่างที่ห้ามย่อต่ำกว่า)
        ถ้ารวมแล้วยังกว้างเกินพื้นที่ ค่อยย่อเฉพาะส่วนที่เกินพื้นล่างลงตามสัดส่วน
        """
        column_count = table.columnCount()
        if column_count == 0:
            return
        header_min = list(table.property("header_min_widths") or [])
        floors = [header_min[col] if col < len(header_min) else 60 for col in range(column_count)]

        if table.rowCount():
            table.resizeColumnsToContents()
            widths = [min(max(table.columnWidth(col) + 12, floors[col]), self.MAX_AUTO_COL_WIDTH)
                      for col in range(column_count)]
        else:
            widths = [max(table.columnWidth(col), floors[col]) for col in range(column_count)]

        available = table.viewport().width()
        overflow = sum(widths) - available
        if overflow > 0 and available > 0:
            # ย่อได้เท่าที่แต่ละคอลัมน์เกินพื้นล่างของมันเท่านั้น
            slack = [widths[col] - floors[col] for col in range(column_count)]
            total_slack = sum(slack)
            if total_slack > 0:
                shrink = min(overflow, total_slack)
                for col in range(column_count):
                    if slack[col] > 0:
                        widths[col] -= round(shrink * slack[col] / total_slack)
                # เก็บเศษจากการปัดเศษ ไม่ให้เหลือเกินมา 1-2px จนมี scrollbar
                leftover = sum(widths) - available
                while leftover > 0:
                    widest = max(range(column_count), key=lambda c: widths[c] - floors[c])
                    if widths[widest] - floors[widest] <= 0:
                        break
                    take = min(leftover, widths[widest] - floors[widest])
                    widths[widest] -= take
                    leftover -= take

        for col in range(column_count):
            table.setColumnWidth(col, widths[col])

    def resizeEvent(self, event):
        super().resizeEvent(event)
        # ย่อ/ขยายคอลัมน์ตามขนาดหน้าต่างที่เปลี่ยนไป
        for table in (getattr(self, 'survey_tree', None), getattr(self, 'making_tree', None)):
            if table is not None and table.rowCount():
                self._fit_columns_to_contents(table)

    # ------------------------------------------------------------------
    #  ตัวช่วยเข้าถึงตาราง (แทน API ของ ttk.Treeview เดิม)
    # ------------------------------------------------------------------
    def _row_map(self, tree):
        return self._survey_rows if tree is self.survey_tree else self._making_rows

    def _columns_of(self, tree):
        return self.SURVEY_COLUMNS if tree is self.survey_tree else self.MAKING_COLUMNS

    def tree_exists(self, tree, qid):
        return qid in self._row_map(tree)

    def tree_set(self, tree, qid, column_name, value):
        row_map = self._row_map(tree)
        if qid not in row_map:
            return
        col = self._columns_of(tree).index(column_name)
        item = tree.item(row_map[qid], col)
        if item is not None:
            item.setText(str(value))

    def tree_get(self, tree, qid, column_name):
        row_map = self._row_map(tree)
        if qid not in row_map:
            return ""
        col = self._columns_of(tree).index(column_name)
        item = tree.item(row_map[qid], col)
        return item.text() if item is not None else ""

    def tree_clear(self, tree):
        tree.setRowCount(0)
        self._row_map(tree).clear()

    # ------------------------------------------------------------------
    #  Logging  (พฤติกรรมเหมือนเดิม)
    # ------------------------------------------------------------------
    def log_message(self, message, level="INFO"):
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        formatted_message = f"[{timestamp} {level}] {message}"
        color = LOG_COLORS.get(level, LOG_COLORS["INFO"])
        safe = (str(formatted_message)
                .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
        self.log_textbox.appendHtml(f'<span style="color:{color};">{safe}</span>')
        self.log_textbox.ensureCursorVisible()

        status_prefix = f"[{level}] " if level != "INFO" else ""
        self.set_status(f"{status_prefix}{message}", LEVEL_COLORS.get(level, LEVEL_COLORS["INFO"]))
        print(formatted_message)

    def set_status(self, text, color=None):
        self.status_label.setText(text)
        if color:
            self.status_label.setStyleSheet(f"color:{color}; font-weight:600; padding:2px 4px;")
        QtWidgets.QApplication.processEvents()

    # ------------------------------------------------------------------
    #  UX Helpers
    # ------------------------------------------------------------------
    def _start_processing(self, status_message="Processing..."):
        QtWidgets.QApplication.setOverrideCursor(QtGui.QCursor(Qt.CursorShape.WaitCursor))
        self.load_button.setEnabled(False)
        self.generate_button.setEnabled(False)
        self.save_button.setEnabled(False)
        self.set_status(status_message, LEVEL_COLORS["INFO"])
        self.progress_bar.setVisible(True)
        QtWidgets.QApplication.processEvents()

    def _end_processing(self, final_status_message=None, level="INFO"):
        self.progress_bar.setVisible(False)
        QtWidgets.QApplication.restoreOverrideCursor()
        if final_status_message:
            self.log_message(final_status_message, level=level)
        self.load_button.setEnabled(True)
        if self.file_path and self.excel_data_original_rows is not None:
            self.generate_button.setEnabled(True)
            self.save_button.setEnabled(bool(self.making_data_generated))
        else:
            self.generate_button.setEnabled(False)
            self.save_button.setEnabled(False)
        QtWidgets.QApplication.processEvents()

    def show_error_popup(self, title, message):
        QtWidgets.QMessageBox.critical(self, title, message)

    def show_warning_popup(self, title, message):
        QtWidgets.QMessageBox.warning(self, title, message)

    # ------------------------------------------------------------------
    #  Load Excel
    # ------------------------------------------------------------------
    def load_excel(self):
        self.set_status("Selecting file...", LEVEL_COLORS["INFO"])
        selected_file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Select Itemdef Excel File", "", "Excel files (*.xlsx *.xls);;All files (*.*)")
        if not selected_file_path:
            self.set_status("Ready. Please load an Excel file.", LEVEL_COLORS["INFO"])
            self.log_message("File selection cancelled.", level="INFO")
            return

        selected_file_path = os.path.normpath(selected_file_path)
        try:
            base_filename = os.path.basename(selected_file_path)
        except Exception as e:
            self.log_message(f"Error getting basename from selected path: {e}", level="WARNING")
            base_filename = "Selected File"

        self.clear_results()
        self.file_path = selected_file_path
        self._start_processing(f"Loading: {base_filename}...")
        success = False
        final_message = ""
        level = "INFO"

        try:
            self.log_message(f"Reading Excel file: {self.file_path}", level="INFO")
            start_time = time.time()
            try:
                df_full = pd.read_excel(self.file_path, header=None, sheet_name=0,
                                        keep_default_na=False, na_values=['']).fillna('')
            except Exception as read_err:
                raise ValueError(f"Cannot read Excel file: [{type(read_err).__name__}] {read_err}")
            read_time = time.time()

            if len(df_full.columns) < EXPECTED_COL_COUNT:
                raise ValueError(f"File structure error: Expected at least {EXPECTED_COL_COUNT} "
                                 f"columns, found {len(df_full.columns)}.")

            self.excel_data_original_rows = df_full.values.tolist()
            self.log_message(f"Excel file read successfully ({read_time - start_time:.2f}s).", level="INFO")
            self.set_status(f"Processing data from '{base_filename}'...")

            self.log_message("Processing data...", level="INFO")
            self.process_and_find_indices_combined()
            process_time = time.time()
            self.log_message(f"Data processed ({process_time - read_time:.2f}s). "
                             f"Found {len(self.ordered_survey_qids)} potential scales.", level="INFO")

            self.set_status("Populating survey view...")
            self.log_message("Populating ตัวแปร Original view...", level="INFO")

            survey_items_ordered = []
            for qid in self.ordered_survey_qids:
                if qid in self.scales_data_store:
                    data = self.scales_data_store[qid]
                    if not data.get('is_making'):
                        survey_items_ordered.append((qid, data))
                else:
                    self.log_message(f"Warning: Ordered QID '{qid}' not found in data store during display.",
                                     "WARNING")

            self._populate_treeview(survey_items_ordered, target_tree=self.survey_tree)
            self.tab_view.setCurrentIndex(0)

            if self.survey_tree.rowCount():
                num_found = self.survey_tree.rowCount()
                plural = "s" if num_found > 1 else ""
                final_message = (f"Load complete. Displayed {num_found} survey scale{plural}. "
                                 "Right-click 'Direction' cell to edit.")
                level = "SUCCESS"
            else:
                final_message = "Load complete. No valid ตัวแปร Original (5, 7, or 10 options) found."
                level = "WARNING"
            success = True
            self.making_data_generated = False

        except FileNotFoundError:
            final_message = f"Error: File '{base_filename}' not found."
            level = "ERROR"
            self.show_error_popup("File Not Found", f"{final_message}\nPlease check the path:\n{selected_file_path}")
            self.clear_results()
        except ValueError as ve:
            final_message = f"Error processing file: {ve}"
            level = "ERROR"
            self.show_error_popup("Data Structure/Format Error",
                                  f"{final_message}\nPlease ensure '{base_filename}' has the correct format.")
            self.clear_results()
        except Exception as e:
            final_message = f"An unexpected error occurred: [{type(e).__name__}] {e}"
            level = "ERROR"
            self.show_error_popup("Loading Error",
                                  f"{final_message}\nCheck log/console for details regarding '{base_filename}'.")
            print(traceback.format_exc())
            self.clear_results()
        finally:
            self._end_processing(final_status_message=final_message,
                                 level=level if success or level == "ERROR" else "INFO")

    # ------------------------------------------------------------------
    #  Clear Results
    # ------------------------------------------------------------------
    def clear_results(self):
        self.tree_clear(self.survey_tree)
        self.tree_clear(self.making_tree)
        self.scales_data_store = {}
        self.excel_data_original_rows = None
        self.survey_main_row_indices = {}
        self.making_data_generated = False
        self.file_path = None
        self.ordered_survey_qids = []
        self.ordered_making_qids = []
        self.log_textbox.clear()
        self.generate_button.setEnabled(False)
        self.save_button.setEnabled(False)
        self.log_message("Results cleared. Please load an Itemdef Excel file.", level="INFO")

    # ------------------------------------------------------------------
    #  Process Survey Data & Find Indices  (เหมือนเดิมทุกบรรทัด)
    # ------------------------------------------------------------------
    def process_and_find_indices_combined(self):
        if self.excel_data_original_rows is None:
            self.log_message("Cannot process data: Original Excel data not loaded.", level="ERROR")
            return

        self.scales_data_store = defaultdict(lambda: {
            'type': 'N/A', 'sub_labels': [], 'scale_options': [], 'direction': '',
            'is_making': False, 'conditions': None, 'condition_status': ''
        })
        self.survey_main_row_indices = {}
        potential_main_row_indices = {}
        self.ordered_survey_qids = []
        self.ordered_making_qids = []
        current_main_id = None
        current_main_is_loop = False
        row_errors = []

        for idx, row in enumerate(self.excel_data_original_rows):
            if idx < 2:
                continue   # Skip headers

            try:
                if len(row) <= max(COL_IDX.values()):
                    row_errors.append(f"Row {idx + 1}: Insufficient columns "
                                      f"(found {len(row)}, expected {EXPECTED_COL_COUNT}).")
                    continue

                id_val = str(row[COL_IDX['ID']]).strip()
                item_type_val = str(row[COL_IDX['ItemType']]).strip()
                label_val = str(row[COL_IDX['Label']]).strip()
                loop_sub_val = str(row[COL_IDX['LoopSub']]).strip()

                is_main_id = (id_val and not id_val.isdigit() and '(' not in id_val
                              and id_val not in ['SbjNum', 'F_Status', 'team_FW'])

                if is_main_id:
                    current_main_id = id_val
                    stored_type = item_type_val if item_type_val else \
                        self.scales_data_store[current_main_id].get('type', 'N/A')
                    self.scales_data_store[current_main_id]['type'] = stored_type
                    self.scales_data_store[current_main_id]['is_making'] = False
                    current_main_is_loop = 'Loop' in stored_type
                    if item_type_val:
                        potential_main_row_indices[current_main_id] = idx

                elif current_main_id and label_val:
                    is_sub_label = current_main_is_loop and '(' in id_val and ')' in id_val
                    is_option = (not id_val or id_val.isdigit()) and \
                                not (current_main_is_loop and loop_sub_val == "Loop sub")

                    if is_sub_label:
                        self.scales_data_store[current_main_id]['sub_labels'].append(label_val)
                    elif is_option:
                        # FIX: เก็บ Label ตามต้นฉบับทั้งหมด ห้ามตัดเลข Code นำหน้าออก
                        # (ของเดิมตัด "1 " ออกจาก "1 - Don't like at all" ทำให้เลขหายไปในไฟล์ผลลัพธ์)
                        scale_text = label_val
                        if scale_text and str(scale_text).strip():
                            self.scales_data_store[current_main_id]['scale_options'].append(
                                str(scale_text).strip())

            except IndexError:
                row_errors.append(f"Row {idx + 1}: Data access error (IndexError).")
                continue
            except Exception as ex:
                row_errors.append(f"Row {idx + 1}: Unexpected error ({type(ex).__name__}).")
                continue

        validated_qids_in_order = []
        for qid, idx in potential_main_row_indices.items():
            if qid in self.scales_data_store:
                num_options = len(self.scales_data_store[qid].get('scale_options', []))
                if num_options in [5, 7, 10]:
                    self.survey_main_row_indices[qid] = idx
                    validated_qids_in_order.append(qid)

        self.ordered_survey_qids = validated_qids_in_order

        if row_errors:
            self.log_message(f"Processing completed with {len(row_errors)} potential issues.", level="WARNING")

    # ------------------------------------------------------------------
    #  Populate table
    # ------------------------------------------------------------------
    def _populate_treeview(self, items_to_display, target_tree):
        tree_name = 'Survey' if target_tree is self.survey_tree else 'Making'
        self.log_message(f"Populating {tree_name} view...", level="INFO")
        displayed_count = 0

        self.tree_clear(target_tree)
        row_map = self._row_map(target_tree)
        center_cols = set(target_tree.property("center_cols") or [])
        row_bg = QtGui.QBrush(QtGui.QColor("#FFFFFF" if target_tree is self.survey_tree else "#E6F2FF"))

        target_tree.setUpdatesEnabled(False)
        try:
            for q_id, data in items_to_display:
                is_making = data.get('is_making', False)
                first_label = "N/A"
                last_label = "N/A"
                scale_options = data.get('scale_options', [])
                valid_scale_options = [opt for opt in scale_options if opt is not None and str(opt).strip()]
                original_num_scale_options = len(valid_scale_options)

                if valid_scale_options:
                    # แสดงผลบนตารางเท่านั้น -> ตัดเลข Code นำหน้าออกเพื่อไม่ให้ซ้ำกับ "1 = ..."
                    first_label = strip_leading_code(valid_scale_options[0]).replace('\t', ' ')
                    last_label = strip_leading_code(valid_scale_options[-1]).replace('\t', ' ')

                if original_num_scale_options not in [5, 7, 10]:
                    continue

                displayed_count += 1
                item_type = data.get('type', "N/A")
                condition_status = data.get('condition_status', '')
                first_val_text = f"1 = {first_label}"
                last_val_text = f"{original_num_scale_options} = {last_label}"

                if target_tree is self.survey_tree:
                    direction = data.get('direction', '')
                    if not is_making and not direction and condition_status != "Making Generated":
                        condition_status = "ไม่ทำ Making"
                    display_values = (item_type, q_id, first_val_text, last_val_text,
                                      direction, condition_status)
                else:
                    original_q_id = data.get('original_q_id', 'N/A')
                    display_values = (item_type, q_id, original_q_id, first_val_text,
                                      last_val_text, condition_status)

                row = target_tree.rowCount()
                target_tree.insertRow(row)
                for col, value in enumerate(display_values):
                    cell = QtWidgets.QTableWidgetItem(str(value))
                    if col in center_cols:
                        cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    cell.setBackground(row_bg)
                    target_tree.setItem(row, col, cell)
                row_map[q_id] = row
        finally:
            target_tree.setUpdatesEnabled(True)

        self._fit_columns_to_contents(target_tree)
        self.log_message(f"Finished populating {tree_name} view ({displayed_count} items displayed).",
                         level="INFO")

    # ------------------------------------------------------------------
    #  Generate Conditions  (เหมือนเดิมทุกบรรทัด)
    # ------------------------------------------------------------------
    def _generate_conditions(self, q_id, direction, num_labels):
        conditions = []
        status = "OK"
        base_q_id = q_id
        for i in range(1, num_labels + 1):
            conditions.append(f"{base_q_id}={i}")

        if num_labels == 5:
            num_new_options = 4
        elif num_labels in [7, 10]:
            num_new_options = 6
        else:
            return [], f"Invalid Scale ({num_labels})"

        no_dir_placeholder = ["NO_DIRECTION"] * num_new_options
        generated_new_conditions = []

        if direction == DIR_LOW_GOOD:
            if num_labels == 5:
                l1, l2, h1, h2 = 1, 2, 5, 4
                generated_new_conditions = [
                    f"{base_q_id}={l1}",
                    f"{base_q_id}={l1}|{base_q_id}={l2}",
                    f"{base_q_id}={h1}",
                    f"{base_q_id}={h2}|{base_q_id}={h1}"]
            elif num_labels in (7, 10):
                l1, l2, l3 = 1, 2, 3
                h1, h2, h3 = (7, 6, 5) if num_labels == 7 else (10, 9, 8)
                generated_new_conditions = [
                    f"{base_q_id}={l1}",
                    f"{base_q_id}={l1}|{base_q_id}={l2}",
                    f"{base_q_id}={l1}|{base_q_id}={l2}|{base_q_id}={l3}",
                    f"{base_q_id}={h1}",
                    f"{base_q_id}={h2}|{base_q_id}={h1}",
                    f"{base_q_id}={h3}|{base_q_id}={h2}|{base_q_id}={h1}"]
        elif direction == DIR_HIGH_GOOD:
            if num_labels == 5:
                l1, l2, h1, h2 = 1, 2, 5, 4
                generated_new_conditions = [
                    f"{base_q_id}={h1}",
                    f"{base_q_id}={h1}|{base_q_id}={h2}",
                    f"{base_q_id}={l1}",
                    f"{base_q_id}={l1}|{base_q_id}={l2}"]
            elif num_labels in (7, 10):
                l1, l2, l3 = 1, 2, 3
                h1, h2, h3 = (7, 6, 5) if num_labels == 7 else (10, 9, 8)
                generated_new_conditions = [
                    f"{base_q_id}={h1}",
                    f"{base_q_id}={h1}|{base_q_id}={h2}",
                    f"{base_q_id}={h1}|{base_q_id}={h2}|{base_q_id}={h3}",
                    f"{base_q_id}={l1}",
                    f"{base_q_id}={l1}|{base_q_id}={l2}",
                    f"{base_q_id}={l1}|{base_q_id}={l2}|{base_q_id}={l3}"]
        else:
            status = "No Direction"
            generated_new_conditions.extend(no_dir_placeholder)

        conditions.extend(generated_new_conditions)
        expected_len = num_labels + num_new_options
        if len(conditions) != expected_len:
            self.log_message(f"CRITICAL ERROR: Condition length mismatch for {q_id}! "
                             f"Expected {expected_len}, got {len(conditions)}.", level="ERROR")
            status = "Length Error"
            while len(conditions) < expected_len:
                conditions.append("ERR_LEN_COND")
            conditions = conditions[:expected_len]
        return conditions, status

    # ------------------------------------------------------------------
    #  Run Making Generation
    # ------------------------------------------------------------------
    def run_making_generation(self):
        if not self.scales_data_store or not self.survey_main_row_indices:
            self.log_message("Cannot Generate: No survey data loaded or no valid scales found.", level="WARNING")
            self.show_warning_popup("Cannot Generate",
                                    "No survey data loaded or no valid scales found. Please load a file first.")
            return

        items_requiring_direction = []
        q_ids_to_process = []

        for q_id in self.ordered_survey_qids:
            if q_id in self.scales_data_store:
                data = self.scales_data_store[q_id]
                if not data.get('is_making'):
                    if not data.get('direction'):
                        items_requiring_direction.append(q_id)
                    else:
                        q_ids_to_process.append(q_id)

        if items_requiring_direction:
            num_missing = len(items_requiring_direction)
            plural = "s" if num_missing > 1 else ""
            self.log_message(f"{num_missing} item{plural} need direction.", level="WARNING")
            self.show_warning_popup("Direction Needed",
                                    f"{num_missing} ตัวแปร 5/7/10 Scale.\n\nที่ไม่ได้ทำ Marking")
            for qid in items_requiring_direction:
                if self.tree_exists(self.survey_tree, qid):
                    self.tree_set(self.survey_tree, qid, "Condition Status", "ไม่ทำ Making")

        if not q_ids_to_process:
            self.log_message("No items ready for Making generation (check directions).", level="WARNING")
            return

        self._start_processing(f"Generating Making data for {len(q_ids_to_process)} item(s)...")

        making_items_store = {}
        items_with_errors = []
        generated_ids_this_run = set()
        generated_making_qids_ordered_this_run = []
        final_status_message = ""
        level = "INFO"

        try:
            total_items = len(q_ids_to_process)
            for i, q_id in enumerate(q_ids_to_process):
                if i % 5 == 0 or i == total_items - 1:
                    self.set_status(f"Processing item {i + 1} of {total_items}...")

                data = self.scales_data_store.get(q_id)
                if not data or data.get('is_making', False) or not data.get('direction'):
                    self.log_message(f"Skipping {q_id} during generation (unexpected state).", level="WARNING")
                    continue

                original_scale_options = data.get('scale_options', [])
                valid_original_scale_options = [opt for opt in original_scale_options
                                                if opt is not None and str(opt).strip()]
                original_num_scale_options = len(valid_original_scale_options)

                if original_num_scale_options not in [5, 7, 10]:
                    self.log_message(f"Skipping {q_id}: Invalid number of valid scale options "
                                     f"({original_num_scale_options}) found during generation.", level="WARNING")
                    items_with_errors.append(f"{q_id} (Invalid Scale Opts: {original_num_scale_options})")
                    self.tree_set(self.survey_tree, q_id, "Condition Status",
                                  f"Err: Invalid Opts ({original_num_scale_options})")
                    continue

                direction = data.get('direction')
                making_q_id = f"N{q_id}"

                try:
                    original_type = data.get('type', 'N/A')
                    new_type = original_type
                    if 'Loop(SA)' in original_type:
                        new_type = original_type.replace('Loop(SA)', 'Loop(MA)', 1)
                    elif 'SA' in original_type:
                        new_type = original_type.replace('SA', 'MA', 1)

                    current_new_options_list = (NEW_MAKING_OPTIONS_7_10
                                                if original_num_scale_options in [7, 10]
                                                else NEW_MAKING_OPTIONS)
                    sub_labels = data.get('sub_labels', [])
                    new_labels = sub_labels + valid_original_scale_options + current_new_options_list
                    conditions_list, condition_status = self._generate_conditions(
                        q_id, direction, original_num_scale_options)

                    if condition_status != "OK":
                        self.log_message(f"ERROR generating conditions for {q_id}: {condition_status}",
                                         level="ERROR")
                        items_with_errors.append(f"{q_id} (Condition: {condition_status})")
                        self.tree_set(self.survey_tree, q_id, "Condition Status",
                                      f"Cond. Error: {condition_status}")
                        continue

                    making_items_store[making_q_id] = {
                        'type': new_type,
                        'sub_labels': sub_labels,
                        'scale_options': valid_original_scale_options,
                        'labels': new_labels,
                        'direction': '',
                        'conditions': conditions_list,
                        'condition_status': condition_status,
                        'is_making': True,
                        'original_q_id': q_id
                    }
                    if making_q_id not in generated_ids_this_run:
                        generated_ids_this_run.add(making_q_id)
                        generated_making_qids_ordered_this_run.append(making_q_id)
                    self.tree_set(self.survey_tree, q_id, "Condition Status", "Making Generated")

                except Exception as gen_err:
                    error_type_name = type(gen_err).__name__
                    self.log_message(f"CRITICAL ERROR generating {making_q_id}: {error_type_name} - {gen_err}",
                                     level="ERROR")
                    items_with_errors.append(f"{q_id} (Runtime Error: {error_type_name})")
                    self.tree_set(self.survey_tree, q_id, "Condition Status", f"Gen Error: {error_type_name}")

            self.set_status("Processing complete. Preparing results...")
            self.scales_data_store.update(making_items_store)
            total_generated = len(generated_ids_this_run)
            self.ordered_making_qids = generated_making_qids_ordered_this_run

            if total_generated > 0:
                self.making_data_generated = True
                all_making_items_ordered = []
                for making_qid in self.ordered_making_qids:
                    if making_qid in self.scales_data_store:
                        data = self.scales_data_store[making_qid]
                        if data.get('is_making'):
                            all_making_items_ordered.append((making_qid, data))
                    else:
                        self.log_message(f"Warning: Ordered Making QID '{making_qid}' not found "
                                         "in data store during making display.", "WARNING")
                self._populate_treeview(all_making_items_ordered, target_tree=self.making_tree)
                final_status_message = f"Generated/Updated {total_generated} 'Making' items."
                level = "SUCCESS"
                self.tab_view.setCurrentIndex(1)
            else:
                if not items_with_errors:
                    final_status_message = "No new 'Making' items were generated."
                if not any(d.get('is_making') for d in self.scales_data_store.values()):
                    self.making_data_generated = False

            if items_with_errors:
                final_status_message += f" Encountered errors on {len(items_with_errors)} item(s)."
                level = "WARNING"

        except Exception as e:
            final_status_message = f"Generation process failed: [{type(e).__name__}] {e}"
            level = "ERROR"
            self.show_error_popup("Generation Error", f"{final_status_message}\nCheck log for details.")
            print(traceback.format_exc())
            self.making_data_generated = False
        finally:
            self._end_processing(final_status_message=final_status_message, level=level)

    # ------------------------------------------------------------------
    #  Generate Making Rows for Excel  (เหมือนเดิมทุกบรรทัด)
    # ------------------------------------------------------------------
    def _generate_making_rows_for_excel(self, original_q_id, making_data, original_survey_row):
        output_rows = []
        making_q_id = f"N{original_q_id}"
        sub_labels = making_data.get('sub_labels', [])
        scale_options = making_data.get('scale_options', [])
        conditions = making_data.get('conditions', [])
        making_type = making_data.get('type', 'N/A')

        num_sub_labels = len(sub_labels)
        original_num_scale_options = len(scale_options)
        is_loop_making = 'Loop' in making_type

        if original_num_scale_options in [7, 10]:
            current_new_options_list = NEW_MAKING_OPTIONS_7_10
        elif original_num_scale_options == 5:
            current_new_options_list = NEW_MAKING_OPTIONS
        else:
            self.log_message(f"ERROR: Cannot generate Excel rows for {making_q_id}. "
                             f"Invalid scale option count ({original_num_scale_options}).", level="ERROR")
            return []
        num_new_options = len(current_new_options_list)

        expected_conditions_count = original_num_scale_options + num_new_options
        if len(conditions) != expected_conditions_count:
            self.log_message(f"WARNING: Condition count mismatch for {making_q_id} "
                             f"({len(conditions)}/{expected_conditions_count}).", level="WARNING")

        # ----- แถวหลักของ Making -----
        main_making_row = [''] * EXPECTED_COL_COUNT
        main_making_row[COL_IDX['Segment']] = (original_survey_row[COL_IDX['Segment']]
                                               if len(original_survey_row) > COL_IDX['Segment'] else "Item")
        main_making_row[COL_IDX['Format']] = "Making"
        main_making_row[COL_IDX['ItemType']] = making_type
        main_making_row[COL_IDX['Display']] = (original_survey_row[COL_IDX['Display']]
                                               if len(original_survey_row) > COL_IDX['Display'] else "O")
        main_making_row[COL_IDX['ID']] = making_q_id

        # คัดลอก Label จากคำถาม Survey ต้นฉบับ
        if len(original_survey_row) > COL_IDX['Label']:
            main_making_row[COL_IDX['Label']] = original_survey_row[COL_IDX['Label']]

        main_making_row[COL_IDX['Statistic']] = str(original_q_id).upper()
        main_making_row[COL_IDX['Conditions']] = ''
        main_making_row[COL_IDX['BaseType']] = "Follow the condition items"
        if is_loop_making and len(original_survey_row) > COL_IDX['LoopSub'] and original_survey_row[COL_IDX['LoopSub']]:
            main_making_row[COL_IDX['LoopSub']] = original_survey_row[COL_IDX['LoopSub']]
        output_rows.append(main_making_row)

        # ----- หา Category Type จากตัวเลือกแรกของคำถามต้นฉบับ -----
        option_category_type = DEFAULT_MAKING_OPTION_CATEGORY
        if original_q_id in self.survey_main_row_indices:
            survey_main_idx = self.survey_main_row_indices[original_q_id]
            first_option_idx_0based = survey_main_idx + 1 + num_sub_labels
            if first_option_idx_0based < len(self.excel_data_original_rows or []):
                first_option_row = self.excel_data_original_rows[first_option_idx_0based]
                if len(first_option_row) > max(COL_IDX['ID'], COL_IDX['Label'], COL_IDX['CategoryType']):
                    is_option_row = (str(first_option_row[COL_IDX['ID']]).strip().isdigit() or
                                     (not str(first_option_row[COL_IDX['ID']]).strip()
                                      and str(first_option_row[COL_IDX['Label']]).strip()))
                    if is_option_row and first_option_row[COL_IDX['CategoryType']]:
                        option_category_type = first_option_row[COL_IDX['CategoryType']]

        # ----- Loop sub -----
        if is_loop_making:
            for i, sub_label in enumerate(sub_labels):
                sub_label_row = [''] * EXPECTED_COL_COUNT
                sub_label_row[COL_IDX['LoopSub']] = "Loop sub"
                sub_label_row[COL_IDX['ID']] = f"{making_q_id}({i + 1})"
                sub_label_row[COL_IDX['Label']] = sub_label
                output_rows.append(sub_label_row)

        # ----- ตัวเลือกเดิม -----
        overall_option_index = 0
        for i in range(original_num_scale_options):
            overall_option_index += 1
            option_row = [''] * EXPECTED_COL_COUNT
            option_row[COL_IDX['ID']] = overall_option_index
            option_row[COL_IDX['Label']] = scale_options[i]
            option_row[COL_IDX['CategoryType']] = option_category_type
            option_row[COL_IDX['Conditions']] = conditions[i] if i < len(conditions) else "ERR_MISSING_COND"
            if is_loop_making:
                option_row[COL_IDX['LoopSub']] = ""
            output_rows.append(option_row)

        # ----- ตัวเลือกใหม่ TB/T2B/... -----
        start_index_new_conditions = original_num_scale_options
        for i, new_label in enumerate(current_new_options_list):
            overall_option_index += 1
            option_row = [''] * EXPECTED_COL_COUNT
            option_row[COL_IDX['ID']] = overall_option_index
            option_row[COL_IDX['Label']] = new_label
            option_row[COL_IDX['CategoryType']] = option_category_type
            condition_index = start_index_new_conditions + i
            option_row[COL_IDX['Conditions']] = (conditions[condition_index]
                                                 if condition_index < len(conditions) else "ERR_MISSING_COND")
            if is_loop_making:
                option_row[COL_IDX['LoopSub']] = ""
            output_rows.append(option_row)

        return output_rows

    # ------------------------------------------------------------------
    #  Save As New File
    # ------------------------------------------------------------------
    def save_as_new_file(self):
        if not self.making_data_generated:
            self.log_message("Cannot Save: No 'Making' data generated.", level="WARNING")
            self.show_error_popup("Cannot Save", "No 'Making' data has been generated yet.")
            return
        if not self.file_path:
            self.log_message("Cannot Save: Original file path missing.", level="ERROR")
            self.show_error_popup("Cannot Save", "Original file path is missing.")
            return
        if not self.excel_data_original_rows:
            self.log_message("Cannot Save: Internal data missing (original rows).", level="ERROR")
            self.show_error_popup("Cannot Save", "Internal data missing (original rows). Please reload the file.")
            return
        if not self.survey_main_row_indices:
            self.log_message("Cannot Save: Internal data missing (survey indices).", level="ERROR")
            self.show_error_popup("Cannot Save",
                                  "Internal data missing (survey indices). Please reload and regenerate.")
            return

        try:
            if not self.template_file_path or not os.path.exists(self.template_file_path):
                missing_path = self.template_file_path or "Not Set"
                self.log_message(f"Template file not found or path not set: {missing_path}", level="ERROR")
                self.show_error_popup("Template Not Found",
                                      f"Template file not found or path invalid:\n{missing_path}")
                return
            directory = os.path.dirname(self.file_path)
            original_filename = os.path.basename(self.file_path)
            original_filename_no_ext, original_extension = os.path.splitext(original_filename)
            if not original_extension.lower().startswith(".xls"):
                original_extension = ".xlsx"
            new_filename = f"{original_filename_no_ext}_making{original_extension}"
            new_file_path = os.path.join(directory, new_filename)
        except Exception as path_err:
            self.log_message(f"Error preparing file paths: {path_err}", level="ERROR")
            self.show_error_popup("Path Error", f"Error preparing file paths:\n\n{path_err}")
            return

        confirm = QtWidgets.QMessageBox.question(
            self, "Confirm Save As New File",
            f"Create NEW file:\n'{new_filename}'?\n\n"
            f"Original file ('{original_filename}') will NOT be changed.")
        if confirm != QtWidgets.QMessageBox.StandardButton.Yes:
            self.log_message("Save As operation cancelled by user.", level="INFO")
            return

        self._start_processing(f"Saving data to '{new_filename}'...")
        total_start_time = time.time()
        workbook = None
        final_rows_to_write_local = []
        final_message = "Save process finished."
        level = "INFO"

        try:
            self.set_status("Building data in memory...")
            start_prep_time = time.time()
            self.log_message("Building final data rows in memory...", level="INFO")

            if not self.excel_data_original_rows:
                raise ValueError("Original data missing before prep.")
            current_rows = [row[:] for row in self.excel_data_original_rows]
            making_blocks_info = {}
            original_qids_processed_for_making = set()
            data_prep_errors = []

            items_to_process_for_save = {qid: data for qid, data in self.scales_data_store.items()
                                         if data.get('is_making')}
            total_items = len(items_to_process_for_save)

            for i, (making_q_id, data) in enumerate(items_to_process_for_save.items()):
                if i % 10 == 0 or i == total_items - 1:
                    self.set_status(f"Preparing item {i + 1} of {total_items}...")

                original_q_id = data.get('original_q_id')
                if not original_q_id:
                    data_prep_errors.append(f"{making_q_id}: Missing original ID.")
                    continue

                if original_q_id in self.survey_main_row_indices:
                    original_row_idx = self.survey_main_row_indices[original_q_id]
                    if original_row_idx < len(current_rows):
                        try:
                            making_rows = self._generate_making_rows_for_excel(
                                original_q_id, data, current_rows[original_row_idx])
                            if not making_rows:
                                data_prep_errors.append(f"{making_q_id}: Row generation returned empty list.")
                                continue
                            num_subs = len(data.get('sub_labels', []))
                            num_opts = len(data.get('scale_options', []))
                            insert_idx = original_row_idx + 1 + num_subs + num_opts
                            making_blocks_info[original_q_id] = {'rows': making_rows,
                                                                 'insert_at_idx_0based': insert_idx}
                            original_qids_processed_for_making.add(original_q_id)
                        except Exception as e:
                            data_prep_errors.append(f"{making_q_id}: Row gen failed ({type(e).__name__} - {e}).")
                    else:
                        data_prep_errors.append(f"{original_q_id}: Original index {original_row_idx} "
                                                f"out of bounds ({len(current_rows)}).")
                else:
                    data_prep_errors.append(f"{making_q_id}: Linked survey index missing for '{original_q_id}'.")

            # ----- ใส่ Weight ให้ตัวเลือกของคำถามต้นฉบับ -----
            self.set_status("Processing weights...")
            TARGET_WEIGHT_COL_IDX = COL_IDX['CategoryWeight']
            weights_applied_count = 0

            for original_q_id in original_qids_processed_for_making:
                survey_data = self.scales_data_store.get(original_q_id)
                if not survey_data:
                    continue
                direction = survey_data.get('direction', '')
                scale_options_for_weight = survey_data.get('scale_options', [])
                num_options = len(scale_options_for_weight)

                if num_options > 0 and direction in [DIR_LOW_GOOD, DIR_HIGH_GOOD] \
                        and original_q_id in self.survey_main_row_indices:
                    num_sub_labels = len(survey_data.get('sub_labels', []))
                    main_row_idx = self.survey_main_row_indices[original_q_id]
                    first_option_row_idx = main_row_idx + 1 + num_sub_labels
                    weights = (list(range(num_options, 0, -1)) if direction == DIR_LOW_GOOD
                               else list(range(1, num_options + 1)))
                    applied_flag = False

                    for i in range(num_options):
                        target_row_idx = first_option_row_idx + i
                        if target_row_idx < len(current_rows) and \
                                len(current_rows[target_row_idx]) > TARGET_WEIGHT_COL_IDX:
                            try:
                                current_rows[target_row_idx][TARGET_WEIGHT_COL_IDX] = weights[i]
                                applied_flag = True
                            except IndexError:
                                data_prep_errors.append(
                                    f"{original_q_id}: Weight apply IndexError row {target_row_idx + 1}.")
                            except TypeError:
                                data_prep_errors.append(
                                    f"{original_q_id}: Weight apply TypeError row {target_row_idx + 1}.")
                        else:
                            data_prep_errors.append(
                                f"{original_q_id}: Weight target row {target_row_idx + 1} invalid or too short.")
                    if applied_flag:
                        weights_applied_count += 1

            # ----- แทรกบล็อก Making -----
            self.set_status("Finalizing data structure...")
            if making_blocks_info:
                final_rows_after_insertion = []
                sorted_keys = sorted(making_blocks_info.keys(),
                                     key=lambda q: making_blocks_info[q]['insert_at_idx_0based'])
                last_copied_idx = 0
                for qid in sorted_keys:
                    info = making_blocks_info[qid]
                    insert_idx = info['insert_at_idx_0based']
                    final_rows_after_insertion.extend(current_rows[last_copied_idx:insert_idx])
                    final_rows_after_insertion.extend(info['rows'])
                    last_copied_idx = insert_idx
                final_rows_after_insertion.extend(current_rows[last_copied_idx:])
                final_rows_to_write_local = final_rows_after_insertion
            else:
                final_rows_to_write_local = current_rows

            if not final_rows_to_write_local and self.excel_data_original_rows:
                self.log_message("Warning: Final row list was empty after processing, "
                                 "reverting to original data.", level="WARNING")
                final_rows_to_write_local = [row[:] for row in self.excel_data_original_rows]

            prep_time = time.time() - start_prep_time
            self.log_message(f"Data preparation complete ({prep_time:.2f}s). "
                             f"Final total rows: {len(final_rows_to_write_local)}.", level="INFO")
            if weights_applied_count > 0:
                self.log_message(f"Applied weights to {weights_applied_count} original survey items.", level="INFO")
            if data_prep_errors:
                self.log_message(f"Encountered {len(data_prep_errors)} issues during data prep/weighting.",
                                 level="WARNING")

            # ----- เขียนลง Template -----
            self.set_status("Creating Excel workbook...")
            self.log_message(f"Loading template: {self.template_file_path}", level="INFO")
            workbook = openpyxl.load_workbook(self.template_file_path)
            ws = workbook.active

            self.log_message(f"Writing {len(final_rows_to_write_local)} rows to worksheet...", level="INFO")
            write_start_time = time.time()

            ID_COL = COL_IDX['ID'] + 1
            LABEL_COL = COL_IDX['Label'] + 1
            WEIGHT_COL = COL_IDX['CategoryWeight'] + 1
            FORMAT_COL = COL_IDX['Format'] + 1
            making_fill = PatternFill(start_color='FFE6F2FF', end_color='FFE6F2FF', fill_type='solid')
            wrap_align = Alignment(wrap_text=True, vertical='top', horizontal='left')
            right_align = Alignment(horizontal='right', vertical='center')
            total_rows = len(final_rows_to_write_local)
            update_interval = 50

            for r_idx, row_data in enumerate(final_rows_to_write_local):
                if r_idx % update_interval == 0 or r_idx == total_rows - 1:
                    percent_done = min(100, int((r_idx + 1) / total_rows * 100))
                    self.set_status(f"Writing data: {percent_done}% ({r_idx + 1}/{total_rows} rows)")

                row_num_excel = r_idx + 1
                is_main_making = (len(row_data) > FORMAT_COL - 1
                                  and str(row_data[FORMAT_COL - 1]).strip() == "Making")

                for c_idx, cell_value in enumerate(row_data):
                    col_idx = c_idx + 1
                    try:
                        cell = ws.cell(row=row_num_excel, column=col_idx)
                        if isinstance(cell, openpyxl.cell.cell.MergedCell):
                            for merged_range in ws.merged_cells.ranges:
                                if cell.coordinate in merged_range:
                                    top_left_cell = ws.cell(row=merged_range.min_row,
                                                            column=merged_range.min_col)
                                    if cell.coordinate == top_left_cell.coordinate:
                                        top_left_cell.value = cell_value
                                    break
                            continue

                        cell.value = cell_value
                        if col_idx == LABEL_COL:
                            cell.alignment = wrap_align
                        if is_main_making and col_idx == LABEL_COL:
                            cell.fill = making_fill
                        if isinstance(cell_value, (int, float)) and col_idx in [ID_COL, WEIGHT_COL]:
                            cell.alignment = right_align
                            if isinstance(cell_value, int):
                                cell.number_format = '0'
                    except Exception as write_err:
                        self.log_message(f"Warning: Error writing/styling cell "
                                         f"({row_num_excel},{col_idx}): {type(write_err).__name__}",
                                         level="WARNING")

            label_col_letter = get_column_letter(LABEL_COL)
            if label_col_letter:
                ws.column_dimensions[label_col_letter].width = 45

            self.log_message(f"Finished writing data ({time.time() - write_start_time:.2f}s).", level="INFO")

            self.set_status("Saving workbook to file...")
            self.log_message(f"Saving workbook to: {new_file_path}", level="INFO")
            save_start_time = time.time()
            try:
                workbook.save(new_file_path)
            except Exception as save_err:
                try:
                    workbook.close()
                except Exception:
                    pass
                raise save_err

            self.log_message(f"Workbook saved successfully ({time.time() - save_start_time:.2f}s).",
                             level="SUCCESS")
            final_message = (f"File saved successfully ({time.time() - total_start_time:.2f}s): "
                             f"{new_filename}")
            level = "SUCCESS"
            QtWidgets.QMessageBox.information(self, "Save Successful", f"File saved as:\n{new_file_path}")

        except PermissionError:
            msg = f"Cannot save '{new_filename}'. Permission Denied. (Is the file open or write-protected?)"
            self.log_message(msg, level="ERROR")
            self.show_error_popup("Permission Error", msg)
            final_message = "Save failed: Permission Error."
            level = "ERROR"
        except Exception as e:
            msg = f"Error during Save As: [{type(e).__name__}] {e}"
            self.log_message(msg, level="ERROR")
            self.show_error_popup("Save Error", f"{msg}\nCheck log/console.")
            print(traceback.format_exc())
            final_message = "Save failed: Unexpected Error."
            level = "ERROR"
        finally:
            if workbook:
                try:
                    workbook.close()
                except Exception:
                    pass
            self._end_processing(final_status_message=final_message, level=level)

    # ------------------------------------------------------------------
    #  คลิกขวาเลือก Direction
    # ------------------------------------------------------------------
    def on_survey_tree_right_click(self, pos):
        index = self.survey_tree.indexAt(pos)
        if not index.isValid():
            return
        if index.column() != self.DIRECTION_COL_INDEX:
            return

        row = index.row()
        qid_item = self.survey_tree.item(row, self.SURVEY_COLUMNS.index("ข้อแบบสอบถาม"))
        if qid_item is None:
            return
        row_id = qid_item.text()

        self.survey_tree.selectRow(row)
        menu = QtWidgets.QMenu(self)
        act_low = menu.addAction(DIR_LOW_GOOD)
        act_high = menu.addAction(DIR_HIGH_GOOD)
        menu.addSeparator()
        act_clear = menu.addAction("Clear Direction")
        chosen = menu.exec(self.survey_tree.viewport().mapToGlobal(pos))
        if chosen is act_low:
            self._set_direction(row_id, DIR_LOW_GOOD)
        elif chosen is act_high:
            self._set_direction(row_id, DIR_HIGH_GOOD)
        elif chosen is act_clear:
            self._set_direction(row_id, "")

    def _set_direction(self, row_id, new_direction):
        if not self.tree_exists(self.survey_tree, row_id):
            return
        self.tree_set(self.survey_tree, row_id, "Direction", new_direction)
        if row_id in self.scales_data_store and not self.scales_data_store[row_id].get('is_making'):
            self.scales_data_store[row_id]['direction'] = new_direction
            self.log_message(f"Set Direction for {row_id} to "
                             f"'{new_direction if new_direction else 'None'}'", level="INFO")
            current_status = self.tree_get(self.survey_tree, row_id, "Condition Status")
            new_status = ""
            if not new_direction and current_status != "Making Generated":
                new_status = "ไม่ทำ Making"
            elif new_direction and current_status == "ไม่ทำ Making":
                new_status = ""
            elif new_direction and current_status != "Making Generated":
                new_status = current_status

            if new_status != current_status and current_status != "Making Generated":
                self.tree_set(self.survey_tree, row_id, "Condition Status", new_status)
        else:
            self.log_message(f"Could not set direction for {row_id} "
                             "(not found in store or is making item).", level="WARNING")


# =========================================================================
#  Entry point
# =========================================================================
_MAIN_WINDOW = None   # กัน QMainWindow ถูก GC เมื่อถูกเรียกจาก QApplication ที่มีอยู่แล้ว


def run_this_app(working_dir=None):
    global _MAIN_WINDOW
    print("--- T2B_INFO: Starting 'Program_T2B_Itemdef_Qt' via run_this_app() ---")

    if working_dir and os.path.isdir(working_dir):
        try:
            os.chdir(working_dir)
        except Exception as e:
            print(f"T2B_WARNING: Could not chdir to {working_dir}: {e}")

    app = QtWidgets.QApplication.instance()
    owns_app = app is None
    if owns_app:
        app = QtWidgets.QApplication(sys.argv)
    app.setStyleSheet(STYLESHEET)

    try:
        if getattr(sys, 'frozen', False):
            application_path = sys._MEIPASS
        else:
            application_path = os.path.dirname(os.path.abspath(__file__))
        print(f"Application path: {application_path}")

        template_filename = "template.xlsx"
        template_file_path_main = os.path.join(application_path, template_filename)
        print(f"Attempting to use template file at: {template_file_path_main}")

        if not os.path.exists(template_file_path_main):
            QtWidgets.QMessageBox.critical(
                None, "Template Missing",
                f"Template file '{template_filename}' not found in the application directory:\n"
                f"{application_path}\n\nThe application cannot continue.")
            print(f"Error: Template file '{template_filename}' not found.")
            return

        icon_path = resource_path("T2B.ico")
        if os.path.exists(icon_path):
            app.setWindowIcon(QtGui.QIcon(icon_path))

        window = ExcelScaleExtractorApp()
        window.template_file_path = template_file_path_main
        _MAIN_WINDOW = window

        screen = app.primaryScreen()
        if screen is not None:
            geo = screen.availableGeometry()
            frame = window.frameGeometry()
            frame.moveCenter(geo.center())
            window.move(frame.topLeft())

        window.show()
        if owns_app:
            app.exec()
        print("--- T2B_INFO: run_this_app() finished. ---")

    except Exception as e:
        print(f"T2B_ERROR: An error occurred during application execution: {e}")
        print(traceback.format_exc())
        try:
            QtWidgets.QMessageBox.critical(None, "Application Error (T2B Itemdef)",
                                           f"An unexpected error occurred:\n{e}")
        except Exception as popup_err:
            print(f"T2B_ERROR: Could not show error popup: {popup_err}")


if __name__ == "__main__":
    print("--- Running Program_T2B_Itemdef_Qt.py directly for testing ---")
    run_this_app()
    print("--- Finished direct execution of Program_T2B_Itemdef_Qt.py ---")
