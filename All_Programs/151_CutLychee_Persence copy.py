from __future__ import annotations

import sys
import traceback
import os
import time
import re
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor, as_completed
from copy import copy
from pathlib import Path
from typing import List, Tuple, Dict

import openpyxl
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.cell import range_boundaries

from PyQt6.QtCore import (
    QEasingCurve,
    QObject,
    QPropertyAnimation,
    QThread,
    Qt,
    pyqtSignal,
)
from PyQt6.QtGui import QColor, QPalette
from PyQt6.QtWidgets import (
    QApplication,
    QFrame,
    QGraphicsOpacityEffect,
    QHBoxLayout,
    QLabel,
    QListWidget,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QTextEdit,
    QVBoxLayout,
    QWidget,
    QFileDialog,
    QProgressBar,
)


# ----------------------------
# Excel processing core
# ----------------------------
def build_output_stem(source_stem: str) -> str:
    today = datetime.now().strftime("%Y%m%d")
    stem = source_stem

    # Remove trailing processed marker if present.
    stem = re.sub(r"_processed(?:_\d+)?$", "", stem, flags=re.IGNORECASE)

    # Normalize any N% token in filename to %.
    stem = re.sub(r"(?i)N%", "%", stem)

    # Normalize trailing date to today's date (8 digits at end).
    if re.search(r"\d{8}$", stem):
        stem = re.sub(r"\d{8}$", today, stem)
    else:
        stem = f"{stem} {today}"

    return stem


def unique_output_path_reserved(
    out_dir: Path, source_path: Path, reserved_names: set[str]
) -> Path:
    base = build_output_stem(source_path.stem)
    candidate_name = f"{base}.xlsx"
    idx = 1
    while True:
        candidate = out_dir / candidate_name
        key = str(candidate.resolve()).lower()
        if key not in reserved_names and not candidate.exists():
            reserved_names.add(key)
            return candidate
        candidate_name = f"{base}_{idx}.xlsx"
        idx += 1


def process_one_file_task(src_path: str, dst_path: str) -> str:
    process_workbook(Path(src_path), Path(dst_path))
    return dst_path


# ----------------------------------------------------------------
#  Row-height helper:  estimate the number of visual lines a cell
#  needs when wrap-text is on, then pick the tallest cell in each
#  row to set a uniform row height.
# ----------------------------------------------------------------
def _estimate_lines(text: str, col_width_chars: float) -> int:
    """Return estimated number of wrapped lines for *text* in a column
    whose width is *col_width_chars* characters."""
    if not text:
        return 1
    lines = text.split("\n")
    total = 0
    for ln in lines:
        length = len(ln)
        if length == 0:
            total += 1
        else:
            # rough: each visual line fits ~col_width_chars characters
            total += max(1, -(-length // max(int(col_width_chars), 1)))  # ceil div
    return total


def _auto_row_height(sheet, row: int, max_col: int, default_line_h: float = 15.0) -> float:
    """Calculate a suitable row height for *row* based on wrap-text content."""
    max_lines = 1
    for c in range(1, max_col + 1):
        cell = sheet.cell(row=row, column=c)
        val = cell.value
        if val is None:
            continue
        text = str(val)
        # get column width (openpyxl stores it per column-letter)
        col_letter = openpyxl.utils.get_column_letter(c)
        dim = sheet.column_dimensions.get(col_letter)
        col_w = dim.width if (dim and dim.width) else 10.0
        lines = _estimate_lines(text, col_w)
        if lines > max_lines:
            max_lines = lines
    return max_lines * default_line_h


# ----------------------------------------------------------------
#  Core workbook processor  (single-pass, optimised)
# ----------------------------------------------------------------
def process_workbook(input_path: Path, save_path: Path) -> None:
    wb = openpyxl.load_workbook(input_path)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet_name.strip().lower() == "contents":
            continue
        if sheet.max_row < 7:
            continue

        # --- detect effective max_col (fast sample) ----------------
        max_col = _detect_effective_max_col(sheet)

        # --- label row-6 ------------------------------------------
        from openpyxl.cell.cell import MergedCell
        if not isinstance(sheet["A6"], MergedCell):
            sheet["A6"] = "1st row: Column %"

        # --- unmerge (data region, row >= 7) -----------------------
        _unmerge_data_region(sheet)

        # --- single-pass: identify rows to delete ------------------
        rows_to_delete: List[int] = []
        current_max_row = sheet.max_row

        # Cache column-B values for fast lookup
        col_b: Dict[int, str] = {}
        for r in range(7, current_max_row + 1):
            v = sheet.cell(row=r, column=2).value
            col_b[r] = str(v).strip() if v is not None else ""

        # Walk bottom-up once  →  mark regular-pair rows AND total-% rows
        for i in range(current_max_row, 7, -1):
            cur_text = col_b.get(i, "")
            above_text = col_b.get(i - 1, "")

            # --- regular pair (same non-empty, non-TOTAL label) ---
            is_regular = (
                cur_text != ""
                and cur_text.upper() != "TOTAL"
                and cur_text == above_text
            )
            if is_regular:
                # Copy bottom border up
                for col in range(1, max_col + 1):
                    target = sheet.cell(row=i - 1, column=col)
                    source = sheet.cell(row=i, column=col)
                    target.border = Border(
                        left=target.border.left,
                        right=target.border.right,
                        top=target.border.top,
                        bottom=source.border.bottom,
                    )
                # Choose best value for data columns (col 3+)
                for col in range(3, max_col + 1):
                    target = sheet.cell(row=i - 1, column=col)
                    source = sheet.cell(row=i, column=col)
                    chosen_val, chosen_fmt = _choose_value_and_format(target, source)
                    target.value = chosen_val
                    target.number_format = chosen_fmt

                rows_to_delete.append(i)
                continue

            # --- TOTAL % row (above is TOTAL, current is TOTAL or blank) ---
            if above_text.upper() == "TOTAL" and (cur_text.upper() == "TOTAL" or cur_text == ""):
                for col in range(1, max_col + 1):
                    target = sheet.cell(row=i - 1, column=col)
                    source = sheet.cell(row=i, column=col)
                    target.border = Border(
                        left=target.border.left,
                        right=target.border.right,
                        top=target.border.top,
                        bottom=source.border.bottom,
                    )
                rows_to_delete.append(i)

        # --- bulk delete (descending, batched) ---------------------
        _delete_rows_desc(sheet, rows_to_delete)

        # --- bottom border on last data row -----------------------
        last_row = _find_last_data_row(sheet, max_col)
        thin = Side(border_style="thin", color="000000")
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=last_row, column=col)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=thin,
            )

        # --- normalise wrap-text & row heights --------------------
        _normalise_row_heights(sheet, max_col, last_row)

    wb.save(save_path)


# ----------------------------------------------------------------
#  Helper: detect effective max column (fast sampling)
# ----------------------------------------------------------------
def _detect_effective_max_col(sheet) -> int:
    dim = sheet.calculate_dimension()
    _, _, dim_max_col, _ = range_boundaries(dim)
    max_row = sheet.max_row

    sample_rows = set(range(1, min(max_row, 30) + 1))
    sample_rows.update({4, 5, 6, 7, 8, 9, 10, max_row})
    if max_row > 30:
        sample_rows.update(range(max(1, max_row - 9), max_row + 1))
    sample_rows = sorted(r for r in sample_rows if 1 <= r <= max_row)

    for col in range(dim_max_col, 2, -1):
        for r in sample_rows:
            value = sheet.cell(row=r, column=col).value
            if value is not None and str(value).strip() != "":
                return col
    return max(3, dim_max_col)


# ----------------------------------------------------------------
#  Helper: unmerge data region (rows >= 7) and fill values/styles
# ----------------------------------------------------------------
def _unmerge_data_region(sheet) -> None:
    merged_ranges = list(sheet.merged_cells.ranges)
    for m_range in merged_ranges:
        min_col, min_row, m_max_col, m_max_row = m_range.bounds
        if min_row < 7:
            continue

        top = sheet.cell(row=min_row, column=min_col)
        tl_val = top.value
        tl_border = copy(top.border)
        tl_fill = copy(top.fill)
        tl_font = copy(top.font)
        tl_alignment = copy(top.alignment)
        tl_number_format = top.number_format

        sheet.unmerge_cells(str(m_range))

        for r in range(min_row, m_max_row + 1):
            for c in range(min_col, m_max_col + 1):
                cell = sheet.cell(row=r, column=c)
                cell.value = tl_val
                cell.border = copy(tl_border)
                cell.fill = copy(tl_fill)
                cell.font = copy(tl_font)
                cell.alignment = copy(tl_alignment)
                cell.number_format = tl_number_format


# ----------------------------------------------------------------
#  Helper: choose value between upper/lower cell
# ----------------------------------------------------------------
def _choose_value_and_format(upper_cell, lower_cell):
    upper_val = upper_cell.value
    lower_val = lower_cell.value

    if lower_val is None or str(lower_val).strip() == "":
        return upper_val, upper_cell.number_format

    if isinstance(upper_val, (int, float)) and isinstance(lower_val, (int, float)):
        upper_has_frac = abs(float(upper_val) - int(float(upper_val))) > 1e-12
        lower_has_frac = abs(float(lower_val) - int(float(lower_val))) > 1e-12
        if upper_has_frac and not lower_has_frac:
            return upper_val, upper_cell.number_format
        return lower_val, lower_cell.number_format

    return lower_val, lower_cell.number_format


# ----------------------------------------------------------------
#  Helper: delete rows in descending order (batched consecutive)
# ----------------------------------------------------------------
def _delete_rows_desc(sheet, row_indexes: List[int]) -> None:
    if not row_indexes:
        return
    rows = sorted(set(row_indexes), reverse=True)
    start = rows[0]
    count = 1
    prev = rows[0]
    for r in rows[1:]:
        if r == prev - 1:
            count += 1
        else:
            sheet.delete_rows(start - count + 1, count)
            start = r
            count = 1
        prev = r
    sheet.delete_rows(start - count + 1, count)


# ----------------------------------------------------------------
#  Helper: find the last row that contains data
# ----------------------------------------------------------------
def _find_last_data_row(sheet, max_col: int) -> int:
    for r in range(sheet.max_row, 6, -1):
        for c in range(1, max_col + 1):
            v = sheet.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                return r
    return 7


# ----------------------------------------------------------------
#  Helper: normalise wrap-text alignment & equalise row heights
# ----------------------------------------------------------------
def _normalise_row_heights(sheet, max_col: int, last_row: int) -> None:
    """Ensure every data row (7 → last_row) has wrap_text enabled on
    column B and a consistent row height that fits the tallest cell."""
    for r in range(7, last_row + 1):
        # Enable wrap_text on column B (the label column) so text doesn't overflow
        b_cell = sheet.cell(row=r, column=2)
        if b_cell.alignment:
            b_cell.alignment = Alignment(
                horizontal=b_cell.alignment.horizontal,
                vertical=b_cell.alignment.vertical or "center",
                wrap_text=True,
            )
        else:
            b_cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Calculate & set row height based on content
        height = _auto_row_height(sheet, r, max_col)
        sheet.row_dimensions[r].height = height


# ----------------------------
# Input resolution
# ----------------------------
def resolve_input_items(items: List[str]) -> Tuple[List[Path], List[str]]:
    valid_paths: List[Path] = []
    invalid_items: List[str] = []

    for raw in items:
        p = raw.strip().strip('"').strip("'")
        if not p:
            continue
        path_obj = Path(p)

        if path_obj.is_file() and path_obj.suffix.lower() == ".xlsx":
            valid_paths.append(path_obj)
        elif path_obj.is_dir():
            files = sorted(path_obj.glob("*.xlsx"))
            if files:
                valid_paths.extend(files)
            else:
                invalid_items.append(f"{p} (no .xlsx inside folder)")
        else:
            invalid_items.append(p)

    seen = set()
    deduped: List[Path] = []
    for p in valid_paths:
        key = str(p.resolve()).lower()
        if key not in seen:
            seen.add(key)
            deduped.append(p)

    return deduped, invalid_items


# ----------------------------
# UI components
# ----------------------------
class AnimatedBackground(QWidget):
    pass


class FancyButton(QPushButton):
    def __init__(self, text: str) -> None:
        super().__init__(text)
        self.setCursor(Qt.CursorShape.PointingHandCursor)


class ProcessingWorker(QObject):
    progress = pyqtSignal(int, str)
    done = pyqtSignal(list)
    failed = pyqtSignal(str)

    def __init__(self, files: List[Path], output_dir: Path) -> None:
        super().__init__()
        self.files = files
        self.output_dir = output_dir

    def run(self) -> None:
        try:
            saved_files: List[str] = []
            total = len(self.files)
            if total == 0:
                self.done.emit([])
                return

            reserved_names: set[str] = set()
            plans: List[Tuple[Path, Path]] = []
            for src in self.files:
                dst = unique_output_path_reserved(self.output_dir, src, reserved_names)
                plans.append((src, dst))

            if total == 1:
                src, dst = plans[0]
                process_workbook(src, dst)
                saved_files.append(str(dst))
                self.progress.emit(100, f"1/1 processed: {src.name}")
            else:
                max_workers = max(2, min(total, os.cpu_count() or 2))
                with ProcessPoolExecutor(max_workers=max_workers) as executor:
                    future_map = {
                        executor.submit(process_one_file_task, str(src), str(dst)): (src, dst)
                        for src, dst in plans
                    }
                    completed = 0
                    for fut in as_completed(future_map):
                        src, dst = future_map[fut]
                        try:
                            fut.result()
                            saved_files.append(str(dst))
                        except Exception as e:
                            self.progress.emit(
                                int(((completed + 1) / total) * 100),
                                f"ERROR: {src.name} — {e}",
                            )
                        completed += 1
                        pct = int((completed / total) * 100)
                        self.progress.emit(pct, f"{completed}/{total} processed: {src.name}")

            # Keep output list in input order for easier review.
            ordered_set = set(saved_files)
            ordered = [str(dst) for _, dst in plans if str(dst) in ordered_set]
            self.done.emit(ordered)
        except Exception as exc:
            self.failed.emit(str(exc))


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.selected_files: List[Path] = []
        self.output_dir: Path | None = None
        self.worker_thread: QThread | None = None
        self.started_at: float | None = None

        self.setWindowTitle("Excel Smart Formatter Pro")
        self.resize(980, 700)

        self.bg = AnimatedBackground()
        self.bg.setObjectName("bg")
        self.setCentralWidget(self.bg)

        self._build_ui()
        self._animate_intro()

    def _build_ui(self) -> None:
        root_layout = QVBoxLayout(self.bg)
        root_layout.setContentsMargins(40, 30, 40, 30)

        self.card = QFrame()
        self.card.setObjectName("card")
        card_layout = QVBoxLayout(self.card)
        card_layout.setSpacing(14)
        card_layout.setContentsMargins(24, 24, 24, 24)

        title = QLabel("Excel %/N Cleaner")
        title.setObjectName("title")
        subtitle = QLabel("Professional PyQt6 UI with multi-file workflow, folder/path paste, and smooth motion")
        subtitle.setObjectName("subtitle")
        subtitle.setWordWrap(True)

        self.btn_pick_files = FancyButton("Select .xlsx Files")
        self.btn_pick_files.clicked.connect(self.pick_files)

        self.btn_pick_output = FancyButton("Select Output Folder")
        self.btn_pick_output.clicked.connect(self.pick_output_folder)

        self.paths_edit = QTextEdit()
        self.paths_edit.setPlaceholderText("Paste file/folder paths here. One per line (or separated by semicolon).")
        self.paths_edit.setFixedHeight(120)

        self.btn_apply_paths = FancyButton("Apply Pasted Paths")
        self.btn_apply_paths.clicked.connect(self.apply_pasted_paths)

        self.list_files = QListWidget()
        self.list_files.setMinimumHeight(160)

        self.lbl_output = QLabel("Output folder: not selected")
        self.lbl_output.setObjectName("meta")
        self.lbl_status = QLabel("Ready")
        self.lbl_status.setObjectName("status")

        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setValue(0)

        self.btn_start = FancyButton("Process All Files")
        self.btn_start.clicked.connect(self.start_processing)
        self.btn_start.setEnabled(False)

        row = QHBoxLayout()
        row.addWidget(self.btn_pick_files)
        row.addWidget(self.btn_pick_output)

        card_layout.addWidget(title)
        card_layout.addWidget(subtitle)
        card_layout.addLayout(row)
        card_layout.addWidget(self.paths_edit)
        card_layout.addWidget(self.btn_apply_paths)
        card_layout.addWidget(self.list_files)
        card_layout.addWidget(self.lbl_output)
        card_layout.addWidget(self.progress)
        card_layout.addWidget(self.lbl_status)
        card_layout.addWidget(self.btn_start)

        root_layout.addWidget(self.card)

        self.setStyleSheet(
            """
            QMainWindow, QWidget { color: #eaf0ff; font-family: 'Segoe UI', 'Tahoma'; font-size: 14px; }
            #bg {
                background: qradialgradient(cx:0.2, cy:0.1, radius:1.25,
                    fx:0.15, fy:0.05,
                    stop:0 #1f3866, stop:0.45 #142746, stop:1 #0b1321);
            }
            #card {
                background: rgba(11, 19, 31, 0.80);
                border: 1px solid rgba(255,255,255,0.16);
                border-radius: 20px;
            }
            #title { font-size: 34px; font-weight: 700; color: #ffffff; }
            #subtitle { font-size: 13px; color: #bfd3ff; }
            #meta { color: #98b7ff; font-size: 12px; }
            #status { color: #ffffff; font-size: 12px; }
            QTextEdit, QListWidget {
                background: rgba(255, 255, 255, 0.08);
                border: 1px solid rgba(255, 255, 255, 0.2);
                border-radius: 12px;
                padding: 8px;
                color: #f2f6ff;
            }
            QProgressBar {
                background: rgba(255, 255, 255, 0.10);
                border: 1px solid rgba(255,255,255,0.2);
                border-radius: 9px;
                text-align: center;
                height: 18px;
            }
            QProgressBar::chunk {
                border-radius: 8px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2dc3ff, stop:1 #4df0b7);
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #0ea5e9, stop:1 #10b981);
                border: none;
                border-radius: 12px;
                color: #062033;
                font-weight: 700;
                padding: 12px 16px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #38bdf8, stop:1 #34d399);
            }
            QPushButton:pressed {
                padding-top: 13px;
                padding-bottom: 11px;
            }
            QPushButton:disabled {
                background: rgba(255,255,255,0.25);
                color: rgba(255,255,255,0.6);
            }
            """
        )

    def _animate_intro(self) -> None:
        effect = QGraphicsOpacityEffect(self.card)
        self.card.setGraphicsEffect(effect)
        anim = QPropertyAnimation(effect, b"opacity", self)
        anim.setDuration(650)
        anim.setStartValue(0.0)
        anim.setEndValue(1.0)
        anim.setEasingCurve(QEasingCurve.Type.OutCubic)
        anim.start()
        self._intro_anim = anim

    def refresh_file_list(self) -> None:
        self.list_files.clear()
        for p in self.selected_files:
            self.list_files.addItem(str(p))
        self.btn_start.setEnabled(bool(self.selected_files and self.output_dir))

    def pick_files(self) -> None:
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Excel Files",
            "",
            "Excel Files (*.xlsx);;All Files (*.*)",
        )
        if not files:
            return

        parsed, invalid = resolve_input_items(files)
        self.selected_files = parsed
        self.refresh_file_list()

        if invalid:
            QMessageBox.warning(self, "Invalid paths", "\n".join(invalid[:5]))
        self.lbl_status.setText(f"Loaded {len(self.selected_files)} files")

    def pick_output_folder(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if not folder:
            return

        self.output_dir = Path(folder)
        self.lbl_output.setText(f"Output folder: {self.output_dir}")
        self.refresh_file_list()

    def apply_pasted_paths(self) -> None:
        raw = self.paths_edit.toPlainText().replace(";", "\n")
        items = [line for line in raw.splitlines() if line.strip()]
        files, invalid = resolve_input_items(items)

        self.selected_files = files
        self.refresh_file_list()

        if invalid:
            show = "\n".join(invalid[:7])
            more = "" if len(invalid) <= 7 else f"\n... and {len(invalid) - 7} more"
            QMessageBox.warning(self, "Some inputs were not valid", f"{show}{more}")

        self.lbl_status.setText(f"Loaded {len(self.selected_files)} files from pasted paths")

    def start_processing(self) -> None:
        if not self.selected_files:
            QMessageBox.warning(self, "No files", "Please select files first")
            return
        if not self.output_dir:
            QMessageBox.warning(self, "No output folder", "Please choose an output folder")
            return

        self.btn_start.setEnabled(False)
        self.progress.setValue(0)
        self.lbl_status.setText("Starting processing...")
        self.started_at = time.perf_counter()

        self.worker_thread = QThread()
        self.worker = ProcessingWorker(self.selected_files, self.output_dir)
        self.worker.moveToThread(self.worker_thread)

        self.worker_thread.started.connect(self.worker.run)
        self.worker.progress.connect(self.on_progress)
        self.worker.done.connect(self.on_done)
        self.worker.failed.connect(self.on_failed)
        self.worker.done.connect(self.worker_thread.quit)
        self.worker.failed.connect(self.worker_thread.quit)
        self.worker_thread.finished.connect(self.worker_thread.deleteLater)
        self.worker_thread.finished.connect(self.worker.deleteLater)
        self.worker_thread.finished.connect(lambda: self.btn_start.setEnabled(True))

        self.worker_thread.start()

    def on_progress(self, pct: int, text: str) -> None:
        self.progress.setValue(pct)
        self.lbl_status.setText(text)

    def on_done(self, saved_files: List[str]) -> None:
        self.progress.setValue(100)
        self.lbl_status.setText("Completed successfully")
        elapsed = 0.0
        if self.started_at is not None:
            elapsed = time.perf_counter() - self.started_at
        avg = (elapsed / len(saved_files)) if saved_files else 0.0

        preview = "\n".join(Path(p).name for p in saved_files[:8])
        more = "" if len(saved_files) <= 8 else f"\n...และอีก {len(saved_files) - 8} ไฟล์"
        summary = (
            f"ประมวลผลเสร็จ {len(saved_files)} ไฟล์\n"
            f"ใช้เวลา {elapsed:.1f} วินาที\n"
            f"เฉลี่ย {avg:.1f} วินาที/ไฟล์"
        )
        detail = f"ไฟล์ที่สร้าง:\n{preview}{more}"

        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setWindowTitle("Done")
        msg.setText(summary)
        msg.setInformativeText(detail)
        msg.setStyleSheet(
            "QMessageBox { background: #f7fafc; }"
            "QLabel { color: #0f172a; font-size: 12px; }"
            "QPushButton { min-width: 84px; color: #0f172a; background: #bae6fd; border: 1px solid #7dd3fc; border-radius: 8px; padding: 6px 12px; }"
            "QPushButton:hover { background: #7dd3fc; }"
        )
        msg.exec()

    def on_failed(self, message: str) -> None:
        self.lbl_status.setText("Error occurred")
        QMessageBox.critical(self, "Processing failed", message)


def main() -> None:
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    pal = QPalette()
    pal.setColor(QPalette.ColorRole.WindowText, QColor("#eaf0ff"))
    app.setPalette(pal)

    def handle_uncaught_exception(exc_type, exc_value, exc_tb):
        err_text = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
        log_path = Path(__file__).with_name("6_TEST_error.log")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with log_path.open("a", encoding="utf-8") as f:
            f.write(f"\n[{timestamp}]\n{err_text}\n")
        QMessageBox.critical(
            None,
            "Application Error",
            f"โปรแกรมเกิดข้อผิดพลาดและถูกหยุดการทำงาน\nบันทึก log ที่:\n{log_path}\n\n{exc_value}",
        )

    sys.excepthook = handle_uncaught_exception

    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
