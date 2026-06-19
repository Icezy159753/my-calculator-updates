# -*- coding: utf-8 -*-
import os
import re
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from PyQt6.QtCore import Qt, QLocale
from PyQt6.QtGui import QColor, QFont, QIcon, QPixmap
from PyQt6.QtWidgets import (
    QAbstractItemView, QApplication, QComboBox, QDialog, QDialogButtonBox,
    QFileDialog, QGroupBox, QHBoxLayout, QHeaderView, QInputDialog, QLabel,
    QListWidget, QListWidgetItem, QMenu, QMessageBox, QPushButton, QScrollArea,
    QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget,
)

# ชื่อคอลัมน์ open-end:  Q<ข้อ>_<concept>_O<mention>  เช่น Q1_2_O1  -> concept = 2
COL_PAT = re.compile(r"^(Q\d+)_(\d+)_O(\d+)$", re.IGNORECASE)

# ไอคอนลูกศร dropdown — ฝังไว้ในโปรแกรม (เขียนลงไฟล์ temp ตอนรัน ไม่ต้องพึ่งไฟล์ข้างนอก)
_CHEVRON_SVG = (
    '<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" '
    'viewBox="0 0 16 16" fill="none">'
    '<path d="M4 6 L8 10.5 L12 6" stroke="#64748b" stroke-width="1.9" '
    'stroke-linecap="round" stroke-linejoin="round"/></svg>'
)


def _arrow_url():
    import tempfile
    p = os.path.join(tempfile.gettempdir(), "_concept_recoder_chevron.svg")
    try:
        with open(p, "w", encoding="utf-8") as f:
            f.write(_CHEVRON_SVG)
    except Exception:
        pass
    return p.replace("\\", "/")


ARROW_URL = _arrow_url()

# โลโก้โปรแกรม (ฝังในตัว) — สื่อ "1 codeframe แตกเป็นหลาย Concept" โทน teal
_LOGO_SVG = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 256 256">
  <defs>
    <linearGradient id="g" x1="0" y1="0" x2="0" y2="1">
      <stop offset="0" stop-color="#2dd4bf"/>
      <stop offset="1" stop-color="#0f766e"/>
    </linearGradient>
  </defs>
  <rect x="18" y="18" width="220" height="220" rx="56" fill="url(#g)"/>
  <g fill="none" stroke="#ffffff" stroke-width="13" stroke-linecap="round">
    <path d="M88 128 H122"/>
    <path d="M122 128 C154 128 152 80 182 80"/>
    <path d="M122 128 H182"/>
    <path d="M122 128 C154 128 152 176 182 176"/>
  </g>
  <circle cx="88" cy="128" r="17" fill="#ffffff"/>
  <circle cx="184" cy="80" r="13.5" fill="#ffffff"/>
  <circle cx="184" cy="128" r="13.5" fill="#ffffff"/>
  <circle cx="184" cy="176" r="13.5" fill="#ffffff"/>
</svg>"""


def _logo_icon():
    """คืน QIcon ของโลโก้ (เขียน SVG ลง temp แล้วโหลด)"""
    import tempfile
    p = os.path.join(tempfile.gettempdir(), "_concept_recoder_logo.svg")
    try:
        with open(p, "w", encoding="utf-8") as f:
            f.write(_LOGO_SVG)
    except Exception:
        return QIcon()
    return QIcon(p)

HEADER = "BDD7EE"       # แถวหัวตาราง Code/English
# สีพื้นตามระดับชั้น (kind = ชื่อระดับตัวพิมพ์เล็ก)
KIND_FILL = {
    "grandnet":  "C6E0B4",   # เขียว
    "net":       "DDEBF7",   # ฟ้า
    "subnet":    "FFF2CC",   # ครีม
    "subsubnet": "FCE4D6",   # ส้มอ่อน
}
HEADER_DEFAULT_FILL = "EDEDED"   # ชั้นที่ลึกกว่านั้น


def level_name(n):
    """0=Grandnet, 1=Net, 2=Subnet, 3=Subsubnet, ..."""
    if n <= 0:
        return "Grandnet"
    if n == 1:
        return "Net"
    return "Sub" + "sub" * (n - 2) + "net"


def header_level(text):
    """อ่านระดับชั้นจากข้อความหัว เช่น 'Subnet : X' -> 2"""
    t = str(text).split(":")[0].strip().lower().replace(" ", "")
    if t.startswith("grandnet"):
        return 0
    n = 0
    while t.startswith("sub"):
        n += 1
        t = t[3:]
    return n + 1 if t.startswith("net") else 1   # net=1, subnet=2, ...


def fill_for_kind(kind):
    """คืนรหัสสีพื้นตาม kind (None = ไม่มีสี เช่น concept)"""
    if kind == "concept":
        return None
    return KIND_FILL.get(kind, HEADER_DEFAULT_FILL)


# ----------------------------------------------------------------------
# ส่วนตรรกะประมวลผล (แยกจาก GUI เพื่อให้ทดสอบ/รันซ้ำได้)
# ----------------------------------------------------------------------
def find_columns(df):
    """หาแถวหัวตาราง + ตำแหน่งคอลัมน์ 'Code' และ 'English' อัตโนมัติ
    คืนค่า (header_row, code_col, english_col)"""
    for r in range(min(15, len(df))):
        vals = [str(x).strip().lower() if pd.notna(x) else "" for x in df.iloc[r]]
        if "english" in vals:
            eng_col = vals.index("english")
            code_col = vals.index("code") if "code" in vals else 0
            return r, code_col, eng_col
    # ไม่พบหัว -> เดา: คอลัมน์ 0 = Code, คอลัมน์สุดท้าย = English
    return -1, 0, df.shape[1] - 1


def parse_codeframe(df):
    """อ่าน codeframe -> (rows, group_headers)
       rows : list ของ dict {orig, english, kind, level, groups}
       group_headers : ชื่อคอลัมน์กลุ่ม (เช่น Thaigroup1/Thaigroup2) ที่จะ carry ไปด้วย
       หาคอลัมน์ Code/English อัตโนมัติจากหัวตาราง"""
    header_row, code_col, eng_col = find_columns(df)
    # คอลัมน์กลุ่ม = คอลัมน์ที่มีชื่อหัว และไม่ใช่ Code/English
    # (c = index คอลัมน์ในไฟล์ ; None = คอลัมน์ที่เติมเพิ่มให้เปล่าๆ)
    group_cols, group_headers = [], []
    if header_row >= 0:
        for c in range(df.shape[1]):
            if c in (code_col, eng_col):
                continue
            h = df.iat[header_row, c]
            if pd.notna(h) and str(h).strip():
                group_cols.append(c)
                group_headers.append(str(h).strip())

    # การันตีต้องมีคอลัมน์ Thaigroup1 / Thaigroup2 เสมอ (เติมเปล่าถ้าไม่มี)
    norm = [h.strip().lower().replace(" ", "") for h in group_headers]
    for need, label in (("thaigroup1", "Thaigroup1"), ("thaigroup2", "Thaigroup2")):
        if need not in norm:
            group_cols.append(None)
            group_headers.append(label)

    rows = []
    for r in range(header_row + 1, len(df)):
        code = df.iat[r, code_col]
        eng = df.iat[r, eng_col]
        if pd.isna(eng):
            continue
        groups = [(None if c is None or pd.isna(df.iat[r, c]) else df.iat[r, c])
                  for c in group_cols]
        if pd.isna(code):                      # แถวหัว (Grandnet / Net / Subnet / ...)
            text = str(eng).strip()
            lv = header_level(text)
            rows.append({"orig": None, "english": text, "kind": level_name(lv).lower(),
                         "level": lv, "groups": groups})
        else:
            rows.append({"orig": int(float(code)), "english": str(eng).strip(),
                         "kind": "code", "level": None, "groups": groups})
    return rows, group_headers


def _fmt_cell(v):
    """แสดงค่าของ filter ให้สวย (1.0 -> '1')"""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def build_codeframe(data, concept_cols, cf_rows, group_headers=None, filter_col=None):
    """
    โครงสร้างแต่ละแถว = dict:
      {code, english, kind, tg(list กลุ่ม), orig(CodeOriginal), freq(ความถี่)}

    สูตร:  block = n_concept * n_cell
           NewCode = (CodeEดิม-1)*block + (ลำดับCell-1)*n_concept + ลำดับConcept
    คืนค่า: rows, data_new, info(dict)
    """
    concepts = list(concept_cols)          # ชื่อ Concept ตามลำดับที่ผู้ใช้กำหนด
    n = len(concepts)
    crank = {con: i + 1 for i, con in enumerate(concepts)}
    ngroups = len(group_headers) if group_headers else (
        len(cf_rows[0]["groups"]) if cf_rows and "groups" in cf_rows[0] else 0)
    empty_tg = [None] * ngroups

    # ค่าของ Cell (distinct, เรียง)
    if filter_col and filter_col in data.columns:
        cells = sorted((v for v in pd.unique(data[filter_col]) if pd.notna(v)),
                       key=lambda x: (str(type(x)), x))
    else:
        cells = [None]
    ncell = len(cells)
    cellrank = {v: i + 1 for i, v in enumerate(cells)}
    block = n * ncell

    def newcode(orig, con, cell):
        cr = cellrank[cell] if cell is not None else 1
        return (int(orig) - 1) * block + (cr - 1) * n + crank[con]

    def concept_label(con, cell):
        if cell is None:
            return str(con)
        return f"{con} - {filter_col}{_fmt_cell(cell)}"

    # ---- recode data ก่อน เพื่อใช้คำนวณความถี่ ----
    data_new = data.copy()
    if cells == [None]:
        for con in concepts:
            for col in concept_cols[con]:
                data_new[col] = data[col].map(
                    lambda v, c=con: newcode(v, c, None) if pd.notna(v) else pd.NA
                ).astype("Int64")
    else:
        cellseries = data[filter_col]
        for con in concepts:
            for col in concept_cols[con]:
                vals = []
                for v, cv in zip(data[col], cellseries):
                    if pd.isna(v) or pd.isna(cv) or cv not in cellrank:
                        vals.append(pd.NA)
                    else:
                        vals.append(newcode(v, con, cv))
                data_new[col] = pd.array(vals, dtype="Int64")

    # ---- ความถี่ของแต่ละ code ใหม่ (นับจากทุกคอลัมน์ concept ที่ recode แล้ว) ----
    freq = {}
    for con in concepts:
        for col in concept_cols[con]:
            for v in data_new[col].dropna():
                iv = int(v)
                freq[iv] = freq.get(iv, 0) + 1

    def mk(code, english, kind, tg, orig, fq):
        return {"code": code, "english": english, "kind": kind,
                "tg": tg, "orig": orig, "freq": fq}

    rows = []
    cur_level = 1                  # ระดับชั้นล่าสุดของหัวต้นทาง (ค่าเริ่ม = Net)
    for r in cf_rows:
        tg = r.get("groups") or empty_tg
        if r["orig"] is None:                               # หัวต้นทาง (Grandnet/Net/Subnet/..)
            cur_level = r.get("level", header_level(r["english"]))
            rows.append(mk(None, str(r["english"]).strip(), r.get("kind", "net"),
                           tg, None, None))
        else:                                               # code เดิม -> หัวชั้นถัดลงไป 1 ขั้น
            gl = cur_level + 1
            gkind = level_name(gl).lower()
            # แถว Subnet: ใส่ Thai Group + CodeOriginal (เลข code เดิม)
            rows.append(mk(None, f"{level_name(gl)} : {str(r['english']).strip()}",
                           gkind, tg, r["orig"], None))
            for cell in cells:                              # เรียง Cell ก่อน
                for con in concepts:                        # แล้ว Concept
                    nc = newcode(r["orig"], con, cell)
                    rows.append(mk(nc, concept_label(con, cell), "concept",
                                   empty_tg, None, freq.get(nc, 0)))

    info = {"n_concept": n, "n_cell": ncell, "cells": cells, "block": block,
            "group_headers": list(group_headers or [])}
    return rows, data_new, info


def write_codeframe_ws(ws, rows, group_headers=None):
    """เขียน Codeframe ลงในเวิร์กชีท ws ที่ให้มา
       Code | <Thai Group...> | English | CodeOriginal | (ว่าง) | Frequenzy"""
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ghs = list(group_headers or [])
    ng = len(ghs)
    col_code = 1
    col_eng = 2 + ng
    col_orig = col_eng + 1
    col_blank = col_orig + 1          # ช่องว่างคั่น (ตามภาพ)
    col_freq = col_blank + 1
    ncols = col_freq

    # หัวตาราง
    headers = {col_code: "Code", col_eng: "English",
               col_orig: "CodeOriginal", col_freq: "Frequenzy"}
    for j, gh in enumerate(ghs):
        headers[2 + j] = gh
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c, value=headers.get(c, None))
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.font = Font(bold=True)
        cell.border = border

    for i, r in enumerate(rows, start=2):
        a = ws.cell(row=i, column=col_code, value=r["code"])
        a.alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=col_eng, value=r["english"])
        for j, gv in enumerate(r.get("tg") or []):
            ws.cell(row=i, column=2 + j, value=gv)
        if r.get("orig") is not None:
            o = ws.cell(row=i, column=col_orig, value=r["orig"])
            o.alignment = Alignment(horizontal="center")
            o.font = Font(bold=True)
        if r.get("freq") is not None:
            f = ws.cell(row=i, column=col_freq, value=r["freq"])
            f.alignment = Alignment(horizontal="center")
            f.font = Font(color="C00000")          # ความถี่ตัวสีแดง
        # เส้นขอบ + สีพื้นแถวหัวทุกระดับชั้น
        color = fill_for_kind(r["kind"])
        for c in range(1, ncols + 1):
            cell = ws.cell(row=i, column=c)
            cell.border = border
            if color:
                cell.fill = PatternFill("solid", fgColor=color)
        if color:
            ws.cell(row=i, column=col_eng).font = Font(bold=True)

    ws.column_dimensions[get_column_letter(col_code)].width = 8
    for j in range(ng):
        ws.column_dimensions[get_column_letter(2 + j)].width = 14
    ws.column_dimensions[get_column_letter(col_eng)].width = 60
    ws.column_dimensions[get_column_letter(col_orig)].width = 12
    ws.column_dimensions[get_column_letter(col_blank)].width = 3
    ws.column_dimensions[get_column_letter(col_freq)].width = 12
    ws.freeze_panes = "A2"


def write_codeframe_book(sheets, path):
    """เขียน Codeframe หลายชีทลงไฟล์เดียว
       sheets = list ของ (sheet_name, rows, group_headers)"""
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    for name, rows, ghs in sheets:
        safe = (name or "Codeframe")[:31]
        base, k = safe, 2
        while safe in used:                       # กันชื่อชีทซ้ำ
            safe = f"{base[:28]}_{k}"; k += 1
        used.add(safe)
        write_codeframe_ws(wb.create_sheet(safe), rows, ghs)
    wb.save(path)


def style_data_sheet(ws, columns, r_names, concept_groups):
    """ใส่สีหัวคอลัมน์ในไฟล์ Data export + ตรึงที่ B3
       - R1/R2/R3 : ส้ม
       - Concept code : เขียวเข้ม/เขียวอ่อน สลับกันแต่ละ Concept"""
    orange = PatternFill("solid", fgColor="FFC000")
    dark = PatternFill("solid", fgColor="548235")
    light = PatternFill("solid", fgColor="A9D08E")
    idx = {name: i + 1 for i, name in enumerate(columns)}     # ชื่อ -> เลขคอลัมน์

    for name in r_names:
        c = ws.cell(row=1, column=idx[name])
        c.fill = orange
        c.font = Font(bold=True)

    for gi, grp in enumerate(concept_groups, start=1):
        is_dark = (gi % 2 == 1)
        fill = dark if is_dark else light
        font = Font(bold=True, color="FFFFFF") if is_dark else Font(bold=True)
        for name in grp:
            c = ws.cell(row=1, column=idx[name])
            c.fill = fill
            c.font = font

    ws.freeze_panes = "B3"


# ----------------------------------------------------------------------
# GUI
# ----------------------------------------------------------------------
def center_on_screen(widget):
    """ย้ายหน้าต่างให้อยู่กึ่งกลางจอ + ย่อให้พอดีจอถ้าใหญ่เกิน"""
    scr = QApplication.primaryScreen().availableGeometry()
    w = min(widget.width(), scr.width() - 60)
    h = min(widget.height(), scr.height() - 90)   # เผื่อ title bar + taskbar
    if w != widget.width() or h != widget.height():
        widget.resize(w, h)
    fg = widget.frameGeometry()
    fg.moveCenter(scr.center())
    widget.move(max(scr.left(), fg.left()), max(scr.top(), fg.top()))


class DropList(QListWidget):
    """QListWidget ที่รับ/ส่งไอเทมแบบลากวาง (ย้ายข้ามกล่องได้)"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setDragEnabled(True)
        self.setAcceptDrops(True)
        self.setDragDropMode(QAbstractItemView.DragDropMode.DragDrop)
        self.setDefaultDropAction(Qt.DropAction.MoveAction)
        self.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.setMinimumHeight(60)


class ConceptDialog(QDialog):
    """
    หน้าต่างกำหนด Concept แบบลากวาง
      - ซ้าย: รายชื่อตัวแปร (คอลัมน์) ต้นทาง
      - ขวา: กล่อง Concept (กด Add เพิ่มได้) ลากตัวแปรจากซ้ายมาวางในแต่ละกล่อง
    คืนผ่าน .result_map() -> {ชื่อ:[cols]}  และ .filter_column() -> ชื่อคอลัมน์ Filter | None
    """
    def __init__(self, columns, current_map=None, parent=None,
                 current_filter=None, cell_values_fn=None,
                 current_verbatim=None, current_range=None):
        super().__init__(parent)
        self.setWindowTitle("กำหนด Concept — ลากตัวแปรเข้ากล่อง")
        self.resize(1120, 820)
        self._centered = False
        self.concepts = []               # list ของ {"name", "cols"}
        self._cell_values_fn = cell_values_fn

        root = QVBoxLayout(self)
        root.setContentsMargins(18, 18, 18, 18)
        root.setSpacing(12)

        head = QLabel("กำหนด Concept")
        head.setObjectName("dlgtitle")
        root.addWidget(head)
        hint = QLabel("①  ทางขวา กด “＋ เพิ่มกล่อง Concept” ตามจำนวนที่ต้องการ\n"
                      "②  ลากชื่อตัวแปรจากกล่องซ้าย  →  วางในกล่อง Concept ทางขวา")
        hint.setWordWrap(True)
        hint.setObjectName("hintbox")
        root.addWidget(hint)

        # ----- แถบเลือกตัวแปร Filter (เช่น Cell) -----
        fbar = QWidget()
        fbar.setObjectName("filterbar")
        fl = QHBoxLayout(fbar)
        fl.setContentsMargins(12, 8, 12, 8)
        fl.addWidget(QLabel("ตัวแปร Filter (เช่น Cell):"))
        self.cmb_filter = QComboBox()
        self.cmb_filter.addItem("(ไม่มี)")
        self.cmb_filter.addItems([str(c) for c in columns])
        if current_filter:
            self.cmb_filter.setCurrentText(str(current_filter))
        self.cmb_filter.currentTextChanged.connect(self._on_filter)
        fl.addWidget(self.cmb_filter)
        self.lbl_cells = QLabel("")
        self.lbl_cells.setObjectName("cellcount")
        fl.addWidget(self.lbl_cells, 1)
        root.addWidget(fbar)

        body = QHBoxLayout()
        body.setSpacing(14)
        root.addLayout(body, 1)

        # ----- ซ้าย: ตัวแปรต้นทาง -----
        lwrap = QVBoxLayout()
        lwrap.setSpacing(6)
        self.lbl_src = QLabel("ตัวแปรทั้งหมด (ต้นทาง)")
        self.lbl_src.setObjectName("colhead")
        lwrap.addWidget(self.lbl_src)
        self.var_list = DropList()
        self.var_list.setObjectName("srclist")
        self.var_list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.var_list.customContextMenuRequested.connect(self._var_menu)
        lwrap.addWidget(self.var_list, 1)
        self.var_list.model().rowsInserted.connect(self._upd_src_count)
        self.var_list.model().rowsRemoved.connect(self._upd_src_count)
        body.addLayout(lwrap, 2)

        # ----- ขวา: Verbatim / Filter_Range / กล่อง Concept -----
        rwrap = QVBoxLayout()
        rwrap.setSpacing(6)

        # กล่อง Verbatim (Export อย่างเดียว ไม่ recode)
        rwrap.addWidget(QLabel("📝 Verbatim (Export เฉยๆ — ไม่ recode)"))
        self.verbatim_list = DropList()
        self.verbatim_list.setObjectName("vbox")
        self.verbatim_list.setMinimumHeight(70)
        self.verbatim_list.setMaximumHeight(160)
        rwrap.addWidget(self.verbatim_list)

        # กล่อง Filter_Range (Export อย่างเดียว)
        rwrap.addWidget(QLabel("🔎 Filter_Range (Export เฉยๆ)"))
        self.range_list = DropList()
        self.range_list.setObjectName("rbox")
        self.range_list.setMinimumHeight(55)
        self.range_list.setMaximumHeight(130)
        rwrap.addWidget(self.range_list)

        rhead = QHBoxLayout()
        lab = QLabel("🧩 กล่อง Concept (recode)")
        lab.setObjectName("colhead")
        rhead.addWidget(lab)
        rhead.addStretch(1)
        self.btn_auto = QPushButton("⚡ ตรวจจับอัตโนมัติ")
        self.btn_auto.setToolTip("สร้างกล่องจากชื่อคอลัมน์รูปแบบ Q#_<concept>_O#")
        self.btn_auto.clicked.connect(self.auto_detect)
        rhead.addWidget(self.btn_auto)
        rwrap.addLayout(rhead)

        self.btn_add = QPushButton("＋  เพิ่มกล่อง Concept")
        self.btn_add.setObjectName("addbtn")
        self.btn_add.clicked.connect(lambda: self.add_concept())
        rwrap.addWidget(self.btn_add)

        self.ctable = QTableWidget(0, 3)
        self.ctable.setObjectName("ctable")
        self.ctable.setHorizontalHeaderLabels(["Concept", "ตัวแปร (คอลัมน์)", ""])
        self.ctable.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents)
        self.ctable.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self.ctable.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.ResizeToContents)
        self.ctable.verticalHeader().setVisible(False)
        self.ctable.setMinimumHeight(150)
        self.ctable.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ctable.customContextMenuRequested.connect(self._table_menu)
        self.ctable.itemChanged.connect(self._on_name_edited)
        rwrap.addWidget(self.ctable, 1)
        hint2 = QLabel("เคล็ดลับ: เลือกตัวแปรทางซ้าย → คลิกขวา → สร้าง/ย้ายเข้า Concept / Verbatim / "
                       "Filter_Range  •  ลากเข้ากล่องได้  •  ดับเบิลคลิกชื่อเพื่อแก้ไข")
        hint2.setObjectName("tip")
        hint2.setWordWrap(True)
        rwrap.addWidget(hint2)
        body.addLayout(rwrap, 3)

        # ปุ่ม OK/Cancel
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok
                              | QDialogButtonBox.StandardButton.Cancel)
        bb.button(QDialogButtonBox.StandardButton.Ok).setText("ตกลง")
        bb.button(QDialogButtonBox.StandardButton.Cancel).setText("ยกเลิก")
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)
        root.addWidget(bb)

        self._all_columns = [str(c) for c in columns]
        self._populate(current_map)
        self._seed_box(self.verbatim_list, current_verbatim)
        self._seed_box(self.range_list, current_range)
        self._style()
        self._upd_src_count()
        self._on_filter(self.cmb_filter.currentText())

    def showEvent(self, e):
        super().showEvent(e)
        if not self._centered:           # จัดกึ่งกลาง + ย่อพอดีจอ หลังหน้าต่างแสดงจริง
            self._centered = True
            center_on_screen(self)

    def _seed_box(self, box, cols):
        """ย้ายคอลัมน์ที่กำหนดจากต้นทาง -> กล่อง (Verbatim/Range) ตอนเปิด"""
        for col in (cols or []):
            for r in range(self.var_list.count()):
                if self.var_list.item(r).text() == str(col):
                    self.var_list.takeItem(r)
                    break
            box.addItem(str(col))

    def verbatim_cols(self):
        return [self.verbatim_list.item(r).text()
                for r in range(self.verbatim_list.count())]

    def filter_range_cols(self):
        return [self.range_list.item(r).text()
                for r in range(self.range_list.count())]

    # ---------- Filter (Cell) ----------
    def _on_filter(self, name):
        if not name or name == "(ไม่มี)" or not self._cell_values_fn:
            self.lbl_cells.setText("")
            return
        try:
            cells = [str(_fmt_cell(v)) for v in self._cell_values_fn(name)]
        except Exception:
            cells = []
        if cells:
            self.lbl_cells.setText(
                f"พบ {len(cells)} Cell: {', '.join(cells)}  →  Concept จะถูกเบิ้ล ×{len(cells)}")
        else:
            self.lbl_cells.setText("")

    def filter_column(self):
        name = self.cmb_filter.currentText()
        return None if name == "(ไม่มี)" else name

    # ---------- โมเดล Concept = list ของ {"name","cols"} ----------
    def add_concept(self, name=None, items=None):
        if not name:
            name = f"Concept {len(self.concepts) + 1}"
        self.concepts.append({"name": name, "cols": list(items or [])})
        self._render_table()
        return len(self.concepts) - 1

    def remove_concept(self, idx):
        # คืนตัวแปรกลับต้นทาง (ซ้าย)
        for t in self.concepts[idx]["cols"]:
            self.var_list.addItem(t)
        del self.concepts[idx]
        self._render_table()

    def rename_concept(self, idx):
        name, ok = QInputDialog.getText(self, "เปลี่ยนชื่อ Concept",
                                        "ชื่อ Concept:", text=self.concepts[idx]["name"])
        if ok and name.strip():
            self.concepts[idx]["name"] = name.strip()
            self._render_table()

    def remove_var(self, idx, col):
        self.concepts[idx]["cols"].remove(col)
        self.var_list.addItem(col)
        self._render_table()

    def _move_items(self, items, idx):
        texts = [it.text() for it in items]
        for t in texts:
            for r in range(self.var_list.count()):
                if self.var_list.item(r).text() == t:
                    self.var_list.takeItem(r)
                    break
        # กันซ้ำ
        for t in texts:
            if t not in self.concepts[idx]["cols"]:
                self.concepts[idx]["cols"].append(t)
        self._render_table()

    def _upd_src_count(self, *args):
        self.lbl_src.setText(f"ตัวแปรทั้งหมด (ต้นทาง)  •  {self.var_list.count()} ตัว")

    # ---------- ตารางฝั่งขวา ----------
    def _render_table(self):
        self.ctable.blockSignals(True)
        self.ctable.setRowCount(len(self.concepts))
        for i, c in enumerate(self.concepts):
            name_it = QTableWidgetItem(c["name"])
            name_it.setFlags(name_it.flags() | Qt.ItemFlag.ItemIsEditable)
            f = name_it.font(); f.setBold(True); name_it.setFont(f)
            self.ctable.setItem(i, 0, name_it)

            vars_it = QTableWidgetItem(", ".join(c["cols"]) if c["cols"]
                                       else "(ลากหรือคลิกขวาเพื่อใส่ตัวแปร)")
            vars_it.setFlags(Qt.ItemFlag.ItemIsEnabled)
            self.ctable.setItem(i, 1, vars_it)

            btn = QPushButton("✕")
            btn.setObjectName("rmbtn")
            btn.setToolTip("ลบกล่องนี้")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.clicked.connect(lambda _=False, idx=i: self.remove_concept(idx))
            self.ctable.setCellWidget(i, 2, btn)
        self.ctable.blockSignals(False)

    def _on_name_edited(self, item):
        if item.column() == 0 and 0 <= item.row() < len(self.concepts):
            txt = item.text().strip()
            if txt:
                self.concepts[item.row()]["name"] = txt
            else:
                self._render_table()      # ชื่อว่าง -> คืนค่าเดิม

    def _table_menu(self, pos):
        row = self.ctable.rowAt(pos.y())
        if row < 0 or row >= len(self.concepts):
            return
        c = self.concepts[row]
        menu = QMenu(self)
        a_rename = menu.addAction("✏ เปลี่ยนชื่อ")
        a_del = menu.addAction("🗑 ลบกล่องนี้")
        rm_acts = []
        if c["cols"]:
            sub = menu.addMenu("⬅ เอาตัวแปรออก (คืนต้นทาง)")
            for col in c["cols"]:
                rm_acts.append((sub.addAction(col), col))
        chosen = menu.exec(self.ctable.viewport().mapToGlobal(pos))
        if chosen is None:
            return
        if chosen is a_rename:
            self.rename_concept(row)
        elif chosen is a_del:
            self.remove_concept(row)
        else:
            for act, col in rm_acts:
                if chosen is act:
                    self.remove_var(row, col)
                    break

    # ---------- คลิกขวาที่ตัวแปร (ฝั่งซ้าย) ----------
    def _var_menu(self, pos):
        items = self.var_list.selectedItems()
        if not items:
            return
        menu = QMenu(self)
        act_new = menu.addAction(f"➕ สร้างกล่อง Concept ใหม่จาก {len(items)} ตัวแปรที่เลือก…")
        move_acts = []
        if self.concepts:
            sub = menu.addMenu("➡ ย้ายเข้ากล่อง Concept ที่มีอยู่")
            for i, c in enumerate(self.concepts):
                move_acts.append((sub.addAction(c["name"]), i))
        menu.addSeparator()
        act_vb = menu.addAction("📝 เพิ่มเป็น Verbatim")
        act_rg = menu.addAction("🔎 เพิ่มเป็น Filter_Range")
        chosen = menu.exec(self.var_list.mapToGlobal(pos))
        if chosen is None:
            return
        if chosen is act_new:
            default = f"Concept {len(self.concepts) + 1}"
            name, ok = QInputDialog.getText(self, "สร้างกล่อง Concept",
                                            "ตั้งชื่อ Concept:", text=default)
            if not ok:
                return
            idx = self.add_concept(name=name.strip() or default)
            self._move_items(items, idx)
        elif chosen is act_vb:
            self._move_to_box(items, self.verbatim_list)
        elif chosen is act_rg:
            self._move_to_box(items, self.range_list)
        else:
            for act, idx in move_acts:
                if chosen is act:
                    self._move_items(items, idx)
                    break

    def _move_to_box(self, items, box):
        texts = [it.text() for it in items]
        for t in texts:
            for r in range(self.var_list.count()):
                if self.var_list.item(r).text() == t:
                    self.var_list.takeItem(r)
                    break
            box.addItem(t)

    def auto_detect(self):
        groups = {}
        for c in self._all_columns:
            m = COL_PAT.match(c)
            if m:
                groups.setdefault(int(m.group(2)), []).append(c)
        if not groups:
            QMessageBox.information(self, "ตรวจจับอัตโนมัติ",
                                    "ไม่พบคอลัมน์รูปแบบ Q#_<concept>_O#")
            return
        self._populate({f"Concept {con}": groups[con] for con in sorted(groups)})

    # ---------- โหลด/อ่านสถานะ ----------
    def _populate(self, mapping):
        self.concepts = []
        self.var_list.clear()
        used = set(self.verbatim_cols()) | set(self.filter_range_cols())
        if mapping:
            for name, cols in mapping.items():     # รักษาลำดับเดิม
                cols = [str(c) for c in cols]
                self.concepts.append({"name": str(name), "cols": cols})
                used.update(cols)
        self.var_list.addItems([c for c in self._all_columns if c not in used])
        self._render_table()

    def result_map(self):
        out = {}
        for c in self.concepts:
            if not c["cols"]:
                continue
            name = (c["name"] or "Concept").strip()
            base, k = name, 2                       # กันชื่อซ้ำ
            while name in out:
                name = f"{base} ({k})"; k += 1
            out[name] = c["cols"]
        return out

    def _style(self):
        self.setStyleSheet(("""
            QDialog { background:#eef1f7; font-family:'Segoe UI Variable','Segoe UI','Tahoma';
                      font-size:13px; }
            QLabel { color:#334155; }
            #dlgtitle { font-size:22px; font-weight:800; color:#0f766e; letter-spacing:0.3px; }
            #hintbox { color:#0f5e57; background:#e3f4f0; border:1px solid #bce3da;
                       border-radius:12px; padding:12px 14px; line-height:160%; }
            #colhead { font-weight:800; color:#243044; font-size:14px; }

            /* รายการตัวแปรต้นทาง */
            #srclist { border:1px solid #e0e6ef; border-radius:14px; background:#fafbfd;
                       padding:5px; outline:none; }
            #srclist::item { padding:8px 11px; border-radius:9px; margin:2px;
                             background:#ffffff; border:1px solid #eef1f6; }
            #srclist::item:hover { background:#e8f6f3; border-color:#a7e0d6; }
            #srclist::item:selected { background:#0f766e; color:white; border-color:#0f766e; }

            /* ตาราง Concept (ปลายทาง) */
            #ctable { border:1px solid #e0e6ef; border-radius:14px; background:white;
                      gridline-color:transparent; outline:none; }
            #ctable::item { padding:8px; }
            #ctable::item:selected { background:#e8f6f3; color:#0f172a; }
            QHeaderView::section { background:qlineargradient(x1:0,y1:0,x2:0,y2:1,
                                   stop:0 #14b8a6, stop:1 #0f766e);
                                   color:white; padding:9px 8px; border:none; font-weight:800; }
            QHeaderView::section:first { border-top-left-radius:13px; }
            QHeaderView::section:last  { border-top-right-radius:13px; }
            #tip { color:#8a94a3; font-size:12px; }

            #addbtn { color:white; border:none; border-radius:12px; padding:11px; font-weight:800;
                      background:qlineargradient(x1:0,y1:0,x2:0,y2:1,
                                 stop:0 #14b8a6, stop:1 #0d9488); }
            #addbtn:hover { background:qlineargradient(x1:0,y1:0,x2:0,y2:1,
                                 stop:0 #0d9488, stop:1 #0f766e); }
            QPushButton { background:white; border:1px solid #d4dbe6; border-radius:10px;
                          padding:8px 14px; color:#334155; font-weight:600; }
            QPushButton:hover { background:#f0fdfa; border-color:#5eead4; color:#0f766e; }
            QPushButton:pressed { background:#ccfbf1; }
            #rmbtn { color:#9aa3b2; border:none; padding:2px 8px; font-size:15px;
                     font-weight:800; border-radius:6px; }
            #rmbtn:hover { background:#fee2e2; color:#dc2626; }

            #filterbar { background:#e8f6f3; border:1px solid #c7ebe4; border-radius:12px; }
            #filterbar QLabel { color:#0f5e57; font-weight:700; }
            #cellcount { color:#7b8694; font-weight:400; }

            #vbox, #rbox { border:1px dashed #c8d2e0; border-radius:12px; background:#f8fafc;
                           padding:4px; outline:none; }
            #vbox::item, #rbox::item { padding:6px 9px; border-radius:8px; margin:2px;
                                       background:#eef2ff; }
            #vbox::item:selected, #rbox::item:selected { background:#0f766e; color:white; }

            QComboBox { padding:7px 11px; border:1px solid #d4dbe6; border-radius:10px;
                        background:white; }
            QComboBox:hover { border-color:#5eead4; }
            QComboBox::drop-down { subcontrol-origin:padding; subcontrol-position:center right;
                                   width:30px; border:none; }
            QComboBox::down-arrow { image: url("__ARROW__"); width:15px; height:15px; }
            QComboBox QAbstractItemView { border:1px solid #d4dbe6; border-radius:8px;
                                          background:white; selection-background-color:#14b8a6;
                                          selection-color:white; padding:4px; outline:none; }

            QDialogButtonBox QPushButton { min-width:96px; padding:9px 18px; border-radius:11px; }

            QScrollBar:vertical { background:transparent; width:11px; margin:2px; }
            QScrollBar::handle:vertical { background:#c4cdda; border-radius:5px; min-height:30px; }
            QScrollBar::handle:vertical:hover { background:#9aa6b8; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height:0; }
            QScrollBar:horizontal { background:transparent; height:11px; margin:2px; }
            QScrollBar::handle:horizontal { background:#c4cdda; border-radius:5px; min-width:30px; }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width:0; }
        """).replace("__ARROW__", ARROW_URL))


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Concept Recoder  •  แตก Codeframe ตาม Concept")
        self.setWindowIcon(_logo_icon())
        self.resize(1160, 720)
        self.setMinimumSize(900, 560)

        self.raw_path = None
        self.cf_path = None
        self.rows = None
        self.data_new = None
        self.all_columns = []        # คอลัมน์ทั้งหมดในชีท Rawdata
        self.concept_map = {}        # {ชื่อ:[cols], ...}
        self.filter_col = None       # คอลัมน์ Filter (เช่น Cell)
        self.verbatim = []           # คอลัมน์ Verbatim (export อย่างเดียว)
        self.filter_range = []       # คอลัมน์ Filter_Range (export อย่างเดียว)
        self.group_headers = []      # ชื่อคอลัมน์กลุ่มจาก codeframe (Thai Group1/2)
        self.sets = []               # งานหลาย Set ที่บันทึกไว้
        self.preview_cfs = []        # [(name, rows, group_headers)] สำหรับสลับดูพรีวิว

        self._build_ui()
        self._apply_style()

    # ---------- UI ----------
    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(10)

        head = QHBoxLayout()
        head.setSpacing(12)
        logo = QLabel()
        logo.setPixmap(_logo_icon().pixmap(44, 44))
        logo.setFixedSize(44, 44)
        title = QLabel("Concept Recoder")
        title.setObjectName("title")
        head.addWidget(logo)
        head.addWidget(title)
        head.addStretch(1)
        root.addLayout(head)
        sub = QLabel("แตก Codeframe ชุดเดียวออกเป็นหลาย Concept • recode Rawdata อัตโนมัติ")
        sub.setObjectName("subtitle")
        root.addWidget(sub)

        # ===== แบ่งซ้าย (ตัวเลือก) / ขวา (พรีวิวเต็มจอ) =====
        main = QHBoxLayout()
        main.setSpacing(14)
        root.addLayout(main, 1)

        # ---------- คอลัมน์ซ้าย (เลื่อนได้ ถ้าจอเตี้ย) ----------
        left = QVBoxLayout()
        left.setSpacing(12)
        left.setContentsMargins(0, 0, 8, 0)
        left_panel = QWidget()
        left_panel.setLayout(left)
        left_panel.setObjectName("leftpanel")
        left_panel.setFixedWidth(424)          # ความกว้างตายตัว -> กล่องในไม่ขยายออกข้าง
        self.left_panel = left_panel
        left_scroll = QScrollArea()
        left_scroll.setWidget(left_panel)
        left_scroll.setWidgetResizable(True)   # คุมความสูง (width ตรึงด้วย fixedWidth ข้างบน)
        left_scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        left_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        left_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        left_scroll.setFixedWidth(446)
        left_scroll.setStyleSheet("QScrollArea, #leftpanel { background:transparent; }")
        main.addWidget(left_scroll)

        # กล่อง Rawdata
        g_raw = QGroupBox("1 • Rawdata")
        v = QVBoxLayout(g_raw)
        h = QHBoxLayout()
        self.btn_raw = QPushButton("📂  เลือกไฟล์ Rawdata")
        self.btn_raw.clicked.connect(self.load_raw)
        self.lbl_raw = QLabel("ยังไม่ได้เลือกไฟล์")
        self.lbl_raw.setObjectName("pathlbl")
        h.addWidget(self.btn_raw)
        h.addWidget(self.lbl_raw, 1)
        v.addLayout(h)

        hs = QHBoxLayout()
        hs.addWidget(QLabel("ชีท:"))
        self.cmb_raw_sheet = QComboBox()
        self._lock_combo(self.cmb_raw_sheet)
        self.cmb_raw_sheet.currentTextChanged.connect(self.on_raw_sheet)
        hs.addWidget(self.cmb_raw_sheet, 1)
        v.addLayout(hs)

        hc = QHBoxLayout()
        hc.addWidget(QLabel("กำหนด Concept:"))
        self.btn_concept = QPushButton("🧩  กำหนด Concept + Filter")
        self.btn_concept.clicked.connect(self.open_concept_dialog)
        hc.addWidget(self.btn_concept, 1)
        v.addLayout(hc)
        self.lst_concepts = QListWidget()   # สรุปผลแบบอ่านอย่างเดียว
        self.lst_concepts.setSelectionMode(QListWidget.SelectionMode.NoSelection)
        self.lst_concepts.setMinimumHeight(80)
        self.lst_concepts.setMaximumHeight(150)
        self._lock_list(self.lst_concepts)
        v.addWidget(self.lst_concepts)
        left.addWidget(g_raw)

        # กล่อง Codeframe
        g_cf = QGroupBox("2 • Codeframe")
        v2 = QVBoxLayout(g_cf)
        h2 = QHBoxLayout()
        self.btn_cf = QPushButton("📂  เลือกไฟล์ Codeframe")
        self.btn_cf.clicked.connect(self.load_cf)
        self.lbl_cf = QLabel("ยังไม่ได้เลือกไฟล์")
        self.lbl_cf.setObjectName("pathlbl")
        h2.addWidget(self.btn_cf)
        h2.addWidget(self.lbl_cf, 1)
        v2.addLayout(h2)

        hs2 = QHBoxLayout()
        hs2.addWidget(QLabel("ชีท:"))
        self.cmb_cf_sheet = QComboBox()
        self._lock_combo(self.cmb_cf_sheet)
        hs2.addWidget(self.cmb_cf_sheet, 1)
        v2.addLayout(hs2)

        self.lbl_block = QLabel("Code ใหม่ = (Codeเดิม−1) × จำนวนConcept + ลำดับConcept")
        self.lbl_block.setObjectName("pathlbl")
        self.lbl_block.setWordWrap(True)
        v2.addWidget(self.lbl_block)
        left.addWidget(g_cf)

        # กล่อง Sets (งานหลายชุด)
        g_set = QGroupBox("3 • Sets (งานหลายชุด)")
        v3 = QVBoxLayout(g_set)
        self.btn_save_set = QPushButton("💾  บันทึก Set นี้ + เริ่ม Set ใหม่")
        self.btn_save_set.clicked.connect(self.save_set)
        v3.addWidget(self.btn_save_set)
        self.lst_sets = QListWidget()
        self.lst_sets.setMaximumHeight(110)
        self.lst_sets.setToolTip("ดับเบิลคลิกเพื่อลบ Set")
        self.lst_sets.itemDoubleClicked.connect(self.remove_set)
        self._lock_list(self.lst_sets)
        v3.addWidget(self.lst_sets)
        self.lbl_sets = QLabel("ยังไม่มี Set ที่บันทึก — กด RUN จะรันชุดปัจจุบัน")
        self.lbl_sets.setObjectName("pathlbl")
        self.lbl_sets.setWordWrap(True)
        v3.addWidget(self.lbl_sets)
        left.addWidget(g_set)
        left.addStretch(1)

        # ---------- คอลัมน์ขวา (RUN ด้านบน + พรีวิว) ----------
        right = QVBoxLayout()
        right.setSpacing(8)

        # ปุ่ม RUN (เล็ก ชิดขวา) + status อยู่บนสุดฝั่งขวา
        runrow = QHBoxLayout()
        self.status = QLabel("พร้อมใช้งาน")
        self.status.setObjectName("status")
        self.status.setWordWrap(True)
        runrow.addWidget(self.status, 1)
        self.btn_run = QPushButton("▶  RUN ทุก Set")
        self.btn_run.setObjectName("run")
        self.btn_run.setFixedWidth(200)
        self.btn_run.clicked.connect(self.run)
        runrow.addWidget(self.btn_run)
        right.addLayout(runrow)

        ph = QHBoxLayout()
        prev_lbl = QLabel("Preview Codeframe")
        prev_lbl.setObjectName("section")
        ph.addWidget(prev_lbl)
        ph.addStretch(1)
        self.cmb_preview = QComboBox()
        self.cmb_preview.setMinimumWidth(220)
        self.cmb_preview.currentIndexChanged.connect(self.on_preview_changed)
        self.cmb_preview.hide()
        ph.addWidget(self.cmb_preview)
        right.addLayout(ph)
        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["Code", "English"])
        self.table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setShowGrid(False)
        self.table.horizontalHeader().setHighlightSections(False)
        # เลือก/ไฮไลต์ทั้งบรรทัด
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        right.addWidget(self.table, 1)
        main.addLayout(right, 1)

        # กันไม่ให้ป้ายชื่อไฟล์ดันกล่องออกข้าง
        self.lbl_raw.setWordWrap(True)
        self.lbl_cf.setWordWrap(True)
        self._lock_combo(self.cmb_preview)

    def _lock_combo(self, cmb):
        """ไม่ให้ combobox ขยายกว้างตามความยาวข้อความ"""
        cmb.setSizeAdjustPolicy(
            QComboBox.SizeAdjustPolicy.AdjustToMinimumContentsLengthWithIcon)
        cmb.setMinimumContentsLength(1)

    def _lock_list(self, lst):
        """ไม่ให้ list ขยายออกข้าง — ตัดบรรทัดแทน + ปิด scroll แนวนอน"""
        lst.setWordWrap(True)
        lst.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

    def _apply_style(self):
        self.setStyleSheet(("""
            QWidget { font-family: 'Segoe UI Variable','Segoe UI','Tahoma';
                      font-size: 13px; color:#243044; background:#eef1f7; }
            QLabel { background:transparent; }   /* กันไม่ให้ label มีพื้นเทา */

            #title { font-size: 26px; font-weight: 800; color:#0f766e;
                     letter-spacing:0.3px; }
            #subtitle { color:#7b8694; font-size:13px; }
            #section { font-weight:800; color:#334155; font-size:14px; }

            /* การ์ดแต่ละกล่อง */
            QGroupBox { font-weight:800; color:#0f766e; border:1px solid #e4e9f2;
                        border-radius:16px; margin-top:16px; background:#ffffff;
                        padding:14px 12px 12px 12px; }
            QGroupBox::title { subcontrol-origin: margin; left:16px; top:2px;
                               padding:2px 10px; background:#e8f6f3; border-radius:9px;
                               color:#0f766e; }

            /* ปุ่มทั่วไป */
            QPushButton { background:#ffffff; border:1px solid #d4dbe6; border-radius:11px;
                          padding:9px 16px; color:#334155; font-weight:600; }
            QPushButton:hover { background:#f0fdfa; border-color:#5eead4; color:#0f766e; }
            QPushButton:pressed { background:#ccfbf1; }

            /* ปุ่ม RUN ไล่เฉดสี */
            #run { color:white; border:none; font-size:14px; font-weight:800;
                   padding:10px 14px; border-radius:11px;
                   background:qlineargradient(x1:0,y1:0,x2:0,y2:1,
                              stop:0 #14b8a6, stop:1 #0d9488); }
            #run:hover { background:qlineargradient(x1:0,y1:0,x2:0,y2:1,
                              stop:0 #0d9488, stop:1 #0f766e); }
            #run:pressed { background:#0f766e; }

            #pathlbl { color:#8a94a3; }
            #status { color:#0f5e57; padding:10px 12px; background:#e8f6f3;
                      border:1px solid #c7ebe4; border-radius:11px; }

            /* อินพุต */
            QComboBox { padding:8px 12px; border:1px solid #d4dbe6; border-radius:11px;
                        background:white; min-height:18px; }
            QComboBox:hover { border-color:#5eead4; }
            QComboBox:focus { border-color:#14b8a6; }
            QComboBox::drop-down { subcontrol-origin:padding; subcontrol-position:center right;
                                   width:30px; border:none; }
            QComboBox::down-arrow { image: url("__ARROW__"); width:15px; height:15px; }
            QComboBox QAbstractItemView { border:1px solid #d4dbe6; border-radius:8px;
                                          background:white; selection-background-color:#14b8a6;
                                          selection-color:white; padding:4px; outline:none; }

            /* รายการ/ตาราง */
            QListWidget { border:1px solid #e4e9f2; border-radius:12px; background:#fafbfd;
                          padding:4px; outline:none; }
            QListWidget::item { padding:6px 8px; border-radius:7px; }
            QTableWidget { border:1px solid #e4e9f2; border-radius:12px; background:white;
                           alternate-background-color:#f6f8fc; gridline-color:transparent;
                           outline:none; selection-background-color:#bfe9e1;
                           selection-color:#0f172a; }
            QTableWidget::item { padding:7px 8px; }
            QTableWidget::item:selected { background:#bfe9e1; color:#0f172a; }
            QHeaderView::section { background:qlineargradient(x1:0,y1:0,x2:0,y2:1,
                                   stop:0 #14b8a6, stop:1 #0f766e);
                                   color:white; padding:10px 8px; border:none;
                                   font-weight:800; }
            QHeaderView::section:first { border-top-left-radius:11px; }
            QHeaderView::section:last { border-top-right-radius:11px; }

            QScrollBar:vertical { background:transparent; width:11px; margin:2px; }
            QScrollBar::handle:vertical { background:#c4cdda; border-radius:5px; min-height:30px; }
            QScrollBar::handle:vertical:hover { background:#9aa6b8; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height:0; }
            QScrollBar:horizontal { background:transparent; height:11px; margin:2px; }
            QScrollBar::handle:horizontal { background:#c4cdda; border-radius:5px; min-width:30px; }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width:0; }
        """).replace("__ARROW__", ARROW_URL))

    # ---------- โหลดไฟล์ ----------
    def load_raw(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "เลือกไฟล์ Rawdata", "", "Excel (*.xlsx *.xls)")
        if not path:
            return
        self.raw_path = path
        self.lbl_raw.setText(os.path.basename(path))
        try:
            sheets = pd.ExcelFile(path).sheet_names
        except Exception as e:
            QMessageBox.critical(self, "ผิดพลาด", f"เปิดไฟล์ไม่ได้:\n{e}")
            return
        self.sets = []                     # โหลดไฟล์ใหม่ -> ล้าง Set เดิม
        self.refresh_sets_ui()
        self.cmb_raw_sheet.blockSignals(True)
        self.cmb_raw_sheet.clear()
        self.cmb_raw_sheet.addItems(sheets)
        self.cmb_raw_sheet.blockSignals(False)
        self.on_raw_sheet(self.cmb_raw_sheet.currentText())

    def on_raw_sheet(self, sheet):
        if not self.raw_path or not sheet:
            return
        try:
            df = pd.read_excel(self.raw_path, sheet_name=sheet, nrows=0)
        except Exception as e:
            QMessageBox.critical(self, "ผิดพลาด", f"อ่านชีทไม่ได้:\n{e}")
            return
        self.all_columns = [str(c) for c in df.columns]
        # ไม่ auto-detect — ให้ผู้ใช้กำหนดเองในหน้าต่าง Concept
        self.concept_map = {}
        self.filter_col = None
        self.verbatim = []
        self.filter_range = []
        self.update_concept_summary()

    def _cell_values(self, name):
        """คืนค่า distinct ของคอลัมน์ (ใช้แสดงจำนวน Cell ในหน้าต่าง Concept)"""
        df = pd.read_excel(self.raw_path,
                           sheet_name=self.cmb_raw_sheet.currentText(),
                           usecols=[name])
        return sorted((v for v in pd.unique(df[name]) if pd.notna(v)),
                      key=lambda x: (str(type(x)), x))

    def update_concept_summary(self):
        """อัปเดตรายการสรุป Concept ในหน้าหลัก"""
        self.lst_concepts.clear()
        if not self.concept_map:
            it = QListWidgetItem("ยังไม่ได้กำหนด — กดปุ่มกำหนด Concept + Filter")
            it.setFlags(Qt.ItemFlag.NoItemFlags)
            self.lst_concepts.addItem(it)
            return
        def info_row(text):
            it = QListWidgetItem(text)
            it.setFlags(Qt.ItemFlag.NoItemFlags)
            self.lst_concepts.addItem(it)
        if self.filter_col:
            info_row(f"⚙ Filter (Cell): {self.filter_col}")
        if self.filter_range:
            info_row(f"🔎 Filter_Range: {', '.join(self.filter_range)}")
        if self.verbatim:
            info_row(f"📝 Verbatim: {', '.join(self.verbatim)}")
        for name, cols in self.concept_map.items():     # ตามลำดับที่กำหนด
            self.lst_concepts.addItem(QListWidgetItem(f"🧩 {name}:  {', '.join(cols)}"))

    def open_concept_dialog(self):
        if not self.all_columns:
            QMessageBox.warning(self, "ยังไม่มีข้อมูล",
                                "กรุณาเลือกไฟล์ Rawdata และชีทก่อน")
            return
        dlg = ConceptDialog(self.all_columns, self.concept_map, self,
                            current_filter=self.filter_col,
                            cell_values_fn=self._cell_values,
                            current_verbatim=self.verbatim,
                            current_range=self.filter_range)
        if dlg.exec():
            self.concept_map = dlg.result_map()
            self.filter_col = dlg.filter_column()
            self.verbatim = dlg.verbatim_cols()
            self.filter_range = dlg.filter_range_cols()
            self.update_concept_summary()

    def load_cf(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "เลือกไฟล์ Codeframe", "", "Excel (*.xlsx *.xls)")
        if not path:
            return
        self.cf_path = path
        self.lbl_cf.setText(os.path.basename(path))
        try:
            sheets = pd.ExcelFile(path).sheet_names
        except Exception as e:
            QMessageBox.critical(self, "ผิดพลาด", f"เปิดไฟล์ไม่ได้:\n{e}")
            return
        self.cmb_cf_sheet.clear()
        self.cmb_cf_sheet.addItems(sheets)

    # ---------- Sets ----------
    def selected_concept_cols(self):
        return {con: cols for con, cols in self.concept_map.items() if cols}

    def current_config(self, name):
        """snapshot การตั้งค่าปัจจุบันเป็น 1 Set"""
        return {
            "name": name,
            "raw_sheet": self.cmb_raw_sheet.currentText(),
            "cf_path": self.cf_path,
            "cf_sheet": self.cmb_cf_sheet.currentText(),
            "concept_map": {k: list(v) for k, v in self.concept_map.items() if v},
            "filter_col": self.filter_col,
            "verbatim": list(self.verbatim),
            "filter_range": list(self.filter_range),
        }

    def save_set(self):
        if not self.selected_concept_cols():
            QMessageBox.warning(self, "ยังไม่ครบ", "กรุณากำหนด Concept ก่อนบันทึก Set")
            return
        if not self.cf_path or not self.cmb_cf_sheet.currentText():
            QMessageBox.warning(self, "ยังไม่ครบ", "กรุณาเลือกไฟล์ Codeframe + ชีท ก่อนบันทึก Set")
            return
        self.sets.append(self.current_config(f"Set {len(self.sets) + 1}"))
        # เคลียร์เพื่อเริ่ม Set ใหม่ (คง Rawdata/Codeframe ไว้ให้สะดวก)
        self.concept_map = {}
        self.filter_col = None
        self.verbatim = []
        self.filter_range = []
        self.update_concept_summary()
        self.refresh_sets_ui()

    def remove_set(self, item):
        row = self.lst_sets.row(item)
        if 0 <= row < len(self.sets):
            del self.sets[row]
            for i, s in enumerate(self.sets, start=1):     # ตั้งชื่อใหม่
                s["name"] = f"Set {i}"
            self.refresh_sets_ui()

    def refresh_sets_ui(self):
        self.lst_sets.clear()
        for s in self.sets:
            cfname = os.path.basename(s["cf_path"]) if s["cf_path"] else "-"
            self.lst_sets.addItem(
                f"{s['name']}: {len(s['concept_map'])} Concept • CF={cfname}#{s['cf_sheet']}")
        n = len(self.sets)
        self.lbl_sets.setText(
            f"มี {n} Set ที่บันทึก — กด RUN จะรวมชุดปัจจุบันด้วย (ดับเบิลคลิกเพื่อลบ)"
            if n else "ยังไม่มี Set ที่บันทึก — กด RUN จะรันชุดปัจจุบัน")

    # ---------- ประมวลผล 1 Set ----------
    def _build_export(self, dn, cfg):
        """สร้าง DataFrame สำหรับ export ของ 1 Set
           คืน (export_df, order, r_names, concept_groups)"""
        verbatim = [c for c in cfg["verbatim"] if c in dn.columns]
        franges = [c for c in cfg["filter_range"] if c in dn.columns]

        def to_int(v):
            try:
                return int(float(v))
            except (TypeError, ValueError):
                return None

        def join_verbatim(i):
            parts = []
            for c in verbatim:
                v = dn[c].iat[i]
                if pd.notna(v) and str(v).strip():
                    parts.append(str(v).strip())
            return " / ".join(parts)

        range_col = franges[0] if franges else cfg["filter_col"]
        r1, r2, r3 = [], [], []
        if verbatim and range_col in dn.columns:
            for i in range(len(dn)):
                jv = join_verbatim(i)
                rv = to_int(dn[range_col].iat[i])
                r1.append(jv if rv in (1, 2) else "")
                r2.append(jv if rv == 3 else "")
                r3.append(jv if rv in (4, 5) else "")

        order, out, seen = [], {}, set()

        def put(name, series):
            if name in seen:
                return
            seen.add(name); order.append(name); out[name] = list(series)

        put(dn.columns[0], dn[dn.columns[0]])
        if cfg["filter_col"] and cfg["filter_col"] in dn.columns:
            put(cfg["filter_col"], dn[cfg["filter_col"]])
        for c in franges:
            put(c, dn[c])
        r_names = []
        if r1:
            put("R1(1-2)", r1); put("R2(3)", r2); put("R3(4-5)", r3)
            r_names = ["R1(1-2)", "R2(3)", "R3(4-5)"]
        for c in verbatim:
            put(c, dn[c])
        concept_groups = []
        for ci, cols in enumerate(cfg["concept_map"].values(), start=1):
            grp = []
            for cj, c in enumerate(cols, start=1):
                if c in dn.columns:
                    nm = f"C{ci}_Code{cj}"
                    put(nm, dn[c]); grp.append(nm)
            if grp:
                concept_groups.append(grp)
        return pd.DataFrame(out)[order], order, r_names, concept_groups

    def _process_job(self, cfg):
        """อ่านไฟล์ + build codeframe + export ของ 1 Set"""
        data = pd.read_excel(self.raw_path, sheet_name=cfg["raw_sheet"])
        cf = pd.read_excel(cfg["cf_path"], sheet_name=cfg["cf_sheet"], header=None)
        cf_rows, ghs = parse_codeframe(cf)
        rows, data_new, info = build_codeframe(
            data, cfg["concept_map"], cf_rows, group_headers=ghs,
            filter_col=cfg["filter_col"])
        export_df, order, r_names, groups = self._build_export(data_new, cfg)
        return {"cfg": cfg, "rows": rows, "ghs": ghs, "info": info,
                "export_df": export_df, "order": order,
                "r_names": r_names, "groups": groups,
                "cf_key": (cfg["cf_path"], cfg["cf_sheet"])}

    # ---------- RUN ----------
    def run(self):
        if not self.raw_path:
            QMessageBox.warning(self, "ยังไม่ครบ", "กรุณาเลือกไฟล์ Rawdata")
            return
        # รวม Set ที่บันทึก + ชุดปัจจุบัน (ถ้ามี)
        jobs = list(self.sets)
        if self.selected_concept_cols():
            if not self.cf_path:
                QMessageBox.warning(self, "ยังไม่ครบ", "กรุณาเลือกไฟล์ Codeframe")
                return
            jobs.append(self.current_config(f"Set {len(self.sets) + 1}"))
        if not jobs:
            QMessageBox.warning(self, "ยังไม่ครบ",
                                "กรุณากำหนด Concept หรือบันทึก Set อย่างน้อย 1 ชุด")
            return
        try:
            results = [self._process_job(c) for c in jobs]
        except Exception as e:
            QMessageBox.critical(self, "ประมวลผลผิดพลาด", str(e))
            return

        # Codeframe: รวมชีท โดยรวม Set ที่ codeframe (ไฟล์+ชีท) เดียวกันเข้าชีทเดียว
        cf_sheets, cf_index = [], {}
        for res in results:
            key = res["cf_key"]
            if key in cf_index:
                cf_index[key][0].append(res["cfg"]["name"])
            else:
                cf_index[key] = ([res["cfg"]["name"]], res)
                cf_sheets.append(key)
        codeframe_sheets = []
        for key in cf_sheets:
            names, res = cf_index[key]
            sheet_name = "+".join(n.replace("Set ", "S") for n in names)
            codeframe_sheets.append((sheet_name, res["rows"], res["ghs"]))

        out_dir = os.path.dirname(self.raw_path)
        cf_out = os.path.join(out_dir, "Codeframe_byConcept.xlsx")
        data_out = os.path.join(out_dir, "Data_byConcept.xlsx")
        try:
            write_codeframe_book(codeframe_sheets, cf_out)
            with pd.ExcelWriter(data_out, engine="openpyxl") as w:
                for res in results:
                    sn = res["cfg"]["name"].replace("Set ", "Set")
                    res["export_df"].to_excel(w, sheet_name=sn, index=False)
                    style_data_sheet(w.sheets[sn], res["order"],
                                     res["r_names"], res["groups"])
        except PermissionError:
            QMessageBox.critical(
                self, "บันทึกไม่ได้",
                "ไฟล์ผลลัพธ์อาจเปิดค้างอยู่ใน Excel — กรุณาปิดแล้วลองใหม่")
            return
        except Exception as e:
            QMessageBox.critical(self, "บันทึกผิดพลาด", str(e))
            return

        # พรีวิว: เก็บ codeframe ทุกชีทไว้สลับดู
        self.preview_cfs = [(nm, rows, ghs) for nm, rows, ghs in codeframe_sheets]
        self.cmb_preview.blockSignals(True)
        self.cmb_preview.clear()
        self.cmb_preview.addItems([nm for nm, _, _ in self.preview_cfs])
        self.cmb_preview.blockSignals(False)
        self.cmb_preview.setVisible(len(self.preview_cfs) > 1)
        if self.preview_cfs:
            self.fill_preview(self.preview_cfs[0][1], self.preview_cfs[0][2])

        tot_codes = sum(sum(1 for r in res["rows"] if r["kind"] == "concept")
                        for res in results)
        self.status.setText(
            f"✓ สำเร็จ • {len(results)} Set • {len(codeframe_sheets)} Codeframe sheet • "
            f"{tot_codes} code ใหม่ • บันทึก: {os.path.basename(cf_out)} "
            f"({len(codeframe_sheets)} ชีท), {os.path.basename(data_out)} ({len(results)} ชีท)")
        QMessageBox.information(
            self, "เสร็จแล้ว",
            f"ประมวลผล {len(results)} Set เรียบร้อย:\n\n"
            f"• {data_out}  ({len(results)} ชีท)\n"
            f"• {cf_out}  ({len(codeframe_sheets)} ชีท)")

    def on_preview_changed(self, idx):
        if 0 <= idx < len(self.preview_cfs):
            _, rows, ghs = self.preview_cfs[idx]
            self.fill_preview(rows, ghs)

    def fill_preview(self, rows, group_headers=None):
        ghs = list(group_headers or [])
        cols = ["Code"] + ghs + ["English", "CodeOriginal", "Frequenzy"]
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)
        hdr = self.table.horizontalHeader()
        eng_idx = 1 + len(ghs)
        # English = ยืดกว้างสุด ; คอลัมน์อื่นกว้างคงที่ (กันคอลัมน์กลุ่มกินพื้นที่)
        widths = {0: 55, eng_idx + 1: 95, eng_idx + 2: 90}   # Code, CodeOriginal, Frequenzy
        for gi in range(len(ghs)):
            widths[1 + gi] = 110                             # Thaigroup*
        for c in range(len(cols)):
            if c == eng_idx:
                hdr.setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)
            else:
                hdr.setSectionResizeMode(c, QHeaderView.ResizeMode.Interactive)
                self.table.setColumnWidth(c, widths.get(c, 100))

        self.table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            color = fill_for_kind(r["kind"])
            vals = ([("" if r["code"] is None else str(r["code"]))]
                    + [("" if g is None else str(g)) for g in (r.get("tg") or [])]
                    + [str(r["english"]),
                       "" if r.get("orig") is None else str(r["orig"]),
                       "" if r.get("freq") is None else str(r["freq"])])
            for c, val in enumerate(vals):
                it = QTableWidgetItem(val)
                if c != eng_idx:
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if color:
                    it.setBackground(QColor("#" + color))
                    f = it.font(); f.setBold(True); it.setFont(f)
                elif c == len(cols) - 1 and val:        # Frequenzy สีแดง
                    it.setForeground(QColor("#C00000"))
                self.table.setItem(i, c, it)


def main():
    # บังคับใช้เลขอารบิก (0-9) ไม่ให้ขึ้นเลขไทยตาม locale ระบบ
    QLocale.setDefault(QLocale(QLocale.Language.English, QLocale.Country.UnitedStates))
    app = QApplication(sys.argv)
    app.setWindowIcon(_logo_icon())
    w = MainWindow()
    # ขนาดให้พอดีจอ: กว้างพอ + สูงเท่าที่เนื้อหาฝั่งซ้ายต้องการ (ไม่เกินจอ)
    scr = app.primaryScreen().availableGeometry()
    need_h = w.left_panel.sizeHint().height() + 90
    width = min(1200, scr.width() - 80)
    height = min(max(need_h, 600), scr.height() - 80)
    w.resize(width, height)
    w.show()
    center_on_screen(w)
    sys.exit(app.exec())


#if __name__ == "__main__":
    #main()


# <<< START OF CHANGES >>>
# --- ฟังก์ชัน Entry Point ใหม่ (สำหรับให้ Launcher เรียก) ---
def run_this_app(working_dir=None): # ชื่อฟังก์ชันนี้จะถูกใช้ใน Launcher
    """
    ฟังก์ชันหลักสำหรับสร้างและรัน QuotaSamplerApp.
    """
    print(f"--- QUOTA_SAMPLER_INFO: Starting 'QuotaSamplerApp' via run_this_app() ---")
    try:
    # --- ส่วนที่ใช้รันโปรแกรม ---
    #if __name__ == "__main__":
    #if __name__ == "__main__":
        main()


        print(f"--- QUOTA_SAMPLER_INFO: QuotaSamplerApp mainloop finished. ---")

    except Exception as e:
        # ดักจับ Error ที่อาจเกิดขึ้นระหว่างการสร้างหรือรัน App
        print(f"QUOTA_SAMPLER_ERROR: An error occurred during QuotaSamplerApp execution: {e}")
        # แสดง Popup ถ้ามีปัญหา
        if 'root' not in locals() or not root.winfo_exists(): # สร้าง root ชั่วคราวถ้ายังไม่มี
            root_temp = tk.Tk()
            root_temp.withdraw()
            messagebox.showerror("Application Error (Quota Sampler)",
                                f"An unexpected error occurred:\n{e}", parent=root_temp)
            root_temp.destroy()
        else:
            messagebox.showerror("Application Error (Quota Sampler)",
                                f"An unexpected error occurred:\n{e}", parent=root) # ใช้ root ที่มีอยู่ถ้าเป็นไปได้
        sys.exit(f"Error running QuotaSamplerApp: {e}") # อาจจะ exit หรือไม่ก็ได้ ขึ้นกับการออกแบบ


# --- ส่วน Run Application เมื่อรันไฟล์นี้โดยตรง (สำหรับ Test) ---
if __name__ == "__main__":
    print("--- Running QuotaSamplerApp.py directly for testing ---")
    # (ถ้ามีการตั้งค่า DPI ด้านบน มันจะทำงานอัตโนมัติ)

    # เรียกฟังก์ชัน Entry Point ที่เราสร้างขึ้น
    run_this_app()

    print("--- Finished direct execution of QuotaSamplerApp.py ---")
# <<< END OF CHANGES >>>